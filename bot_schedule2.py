import logging
import os
import threading
import asyncio
import requests
import calendar
import csv
import hashlib
import base64
import hmac
import secrets
import jwt
from datetime import datetime
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from io import BytesIO, StringIO
from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.dispatcher import FSMContext
from flask import Flask, request, jsonify, send_file, g, redirect
from flask_cors import CORS
from functools import wraps
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import xlsxwriter
import json
import html
from concurrent.futures import ThreadPoolExecutor
from database import db, TECHNICAL_ISSUE_REASONS
import uuid
from passlib.hash import pbkdf2_sha256
from werkzeug.utils import secure_filename
from google.cloud import storage as gcs_storage
import tempfile
from datetime import datetime, timedelta, date as dt_date, timezone
import time
import math
from urllib.parse import quote, urlparse, parse_qs
from zoneinfo import ZoneInfo
from ai_feed_back_service import generate_monthly_feedback_with_ai, generate_birthday_greeting_with_ai
try:
    from PIL import Image, ImageOps
except Exception:
    Image = None
    ImageOps = None

os.environ['TZ'] = 'Asia/Almaty'
time.tzset()

# === Логирование =====================================================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# === Переменные окружения =========================================================================================
API_TOKEN = os.getenv('BOT_TOKEN')
FLASK_API_KEY = os.getenv('FLASK_API_KEY')
super_admin_id = int(os.getenv('SUPER_ADMIN_ID', '0'))
super_admin_login = os.getenv('SUPER_ADMIN_LOGIN', 'admin4')
super_admin_password = os.getenv('SUPER_ADMIN_PASSWORD', 'admin1234')

if not API_TOKEN:
    raise Exception("Переменная окружения BOT_TOKEN обязательна.")
if not FLASK_API_KEY:
    raise Exception("Переменная окружения FLASK_API_KEY обязательна.")

# === Инициализация бота и диспетчера =============================================================================
bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot=bot, storage=storage)

# === Блокировки ==================================================================================================
report_lock = threading.Lock()
executor_pool = ThreadPoolExecutor(max_workers=4)

# === Flask-сервер ================================================================================================
app = Flask(__name__)
JWT_SECRET = os.getenv('JWT_SECRET') or FLASK_API_KEY
JWT_ALGORITHM = "HS256"
JWT_ACCESS_TOKEN_MINUTES = int(os.getenv('JWT_ACCESS_TOKEN_MINUTES', '30'))
JWT_REFRESH_TOKEN_DAYS = int(os.getenv('JWT_REFRESH_TOKEN_DAYS', '30'))
JWT_ACCESS_COOKIE_NAME = os.getenv('JWT_ACCESS_COOKIE_NAME', 'otp_access_token')
JWT_REFRESH_COOKIE_NAME = os.getenv('JWT_REFRESH_COOKIE_NAME', 'otp_refresh_token')
JWT_REFRESH_HEADER_NAME = os.getenv('JWT_REFRESH_HEADER_NAME', 'X-Refresh-Token')
JWT_COOKIE_DOMAIN = os.getenv('JWT_COOKIE_DOMAIN', None)
JWT_COOKIE_SAMESITE = os.getenv('JWT_COOKIE_SAMESITE', 'None')
JWT_COOKIE_SECURE = os.getenv('JWT_COOKIE_SECURE', 'true').lower() == 'true'
JWT_COOKIE_PARTITIONED = os.getenv('JWT_COOKIE_PARTITIONED', 'true').lower() == 'true'
JWT_TOKEN_PEPPER = os.getenv('JWT_TOKEN_PEPPER', JWT_SECRET)
SENSITIVE_QR_SECRET = os.getenv('SENSITIVE_QR_SECRET', JWT_SECRET)
SENSITIVE_QR_TTL_SECONDS = int(os.getenv('SENSITIVE_QR_TTL_SECONDS', '300'))
LMS_HEARTBEAT_SECONDS = int(os.getenv('LMS_HEARTBEAT_SECONDS', '15'))
LMS_STALE_GAP_SECONDS = int(os.getenv('LMS_STALE_GAP_SECONDS', '45'))
LMS_COMPLETION_THRESHOLD = float(os.getenv('LMS_COMPLETION_THRESHOLD', '95'))
LMS_DEFAULT_PASS_THRESHOLD = float(os.getenv('LMS_DEFAULT_PASS_THRESHOLD', '80'))
LMS_DEFAULT_ATTEMPT_LIMIT = int(os.getenv('LMS_DEFAULT_ATTEMPT_LIMIT', '3'))
LMS_CERTIFICATE_STORAGE = (os.getenv('LMS_CERTIFICATE_STORAGE') or 'db').strip().lower()

ALLOWED_ORIGINS = {
    "https://alfa330.github.io",
    "https://call-evalution.pages.dev",
    "https://szov.pages.dev",
    "https://moders.pages.dev",
    "https://table-7kx.pages.dev",
    "https://base-pmy9.onrender.com"
}

if not os.getenv('JWT_SECRET'):
    logging.warning("JWT_SECRET is not set. Falling back to FLASK_API_KEY for signing tokens.")


def _normalize_origin(origin):
    if not origin:
        return ""
    return str(origin).strip().rstrip("/")


def _is_allowed_origin(origin):
    origin = _normalize_origin(origin)
    if not origin:
        return False
    if origin in ALLOWED_ORIGINS:
        return True
    return origin.startswith("http://localhost:") or origin.startswith("http://127.0.0.1:")


CORS(app, resources={
    r"/api/*": {
        "origins": list(ALLOWED_ORIGINS) + [r"http://localhost:\d+", r"http://127\.0\.0\.1:\d+"],
        "methods": ["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "X-API-Key", "X-User-Id", "Authorization", "X-Refresh-Token"],
        "supports_credentials": True,
        "max_age": 86400
    }
})

def get_gcs_client():
    # Если используется JSON из переменной окружения
    credentials_content = os.getenv('GOOGLE_APPLICATION_CREDENTIALS_CONTENT')
    if credentials_content:
        import json
        from google.oauth2 import service_account
        credentials_info = json.loads(credentials_content)
        credentials = service_account.Credentials.from_service_account_info(credentials_info)
        return gcs_storage.Client(credentials=credentials)
    return gcs_storage.Client()


class AuthError(Exception):
    def __init__(self, code, message):
        super().__init__(message)
        self.code = code
        self.message = message


def _client_ip():
    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.remote_addr


def _hash_token(raw_token):
    value = f"{raw_token}:{JWT_TOKEN_PEPPER}"
    return hashlib.sha256(value.encode('utf-8')).hexdigest()


def _get_user_payload(user):
    avatar_bucket = None
    avatar_blob_path = None
    avatar_updated_at = None
    gender = None
    role = None
    user_id = None
    name = None
    telegram_id = None
    if isinstance(user, (tuple, list)):
        role = user[3]
        user_id = user[0]
        name = user[2]
        telegram_id = user[1]
        if len(user) >= 20:
            gender = user[13]
            avatar_bucket = user[15]
            avatar_blob_path = user[16]
            avatar_updated_at = user[19]
        elif len(user) >= 17:
            gender = user[16]
            avatar_bucket = user[11]
            avatar_blob_path = user[12]
            avatar_updated_at = user[15]
        elif len(user) >= 16:
            avatar_bucket = user[11]
            avatar_blob_path = user[12]
            avatar_updated_at = user[15]
    elif isinstance(user, dict):
        role = user.get("role")
        user_id = user.get("id")
        name = user.get("name")
        telegram_id = user.get("telegram_id")
        gender = user.get("gender")
        avatar_bucket = user.get("avatar_bucket")
        avatar_blob_path = user.get("avatar_blob_path")
        avatar_updated_at = user.get("avatar_updated_at")
    return {
        "role": role,
        "id": user_id,
        "name": name,
        "telegram_id": telegram_id,
        "gender": gender,
        "avatar_url": _build_avatar_signed_url(avatar_bucket, avatar_blob_path),
        "avatar_updated_at": avatar_updated_at.isoformat() if hasattr(avatar_updated_at, "isoformat") else avatar_updated_at
    }


def _set_request_auth_context(user_id):
    request.environ['HTTP_X_USER_ID'] = str(user_id)
    request.environ['HTTP_X_API_KEY'] = FLASK_API_KEY
    request.user_id = int(user_id)
    g.user_id = int(user_id)


def _decode_token(token, expected_type, verify_exp=True):
    try:
        payload = jwt.decode(
            token,
            JWT_SECRET,
            algorithms=[JWT_ALGORITHM],
            options={"verify_exp": verify_exp}
        )
    except jwt.ExpiredSignatureError as e:
        raise AuthError("TOKEN_EXPIRED", "JWT token expired") from e
    except jwt.InvalidTokenError as e:
        raise AuthError("INVALID_TOKEN", "Invalid JWT token") from e

    token_type = payload.get("type")
    if token_type != expected_type:
        raise AuthError("INVALID_TOKEN_TYPE", "Invalid token type")
    if not payload.get("sub") or not payload.get("sid"):
        raise AuthError("INVALID_TOKEN", "Malformed JWT token")
    return payload


def _build_access_token(user, session_id):
    now = datetime.utcnow()
    payload = {
        "sub": str(user[0]),
        "role": user[3],
        "name": user[2],
        "sid": str(session_id),
        "type": "access",
        "iat": int(now.timestamp()),
        "exp": int((now + timedelta(minutes=JWT_ACCESS_TOKEN_MINUTES)).timestamp())
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def _build_refresh_token(user_id, session_id):
    now = datetime.utcnow()
    payload = {
        "sub": str(user_id),
        "sid": str(session_id),
        "type": "refresh",
        "iat": int(now.timestamp()),
        "exp": int((now + timedelta(days=JWT_REFRESH_TOKEN_DAYS)).timestamp())
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def _cookie_options():
    secure = JWT_COOKIE_SECURE
    origin = request.headers.get('Origin', '')
    if origin.startswith("http://localhost:") or origin.startswith("http://127.0.0.1:"):
        secure = False

    samesite = JWT_COOKIE_SAMESITE
    if not secure and str(samesite).lower() == 'none':
        samesite = 'Lax'

    partitioned = bool(JWT_COOKIE_PARTITIONED and secure and str(samesite).lower() == 'none')

    return {
        "httponly": True,
        "secure": secure,
        "samesite": samesite,
        "domain": JWT_COOKIE_DOMAIN,
        "path": "/",
        "partitioned": partitioned
    }


def _serialize_cookie(name, value, max_age, cookie_kwargs, expires=None):
    safe_value = quote(value, safe="!#$%&'()*+-./:<=>?@[]^_`{|}~")
    parts = [f"{name}={safe_value}"]
    parts.append(f"Path={cookie_kwargs['path']}")
    parts.append(f"Max-Age={int(max_age)}")
    if cookie_kwargs.get('domain'):
        parts.append(f"Domain={cookie_kwargs['domain']}")
    if expires:
        parts.append(f"Expires={expires}")
    if cookie_kwargs.get('httponly'):
        parts.append("HttpOnly")
    if cookie_kwargs.get('secure'):
        parts.append("Secure")
    if cookie_kwargs.get('samesite'):
        parts.append(f"SameSite={cookie_kwargs['samesite']}")
    if cookie_kwargs.get('partitioned'):
        parts.append("Partitioned")
    return "; ".join(parts)


def _set_cookie_header(response, name, value, max_age):
    cookie_kwargs = _cookie_options()
    cookie_header = _serialize_cookie(name, value, max_age=max_age, cookie_kwargs=cookie_kwargs)
    response.headers.add("Set-Cookie", cookie_header)


def _clear_cookie_header(response, name):
    cookie_kwargs = _cookie_options()
    cookie_header = _serialize_cookie(
        name,
        '',
        max_age=0,
        cookie_kwargs=cookie_kwargs,
        expires="Thu, 01 Jan 1970 00:00:00 GMT"
    )
    response.headers.add("Set-Cookie", cookie_header)


def _set_auth_cookies(response, access_token, refresh_token):
    _set_cookie_header(response, JWT_ACCESS_COOKIE_NAME, access_token, JWT_ACCESS_TOKEN_MINUTES * 60)
    _set_cookie_header(response, JWT_REFRESH_COOKIE_NAME, refresh_token, JWT_REFRESH_TOKEN_DAYS * 24 * 60 * 60)


def _clear_auth_cookies(response):
    _clear_cookie_header(response, JWT_ACCESS_COOKIE_NAME)
    _clear_cookie_header(response, JWT_REFRESH_COOKIE_NAME)


def _get_cookie_values(cookie_name):
    # request.cookies may behave inconsistently if duplicate cookie names exist.
    # Parse raw Cookie header and keep all occurrences.
    raw_cookie = request.headers.get('Cookie', '') or ''
    values = []
    for chunk in raw_cookie.split(';'):
        part = chunk.strip()
        if not part or '=' not in part:
            continue
        key, val = part.split('=', 1)
        if key.strip() == cookie_name:
            values.append(val.strip())
    if not values:
        fallback = request.cookies.get(cookie_name)
        if fallback:
            values.append(fallback)
    return values


def _get_cookie_value(cookie_name):
    values = _get_cookie_values(cookie_name)
    if not values:
        return None
    return values[-1]


def _get_bearer_access_token():
    auth_header = (request.headers.get('Authorization') or '').strip()
    if not auth_header:
        return None
    parts = auth_header.split(' ', 1)
    if len(parts) != 2 or parts[0].strip().lower() != 'bearer':
        return None
    token = parts[1].strip()
    return token or None


def _get_refresh_header_token():
    token = (request.headers.get(JWT_REFRESH_HEADER_NAME) or '').strip()
    return token or None


def _get_access_token_values():
    tokens = _get_cookie_values(JWT_ACCESS_COOKIE_NAME)
    header_token = _get_bearer_access_token()
    if header_token:
        tokens.append(header_token)
    return tokens


def _get_refresh_token_values():
    tokens = _get_cookie_values(JWT_REFRESH_COOKIE_NAME)
    header_token = _get_refresh_header_token()
    if header_token:
        tokens.append(header_token)
    return tokens


def _authenticate_access_cookie(optional=True, touch_session=True):
    access_tokens = _get_access_token_values()
    if not access_tokens:
        if optional:
            return None
        raise AuthError("MISSING_TOKEN", "Missing access token")

    last_error = AuthError("INVALID_TOKEN", "Invalid access token")
    for access_token in reversed(access_tokens):
        try:
            payload = _decode_token(access_token, expected_type="access", verify_exp=True)
            user_id = int(payload["sub"])
            session_id = str(payload["sid"])
            session = db.get_user_session(session_id=session_id, user_id=user_id)
            if not session:
                raise AuthError("SESSION_NOT_FOUND", "Session not found")
            if session["revoked_at"] is not None:
                raise AuthError("SESSION_REVOKED", "Session revoked")
            if session["expires_at"] and session["expires_at"] < datetime.utcnow():
                raise AuthError("SESSION_EXPIRED", "Session expired")

            if touch_session:
                db.touch_user_session(
                    session_id=session_id,
                    user_id=user_id,
                    ip_address=_client_ip(),
                    user_agent=request.headers.get('User-Agent')
                )

            _set_request_auth_context(user_id)
            return payload
        except AuthError as auth_error:
            last_error = auth_error
            continue

    raise last_error


def _authenticate_refresh_cookie(optional=True, rotate_tokens=True):
    refresh_tokens = _get_refresh_token_values()
    if not refresh_tokens:
        if optional:
            return None
        raise AuthError("MISSING_REFRESH_TOKEN", "Missing refresh token")

    last_error = AuthError("INVALID_TOKEN", "Invalid refresh token")
    for refresh_token in reversed(refresh_tokens):
        try:
            payload = _decode_token(refresh_token, expected_type="refresh", verify_exp=True)
            user_id = int(payload["sub"])
            session_id = str(payload["sid"])

            session = db.get_user_session(session_id=session_id, user_id=user_id)
            if not session:
                raise AuthError("SESSION_NOT_FOUND", "Session not found")
            if session["revoked_at"] is not None:
                raise AuthError("SESSION_REVOKED", "Session revoked")
            if session["expires_at"] and session["expires_at"] < datetime.utcnow():
                raise AuthError("SESSION_EXPIRED", "Session expired")
            if session["refresh_token_hash"] != _hash_token(refresh_token):
                raise AuthError("REFRESH_TOKEN_MISMATCH", "Refresh token mismatch")

            user = db.get_user(id=user_id)
            if not user:
                raise AuthError("USER_NOT_FOUND", "User not found")

            if rotate_tokens:
                new_access_token = _build_access_token(user, session_id)
                new_refresh_token = _build_refresh_token(user_id, session_id)
                db.rotate_user_session_token(
                    session_id=session_id,
                    user_id=user_id,
                    refresh_token_hash=_hash_token(new_refresh_token),
                    expires_at=datetime.utcnow() + timedelta(days=JWT_REFRESH_TOKEN_DAYS)
                )
                g.pending_auth_tokens = (new_access_token, new_refresh_token)
            else:
                db.touch_user_session(
                    session_id=session_id,
                    user_id=user_id,
                    ip_address=_client_ip(),
                    user_agent=request.headers.get('User-Agent')
                )

            _set_request_auth_context(user_id)
            return payload
        except AuthError as auth_error:
            last_error = auth_error
            continue

    raise last_error


def _current_session_id_from_access_token():
    access_token = _get_cookie_value(JWT_ACCESS_COOKIE_NAME) or _get_bearer_access_token()
    if not access_token:
        return None
    try:
        payload = _decode_token(access_token, expected_type="access", verify_exp=False)
        return str(payload.get("sid"))
    except AuthError:
        return None


def _base64url_encode(value: bytes) -> str:
    return base64.urlsafe_b64encode(value).decode('utf-8').rstrip('=')


def _base64url_decode(value: str) -> bytes:
    padding = '=' * (-len(value) % 4)
    return base64.urlsafe_b64decode((value + padding).encode('utf-8'))


def _normalize_user_role(role) -> str:
    role_norm = str(role or '').strip().lower()
    if role_norm == 'supervisor':
        return 'sv'
    return role_norm


ROLE_HIERARCHY = {
    'operator': 10,
    'trainee': 10,
    'trainer': 20,
    'sv': 30,
    'admin': 40,
    'super_admin': 50
}


def _get_role_level(role) -> int:
    return int(ROLE_HIERARCHY.get(_normalize_user_role(role), 0))


def _has_min_role(role, required_role) -> bool:
    required_level = _get_role_level(required_role)
    if required_level <= 0:
        return False
    return _get_role_level(role) >= required_level


def _has_any_role(role, allowed_roles) -> bool:
    role_norm = _normalize_user_role(role)
    if not role_norm:
        return False
    normalized_allowed = {_normalize_user_role(item) for item in (allowed_roles or [])}
    return role_norm in normalized_allowed


def _is_super_admin_role(role) -> bool:
    return _normalize_user_role(role) == 'super_admin'


def _is_admin_role(role) -> bool:
    return _has_min_role(role, 'admin')


def _is_privileged_role(role: str) -> bool:
    return _has_min_role(role, 'sv')


def _mask_phone_number(phone_number):
    if phone_number is None:
        return None
    raw = str(phone_number).strip()
    if not raw:
        return raw
    digits = re.sub(r'\D', '', raw)
    if not digits:
        return '*' * min(len(raw), 6)
    if len(digits) <= 4:
        return '*' * len(digits)
    return ('*' * (len(digits) - 4)) + digits[-4:]


KZ_PHONE_REGEX = re.compile(r'^\+7\d{10}$')


def _is_valid_kz_phone(phone_number):
    if phone_number is None:
        return True
    value = str(phone_number).strip()
    if not value:
        return True
    return bool(KZ_PHONE_REGEX.match(value))


def _build_sensitive_qr_token(session_id, user_id):
    expires_at = datetime.utcnow() + timedelta(seconds=SENSITIVE_QR_TTL_SECONDS)
    payload = {
        "sid": str(session_id),
        "uid": int(user_id),
        "exp": int(expires_at.timestamp()),
        "nonce": uuid.uuid4().hex
    }
    payload_json = json.dumps(payload, separators=(',', ':'), sort_keys=True).encode('utf-8')
    payload_b64 = _base64url_encode(payload_json)
    signature = hmac.new(
        SENSITIVE_QR_SECRET.encode('utf-8'),
        payload_b64.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    token = f"{payload_b64}.{signature}"
    return token, expires_at


def _decode_sensitive_qr_token(token):
    if not token or '.' not in token:
        raise ValueError("Invalid QR token format")

    payload_b64, signature = token.rsplit('.', 1)
    expected_signature = hmac.new(
        SENSITIVE_QR_SECRET.encode('utf-8'),
        payload_b64.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()

    if not hmac.compare_digest(signature, expected_signature):
        raise ValueError("Invalid QR token signature")

    try:
        payload_raw = _base64url_decode(payload_b64)
        payload = json.loads(payload_raw.decode('utf-8'))
    except Exception as exc:
        raise ValueError("Invalid QR token payload") from exc

    session_id = str(payload.get('sid') or '').strip()
    user_id = payload.get('uid')
    exp_ts = payload.get('exp')

    try:
        user_id = int(user_id)
        exp_ts = int(exp_ts)
    except Exception as exc:
        raise ValueError("Invalid QR token claims") from exc

    if not session_id:
        raise ValueError("Invalid QR token session")

    now_ts = int(datetime.utcnow().timestamp())
    if exp_ts <= now_ts:
        raise ValueError("QR token expired")

    return {
        "session_id": session_id,
        "user_id": user_id,
        "expires_at": datetime.utcfromtimestamp(exp_ts)
    }


def _is_sensitive_access_unlocked(user_id, session_id):
    if not user_id or not session_id:
        return False
    try:
        return db.is_session_sensitive_access_unlocked(session_id=session_id, user_id=user_id)
    except Exception as exc:
        logging.error(f"Failed to check sensitive session access: {exc}")
        return False


def _sanitize_evaluations_for_access(evaluations, reveal_sensitive, hide_hidden_operator_comments=False):
    result = []
    for ev in (evaluations or []):
        if not isinstance(ev, dict):
            result.append(ev)
            continue

        item = dict(ev)
        phone_masked = _mask_phone_number(item.get('phone_number'))
        item['phone_number_masked'] = phone_masked
        if not reveal_sensitive:
            item['phone_number'] = phone_masked
            item['audio_path'] = None
        if hide_hidden_operator_comments and item.get('comment_visible_to_operator') is False:
            item['comment'] = None
        result.append(item)
    return result


def _authorize_operator_scope(requester, requester_id, operator_id):
    role = _normalize_user_role(requester[3])
    if _is_admin_role(role) or role == 'sv':
        return True
    if role == 'operator':
        return requester_id == operator_id
    return False


def _ensure_call_access_for_requester(call_operator_id, requester, requester_id):
    role = _normalize_user_role(requester[3])
    if _is_admin_role(role):
        return True
    if role == 'operator':
        return requester_id == call_operator_id
    if role == 'sv':
        operator = db.get_user(id=call_operator_id)
        return bool(operator and operator[3] == 'operator' and operator[6] == requester_id)
    return False


def _is_supervisor_role(role: str) -> bool:
    return _normalize_user_role(role) == 'sv'


def _normalize_management_role(role):
    return _normalize_user_role(role)


def _normalize_int_id_list(values):
    if values is None:
        return []
    if isinstance(values, (list, tuple, set)):
        raw_values = list(values)
    else:
        raw_values = [values]

    result = []
    seen = set()
    for item in raw_values:
        if item is None:
            continue
        if isinstance(item, str) and not item.strip():
            continue
        try:
            ivalue = int(item)
        except (TypeError, ValueError):
            continue
        if ivalue <= 0 or ivalue in seen:
            continue
        seen.add(ivalue)
        result.append(ivalue)
    return result


def _normalize_calibration_score_value(value):
    raw = str(value or '').strip().lower()
    mapping = {
        'correct': 'Correct',
        'ok': 'Correct',
        'n/a': 'N/A',
        'na': 'N/A',
        'n\\a': 'N/A',
        'incorrect': 'Incorrect',
        'deficiency': 'Deficiency',
        'error': 'Error'
    }
    return mapping.get(raw, str(value or 'Correct').strip() or 'Correct')


def _calibration_label_score(value):
    normalized = _normalize_calibration_score_value(value)
    labels = {
        'Correct': 'Корректно',
        'N/A': 'N/A',
        'Incorrect': 'Ошибка',
        'Deficiency': 'Недочет',
        'Error': 'Критич. ошибка'
    }
    return labels.get(normalized, normalized)


def _compute_total_score_from_criteria(criteria, scores):
    criteria = criteria if isinstance(criteria, list) else []
    scores = scores if isinstance(scores, list) else []
    has_critical_error = False
    total = 0.0

    for idx, criterion in enumerate(criteria):
        criterion = criterion if isinstance(criterion, dict) else {}
        score_value = _normalize_calibration_score_value(scores[idx] if idx < len(scores) else 'Correct')
        is_critical = bool(criterion.get('isCritical'))
        if is_critical:
            if score_value == 'Error':
                has_critical_error = True
            continue
        try:
            weight = float(criterion.get('weight') or 0)
        except Exception:
            weight = 0.0
        if score_value in ('Correct', 'N/A'):
            total += weight
        elif score_value == 'Deficiency':
            deficiency = criterion.get('deficiency') if isinstance(criterion.get('deficiency'), dict) else {}
            try:
                total += float(deficiency.get('weight') or 0)
            except Exception:
                pass

    if has_critical_error:
        return 0.0, True
    return round(float(total), 2), False


def _build_calibration_results(criteria, etalon_scores, etalon_comments, evaluations, admin_scores=None, admin_comments=None):
    criteria = criteria if isinstance(criteria, list) else []
    etalon_scores = etalon_scores if isinstance(etalon_scores, list) else []
    etalon_comments = etalon_comments if isinstance(etalon_comments, list) else []
    admin_scores = admin_scores if isinstance(admin_scores, list) else etalon_scores
    admin_comments = admin_comments if isinstance(admin_comments, list) else etalon_comments
    evaluations = evaluations if isinstance(evaluations, list) else []

    eval_count = len(evaluations)
    critical_mismatch = False
    rows = []

    for idx, criterion in enumerate(criteria):
        criterion = criterion if isinstance(criterion, dict) else {}
        is_critical = bool(criterion.get('isCritical'))
        admin_val = _normalize_calibration_score_value(admin_scores[idx] if idx < len(admin_scores) else 'Correct')
        etalon_val = _normalize_calibration_score_value(etalon_scores[idx] if idx < len(etalon_scores) else admin_val)
        # Percent is based on overall agreement between all participants:
        # admin + all supervisors. Etalon is displayed separately and does not
        # affect calibration percent.
        score_counts = {admin_val: 1}
        by_evaluator = []

        for evaluation in evaluations:
            ev_scores = evaluation.get('scores') if isinstance(evaluation.get('scores'), list) else []
            ev_comments = evaluation.get('criterion_comments') if isinstance(evaluation.get('criterion_comments'), list) else []
            ev_val = _normalize_calibration_score_value(ev_scores[idx] if idx < len(ev_scores) else 'Correct')
            score_counts[ev_val] = int(score_counts.get(ev_val, 0)) + 1
            is_match = ev_val == admin_val

            by_evaluator.append({
                "evaluator_id": evaluation.get('evaluator_id'),
                "score": ev_val,
                "score_label": _calibration_label_score(ev_val),
                "comment": (ev_comments[idx] if idx < len(ev_comments) else None) or None,
                "is_match": bool(is_match)
            })

        percent = None
        if eval_count > 0:
            participants_count = 1 + eval_count
            dominant_count = max(score_counts.values()) if score_counts else 0
            percent = round((dominant_count * 100.0) / float(participants_count), 1)
            if is_critical and dominant_count < participants_count:
                critical_mismatch = True
                percent = 0.0

        rows.append({
            "criterion_index": idx,
            "criterion_name": criterion.get('name') or f"Критерий {idx + 1}",
            "is_critical": is_critical,
            "percent": percent,
            "etalon": {
                "score": etalon_val,
                "score_label": _calibration_label_score(etalon_val),
                "comment": (etalon_comments[idx] if idx < len(etalon_comments) else None) or None
            },
            "admin": {
                "score": admin_val,
                "score_label": _calibration_label_score(admin_val),
                "comment": (admin_comments[idx] if idx < len(admin_comments) else None) or None
            },
            "benchmark": {
                "score": etalon_val,
                "score_label": _calibration_label_score(etalon_val),
                "comment": (etalon_comments[idx] if idx < len(etalon_comments) else None) or None
            },
            "by_evaluator": by_evaluator
        })

    overall_percent = None
    if rows and eval_count > 0:
        perc_values = [r.get('percent') for r in rows if r.get('percent') is not None]
        if perc_values:
            overall_percent = round(sum(perc_values) / float(len(perc_values)), 1)
    if critical_mismatch and overall_percent is not None:
        overall_percent = 0.0

    return {
        "criteria_rows": rows,
        "evaluated_count": eval_count,
        "critical_mismatch": critical_mismatch,
        "overall_percent": overall_percent
    }


def _build_signed_audio_url(audio_path):
    if not audio_path:
        return None
    try:
        path_parts = str(audio_path).split('/', 1)
        if len(path_parts) != 2:
            return None
        bucket_name, blob_path = path_parts
        bucket = get_gcs_client().bucket(bucket_name)
        blob = bucket.blob(blob_path)
        if not blob.exists():
            return None
        return blob.generate_signed_url(
            version="v4",
            expiration=timedelta(minutes=15),
            method="GET",
            response_type='audio/mpeg'
        )
    except Exception as e:
        logging.error("Error building signed audio URL: %s", e)
        return None


TASK_ALLOWED_TAGS = {'task', 'problem', 'suggestion'}
TASK_ALLOWED_ACTIONS = {'in_progress', 'completed', 'accepted', 'returned', 'reopened'}
TASK_MAX_FILES = 10
TASK_MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
TASK_ATTACHMENTS_UPLOAD_FOLDER = (os.getenv('TASK_ATTACHMENTS_UPLOAD_FOLDER') or 'TaskAttachments/').strip()
TELEGRAM_MAX_MESSAGE_CHARS = 4096
TELEGRAM_MAX_CAPTION_CHARS = 1024
TASK_TELEGRAM_TIMEOUT_SECONDS = 20
AVATAR_MAX_UPLOAD_BYTES = int(os.getenv('AVATAR_MAX_UPLOAD_BYTES', str(2 * 1024 * 1024)))
AVATAR_MAX_ORIGINAL_UPLOAD_BYTES = int(os.getenv('AVATAR_MAX_ORIGINAL_UPLOAD_BYTES', str(10 * 1024 * 1024)))
AVATAR_SIGNED_URL_TTL_SECONDS = int(os.getenv('AVATAR_SIGNED_URL_TTL_SECONDS', str(24 * 60 * 60)))
AVATAR_SIGNED_URL_CACHE_MAX_ITEMS = int(os.getenv('AVATAR_SIGNED_URL_CACHE_MAX_ITEMS', '20000'))
AVATAR_SIGNED_URL_CACHE_MIN_REMAINING_SECONDS = int(os.getenv('AVATAR_SIGNED_URL_CACHE_MIN_REMAINING_SECONDS', '60'))
AVATAR_UPLOAD_FOLDER = (os.getenv('AVATAR_UPLOAD_FOLDER') or 'AvatarUploads/').strip()
AVATAR_THUMBNAIL_SUFFIX = (os.getenv('AVATAR_THUMBNAIL_SUFFIX') or '128').strip() or '128'
AVATAR_SIGNED_URL_CACHE = {}
AVATAR_SIGNED_URL_CACHE_LOCK = threading.Lock()
TASK_TAG_LABELS = {
    'task': 'Задача',
    'problem': 'Проблема',
    'suggestion': 'Предложение'
}
SENSITIVE_ACCESS_ROLE_LABELS = {
    'super_admin': 'Супер админ',
    'admin': 'Админ',
    'sv': 'Супервайзер',
    'operator': 'Оператор',
    'trainee': 'Стажер',
    'trainer': 'Тренер'
}
try:
    SENSITIVE_ACCESS_NOTIFICATION_TZ = ZoneInfo('Asia/Almaty')
except Exception:
    SENSITIVE_ACCESS_NOTIFICATION_TZ = None


def _truncate_for_telegram(text, max_chars):
    value = str(text or '').strip()
    if not value:
        return ''
    if len(value) <= max_chars:
        return value
    suffix = '...'
    available = max(0, max_chars - len(suffix))
    return f"{value[:available].rstrip()}{suffix}"


def _escape_telegram_html(text, max_chars=None):
    value = str(text or '').strip()
    if max_chars is not None:
        value = _truncate_for_telegram(value, max_chars)
    return html.escape(value)


def _get_telegram_error_text(response):
    try:
        payload = response.json()
        if isinstance(payload, dict):
            description = payload.get('description')
            if description:
                return str(description)
    except Exception:
        pass
    raw_text = (response.text or '').strip()
    if raw_text:
        return _truncate_for_telegram(raw_text, 300)
    return f"HTTP {response.status_code}"


def _send_telegram_text_message(chat_id, text, parse_mode='HTML'):
    telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
    message_text = str(text or '')
    if parse_mode != 'HTML':
        message_text = _truncate_for_telegram(message_text, TELEGRAM_MAX_MESSAGE_CHARS)
    payload = {
        "chat_id": int(chat_id),
        "text": message_text,
        "disable_web_page_preview": True
    }
    if parse_mode:
        payload["parse_mode"] = parse_mode
    return requests.post(telegram_url, json=payload, timeout=TASK_TELEGRAM_TIMEOUT_SECONDS)


def _coerce_sensitive_access_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace('Z', '+00:00')
    try:
        return datetime.fromisoformat(normalized)
    except Exception:
        pass

    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d.%m.%Y %H:%M:%S'):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def _format_sensitive_access_notification_dt(value, assume_utc=False):
    dt_value = _coerce_sensitive_access_datetime(value)
    if dt_value is None:
        text = str(value or '').strip()
        return text or '—'

    try:
        tz = SENSITIVE_ACCESS_NOTIFICATION_TZ
        if tz is not None:
            if dt_value.tzinfo is None:
                if assume_utc:
                    dt_value = dt_value.replace(tzinfo=timezone.utc)
                else:
                    dt_value = dt_value.replace(tzinfo=tz)
            dt_value = dt_value.astimezone(tz)
        return dt_value.strftime('%d.%m.%Y %H:%M:%S (GMT+5)')
    except Exception:
        return dt_value.strftime('%d.%m.%Y %H:%M:%S')


def _parse_user_agent_details(user_agent):
    ua_raw = str(user_agent or '').strip()
    if not ua_raw:
        return {
            "device": "—",
            "os": "—",
            "browser": "—",
            "raw": "—"
        }

    ua = ua_raw.lower()

    device = 'ПК/ноутбук'
    if 'ipad' in ua or 'tablet' in ua:
        device = 'Планшет'
    elif 'iphone' in ua or 'ipod' in ua:
        device = 'Смартфон (iPhone)'
    elif 'android' in ua and 'mobile' in ua:
        device = 'Смартфон (Android)'
    elif 'android' in ua:
        device = 'Планшет/Android-устройство'
    elif 'mobile' in ua:
        device = 'Мобильное устройство'

    os_name = 'Не определена'
    if 'windows nt 10.0' in ua:
        os_name = 'Windows 10/11'
    elif 'windows nt 6.3' in ua:
        os_name = 'Windows 8.1'
    elif 'windows nt 6.2' in ua:
        os_name = 'Windows 8'
    elif 'windows nt 6.1' in ua:
        os_name = 'Windows 7'
    elif 'android' in ua:
        match = re.search(r'android\s([\d\.]+)', ua, re.IGNORECASE)
        os_name = f"Android {match.group(1)}" if match else 'Android'
    elif 'iphone' in ua or 'ipad' in ua or 'ios' in ua:
        match = re.search(r'os\s(\d+(?:[_\.]\d+)*)', ua, re.IGNORECASE)
        ios_version = match.group(1).replace('_', '.') if match else None
        os_name = f"iOS {ios_version}" if ios_version else 'iOS'
    elif 'mac os x' in ua or 'macintosh' in ua:
        match = re.search(r'mac os x\s(\d+(?:[_\.]\d+)*)', ua, re.IGNORECASE)
        mac_version = match.group(1).replace('_', '.') if match else None
        os_name = f"macOS {mac_version}" if mac_version else 'macOS'
    elif 'cros' in ua:
        os_name = 'ChromeOS'
    elif 'linux' in ua or 'x11' in ua:
        os_name = 'Linux'

    browser = 'Не определен'
    if 'yabrowser/' in ua:
        match = re.search(r'yabrowser/([0-9\.]+)', ua, re.IGNORECASE)
        browser = f"Yandex Browser {match.group(1)}" if match else 'Yandex Browser'
    elif 'edg/' in ua:
        match = re.search(r'edg/([0-9\.]+)', ua, re.IGNORECASE)
        browser = f"Microsoft Edge {match.group(1)}" if match else 'Microsoft Edge'
    elif 'opr/' in ua or 'opera' in ua:
        match = re.search(r'(?:opr|opera)/([0-9\.]+)', ua, re.IGNORECASE)
        browser = f"Opera {match.group(1)}" if match else 'Opera'
    elif 'samsungbrowser/' in ua:
        match = re.search(r'samsungbrowser/([0-9\.]+)', ua, re.IGNORECASE)
        browser = f"Samsung Browser {match.group(1)}" if match else 'Samsung Browser'
    elif 'crios/' in ua:
        match = re.search(r'crios/([0-9\.]+)', ua, re.IGNORECASE)
        browser = f"Chrome (iOS) {match.group(1)}" if match else 'Chrome (iOS)'
    elif 'chrome/' in ua and 'chromium' not in ua:
        match = re.search(r'chrome/([0-9\.]+)', ua, re.IGNORECASE)
        browser = f"Google Chrome {match.group(1)}" if match else 'Google Chrome'
    elif 'fxios/' in ua:
        match = re.search(r'fxios/([0-9\.]+)', ua, re.IGNORECASE)
        browser = f"Firefox (iOS) {match.group(1)}" if match else 'Firefox (iOS)'
    elif 'firefox/' in ua:
        match = re.search(r'firefox/([0-9\.]+)', ua, re.IGNORECASE)
        browser = f"Mozilla Firefox {match.group(1)}" if match else 'Mozilla Firefox'
    elif 'safari/' in ua and 'chrome/' not in ua and 'crios/' not in ua and 'edg/' not in ua:
        match = re.search(r'version/([0-9\.]+)', ua, re.IGNORECASE)
        browser = f"Safari {match.group(1)}" if match else 'Safari'
    elif 'telegrambot' in ua:
        browser = 'Telegram Bot'

    return {
        "device": device,
        "os": os_name,
        "browser": browser,
        "raw": ua_raw
    }


def _resolve_super_admin_chat_ids():
    chat_ids = []
    seen = set()

    def _add_chat_id(raw_chat_id):
        if raw_chat_id in (None, ''):
            return
        try:
            chat_id = int(str(raw_chat_id).strip())
        except Exception:
            return
        if chat_id <= 0 or chat_id in seen:
            return
        seen.add(chat_id)
        chat_ids.append(chat_id)

    _add_chat_id(super_admin_id)

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT telegram_id
                FROM users
                WHERE role = 'super_admin'
                  AND telegram_id IS NOT NULL
            """)
            for row in cursor.fetchall() or []:
                _add_chat_id(row[0] if isinstance(row, (tuple, list)) and row else row)
    except Exception as e:
        logging.error(f"Failed to resolve super admin chat ids: {e}", exc_info=True)

    return chat_ids


def _build_sensitive_access_approved_message_html(
    approver,
    operator,
    claims,
    operator_session,
    approval_context,
    operator_supervisor_name
):
    approval_context = approval_context or {}
    session = operator_session or {}
    claims = claims or {}

    operator_id = operator[0] if isinstance(operator, (tuple, list)) and len(operator) > 0 else None
    operator_tg = operator[1] if isinstance(operator, (tuple, list)) and len(operator) > 1 else None
    operator_name = operator[2] if isinstance(operator, (tuple, list)) and len(operator) > 2 else None
    operator_role = _normalize_user_role(operator[3] if isinstance(operator, (tuple, list)) and len(operator) > 3 else None)
    operator_direction = operator[4] if isinstance(operator, (tuple, list)) and len(operator) > 4 else None
    operator_supervisor_id = operator[6] if isinstance(operator, (tuple, list)) and len(operator) > 6 else None
    operator_login = operator[7] if isinstance(operator, (tuple, list)) and len(operator) > 7 else None

    approver_id = approver[0] if isinstance(approver, (tuple, list)) and len(approver) > 0 else None
    approver_tg = approver[1] if isinstance(approver, (tuple, list)) and len(approver) > 1 else None
    approver_name = approver[2] if isinstance(approver, (tuple, list)) and len(approver) > 2 else None
    approver_role = _normalize_user_role(approver[3] if isinstance(approver, (tuple, list)) and len(approver) > 3 else None)
    approver_login = approver[7] if isinstance(approver, (tuple, list)) and len(approver) > 7 else None

    operator_role_label = SENSITIVE_ACCESS_ROLE_LABELS.get(operator_role, operator_role or '—')
    approver_role_label = SENSITIVE_ACCESS_ROLE_LABELS.get(approver_role, approver_role or '—')
    operator_device = _parse_user_agent_details(session.get('user_agent'))
    approver_device = _parse_user_agent_details(approval_context.get('request_user_agent'))

    lines = [
        "<b>🔐 QR-доступ к данным оценок открыт</b>",
        "",
        f"<b>Время события:</b> {_escape_telegram_html(_format_sensitive_access_notification_dt(datetime.now()), 40)}",
        f"<b>ID сессии:</b> <code>{_escape_telegram_html(session.get('session_id') or claims.get('session_id') or '—', 180)}</code>",
        f"<b>QR токен действителен до:</b> {_escape_telegram_html(_format_sensitive_access_notification_dt(claims.get('expires_at'), assume_utc=True), 40)}",
        "",
        "<b>Оператор</b>",
        f"<b>ФИО:</b> {_escape_telegram_html(operator_name or '—', 120)}",
        f"<b>ID:</b> {_escape_telegram_html(operator_id or '—', 40)}",
        f"<b>Роль:</b> {_escape_telegram_html(operator_role_label, 40)}",
        f"<b>Логин:</b> {_escape_telegram_html(operator_login or '—', 80)}",
        f"<b>Telegram ID:</b> {_escape_telegram_html(operator_tg or '—', 40)}",
        f"<b>Направление:</b> {_escape_telegram_html(operator_direction or '—', 80)}",
        f"<b>Супервайзер:</b> {_escape_telegram_html(operator_supervisor_name or '—', 120)} (ID: {_escape_telegram_html(operator_supervisor_id or '—', 40)})",
        "",
        "<b>Кто подтвердил доступ</b>",
        f"<b>ФИО:</b> {_escape_telegram_html(approver_name or '—', 120)}",
        f"<b>ID:</b> {_escape_telegram_html(approver_id or '—', 40)}",
        f"<b>Роль:</b> {_escape_telegram_html(approver_role_label, 40)}",
        f"<b>Логин:</b> {_escape_telegram_html(approver_login or '—', 80)}",
        f"<b>Telegram ID:</b> {_escape_telegram_html(approver_tg or '—', 40)}",
        "",
        "<b>Сессия оператора</b>",
        f"<b>IP:</b> {_escape_telegram_html(session.get('ip_address') or '—', 64)}",
        f"<b>Устройство:</b> {_escape_telegram_html(operator_device.get('device') or '—', 80)}",
        f"<b>ОС:</b> {_escape_telegram_html(operator_device.get('os') or '—', 80)}",
        f"<b>Браузер:</b> {_escape_telegram_html(operator_device.get('browser') or '—', 100)}",
        f"<b>User-Agent:</b> {_escape_telegram_html(operator_device.get('raw') or '—', 240)}",
        f"<b>Создана:</b> {_escape_telegram_html(_format_sensitive_access_notification_dt(session.get('created_at')), 40)}",
        f"<b>Последняя активность:</b> {_escape_telegram_html(_format_sensitive_access_notification_dt(session.get('last_seen_at')), 40)}",
        f"<b>Истекает:</b> {_escape_telegram_html(_format_sensitive_access_notification_dt(session.get('expires_at')), 40)}",
        f"<b>Разблокирована в:</b> {_escape_telegram_html(_format_sensitive_access_notification_dt(session.get('sensitive_data_unlocked_at')), 40)}",
        "",
        "<b>Контекст подтверждения</b>",
        f"<b>IP подтверждающего:</b> {_escape_telegram_html(approval_context.get('request_ip') or '—', 64)}",
        f"<b>Origin:</b> {_escape_telegram_html(approval_context.get('request_origin') or '—', 120)}",
        f"<b>Устройство подтверждающего:</b> {_escape_telegram_html(approver_device.get('device') or '—', 80)}",
        f"<b>ОС подтверждающего:</b> {_escape_telegram_html(approver_device.get('os') or '—', 80)}",
        f"<b>Браузер подтверждающего:</b> {_escape_telegram_html(approver_device.get('browser') or '—', 100)}",
        f"<b>User-Agent подтверждающего:</b> {_escape_telegram_html(approver_device.get('raw') or '—', 240)}"
    ]
    message = "\n".join(lines)
    if len(message) > TELEGRAM_MAX_MESSAGE_CHARS:
        message = message[:TELEGRAM_MAX_MESSAGE_CHARS]
    return message


def _notify_super_admin_sensitive_access_approved(
    approver,
    operator,
    claims,
    operator_session,
    approval_context,
    operator_supervisor_name
):
    try:
        super_admin_chat_ids = _resolve_super_admin_chat_ids()
        if not super_admin_chat_ids:
            logging.warning("Sensitive access approved, but super admin chat_id is not configured")
            return

        message = _build_sensitive_access_approved_message_html(
            approver=approver,
            operator=operator,
            claims=claims,
            operator_session=operator_session,
            approval_context=approval_context,
            operator_supervisor_name=operator_supervisor_name
        )

        for chat_id in super_admin_chat_ids:
            try:
                response = _send_telegram_text_message(chat_id=chat_id, text=message, parse_mode='HTML')
                if response.status_code != 200:
                    logging.error(
                        "Failed to send sensitive access approval notification to super admin %s: %s",
                        chat_id,
                        _get_telegram_error_text(response)
                    )
            except Exception as send_error:
                logging.error(
                    "Error sending sensitive access approval notification to super admin %s: %s",
                    chat_id,
                    send_error
                )
    except Exception as notify_error:
        logging.error(f"notify_super_admin_sensitive_access_approved error: {notify_error}", exc_info=True)


def _fetch_task_notification_context(task_id):
    with db._get_cursor() as cursor:
        cursor.execute("""
            SELECT
                t.id, t.subject, t.tag, t.created_by,
                creator.telegram_id, creator.name,
                t.assigned_to, assignee.telegram_id, assignee.name
            FROM tasks t
            LEFT JOIN users creator ON creator.id = t.created_by
            LEFT JOIN users assignee ON assignee.id = t.assigned_to
            WHERE t.id = %s
            LIMIT 1
        """, (int(task_id),))
        row = cursor.fetchone()
    if not row:
        return None
    return {
        "id": row[0],
        "subject": row[1],
        "tag": row[2],
        "created_by": row[3],
        "creator_telegram_id": row[4],
        "creator_name": row[5],
        "assigned_to": row[6],
        "assignee_telegram_id": row[7],
        "assignee_name": row[8]
    }


def _collect_task_notification_recipients(task_ctx, actor_user_id):
    recipients = []
    seen_chat_ids = set()

    creator_chat_id = task_ctx.get('creator_telegram_id')
    creator_id = task_ctx.get('created_by')
    if creator_chat_id and (creator_id is None or int(creator_id) != int(actor_user_id)):
        chat_key = str(creator_chat_id)
        if chat_key not in seen_chat_ids:
            seen_chat_ids.add(chat_key)
            recipients.append({
                "kind": "creator",
                "chat_id": creator_chat_id,
                "name": task_ctx.get('creator_name') or 'Постановщик'
            })

    assignee_chat_id = task_ctx.get('assignee_telegram_id')
    assignee_id = task_ctx.get('assigned_to')
    if assignee_chat_id and (assignee_id is None or int(assignee_id) != int(actor_user_id)):
        chat_key = str(assignee_chat_id)
        if chat_key not in seen_chat_ids:
            seen_chat_ids.add(chat_key)
            recipients.append({
                "kind": "assignee",
                "chat_id": assignee_chat_id,
                "name": task_ctx.get('assignee_name') or 'Исполнитель'
            })

    return recipients


def _build_task_status_notification_html(
    action,
    task_ctx,
    actor_name,
    recipient_kind,
    comment=None,
    completion_summary=None,
    completion_files_count=0
):
    action_norm = (action or '').strip().lower()
    tag_label = TASK_TAG_LABELS.get((task_ctx.get('tag') or 'task').strip().lower(), 'Задача')
    subject_safe = _escape_telegram_html(task_ctx.get('subject') or f"Задача #{task_ctx.get('id')}", 220)
    tag_safe = _escape_telegram_html(tag_label, 60)
    actor_safe = _escape_telegram_html(actor_name or 'Сотрудник', 80)

    header_map = {
        'in_progress': '📥 Задача принята в работу',
        'completed': '✅ Задача отмечена выполненной',
        'accepted': '✅ Результат принят',
        'returned': '↩️ Задача возвращена на доработку',
        'reopened': '🔄 Задача возобновлена'
    }
    status_map = {
        'in_progress': 'В работе',
        'completed': 'Ожидает проверки',
        'accepted': 'Принята',
        'returned': 'Возвращена',
        'reopened': 'Снова в работе'
    }

    lines = [
        f"<b>{header_map.get(action_norm, '🔔 Обновление статуса задачи')}</b>",
        "",
        f"<b>Тип:</b> {tag_safe}",
        f"<b>Тема:</b> {subject_safe}",
        f"<b>Статус:</b> {status_map.get(action_norm, 'Обновлён')}",
        f"<b>Кто изменил:</b> {actor_safe}"
    ]

    if comment:
        lines.extend([
            "",
            "<b>Комментарий:</b>",
            _escape_telegram_html(comment, 800)
        ])

    if action_norm == 'completed':
        lines.append(f"<b>Файлов:</b> {int(completion_files_count or 0)}")
        lines.extend([
            "",
            "<b>Итоги выполнения:</b>",
            _escape_telegram_html(completion_summary or "Итоги выполнения не указаны.", 2500)
        ])

    message = "\n".join(lines)
    if len(message) > TELEGRAM_MAX_MESSAGE_CHARS:
        message = message[:TELEGRAM_MAX_MESSAGE_CHARS]
    return message


def _send_task_completion_attachments_to_telegram(chat_id, task_subject, attachments):
    warnings = []
    if not chat_id or not attachments:
        return warnings

    telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendDocument"
    gcs_client = None

    for idx, attachment in enumerate(attachments, start=1):
        try:
            storage_type = str(attachment.get('storage_type') or 'gcs').strip().lower() or 'gcs'
            file_name = (attachment.get('file_name') or f'attachment_{idx}').strip() or f'attachment_{idx}'
            content_type = (attachment.get('content_type') or '').strip() or 'application/octet-stream'

            if storage_type == 'gcs':
                bucket_name = (attachment.get('gcs_bucket') or '').strip()
                blob_path = (attachment.get('gcs_blob_path') or '').strip()
                if not bucket_name or not blob_path:
                    warnings.append(f"Не удалось отправить файл '{file_name}': отсутствуют метаданные хранилища")
                    continue
                if gcs_client is None:
                    gcs_client = get_gcs_client()
                blob = gcs_client.bucket(bucket_name).blob(blob_path)
                if not blob.exists():
                    warnings.append(f"Не удалось отправить файл '{file_name}': файл не найден в хранилище")
                    continue
                file_bytes = blob.download_as_bytes()
            else:
                file_data = attachment.get('file_data')
                if isinstance(file_data, memoryview):
                    file_bytes = file_data.tobytes()
                elif isinstance(file_data, bytearray):
                    file_bytes = bytes(file_data)
                elif isinstance(file_data, bytes):
                    file_bytes = file_data
                else:
                    warnings.append(f"Не удалось отправить файл '{file_name}': неподдерживаемый формат данных")
                    continue

            files = {
                'document': (file_name, file_bytes, content_type)
            }
            data = {
                'chat_id': int(chat_id)
            }
            if idx == 1:
                caption = (
                    "<b>📎 Файлы по задаче</b>\n"
                    f"{_escape_telegram_html(task_subject or 'Без названия', 180)}"
                )
                if len(caption) > TELEGRAM_MAX_CAPTION_CHARS:
                    caption = caption[:TELEGRAM_MAX_CAPTION_CHARS]
                if caption:
                    data['caption'] = caption
                    data['parse_mode'] = 'HTML'

            response = requests.post(
                telegram_url,
                files=files,
                data=data,
                timeout=TASK_TELEGRAM_TIMEOUT_SECONDS
            )
            if response.status_code != 200:
                warnings.append(f"Не удалось отправить файл '{file_name}': {_get_telegram_error_text(response)}")
        except Exception as attachment_error:
            warnings.append(f"Не удалось отправить файл '{attachment.get('file_name') or idx}': {attachment_error}")

    return warnings


def _build_avatar_signed_url(bucket_name, blob_path):
    bucket_name = (bucket_name or '').strip()
    blob_path = (blob_path or '').strip()
    if not bucket_name or not blob_path:
        return None
    cache_key = (bucket_name, blob_path)
    now_ts = time.time()
    with AVATAR_SIGNED_URL_CACHE_LOCK:
        cached = AVATAR_SIGNED_URL_CACHE.get(cache_key)
        if cached and cached.get('expires_at', 0) > (now_ts + AVATAR_SIGNED_URL_CACHE_MIN_REMAINING_SECONDS):
            return cached.get('url')
        if cached:
            AVATAR_SIGNED_URL_CACHE.pop(cache_key, None)
    try:
        gcs_client = get_gcs_client()
        bucket = gcs_client.bucket(bucket_name)
        blob = bucket.blob(blob_path)
        signed_url = blob.generate_signed_url(
            version="v4",
            expiration=timedelta(seconds=AVATAR_SIGNED_URL_TTL_SECONDS),
            method="GET"
        )
        expires_at = now_ts + AVATAR_SIGNED_URL_TTL_SECONDS
        with AVATAR_SIGNED_URL_CACHE_LOCK:
            AVATAR_SIGNED_URL_CACHE[cache_key] = {
                "url": signed_url,
                "expires_at": expires_at
            }
            if len(AVATAR_SIGNED_URL_CACHE) > AVATAR_SIGNED_URL_CACHE_MAX_ITEMS:
                stale_keys = [k for k, v in AVATAR_SIGNED_URL_CACHE.items() if v.get('expires_at', 0) <= now_ts]
                for stale_key in stale_keys:
                    AVATAR_SIGNED_URL_CACHE.pop(stale_key, None)
                overflow = len(AVATAR_SIGNED_URL_CACHE) - AVATAR_SIGNED_URL_CACHE_MAX_ITEMS
                if overflow > 0:
                    oldest_keys = sorted(
                        AVATAR_SIGNED_URL_CACHE.items(),
                        key=lambda item: item[1].get('expires_at', 0)
                    )[:overflow]
                    for oldest_key, _ in oldest_keys:
                        AVATAR_SIGNED_URL_CACHE.pop(oldest_key, None)
        return signed_url
    except Exception as e:
        logging.warning(f"Failed to build avatar signed URL for bucket={bucket_name}: {e}")
        return None


def _sanitize_avatar_extension(source_filename, fallback='.webp'):
    safe_name = secure_filename(source_filename or "")
    extension = os.path.splitext(safe_name)[1].lower()
    if not extension or not re.match(r'^\.[a-z0-9]{1,10}$', extension):
        return fallback
    return extension


def _build_avatar_blob_paths(user_id, thumbnail_filename, original_filename=None):
    upload_folder = AVATAR_UPLOAD_FOLDER.strip('/')
    upload_prefix = f"{upload_folder}/" if upload_folder else ""
    thumbnail_ext = _sanitize_avatar_extension(thumbnail_filename, '.webp')
    original_ext = _sanitize_avatar_extension(original_filename or thumbnail_filename, thumbnail_ext)
    date_prefix = datetime.utcnow().strftime('%Y/%m/%d')
    random_id = uuid.uuid4().hex
    base_prefix = f"{upload_prefix}avatars/{int(user_id)}/{date_prefix}/{random_id}"
    return {
        "thumbnail_path": f"{base_prefix}_{AVATAR_THUMBNAIL_SUFFIX}{thumbnail_ext}",
        "original_path": f"{base_prefix}_original{original_ext}"
    }


def _delete_avatar_blobs(bucket_name, *blob_paths):
    bucket_name = (bucket_name or '').strip()
    unique_paths = []
    for item in blob_paths:
        path = (item or '').strip()
        if path and path not in unique_paths:
            unique_paths.append(path)
    if not bucket_name or not unique_paths:
        return
    try:
        client = get_gcs_client()
        bucket = client.bucket(bucket_name)
    except Exception:
        return
    for blob_path in unique_paths:
        try:
            bucket.blob(blob_path).delete()
        except Exception:
            pass


def _resolve_avatar_target_user(requester, requester_id, target_user_id):
    target_user = db.get_user(id=target_user_id)
    if not target_user:
        return None, ("User not found", 404)
    requester_role = _normalize_user_role(requester[3])
    if _is_admin_role(requester_role):
        return target_user, None
    if requester_id == target_user[0]:
        return target_user, None
    if _is_supervisor_role(requester_role) and _normalize_user_role(target_user[3]) == 'operator' and target_user[6] == requester_id:
        return target_user, None
    return None, ("You do not have access to manage this avatar", 403)


def _resolve_requester():
    requester_id_raw = request.headers.get('X-User-Id') or getattr(g, 'user_id', None)
    if not requester_id_raw:
        return None, None, ("X-User-Id header required", 400)
    try:
        requester_id = int(requester_id_raw)
    except Exception:
        return None, None, ("Invalid X-User-Id", 400)

    requester = db.get_user(id=requester_id)
    if not requester:
        return None, None, ("Requester not found", 404)
    return requester_id, requester, None


def _task_route_guard():
    requester_id, requester, error = _resolve_requester()
    if error:
        message, status_code = error
        return None, None, jsonify({"error": message}), status_code
    if not (_is_admin_role(requester[3]) or _is_supervisor_role(requester[3])):
        return None, None, jsonify({"error": "Only admin and sv can access tasks"}), 403
    return requester_id, requester, None, None


def _normalize_surveys_role(role):
    return _normalize_user_role(role)


def _surveys_route_guard():
    requester_id, requester, error = _resolve_requester()
    if error:
        message, status_code = error
        return None, None, None, jsonify({"error": message}), status_code

    role = _normalize_surveys_role(requester[3])
    if not _has_any_role(role, ('operator', 'trainer', 'sv', 'admin', 'super_admin')):
        return None, None, None, jsonify({"error": "Only admin, sv, trainer and operator can access surveys"}), 403

    return requester_id, requester, role, None, None


def _cleanup_task_uploaded_blobs(gcs_bucket, uploaded_blob_paths):
    if not gcs_bucket or not uploaded_blob_paths:
        return
    for uploaded_path in uploaded_blob_paths:
        try:
            gcs_bucket.blob(uploaded_path).delete()
        except Exception:
            pass


def _upload_task_attachments_to_gcs(files, stage='initial'):
    file_items = [item for item in (files or []) if item and item.filename]
    if len(file_items) > TASK_MAX_FILES:
        raise ValueError(f"Too many files. Max allowed: {TASK_MAX_FILES}")
    if not file_items:
        return [], [], None

    bucket_name = (os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET_TASKS') or '').strip()
    if not bucket_name:
        raise RuntimeError("GOOGLE_CLOUD_STORAGE_BUCKET_TASKS is not configured")

    upload_folder = TASK_ATTACHMENTS_UPLOAD_FOLDER.strip('/')
    upload_prefix = f"{upload_folder}/" if upload_folder else ""
    gcs_client = get_gcs_client()
    gcs_bucket = gcs_client.bucket(bucket_name)

    attachments = []
    uploaded_blob_paths = []

    for idx, file_storage in enumerate(file_items):
        file_data = file_storage.read() or b''
        if len(file_data) > TASK_MAX_FILE_SIZE_BYTES:
            raise ValueError(f"File '{file_storage.filename}' exceeds 10 MB limit")

        safe_name = secure_filename(file_storage.filename) or f'attachment_{idx + 1}'
        content_type = (file_storage.mimetype or '').strip() or 'application/octet-stream'
        blob_path = (
            f"{upload_prefix}tasks/{stage}/"
            f"{datetime.utcnow().strftime('%Y/%m/%d')}/"
            f"{uuid.uuid4().hex}_{safe_name}"
        )

        blob = gcs_bucket.blob(blob_path)
        blob.upload_from_string(file_data, content_type=content_type)
        uploaded_blob_paths.append(blob_path)

        attachments.append({
            "file_name": safe_name,
            "content_type": content_type,
            "file_size": len(file_data),
            "storage_type": "gcs",
            "gcs_bucket": bucket_name,
            "gcs_blob_path": blob_path
        })

    return attachments, uploaded_blob_paths, gcs_bucket


@app.before_request
def hydrate_user_context_from_jwt():
    if not request.path.startswith('/api/'):
        return None
    if request.method == 'OPTIONS':
        return None

    if request.path == '/api/auth/refresh':
        return None

    try:
        payload = _authenticate_access_cookie(optional=True)
        if payload:
            return None
    except AuthError as auth_error:
        request.environ['JWT_AUTH_ERROR_CODE'] = auth_error.code

    try:
        _authenticate_refresh_cookie(optional=True, rotate_tokens=False)
    except AuthError as refresh_auth_error:
        request.environ['JWT_AUTH_ERROR_CODE'] = refresh_auth_error.code
    return None

def require_api_key(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        if getattr(g, 'user_id', None):
            return f(*args, **kwargs)

        try:
            if _authenticate_access_cookie(optional=True):
                return f(*args, **kwargs)
        except AuthError as auth_error:
            request.environ['JWT_AUTH_ERROR_CODE'] = auth_error.code

        try:
            if _authenticate_refresh_cookie(optional=True, rotate_tokens=False):
                return f(*args, **kwargs)
        except AuthError as refresh_auth_error:
            request.environ['JWT_AUTH_ERROR_CODE'] = refresh_auth_error.code

        auth_error_code = request.environ.get('JWT_AUTH_ERROR_CODE')
        if auth_error_code:
            return jsonify({"error": "JWT authentication failed", "code": auth_error_code}), 401

        if _get_refresh_token_values():
            return jsonify({"error": "Access token required", "code": "TOKEN_EXPIRED"}), 401

        logging.warning("Unauthorized request: no valid JWT token")
        return jsonify({"error": "Unauthorized"}), 401
    return decorated

def _build_cors_preflight_response():
    origin = _normalize_origin(request.headers.get("Origin"))
    response = jsonify({"status": "ok"})

    if _is_allowed_origin(origin):
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Vary"] = "Origin"

    response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-API-Key, X-User-Id, Authorization, X-Refresh-Token"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, PATCH, DELETE, OPTIONS"
    response.headers["Access-Control-Max-Age"] = "86400"
    return response

@app.route('/')
def index():
    return "Bot is alive!", 200

@app.after_request
def after_request(response):
    origin = _normalize_origin(request.headers.get('Origin'))
    if _is_allowed_origin(origin):
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Vary'] = 'Origin'

    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, X-API-Key, X-User-Id, Authorization, X-Refresh-Token'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS, DELETE, PUT, PATCH'

    pending_tokens = getattr(g, 'pending_auth_tokens', None)
    if pending_tokens:
        access_token, refresh_token = pending_tokens
        _set_auth_cookies(response, access_token, refresh_token)

    return response

@app.route('/api/health', methods=['GET'])
def health_check():
    try:
        # Test database connectivity
        with db._get_cursor() as cursor:
            cursor.execute("SELECT 1")
        return jsonify({"status": "healthy", "database": "connected"}), 200
    except Exception as e:
        logging.error(f"Health check failed: {e}")
        return jsonify({"status": "unhealthy", "error": str(e)}), 500

@app.route('/api/login', methods=['POST', 'OPTIONS'])
def login():
    try:
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        data = request.get_json() or {}
        login_value = data.get('login', '').strip()
        password = data.get('password', '')
        if not login_value or not password:
            return jsonify({"error": "Missing credentials"}), 400

        user = db.get_user_by_login(login_value)
        if user and db.verify_password(user[0], password):
            user_profile = db.get_user(id=user[0])
            if not user_profile:
                return jsonify({"error": "Invalid credentials"}), 401

            user_status = str(user_profile[11] or '').strip().lower()
            if user_status in ('fired', 'dismissal'):
                db.revoke_all_user_sessions(user_id=user[0])
                return jsonify({"error": "User account is inactive"}), 403

            session_id = str(uuid.uuid4())
            access_token = _build_access_token(user_profile, session_id)
            refresh_token = _build_refresh_token(user[0], session_id)
            db.create_user_session(
                session_id=session_id,
                user_id=user[0],
                refresh_token_hash=_hash_token(refresh_token),
                expires_at=datetime.utcnow() + timedelta(days=JWT_REFRESH_TOKEN_DAYS),
                user_agent=request.headers.get('User-Agent'),
                ip_address=_client_ip()
            )
            _set_request_auth_context(user[0])
            payload = _get_user_payload(user_profile)
            response = jsonify({
                "status": "success",
                "user": payload,
                **payload,
                "access_token": access_token,
                "refresh_token": refresh_token
            })
            _set_auth_cookies(response, access_token, refresh_token)
            return response, 200

        return jsonify({"error": "Invalid credentials"}), 401
    except Exception as e:
        logging.error(f"Login error: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/auth/me', methods=['GET'])
@require_api_key
def auth_me():
    try:
        user_id = getattr(g, 'user_id', None)
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401
        user = db.get_user(id=user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        payload = _get_user_payload(user)
        return jsonify({"status": "success", "user": payload, **payload}), 200
    except Exception as e:
        logging.error(f"auth_me error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/auth/refresh', methods=['POST', 'OPTIONS'])
def refresh_auth():
    try:
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        _authenticate_refresh_cookie(optional=False, rotate_tokens=True)
        user_id = getattr(g, 'user_id', None)
        if not user_id:
            raise AuthError("UNAUTHORIZED", "Unauthorized")
        user = db.get_user(id=user_id)
        if not user:
            raise AuthError("USER_NOT_FOUND", "User not found")

        payload_user = _get_user_payload(user)
        response_payload = {"status": "success", "user": payload_user, **payload_user}
        pending_tokens = getattr(g, 'pending_auth_tokens', None)
        if pending_tokens:
            response_payload["access_token"] = pending_tokens[0]
            response_payload["refresh_token"] = pending_tokens[1]
        response = jsonify(response_payload)
        return response, 200
    except AuthError as auth_error:
        response = jsonify({"error": auth_error.message, "code": auth_error.code})
        _clear_auth_cookies(response)
        return response, 401
    except Exception as e:
        logging.error(f"refresh_auth error: {e}", exc_info=True)
        response = jsonify({"error": "Internal server error"})
        _clear_auth_cookies(response)
        return response, 500


@app.route('/api/logout', methods=['POST', 'OPTIONS'])
def logout():
    try:
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        session_id = _current_session_id_from_access_token()
        user_id = getattr(g, 'user_id', None)
        if session_id and user_id:
            db.revoke_user_session(session_id=session_id, user_id=user_id)
        elif session_id:
            db.revoke_user_session(session_id=session_id)

        response = jsonify({"status": "success"})
        _clear_auth_cookies(response)
        return response, 200
    except Exception as e:
        logging.error(f"logout error: {e}", exc_info=True)
        response = jsonify({"status": "success"})
        _clear_auth_cookies(response)
        return response, 200


@app.route('/api/auth/logout_all', methods=['POST', 'OPTIONS'])
@require_api_key
def logout_all():
    try:
        user_id = getattr(g, 'user_id', None)
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401
        db.revoke_all_user_sessions(user_id=user_id)
        response = jsonify({"status": "success"})
        _clear_auth_cookies(response)
        return response, 200
    except Exception as e:
        logging.error(f"logout_all error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/auth/sessions', methods=['GET', 'OPTIONS'])
@require_api_key
def list_auth_sessions():
    try:
        user_id = getattr(g, 'user_id', None)
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401
        current_session_id = _current_session_id_from_access_token()
        sessions = db.list_user_sessions(user_id=user_id)
        serialized = []
        for item in sessions:
            serialized.append({
                "session_id": item["session_id"],
                "is_current": item["session_id"] == current_session_id,
                "user_agent": item["user_agent"],
                "ip_address": item["ip_address"],
                "created_at": item["created_at"].isoformat() if item["created_at"] else None,
                "last_seen_at": item["last_seen_at"].isoformat() if item["last_seen_at"] else None,
                "expires_at": item["expires_at"].isoformat() if item["expires_at"] else None,
                "revoked_at": item["revoked_at"].isoformat() if item["revoked_at"] else None
            })
        return jsonify({"status": "success", "sessions": serialized}), 200
    except Exception as e:
        logging.error(f"list_auth_sessions error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/auth/sessions/<session_id>/revoke', methods=['POST', 'OPTIONS'])
@require_api_key
def revoke_auth_session(session_id):
    try:
        user_id = getattr(g, 'user_id', None)
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401
        revoked = db.revoke_user_session(session_id=session_id, user_id=user_id)
        if not revoked:
            return jsonify({"error": "Session not found"}), 404

        response = jsonify({"status": "success"})
        if _current_session_id_from_access_token() == session_id:
            _clear_auth_cookies(response)
        return response, 200
    except Exception as e:
        logging.error(f"revoke_auth_session error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/admin/sessions', methods=['GET', 'OPTIONS'])
@require_api_key
def list_admin_sessions():
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester or not _is_admin_role(requester[3]):
            return jsonify({"error": "Forbidden: only admins can access"}), 403

        raw_limit = request.args.get('limit')
        raw_offset = request.args.get('offset')
        query_text = (request.args.get('q') or '').strip()

        try:
            limit = int(raw_limit) if raw_limit is not None else 100
            offset = int(raw_offset) if raw_offset is not None else 0
        except Exception:
            return jsonify({"error": "Invalid pagination params"}), 400

        if limit < 1 or limit > 500:
            return jsonify({"error": "limit must be in range 1..500"}), 400
        if offset < 0:
            return jsonify({"error": "offset must be >= 0"}), 400

        current_session_id = _current_session_id_from_access_token()
        sessions = db.list_all_active_sessions(limit=limit, offset=offset, search=query_text)
        summary = db.get_all_active_sessions_summary(search=query_text)
        serialized = []
        for item in sessions:
            serialized.append({
                "session_id": item["session_id"],
                "user_id": item["user_id"],
                "user_name": item["user_name"],
                "user_role": item["user_role"],
                "user_login": item["user_login"],
                "supervisor_id": item["supervisor_id"],
                "supervisor_name": item["supervisor_name"],
                "avatar_url": _build_avatar_signed_url(item.get("avatar_bucket"), item.get("avatar_blob_path")),
                "is_current": item["session_id"] == current_session_id,
                "user_agent": item["user_agent"],
                "ip_address": item["ip_address"],
                "created_at": item["created_at"].isoformat() if item["created_at"] else None,
                "last_seen_at": item["last_seen_at"].isoformat() if item["last_seen_at"] else None,
                "expires_at": item["expires_at"].isoformat() if item["expires_at"] else None,
                "sensitive_data_unlocked": bool(item.get("sensitive_data_unlocked", False)),
                "sensitive_data_unlocked_at": (
                    item["sensitive_data_unlocked_at"].isoformat()
                    if item.get("sensitive_data_unlocked_at")
                    else None
                )
            })

        total_sessions = int(summary.get("total_sessions", 0))
        returned = len(serialized)
        has_more = (offset + returned) < total_sessions

        return jsonify({
            "status": "success",
            "sessions": serialized,
            "summary": summary,
            "pagination": {
                "limit": limit,
                "offset": offset,
                "returned": returned,
                "has_more": has_more
            },
            "query": query_text
        }), 200
    except Exception as e:
        logging.error(f"list_admin_sessions error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/admin/sessions/<session_id>/revoke', methods=['POST', 'OPTIONS'])
@require_api_key
def revoke_admin_session(session_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester or not _is_admin_role(requester[3]):
            return jsonify({"error": "Forbidden: only admins can access"}), 403

        session = db.get_user_session(session_id=session_id)
        if not session or session["revoked_at"] is not None:
            return jsonify({"error": "Session not found"}), 404

        revoked = db.revoke_user_session(session_id=session_id)
        if not revoked:
            return jsonify({"error": "Session not found"}), 404

        current_session_revoked = _current_session_id_from_access_token() == session_id
        response = jsonify({
            "status": "success",
            "current_session_revoked": current_session_revoked
        })
        if current_session_revoked:
            _clear_auth_cookies(response)
        return response, 200
    except Exception as e:
        logging.error(f"revoke_admin_session error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/sensitive-access/qr/request', methods=['POST', 'OPTIONS'])
@require_api_key
def request_sensitive_access_qr():
    try:
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester or requester[3] != 'operator':
            return jsonify({"error": "Forbidden: only operator can generate QR access token"}), 403

        session_id = _current_session_id_from_access_token()
        if not session_id:
            return jsonify({"error": "Active session not found"}), 401

        session = db.get_user_session(session_id=session_id, user_id=requester_id)
        if not session or session["revoked_at"] is not None:
            return jsonify({"error": "Session not found or revoked"}), 401

        token, expires_at = _build_sensitive_qr_token(session_id=session_id, user_id=requester_id)
        qr_payload = f"OTP-SENSITIVE:{token}"

        return jsonify({
            "status": "success",
            "qr_payload": qr_payload,
            "token": token,
            "token_expires_at": expires_at.isoformat() + "Z",
            "granted": _is_sensitive_access_unlocked(requester_id, session_id),
            "required": True
        }), 200
    except Exception as e:
        logging.error(f"request_sensitive_access_qr error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/sensitive-access/status', methods=['GET', 'OPTIONS'])
@require_api_key
def get_sensitive_access_status():
    try:
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        role = _normalize_user_role(requester[3])
        session_id = _current_session_id_from_access_token()
        required = role == 'operator'
        if required:
            granted = bool(_is_sensitive_access_unlocked(requester_id, session_id))
        else:
            granted = True

        return jsonify({
            "status": "success",
            "required": required,
            "granted": granted,
            "session_id": session_id
        }), 200
    except Exception as e:
        logging.error(f"get_sensitive_access_status error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/sensitive-access/revoke', methods=['POST', 'OPTIONS'])
@require_api_key
def revoke_sensitive_access():
    try:
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        session_id = _current_session_id_from_access_token()
        if not session_id:
            return jsonify({"error": "Active session not found"}), 401

        db.set_session_sensitive_access(session_id=session_id, user_id=requester_id, unlocked=False)
        return jsonify({"status": "success", "granted": False}), 200
    except Exception as e:
        logging.error(f"revoke_sensitive_access error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/sensitive-access/approve', methods=['POST', 'OPTIONS'])
@require_api_key
def approve_sensitive_access():
    try:
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        approver_id = getattr(g, 'user_id', None)
        if not approver_id:
            return jsonify({"error": "Unauthorized"}), 401

        approver = db.get_user(id=approver_id)
        if not approver or not _is_privileged_role(approver[3]):
            return jsonify({"error": "Forbidden: only admin or supervisor can approve QR"}), 403

        data = request.get_json(silent=True) or {}
        raw_token = (data.get('token') or '').strip()
        if not raw_token:
            return jsonify({"error": "Missing token"}), 400

        token = raw_token
        if raw_token.upper().startswith('OTP-SENSITIVE:'):
            token = raw_token.split(':', 1)[1].strip()
        if 'token=' in token and ('http://' in token or 'https://' in token):
            try:
                parsed = urlparse(token)
                query_token = parse_qs(parsed.query).get('token', [''])[0]
                if query_token:
                    token = query_token.strip()
            except Exception:
                pass

        claims = _decode_sensitive_qr_token(token)
        operator_id = claims["user_id"]
        operator = db.get_user(id=operator_id)
        if not operator or operator[3] != 'operator':
            return jsonify({"error": "QR token does not belong to operator session"}), 400

        session = db.get_user_session(session_id=claims["session_id"], user_id=operator_id)
        if not session or session["revoked_at"] is not None:
            return jsonify({"error": "Operator session not found or revoked"}), 410

        if approver[3] == 'sv' and operator[6] != approver_id:
            return jsonify({"error": "Supervisor can approve only own operators"}), 403

        updated = db.set_session_sensitive_access(
            session_id=claims["session_id"],
            user_id=operator_id,
            unlocked=True
        )
        if not updated:
            return jsonify({"error": "Failed to activate access for operator session"}), 409

        operator_session = db.get_user_session(session_id=claims["session_id"], user_id=operator_id)
        operator_supervisor_name = None
        operator_supervisor_id = operator[6] if len(operator) > 6 else None
        if operator_supervisor_id:
            supervisor = db.get_user(id=operator_supervisor_id)
            if supervisor:
                operator_supervisor_name = supervisor[2]

        approval_context = {
            "request_ip": _client_ip(),
            "request_user_agent": request.headers.get('User-Agent'),
            "request_origin": request.headers.get('Origin')
        }

        threading.Thread(
            target=_notify_super_admin_sensitive_access_approved,
            args=(approver, operator, claims, operator_session, approval_context, operator_supervisor_name),
            daemon=True
        ).start()

        return jsonify({
            "status": "success",
            "operator_id": operator_id,
            "operator_name": operator[2],
            "session_id": claims["session_id"],
            "approved_by": approver[2],
            "approved_by_role": approver[3]
        }), 200
    except ValueError as token_error:
        return jsonify({"error": str(token_error)}), 400
    except Exception as e:
        logging.error(f"approve_sensitive_access error: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/user/profile', methods=['GET'])
@require_api_key
def get_user_profile():
    try:
        user_id = request.args.get('user_id')
        if not user_id:
            logging.warning("Missing user_id parameter in profile request")
            return jsonify({"error": "Missing user_id parameter"}), 400

        try:
            user_id = int(user_id)
        except ValueError:
            logging.error(f"Invalid user_id format: {user_id}")
            return jsonify({"error": "Invalid user_id format"}), 400

        user = db.get_user(id=user_id)
        if not user:
            logging.warning(f"User not found with ID: {user_id}")
            return jsonify({"error": "User not found"}), 404

        supervisor_name = None
        if user[6]:
            supervisor = db.get_user(id=user[6])
            supervisor_name = supervisor[2] if supervisor else None

        hire_date = user[5].strftime('%Y-%m-%d') if user[5] else None

        profile_data = {
            "id": user[0],
            "name": user[2] or "Unknown",
            "role": user[3] or "Unknown",
            "direction": user[4],  # Now returns direction name from joined directions table
            "hire_date": hire_date,
            "supervisor_name": supervisor_name,
            "telegram_id": user[1] or None,
            "scores_table_url": user[9] or None,
            "hours_table_url": user[8] or None,
            "status": user[11],  # Adjust index based on query
            "rate": float(user[12]) if user[12] else 1.0,
            "avatar_url": _build_avatar_signed_url(user[15], user[16]),
            "avatar_updated_at": user[19].isoformat() if len(user) > 19 and user[19] else None
        }

        logging.info(f"Profile data fetched successfully for user_id: {user_id}")
        return jsonify({"status": "success", "profile": profile_data}), 200
    except Exception as e:
        logging.error(f"Error fetching user profile for user_id {user_id}: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/user/avatar', methods=['POST', 'DELETE', 'OPTIONS'])
@require_api_key
def manage_user_avatar():
    try:
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        requester_id_raw = request.headers.get('X-User-Id') or getattr(g, 'user_id', None)
        if not requester_id_raw:
            return jsonify({"error": "Missing X-User-Id header"}), 400
        try:
            requester_id = int(requester_id_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid X-User-Id"}), 400

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404

        json_payload = request.get_json(silent=True) or {}
        target_user_id_raw = request.form.get('target_user_id') if request.method == 'POST' else json_payload.get('target_user_id')
        if target_user_id_raw in (None, ''):
            target_user_id = requester_id
        else:
            try:
                target_user_id = int(target_user_id_raw)
            except (TypeError, ValueError):
                return jsonify({"error": "Invalid target_user_id"}), 400

        target_user, access_error = _resolve_avatar_target_user(requester, requester_id, target_user_id)
        if access_error:
            message, status_code = access_error
            return jsonify({"error": message}), status_code

        remove_flag_raw = request.form.get('remove') if request.method == 'POST' else json_payload.get('remove')
        remove_flag = request.method == 'DELETE' or str(remove_flag_raw or '').strip().lower() in ('1', 'true', 'yes')
        previous_avatar = db.get_user_avatar_storage(target_user_id) or {}

        if remove_flag:
            if not db.clear_user_avatar(target_user_id):
                return jsonify({"error": "Failed to clear avatar"}), 500

            prev_bucket = (previous_avatar.get('avatar_bucket') or '').strip()
            prev_blob_path = (previous_avatar.get('avatar_blob_path') or '').strip()
            prev_original_blob_path = (previous_avatar.get('avatar_original_blob_path') or '').strip()
            _delete_avatar_blobs(prev_bucket, prev_blob_path, prev_original_blob_path)

            updated_avatar = db.get_user_avatar_storage(target_user_id) or {}
            return jsonify({
                "status": "success",
                "message": "Avatar removed",
                "target_user_id": target_user[0],
                "avatar_url": None,
                "avatar_updated_at": updated_avatar.get("avatar_updated_at")
            }), 200

        file_storage = request.files.get('avatar')
        if not file_storage or not file_storage.filename:
            return jsonify({"error": "Missing avatar file"}), 400

        avatar_bytes = file_storage.read() or b''
        if not avatar_bytes:
            return jsonify({"error": "Uploaded avatar is empty"}), 400
        if len(avatar_bytes) > AVATAR_MAX_UPLOAD_BYTES:
            return jsonify({"error": f"Avatar exceeds {AVATAR_MAX_UPLOAD_BYTES // 1024} KB limit"}), 400

        content_type = (file_storage.mimetype or '').strip().lower()
        if not content_type.startswith('image/'):
            return jsonify({"error": "Only image files are allowed"}), 400

        original_file_storage = request.files.get('avatar_original')
        original_bytes = avatar_bytes
        original_content_type = content_type
        original_filename = file_storage.filename
        if original_file_storage and original_file_storage.filename:
            candidate_original_bytes = original_file_storage.read() or b''
            if not candidate_original_bytes:
                return jsonify({"error": "Uploaded original avatar is empty"}), 400
            if len(candidate_original_bytes) > AVATAR_MAX_ORIGINAL_UPLOAD_BYTES:
                return jsonify({"error": f"Original avatar exceeds {AVATAR_MAX_ORIGINAL_UPLOAD_BYTES // 1024} KB limit"}), 400
            candidate_original_content_type = (original_file_storage.mimetype or '').strip().lower()
            if not candidate_original_content_type.startswith('image/'):
                return jsonify({"error": "Only image files are allowed for original avatar"}), 400
            original_bytes = candidate_original_bytes
            original_content_type = candidate_original_content_type
            original_filename = original_file_storage.filename

        bucket_name = (
            os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET_AVATARS')
            or os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET_TASKS')
            or os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET')
            or ''
        ).strip()
        if not bucket_name:
            return jsonify({"error": "Avatar bucket is not configured"}), 500

        avatar_blob_paths = _build_avatar_blob_paths(target_user_id, file_storage.filename, original_filename)
        blob_path = avatar_blob_paths['thumbnail_path']
        original_blob_path = avatar_blob_paths['original_path']
        gcs_client = get_gcs_client()
        gcs_bucket = gcs_client.bucket(bucket_name)
        blob = gcs_bucket.blob(blob_path)
        blob.cache_control = "public, max-age=31536000, immutable"
        blob.upload_from_string(avatar_bytes, content_type=content_type)
        original_blob = gcs_bucket.blob(original_blob_path)
        original_blob.cache_control = "private, max-age=0, no-cache"
        original_blob.upload_from_string(original_bytes, content_type=original_content_type)

        saved_avatar = db.set_user_avatar(
            user_id=target_user_id,
            bucket_name=bucket_name,
            blob_path=blob_path,
            original_blob_path=original_blob_path,
            content_type=content_type,
            file_size=len(avatar_bytes)
        )
        if not saved_avatar:
            _delete_avatar_blobs(bucket_name, blob_path, original_blob_path)
            return jsonify({"error": "Failed to save avatar metadata"}), 500

        prev_bucket = (previous_avatar.get('avatar_bucket') or '').strip()
        prev_blob_path = (previous_avatar.get('avatar_blob_path') or '').strip()
        prev_original_blob_path = (previous_avatar.get('avatar_original_blob_path') or '').strip()
        is_same_thumbnail = prev_bucket == bucket_name and prev_blob_path == blob_path
        is_same_original = prev_bucket == bucket_name and prev_original_blob_path == original_blob_path
        if prev_bucket and (not is_same_thumbnail or not is_same_original):
            _delete_avatar_blobs(prev_bucket, prev_blob_path, prev_original_blob_path)

        return jsonify({
            "status": "success",
            "message": "Avatar updated",
            "target_user_id": target_user[0],
            "avatar_url": _build_avatar_signed_url(saved_avatar.get("avatar_bucket"), saved_avatar.get("avatar_blob_path")),
            "avatar_updated_at": saved_avatar.get("avatar_updated_at"),
            "avatar_file_size": saved_avatar.get("avatar_file_size"),
            "avatar_original_file_size": len(original_bytes)
        }), 200
    except Exception as e:
        logging.error(f"Error in /api/user/avatar: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/average_scores', methods=['GET'])
@require_api_key
def api_average_scores():
    """Endpoint: /api/average_scores?start=YYYY-MM-DD&end=YYYY-MM-DD&operator_ids=1,2,3

    Returns per-operator average score and overall average for the period.
    """
    try:
        start = request.args.get('start')
        end = request.args.get('end')
        op_ids_raw = request.args.get('operator_ids')

        if not start or not end:
            return jsonify({"error": "Missing required parameters 'start' and 'end'"}), 400

        # parse optional operator ids
        operator_ids = None
        if op_ids_raw:
            try:
                operator_ids = [int(x) for x in re.split(r'\s*,\s*', op_ids_raw.strip()) if x]
            except Exception:
                return jsonify({"error": "Invalid operator_ids format; expected comma-separated integers"}), 400

        # normalize dates: accept YYYY-MM-DD or full ISO timestamps
        def _parse_date_param(s, end_of_day=False):
            try:
                if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
                    dt = datetime.strptime(s, '%Y-%m-%d')
                    if end_of_day:
                        return datetime.combine(dt.date(), datetime.max.time())
                    return datetime.combine(dt.date(), datetime.min.time())
                # try ISO parse
                return datetime.fromisoformat(s)
            except Exception:
                return None

        sd = _parse_date_param(start, end_of_day=False)
        ed = _parse_date_param(end, end_of_day=True)
        if not sd or not ed:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD or ISO format."}), 400

        result = db.get_average_scores_for_period(sd, ed, operator_ids)
        return jsonify({"status": "success", "data": result}), 200
    except Exception as e:
        logging.exception("Error in /api/average_scores: %s", e)
        return jsonify({"error": str(e)}), 500


@app.route('/api/baiga/upload_day_csv', methods=['POST', 'OPTIONS'])
@require_api_key
def api_baiga_upload_day_csv():
    # Upload CSV file for a specific day, process server-side and save counts.
    try:
        if request.method == 'OPTIONS':
            return _build_cors_preflight_response()

        # expected: form-data with file field 'file' and 'day' in YYYY-MM-DD
        day = request.form.get('day') or request.args.get('day')
        if not day:
            return jsonify({"error": "Missing 'day' parameter (YYYY-MM-DD)"}), 400

        if 'file' not in request.files:
            return jsonify({"error": "Missing file field"}), 400

        f = request.files['file']
        raw = f.read()
        try:
            text = raw.decode('utf-8')
        except Exception:
            try:
                text = raw.decode('cp1251')
            except Exception:
                text = raw.decode('utf-8', errors='ignore')

        # parse CSV robustly using csv module with ';' as delimiter
        sio = BytesIO(text.encode('utf-8'))
        # Use text IO wrapper
        import io, csv
        si = io.StringIO(text)
        reader = csv.reader(si, delimiter=';', quotechar='"')

        rows = list(reader)
        if not rows:
            return jsonify({"error": "Empty CSV"}), 400

        # Skip header (first row)
        score_counts = {}
        for i in range(1, len(rows)):
            row = rows[i]
            if len(row) <= 7:
                continue
            fio_raw = row[5]
            score_raw = row[7]
            fio = (fio_raw or '').strip().strip('"')
            score_val = (score_raw or '').strip().strip('"')
            if not fio:
                continue
            try:
                num = float(score_val.replace(',', '.'))
                is_five = (int(round(num)) == 5)
            except Exception:
                is_five = (score_val == '5')

            if is_five:
                score_counts[fio] = score_counts.get(fio, 0) + 1

        # Resolve to operators
        operator_scores = {}
        unmatched = []
        total_found = sum(score_counts.values())
        for fio, cnt in score_counts.items():
            user = db.find_operator_by_name_fuzzy(fio)
            if user and user[0]:
                op_id = int(user[0])
                operator_scores[op_id] = operator_scores.get(op_id, 0) + int(cnt)
            else:
                unmatched.append({"fio": fio, "count": int(cnt)})

        # Replace day's baiga scores
        db.replace_baiga_scores_for_day(day, operator_scores)

        return jsonify({"status": "success", "totalFound": total_found, "matched": len(operator_scores), "unmatched": unmatched}), 200
    except Exception as e:
        logging.exception("Error in /api/baiga/upload_day_csv: %s", e)
        return jsonify({"error": str(e)}), 500


@app.route('/api/baiga/day_scores', methods=['GET'])
@require_api_key
def api_baiga_day_scores():
    try:
        day = request.args.get('day')
        if not day:
            return jsonify({"error": "Missing 'day' parameter"}), 400
        rows = db.get_baiga_scores_for_day(day)
        return jsonify({"status": "success", "scores": rows}), 200
    except Exception as e:
        logging.exception("Error in /api/baiga/day_scores: %s", e)
        return jsonify({"error": str(e)}), 500


@app.route('/api/baiga/period_scores', methods=['GET'])
@require_api_key
def api_baiga_period_scores():
    try:
        # expects start and end in YYYY-MM-DD
        start = request.args.get('start')
        end = request.args.get('end')
        if not start or not end:
            return jsonify({"error": "Missing 'start' and 'end' parameters (YYYY-MM-DD)"}), 400

        # validate dates loosely
        try:
            from datetime import datetime as _dt
            _dt.strptime(start, '%Y-%m-%d')
            _dt.strptime(end, '%Y-%m-%d')
        except Exception:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD."}), 400

        data = db.get_baiga_scores_for_range(start, end)
        # Transform into array of days preserving order from start to end
        from datetime import datetime as _dt, timedelta as _td
        s = _dt.strptime(start, '%Y-%m-%d').date()
        e = _dt.strptime(end, '%Y-%m-%d').date()
        days = []
        cur = s
        while cur <= e:
            ds = cur.isoformat()
            days.append({ 'date': ds, 'scores': data.get(ds, []) })
            cur = cur + _td(days=1)

        return jsonify({"status": "success", "days": days}), 200
    except Exception as e:
        logging.exception("Error in /api/baiga/period_scores: %s", e)
        return jsonify({"error": str(e)}), 500

@app.route('/api/admin/users', methods=['GET'])
@require_api_key
def get_admin_users():
    try:
        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        requester_role = _normalize_user_role(requester[3]) if requester else ''
        visible_roles = ['operator', 'trainee', 'trainer']
        if requester_role == 'super_admin':
            visible_roles.append('admin')
        with db._get_cursor() as cursor:
            cursor.execute("""
                    SELECT
                        u.id,
                        u.name,
                        d.name as direction,
                        s.name as supervisor_name,
                        u.direction_id,
                        u.supervisor_id,
                        u.role,
                        u.status,
                        u.rate,
                        u.hire_date,
                        u.gender,
                        u.birth_date,
                        u.avatar_bucket,
                        u.avatar_blob_path,
                        u.avatar_updated_at,
                        sp.status_code as status_period_status_code,
                        sp.start_date as status_period_start_date,
                        sp.end_date as status_period_end_date,
                        sp.dismissal_reason as status_period_dismissal_reason,
                        COALESCE(sp.is_blacklist, FALSE) as status_period_is_blacklist,
                        sp.comment as status_period_comment,
                        u.phone,
                        u.email,
                        u.instagram,
                        u.telegram_nick,
                        u.company_name,
                        u.employment_type,
                        COALESCE(u.has_proxy, FALSE) as has_proxy,
                        u.sip_number,
                        u.study_place,
                        u.study_course,
                        u.close_contact_1_relation,
                        u.close_contact_1_full_name,
                        u.close_contact_1_phone,
                        u.close_contact_2_relation,
                        u.close_contact_2_full_name,
                        u.close_contact_2_phone,
                        u.card_number,
                        COALESCE(u.internship_in_company, FALSE) as internship_in_company,
                        COALESCE(u.front_office_training, FALSE) as front_office_training,
                        u.front_office_training_date,
                        u.taxipro_id
                    FROM users u
                    LEFT JOIN directions d ON u.direction_id = d.id
                    LEFT JOIN users s ON u.supervisor_id = s.id
                    LEFT JOIN LATERAL (
                        SELECT
                            p.status_code,
                            p.start_date,
                            p.end_date,
                            p.dismissal_reason,
                            p.is_blacklist,
                            p.comment
                        FROM operator_schedule_status_periods p
                        WHERE p.operator_id = u.id
                          AND p.status_code = (
                              CASE
                                  WHEN u.status = 'fired' THEN 'dismissal'
                                  WHEN u.status = 'dismissal' THEN 'dismissal'
                                  WHEN u.status = 'unpaid_leave' THEN 'bs'
                                  ELSE u.status
                              END
                          )
                        ORDER BY
                            CASE
                                WHEN p.start_date <= CURRENT_DATE
                                 AND COALESCE(p.end_date, DATE '9999-12-31') >= CURRENT_DATE
                                THEN 0
                                ELSE 1
                            END,
                            p.start_date DESC,
                            p.id DESC
                        LIMIT 1
                    ) sp ON TRUE
                    WHERE u.role = ANY(%s)
                """, (visible_roles,))
            users = []
            for row in cursor.fetchall():
                users.append({
                        "id": row[0],
                        "name": row[1],
                        "direction": row[2],
                        "supervisor_name": row[3],
                        "direction_id": row[4],
                        "supervisor_id": row[5],
                        "role": row[6],
                        "status": row[7],
                        "rate": float(row[8]),
                        "hire_date": row[9].strftime('%d-%m-%Y') if row[9] else None,
                        "gender": row[10],
                        "birth_date": row[11].strftime('%d-%m-%Y') if row[11] else None,
                        "avatar_url": _build_avatar_signed_url(row[12], row[13]),
                        "avatar_updated_at": row[14].isoformat() if row[14] else None,
                        "status_period_status_code": row[15],
                        "status_period_start_date": row[16].strftime('%Y-%m-%d') if row[16] else None,
                        "status_period_end_date": row[17].strftime('%Y-%m-%d') if row[17] else None,
                        "status_period_dismissal_reason": row[18] or "",
                        "status_period_is_blacklist": bool(row[19]) if row[19] is not None else False,
                        "status_period_comment": row[20] or "",
                        "phone": row[21] or "",
                        "email": row[22] or "",
                        "instagram": row[23] or "",
                        "telegram_nick": row[24] or "",
                        "company_name": row[25] or "",
                        "employment_type": row[26] or "",
                        "has_proxy": bool(row[27]) if row[27] is not None else False,
                        "sip_number": row[28] or "",
                        "study_place": row[29] or "",
                        "study_course": row[30] or "",
                        "close_contact_1_relation": row[31] or "",
                        "close_contact_1_full_name": row[32] or "",
                        "close_contact_1_phone": row[33] or "",
                        "close_contact_2_relation": row[34] or "",
                        "close_contact_2_full_name": row[35] or "",
                        "close_contact_2_phone": row[36] or "",
                        "card_number": row[37] or "",
                        "internship_in_company": bool(row[38]) if row[38] is not None else False,
                        "front_office_training": bool(row[39]) if row[39] is not None else False,
                        "front_office_training_date": row[40].strftime('%Y-%m-%d') if row[40] else None,
                        "taxipro_id": row[41] or ""
                    })
        return jsonify({"status": "success", "users": users}), 200
    except Exception as e:
        logging.error(f"Error fetching users: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/update_user', methods=['POST'])
@require_api_key
def admin_update_user():
    try:
        data = request.get_json()
        required_fields = ['user_id', 'field', 'value']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        user_id = int(data['user_id'])
        target_user = db.get_user(id=user_id)
        if not target_user:
            return jsonify({"error": "User not found"}), 404
        target_role = str(target_user[3] or '').strip().lower()

        field = data['field']
        value = data['value']
        if field in ['direction_id', 'supervisor_id']:
            if target_role == 'trainer':
                value = None
            else:
                try:
                    value = int(value) if value else None
                except (TypeError, ValueError):
                    return jsonify({"error": f"Invalid {field}"}), 400
        elif field == 'name':
            # Validate name and prevent duplicates
            if not value or not str(value).strip():
                return jsonify({"error": "Name cannot be empty"}), 400
            new_name = str(value).strip()
            # check duplicate by name (case-insensitive)
            existing = db.get_user_by_name(new_name)
            if existing and int(existing[0]) != user_id:
                return jsonify({"error": "Name already in use by another user"}), 400
            value = new_name
        elif field == 'status':
            if value not in ['working', 'fired', 'unpaid_leave', 'bs', 'sick_leave', 'annual_leave', 'dismissal']:
                return jsonify({"error": "Invalid status value"}), 400
        elif field == 'rate':
            try:
                value = float(value)
                if value not in [1.0, 0.75, 0.5]:
                    return jsonify({"error": "Invalid rate value"}), 400
            except ValueError:
                return jsonify({"error": "Invalid rate format"}), 400
        elif field in ['hire_date', 'birth_date', 'front_office_training_date']:
            if value:
                try:
                    datetime.strptime(value, '%Y-%m-%d')
                except ValueError:
                    return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
            else:
                value = None  # Allow clearing the date
        elif field == 'gender':
            if value in [None, '']:
                value = None
            elif value not in ['male', 'female']:
                return jsonify({"error": "Invalid gender value"}), 400
        elif field in [
            'phone',
            'email',
            'instagram',
            'telegram_nick',
            'company_name',
            'sip_number',
            'study_place',
            'study_course',
            'close_contact_1_relation',
            'close_contact_1_full_name',
            'close_contact_1_phone',
            'close_contact_2_relation',
            'close_contact_2_full_name',
            'close_contact_2_phone',
            'card_number',
            'taxipro_id'
        ]:
            value = str(value).strip() if value is not None else ''
            value = value or None
            if field == 'sip_number' and target_role != 'operator':
                value = None
            if field in ['phone', 'close_contact_1_phone', 'close_contact_2_phone'] and value and not _is_valid_kz_phone(value):
                return jsonify({"error": f"Invalid {field} format. Use +7XXXXXXXXXX"}), 400
            if field == 'email' and value and '@' not in value:
                return jsonify({"error": "Invalid email value"}), 400
        elif field == 'employment_type':
            value = str(value).strip().lower() if value is not None else ''
            value = value or None
            if value not in [None, 'gph', 'of']:
                return jsonify({"error": "Invalid employment_type value"}), 400
        elif field in ['has_proxy', 'internship_in_company', 'front_office_training']:
            if isinstance(value, bool):
                pass
            elif isinstance(value, (int, float)):
                value = bool(value)
            elif isinstance(value, str):
                value_normalized = value.strip().lower()
                if value_normalized in ['1', 'true', 'yes', 'y', 'on']:
                    value = True
                elif value_normalized in ['0', 'false', 'no', 'n', 'off', '']:
                    value = False
                else:
                    return jsonify({"error": f"Invalid {field} value"}), 400
            elif value is None:
                value = False
            else:
                return jsonify({"error": f"Invalid {field} value"}), 400
        else:
            return jsonify({"error": "Invalid field"}), 400

        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        requester_role = _normalize_user_role(requester[3]) if requester else ''
        if requester_role not in ('super_admin', 'admin', 'sv'):
            return jsonify({"error": "Only admins can update users"}), 403
        if target_role in ('admin', 'super_admin') and requester_role != 'super_admin':
            return jsonify({"error": "Only super admins can update admin users"}), 403

        success = db.update_user(user_id, field, value, changed_by=requester_id)  # Pass changed_by
        if not success:
            return jsonify({"error": "Failed to update user"}), 500

        return jsonify({"status": "success", "message": "User updated"}), 200
    except Exception as e:
        logging.error(f"Error updating user: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/users/bulk_update', methods=['POST'])
@require_api_key
def admin_bulk_update_users():
    try:
        data = request.get_json() or {}
        user_ids_raw = data.get('user_ids')
        changes_raw = data.get('changes')

        if not isinstance(user_ids_raw, list) or not user_ids_raw:
            return jsonify({"error": "user_ids must be a non-empty list"}), 400
        if not isinstance(changes_raw, dict) or not changes_raw:
            return jsonify({"error": "changes must be a non-empty object"}), 400

        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        requester_role = _normalize_user_role(requester[3]) if requester else ''
        if requester_role not in ('super_admin', 'admin', 'sv'):
            return jsonify({"error": "Only admins can update users"}), 403

        user_ids = []
        for raw_id in user_ids_raw:
            try:
                user_id = int(raw_id)
            except (TypeError, ValueError):
                return jsonify({"error": f"Invalid user id: {raw_id}"}), 400
            if user_id not in user_ids:
                user_ids.append(user_id)

        allowed_fields = {'direction_id', 'supervisor_id', 'rate'}
        unknown_fields = [field for field in changes_raw.keys() if field not in allowed_fields]
        if unknown_fields:
            return jsonify({"error": f"Unsupported fields: {', '.join(unknown_fields)}"}), 400

        updates = {}
        if 'direction_id' in changes_raw:
            direction_value = changes_raw.get('direction_id')
            if direction_value in [None, '']:
                return jsonify({"error": "direction_id cannot be empty"}), 400
            try:
                updates['direction_id'] = int(direction_value)
            except (TypeError, ValueError):
                return jsonify({"error": "Invalid direction_id"}), 400

        if 'supervisor_id' in changes_raw:
            supervisor_value = changes_raw.get('supervisor_id')
            if supervisor_value in [None, '']:
                return jsonify({"error": "supervisor_id cannot be empty"}), 400
            try:
                updates['supervisor_id'] = int(supervisor_value)
            except (TypeError, ValueError):
                return jsonify({"error": "Invalid supervisor_id"}), 400

        if 'rate' in changes_raw:
            try:
                rate_value = float(changes_raw.get('rate'))
            except (TypeError, ValueError):
                return jsonify({"error": "Invalid rate format"}), 400
            if rate_value not in [1.0, 0.75, 0.5]:
                return jsonify({"error": "Invalid rate value"}), 400
            updates['rate'] = rate_value

        if not updates:
            return jsonify({"error": "No valid changes provided"}), 400

        updated_count = 0
        failed_user_ids = []
        for target_user_id in user_ids:
            try:
                target_user = db.get_user(id=target_user_id)
                if not target_user:
                    failed_user_ids.append(target_user_id)
                    continue
                target_role = str(target_user[3] or '').strip().lower()
                update_ok = True
                for field, value in updates.items():
                    value_to_apply = None if (target_role == 'trainer' and field in ('direction_id', 'supervisor_id')) else value
                    if not db.update_user(target_user_id, field, value_to_apply, changed_by=requester_id):
                        update_ok = False
                        break
                if update_ok:
                    updated_count += 1
                else:
                    failed_user_ids.append(target_user_id)
            except Exception:
                logging.exception("Bulk update failed for user_id=%s", target_user_id)
                failed_user_ids.append(target_user_id)

        return jsonify({
            "status": "success",
            "updated_count": updated_count,
            "failed_user_ids": failed_user_ids,
            "applied_fields": list(updates.keys())
        }), 200
    except Exception as e:
        logging.error(f"Error bulk updating users: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/promote_to_supervisor', methods=['POST'])
@require_api_key
def admin_promote_to_supervisor():
    try:
        data = request.get_json() or {}
        user_id_raw = data.get('user_id')
        if user_id_raw in [None, '']:
            return jsonify({"error": "Missing user_id"}), 400

        try:
            target_user_id = int(user_id_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid user_id"}), 400

        requester_id_raw = request.headers.get('X-User-Id')
        if requester_id_raw in [None, '']:
            return jsonify({"error": "Missing X-User-Id header"}), 400

        try:
            requester_id = int(requester_id_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid X-User-Id header"}), 400

        requester = db.get_user(id=requester_id)
        if not requester or not _is_admin_role(requester[3]):
            return jsonify({"error": "Only admins can promote users"}), 403

        promoted_user = db.promote_operator_to_supervisor(target_user_id, changed_by=requester_id)
        return jsonify({
            "status": "success",
            "message": "User promoted to supervisor",
            "user": promoted_user
        }), 200
    except ValueError as e:
        error_message = str(e)
        if error_message == "User not found":
            return jsonify({"error": error_message}), 404
        if error_message in ("User is already a supervisor", "Only operators can be promoted to supervisor"):
            return jsonify({"error": error_message}), 400
        return jsonify({"error": error_message}), 400
    except Exception as e:
        if 'unique_name_role' in str(e):
            return jsonify({"error": "Cannot promote user: a supervisor with this name already exists"}), 409
        logging.error(f"Error promoting user to supervisor: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500
    
@app.route('/api/user/history', methods=['GET'])
@require_api_key
def get_user_history():
    try:
        user_id = request.args.get('user_id')
        if not user_id:
            return jsonify({"error": "Missing user_id parameter"}), 400
        user_id = int(user_id)

        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404

        target_user = db.get_user(id=user_id)
        if not target_user:
            return jsonify({"error": "User not found"}), 404

        # Permissions: Admin can view any, SV only their operators
        requester_role = _normalize_user_role(requester[3])
        if (not _is_admin_role(requester_role)) and (not _is_supervisor_role(requester_role) or target_user[6] != requester_id):
            return jsonify({"error": "Unauthorized to view this user's history"}), 403

        history = db.get_user_history(user_id)
        return jsonify({"status": "success", "history": history}), 200
    except Exception as e:
        logging.error(f"Error fetching user history: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/user/change_password', methods=['POST'])
@require_api_key
def change_password():
    try:
        data = request.get_json()
        required_fields = ['user_id', 'new_password']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields (user_id, new_password)"}), 400

        user_id = int(data['user_id'])
        new_password = data['new_password']
        requester_id = int(request.headers.get('X-User-Id', user_id))  # Assume user ID is passed in headers for authentication

        if not new_password or len(new_password) < 6:
            return jsonify({"error": "New password must be at least 6 characters long"}), 400

        # Fetch requester and target user
        requester = db.get_user(id=requester_id)
        target_user = db.get_user(id=user_id)

        if not requester or not target_user:
            return jsonify({"error": "User not found"}), 404

        # Authorization checks
        requester_role = _normalize_user_role(requester[3])
        target_role = _normalize_user_role(target_user[3])
        if requester_role == 'operator' and requester_id != user_id:
            return jsonify({"error": "Operators can only change their own password"}), 403

        if requester_role == 'sv' and (target_role != 'operator' or target_user[6] != requester_id) and requester_id != user_id:
            return jsonify({"error": "Supervisors can only change passwords for their operators"}), 403

        # Admins can change any password
        if (not _is_admin_role(requester_role)) and requester_id != user_id and not (requester_role == 'sv' and target_user[6] == requester_id):
            return jsonify({"error": "Unauthorized to change this user's password"}), 403

        # Update password
        success = db.update_user_password(user_id, new_password)
        if not success:
            return jsonify({"error": "Failed to update password"}), 500

        # Notify user via Telegram
        if target_user[1]:  # Check if telegram_id exists
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
            payload = {
                "chat_id": target_user[1],
                "text": f"Ваш пароль был успешно изменён. Новый пароль: <b>{new_password}</b>",
                "parse_mode": "HTML"
            }
            response = requests.post(telegram_url, json=payload, timeout=10)
            if response.status_code != 200:
                error_detail = response.json().get('description', 'Unknown error')
                logging.error(f"Telegram API error: {error_detail}")

        return jsonify({"status": "success", "message": f"Password updated for user {target_user[2]}"})

    except Exception as e:
        logging.error(f"Error changing password: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/sv/preview_calls_table', methods=['POST'])
@require_api_key
def preview_calls_table():
    try:
        # Проверяем наличие файла
        if 'file' not in request.files:
            return jsonify({"error": "Файл не был загружен"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "Файл не выбран"}), 400

        if not file.filename.lower().endswith(('.csv', '.xls', '.xlsx')):
            return jsonify({"error": "Неверный формат файла. Допустимы .csv, .xls, .xlsx"}), 400

        # Получаем user_id из заголовка
        user_id_header = request.headers.get('X-User-Id')
        if not user_id_header or not user_id_header.isdigit():
            return jsonify({"error": "Некорректный X-User-Id"}), 400
        user_id = int(user_id_header)

        user = db.get_user(id=user_id)
        if not user:
            return jsonify({"error": "Requester not found"}), 403

        requester_role = _normalize_user_role(user[3])
        if not (_is_supervisor_role(requester_role) or _is_admin_role(requester_role)):
            return jsonify({"error": "Unauthorized: только супервайзеры или администраторы могут загружать таблицы"}), 403

        selected_sv_id = None
        if _is_supervisor_role(requester_role):
            selected_sv_id = user_id
        else:
            sv_id_raw = (request.form.get('sv_id') or '').strip()
            if not sv_id_raw:
                return jsonify({"error": "Для администратора требуется выбранный супервайзер (sv_id)"}), 400
            try:
                selected_sv_id = int(sv_id_raw)
            except Exception:
                return jsonify({"error": "Некорректный sv_id"}), 400

            selected_sv = db.get_user(id=selected_sv_id)
            selected_sv_role = str(selected_sv[3] or '').strip().lower() if selected_sv else ''
            if not selected_sv or selected_sv_role not in ('sv', 'supervisor'):
                return jsonify({"error": "Супервайзер не найден"}), 404 

        # Только парсим файл, не сохраняем.
        # date (YYYY-MM-DD) опционально: если передан, вернем только строки этой даты.
        target_date = (request.form.get('date') or '').strip() or None
        sheet_name, operators, error = db.parse_calls_file(file, target_date=target_date)
        if error:
            return jsonify({"error": error}), 400

        return jsonify({
            "status": "success",
            "sheet_name": sheet_name,
            "sv_id": selected_sv_id,
            "operators": operators
        }), 200

    except Exception as e:
        logging.error(f"Ошибка при предпросмотре таблицы звонков: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/sv/update_norm_hours', methods=['POST'])
@require_api_key
def update_norm_hours():
    try:
        data = request.get_json()
        operator_id = data.get("operator_id")
        month = data.get("month")
        norm_hours = data.get("norm_hours")

        if not operator_id or not month:
            return jsonify({"error": "operator_id and month required"}), 400

        with db._get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO work_hours (operator_id, month, norm_hours)
                VALUES (%s, %s, %s)
                ON CONFLICT (operator_id, month)
                DO UPDATE SET norm_hours = EXCLUDED.norm_hours
            """, (operator_id, month, norm_hours))

        return jsonify({"status": "success"}), 200
    except Exception as e:
        logging.exception("update_norm_hours error")
        return jsonify({"error": str(e)}), 500

@app.route('/api/sv/daily_hours', methods=['GET'])
@require_api_key
def sv_daily_hours():
    """
    Возвращает daily_hours за месяц для всех операторов текущего супервайзера.
    Если запрос делает обычный оператор — возвращаем только его daily_hours.
    Параметры:
      - month (query param) — YYYY-MM (опционально, по умолчанию текущий месяц)
      - id (query param) — (только для admin) id нужного supervisor
    Заголовки:
      - X-User-Id (обязательно) — id супервайзера/оператора (проверяется роль)
    """
    try:
        # parse month
        month = request.args.get('month')
        if not month:
            month = datetime.now().strftime('%Y-%m')

        # requester id header
        requester_header = request.headers.get('X-User-Id')
        if not requester_header or not requester_header.isdigit():
            return jsonify({"error": "Invalid or missing X-User-Id header"}), 400
        requester_id = int(requester_header)

        # get requester from db
        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 403

        # try to read role in a robust way (support tuple or dict)
        role = None
        if isinstance(requester, dict):
            role = requester.get('role')
        elif isinstance(requester, (list, tuple)) and len(requester) > 3:
            role = requester[3]
        else:
            # fallback: try attribute access
            role = getattr(requester, 'role', None)

        if not role:
            # если роль не определена — ошибка
            return jsonify({"error": "User role not determined"}), 500
        role = _normalize_user_role(role)

        # -----------------------
        # Behavior for supervisors
        # -----------------------
        if _is_supervisor_role(role):
            # Allow optional ?id=<supervisor_id> to view another supervisor's data
            # (frontend may pass `id` when a supervisor selects another supervisor).
            # If no id param provided, fall back to requester_id (own team).
            user_param = request.args.get('id')
            if user_param and str(user_param).isdigit():
                supervisor_id = int(user_param)
            else:
                supervisor_id = requester_id

            try:
                result = db.get_daily_hours_by_supervisor_month(supervisor_id, month)
            except ValueError as e:
                return jsonify({"error": str(e)}), 400
            # result expected to contain "operators" list
            return jsonify({
                "status": "success",
                "month": result.get("month", month),
                "days_in_month": result.get("days_in_month"),
                "operators": result.get("operators", [])
            }), 200

        # -----------------------
        # Behavior for operators
        # -----------------------
        elif role == 'operator':
            try:
                result = db.get_daily_hours_for_operator_month(requester_id, month)
            except Exception as e:
                logging.exception("Error fetching daily hours for operator")
                return jsonify({"error": "Failed to fetch daily hours"}), 500

            operator_obj = result.get("operator") if isinstance(result, dict) else None
            return jsonify({
                "status": "success",
                "month": result.get("month", month) if isinstance(result, dict) else month,
                "days_in_month": result.get("days_in_month") if isinstance(result, dict) else None,
                "operators": [operator_obj] if operator_obj else []
            }), 200

        # -----------------------
        # Behavior for admins
        # -----------------------
        elif _is_admin_role(role):
            # admin must pass id (supervisor id) as query param "id"
            user_id = request.args.get('id')
            if not user_id or not str(user_id).isdigit():
                return jsonify({"error": "Missing or invalid 'id' parameter (supervisor id)"}), 400

            supervisor_id = int(user_id)
            try:
                result = db.get_daily_hours_by_supervisor_month(supervisor_id, month)
            except Exception as e:
                logging.exception("Error fetching daily hours for supervisor (admin request)")
                return jsonify({"error": str(e)}), 400

            # Normalize: expect result to contain "operators" list
            operators = result.get("operators") if isinstance(result, dict) else None
            if operators is None:
                # if older API returned single operator under "operator", handle it
                single = result.get("operator") if isinstance(result, dict) else None
                operators = [single] if single else []

            return jsonify({
                "status": "success",
                "month": result.get("month", month) if isinstance(result, dict) else month,
                "days_in_month": result.get("days_in_month"),
                "operators": operators
            }), 200

        else:
            return jsonify({"error": "Only supervisors, operators and admins can request this endpoint"}), 403

    except Exception as e:
        logging.exception("Error in sv_daily_hours")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/hours/upload_group_day', methods=['POST'])
@require_api_key
def upload_group_day():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Empty payload"}), 400

        # Парсим базовые поля
        date_str = data.get('date')  # ожидаем YYYY-MM-DD
        if not date_str:
            return jsonify({"error": "Field 'date' is required (YYYY-MM-DD)"}), 400

        try:
            # проверяем формат даты
            day_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400

        default_month = date_str[:7]  # YYYY-MM

        sv_id_payload = data.get('sv_id')  # может быть None
        operators = data.get('operators')
        if not operators or not isinstance(operators, list):
            return jsonify({"error": "Field 'operators' must be a non-empty array"}), 400

        # Получаем id запроса (и проверяем пользователя)
        requester_header = request.headers.get('X-User-Id')
        if not requester_header or not requester_header.isdigit():
            return jsonify({"error": "Invalid or missing X-User-Id"}), 400
        requester_id = int(requester_header)
        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 403

        requester_role = _normalize_user_role(requester[3])  # согласно get_user селекту: u.role в позиции 3

        # Если запрос делает супервайзер — разрешаем только для его id (или если payload sv_id указан, он должен совпадать)
        if _is_supervisor_role(requester_role):
            sv_id = requester_id
        elif _is_admin_role(requester_role):
            if sv_id_payload in (None, ''):
                return jsonify({"error": "Field 'sv_id' is required for admin requests"}), 400
            try:
                sv_id = int(sv_id_payload)
            except Exception:
                return jsonify({"error": "Invalid 'sv_id' value"}), 400

            supervisor_user = db.get_user(id=sv_id)
            supervisor_role = str(supervisor_user[3] or '').strip().lower() if supervisor_user else ''
            if not supervisor_user or supervisor_role not in ('sv', 'supervisor'):
                return jsonify({"error": "Selected supervisor not found"}), 404
        else:
            return jsonify({"error": "Unauthorized: only supervisors or admins can upload hours"}), 403

        processed = []
        skipped = []
        errors = []
        processed_operator_ids = set()
        processed_months_by_operator = {}

        for idx, row in enumerate(operators):
            try:
                # Normalize fields
                op_id = row.get('operator_id')
                name = str(row.get('name') or '').strip()

                row_date_str = str(row.get('date') or date_str).strip()
                try:
                    row_day_obj = datetime.strptime(row_date_str, "%Y-%m-%d").date()
                except Exception:
                    skipped.append({"row": idx, "reason": "invalid row date format", "date": row.get('date')})
                    continue

                row_month = str(row.get('month') or row_date_str[:7]).strip()
                if not re.match(r'^\d{4}-\d{2}$', row_month):
                    row_month = row_date_str[:7]

                def _optional_float(field_name):
                    if field_name not in row:
                        return None
                    raw_val = row.get(field_name)
                    if raw_val is None or raw_val == '':
                        return None
                    try:
                        return float(raw_val)
                    except Exception:
                        return None

                def _optional_int(field_name):
                    if field_name not in row:
                        return None
                    raw_val = row.get(field_name)
                    if raw_val is None or raw_val == '':
                        return 0
                    try:
                        return max(int(float(raw_val)), 0)
                    except Exception:
                        return 0

                # Поля, которых нет в payload, должны быть сохранены как есть из текущей записи дня.
                # Это позволяет при загрузке файла обновлять только calls, не затирая часы.
                work_time = _optional_float('work_time')
                break_time = _optional_float('break_time')
                talk_time = _optional_float('talk_time')
                calls = _optional_int('calls')
                efficiency = _optional_float('efficiency')
                fine_amount = _optional_float('fine_amount')

                fine_reason_provided = 'fine_reason' in row
                fine_comment_provided = 'fine_comment' in row
                fine_reason = (row.get('fine_reason') if fine_reason_provided else None)
                fine_comment = (row.get('fine_comment') if fine_comment_provided else None)
                if fine_reason is not None:
                    fine_reason = str(fine_reason).strip() or None
                if fine_comment is not None:
                    fine_comment = str(fine_comment).strip() or None

                # защита: принимаем только разрешённые причины (опционально)
                ALLOWED_FINE_REASONS = ['Корп такси', 'Опоздание', 'Прокси карта', 'Не выход', 'Другое']
                if fine_reason and fine_reason not in ALLOWED_FINE_REASONS:
                    # если невалидная причина — помечаем как "Другое" или сохраняем как есть, в примере — нормализуем в 'Другое'
                    fine_reason = 'Другое'
                # accept fines array (new format): list of {amount, reason, comment}
                fines_arr = row.get('fines') if ('fines' in row and isinstance(row.get('fines'), list)) else None

                # resolve operator_id if not provided
                resolved_operator_id = None
                if op_id:
                    # verify operator exists
                    try:
                        op_id_int = int(op_id)
                        user_row = db.get_user(id=op_id_int)
                        if user_row and user_row[3] == 'operator':  # role == operator
                            # Ensure operator belongs to the selected/current supervisor.
                            if user_row[6] != sv_id:  # user_row[6] == supervisor_id in get_user select
                                skipped.append({"row": idx, "reason": "operator not under selected supervisor", "name": name, "operator_id": op_id})
                                continue
                            resolved_operator_id = op_id_int
                        else:
                            skipped.append({"row": idx, "reason": "operator_id not found or not operator", "name": name, "operator_id": op_id})
                            continue
                    except Exception:
                        skipped.append({"row": idx, "reason": "invalid operator_id", "value": op_id})
                        continue
                else:
                    # try find by name under sv_id (if sv_id is known)
                    if sv_id:
                        with db._get_cursor() as cursor:
                            cursor.execute("""
                                SELECT id FROM users
                                WHERE name = %s AND role = 'operator' AND supervisor_id = %s
                                LIMIT 1
                            """, (name, sv_id))
                            res = cursor.fetchone()
                            if res:
                                resolved_operator_id = res[0]
                    # if still not found, try global lookup (supervisor requests only)
                    if not resolved_operator_id:
                        if _is_admin_role(requester_role):
                            skipped.append({"row": idx, "reason": "operator not found under selected supervisor", "name": name})
                            continue
                        with db._get_cursor() as cursor:
                            cursor.execute("""
                                SELECT id FROM users
                                WHERE name = %s AND role = 'operator'
                                LIMIT 1
                            """, (name,))
                            res2 = cursor.fetchone()
                            if res2:
                                # if requester is sv, ensure operator belongs to them
                                if _is_supervisor_role(requester_role):
                                    # check supervisor
                                    with db._get_cursor() as c2:
                                        c2.execute("SELECT supervisor_id FROM users WHERE id = %s", (res2[0],))
                                        sp = c2.fetchone()
                                        if sp and sp[0] == sv_id:
                                            resolved_operator_id = res2[0]
                                        else:
                                            skipped.append({"row": idx, "reason": "operator found but not under this supervisor", "name": name})
                                            continue
                                else:
                                    resolved_operator_id = res2[0]

                if not resolved_operator_id:
                    skipped.append({"row": idx, "reason": "operator not found", "name": name})
                    continue

                # Подтягиваем текущие значения дня, если какое-то поле не передано в payload.
                existing_daily = None
                needs_existing = any(v is None for v in (work_time, break_time, talk_time, calls, efficiency, fine_amount))
                needs_existing = needs_existing or (not fine_reason_provided) or (not fine_comment_provided)
                if needs_existing:
                    with db._get_cursor() as cursor:
                        cursor.execute("""
                            SELECT
                                COALESCE(work_time, 0),
                                COALESCE(break_time, 0),
                                COALESCE(talk_time, 0),
                                COALESCE(calls, 0),
                                COALESCE(efficiency, 0),
                                COALESCE(fine_amount, 0),
                                fine_reason,
                                fine_comment
                            FROM daily_hours
                            WHERE operator_id = %s AND day = %s
                            LIMIT 1
                        """, (resolved_operator_id, row_day_obj))
                        existing_daily = cursor.fetchone()

                existing_work = float(existing_daily[0]) if existing_daily else 0.0
                existing_break = float(existing_daily[1]) if existing_daily else 0.0
                existing_talk = float(existing_daily[2]) if existing_daily else 0.0
                existing_calls = int(existing_daily[3]) if existing_daily else 0
                existing_eff = float(existing_daily[4]) if existing_daily else 0.0
                existing_fine_amount = float(existing_daily[5]) if existing_daily else 0.0
                existing_fine_reason = (existing_daily[6] if existing_daily else None)
                existing_fine_comment = (existing_daily[7] if existing_daily else None)

                work_time = float(existing_work if work_time is None else work_time)
                break_time = float(existing_break if break_time is None else break_time)
                talk_time = float(existing_talk if talk_time is None else talk_time)
                calls = max(int(existing_calls if calls is None else calls), 0)
                efficiency = float(existing_eff if efficiency is None else efficiency)
                fine_amount = float(existing_fine_amount if fine_amount is None else fine_amount)

                if not fine_reason_provided:
                    fine_reason = existing_fine_reason
                if not fine_comment_provided:
                    fine_comment = existing_fine_comment

                # Insert/update daily_hours
                try:
                    logging.info(
                        "Fine data: amount=%s, reason=%s, comment=%s",
                        fine_amount, fine_reason, fine_comment
                    )
                    # use the Database helper (assumes method exists)
                    db.insert_or_update_daily_hours(resolved_operator_id, row_date_str,
                                                    work_time=work_time,
                                                    break_time=break_time,
                                                    talk_time=talk_time,
                                                    calls=calls,
                                                    efficiency=efficiency,
                                                    fine_amount=fine_amount,
                                                    fine_reason=fine_reason,
                                                    fine_comment=fine_comment,
                                                    fines=fines_arr)
                    processed.append({"row": idx, "operator_id": resolved_operator_id, "name": name, "date": row_date_str})
                    processed_operator_ids.add(resolved_operator_id)
                    processed_months_by_operator.setdefault(resolved_operator_id, set()).add(row_month)
                except Exception as e:
                    errors.append({"row": idx, "operator_id": resolved_operator_id, "error": str(e)})
                    continue

            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
                continue

        # После всех вставок — агрегируем месяц для каждого обработанного оператора
        aggregations = {}
        aggregations_by_month = {}
        for opid in processed_operator_ids:
            months_for_op = sorted(processed_months_by_operator.get(opid) or {default_month})
            aggregations_by_month[opid] = {}
            for month_value in months_for_op:
                try:
                    agg = db.aggregate_month_from_daily(opid, month_value)
                    aggregations_by_month[opid][month_value] = agg
                except Exception as e:
                    aggregations_by_month[opid][month_value] = {"error": str(e)}
            if months_for_op:
                last_month = months_for_op[-1]
                aggregations[opid] = aggregations_by_month[opid].get(last_month)

        return jsonify({
            "status": "success",
            "date": date_str,
            "month": default_month,
            "processed_count": len(processed),
            "processed": processed,
            "skipped_count": len(skipped),
            "skipped": skipped,
            "errors": errors,
            "aggregations": aggregations,
            "aggregations_by_month": aggregations_by_month
        }), 200

    except Exception as e:
        logging.exception("Error in upload_group_day")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/sv/save_table', methods=['POST'])
@require_api_key
def save_table():
    try:
        data = request.get_json()
        required_fields = ['table_url', 'direction_id', 'operators']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields: table_url, direction_id, operators"}), 400

        table_url = data['table_url']
        direction_id = int(data['direction_id'])
        operators = data['operators']  # List of dicts with 'name'

        user_id = int(request.headers.get('X-User-Id'))
        user = db.get_user(id=user_id)
        if not user or user[3] != 'sv':
            return jsonify({"error": "Unauthorized: Only supervisors can save tables"}), 403

        # Validate direction
        directions = db.get_directions()
        if not any(d['id'] == direction_id for d in directions):
            return jsonify({"error": "Invalid direction_id"}), 400

        # Update supervisor's hours table
        db.update_user_table(user_id=user_id, hours_table_url=table_url)

        # Create/update operators with direction and supervisor
        for op in operators:
            db.create_user(
                telegram_id=None,
                name=op['name'],
                role='operator',
                direction_id=direction_id,
                supervisor_id=user_id,
                hours_table_url=table_url
            )

        # Optional: Send Telegram notification to supervisor
        telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
        payload = {
            "chat_id": user[1],
            "text": f"Таблица часов успешно обновлена и операторы сохранены с направлением ID {direction_id} ✅",
            "parse_mode": "HTML"
        }
        response = requests.post(telegram_url, json=payload, timeout=10)
        if response.status_code != 200:
            logging.error(f"Telegram notification failed: {response.json().get('description')}")

        return jsonify({
            "status": "success",
            "message": "Table updated, operators saved with selected direction"
        }), 200
    except Exception as e:
        logging.error(f"Error saving table: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/user/change_login', methods=['POST'])
@require_api_key
def change_login():
    try:
        data = request.get_json()
        required_fields = ['user_id', 'new_login']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields (user_id, new_login)"}), 400

        user_id = int(data['user_id'])
        new_login = data['new_login']
        requester_id = int(request.headers.get('X-User-Id', user_id))

        if not new_login or len(new_login) < 4:
            return jsonify({"error": "New login must be at least 4 characters long"}), 400

        # Check if new login is already taken
        if db.get_user_by_login(new_login):
            return jsonify({"error": "Login is already in use"}), 400

        # Fetch requester and target user
        requester = db.get_user(id=requester_id)
        target_user = db.get_user(id=user_id)

        if not requester or not target_user:
            return jsonify({"error": "User not found"}), 404

        # Authorization checks
        requester_role = _normalize_user_role(requester[3])
        target_role = _normalize_user_role(target_user[3])
        if requester_role == 'operator' and requester_id != user_id:
            return jsonify({"error": "Operators can only change their own login"}), 403

        if requester_role == 'sv' and (target_role != 'operator' or target_user[6] != requester_id) and requester_id != user_id:
            return jsonify({"error": "Supervisors can only change logins for their operators"}), 403

        # Admins can change any login
        if (not _is_admin_role(requester_role)) and requester_id != user_id and not (requester_role == 'sv' and target_user[6] == requester_id):
            return jsonify({"error": "Unauthorized to change this user's login"}), 403

        # Update login
        success = db.update_operator_login(user_id, target_user[6] if requester_role in ('sv', 'admin', 'super_admin') else None, new_login)
        if not success:
            return jsonify({"error": "Failed to update login"}), 500

        # Notify user via Telegram
        if target_user[1]:
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
            payload = {
                "chat_id": target_user[1],
                "text": f"Ваш логин был успешно изменён. Новый логин: <b>{new_login}</b>",
                "parse_mode": "HTML"
            }
            response = requests.post(telegram_url, json=payload, timeout=10)
            if response.status_code != 200:
                error_detail = response.json().get('description', 'Unknown error')
                logging.error(f"Telegram API error: {error_detail}")

        return jsonify({"status": "success", "message": f"Login updated for user {target_user[2]}"})

    except Exception as e:
        logging.error(f"Error changing login: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/user/hours', methods=['GET'])
@require_api_key
def get_user_hours():
    try:
        operator_id = request.args.get('operator_id')
        month = request.args.get('month')
        if not operator_id:
            return jsonify({"error": "Missing operator_id parameter"}), 400
        if not month:
            month = datetime.now().strftime('%Y-%m')
        operator_id = int(operator_id)
        hours_summary = db.get_hours_summary(operator_id=operator_id, month=month)
        
        if not hours_summary:
            return jsonify({"status": "success", "hours": None}), 200

        # Assuming get_hours_summary returns a list, take the first item
        hours_data = hours_summary[0]

        return jsonify({"status": "success", "hours": hours_data}), 200
    except Exception as e:
        logging.error(f"Error fetching hours data: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/sv_list', methods=['GET'])
@require_api_key
def get_sv_list():
    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, name, hours_table_url, role, hire_date, status, gender, birth_date, avatar_bucket, avatar_blob_path
                FROM users
                WHERE role = 'sv'
                ORDER BY name
            """)
            supervisors = cursor.fetchall()

        logging.info(f"Fetched {len(supervisors)} supervisors")
        sv_data = [
            {
                "id": sv[0],
                "name": sv[1],
                "table": sv[2],
                "role": sv[3],
                "hire_date": sv[4].strftime('%d-%m-%Y') if sv[4] else None,
                "status": sv[5],
                "gender": sv[6],
                "birth_date": sv[7].strftime('%d-%m-%Y') if sv[7] else None,
                "avatar_url": _build_avatar_signed_url(sv[8], sv[9])
            }
            for sv in supervisors
        ]
        return jsonify({"status": "success", "sv_list": sv_data})
    except Exception as e:
        logging.error(f"Error fetching SV list: {e}", exc_info=True)
        return jsonify({"error": f"Failed to fetch supervisors: {str(e)}"}), 500

@app.route('/api/call_evaluations', methods=['GET'])
@require_api_key
def get_call_evaluations():
    try:
        operator_id = request.args.get('operator_id')
        month = request.args.get('month')
        if not operator_id:
            return jsonify({"error": "Missing operator_id parameter"}), 400
        if not month:
            month = None
        operator_id = int(operator_id)
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404

        if not _authorize_operator_scope(requester, requester_id, operator_id):
            return jsonify({"error": "Unauthorized to view this operator evaluations"}), 403

        evaluations = db.get_call_evaluations(operator_id, month=month)
        session_id = _current_session_id_from_access_token()
        role = requester[3]
        if role == 'operator':
            reveal_sensitive = bool(_is_sensitive_access_unlocked(requester_id, session_id))
        else:
            reveal_sensitive = True
        evaluations = _sanitize_evaluations_for_access(
            evaluations,
            reveal_sensitive,
            hide_hidden_operator_comments=(role == 'operator')
        )

        # Получаем информацию о супервайзере для dispute button
        operator = db.get_user(id=operator_id)
        supervisor = db.get_user(id=operator[6]) if operator and operator[6] else None

        # Просто возвращаем массивы из базы, не пересоздаём их из комментариев
        return jsonify({
            "status": "success", 
            "evaluations": evaluations,
            "supervisor": {
                "id": supervisor[0] if supervisor else None,
                "name": supervisor[2] if supervisor else None
            },
            "sensitive_access": {
                "required": role == 'operator',
                "granted": reveal_sensitive
            }
        })
    except Exception as e:
        logging.error(f"Error fetching evaluations: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms', methods=['GET'])
@require_api_key
def list_call_calibration_rooms():
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        role = _normalize_user_role(requester[3])
        if not _is_admin_role(role) and not _is_supervisor_role(role):
            return jsonify({"error": "Forbidden: only admin and supervisors can access calibration rooms"}), 403

        month = (request.args.get('month') or '').strip()
        if month and not re.match(r'^\d{4}-\d{2}$', month):
            return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400

        with db._get_cursor() as cursor:
            params = [requester_id, requester_id]
            query = """
                SELECT
                    r.id,
                    r.month,
                    r.room_title,
                    r.created_at,
                    adm.id AS admin_id,
                    adm.name AS admin_name,
                    m.id AS member_id,
                    COALESCE(stats.calls_count, 0) AS calls_count,
                    COALESCE(stats.evaluators_count, 0) AS evaluators_count,
                    COALESCE(stats.evaluations_count, 0) AS evaluations_count,
                    COALESCE(my_stats.my_evaluated_calls, 0) AS my_evaluated_calls
                FROM calibration_rooms r
                JOIN users adm ON adm.id = r.created_by_admin_id
                LEFT JOIN calibration_room_members m
                    ON m.room_id = r.id
                   AND m.supervisor_id = %s
                LEFT JOIN (
                    SELECT
                        c.room_id,
                        COUNT(*) AS calls_count,
                        COUNT(e.id) AS evaluations_count,
                        COUNT(DISTINCT e.evaluator_id) AS evaluators_count
                    FROM calibration_room_calls c
                    LEFT JOIN calibration_room_call_evaluations e ON e.room_call_id = c.id
                    GROUP BY c.room_id
                ) stats ON stats.room_id = r.id
                LEFT JOIN (
                    SELECT
                        c.room_id,
                        COUNT(DISTINCT c.id) AS my_evaluated_calls
                    FROM calibration_room_calls c
                    JOIN calibration_room_call_evaluations e
                      ON e.room_call_id = c.id
                     AND e.evaluator_id = %s
                    GROUP BY c.room_id
                ) my_stats ON my_stats.room_id = r.id
                WHERE 1=1
            """
            if month:
                query += " AND r.month = %s"
                params.append(month)
            query += " ORDER BY r.created_at DESC"
            cursor.execute(query, params)
            rows = cursor.fetchall()

        rooms = []
        for row in rows:
            calls_count = int(row[7] or 0)
            my_evaluated_calls = int(row[10] or 0)
            is_creator = bool(row[4] and int(row[4]) == int(requester_id))
            rooms.append({
                "id": row[0],
                "month": row[1],
                "room_title": row[2] or f"Комната #{row[0]}",
                "created_at": row[3].strftime('%Y-%m-%d %H:%M:%S') if row[3] and hasattr(row[3], "strftime") else None,
                "benchmark_admin": {"id": row[4], "name": row[5]},
                "joined": True if _is_admin_role(role) else bool(row[6] or is_creator),
                "calls_count": calls_count,
                "evaluated_count": int(row[8] or 0),
                "evaluation_rows_count": int(row[9] or 0),
                "my_evaluated_calls": my_evaluated_calls,
                "my_evaluated": bool(calls_count > 0 and my_evaluated_calls >= calls_count),
            })

        return jsonify({"status": "success", "rooms": rooms}), 200
    except Exception as e:
        logging.exception("Error listing calibration rooms")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms', methods=['POST'])
@require_api_key
def create_call_calibration_room():
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        if not _is_admin_role(requester[3]) and not _is_supervisor_role(requester[3]):
            return jsonify({"error": "Only admin and supervisors can create calibration rooms"}), 403

        data = request.form if request.form else (request.get_json(silent=True) or {})
        month = str(data.get('month') or '').strip()
        if not re.match(r'^\d{4}-\d{2}$', month):
            return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400

        room_title = str(data.get('room_title') or data.get('title') or '').strip() or None
        if room_title and len(room_title) > 255:
            return jsonify({"error": "room_title is too long (max 255)"}), 400

        with db._get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO calibration_rooms (
                    created_by_admin_id,
                    month,
                    room_title,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id
            """, (
                requester_id,
                month,
                room_title,
            ))
            room_id = cursor.fetchone()[0]

        return jsonify({"status": "success", "room_id": room_id}), 201
    except Exception as e:
        logging.exception("Error creating calibration room")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>', methods=['DELETE', 'OPTIONS'])
@require_api_key
def delete_calibration_room(room_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        if not _is_admin_role(requester[3]):
            return jsonify({"error": "Only admin can delete calibration rooms"}), 403

        with db._get_cursor() as cursor:
            cursor.execute("SELECT id FROM calibration_rooms WHERE id = %s", (room_id,))
            if not cursor.fetchone():
                return jsonify({"error": "Комната не найдена"}), 404
            cursor.execute("DELETE FROM calibration_rooms WHERE id = %s", (room_id,))

        return jsonify({"status": "success"}), 200
    except Exception as e:
        logging.exception("Error deleting calibration room")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>', methods=['PATCH', 'OPTIONS'])
@require_api_key
def update_calibration_room(room_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        requester_role = _normalize_user_role(requester[3])
        if not _is_admin_role(requester_role) and not _is_supervisor_role(requester_role):
            return jsonify({"error": "Only admin and supervisors can update calibration rooms"}), 403

        data = request.form if request.form else (request.get_json(silent=True) or {})
        room_title = str(data.get('room_title') or data.get('title') or '').strip()
        if len(room_title) > 255:
            return jsonify({"error": "room_title is too long (max 255)"}), 400

        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, created_by_admin_id
                FROM calibration_rooms
                WHERE id = %s
            """, (room_id,))
            room_row = cursor.fetchone()
            if not room_row:
                return jsonify({"error": "Комната не найдена"}), 404

            if _is_supervisor_role(requester_role):
                room_creator_id = int(room_row[1]) if room_row[1] is not None else None
                if int(requester_id) != int(room_creator_id or 0):
                    cursor.execute("""
                        SELECT id
                        FROM calibration_room_members
                        WHERE room_id = %s AND supervisor_id = %s
                    """, (room_id, requester_id))
                    if not cursor.fetchone():
                        return jsonify({"error": "Join the room first"}), 403

            cursor.execute("""
                UPDATE calibration_rooms
                SET
                    room_title = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                RETURNING id, room_title
            """, (room_title or None, room_id))
            updated = cursor.fetchone()
            if not updated:
                return jsonify({"error": "Комната не найдена"}), 404

        return jsonify({
            "status": "success",
            "room": {
                "id": updated[0],
                "room_title": updated[1] or f"Комната #{updated[0]}"
            }
        }), 200
    except Exception as e:
        logging.exception("Error updating calibration room")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>/calls', methods=['POST'])
@require_api_key
def add_call_to_calibration_room(room_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        requester_role = _normalize_user_role(requester[3])
        if not _is_admin_role(requester_role) and not _is_supervisor_role(requester_role):
            return jsonify({"error": "Only admin and supervisors can add calls to calibration room"}), 403

        data = request.form if request.form else (request.get_json(silent=True) or {})
        required_fields = ['operator_id', 'phone_number', 'appeal_date', 'direction', 'scores', 'criterion_comments']
        missing = [f for f in required_fields if f not in data]
        if missing:
            return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400

        try:
            operator_id = int(data.get('operator_id'))
            direction_id = int(data.get('direction'))
        except Exception:
            return jsonify({"error": "operator_id and direction must be integers"}), 400

        month = str(data.get('month') or '').strip()
        if month and not re.match(r'^\d{4}-\d{2}$', month):
            return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400

        phone_number = str(data.get('phone_number') or '').strip()
        if len(phone_number) < 5:
            return jsonify({"error": "Invalid phone number"}), 400

        appeal_date_raw = str(data.get('appeal_date') or '').strip()
        try:
            appeal_date = datetime.fromisoformat(appeal_date_raw.replace('Z', '+00:00'))
        except Exception:
            return jsonify({"error": "Invalid appeal_date format. Use ISO datetime"}), 400

        raw_scores = data.get('scores', '[]')
        raw_comments = data.get('criterion_comments', '[]')
        try:
            scores = json.loads(raw_scores) if isinstance(raw_scores, str) else raw_scores
            criterion_comments = json.loads(raw_comments) if isinstance(raw_comments, str) else raw_comments
        except Exception:
            return jsonify({"error": "scores and criterion_comments must be valid JSON arrays"}), 400

        if not isinstance(scores, list) or not isinstance(criterion_comments, list):
            return jsonify({"error": "scores and criterion_comments must be arrays"}), 400

        operator = db.get_user(id=operator_id)
        if not operator or operator[3] != 'operator':
            return jsonify({"error": "Operator not found"}), 404
        if not _authorize_operator_scope(requester, requester_id, operator_id):
            return jsonify({"error": "Forbidden: you cannot add calls for this operator"}), 403

        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, month, created_by_admin_id
                FROM calibration_rooms
                WHERE id = %s
            """, (room_id,))
            room_row = cursor.fetchone()
            if not room_row:
                return jsonify({"error": "Calibration room not found"}), 404

            room_creator_id = room_row[2]
            if _is_supervisor_role(requester_role) and int(room_creator_id or 0) != int(requester_id):
                cursor.execute("""
                    SELECT id
                    FROM calibration_room_members
                    WHERE room_id = %s AND supervisor_id = %s
                """, (room_id, requester_id))
                if not cursor.fetchone():
                    return jsonify({"error": "Join the room first"}), 403

            room_month = str(room_row[1] or '').strip()
            if room_month and month and room_month != month:
                return jsonify({"error": "Call month must match room month"}), 400
            month = room_month or month
            if not month:
                return jsonify({"error": "Calibration room month is missing"}), 400

            cursor.execute("""
                SELECT id, criteria, has_file_upload
                FROM directions
                WHERE id = %s AND is_active = TRUE
            """, (direction_id,))
            direction = cursor.fetchone()
            if not direction:
                return jsonify({"error": "Direction not found"}), 404

            criteria = direction[1] if isinstance(direction[1], list) else []
            has_file_upload = bool(direction[2])

            normalized_scores = []
            normalized_comments = []
            for idx in range(len(criteria)):
                src_score = scores[idx] if idx < len(scores) else 'Correct'
                src_comment = criterion_comments[idx] if idx < len(criterion_comments) else ''
                normalized_scores.append(_normalize_calibration_score_value(src_score))
                normalized_comments.append(str(src_comment or ''))

            total_score, _ = _compute_total_score_from_criteria(criteria, normalized_scores)
            comment = str(data.get('comment') or '').strip()
            if not comment:
                comment = '; '.join(
                    f"{(criteria[i] or {}).get('name', f'Критерий {i + 1}')}: {text.strip()}"
                    for i, text in enumerate(normalized_comments)
                    if str(text or '').strip()
                )

            audio_path = None
            if has_file_upload:
                file = request.files.get('audio_file')
                if not file or not file.filename:
                    return jsonify({"error": "Audio file is required for this direction"}), 400
                audio_data = file.read()
                if not audio_data:
                    return jsonify({"error": "Audio file is empty"}), 400
                bucket_name = os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET')
                if not bucket_name:
                    return jsonify({"error": "GOOGLE_CLOUD_STORAGE_BUCKET is not configured"}), 500
                upload_folder = os.getenv('UPLOAD_FOLDER', 'Uploads/')
                filename = secure_filename(f"{uuid.uuid4()}.mp3")
                blob_path = f"{upload_folder}{filename}"
                gcs_client = get_gcs_client()
                bucket = gcs_client.bucket(bucket_name)
                blob = bucket.blob(blob_path)
                blob.upload_from_string(audio_data, content_type='audio/mpeg')
                audio_path = f"{bucket_name}/{blob_path}"

            cursor.execute("""
                INSERT INTO calibration_room_calls (
                    room_id,
                    created_by_admin_id,
                    operator_id,
                    month,
                    phone_number,
                    appeal_date,
                    direction_id,
                    score,
                    comment,
                    audio_path,
                    scores,
                    criterion_comments,
                    etalon_scores,
                    etalon_criterion_comments,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, %s::jsonb, %s::jsonb, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id
            """, (
                room_id,
                requester_id,
                operator_id,
                month,
                phone_number,
                appeal_date,
                direction_id,
                total_score,
                comment,
                audio_path,
                json.dumps(normalized_scores, ensure_ascii=False),
                json.dumps(normalized_comments, ensure_ascii=False),
                json.dumps(normalized_scores, ensure_ascii=False),
                json.dumps(normalized_comments, ensure_ascii=False),
            ))
            room_call_id = cursor.fetchone()[0]

        return jsonify({"status": "success", "room_call_id": room_call_id, "score": total_score}), 201
    except Exception as e:
        logging.exception("Error adding call to calibration room")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>/calls/<int:call_id>', methods=['DELETE', 'OPTIONS'])
@require_api_key
def delete_call_from_calibration_room(room_id, call_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        if not _is_admin_role(requester[3]):
            return jsonify({"error": "Only admin can delete calls from calibration room"}), 403

        with db._get_cursor() as cursor:
            cursor.execute(
                "SELECT id FROM calibration_room_calls WHERE id = %s AND room_id = %s",
                (call_id, room_id)
            )
            if not cursor.fetchone():
                return jsonify({"error": "Звонок не найден"}), 404
            cursor.execute(
                "DELETE FROM calibration_room_calls WHERE id = %s AND room_id = %s",
                (call_id, room_id)
            )

        return jsonify({"status": "success"}), 200
    except Exception as e:
        logging.exception("Error deleting call from calibration room")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>/join', methods=['POST'])
@require_api_key
def join_call_calibration_room(room_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        if not _is_supervisor_role(requester[3]):
            return jsonify({"error": "Only supervisors can join calibration rooms"}), 403

        with db._get_cursor() as cursor:
            cursor.execute("SELECT id FROM calibration_rooms WHERE id = %s", (room_id,))
            if not cursor.fetchone():
                return jsonify({"error": "Calibration room not found"}), 404

            cursor.execute("""
                INSERT INTO calibration_room_members (room_id, supervisor_id)
                VALUES (%s, %s)
                ON CONFLICT (room_id, supervisor_id) DO NOTHING
            """, (room_id, requester_id))

        return jsonify({"status": "success"}), 200
    except Exception as e:
        logging.exception("Error joining calibration room")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>', methods=['GET'])
@require_api_key
def get_call_calibration_room(room_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        role = _normalize_user_role(requester[3])
        if not _is_admin_role(role) and not _is_supervisor_role(role):
            return jsonify({"error": "Forbidden: only admin and supervisors can access calibration room"}), 403

        requested_call_id = None
        requested_call_id_raw = request.args.get('call_id')
        if requested_call_id_raw not in (None, ''):
            try:
                requested_call_id = int(requested_call_id_raw)
            except Exception:
                return jsonify({"error": "Invalid call_id"}), 400

        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    r.id,
                    r.month,
                    r.room_title,
                    r.created_at,
                    adm.id AS admin_id,
                    adm.name AS admin_name
                FROM calibration_rooms r
                JOIN users adm ON adm.id = r.created_by_admin_id
                WHERE r.id = %s
            """, (room_id,))
            room_row = cursor.fetchone()
            if not room_row:
                return jsonify({"error": "Calibration room not found"}), 404

            if _is_supervisor_role(role):
                is_creator = int(room_row[4] or 0) == int(requester_id)
                cursor.execute("""
                    SELECT id FROM calibration_room_members
                    WHERE room_id = %s AND supervisor_id = %s
                """, (room_id, requester_id))
                joined = bool(cursor.fetchone()) or bool(is_creator)
                if not joined:
                    return jsonify({"error": "Join the room first"}), 403
            else:
                joined = True

            cursor.execute("""
                SELECT
                    c.id,
                    c.operator_id,
                    op.name AS operator_name,
                    c.phone_number,
                    c.appeal_date,
                    c.score,
                    c.created_at,
                    c.direction_id,
                    d.name AS direction_name,
                    (
                        SELECT COUNT(*)
                        FROM calibration_room_call_evaluations e
                        WHERE e.room_call_id = c.id
                    ) AS evaluated_count,
                    EXISTS(
                        SELECT 1
                        FROM calibration_room_call_evaluations me
                        WHERE me.room_call_id = c.id
                          AND me.evaluator_id = %s
                    ) AS my_evaluated
                FROM calibration_room_calls c
                JOIN users op ON op.id = c.operator_id
                LEFT JOIN directions d ON d.id = c.direction_id
                WHERE c.room_id = %s
                ORDER BY c.created_at DESC, c.id DESC
            """, (requester_id, room_id))
            call_rows = cursor.fetchall()

            selected_call_id = None
            available_call_ids = {int(c[0]) for c in call_rows}
            if requested_call_id is not None:
                if requested_call_id not in available_call_ids:
                    return jsonify({"error": "Calibration call not found in room"}), 404
                selected_call_id = requested_call_id
            elif call_rows:
                selected_call_id = int(call_rows[0][0])

            selected_call_row = None
            if selected_call_id is not None:
                cursor.execute("""
                    SELECT
                        c.id,
                        c.room_id,
                        c.month,
                        c.phone_number,
                        c.appeal_date,
                        c.score,
                        c.comment,
                        c.audio_path,
                        c.scores,
                        c.criterion_comments,
                        c.created_at,
                        c.updated_at,
                        op.id AS operator_id,
                        op.name AS operator_name,
                        d.id AS direction_id,
                        d.name AS direction_name,
                        d.criteria AS direction_criteria,
                        d.has_file_upload AS direction_has_file_upload,
                        c.etalon_scores,
                        c.etalon_criterion_comments,
                        c.etalon_updated_at,
                        eu.id AS etalon_updated_by,
                        eu.name AS etalon_updated_by_name,
                        c.general_comment,
                        c.general_comment_updated_at,
                        gu.id AS general_comment_updated_by_id,
                        gu.name AS general_comment_updated_by_name
                    FROM calibration_room_calls c
                    JOIN users op ON op.id = c.operator_id
                    LEFT JOIN directions d ON d.id = c.direction_id
                    LEFT JOIN users eu ON eu.id = c.etalon_updated_by
                    LEFT JOIN users gu ON gu.id = c.general_comment_updated_by
                    WHERE c.id = %s AND c.room_id = %s
                """, (selected_call_id, room_id))
                selected_call_row = cursor.fetchone()
                if not selected_call_row:
                    return jsonify({"error": "Calibration call not found"}), 404

            my_eval_row = None
            if selected_call_id is not None:
                cursor.execute("""
                    SELECT id, score, comment, scores, criterion_comments, created_at, updated_at
                    FROM calibration_room_call_evaluations
                    WHERE room_call_id = %s AND evaluator_id = %s
                """, (selected_call_id, requester_id))
                my_eval_row = cursor.fetchone()

            can_view_results = _is_admin_role(role) or bool(my_eval_row)
            evaluation_rows = []
            evaluator_columns = []
            if selected_call_id is not None and can_view_results:
                cursor.execute("""
                    SELECT
                        e.id,
                        e.evaluator_id,
                        u.name AS evaluator_name,
                        e.score,
                        e.comment,
                        e.scores,
                        e.criterion_comments,
                        e.created_at,
                        e.updated_at
                    FROM calibration_room_call_evaluations e
                    JOIN users u ON u.id = e.evaluator_id
                    WHERE e.room_call_id = %s
                    ORDER BY u.name
                """, (selected_call_id,))
                fetched = cursor.fetchall()
                for e in fetched:
                    payload = {
                        "id": e[0],
                        "evaluator_id": e[1],
                        "evaluator_name": e[2],
                        "score": float(e[3]) if e[3] is not None else 0.0,
                        "comment": e[4],
                        "scores": e[5] if isinstance(e[5], list) else [],
                        "criterion_comments": e[6] if isinstance(e[6], list) else [],
                        "created_at": e[7].strftime('%Y-%m-%d %H:%M:%S') if e[7] and hasattr(e[7], "strftime") else None,
                        "updated_at": e[8].strftime('%Y-%m-%d %H:%M:%S') if e[8] and hasattr(e[8], "strftime") else None,
                    }
                    evaluation_rows.append(payload)
                    evaluator_columns.append({"id": payload["evaluator_id"], "name": payload["evaluator_name"]})

        room_payload = {
            "id": room_row[0],
            "month": room_row[1],
            "room_title": room_row[2] or f"Комната #{room_row[0]}",
            "created_at": room_row[3].strftime('%Y-%m-%d %H:%M:%S') if room_row[3] and hasattr(room_row[3], "strftime") else None,
            "benchmark_admin": {"id": room_row[4], "name": room_row[5]}
        }

        calls_payload = []
        for call_row in call_rows:
            calls_payload.append({
                "id": call_row[0],
                "operator": {"id": call_row[1], "name": call_row[2]},
                "phone_number": call_row[3],
                "appeal_date": call_row[4].strftime('%Y-%m-%d %H:%M:%S') if call_row[4] and hasattr(call_row[4], "strftime") else None,
                "score": float(call_row[5]) if call_row[5] is not None else 0.0,
                "created_at": call_row[6].strftime('%Y-%m-%d %H:%M:%S') if call_row[6] and hasattr(call_row[6], "strftime") else None,
                "direction": {"id": call_row[7], "name": call_row[8]},
                "evaluated_count": int(call_row[9] or 0),
                "my_evaluated": bool(call_row[10]),
            })

        selected_call_payload = None
        results = None
        if selected_call_row:
            direction_criteria = selected_call_row[16] if isinstance(selected_call_row[16], list) else []
            admin_scores = selected_call_row[8] if isinstance(selected_call_row[8], list) else []
            admin_comments = selected_call_row[9] if isinstance(selected_call_row[9], list) else []
            etalon_scores_raw = selected_call_row[18] if isinstance(selected_call_row[18], list) else []
            etalon_comments_raw = selected_call_row[19] if isinstance(selected_call_row[19], list) else []
            etalon_scores = etalon_scores_raw if etalon_scores_raw else admin_scores
            etalon_comments = etalon_comments_raw if etalon_comments_raw else admin_comments
            etalon_total, _ = _compute_total_score_from_criteria(direction_criteria, etalon_scores)
            selected_call_payload = {
                "id": selected_call_row[0],
                "room_id": selected_call_row[1],
                "month": selected_call_row[2],
                "phone_number": selected_call_row[3],
                "appeal_date": selected_call_row[4].strftime('%Y-%m-%d %H:%M:%S') if selected_call_row[4] and hasattr(selected_call_row[4], "strftime") else None,
                "score": float(selected_call_row[5]) if selected_call_row[5] is not None else 0.0,
                "comment": selected_call_row[6],
                "audio_url": _build_signed_audio_url(selected_call_row[7]),
                "scores": admin_scores,
                "criterion_comments": admin_comments,
                "admin_scores": admin_scores,
                "admin_criterion_comments": admin_comments,
                "etalon_scores": etalon_scores,
                "etalon_criterion_comments": etalon_comments,
                "etalon_score": etalon_total,
                "etalon_updated_at": selected_call_row[20].strftime('%Y-%m-%d %H:%M:%S') if selected_call_row[20] and hasattr(selected_call_row[20], "strftime") else None,
                "etalon_updated_by": {
                    "id": selected_call_row[21],
                    "name": selected_call_row[22]
                } if selected_call_row[21] else None,
                "general_comment": selected_call_row[23] or '',
                "general_comment_updated_at": selected_call_row[24].strftime('%Y-%m-%d %H:%M:%S') if selected_call_row[24] and hasattr(selected_call_row[24], "strftime") else None,
                "general_comment_updated_by": {
                    "id": selected_call_row[25],
                    "name": selected_call_row[26]
                } if selected_call_row[25] else None,
                "created_at": selected_call_row[10].strftime('%Y-%m-%d %H:%M:%S') if selected_call_row[10] and hasattr(selected_call_row[10], "strftime") else None,
                "updated_at": selected_call_row[11].strftime('%Y-%m-%d %H:%M:%S') if selected_call_row[11] and hasattr(selected_call_row[11], "strftime") else None,
                "operator": {"id": selected_call_row[12], "name": selected_call_row[13]},
                "direction": {
                    "id": selected_call_row[14],
                    "name": selected_call_row[15],
                    "criteria": direction_criteria,
                    "hasFileUpload": bool(selected_call_row[17]) if selected_call_row[17] is not None else True
                }
            }
            if can_view_results:
                results = _build_calibration_results(
                    direction_criteria,
                    etalon_scores,
                    etalon_comments,
                    evaluation_rows,
                    admin_scores=admin_scores,
                    admin_comments=admin_comments
                )

        my_evaluation = None
        if my_eval_row:
            my_evaluation = {
                "id": my_eval_row[0],
                "score": float(my_eval_row[1]) if my_eval_row[1] is not None else 0.0,
                "comment": my_eval_row[2],
                "scores": my_eval_row[3] if isinstance(my_eval_row[3], list) else [],
                "criterion_comments": my_eval_row[4] if isinstance(my_eval_row[4], list) else [],
                "created_at": my_eval_row[5].strftime('%Y-%m-%d %H:%M:%S') if my_eval_row[5] and hasattr(my_eval_row[5], "strftime") else None,
                "updated_at": my_eval_row[6].strftime('%Y-%m-%d %H:%M:%S') if my_eval_row[6] and hasattr(my_eval_row[6], "strftime") else None,
            }

        return jsonify({
            "status": "success",
            "room": room_payload,
            "calls": calls_payload,
            "selected_call_id": selected_call_id,
            "selected_call": selected_call_payload,
            "joined": joined,
            "can_evaluate": bool(_is_supervisor_role(role) and selected_call_id is not None and not my_eval_row),
            "can_view_results": bool(can_view_results and selected_call_id is not None),
            "my_evaluation": my_evaluation,
            "results": results,
            "evaluators": evaluator_columns
        }), 200
    except Exception as e:
        logging.exception("Error getting calibration room")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>/export_excel', methods=['GET'])
@require_api_key
def export_call_calibration_room_excel(room_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        role = _normalize_user_role(requester[3])
        if not _is_admin_role(role) and not _is_supervisor_role(role):
            return jsonify({"error": "Forbidden: only admin and supervisors can export calibration room"}), 403

        def _fmt_dt(value):
            if value is None:
                return '—'
            if hasattr(value, 'strftime'):
                return value.strftime('%Y-%m-%d %H:%M:%S')
            return str(value)

        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    r.id,
                    r.month,
                    r.room_title,
                    r.created_at,
                    adm.id AS creator_id,
                    adm.name AS creator_name
                FROM calibration_rooms r
                JOIN users adm ON adm.id = r.created_by_admin_id
                WHERE r.id = %s
            """, (room_id,))
            room_row = cursor.fetchone()
            if not room_row:
                return jsonify({"error": "Calibration room not found"}), 404

            if _is_supervisor_role(role):
                is_creator = int(room_row[4] or 0) == int(requester_id)
                cursor.execute("""
                    SELECT id
                    FROM calibration_room_members
                    WHERE room_id = %s AND supervisor_id = %s
                """, (room_id, requester_id))
                joined = bool(cursor.fetchone()) or bool(is_creator)
                if not joined:
                    return jsonify({"error": "Join the room first"}), 403

            cursor.execute("""
                SELECT
                    c.id,
                    c.operator_id,
                    op.name AS operator_name,
                    c.phone_number,
                    c.appeal_date,
                    c.score,
                    d.name AS direction_name,
                    c.scores,
                    c.criterion_comments,
                    c.etalon_scores,
                    c.etalon_criterion_comments,
                    c.created_at,
                    c.updated_at,
                    d.criteria
                FROM calibration_room_calls c
                JOIN users op ON op.id = c.operator_id
                LEFT JOIN directions d ON d.id = c.direction_id
                WHERE c.room_id = %s
                ORDER BY c.created_at DESC, c.id DESC
            """, (room_id,))
            call_rows = cursor.fetchall()

            call_ids = [int(row[0]) for row in call_rows]
            evaluations_by_call = {}
            if call_ids:
                cursor.execute("""
                    SELECT
                        e.room_call_id,
                        e.evaluator_id,
                        u.name AS evaluator_name,
                        e.score,
                        e.comment,
                        e.scores,
                        e.criterion_comments
                    FROM calibration_room_call_evaluations e
                    JOIN users u ON u.id = e.evaluator_id
                    WHERE e.room_call_id = ANY(%s)
                    ORDER BY e.room_call_id, u.name
                """, (call_ids,))
                for ev in cursor.fetchall():
                    room_call_id = int(ev[0])
                    evaluations_by_call.setdefault(room_call_id, []).append({
                        "id": None,
                        "evaluator_id": ev[1],
                        "evaluator_name": ev[2],
                        "score": float(ev[3]) if ev[3] is not None else 0.0,
                        "comment": ev[4],
                        "scores": ev[5] if isinstance(ev[5], list) else [],
                        "criterion_comments": ev[6] if isinstance(ev[6], list) else [],
                    })

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Сводка'
        ws_details = wb.create_sheet('Критерии')

        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        wrap_alignment = Alignment(vertical='top', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        ws_summary.append(['Параметр', 'Значение'])
        ws_summary.append(['Комната ID', room_row[0]])
        ws_summary.append(['Название комнаты', room_row[2] or f'Комната #{room_row[0]}'])
        ws_summary.append(['Месяц', room_row[1] or '—'])
        ws_summary.append(['Создал', room_row[5] or '—'])
        ws_summary.append(['Создано', _fmt_dt(room_row[3])])
        ws_summary.append(['Количество звонков', len(call_rows)])
        ws_summary.append([])

        summary_headers = [
            'Звонок ID',
            'Оператор',
            'Телефон',
            'Дата обращения',
            'Направление',
            f'Оценка автора ({room_row[5] or "—"})',
            'Текущий эталон',
            'Оценивших',
            'Общий % калибровки',
            'Критич. расхождение'
        ]
        ws_summary.append(summary_headers)
        summary_header_row = ws_summary.max_row

        details_headers = [
            'Звонок ID',
            'Оператор',
            'Критерий',
            'Критичный',
            '% калибровки',
            'Эталон',
            'Комментарий эталона',
            f'Оценка автора ({room_row[5] or "—"})',
            'Комментарий автора',
            'Оценки супервайзеров'
        ]
        ws_details.append(details_headers)
        details_header_row = ws_details.max_row

        for call_row in call_rows:
            call_id = int(call_row[0])
            operator_name = call_row[2] or '—'
            phone_number = call_row[3] or '—'
            appeal_date = _fmt_dt(call_row[4])
            admin_total = float(call_row[5]) if call_row[5] is not None else 0.0
            direction_name = call_row[6] or '—'
            admin_scores = call_row[7] if isinstance(call_row[7], list) else []
            admin_comments = call_row[8] if isinstance(call_row[8], list) else []
            etalon_scores_raw = call_row[9] if isinstance(call_row[9], list) else []
            etalon_comments_raw = call_row[10] if isinstance(call_row[10], list) else []
            direction_criteria = call_row[13] if isinstance(call_row[13], list) else []

            etalon_scores = etalon_scores_raw if etalon_scores_raw else admin_scores
            etalon_comments = etalon_comments_raw if etalon_comments_raw else admin_comments
            etalon_total, _ = _compute_total_score_from_criteria(direction_criteria, etalon_scores)

            evaluations = evaluations_by_call.get(call_id, [])
            eval_name_by_id = {int(ev.get('evaluator_id')): ev.get('evaluator_name') or 'Супервайзер' for ev in evaluations}
            results = _build_calibration_results(
                direction_criteria,
                etalon_scores,
                etalon_comments,
                evaluations,
                admin_scores=admin_scores,
                admin_comments=admin_comments
            )
            overall_percent = results.get('overall_percent')
            overall_percent_text = '—' if overall_percent is None else f"{overall_percent:.1f}%"
            critical_mismatch = bool(results.get('critical_mismatch'))

            ws_summary.append([
                call_id,
                operator_name,
                phone_number,
                appeal_date,
                direction_name,
                round(admin_total, 2),
                round(float(etalon_total or 0.0), 2),
                int(results.get('evaluated_count') or 0),
                overall_percent_text,
                'Да' if critical_mismatch else 'Нет'
            ])

            criteria_rows = results.get('criteria_rows') or []
            if not criteria_rows:
                ws_details.append([
                    call_id,
                    operator_name,
                    'Нет критериев',
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    ''
                ])
            else:
                for criterion_row in criteria_rows:
                    sv_values = []
                    for cell in criterion_row.get('by_evaluator') or []:
                        evaluator_id = int(cell.get('evaluator_id') or 0)
                        evaluator_name = eval_name_by_id.get(evaluator_id) or f"SV #{evaluator_id}"
                        item = f"{evaluator_name}: {cell.get('score_label') or _calibration_label_score(cell.get('score'))}"
                        if cell.get('is_match') is not None:
                            item += ' (совпало)' if bool(cell.get('is_match')) else ' (не совпало)'
                        if cell.get('comment'):
                            item += f" | {cell.get('comment')}"
                        sv_values.append(item)

                    percent = criterion_row.get('percent')
                    percent_text = '—' if percent is None else f"{float(percent):.1f}%"
                    ws_details.append([
                        call_id,
                        operator_name,
                        criterion_row.get('criterion_name') or '',
                        'Да' if criterion_row.get('is_critical') else 'Нет',
                        percent_text,
                        criterion_row.get('etalon', {}).get('score_label') or _calibration_label_score(criterion_row.get('etalon', {}).get('score')),
                        criterion_row.get('etalon', {}).get('comment') or '',
                        criterion_row.get('admin', {}).get('score_label') or _calibration_label_score(criterion_row.get('admin', {}).get('score')),
                        criterion_row.get('admin', {}).get('comment') or '',
                        '\n'.join(sv_values) if sv_values else '—'
                    ])

            ws_details.append([
                call_id,
                operator_name,
                'ИТОГ ПО ЗВОНКУ',
                '',
                '',
                '',
                '',
                '',
                '',
                f"Общий %: {overall_percent_text}; Критич. расхождение: {'Да' if critical_mismatch else 'Нет'}"
            ])
            ws_details.append([])

        for col in range(1, len(summary_headers) + 1):
            cell = ws_summary.cell(row=summary_header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for col in range(1, len(details_headers) + 1):
            cell = ws_details.cell(row=details_header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        for row in ws_summary.iter_rows(min_row=summary_header_row + 1, max_row=ws_summary.max_row, min_col=1, max_col=len(summary_headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = wrap_alignment

        for row in ws_details.iter_rows(min_row=details_header_row + 1, max_row=ws_details.max_row, min_col=1, max_col=len(details_headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = wrap_alignment

        summary_widths = [12, 22, 18, 20, 20, 19, 14, 11, 19, 18]
        for i, width in enumerate(summary_widths, start=1):
            ws_summary.column_dimensions[get_column_letter(i)].width = width

        details_widths = [12, 22, 36, 10, 15, 15, 28, 19, 28, 54]
        for i, width in enumerate(details_widths, start=1):
            ws_details.column_dimensions[get_column_letter(i)].width = width

        ws_summary.freeze_panes = 'A10'
        ws_details.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_month = re.sub(r'[^0-9-]', '', str(room_row[1] or ''))
        filename = f"calibration_room_{room_id}_{safe_month or 'export'}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logging.exception("Error exporting calibration room")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>/calls/<int:call_id>/etalon', methods=['PATCH'])
@require_api_key
def update_call_calibration_etalon(room_id, call_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        requester_role = _normalize_user_role(requester[3])
        if not _is_admin_role(requester_role) and not _is_supervisor_role(requester_role):
            return jsonify({"error": "Only admin and supervisors can update etalon"}), 403

        data = request.get_json() or {}
        scores = data.get('scores')
        criterion_comments = data.get('criterion_comments')
        if not isinstance(scores, list) or not isinstance(criterion_comments, list):
            return jsonify({"error": "scores and criterion_comments must be arrays"}), 400

        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT d.criteria, c.operator_id, r.created_by_admin_id
                FROM calibration_room_calls c
                JOIN calibration_rooms r ON r.id = c.room_id
                LEFT JOIN directions d ON d.id = c.direction_id
                WHERE c.id = %s AND c.room_id = %s
            """, (call_id, room_id))
            call_row = cursor.fetchone()
            if not call_row:
                return jsonify({"error": "Calibration call not found"}), 404
            criteria = call_row[0] if isinstance(call_row[0], list) else []
            call_operator_id = int(call_row[1]) if call_row[1] is not None else None
            room_creator_id = int(call_row[2]) if call_row[2] is not None else None

            if _is_supervisor_role(requester_role):
                if call_operator_id is None or not _authorize_operator_scope(requester, requester_id, call_operator_id):
                    return jsonify({"error": "Forbidden: you cannot update etalon for this operator"}), 403
                if int(requester_id) != room_creator_id:
                    cursor.execute("""
                        SELECT id
                        FROM calibration_room_members
                        WHERE room_id = %s AND supervisor_id = %s
                    """, (room_id, requester_id))
                    if not cursor.fetchone():
                        return jsonify({"error": "Join the room first"}), 403

            normalized_scores = []
            normalized_comments = []
            for idx in range(len(criteria)):
                src_score = scores[idx] if idx < len(scores) else 'Correct'
                src_comment = criterion_comments[idx] if idx < len(criterion_comments) else ''
                normalized_scores.append(_normalize_calibration_score_value(src_score))
                normalized_comments.append(str(src_comment or ''))

            etalon_total, _ = _compute_total_score_from_criteria(criteria, normalized_scores)

            cursor.execute("""
                UPDATE calibration_room_calls
                SET
                    etalon_scores = %s::jsonb,
                    etalon_criterion_comments = %s::jsonb,
                    etalon_updated_by = %s,
                    etalon_updated_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND room_id = %s
                RETURNING id
            """, (
                json.dumps(normalized_scores, ensure_ascii=False),
                json.dumps(normalized_comments, ensure_ascii=False),
                requester_id,
                call_id,
                room_id
            ))
            updated = cursor.fetchone()
            if not updated:
                return jsonify({"error": "Calibration call not found"}), 404

        return jsonify({
            "status": "success",
            "call_id": call_id,
            "etalon_score": etalon_total
        }), 200
    except Exception as e:
        logging.exception("Error updating calibration etalon")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>/calls/<int:call_id>/general_comment', methods=['PATCH'])
@require_api_key
def update_call_calibration_general_comment(room_id, call_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        requester_role = _normalize_user_role(requester[3])
        if not _is_admin_role(requester_role) and not _is_supervisor_role(requester_role):
            return jsonify({"error": "Only admin and supervisors can update general comment"}), 403

        data = request.get_json() or {}
        general_comment = str(data.get('general_comment') or '').strip()

        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT c.operator_id, r.created_by_admin_id
                FROM calibration_room_calls c
                JOIN calibration_rooms r ON r.id = c.room_id
                WHERE c.id = %s AND c.room_id = %s
            """, (call_id, room_id))
            call_row = cursor.fetchone()
            if not call_row:
                return jsonify({"error": "Calibration call not found"}), 404

            if _is_supervisor_role(requester_role):
                call_operator_id = int(call_row[0]) if call_row[0] is not None else None
                room_creator_id = int(call_row[1]) if call_row[1] is not None else None
                if call_operator_id is None or not _authorize_operator_scope(requester, requester_id, call_operator_id):
                    return jsonify({"error": "Forbidden: you cannot update comment for this operator"}), 403
                if int(requester_id) != room_creator_id:
                    cursor.execute("""
                        SELECT id FROM calibration_room_members
                        WHERE room_id = %s AND supervisor_id = %s
                    """, (room_id, requester_id))
                    if not cursor.fetchone():
                        return jsonify({"error": "Join the room first"}), 403

            cursor.execute("""
                UPDATE calibration_room_calls
                SET
                    general_comment = %s,
                    general_comment_updated_by = %s,
                    general_comment_updated_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s AND room_id = %s
                RETURNING id
            """, (general_comment or None, requester_id, call_id, room_id))
            if not cursor.fetchone():
                return jsonify({"error": "Calibration call not found"}), 404

        return jsonify({"status": "success", "call_id": call_id}), 200
    except Exception as e:
        logging.exception("Error updating calibration general comment")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>/evaluate', methods=['POST'])
@require_api_key
def evaluate_call_calibration_room(room_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        if not _is_supervisor_role(requester[3]):
            return jsonify({"error": "Only supervisors can submit calibration results"}), 403

        data = request.get_json() or {}
        call_id_raw = data.get('call_id')
        try:
            call_id = int(call_id_raw)
        except Exception:
            return jsonify({"error": "call_id is required and must be integer"}), 400

        scores = data.get('scores')
        criterion_comments = data.get('criterion_comments')
        comment = str(data.get('comment') or '').strip()
        if not isinstance(scores, list) or not isinstance(criterion_comments, list):
            return jsonify({"error": "scores and criterion_comments must be arrays"}), 400

        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT created_by_admin_id
                FROM calibration_rooms
                WHERE id = %s
            """, (room_id,))
            room_row = cursor.fetchone()
            if not room_row:
                return jsonify({"error": "Calibration room not found"}), 404

            room_creator_id = int(room_row[0]) if room_row[0] is not None else None
            if int(requester_id) != int(room_creator_id or 0):
                cursor.execute("""
                    SELECT id
                    FROM calibration_room_members
                    WHERE room_id = %s AND supervisor_id = %s
                """, (room_id, requester_id))
                if not cursor.fetchone():
                    return jsonify({"error": "Join the room first"}), 403

            cursor.execute("""
                SELECT c.operator_id, d.criteria
                FROM calibration_room_calls c
                LEFT JOIN directions d ON d.id = c.direction_id
                WHERE c.id = %s AND c.room_id = %s
            """, (call_id, room_id))
            call_row = cursor.fetchone()
            if not call_row:
                return jsonify({"error": "Calibration call not found"}), 404
            call_operator_id = int(call_row[0]) if call_row[0] is not None else None
            criteria = call_row[1] if isinstance(call_row[1], list) else []
            if call_operator_id is None or not _authorize_operator_scope(requester, requester_id, call_operator_id):
                return jsonify({"error": "Forbidden: you cannot evaluate this operator"}), 403

            normalized_scores = []
            normalized_comments = []
            for idx in range(len(criteria)):
                src_score = scores[idx] if idx < len(scores) else 'Correct'
                src_comment = criterion_comments[idx] if idx < len(criterion_comments) else ''
                normalized_scores.append(_normalize_calibration_score_value(src_score))
                normalized_comments.append(str(src_comment or ''))

            total_score, _ = _compute_total_score_from_criteria(criteria, normalized_scores)
            if not comment:
                comment = '; '.join(
                    f"{(criteria[i] or {}).get('name', f'Критерий {i + 1}')}: {text.strip()}"
                    for i, text in enumerate(normalized_comments)
                    if str(text or '').strip()
                )

            cursor.execute("""
                INSERT INTO calibration_room_call_evaluations (
                    room_call_id,
                    evaluator_id,
                    score,
                    comment,
                    scores,
                    criterion_comments,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, %s::jsonb, %s::jsonb, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT (room_call_id, evaluator_id)
                DO UPDATE SET
                    score = EXCLUDED.score,
                    comment = EXCLUDED.comment,
                    scores = EXCLUDED.scores,
                    criterion_comments = EXCLUDED.criterion_comments,
                    updated_at = CURRENT_TIMESTAMP
                RETURNING id
            """, (
                call_id,
                requester_id,
                total_score,
                comment,
                json.dumps(normalized_scores, ensure_ascii=False),
                json.dumps(normalized_comments, ensure_ascii=False),
            ))
            evaluation_id = cursor.fetchone()[0]

        return jsonify({
            "status": "success",
            "evaluation_id": evaluation_id,
            "call_id": call_id,
            "score": total_score
        }), 200
    except Exception as e:
        logging.exception("Error submitting calibration evaluation")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/call_calibration/rooms/<int:room_id>/history', methods=['GET', 'OPTIONS'])
@require_api_key
def get_calibration_room_history(room_id):
    if request.method == 'OPTIONS':
        return jsonify({}), 200
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404
        role = _normalize_user_role(requester[3])
        if not _is_admin_role(role) and not _is_supervisor_role(role):
            return jsonify({"error": "Forbidden"}), 403

        events = []

        with db._get_cursor() as cursor:
            # Verify room exists and requester has access
            cursor.execute("""
                SELECT r.id, r.created_at, r.room_title, u.name AS creator_name
                FROM calibration_rooms r
                JOIN users u ON u.id = r.created_by_admin_id
                WHERE r.id = %s
            """, (room_id,))
            room_row = cursor.fetchone()
            if not room_row:
                return jsonify({"error": "Room not found"}), 404

            # Room creation
            events.append({
                "type": "room_created",
                "timestamp": room_row[1].strftime('%Y-%m-%d %H:%M:%S') if room_row[1] else None,
                "actor": room_row[3],
                "description": f"Комната создана: «{room_row[2] or f'Комната #{room_row[0]}'}»"
            })

            # Member joins
            cursor.execute("""
                SELECT m.joined_at, u.name
                FROM calibration_room_members m
                JOIN users u ON u.id = m.supervisor_id
                WHERE m.room_id = %s
                ORDER BY m.joined_at
            """, (room_id,))
            for row in cursor.fetchall():
                events.append({
                    "type": "member_joined",
                    "timestamp": row[0].strftime('%Y-%m-%d %H:%M:%S') if row[0] else None,
                    "actor": row[1],
                    "description": f"Вошёл в комнату"
                })

            # Calls added
            cursor.execute("""
                SELECT c.id, c.created_at, adm.name AS adder_name, op.name AS operator_name, c.phone_number
                FROM calibration_room_calls c
                JOIN users adm ON adm.id = c.created_by_admin_id
                JOIN users op ON op.id = c.operator_id
                WHERE c.room_id = %s
                ORDER BY c.created_at
            """, (room_id,))
            for row in cursor.fetchall():
                events.append({
                    "type": "call_added",
                    "timestamp": row[1].strftime('%Y-%m-%d %H:%M:%S') if row[1] else None,
                    "actor": row[2],
                    "description": f"Добавлен звонок #{row[0]}: оператор {row[3]}, тел. {row[4]}"
                })

            # Etalon updates
            cursor.execute("""
                SELECT c.id, c.etalon_updated_at, u.name, op.name AS operator_name
                FROM calibration_room_calls c
                JOIN users u ON u.id = c.etalon_updated_by
                JOIN users op ON op.id = c.operator_id
                WHERE c.room_id = %s AND c.etalon_updated_at IS NOT NULL
                ORDER BY c.etalon_updated_at
            """, (room_id,))
            for row in cursor.fetchall():
                events.append({
                    "type": "etalon_updated",
                    "timestamp": row[1].strftime('%Y-%m-%d %H:%M:%S') if row[1] else None,
                    "actor": row[2],
                    "description": f"Обновлён эталон для звонка #{row[0]} (оператор {row[3]})"
                })

            # Evaluations submitted / updated
            cursor.execute("""
                SELECT e.id, e.created_at, e.updated_at, u.name AS evaluator_name,
                       c.id AS call_id, op.name AS operator_name
                FROM calibration_room_call_evaluations e
                JOIN users u ON u.id = e.evaluator_id
                JOIN calibration_room_calls c ON c.id = e.room_call_id
                JOIN users op ON op.id = c.operator_id
                WHERE c.room_id = %s
                ORDER BY e.created_at
            """, (room_id,))
            for row in cursor.fetchall():
                events.append({
                    "type": "evaluated",
                    "timestamp": row[1].strftime('%Y-%m-%d %H:%M:%S') if row[1] else None,
                    "actor": row[3],
                    "description": f"Оценил звонок #{row[4]} (оператор {row[5]})"
                })
                created_ts = row[1].strftime('%Y-%m-%d %H:%M:%S') if row[1] else None
                updated_ts = row[2].strftime('%Y-%m-%d %H:%M:%S') if row[2] else None
                if updated_ts and updated_ts != created_ts:
                    events.append({
                        "type": "evaluation_updated",
                        "timestamp": updated_ts,
                        "actor": row[3],
                        "description": f"Изменил оценку звонка #{row[4]} (оператор {row[5]})"
                    })

        events.sort(key=lambda x: x["timestamp"] or "")
        return jsonify({"status": "success", "events": events}), 200
    except Exception as e:
        logging.exception("Error getting calibration room history")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/birthdays/today', methods=['GET'])
@require_api_key
def birthdays_today():
    try:
        today = datetime.now().date()
        birthdays = db.get_birthdays_for_date(today)
        return jsonify({"status": "success", "date": today.isoformat(), "birthdays": birthdays}), 200
    except Exception as e:
        logging.exception("Error in /api/birthdays/today")
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/ai/birthday_greeting', methods=['POST'])
@require_api_key
def ai_birthday_greeting():
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"status": "error", "error": "Missing X-User-Id header"}), 400

        try:
            requester_id = int(requester_id)
        except (ValueError, TypeError):
            return jsonify({"status": "error", "error": "Invalid X-User-Id format"}), 400

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"status": "error", "error": "User not found"}), 404

        status = str(requester[11] or '').strip().lower()
        if status in ('fired', 'dismissal'):
            return jsonify({"status": "error", "error": "User is not active"}), 403

        birth_date = requester[14]
        if not birth_date:
            return jsonify({"status": "error", "error": "Birth date not set"}), 404

        today = datetime.now().date()
        if birth_date.month != today.month or birth_date.day != today.day:
            return jsonify({"status": "error", "error": "Not birthday"}), 409

        user_payload = {
            "id": requester_id,
            "name": requester[2],
            "role": requester[3],
            "direction": requester[4],
            "gender": requester[13],
            "hire_date": requester[5].strftime('%Y-%m-%d') if requester[5] else None
        }

        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            loop = None

        if loop and loop.is_running():
            fut = asyncio.run_coroutine_threadsafe(
                generate_birthday_greeting_with_ai(user_payload=user_payload, for_date=today.isoformat()),
                loop,
            )
            result = fut.result(timeout=120)
        else:
            result = asyncio.run(generate_birthday_greeting_with_ai(user_payload=user_payload, for_date=today.isoformat()))

        if result is None:
            return jsonify({"status": "error", "error": "ai_failed"}), 500

        if isinstance(result, dict) and result.get('error'):
            return jsonify({"status": "error", "error": result}), 200

        return jsonify({"status": "success", "result": result}), 200

    except Exception as e:
        logging.exception("Error in /api/ai/birthday_greeting")
        return jsonify({"status": "error", "error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/ai/monthly_feedback', methods=['POST'])
@require_api_key
def ai_monthly_feedback():
    try:
        # Проверка прав - admin, sv и operator (operator только для чтения из кэша)
        requester_id = request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"status": "error", "error": "Missing X-User-Id header"}), 400
        
        try:
            requester_id = int(requester_id)
        except (ValueError, TypeError):
            return jsonify({"status": "error", "error": "Invalid X-User-Id format"}), 400
        
        requester = db.get_user(id=requester_id)
        requester_role = _normalize_user_role(requester[3]) if requester else ''
        if not requester or (not _is_admin_role(requester_role) and requester_role not in ('sv', 'operator')):
            return jsonify({"status": "error", "error": "Access denied"}), 403

        is_operator = requester_role == 'operator'

        data = request.get_json() or {}
        operator_id = data.get('operator_id')
        month = data.get('month')

        if operator_id is None or month is None:
            return jsonify({"status": "error", "error": "Missing operator_id or month"}), 400

        try:
            operator_id = int(operator_id)
        except Exception:
            return jsonify({"status": "error", "error": "Invalid operator_id"}), 400

        try:
            datetime.strptime(str(month), "%Y-%m")
        except Exception:
            return jsonify({"status": "error", "error": "Invalid month format. Use YYYY-MM"}), 400

        # Для операторов проверяем только кэш
        if is_operator:
            cached_feedback = db.get_ai_feedback_cache(requester_id, str(month))
            if cached_feedback:
                return jsonify({"status": "success", "result": cached_feedback['feedback_data']}), 200
            else:
                return jsonify({"status": "error", "error": "No cached feedback available"}), 404

        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            loop = None

        if loop and loop.is_running():
            fut = asyncio.run_coroutine_threadsafe(
                generate_monthly_feedback_with_ai(operator_id=operator_id, month=str(month)),
                loop,
            )
            result = fut.result(timeout=180)
        else:
            result = asyncio.run(generate_monthly_feedback_with_ai(operator_id=operator_id, month=str(month)))

        if result is None:
            return jsonify({"status": "error", "error": "ai_failed"}), 500

        if isinstance(result, dict) and result.get('error'):
            return jsonify({"status": "error", "error": result}), 200

        return jsonify({"status": "success", "result": result}), 200

    except Exception as e:
        logging.exception("Error in /api/ai/monthly_feedback")
        return jsonify({"status": "error", "error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/operators_summary', methods=['GET'])
@require_api_key
def operators_summary():
    try:
        # month required (YYYY-MM)
        month = request.args.get('month')
        if not month:
            return jsonify({"error": "Missing month parameter (YYYY-MM)"}), 400
        from datetime import datetime
        try:
            datetime.strptime(month, "%Y-%m")
        except ValueError:
            return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400

        # requester identity + role check
        user_id_header = request.headers.get('X-User-Id') or request.args.get('requester_id')
        if not user_id_header:
            return jsonify({"error": "Missing X-User-Id header"}), 400
        try:
            requester_id = int(user_id_header)
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid X-User-Id format"}), 400

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404

        # requester role: support tuple/dict user shapes (following your existing pattern)
        role = None
        if isinstance(requester, dict):
            role = requester.get('role')
        else:
            # assume role in pos 3 like in other parts of code
            role = requester[3] if len(requester) > 3 else None
        role = _normalize_user_role(role)

        if not (_is_admin_role(role) or _is_supervisor_role(role)):
            return jsonify({"error": "Forbidden: only admins or supervisors can access"}), 403

        # optional supervisor_id param: admin can request any sv; sv will be forced to their id
        sv_param = request.args.get('supervisor_id')
        supervisor_id = None
        if _is_supervisor_role(role):
            # if requester is SV — restrict to their operators only
            supervisor_id = requester_id
        else:
            # admin: allow optional supervisor filter
            if sv_param:
                try:
                    supervisor_id = int(sv_param)
                except ValueError:
                    return jsonify({"error": "Invalid supervisor_id"}), 400

        operators = db.get_operators_summary_for_month(month=month, supervisor_id=supervisor_id)

        return jsonify({"status": "success", "month": month, "operators": operators}), 200

    except Exception as e:
        logging.exception("Error in /api/admin/operators_summary")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/change_sv_table', methods=['POST'])
@require_api_key
def change_sv_table():
    try:
        data = request.get_json()
        required_fields = ['id', 'table_url']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        sv_id = int(data['id'])
        table_url = data['table_url']
        
        sheet_name, operators, error = extract_fio_and_links(table_url)
        if error:
            return jsonify({"error": error}), 400
        
        db.update_user_table(sv_id, table_url)
        
        telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
        payload = {
            "chat_id": db.get_user(id=sv_id)[1],
            "text": f"Таблица успешно обновлена администратором <b>успешно✅</b>",
            "parse_mode": "HTML"
        }
        response = requests.post(telegram_url, json=payload, timeout=10)
        if response.status_code != 200:
            error_detail = response.json().get('description', 'Unknown error')
            logging.error(f"Telegram API error: {error_detail}")
        
        return jsonify({
            "status": "success",
            "message": "Table updated",
            "operators": operators
        })
    except Exception as e:
        logging.error(f"Error updating SV table: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/call_evaluation/dispute', methods=['POST'])
@require_api_key
def dispute_call_evaluation():
    try:
        data = request.get_json()
        required_fields = ['operator_id', 'id', 'month', 'dispute_text']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        operator_id = int(data['operator_id'])
        operator = db.get_user(id=operator_id)
        if not operator:
            return jsonify({"error": "Operator not found"}), 404

        supervisor = db.get_user(id=operator[6]) if operator[6] else None
        if not supervisor:
            return jsonify({"error": "Supervisor not found for this operator"}), 404

        # Get call details
        evaluations = db.get_call_evaluations(operator_id, data['month'])
        call = next((e for e in evaluations if e['id'] == data['id']), None)
        if not call:
            return jsonify({"error": "Call evaluation not found"}), 404

        # --- NEW: Try to get audio signed URL ---
        audio_url = None
        if call.get('audio_path'):
            try:
                gcs_client = get_gcs_client()
                path_parts = call['audio_path'].split('/', 1)
                if len(path_parts) == 2:
                    bucket_name, blob_path = path_parts
                    bucket = gcs_client.bucket(bucket_name)
                    blob = bucket.blob(blob_path)
                    if blob.exists():
                        audio_url = blob.generate_signed_url(
                            version="v4",
                            expiration=timedelta(minutes=15),
                            method="GET",
                            response_type='audio/mpeg'
                        )
            except Exception as e:
                logging.error(f"Error generating signed audio URL for dispute: {e}")
        # --- END NEW ---

        # Сообщение для супервайзера
        supervisor_message = (
            f"⚠️ <b>Запрос на пересмотр оценки</b>\n\n"
            f"👤 Оператор: <b>{operator[2]}</b>\n"
            f"📞 Звонок ID: {call['id']}\n"
            f"📱 Номер: {call['phone_number']}\n"
            f"📅 Дата обращения: {' '.join(call['appeal_date'].split('T'))}\n"
            f"📅 Дата оценки: {call['created_at']}\n"
            f"💯 Оценка: {call['score']}\n"
            f"📅 За месяц: {call['month']}\n\n"
            f"📝 <b>Сообщение от оператора:</b>\n"
            f"{data['dispute_text']}"
        )

        # Сообщение для админа
        admin_message = (
            f"⚠️ <b>Запрос на пересмотр оценки</b>\n\n"
            f"💬 Супервайзер: <b>{supervisor[2]}</b>\n"
            f"👤 Оператор: <b>{operator[2]}</b>\n"
            f"📞 Звонок ID: {call['id']}\n"
            f"📅 Дата обращения: {' '.join(call['appeal_date'].split('T'))}\n"
            f"📅 Дата оценки: {call['created_at']}\n"
            f"📱 Номер: {call['phone_number']}\n"
            f"💯 Оценка: {call['score']}\n"
            f"📅 За месяц: {call['month']}\n\n"
            f"📝 <b>Сообщение от оператора:</b>\n"
            f"{data['dispute_text']}"
        )

        telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
        telegram_audio_url = f"https://api.telegram.org/bot{API_TOKEN}/sendAudio"

        # --- NEW: Send audio if available ---
        def send_telegram_message(chat_id, text, audio_url=None):
            if audio_url:
                payload = {
                    "chat_id": chat_id,
                    "audio": audio_url,
                    "caption": text,
                    "parse_mode": "HTML"
                }
                resp = requests.post(telegram_audio_url, json=payload, timeout=10)
            else:
                payload = {
                    "chat_id": chat_id,
                    "text": text,
                    "parse_mode": "HTML"
                }
                resp = requests.post(telegram_url, json=payload, timeout=10)
            return resp

        # Отправка супервайзеру
        supervisor_response = send_telegram_message(supervisor[1], supervisor_message, audio_url)
        if supervisor_response.status_code != 200:
            error_detail = supervisor_response.json().get('description', 'Unknown error')
            logging.error(f"Telegram API error (supervisor): {error_detail}")
            return jsonify({"error": f"Failed to send dispute message to supervisor: {error_detail}"}), 500

        # Отправка админу
        admin_response = send_telegram_message(os.getenv("ADMIN_ID"), admin_message, audio_url)
        if admin_response.status_code != 200:
            error_detail = admin_response.json().get('description', 'Unknown error')
            logging.error(f"Telegram API error (admin): {error_detail}")
            return jsonify({"error": f"Failed to send dispute message to admin: {error_detail}"}), 500
        # --- END NEW ---

        return jsonify({"status": "success", "message": "Dispute sent to supervisor and admin"})
    except Exception as e:
        logging.error(f"Error processing dispute: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/hours/send_request', methods=['POST'])
@require_api_key
def send_request():
    try:
        data = request.get_json()
        required_fields = ['date', 'message']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields: date, message"}), 400

        operator_id = int(request.headers.get('X-User-Id'))
        hours = data.get('hours', 0)
        date = data['date']
        message = data['message']

        operator = db.get_user(id=operator_id)
        if not operator:
            return jsonify({"error": "Operator not found"}), 404

        supervisor = db.get_user(id=operator[6]) if operator[6] else None
        if not supervisor:
            return jsonify({"error": "Supervisor not found for this operator"}), 404

        telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
        payload = {
            "chat_id": supervisor[1],
            "text": (
                f"⚠️ <b>Запрос по рабочим часам</b>\n\n"
                f"👤 Оператор: <b>{operator[2]}</b>\n"
                f"⏰ Часы: <b>{hours} ч</b>\n"
                f"📅 Дата: <b>{date}</b>\n"
                f"📝 <b>Сообщение:</b>\n{message}"
            ),
            "parse_mode": "HTML"
        }
        response = requests.post(telegram_url, json=payload, timeout=10)
        if response.status_code != 200:
            error_detail = response.json().get('description', 'Unknown error')
            logging.error(f"Telegram API error: {error_detail}")
            return jsonify({"error": f"Failed to send request: {error_detail}"}), 500

        return jsonify({"status": "success", "message": "Request sent to supervisor"}), 200
    except Exception as e:
        logging.error(f"Error sending request: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/add_sv', methods=['POST'])
@require_api_key
def add_sv():
    try:
        data = request.get_json()
        required_fields = ['name']  # Removed 'telegram_id' from required fields
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required field: name"}), 400

        name = data['name']
        
        login = f"sv_{str(uuid.uuid4())[:8]}"
        password = str(uuid.uuid4())[:8]  
        
        # Create supervisor with login and hashed_password, telegram_id set to None
        sv_id = db.create_user(
            telegram_id=None,  # Explicitly set to None
            name=name,
            role='sv',
            login=login,
            password=password,
            rate=1.0  # Default rate for SV
        )
        
        return jsonify({
            "status": "success",
            "message": f"SV {name} added",
            "id": sv_id,
            "login": login,  # Return plain login
            "password": password  # Return plain password for admin to share
        })
    except Exception as e:
        logging.error(f"Error adding SV: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/add_user', methods=['POST'])
@require_api_key
def add_user():
    try:
        data = request.get_json() or {}
        requester_id_raw = request.headers.get('X-User-Id')
        if requester_id_raw in [None, '']:
            return jsonify({"error": "Missing X-User-Id header"}), 400
        try:
            requester_id = int(requester_id_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "Invalid X-User-Id header"}), 400
        requester = db.get_user(id=requester_id)
        requester_role = _normalize_user_role(requester[3]) if requester else ''
        if requester_role not in ('super_admin', 'admin', 'sv'):
            return jsonify({"error": "Only admins can add users"}), 403

        name = str(data.get('name') or '').strip()
        if not name:
            return jsonify({"error": "Name cannot be empty"}), 400

        role = str(data.get('role') or 'operator').strip().lower()
        if role not in ('operator', 'trainee', 'trainer', 'admin'):
            return jsonify({"error": "Unsupported role. Allowed: operator, trainee, trainer, admin"}), 400
        if role == 'admin' and requester_role != 'super_admin':
            return jsonify({"error": "Only super admins can create admins"}), 403
        if requester_role == 'sv' and role not in ('operator', 'trainee'):
            return jsonify({"error": "Supervisors can create only operators or trainees"}), 403

        supervisor_id = None
        direction_id = None

        hire_date = data.get('hire_date')
        if not hire_date:
            return jsonify({"error": "Missing required field: hire_date"}), 400
        try:
            datetime.strptime(hire_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({"error": "Invalid hire_date format. Use YYYY-MM-DD"}), 400

        if role in ('operator', 'trainee'):
            supervisor_raw = data.get('supervisor_id')
            if supervisor_raw not in [None, '']:
                try:
                    supervisor_id = int(supervisor_raw)
                except (TypeError, ValueError):
                    return jsonify({"error": "Invalid supervisor_id"}), 400
            if role == 'operator' and not data.get('direction_id'):
                return jsonify({"error": "Missing required field: direction_id"}), 400
            if role == 'operator' and not data.get('rate'):
                return jsonify({"error": "Missing required field: rate"}), 400
            if role == 'operator':
                try:
                    direction_id = int(data['direction_id'])
                except (TypeError, ValueError):
                    return jsonify({"error": "Invalid direction_id"}), 400
                try:
                    rate = float(data['rate'])
                except (TypeError, ValueError):
                    return jsonify({"error": "Invalid rate"}), 400
            else:
                direction_id = None
                try:
                    rate = float(data['rate']) if data.get('rate') else 1.0
                except (TypeError, ValueError):
                    return jsonify({"error": "Invalid rate"}), 400
        elif role == 'trainer':
            # Trainers are never tied to a direction or supervisor.
            supervisor_id = None
            direction_id = None
            try:
                rate = float(data['rate']) if data.get('rate') else 1.0
            except (TypeError, ValueError):
                return jsonify({"error": "Invalid rate"}), 400
        else:
            # Admins are not tied to direction/supervisor.
            supervisor_id = None
            direction_id = None
            try:
                rate = float(data['rate']) if data.get('rate') else 1.0
            except (TypeError, ValueError):
                return jsonify({"error": "Invalid rate"}), 400

        gender = data.get('gender')
        if gender in [None, '']:
            gender = None
        elif gender not in ['male', 'female']:
            return jsonify({"error": "Invalid gender value"}), 400

        birth_date = data.get('birth_date')
        if birth_date:
            try:
                datetime.strptime(birth_date, '%Y-%m-%d')
            except ValueError:
                return jsonify({"error": "Invalid birth_date format. Use YYYY-MM-DD"}), 400
        else:
            birth_date = None

        phone = str(data.get('phone') or '').strip() or None
        email = str(data.get('email') or '').strip() or None
        if email and '@' not in email:
            return jsonify({"error": "Invalid email value"}), 400
        instagram = str(data.get('instagram') or '').strip() or None
        telegram_nick = str(data.get('telegram_nick') or '').strip() or None
        company_name = str(data.get('company_name') or '').strip() or None
        study_place = str(data.get('study_place') or '').strip() or None
        study_course = str(data.get('study_course') or '').strip() or None
        close_contact_1_relation = str(data.get('close_contact_1_relation') or '').strip() or None
        close_contact_1_full_name = str(data.get('close_contact_1_full_name') or '').strip() or None
        close_contact_1_phone = str(data.get('close_contact_1_phone') or '').strip() or None
        close_contact_2_relation = str(data.get('close_contact_2_relation') or '').strip() or None
        close_contact_2_full_name = str(data.get('close_contact_2_full_name') or '').strip() or None
        close_contact_2_phone = str(data.get('close_contact_2_phone') or '').strip() or None
        card_number = str(data.get('card_number') or '').strip() or None
        taxipro_id = str(data.get('taxipro_id') or '').strip() or None

        if phone and not _is_valid_kz_phone(phone):
            return jsonify({"error": "Invalid phone format. Use +7XXXXXXXXXX"}), 400
        if close_contact_1_phone and not _is_valid_kz_phone(close_contact_1_phone):
            return jsonify({"error": "Invalid close_contact_1_phone format. Use +7XXXXXXXXXX"}), 400
        if close_contact_2_phone and not _is_valid_kz_phone(close_contact_2_phone):
            return jsonify({"error": "Invalid close_contact_2_phone format. Use +7XXXXXXXXXX"}), 400

        internship_in_company_raw = data.get('internship_in_company')
        if isinstance(internship_in_company_raw, bool):
            internship_in_company = internship_in_company_raw
        elif isinstance(internship_in_company_raw, (int, float)):
            internship_in_company = bool(internship_in_company_raw)
        elif isinstance(internship_in_company_raw, str):
            internship_in_company_value = internship_in_company_raw.strip().lower()
            if internship_in_company_value in ['1', 'true', 'yes', 'y', 'on']:
                internship_in_company = True
            elif internship_in_company_value in ['0', 'false', 'no', 'n', 'off', '']:
                internship_in_company = False
            else:
                return jsonify({"error": "Invalid internship_in_company value"}), 400
        elif internship_in_company_raw is None:
            internship_in_company = False
        else:
            return jsonify({"error": "Invalid internship_in_company value"}), 400

        front_office_training_raw = data.get('front_office_training')
        if isinstance(front_office_training_raw, bool):
            front_office_training = front_office_training_raw
        elif isinstance(front_office_training_raw, (int, float)):
            front_office_training = bool(front_office_training_raw)
        elif isinstance(front_office_training_raw, str):
            front_office_training_value = front_office_training_raw.strip().lower()
            if front_office_training_value in ['1', 'true', 'yes', 'y', 'on']:
                front_office_training = True
            elif front_office_training_value in ['0', 'false', 'no', 'n', 'off', '']:
                front_office_training = False
            else:
                return jsonify({"error": "Invalid front_office_training value"}), 400
        elif front_office_training_raw is None:
            front_office_training = False
        else:
            return jsonify({"error": "Invalid front_office_training value"}), 400

        front_office_training_date = data.get('front_office_training_date')
        if front_office_training_date:
            try:
                datetime.strptime(front_office_training_date, '%Y-%m-%d')
            except ValueError:
                return jsonify({"error": "Invalid front_office_training_date format. Use YYYY-MM-DD"}), 400
        else:
            front_office_training_date = None
        if not front_office_training:
            front_office_training_date = None

        employment_type = str(data.get('employment_type') or '').strip().lower() or None
        if employment_type not in [None, 'gph', 'of']:
            return jsonify({"error": "Invalid employment_type value"}), 400

        has_proxy_raw = data.get('has_proxy')
        if isinstance(has_proxy_raw, bool):
            has_proxy = has_proxy_raw
        elif isinstance(has_proxy_raw, (int, float)):
            has_proxy = bool(has_proxy_raw)
        elif isinstance(has_proxy_raw, str):
            has_proxy_value = has_proxy_raw.strip().lower()
            if has_proxy_value in ['1', 'true', 'yes', 'y', 'on']:
                has_proxy = True
            elif has_proxy_value in ['0', 'false', 'no', 'n', 'off', '']:
                has_proxy = False
            else:
                return jsonify({"error": "Invalid has_proxy value"}), 400
        elif has_proxy_raw is None:
            has_proxy = False
        else:
            return jsonify({"error": "Invalid has_proxy value"}), 400

        sip_number = str(data.get('sip_number') or '').strip() or None
        if role != 'operator':
            sip_number = None

        if role == 'trainer':
            login_prefix = 'trainer'
        elif role == 'admin':
            login_prefix = 'admin'
        elif role == 'trainee':
            login_prefix = 'trainee'
        else:
            login_prefix = 'user'
        login = f"{login_prefix}_{str(uuid.uuid4())[:8]}"
        password = str(uuid.uuid4())[:8]

        user_id = db.create_user(
            telegram_id=None,
            name=name,
            role=role,
            hire_date=hire_date,
            supervisor_id=supervisor_id,
            rate=rate,
            direction_id=direction_id,
            login=login,
            password=password,
            gender=gender,
            birth_date=birth_date,
            phone=phone,
            email=email,
            instagram=instagram,
            telegram_nick=telegram_nick,
            company_name=company_name,
            employment_type=employment_type,
            has_proxy=has_proxy,
            sip_number=sip_number,
            study_place=study_place,
            study_course=study_course,
            close_contact_1_relation=close_contact_1_relation,
            close_contact_1_full_name=close_contact_1_full_name,
            close_contact_1_phone=close_contact_1_phone,
            close_contact_2_relation=close_contact_2_relation,
            close_contact_2_full_name=close_contact_2_full_name,
            close_contact_2_phone=close_contact_2_phone,
            card_number=card_number,
            internship_in_company=internship_in_company,
            front_office_training=front_office_training,
            front_office_training_date=front_office_training_date,
            taxipro_id=taxipro_id
        )

        changed_by = requester_id
        if role == 'trainer':
            db.update_user(user_id, 'direction_id', None, changed_by=changed_by)
            db.update_user(user_id, 'supervisor_id', None, changed_by=changed_by)

        if role == 'trainer':
            role_label = 'Тренер'
        elif role == 'admin':
            role_label = 'Админ'
        elif role == 'trainee':
            role_label = 'Стажер'
        else:
            role_label = 'Оператор'
        return jsonify({
            "status": "success",
            "message": f"{role_label} {name} добавлен",
            "id": user_id,
            "role": role,
            "login": login,
            "password": password
        })
    except Exception as e:
        logging.error(f"Error adding user: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/directions', methods=['GET'])
@require_api_key
def get_directions():
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            logging.warning("Missing X-User-Id header in /api/admin/directions request")
            return jsonify({"error": "Missing X-User-Id header"}), 400

        try:
            requester_id = int(user_id)
        except (ValueError, TypeError):
            logging.error(f"Invalid X-User-Id format: {user_id}")
            return jsonify({"error": "Invalid X-User-Id format"}), 400

        requester = db.get_user(id=requester_id)

        directions = db.get_directions()
        return jsonify({"status": "success", "directions": directions}), 200
    except Exception as e:
        logging.error(f"Error fetching directions: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/save_directions', methods=['POST'])
@require_api_key
def save_directions():
    try:
        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        if not requester or not _is_admin_role(requester[3]):
            return jsonify({"error": "Only admins can save directions"}), 403

        data = request.get_json()
        if not data or 'directions' not in data:
            return jsonify({"error": "Missing directions data"}), 400

        directions = data['directions']
        db.save_directions(directions)

        # Получаем обновлённый список направлений с id
        updated_directions = db.get_directions()

        return jsonify({
            "status": "success",
            "message": "Directions saved successfully",
            "directions": updated_directions
        }), 200

    except Exception as e:
        logging.error(f"Error saving directions: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/operator/activity', methods=['GET'])
@require_api_key
def get_operator_activity():
    try:
        operator_id = request.args.get('operator_id')
        date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        if not operator_id:
            return jsonify({"error": "Missing operator_id"}), 400
        
        operator_id = int(operator_id)
        supervisor_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=supervisor_id)
        
        # Проверка, что оператор принадлежит супервайзеру
        operator = db.get_user(id=operator_id)
        if requester[3]=="sv" and (not operator or operator[6] != supervisor_id):  # operator[6] - supervisor_id
            return jsonify({"error": "Unauthorized: This operator does not belong to you"}), 403
        
        logs = db.get_activity_logs(operator_id, date_str)
        return jsonify({"status": "success", "logs": logs}), 200
    except Exception as e:
        logging.error(f"Error fetching operator activity: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/sv/operators', methods=['GET'])
@require_api_key
def get_sv_operators_moderka():
    try:
        supervisor_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=supervisor_id)
        if not requester or requester[3] != 'sv':
            return jsonify({"error": "Unauthorized: Only supervisors can access this"}), 403
        
        operators = db.get_operators_by_supervisor(supervisor_id)
        for operator in operators:
            operator['avatar_url'] = _build_avatar_signed_url(
                operator.get('avatar_bucket'),
                operator.get('avatar_blob_path')
            )
            operator.pop('avatar_bucket', None)
            operator.pop('avatar_blob_path', None)
        return jsonify({"status": "success", "operators": operators}), 200
    except Exception as e:
        logging.error(f"Error fetching operators: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/sv/data', methods=['GET'])
@require_api_key
def get_sv_data():
    try:
        # id
        user_id_raw = request.args.get('id')
        if not user_id_raw:
            return jsonify({"error": "Missing ID parameter"}), 400
        try:
            user_id = int(user_id_raw)
        except ValueError:
            return jsonify({"error": "Invalid ID parameter"}), 400

        # optional month YYYY-MM
        month = request.args.get('month')
        if month:
            try:
                datetime.strptime(month, "%Y-%m")
            except ValueError:
                return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400
        else:
            month = datetime.now().strftime("%Y-%m")

        # fetch user (accept any caller role; admin can call this endpoint)
        user = db.get_user(id=user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # normalize user name (support tuple/list or dict)
        if isinstance(user, dict):
            user_name = user.get("name") or user.get("full_name") or ""
        else:
            user_name = user[2] if len(user) > 2 else ""

        response_data = {
            "status": "success",
            "name": user_name,
            "table": user[9] if (not isinstance(user, dict) and len(user) > 9) else (user.get("scores_table_url") if isinstance(user, dict) else None),
            "requested_month": month,
            "operators": []
        }

        # try to get operators for this supervisor id (works when admin requests data for some sv)
        try:
            operators = db.get_operators_by_supervisor(user_id) or []
        except Exception:
            operators = []

        def _extract_score_and_imported(ev):
            """
            Возвращает (score_or_None, is_imported_bool).
            Поддерживает dict, sequence (tuple/list) и objects with attributes.
            For tuple/list we assume score index 4, is_imported last index (if present).
            """
            # dict-like
            try:
                if isinstance(ev, dict):
                    score = ev.get("score")
                    is_imported = bool(ev.get("is_imported")) if "is_imported" in ev else False
                    return (None if score is None else float(score), is_imported)
                # object with attributes
                if hasattr(ev, "__dict__") and not isinstance(ev, (list, tuple)):
                    score = getattr(ev, "score", None)
                    is_imported = bool(getattr(ev, "is_imported", False))
                    return (None if score is None else float(score), is_imported)
                # sequence-like
                if isinstance(ev, (list, tuple)):
                    # common positions from get_call_evaluations: score at index 4; is_imported at index 27 (if exists)
                    score = None
                    is_imported = False
                    if len(ev) > 4:
                        score = ev[4]
                    # try last position for is_imported
                    if len(ev) > 27:
                        is_imported = bool(ev[27])
                    else:
                        # fallback: if last element is bool, consider it
                        last = ev[-1]
                        if isinstance(last, bool):
                            is_imported = bool(last)
                    return (None if score is None else float(score), is_imported)
            except Exception:
                pass
            # cannot parse
            return (None, False)

        for op in operators:
            # flexible unpacking: support dict or sequence rows
            if isinstance(op, dict):
                operator_id = op.get("id") or op.get("operator_id") or op.get("user_id")
                operator_name = op.get("name") or op.get("operator_name")
                direction_id = op.get("direction_id")
                hire_date = op.get("hire_date")
                birth_date = op.get("birth_date")
                hours_table_url = op.get("hours_table_url")
                scores_table_url = op.get("scores_table_url")
                status = op.get("status")
                rate = op.get("rate")
                gender = op.get("gender")
                avatar_url = op.get("avatar_url")
                avatar_bucket = op.get("avatar_bucket")
                avatar_blob_path = op.get("avatar_blob_path")
                status_period_status_code = op.get("status_period_status_code")
                status_period_start_date = op.get("status_period_start_date")
                status_period_end_date = op.get("status_period_end_date")
                status_period_dismissal_reason = op.get("status_period_dismissal_reason")
                status_period_is_blacklist = op.get("status_period_is_blacklist")
                status_period_comment = op.get("status_period_comment")
            else:
                operator_id = op[0] if len(op) > 0 else None
                operator_name = op[1] if len(op) > 1 else None
                direction_id = op[2] if len(op) > 2 else None
                hire_date = op[3] if len(op) > 3 else None
                birth_date = op[10] if len(op) > 10 else None
                hours_table_url = op[4] if len(op) > 4 else None
                scores_table_url = op[5] if len(op) > 5 else None
                status = op[7] if len(op) > 7 else None
                rate = op[8] if len(op) > 8 else None
                gender = op[9] if len(op) > 9 else None
                avatar_url = None
                avatar_bucket = op[11] if len(op) > 11 else None
                avatar_blob_path = op[12] if len(op) > 12 else None
                status_period_status_code = None
                status_period_start_date = None
                status_period_end_date = None
                status_period_dismissal_reason = ""
                status_period_is_blacklist = False
                status_period_comment = ""

            # skip invalid rows
            if not operator_id:
                continue

            # get evaluations for requested month (be tolerant to db method return shape)
            try:
                evaluations = db.get_call_evaluations(operator_id, month=month) or []
            except Exception:
                evaluations = []

            # compute count and average only for реально оценённых записей
            scores = []
            eval_count = 0
            for ev in evaluations:
                score_val, is_imported = _extract_score_and_imported(ev)
                # count only non-imported and with a numeric score
                if not is_imported and score_val is not None:
                    scores.append(score_val)
                    eval_count += 1

            avg_score = round(sum(scores) / len(scores), 2) if len(scores) > 0 else None

            # ensure rate is numeric and reasonable
            try:
                rate_val = float(rate) if rate is not None else 1.0
            except Exception:
                rate_val = 1.0

            response_data["operators"].append({
                "id": operator_id,
                "name": operator_name,
                "hire_date": hire_date,
                "birth_date": (birth_date.strftime('%d-%m-%Y') if hasattr(birth_date, 'strftime') else birth_date),
                "direction_id": direction_id,
                # number of actual evaluated calls (with scores)
                "call_count": eval_count,
                "avg_score": avg_score,
                "scores_table_url": scores_table_url,
                "status": status,
                "rate": rate_val,
                "gender": gender,
                "avatar_url": avatar_url or _build_avatar_signed_url(avatar_bucket, avatar_blob_path),
                "status_period_status_code": status_period_status_code,
                "status_period_start_date": status_period_start_date,
                "status_period_end_date": status_period_end_date,
                "status_period_dismissal_reason": status_period_dismissal_reason or "",
                "status_period_is_blacklist": bool(status_period_is_blacklist),
                "status_period_comment": status_period_comment or ""
            })

        return jsonify(response_data), 200

    except Exception:
        logging.exception("Error fetching SV data")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/notify_sv', methods=['POST'])
@require_api_key
def notify_supervisor():
    try:
        data = request.get_json()
        required_fields = ['sv_id', 'operator_name']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        sv_id = int(data['sv_id'])
        operator_name = data['operator_name']
        
        current_week = get_current_week_of_month()
        expected_calls = get_expected_calls(current_week)
        
        user = db.get_user(id=sv_id)
        if not user or user[3] != 'sv':
            return jsonify({"error": "SV not found"}), 404

        notification_text = (
            f"⚠️ <b>Требуется внимание!</b>\n\n"
            f"У оператора <b>{operator_name}</b> недостаточно прослушанных звонков.\n"
            f"Текущая неделя: {current_week}\n"
            f"Ожидается: {expected_calls} звонков (по 5 в неделю)\n\n"
            f"Пожалуйста, проверьте и прослушайте недостающие звонки."
        )
        
        telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
        payload = {
            "chat_id": user[1],
            "text": notification_text,
            "parse_mode": "HTML"
        }
        response = requests.post(telegram_url, json=payload, timeout=10)
        
        if response.status_code != 200:
            error_detail = response.json().get('description', 'Unknown error')
            logging.error(f"Telegram API error: {error_detail}")
            return jsonify({"error": f"Failed to send Telegram message: {error_detail}"}), 500

        return jsonify({"status": "success", "message": f"Notification sent to {user[2]}"}), 200
    except Exception as e:
        logging.error(f"Error in notify_supervisor: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/surveys', methods=['GET', 'POST', 'OPTIONS'])
@require_api_key
def handle_surveys():
    try:
        requester_id, requester, requester_role, guard_response, guard_status = _surveys_route_guard()
        if guard_response is not None:
            return guard_response, guard_status

        if request.method == 'GET':
            if requester_role == 'operator':
                surveys = db.get_surveys_for_operator(requester_id)
            else:
                surveys = db.get_surveys_for_management(requester_id, requester_role)

            return jsonify({
                "status": "success",
                "surveys": surveys
            }), 200

        if not (_is_admin_role(requester_role) or requester_role in ('sv', 'trainer')):
            return jsonify({"error": "Only admin, sv and trainer can create surveys"}), 403

        data = request.get_json() or {}
        title = data.get('title')
        description = data.get('description')
        assignment = data.get('assignment') or {}
        questions = data.get('questions') or []
        is_test_raw = data.get('is_test', False)
        if isinstance(is_test_raw, str):
            is_test = is_test_raw.strip().lower() in ('1', 'true', 'yes', 'on')
        else:
            is_test = bool(is_test_raw)
        repeat_from_survey_id_raw = data.get('repeat_from_survey_id')
        repeat_from_survey_id = None
        if repeat_from_survey_id_raw not in (None, ''):
            try:
                repeat_from_survey_id = int(repeat_from_survey_id_raw)
            except Exception:
                return jsonify({"error": "Invalid repeat_from_survey_id"}), 400
            if repeat_from_survey_id <= 0:
                return jsonify({"error": "Invalid repeat_from_survey_id"}), 400

        operator_ids_raw = assignment.get('operator_ids') or []
        operator_ids = []
        for op_id in operator_ids_raw:
            try:
                parsed = int(op_id)
            except Exception:
                continue
            if parsed > 0 and parsed not in operator_ids:
                operator_ids.append(parsed)

        if not operator_ids:
            return jsonify({"error": "At least one operator must be assigned"}), 400

        visible_operator_ids = set(db.get_visible_operator_ids_for_requester(requester_id, requester_role))
        forbidden = [op_id for op_id in operator_ids if op_id not in visible_operator_ids]
        if forbidden:
            return jsonify({"error": "You cannot assign surveys to these operators", "operator_ids": forbidden}), 403

        if repeat_from_survey_id is not None:
            visible_surveys = db.get_surveys_for_management(requester_id, requester_role)
            visible_survey_ids = set()
            for survey_item in visible_surveys:
                try:
                    visible_survey_ids.add(int(survey_item.get('id')))
                except Exception:
                    continue
            if repeat_from_survey_id not in visible_survey_ids:
                return jsonify({"error": "Source survey for repeat not found"}), 404

        created = db.create_survey(
            title=title,
            description=description,
            created_by=requester_id,
            assignment=assignment,
            questions=questions,
            operator_ids=operator_ids,
            repeat_from_survey_id=repeat_from_survey_id,
            is_test=is_test
        )

        return jsonify({
            "status": "success",
            "message": "Survey created successfully",
            "survey_id": created.get('id'),
            "created_at": created.get('created_at'),
            "repeat_root_id": created.get('repeat_root_id'),
            "repeat_iteration": created.get('repeat_iteration'),
            "is_test": bool(created.get('is_test'))
        }), 201

    except ValueError as value_error:
        code = str(value_error)
        if code == 'SURVEY_TITLE_REQUIRED':
            return jsonify({"error": "Survey title is required"}), 400
        if code == 'SURVEY_QUESTIONS_REQUIRED':
            return jsonify({"error": "At least one question is required"}), 400
        if code == 'SURVEY_OPERATORS_REQUIRED':
            return jsonify({"error": "At least one operator is required"}), 400
        if code in ('SURVEY_INVALID_TENURE_MIN', 'SURVEY_INVALID_TENURE_MAX', 'SURVEY_INVALID_TENURE_RANGE'):
            return jsonify({"error": "Invalid tenure range"}), 400
        if code.startswith('SURVEY_QUESTION_TEXT_REQUIRED_'):
            return jsonify({"error": "Question text is required"}), 400
        if code.startswith('SURVEY_INVALID_QUESTION_TYPE_'):
            return jsonify({"error": "Invalid question type"}), 400
        if code.startswith('SURVEY_OPTIONS_REQUIRED_'):
            return jsonify({"error": "Question must have at least 2 options"}), 400
        if code.startswith('SURVEY_TEST_RATING_NOT_ALLOWED_'):
            return jsonify({"error": "Rating questions are not allowed in test mode"}), 400
        if code.startswith('SURVEY_CORRECT_OPTIONS_REQUIRED_'):
            return jsonify({"error": "Each test question must have at least one correct option"}), 400
        if code.startswith('SURVEY_CORRECT_OPTION_INVALID_'):
            return jsonify({"error": "Correct options must match question options"}), 400
        if code.startswith('SURVEY_SINGLE_CORRECT_OPTION_REQUIRED_'):
            return jsonify({"error": "Single-choice test questions must have exactly one correct option"}), 400
        if code == 'SURVEY_REPEAT_SOURCE_NOT_FOUND':
            return jsonify({"error": "Source survey for repeat not found"}), 404
        return jsonify({"error": code}), 400
    except Exception as e:
        logging.error(f"Error in handle_surveys: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/surveys/<int:survey_id>', methods=['DELETE', 'OPTIONS'])
@require_api_key
def delete_survey(survey_id):
    try:
        requester_id, requester, requester_role, guard_response, guard_status = _surveys_route_guard()
        if guard_response is not None:
            return guard_response, guard_status

        if not (_is_admin_role(requester_role) or requester_role in ('sv', 'trainer')):
            return jsonify({"error": "Only admin, sv and trainer can delete surveys"}), 403

        db.delete_survey(
            survey_id=survey_id,
            requester_id=requester_id,
            requester_role=requester_role
        )
        return jsonify({"status": "success", "message": "Survey deleted"}), 200
    except ValueError as value_error:
        if str(value_error) == 'SURVEY_NOT_FOUND':
            return jsonify({"error": "Survey not found"}), 404
        return jsonify({"error": str(value_error)}), 400
    except PermissionError:
        return jsonify({"error": "You do not have access to delete this survey"}), 403
    except Exception as e:
        logging.error(f"Error in delete_survey: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/surveys/<int:survey_id>/submit', methods=['POST', 'OPTIONS'])
@require_api_key
def submit_survey(survey_id):
    try:
        requester_id, requester, requester_role, guard_response, guard_status = _surveys_route_guard()
        if guard_response is not None:
            return guard_response, guard_status

        if requester_role != 'operator':
            return jsonify({"error": "Only operators can submit surveys"}), 403

        data = request.get_json() or {}
        answers = data.get('answers')

        result = db.submit_survey_response(
            survey_id=survey_id,
            operator_id=requester_id,
            answers=answers
        )

        return jsonify({
            "status": "success",
            "message": "Survey completed",
            "result": result
        }), 200
    except ValueError as value_error:
        code = str(value_error)
        if code == 'SURVEY_NOT_FOUND':
            return jsonify({"error": "Survey not found"}), 404
        if code == 'SURVEY_ALREADY_COMPLETED':
            return jsonify({"error": "Survey already completed"}), 409
        if code == 'SURVEY_HAS_NO_QUESTIONS':
            return jsonify({"error": "Survey has no questions"}), 400
        if code == 'SURVEY_ANSWERS_REQUIRED':
            return jsonify({"error": "Answers are required"}), 400
        if code == 'SURVEY_EMPTY_RESPONSE':
            return jsonify({"error": "Please answer at least one question"}), 400
        if code.startswith('SURVEY_REQUIRED_QUESTION_'):
            return jsonify({"error": "Please answer all required questions"}), 400
        if code.startswith('SURVEY_INVALID_RATING_'):
            return jsonify({"error": "Rating answer must be between 1 and 5"}), 400
        if code.startswith('SURVEY_INVALID_OPTION_') or code.startswith('SURVEY_TOO_MANY_OPTIONS_'):
            return jsonify({"error": "Invalid selected option"}), 400
        if code.startswith('SURVEY_TEST_RATING_NOT_ALLOWED_'):
            return jsonify({"error": "Test configuration is invalid for this survey"}), 400
        if code.startswith('SURVEY_OTHER_TEXT_TOO_LONG_'):
            return jsonify({"error": "Other answer must not exceed 500 characters"}), 400
        return jsonify({"error": code}), 400
    except PermissionError as permission_error:
        if str(permission_error) == 'SURVEY_NOT_ASSIGNED':
            return jsonify({"error": "Survey is not assigned to this operator"}), 403
        return jsonify({"error": str(permission_error)}), 403
    except Exception as e:
        logging.error(f"Error in submit_survey: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


def _survey_export_format_dt(value):
    if value in (None, ''):
        return ''
    text = str(value).strip()
    if not text:
        return ''
    try:
        if text.endswith('Z'):
            text = text[:-1] + '+00:00'
        dt_value = datetime.fromisoformat(text)
        return dt_value.strftime('%d.%m.%Y %H:%M')
    except Exception:
        if 'T' in text:
            return text.replace('T', ' ')[:16]
        return text


def _survey_export_answer_text(question, answer):
    if not question or not answer:
        return '—'

    qtype = str((question or {}).get('type') or '').strip().lower()
    if qtype == 'rating':
        rating_value = answer.get('rating_value')
        try:
            return str(int(rating_value))
        except Exception:
            return '—'

    selected = []
    for item in (answer.get('selected_options') or []):
        text = str(item or '').strip()
        if text and text not in selected:
            selected.append(text)

    other_text = str(answer.get('answer_text') or '').strip()
    if selected and other_text:
        return f"{', '.join(selected)}; Другое: {other_text}"
    if selected:
        return ', '.join(selected)
    if other_text:
        return f"Другое: {other_text}"
    return '—'


def _survey_export_unique_trimmed_list(values):
    source = values if isinstance(values, list) else []
    normalized = []
    for value in source:
        text = str(value or '').strip()
        if text and text not in normalized:
            normalized.append(text)
    return normalized


def _survey_export_expected_options(question, answer):
    expected_from_answer = _survey_export_unique_trimmed_list((answer or {}).get('expected_options'))
    if expected_from_answer:
        return expected_from_answer
    return _survey_export_unique_trimmed_list((question or {}).get('correct_options'))


def _survey_export_is_answer_present(question, answer):
    if not question or not answer:
        return False

    qtype = str((question or {}).get('type') or '').strip().lower()
    if qtype == 'rating':
        try:
            return answer.get('rating_value') is not None and str(answer.get('rating_value')).strip() != ''
        except Exception:
            return False

    selected = _survey_export_unique_trimmed_list((answer or {}).get('selected_options'))
    answer_text = str((answer or {}).get('answer_text') or '').strip()
    return bool(selected or answer_text)


def _survey_export_is_test_answer_correct(question, answer):
    if not question or not answer:
        return False
    if isinstance(answer.get('is_correct'), bool):
        return bool(answer.get('is_correct'))

    qtype = str((question or {}).get('type') or '').strip().lower()
    selected_options = _survey_export_unique_trimmed_list((answer or {}).get('selected_options'))
    answer_text = str((answer or {}).get('answer_text') or '').strip()
    expected_options = _survey_export_expected_options(question, answer)

    if qtype == 'single':
        return (
            len(expected_options) == 1
            and len(selected_options) == 1
            and selected_options[0] == expected_options[0]
            and not answer_text
        )
    if qtype == 'multiple':
        return (
            len(expected_options) > 0
            and len(selected_options) == len(expected_options)
            and all(option in selected_options for option in expected_options)
            and not answer_text
        )
    return False


def _survey_export_answer_text_for_test(question, answer):
    base_text = _survey_export_answer_text(question, answer)
    has_answer = _survey_export_is_answer_present(question, answer)
    if not has_answer:
        return f"{base_text} [Нет ответа]"

    is_correct = _survey_export_is_test_answer_correct(question, answer)
    status_text = 'Верно' if is_correct else 'Неверно'
    expected_options = _survey_export_expected_options(question, answer)
    if expected_options:
        return f"{base_text} [{status_text}] (Правильный: {', '.join(expected_options)})"
    return f"{base_text} [{status_text}]"


def _survey_export_test_row_metrics(row):
    summary = row.get('test_summary') or {}
    total_questions = int(summary.get('total_questions') or 0)
    answered_questions = int(summary.get('answered_questions') or 0)
    correct_answers = int(summary.get('correct_answers') or 0)

    score_percent = None
    try:
        if summary.get('score_percent') is not None:
            score_percent = float(summary.get('score_percent'))
    except Exception:
        score_percent = None

    if total_questions > 0:
        correct_ratio = f"{correct_answers}/{total_questions}"
        answered_ratio = f"{answered_questions}/{total_questions}"
    else:
        correct_ratio = '—'
        answered_ratio = '—'

    return {
        'score_percent': score_percent,
        'correct_ratio': correct_ratio,
        'answered_ratio': answered_ratio
    }


def _survey_export_answer_text_with_mode(question, answer, is_test=False):
    if is_test:
        return _survey_export_answer_text_for_test(question, answer)
    return _survey_export_answer_text(question, answer)


@app.route('/api/surveys/<int:survey_id>/export_excel', methods=['GET', 'OPTIONS'])
@require_api_key
def export_survey_statistics_excel(survey_id):
    try:
        requester_id, requester, requester_role, guard_response, guard_status = _surveys_route_guard()
        if guard_response is not None:
            return guard_response, guard_status

        if not (_is_admin_role(requester_role) or requester_role in ('sv', 'trainer')):
            return jsonify({"error": "Only admin, sv and trainer can export survey statistics"}), 403

        surveys = db.get_surveys_for_management(requester_id, requester_role)
        selected_survey = None
        for survey_item in surveys:
            try:
                if int(survey_item.get('id')) == int(survey_id):
                    selected_survey = survey_item
                    break
            except Exception:
                continue

        if not selected_survey:
            return jsonify({"error": "Survey not found"}), 404

        title = str(selected_survey.get('title') or f'Опрос #{survey_id}')
        description = str(selected_survey.get('description') or '')
        statistics = selected_survey.get('statistics') or {}
        questions = list(selected_survey.get('questions') or [])
        is_test = bool(selected_survey.get('is_test'))
        question_stats = list(statistics.get('question_stats') or [])
        responses_detailed_all = list(statistics.get('responses_detailed_all_repetitions') or [])
        responses_detailed = responses_detailed_all if responses_detailed_all else list(statistics.get('responses_detailed') or [])
        repeat_info = selected_survey.get('repeat') or {}
        try:
            repeat_iteration = int(repeat_info.get('iteration') or 1)
        except Exception:
            repeat_iteration = 1
        repeat_iteration = repeat_iteration if repeat_iteration >= 1 else 1

        question_type_labels = {
            'single': 'Один вариант',
            'multiple': 'Несколько вариантов',
            'rating': 'Рейтинг 1-5'
        }
        question_stats_by_id = {}
        for stat in question_stats:
            try:
                qid = int(stat.get('question_id'))
            except Exception:
                continue
            question_stats_by_id[qid] = stat

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Сводка'

        ws_summary.append(['Отчет', 'Статистика по опросу'])
        ws_summary.append(['Опрос', title])
        ws_summary.append(['Описание', description])
        ws_summary.append(['Повторение', f"#{repeat_iteration}"])
        ws_summary.append(['Создан', _survey_export_format_dt(selected_survey.get('created_at'))])
        ws_summary.append(['Обновлен', _survey_export_format_dt(selected_survey.get('updated_at'))])
        ws_summary.append(['Сформирован', datetime.now().strftime('%d.%m.%Y %H:%M')])
        ws_summary.append(['Назначено', int(statistics.get('assigned_count') or 0)])
        ws_summary.append(['Пройдено', int(statistics.get('completed_count') or 0)])
        ws_summary.append(['Ожидают', int(statistics.get('pending_count') or 0)])
        ws_summary.append(['Процент прохождения, %', float(statistics.get('completion_rate') or 0)])
        if is_test:
            score_values = []
            for response_row in responses_detailed:
                metrics = _survey_export_test_row_metrics(response_row)
                score_percent = metrics.get('score_percent')
                if isinstance(score_percent, (int, float)):
                    score_values.append(float(score_percent))
            ws_summary.append(['Средний балл теста, %', round(sum(score_values) / len(score_values), 1) if score_values else '—'])
            ws_summary.append(['Лучший балл теста, %', max(score_values) if score_values else '—'])
            ws_summary.append(['Минимальный балл теста, %', min(score_values) if score_values else '—'])
        ws_summary.append([])
        ws_summary.append(['Статистика по вопросам'])
        ws_summary.append([
            '№',
            'Вопрос',
            'Тип',
            'Ответили',
            'Респондентов (по вопросу)',
            'Доля ответивших, %',
            'Метрика',
            'Значение',
            'Количество',
            '% от ответивших',
            '% от респондентов'
        ])
        summary_correct_option_rows = set()

        for idx, question in enumerate(questions, start=1):
            question_id = int(question.get('id') or 0)
            qtype = str(question.get('type') or '').strip().lower()
            qtype_label = question_type_labels.get(qtype, qtype or '—')
            question_text = str(question.get('text') or f'Вопрос #{idx}')
            stat = question_stats_by_id.get(question_id)

            if not stat:
                ws_summary.append([idx, question_text, qtype_label, 0, 0, 0, 'Нет данных', '', '', '', ''])
                continue

            answered = int(stat.get('responses_with_answer') or 0)
            respondents = int(stat.get('question_respondents_total') or answered)
            response_rate = float(stat.get('response_rate') or 0)
            correct_options_set = set(_survey_export_unique_trimmed_list((stat or {}).get('correct_options') or question.get('correct_options')))

            if qtype == 'rating':
                ws_summary.append([
                    idx, question_text, qtype_label, answered, respondents, response_rate,
                    'Среднее', stat.get('average_rating') if stat.get('average_rating') is not None else '—', '', '', ''
                ])
                ws_summary.append([
                    idx, question_text, qtype_label, answered, respondents, response_rate,
                    'Медиана', stat.get('median_rating') if stat.get('median_rating') is not None else '—', '', '', ''
                ])
                min_rating = stat.get('min_rating')
                max_rating = stat.get('max_rating')
                rating_range = '—'
                if min_rating is not None or max_rating is not None:
                    rating_range = f"{min_rating if min_rating is not None else '—'}-{max_rating if max_rating is not None else '—'}"
                ws_summary.append([
                    idx, question_text, qtype_label, answered, respondents, response_rate,
                    'Диапазон', rating_range, '', '', ''
                ])
                for bucket in (stat.get('ratings_distribution_detailed') or []):
                    ws_summary.append([
                        idx,
                        question_text,
                        qtype_label,
                        answered,
                        respondents,
                        response_rate,
                        f"Оценка {bucket.get('value')}",
                        '',
                        int(bucket.get('count') or 0),
                        float(bucket.get('percent_of_answers') or 0),
                        float(bucket.get('percent_of_respondents') or 0)
                    ])
                continue

            options = list(stat.get('options') or [])
            if not options:
                ws_summary.append([idx, question_text, qtype_label, answered, respondents, response_rate, 'Нет данных', '', '', '', ''])
                continue

            for option_item in options:
                option_text = str(option_item.get('option') or '').strip()
                is_correct_option = is_test and option_text and option_text in correct_options_set
                ws_summary.append([
                    idx,
                    question_text,
                    qtype_label,
                    answered,
                    respondents,
                    response_rate,
                    'Вариант (правильный)' if is_correct_option else 'Вариант',
                    option_text,
                    int(option_item.get('count') or 0),
                    float(option_item.get('percent_of_answers') or option_item.get('percent') or 0),
                    float(option_item.get('percent_of_respondents') or option_item.get('percent') or 0)
                ])
                if is_correct_option:
                    summary_correct_option_rows.add(ws_summary.max_row)

        ws_scores = None
        if is_test:
            ws_scores = wb.create_sheet('Баллы теста')
            ws_scores.append(['Сотрудник', 'ID', 'Статус', 'Отправлено', 'Повторение', 'Общий балл, %', 'Верно', 'Отвечено'])
            for row in responses_detailed:
                status_raw = str(row.get('status') or '').strip().lower()
                status_label = 'Пройден' if status_raw == 'completed' else 'Назначен'
                test_metrics = _survey_export_test_row_metrics(row)
                ws_scores.append([
                    str(row.get('operator_name') or f"#{row.get('operator_id') or ''}"),
                    int(row.get('operator_id') or 0),
                    status_label,
                    _survey_export_format_dt(row.get('submitted_at')),
                    int(row.get('repeat_iteration') or repeat_iteration),
                    test_metrics.get('score_percent') if test_metrics.get('score_percent') is not None else '—',
                    test_metrics.get('correct_ratio'),
                    test_metrics.get('answered_ratio')
                ])

        ws_answers = wb.create_sheet('Ответы сотрудников')
        answer_headers = ['Сотрудник', 'ID', 'Статус', 'Отправлено', 'Повторение']
        if is_test:
            answer_headers.extend(['Общий балл, %', 'Верно', 'Отвечено'])
        for idx, question in enumerate(questions, start=1):
            question_text = str(question.get('text') or f'Вопрос #{idx}').replace('\n', ' ').strip()
            if len(question_text) > 100:
                question_text = question_text[:97] + '...'
            answer_headers.append(f"Q{idx}: {question_text}")
        ws_answers.append(answer_headers)

        for row in responses_detailed:
            answers_by_question = row.get('answers_by_question') or {}
            status_raw = str(row.get('status') or '').strip().lower()
            status_label = 'Пройден' if status_raw == 'completed' else 'Назначен'
            row_values = [
                str(row.get('operator_name') or f"#{row.get('operator_id') or ''}"),
                int(row.get('operator_id') or 0),
                status_label,
                _survey_export_format_dt(row.get('submitted_at')),
                int(row.get('repeat_iteration') or repeat_iteration)
            ]
            if is_test:
                test_metrics = _survey_export_test_row_metrics(row)
                row_values.extend([
                    test_metrics.get('score_percent') if test_metrics.get('score_percent') is not None else '—',
                    test_metrics.get('correct_ratio'),
                    test_metrics.get('answered_ratio')
                ])

            for question in questions:
                question_id = int(question.get('id') or 0)
                answer = answers_by_question.get(str(question_id)) or answers_by_question.get(question_id)
                row_values.append(_survey_export_answer_text_with_mode(question, answer, is_test=is_test))

            ws_answers.append(row_values)

        # Visual formatting for a cleaner, management-friendly export
        title_fill = PatternFill(fill_type='solid', start_color='1F4E78', end_color='1F4E78')
        section_fill = PatternFill(fill_type='solid', start_color='2F75B5', end_color='2F75B5')
        header_fill = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')
        label_fill = PatternFill(fill_type='solid', start_color='E9EFFB', end_color='E9EFFB')
        even_row_fill = PatternFill(fill_type='solid', start_color='F8FAFF', end_color='F8FAFF')
        status_done_fill = PatternFill(fill_type='solid', start_color='E6F4EA', end_color='E6F4EA')
        status_pending_fill = PatternFill(fill_type='solid', start_color='FFF4E5', end_color='FFF4E5')
        correct_option_fill = PatternFill(fill_type='solid', start_color='DCFCE7', end_color='DCFCE7')
        answer_correct_fill = PatternFill(fill_type='solid', start_color='DCFCE7', end_color='DCFCE7')
        answer_incorrect_fill = PatternFill(fill_type='solid', start_color='FEF2F2', end_color='FEF2F2')
        answer_empty_fill = PatternFill(fill_type='solid', start_color='F3F4F6', end_color='F3F4F6')

        white_bold_font = Font(color='FFFFFF', bold=True)
        dark_bold_font = Font(color='1F4E78', bold=True)
        regular_font = Font(color='1F2937')
        correct_option_font = Font(color='166534', bold=True)
        answer_correct_font = Font(color='166534')
        answer_incorrect_font = Font(color='991B1B')
        answer_empty_font = Font(color='6B7280')

        thin_side = Side(style='thin', color='D9E2F3')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # ---- Summary sheet styles ----
        summary_header_row = None
        for row_idx in range(1, ws_summary.max_row + 1):
            if ws_summary.cell(row=row_idx, column=1).value == '№':
                summary_header_row = row_idx
                break

        summary_section_row = summary_header_row - 1 if summary_header_row else None
        if summary_section_row and summary_section_row >= 1:
            ws_summary.merge_cells(start_row=summary_section_row, start_column=1, end_row=summary_section_row, end_column=11)
            section_cell = ws_summary.cell(row=summary_section_row, column=1)
            section_cell.fill = section_fill
            section_cell.font = white_bold_font
            section_cell.alignment = Alignment(horizontal='left', vertical='center')

        # Report title row
        ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        ws_summary.cell(row=1, column=1).value = 'Статистика по опросу'
        ws_summary.cell(row=1, column=1).fill = title_fill
        ws_summary.cell(row=1, column=1).font = white_bold_font
        ws_summary.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws_summary.row_dimensions[1].height = 24

        # Meta rows (key-value)
        meta_last_row = (summary_section_row - 2) if summary_section_row else 1
        for row_idx in range(2, max(2, meta_last_row + 1)):
            key_cell = ws_summary.cell(row=row_idx, column=1)
            val_cell = ws_summary.cell(row=row_idx, column=2)
            key_cell.fill = label_fill
            key_cell.font = dark_bold_font
            key_cell.alignment = Alignment(horizontal='left', vertical='center')
            val_cell.font = regular_font
            val_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            key_cell.border = thin_border
            val_cell.border = thin_border

        if summary_header_row:
            for col_idx in range(1, 12):
                cell = ws_summary.cell(row=summary_header_row, column=col_idx)
                cell.fill = header_fill
                cell.font = white_bold_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            for row_idx in range(summary_header_row + 1, ws_summary.max_row + 1):
                for col_idx in range(1, 12):
                    cell = ws_summary.cell(row=row_idx, column=col_idx)
                    cell.font = regular_font
                    cell.border = thin_border
                    if col_idx in (2, 8):
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    elif col_idx in (1, 3, 4, 5, 6, 7, 9, 10, 11):
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    if isinstance(cell.value, (int, float)):
                        if col_idx in (6, 10, 11):
                            cell.number_format = '0.0"%"'
                        elif col_idx in (4, 5, 9):
                            cell.number_format = '0'

            for row_idx in summary_correct_option_rows:
                if row_idx <= summary_header_row or row_idx > ws_summary.max_row:
                    continue
                for col_idx in range(7, 12):
                    marked_cell = ws_summary.cell(row=row_idx, column=col_idx)
                    marked_cell.fill = correct_option_fill
                    if col_idx in (7, 8):
                        marked_cell.font = correct_option_font

        summary_widths = {
            1: 6,   # №
            2: 56,  # Вопрос
            3: 22,  # Тип
            4: 12,  # Ответили
            5: 14,  # Респондентов
            6: 18,  # Доля ответивших
            7: 20,  # Метрика
            8: 24,  # Значение
            9: 14,  # Количество
            10: 18, # % от ответивших
            11: 20  # % от респондентов
        }
        for col_idx, width in summary_widths.items():
            ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

        if summary_header_row:
            ws_summary.auto_filter.ref = f"A{summary_header_row}:K{ws_summary.max_row}"
            ws_summary.freeze_panes = f"A{summary_header_row + 1}"
            ws_summary.row_dimensions[summary_header_row].height = 30

        # ---- Answers sheet styles ----
        answers_question_start_col = 9 if is_test else 6
        last_answers_col = ws_answers.max_column
        last_answers_row = ws_answers.max_row

        for col_idx in range(1, last_answers_col + 1):
            header_cell = ws_answers.cell(row=1, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = white_bold_font
            header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_cell.border = thin_border
        ws_answers.row_dimensions[1].height = 28

        for row_idx in range(2, last_answers_row + 1):
            for col_idx in range(1, last_answers_col + 1):
                cell = ws_answers.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = even_row_fill

                if col_idx >= answers_question_start_col:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                elif col_idx in (2, 3, 4, 5, 6, 7, 8):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                if isinstance(cell.value, (int, float)) and (
                    (is_test and col_idx == 6)
                ):
                    cell.number_format = '0.0"%"'

                if is_test and col_idx >= answers_question_start_col:
                    answer_text = str(cell.value or '')
                    if '[Верно]' in answer_text:
                        cell.fill = answer_correct_fill
                        cell.font = answer_correct_font
                    elif '[Неверно]' in answer_text:
                        cell.fill = answer_incorrect_fill
                        cell.font = answer_incorrect_font
                    elif '[Нет ответа]' in answer_text:
                        cell.fill = answer_empty_fill
                        cell.font = answer_empty_font

            status_cell = ws_answers.cell(row=row_idx, column=3)
            status_text = str(status_cell.value or '').strip().lower()
            if status_text == 'пройден':
                status_cell.fill = status_done_fill
            else:
                status_cell.fill = status_pending_fill
            status_cell.font = Font(color='1F2937', bold=True)

        ws_answers.column_dimensions['A'].width = 28
        ws_answers.column_dimensions['B'].width = 10
        ws_answers.column_dimensions['C'].width = 14
        ws_answers.column_dimensions['D'].width = 20
        ws_answers.column_dimensions['E'].width = 12
        if is_test:
            ws_answers.column_dimensions['F'].width = 14
            ws_answers.column_dimensions['G'].width = 12
            ws_answers.column_dimensions['H'].width = 12
        for col_idx in range(answers_question_start_col, last_answers_col + 1):
            ws_answers.column_dimensions[get_column_letter(col_idx)].width = 38

        ws_answers.freeze_panes = 'I2' if is_test else 'F2'
        ws_answers.auto_filter.ref = f"A1:{get_column_letter(last_answers_col)}{max(1, last_answers_row)}"

        # ---- Test scores sheet styles ----
        if ws_scores is not None:
            last_scores_col = ws_scores.max_column
            last_scores_row = ws_scores.max_row

            for col_idx in range(1, last_scores_col + 1):
                header_cell = ws_scores.cell(row=1, column=col_idx)
                header_cell.fill = header_fill
                header_cell.font = white_bold_font
                header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                header_cell.border = thin_border
            ws_scores.row_dimensions[1].height = 28

            for row_idx in range(2, last_scores_row + 1):
                for col_idx in range(1, last_scores_col + 1):
                    cell = ws_scores.cell(row=row_idx, column=col_idx)
                    cell.font = regular_font
                    cell.border = thin_border
                    if row_idx % 2 == 0:
                        cell.fill = even_row_fill

                    if col_idx in (1,):
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    if col_idx == 6 and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0"%"'

                status_cell = ws_scores.cell(row=row_idx, column=3)
                status_text = str(status_cell.value or '').strip().lower()
                if status_text == 'пройден':
                    status_cell.fill = status_done_fill
                else:
                    status_cell.fill = status_pending_fill
                status_cell.font = Font(color='1F2937', bold=True)

            ws_scores.column_dimensions['A'].width = 28
            ws_scores.column_dimensions['B'].width = 10
            ws_scores.column_dimensions['C'].width = 14
            ws_scores.column_dimensions['D'].width = 20
            ws_scores.column_dimensions['E'].width = 12
            ws_scores.column_dimensions['F'].width = 14
            ws_scores.column_dimensions['G'].width = 12
            ws_scores.column_dimensions['H'].width = 12
            ws_scores.freeze_panes = 'F2'
            ws_scores.auto_filter.ref = f"A1:{get_column_letter(last_scores_col)}{max(1, last_scores_row)}"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_title = re.sub(r'[^A-Za-zА-Яа-я0-9 _-]+', '', title).strip().replace(' ', '_')
        if not safe_title:
            safe_title = f'survey_{survey_id}'
        filename = f"{safe_title[:64]}_stats.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logging.error(f"Error in export_survey_statistics_excel: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/tasks/recipients', methods=['GET', 'OPTIONS'])
@require_api_key
def get_task_recipients():
    try:
        requester_id, requester, guard_response, guard_status = _task_route_guard()
        if guard_response is not None:
            return guard_response, guard_status

        recipients = db.get_task_recipients(requester_id, requester[3])
        return jsonify({
            "status": "success",
            "recipients": recipients
        }), 200
    except Exception as e:
        logging.error(f"Error in get_task_recipients: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/tasks', methods=['GET', 'POST', 'OPTIONS'])
@require_api_key
def handle_tasks():
    try:
        requester_id, requester, guard_response, guard_status = _task_route_guard()
        if guard_response is not None:
            return guard_response, guard_status

        requester_role = requester[3]

        if request.method == 'GET':
            tasks = db.get_tasks_for_requester(requester_id, requester_role)
            return jsonify({
                "status": "success",
                "tasks": tasks
            }), 200

        subject = (request.form.get('subject') or '').strip()
        description = (request.form.get('description') or '').strip()
        tag = (request.form.get('tag') or 'task').strip().lower() or 'task'
        assigned_to_raw = request.form.get('assigned_to')

        if not subject:
            return jsonify({"error": "subject is required"}), 400
        if tag not in TASK_ALLOWED_TAGS:
            return jsonify({"error": "Invalid tag"}), 400
        if not assigned_to_raw:
            return jsonify({"error": "assigned_to is required"}), 400

        try:
            assigned_to = int(assigned_to_raw)
        except Exception:
            return jsonify({"error": "assigned_to must be an integer"}), 400

        allowed_recipients = db.get_task_recipients(requester_id, requester_role)
        allowed_ids = {int(item['id']) for item in allowed_recipients if item.get('id') is not None}
        if assigned_to not in allowed_ids:
            return jsonify({"error": "You cannot assign a task to this user"}), 403

        files = request.files.getlist('files')
        try:
            attachments, uploaded_blob_paths, gcs_bucket = _upload_task_attachments_to_gcs(files, stage='initial')
        except ValueError as upload_error:
            return jsonify({"error": str(upload_error)}), 400
        except RuntimeError as upload_error:
            return jsonify({"error": str(upload_error)}), 500
        except Exception as upload_error:
            logging.error(f"Task attachment upload failed: {upload_error}")
            return jsonify({"error": "Failed to upload task attachments"}), 500

        try:
            created = db.create_task(
                subject=subject,
                description=description,
                tag=tag,
                assigned_to=assigned_to,
                created_by=requester_id,
                attachments=attachments
            )
        except Exception:
            # DB write failed after upload -> cleanup uploaded blobs
            _cleanup_task_uploaded_blobs(gcs_bucket, uploaded_blob_paths)
            raise

        telegram_warning = None
        try:
            assignee = db.get_user(id=assigned_to)
            assignee_chat_id = assignee[1] if assignee else None
            if assignee_chat_id:
                tag_label = TASK_TAG_LABELS.get(tag, 'Задача')
                requester_name = requester[2] if requester else 'Система'
                message = (
                    "<b>🆕 Новая задача</b>\n\n"
                    f"<b>Тип:</b> {_escape_telegram_html(tag_label, 60)}\n"
                    f"<b>Тема:</b> {_escape_telegram_html(subject, 220)}\n"
                    f"<b>От:</b> {_escape_telegram_html(requester_name, 80)}\n\n"
                    "<b>Откройте раздел «Задачи», чтобы посмотреть детали.</b>"
                )
                tg_response = _send_telegram_text_message(assignee_chat_id, message, parse_mode='HTML')
                if tg_response.status_code != 200:
                    error_detail = _get_telegram_error_text(tg_response)
                    telegram_warning = f"Task created, but Telegram notification failed: {error_detail}"
                    logging.error(f"Task Telegram API error: {error_detail}")
        except Exception as notify_error:
            telegram_warning = f"Task created, but Telegram notification failed: {str(notify_error)}"
            logging.error(f"Task Telegram notification error: {notify_error}")

        response_payload = {
            "status": "success",
            "message": "Task created successfully",
            "task_id": created.get("id")
        }
        if telegram_warning:
            response_payload["warning"] = telegram_warning

        return jsonify(response_payload), 201

    except Exception as e:
        logging.error(f"Error in handle_tasks: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/tasks/<int:task_id>/status', methods=['POST', 'OPTIONS'])
@require_api_key
def update_task_status(task_id):
    try:
        requester_id, requester, guard_response, guard_status = _task_route_guard()
        if guard_response is not None:
            return guard_response, guard_status

        content_type = (request.content_type or '').lower()
        is_multipart = 'multipart/form-data' in content_type

        if is_multipart:
            action = (request.form.get('action') or '').strip().lower()
            comment = (request.form.get('comment') or '').strip()
            completion_summary = (request.form.get('completion_summary') or '').strip()
            completion_files = request.files.getlist('files')
        else:
            data = request.get_json() or {}
            action = (data.get('action') or '').strip().lower()
            comment = (data.get('comment') or '').strip()
            completion_summary = (data.get('completion_summary') or '').strip()
            completion_files = []

        if action not in TASK_ALLOWED_ACTIONS:
            return jsonify({"error": "Invalid action"}), 400

        completion_attachments = []
        uploaded_blob_paths = []
        gcs_bucket = None

        if action == 'completed':
            try:
                completion_attachments, uploaded_blob_paths, gcs_bucket = _upload_task_attachments_to_gcs(
                    completion_files,
                    stage='result'
                )
            except ValueError as upload_error:
                return jsonify({"error": str(upload_error)}), 400
            except RuntimeError as upload_error:
                return jsonify({"error": str(upload_error)}), 500
            except Exception as upload_error:
                logging.error(f"Task completion attachment upload failed: {upload_error}")
                return jsonify({"error": "Failed to upload completion attachments"}), 500
        elif completion_files:
            return jsonify({"error": "Files can be attached only when completing a task"}), 400

        try:
            result = db.update_task_status(
                task_id=task_id,
                requester_id=requester_id,
                requester_role=requester[3],
                action=action,
                comment=comment,
                completion_summary=completion_summary if action == 'completed' else None,
                completion_attachments=completion_attachments
            )
        except ValueError as value_error:
            _cleanup_task_uploaded_blobs(gcs_bucket, uploaded_blob_paths)
            code = str(value_error)
            if code == 'TASK_NOT_FOUND':
                return jsonify({"error": "Task not found"}), 404
            if code in ('INVALID_ACTION', 'INVALID_TRANSITION'):
                return jsonify({"error": "Invalid task status transition"}), 400
            return jsonify({"error": code}), 400
        except PermissionError as permission_error:
            _cleanup_task_uploaded_blobs(gcs_bucket, uploaded_blob_paths)
            code = str(permission_error)
            if code in ('TASK_FORBIDDEN', 'ONLY_ASSIGNEE', 'ONLY_REVIEWER'):
                return jsonify({"error": "You do not have permission for this action"}), 403
            return jsonify({"error": code}), 403
        except Exception:
            _cleanup_task_uploaded_blobs(gcs_bucket, uploaded_blob_paths)
            raise

        telegram_warnings = []
        try:
            should_notify_participants = bool(result.get('history_id'))
            if should_notify_participants:
                task_ctx = _fetch_task_notification_context(task_id)
                if task_ctx:
                    actor_name = requester[2] if requester and len(requester) > 2 else 'Сотрудник'
                    task_subject = (task_ctx.get('subject') or f"Задача #{task_id}").strip()
                    recipients = _collect_task_notification_recipients(task_ctx, requester_id)

                    for recipient in recipients:
                        message_html = _build_task_status_notification_html(
                            action=action,
                            task_ctx=task_ctx,
                            actor_name=actor_name,
                            recipient_kind=recipient.get('kind') or 'participant',
                            comment=comment,
                            completion_summary=completion_summary,
                            completion_files_count=len(completion_attachments)
                        )
                        response = _send_telegram_text_message(
                            recipient.get('chat_id'),
                            message_html,
                            parse_mode='HTML'
                        )
                        if response.status_code != 200:
                            recipient_name = recipient.get('name') or recipient.get('kind') or 'получатель'
                            telegram_warnings.append(
                                f"Не удалось отправить уведомление ({recipient_name}): {_get_telegram_error_text(response)}"
                            )
                            continue

                        if action == 'completed':
                            telegram_warnings.extend(
                                _send_task_completion_attachments_to_telegram(
                                    recipient.get('chat_id'),
                                    task_subject,
                                    completion_attachments
                                )
                            )
        except Exception as notify_error:
            telegram_warnings.append(f"Ошибка отправки Telegram-уведомления: {notify_error}")

        action_messages = {
            'in_progress': 'Task moved to in progress',
            'completed': 'Task marked as completed',
            'accepted': 'Task accepted',
            'returned': 'Task returned for rework',
            'reopened': 'Task reopened'
        }
        response_payload = {
            "status": "success",
            "message": action_messages.get(action, 'Task status updated'),
            "task": result
        }
        if telegram_warnings:
            response_payload["warning"] = _truncate_for_telegram(" | ".join(telegram_warnings), 1000)
        return jsonify(response_payload), 200

    except Exception as e:
        logging.error(f"Error in update_task_status: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/tasks/attachments/<int:attachment_id>/download', methods=['GET', 'OPTIONS'])
@require_api_key
def download_task_attachment(attachment_id):
    try:
        requester_id, requester, guard_response, guard_status = _task_route_guard()
        if guard_response is not None:
            return guard_response, guard_status

        try:
            attachment = db.get_task_attachment_for_requester(
                attachment_id=attachment_id,
                requester_id=requester_id,
                requester_role=requester[3]
            )
        except PermissionError:
            return jsonify({"error": "You do not have access to this attachment"}), 403

        if not attachment:
            return jsonify({"error": "Attachment not found"}), 404

        if attachment.get('storage_type') == 'gcs':
            bucket_name = (attachment.get('gcs_bucket') or '').strip()
            blob_path = (attachment.get('gcs_blob_path') or '').strip()
            if not bucket_name or not blob_path:
                return jsonify({"error": "Attachment storage metadata is invalid"}), 500
            try:
                gcs_client = get_gcs_client()
                bucket = gcs_client.bucket(bucket_name)
                blob = bucket.blob(blob_path)
                if not blob.exists():
                    return jsonify({"error": "Attachment object not found in storage"}), 404
                file_bytes = blob.download_as_bytes()
            except Exception as storage_error:
                logging.error(f"Error downloading GCS task attachment: {storage_error}")
                return jsonify({"error": "Failed to download attachment from storage"}), 500

            return send_file(
                BytesIO(file_bytes),
                as_attachment=True,
                download_name=attachment['file_name'],
                mimetype=attachment.get('content_type') or 'application/octet-stream'
            )

        return send_file(
            BytesIO(attachment['file_data']),
            as_attachment=True,
            download_name=attachment['file_name'],
            mimetype=attachment.get('content_type') or 'application/octet-stream'
        )
    except Exception as e:
        logging.error(f"Error in download_task_attachment: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/monthly_report', methods=['GET'])
@require_api_key
def handle_monthly_report():
    try:
        month = request.args.get('month')
        user_id = request.args.get('')
        if not month:
            month = datetime.now().strftime('%Y-%m')
        
        # Validate month format
        try:
            month_start = datetime.strptime(month + '-01', '%Y-%m-%d')
        except ValueError:
            return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400
        
        logging.info(f"Generating monthly report for {month}")
        
        user_id = int(request.headers.get('X-User-Id'))
        user = db.get_user(id=user_id)
        if user[3]=="sv":
            # get_user returns tuple: (id, telegram_id, name, role, ... , status)
            if user[5] == "fired":
                return jsonify({"error": "Your status is fired. No report available."}), 403
            svs=[(user[0], user[2], "", "", "", user[5])]
        elif user[3]=="admin":
            svs = [sv for sv in db.get_supervisors() if len(sv) > 5 and sv[5] != "fired"]
        else:
            return jsonify({"error": "Only admin or SV can access to this report"}), 404

        if not svs:
            return jsonify({"error": "No supervisors found"}), 404

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})

        # Styles
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1,
            'align': 'center'
        })
        fio_format = workbook.add_format({'border': 1, 'bold': True, 'align': 'left'})
        int_format = workbook.add_format({'border': 1, 'num_format': '0', 'align': 'center'})
        float_format = workbook.add_format({'border': 1, 'num_format': '0.00', 'align': 'center'})
        # Score color formats
        red_format = workbook.add_format({'border': 1, 'num_format': '0.00', 'bg_color': '#FFCCCC', 'align': 'center'})
        yellow_format = workbook.add_format({'border': 1, 'num_format': '0.00', 'bg_color': '#FFFACD', 'align': 'center'})
        green_format = workbook.add_format({'border': 1, 'num_format': '0.00', 'bg_color': '#C6EFCE', 'align': 'center'})
        # Итоговые колонки
        total_format = workbook.add_format({'border': 1, 'bold': True, 'bg_color': '#E2EFDA', 'num_format': '0.00', 'align': 'center'})
        total_int_format = workbook.add_format({'border': 1, 'bold': True, 'bg_color': '#E2EFDA', 'num_format': '0', 'align': 'center'})

        for sv in svs:
            sv_id, sv_name = sv[0], sv[1]
            # Exclude supervisors with status 'fired' (should already be filtered above, but double check)
            if len(sv) > 5 and sv[5] == "fired":
                continue
            operators = db.get_operators_by_supervisor(sv_id)
            # Exclude operators with status 'fired'
            operators = [op for op in operators if op.get('status', '').lower() != 'fired']

            safe_sheet_name = sv_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
            worksheet = workbook.add_worksheet(safe_sheet_name)

            special_evaluator_id = 169
            special_table_header_row = len(operators) + 2

            # Headers
            headers = ['ФИО']
            for i in range(1, 21):
                headers.append(f'{i}')
            headers.append('Средний балл')
            headers.append('Кол-во оцененных звонков')

            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)

            for col, header in enumerate(headers):
                worksheet.write(special_table_header_row, col, header, header_format)

            for row_idx, op in enumerate(operators, start=1):
                op_id = op['id']
                op_name = op['name']

                # Get scores
                with db._get_cursor() as cursor:
                    cursor.execute("""
                        SELECT c.score
                        FROM calls c
                        JOIN (
                            SELECT phone_number, MAX(created_at) as max_date
                            FROM calls 
                            WHERE operator_id = %s AND month = %s AND is_draft = FALSE 
                            GROUP BY phone_number
                        ) lv ON c.phone_number = lv.phone_number AND c.created_at = lv.max_date
                        WHERE c.is_draft = FALSE
                        ORDER BY c.created_at ASC
                    """, (op_id, month))
                    scores = [row[0] for row in cursor.fetchall()]

                with db._get_cursor() as cursor:
                    cursor.execute("""
                        SELECT c.score
                        FROM calls c
                        JOIN (
                            SELECT phone_number, MAX(created_at) as max_date
                            FROM calls
                            WHERE operator_id = %s AND month = %s AND is_draft = FALSE AND evaluator_id = %s
                            GROUP BY phone_number
                        ) lv ON c.phone_number = lv.phone_number AND c.created_at = lv.max_date
                        WHERE c.is_draft = FALSE AND c.operator_id = %s AND c.month = %s AND c.evaluator_id = %s
                        ORDER BY c.created_at ASC
                    """, (op_id, month, special_evaluator_id, op_id, month, special_evaluator_id))
                    special_scores = [row[0] for row in cursor.fetchall()]

                count = len(scores)
                avg_score = sum(scores) / count if count > 0 else 0.0

                special_count = len(special_scores)
                special_avg_score = sum(special_scores) / special_count if special_count > 0 else 0.0

                # ФИО
                worksheet.write(row_idx, 0, op_name, fio_format)

                special_row_idx = special_table_header_row + row_idx
                worksheet.write(special_row_idx, 0, op_name, fio_format)

                # Оценки с цветом
                for col in range(1, 21):
                    if col-1 < count:
                        score = scores[col-1]
                        if score < 80:
                            fmt = red_format
                        elif score < 95:
                            fmt = yellow_format
                        else:
                            fmt = green_format
                        worksheet.write(row_idx, col, score, fmt)
                    else:
                        worksheet.write(row_idx, col, '', float_format)

                    if col-1 < special_count:
                        score = special_scores[col-1]
                        if score < 80:
                            fmt = red_format
                        elif score < 95:
                            fmt = yellow_format
                        else:
                            fmt = green_format
                        worksheet.write(special_row_idx, col, score, fmt)
                    else:
                        worksheet.write(special_row_idx, col, '', float_format)

                # Итоговые колонки
                worksheet.write(row_idx, 21, avg_score, total_format)
                worksheet.write(row_idx, 22, count, total_int_format)

                worksheet.write(special_row_idx, 21, special_avg_score, total_format)
                worksheet.write(special_row_idx, 22, special_count, total_int_format)

            worksheet.set_column('A:A', 30)
            for i in range(1, 21):
                worksheet.set_column(i, i, 12)
            worksheet.set_column(21, 21, 15)
            worksheet.set_column(22, 22, 20)

        workbook.close()
        output.seek(0)

        if output.getvalue():
            filename = f"Monthly_Report_{month}.xlsx"
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            logging.error("Report is empty")
            return jsonify({"error": "Generated report is empty"}), 500

    except Exception as e:
        logging.error(f"Error in monthly_report: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/admin/users_report', methods=['GET'])
@require_api_key
def get_users_report():
    try:
        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        if not requester or not _is_admin_role(requester[3]):
            return jsonify({"error": "Only admins can generate users report"}), 403
        
        filename, content = db.generate_users_report()
        if not filename or not content:
            return jsonify({"error": "Failed to generate report"}), 500
        
        return send_file(
            BytesIO(content),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logging.error(f"Error generating users report: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/call_evaluation/sv_request', methods=['POST'])
@require_api_key
def sv_request_call():
    """
    Супервайзер отправляет запрос на переоценку по конкретному звонку (call_id).
    Тело JSON: { "call_id": int, "comment": "..." }
    Заголовки: X-User-Id, X-API-Key (require_api_key обеспечит авторизацию)
    """
    try:
        data = request.get_json()
        if not data or 'call_id' not in data:
            return jsonify({"error": "Missing call_id"}), 400

        call_id = int(data['call_id'])
        comment = data.get('comment', '').strip()
        requester_id = int(request.headers.get('X-User-Id'))

        # Проверка: существует ли звонок и принадлежит ли оператор супервайзеру
        call = db.get_call_by_id(call_id)
        if not call:
            return jsonify({"error": "Call not found"}), 404

        # получить запись звонка с operator_id
        with db._get_cursor() as cursor:
            cursor.execute("SELECT operator_id, evaluator_id, month FROM calls WHERE id = %s", (call_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({"error": "Call not found"}), 404
            operator_id, evaluator_id, call_month = row

        operator = db.get_user(id=operator_id)
        if not operator:
            return jsonify({"error": "Operator not found"}), 404

        # Проверяем что requester — супервайзер этого оператора (или админ)
        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 403

        if not _is_admin_role(requester[3]) and operator[6] != requester_id:
            return jsonify({"error": "Only the operator's supervisor or admin can request reevaluation"}), 403

        # Сохраняем заявку в calls
        with db._get_cursor() as cursor:
            cursor.execute("""
                UPDATE calls
                SET sv_request = TRUE,
                    sv_request_comment = %s,
                    sv_request_by = %s,
                    sv_request_at = %s,
                    sv_request_approved = FALSE,
                    sv_request_approved_by = NULL,
                    sv_request_approved_at = NULL
                WHERE id = %s
            """, (comment, requester_id, datetime.utcnow(), call_id))

        # Уведомляем админа через Telegram с inline-кнопкой
        API_TOKEN = os.getenv('BOT_TOKEN')
        admin_chat_id = os.getenv('ADMIN_ID')
        if API_TOKEN and admin_chat_id:
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
            text = (
                f"⚠️ <b>Запрос на переоценку (от СВ)</b>\n\n"
                f"📞 Call ID: <b>{call_id}</b>\n"
                f"👤 Оператор ID: <b>{operator_id}</b> / {operator[2]}\n"
                f"📄 Месяц: <b>{call_month}</b>\n"
                f"📝 Комментарий: {comment or '-'}\n\n"
                f"Нажмите кнопку, чтобы одобрить переоценку и открыть звонок для переоценки."
            )
            reply_markup = {
                "inline_keyboard": [
                    [
                        {"text": "Одобрить переоценку ✅", "callback_data": f"approve_reval:{call_id}"}
                    ]
                ]
            }
            try:
                requests.post(telegram_url, json={
                    "chat_id": admin_chat_id,
                    "text": text,
                    "parse_mode": "HTML",
                    "reply_markup": json.dumps(reply_markup)
                }, timeout=10)
            except Exception as e:
                logging.error(f"Failed to send telegram sv_request notification: {e}")

        return jsonify({"status": "success", "message": "sv_request created"}), 200

    except Exception as e:
        logging.error(f"Error in sv_request_call: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/call_versions/<int:call_id>', methods=['GET'])
@require_api_key
def get_call_versions(call_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        if not requester or not _is_privileged_role(requester[3]):
            return jsonify({"error": "Forbidden: only admin or supervisor can access call versions"}), 403

        head_call = db.get_call_by_id(call_id)
        if not head_call:
            return jsonify({"error": "Call not found"}), 404

        if not _ensure_call_access_for_requester(head_call["operator_id"], requester, requester_id):
            return jsonify({"error": "Unauthorized to access this call versions"}), 403

        with db._get_cursor() as cursor:
            # Получаем все версии оценки, начиная с текущей и идя назад по previous_version_id
            versions = []
            current_id = call_id
            
            while current_id:
                cursor.execute("""
                    SELECT 
                        c.id, c.score, c.comment, c.phone_number, c.month, 
                        c.audio_path, c.created_at, c.is_correction,
                        u.name as evaluator_name,
                        c.previous_version_id  
                    FROM calls c
                    JOIN users u ON c.evaluator_id = u.id
                    WHERE c.id = %s
                """, (current_id,))
                version = cursor.fetchone()
                if not version:
                    break
                    
                versions.append({
                    "id": version[0],
                    "score": float(version[1]),
                    "comment": version[2],
                    "phone_number": version[3],
                    "month": version[4],
                    "audio_path": version[5],
                    "evaluation_date": version[6].strftime('%Y-%m-%d %H:%M'),
                    "is_correction": version[7],
                    "evaluator_name": version[8],
                    "audio_url": None  # Будет заполнено позже
                })
                
                # Переходим к предыдущей версии
                current_id = version[9]  # previous_version_id

            # Генерируем signed URLs для аудиофайлов
            gcs_client = get_gcs_client()
            for version in versions:
                if version['audio_path']:
                    try:
                        path_parts = version['audio_path'].split('/', 1)
                        if len(path_parts) == 2:
                            bucket_name, blob_path = path_parts
                            bucket = gcs_client.bucket(bucket_name)
                            blob = bucket.blob(blob_path)
                            if blob.exists():
                                version['audio_url'] = blob.generate_signed_url(
                                    version="v4",
                                    expiration=timedelta(minutes=15),
                                    method="GET",
                                    response_type='audio/mpeg'
                                )
                    except Exception as e:
                        logging.error(f"Error generating signed URL for version {version['id']}: {e}")

            return jsonify({"status": "success", "versions": versions})
    except Exception as e:
        logging.error(f"Error fetching call versions: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/call_evaluations/<int:evaluation_id>', methods=['DELETE'])
@require_api_key
def delete_draft_evaluation(evaluation_id):
    try:
        # Now delete from imported_calls by id (admin or supervisor of the operator)
        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 403

        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT operator_id, status FROM imported_calls
                WHERE id = %s
            """, (evaluation_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({"error": "Imported call not found"}), 404

            operator_id, status = row

            # Prevent deleting already evaluated calls
            if status == 'evaluated':
                return jsonify({"error": "Cannot delete evaluated imported call"}), 400

            # Authorization: admins can delete anything; supervisors can delete calls for their operators
            allowed = False
            try:
                role = requester[3]
            except Exception:
                role = None

            if _is_admin_role(role):
                allowed = True

            if not allowed:
                return jsonify({"error": "Unauthorized to delete this imported call"}), 403

            cursor.execute("""
                DELETE FROM imported_calls WHERE id = %s
            """, (evaluation_id,))

        return jsonify({"status": "success", "message": "Imported call deleted"}), 200
    except Exception as e:
        logging.error(f"Error deleting draft evaluation: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/audio/<int:evaluation_id>', methods=['GET'])
@require_api_key
def get_audio_file(evaluation_id):
    try:
        requester_id = getattr(g, 'user_id', None)
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester = db.get_user(id=requester_id)
        requester_role = _normalize_user_role(requester[3]) if requester else ''
        if not requester or (not _is_admin_role(requester_role) and requester_role not in ('sv', 'operator')):
            return jsonify({"error": "Audio is not available for this role"}), 403

        if requester_role == 'operator':
            session_id = _current_session_id_from_access_token()
            if not _is_sensitive_access_unlocked(requester_id, session_id):
                return jsonify({
                    "error": "Sensitive data access requires QR confirmation in current session",
                    "code": "SENSITIVE_ACCESS_REQUIRED"
                }), 403

        call = db.get_call_by_id(evaluation_id)
        if not call or not call['audio_path']:
            return jsonify({"error": "Audio file not found"}), 404

        if not _ensure_call_access_for_requester(call["operator_id"], requester, requester_id):
            return jsonify({"error": "Unauthorized to access this audio"}), 403

        gcs_client = get_gcs_client()

        # Разбиваем путь: 'bucket_name/folder/file.mp3'
        path_parts = call['audio_path'].split('/', 1)
        if len(path_parts) != 2:
            return jsonify({"error": "Invalid GCS path format"}), 400

        bucket_name, blob_path = path_parts
        bucket = gcs_client.bucket(bucket_name)
        blob = bucket.blob(blob_path)

        if not blob.exists():
            return jsonify({"error": "Audio file not found in GCS"}), 404

        # Генерируем временный Signed URL
        signed_url = blob.generate_signed_url(
            version="v4",
            expiration=timedelta(minutes=15),
            method="GET",
            response_type='audio/mpeg'
        )

        return jsonify({"status": "success", "url": signed_url})
    except Exception as e:
        logging.error(f"Error generating signed audio URL: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/shuffle', methods=['POST'])
@require_api_key
def shuffle_imported_calls():
    """
    Импорт звонков для распределения по операторам (admin shuffle).
    Формат входных данных:
    {
        "month": "2025-11",
        "distribution": [
            {
                "operator": "ФИО",
                "desired": 3,
                "available": 17,
                "calls": [
                    {
                        "id": "uuid",
                        "datetimeRaw": "22.10.2025 13:56:59",
                        "phone": "77084919987",
                        "durationSec": 136.28
                    },
                    ...
                ]
            },
            ...
        ]
    }
    """
    try:
        payload = request.get_json(force=True)
    except Exception:
        return jsonify({"error": "Invalid or missing JSON body"}), 400

    if not payload or "month" not in payload or "distribution" not in payload:
        return jsonify({"error": "Missing required fields: month and distribution"}), 400

    try:
        # ID импортирующего администратора / супервизора
        importer_id = getattr(request, "user_id", None)
        result = db.import_calls_from_distribution(payload, importer_id=importer_id)

        return jsonify({
            "status": "ok",
            "month": payload.get("month"),
            "imported": result.get("imported"),
            "updated": result.get("updated"),
            "skipped": result.get("skipped"),
            "missing_operators": result.get("missing_operators"),
            "errors": result.get("errors"),
            "timestamp": datetime.now().isoformat()
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

@app.route('/api/call_evaluation', methods=['POST'])
@require_api_key
def receive_call_evaluation():
    try:
        if not request.form:
            return jsonify({"error": "Missing form data"}), 400

        required_fields = ['evaluator', 'operator', 'phone_number', 'appeal_date', 'score', 'comment', 'month', 'is_draft']
        missing_fields = [field for field in required_fields if field not in request.form]
        if missing_fields:
            return jsonify({"error": f"Missing required fields: {', '.join(missing_fields)}"}), 400

        evaluator_name = request.form['evaluator']
        operator_name = request.form['operator']
        phone_number = request.form['phone_number']
        appeal_date = request.form['appeal_date']
        score = float(request.form['score'])
        comment = request.form['comment']
        month = request.form['month'] or datetime.now().strftime('%Y-%m')
        is_draft = request.form['is_draft'].lower() == 'true'
        comment_visible_to_operator_raw = str(request.form.get('comment_visible_to_operator', 'true')).strip().lower()
        comment_visible_to_operator = comment_visible_to_operator_raw in ('1', 'true', 'yes', 'on')
        scores = json.loads(request.form.get('scores', '[]'))
        criterion_comments = json.loads(request.form.get('criterion_comments', '[]'))
        direction_id = request.form.get('direction')
        previous_version_id = request.form.get('previous_version_id')
        is_correction = request.form.get('is_correction', 'false').lower() == 'true'

        evaluator = db.get_user(name=evaluator_name)
        operator = db.get_user(name=operator_name)
        if not evaluator or not operator:
            return jsonify({"error": "Evaluator or operator not found"}), 404

        if is_correction:
            requester_id = int(request.headers.get('X-User-Id'))
            requester = db.get_user(id=requester_id)
            if not requester:
                return jsonify({"error": "Requester not found"}), 403

            # admins всегда могут делать переоценки
            if _is_admin_role(requester[3]):
                allowed = True
            else:
                # non-admin может переоценивать ТОЛЬКО если:
                # - передан previous_version_id
                # - previous call exists AND sv_request == TRUE AND sv_request_approved == TRUE
                # - и sv_request_by == requester_id (т.е. именно этот СВ отправил запрос)
                allowed = False
                if previous_version_id:
                    try:
                        prev_id = int(previous_version_id)
                    except Exception:
                        prev_id = None
                    if prev_id:
                        try:
                            with db._get_cursor() as cursor:
                                cursor.execute("SELECT sv_request, sv_request_approved, sv_request_by FROM calls WHERE id = %s", (prev_id,))
                                prev = cursor.fetchone()
                        except Exception as e:
                            logging.error(f"Error checking previous call for re-eval permissions: {e}")
                            prev = None

                        if prev:
                            sv_request_flag = bool(prev[0])
                            sv_request_approved_flag = bool(prev[1])
                            sv_request_by_id = prev[2]
                            if sv_request_flag and sv_request_approved_flag and sv_request_by_id == requester_id:
                                allowed = True

            if not allowed:
                return jsonify({"error": "Only admins or the supervisor who requested and was approved can perform re-evaluations"}), 403


        audio_path = None
        audio_data = None
        blob_path = None
        bucket_name = None
        has_new_audio = False

        if 'audio_file' in request.files:
            file = request.files['audio_file']
            if file and file.filename:
                has_new_audio = True
                try:
                    audio_data = file.read()
                except Exception as e:
                    logging.error(f"Error reading audio file: {e}")
                    return jsonify({"error": f"Failed to process audio file: {str(e)}"}), 500
                filename = secure_filename(f"{uuid.uuid4()}.mp3")
                bucket_name = os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET')
                upload_folder = os.getenv('UPLOAD_FOLDER', 'Uploads/')
                blob_path = f"{upload_folder}{filename}"
                audio_path = "my-app-audio-uploads/" + blob_path

        elif is_correction and previous_version_id:
            try:
                with db._get_cursor() as cursor:
                    cursor.execute("SELECT audio_path FROM calls WHERE id = %s", (previous_version_id,))
                    result = cursor.fetchone()
                    if result and result[0]:
                        audio_path = result[0]
            except Exception as e:
                logging.error(f"Error retrieving audio_path from previous version: {e}")
                return jsonify({"error": f"Failed to retrieve audio from previous version: {str(e)}"}), 500

        evaluation_id = db.add_call_evaluation(
            evaluator_id=evaluator[0],
            operator_id=operator[0],
            phone_number=phone_number,
            score=score,
            comment=comment,
            comment_visible_to_operator=comment_visible_to_operator,
            month=month,
            audio_path=audio_path,
            is_draft=is_draft,
            scores=scores,
            criterion_comments=criterion_comments,
            direction_id=direction_id,
            is_correction=is_correction,
            previous_version_id=previous_version_id if previous_version_id else None,
            appeal_date=appeal_date
        )

        if has_new_audio or (not is_draft and audio_path):
            threading.Thread(target=background_upload_and_notify, args=(
                audio_data, bucket_name, blob_path, evaluation_id, is_draft,
                evaluator[2], operator[2], month, phone_number, score, comment,
                is_correction, previous_version_id, audio_path if not has_new_audio else None,
                appeal_date
            )).start()
        elif not is_draft:
            threading.Thread(target=send_telegram_notification, args=(
                evaluator[2], operator[2], month, phone_number, score, comment,
                is_correction, previous_version_id, None, appeal_date
            )).start()

        return jsonify({"status": "success", "evaluation_id": evaluation_id}), 200
    except Exception as e:
        logging.error(f"Error processing call evaluation: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

def background_upload_and_notify(audio_data, bucket_name, blob_path, evaluation_id, is_draft,
                                 evaluator_name, operator_name, month, phone_number, score, comment,
                                 is_correction, previous_version_id, existing_audio_path, appeal_date):
    audio_path = existing_audio_path or ("my-app-audio-uploads/" + blob_path if blob_path else None)
    upload_success = False
    if audio_data:
        try:
            client = get_gcs_client()
            bucket = client.bucket(bucket_name)
            blob = bucket.blob(blob_path)
            blob.upload_from_string(audio_data, content_type='audio/mpeg')
            upload_success = True
        except Exception as e:
            logging.error(f"Error uploading file to GCS in background: {e}")
            try:
                with db._get_connection() as conn:
                    with conn.cursor() as cursor:
                        cursor.execute("UPDATE calls SET audio_path = NULL WHERE id = %s", (evaluation_id,))
                        conn.commit()
            except Exception as update_e:
                logging.error(f"Error updating DB after failed upload: {update_e}")
            audio_path = None
    if not is_draft:
        send_telegram_notification(evaluator_name, operator_name, month, phone_number, score, comment,
                                   is_correction, previous_version_id, audio_path if (upload_success or existing_audio_path) else None, appeal_date)

def send_telegram_notification(evaluator_name, operator_name, month, phone_number, score, comment, is_correction, previous_version_id, audio_path, appeal_date):
    try:
        API_TOKEN = os.getenv('BOT_TOKEN')
        admin = os.getenv('ADMIN_ID')
        if not admin:
            return

        message = (
            f"👤 Оценивающий: <b>{evaluator_name}</b>\n"
            f"📋 Оператор: <b>{operator_name}</b>\n"
            f"📄 За месяц: <b>{month}</b>\n"
            f"📱 Номер телефона: <b>{phone_number}</b>\n"
            f"📅 Дата обращения: <b>{' '.join(appeal_date.split('T'))}</b>\n"
            f"💯 Оценка: <b>{score}</b>\n"
        )
        if is_correction:
            message += f"🔄 <b>Переоценка звонка (ID предыдущей версии: {previous_version_id})</b>\n"
        if score < 100 and comment:
            message += f"\n💬 Комментарий: \n{comment}\n"

        audio_signed_url = None
        if audio_path:
            try:
                client = get_gcs_client()
                path_parts = audio_path.split('/', 1)
                if len(path_parts) == 2:
                    bucket_name, blob_path = path_parts
                    bucket = client.bucket(bucket_name)
                    blob = bucket.blob(blob_path)
                    if blob.exists():
                        expiration = datetime.utcnow() + timedelta(minutes=15)
                        audio_signed_url = blob.generate_signed_url(
                            expiration=expiration,
                            method='GET',
                            version='v4'
                        )
            except Exception as e:
                logging.error(f"Error generating signed URL for audio: {e}")

        telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendAudio" if audio_signed_url else f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
        payload = {
            "chat_id": admin,
            "parse_mode": "HTML"
        }
        if audio_signed_url:
            payload["audio"] = audio_signed_url
            payload["caption"] = f"📞 <b>{'Переоценка звонка' if is_correction else 'Оценка звонка'}</b>\n" + message
        else:
            payload["text"] = f"💬 <b>{'Переоценка чата' if is_correction else 'Оценка чата'}</b>\n" + message

        response = requests.post(telegram_url, json=payload, timeout=10)
        if response.status_code != 200:
            error_detail = response.json().get('description', 'Unknown error')
            logging.error(f"Telegram API error: {error_detail}")
    except Exception as e:
        logging.error(f"Error sending Telegram notification: {e}")

@app.route('/api/admin/add_operator', methods=['POST'])
@require_api_key
def add_operator():
    try:
        data = request.get_json()
        required_fields = ['name', 'telegram_id', 'supervisor_id']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        name = data['name']
        telegram_id = int(data['telegram_id'])
        supervisor_id = int(data['supervisor_id'])
        
        # Проверяем существование пользователя
        if db.get_user(telegram_id=telegram_id):
            return jsonify({"error": "Operator with this Telegram ID already exists"}), 400
        
        # Создаем оператора
        operator_id = db.create_user(
            telegram_id=telegram_id,
            name=name,
            role='operator',
            supervisor_id=supervisor_id
        )
        
        telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
        payload = {
            "chat_id": telegram_id,
            "text": f"Вы добавлены как оператор в команду <b>успешно✅</b>",
            "parse_mode": "HTML"
        }
        response = requests.post(telegram_url, json=payload, timeout=10)
        if response.status_code != 200:
            error_detail = response.json().get('description', 'Unknown error')
            logging.error(f"Telegram API error: {error_detail}")
        
        return jsonify({"status": "success", "message": f"Operator {name} added", "id": operator_id})
    except Exception as e:
        logging.error(f"Error adding operator: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/user/toggle_active', methods=['POST'])
@require_api_key
def toggle_user_active():
    try:
        data = request.get_json()
        if not data or 'status' not in data:
            return jsonify({"error": "Missing status field"}), 400

        new_status = data['status']
        allowed_statuses = {"active", "break", "training", "inactive", "tech","iesigning"}
        if new_status not in allowed_statuses:
            return jsonify({"error": f"Invalid status. Allowed: {', '.join(allowed_statuses)}"}), 400

        user_id = int(request.headers.get('X-User-Id'))
        user = db.get_user(id=user_id)
        if not user:
            return jsonify({"error": "Unauthurized"}), 403

        # здесь предполагаем, что user[10] хранит BOOLEAN (is_active)
        current_active = user[10]
        new_active_bool = True if new_status == "active" else False

        # если статус active ↔ True/False не меняется — нет смысла обновлять
        if current_active == new_active_bool:
            # но дополнительно можно проверить, не был ли изменён текстовый статус (например break/training)
            last_log_status = db.get_last_activity_status(user_id)  # нужна функция в db
            if last_log_status == new_status:
                return jsonify({"status": "unchanged", "message": "Status is already set to the requested value"})

        # обновляем users.is_active (bool)
        success = db.set_user_active(user_id, new_status)
        if not success:
            return jsonify({"error": "Failed to update user active flag"}), 500

        # записываем строковый статус в логи
        log_success = db.log_activity(user_id, new_status)
        if not log_success:
            logging.warning("User active flag updated but logging failed")
            return jsonify({
                "status": "partial_success",
                "message": "User active flag updated, but logging failed"
            }), 500

        return jsonify({"status": "success", "message": f"Status updated to '{new_status}'"}), 200

    except Exception as e:
        logging.error(f"Error toggling status: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/active_operators', methods=['GET'])
@require_api_key
def get_active_operators():
    try:
        direction_name = request.args.get("direction")  # например: /api/active_operators?direction=Sales
        operators = db.get_active_operators(direction_name)
        return jsonify({"status": "success", "operators": operators})
    except Exception as e:
        logging.error(f"Error fetching active operators: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500
    
@app.route('/api/report/monthly', methods=['GET'])
@require_api_key
def get_monthly_report():
    try:
        supervisor_id = request.args.get('supervisor_id')
        month = request.args.get('month')
        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404

        requester_role = _normalize_user_role(requester[3])
        if not supervisor_id:
            if _is_supervisor_role(requester_role):
                supervisor_id = requester_id
            else:
                return jsonify({"error": "supervisor_id required"}), 400
        else:
            supervisor_id = int(supervisor_id)

        if (not _is_admin_role(requester_role)) and (not _is_supervisor_role(requester_role) or supervisor_id != requester_id):
            return jsonify({"error": "Unauthorized to access this report"}), 403

        logging.info("Начало генерации отчета")

        filename, content = db.generate_monthly_report(supervisor_id, month)
        if not filename or not content:
            return jsonify({"error": "Failed to generate report"}), 500
        
        return send_file(
            BytesIO(content),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logging.error(f"Error generating monthly report: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

MONTH_RE = re.compile(r'^\d{4}-\d{2}$')

def build_trainings_map(trainings_list):
    """
    Преобразует список тренингов в структуру { operator_id: { dayNum: [training...] } }.
    Ожидается, что у тренинга есть поле 'operator_id' и 'date' в формате 'YYYY-MM-DD'.
    """
    tmap = {}
    if not trainings_list:
        return tmap
    for t in trainings_list:
        op = t.get('operator_id')
        if op is None:
            continue
        dt = t.get('date') or t.get('date_str') or t.get('day')
        day = None
        if isinstance(dt, str) and len(dt) >= 10:
            try:
                day = int(dt[8:10])
            except Exception:
                day = None
        elif isinstance(dt, int):
            day = dt
        if day is None:
            # пропускаем записи без понятного дня
            continue
        tmap.setdefault(op, {}).setdefault(day, []).append(t)
    return tmap

def build_technical_issues_map(issues_list):
    """
    Преобразует список техсбоев в структуру { operator_id: { dayNum: [issue...] } }.
    Ожидается, что у записи есть поле 'operator_id' и 'date' в формате 'YYYY-MM-DD'.
    """
    imap = {}
    if not issues_list:
        return imap
    for item in issues_list:
        op = item.get('operator_id')
        if op is None:
            continue
        dt = item.get('date') or item.get('issue_date') or item.get('day')
        day = None
        if isinstance(dt, str) and len(dt) >= 10:
            try:
                day = int(dt[8:10])
            except Exception:
                day = None
        elif isinstance(dt, int):
            day = dt
        if day is None:
            continue
        imap.setdefault(op, {}).setdefault(day, []).append(item)
    return imap

def build_offline_activities_map(activities_list):
    """
    Преобразует список офлайн-активностей в структуру { operator_id: { dayNum: [activity...] } }.
    Ожидается, что у записи есть поле 'operator_id' и 'date' в формате 'YYYY-MM-DD'.
    """
    amap = {}
    if not activities_list:
        return amap
    for item in activities_list:
        op = item.get('operator_id')
        if op is None:
            continue
        dt = item.get('date') or item.get('activity_date') or item.get('day')
        day = None
        if isinstance(dt, str) and len(dt) >= 10:
            try:
                day = int(dt[8:10])
            except Exception:
                day = None
        elif isinstance(dt, int):
            day = dt
        if day is None:
            continue
        amap.setdefault(op, {}).setdefault(day, []).append(item)
    return amap

@app.route('/api/report/monthly_hours', methods=['GET'])
@require_api_key
def get_monthly_report_hours():
    try:
        supervisor_id = request.args.get('supervisor_id')
        month = request.args.get('month')
        if not month or not MONTH_RE.match(month):
            return jsonify({"error": "month required in format YYYY-MM"}), 400

        # requester
        try:
            requester_id = int(request.headers.get('X-User-Id'))
        except Exception:
            return jsonify({"error": "X-User-Id header required"}), 400

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "Requester not found"}), 404

        # роль в requester ожидается в requester[3] как в вашем примере
        role = _normalize_user_role(requester[3]) 

        # Если supervisor_id не указан — формируем общий отчёт для всех операторов .
        # Если supervisor_id указан или requester — sv без param, формируем отчёт по конкретному СВ.
        generate_all = False
        if not supervisor_id:
                generate_all = True
        else:
            try:
                supervisor_id = int(supervisor_id)
            except ValueError:
                return jsonify({"error": "supervisor_id must be integer"}), 400

        # проверка прав для случая конкретного SV: admin или сам SV
        if not generate_all:
            if not _is_admin_role(role) and not _is_supervisor_role(role):
                return jsonify({"error": "Unauthorized to access this report"}), 403

        logging.info("Начало генерации отчета: supervisor_id=%s month=%s generate_all=%s", supervisor_id, month, generate_all)

        try:
            if generate_all:
                operators = db.get_daily_hours_for_all_month(month)
            else:
                operators = db.get_daily_hours_by_supervisor_month(supervisor_id, month)
        except Exception as e:
            logging.exception("Ошибка получения operators из db")
            return jsonify({"error": f"Ошибка получения операторов: {str(e)}"}), 500

        try:
            if generate_all:
                trainings_list = db.get_trainings(None, month)
            else:
                trainings_list = db.get_trainings(supervisor_id, month)
        except Exception as e:
            logging.exception("Ошибка получения trainings из db")
            trainings_list = []

        trainings_map = build_trainings_map(trainings_list)

        try:
            y, m = map(int, month.split('-'))
            month_last_day = calendar.monthrange(y, m)[1]
            date_from = f"{month}-01"
            date_to = f"{month}-{str(month_last_day).zfill(2)}"
            technical_result = db.get_operator_technical_issues(
                requester_id=requester_id,
                requester_role=role,
                date_from=date_from,
                date_to=date_to,
                limit=5000,
                offset=0
            )
            technical_items_raw = technical_result.get('items', []) if isinstance(technical_result, dict) else []
        except Exception:
            logging.exception("Ошибка получения technical issues из db")
            technical_items_raw = []

        visible_operator_ids = set()
        try:
            operators_list_for_scope = operators.get("operators", []) if isinstance(operators, dict) else []
            visible_operator_ids = {
                int(op.get("operator_id"))
                for op in operators_list_for_scope
                if op and op.get("operator_id") is not None
            }
        except Exception:
            visible_operator_ids = set()

        if visible_operator_ids:
            technical_items = []
            for item in technical_items_raw:
                try:
                    if int(item.get("operator_id")) in visible_operator_ids:
                        technical_items.append(item)
                except Exception:
                    continue
        else:
            technical_items = technical_items_raw

        technical_issues_map = build_technical_issues_map(technical_items)

        try:
            y, m = map(int, month.split('-'))
            month_last_day = calendar.monthrange(y, m)[1]
            date_from = f"{month}-01"
            date_to = f"{month}-{str(month_last_day).zfill(2)}"
            offline_result = db.get_operator_offline_activities(
                requester_id=requester_id,
                requester_role=role,
                date_from=date_from,
                date_to=date_to,
                limit=5000,
                offset=0
            )
            offline_items_raw = offline_result.get('items', []) if isinstance(offline_result, dict) else []
        except Exception:
            logging.exception("Ошибка получения offline activities из db")
            offline_items_raw = []

        if visible_operator_ids:
            offline_items = []
            for item in offline_items_raw:
                try:
                    if int(item.get("operator_id")) in visible_operator_ids:
                        offline_items.append(item)
                except Exception:
                    continue
        else:
            offline_items = offline_items_raw

        offline_activities_map = build_offline_activities_map(offline_items)

        if generate_all:
            filename, content = db.generate_excel_report_all_operators_from_view(
                operators,
                trainings_map,
                technical_issues_map,
                month,
                offline_activities_map=offline_activities_map
            )
        else:
            filename, content = db.generate_excel_report_from_view(
                operators,
                trainings_map,
                technical_issues_map,
                month,
                offline_activities_map=offline_activities_map
            )
            
        if not filename or not content:
            logging.error("Генерация отчёта вернула пустой результат")
            return jsonify({"error": "Failed to generate report"}), 500

        return send_file(
            BytesIO(content),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.exception("Error generating monthly report")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/trainings', methods=['GET'])
@require_api_key
def get_trainings():
    
    try:
        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        if not requester:
            logging.warning(f"User not found: {requester_id}")
            return jsonify({"error": "User not found"}), 404

        month = request.args.get('month')
        if month:
            try:
                datetime.strptime(month, '%Y-%m')
            except ValueError:
                logging.warning(f"Invalid month format: {month}")
                return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400

        role = _normalize_user_role(requester[3])
        if not (_is_admin_role(role) or role in ('sv', 'operator')):
            logging.warning(f"Unauthorized role: {role}")
            return jsonify({"error": "Unauthorized"}), 403

        # Allow optional ?id=<supervisor_id> to fetch trainings for a specific supervisor
        # If id is absent, default to requester_id (own data)
        user_param = request.args.get('id')
        if user_param and str(user_param).isdigit():
            target_id = int(user_param)
        else:
            target_id = requester_id

        trainings = db.get_trainings(target_id, month)

        logging.info(f"Trainings fetched for role {role}, month: {month}, by user {requester_id}")
        return jsonify({"status": "success", "trainings": trainings}), 200
    except Exception as e:
        logging.error(f"Error fetching trainings: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/trainings', methods=['POST'])
@require_api_key
def add_training():
    
    try:
        data = request.get_json()
        required_fields = ['operator_id', 'date', 'start_time', 'end_time', 'reason']
        if not data or not all(field in data for field in required_fields):
            logging.warning(f"Missing required fields in add_training: {data}")
            return jsonify({"error": "Missing required fields"}), 400

        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        if not requester:
            logging.warning(f"User not found: {requester_id}")
            return jsonify({"error": "User not found"}), 404

        # Проверка прав: админ или супервайзер оператора
        operator = db.get_user(id=data['operator_id'])
        if not operator:
            logging.warning(f"Operator not found: {data['operator_id']}")
            return jsonify({"error": "Operator not found"}), 404
        
        if not _is_admin_role(requester[3]) and not _is_supervisor_role(requester[3]):
            logging.warning(f"Unauthorized attempt to add training by user {requester_id}")
            return jsonify({"error": "Unauthorized"}), 403

        # Валидация данных
        try:
            datetime.strptime(data['date'], '%Y-%m-%d')
            datetime.strptime(data['start_time'], '%H:%M')
            datetime.strptime(data['end_time'], '%H:%M')
        except ValueError:
            logging.warning(f"Invalid date/time format in add_training: {data}")
            return jsonify({"error": "Invalid date or time format"}), 400

        allowed_reasons = [
            "Обратная связь", "Собрание", "Тех. сбой", "Мотивационная беседа",
            "Дисциплинарный тренинг", "Тренинг по качеству. Разбор ошибок",
            "Тренинг по качеству. Объяснение МШ", "Тренинг по продукту",
            "Мониторинг", "Практика в офисе таксопарка", "Другое"
        ]
        if data['reason'] not in allowed_reasons:
            logging.warning(f"Invalid training reason: {data['reason']}")
            return jsonify({"error": "Invalid training reason"}), 400

        training_id = db.add_training(
            operator_id=data['operator_id'],
            training_date=data['date'],
            start_time=data['start_time'],
            end_time=data['end_time'],
            reason=data['reason'],
            comment=data.get('comment'),
            created_by=requester_id,
            count_in_hours=data.get("count_in_hours", True)
        )
        logging.info(f"Training added: ID {training_id} for operator {data['operator_id']}")
        return jsonify({"status": "success", "id": training_id}), 201
    except Exception as e:
        logging.error(f"Error adding training: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/trainings/<int:training_id>', methods=['PUT'])
@require_api_key
def update_training(training_id):
    
    try:
        data = request.get_json()
        if not data:
            logging.warning("No data provided for update_training")
            return jsonify({"error": "No data provided"}), 400

        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        if not requester:
            logging.warning(f"User not found: {requester_id}")
            return jsonify({"error": "User not found"}), 404

        # Проверка прав: админ или создатель тренинга
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT created_by FROM trainings WHERE id = %s
            """, (training_id,))
            training = cursor.fetchone()
            if not training:
                logging.warning(f"Training not found: {training_id}")
                return jsonify({"error": "Training not found"}), 404

            if not _is_admin_role(requester[3]) and not _is_supervisor_role(requester[3]):
                logging.warning(f"Unauthorized attempt to update training {training_id} by user {requester_id}")
                return jsonify({"error": "Unauthorized"}), 403

        # Валидация данных, если они предоставлены
        if 'date' in data:
            try:
                datetime.strptime(data['date'], '%Y-%m-%d')
            except ValueError:
                logging.warning(f"Invalid date format: {data['date']}")
                return jsonify({"error": "Invalid date format"}), 400
        if 'start_time' in data:
            try:
                datetime.strptime(data['start_time'], '%H:%M')
            except ValueError:
                logging.warning(f"Invalid start_time format: {data['start_time']}")
                return jsonify({"error": "Invalid start time format"}), 400
        if 'end_time' in data:
            try:
                datetime.strptime(data['end_time'], '%H:%M')
            except ValueError:
                logging.warning(f"Invalid end_time format: {data['end_time']}")
                return jsonify({"error": "Invalid end time format"}), 400
        if 'reason' in data:
            allowed_reasons = [
                "Обратная связь", "Собрание", "Тех. сбой", "Мотивационная беседа",
                "Дисциплинарный тренинг", "Тренинг по качеству. Разбор ошибок",
                "Тренинг по качеству. Объяснение МШ", "Тренинг по продукту",
                "Мониторинг", "Практика в офисе таксопарка", "Другое"
            ]
            if data['reason'] not in allowed_reasons:
                logging.warning(f"Invalid training reason: {data['reason']}")
                return jsonify({"error": "Invalid training reason"}), 400

        success = db.update_training(
            training_id=training_id,
            training_date=data.get('date'),
            start_time=data.get('start_time'),
            end_time=data.get('end_time'),
            reason=data.get('reason'),
            comment=data.get('comment'),
            count_in_hours=data.get('count_in_hours') 
        )
        if not success:
            logging.warning(f"Failed to update training {training_id}")
            return jsonify({"error": "Failed to update training"}), 500

        logging.info(f"Training updated: ID {training_id}")
        return jsonify({"status": "success", "message": "Training updated"}), 200
    except Exception as e:
        logging.error(f"Error updating training {training_id}: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/trainings/<int:training_id>', methods=['DELETE'])
@require_api_key
def delete_training(training_id):
    try:
        requester_id = int(request.headers.get('X-User-Id'))
        requester = db.get_user(id=requester_id)
        if not requester:
            logging.warning(f"User not found: {requester_id}")
            return jsonify({"error": "User not found"}), 404

        # Проверка прав: админ или создатель тренинга
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT created_by FROM trainings WHERE id = %s
            """, (training_id,))
            training = cursor.fetchone()
            if not training:
                logging.warning(f"Training not found: {training_id}")
                return jsonify({"error": "Training not found"}), 404
            
            if not _is_admin_role(requester[3]) and not _is_supervisor_role(requester[3]):
                logging.warning(f"Unauthorized attempt to delete training {training_id} by user {requester_id}")
                return jsonify({"error": "Unauthorized"}), 403

        success = db.delete_training(training_id)
        if not success:
            logging.warning(f"Failed to delete training {training_id}")
            return jsonify({"error": "Failed to delete training"}), 500

        logging.info(f"Training deleted: ID {training_id}")
        return jsonify({"status": "success", "message": "Training deleted"}), 200
    except Exception as e:
        logging.error(f"Error deleting training {training_id}: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/technical_issues/reasons', methods=['GET'])
@require_api_key
def technical_issue_reasons():
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        requester_role = _normalize_management_role(requester[3])
        if not (_is_admin_role(requester_role) or _is_supervisor_role(requester_role)):
            return jsonify({"error": "Forbidden"}), 403

        return jsonify({
            "status": "success",
            "reasons": list(TECHNICAL_ISSUE_REASONS)
        }), 200
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error fetching technical issue reasons: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/technical_issues', methods=['POST'])
@require_api_key
def create_technical_issue():
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        requester_role = _normalize_management_role(requester[3])
        if not (_is_admin_role(requester_role) or _is_supervisor_role(requester_role)):
            return jsonify({"error": "Forbidden"}), 403

        payload = request.get_json(silent=True) or {}
        issue_date = payload.get('date') or payload.get('issue_date')
        start_time = payload.get('start_time') or payload.get('start')
        end_time = payload.get('end_time') or payload.get('end')
        reason = payload.get('reason')
        comment = payload.get('comment')

        operator_ids = _normalize_int_id_list(payload.get('operator_ids'))
        operator_ids.extend(_normalize_int_id_list(payload.get('operator_id')))
        operator_ids = _normalize_int_id_list(operator_ids)

        direction_ids = _normalize_int_id_list(payload.get('direction_ids'))
        direction_ids.extend(_normalize_int_id_list(payload.get('direction_id')))
        direction_ids = _normalize_int_id_list(direction_ids)

        result = db.create_operator_technical_issues(
            requester_id=requester_id,
            requester_role=requester_role,
            issue_date=issue_date,
            start_time=start_time,
            end_time=end_time,
            reason=reason,
            comment=comment,
            operator_ids=operator_ids,
            direction_ids=direction_ids
        )

        return jsonify({
            "status": "success",
            "message": "Technical issue saved",
            "result": result
        }), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error creating technical issue: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/technical_issues', methods=['GET'])
@require_api_key
def list_technical_issues():
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        requester_role = _normalize_management_role(requester[3])
        if not (_is_admin_role(requester_role) or _is_supervisor_role(requester_role)):
            return jsonify({"error": "Forbidden"}), 403

        issue_date = (request.args.get('date') or '').strip() or None
        date_from = (request.args.get('date_from') or '').strip() or None
        date_to = (request.args.get('date_to') or '').strip() or None
        operator_id = (request.args.get('operator_id') or '').strip() or None
        reason = (request.args.get('reason') or '').strip() or None
        limit = request.args.get('limit', 500)
        offset = request.args.get('offset', 0)

        result = db.get_operator_technical_issues(
            requester_id=requester_id,
            requester_role=requester_role,
            issue_date=issue_date,
            date_from=date_from,
            date_to=date_to,
            operator_id=operator_id,
            reason=reason,
            limit=limit,
            offset=offset
        )

        return jsonify({
            "status": "success",
            "total": int(result.get('total') or 0),
            "items": result.get('items') or [],
            "reasons": list(TECHNICAL_ISSUE_REASONS)
        }), 200
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error listing technical issues: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/technical_issues/<int:issue_id>', methods=['DELETE'])
@require_api_key
def delete_technical_issue(issue_id):
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        requester_role = _normalize_management_role(requester[3])
        if not (_is_admin_role(requester_role) or _is_supervisor_role(requester_role)):
            return jsonify({"error": "Forbidden"}), 403

        result = db.delete_operator_technical_issue(
            requester_id=requester_id,
            requester_role=requester_role,
            issue_id=issue_id
        )

        return jsonify({
            "status": "success",
            "message": "Technical issue deleted",
            "result": result
        }), 200
    except PermissionError:
        return jsonify({"error": "Forbidden"}), 403
    except ValueError as e:
        error_text = str(e)
        if error_text == "Technical issue not found":
            return jsonify({"error": error_text}), 404
        return jsonify({"error": error_text}), 400
    except Exception as e:
        logging.error(f"Error deleting technical issue {issue_id}: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/offline_activities', methods=['POST'])
@require_api_key
def create_offline_activity():
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        requester_role = _normalize_management_role(requester[3])
        if not (_is_admin_role(requester_role) or _is_supervisor_role(requester_role)):
            return jsonify({"error": "Forbidden"}), 403

        payload = request.get_json(silent=True) or {}
        activity_date = payload.get('date') or payload.get('activity_date')
        start_time = payload.get('start_time') or payload.get('start')
        end_time = payload.get('end_time') or payload.get('end')
        comment = payload.get('comment')

        operator_id = payload.get('operator_id')
        if operator_id is None:
            operator_ids = _normalize_int_id_list(payload.get('operator_ids'))
            if operator_ids:
                operator_id = operator_ids[0]

        result = db.create_operator_offline_activity(
            requester_id=requester_id,
            requester_role=requester_role,
            activity_date=activity_date,
            start_time=start_time,
            end_time=end_time,
            comment=comment,
            operator_id=operator_id
        )

        return jsonify({
            "status": "success",
            "message": "Offline activity saved",
            "result": result
        }), 201
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error creating offline activity: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/offline_activities', methods=['GET'])
@require_api_key
def list_offline_activities():
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        requester_role = _normalize_management_role(requester[3])
        if not (_is_admin_role(requester_role) or _is_supervisor_role(requester_role)):
            return jsonify({"error": "Forbidden"}), 403

        activity_date = (request.args.get('date') or '').strip() or None
        date_from = (request.args.get('date_from') or '').strip() or None
        date_to = (request.args.get('date_to') or '').strip() or None
        operator_id = (request.args.get('operator_id') or '').strip() or None
        limit = request.args.get('limit', 500)
        offset = request.args.get('offset', 0)

        result = db.get_operator_offline_activities(
            requester_id=requester_id,
            requester_role=requester_role,
            activity_date=activity_date,
            date_from=date_from,
            date_to=date_to,
            operator_id=operator_id,
            limit=limit,
            offset=offset
        )

        return jsonify({
            "status": "success",
            "total": int(result.get('total') or 0),
            "items": result.get('items') or []
        }), 200
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error listing offline activities: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/offline_activities/<int:activity_id>', methods=['DELETE'])
@require_api_key
def delete_offline_activity(activity_id):
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        requester_role = _normalize_management_role(requester[3])
        if not (_is_admin_role(requester_role) or _is_supervisor_role(requester_role)):
            return jsonify({"error": "Forbidden"}), 403

        result = db.delete_operator_offline_activity(
            requester_id=requester_id,
            requester_role=requester_role,
            activity_id=activity_id
        )

        return jsonify({
            "status": "success",
            "message": "Offline activity deleted",
            "result": result
        }), 200
    except PermissionError:
        return jsonify({"error": "Forbidden"}), 403
    except ValueError as e:
        error_text = str(e)
        if error_text == "Offline activity not found":
            return jsonify({"error": error_text}), 404
        return jsonify({"error": error_text}), 400
    except Exception as e:
        logging.error(f"Error deleting offline activity {activity_id}: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/api/technical_issues/export_excel', methods=['GET'])
@require_api_key
def export_technical_issues_excel():
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        requester = db.get_user(id=requester_id)
        if not requester:
            return jsonify({"error": "User not found"}), 404

        requester_role = _normalize_management_role(requester[3])
        if not _is_admin_role(requester_role):
            return jsonify({"error": "Only admin can export technical issues"}), 403

        issue_date = (request.args.get('date') or '').strip() or None
        date_from = (request.args.get('date_from') or '').strip() or None
        date_to = (request.args.get('date_to') or '').strip() or None
        operator_id = (request.args.get('operator_id') or '').strip() or None
        reason = (request.args.get('reason') or '').strip() or None

        result = db.get_operator_technical_issues(
            requester_id=requester_id,
            requester_role=requester_role,
            issue_date=issue_date,
            date_from=date_from,
            date_to=date_to,
            operator_id=operator_id,
            reason=reason,
            limit=5000,
            offset=0
        )
        rows = result.get('items') or []

        wb = Workbook()
        ws = wb.active
        ws.title = 'Тех причины'

        headers = [
            'Дата проблемы',
            'Время начала',
            'Время окончания',
            'Оператор',
            'Направление оператора',
            'Техническая причина',
            'Комментарий',
            'Выбранные направления',
            'Добавил',
            'Дата фиксации'
        ]
        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = cell.font.copy(bold=True)

        for row_idx, item in enumerate(rows, start=2):
            selected_directions = ', '.join(item.get('selected_direction_names') or [])
            ws.cell(row=row_idx, column=1, value=item.get('date') or '')
            ws.cell(row=row_idx, column=2, value=item.get('start_time') or '')
            ws.cell(row=row_idx, column=3, value=item.get('end_time') or '')
            ws.cell(row=row_idx, column=4, value=item.get('operator_name') or '')
            ws.cell(row=row_idx, column=5, value=item.get('direction_name') or '')
            ws.cell(row=row_idx, column=6, value=item.get('reason') or '')
            ws.cell(row=row_idx, column=7, value=item.get('comment') or '')
            ws.cell(row=row_idx, column=8, value=selected_directions)
            ws.cell(row=row_idx, column=9, value=item.get('created_by_name') or '')
            ws.cell(row=row_idx, column=10, value=item.get('created_at') or '')

        for column_cells in ws.columns:
            max_len = 0
            for cell in column_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    continue
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 60)

        filename = f"technical_issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error exporting technical issues: {e}", exc_info=True)
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


# ==================== Work Schedules API ====================

WORK_SCHEDULE_EXCEL_SHIFT_RE = re.compile(
    r'(?P<sh>\d{1,2})(?:[:/](?P<sm>\d{1,2}))?\s*\*\s*(?P<eh>\d{1,2})(?:[:/](?P<em>\d{1,2}))?'
)


def _normalize_schedule_excel_name_key(value):
    return re.sub(r'\s+', ' ', str(value or '').strip()).replace('ё', 'е').replace('Ё', 'Е').lower()


def _parse_schedule_excel_header_date(cell):
    value = cell.value
    if value is None:
        return None

    # openpyxl для xlsx чаще всего уже возвращает datetime/date при is_date=True
    if getattr(cell, 'is_date', False):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, dt_date):
            return value

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, dt_date):
        return value

    text = str(value).strip()
    if not text:
        return None
    for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            continue
    return None


def _format_schedule_excel_compact_time(hhmm):
    try:
        hh, mm = str(hhmm).split(':', 1)
        h = int(hh)
        m = int(mm)
    except Exception:
        return str(hhmm or '')
    if m == 0:
        return str(h)
    return f"{h}/{m:02d}"


def _parse_schedule_excel_cell_value(raw_value):
    """
    Возвращает:
      {'kind': 'empty'}
      {'kind': 'day_off'}
      {'kind': 'shifts', 'shifts': [{'start': 'HH:MM', 'end': 'HH:MM'}, ...]}
      {'kind': 'invalid', 'error': '...'}
    """
    if raw_value is None:
        return {'kind': 'empty'}

    text = str(raw_value).strip()
    if not text:
        return {'kind': 'empty'}

    lowered = text.lower()
    shift_matches = list(WORK_SCHEDULE_EXCEL_SHIFT_RE.finditer(text))

    if not shift_matches and 'выходн' in lowered:
        return {'kind': 'day_off'}

    shifts = []
    seen = set()
    for m in shift_matches:
        try:
            sh = int(m.group('sh'))
            sm = int(m.group('sm') or 0)
            eh = int(m.group('eh'))
            em = int(m.group('em') or 0)
        except Exception:
            continue

        if not (0 <= sh <= 23 and 0 <= eh <= 23 and 0 <= sm <= 59 and 0 <= em <= 59):
            return {'kind': 'invalid', 'error': f'Некорректное время: {m.group(0)}'}

        start = f"{sh:02d}:{sm:02d}"
        end = f"{eh:02d}:{em:02d}"
        key = (start, end)
        if key in seen:
            continue
        seen.add(key)
        shifts.append({'start': start, 'end': end})

    if shifts:
        return {'kind': 'shifts', 'shifts': shifts}

    if 'выходн' in lowered:
        return {'kind': 'day_off'}

    return {'kind': 'invalid', 'error': 'Ячейка не похожа на смену или "Выходной"'}


STATUS_IMPORT_MAX_FILE_SIZE_MB = max(1, int(os.getenv('STATUS_IMPORT_MAX_FILE_SIZE_MB', '12')))
STATUS_IMPORT_MAX_FILE_SIZE_BYTES = STATUS_IMPORT_MAX_FILE_SIZE_MB * 1024 * 1024
STATUS_IMPORT_MAX_SOURCE_ROWS = max(1, int(os.getenv('STATUS_IMPORT_MAX_SOURCE_ROWS', '120000')))
STATUS_IMPORT_INVALID_ROWS_PREVIEW_LIMIT = max(1, int(os.getenv('STATUS_IMPORT_INVALID_ROWS_PREVIEW_LIMIT', '30')))
STATUS_IMPORT_LOCK = threading.Lock()


def _status_import_normalize_key(value):
    return ' '.join(str(value or '').strip().lower().split())


def _status_import_normalize_header(value):
    return re.sub(r'[\s_-]+', '', str(value or '').replace('\ufeff', '').strip().lower())


def _status_import_normalize_operator_name(value):
    return re.sub(r'\s+', ' ', str(value or '').strip()).replace('ё', 'е').replace('Ё', 'Е').lower()


def _status_import_parse_datetime(value):
    text = str(value or '').strip()
    if not text:
        return None
    formats = (
        '%d.%m.%Y %H:%M:%S',
        '%d.%m.%Y %H:%M',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%dT%H:%M'
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def _status_import_resolve_break_note_label(state_note_raw):
    note_key = _status_import_normalize_key(state_note_raw)
    if not note_key:
        return 'Перерыв'
    compact_note_key = re.sub(r'[\s._-]+', '', note_key)
    if note_key == 'вышел':
        return 'Вышел'
    # "Авто" в рамках "Перерыв" трактуем как обычный перерыв.
    if note_key == 'авто':
        return 'Перерыв'
    if note_key == 'перезвон':
        return 'Перезвон'
    if compact_note_key == 'техпричина':
        return 'Тех причина'
    if note_key == 'тренинг':
        return 'Тренинг'
    return str(state_note_raw or '').strip() or 'Перерыв'


def _status_import_resolve_display_state(state_name_raw, state_note_raw):
    base_name = str(state_name_raw or '').strip()
    base_key = _status_import_normalize_key(base_name)
    if base_key == 'перерыв':
        label = _status_import_resolve_break_note_label(state_note_raw)
        return {
            'label': label,
            'key': _status_import_normalize_key(label),
            'base_key': base_key,
            'base_name': base_name
        }
    return {
        'label': base_name or '—',
        'key': base_key,
        'base_key': base_key,
        'base_name': base_name
    }


def _status_import_split_segment_by_day(start_dt, end_dt):
    if not isinstance(start_dt, datetime) or not isinstance(end_dt, datetime):
        return []
    if end_dt <= start_dt:
        return []

    result = []
    cursor = start_dt
    while cursor < end_dt:
        day_start = cursor.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        seg_end = min(end_dt, day_end)
        duration_sec = int((seg_end - cursor).total_seconds())
        if duration_sec > 0:
            result.append({
                'date': cursor.date(),
                'start': cursor,
                'end': seg_end,
                'duration_sec': duration_sec
            })
        cursor = seg_end
    return result


def _status_import_build_operator_lookup():
    lookup = {}
    for row in (db.get_all_operators() or []):
        try:
            operator_id = int(row[0])
        except Exception:
            continue
        operator_name = str(row[1] or '').strip()
        key = _status_import_normalize_operator_name(operator_name)
        if not key:
            continue
        lookup.setdefault(key, []).append({
            'id': operator_id,
            'name': operator_name
        })
    return lookup


def _status_import_parse_csv(csv_text, operator_lookup, max_source_rows=None, invalid_rows_preview_limit=None):
    if not isinstance(csv_text, str):
        raise ValueError("CSV text is required")

    reader = csv.reader(StringIO(csv_text), delimiter=';', quotechar='"')
    header = next(reader, None)
    if not header:
        raise ValueError("Файл пустой")

    normalized_header = [_status_import_normalize_header(h) for h in header]

    def _find_col(candidates):
        for candidate in candidates:
            if candidate in normalized_header:
                return normalized_header.index(candidate)
        return None

    operator_col = _find_col(['operatorname', 'operator', 'name'])
    state_col = _find_col(['statename', 'state', 'status'])
    time_col = _find_col(['timechange', 'time', 'datetime'])
    note_col = _find_col(['statenote', 'statusnote', 'note'])

    if operator_col is None or state_col is None or time_col is None:
        raise ValueError("CSV должен содержать колонки OperatorName;StateName;TimeChange")

    max_rows_value = None
    if max_source_rows is not None:
        try:
            max_rows_value = int(max_source_rows)
        except Exception:
            max_rows_value = None
        if max_rows_value is not None and max_rows_value <= 0:
            max_rows_value = None

    preview_limit_value = STATUS_IMPORT_INVALID_ROWS_PREVIEW_LIMIT
    if invalid_rows_preview_limit is not None:
        try:
            preview_limit_value = max(1, int(invalid_rows_preview_limit))
        except Exception:
            preview_limit_value = STATUS_IMPORT_INVALID_ROWS_PREVIEW_LIMIT

    def _cell(cols, idx):
        if idx is None:
            return ''
        if idx < 0 or idx >= len(cols):
            return ''
        return str(cols[idx] or '').strip()

    source_rows = 0
    valid_events = 0
    invalid_rows_count = 0
    invalid_rows_preview = []
    events_by_operator = {}
    source_order_stats = {}

    def _push_invalid(row_num, reason, operator_name, source_state_name, state_note, time_change):
        nonlocal invalid_rows_count
        invalid_rows_count += 1
        if len(invalid_rows_preview) >= preview_limit_value:
            return
        invalid_rows_preview.append({
            'row': int(row_num),
            'reason': str(reason or '').strip() or 'Некорректная строка',
            'operator_name': str(operator_name or '').strip(),
            'state_name': str(source_state_name or '').strip(),
            'state_note': str(state_note or '').strip(),
            'time_change': str(time_change or '').strip()
        })

    for row_num, cols in enumerate(reader, start=2):
        source_rows += 1
        if max_rows_value is not None and source_rows > max_rows_value:
            raise OverflowError(f"Лимит строк CSV превышен ({max_rows_value})")

        operator_name = _cell(cols, operator_col)
        source_state_name = _cell(cols, state_col)
        state_note = _cell(cols, note_col)
        time_change = _cell(cols, time_col)

        if not operator_name and not source_state_name and not state_note and not time_change:
            continue

        if not operator_name or not source_state_name or not time_change:
            _push_invalid(
                row_num=row_num,
                reason='Отсутствуют обязательные поля',
                operator_name=operator_name,
                source_state_name=source_state_name,
                state_note=state_note,
                time_change=time_change
            )
            continue

        ts = _status_import_parse_datetime(time_change)
        if not ts:
            _push_invalid(
                row_num=row_num,
                reason='Некорректный формат TimeChange',
                operator_name=operator_name,
                source_state_name=source_state_name,
                state_note=state_note,
                time_change=time_change
            )
            continue

        valid_events += 1

        operator_key = _status_import_normalize_operator_name(operator_name)
        operator_matches = operator_lookup.get(operator_key) or []
        if len(operator_matches) != 1:
            _push_invalid(
                row_num=row_num,
                reason='Оператор не найден' if len(operator_matches) == 0 else 'Найдено несколько операторов с таким именем',
                operator_name=operator_name,
                source_state_name=source_state_name,
                state_note=state_note,
                time_change=time_change
            )
            continue

        resolved = _status_import_resolve_display_state(source_state_name, state_note)
        operator_info = operator_matches[0]
        operator_id = int(operator_info['id'])

        order_stats = source_order_stats.setdefault(operator_id, {'asc': 0, 'desc': 0, 'last_ts': None})
        prev_ts = order_stats.get('last_ts')
        if isinstance(prev_ts, datetime):
            diff = (ts - prev_ts).total_seconds()
            if diff > 0:
                order_stats['asc'] = int(order_stats.get('asc') or 0) + 1
            elif diff < 0:
                order_stats['desc'] = int(order_stats.get('desc') or 0) + 1
        order_stats['last_ts'] = ts

        events_by_operator.setdefault(operator_id, []).append({
            'operator_id': operator_id,
            'event_at': ts,
            'status_key': resolved.get('key') or _status_import_normalize_key(source_state_name),
            'state_note': state_note,
            'source_row': int(row_num)
        })

    open_tail_events = 0
    zero_or_negative_transitions = 0
    segments = []
    events_for_db = []
    now_dt = datetime.now()

    for operator_id, events_list in events_by_operator.items():
        if not events_list:
            continue

        order_stats = source_order_stats.get(int(operator_id)) or {}
        prefer_desc = int(order_stats.get('desc') or 0) > int(order_stats.get('asc') or 0)
        if len(events_list) > 1:
            events_list.sort(
                key=lambda ev: (
                    ev['event_at'],
                    -int(ev.get('source_row') or 0) if prefer_desc else int(ev.get('source_row') or 0)
                )
            )

        events_for_db.extend(events_list)

        for idx in range(len(events_list) - 1):
            cur = events_list[idx]
            nxt = events_list[idx + 1]
            cur_ts = cur['event_at']
            next_ts = nxt['event_at']
            if next_ts <= cur_ts:
                zero_or_negative_transitions += 1
                continue

            for part in _status_import_split_segment_by_day(cur_ts, next_ts):
                segments.append({
                    'operator_id': int(operator_id),
                    'status_date': part['date'].strftime('%Y-%m-%d'),
                    'start_at': part['start'],
                    'end_at': part['end'],
                    'duration_sec': int(part['duration_sec']),
                    'status_key': cur.get('status_key'),
                    'state_note': cur.get('state_note')
                })

        last_event = events_list[-1]
        last_event_at = last_event.get('event_at')
        tail_anchor_dt = last_event_at
        if isinstance(last_event_at, datetime) and (now_dt - last_event_at) <= timedelta(hours=48):
            tail_anchor_dt = now_dt
        if (
            isinstance(last_event_at, datetime)
            and isinstance(tail_anchor_dt, datetime)
            and tail_anchor_dt > last_event_at
        ):
            for part in _status_import_split_segment_by_day(last_event_at, tail_anchor_dt):
                segments.append({
                    'operator_id': int(operator_id),
                    'status_date': part['date'].strftime('%Y-%m-%d'),
                    'start_at': part['start'],
                    'end_at': part['end'],
                    'duration_sec': int(part['duration_sec']),
                    'status_key': last_event.get('status_key'),
                    'state_note': last_event.get('state_note')
                })

        open_tail_events += 1

    return {
        'source_rows': int(source_rows),
        'valid_events': int(valid_events),
        'matched_events': len(events_for_db),
        'invalid_rows_count': int(invalid_rows_count),
        'invalid_rows_preview': invalid_rows_preview,
        'parse_errors_count': 0,
        'operators_count': len(events_by_operator),
        'open_tail_events': int(open_tail_events),
        'zero_or_negative_transitions': int(zero_or_negative_transitions),
        'events': events_for_db,
        'segments': segments
    }


def _ws_time_to_minutes(value):
    if value is None:
        return 0
    parts = str(value).strip().split(':')
    try:
        hh = int(parts[0])
        mm = int(parts[1]) if len(parts) > 1 else 0
    except Exception:
        return 0
    return hh * 60 + mm


def _ws_minutes_to_time(minutes):
    minutes = int(round(minutes)) % (24 * 60)
    hh = minutes // 60
    mm = minutes % 60
    return f"{hh:02d}:{mm:02d}"


def _ws_normalize_direction_key(value):
    return ' '.join(str(value or '').strip().lower().split())


def _ws_is_chat_manager_direction(direction_value):
    key = _ws_normalize_direction_key(direction_value)
    return key in ('чат менеджер', 'chat manager')


def _ws_normalize_break_durations(value):
    if value is None:
        return []
    raw = value if isinstance(value, list) else []
    result = []
    for item in raw:
        try:
            minutes = int(item)
        except Exception:
            continue
        if minutes > 0:
            result.append(minutes)
    return result


def _ws_pick_break_durations_for_shift(duration_minutes, direction_value=None, break_rules_map=None):
    dur = int(duration_minutes)
    rules_map = break_rules_map if isinstance(break_rules_map, dict) else {}
    selected_custom = None
    direction_key = _ws_normalize_direction_key(direction_value)
    direction_rules = rules_map.get(direction_key) or []
    for rule in sorted(
        [r for r in direction_rules if isinstance(r, dict)],
        key=lambda x: (int(x.get('minMinutes', 0)), int(x.get('maxMinutes', 0)))
    ):
        try:
            min_minutes = int(rule.get('minMinutes'))
            max_minutes = int(rule.get('maxMinutes'))
        except Exception:
            continue
        if max_minutes <= min_minutes:
            continue
        if dur >= min_minutes and dur <= max_minutes:
            if selected_custom is None or min_minutes >= int(selected_custom.get('minMinutes', -1)):
                selected_custom = rule

    if selected_custom is not None:
        return _ws_normalize_break_durations(selected_custom.get('breakDurations'))

    if _ws_is_chat_manager_direction(direction_value):
        if dur >= 6 * 60 and dur < 9 * 60:
            return [30]
        if dur >= 9 * 60 and dur <= 12 * 60:
            return [30, 30]

    if dur >= 5 * 60 and dur < 6 * 60:
        return [15]
    if dur >= 6 * 60 and dur < 8 * 60:
        return [15, 15]
    if dur >= 8 * 60 and dur < 11 * 60:
        return [15, 30, 15]
    if dur >= 11 * 60:
        return [15, 30, 15, 15]
    return []


def _ws_compute_breaks_for_shift_minutes(start_min, end_min, direction_value=None, break_rules_map=None):
    dur = int(end_min) - int(start_min)
    if dur <= 0:
        return []
    breaks = []

    def snap5(x):
        return int(round(float(x) / 5.0) * 5)

    def push_centered(center, size):
        center = snap5(center)
        s = snap5(center - (size / 2))
        e = s + int(size)
        s = max(int(start_min), min(int(end_min), s))
        e = max(int(start_min), min(int(end_min), e))
        if e > s:
            breaks.append({'start': s, 'end': e})

    break_durations = _ws_pick_break_durations_for_shift(
        duration_minutes=dur,
        direction_value=direction_value,
        break_rules_map=break_rules_map
    )
    if break_durations:
        count = len(break_durations)
        for idx, size in enumerate(break_durations):
            center = int(start_min) + (dur * ((idx + 1) / (count + 1)))
            push_centered(center, int(size))

    normalized = []
    seen = set()
    for b in sorted(breaks, key=lambda x: (int(x.get('start', 0)), int(x.get('end', 0)))):
        s = int(b.get('start', 0))
        e = int(b.get('end', 0))
        if e <= s:
            continue
        key = (s, e)
        if key in seen:
            continue
        seen.add(key)
        normalized.append({'start': s, 'end': e})
    return normalized


def _ws_intervals_overlap(a, b):
    return int(a['start']) < int(b['end']) and int(b['start']) < int(a['end'])


def _ws_merge_intervals(items):
    if not items:
        return []
    sorted_items = sorted(
        [{'start': int(x['start']), 'end': int(x['end'])} for x in items if int(x.get('end', 0)) > int(x.get('start', 0))],
        key=lambda x: (x['start'], x['end'])
    )
    if not sorted_items:
        return []
    res = [dict(sorted_items[0])]
    for cur in sorted_items[1:]:
        last = res[-1]
        if cur['start'] <= last['end']:
            last['end'] = max(last['end'], cur['end'])
        else:
            res.append(dict(cur))
    return res


def _ws_parse_date_str(value):
    return datetime.strptime(str(value), '%Y-%m-%d').date()


def _ws_date_str(value):
    if isinstance(value, datetime):
        return value.date().strftime('%Y-%m-%d')
    if isinstance(value, dt_date):
        return value.strftime('%Y-%m-%d')
    return str(value)


def _ws_add_days_str(date_str, delta_days):
    return (_ws_parse_date_str(date_str) + timedelta(days=int(delta_days))).strftime('%Y-%m-%d')


def _ws_clone_ops_for_break_simulation(operators):
    cloned = []
    for op in (operators or []):
        shifts = {}
        for d, segs in (op.get('shifts') or {}).items():
            cloned_segs = []
            for seg in (segs or []):
                seg_copy = {
                    'start': str(seg.get('start') or ''),
                    'end': str(seg.get('end') or ''),
                    'breaks': [ {'start': int(b.get('start')), 'end': int(b.get('end'))} for b in (seg.get('breaks') or []) if b is not None ]
                }
                smin = _ws_time_to_minutes(seg_copy['start'])
                emin = _ws_time_to_minutes(seg_copy['end'])
                if emin <= smin:
                    emin += 1440
                seg_copy['__startMin'] = smin
                seg_copy['__endMin'] = emin
                cloned_segs.append(seg_copy)
            shifts[str(d)] = cloned_segs
        cloned.append({
            'id': op.get('id'),
            'direction': op.get('direction'),
            'shifts': shifts,
            'daysOff': list(op.get('daysOff') or [])
        })
    return cloned


def _ws_sanitize_break_direction_groups(groups):
    if not isinstance(groups, list):
        return []
    used = set()
    seen_groups = set()
    result = []
    for group in groups:
        if not isinstance(group, list):
            continue
        normalized = []
        local_keys = set()
        for raw in group:
            label = str(raw or '').strip()
            key = label.lower()
            if not key or key in used or key in local_keys:
                continue
            local_keys.add(key)
            normalized.append(label)
        if len(normalized) < 2:
            continue
        sig = '|'.join(sorted(x.lower() for x in normalized))
        if sig in seen_groups:
            continue
        seen_groups.add(sig)
        for x in normalized:
            used.add(x.lower())
        result.append(normalized)
    return result


def _ws_make_direction_scope_resolver(break_direction_groups):
    scope_map = {}
    for group in _ws_sanitize_break_direction_groups(break_direction_groups):
        keys = sorted({str(x or '').strip().lower() for x in group if str(x or '').strip()})
        if len(keys) < 2:
            continue
        scope_key = f"group:{'|'.join(keys)}"
        for k in keys:
            scope_map[k] = scope_key

    def _resolver(direction_value):
        key = str(direction_value or '').strip().lower()
        if not key:
            return 'dir:'
        return scope_map.get(key) or f"dir:{key}"

    return _resolver


def _ws_build_occupied_intervals_for_date(all_operators, date_str, exclude_op_id, direction_scope, get_scope_key):
    occupied = []
    prev_str = _ws_add_days_str(date_str, -1)
    next_str = _ws_add_days_str(date_str, 1)
    target_scope = str(get_scope_key(direction_scope))

    def clamp_push(s, e):
        ns = max(0, min(2880, int(s)))
        ne = max(0, min(2880, int(e)))
        if ne > ns:
            occupied.append({'start': ns, 'end': ne})

    for op in (all_operators or []):
        if int(op.get('id') or 0) == int(exclude_op_id or 0):
            continue
        if str(get_scope_key(op.get('direction'))) != target_scope:
            continue

        for seg in (op.get('shifts', {}).get(date_str) or []):
            for b in (seg.get('breaks') or []):
                clamp_push(b.get('start', 0), b.get('end', 0))

        for seg in (op.get('shifts', {}).get(prev_str) or []):
            for b in (seg.get('breaks') or []):
                clamp_push(int(b.get('start', 0)) - 1440, int(b.get('end', 0)) - 1440)

        for seg in (op.get('shifts', {}).get(next_str) or []):
            for b in (seg.get('breaks') or []):
                clamp_push(int(b.get('start', 0)) + 1440, int(b.get('end', 0)) + 1440)

    return _ws_merge_intervals(occupied)


def _ws_find_non_overlapping_start(desired_start, length, seg_start, seg_end, occupied_intervals):
    step = 5
    def snap(v):
        return int(round(float(v) / step) * step)

    desired_start = snap(desired_start)
    candidates = [0]
    max_shift = max(int(seg_end) - int(seg_start), 60)
    for d in range(step, max_shift + step, step):
        candidates.append(d)
        candidates.append(-d)

    for delta in candidates:
        s = desired_start + delta
        start = max(int(seg_start), min(int(seg_end) - int(length), s))
        end = start + int(length)
        if start < int(seg_start) or end > int(seg_end):
            continue
        test_iv = {'start': start, 'end': end}
        if any(_ws_intervals_overlap(test_iv, occ) for occ in (occupied_intervals or [])):
            continue
        return start
    return None


def _ws_adjust_breaks_for_operator_on_date(op, date_str, all_operators, get_direction_scope_key, break_rules_map=None):
    segs = (op or {}).get('shifts', {}).get(date_str) or []
    if not segs:
        return
    occupied = _ws_build_occupied_intervals_for_date(
        all_operators=all_operators,
        date_str=date_str,
        exclude_op_id=op.get('id'),
        direction_scope=op.get('direction'),
        get_scope_key=get_direction_scope_key
    )
    for seg in segs:
        raw_seg_start = int(seg.get('__startMin', _ws_time_to_minutes(seg.get('start'))))
        seg_start_base = _ws_time_to_minutes(seg.get('start'))
        seg_end_base = _ws_time_to_minutes(seg.get('end'))
        raw_seg_end = int(seg.get('__endMin', seg_end_base + (1440 if seg_end_base <= seg_start_base else 0)))
        if not isinstance(seg.get('breaks'), list):
            seg['breaks'] = []
        if len(seg['breaks']) == 0:
            seg['breaks'] = _ws_compute_breaks_for_shift_minutes(
                raw_seg_start,
                raw_seg_end,
                op.get('direction'),
                break_rules_map=break_rules_map
            )

        seg_start = max(0, raw_seg_start)
        seg_end = max(seg_start, min(2880, raw_seg_end))
        new_breaks = []
        for b in (seg.get('breaks') or []):
            b_start = int(b.get('start', 0))
            b_end = int(b.get('end', 0))
            length = b_end - b_start
            if length <= 0 or (seg_end - seg_start) <= 0:
                continue
            desired_start = max(seg_start, min(seg_end - length, b_start))
            found = _ws_find_non_overlapping_start(desired_start, length, seg_start, seg_end, occupied)
            if found is not None:
                nb = {'start': int(found), 'end': int(found) + int(length)}
            else:
                clamped_start = max(seg_start, min(seg_end - length, b_start))
                clamped_end = max(seg_start, min(seg_end, b_end))
                if clamped_end <= clamped_start:
                    continue
                nb = {'start': clamped_start, 'end': clamped_end}
            new_breaks.append(nb)
            occupied.append(nb)
            occupied = _ws_merge_intervals(occupied)
        seg['breaks'] = new_breaks


def _iter_schedule_excel_dates(start_date_obj, end_date_obj):
    cur = start_date_obj
    while cur <= end_date_obj:
        yield cur
        cur = cur + timedelta(days=1)


def _work_schedule_operator_requester():
    requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
    if not requester_id:
        return None, None, ("Unauthorized", 401)

    try:
        requester_id = int(requester_id)
    except Exception:
        return None, None, ("Invalid requester id", 400)

    user_data = db.get_user(id=requester_id)
    if not user_data:
        return None, None, ("User not found", 404)
    if user_data[3] != 'operator':
        return None, None, ("Forbidden", 403)

    return requester_id, user_data, None


def _ws_query_bool(name, default=False):
    raw = request.args.get(name)
    if raw is None:
        return bool(default)
    return str(raw).strip().lower() in ('1', 'true', 'yes', 'on', 'y')


@app.route('/api/work_schedules/my', methods=['GET'])
@require_api_key
def get_my_work_schedules():
    """
    Получить смены/перерывы/выходные текущего оператора.
    Query params: start_date, end_date (optional, format: YYYY-MM-DD)
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404

        if user_data[3] != 'operator':
            return jsonify({"error": "Forbidden"}), 403

        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        include_imported_statuses = _ws_query_bool('include_imported_statuses', default=False)
        include_technical_issues = _ws_query_bool('include_technical_issues', default=True)
        include_offline_activities = _ws_query_bool('include_offline_activities', default=True)
        operator_schedule = db.get_operator_with_shifts(
            requester_id,
            start_date,
            end_date,
            include_imported_statuses=include_imported_statuses,
            include_technical_issues=include_technical_issues,
            include_offline_activities=include_offline_activities
        )
        if not operator_schedule:
            return jsonify({"error": "Operator not found"}), 404

        return jsonify({"operator": operator_schedule}), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error getting my work schedules: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/direction', methods=['GET'])
@require_api_key
def get_direction_work_schedules():
    """
    Получить графики операторов своего направления за период.
    Query params: start_date, end_date (YYYY-MM-DD)
    Доступно только для операторов.
    """
    try:
        requester_id, user_data, err = _work_schedule_operator_requester()
        if err:
            return jsonify({"error": err[0]}), err[1]

        direction_name = user_data[4]  # d.name from get_user()
        if not direction_name:
            return jsonify({"operators": [], "direction": None}), 200

        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        include_imported_statuses = _ws_query_bool('include_imported_statuses', default=False)
        include_technical_issues = _ws_query_bool('include_technical_issues', default=True)
        include_offline_activities = _ws_query_bool('include_offline_activities', default=True)

        operators = db.get_operators_with_shifts(
            start_date=start_date,
            end_date=end_date,
            direction_name=direction_name,
            include_imported_statuses=include_imported_statuses,
            include_technical_issues=include_technical_issues,
            include_offline_activities=include_offline_activities
        )
        return jsonify({"operators": operators, "direction": direction_name}), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error getting direction work schedules: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/shift_swap/candidates', methods=['GET'])
@require_api_key
def get_shift_swap_candidates():
    """
    Вернуть операторов-кандидатов для обмена:
    - в совместимом направлении с текущим оператором (СМЗ и Основа взаимозаменяемы)
    - в режиме обмена: только с пересечением по выбранному интервалу
    - в режиме обычной замены: только без пересечения по выбранному интервалу
    Query params:
      swap_date=YYYY-MM-DD
      end_date=YYYY-MM-DD (optional, default: swap_date)
      start_time=HH:MM
      end_time=HH:MM
      overlap_only=0|1 (optional, default 0)
      non_overlap_only=0|1 (optional, default 0)
    """
    try:
        requester_id, _user_data, err = _work_schedule_operator_requester()
        if err:
            return jsonify({"error": err[0]}), err[1]

        swap_date = request.args.get('swap_date')
        end_date = request.args.get('end_date')
        start_time = request.args.get('start_time')
        end_time = request.args.get('end_time')
        overlap_only_raw = request.args.get('overlap_only', '0')
        non_overlap_only_raw = request.args.get('non_overlap_only', '0')
        overlap_only = str(overlap_only_raw).strip().lower() in ('1', 'true', 'yes', 'y')
        non_overlap_only = str(non_overlap_only_raw).strip().lower() in ('1', 'true', 'yes', 'y')
        if overlap_only:
            non_overlap_only = False
        if not swap_date or not start_time or not end_time:
            return jsonify({"error": "swap_date, start_time and end_time are required"}), 400

        candidates = db.get_shift_swap_candidates(
            requester_operator_id=requester_id,
            swap_date=swap_date,
            end_date=end_date,
            start_time=start_time,
            end_time=end_time,
            overlap_only=overlap_only,
            non_overlap_only=non_overlap_only
        )
        return jsonify({
            "candidates": candidates,
            "swap_date": swap_date,
            "end_date": end_date or swap_date,
            "start_time": start_time,
            "end_time": end_time,
            "overlap_only": overlap_only,
            "non_overlap_only": non_overlap_only
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error getting shift swap candidates: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/shift_swap/requests', methods=['GET', 'POST'])
@require_api_key
def shift_swap_requests():
    """
    GET: входящие/исходящие запросы на замену для текущего оператора.
    POST: создать запрос на обмен сменами.
    """
    try:
        requester_id, _user_data, err = _work_schedule_operator_requester()
        if err:
            return jsonify({"error": err[0]}), err[1]

        if request.method == 'GET':
            raw_limit = request.args.get('limit')
            try:
                limit = int(raw_limit) if raw_limit is not None else 200
            except Exception:
                return jsonify({"error": "Invalid limit"}), 400

            payload = db.get_shift_swap_requests_for_operator(
                operator_id=requester_id,
                limit=limit
            )
            return jsonify(payload), 200

        data = request.get_json(silent=True) or {}
        target_operator_id = data.get('target_operator_id')
        swap_date = data.get('swap_date')
        end_date = data.get('end_date')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        request_comment = data.get('comment')
        target_segments = data.get('target_segments')
        if not target_operator_id or not swap_date or not start_time or not end_time:
            return jsonify({"error": "target_operator_id, swap_date, start_time and end_time are required"}), 400

        created_request = db.create_shift_swap_request(
            requester_operator_id=requester_id,
            target_operator_id=target_operator_id,
            swap_date=swap_date,
            end_date=end_date,
            start_time=start_time,
            end_time=end_time,
            request_comment=request_comment,
            target_segments=target_segments
        )
        return jsonify({
            "message": "Swap request created",
            "request": created_request
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error in shift swap requests endpoint: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/shift_swap/respond', methods=['POST'])
@require_api_key
def respond_shift_swap_request():
    """
    Ответ на запрос замены (только получатель):
    Body:
      {
        "request_id": int,
        "action": "accept" | "reject",
        "comment": "optional"
      }
    """
    try:
        requester_id, _user_data, err = _work_schedule_operator_requester()
        if err:
            return jsonify({"error": err[0]}), err[1]

        data = request.get_json(silent=True) or {}
        request_id = data.get('request_id')
        action = data.get('action')
        response_comment = data.get('comment')
        if not request_id or not action:
            return jsonify({"error": "request_id and action are required"}), 400

        updated_request = db.respond_shift_swap_request(
            request_id=request_id,
            responder_operator_id=requester_id,
            action=action,
            response_comment=response_comment
        )
        action_norm = str(action).strip().lower()
        message = "Swap request accepted" if action_norm == 'accept' else "Swap request rejected"
        return jsonify({
            "message": message,
            "request": updated_request
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error responding shift swap request: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/shift_swap/journal', methods=['GET'])
@require_api_key
def get_shift_swap_journal():
    """
    Журнал замен за месяц для admin/sv.
    Query params:
      month=YYYY-MM (required)
      limit=int (optional, default 500)
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        role = str(user_data[3] or '')
        if not (_is_admin_role(role) or _is_supervisor_role(role)):
            return jsonify({"error": "Forbidden"}), 403

        month = request.args.get('month')
        if not month:
            return jsonify({"error": "month is required (YYYY-MM)"}), 400

        raw_limit = request.args.get('limit')
        try:
            limit = int(raw_limit) if raw_limit is not None else 500
        except Exception:
            return jsonify({"error": "Invalid limit"}), 400

        payload = db.get_shift_swap_journal_for_month(
            month=month,
            requester_role=role,
            requester_id=requester_id,
            limit=limit
        )
        return jsonify(payload), 200
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error getting shift swap journal: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/operators', methods=['GET'])
@require_api_key
def get_operators_with_schedules():
    """
    Получить всех операторов с их сменами и выходными днями.
    Query params:
      - start_date, end_date (optional, format: YYYY-MM-DD)
      - anchor_date (optional, format: YYYY-MM-DD; fallback for default range)
      - view_mode (optional: day/week/month; fallback for default range)
      - include_imported_statuses (optional: 1/true/yes/on)
      - include_technical_issues (optional: 1/true/yes/on, default: true)
      - include_offline_activities (optional: 1/true/yes/on, default: true)
    """
    try:
        user = request.headers.get('X-User-Id')
        if not user:
            return jsonify({"error": "Unauthorized"}), 401
        
        user_id = int(user)
        user_data = db.get_user(id=user_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        role = user_data[3]
        
        # Только admin и sv могут видеть планировщик смен
        if not (_is_admin_role(role) or _is_supervisor_role(role)):
            return jsonify({"error": "Forbidden"}), 403
        
        start_date = (request.args.get('start_date') or '').strip()
        end_date = (request.args.get('end_date') or '').strip()

        if start_date and not end_date:
            end_date = start_date
        elif end_date and not start_date:
            start_date = end_date

        if not start_date and not end_date:
            anchor_date_raw = (request.args.get('anchor_date') or '').strip()
            view_mode = (request.args.get('view_mode') or 'month').strip().lower()
            if view_mode not in {'day', 'week', 'month'}:
                view_mode = 'month'

            if anchor_date_raw:
                anchor_date = datetime.strptime(anchor_date_raw, '%Y-%m-%d').date()
            else:
                anchor_date = datetime.now().date()

            if view_mode == 'day':
                range_start_obj = anchor_date
                range_end_obj = anchor_date
            elif view_mode == 'week':
                # Понедельник - начало недели
                week_day = anchor_date.isoweekday()  # 1..7
                range_start_obj = anchor_date - timedelta(days=week_day - 1)
                range_end_obj = range_start_obj + timedelta(days=6)
            else:
                range_start_obj = anchor_date.replace(day=1)
                next_month = (range_start_obj.replace(day=28) + timedelta(days=4)).replace(day=1)
                range_end_obj = next_month - timedelta(days=1)

            start_date = range_start_obj.strftime('%Y-%m-%d')
            end_date = range_end_obj.strftime('%Y-%m-%d')
        else:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            if end_date_obj < start_date_obj:
                return jsonify({"error": "end_date must be >= start_date"}), 400
        
        include_imported_statuses = _ws_query_bool('include_imported_statuses', default=False)
        include_technical_issues = _ws_query_bool('include_technical_issues', default=True)
        include_offline_activities = _ws_query_bool('include_offline_activities', default=True)
        operators = db.get_operators_with_shifts(
            start_date,
            end_date,
            include_imported_statuses=include_imported_statuses,
            include_technical_issues=include_technical_issues,
            include_offline_activities=include_offline_activities
        )
        
        return jsonify({
            "operators": operators,
            "range": {
                "start_date": start_date,
                "end_date": end_date
            }
        }), 200
    
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error getting operators with schedules: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/break_rules', methods=['GET'])
@require_api_key
def get_work_schedule_break_rules():
    """
    Получить правила автоперерывов по направлениям.
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404

        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        direction_rules = db.get_work_schedule_break_rules()
        return jsonify({"direction_rules": direction_rules}), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error getting work schedule break rules: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/break_rules', methods=['POST'])
@require_api_key
def save_work_schedule_break_rules():
    """
    Сохранить правила автоперерывов по направлениям.
    Body: {
        "direction_rules": [
            {
                "direction": "Название направления",
                "rules": [
                    {"minMinutes": 360, "maxMinutes": 540, "breakDurations": [30]}
                ]
            }
        ]
    }
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404

        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        data = request.get_json(silent=True) or {}
        direction_rules = (
            data.get('direction_rules')
            if 'direction_rules' in data else (
                data.get('directionRules')
                if 'directionRules' in data else data.get('rules')
            )
        )
        if not isinstance(direction_rules, list):
            return jsonify({"error": "direction_rules must be a list"}), 400

        saved = db.save_work_schedule_break_rules(direction_rules)
        return jsonify({
            "message": "Break rules saved successfully",
            "direction_rules": saved
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error saving work schedule break rules: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/shift', methods=['POST'])
@require_api_key
def save_work_shift():
    """
    Сохранить смену для оператора.
    Body: {
        "operator_id": int,
        "shift_date": "YYYY-MM-DD",
        "start_time": "HH:MM",
        "end_time": "HH:MM",
        "breaks": [{"start": minutes, "end": minutes}, ...],  // optional
        "previous_start_time": "HH:MM",  // optional, for edit without duplicates
        "previous_end_time": "HH:MM"     // optional, for edit without duplicates
    }
    """
    try:
        user = request.headers.get('X-User-Id')
        if not user:
            return jsonify({"error": "Unauthorized"}), 401
        
        user_id = int(user)
        user_data = db.get_user(id=user_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        role = user_data[3]
        
        if not (_is_admin_role(role) or _is_supervisor_role(role)):
            return jsonify({"error": "Forbidden"}), 403
        
        data = request.get_json(silent=True) or {}
        operator_id = data.get('operator_id')
        shift_date = data.get('shift_date')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        breaks = data.get('breaks')
        previous_start_time = data.get('previous_start_time')
        previous_end_time = data.get('previous_end_time')
        
        if not all([operator_id, shift_date, start_time, end_time]):
            return jsonify({"error": "Missing required fields"}), 400
        
        shift_id = db.save_shift(
            operator_id,
            shift_date,
            start_time,
            end_time,
            breaks,
            previous_start_time=previous_start_time,
            previous_end_time=previous_end_time
        )
        
        return jsonify({"message": "Shift saved successfully", "shift_id": shift_id}), 200
    
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error saving shift: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/shift', methods=['DELETE'])
@require_api_key
def delete_work_shift():
    """
    Удалить смену оператора.
    Body: {
        "operator_id": int,
        "shift_date": "YYYY-MM-DD",
        "start_time": "HH:MM",
        "end_time": "HH:MM"
    }
    """
    try:
        user = request.headers.get('X-User-Id')
        if not user:
            return jsonify({"error": "Unauthorized"}), 401
        
        user_id = int(user)
        user_data = db.get_user(id=user_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        role = user_data[3]
        
        if not (_is_admin_role(role) or _is_supervisor_role(role)):
            return jsonify({"error": "Forbidden"}), 403
        
        data = request.get_json(silent=True) or {}
        operator_id = data.get('operator_id')
        shift_date = data.get('shift_date')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        
        if not all([operator_id, shift_date, start_time, end_time]):
            return jsonify({"error": "Missing required fields"}), 400
        
        success = db.delete_shift(operator_id, shift_date, start_time, end_time)
        
        if success:
            return jsonify({"message": "Shift deleted successfully"}), 200
        else:
            return jsonify({"error": "Shift not found"}), 404
    
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error deleting shift: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/day_off', methods=['POST'])
@require_api_key
def toggle_work_day_off():
    """
    Переключить выходной день для оператора.
    Body: {
        "operator_id": int,
        "day_off_date": "YYYY-MM-DD"
    }
    """
    try:
        user = request.headers.get('X-User-Id')
        if not user:
            return jsonify({"error": "Unauthorized"}), 401
        
        user_id = int(user)
        user_data = db.get_user(id=user_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        role = user_data[3]
        
        if not (_is_admin_role(role) or _is_supervisor_role(role)):
            return jsonify({"error": "Forbidden"}), 403
        
        data = request.get_json(silent=True) or {}
        operator_id = data.get('operator_id')
        day_off_date = data.get('day_off_date')
        
        if not all([operator_id, day_off_date]):
            return jsonify({"error": "Missing required fields"}), 400
        
        is_day_off = db.toggle_day_off(operator_id, day_off_date)
        
        return jsonify({
            "message": "Day off toggled successfully",
            "is_day_off": is_day_off
        }), 200
    
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error toggling day off: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/status_period', methods=['POST'])
@require_api_key
def save_work_schedule_status_period():
    """
    Сохранить специальный статус оператора на период.
    Body: {
        "operator_id": int,
        "status_code": "bs" | "sick_leave" | "annual_leave" | "dismissal",
        "start_date": "YYYY-MM-DD",
        "end_date": "YYYY-MM-DD",              # required except dismissal; for dismissal optional (прерывание)
        "dismissal_reason": "...",             # required for dismissal
        "is_blacklist": true|false,            # optional: only for dismissal (ЧС / без восстановления)
        "comment": "..."                       # required for dismissal
        "range_start": "YYYY-MM-DD",           # optional: return operator snapshot for range
        "range_end": "YYYY-MM-DD"              # optional
    }
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404

        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        data = request.get_json(silent=True) or {}
        operator_id = data.get('operator_id')
        status_code = data.get('status_code')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        dismissal_reason = data.get('dismissal_reason')
        is_blacklist = data.get('is_blacklist')
        comment = data.get('comment')
        range_start = data.get('range_start')
        range_end = data.get('range_end')

        if not operator_id or not status_code or not start_date:
            return jsonify({"error": "Missing required fields"}), 400

        status_period = db.save_schedule_status_period(
            operator_id=operator_id,
            status_code=status_code,
            start_date=start_date,
            end_date=end_date,
            dismissal_reason=dismissal_reason,
            is_blacklist=is_blacklist,
            comment=comment,
            created_by=requester_id
        )

        operator_snapshot = None
        if range_start and range_end:
            operator_snapshot = db.get_operator_with_shifts(operator_id, range_start, range_end)

        return jsonify({
            "message": "Status period saved successfully",
            "status_period": status_period,
            "operator": operator_snapshot
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error saving status period: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/status_period', methods=['DELETE'])
@require_api_key
def delete_work_schedule_status_period():
    """
    Удалить специальный статус оператора (период) по id.
    Body: {
        "status_period_id": int,
        "operator_id": int,        # optional but recommended
        "range_start": "YYYY-MM-DD",
        "range_end": "YYYY-MM-DD"
    }
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404

        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        data = request.get_json(silent=True) or {}
        status_period_id = data.get('status_period_id')
        operator_id = data.get('operator_id')
        range_start = data.get('range_start')
        range_end = data.get('range_end')

        if not status_period_id:
            return jsonify({"error": "Missing status_period_id"}), 400

        deleted_period = db.delete_schedule_status_period(
            status_period_id=status_period_id,
            operator_id=operator_id
        )
        if not deleted_period:
            return jsonify({"error": "Status period not found"}), 404

        target_operator_id = operator_id or deleted_period.get('operatorId')
        operator_snapshot = None
        if target_operator_id and range_start and range_end:
            operator_snapshot = db.get_operator_with_shifts(target_operator_id, range_start, range_end)

        return jsonify({
            "message": "Status period deleted successfully",
            "deleted_status_period": deleted_period,
            "operator": operator_snapshot
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error deleting status period: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/shifts_bulk', methods=['POST'])
@require_api_key
def save_shifts_bulk():
    """
    Массовое сохранение смен для оператора.
    Body: {
        "operator_id": int,
        "shifts": [
            {
                "date": "YYYY-MM-DD",
                "start": "HH:MM",
                "end": "HH:MM",
                "breaks": [{"start": minutes, "end": minutes}, ...]  // optional
            },
            ...
        ]
    }
    """
    try:
        user = request.headers.get('X-User-Id')
        if not user:
            return jsonify({"error": "Unauthorized"}), 401
        
        user_id = int(user)
        user_data = db.get_user(id=user_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        role = user_data[3]
        
        if not (_is_admin_role(role) or _is_supervisor_role(role)):
            return jsonify({"error": "Forbidden"}), 403
        
        data = request.get_json(silent=True) or {}
        operator_id = data.get('operator_id')
        shifts = data.get('shifts', [])
        
        if not operator_id:
            return jsonify({"error": "Missing operator_id"}), 400
        
        if not shifts:
            return jsonify({"error": "No shifts provided"}), 400
        
        shift_ids = db.save_shifts_bulk(operator_id, shifts)
        
        return jsonify({
            "message": f"Successfully saved {len(shift_ids)} shifts",
            "shift_ids": shift_ids
        }), 200
    
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error saving shifts bulk: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/bulk_actions', methods=['POST'])
@require_api_key
def apply_work_schedule_bulk_actions():
    """
    Атомарные массовые действия по графикам в одном запросе.
    Body: {
        "actions": [
            {"action": "set_shift", "operator_id": 1, "date": "YYYY-MM-DD", "start": "HH:MM", "end": "HH:MM", "breaks": [...]},
            {"action": "set_day_off", "operator_id": 1, "date": "YYYY-MM-DD"},
            {"action": "delete_shifts", "operator_id": 1, "date": "YYYY-MM-DD"}  # очистка дня: смены + выходной
        ]
    }
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404

        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        data = request.get_json(silent=True) or {}
        actions = data.get('actions')
        result = db.apply_work_schedule_bulk_actions(actions)
        return jsonify({
            "message": "Bulk work schedule actions applied successfully",
            "result": result
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error applying bulk work schedule actions: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/aggregate', methods=['POST'])
@require_api_key
def aggregate_work_schedule_metrics():
    """
    Ручной запуск автоагрегации метрик по графикам/статусам.
    Body: {
        "operator_id": int,                 # optional if operator_ids provided
        "operator_ids": [int, ...],         # optional
        "start_date": "YYYY-MM-DD",         # required (or date)
        "end_date": "YYYY-MM-DD",           # optional (default=start_date)
        "date": "YYYY-MM-DD"                # optional alias for one-day range
    }
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        data = request.get_json(silent=True) or {}
        operator_id = data.get('operator_id')
        operator_ids_raw = data.get('operator_ids')
        day_value = data.get('date')
        start_date_input = data.get('start_date')
        end_date_input = data.get('end_date')
        start_date = start_date_input or day_value
        end_date = end_date_input or start_date

        # Для запуска "агрегировать день" считаем расширенное окно:
        # день-1..день+1, чтобы корректно обработать ночные смены и статусы,
        # которые перетекают через 00:00.
        expanded_cross_day = False
        if day_value and not start_date_input and not end_date_input:
            try:
                day_obj = datetime.strptime(str(day_value), '%Y-%m-%d').date()
            except Exception:
                return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
            start_date = (day_obj - timedelta(days=1)).strftime('%Y-%m-%d')
            end_date = (day_obj + timedelta(days=1)).strftime('%Y-%m-%d')
            expanded_cross_day = True

        operator_ids = []
        if isinstance(operator_ids_raw, list) and operator_ids_raw:
            for value in operator_ids_raw:
                operator_ids.append(int(value))
        elif operator_id is not None:
            operator_ids = [int(operator_id)]
        else:
            return jsonify({"error": "operator_id or operator_ids is required"}), 400

        if not start_date:
            return jsonify({"error": "start_date (or date) is required"}), 400
        if not end_date:
            end_date = start_date

        result = db.recalculate_auto_daily_hours(
            operator_ids=operator_ids,
            start_date=start_date,
            end_date=end_date
        )

        return jsonify({
            "message": "Aggregation completed",
            "result": result,
            "range": {
                "start_date": start_date,
                "end_date": end_date
            },
            "expanded_cross_day": bool(expanded_cross_day),
            "operator_ids": operator_ids
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error aggregating work schedule metrics: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/auto_flags/resolve', methods=['POST'])
@require_api_key
def resolve_work_schedule_auto_flag():
    """
    Подтверждение/отклонение авто-флага по дню.
    Body: {
        "operator_id": int,
        "day": "YYYY-MM-DD",
        "flag_type": "late" | "early_leave" | "training",
        "action": "confirm" | "reject" | "pending",
        "range_start": "YYYY-MM-DD",   # optional
        "range_end": "YYYY-MM-DD"      # optional
    }
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        data = request.get_json(silent=True) or {}
        operator_id = data.get('operator_id')
        day = data.get('day') or data.get('date')
        flag_type = data.get('flag_type')
        action = data.get('action')
        range_start = data.get('range_start')
        range_end = data.get('range_end')

        if not operator_id or not day or not flag_type or not action:
            return jsonify({"error": "Missing required fields"}), 400

        result = db.resolve_auto_schedule_flag(
            operator_id=operator_id,
            day=day,
            flag_type=flag_type,
            action=action
        )

        operator_snapshot = None
        if range_start and range_end:
            operator_snapshot = db.get_operator_with_shifts(
                int(operator_id),
                range_start,
                range_end,
                include_imported_statuses=False
            )

        return jsonify({
            "message": "Auto flag resolved successfully",
            "result": result,
            "operator": operator_snapshot
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error resolving work schedule auto flag: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/fines', methods=['POST'])
@require_api_key
def add_work_schedule_fine():
    """
    Добавить ручной штраф за день оператору из интерфейса графиков.
    Body: {
        "operator_id": int,
        "day": "YYYY-MM-DD" | "date": "YYYY-MM-DD",
        "amount": number,
        "reason": string,
        "comment": string(optional)
    }
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404

        requester_role = _normalize_user_role(user_data[3])
        if not (_is_admin_role(requester_role) or _is_supervisor_role(requester_role)):
            return jsonify({"error": "Forbidden"}), 403

        data = request.get_json(silent=True) or {}
        operator_id = data.get('operator_id')
        day = data.get('day') or data.get('date')
        amount = data.get('amount')
        reason = data.get('reason')
        comment = data.get('comment')

        if operator_id is None or not day:
            return jsonify({"error": "operator_id and day are required"}), 400

        operator_id_int = int(operator_id)
        operator_data = db.get_user(id=operator_id_int)
        if not operator_data or str(operator_data[3] or '').strip().lower() != 'operator':
            return jsonify({"error": "Operator not found"}), 404

        # supervisor_id в get_user находится в позиции 6
        operator_supervisor_id = operator_data[6]
        if _is_supervisor_role(requester_role) and operator_supervisor_id != requester_id:
            return jsonify({"error": "Forbidden for this operator"}), 403

        allowed_reasons = {'Корп такси', 'Опоздание', 'Прокси карта', 'Не выход', 'Другое'}
        reason_text = str(reason or '').strip()
        if reason_text not in allowed_reasons:
            return jsonify({"error": "Invalid fine reason"}), 400

        result = db.add_manual_fine_for_day(
            operator_id=operator_id_int,
            day=day,
            amount=amount,
            reason=reason_text,
            comment=comment
        )

        return jsonify({
            "message": "Fine added successfully",
            "result": result
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error adding work schedule fine: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/export_excel', methods=['GET'])
@require_api_key
def export_work_schedules_excel():
    """
    Экспорт графика в Excel-матрицу:
    ФИО | Ставка | YYYY-MM-DD даты (в файле выводятся как dd.mm.yyyy)
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        if not start_date or not end_date:
            return jsonify({"error": "start_date and end_date are required"}), 400

        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        if end_date_obj < start_date_obj:
            return jsonify({"error": "end_date must be >= start_date"}), 400

        operators = db.get_operators_with_shifts(start_date, end_date) or []
        operators = sorted(operators, key=lambda op: str(op.get('name') or '').lower())

        wb = Workbook()
        ws = wb.active
        ws.title = 'График'

        ws.cell(row=1, column=1, value='ФИО')
        ws.cell(row=1, column=2, value='Ставка')
        date_list = list(_iter_schedule_excel_dates(start_date_obj, end_date_obj))
        for idx, day_obj in enumerate(date_list, start=3):
            cell = ws.cell(row=1, column=idx, value=day_obj)
            cell.number_format = 'DD.MM.YYYY'

        for col_idx in range(1, 3 + len(date_list)):
            ws.cell(row=1, column=col_idx).font = ws.cell(row=1, column=col_idx).font.copy(bold=True)

        def _format_shift_cell_segment(shift_item):
            start = _format_schedule_excel_compact_time(shift_item.get('start'))
            end = _format_schedule_excel_compact_time(shift_item.get('end'))
            return f"{start}*{end}"

        for row_idx, op in enumerate(operators, start=2):
            ws.cell(row=row_idx, column=1, value=op.get('name') or '')
            rate_val = op.get('rate')
            ws.cell(row=row_idx, column=2, value=rate_val if rate_val is not None else '')

            shifts_by_date = op.get('shifts') or {}
            days_off_set = set(op.get('daysOff') or [])

            for col_idx, day_obj in enumerate(date_list, start=3):
                day_key = day_obj.strftime('%Y-%m-%d')
                shifts = shifts_by_date.get(day_key) or []
                if shifts:
                    ws.cell(row=row_idx, column=col_idx, value=','.join(_format_shift_cell_segment(s) for s in shifts))
                elif day_key in days_off_set:
                    ws.cell(row=row_idx, column=col_idx, value='Выходной')
                else:
                    ws.cell(row=row_idx, column=col_idx, value='')

        ws.freeze_panes = 'C2'
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 10
        for i in range(len(date_list)):
            col_letter = ws.cell(row=1, column=3 + i).column_letter
            ws.column_dimensions[col_letter].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"work_schedule_{start_date_obj.strftime('%Y%m%d')}_{end_date_obj.strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error exporting work schedules excel: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/import_excel', methods=['POST'])
@require_api_key
def import_work_schedules_excel():
    """
    Импорт графика из Excel-матрицы:
    - Row 1: ФИО, Ставка, даты...
    - Cells: "9*18", "10*16/30", "9*13,14*18" или "Выходной"
    """
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401

        requester_id = int(requester_id)
        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        file_storage = request.files.get('file')
        if not file_storage:
            return jsonify({"error": "file is required"}), 400

        file_name = secure_filename(file_storage.filename or 'schedule.xlsx')
        if not file_name.lower().endswith(('.xlsx', '.xlsm')):
            return jsonify({"error": "Only .xlsx/.xlsm files are supported"}), 400

        wb = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
        ws = wb.active
        if ws.max_row < 2 or ws.max_column < 3:
            return jsonify({"error": "Excel file is too small for schedule import"}), 400

        header_cells = list(ws[1])
        fio_col_idx = None
        for cell in header_cells:
            header_text = str(cell.value or '').strip().lower()
            if header_text == 'фио' or 'фио' in header_text:
                fio_col_idx = cell.column
                break
        if fio_col_idx is None:
            fio_col_idx = 1

        date_columns = []
        for cell in header_cells:
            parsed_date = _parse_schedule_excel_header_date(cell)
            if not parsed_date:
                continue
            date_columns.append((cell.column, parsed_date))

        if not date_columns:
            return jsonify({"error": "No date columns found in header row"}), 400

        date_columns.sort(key=lambda item: item[0])
        import_start = min(d for _, d in date_columns).strftime('%Y-%m-%d')
        import_end = max(d for _, d in date_columns).strftime('%Y-%m-%d')

        operators = db.get_operators_with_shifts(import_start, import_end) or []
        operator_exact_map = {}
        for op in operators:
            key = _normalize_schedule_excel_name_key(op.get('name'))
            if not key:
                continue
            operator_exact_map.setdefault(key, []).append(op)

        parsed_entries = []
        skipped_rows = 0
        skipped_empty_cells = 0
        unmatched_rows = []
        ambiguous_rows = []
        invalid_cells = []

        for row_idx in range(2, ws.max_row + 1):
            fio_value = ws.cell(row=row_idx, column=fio_col_idx).value
            fio_text = str(fio_value or '').strip()
            if not fio_text:
                continue

            op_match = None
            exact_candidates = operator_exact_map.get(_normalize_schedule_excel_name_key(fio_text), [])
            if len(exact_candidates) == 1:
                op_match = exact_candidates[0]
            elif len(exact_candidates) > 1:
                ambiguous_rows.append({'row': row_idx, 'name': fio_text, 'count': len(exact_candidates)})
                skipped_rows += 1
                continue
            else:
                # fallback на существующий fuzzy finder
                fuzzy = db.find_operator_by_name_fuzzy(fio_text)
                if fuzzy:
                    op_match = {
                        'id': fuzzy[0],
                        'name': fuzzy[2]
                    }

            if not op_match:
                unmatched_rows.append({'row': row_idx, 'name': fio_text})
                skipped_rows += 1
                continue

            operator_id = int(op_match['id'])
            for col_idx, day_obj in date_columns:
                raw_cell = ws.cell(row=row_idx, column=col_idx).value
                parsed_cell = _parse_schedule_excel_cell_value(raw_cell)
                kind = parsed_cell.get('kind')
                if kind == 'empty':
                    skipped_empty_cells += 1
                    continue
                if kind == 'invalid':
                    invalid_cells.append({
                        'row': row_idx,
                        'column': int(col_idx),
                        'date': day_obj.strftime('%Y-%m-%d'),
                        'name': fio_text,
                        'value': str(raw_cell or ''),
                        'error': parsed_cell.get('error') or 'Invalid cell format'
                    })
                    continue

                entry = {
                    'operator_id': operator_id,
                    'operator_name': str(op_match.get('name') or fio_text or '').strip(),
                    'date': day_obj.strftime('%Y-%m-%d'),
                    'is_day_off': (kind == 'day_off'),
                    'shifts': []
                }
                if kind == 'shifts':
                    entry['shifts'] = parsed_cell.get('shifts') or []
                parsed_entries.append(entry)

        if not parsed_entries and not invalid_cells:
            return jsonify({
                "message": "No schedule cells found to import",
                "result": {
                    "days_processed": 0,
                    "set_day_off_days": 0,
                    "set_shift_days": 0,
                    "shift_rows_saved": 0,
                    "deleted_shift_rows": 0,
                    "deleted_day_off_rows": 0,
                    "blacklist_skipped_entries": [],
                    "blacklist_skipped_total": 0,
                    "affected_operator_ids": []
                },
                "warnings": {
                    "skipped_rows": skipped_rows,
                    "skipped_empty_cells": skipped_empty_cells,
                    "unmatched_rows": unmatched_rows,
                    "ambiguous_rows": ambiguous_rows,
                    "invalid_cells": invalid_cells,
                    "blacklist_skipped_entries": [],
                    "blacklist_skipped_total": 0
                }
            }), 200

        if not parsed_entries and invalid_cells:
            return jsonify({
                "message": "No valid schedule cells found to import",
                "result": {
                    "days_processed": 0,
                    "set_day_off_days": 0,
                    "set_shift_days": 0,
                    "shift_rows_saved": 0,
                    "deleted_shift_rows": 0,
                    "deleted_day_off_rows": 0,
                    "blacklist_skipped_entries": [],
                    "blacklist_skipped_total": 0,
                    "affected_operator_ids": []
                },
                "warnings": {
                    "skipped_rows": skipped_rows,
                    "skipped_empty_cells": skipped_empty_cells,
                    "unmatched_rows": unmatched_rows,
                    "ambiguous_rows": ambiguous_rows,
                    "invalid_cells": invalid_cells[:30],
                    "invalid_cells_total": len(invalid_cells),
                    "blacklist_skipped_entries": [],
                    "blacklist_skipped_total": 0
                }
            }), 200

        break_direction_groups = []
        try:
            raw_groups = request.form.get('break_direction_groups')
            if raw_groups:
                parsed_groups = json.loads(raw_groups)
                if isinstance(parsed_groups, list):
                    break_direction_groups = parsed_groups
        except Exception:
            # Не валим импорт из-за битого optional-поля, просто игнорируем группы.
            break_direction_groups = []

        # Подготовка перерывов как во фронте: автогенерация + анти-пересечение между операторами
        # (по направлению / группе направлений) на симуляции перед сохранением.
        import_start_obj = datetime.strptime(import_start, '%Y-%m-%d').date()
        import_end_obj = datetime.strptime(import_end, '%Y-%m-%d').date()
        sim_start = (import_start_obj - timedelta(days=1)).strftime('%Y-%m-%d')
        sim_end = (import_end_obj + timedelta(days=1)).strftime('%Y-%m-%d')
        sim_source_operators = db.get_operators_with_shifts(sim_start, sim_end) or []
        sim_operators = _ws_clone_ops_for_break_simulation(sim_source_operators)
        sim_op_by_id = {int(op.get('id')): op for op in sim_operators if op.get('id') is not None}
        scope_resolver = _ws_make_direction_scope_resolver(break_direction_groups)
        break_rules_map = db.get_work_schedule_break_rules_map() or {}

        for entry in parsed_entries:
            op_id = int(entry.get('operator_id'))
            date_str = str(entry.get('date'))
            sim_op = sim_op_by_id.get(op_id)
            if not sim_op:
                sim_op = {'id': op_id, 'direction': None, 'shifts': {}, 'daysOff': []}
                sim_operators.append(sim_op)
                sim_op_by_id[op_id] = sim_op

            sim_op['daysOff'] = [d for d in (sim_op.get('daysOff') or []) if str(d) != date_str]
            sim_op.setdefault('shifts', {})
            sim_op['shifts'].pop(date_str, None)

            if entry.get('is_day_off'):
                if date_str not in sim_op['daysOff']:
                    sim_op['daysOff'].append(date_str)
                entry['shifts'] = []
                continue

            prepared_shifts = []
            for raw_shift in (entry.get('shifts') or []):
                start_str = str(raw_shift.get('start') or raw_shift.get('start_time') or '').strip()
                end_str = str(raw_shift.get('end') or raw_shift.get('end_time') or '').strip()
                if not start_str or not end_str:
                    continue
                smin = _ws_time_to_minutes(start_str)
                emin = _ws_time_to_minutes(end_str)
                if emin <= smin:
                    emin += 1440
                prepared_shifts.append({
                    'start': _ws_minutes_to_time(smin),
                    'end': _ws_minutes_to_time(emin),
                    '__startMin': smin,
                    '__endMin': emin,
                    'breaks': _ws_compute_breaks_for_shift_minutes(
                        smin,
                        emin,
                        sim_op.get('direction'),
                        break_rules_map=break_rules_map
                    )
                })

            sim_op['shifts'][date_str] = prepared_shifts
            if prepared_shifts:
                _ws_adjust_breaks_for_operator_on_date(
                    sim_op,
                    date_str,
                    sim_operators,
                    scope_resolver,
                    break_rules_map=break_rules_map
                )

            # Сохраняем в entry уже финальные перерывы после анти-пересечения
            final_shifts = []
            for seg in (sim_op.get('shifts', {}).get(date_str) or []):
                final_shifts.append({
                    'start': str(seg.get('start') or ''),
                    'end': str(seg.get('end') or ''),
                    'breaks': [
                        {'start': int(b.get('start')), 'end': int(b.get('end'))}
                        for b in (seg.get('breaks') or [])
                        if b is not None and int(b.get('end', 0)) > int(b.get('start', 0))
                    ]
                })
            entry['shifts'] = final_shifts

        result = db.import_work_schedule_excel_entries(parsed_entries)
        return jsonify({
            "message": "Work schedules imported from Excel successfully",
            "result": result,
            "breakAntiOverlapApplied": True,
            "warnings": {
                "skipped_rows": skipped_rows,
                "skipped_empty_cells": skipped_empty_cells,
                "unmatched_rows": unmatched_rows,
                "ambiguous_rows": ambiguous_rows,
                "invalid_cells": invalid_cells[:30],
                "invalid_cells_total": len(invalid_cells),
                "blacklist_skipped_entries": result.get('blacklist_skipped_entries') or [],
                "blacklist_skipped_total": int(result.get('blacklist_skipped_total') or 0)
            }
        }), 200

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error importing work schedules excel: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/api/work_schedules/import_statuses_csv', methods=['POST'])
@require_api_key
def import_work_schedules_statuses_csv():
    """
    Импорт CSV переключений статусов операторов с сохранением в БД:
    - operator_status_events (сырые события)
    - operator_status_segments (дневные интервалы для аналитики)
    """
    lock_acquired = False
    raw_bytes = b''
    try:
        requester_id = getattr(g, 'user_id', None) or request.headers.get('X-User-Id')
        if not requester_id:
            return jsonify({"error": "Unauthorized"}), 401
        requester_id = int(requester_id)

        user_data = db.get_user(id=requester_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        if not (_is_admin_role(user_data[3]) or _is_supervisor_role(user_data[3])):
            return jsonify({"error": "Forbidden"}), 403

        file_storage = request.files.get('file')
        if not file_storage:
            return jsonify({"error": "file is required"}), 400

        file_name = secure_filename(file_storage.filename or 'statuses.csv')
        if not file_name.lower().endswith('.csv'):
            return jsonify({"error": "Only .csv files are supported"}), 400

        content_length = request.content_length
        if content_length is not None and int(content_length) > STATUS_IMPORT_MAX_FILE_SIZE_BYTES:
            return jsonify({
                "error": f"Файл слишком большой. Лимит: {STATUS_IMPORT_MAX_FILE_SIZE_MB} MB"
            }), 413

        if not STATUS_IMPORT_LOCK.acquire(blocking=False):
            return jsonify({
                "error": "Импорт статусов уже выполняется другим пользователем. Повторите через несколько секунд."
            }), 429
        lock_acquired = True

        read_limit = int(STATUS_IMPORT_MAX_FILE_SIZE_BYTES) + 1
        if getattr(file_storage, 'stream', None) is not None:
            raw_bytes = file_storage.stream.read(read_limit)
        else:
            raw_bytes = file_storage.read(read_limit)
        if not raw_bytes:
            return jsonify({"error": "Файл пустой"}), 400
        if len(raw_bytes) > STATUS_IMPORT_MAX_FILE_SIZE_BYTES:
            return jsonify({
                "error": f"Файл слишком большой. Лимит: {STATUS_IMPORT_MAX_FILE_SIZE_MB} MB"
            }), 413

        csv_text = None
        for enc in ('utf-8-sig', 'cp1251'):
            try:
                csv_text = raw_bytes.decode(enc)
                break
            except Exception:
                continue
        if csv_text is None:
            return jsonify({"error": "Не удалось декодировать CSV (поддерживаются UTF-8 и CP1251)"}), 400

        operator_lookup = _status_import_build_operator_lookup()
        parsed = _status_import_parse_csv(
            csv_text,
            operator_lookup,
            max_source_rows=STATUS_IMPORT_MAX_SOURCE_ROWS,
            invalid_rows_preview_limit=STATUS_IMPORT_INVALID_ROWS_PREVIEW_LIMIT
        )

        if not parsed.get('events'):
            return jsonify({
                "message": "Нет валидных событий для сохранения",
                "import": {
                    "batch_id": None,
                    "source_rows": int(parsed.get('source_rows') or 0),
                    "valid_events": int(parsed.get('valid_events') or 0),
                    "matched_events": 0,
                    "segments_saved": 0,
                    "deleted_events": 0,
                    "deleted_segments": 0,
                    "invalid_rows_count": int(parsed.get('invalid_rows_count') or 0),
                    "parse_errors_count": int(parsed.get('parse_errors_count') or 0),
                    "operators_count": 0,
                    "open_tail_events": int(parsed.get('open_tail_events') or 0),
                    "zero_or_negative_transitions": int(parsed.get('zero_or_negative_transitions') or 0),
                    "date_from": None,
                    "date_to": None
                },
                "warnings": {
                    "invalid_rows_preview": parsed.get('invalid_rows_preview') or []
                }
            }), 200

        save_summary = db.save_operator_status_import(
            events=parsed.get('events') or [],
            segments=parsed.get('segments') or [],
            imported_by=requester_id,
            summary={
                'source_rows': int(parsed.get('source_rows') or 0),
                'valid_events': int(parsed.get('valid_events') or 0),
                'invalid_rows_count': int(parsed.get('invalid_rows_count') or 0),
                'parse_errors_count': int(parsed.get('parse_errors_count') or 0),
                'open_tail_events': int(parsed.get('open_tail_events') or 0),
                'zero_or_negative_transitions': int(parsed.get('zero_or_negative_transitions') or 0),
                'meta': {
                    'api': 'import_statuses_csv',
                    'file_size_bytes': len(raw_bytes or b'')
                }
            }
        )

        return jsonify({
            "message": "Статусы операторов сохранены в БД",
            "import": save_summary,
            "warnings": {
                "invalid_rows_preview": parsed.get('invalid_rows_preview') or []
            }
        }), 200

    except OverflowError as e:
        return jsonify({"error": str(e)}), 413
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.error(f"Error importing statuses csv: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500
    finally:
        if lock_acquired:
            STATUS_IMPORT_LOCK.release()


def run_flask():
    app.run(host='0.0.0.0', port=int(os.getenv('PORT', 8080)), debug=False, use_reloader=False)

# === Классы =====================================================================================================
class new_sv(StatesGroup):
    svname = State()
    svid = State()

class new_operator(StatesGroup):
    opname = State()
    opid = State()
    svselect = State()

class sv(StatesGroup):
    delete = State()
    delete_operator = State()  # Новое состояние
    view_evaluations = State()
    change_table = State()

class ChangeCredentials(StatesGroup):
    waiting_for_value = State()
    waiting_for_new_login = State()
    waiting_for_new_password = State()
    waiting_for_current_password = State()
    

class Auth(StatesGroup):
    login = State()
    password = State()


class sv_edit(StatesGroup):
    choose_action = State()
    edit_login = State()
    edit_password = State()
    edit_status = State()


class operator_edit(StatesGroup):
    choose_action = State()
    edit_login = State()
    edit_password = State()
    change_sv = State()

MAX_LOGIN_ATTEMPTS = 3

def get_access_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Сменить логин'))
    kb.insert(KeyboardButton('Сменить пароль'))
    kb.add(KeyboardButton('Выход🚪')) 
    kb.add(KeyboardButton('Назад 🔙'))
    return kb

def get_current_week_of_month():
    today = datetime.now()
    week_number = (today.day - 1) // 7 + 1
    return week_number

def get_expected_calls(week_number):
    return week_number * 5

def get_cancel_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(KeyboardButton('Отмена ❌'))
    return kb

def get_admin_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Редактор СВ📝'))
    kb.insert(KeyboardButton('Операторы👷'))
    kb.add(KeyboardButton('Данные📈'))
    kb.add(KeyboardButton('Доступ🔑'))
    return kb

def get_data_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Оценки📊'))
    kb.insert(KeyboardButton('Часы⏱️'))
    kb.add(KeyboardButton('Назад 🔙'))
    return kb

def get_evaluations_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Отчет за месяц📅'))
    kb.add(KeyboardButton('Назад 🔙'))
    return kb

def get_sv_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Добавить таблицу часов📊'))
    kb.add(KeyboardButton('Управление операторами🔑'))
    kb.add(KeyboardButton('Доступ🔑'))
    return kb

def get_verify_keyboard():
    ikb = InlineKeyboardMarkup(row_width=2)
    ikb.add(
        InlineKeyboardButton("Да ✅", callback_data="verify_yes"),
        InlineKeyboardButton("Нет ❌", callback_data="verify_no")
    )
    return ikb

def get_direction_keyboard():
    keyboard = types.InlineKeyboardMarkup(row_width=3)
    directions = db.get_directions()  # Запрашиваем направления из базы данных
    if not directions:
        return None  # Если направлений нет, возвращаем None
    buttons = [
        types.InlineKeyboardButton(
            f"{direction['name']} {'📄' if direction['hasFileUpload'] else '📝'}",
            callback_data=f"dir_{direction['id']}"
        )
        for direction in directions
    ]
    keyboard.add(*buttons)
    return keyboard

def get_editor_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Добавить СВ➕'))
    kb.insert(KeyboardButton('Редактировать СВ✏️'))
    kb.add(KeyboardButton('Назад 🔙'))
    return kb

def get_operators_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Добавить оператора👷‍♂️'))
    kb.insert(KeyboardButton('Редактировать Оператора✏️'))
    kb.add(KeyboardButton('Назад 🔙'))
    return kb

def get_operator_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Моя статистика📊'))
    kb.add(KeyboardButton('Мои оценки📝'))
    kb.add(KeyboardButton('Доступ🔑'))
    return kb

@dp.message_handler(regexp='Отмена ❌', state='*')
async def cancel_handler(message: types.Message, state: FSMContext):
    logging.info(f"User {message.from_user.id} canceled the operation.")
    current_state = await state.get_state()
    if current_state is None:
        return
    await state.finish()
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        kb = get_admin_keyboard()
    elif user and user[3] == 'sv':
        kb = get_sv_keyboard()
    elif user and user[3] == 'operator':
        kb = get_operator_keyboard()
    else:
        kb = ReplyKeyboardRemove()
    await bot.send_message(
        chat_id=message.from_user.id,
        text="Действие отменено.",
        parse_mode='HTML',
        reply_markup=kb
    )
    await message.delete()

# === Команды ====================================================================================================
@dp.message_handler(commands=['start'])
async def start_command(message: types.Message):
    await message.delete()
    user = db.get_user(telegram_id=message.from_user.id)
    if user:
        if _is_admin_role(user[3]):
            await bot.send_message(
                chat_id=message.from_user.id,
                text="<b>Бобро пожаловать!</b>\nЭто бот для прослушки прослушек.",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
        elif user[3] == 'sv':
            await bot.send_message(
                chat_id=message.from_user.id,
                text=f"<b>Бобро пожаловать, {user[2]}!</b>",
                parse_mode='HTML',
                reply_markup=get_sv_keyboard()
            )
        elif user[3] == 'operator':
            await bot.send_message(
                chat_id=message.from_user.id,
                text=f"<b>Добро пожаловать, оператор {user[2]}!</b>\n\n"
                    "Используйте кнопки ниже для просмотра вашей статистики.",
                parse_mode='HTML',
                reply_markup=get_operator_keyboard()
            )
    else:
        kb = ReplyKeyboardMarkup(resize_keyboard=True)
        kb.add(KeyboardButton('Вход👤'))
        await bot.send_message(
            chat_id=message.from_user.id,
            text=f"<b>Бобро пожаловать!</b>\n\nНажмите <b>Вход👤</b>, чтобы подключиться к <b>OTP dashboard</b>. 👥",
            parse_mode='HTML',
            reply_markup=kb
        )

@dp.message_handler(regexp='Вход👤')
async def start_auth(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user:
        await bot.send_message(
            chat_id=message.from_user.id,
            text=f"<b>Вы уже авторизованы как {user[2]} ({user[3]})</b>.\n\n",
            parse_mode='HTML',
            reply_markup=ReplyKeyboardRemove()
        )
        await message.delete()
        return
    
    await message.delete()
    await message.answer(
        "<b>Введите ваш логин:</b>",
        parse_mode='HTML',
        reply_markup=get_cancel_keyboard()
    )
    await Auth.login.set()
    # Инициализация счетчика попыток
    await dp.storage.set_data(chat=message.chat.id, data={'attempts': 0})

@dp.message_handler(state=Auth.login)
async def process_login(message: types.Message, state: FSMContext):
    login = message.text.strip()
    try:
        user = db.get_user_by_login(login)
        if not user:
            await message.delete()
            await message.answer(
                "<b>Неверный логин. Попробуйте снова🔁</b>",
                parse_mode='HTML'
            )
            return

        # Сохраняем данные пользователя в состоянии
        await state.update_data({
            'user': {
                'id': user[0],
                'telegram_id': user[1],
                'name': user[2],
                'role': user[3],
                'direction': user[4],
                'hire_date': user[5],
                'supervisor_id': user[6],
                'login': user[7]
            }
        })
        await message.delete()
        await message.answer(
            "<b>Введите ваш пароль:</b>",
            parse_mode='HTML'
        )
        await Auth.password.set()
    except Exception as e:
        logging.error(f"Ошибка при обработке логина {login}: {str(e)}")
        await message.delete()
        await message.answer(
            "<b>Произошла ошибка. Попробуйте снова или свяжитесь с поддержкой.</b>",
            parse_mode='HTML'
        )
        await state.finish()
        
@dp.message_handler(state=Auth.password)
async def process_password(message: types.Message, state: FSMContext):
    password = message.text.strip()
    user_data = await state.get_data()
    user = user_data.get('user')
    attempts = (await dp.storage.get_data(chat=message.chat.id)).get('attempts', 0)

    try:
        if attempts >= MAX_LOGIN_ATTEMPTS:
            await message.delete()
            await message.answer(
                "<b>Слишком много попыток. Авторизация заблокирована. Свяжитесь с поддержкой.</b>",
                parse_mode='HTML'
            )
            await state.finish()
            await dp.storage.reset_data(chat=message.chat.id)
            logging.warning(f"Превышено количество попыток авторизации для chat_id {message.chat.id}")
            return

        if not user or not db.verify_password(user['id'], password):
            attempts += 1
            await dp.storage.set_data(chat=message.chat.id, data={'attempts': attempts})
            await message.delete()
            await message.answer(
                f"<b>Неверный пароль. Осталось попыток: {MAX_LOGIN_ATTEMPTS - attempts}. Попробуйте снова.</b>",
                parse_mode='HTML'
            )
            logging.warning(f"Неверный пароль для логина {user.get('login')} (попытка {attempts})")
            return

        # Успешная авторизация
        db.update_telegram_id(user['id'], message.from_user.id)
        await message.delete()

        # Формируем приветственное сообщение в зависимости от роли
        role = user['role']
        name = user['name']
        if _is_admin_role(role):
            await message.answer(
                "<b>Бобро пожаловать!</b>\nЭто бот для управления прослушками.",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
        elif role == 'sv':
            await message.answer(
                f"<b>Бобро пожаловать, {name}!</b>",
                parse_mode='HTML',
                reply_markup=get_sv_keyboard()
            )
        elif role == 'operator':
            await message.answer(
                text=f"<b>Добро пожаловать, оператор {name}!</b>\n\n"
                    "Используйте кнопки ниже для просмотра вашей статистики.",
                parse_mode='HTML',
                reply_markup=get_operator_keyboard()
            )

        logging.info(f"Успешная авторизация: login={user['login']}, role={role}, chat_id={message.chat.id}")
        await state.finish()
        await dp.storage.reset_data(chat=message.chat.id)

    except Exception as e:
        logging.error(f"Ошибка при обработке пароля для login {user.get('login')}: {str(e)}")
        await message.delete()
        await message.answer(
            "<b>Произошла ошибка. Попробуйте снова или свяжитесь с СВ.</b>",
            parse_mode='HTML'
        )
        await state.finish()

# === Админка ===================================================================================================

@dp.message_handler(regexp='Редактор СВ📝')
async def editor_sv(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        await bot.send_message(
            chat_id=message.from_user.id,
            text='<b>Редактор супервайзеров</b>',
            parse_mode='HTML',
            reply_markup=get_editor_keyboard()
        )
    await message.delete()

@dp.message_handler(regexp='Операторы👷')
async def operators_menu(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        await bot.send_message(
            chat_id=message.from_user.id,
            text='<b>Управление операторами</b>',
            parse_mode='HTML',
            reply_markup=get_operators_keyboard()
        )
    await message.delete()

@dp.message_handler(regexp='Назад 🔙')
async def back_to_admin(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        await bot.send_message(
            chat_id=message.from_user.id,
            text='<b>Главное меню</b>',
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )
    elif user and user[3] == 'sv':
        await bot.send_message(
            chat_id=message.from_user.id,
            text=f'<b>Главное меню</b>',
            parse_mode='HTML',
            reply_markup=get_sv_keyboard()
        )
    elif user and user[3] == 'operator':
        await bot.send_message(
            chat_id=message.from_user.id,
            text=f'<b>Главное меню</b>',
            parse_mode='HTML',
            reply_markup=get_operator_keyboard()
        )
    await message.delete()


@dp.message_handler(regexp='Добавить СВ➕')
async def newSv(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        await bot.send_message(
            text='<b>Добавление СВ, этап</b>: 1 из 2📍\n\nФИО нового СВ🖊',
            chat_id=message.from_user.id,
            parse_mode='HTML',
            reply_markup=get_cancel_keyboard()
        )
        await new_sv.svname.set()
    await message.delete()

@dp.message_handler(state=new_sv.svname)
async def newSVname(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['svname'] = message.text
    
    # Генерируем случайные логин и пароль
    login = f"sv_{str(uuid.uuid4())[:8]}"
    password = str(uuid.uuid4())[:8]
    
    async with state.proxy() as data:
        data['login'] = login
        data['password'] = password
    
    await message.answer(
        text=f'<b>Данные для нового СВ:</b>\n\n'
             f'ФИО: <b>{message.text}</b>\n'
             f'Логин: <code>{login}</code>\n'
             f'Пароль: <code>{password}</code>\n\n'
             f'Передайте эти данные супервайзеру. Он сможет войти в систему и добавить '
             f'остальную информацию самостоятельно.\n\n'
             f'<b>Хотите сохранить этого супервайзера?</b>',
        parse_mode='HTML',
        reply_markup=get_verify_keyboard()
    )
    await new_sv.next()
    await message.delete()

@dp.callback_query_handler(state=new_sv.svid)
async def newSVid(callback: types.CallbackQuery, state: FSMContext):
    if callback.data == "verify_yes":
        async with state.proxy() as data:
            sv_name = data['svname']
            login = data['login']
            password = data['password']
        
        # Создаем супервайзера без telegram_id (он добавит его при первом входе)
        sv_id = db.create_user(
            telegram_id=None,  # Будет установлен при первом входе
            name=sv_name,
            role='sv',
            login=login,
            password=password
        )
        
        await bot.send_message(
            chat_id=callback.from_user.id,
            text=f'<b>Супервайзер {sv_name} успешно добавлен!</b>\n\n'
                 f'Логин: <code>{login}</code>\n'
                 f'Пароль: <code>{password}</code>\n\n'
                 f'Передайте эти данные супервайзеру для входа в систему.',
            parse_mode='HTML',
            reply_markup=get_editor_keyboard()
        )
    else:
        await bot.send_message(
            chat_id=callback.from_user.id,
            text='Добавление супервайзера отменено.',
            parse_mode='HTML',
            reply_markup=get_editor_keyboard()
        )
    
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
    await state.finish()


@dp.message_handler(regexp='Добавить оператора👷‍♂️')
async def newOperator(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        await bot.send_message(
            text='<b>Добавление оператора, этап</b>: 1 из 3📍\n\nФИО нового оператора🖊',
            chat_id=message.from_user.id,
            parse_mode='HTML',
            reply_markup=get_cancel_keyboard()
        )
        await new_operator.opname.set()
    await message.delete()

@dp.message_handler(state=new_operator.opname)
async def newOperatorName(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['opname'] = message.text
    await message.answer(
        text=f'Класс, ФИО - <b>{message.text}</b>\n\n<b>Добавление оператора, этап</b>: 2 из 3📍\n\nНапишите <b>ID</b> нового оператора🆔',
        parse_mode='HTML',
        reply_markup=get_cancel_keyboard()
    )
    await new_operator.next()
    await message.delete()

@dp.message_handler(state=new_operator.opid)
async def newOperatorId(message: types.Message, state: FSMContext):
    try:
        op_id = int(message.text)
        async with state.proxy() as data:
            data['opid'] = op_id
            
        supervisors = db.get_supervisors()
        if not supervisors:
            await message.answer(
                text='Нет доступных супервайзеров! Сначала добавьте СВ.',
                parse_mode='HTML',
                reply_markup=get_cancel_keyboard()
            )
            await state.finish()
            return
            
        ikb = InlineKeyboardMarkup(row_width=1)
        for sv in supervisors:
            ikb.insert(InlineKeyboardButton(text=sv[1], callback_data=str(sv[0])))
            
        await message.answer(
            text=f'Отлично, ID - <b>{message.text}</b>\n\n<b>Добавление оператора, этап</b>: 3 из 3📍\n\nВыберите супервайзера для этого оператора:',
            parse_mode='HTML',
            reply_markup=ikb
        )
        await new_operator.next()
    except:
        await message.answer(
            text='Ой, похоже вы отправили не тот <b>ID</b>❌\n\n<b>Пожалуйста повторите попытку!</b>',
            parse_mode='HTML',
            reply_markup=get_cancel_keyboard()
        )
    await message.delete()

@dp.callback_query_handler(state=new_operator.svselect)
async def newOperatorSV(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        op_name = data['opname']
        op_id = data['opid']
        sv_id = int(callback.data)
        
        # Создаем оператора
        db.create_user(
            telegram_id=op_id,
            name=op_name,
            role='operator',
            supervisor_id=sv_id
        )
        
        await bot.send_message(
            chat_id=op_id,
            text=f"Вы добавлены как оператор в команду <b>успешно✅</b>",
            parse_mode='HTML'
        )
        
        await bot.send_message(
            chat_id=callback.from_user.id,
            text=f'Оператор <b>{op_name}</b> успешно добавлен✅',
            parse_mode='HTML',
            reply_markup=get_operators_keyboard()
        )
    await state.finish()
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.message_handler(regexp='Редактировать СВ✏️')
async def editSv(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        supervisors = db.get_supervisors()
        if not supervisors:
            await bot.send_message(
                chat_id=message.from_user.id,
                text="<b>Нет доступных супервайзеров</b>",
                parse_mode='HTML',
                reply_markup=get_editor_keyboard()
            )
            return

        ikb = InlineKeyboardMarkup(row_width=1)
        for sv_id, sv_name, *_ in supervisors:
            ikb.insert(InlineKeyboardButton(text=sv_name, callback_data=f"editsv_{sv_id}"))

        await bot.send_message(
            chat_id=message.from_user.id,
            text="<b>Выберите супервайзера для редактирования</b>",
            parse_mode='HTML',
            reply_markup=ikb
        )
    await message.delete()


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editsv_'))
async def editSV_select(callback: types.CallbackQuery, state: FSMContext):
    try:
        sv_id = int(callback.data.split('_')[1])
        user = db.get_user(id=sv_id)
        if not user:
            await bot.answer_callback_query(callback.id, text="Супервайзер не найден")
            return

        ikb = InlineKeyboardMarkup(row_width=1)
        ikb.add(
            InlineKeyboardButton('Сменить логин', callback_data=f'edit_login_{sv_id}'),
            InlineKeyboardButton('Сменить пароль', callback_data=f'edit_pass_{sv_id}'),
        )
        ikb.add(InlineKeyboardButton('Изменить статус', callback_data=f'edit_status_{sv_id}'))
        ikb.add(InlineKeyboardButton('Отмена', callback_data='edit_cancel'))

        await bot.send_message(
            chat_id=callback.from_user.id,
            text=(f"<b>Супервайзер:</b> {user[2]}\n" 
                  f"<b>Текущий статус:</b> {user[9] if len(user) > 9 else 'unknown'}\n\n"
                  "Выберите действие:"),
            parse_mode='HTML',
            reply_markup=ikb
        )
    finally:
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('edit_login_'))
async def editSV_login_start(callback: types.CallbackQuery, state: FSMContext):
    sv_id = int(callback.data.split('_')[2])
    await state.update_data({'sv_edit_id': sv_id})
    await sv_edit.edit_login.set()
    await bot.send_message(chat_id=callback.from_user.id, text='Введите новый логин супервайзера:')
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.message_handler(state=sv_edit.edit_login)
async def process_sv_login_change(message: types.Message, state: FSMContext):
    data = await state.get_data()
    sv_id = data.get('sv_edit_id')
    new_login = message.text.strip()
    try:
        success = db.update_operator_login(sv_id, None, new_login)
        if success:
            await bot.send_message(chat_id=message.from_user.id, text=f'Логин супервайзера обновлён на: <code>{new_login}</code>', parse_mode='HTML', reply_markup=get_editor_keyboard())
        else:
            await bot.send_message(chat_id=message.from_user.id, text='Не удалось изменить логин супервайзера')
    except Exception as e:
        logging.error(f"Error updating supervisor login: {e}")
        await bot.send_message(chat_id=message.from_user.id, text='Ошибка при обновлении логина')
    finally:
        await state.finish()
        await message.delete()


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('edit_pass_'))
async def editSV_pass_start(callback: types.CallbackQuery, state: FSMContext):
    sv_id = int(callback.data.split('_')[2])
    await state.update_data({'sv_edit_id': sv_id})
    await sv_edit.edit_password.set()
    await bot.send_message(chat_id=callback.from_user.id, text='Введите новый пароль супервайзера:')
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.message_handler(state=sv_edit.edit_password)
async def process_sv_password_change(message: types.Message, state: FSMContext):
    data = await state.get_data()
    sv_id = data.get('sv_edit_id')
    new_pass = message.text.strip()
    try:
        success = db.update_user_password(sv_id, new_pass)
        if success:
            await bot.send_message(chat_id=message.from_user.id, text='Пароль супервайзера успешно обновлён', reply_markup=get_editor_keyboard())
        else:
            await bot.send_message(chat_id=message.from_user.id, text='Не удалось изменить пароль супервайзера')
    except Exception as e:
        logging.error(f"Error updating supervisor password: {e}")
        await bot.send_message(chat_id=message.from_user.id, text='Ошибка при обновлении пароля')
    finally:
        await state.finish()
        await message.delete()


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('edit_status_'))
async def editSV_status_menu(callback: types.CallbackQuery):
    sv_id = int(callback.data.split('_')[2])
    ikb = InlineKeyboardMarkup(row_width=2)
    ikb.add(
        InlineKeyboardButton('Работает', callback_data=f'edit4_status_set_{sv_id}_working'),
        InlineKeyboardButton('Уволен', callback_data=f'edit4_status_set_{sv_id}_fired')
    )
    ikb.add(InlineKeyboardButton('БС', callback_data=f'edit4_status_set_{sv_id}_unpaid_leave'))
    ikb.add(InlineKeyboardButton('Отмена', callback_data='edit_cancel'))
    await bot.send_message(chat_id=callback.from_user.id, text='Выберите новый статус для СВ:', reply_markup=ikb)
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('edit4_status_set_'))
async def editSV_status_set(callback: types.CallbackQuery):
    parts = callback.data.split('_')
    sv_id = int(parts[3])
    new_status = parts[4]+'_' + parts[5] if len(parts) > 5 else parts[4]
    try:
        requester = db.get_user(telegram_id=callback.from_user.id)
        changed_by = requester[0] if requester else None
        success = db.update_user(sv_id, 'status', new_status, changed_by=changed_by)
        if success:
            await bot.send_message(chat_id=callback.from_user.id, text=f'Статус супервайзера обновлён на: {new_status}', reply_markup=get_editor_keyboard())
        else:
            await bot.send_message(chat_id=callback.from_user.id, text='Не удалось изменить статус супервайзера')
    except Exception as e:
        logging.error(f"Error updating supervisor status: {e}")
        await bot.send_message(chat_id=callback.from_user.id, text='Ошибка при обновлении статуса')
    finally:
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data == 'edit_cancel')
async def editSV_cancel(callback: types.CallbackQuery):
    await bot.answer_callback_query(callback.id, text='Действие отменено')
    await bot.send_message(chat_id=callback.from_user.id, text='Отмена', reply_markup=get_editor_keyboard())
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)

@dp.message_handler(regexp='Данные📈')
async def view_data_menu(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        await bot.send_message(
            chat_id=message.from_user.id,
            text='<b>Меню данных</b>',
            parse_mode='HTML',
            reply_markup=get_data_keyboard()
        )
    await message.delete()

@dp.message_handler(regexp='Часы⏱️')
async def view_hours_data(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        supervisors = db.get_supervisors()
        if not supervisors:
            await bot.send_message(
                chat_id=message.from_user.id,
                text="<b>Нет доступных супервайзеров</b>",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
            return
        
        ikb = InlineKeyboardMarkup(row_width=1)
        for sv_id, sv_name, _, _, _, status in supervisors:
            if status != 'fired':
                ikb.insert(InlineKeyboardButton(text=sv_name, callback_data=f"hours_{sv_id}"))
        
        await bot.send_message(
            chat_id=message.from_user.id,
            text="<b>Выберите супервайзера для просмотра часов операторов:</b>",
            parse_mode='HTML',
            reply_markup=ikb
        )
    await message.delete()

# Обработчик для отображения часов операторов
@dp.callback_query_handler(lambda c: c.data.startswith('hours_'))
async def show_operator_hours(callback: types.CallbackQuery):
    sv_id = int(callback.data.split('_')[1])
    user = db.get_user(id=sv_id)
    
    if not user or user[3] != 'sv':
        await bot.answer_callback_query(callback.id, text="Ошибка: СВ не найден")
        return

    operators = db.get_operators_by_supervisor(sv_id)
    current_month = datetime.now().strftime('%Y-%m')
    
    message_text = f"<b>Часы операторов {user[2]} за {current_month}:</b>\n\n"
    
    for op in operators:
        op_id = op.get('id')
        op_name = op.get('name')
        hours_data = db.get_hours_summary(op_id, current_month)
        if hours_data:
            hours = hours_data[0]
            regular_hours = round(hours.get('regular_hours', 0) or 0, 2)
            training_hours = round(hours.get('training_hours', 0) or 0, 2)
            technical_issue_hours = round(hours.get('technical_issue_hours', 0) or 0, 2)
            offline_activity_hours = round(hours.get('offline_activity_hours', 0) or 0, 2)
            accounted_hours = round(hours.get('accounted_hours', regular_hours + training_hours + technical_issue_hours + offline_activity_hours) or 0, 2)
            norm_hours = hours.get('norm_hours', 0) or 0
            percent_complete = 0
            if norm_hours > 0:
                percent_complete = round(accounted_hours / norm_hours * 100, 1)
            message_text += (
                f"👤 <b>{op_name}</b>\n"
                f"   ⏱️ Часы работы: {accounted_hours} из {norm_hours}\n"
                f"   📈 Процент выполнения: {percent_complete}%\n"
                f"   📚 Часы тренинга: {training_hours}\n"
                f"   🛠 Тех. сбои: {technical_issue_hours}\n"
                f"   👤 Офлайн активность: {offline_activity_hours}\n"
                f"   💸 Штрафы: {hours.get('fines', 0)}\n\n"
            )
        else:
            message_text += f"👤 <b>{op_name}</b> - данные отсутствуют\n\n"
    logging.info(message_text)
    await bot.send_message(
        chat_id=callback.from_user.id,
        text=message_text,
        parse_mode='HTML'
    )
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.message_handler(regexp='Оценки📊')
async def view_evaluations(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        supervisors = db.get_supervisors()
        if not supervisors:
            await bot.send_message(
                chat_id=message.from_user.id,
                text="<b>Нет доступных супервайзеров</b>",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
            return
        
        await bot.send_message(
                    text='<b>Выберите чьи оценки просмотреть или сгенерируйте отчет</b>',
                    chat_id=user[1],
                    parse_mode='HTML',
                    reply_markup=get_evaluations_keyboard()
                )
        ikb = InlineKeyboardMarkup(row_width=1)
        for sv_id, sv_name, _, _, _, status in supervisors:
            if status != 'fired':
                ikb.insert(InlineKeyboardButton(text=sv_name, callback_data=f"eval_{sv_id}"))
        
        await bot.send_message(
            chat_id=message.from_user.id,
            text="<b>Лист СВ:</b>",
            parse_mode='HTML',
            reply_markup=ikb
        )
        await sv.view_evaluations.set()
    await message.delete()

@dp.message_handler(regexp='Редактировать Оператора✏️')
async def edit_operator_menu(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        supervisors = db.get_supervisors()
        if not supervisors:
            await bot.send_message(
                chat_id=message.from_user.id,
                text="<b>Нет доступных супервайзеров</b>",
                parse_mode='HTML',
                reply_markup=get_operators_keyboard()
            )
            return

        ikb = InlineKeyboardMarkup(row_width=1)
        for sv_id, sv_name, *_ in supervisors:
            ikb.insert(InlineKeyboardButton(text=sv_name, callback_data=f'editop_sv_{sv_id}'))

        await bot.send_message(
            chat_id=message.from_user.id,
            text="<b>Выберите СВ, чтобы показать его операторов</b>",
            parse_mode='HTML',
            reply_markup=ikb
        )
    await message.delete()


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop_sv_'))
async def edit_operator_by_sv(callback: types.CallbackQuery, state: FSMContext):
    sv_id = int(callback.data.split('_')[2])
    sv = db.get_user(id=sv_id)
    if not sv or sv[3] != 'sv':
        await bot.answer_callback_query(callback.id, text='СВ не найден')
        return

    operators = db.get_operators_by_supervisor(sv_id)
    if not operators:
        await bot.send_message(chat_id=callback.from_user.id, text='У выбранного СВ нет операторов')
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
        return

    ikb = InlineKeyboardMarkup(row_width=1)
    for op in operators:
        op_id = op.get('id')
        op_name = op.get('name')
        ikb.insert(InlineKeyboardButton(text=op_name, callback_data=f'editop1_{op_id}'))

    await bot.send_message(chat_id=callback.from_user.id, text=f"Операторы СВ: <b>{sv[2]}</b>", parse_mode='HTML', reply_markup=ikb)
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop1_'))
async def edit_operator_select(callback: types.CallbackQuery, state: FSMContext):
    op_id = int(callback.data.split('_')[1])
    user = db.get_user(id=op_id)
    if not user or user[3] != 'operator':
        await bot.answer_callback_query(callback.id, text='Оператор не найден')
        return

    ikb = InlineKeyboardMarkup(row_width=1)
    ikb.add(
        InlineKeyboardButton('Сменить логин', callback_data=f'editop_login_{op_id}'),
        InlineKeyboardButton('Сменить пароль', callback_data=f'editop_pass_{op_id}'),
        InlineKeyboardButton('Сменить ставку', callback_data=f'editop_rate_{op_id}'),
        InlineKeyboardButton('Изменить статус', callback_data=f'editop_status_{op_id}')
    )
    ikb.add(InlineKeyboardButton('Отмена', callback_data='editop_cancel'))

    await bot.send_message(chat_id=callback.from_user.id, text=(f"Оператор: <b>{user[2]}</b>\nВыберите действие:"), parse_mode='HTML', reply_markup=ikb)
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop_login_'))
async def edit_operator_login_start(callback: types.CallbackQuery, state: FSMContext):
    op_id = int(callback.data.split('_')[2])
    await state.update_data({'edit_op_id': op_id})
    await operator_edit.edit_login.set()
    await bot.send_message(chat_id=callback.from_user.id, text='Введите новый логин оператора:')
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.message_handler(state=operator_edit.edit_login)
async def process_operator_login_change(message: types.Message, state: FSMContext):
    data = await state.get_data()
    op_id = data.get('edit_op_id')
    new_login = message.text.strip()
    try:
        success = db.update_operator_login(op_id, None, new_login)
        if success:
            await bot.send_message(chat_id=message.from_user.id, text=f'Логин оператора обновлён на: <code>{new_login}</code>', parse_mode='HTML', reply_markup=get_operators_keyboard())
        else:
            await bot.send_message(chat_id=message.from_user.id, text='Не удалось изменить логин оператора')
    except Exception as e:
        logging.error(f"Error updating operator login: {e}")
        await bot.send_message(chat_id=message.from_user.id, text='Ошибка при обновлении логина')
    finally:
        await state.finish()
        await message.delete()


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop_pass_'))
async def edit_operator_pass_start(callback: types.CallbackQuery, state: FSMContext):
    op_id = int(callback.data.split('_')[2])
    await state.update_data({'edit_op_id': op_id})
    await operator_edit.edit_password.set()
    await bot.send_message(chat_id=callback.from_user.id, text='Введите новый пароль оператора:')
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.message_handler(state=operator_edit.edit_password)
async def process_operator_password_change(message: types.Message, state: FSMContext):
    data = await state.get_data()
    op_id = data.get('edit_op_id')
    new_pass = message.text.strip()
    try:
        success = db.update_user_password(op_id, new_pass)
        if success:
            await bot.send_message(chat_id=message.from_user.id, text='Пароль оператора успешно обновлён', reply_markup=get_operators_keyboard())
        else:
            await bot.send_message(chat_id=message.from_user.id, text='Не удалось изменить пароль оператора')
    except Exception as e:
        logging.error(f"Error updating operator password: {e}")
        await bot.send_message(chat_id=message.from_user.id, text='Ошибка при обновлении пароля')
    finally:
        await state.finish()
        await message.delete()


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop_rate_'))
async def edit_operator_rate_menu(callback: types.CallbackQuery):
    op_id = int(callback.data.split('_')[2])
    ikb = InlineKeyboardMarkup(row_width=3)
    ikb.add(
        InlineKeyboardButton('0.5', callback_data=f'editop2_rate_set_{op_id}_0.5'),
        InlineKeyboardButton('0.75', callback_data=f'editop2_rate_set_{op_id}_0.75'),
        InlineKeyboardButton('1', callback_data=f'editop2_rate_set_{op_id}_1')
    )
    ikb.add(InlineKeyboardButton('Отмена', callback_data='editop_cancel'))
    await bot.send_message(chat_id=callback.from_user.id, text='Выберите новую ставку для оператора:', reply_markup=ikb)
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop2_rate_set_'))
async def edit_operator_rate_set(callback: types.CallbackQuery):
    parts = callback.data.split('_')
    op_id = int(parts[3])
    rate = float(parts[4])
    try:
        requester = db.get_user(telegram_id=callback.from_user.id)
        changed_by = requester[0] if requester else None
        success = db.update_user(op_id, 'rate', rate, changed_by=changed_by)
        if success:
            await bot.send_message(chat_id=callback.from_user.id, text=f'Ставка оператора обновлена на {rate}', reply_markup=get_operators_keyboard())
        else:
            await bot.send_message(chat_id=callback.from_user.id, text='Не удалось изменить ставку оператора')
    except Exception as e:
        logging.error(f"Error updating operator rate: {e}")
        await bot.send_message(chat_id=callback.from_user.id, text='Ошибка при обновлении ставки')
    finally:
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop_status_'))
async def edit_operator_status_menu(callback: types.CallbackQuery):
    op_id = int(callback.data.split('_')[2])
    ikb = InlineKeyboardMarkup(row_width=2)
    ikb.add(
        InlineKeyboardButton('Работает', callback_data=f'editop2_status_set_{op_id}_working'),
        InlineKeyboardButton('Уволен', callback_data=f'editop2_status_set_{op_id}_fired')
    )
    ikb.add(InlineKeyboardButton('БС', callback_data=f'editop2_status_set_{op_id}_unpaid_leave'))
    ikb.add(InlineKeyboardButton('Отмена', callback_data='editop_cancel'))
    await bot.send_message(chat_id=callback.from_user.id, text='Выберите новый статус для оператора:', reply_markup=ikb)
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop2_status_set_'))
async def edit_operator_status_set(callback: types.CallbackQuery):
    parts = callback.data.split('_')
    op_id = int(parts[3])
    new_status = parts[4]+'_' + parts[5] if len(parts) > 5 else parts[4]
    try:
        requester = db.get_user(telegram_id=callback.from_user.id)
        changed_by = requester[0] if requester else None
        success = db.update_user(op_id, 'status', new_status, changed_by=changed_by)
        if success:
            await bot.send_message(chat_id=callback.from_user.id, text=f'Статус оператора обновлён на: {new_status}', reply_markup=get_operators_keyboard())
        else:
            await bot.send_message(chat_id=callback.from_user.id, text='Не удалось изменить статус оператора')
    except Exception as e:
        logging.error(f"Error updating operator status: {e}")
        await bot.send_message(chat_id=callback.from_user.id, text='Ошибка при обновлении статуса')
    finally:
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop_change_sv_'))
async def edit_operator_change_sv_start(callback: types.CallbackQuery, state: FSMContext):
    op_id = int(callback.data.split('_')[2])
    supervisors = db.get_supervisors()
    if not supervisors:
        await bot.answer_callback_query(callback.id, text='Нет доступных СВ')
        return

    ikb = InlineKeyboardMarkup(row_width=1)
    for sv_id, sv_name, *_ in supervisors:
        ikb.insert(InlineKeyboardButton(text=sv_name, callback_data=f'editop_setsv_{op_id}_{sv_id}'))

    ikb.add(InlineKeyboardButton('Отмена', callback_data='editop_cancel'))
    await bot.send_message(chat_id=callback.from_user.id, text='Выберите нового СВ для оператора:', reply_markup=ikb)
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data and c.data.startswith('editop_setsv_'))
async def edit_operator_set_sv(callback: types.CallbackQuery):
    parts = callback.data.split('_')
    op_id = int(parts[2])
    sv_id = int(parts[3])
    try:
        requester = db.get_user(telegram_id=callback.from_user.id)
        changed_by = requester[0] if requester else None
        success = db.update_user(op_id, 'supervisor_id', sv_id, changed_by=changed_by)
        if success:
            await bot.send_message(chat_id=callback.from_user.id, text='Супервайзер оператора успешно обновлён', reply_markup=get_operators_keyboard())
        else:
            await bot.send_message(chat_id=callback.from_user.id, text='Не удалось сменить супервайзера')
    except Exception as e:
        logging.error(f"Error setting operator supervisor: {e}")
        await bot.send_message(chat_id=callback.from_user.id, text='Ошибка при смене СВ')
    finally:
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.callback_query_handler(lambda c: c.data in ['editop_cancel'])
async def editop_cancel(callback: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback.id, text='Действие отменено')
    await bot.send_message(chat_id=callback.from_user.id, text='Отмена', reply_markup=get_operators_keyboard())
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)


@dp.message_handler(regexp='Отчет за месяц📅')
async def handle_monthly_report(message: types.Message, state: FSMContext):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and _is_admin_role(user[3]):
        try:
            await bot.send_message(
                chat_id=message.from_user.id,
                text="📊 Генерирую отчет за месяц...",
                parse_mode='HTML'
            )
            await generate_weekly_report()
            await bot.send_message(
                chat_id=message.from_user.id,
                text="✅ Отчет успешно сгенерирован и отправлен!",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
        except Exception as e:
            logging.error(f"Ошибка генерации отчета: {e}")
            await bot.send_message(
                chat_id=message.from_user.id,
                text=f"❌ Ошибка генерации отчета: {str(e)}",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
    else:
        await bot.send_message(
            chat_id=message.from_user.id,
            text="⚠️ Команда доступна только администратору",
            parse_mode='HTML'
        )
    await state.finish()
    await message.delete()

@dp.message_handler(regexp='Назад 🔙', state=sv.view_evaluations)
async def back_from_evaluations(message: types.Message, state: FSMContext):
    await state.finish()
    await back_to_admin(message)


@dp.callback_query_handler(lambda c: c.data.startswith('eval_'), state=sv.view_evaluations)
async def show_sv_evaluations(callback: types.CallbackQuery, state: FSMContext):
    sv_id = int(callback.data.split('_')[1])
    user = db.get_user(id=sv_id)
    
    if not user or user[3] != 'sv':
        await bot.answer_callback_query(callback.id, text="Ошибка: СВ не найден")
        return

    # Получаем операторов этого СВ
    operators = db.get_operators_by_supervisor(sv_id)
    current_week = get_current_week_of_month()
    expected_calls = get_expected_calls(current_week)
    
    message_text = (
        f"<b>Оценки {user[2]} (неделя {current_week}):</b>\n"
        f"<i>Ожидается: {expected_calls} звонков (по 5 в неделю)</i>\n\n"
    )
    
    operators_with_issues = []
    
    for op in operators:
        # Для каждого оператора получаем статистику звонков
        if op.get('status') == 'fired':
            continue
        op_id = op.get('id')
        op_name = op.get('name')
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*), AVG(score) 
                FROM calls 
                WHERE operator_id = %s AND month=%s
            """, (op_id, datetime.now().strftime('%Y-%m')))
            result = cursor.fetchone()
        
        call_count = result[0] or 0
        avg_score = result[1] or 0
        
        # Форматируем данные для отображения
        if call_count < expected_calls:
            operators_with_issues.append({
                'name': op_name,
                'call_count': call_count,
                'expected': expected_calls
            })
            
        # Добавляем в общее сообщение
        message_text += f"👤 {op_name}\n"
        message_text += f"   📞 Звонков: {call_count}/{expected_calls}\n"
        message_text += f"   ⭐ Средний балл: {avg_score:.2f}\n\n"
    
    # Клавиатура для уведомлений
    ikb = InlineKeyboardMarkup(row_width=1)
    for op in operators_with_issues:
        ikb.add(InlineKeyboardButton(
            text=f"Уведомить о {op['name']}",
            callback_data=f"notify_{sv_id}_{1}"))
    
    await bot.send_message(
        chat_id=callback.from_user.id,
        text=message_text,
        parse_mode='HTML',
        reply_markup=ikb
    )
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
    await state.finish()

@dp.callback_query_handler(lambda c: c.data.startswith('notify_'))
async def notify_supervisor_handler(callback: types.CallbackQuery):
    try:
        _, sv_id, op_name = callback.data.split('_', 2)
        sv_id = int(sv_id)
        
        user = db.get_user(id=sv_id)
        if not user or user[3] != 'sv':
            await bot.answer_callback_query(callback.id, text="Ошибка: СВ не найден")
            return

        current_week = get_current_week_of_month()
        expected_calls = get_expected_calls(current_week)
        
        notification_text = (
            f"⚠️ <b>Требуется внимание!</b>\n\n"
            f"У оператора <b>{op_name}</b> недостаточно прослушанных звонков.\n"
            f"Текущая неделя: {current_week}\n"
            f"Ожидается: {expected_calls} звонков\n\n"
            f"Пожалуйста, проверьте и прослушайте недостающие звонки."
        )
        
        await bot.send_message(
            chat_id=user[1],
            text=notification_text,
            parse_mode='HTML'
        )
        
        await bot.answer_callback_query(
            callback.id,
            text=f"Уведомление отправлено СВ {user[2]}",
            show_alert=False
        )
    except Exception as e:
        logging.error(f"Ошибка уведомления: {e}")
        await bot.answer_callback_query(
            callback.id,
            text="Ошибка отправки уведомления",
            show_alert=True
        )

async def _is_admin_user(tg_id):
    """Проверка админа: сначала сравнение с переменной admin, затем fallback — проверка в БД."""
    try:
        if os.getenv("ADMIN_ID") is not None and str(tg_id) == str(os.getenv("ADMIN_ID")):
            return True

        # fallback: проверить в users по telegram_id и role='admin'
        loop = asyncio.get_event_loop()
        def _check():
            with db._get_cursor() as cur:
                cur.execute("SELECT role FROM users WHERE telegram_id = %s LIMIT 1", (tg_id,))
                row = cur.fetchone()
                return bool(row and row[0] == 'admin')
        return await loop.run_in_executor(None, _check)
    except Exception as e:
        logging.exception("Error checking admin: %s", e)
        return False

async def _approve_call_and_get_notify_row(call_id, approver_tg_id):
    """
    Пометить звонок одобренным и вернуть данные для уведомлений:
    (sv_request_by, sv_tg, evaluator_id, eval_tg)
    """
    loop = asyncio.get_event_loop()

    def _update():
        with db._get_cursor() as cur:
            # получить internal user id админа (если есть)
            cur.execute("SELECT id FROM users WHERE telegram_id = %s LIMIT 1", (approver_tg_id,))
            row = cur.fetchone()
            approver_user_id = row[0] if row else None

            cur.execute("""
                UPDATE calls
                SET sv_request_approved = TRUE,
                    sv_request_approved_by = %s,
                    sv_request_approved_at = %s
                WHERE id = %s
            """, (approver_user_id, datetime.utcnow(), call_id))

            cur.execute("""
                SELECT c.sv_request_by, u_super.telegram_id AS sv_tg, c.evaluator_id, u_eval.telegram_id AS eval_tg
                FROM calls c
                LEFT JOIN users u_super ON u_super.id = c.sv_request_by
                LEFT JOIN users u_eval ON u_eval.id = c.evaluator_id
                WHERE c.id = %s
            """, (call_id,))
            return cur.fetchone()

    result = await loop.run_in_executor(None, _update)
    return result  # None или кортеж (sv_request_by, sv_tg, evaluator_id, eval_tg)

@dp.callback_query_handler(lambda c: c.data and c.data.startswith('approve_reval:'))
async def handle_approve_reval(callback_query: types.CallbackQuery):
    """
    Обработка approve_reval:{call_id}.
    Требования:
      - admin (telegram id) доступен в переменной `admin`.
      - db доступен и синхронен (используем run_in_executor).
    """
    cq = callback_query
    user = cq.from_user
    data = cq.data

    # извлечь call_id
    try:
        _, call_id_str = data.split(':', 1)
        call_id = int(call_id_str)
    except Exception:
        await cq.answer("Неверный идентификатор звонка", show_alert=True)
        return

    # проверить, что нажал админ
    try:
        is_admin = await _is_admin_user(user.id)
    except Exception as e:
        logging.exception("Ошибка проверки администратора: %s", e)
        is_admin = False

    if not is_admin:
        await cq.answer("Только администратор может одобрять переоценку", show_alert=True)
        return

    # пометим звонок как одобренный и получим данные для уведомлений
    try:
        notify_row = await _approve_call_and_get_notify_row(call_id, user.id)
    except Exception as e:
        logging.exception("Ошибка при пометке звонка: %s", e)
        await cq.answer("Ошибка сервера при одобрении", show_alert=True)
        return

    # ответ на callback (чтобы кнопка показала результат)
    try:
        await cq.answer("Переоценка одобрена администратором", show_alert=False)
    except Exception as e:
        logging.debug("Не удалось послать answerCallbackQuery: %s", e)

    # убрать inline-кнопку (если возможно)
    try:
        if cq.message:
            await cq.bot.edit_message_reply_markup(chat_id=cq.message.chat.id, message_id=cq.message.message_id, reply_markup=None)
    except Exception as e:
        logging.debug("Не удалось удалить reply_markup: %s", e)

    # уведомить супервайзера и оценивающего (если есть tg id)
    if notify_row:
        try:
            sv_request_by, sv_tg, evaluator_id, eval_tg = notify_row
            if sv_tg:
                try:
                    await cq.bot.send_message(int(sv_tg),
                        f"✅ Ваша заявка на переоценку для Call ID {call_id} одобрена администратором. Можете инициировать переоценку.")
                except Exception as e:
                    logging.debug("Не удалось уведомить супервайзера: %s", e)
            if eval_tg:
                try:
                    await cq.bot.send_message(int(eval_tg),
                        f"ℹ️ Админ одобрил запрос на переоценку для Call ID {call_id}. Для переоценки создайте новую оценку с is_correction=true и previous_version_id={call_id}.")
                except Exception as e:
                    logging.debug("Не удалось уведомить оценивающего: %s", e)
        except Exception as e:
            logging.exception("Ошибка при отправке уведомлений после approve: %s", e)

    # (опционально) логирование
    logging.info("Call %s approved by tg_user %s", call_id, user.id)


# === Супервайзерам =============================================================================================

@dp.message_handler(regexp='Управление операторами🔑')
async def manage_operators_credentials(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and user[3] == 'sv':
        operators = db.get_operators_by_supervisor(user[0])
        if not operators:
            await bot.send_message(
                chat_id=message.from_user.id,
                text="<b>У вас нет операторов</b>",
                parse_mode='HTML',
                reply_markup=ReplyKeyboardMarkup(resize_keyboard=True).add(KeyboardButton('Назад 🔙'))
            )
            return

        ikb = InlineKeyboardMarkup(row_width=1)
        for op in operators:
            ikb.insert(InlineKeyboardButton(text=op['name'], callback_data=f"cred_{op['id']}"))
        
        await bot.send_message(
            chat_id=message.from_user.id,
            text="<b>Выберите оператора для управления доступом:</b>",
            parse_mode='HTML',
            reply_markup=ikb
        )
    await message.delete()

@dp.callback_query_handler(lambda c: c.data.startswith('cred_'))
async def operator_credentials_menu(callback: types.CallbackQuery):
    operator_id = int(callback.data.split('_')[1])
    user = db.get_user(telegram_id=callback.from_user.id)
    
    if user and user[3] == 'sv':
        operator = db.get_operator_credentials(operator_id, user[0])
        if not operator:
            await bot.answer_callback_query(callback.id, text="Оператор не найден")
            return

        ikb = InlineKeyboardMarkup(row_width=2)
        ikb.add(
            InlineKeyboardButton("Изменить логин", callback_data=f"chlogin_{operator_id}"),
            InlineKeyboardButton("Изменить пароль", callback_data=f"chpass_{operator_id}")
        )
        ikb.add(InlineKeyboardButton("Назад", callback_data="cred_back"))

        await bot.edit_message_text(
            chat_id=callback.from_user.id,
            message_id=callback.message.message_id,
            text=f"<b>Управление доступом оператора</b>\n\nЛогин: <code>{operator[1]}</code>",
            parse_mode='HTML',
            reply_markup=ikb
        )

@dp.callback_query_handler(lambda c: c.data.startswith('chlogin_'))
async def change_login_start(callback: types.CallbackQuery, state: FSMContext):
    operator_id = int(callback.data.split('_')[1])
    await state.update_data(operator_id=operator_id, action="login")
    await bot.edit_message_text(
        chat_id=callback.from_user.id,
        message_id=callback.message.message_id,
        text="Введите новый логин для оператора:",
        reply_markup=None
    )
    await ChangeCredentials.waiting_for_value.set()

@dp.callback_query_handler(lambda c: c.data.startswith('chpass_'))
async def change_password_start(callback: types.CallbackQuery, state: FSMContext):
    operator_id = int(callback.data.split('_')[1])
    await state.update_data(operator_id=operator_id, action="password")
    await bot.edit_message_text(
        chat_id=callback.from_user.id,
        message_id=callback.message.message_id,
        text="Введите новый пароль для оператора:",
        reply_markup=None
    )
    await ChangeCredentials.waiting_for_value.set()

@dp.message_handler(state=ChangeCredentials.waiting_for_value)
async def process_credential_change(message: types.Message, state: FSMContext):
    user_data = await state.get_data()
    operator_id = user_data['operator_id']
    action = user_data['action']
    value = message.text.strip()
    user = db.get_user(telegram_id=message.from_user.id)

    try:
        if action == "login":
            success = db.update_operator_login(operator_id, user[0], value)
            msg = f"Логин оператора изменён на: <code>{value}</code>"
        else:
            success = db.update_operator_password(operator_id, user[0], value)
            msg = "Пароль оператора успешно изменён"
        
        if success:
            operator = db.get_user(id=operator_id)
            await bot.send_message(
                chat_id=message.from_user.id,
                text=f"✅ {msg}\n\nОператор: {operator[2]}",
                parse_mode='HTML',
                reply_markup=get_sv_keyboard()
            )
        else:
            await bot.send_message(
                chat_id=message.from_user.id,
                text="❌ Не удалось изменить данные оператора",
                parse_mode='HTML'
            )
    except Exception as e:
        logging.error(f"Error changing operator {action}: {e}")
        await bot.send_message(
            chat_id=message.from_user.id,
            text=f"❌ Ошибка при изменении: {str(e)}",
            parse_mode='HTML'
        )
    
    await state.finish()
    await message.delete()

@dp.callback_query_handler(lambda c: c.data == "cred_back", state="*")
async def credentials_back(callback: types.CallbackQuery, state: FSMContext):
    await state.finish()
    await manage_operators_credentials(callback.message)


@dp.message_handler(regexp='Добавить таблицу часов📊')
async def add_hours_table(message: types.Message, state: FSMContext):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and user[3] == 'sv':
        await bot.send_message(
            text='<b>Отправьте ссылку на таблицу с часами работы операторов:</b>',
            chat_id=message.from_user.id,
            parse_mode='HTML',
            reply_markup=get_cancel_keyboard()
        )
        await state.set_state("waiting_for_hours_table")

# Обработчик для сохранения таблицы часов
@dp.message_handler(state="waiting_for_hours_table")
async def save_hours_table(message: types.Message, state: FSMContext):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and user[3] == 'sv':
        try:
            sheet_name, operators, error = extract_fio_and_links(message.text)
            if error:
                await bot.send_message(
                    chat_id=message.from_user.id,
                    text=f"{error}\n\n<b>Пожалуйста, отправьте корректную ссылку на таблицу.</b>",
                    parse_mode="HTML",
                    reply_markup=get_cancel_keyboard()
                )
                return

            async with state.proxy() as data:
                data['hours_table_url'] = message.text
                data['operators'] = operators
                data['sheet_name'] = sheet_name
                data['sv_id'] = user[0]

            message_text = f"<b>Название листа:</b> {sheet_name}\n\n<b>ФИО операторов:</b>\n"
            for op in operators:
                message_text += f"👤 {op['name']}\n"
            message_text += "\n<b>Это все ваши операторы?</b>"

            await bot.send_message(
                chat_id=message.from_user.id,
                text=message_text,
                parse_mode="HTML",
                reply_markup=get_verify_keyboard(),
                disable_web_page_preview=True
            )
            await state.set_state("verify_hours_table")
            await message.delete()
        except Exception as e:
            await bot.send_message(
                chat_id=message.from_user.id,
                text=f"<b>Ошибка при обработке таблицы: {str(e)}</b>",
                parse_mode='HTML'
            )
            await state.finish()

@dp.callback_query_handler(state="verify_hours_table")
async def verify_hours_table(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        hours_table_url = data.get('hours_table_url')
        sv_id = data.get('sv_id')
        operators = data.get('operators')
        sheet_name = data.get('sheet_name')

    user = db.get_user(id=sv_id)
    if not user:
        await bot.answer_callback_query(callback.id, text="Ошибка: СВ не найден")
        await state.finish()
        return

    if callback.data == "verify_yes":
        await bot.send_message(
            chat_id=callback.from_user.id,
            text="Выберите направление для операторов:",
            reply_markup=get_direction_keyboard()
        )
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
        await state.set_state("select_hours_direction")
    elif callback.data == "verify_no":
        await bot.send_message(
            chat_id=callback.from_user.id,
            text=f'<b>Отправьте корректную таблицу часов для {user[2]}🖊</b>',
            parse_mode='HTML',
            reply_markup=get_cancel_keyboard()
        )
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
        await state.set_state("waiting_for_hours_table")

@dp.callback_query_handler(state="select_hours_direction")
async def select_hours_direction(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        hours_table_url = data.get('hours_table_url')
        sv_id = data.get('sv_id')
        operators = data.get('operators')

    # Извлекаем direction_id из callback_data
    direction_id = None
    if callback.data.startswith("dir_"):
        try:
            direction_id = int(callback.data.replace("dir_", ""))
        except ValueError:
            await bot.answer_callback_query(callback.id, text="Ошибка: Неверный формат направления")
            return

    direction = next((d for d in db.get_directions() if d['id'] == direction_id), None)
    if not direction:
        await bot.answer_callback_query(callback.id, text="Ошибка: Направление не найдено")
        await state.finish()
        return

    user = db.get_user(id=sv_id)
    if not user:
        await bot.answer_callback_query(callback.id, text="Ошибка: СВ не найден")
        await state.finish()
        return

    # Обновляем таблицу часов супервайзера
    db.update_user_table(user[0], hours_table_url=hours_table_url)

    # Создаём/обновляем операторов с direction_id
    for op in operators:
        db.create_user(
            telegram_id=None,
            name=op['name'],
            role='operator',
            direction_id=direction_id,
            supervisor_id=user[0]
        )

    await bot.send_message(
        chat_id=callback.from_user.id,
        text=f"""<b>Таблица часов сохранена, операторы добавлены/обновлены с направлением "{direction['name']}"✅</b>""",
        parse_mode='HTML',
        reply_markup=get_sv_keyboard()
    )
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
    await state.finish()


@dp.message_handler(regexp='Доступ🔑')
async def change_credentials_menu(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user:        
        await bot.send_message(
            chat_id=message.from_user.id,
            text="<b>Выберите что хотите изменить:</b>",
            parse_mode='HTML',
            reply_markup=get_access_keyboard()
        )
    await message.delete()

@dp.message_handler(regexp='Сменить логин')
async def change_login_start(message: types.Message):
    await bot.send_message(
        chat_id=message.from_user.id,
        text="<b>Введите новый логин:</b>",
        parse_mode='HTML',
        reply_markup=get_cancel_keyboard()
    )
    await ChangeCredentials.waiting_for_new_login.set()
    await message.delete()

@dp.message_handler(state=ChangeCredentials.waiting_for_new_login)
async def process_new_login(message: types.Message, state: FSMContext):
    new_login = message.text.strip()
    if len(new_login) < 4:
        await bot.send_message(
            chat_id=message.from_user.id,
            text="Логин должен быть не менее 4 символов. Попробуйте еще раз:",
            reply_markup=get_cancel_keyboard()
        )
        return
    
    user = db.get_user(telegram_id=message.from_user.id)
    if user:
        try:
            # Проверяем, не занят ли логин
            existing_user = db.get_user_by_login(new_login)
            if existing_user and existing_user[0] != user[0]:
                await bot.send_message(
                    chat_id=message.from_user.id,
                    text="Этот логин уже занят. Попробуйте другой:",
                    reply_markup=get_cancel_keyboard()
                )
                return
            
            # Обновляем логин
            with db._get_cursor() as cursor:
                cursor.execute("UPDATE users SET login = %s WHERE id = %s", (new_login, user[0]))
            
            await bot.send_message(
                chat_id=message.from_user.id,
                text=f"✅ Логин успешно изменен на: <code>{new_login}</code>",
                parse_mode='HTML'
            )
        except Exception as e:
            logging.error(f"Error changing login: {e}")
            await bot.send_message(
                chat_id=message.from_user.id,
                text="❌ Произошла ошибка при изменении логина. Попробуйте позже."
            )
    
    await state.finish()
    await message.delete()

@dp.message_handler(regexp='Сменить пароль')
async def change_password_start(message: types.Message):
    await bot.send_message(
        chat_id=message.from_user.id,
        text="<b>Введите текущий пароль для подтверждения:</b>",
        parse_mode='HTML',
        reply_markup=get_cancel_keyboard()
    )
    await ChangeCredentials.waiting_for_current_password.set()
    await message.delete()

@dp.message_handler(state=ChangeCredentials.waiting_for_current_password)
async def verify_current_password(message: types.Message, state: FSMContext):
    current_password = message.text.strip()
    user = db.get_user(telegram_id=message.from_user.id)
    
    if user and db.verify_password(user[0], current_password):
        await state.update_data(user_id=user[0])
        await bot.send_message(
            chat_id=message.from_user.id,
            text="<b>Введите новый пароль:</b>",
            parse_mode='HTML',
            reply_markup=get_cancel_keyboard()
        )
        await ChangeCredentials.waiting_for_new_password.set()
    else:
        await bot.send_message(
            chat_id=message.from_user.id,
            text="❌ Неверный текущий пароль. Попробуйте еще раз:",
            reply_markup=get_cancel_keyboard()
        )
    await message.delete()

@dp.message_handler(state=ChangeCredentials.waiting_for_new_password)
async def process_new_password(message: types.Message, state: FSMContext):
    new_password = message.text.strip()
    if len(new_password) < 6:
        await bot.send_message(
            chat_id=message.from_user.id,
            text="Пароль должен быть не менее 6 символов. Попробуйте еще раз:",
            reply_markup=get_cancel_keyboard()
        )
        return
    
    user_data = await state.get_data()
    user_id = user_data.get('user_id')
    
    if user_id:
        try:
            password_hash = pbkdf2_sha256.hash(new_password)
            with db._get_cursor() as cursor:
                cursor.execute("UPDATE users SET password_hash = %s WHERE id = %s", (password_hash, user_id))
            
            await bot.send_message(
                chat_id=message.from_user.id,
                text="✅ Пароль успешно изменен!"
            )
        except Exception as e:
            logging.error(f"Error changing password: {e}")
            await bot.send_message(
                chat_id=message.from_user.id,
                text="❌ Произошла ошибка при изменении пароля. Попробуйте позже."
            )
    
    await state.finish()
    await message.delete()

@dp.message_handler(regexp='Выход🚪')
async def logout_user(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user:
        # Обнуляем telegram_id в базе данных
        with db._get_cursor() as cursor:
            cursor.execute("UPDATE users SET telegram_id = NULL WHERE id = %s", (user[0],))
        await bot.send_message(  
            chat_id=message.from_user.id,
            text="✅ <b>Вы успешно вышли из системы.</b>Для входа снова нажмите 'Вход👤'.",
            parse_mode='HTML',
            reply_markup=ReplyKeyboardMarkup(resize_keyboard=True).add(KeyboardButton('Вход👤'))
            )
    else:
        await bot.send_message(
            chat_id=message.from_user.id,
            text="❌ Вы не вошли в систему.",
            parse_mode='HTML'
        )
    await message.delete()

# === Операторам =============================================================================================

@dp.message_handler(regexp='Моя статистика📊')
async def show_operator_stats(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and user[3] == 'operator':
        stats = db.get_operator_stats(user[0])
        current_month = datetime.now().strftime('%B %Y')
        accounted_hours = stats.get('accounted_hours', (stats.get('regular_hours', 0) or 0) + (stats.get('training_hours', 0) or 0) + (stats.get('technical_issue_hours', 0) or 0) + (stats.get('offline_activity_hours', 0) or 0))
        
        message_text = (
            f"<b>Ваша статистика за {current_month}:</b>\n\n"
            f"⏱ <b>Общие часы работы:</b> {accounted_hours} из {stats['norm_hours']} ({stats['percent_complete']}%)\n"
            f"📚 <b>Часы тренинга:</b> {stats['training_hours']}\n"
            f"🛠 <b>Тех. сбои:</b> {stats.get('technical_issue_hours', 0)}\n"
            f"👤 <b>Офлайн активность:</b> {stats.get('offline_activity_hours', 0)}\n"
            f"📞 <b>Количество звонков в час:</b> {stats['calls_per_hour']}\n"
            f"💸 <b>Штрафы:</b> {stats['fines']}\n\n"
            f"📞 <b>Прослушано звонков:</b> {stats['call_count']}\n"
            f"⭐ <b>Средний балл:</b> {stats['avg_score']:.2f}"
        )
        
        await bot.send_message(
            chat_id=message.from_user.id,
            text=message_text,
            parse_mode='HTML'
        )
    await message.delete()

@dp.message_handler(regexp='Мои оценки📝')
async def show_operator_evaluations(message: types.Message):
    user = db.get_user(telegram_id=message.from_user.id)
    if user and user[3] == 'operator':
        evaluations = db.get_call_evaluations(user[0], month=datetime.now().strftime('%Y-%m'))
        
        if not evaluations:
            await bot.send_message(
                chat_id=message.from_user.id,
                text="<b>У вас пока нет оценок за текущий месяц.</b>",
                parse_mode='HTML'
            )
            return
        
        message_text = "<b>Ваши последние оценки:</b>\n\n"
        for eval in evaluations[:5]:  # Показываем последние 5 оценок (уже только последние версии)
            correction_mark = " (корректировка)" if eval['is_correction'] else ""
            masked_phone = eval.get('phone_number_masked') or _mask_phone_number(eval.get('phone_number'))
            message_text += (
                f"📞 <b>Звонок {eval['id']}{correction_mark}</b>\n"
                f"   📅 {eval['month']}\n"
                f"   📱 {masked_phone}\n"
                f"   ⭐ Оценка: <b>{eval['score']}</b>\n"
                f"   🕒 Дата оценки: {eval['evaluation_date']}\n"
            )
            if eval['comment']:
                message_text += f"   💬 Комментарий: {eval['comment']}\n"
            message_text += "\n"
        
        await bot.send_message(
            chat_id=message.from_user.id,
            text=message_text,
            parse_mode='HTML'
        )
    await message.delete()

# Остальной код остается аналогичным предыдущей версии, но с использованием базы данных
# ... (код для удаления СВ, изменения таблиц, просмотра оценок и т.д.) ...

LMS_LEARNER_ROLES = ('operator', 'trainee')
LMS_MANAGER_ROLES = ('sv', 'trainer', 'admin', 'super_admin')
LMS_FULL_ADMIN_ROLES = ('admin', 'super_admin')
LMS_ALLOWED_ACCOUNTS = (
    ('super_admin', 2),
    ('operator', 56)
)


def _lms_is_allowed_account(user_id, role):
    role_norm = _normalize_user_role(role)
    try:
        uid = int(user_id)
    except Exception:
        return False
    return any(role_norm == item_role and uid == item_id for item_role, item_id in LMS_ALLOWED_ACCOUNTS)


def _lms_now():
    return datetime.now()


def _lms_parse_json(value, default):
    if value is None:
        return default
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return default
        try:
            return json.loads(raw)
        except Exception:
            return default
    return default


def _lms_to_float(value, default=0.0):
    try:
        return float(value)
    except Exception:
        return float(default)


def _lms_to_int(value, default=0):
    try:
        return int(value)
    except Exception:
        return int(default)


def _lms_normalize_skills(value, limit=30):
    if not isinstance(value, list):
        return []
    out = []
    for item in value:
        text = str(item or '').strip()
        if not text:
            continue
        if len(text) > 80:
            text = text[:80].strip()
        if text and text not in out:
            out.append(text)
        if len(out) >= limit:
            break
    return out


def _lms_parse_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        norm = value.strip().lower()
        if norm in ('1', 'true', 'yes', 'y', 'on'):
            return True
        if norm in ('0', 'false', 'no', 'n', 'off'):
            return False
    return bool(default)


def _lms_normalize_text(value):
    text = str(value or '').strip().lower()
    return re.sub(r'\s+', ' ', text)


def _lms_normalize_answer_text(value):
    text = _lms_normalize_text(value)
    text = re.sub(r'[^0-9a-zа-яё\s]+', ' ', text, flags=re.IGNORECASE)
    return re.sub(r'\s+', ' ', text).strip()


def _lms_parse_datetime(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    raw = str(value).strip()
    if not raw:
        return None
    direct_formats = (
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d'
    )
    for fmt in direct_formats:
        try:
            return datetime.strptime(raw, fmt)
        except Exception:
            continue
    try:
        parsed = datetime.fromisoformat(raw.replace('Z', '+00:00'))
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
        return parsed
    except Exception:
        return None


def _lms_deadline_status(due_at, completed_at):
    if completed_at:
        if due_at and completed_at > due_at:
            return 'orange'
        return 'green'
    if due_at and _lms_now() > due_at:
        return 'red'
    return None


def _lms_bucket_name():
    return (
        (os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET_LMS') or '').strip()
        or (os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET_TASKS') or '').strip()
        or (os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET') or '').strip()
    )


def _lms_signed_url(bucket_name, blob_path, expires_minutes=120):
    bucket_name = str(bucket_name or '').strip()
    blob_path = str(blob_path or '').strip()
    if not bucket_name or not blob_path:
        return None
    try:
        client = get_gcs_client()
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(blob_path)
        return blob.generate_signed_url(
            version='v4',
            expiration=timedelta(minutes=max(1, int(expires_minutes))),
            method='GET'
        )
    except Exception:
        return None


def _lms_collect_course_blob_refs_tx(cursor, course_id):
    refs = set()

    cursor.execute("""
        SELECT anti_cheat_settings
        FROM lms_course_versions
        WHERE course_id = %s
    """, (course_id,))
    for row in cursor.fetchall():
        settings = _lms_parse_json(row[0], {})
        if not isinstance(settings, dict):
            continue
        cover_bucket = str(settings.get('cover_bucket') or '').strip()
        cover_blob_path = str(settings.get('cover_blob_path') or '').strip()
        if cover_bucket and cover_blob_path:
            refs.add((cover_bucket, cover_blob_path))

    cursor.execute("""
        SELECT DISTINCT lm.gcs_bucket, lm.gcs_blob_path
        FROM lms_lesson_materials lm
        JOIN lms_lessons l ON l.id = lm.lesson_id
        JOIN lms_modules m ON m.id = l.module_id
        JOIN lms_course_versions cv ON cv.id = m.course_version_id
        WHERE cv.course_id = %s
          AND lm.gcs_bucket IS NOT NULL
          AND lm.gcs_blob_path IS NOT NULL
    """, (course_id,))
    for row in cursor.fetchall():
        bucket_name = str(row[0] or '').strip()
        blob_path = str(row[1] or '').strip()
        if bucket_name and blob_path:
            refs.add((bucket_name, blob_path))

    cursor.execute("""
        SELECT DISTINCT gcs_bucket, gcs_blob_path
        FROM lms_certificates
        WHERE course_id = %s
          AND pdf_storage_type = 'gcs'
          AND gcs_bucket IS NOT NULL
          AND gcs_blob_path IS NOT NULL
    """, (course_id,))
    for row in cursor.fetchall():
        bucket_name = str(row[0] or '').strip()
        blob_path = str(row[1] or '').strip()
        if bucket_name and blob_path:
            refs.add((bucket_name, blob_path))

    return sorted(refs)


def _lms_delete_blob_refs(blob_refs):
    unique_refs = []
    seen = set()
    for item in (blob_refs or []):
        if not isinstance(item, (tuple, list)) or len(item) < 2:
            continue
        bucket_name = str(item[0] or '').strip()
        blob_path = str(item[1] or '').strip()
        if not bucket_name or not blob_path:
            continue
        key = (bucket_name, blob_path)
        if key in seen:
            continue
        seen.add(key)
        unique_refs.append(key)

    if not unique_refs:
        return {"attempted": 0, "deleted": 0, "failed": []}

    client = get_gcs_client()
    bucket_cache = {}
    failed = []
    deleted_count = 0

    for bucket_name, blob_path in unique_refs:
        try:
            bucket = bucket_cache.get(bucket_name)
            if bucket is None:
                bucket = client.bucket(bucket_name)
                bucket_cache[bucket_name] = bucket
            bucket.blob(blob_path).delete()
            deleted_count += 1
        except Exception as delete_error:
            error_text = str(delete_error or '').lower()
            if '404' in error_text or 'not found' in error_text or 'no such object' in error_text:
                continue
            failed.append({
                "bucket": bucket_name,
                "blob_path": blob_path,
                "error": str(delete_error)
            })

    return {
        "attempted": len(unique_refs),
        "deleted": deleted_count,
        "failed": failed
    }


def _lms_convert_image_to_webp(raw_bytes, max_side=1600, quality=88):
    if Image is None:
        return None
    if not raw_bytes:
        return None
    try:
        with Image.open(BytesIO(raw_bytes)) as img:
            if ImageOps is not None:
                img = ImageOps.exif_transpose(img)
            if img.mode not in ('RGB', 'RGBA'):
                img = img.convert('RGBA' if 'A' in img.getbands() else 'RGB')
            width, height = img.size
            max_dim = max(int(width or 0), int(height or 0))
            if max_dim > int(max_side):
                scale = float(max_side) / float(max_dim)
                resize_to = (
                    max(1, int(round(width * scale))),
                    max(1, int(round(height * scale)))
                )
                resample = Image.Resampling.LANCZOS if hasattr(Image, 'Resampling') else Image.LANCZOS
                img = img.resize(resize_to, resample)
            out = BytesIO()
            img.save(out, format='WEBP', quality=max(40, min(95, int(quality))), method=6)
            return out.getvalue()
    except Exception:
        return None


def _lms_resolve_cover_payload(version_settings):
    settings = version_settings if isinstance(version_settings, dict) else {}
    cover_url = (str(settings.get("cover_url") or "").strip() or None)
    cover_bucket = (str(settings.get("cover_bucket") or "").strip() or None)
    cover_blob_path = (str(settings.get("cover_blob_path") or "").strip() or None)
    signed_cover_url = _lms_signed_url(cover_bucket, cover_blob_path, expires_minutes=240) if cover_bucket and cover_blob_path else None
    return {
        "cover_url": signed_cover_url or cover_url,
        "cover_bucket": cover_bucket,
        "cover_blob_path": cover_blob_path
    }


def _lms_emit_notification(cursor, user_id, notification_type, title, message=None, payload=None):
    cursor.execute("""
        INSERT INTO lms_notifications (user_id, notification_type, title, message, payload)
        VALUES (%s, %s, %s, %s, %s::jsonb)
    """, (
        int(user_id),
        str(notification_type or 'info').strip()[:50] or 'info',
        str(title or '').strip()[:255] or 'Уведомление LMS',
        (str(message or '').strip() or None),
        json.dumps(payload or {}, ensure_ascii=False)
    ))


def _lms_audit(cursor, actor_id, actor_role, action, entity_type, entity_id=None, details=None):
    cursor.execute("""
        INSERT INTO lms_admin_audit_log (actor_id, actor_role, action, entity_type, entity_id, details)
        VALUES (%s, %s, %s, %s, %s, %s::jsonb)
    """, (
        int(actor_id) if actor_id is not None else None,
        str(actor_role or '').strip()[:20] or None,
        str(action or '').strip()[:100] or 'unknown',
        str(entity_type or '').strip()[:50] or 'unknown',
        int(entity_id) if entity_id is not None else None,
        json.dumps(details or {}, ensure_ascii=False)
    ))


def _lms_verify_url(verify_token):
    base = (request.url_root or '').rstrip('/')
    return f"{base}/api/lms/certificates/verify/{verify_token}"


def _lms_escape_pdf_text(text):
    value = str(text or '')
    value = value.replace('\\', '\\\\')
    value = value.replace('(', '\\(')
    value = value.replace(')', '\\)')
    return value


def _lms_build_simple_pdf(lines):
    stream_lines = []
    y = 780
    for idx, line in enumerate(lines):
        font_size = 20 if idx == 0 else 12
        stream_lines.append(f"BT /F1 {font_size} Tf 50 {y} Td ({_lms_escape_pdf_text(line)}) Tj ET")
        y -= 28 if idx == 0 else 22
    stream = '\n'.join(stream_lines).encode('latin-1', errors='replace')

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Count 1 /Kids [3 0 R] >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>",
        b"<< /Length " + str(len(stream)).encode('ascii') + b" >>\nstream\n" + stream + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"
    ]

    data = bytearray()
    data.extend(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(data))
        data.extend(f"{index} 0 obj\n".encode('ascii'))
        data.extend(obj)
        data.extend(b"\nendobj\n")

    xref_offset = len(data)
    total_objects = len(objects) + 1
    data.extend(f"xref\n0 {total_objects}\n".encode('ascii'))
    data.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        data.extend(f"{offset:010d} 00000 n \n".encode('ascii'))
    data.extend(
        f"trailer\n<< /Size {total_objects} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode('ascii')
    )
    return bytes(data)


def _lms_generate_certificate_pdf(certificate_number, learner_name, course_title, issued_at, score_percent, verify_token):
    issue_dt = issued_at if isinstance(issued_at, datetime) else _lms_now()
    verify_url = _lms_verify_url(verify_token)
    lines = [
        "OTP LMS CERTIFICATE",
        f"Certificate: {certificate_number}",
        f"Learner: {learner_name or '-'}",
        f"Course: {course_title or '-'}",
        f"Issued at: {issue_dt.strftime('%Y-%m-%d %H:%M:%S')}",
        f"Score: {float(score_percent or 0):.2f}%",
        f"QR verify token: {verify_token}",
        f"Verify URL: {verify_url}",
    ]
    return _lms_build_simple_pdf(lines)


def _lms_resolve_request(required_scope='learner'):
    requester_id, requester, error = _resolve_requester()
    if error:
        message, status_code = error
        return None, None, None, jsonify({"error": message}), status_code

    requester_role = _normalize_user_role(requester[3])
    if not _lms_is_allowed_account(requester_id, requester_role):
        return None, None, None, jsonify({"error": "LMS section is restricted for this account"}), 403

    if required_scope == 'learner':
        if requester_role not in LMS_LEARNER_ROLES:
            return None, None, None, jsonify({"error": "Learner role required"}), 403
    elif required_scope == 'manager':
        if requester_role not in LMS_MANAGER_ROLES:
            return None, None, None, jsonify({"error": "Manager role required"}), 403
    elif required_scope == 'full_admin':
        if requester_role not in LMS_FULL_ADMIN_ROLES:
            return None, None, None, jsonify({"error": "Admin role required"}), 403
    elif required_scope == 'manager_or_learner':
        if requester_role not in LMS_MANAGER_ROLES and requester_role not in LMS_LEARNER_ROLES:
            return None, None, None, jsonify({"error": "LMS role required"}), 403

    return requester_id, requester, requester_role, None, None


def _lms_visible_learner_ids(requester_id, requester_role):
    try:
        return db.get_visible_lms_learner_ids_for_requester(requester_id, requester_role)
    except Exception:
        return []


def _lms_material_row_to_payload(row):
    bucket = row[5]
    blob_path = row[6]
    signed_url = _lms_signed_url(bucket, blob_path, expires_minutes=240)
    return {
        "id": int(row[0]),
        "lesson_id": int(row[1]),
        "title": row[2],
        "material_type": row[3],
        "content_text": row[4],
        "url": signed_url or row[7],
        "content_url": row[7],
        "signed_url": signed_url,
        "mime_type": row[8],
        "metadata": _lms_parse_json(row[9], {}),
        "position": int(row[10] or 1)
    }


def _lms_ensure_progress_and_session_tx(cursor, assignment_id, lesson_id, user_id):
    now = _lms_now()
    cursor.execute("""
        SELECT
            id, status, max_position_seconds, confirmed_seconds, completion_ratio,
            active_seconds, last_heartbeat_at, tab_hidden_count, stale_gap_count, started_at, completed_at
        FROM lms_lesson_progress
        WHERE assignment_id = %s AND lesson_id = %s AND user_id = %s
        LIMIT 1
    """, (assignment_id, lesson_id, user_id))
    progress = cursor.fetchone()
    if not progress:
        cursor.execute("""
            INSERT INTO lms_lesson_progress (
                assignment_id, lesson_id, user_id, status, started_at, updated_at
            )
            VALUES (%s, %s, %s, 'in_progress', %s, %s)
            RETURNING
                id, status, max_position_seconds, confirmed_seconds, completion_ratio,
                active_seconds, last_heartbeat_at, tab_hidden_count, stale_gap_count, started_at, completed_at
        """, (assignment_id, lesson_id, user_id, now, now))
        progress = cursor.fetchone()
    elif progress[1] == 'not_started':
        cursor.execute("""
            UPDATE lms_lesson_progress
            SET status = 'in_progress',
                started_at = COALESCE(started_at, %s),
                updated_at = %s
            WHERE id = %s
            RETURNING
                id, status, max_position_seconds, confirmed_seconds, completion_ratio,
                active_seconds, last_heartbeat_at, tab_hidden_count, stale_gap_count, started_at, completed_at
        """, (now, now, progress[0]))
        progress = cursor.fetchone()

    cursor.execute("""
        SELECT
            id, started_at, ended_at, last_heartbeat_at, is_active, max_position_seconds,
            confirmed_seconds, active_seconds, tab_hidden_count, stale_gap_count
        FROM lms_learning_sessions
        WHERE assignment_id = %s AND lesson_id = %s AND user_id = %s AND is_active = TRUE
        ORDER BY id DESC
        LIMIT 1
    """, (assignment_id, lesson_id, user_id))
    session = cursor.fetchone()
    if not session:
        cursor.execute("""
            INSERT INTO lms_learning_sessions (
                assignment_id, lesson_id, user_id, started_at, last_heartbeat_at, last_visible_at, is_active
            )
            VALUES (%s, %s, %s, %s, %s, %s, TRUE)
            RETURNING
                id, started_at, ended_at, last_heartbeat_at, is_active, max_position_seconds,
                confirmed_seconds, active_seconds, tab_hidden_count, stale_gap_count
        """, (assignment_id, lesson_id, user_id, now, now, now))
        session = cursor.fetchone()

    return progress, session


def _lms_get_lesson_context_tx(cursor, user_id, lesson_id):
    cursor.execute("""
        SELECT
            a.id as assignment_id,
            a.course_id,
            a.course_version_id,
            a.status,
            a.due_at,
            l.id as lesson_id,
            l.title,
            l.description,
            l.duration_seconds,
            COALESCE(l.allow_fast_forward, FALSE),
            COALESCE(l.completion_threshold, %s),
            m.title as module_title
        FROM lms_lessons l
        JOIN lms_modules m ON m.id = l.module_id
        JOIN lms_course_versions cv ON cv.id = m.course_version_id
        JOIN lms_course_assignments a
            ON a.course_version_id = cv.id
           AND a.user_id = %s
        WHERE l.id = %s
        LIMIT 1
    """, (LMS_COMPLETION_THRESHOLD, user_id, lesson_id))
    row = cursor.fetchone()
    if not row:
        return None

    assignment_id = int(row[0])
    lesson_id = int(row[5])
    progress, session = _lms_ensure_progress_and_session_tx(cursor, assignment_id, lesson_id, int(user_id))
    return {
        "assignment_id": assignment_id,
        "course_id": int(row[1]),
        "course_version_id": int(row[2]),
        "assignment_status": row[3],
        "due_at": row[4],
        "lesson_id": lesson_id,
        "lesson_title": row[6],
        "lesson_description": row[7],
        "duration_seconds": int(row[8] or 0),
        "allow_fast_forward": bool(row[9]),
        "completion_threshold": float(row[10] or LMS_COMPLETION_THRESHOLD),
        "module_title": row[11],
        "progress": progress,
        "session": session
    }


def _lms_try_issue_certificate_tx(cursor, assignment_id, user_id, score_percent=None, test_attempt_id=None):
    cursor.execute("""
        SELECT id, certificate_number, verify_token, status, issued_at
        FROM lms_certificates
        WHERE assignment_id = %s
        ORDER BY id DESC
        LIMIT 1
    """, (assignment_id,))
    existing = cursor.fetchone()
    if existing and existing[3] == 'active':
        return {
            "id": int(existing[0]),
            "certificate_number": existing[1],
            "verify_token": existing[2],
            "status": existing[3],
            "issued_at": existing[4]
        }

    cursor.execute("""
        SELECT a.course_id, c.title, u.name
        FROM lms_course_assignments a
        JOIN lms_courses c ON c.id = a.course_id
        JOIN users u ON u.id = a.user_id
        WHERE a.id = %s AND a.user_id = %s
        LIMIT 1
    """, (assignment_id, user_id))
    base = cursor.fetchone()
    if not base:
        return None

    course_id = int(base[0])
    course_title = base[1]
    learner_name = base[2]
    issued_at = _lms_now()

    certificate_number = None
    for _ in range(5):
        candidate = f"OTP-LMS-{issued_at.strftime('%Y%m%d')}-{secrets.token_hex(4).upper()}"
        cursor.execute("SELECT id FROM lms_certificates WHERE certificate_number = %s LIMIT 1", (candidate,))
        if not cursor.fetchone():
            certificate_number = candidate
            break
    if not certificate_number:
        certificate_number = f"OTP-LMS-{issued_at.strftime('%Y%m%d%H%M%S')}-{secrets.token_hex(3).upper()}"

    verify_token = secrets.token_urlsafe(24).replace('-', '').replace('_', '')
    pdf_bytes = _lms_generate_certificate_pdf(
        certificate_number=certificate_number,
        learner_name=learner_name,
        course_title=course_title,
        issued_at=issued_at,
        score_percent=score_percent,
        verify_token=verify_token
    )

    storage_type = 'db'
    pdf_data = pdf_bytes
    gcs_bucket = None
    gcs_blob_path = None
    if LMS_CERTIFICATE_STORAGE == 'gcs':
        bucket_name = _lms_bucket_name()
        if bucket_name:
            try:
                client = get_gcs_client()
                bucket = client.bucket(bucket_name)
                blob_path = (
                    f"lms/certificates/{issued_at.strftime('%Y/%m/%d')}/"
                    f"{certificate_number}.pdf"
                )
                blob = bucket.blob(blob_path)
                blob.upload_from_string(pdf_bytes, content_type='application/pdf')
                storage_type = 'gcs'
                pdf_data = None
                gcs_bucket = bucket_name
                gcs_blob_path = blob_path
            except Exception:
                storage_type = 'db'
                pdf_data = pdf_bytes
                gcs_bucket = None
                gcs_blob_path = None

    cursor.execute("""
        INSERT INTO lms_certificates (
            assignment_id, course_id, user_id, test_attempt_id,
            certificate_number, verify_token, score_percent, status, issued_at,
            pdf_storage_type, pdf_data, gcs_bucket, gcs_blob_path, metadata
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, 'active', %s, %s, %s, %s, %s, %s::jsonb)
        RETURNING id
    """, (
        assignment_id,
        course_id,
        user_id,
        (int(test_attempt_id) if test_attempt_id is not None else None),
        certificate_number,
        verify_token,
        (float(score_percent) if score_percent is not None else None),
        issued_at,
        storage_type,
        pdf_data,
        gcs_bucket,
        gcs_blob_path,
        json.dumps({
            "verify_url": _lms_verify_url(verify_token)
        }, ensure_ascii=False)
    ))
    cert_id = int(cursor.fetchone()[0])

    _lms_emit_notification(
        cursor,
        user_id=user_id,
        notification_type='certificate',
        title='Сертификат сформирован',
        message=f"Курс «{course_title}» завершен. Сертификат #{certificate_number} доступен для скачивания.",
        payload={"certificate_id": cert_id, "verify_token": verify_token}
    )
    return {
        "id": cert_id,
        "certificate_number": certificate_number,
        "verify_token": verify_token,
        "status": "active",
        "issued_at": issued_at
    }


def _lms_try_complete_assignment_tx(cursor, assignment_id, user_id, score_percent=None, test_attempt_id=None):
    cursor.execute("""
        SELECT id, course_id, course_version_id, due_at, status, completed_at
        FROM lms_course_assignments
        WHERE id = %s AND user_id = %s
        LIMIT 1
    """, (assignment_id, user_id))
    assignment = cursor.fetchone()
    if not assignment:
        return {"completed": False, "reason": "ASSIGNMENT_NOT_FOUND"}

    course_version_id = assignment[2]
    due_at = assignment[3]
    existing_completed_at = assignment[5]

    cursor.execute("""
        SELECT COUNT(*)
        FROM lms_lessons l
        JOIN lms_modules m ON m.id = l.module_id
        WHERE m.course_version_id = %s
    """, (course_version_id,))
    total_lessons = int(cursor.fetchone()[0] or 0)

    cursor.execute("""
        SELECT COUNT(*)
        FROM lms_lesson_progress
        WHERE assignment_id = %s
          AND user_id = %s
          AND status = 'completed'
    """, (assignment_id, user_id))
    completed_lessons = int(cursor.fetchone()[0] or 0)

    if total_lessons > 0 and completed_lessons < total_lessons:
        return {
            "completed": False,
            "reason": "LESSONS_INCOMPLETE",
            "total_lessons": total_lessons,
            "completed_lessons": completed_lessons
        }

    cursor.execute("""
        SELECT
            COUNT(*) AS total_tests,
            COUNT(*) FILTER (WHERE COALESCE(is_final, FALSE) = TRUE) AS total_final_tests
        FROM lms_tests
        WHERE course_version_id = %s
          AND status <> 'archived'
    """, (course_version_id,))
    tests_count_row = cursor.fetchone() or (0, 0)
    total_tests = int(tests_count_row[0] or 0)
    total_final_tests = int(tests_count_row[1] or 0)
    require_final_only = total_final_tests > 0
    required_tests = total_final_tests if require_final_only else total_tests

    cursor.execute("""
        SELECT COUNT(DISTINCT ta.test_id)
        FROM lms_test_attempts ta
        JOIN lms_tests t ON t.id = ta.test_id
        WHERE ta.assignment_id = %s
          AND ta.user_id = %s
          AND ta.passed = TRUE
          AND t.course_version_id = %s
          AND (%s::boolean = FALSE OR COALESCE(t.is_final, FALSE) = TRUE)
    """, (assignment_id, user_id, course_version_id, require_final_only))
    passed_tests = int(cursor.fetchone()[0] or 0)

    if required_tests > 0 and passed_tests < required_tests:
        return {
            "completed": False,
            "reason": "TESTS_INCOMPLETE",
            "total_tests": required_tests,
            "passed_tests": passed_tests
        }

    completed_at = existing_completed_at or _lms_now()
    deadline_status = _lms_deadline_status(due_at, completed_at)
    cursor.execute("""
        UPDATE lms_course_assignments
        SET status = 'completed',
            completed_at = %s,
            completion_color_status = %s,
            updated_at = %s
        WHERE id = %s
    """, (completed_at, deadline_status, _lms_now(), assignment_id))

    certificate = _lms_try_issue_certificate_tx(
        cursor,
        assignment_id=assignment_id,
        user_id=user_id,
        score_percent=score_percent,
        test_attempt_id=test_attempt_id
    )

    return {
        "completed": True,
        "completed_at": completed_at,
        "deadline_status": deadline_status,
        "certificate": certificate
    }


def _lms_fetch_test_questions_tx(cursor, test_id, include_correct=False):
    cursor.execute("""
        SELECT id, question_type, prompt, points, position, required, metadata, correct_text_answers
        FROM lms_questions
        WHERE test_id = %s
        ORDER BY position ASC, id ASC
    """, (test_id,))
    question_rows = cursor.fetchall()
    questions = []
    for q in question_rows:
        q_id = int(q[0])
        q_type = str(q[1] or '').strip().lower()
        cursor.execute("""
            SELECT id, option_key, option_text, position, is_correct, match_key, metadata
            FROM lms_question_options
            WHERE question_id = %s
            ORDER BY position ASC, id ASC
        """, (q_id,))
        option_rows = cursor.fetchall()
        options = []
        for opt in option_rows:
            item = {
                "id": int(opt[0]),
                "key": opt[1],
                "text": opt[2],
                "position": int(opt[3] or 1),
                "metadata": _lms_parse_json(opt[6], {})
            }
            if include_correct:
                item["is_correct"] = bool(opt[4])
                item["match_key"] = opt[5]
            options.append(item)

        question_payload = {
            "id": q_id,
            "type": q_type,
            "prompt": q[2],
            "points": float(q[3] or 1.0),
            "position": int(q[4] or 1),
            "required": bool(q[5]),
            "metadata": _lms_parse_json(q[6], {}),
            "options": options
        }
        if include_correct:
            question_payload["correct_text_answers"] = _lms_parse_json(q[7], [])
        questions.append(question_payload)
    return questions


def _lms_extract_selected_option_ids(answer_payload):
    selected = []
    payload = answer_payload
    if isinstance(payload, dict):
        if isinstance(payload.get('option_ids'), list):
            selected.extend(payload.get('option_ids') or [])
        elif payload.get('option_id') is not None:
            selected.append(payload.get('option_id'))
        elif isinstance(payload.get('value'), list):
            selected.extend(payload.get('value') or [])
        elif payload.get('value') is not None:
            selected.append(payload.get('value'))
    elif isinstance(payload, list):
        selected.extend(payload)
    elif payload is not None:
        selected.append(payload)

    out = []
    for value in selected:
        try:
            parsed = int(value)
        except Exception:
            continue
        if parsed not in out:
            out.append(parsed)
    return out


def _lms_answer_to_matching_map(answer_payload):
    if isinstance(answer_payload, dict):
        pairs = answer_payload.get('pairs')
        if isinstance(pairs, dict):
            return {str(k): str(v) for k, v in pairs.items()}
        if isinstance(pairs, list):
            mapped = {}
            for item in pairs:
                if not isinstance(item, dict):
                    continue
                left = item.get('left')
                right = item.get('right')
                if left in (None, ''):
                    continue
                mapped[str(left)] = str(right) if right is not None else ''
            return mapped
    if isinstance(answer_payload, list):
        mapped = {}
        for item in answer_payload:
            if not isinstance(item, dict):
                continue
            left = item.get('left')
            right = item.get('right')
            if left in (None, ''):
                continue
            mapped[str(left)] = str(right) if right is not None else ''
        return mapped
    return {}


def _lms_finalize_attempt_tx(cursor, attempt_id, user_id):
    cursor.execute("""
        SELECT
            ta.id, ta.assignment_id, ta.test_id, ta.user_id, ta.status, ta.started_at,
            t.title, t.pass_threshold, t.attempt_limit
        FROM lms_test_attempts ta
        JOIN lms_tests t ON t.id = ta.test_id
        WHERE ta.id = %s AND ta.user_id = %s
        LIMIT 1
    """, (attempt_id, user_id))
    attempt = cursor.fetchone()
    if not attempt:
        return None

    if attempt[4] != 'in_progress':
        cursor.execute("""
            SELECT score_percent, passed, finished_at, status
            FROM lms_test_attempts
            WHERE id = %s
        """, (attempt_id,))
        ready = cursor.fetchone()
        return {
            "attempt_id": int(attempt_id),
            "assignment_id": int(attempt[1]),
            "test_id": int(attempt[2]),
            "status": ready[3] if ready else attempt[4],
            "score_percent": float(ready[0] or 0.0) if ready else 0.0,
            "passed": bool(ready[1]) if ready and ready[1] is not None else False,
            "finished_at": ready[2] if ready else None,
            "already_finished": True
        }

    questions = _lms_fetch_test_questions_tx(cursor, int(attempt[2]), include_correct=True)
    cursor.execute("""
        SELECT question_id, answer_payload
        FROM lms_test_attempt_answers
        WHERE attempt_id = %s
    """, (attempt_id,))
    raw_answers = cursor.fetchall()
    answer_by_q = {}
    for row in raw_answers:
        answer_by_q[int(row[0])] = _lms_parse_json(row[1], {})

    total_points = 0.0
    scored_points = 0.0
    breakdown = []

    for question in questions:
        q_id = int(question['id'])
        q_type = str(question.get('type') or '').strip().lower()
        q_points = float(question.get('points') or 1.0)
        total_points += q_points
        payload = answer_by_q.get(q_id, {})
        is_correct = False

        if q_type in ('single', 'multiple'):
            expected = [int(opt['id']) for opt in question.get('options', []) if opt.get('is_correct')]
            expected_set = set(expected)
            selected_set = set(_lms_extract_selected_option_ids(payload))
            is_correct = selected_set == expected_set and (q_type != 'single' or len(selected_set) == 1)
        elif q_type == 'true_false':
            expected_options = [opt for opt in question.get('options', []) if opt.get('is_correct')]
            expected_bool = None
            if expected_options:
                text = _lms_normalize_text(expected_options[0].get('text'))
                expected_bool = text in ('true', 'верно', 'истина', 'да')
            if expected_bool is None:
                expected_bool = _lms_parse_bool((question.get('metadata') or {}).get('correct'), False)

            selected_bool = None
            if isinstance(payload, dict):
                if 'value' in payload:
                    selected_bool = _lms_parse_bool(payload.get('value'), False)
                elif 'answer' in payload:
                    selected_bool = _lms_parse_bool(payload.get('answer'), False)
                elif payload.get('option_id') is not None:
                    selected = _lms_extract_selected_option_ids(payload)
                    if selected:
                        option_map = {int(opt['id']): opt for opt in question.get('options', [])}
                        selected_option = option_map.get(int(selected[0]))
                        if selected_option:
                            selected_bool = _lms_normalize_text(selected_option.get('text')) in ('true', 'верно', 'истина', 'да')
            elif isinstance(payload, bool):
                selected_bool = payload

            if selected_bool is not None:
                is_correct = (selected_bool == expected_bool)
        elif q_type == 'matching':
            expected_map = {}
            for opt in question.get('options', []):
                key = opt.get('key')
                match_key = opt.get('match_key')
                if key not in (None, '') and match_key not in (None, ''):
                    expected_map[str(key)] = str(match_key)
            if not expected_map:
                for pair in (question.get('metadata') or {}).get('pairs', []):
                    if not isinstance(pair, dict):
                        continue
                    left = pair.get('left')
                    right = pair.get('right')
                    if left in (None, '') or right in (None, ''):
                        continue
                    expected_map[str(left)] = str(right)
            answer_map = _lms_answer_to_matching_map(payload)
            is_correct = bool(expected_map) and answer_map == expected_map
        elif q_type == 'text':
            accepted = question.get('correct_text_answers') or []
            if not accepted:
                accepted = (question.get('metadata') or {}).get('accepted_answers') or []
            normalized_accepted = {_lms_normalize_answer_text(item) for item in accepted if str(item or '').strip()}
            submitted = ''
            if isinstance(payload, dict):
                submitted = payload.get('text') or payload.get('value') or payload.get('answer') or ''
            else:
                submitted = payload
            submitted_norm = _lms_normalize_answer_text(submitted)
            is_correct = bool(submitted_norm) and submitted_norm in normalized_accepted

        points_awarded = q_points if is_correct else 0.0
        scored_points += points_awarded
        cursor.execute("""
            INSERT INTO lms_test_attempt_answers (
                attempt_id, question_id, answer_payload, is_correct, points_awarded, answered_at
            )
            VALUES (%s, %s, %s::jsonb, %s, %s, %s)
            ON CONFLICT (attempt_id, question_id)
            DO UPDATE SET
                answer_payload = EXCLUDED.answer_payload,
                is_correct = EXCLUDED.is_correct,
                points_awarded = EXCLUDED.points_awarded,
                answered_at = EXCLUDED.answered_at
        """, (
            attempt_id,
            q_id,
            json.dumps(payload, ensure_ascii=False),
            is_correct,
            points_awarded,
            _lms_now()
        ))

        breakdown.append({
            "question_id": q_id,
            "type": q_type,
            "is_correct": is_correct,
            "points_awarded": points_awarded,
            "points_total": q_points
        })

    score_percent = 0.0
    if total_points > 0:
        score_percent = round((scored_points / total_points) * 100.0, 2)

    pass_threshold = float(attempt[7] if attempt[7] is not None else LMS_DEFAULT_PASS_THRESHOLD)
    passed = score_percent >= pass_threshold
    finished_at = _lms_now()
    started_at = attempt[5]
    duration_seconds = None
    if isinstance(started_at, datetime):
        duration_seconds = max(0, int((finished_at - started_at).total_seconds()))

    cursor.execute("""
        UPDATE lms_test_attempts
        SET status = 'finished',
            score_percent = %s,
            passed = %s,
            finished_at = %s,
            duration_seconds = %s
        WHERE id = %s
    """, (score_percent, passed, finished_at, duration_seconds, attempt_id))

    completion = _lms_try_complete_assignment_tx(
        cursor,
        assignment_id=int(attempt[1]),
        user_id=int(user_id),
        score_percent=score_percent if passed else None,
        test_attempt_id=int(attempt_id) if passed else None
    )

    title = 'Тест пройден' if passed else 'Тест не пройден'
    text = (
        f"Результат: {score_percent:.2f}% (порог {pass_threshold:.2f}%)."
        + (" Поздравляем!" if passed else " Можно попробовать снова, если есть попытки.")
    )
    _lms_emit_notification(
        cursor,
        user_id=user_id,
        notification_type='test_result',
        title=title,
        message=text,
        payload={
            "attempt_id": int(attempt_id),
            "test_id": int(attempt[2]),
            "score_percent": score_percent,
            "passed": passed
        }
    )

    return {
        "attempt_id": int(attempt_id),
        "assignment_id": int(attempt[1]),
        "test_id": int(attempt[2]),
        "status": "finished",
        "score_percent": score_percent,
        "pass_threshold": pass_threshold,
        "passed": passed,
        "finished_at": finished_at,
        "duration_seconds": duration_seconds,
        "total_points": round(total_points, 2),
        "scored_points": round(scored_points, 2),
        "breakdown": breakdown,
        "assignment_completion": completion
    }


def _lms_course_structure_tx(cursor, course_id, course_version_id):
    cursor.execute("""
        SELECT
            c.id, c.title, c.description, c.category, c.status,
            c.default_pass_threshold, c.default_attempt_limit,
            cv.id, cv.version_number, cv.status, cv.pass_threshold, cv.attempt_limit,
            cv.anti_cheat_settings
        FROM lms_courses c
        JOIN lms_course_versions cv ON cv.id = %s AND cv.course_id = c.id
        WHERE c.id = %s
        LIMIT 1
    """, (course_version_id, course_id))
    header = cursor.fetchone()
    if not header:
        return None

    cursor.execute("""
        SELECT id, title, description, position
        FROM lms_modules
        WHERE course_version_id = %s
        ORDER BY position ASC, id ASC
    """, (course_version_id,))
    modules_rows = cursor.fetchall()

    cursor.execute("""
        SELECT
            l.id, l.module_id, l.title, l.description, l.position,
            l.duration_seconds, COALESCE(l.allow_fast_forward, FALSE), COALESCE(l.completion_threshold, %s)
        FROM lms_lessons l
        JOIN lms_modules m ON m.id = l.module_id
        WHERE m.course_version_id = %s
        ORDER BY m.position ASC, l.position ASC, l.id ASC
    """, (LMS_COMPLETION_THRESHOLD, course_version_id))
    lessons_rows = cursor.fetchall()

    cursor.execute("""
        SELECT
            id, lesson_id, title, material_type, content_text, gcs_bucket, gcs_blob_path,
            content_url, mime_type, metadata, position
        FROM lms_lesson_materials
        WHERE lesson_id IN (
            SELECT l.id
            FROM lms_lessons l
            JOIN lms_modules m ON m.id = l.module_id
            WHERE m.course_version_id = %s
        )
        ORDER BY position ASC, id ASC
    """, (course_version_id,))
    material_rows = cursor.fetchall()

    cursor.execute("""
        SELECT id, module_id, title, description, pass_threshold, attempt_limit, is_final, status
        FROM lms_tests
        WHERE course_version_id = %s
          AND status <> 'archived'
        ORDER BY id ASC
    """, (course_version_id,))
    tests_rows = cursor.fetchall()

    material_by_lesson = {}
    for row in material_rows:
        material_by_lesson.setdefault(int(row[1]), []).append(_lms_material_row_to_payload(row))

    lessons_by_module = {}
    for row in lessons_rows:
        lesson_id = int(row[0])
        module_id = int(row[1])
        lessons_by_module.setdefault(module_id, []).append({
            "id": lesson_id,
            "module_id": module_id,
            "title": row[2],
            "description": row[3],
            "position": int(row[4] or 1),
            "duration_seconds": int(row[5] or 0),
            "allow_fast_forward": bool(row[6]),
            "completion_threshold": float(row[7] or LMS_COMPLETION_THRESHOLD),
            "materials": material_by_lesson.get(lesson_id, [])
        })

    modules = []
    for row in modules_rows:
        module_id = int(row[0])
        modules.append({
            "id": module_id,
            "title": row[1],
            "description": row[2],
            "position": int(row[3] or 1),
            "lessons": lessons_by_module.get(module_id, [])
        })

    tests = []
    for row in tests_rows:
        test_id = int(row[0])
        cursor.execute("SELECT COUNT(*) FROM lms_questions WHERE test_id = %s", (test_id,))
        question_count = int(cursor.fetchone()[0] or 0)
        tests.append({
            "id": test_id,
            "module_id": int(row[1]) if row[1] is not None else None,
            "title": row[2],
            "description": row[3],
            "pass_threshold": float(row[4] if row[4] is not None else LMS_DEFAULT_PASS_THRESHOLD),
            "attempt_limit": int(row[5] if row[5] is not None else LMS_DEFAULT_ATTEMPT_LIMIT),
            "is_final": bool(row[6]),
            "status": row[7],
            "question_count": question_count
        })

    version_settings = _lms_parse_json(header[12], {})
    cover_payload = _lms_resolve_cover_payload(version_settings)
    return {
        "id": int(header[0]),
        "title": header[1],
        "description": header[2],
        "category": header[3],
        "status": header[4],
        "default_pass_threshold": float(header[5] if header[5] is not None else LMS_DEFAULT_PASS_THRESHOLD),
        "default_attempt_limit": int(header[6] if header[6] is not None else LMS_DEFAULT_ATTEMPT_LIMIT),
        "course_version": {
            "id": int(header[7]),
            "version_number": int(header[8] or 1),
            "status": header[9],
            "pass_threshold": float(header[10] if header[10] is not None else LMS_DEFAULT_PASS_THRESHOLD),
            "attempt_limit": int(header[11] if header[11] is not None else LMS_DEFAULT_ATTEMPT_LIMIT),
            "cover_url": cover_payload["cover_url"],
            "cover_bucket": cover_payload["cover_bucket"],
            "cover_blob_path": cover_payload["cover_blob_path"],
            "skills": _lms_normalize_skills(version_settings.get("skills"))
        },
        "modules": modules,
        "tests": tests
    }


@app.route('/api/lms/home', methods=['GET'])
@require_api_key
def lms_home():
    requester_id, requester, requester_role, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    a.id, a.course_id, a.course_version_id, a.status, a.due_at, a.started_at, a.completed_at,
                    c.title, c.description, c.category,
                    cv.anti_cheat_settings
                FROM lms_course_assignments a
                JOIN lms_courses c ON c.id = a.course_id
                LEFT JOIN lms_course_versions cv ON cv.id = a.course_version_id
                WHERE a.user_id = %s
                ORDER BY COALESCE(a.due_at, a.created_at) ASC, a.id DESC
            """, (requester_id,))
            assignment_rows = cursor.fetchall()

            courses = []
            for row in assignment_rows:
                assignment_id = int(row[0])
                course_version_id = int(row[2]) if row[2] is not None else None

                cursor.execute("""
                    SELECT COUNT(*)
                    FROM lms_lessons l
                    JOIN lms_modules m ON m.id = l.module_id
                    WHERE m.course_version_id = %s
                """, (course_version_id,))
                total_lessons = int(cursor.fetchone()[0] or 0)

                cursor.execute("""
                    SELECT COUNT(*)
                    FROM lms_lesson_progress
                    WHERE assignment_id = %s
                      AND user_id = %s
                      AND status = 'completed'
                """, (assignment_id, requester_id))
                completed_lessons = int(cursor.fetchone()[0] or 0)

                cursor.execute("""
                    SELECT COUNT(*)
                    FROM lms_tests
                    WHERE course_version_id = %s
                      AND status <> 'archived'
                      AND COALESCE(is_final, FALSE) = FALSE
                """, (course_version_id,))
                total_intermediate_tests = int(cursor.fetchone()[0] or 0)

                cursor.execute("""
                    SELECT COUNT(DISTINCT ta.test_id)
                    FROM lms_test_attempts ta
                    JOIN lms_tests t ON t.id = ta.test_id
                    WHERE ta.assignment_id = %s
                      AND ta.user_id = %s
                      AND ta.status = 'finished'
                      AND ta.passed = TRUE
                      AND t.course_version_id = %s
                      AND COALESCE(t.is_final, FALSE) = FALSE
                """, (assignment_id, requester_id, course_version_id))
                completed_intermediate_tests = int(cursor.fetchone()[0] or 0)

                cursor.execute("""
                    SELECT MAX(score_percent)
                    FROM lms_test_attempts
                    WHERE assignment_id = %s
                      AND user_id = %s
                      AND status = 'finished'
                """, (assignment_id, requester_id))
                best_score_raw = cursor.fetchone()[0]

                cursor.execute("""
                    SELECT id, status
                    FROM lms_certificates
                    WHERE assignment_id = %s
                    ORDER BY id DESC
                    LIMIT 1
                """, (assignment_id,))
                cert_row = cursor.fetchone()

                progress_total_items = total_lessons + total_intermediate_tests
                progress_completed_items = completed_lessons + completed_intermediate_tests

                progress_percent = 0.0
                if row[3] == 'completed':
                    progress_percent = 100.0
                elif progress_total_items > 0:
                    progress_percent = round((progress_completed_items / progress_total_items) * 100.0, 2)

                deadline_status = _lms_deadline_status(row[4], row[6])
                version_settings = _lms_parse_json(row[10], {})
                cover_payload = _lms_resolve_cover_payload(version_settings)
                courses.append({
                    "assignment_id": assignment_id,
                    "course_id": int(row[1]),
                    "course_version_id": course_version_id,
                    "title": row[7],
                    "description": row[8],
                    "category": row[9],
                    "cover_url": cover_payload["cover_url"],
                    "cover_bucket": cover_payload["cover_bucket"],
                    "cover_blob_path": cover_payload["cover_blob_path"],
                    "skills": _lms_normalize_skills(version_settings.get("skills")),
                    "status": row[3],
                    "due_at": row[4].isoformat() if row[4] else None,
                    "started_at": row[5].isoformat() if row[5] else None,
                    "completed_at": row[6].isoformat() if row[6] else None,
                    "deadline_status": deadline_status,
                    "progress_percent": progress_percent,
                    "completed_lessons": completed_lessons,
                    "total_lessons": total_lessons,
                    "completed_intermediate_tests": completed_intermediate_tests,
                    "total_intermediate_tests": total_intermediate_tests,
                    "best_score": float(best_score_raw) if best_score_raw is not None else None,
                    "certificate": (
                        {
                            "id": int(cert_row[0]),
                            "status": cert_row[1]
                        } if cert_row else None
                    )
                })

            cursor.execute("""
                SELECT COUNT(*)
                FROM lms_notifications
                WHERE user_id = %s
                  AND is_read = FALSE
            """, (requester_id,))
            unread_count = int(cursor.fetchone()[0] or 0)

        return jsonify({
            "status": "success",
            "role": requester_role,
            "heartbeat_seconds": LMS_HEARTBEAT_SECONDS,
            "stale_gap_seconds": LMS_STALE_GAP_SECONDS,
            "completion_threshold_percent": LMS_COMPLETION_THRESHOLD,
            "unread_notifications": unread_count,
            "courses": courses
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/home")
        return jsonify({"error": f"Failed to load LMS home: {str(e)}"}), 500


@app.route('/api/lms/courses', methods=['GET'])
@require_api_key
def lms_courses():
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    a.id, a.course_id, a.course_version_id, a.status, a.due_at, a.started_at, a.completed_at,
                    c.title, c.description, c.category
                FROM lms_course_assignments a
                JOIN lms_courses c ON c.id = a.course_id
                WHERE a.user_id = %s
                ORDER BY c.title ASC
            """, (requester_id,))
            rows = cursor.fetchall()

            items = []
            for row in rows:
                due_at = row[4]
                completed_at = row[6]
                items.append({
                    "assignment_id": int(row[0]),
                    "course_id": int(row[1]),
                    "course_version_id": int(row[2]) if row[2] is not None else None,
                    "status": row[3],
                    "title": row[7],
                    "description": row[8],
                    "category": row[9],
                    "due_at": due_at.isoformat() if due_at else None,
                    "started_at": row[5].isoformat() if row[5] else None,
                    "completed_at": completed_at.isoformat() if completed_at else None,
                    "deadline_status": _lms_deadline_status(due_at, completed_at)
                })

        return jsonify({"status": "success", "courses": items}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/courses")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/courses/<int:course_id>', methods=['GET'])
@require_api_key
def lms_course_detail(course_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, course_version_id, status, due_at, started_at, completed_at
                FROM lms_course_assignments
                WHERE user_id = %s
                  AND course_id = %s
                LIMIT 1
            """, (requester_id, course_id))
            assignment = cursor.fetchone()
            if not assignment:
                return jsonify({"error": "Course is not assigned to learner"}), 404

            assignment_id = int(assignment[0])
            course_version_id = int(assignment[1]) if assignment[1] is not None else None
            if course_version_id is None:
                cursor.execute("SELECT current_version_id FROM lms_courses WHERE id = %s", (course_id,))
                row = cursor.fetchone()
                if not row or row[0] is None:
                    return jsonify({"error": "Course version is not available"}), 404
                course_version_id = int(row[0])

            structure = _lms_course_structure_tx(cursor, course_id, course_version_id)
            if not structure:
                return jsonify({"error": "Course not found"}), 404

            cursor.execute("""
                SELECT
                    lesson_id, status, completion_ratio, max_position_seconds,
                    confirmed_seconds, active_seconds, completed_at
                FROM lms_lesson_progress
                WHERE assignment_id = %s
                  AND user_id = %s
            """, (assignment_id, requester_id))
            progress_rows = cursor.fetchall()
            lesson_progress = {}
            for row in progress_rows:
                lesson_progress[int(row[0])] = {
                    "status": row[1],
                    "completion_ratio": float(row[2] or 0.0),
                    "max_position_seconds": float(row[3] or 0.0),
                    "confirmed_seconds": float(row[4] or 0.0),
                    "active_seconds": int(row[5] or 0),
                    "completed_at": row[6].isoformat() if row[6] else None
                }

            cursor.execute("""
                SELECT
                    test_id,
                    MAX(score_percent) FILTER (WHERE status = 'finished') AS best_score,
                    BOOL_OR(COALESCE(passed, FALSE)) FILTER (WHERE status = 'finished') AS passed_any,
                    COUNT(*) AS attempts_used
                FROM lms_test_attempts
                WHERE assignment_id = %s
                  AND user_id = %s
                GROUP BY test_id
            """, (assignment_id, requester_id))
            attempts = {}
            for row in cursor.fetchall():
                attempts[int(row[0])] = {
                    "best_score": float(row[1]) if row[1] is not None else None,
                    "passed_any": bool(row[2]) if row[2] is not None else False,
                    "attempts_used": int(row[3] or 0)
                }

            structure["assignment"] = {
                "id": assignment_id,
                "status": assignment[2],
                "due_at": assignment[3].isoformat() if assignment[3] else None,
                "started_at": assignment[4].isoformat() if assignment[4] else None,
                "completed_at": assignment[5].isoformat() if assignment[5] else None,
                "deadline_status": _lms_deadline_status(assignment[3], assignment[5]),
                "lesson_progress": lesson_progress,
                "tests": attempts
            }

        return jsonify({"status": "success", "course": structure}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/courses/<course_id>")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/courses/<int:course_id>/start', methods=['POST'])
@require_api_key
def lms_start_course(course_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        now = _lms_now()
        with db._get_cursor() as cursor:
            cursor.execute("""
                UPDATE lms_course_assignments
                SET status = CASE WHEN status = 'assigned' THEN 'in_progress' ELSE status END,
                    started_at = COALESCE(started_at, %s),
                    updated_at = %s
                WHERE user_id = %s
                  AND course_id = %s
                RETURNING id, status, started_at, due_at
            """, (now, now, requester_id, course_id))
            updated = cursor.fetchone()
            if not updated:
                return jsonify({"error": "Course is not assigned to learner"}), 404

        return jsonify({
            "status": "success",
            "assignment_id": int(updated[0]),
            "assignment_status": updated[1],
            "started_at": updated[2].isoformat() if updated[2] else None,
            "due_at": updated[3].isoformat() if updated[3] else None
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/courses/<course_id>/start")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/lessons/<int:lesson_id>', methods=['GET'])
@require_api_key
def lms_lesson_detail(lesson_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            context = _lms_get_lesson_context_tx(cursor, requester_id, lesson_id)
            if not context:
                return jsonify({"error": "Lesson is not available for learner"}), 404

            cursor.execute("""
                UPDATE lms_course_assignments
                SET status = CASE WHEN status = 'assigned' THEN 'in_progress' ELSE status END,
                    started_at = COALESCE(started_at, %s),
                    updated_at = %s
                WHERE id = %s
            """, (_lms_now(), _lms_now(), context["assignment_id"]))

            cursor.execute("""
                SELECT
                    id, lesson_id, title, material_type, content_text, gcs_bucket, gcs_blob_path,
                    content_url, mime_type, metadata, position
                FROM lms_lesson_materials
                WHERE lesson_id = %s
                ORDER BY position ASC, id ASC
            """, (lesson_id,))
            materials = [_lms_material_row_to_payload(row) for row in cursor.fetchall()]

            progress = context["progress"]
            session = context["session"]
            completion_ratio = float(progress[4] or 0.0)
            required_ratio = max(float(context["completion_threshold"]), float(LMS_COMPLETION_THRESHOLD))

        return jsonify({
            "status": "success",
            "lesson": {
                "id": context["lesson_id"],
                "title": context["lesson_title"],
                "description": context["lesson_description"],
                "module_title": context["module_title"],
                "duration_seconds": context["duration_seconds"],
                "allow_fast_forward": context["allow_fast_forward"],
                "completion_threshold": required_ratio
            },
            "assignment_id": context["assignment_id"],
            "session": {
                "id": int(session[0]),
                "started_at": session[1].isoformat() if session[1] else None,
                "last_heartbeat_at": session[3].isoformat() if session[3] else None,
                "is_active": bool(session[4]),
                "max_position_seconds": float(session[5] or 0.0),
                "confirmed_seconds": float(session[6] or 0.0),
                "active_seconds": int(session[7] or 0),
                "tab_hidden_count": int(session[8] or 0),
                "stale_gap_count": int(session[9] or 0)
            },
            "progress": {
                "id": int(progress[0]),
                "status": progress[1],
                "max_position_seconds": float(progress[2] or 0.0),
                "confirmed_seconds": float(progress[3] or 0.0),
                "completion_ratio": completion_ratio,
                "active_seconds": int(progress[5] or 0),
                "last_heartbeat_at": progress[6].isoformat() if progress[6] else None,
                "tab_hidden_count": int(progress[7] or 0),
                "stale_gap_count": int(progress[8] or 0),
                "can_complete": completion_ratio >= required_ratio
            },
            "anti_cheat": {
                "heartbeat_seconds": LMS_HEARTBEAT_SECONDS,
                "stale_gap_seconds": LMS_STALE_GAP_SECONDS,
                "completion_threshold_percent": required_ratio
            },
            "materials": materials
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/lessons/<lesson_id>")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/lessons/<int:lesson_id>/event', methods=['POST'])
@require_api_key
def lms_lesson_event(lesson_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    data = request.get_json(silent=True) or {}
    event_type = str(data.get('event_type') or '').strip().lower() or 'unknown'
    payload = data.get('payload') if isinstance(data.get('payload'), dict) else {}
    client_ts = _lms_parse_datetime(data.get('client_ts'))

    try:
        with db._get_cursor() as cursor:
            context = _lms_get_lesson_context_tx(cursor, requester_id, lesson_id)
            if not context:
                return jsonify({"error": "Lesson is not available for learner"}), 404

            progress = context["progress"]
            session = context["session"]
            progress_id = int(progress[0])
            session_id = int(session[0])
            allow_fast_forward = bool(context["allow_fast_forward"])
            progress_max = float(progress[2] or 0.0)

            blocked_seek = False
            allowed_position = None
            if event_type == 'seek':
                to_seconds = _lms_to_float(
                    payload.get('to_seconds')
                    if isinstance(payload, dict) else None,
                    default=_lms_to_float(data.get('to_seconds'), default=0.0)
                )
                from_seconds = _lms_to_float(
                    payload.get('from_seconds')
                    if isinstance(payload, dict) else None,
                    default=_lms_to_float(data.get('from_seconds'), default=0.0)
                )
                payload = {
                    **(payload or {}),
                    "from_seconds": from_seconds,
                    "to_seconds": to_seconds
                }

                if not allow_fast_forward:
                    allowed_position = progress_max + max(3.0, float(LMS_HEARTBEAT_SECONDS) * 2.0)
                    if to_seconds > allowed_position:
                        blocked_seek = True

            if event_type == 'visibility':
                is_visible = _lms_parse_bool(payload.get('is_visible') if isinstance(payload, dict) else None, True)
                if not is_visible:
                    cursor.execute("""
                        UPDATE lms_lesson_progress
                        SET tab_hidden_count = COALESCE(tab_hidden_count, 0) + 1,
                            last_event_at = %s,
                            updated_at = %s
                        WHERE id = %s
                    """, (_lms_now(), _lms_now(), progress_id))
                    cursor.execute("""
                        UPDATE lms_learning_sessions
                        SET tab_hidden_count = COALESCE(tab_hidden_count, 0) + 1
                        WHERE id = %s
                    """, (session_id,))

            if blocked_seek:
                payload = {**(payload or {}), "blocked": True, "allowed_position": allowed_position}
                event_name = 'seek_blocked'
            else:
                event_name = event_type

            cursor.execute("""
                INSERT INTO lms_learning_events (
                    session_id, lesson_id, user_id, event_type, payload, client_ts, created_at
                )
                VALUES (%s, %s, %s, %s, %s::jsonb, %s, %s)
            """, (
                session_id,
                lesson_id,
                requester_id,
                event_name,
                json.dumps(payload or {}, ensure_ascii=False),
                client_ts,
                _lms_now()
            ))

            if blocked_seek:
                return jsonify({
                    "status": "blocked",
                    "reason": "FORWARD_SEEK_NOT_ALLOWED",
                    "allowed_position": allowed_position
                }), 409

        return jsonify({"status": "success"}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/lessons/<lesson_id>/event")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/lessons/<int:lesson_id>/heartbeat', methods=['POST'])
@require_api_key
def lms_lesson_heartbeat(lesson_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    data = request.get_json(silent=True) or {}
    raw_position = data.get('position_seconds')
    tab_visible = _lms_parse_bool(data.get('tab_visible'), True)

    try:
        with db._get_cursor() as cursor:
            context = _lms_get_lesson_context_tx(cursor, requester_id, lesson_id)
            if not context:
                return jsonify({"error": "Lesson is not available for learner"}), 404

            progress = context["progress"]
            session = context["session"]

            progress_id = int(progress[0])
            session_id = int(session[0])
            lesson_duration = float(context["duration_seconds"] or 0.0)
            allow_fast_forward = bool(context["allow_fast_forward"])

            now = _lms_now()
            requested_position = max(0.0, _lms_to_float(raw_position, default=0.0))
            progress_max = float(progress[2] or 0.0)
            progress_confirmed = float(progress[3] or 0.0)
            progress_active = int(progress[5] or 0)
            progress_tab_hidden = int(progress[7] or 0)
            progress_stale = int(progress[8] or 0)

            last_heartbeat = session[3] or progress[6]
            gap_seconds = None
            stale_gap = False
            if isinstance(last_heartbeat, datetime):
                gap_seconds = max(0.0, (now - last_heartbeat).total_seconds())
                stale_gap = gap_seconds > float(LMS_STALE_GAP_SECONDS)

            blocked_forward_seek = False
            allowed_position = None
            effective_position = requested_position
            if not allow_fast_forward:
                allowed_position = progress_max + max(3.0, float(LMS_HEARTBEAT_SECONDS) * 2.0)
                if effective_position > allowed_position:
                    effective_position = allowed_position
                    blocked_forward_seek = True

            next_max = max(progress_max, effective_position)
            next_confirmed = progress_confirmed
            active_increment = 0

            if tab_visible and not stale_gap:
                next_confirmed = max(next_confirmed, effective_position)
                if gap_seconds is not None:
                    active_increment = int(min(gap_seconds, float(LMS_STALE_GAP_SECONDS)))
                else:
                    active_increment = int(LMS_HEARTBEAT_SECONDS)

            next_active = progress_active + active_increment
            next_tab_hidden = progress_tab_hidden + (0 if tab_visible else 1)
            next_stale = progress_stale + (1 if stale_gap else 0)
            completion_ratio = 100.0 if lesson_duration <= 0 else min(100.0, round((next_confirmed / lesson_duration) * 100.0, 2))

            cursor.execute("""
                UPDATE lms_lesson_progress
                SET status = CASE WHEN status = 'not_started' THEN 'in_progress' ELSE status END,
                    max_position_seconds = %s,
                    confirmed_seconds = %s,
                    completion_ratio = %s,
                    active_seconds = %s,
                    last_heartbeat_at = %s,
                    last_event_at = %s,
                    tab_hidden_count = %s,
                    stale_gap_count = %s,
                    started_at = COALESCE(started_at, %s),
                    updated_at = %s
                WHERE id = %s
            """, (
                next_max,
                next_confirmed,
                completion_ratio,
                next_active,
                now,
                now,
                next_tab_hidden,
                next_stale,
                now,
                now,
                progress_id
            ))

            cursor.execute("""
                UPDATE lms_learning_sessions
                SET
                    last_heartbeat_at = %s,
                    last_visible_at = CASE WHEN %s THEN %s ELSE last_visible_at END,
                    max_position_seconds = %s,
                    confirmed_seconds = %s,
                    active_seconds = %s,
                    tab_hidden_count = %s,
                    stale_gap_count = %s
                WHERE id = %s
            """, (
                now,
                tab_visible,
                now,
                max(float(session[5] or 0.0), next_max),
                max(float(session[6] or 0.0), next_confirmed),
                int(session[7] or 0) + active_increment,
                int(session[8] or 0) + (0 if tab_visible else 1),
                int(session[9] or 0) + (1 if stale_gap else 0),
                session_id
            ))

            cursor.execute("""
                INSERT INTO lms_learning_events (
                    session_id, lesson_id, user_id, event_type, payload, client_ts, created_at
                )
                VALUES (%s, %s, %s, 'heartbeat', %s::jsonb, %s, %s)
            """, (
                session_id,
                lesson_id,
                requester_id,
                json.dumps({
                    "requested_position_seconds": requested_position,
                    "effective_position_seconds": effective_position,
                    "tab_visible": tab_visible,
                    "stale_gap": stale_gap,
                    "blocked_forward_seek": blocked_forward_seek
                }, ensure_ascii=False),
                _lms_parse_datetime(data.get('client_ts')),
                now
            ))

        return jsonify({
            "status": "success",
            "position_seconds": effective_position,
            "completion_ratio": completion_ratio,
            "active_seconds": next_active,
            "stale_gap": stale_gap,
            "blocked_forward_seek": blocked_forward_seek,
            "allowed_position": allowed_position if blocked_forward_seek else None
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/lessons/<lesson_id>/heartbeat")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/lessons/<int:lesson_id>/complete', methods=['POST'])
@require_api_key
def lms_lesson_complete(lesson_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            context = _lms_get_lesson_context_tx(cursor, requester_id, lesson_id)
            if not context:
                return jsonify({"error": "Lesson is not available for learner"}), 404

            progress = context["progress"]
            progress_id = int(progress[0])
            current_ratio = float(progress[4] or 0.0)
            last_heartbeat_at = progress[6]
            required_ratio = max(float(context["completion_threshold"]), float(LMS_COMPLETION_THRESHOLD))
            cursor.execute("""
                SELECT EXISTS(
                    SELECT 1
                    FROM lms_lesson_materials
                    WHERE lesson_id = %s
                      AND material_type = 'video'
                )
            """, (lesson_id,))
            has_video_material = bool((cursor.fetchone() or [False])[0])

            if has_video_material and current_ratio < required_ratio:
                return jsonify({
                    "error": "Lesson completion threshold is not reached",
                    "required_ratio": required_ratio,
                    "current_ratio": current_ratio
                }), 409

            if has_video_material and isinstance(last_heartbeat_at, datetime):
                gap = (_lms_now() - last_heartbeat_at).total_seconds()
                if gap > float(LMS_STALE_GAP_SECONDS):
                    return jsonify({
                        "error": "Heartbeat is stale, lesson completion denied",
                        "stale_gap_seconds": LMS_STALE_GAP_SECONDS,
                        "current_gap_seconds": round(gap, 2)
                    }), 409

            now = _lms_now()
            cursor.execute("""
                UPDATE lms_lesson_progress
                SET status = 'completed',
                    completion_ratio = GREATEST(completion_ratio, %s),
                    completed_at = COALESCE(completed_at, %s),
                    updated_at = %s
                WHERE id = %s
                RETURNING completed_at, completion_ratio
            """, (required_ratio, now, now, progress_id))
            updated = cursor.fetchone()

            completion = _lms_try_complete_assignment_tx(
                cursor,
                assignment_id=context["assignment_id"],
                user_id=requester_id
            )

            _lms_emit_notification(
                cursor,
                user_id=requester_id,
                notification_type='lesson_completed',
                title='Урок завершен',
                message=f"Урок «{context['lesson_title']}» отмечен как завершенный.",
                payload={"lesson_id": lesson_id, "assignment_id": context["assignment_id"]}
            )

        return jsonify({
            "status": "success",
            "lesson_id": lesson_id,
            "completed_at": updated[0].isoformat() if updated and updated[0] else None,
            "completion_ratio": float(updated[1] or required_ratio) if updated else required_ratio,
            "assignment_completion": completion
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/lessons/<lesson_id>/complete")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/tests/<int:test_id>/start', methods=['POST'])
@require_api_key
def lms_test_start(test_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    a.id as assignment_id,
                    a.course_id,
                    a.course_version_id,
                    a.status as assignment_status,
                    t.id as test_id,
                    t.title,
                    COALESCE(t.pass_threshold, cv.pass_threshold, c.default_pass_threshold, %s),
                    COALESCE(t.attempt_limit, cv.attempt_limit, c.default_attempt_limit, %s)
                FROM lms_tests t
                JOIN lms_course_versions cv ON cv.id = t.course_version_id
                JOIN lms_courses c ON c.id = cv.course_id
                JOIN lms_course_assignments a
                    ON a.course_version_id = cv.id
                   AND a.user_id = %s
                WHERE t.id = %s
                  AND t.status <> 'archived'
                LIMIT 1
            """, (
                LMS_DEFAULT_PASS_THRESHOLD,
                LMS_DEFAULT_ATTEMPT_LIMIT,
                requester_id,
                test_id
            ))
            row = cursor.fetchone()
            if not row:
                return jsonify({"error": "Test is not available for learner"}), 404

            assignment_id = int(row[0])
            course_version_id = int(row[2])
            pass_threshold = float(row[6] or LMS_DEFAULT_PASS_THRESHOLD)
            attempt_limit = int(row[7] or LMS_DEFAULT_ATTEMPT_LIMIT)

            cursor.execute("""
                SELECT COUNT(*)
                FROM lms_lessons l
                JOIN lms_modules m ON m.id = l.module_id
                WHERE m.course_version_id = %s
            """, (course_version_id,))
            total_lessons = int(cursor.fetchone()[0] or 0)

            cursor.execute("""
                SELECT COUNT(*)
                FROM lms_lesson_progress
                WHERE assignment_id = %s
                  AND user_id = %s
                  AND status = 'completed'
            """, (assignment_id, requester_id))
            completed_lessons = int(cursor.fetchone()[0] or 0)

            if total_lessons > 0 and completed_lessons < total_lessons:
                return jsonify({
                    "error": "Lessons must be completed before test start",
                    "completed_lessons": completed_lessons,
                    "total_lessons": total_lessons
                }), 409

            cursor.execute("""
                SELECT id, attempt_no
                FROM lms_test_attempts
                WHERE assignment_id = %s
                  AND test_id = %s
                  AND user_id = %s
                  AND status = 'in_progress'
                ORDER BY id DESC
                LIMIT 1
            """, (assignment_id, test_id, requester_id))
            active_attempt = cursor.fetchone()
            if active_attempt:
                attempt_id = int(active_attempt[0])
                attempt_no = int(active_attempt[1])
            else:
                cursor.execute("""
                    SELECT COUNT(*)
                    FROM lms_test_attempts
                    WHERE assignment_id = %s
                      AND test_id = %s
                      AND user_id = %s
                """, (assignment_id, test_id, requester_id))
                attempts_used = int(cursor.fetchone()[0] or 0)
                if attempts_used >= attempt_limit:
                    return jsonify({
                        "error": "Attempt limit reached",
                        "attempt_limit": attempt_limit,
                        "attempts_used": attempts_used
                    }), 409

                attempt_no = attempts_used + 1
                cursor.execute("""
                    INSERT INTO lms_test_attempts (
                        assignment_id, test_id, user_id, attempt_no, status, started_at
                    )
                    VALUES (%s, %s, %s, %s, 'in_progress', %s)
                    RETURNING id
                """, (assignment_id, test_id, requester_id, attempt_no, _lms_now()))
                attempt_id = int(cursor.fetchone()[0])

            questions = _lms_fetch_test_questions_tx(cursor, test_id, include_correct=False)

        return jsonify({
            "status": "success",
            "attempt": {
                "id": attempt_id,
                "attempt_no": attempt_no,
                "assignment_id": assignment_id,
                "test_id": int(test_id),
                "pass_threshold": pass_threshold,
                "attempt_limit": attempt_limit
            },
            "questions": questions
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/tests/<test_id>/start")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/tests/attempts/<int:attempt_id>/answer', methods=['PATCH'])
@require_api_key
def lms_test_answer(attempt_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    data = request.get_json(silent=True) or {}
    question_id = data.get('question_id')
    if question_id in (None, ''):
        return jsonify({"error": "question_id is required"}), 400
    try:
        question_id = int(question_id)
    except Exception:
        return jsonify({"error": "Invalid question_id"}), 400

    raw_answer = data.get('answer_payload', data.get('answer'))
    if raw_answer is None:
        raw_answer = {}

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, test_id, status
                FROM lms_test_attempts
                WHERE id = %s AND user_id = %s
                LIMIT 1
            """, (attempt_id, requester_id))
            attempt = cursor.fetchone()
            if not attempt:
                return jsonify({"error": "Attempt not found"}), 404
            if attempt[2] != 'in_progress':
                return jsonify({"error": "Attempt is not in progress"}), 409

            cursor.execute("""
                SELECT id
                FROM lms_questions
                WHERE id = %s AND test_id = %s
                LIMIT 1
            """, (question_id, int(attempt[1])))
            qrow = cursor.fetchone()
            if not qrow:
                return jsonify({"error": "Question does not belong to this attempt"}), 404

            cursor.execute("""
                INSERT INTO lms_test_attempt_answers (
                    attempt_id, question_id, answer_payload, answered_at
                )
                VALUES (%s, %s, %s::jsonb, %s)
                ON CONFLICT (attempt_id, question_id)
                DO UPDATE SET
                    answer_payload = EXCLUDED.answer_payload,
                    answered_at = EXCLUDED.answered_at
            """, (
                attempt_id,
                question_id,
                json.dumps(raw_answer, ensure_ascii=False),
                _lms_now()
            ))

        return jsonify({"status": "success", "attempt_id": attempt_id, "question_id": question_id}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/tests/attempts/<attempt_id>/answer")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/tests/attempts/<int:attempt_id>/finish', methods=['POST'])
@require_api_key
def lms_test_finish(attempt_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            result = _lms_finalize_attempt_tx(cursor, attempt_id, requester_id)
            if not result:
                return jsonify({"error": "Attempt not found"}), 404

        return jsonify({"status": "success", "result": result}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/tests/attempts/<attempt_id>/finish")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/tests/attempts/<int:attempt_id>/result', methods=['GET'])
@require_api_key
def lms_test_result(attempt_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    ta.id, ta.assignment_id, ta.test_id, ta.attempt_no, ta.status,
                    ta.score_percent, ta.passed, ta.started_at, ta.finished_at, ta.duration_seconds,
                    t.title, t.pass_threshold
                FROM lms_test_attempts ta
                JOIN lms_tests t ON t.id = ta.test_id
                WHERE ta.id = %s
                  AND ta.user_id = %s
                LIMIT 1
            """, (attempt_id, requester_id))
            attempt = cursor.fetchone()
            if not attempt:
                return jsonify({"error": "Attempt not found"}), 404

            cursor.execute("""
                SELECT
                    q.id, q.question_type, q.prompt, q.points,
                    a.answer_payload, a.is_correct, a.points_awarded, a.answered_at
                FROM lms_questions q
                LEFT JOIN lms_test_attempt_answers a
                  ON a.question_id = q.id
                 AND a.attempt_id = %s
                WHERE q.test_id = %s
                ORDER BY q.position ASC, q.id ASC
            """, (attempt_id, int(attempt[2])))
            answer_rows = cursor.fetchall()
            answers = []
            for row in answer_rows:
                answers.append({
                    "question_id": int(row[0]),
                    "type": row[1],
                    "prompt": row[2],
                    "points_total": float(row[3] or 1.0),
                    "answer_payload": _lms_parse_json(row[4], {}),
                    "is_correct": (bool(row[5]) if row[5] is not None else None),
                    "points_awarded": (float(row[6]) if row[6] is not None else None),
                    "answered_at": row[7].isoformat() if row[7] else None
                })

            cursor.execute("""
                SELECT id, certificate_number, verify_token, status
                FROM lms_certificates
                WHERE assignment_id = %s
                ORDER BY id DESC
                LIMIT 1
            """, (int(attempt[1]),))
            cert = cursor.fetchone()

        return jsonify({
            "status": "success",
            "attempt": {
                "id": int(attempt[0]),
                "assignment_id": int(attempt[1]),
                "test_id": int(attempt[2]),
                "attempt_no": int(attempt[3]),
                "status": attempt[4],
                "score_percent": float(attempt[5]) if attempt[5] is not None else None,
                "passed": bool(attempt[6]) if attempt[6] is not None else None,
                "started_at": attempt[7].isoformat() if attempt[7] else None,
                "finished_at": attempt[8].isoformat() if attempt[8] else None,
                "duration_seconds": int(attempt[9] or 0),
                "test_title": attempt[10],
                "pass_threshold": float(attempt[11] if attempt[11] is not None else LMS_DEFAULT_PASS_THRESHOLD)
            },
            "answers": answers,
            "certificate": (
                {
                    "id": int(cert[0]),
                    "certificate_number": cert[1],
                    "verify_token": cert[2],
                    "verify_url": _lms_verify_url(cert[2]),
                    "status": cert[3]
                } if cert else None
            )
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/tests/attempts/<attempt_id>/result")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/certificates', methods=['GET'])
@require_api_key
def lms_certificates():
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    c.id, c.assignment_id, c.course_id, c.certificate_number, c.verify_token,
                    c.score_percent, c.status, c.issued_at, c.revoked_at, c.revoke_reason
                FROM lms_certificates c
                WHERE c.user_id = %s
                ORDER BY c.issued_at DESC, c.id DESC
            """, (requester_id,))
            rows = cursor.fetchall()

            items = []
            for row in rows:
                items.append({
                    "id": int(row[0]),
                    "assignment_id": int(row[1]) if row[1] is not None else None,
                    "course_id": int(row[2]),
                    "certificate_number": row[3],
                    "verify_token": row[4],
                    "verify_url": _lms_verify_url(row[4]),
                    "score_percent": float(row[5]) if row[5] is not None else None,
                    "status": row[6],
                    "issued_at": row[7].isoformat() if row[7] else None,
                    "revoked_at": row[8].isoformat() if row[8] else None,
                    "revoke_reason": row[9]
                })

        return jsonify({"status": "success", "certificates": items}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/certificates")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/certificates/<int:certificate_id>/download', methods=['GET'])
@require_api_key
def lms_certificate_download(certificate_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    certificate_number, pdf_storage_type, pdf_data, gcs_bucket, gcs_blob_path, status
                FROM lms_certificates
                WHERE id = %s
                  AND user_id = %s
                LIMIT 1
            """, (certificate_id, requester_id))
            row = cursor.fetchone()
            if not row:
                return jsonify({"error": "Certificate not found"}), 404

            cert_number = row[0] or f"LMS-{certificate_id}"
            storage_type = row[1] or 'db'

            if storage_type == 'gcs' and row[3] and row[4]:
                signed_url = _lms_signed_url(row[3], row[4], expires_minutes=30)
                if signed_url:
                    return redirect(signed_url, code=302)

            pdf_data = row[2]
            if isinstance(pdf_data, memoryview):
                pdf_data = pdf_data.tobytes()
            elif isinstance(pdf_data, bytearray):
                pdf_data = bytes(pdf_data)

            if not pdf_data:
                return jsonify({"error": "Certificate file is unavailable"}), 404

            return send_file(
                BytesIO(pdf_data),
                as_attachment=True,
                download_name=f"{cert_number}.pdf",
                mimetype='application/pdf'
            )
    except Exception as e:
        logging.exception("Error in /api/lms/certificates/<certificate_id>/download")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/notifications', methods=['GET'])
@require_api_key
def lms_notifications():
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    limit = _lms_to_int(request.args.get('limit'), default=100)
    limit = min(max(limit, 1), 500)
    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, notification_type, title, message, payload, is_read, read_at, created_at
                FROM lms_notifications
                WHERE user_id = %s
                ORDER BY created_at DESC, id DESC
                LIMIT %s
            """, (requester_id, limit))
            rows = cursor.fetchall()

            items = []
            for row in rows:
                items.append({
                    "id": int(row[0]),
                    "type": row[1],
                    "title": row[2],
                    "message": row[3],
                    "payload": _lms_parse_json(row[4], {}),
                    "is_read": bool(row[5]),
                    "read_at": row[6].isoformat() if row[6] else None,
                    "created_at": row[7].isoformat() if row[7] else None
                })
        return jsonify({"status": "success", "notifications": items}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/notifications")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/notifications/<int:notification_id>/read', methods=['POST'])
@require_api_key
def lms_notification_read(notification_id):
    requester_id, _, _, error_response, status_code = _lms_resolve_request('learner')
    if error_response:
        return error_response, status_code

    try:
        with db._get_cursor() as cursor:
            now = _lms_now()
            cursor.execute("""
                UPDATE lms_notifications
                SET is_read = TRUE,
                    read_at = COALESCE(read_at, %s)
                WHERE id = %s
                  AND user_id = %s
                RETURNING id
            """, (now, notification_id, requester_id))
            updated = cursor.fetchone()
            if not updated:
                return jsonify({"error": "Notification not found"}), 404
        return jsonify({"status": "success", "notification_id": notification_id}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/notifications/<notification_id>/read")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/certificates/verify/<string:token>', methods=['GET'])
def lms_verify_certificate(token):
    verify_token = str(token or '').strip()
    if not verify_token:
        return jsonify({"error": "Invalid token"}), 400

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    c.id, c.certificate_number, c.verify_token, c.status, c.issued_at,
                    c.revoked_at, c.revoke_reason, c.score_percent,
                    u.id, u.name,
                    cr.id, cr.title
                FROM lms_certificates c
                JOIN users u ON u.id = c.user_id
                JOIN lms_courses cr ON cr.id = c.course_id
                WHERE c.verify_token = %s
                LIMIT 1
            """, (verify_token,))
            row = cursor.fetchone()
            if not row:
                return jsonify({"status": "not_found", "valid": False}), 404

            valid = (row[3] == 'active')
            return jsonify({
                "status": "success",
                "valid": valid,
                "certificate": {
                    "id": int(row[0]),
                    "certificate_number": row[1],
                    "verify_token": row[2],
                    "status": row[3],
                    "issued_at": row[4].isoformat() if row[4] else None,
                    "revoked_at": row[5].isoformat() if row[5] else None,
                    "revoke_reason": row[6],
                    "score_percent": float(row[7]) if row[7] is not None else None,
                    "learner": {
                        "id": int(row[8]),
                        "name": row[9]
                    },
                    "course": {
                        "id": int(row[10]),
                        "title": row[11]
                    }
                }
            }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/certificates/verify/<token>")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/admin/courses', methods=['GET', 'POST', 'PATCH'])
@require_api_key
def lms_admin_courses():
    requester_id, _, requester_role, error_response, status_code = _lms_resolve_request('manager')
    if error_response:
        return error_response, status_code

    try:
        if request.method == 'GET':
            course_id = _lms_to_int(request.args.get('course_id'), default=0)
            with db._get_cursor() as cursor:
                if course_id > 0:
                    cursor.execute("SELECT current_version_id FROM lms_courses WHERE id = %s LIMIT 1", (course_id,))
                    base = cursor.fetchone()
                    if not base:
                        return jsonify({"error": "Course not found"}), 404
                    version_id = int(base[0]) if base[0] is not None else None
                    if not version_id:
                        return jsonify({"error": "Course has no active version"}), 404
                    detail = _lms_course_structure_tx(cursor, course_id, version_id)
                    return jsonify({"status": "success", "course": detail}), 200

                cursor.execute("""
                    SELECT
                        c.id, c.slug, c.title, c.description, c.category, c.status,
                        c.default_pass_threshold, c.default_attempt_limit,
                        c.current_version_id,
                        cv.version_number, cv.status, cv.pass_threshold, cv.attempt_limit,
                        cv.anti_cheat_settings
                    FROM lms_courses c
                    LEFT JOIN lms_course_versions cv ON cv.id = c.current_version_id
                    ORDER BY c.updated_at DESC, c.id DESC
                """)
                rows = cursor.fetchall()
                courses = []
                for row in rows:
                    version_settings = _lms_parse_json(row[13], {})
                    cover_payload = _lms_resolve_cover_payload(version_settings)
                    courses.append({
                        "id": int(row[0]),
                        "slug": row[1],
                        "title": row[2],
                        "description": row[3],
                        "category": row[4],
                        "status": row[5],
                        "cover_url": cover_payload["cover_url"],
                        "cover_bucket": cover_payload["cover_bucket"],
                        "cover_blob_path": cover_payload["cover_blob_path"],
                        "skills": _lms_normalize_skills(version_settings.get("skills")),
                        "default_pass_threshold": float(row[6] if row[6] is not None else LMS_DEFAULT_PASS_THRESHOLD),
                        "default_attempt_limit": int(row[7] if row[7] is not None else LMS_DEFAULT_ATTEMPT_LIMIT),
                        "current_version_id": int(row[8]) if row[8] is not None else None,
                        "current_version": {
                            "version_number": int(row[9]) if row[9] is not None else None,
                            "status": row[10],
                            "pass_threshold": float(row[11]) if row[11] is not None else None,
                            "attempt_limit": int(row[12]) if row[12] is not None else None
                        } if row[8] is not None else None
                    })
                return jsonify({"status": "success", "courses": courses}), 200

        data = request.get_json(silent=True) or {}
        if request.method == 'POST':
            title = str(data.get('title') or '').strip()
            if not title:
                return jsonify({"error": "title is required"}), 400

            description = str(data.get('description') or '').strip() or None
            category = str(data.get('category') or '').strip() or None
            cover_url = str(data.get('cover_url') or '').strip() or None
            cover_bucket = str(data.get('cover_bucket') or '').strip() or None
            cover_blob_path = str(data.get('cover_blob_path') or '').strip() or None
            skills = _lms_normalize_skills(data.get('skills'))
            pass_threshold = min(100.0, max(0.0, _lms_to_float(data.get('pass_threshold'), LMS_DEFAULT_PASS_THRESHOLD)))
            attempt_limit = max(1, _lms_to_int(data.get('attempt_limit'), LMS_DEFAULT_ATTEMPT_LIMIT))
            modules = data.get('modules') if isinstance(data.get('modules'), list) else []
            tests = data.get('tests') if isinstance(data.get('tests'), list) else []
            if not tests and isinstance(data.get('test'), dict):
                tests = [data.get('test')]

            base_slug = re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-') or 'course'
            slug = f"{base_slug[:80]}-{secrets.token_hex(3)}"

            with db._get_cursor() as cursor:
                now = _lms_now()
                cursor.execute("""
                    INSERT INTO lms_courses (
                        slug, title, description, category, status,
                        default_pass_threshold, default_attempt_limit,
                        created_by, updated_by, created_at, updated_at
                    )
                    VALUES (%s, %s, %s, %s, 'draft', %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (slug, title, description, category, pass_threshold, attempt_limit, requester_id, requester_id, now, now))
                course_id = int(cursor.fetchone()[0])

                cursor.execute("""
                    INSERT INTO lms_course_versions (
                        course_id, version_number, title, description, status,
                        pass_threshold, attempt_limit, anti_cheat_settings, created_by, created_at
                    )
                    VALUES (%s, 1, %s, %s, 'draft', %s, %s, %s::jsonb, %s, %s)
                    RETURNING id
                """, (
                    course_id,
                    title,
                    description,
                    pass_threshold,
                    attempt_limit,
                    json.dumps({
                        "heartbeat_seconds": LMS_HEARTBEAT_SECONDS,
                        "stale_gap_seconds": LMS_STALE_GAP_SECONDS,
                        "completion_threshold_percent": LMS_COMPLETION_THRESHOLD,
                        "cover_url": cover_url,
                        "cover_bucket": cover_bucket,
                        "cover_blob_path": cover_blob_path,
                        "skills": skills
                    }, ensure_ascii=False),
                    requester_id,
                    now
                ))
                version_id = int(cursor.fetchone()[0])

                cursor.execute("UPDATE lms_courses SET current_version_id = %s WHERE id = %s", (version_id, course_id))

                created_module_ids = []
                for module_index, module in enumerate(modules, start=1):
                    if not isinstance(module, dict):
                        continue
                    module_title = str(module.get('title') or '').strip()
                    if not module_title:
                        continue
                    module_desc = str(module.get('description') or '').strip() or None
                    module_pos = max(1, _lms_to_int(module.get('position'), module_index))
                    cursor.execute("""
                        INSERT INTO lms_modules (course_version_id, title, description, position, created_at)
                        VALUES (%s, %s, %s, %s, %s)
                        RETURNING id
                    """, (version_id, module_title, module_desc, module_pos, now))
                    module_id = int(cursor.fetchone()[0])
                    created_module_ids.append(module_id)

                    lessons = module.get('lessons') if isinstance(module.get('lessons'), list) else []
                    for lesson_index, lesson in enumerate(lessons, start=1):
                        if not isinstance(lesson, dict):
                            continue
                        lesson_title = str(lesson.get('title') or '').strip()
                        if not lesson_title:
                            continue
                        lesson_desc = str(lesson.get('description') or '').strip() or None
                        lesson_type = str(lesson.get('lesson_type') or lesson.get('type') or '').strip().lower()
                        lesson_materials = lesson.get('materials') if isinstance(lesson.get('materials'), list) else []
                        cursor.execute("""
                            INSERT INTO lms_lessons (
                                module_id, title, description, position, duration_seconds,
                                allow_fast_forward, completion_threshold, created_at
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            RETURNING id
                        """, (
                            module_id,
                            lesson_title,
                            lesson_desc,
                            max(1, _lms_to_int(lesson.get('position'), lesson_index)),
                            max(0, _lms_to_int(lesson.get('duration_seconds'), 0)),
                            _lms_parse_bool(lesson.get('allow_fast_forward'), False),
                            min(100.0, max(0.0, _lms_to_float(lesson.get('completion_threshold'), LMS_COMPLETION_THRESHOLD))),
                            now
                        ))
                        lesson_id = int(cursor.fetchone()[0])

                        material_position = 1
                        has_text_material = False
                        for material in lesson_materials:
                            if not isinstance(material, dict):
                                continue
                            material_type = str(material.get('material_type') or material.get('type') or 'file').strip().lower()
                            if material_type not in ('video', 'pdf', 'link', 'text', 'file'):
                                material_type = 'file'
                            content_url = str(material.get('content_url') or material.get('url') or '').strip() or None
                            content_text = str(material.get('content_text') or '').strip() or None
                            if material_type == 'text' and content_text:
                                has_text_material = True
                            if not content_url and not content_text:
                                continue
                            title_raw = str(material.get('title') or '').strip()
                            material_title = title_raw or ("Текстовый материал" if material_type == 'text' else f"Материал {material_position}")
                            mime_type = str(material.get('mime_type') or material.get('content_type') or '').strip() or None
                            gcs_bucket = str(material.get('bucket') or material.get('gcs_bucket') or '').strip() or None
                            gcs_blob_path = str(material.get('blob_path') or material.get('gcs_blob_path') or '').strip() or None
                            metadata = material.get('metadata') if isinstance(material.get('metadata'), dict) else {}
                            uploaded_name = str(material.get('file_name') or material.get('uploaded_file_name') or '').strip()
                            if uploaded_name:
                                metadata["uploaded_file_name"] = uploaded_name
                            position = max(1, _lms_to_int(material.get('position'), material_position))
                            cursor.execute("""
                                INSERT INTO lms_lesson_materials (
                                    lesson_id, title, material_type, content_text, content_url,
                                    gcs_bucket, gcs_blob_path, mime_type, metadata, position, created_by, created_at
                                )
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s, %s, %s)
                            """, (
                                lesson_id,
                                material_title,
                                material_type,
                                content_text,
                                content_url,
                                gcs_bucket,
                                gcs_blob_path,
                                mime_type,
                                json.dumps(metadata, ensure_ascii=False),
                                position,
                                requester_id,
                                now
                            ))
                            material_position += 1

                        if lesson_type == 'text' and not has_text_material:
                            fallback_text = str(lesson.get('content_text') or lesson_desc or '').strip()
                            if fallback_text:
                                cursor.execute("""
                                    INSERT INTO lms_lesson_materials (
                                        lesson_id, title, material_type, content_text, content_url,
                                        gcs_bucket, gcs_blob_path, mime_type, metadata, position, created_by, created_at
                                    )
                                    VALUES (%s, %s, 'text', %s, NULL, NULL, NULL, 'text/plain', %s::jsonb, %s, %s, %s)
                                """, (
                                    lesson_id,
                                    "Текстовый материал",
                                    fallback_text,
                                    json.dumps({}, ensure_ascii=False),
                                    material_position,
                                    requester_id,
                                    now
                                ))

                for test_index, test in enumerate(tests, start=1):
                    if not isinstance(test, dict):
                        continue
                    test_title = str(test.get('title') or '').strip() or f"Тест {test_index}"
                    test_module_id = test.get('module_id')
                    if test_module_id not in created_module_ids:
                        test_module_id = None
                    cursor.execute("""
                        INSERT INTO lms_tests (
                            course_version_id, module_id, title, description, pass_threshold,
                            attempt_limit, is_final, status, created_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, 'draft', %s)
                        RETURNING id
                    """, (
                        version_id,
                        test_module_id,
                        test_title,
                        str(test.get('description') or '').strip() or None,
                        min(100.0, max(0.0, _lms_to_float(test.get('pass_threshold'), pass_threshold))),
                        max(1, _lms_to_int(test.get('attempt_limit'), attempt_limit)),
                        _lms_parse_bool(test.get('is_final'), True),
                        now
                    ))
                    test_id = int(cursor.fetchone()[0])

                    questions = test.get('questions') if isinstance(test.get('questions'), list) else []
                    for q_index, question in enumerate(questions, start=1):
                        if not isinstance(question, dict):
                            continue
                        prompt = str(question.get('prompt') or question.get('text') or '').strip()
                        if not prompt:
                            continue
                        q_type = str(question.get('type') or 'single').strip().lower()
                        if q_type in ('single_choice',):
                            q_type = 'single'
                        if q_type in ('multiple_choice', 'multi'):
                            q_type = 'multiple'
                        if q_type in ('boolean', 'truefalse'):
                            q_type = 'true_false'
                        if q_type not in ('single', 'multiple', 'true_false', 'matching', 'text'):
                            q_type = 'single'

                        cursor.execute("""
                            INSERT INTO lms_questions (
                                test_id, question_type, prompt, points, position, required, metadata, correct_text_answers, created_at
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, %s)
                            RETURNING id
                        """, (
                            test_id,
                            q_type,
                            prompt,
                            max(0.1, _lms_to_float(question.get('points'), 1.0)),
                            max(1, _lms_to_int(question.get('position'), q_index)),
                            _lms_parse_bool(question.get('required'), True),
                            json.dumps(question.get('metadata') if isinstance(question.get('metadata'), dict) else {}, ensure_ascii=False),
                            json.dumps(question.get('correct_text_answers') if isinstance(question.get('correct_text_answers'), list) else [], ensure_ascii=False),
                            now
                        ))
                        question_id = int(cursor.fetchone()[0])

                        options = question.get('options') if isinstance(question.get('options'), list) else []
                        if q_type == 'true_false' and not options:
                            correct_bool = _lms_parse_bool(question.get('correct'), True)
                            options = [
                                {"text": "True", "is_correct": correct_bool, "key": "true"},
                                {"text": "False", "is_correct": (not correct_bool), "key": "false"}
                            ]

                        for o_index, option in enumerate(options, start=1):
                            if isinstance(option, dict):
                                opt_text = str(option.get('text') or option.get('label') or '').strip()
                                opt_key = str(option.get('key') or '').strip() or None
                                is_correct = _lms_parse_bool(option.get('is_correct'), False)
                                match_key = str(option.get('match_key') or option.get('right') or '').strip() or None
                                opt_meta = option.get('metadata') if isinstance(option.get('metadata'), dict) else {}
                                opt_position = max(1, _lms_to_int(option.get('position'), o_index))
                            else:
                                opt_text = str(option or '').strip()
                                opt_key = None
                                is_correct = False
                                match_key = None
                                opt_meta = {}
                                opt_position = o_index
                            if not opt_text:
                                continue
                            cursor.execute("""
                                INSERT INTO lms_question_options (
                                    question_id, option_key, option_text, position, is_correct, match_key, metadata
                                )
                                VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb)
                            """, (
                                question_id,
                                opt_key,
                                opt_text,
                                opt_position,
                                is_correct,
                                match_key,
                                json.dumps(opt_meta, ensure_ascii=False)
                            ))

                _lms_audit(
                    cursor,
                    actor_id=requester_id,
                    actor_role=requester_role,
                    action='create_course',
                    entity_type='lms_course',
                    entity_id=course_id,
                    details={"course_version_id": version_id}
                )

            return jsonify({"status": "success", "course_id": course_id, "course_version_id": version_id}), 201

        if request.method == 'PATCH':
            course_id = _lms_to_int(data.get('course_id'), default=0)
            if course_id <= 0:
                return jsonify({"error": "course_id is required"}), 400
            with db._get_cursor() as cursor:
                cursor.execute("SELECT id, current_version_id FROM lms_courses WHERE id = %s", (course_id,))
                course_row = cursor.fetchone()
                if not course_row:
                    return jsonify({"error": "Course not found"}), 404

                updates = []
                params = []
                if 'title' in data:
                    updates.append("title = %s")
                    params.append(str(data.get('title') or '').strip())
                if 'description' in data:
                    updates.append("description = %s")
                    params.append(str(data.get('description') or '').strip() or None)
                if 'category' in data:
                    updates.append("category = %s")
                    params.append(str(data.get('category') or '').strip() or None)
                if 'status' in data:
                    status = str(data.get('status') or '').strip().lower()
                    if status not in ('draft', 'published', 'archived'):
                        return jsonify({"error": "Invalid status"}), 400
                    updates.append("status = %s")
                    params.append(status)
                if 'pass_threshold' in data:
                    updates.append("default_pass_threshold = %s")
                    params.append(min(100.0, max(0.0, _lms_to_float(data.get('pass_threshold'), LMS_DEFAULT_PASS_THRESHOLD))))
                if 'attempt_limit' in data:
                    updates.append("default_attempt_limit = %s")
                    params.append(max(1, _lms_to_int(data.get('attempt_limit'), LMS_DEFAULT_ATTEMPT_LIMIT)))
                updates.append("updated_by = %s")
                params.append(requester_id)
                updates.append("updated_at = %s")
                params.append(_lms_now())

                cursor.execute(f"UPDATE lms_courses SET {', '.join(updates)} WHERE id = %s", params + [course_id])

                version_id = _lms_to_int(data.get('course_version_id'), default=int(course_row[1] or 0))
                version_updates = []
                version_params = []
                if version_id > 0:
                    next_version_settings = None
                    if 'version_status' in data:
                        status = str(data.get('version_status') or '').strip().lower()
                        if status not in ('draft', 'published', 'archived'):
                            return jsonify({"error": "Invalid version_status"}), 400
                        version_updates.append("status = %s")
                        version_params.append(status)
                    if 'version_title' in data:
                        version_updates.append("title = %s")
                        version_params.append(str(data.get('version_title') or '').strip() or None)
                    if 'version_description' in data:
                        version_updates.append("description = %s")
                        version_params.append(str(data.get('version_description') or '').strip() or None)
                    if 'pass_threshold' in data:
                        version_updates.append("pass_threshold = %s")
                        version_params.append(min(100.0, max(0.0, _lms_to_float(data.get('pass_threshold'), LMS_DEFAULT_PASS_THRESHOLD))))
                    if 'attempt_limit' in data:
                        version_updates.append("attempt_limit = %s")
                        version_params.append(max(1, _lms_to_int(data.get('attempt_limit'), LMS_DEFAULT_ATTEMPT_LIMIT)))
                    if 'cover_url' in data or 'cover_bucket' in data or 'cover_blob_path' in data or 'skills' in data:
                        cursor.execute("""
                            SELECT anti_cheat_settings
                            FROM lms_course_versions
                            WHERE id = %s AND course_id = %s
                            LIMIT 1
                        """, (version_id, course_id))
                        settings_row = cursor.fetchone()
                        if not settings_row:
                            return jsonify({"error": "Version not found for course"}), 404
                        next_version_settings = _lms_parse_json(settings_row[0], {})
                        if not isinstance(next_version_settings, dict):
                            next_version_settings = {}
                        if 'cover_url' in data:
                            next_version_settings['cover_url'] = (str(data.get('cover_url') or '').strip() or None)
                        if 'cover_bucket' in data:
                            next_version_settings['cover_bucket'] = (str(data.get('cover_bucket') or '').strip() or None)
                        if 'cover_blob_path' in data:
                            next_version_settings['cover_blob_path'] = (str(data.get('cover_blob_path') or '').strip() or None)
                        if 'skills' in data:
                            next_version_settings['skills'] = _lms_normalize_skills(data.get('skills'))
                        version_updates.append("anti_cheat_settings = %s::jsonb")
                        version_params.append(json.dumps(next_version_settings, ensure_ascii=False))
                    if version_updates:
                        cursor.execute(f"""
                            UPDATE lms_course_versions
                            SET {', '.join(version_updates)}
                            WHERE id = %s AND course_id = %s
                        """, version_params + [version_id, course_id])

                _lms_audit(
                    cursor,
                    actor_id=requester_id,
                    actor_role=requester_role,
                    action='update_course',
                    entity_type='lms_course',
                    entity_id=course_id,
                    details={"payload_keys": list(data.keys())}
                )

            return jsonify({"status": "success", "course_id": course_id}), 200

        return jsonify({"error": "Method not allowed"}), 405
    except Exception as e:
        logging.exception("Error in /api/lms/admin/courses")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/admin/courses/<int:course_id>', methods=['DELETE'])
@require_api_key
def lms_admin_delete_course(course_id):
    requester_id, _, requester_role, error_response, status_code = _lms_resolve_request('manager')
    if error_response:
        return error_response, status_code

    try:
        cleanup_result = {"attempted": 0, "deleted": 0, "failed": []}
        with db._get_cursor() as cursor:
            cursor.execute("SELECT id, title FROM lms_courses WHERE id = %s LIMIT 1", (course_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({"error": "Course not found"}), 404

            course_title = str(row[1] or f"Course #{course_id}")
            blob_refs = _lms_collect_course_blob_refs_tx(cursor, course_id)
            cleanup_result = _lms_delete_blob_refs(blob_refs)
            if cleanup_result.get("failed"):
                first_error = cleanup_result["failed"][0]
                raise RuntimeError(
                    "Failed to delete one or more course files from GCS "
                    f"(example: {first_error.get('bucket')}/{first_error.get('blob_path')})"
                )

            cursor.execute("DELETE FROM lms_courses WHERE id = %s", (course_id,))
            if cursor.rowcount <= 0:
                return jsonify({"error": "Course not found"}), 404

            _lms_audit(
                cursor,
                actor_id=requester_id,
                actor_role=requester_role,
                action='delete_course',
                entity_type='lms_course',
                entity_id=course_id,
                details={
                    "course_title": course_title,
                    "gcs_blobs_attempted": int(cleanup_result.get("attempted") or 0),
                    "gcs_blobs_deleted": int(cleanup_result.get("deleted") or 0)
                }
            )

        return jsonify({
            "status": "success",
            "course_id": course_id,
            "deleted_gcs_blobs": int(cleanup_result.get("deleted") or 0)
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/admin/courses/<course_id> DELETE")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/admin/courses/<int:course_id>/publish', methods=['POST'])
@require_api_key
def lms_admin_publish_course(course_id):
    requester_id, _, requester_role, error_response, status_code = _lms_resolve_request('manager')
    if error_response:
        return error_response, status_code

    data = request.get_json(silent=True) or {}
    version_id_raw = data.get('course_version_id')
    try:
        version_id = int(version_id_raw) if version_id_raw not in (None, '') else None
    except Exception:
        return jsonify({"error": "Invalid course_version_id"}), 400

    try:
        with db._get_cursor() as cursor:
            cursor.execute("SELECT id FROM lms_courses WHERE id = %s", (course_id,))
            if not cursor.fetchone():
                return jsonify({"error": "Course not found"}), 404

            if version_id is None:
                cursor.execute("""
                    SELECT id
                    FROM lms_course_versions
                    WHERE course_id = %s
                    ORDER BY version_number DESC, id DESC
                    LIMIT 1
                """, (course_id,))
                row = cursor.fetchone()
                if not row:
                    return jsonify({"error": "Course has no versions"}), 404
                version_id = int(row[0])
            else:
                cursor.execute("""
                    SELECT id
                    FROM lms_course_versions
                    WHERE id = %s
                      AND course_id = %s
                    LIMIT 1
                """, (version_id, course_id))
                if not cursor.fetchone():
                    return jsonify({"error": "Version not found for course"}), 404

            now = _lms_now()
            cursor.execute("""
                UPDATE lms_course_versions
                SET status = 'archived'
                WHERE course_id = %s
                  AND status = 'published'
                  AND id <> %s
            """, (course_id, version_id))
            cursor.execute("""
                UPDATE lms_course_versions
                SET status = 'published',
                    published_by = %s,
                    published_at = %s
                WHERE id = %s
            """, (requester_id, now, version_id))
            cursor.execute("""
                UPDATE lms_tests
                SET status = 'published'
                WHERE course_version_id = %s
            """, (version_id,))
            cursor.execute("""
                UPDATE lms_courses
                SET status = 'published',
                    current_version_id = %s,
                    updated_by = %s,
                    updated_at = %s
                WHERE id = %s
            """, (version_id, requester_id, now, course_id))

            _lms_audit(
                cursor,
                actor_id=requester_id,
                actor_role=requester_role,
                action='publish_course',
                entity_type='lms_course',
                entity_id=course_id,
                details={"course_version_id": version_id}
            )

        return jsonify({"status": "success", "course_id": course_id, "course_version_id": version_id}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/admin/courses/<course_id>/publish")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/admin/courses/<int:course_id>/assignments', methods=['POST'])
@require_api_key
def lms_admin_assign_course(course_id):
    requester_id, _, requester_role, error_response, status_code = _lms_resolve_request('manager')
    if error_response:
        return error_response, status_code

    data = request.get_json(silent=True) or {}
    user_ids_raw = data.get('user_ids') if isinstance(data.get('user_ids'), list) else []
    if not user_ids_raw:
        return jsonify({"error": "user_ids must be a non-empty list"}), 400

    due_at = _lms_parse_datetime(data.get('due_at'))
    version_id_raw = data.get('course_version_id')
    try:
        version_id = int(version_id_raw) if version_id_raw not in (None, '') else None
    except Exception:
        return jsonify({"error": "Invalid course_version_id"}), 400

    user_ids = []
    for item in user_ids_raw:
        try:
            parsed = int(item)
        except Exception:
            continue
        if parsed > 0 and parsed not in user_ids:
            user_ids.append(parsed)
    if not user_ids:
        return jsonify({"error": "No valid user_ids"}), 400

    visible_learner_ids = None
    if requester_role in ('sv', 'trainer'):
        visible_learner_ids = set(_lms_visible_learner_ids(requester_id, requester_role))

    try:
        with db._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, current_version_id
                FROM lms_courses
                WHERE id = %s
                LIMIT 1
            """, (course_id,))
            course_row = cursor.fetchone()
            if not course_row:
                return jsonify({"error": "Course not found"}), 404

            if version_id is None:
                version_id = int(course_row[1]) if course_row[1] is not None else None
            if not version_id:
                cursor.execute("""
                    SELECT id
                    FROM lms_course_versions
                    WHERE course_id = %s
                    ORDER BY version_number DESC, id DESC
                    LIMIT 1
                """, (course_id,))
                row = cursor.fetchone()
                if not row:
                    return jsonify({"error": "Course version not found"}), 404
                version_id = int(row[0])

            cursor.execute("""
                SELECT id
                FROM lms_course_versions
                WHERE id = %s
                  AND course_id = %s
                LIMIT 1
            """, (version_id, course_id))
            if not cursor.fetchone():
                return jsonify({"error": "course_version_id does not belong to course"}), 400

            assigned = []
            skipped = []
            now = _lms_now()
            for user_id in user_ids:
                if visible_learner_ids is not None and user_id not in visible_learner_ids:
                    skipped.append({"user_id": user_id, "reason": "not_visible_for_manager"})
                    continue

                cursor.execute("SELECT id, role FROM users WHERE id = %s LIMIT 1", (user_id,))
                user_row = cursor.fetchone()
                if not user_row:
                    skipped.append({"user_id": user_id, "reason": "user_not_found"})
                    continue
                role = _normalize_user_role(user_row[1])
                if role not in LMS_LEARNER_ROLES:
                    skipped.append({"user_id": user_id, "reason": "user_not_learner"})
                    continue

                cursor.execute("""
                    INSERT INTO lms_course_assignments (
                        course_id, course_version_id, user_id, assigned_by, due_at, status, created_at, updated_at
                    )
                    VALUES (%s, %s, %s, %s, %s, 'assigned', %s, %s)
                    ON CONFLICT (course_id, user_id)
                    DO UPDATE SET
                        course_version_id = EXCLUDED.course_version_id,
                        assigned_by = EXCLUDED.assigned_by,
                        due_at = EXCLUDED.due_at,
                        status = CASE
                            WHEN lms_course_assignments.status = 'completed' THEN 'assigned'
                            ELSE lms_course_assignments.status
                        END,
                        completed_at = CASE
                            WHEN lms_course_assignments.status = 'completed' THEN NULL
                            ELSE lms_course_assignments.completed_at
                        END,
                        completion_color_status = CASE
                            WHEN lms_course_assignments.status = 'completed' THEN NULL
                            ELSE lms_course_assignments.completion_color_status
                        END,
                        updated_at = EXCLUDED.updated_at
                    RETURNING id
                """, (course_id, version_id, user_id, requester_id, due_at, now, now))
                assignment_id = int(cursor.fetchone()[0])

                _lms_emit_notification(
                    cursor,
                    user_id=user_id,
                    notification_type='assignment',
                    title='Назначен новый курс',
                    message='Вам назначен новый курс в разделе Обучение.',
                    payload={
                        "assignment_id": assignment_id,
                        "course_id": course_id,
                        "course_version_id": version_id,
                        "due_at": due_at.isoformat() if due_at else None
                    }
                )
                assigned.append({"user_id": user_id, "assignment_id": assignment_id})

            _lms_audit(
                cursor,
                actor_id=requester_id,
                actor_role=requester_role,
                action='assign_course',
                entity_type='lms_course',
                entity_id=course_id,
                details={
                    "course_version_id": version_id,
                    "assigned_count": len(assigned),
                    "skipped_count": len(skipped)
                }
            )

        return jsonify({
            "status": "success",
            "course_id": course_id,
            "course_version_id": version_id,
            "assigned": assigned,
            "skipped": skipped
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/admin/courses/<course_id>/assignments")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/admin/learners', methods=['GET'])
@require_api_key
def lms_admin_learners():
    requester_id, _, requester_role, error_response, status_code = _lms_resolve_request('manager')
    if error_response:
        return error_response, status_code

    try:
        visible_ids = None
        if requester_role in ('sv', 'trainer'):
            visible_ids = set(_lms_visible_learner_ids(requester_id, requester_role))

        with db._get_cursor() as cursor:
            params = [list(LMS_LEARNER_ROLES)]
            where = ["LOWER(TRIM(COALESCE(role, ''))) = ANY(%s)"]

            if visible_ids is not None:
                if not visible_ids:
                    return jsonify({"status": "success", "learners": []}), 200
                where.append("id = ANY(%s)")
                params.append(list(visible_ids))

            cursor.execute(f"""
                SELECT id, name, role
                FROM users
                WHERE {' AND '.join(where)}
                ORDER BY name ASC, id ASC
            """, params)
            rows = cursor.fetchall()

        learners = [{
            "id": int(row[0]),
            "name": row[1] or f"User #{row[0]}",
            "role": _normalize_user_role(row[2])
        } for row in rows]

        return jsonify({"status": "success", "learners": learners}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/admin/learners")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/admin/materials/upload', methods=['POST'])
@require_api_key
def lms_admin_upload_materials():
    requester_id, _, requester_role, error_response, status_code = _lms_resolve_request('manager')
    if error_response:
        return error_response, status_code

    try:
        files = request.files.getlist('files')
        if not files:
            single = request.files.get('file')
            files = [single] if single else []
        files = [f for f in files if f and f.filename]
        if not files:
            return jsonify({"error": "No files were uploaded"}), 400

        bucket_name = _lms_bucket_name()
        if not bucket_name:
            return jsonify({"error": "LMS GCS bucket is not configured"}), 500

        lesson_id_raw = request.form.get('lesson_id')
        lesson_id = None
        if lesson_id_raw not in (None, ''):
            try:
                lesson_id = int(lesson_id_raw)
            except Exception:
                return jsonify({"error": "Invalid lesson_id"}), 400
        requested_material_type = str(request.form.get('material_type') or 'file').strip().lower()
        if requested_material_type not in ('video', 'pdf', 'link', 'text', 'file', 'cover'):
            requested_material_type = 'file'
        replace_bucket = str(request.form.get('replace_bucket') or '').strip() or None
        replace_blob_path = str(request.form.get('replace_blob_path') or '').strip() or None
        replace_requested = bool(replace_bucket and replace_blob_path)
        if not replace_requested:
            replace_bucket = None
            replace_blob_path = None

        client = get_gcs_client()
        bucket = client.bucket(bucket_name)
        uploaded = []
        replacement_cleanup = {
            "requested": replace_requested,
            "attempted": 0,
            "deleted": 0,
            "failed": []
        }
        now = _lms_now()
        with db._get_cursor() as cursor:
            if lesson_id is not None:
                cursor.execute("SELECT id FROM lms_lessons WHERE id = %s", (lesson_id,))
                if not cursor.fetchone():
                    return jsonify({"error": "Lesson not found"}), 404

            for idx, file_storage in enumerate(files, start=1):
                safe_name = secure_filename(file_storage.filename or f"material_{idx}")
                content = file_storage.read() or b''
                content_type = (file_storage.mimetype or '').strip() or 'application/octet-stream'
                upload_name = safe_name
                if requested_material_type == 'cover' and content_type.startswith('image/'):
                    converted = _lms_convert_image_to_webp(content)
                    if converted:
                        content = converted
                        content_type = 'image/webp'
                        name_root = os.path.splitext(safe_name)[0].strip() or f"cover_{idx}"
                        upload_name = f"{name_root}.webp"
                blob_path = (
                    f"lms/materials/{now.strftime('%Y/%m/%d')}/"
                    f"{uuid.uuid4().hex}_{upload_name}"
                )
                blob = bucket.blob(blob_path)
                blob.upload_from_string(content, content_type=content_type)
                signed_url = _lms_signed_url(bucket_name, blob_path, expires_minutes=240)

                material_id = None
                if lesson_id is not None:
                    title = str(request.form.get('title') or upload_name).strip() or upload_name
                    material_type = requested_material_type
                    if material_type == 'cover':
                        material_type = 'file'
                    if material_type not in ('video', 'pdf', 'link', 'text', 'file'):
                        material_type = 'file'
                    position = max(1, _lms_to_int(request.form.get('position'), idx))
                    cursor.execute("""
                        INSERT INTO lms_lesson_materials (
                            lesson_id, title, material_type, content_url,
                            gcs_bucket, gcs_blob_path, mime_type, metadata, position, created_by, created_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s, %s, %s)
                        RETURNING id
                    """, (
                        lesson_id,
                        title,
                        material_type,
                        signed_url or f"gs://{bucket_name}/{blob_path}",
                        bucket_name,
                        blob_path,
                        content_type,
                        json.dumps({"uploaded_file_name": upload_name}, ensure_ascii=False),
                        position,
                        requester_id,
                        now
                    ))
                    material_id = int(cursor.fetchone()[0])

                uploaded.append({
                    "material_id": material_id,
                    "file_name": upload_name,
                    "content_type": content_type,
                    "size": len(content),
                    "bucket": bucket_name,
                    "blob_path": blob_path,
                    "signed_url": signed_url
                })

            replacement_skipped_reason = None
            if replace_requested:
                if len(uploaded) != 1:
                    replacement_skipped_reason = "multiple_uploads_not_supported"
                elif replace_bucket != bucket_name:
                    replacement_skipped_reason = "bucket_mismatch"
                elif not str(replace_blob_path).startswith("lms/materials/"):
                    replacement_skipped_reason = "blob_path_out_of_scope"
                elif any(
                    (str(item.get("bucket") or "").strip() == replace_bucket)
                    and (str(item.get("blob_path") or "").strip() == replace_blob_path)
                    for item in uploaded
                ):
                    replacement_skipped_reason = "same_as_new_upload"
                else:
                    cleanup_result = _lms_delete_blob_refs([(replace_bucket, replace_blob_path)])
                    replacement_cleanup["attempted"] = int(cleanup_result.get("attempted") or 0)
                    replacement_cleanup["deleted"] = int(cleanup_result.get("deleted") or 0)
                    replacement_cleanup["failed"] = cleanup_result.get("failed") or []
            if replacement_skipped_reason:
                replacement_cleanup["skipped_reason"] = replacement_skipped_reason

            _lms_audit(
                cursor,
                actor_id=requester_id,
                actor_role=requester_role,
                action='upload_material',
                entity_type='lms_lesson',
                entity_id=lesson_id,
                details={
                    "files": [item["file_name"] for item in uploaded],
                    "replacement_cleanup": {
                        "requested": replacement_cleanup.get("requested"),
                        "attempted": replacement_cleanup.get("attempted"),
                        "deleted": replacement_cleanup.get("deleted"),
                        "failed_count": len(replacement_cleanup.get("failed") or []),
                        "skipped_reason": replacement_cleanup.get("skipped_reason")
                    }
                }
            )

        return jsonify({
            "status": "success",
            "uploaded": uploaded,
            "replacement_cleanup": replacement_cleanup
        }), 200
    except Exception as e:
        logging.exception("Error in /api/lms/admin/materials/upload")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/admin/progress', methods=['GET'])
@require_api_key
def lms_admin_progress():
    requester_id, _, requester_role, error_response, status_code = _lms_resolve_request('manager')
    if error_response:
        return error_response, status_code

    try:
        course_id_filter = _lms_to_int(request.args.get('course_id'), default=0)
        user_id_filter = _lms_to_int(request.args.get('user_id'), default=0)
        visible_ids = None
        if requester_role in ('sv', 'trainer'):
            visible_ids = set(_lms_visible_learner_ids(requester_id, requester_role))

        with db._get_cursor() as cursor:
            params = []
            where = ["1=1"]
            if course_id_filter > 0:
                where.append("a.course_id = %s")
                params.append(course_id_filter)
            if user_id_filter > 0:
                where.append("a.user_id = %s")
                params.append(user_id_filter)
            if visible_ids is not None:
                if not visible_ids:
                    return jsonify({"status": "success", "rows": []}), 200
                where.append("a.user_id = ANY(%s)")
                params.append(list(visible_ids))

            cursor.execute(f"""
                SELECT
                    a.id, a.course_id, c.title, a.course_version_id,
                    a.user_id, u.name, u.role,
                    a.status, a.due_at, a.started_at, a.completed_at, a.completion_color_status,
                    COALESCE(lp.total_lessons, 0) AS total_lessons,
                    COALESCE(lp.completed_lessons, 0) AS completed_lessons,
                    COALESCE(tp.total_tests, 0) AS total_tests,
                    COALESCE(tp.passed_tests, 0) AS passed_tests
                FROM lms_course_assignments a
                JOIN lms_courses c ON c.id = a.course_id
                JOIN users u ON u.id = a.user_id
                LEFT JOIN LATERAL (
                    SELECT
                        COUNT(*) AS total_lessons,
                        COUNT(*) FILTER (WHERE p.status = 'completed') AS completed_lessons
                    FROM lms_lessons l
                    JOIN lms_modules m ON m.id = l.module_id
                    LEFT JOIN lms_lesson_progress p
                      ON p.assignment_id = a.id
                     AND p.lesson_id = l.id
                     AND p.user_id = a.user_id
                    WHERE m.course_version_id = a.course_version_id
                ) lp ON TRUE
                LEFT JOIN LATERAL (
                    SELECT
                        COUNT(*) AS total_tests,
                        COUNT(*) FILTER (WHERE EXISTS (
                            SELECT 1
                            FROM lms_test_attempts ta
                            WHERE ta.assignment_id = a.id
                              AND ta.user_id = a.user_id
                              AND ta.test_id = t.id
                              AND ta.passed = TRUE
                        )) AS passed_tests
                    FROM lms_tests t
                    WHERE t.course_version_id = a.course_version_id
                      AND t.status <> 'archived'
                ) tp ON TRUE
                WHERE {" AND ".join(where)}
                ORDER BY a.updated_at DESC, a.id DESC
            """, params)
            rows = cursor.fetchall()

            payload = []
            for row in rows:
                deadline_status = row[11] or _lms_deadline_status(row[8], row[10])
                total_lessons = int(row[12] or 0)
                completed_lessons = int(row[13] or 0)
                progress_percent = 0.0
                if total_lessons > 0:
                    progress_percent = round((completed_lessons / total_lessons) * 100.0, 2)
                elif row[7] == 'completed':
                    progress_percent = 100.0
                payload.append({
                    "assignment_id": int(row[0]),
                    "course_id": int(row[1]),
                    "course_title": row[2],
                    "course_version_id": int(row[3]) if row[3] is not None else None,
                    "user_id": int(row[4]),
                    "user_name": row[5],
                    "user_role": row[6],
                    "status": row[7],
                    "due_at": row[8].isoformat() if row[8] else None,
                    "started_at": row[9].isoformat() if row[9] else None,
                    "completed_at": row[10].isoformat() if row[10] else None,
                    "deadline_status": deadline_status,
                    "total_lessons": total_lessons,
                    "completed_lessons": completed_lessons,
                    "progress_percent": progress_percent,
                    "total_tests": int(row[14] or 0),
                    "passed_tests": int(row[15] or 0)
                })

        return jsonify({"status": "success", "rows": payload}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/admin/progress")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/admin/attempts', methods=['GET'])
@require_api_key
def lms_admin_attempts():
    requester_id, _, requester_role, error_response, status_code = _lms_resolve_request('manager')
    if error_response:
        return error_response, status_code

    limit = min(max(_lms_to_int(request.args.get('limit'), default=200), 1), 1000)
    try:
        visible_ids = None
        if requester_role in ('sv', 'trainer'):
            visible_ids = set(_lms_visible_learner_ids(requester_id, requester_role))

        with db._get_cursor() as cursor:
            params = []
            where = ["1=1"]

            course_id_filter = _lms_to_int(request.args.get('course_id'), default=0)
            if course_id_filter > 0:
                where.append("a.course_id = %s")
                params.append(course_id_filter)

            user_id_filter = _lms_to_int(request.args.get('user_id'), default=0)
            if user_id_filter > 0:
                where.append("ta.user_id = %s")
                params.append(user_id_filter)

            if visible_ids is not None:
                if not visible_ids:
                    return jsonify({"status": "success", "attempts": []}), 200
                where.append("ta.user_id = ANY(%s)")
                params.append(list(visible_ids))

            params.append(limit)
            cursor.execute(f"""
                SELECT
                    ta.id, ta.assignment_id, ta.test_id, ta.user_id, u.name, u.role,
                    ta.attempt_no, ta.status, ta.score_percent, ta.passed,
                    ta.started_at, ta.finished_at, ta.duration_seconds,
                    t.title, a.course_id, c.title
                FROM lms_test_attempts ta
                JOIN users u ON u.id = ta.user_id
                JOIN lms_tests t ON t.id = ta.test_id
                JOIN lms_course_assignments a ON a.id = ta.assignment_id
                JOIN lms_courses c ON c.id = a.course_id
                WHERE {" AND ".join(where)}
                ORDER BY ta.started_at DESC, ta.id DESC
                LIMIT %s
            """, params)
            rows = cursor.fetchall()

            attempts = []
            for row in rows:
                attempts.append({
                    "attempt_id": int(row[0]),
                    "assignment_id": int(row[1]),
                    "test_id": int(row[2]),
                    "user_id": int(row[3]),
                    "user_name": row[4],
                    "user_role": row[5],
                    "attempt_no": int(row[6] or 1),
                    "status": row[7],
                    "score_percent": float(row[8]) if row[8] is not None else None,
                    "passed": bool(row[9]) if row[9] is not None else None,
                    "started_at": row[10].isoformat() if row[10] else None,
                    "finished_at": row[11].isoformat() if row[11] else None,
                    "duration_seconds": int(row[12] or 0) if row[12] is not None else None,
                    "test_title": row[13],
                    "course_id": int(row[14]),
                    "course_title": row[15]
                })

        return jsonify({"status": "success", "attempts": attempts}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/admin/attempts")
        return jsonify({"error": str(e)}), 500


@app.route('/api/lms/admin/certificates/<int:certificate_id>/revoke', methods=['POST'])
@require_api_key
def lms_admin_revoke_certificate(certificate_id):
    requester_id, _, requester_role, error_response, status_code = _lms_resolve_request('full_admin')
    if error_response:
        return error_response, status_code

    data = request.get_json(silent=True) or {}
    reason = str(data.get('reason') or '').strip() or 'Revoked by administrator'

    try:
        with db._get_cursor() as cursor:
            now = _lms_now()
            cursor.execute("""
                UPDATE lms_certificates
                SET status = 'revoked',
                    revoked_at = %s,
                    revoked_by = %s,
                    revoke_reason = %s
                WHERE id = %s
                RETURNING user_id, certificate_number
            """, (now, requester_id, reason, certificate_id))
            row = cursor.fetchone()
            if not row:
                return jsonify({"error": "Certificate not found"}), 404

            user_id = int(row[0])
            certificate_number = row[1]
            _lms_emit_notification(
                cursor,
                user_id=user_id,
                notification_type='certificate_revoked',
                title='Сертификат отозван',
                message=f"Сертификат #{certificate_number} отозван администратором.",
                payload={"certificate_id": certificate_id, "reason": reason}
            )
            _lms_audit(
                cursor,
                actor_id=requester_id,
                actor_role=requester_role,
                action='revoke_certificate',
                entity_type='lms_certificate',
                entity_id=certificate_id,
                details={"reason": reason}
            )

        return jsonify({"status": "success", "certificate_id": certificate_id}), 200
    except Exception as e:
        logging.exception("Error in /api/lms/admin/certificates/<certificate_id>/revoke")
        return jsonify({"error": str(e)}), 500

def extract_fio_and_links(spreadsheet_url):
    try:
        match = re.search(r"/d/([a-zA-Z0-9_-]+)", spreadsheet_url)
        if not match:
            return None, None, "Ошибка: Неверный формат ссылки на Google Sheets."
        file_id = match.group(1)
        export_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"

        response = requests.get(export_url)
        if response.status_code != 200:
            return None, None, "Ошибка: Не удалось скачать таблицу. Проверьте доступность."

        temp_file = f"temp_table_{threading.current_thread().ident}.xlsx"
        with open(temp_file, "wb") as f:
            f.write(response.content)

        wb = load_workbook(temp_file, data_only=True)
        ws = wb.worksheets[-1]
        sheet_name = ws.title

        fio_column = None
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value is not None and "ФИО" in str(cell.value).strip():
                    fio_column = cell.column

        if not fio_column:
            os.remove(temp_file)
            return None, None, "Колонка ФИО не найдена."

        operators = []
        for row in ws.iter_rows(min_row=2):
            fio_cell = row[fio_column - 1]
            if not fio_cell.value:
                break
            operator_info = {
                "name": str(fio_cell.value)
            }
            operators.append(operator_info)

        os.remove(temp_file)
        return sheet_name, operators, None
    except Exception as e:
        if 'temp_file' in locals() and os.path.exists(temp_file):
            os.remove(temp_file)
        logging.error(f"Ошибка обработки таблицы: {str(e)}")
        return None, None, f"Ошибка обработки: {str(e)}"

def sync_generate_weekly_report():
    try:
        if not report_lock.acquire(blocking=False):
            logging.warning("Report generation skipped: already running")
            return False
            
        logging.info("Starting weekly report generation")
        
        current_date = datetime.now()
        current_month = current_date.strftime('%Y-%m')
        current_week = get_current_week_of_month()
        now = current_date.strftime("%Y-%m-%d %H:%M:%S")
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1
        })
        cell_format_int = workbook.add_format({'border': 1, 'num_format': '0'})
        cell_format_float = workbook.add_format({'border': 1, 'num_format': '0.00'})
        
        # Определение недель текущего месяца
        first_day = current_date.replace(day=1)
        weeks = []
        for w in range(1, current_week + 1):
            start_day = (w - 1) * 7 + 1
            start_date = first_day + timedelta(days=start_day - 1)
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_day = min(start_day + 6, current_date.day)
            end_date = first_day + timedelta(days=end_day - 1)
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            weeks.append((w, start_date, end_date))
        
        # Get supervisors from the database
        svs = db.get_supervisors()
        
        for sv_id, sv_name, _, _, _, _ in svs:
            operators = db.get_operators_by_supervisor(sv_id)
            
            safe_sheet_name = sv_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
            worksheet = workbook.add_worksheet(safe_sheet_name)
            
            # Headers
            headers = ['ФИО']
            for w, _, _ in weeks:
                headers.append(f'Неделя {w} Количество звонков')
                headers.append(f'Неделя {w} Средний балл')
            
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
            
            for row_idx, op in enumerate(operators, start=1):
                op_name = op['name']
                worksheet.write(row_idx, 0, op_name, cell_format_int)
                col = 1
                for w, start_date, end_date in weeks:
                    call_count, avg_score = db.get_week_call_stats(op['id'], start_date, end_date)
                    worksheet.write(row_idx, col, call_count, cell_format_int)
                    col += 1
                    worksheet.write(row_idx, col, avg_score, cell_format_float)
                    col += 1
            
            worksheet.set_column('A:A', 30)
            for i in range(1, len(headers)):
                worksheet.set_column(i, i, 20)
        
        workbook.close()
        output.seek(0)
        
        if output.getvalue():
            filename = f"Weekly_Report_{current_month}_{current_date.strftime('%Y%m%d')}.xlsx"
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendDocument"
            
            # Fetch admins from the database
            try:
                admins = []
                with db._get_cursor() as cursor:
                    cursor.execute("SELECT telegram_id FROM users WHERE role = 'admin'")
                    admins = [row[0] for row in cursor.fetchall()]
            except Exception as e:
                logging.error(f"Error fetching admins: {e}")
                return False
            
            # Send report to all admins
            for admin_id in admins:
                files = {'document': (filename, output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {'chat_id': admin_id, 'caption': f"[{now}] 📊 Отчет по неделям за {current_month} (до {current_week}-й недели)"}
                response = requests.post(telegram_url, files=files, data=data)
                
                if response.status_code != 200:
                    logging.error(f"Error sending report to admin {admin_id}: {response.text}")
            
            return True
        else:
            logging.error("Report is empty")
            return False
    except Exception as e:
        logging.error(f"Critical error in report generation: {e}")
        return False
    finally:
        try:
            report_lock.release()
        except:
            pass

async def generate_weekly_report():
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(executor_pool, sync_generate_weekly_report)

def sync_schedule_statuses_to_user_statuses_job():
    """Background job: sync users.status with active schedule status periods."""
    try:
        updated = db.sync_user_statuses_from_schedule_periods()
        if updated:
            logging.info("Auto status sync updated %s operator statuses", updated)
    except Exception as e:
        logging.exception("Error in auto status sync job: %s", e)

# === Главный запуск =============================================================================================
if __name__ == '__main__':
    
    # Запускаем Flask в отдельном потоке
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    
    # Настраиваем и запускаем планировщик
    async def monthly_auto_fill_norm():
        """Coroutine executed on the 1st day of each month to auto-fill norm_hours."""
        try:
            month = datetime.now().strftime('%Y-%m')
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(executor_pool, db.auto_fill_norm_hours, month)
            logging.info(f"auto_fill_norm_hours result: {result}")
        except Exception as e:
            logging.exception(f"Error running monthly_auto_fill_norm: {e}")

    scheduler = AsyncIOScheduler()
    scheduler.add_job(
        generate_weekly_report, 
        CronTrigger(day_of_week='mon', hour=9, minute=0),
        misfire_grace_time=3600
    )

    # Запуск авто-заполнения нормы часов: 1-й день каждого месяца в 03:00
    scheduler.add_job(
        monthly_auto_fill_norm,
        CronTrigger(day='1', hour=3, minute=0),
        misfire_grace_time=3600
    )

    # Автопереключение статусов операторов по периодным статусам графика (ежедневно в полночь)
    scheduler.add_job(
        sync_schedule_statuses_to_user_statuses_job,
        CronTrigger(hour=0, minute=0),
        id='sync_schedule_statuses_to_user_statuses',
        misfire_grace_time=120,
        max_instances=1,
        coalesce=True
    )

    scheduler.start()

    # Стартовый прогон (не ждём первой минуты)
    try:
        sync_schedule_statuses_to_user_statuses_job()
    except Exception:
        logging.exception("Initial auto status sync failed")
    
    logging.info("🔄 Планировщик запущен")
    logging.info("🤖 Бот запущен")
    
    # Запускаем бота
    executor.start_polling(dp, skip_updates=True)

