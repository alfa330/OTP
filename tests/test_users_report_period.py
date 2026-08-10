# -*- coding: utf-8 -*-
"""Regression tests for the employee report calculation-month filter.

``database.py`` is intentionally not imported here: constructing ``Database``
initializes the application schema.  The report method is extracted through the
AST and run with a fake cursor, following the other Excel-report tests.
"""

import ast
import calendar
import logging
import re
import textwrap
import unittest
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATABASE_PATH = ROOT / "database.py"


def _load_users_report_builder():
    source = DATABASE_PATH.read_text(encoding="utf-8-sig")
    module = ast.parse(source)
    database_class = next(
        node
        for node in module.body
        if isinstance(node, ast.ClassDef) and node.name == "Database"
    )
    method = next(
        node
        for node in database_class.body
        if isinstance(node, ast.FunctionDef) and node.name == "generate_users_report"
    )
    method_source = textwrap.dedent(ast.get_source_segment(source, method))
    namespace = {
        "BytesIO": BytesIO,
        "Font": Font,
        "PROXY_STATUS_LABELS": {},
        "Table": Table,
        "TableStyleInfo": TableStyleInfo,
        "Workbook": Workbook,
        "calendar": calendar,
        "date": date,
        "datetime": datetime,
        "defaultdict": defaultdict,
        "get_column_letter": get_column_letter,
        "logging": logging,
        "re": re,
    }
    exec(compile(method_source, str(DATABASE_PATH), "exec"), namespace)
    return namespace["generate_users_report"]


class _CursorContext:
    def __init__(self, cursor):
        self.cursor = cursor

    def __enter__(self):
        return self.cursor

    def __exit__(self, exc_type, exc_value, traceback):
        return False


class _FakeColumn:
    """Минимум от psycopg2.extensions.Column: отчёт читает только .name."""

    def __init__(self, name):
        self.name = name


# Порядок колонок ``users_report_query`` — по нему отчёт находит «Город».
REPORT_COLUMN_NAMES = (
    "name", "login", "role", "department_name", "direction", "supervisor",
    "status", "rate", "hire_date", "phone", "email", "personal_email",
    "instagram", "telegram_nick", "study_place", "study_course", "company_name",
    "employment_type", "has_proxy", "proxy_card_number", "proxy_status",
    "has_driver_license", "sip_number",
    "close_contact_1_relation", "close_contact_1_full_name", "close_contact_1_phone",
    "close_contact_2_relation", "close_contact_2_full_name", "close_contact_2_phone",
    "card_number", "city", "internship_in_company", "front_office_training",
    "front_office_training_date", "taxipro_id", "supervisor_id", "gender",
    "dismissal_start_date", "dismissal_end_date", "dismissal_reason",
    "dismissal_comment", "dismissal_is_blacklist", "status_fired_changed_at",
)


class _FakeCursor:
    def __init__(self, rows=()):
        self.rows = list(rows)
        self.executions = []
        self.description = tuple(_FakeColumn(name) for name in REPORT_COLUMN_NAMES)

    def execute(self, query, params=None):
        self.executions.append((str(query), list(params or [])))

    def fetchall(self):
        return list(self.rows)


class _FakeDatabase:
    generate_users_report = _load_users_report_builder()

    def __init__(self, rows=()):
        self.cursor = _FakeCursor(rows)

    def _get_cursor(self):
        return _CursorContext(self.cursor)


def _operator_row(
    *,
    name="Тестбаев Алан Тестович",
    hire_date=date(2024, 11, 7),
    dismissal_start=date(2026, 2, 25),
    dismissal_end=date(2026, 7, 12),
    city=None,
):
    """Return one row in the shape selected by ``generate_users_report``."""
    row = [None] * len(REPORT_COLUMN_NAMES)

    def _set(column, value):
        row[REPORT_COLUMN_NAMES.index(column)] = value

    _set("name", name)
    _set("login", "alan.erlanov")
    _set("role", "operator")
    _set("department_name", "СЗоВ")
    _set("direction", "Основное")
    _set("supervisor", "Тестовый СВ")
    _set("status", "working")
    _set("rate", 1.0)
    _set("hire_date", hire_date)
    _set("has_proxy", False)
    _set("has_driver_license", False)
    _set("city", city)
    _set("internship_in_company", False)
    _set("front_office_training", False)
    _set("supervisor_id", 126)
    _set("gender", "male")
    _set("dismissal_start_date", dismissal_start)
    _set("dismissal_end_date", dismissal_end)
    _set("dismissal_reason", "Нарушение дисциплины")
    _set("dismissal_comment", "Тестовый комментарий")
    _set("dismissal_is_blacklist", False)
    _set("status_fired_changed_at", datetime(2026, 3, 5, 12, 20))
    return tuple(row)


def _normalized_sql(query):
    return re.sub(r"\s+", " ", str(query)).strip().lower()


class UsersReportPeriodQueryTests(unittest.TestCase):
    def _build(self, rows=(), **overrides):
        database = _FakeDatabase(rows)
        options = {
            "period_month": "2026-07",
            "sheet_mode": "summary",
        }
        options.update(overrides)
        result = database.generate_users_report(**options)
        self.assertIsNotNone(result)
        self.assertEqual(len(database.cursor.executions), 1)
        query, params = database.cursor.executions[0]
        return result, _normalized_sql(query), params

    def test_month_filter_keeps_old_boundaries_and_allows_a_closed_dismissal(self):
        """A closed dismissal must not hide an operator who returned mid-month."""
        _result, sql, _params = self._build()
        period_filter = sql.split("where u.role = 'operator'", 1)[1]

        # Preserve the original selection boundaries: hired by month end and
        # either no dismissal or a dismissal that started during the month.
        self.assertRegex(period_filter, r"u\.hire_date\s*<=\s*%s(?:::\s*date)?")
        self.assertRegex(
            period_filter,
            r"dismissal_info\.dismissal_start_date\s+is\s+null",
        )
        self.assertRegex(
            period_filter,
            r"dismissal_info\.dismissal_start_date\s*>=\s*%s(?:::\s*date)?",
        )

        # Regression for Тестбаев Алан: his dismissal ended on July 12, so he
        # returned before the July 31 period end and must remain in the export.
        self.assertRegex(
            period_filter,
            r"dismissal_info\.dismissal_end_date\s+is\s+not\s+null\s+and\s+"
            r"dismissal_info\.dismissal_end_date\s*<\s*%s(?:::\s*date)?",
        )

    def test_month_window_and_scope_are_bound_parameters(self):
        _result, sql, params = self._build(
            department_ids=[1],
            supervisor_ids=[126],
        )

        self.assertNotIn("date '2026-07-01'", sql)
        self.assertNotIn("date '2026-07-31'", sql)
        self.assertEqual(
            params,
            [
                [1],
                [126],
                date(2026, 7, 31),
                date(2026, 7, 1),
                date(2026, 7, 31),
            ],
        )

    def test_returned_operator_can_be_written_to_period_workbook(self):
        (filename, content), _sql, _params = self._build([_operator_row()])

        self.assertEqual(filename, "users_report_2026-07.xlsx")
        self.assertTrue(content.startswith(b"PK"))
        workbook = load_workbook(BytesIO(content), data_only=True)
        self.assertEqual(workbook.sheetnames, ["Summary"])
        self.assertEqual(workbook["Summary"]["A2"].value, "Тестбаев Алан Тестович")
        self.assertEqual(workbook["Summary"]["J2"].value, "2024-11-07")
        # Ни у кого нет города — колонку не добавляем, чтобы не плодить пустую.
        self.assertNotIn("Город", [cell.value for cell in workbook["Summary"][1]])

    def test_city_column_appears_only_when_someone_has_a_city(self):
        (_filename, content), _sql, _params = self._build([_operator_row(city="Актобе")])

        workbook = load_workbook(BytesIO(content), data_only=True)
        headers = [cell.value for cell in workbook["Summary"][1]]
        self.assertEqual(headers.index("Город"), headers.index("Отдел") + 1)
        self.assertEqual(
            workbook["Summary"].cell(2, headers.index("Город") + 1).value,
            "Актобе",
        )


if __name__ == "__main__":
    unittest.main()
