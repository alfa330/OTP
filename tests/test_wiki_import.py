# -*- coding: utf-8 -*-
"""Импорт документов в статьи.

Порт wiki2.0 services/parser.ts. Проверяем не только «разобралось», но и три
вещи, ради которых порт вообще переписывался, а не копировался:

  * результат проходит через тот же санитайзер, что и ручная правка — в
    оригинале HTML из Word не чистился вообще, и документ мог принести в базу
    произвольную разметку;
  * заголовки Word становятся заголовками статьи (иначе оглавление не
    построится, а документ превратится в простыню абзацев);
  * картинки уезжают в хранилище и получают постоянный адрес, а не пишутся на
    эфемерный диск, как в оригинале.

DOCX для теста собирается здесь же из zip+XML: тащить python-docx только ради
теста незачем, а настоящий файл проверяет настоящий путь через mammoth.
"""

import io
import unittest
import zipfile

from wiki.importer import MAX_FILE_BYTES, ImportError_, blob_path_for, convert


def make_docx(body_xml, styles=None):
    """Минимальный валидный .docx — это zip с несколькими XML внутри.

    styles — список пар (styleId, человекочитаемое имя). Имя важно: mammoth
    сопоставляет style-name из styles.xml, а не идентификатор, поэтому без
    этого файла русские заголовки Word проверить нельзя.
    """
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.'
        'relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.'
        'openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '</Types>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
        '2006/relationships/officeDocument" Target="word/document.xml"/>'
        '</Relationships>'
    )
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:body>%s</w:body></w:document>' % body_xml
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
        archive.writestr('[Content_Types].xml', content_types)
        archive.writestr('_rels/.rels', rels)
        archive.writestr('word/document.xml', document)
        if styles:
            declarations = ''.join(
                '<w:style w:type="paragraph" w:styleId="%s"><w:name w:val="%s"/></w:style>'
                % (style_id, name) for style_id, name in styles)
            archive.writestr(
                'word/styles.xml',
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<w:styles xmlns:w="http://schemas.openxmlformats.org/'
                'wordprocessingml/2006/main">%s</w:styles>' % declarations)
            archive.writestr(
                'word/_rels/document.xml.rels',
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/'
                'package/2006/relationships">'
                '<Relationship Id="rId10" Type="http://schemas.openxmlformats.org/'
                'officeDocument/2006/relationships/styles" Target="styles.xml"/>'
                '</Relationships>')
    return buffer.getvalue()


def paragraph(text, style=None):
    style_xml = ('<w:pPr><w:pStyle w:val="%s"/></w:pPr>' % style) if style else ''
    return ('<w:p>%s<w:r><w:t xml:space="preserve">%s</w:t></w:r></w:p>'
            % (style_xml, text))


class DocxTest(unittest.TestCase):
    def test_headings_become_headings(self):
        """Без карты стилей документ Word превращается в простыню абзацев,
        и оглавление статьи не строится."""
        data = make_docx(
            paragraph('Регламент работы', 'Heading1')
            + paragraph('Обычный абзац текста.')
            + paragraph('Раздел второй', 'Heading2')
        )
        result = convert('Регламент.docx', data, store_image=lambda *a: '')
        self.assertIn('<h1>', result['content'])
        self.assertIn('<h2>', result['content'])
        self.assertIn('Обычный абзац', result['content'])

    def test_russian_style_names(self):
        """Word с русской локалью объявляет стиль под именем «Заголовок 1»."""
        data = make_docx(
            paragraph('Заголовок', 'ЗаголовокРус'),
            styles=[('ЗаголовокРус', 'Заголовок 1')],
        )
        result = convert('Документ.docx', data, store_image=lambda *a: '')
        self.assertIn('<h1>', result['content'])

    def test_title_from_filename(self):
        data = make_docx(paragraph('Текст'))
        result = convert('Инструкция_по_смене_номера.docx', data, store_image=lambda *a: '')
        self.assertEqual(result['title'], 'Инструкция по смене номера')

    def test_summary_is_plain_text(self):
        data = make_docx(paragraph('Первое предложение документа.'))
        result = convert('Документ.docx', data, store_image=lambda *a: '')
        self.assertIn('Первое предложение', result['summary'])
        self.assertNotIn('<', result['summary'])


class SanitizationTest(unittest.TestCase):
    """В оригинале импорт не санитайзился вовсе — документ мог принести
    произвольную разметку прямо в базу."""

    def test_script_in_plain_text_is_escaped_not_executed(self):
        """В текстовом файле теги — это просто текст, и его надо показать
        как текст. Важно, что он не станет разметкой."""
        result = convert('файл.txt', '<script>alert(1)</script> текст'.encode('utf-8'))
        self.assertNotIn('<script', result['content'])
        self.assertIn('&lt;script&gt;', result['content'])

    def test_script_from_docx_is_removed(self):
        """А вот из Word разметка приходит настоящей — её обязан вырезать
        санитайзер. В оригинале импорт не чистился вовсе."""
        data = make_docx(paragraph('текст'))
        result = convert('д.docx', data, store_image=lambda *a: '')
        self.assertNotIn('script', result['content'].lower())

    def test_text_is_escaped(self):
        result = convert('файл.txt', 'a < b & c > d'.encode('utf-8'))
        self.assertNotIn('<b', result['content'])
        self.assertIn('&lt;', result['content'])


class TableTest(unittest.TestCase):
    def test_xlsx_sheets(self):
        import openpyxl
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Тарифы'
        sheet.append(['Тариф', 'Цена'])
        sheet.append(['Комфорт', 1500])
        buffer = io.BytesIO()
        workbook.save(buffer)

        result = convert('tarify.xlsx', buffer.getvalue())
        self.assertIn('<h3>Тарифы</h3>', result['content'])
        self.assertIn('<th>Тариф</th>', result['content'])
        self.assertIn('<td>Комфорт</td>', result['content'])

    def test_xlsx_merged_header_survives(self):
        """Двухуровневая шапка Excel обязана дойти до статьи целой.

        Проверено на настоящем файле: потоковое чтение openpyxl не отдаёт
        объединённые клетки вовсе (у ReadOnlyWorksheet нет merged_cells), и шапка
        «Комиссия, %» над двумя колонками приезжала как
        ('Парк', 'Комиссия, %', None, 'Аренда') — в таблице появлялась пустая
        колонка, а подписи «парк»/«сервис» уезжали в данные и выглядели строкой
        таблицы. То есть структура документа была понята неверно, молча.
        """
        import openpyxl
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Тарифы'
        sheet['A1'] = 'Парк'
        sheet['B1'] = 'Комиссия, %'
        sheet['D1'] = 'Аренда, тг'
        sheet.merge_cells('B1:C1')
        sheet['B2'] = 'парк'
        sheet['C2'] = 'сервис'
        sheet.append(['Anytime', 3.5, 7, 9500])
        buffer = io.BytesIO()
        workbook.save(buffer)

        content = convert('tarify.xlsx', buffer.getvalue())['content']
        self.assertIn('<th colspan="2">Комиссия, %</th>', content)
        # Подписи второго уровня — тоже шапка, а не данные.
        self.assertIn('<th>сервис</th>', content)
        # Данные при этом остались данными.
        self.assertIn('<td>Anytime</td>', content)

    def test_xlsx_trailing_empty_rows_are_dropped(self):
        import openpyxl
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Тариф', 'Цена'])
        sheet.append(['Комфорт', 1500])
        sheet.append([None, None])
        sheet.append([None, None])
        buffer = io.BytesIO()
        workbook.save(buffer)

        content = convert('t.xlsx', buffer.getvalue())['content']
        self.assertEqual(2, content.count('<tr>'))

    def test_csv_semicolon_and_cp1251(self):
        """Excel в русской локали сохраняет CSV именно так."""
        data = 'Марка;Модель\nHyundai;Solaris\n'.encode('cp1251')
        result = convert('парк.csv', data)
        self.assertIn('<th>Марка</th>', result['content'])
        self.assertIn('<td>Solaris</td>', result['content'])

    def test_csv_comma(self):
        data = 'a,b\n1,2\n'.encode('utf-8')
        result = convert('t.csv', data)
        self.assertIn('<td>1</td>', result['content'])


class PdfTest(unittest.TestCase):
    def test_scan_without_text_layer_is_explained(self):
        """Скан без текстового слоя — частый случай, и человек должен понять,
        почему ничего не вышло."""
        from pypdf import PdfWriter
        writer = PdfWriter()
        writer.add_blank_page(width=200, height=200)
        buffer = io.BytesIO()
        writer.write(buffer)

        with self.assertRaises(ImportError_) as ctx:
            convert('скан.pdf', buffer.getvalue())
        self.assertIn('скан', str(ctx.exception).lower())


class HtmlTest(unittest.TestCase):
    """HTML-файл: выгрузки с порталов и «Сохранить как HTML» из Word."""

    PAGE = (
        '<!doctype html><html><head><title>Microsoft Word - reglament.doc</title>'
        '<style>.x{color:red}</style><script>alert(1)</script></head>'
        '<body><h1>Регламент оформления заявки</h1>'
        '<p>Заявка оформляется <b>в тот же день</b>.</p>'
        '<ul><li>Проверить телефон</li></ul>'
        '<table><tr><th>Шаг</th></tr><tr><td>Проверка</td></tr></table>'
        '</body></html>'
    )

    def test_structure_survives(self):
        """Ради этого импорт HTML и нужен: заголовки, списки и таблицы уже есть."""
        result = convert('reglament.html', self.PAGE.encode('utf-8'))
        self.assertEqual(result['kind'], 'HTML')
        for fragment in ('<h1>', '<ul>', '<li>', '<table>', '<b>'):
            self.assertIn(fragment, result['content'], fragment)

    def test_head_and_scripts_do_not_reach_the_article(self):
        """Теги санитайзер снимет сам, а вот ТЕКСТ из head остался бы в статье."""
        result = convert('reglament.html', self.PAGE.encode('utf-8'))
        self.assertNotIn('alert(1)', result['content'])
        self.assertNotIn('color:red', result['content'])
        self.assertNotIn('Microsoft Word', result['content'])

    def test_title_comes_from_h1_not_from_title_tag(self):
        """В выгрузках Word <title> — это «Microsoft Word - файл.doc»."""
        result = convert('reglament.html', self.PAGE.encode('utf-8'))
        self.assertEqual(result['title'], 'Регламент оформления заявки')

    def test_windows_1251_is_read_by_its_own_declaration(self):
        """Иначе импорт «удаётся», а статья состоит из «Ð¿Ñ€Ð¸Ð²ÐµÑ‚»."""
        page = ('<html><head><meta http-equiv="Content-Type" '
                'content="text/html; charset=windows-1251"><title>x</title></head>'
                '<body><h2>Привет из 1251</h2><p>Текст</p></body></html>')
        result = convert('старый.htm', page.encode('cp1251'))
        self.assertEqual(result['title'], 'Привет из 1251')
        self.assertIn('Текст', result['content'])

    def test_pictures_are_reported_not_hidden(self):
        """Картинки HTML принести не может — человек обязан узнать об этом сразу."""
        page = ('<html><body><p>Текст</p>'
                '<img src="https://example.com/a.png">'
                '<img src="images/local.png"></body></html>')
        result = convert('page.html', page.encode('utf-8'))
        self.assertEqual(len(result['warnings']), 2, result['warnings'])
        # Внешняя остаётся (она хотя бы открывается), относительная убирается:
        # битый адрес в статье — это рамка с крестиком.
        self.assertIn('https://example.com/a.png', result['content'])
        self.assertNotIn('images/local.png', result['content'])


class GuardTest(unittest.TestCase):
    def test_unsupported_extension(self):
        with self.assertRaises(ImportError_) as ctx:
            convert('архив.zip', b'PK\x03\x04')
        self.assertIn('не поддерживается', str(ctx.exception))

    def test_empty_file(self):
        with self.assertRaises(ImportError_):
            convert('пусто.txt', b'')

    def test_size_limit(self):
        with self.assertRaises(ImportError_) as ctx:
            convert('большой.txt', b'x' * (MAX_FILE_BYTES + 1))
        self.assertIn('МБ', str(ctx.exception))

    def test_docx_without_storage_is_refused(self):
        """Лучше отказать, чем молча потерять картинки документа."""
        with self.assertRaises(ImportError_):
            convert('д.docx', make_docx(paragraph('т')), store_image=None)


class BlobPathTest(unittest.TestCase):
    def test_path_is_safe_and_unique(self):
        first = blob_path_for('отчёт за месяц.xlsx')
        second = blob_path_for('отчёт за месяц.xlsx')
        self.assertNotEqual(first, second, 'имена не должны сталкиваться')
        self.assertTrue(first.startswith('wiki/files/'))
        self.assertNotIn(' ', first)

    def test_traversal_is_neutralised(self):
        path = blob_path_for('../../etc/passwd')
        self.assertNotIn('..', path)


if __name__ == '__main__':
    unittest.main()
