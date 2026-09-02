# -*- coding: utf-8 -*-
"""Контракт Excel-выгрузки детализации лидов TEZ ОП."""

import ast
import copy
import io
import os
import re
import sys
import unittest
from datetime import date as dt_date, datetime, timedelta
from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tests import source_cache


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
BOT_PATH = ROOT / "bot_schedule2.py"
BOT_SOURCE = BOT_PATH.read_text(encoding="utf-8-sig")
BOT_TREE = source_cache.parse(BOT_SOURCE)
TOP_LEVEL_FUNCTIONS = {
    node.name: node
    for node in BOT_TREE.body
    if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
}
DATABASE_PATH = ROOT / "database.py"
DATABASE_SOURCE = DATABASE_PATH.read_text(encoding="utf-8-sig")
DATABASE_TREE = source_cache.parse(DATABASE_SOURCE)


def _load_export_function(namespace):
    """Загружает route и его локальные helper-функции без импорта bot_schedule2."""
    pending = ["tez_leads_export"]
    selected = {}
    while pending:
        name = pending.pop()
        if name in selected or name in namespace:
            continue
        node = TOP_LEVEL_FUNCTIONS.get(name)
        if node is None:
            continue
        selected[name] = node
        for child in ast.walk(node):
            if isinstance(child, ast.Call) and isinstance(child.func, ast.Name):
                called_name = child.func.id
                if called_name in TOP_LEVEL_FUNCTIONS:
                    pending.append(called_name)

    nodes = []
    for original in BOT_TREE.body:
        if (
            isinstance(original, (ast.FunctionDef, ast.AsyncFunctionDef))
            and original.name in selected
        ):
            cloned = copy.deepcopy(original)
            cloned.decorator_list = []
            nodes.append(cloned)
    module = ast.Module(body=nodes, type_ignores=[])
    ast.fix_missing_locations(module)
    result = dict(namespace)
    exec(compile(module, str(BOT_PATH), "exec"), result)
    return result["tez_leads_export"]


def _load_top_level_function(name, namespace=None):
    original = TOP_LEVEL_FUNCTIONS.get(name)
    if original is None:
        raise AssertionError(f"Missing function in bot_schedule2.py: {name}")
    cloned = copy.deepcopy(original)
    cloned.decorator_list = []
    module = ast.Module(body=[cloned], type_ignores=[])
    ast.fix_missing_locations(module)
    result = dict(namespace or {})
    exec(compile(module, str(BOT_PATH), "exec"), result)
    return result[name]


def _load_database_method(name):
    database_class = next(
        node
        for node in DATABASE_TREE.body
        if isinstance(node, ast.ClassDef) and node.name == "Database"
    )
    original = next(
        node
        for node in database_class.body
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
        and node.name == name
    )
    cloned = copy.deepcopy(original)
    cloned.decorator_list = []
    module = ast.Module(body=[cloned], type_ignores=[])
    ast.fix_missing_locations(module)
    namespace = {}
    exec(compile(module, str(DATABASE_PATH), "exec"), namespace)
    return namespace[name], ast.get_source_segment(DATABASE_SOURCE, original)


class _FakeDb:
    """Выгрузка тянет отчёт периода ОДНИМ запросом: своя база + перенос.

    Разделение на два периода жило в самой выгрузке и давало два листа; теперь
    периметр собирает get_tez_leads_report, а строка сама говорит, своя она или
    перенесённая (is_carried / base_month).
    """

    def __init__(self, rows):
        self.rows = rows
        self.calls = []

    def get_tez_leads_report(self, year, month, **kwargs):
        self.calls.append((year, month, kwargs))
        return self.rows


class _QuietLogging:
    @staticmethod
    def exception(*_args, **_kwargs):
        return None


class _FakeCursor:
    def __init__(self, rows):
        self.rows = rows
        self.sql = None
        self.params = None

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def execute(self, sql, params):
        self.sql = sql
        self.params = params

    def fetchall(self):
        return self.rows


class _FakeDatabase:
    def __init__(self, row):
        self.cursor = _FakeCursor([row])

    @staticmethod
    def _tez_leads_detail_filters(_status=None, _operator_id=None, _search=None):
        return "", []

    def _get_cursor(self):
        return self.cursor


class TezLeadsExcelExportTests(unittest.TestCase):
    def _export_bytes(self, rows):
        db = _FakeDb(rows)

        def send_file(stream, **kwargs):
            return {
                "content": stream.getvalue(),
                "kwargs": kwargs,
            }

        export = _load_export_function(
            {
                "_tez_leads_require_manager": lambda: (17, None),
                "_tez_leads_period_from_request": lambda: (2026, 7),
                "db": db,
                "jsonify": lambda value: value,
                "logging": _QuietLogging,
                "Workbook": Workbook,
                "Font": Font,
                "PatternFill": PatternFill,
                "Alignment": Alignment,
                "BytesIO": io.BytesIO,
                "datetime": datetime,
                "dt_date": dt_date,
                "timedelta": timedelta,
                "send_file": send_file,
                "get_column_letter": get_column_letter,
                "ZipFile": ZipFile,
                "ZIP_DEFLATED": ZIP_DEFLATED,
            }
        )
        response = export()
        return db, response["content"]

    def _export(self, rows):
        db, content = self._export_bytes(rows)
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        return db, workbook.active

    def _export_book(self, rows):
        db, content = self._export_bytes(rows)
        return db, load_workbook(io.BytesIO(content), data_only=True)

    @staticmethod
    def _detail_row(**overrides):
        row = {
            "lead_id": "8d0c0346-9563-46c3-b757-696414b19bf2",
            "phone": "77010000001",
            "full_name": "Тестовый водитель",
            "status": "success",
            "status_rule": "same_month",
            "upload_count": 1,
            "first_order_at": "2026-07-20T12:00:00",
            "operator_id": 17,
            "operator_name": "Оператор ТЕЗ",
            "call_at": "2026-07-18T10:15:00",
            "success_date": "2026-07-20",
            "rule_code": "same_month",
            "month_first_order_at": "2026-07-20T12:00:00",
            "prev_month_first_order_at": None,
            "source_file_name": "Лиды Алматы июль.xlsx",
            "talk_duration_seconds": 3723,
            "is_carried": False,
            "base_year": 2026,
            "base_month": 7,
        }
        row.update(overrides)
        return row

    @classmethod
    def _carried_row(cls, **overrides):
        """Строка переноса: лид ИЮНЬСКОЙ базы в отчёте за июль.

        Все её колонки метод отчёта уже пересчитал на июль, поэтому фикстура
        задаёт июльские даты и июльский статус — ровно то, что видит выгрузка.
        """
        row = cls._detail_row(
            phone="77010000009",
            full_name="Перенесённый водитель",
            source_file_name="Лиды Алматы июнь.xlsx",
            is_carried=True,
            base_month=6,
        )
        row.update(overrides)
        return row

    def test_export_contains_source_and_native_talk_duration_from_detail_payload(self):
        db, sheet = self._export([self._detail_row()])

        self.assertEqual(
            db.calls,
            [(2026, 7, {"limit": 50000})],
            "Лист один, и собирается он одним запросом отчёта за период",
        )
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Источник", headers)
        self.assertIn("Время разговора в отчётном месяце", headers)

        source_column = headers.index("Источник") + 1
        duration_column = headers.index("Время разговора в отчётном месяце") + 1
        self.assertEqual(sheet.cell(2, source_column).value, "Лиды Алматы июль.xlsx")
        # Длительность — число секунд, а не формат времени: отчёт считают и
        # фильтруют в секундах, в них же задан порог успешки.
        self.assertEqual(sheet.cell(2, duration_column).value, 3723)
        self.assertEqual(sheet.cell(2, duration_column).number_format, "0")

    # ── Один лист: своя база месяца плюс перенос с прошлого ─────────────────

    def test_report_is_a_single_sheet_marked_by_base_month(self):
        """Владелец: лиды двух месяцев — на ОДНОМ листе, и по строке должно быть
        видно, из чьей базы она пришла."""
        db, book = self._export_book([self._detail_row(), self._carried_row()])

        self.assertEqual(book.sheetnames, ["Лиды 07.2026"])
        sheet = book.active
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers[0], "Месяц базы")
        self.assertEqual(sheet.cell(2, 1).value, "Июль 2026")
        # У переноса месяц базы — прошлый, и это сказано словом: читателю не
        # приходится помнить, что июнь на июльском листе и означает перенос.
        self.assertEqual(sheet.cell(3, 1).value, "Июнь 2026 (перенос)")
        self.assertEqual(sheet.cell(3, headers.index("Телефон") + 1).value, "77010000009")

    def test_every_success_of_the_period_is_on_the_sheet(self):
        """Тот самый расчёт, из-за которого файл расходился с разделом: успешки
        периода = успешки своей базы ПЛЮС успешки переноса. За август в разделе
        стояло 245, в файле — 134, потому что 111 переносов лежали отдельно."""
        own = [self._detail_row(phone=f"7701000{idx:04d}") for idx in range(3)]
        carried = [
            self._carried_row(phone=f"7702000{idx:04d}", success_date="2026-07-05")
            for idx in range(2)
        ]
        _, sheet = self._export(own + carried)
        headers = [cell.value for cell in sheet[1]]
        status_column = headers.index("Статус") + 1

        statuses = [sheet.cell(row, status_column).value for row in range(2, 7)]
        self.assertEqual(statuses, ["Успешка"] * 5)
        self.assertEqual(sheet.max_row, 6, "лист один, строки обеих баз в нём")

    def test_carried_lead_shows_its_reporting_month_status(self):
        """Главное: у перенесённой строки статус и правило — ОТЧЁТНОГО месяца.

        Раньше здесь стоял статус прошлого месяца («В работе»), и понять, стал
        лид успешкой в этом месяце или нет, было нельзя.
        """
        carried = self._carried_row(
            status="success",
            status_rule="reactivated_30d",
            success_date="2026-07-14",
            month_first_order_at="2026-07-14T09:00:00",
            first_order_at="2026-07-14T09:00:00",
            prev_month_first_order_at=None,
        )
        _, sheet = self._export([carried])
        headers = [cell.value for cell in sheet[1]]

        self.assertEqual(sheet.cell(2, headers.index("Статус") + 1).value, "Успешка")
        self.assertEqual(
            sheet.cell(2, headers.index("Правило") + 1).value,
            "Не работал 30+ дней и вернулся после звонка",
        )
        self.assertEqual(sheet.cell(2, headers.index("Дата успешки") + 1).value, "2026-07-14")
        self.assertEqual(
            sheet.cell(2, headers.index("Поездка в отчётном месяце") + 1).value,
            "2026-07-14 09:00:00",
        )

    def test_carried_lead_without_success_shows_why_it_was_not_counted(self):
        """Перенос, который в отчётном месяце успешкой не стал: видно и статус,
        и причину — иначе строка ничем не отличается от «просто базы»."""
        carried = self._carried_row(
            status="already_working",
            status_rule="gap_under_30d",
            success_date=None,
            rule_code=None,
            month_first_order_at="2026-07-06T08:00:00",
            first_order_at="2026-07-06T08:00:00",
            prev_month_first_order_at="2026-06-28T19:00:00",
        )
        _, sheet = self._export([carried])
        headers = [cell.value for cell in sheet[1]]

        self.assertEqual(
            sheet.cell(2, headers.index("Статус") + 1).value, "Уже работающий"
        )
        self.assertIn(
            "30 дней", sheet.cell(2, headers.index("Правило") + 1).value
        )
        self.assertFalse(sheet.cell(2, headers.index("Дата успешки") + 1).value)
        # Поездка месяца базы у переноса становится «поездкой в прошлом месяце»:
        # месяц базы для отчётного месяца и есть прошлый.
        self.assertEqual(
            sheet.cell(2, headers.index("Поездка в прошлом месяце") + 1).value,
            "2026-06-28 19:00:00",
        )

    def test_missing_carry_status_is_named_rather_than_left_blank(self):
        """Статуса доработки может не быть (пересчёт по периоду после появления
        колонки не проходил). Пустая ячейка читалась бы как потерянный статус,
        поэтому у состояния есть имя, и оно не врёт про «уже пересчитали»."""
        _, sheet = self._export([self._carried_row(status="carry_pending",
                                                   status_rule=None,
                                                   success_date=None)])
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            sheet.cell(2, headers.index("Статус") + 1).value, "Статус не рассчитан"
        )

    def test_status_success_without_a_success_row_is_not_counted_as_one(self):
        """Обратная страховка того же расхождения: успешек в файле не должно
        стать БОЛЬШЕ, чем в разделе. Успешкой строку делает запись об успешке
        за период, а не залипший статус."""
        _, sheet = self._export([self._carried_row(status="stale_success",
                                                   status_rule="reactivated_30d",
                                                   success_date=None)])
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            sheet.cell(2, headers.index("Статус") + 1).value, "Требует сверки"
        )
        self.assertNotEqual(sheet.cell(2, headers.index("Статус") + 1).value, "Успешка")

    def test_carried_success_shows_its_call_like_an_ordinary_one(self):
        """Перенесённая успешка: лид июня, поездка и звонок — в июле. Звонок
        ложится в те же колонки, что и у обычной июльской успешки, потому что
        лист один и месяц раскладки у всех строк отчётный."""
        carried = self._carried_row(
            success_date="2026-07-14",
            call_at="2026-07-09T11:20:00",
            talk_duration_seconds=95,
        )
        _, sheet = self._export([carried])
        col = self._call_columns(sheet)

        self.assertEqual(sheet.cell(2, col["month_call"]).value, "2026-07-09 11:20:00")
        self.assertEqual(sheet.cell(2, col["month_duration"]).value, 95)
        self.assertFalse(sheet.cell(2, col["prev_call"]).value)
        self.assertFalse(sheet.cell(2, col["prev_duration"]).value)

    def test_carried_success_on_the_month_seam_fills_the_previous_pair(self):
        """Звонок 25 июня при поездке 3 июля — окно «последние 7 дней прошлого
        месяца» относительно ОТЧЁТНОГО месяца, для переноса тоже."""
        carried = self._carried_row(
            success_date="2026-07-03",
            status_rule="prev_month_last7",
            rule_code="prev_month_last7",
            call_at="2026-06-25T10:00:00",
            talk_duration_seconds=61,
        )
        _, sheet = self._export([carried])
        col = self._call_columns(sheet)

        self.assertEqual(sheet.cell(2, col["prev_call"]).value, "2026-06-25 10:00:00")
        self.assertEqual(sheet.cell(2, col["prev_duration"]).value, 61)
        self.assertFalse(sheet.cell(2, col["month_call"]).value)

    def test_missing_talk_duration_is_an_empty_excel_cell(self):
        _, sheet = self._export(
            [self._detail_row(source_file_name="", talk_duration_seconds=None)]
        )
        headers = [cell.value for cell in sheet[1]]
        source_column = headers.index("Источник") + 1
        duration_column = headers.index("Время разговора в отчётном месяце") + 1

        self.assertIsNone(sheet.cell(2, duration_column).value)
        self.assertIn(sheet.cell(2, source_column).value, (None, ""))

    def test_user_controlled_text_is_not_exported_as_an_excel_formula(self):
        _, sheet = self._export(
            [
                self._detail_row(
                    full_name="=1+1",
                    source_file_name='=HYPERLINK("https://example.invalid")',
                )
            ]
        )
        headers = [cell.value for cell in sheet[1]]
        source_column = headers.index("Источник") + 1
        name_column = headers.index("ФИО") + 1

        self.assertEqual(sheet.cell(2, name_column).value, "=1+1")
        self.assertEqual(sheet.cell(2, name_column).data_type, "s")
        self.assertEqual(
            sheet.cell(2, source_column).value,
            '=HYPERLINK("https://example.invalid")',
        )
        self.assertEqual(sheet.cell(2, source_column).data_type, "s")

    def test_phone_column_is_text_formatted(self):
        """«Телефон» должен быть текстом и по типу, и по формату ячейки.

        С форматом «Общий» Excel при первой же правке превращает 77012345678 в
        число (длинные номера — в экспоненту), и номер перестаёт совпадать при
        поиске и сверке с базой.
        """
        _, sheet = self._export([self._detail_row()])
        headers = [cell.value for cell in sheet[1]]
        phone_column = headers.index("Телефон") + 1
        cell = sheet.cell(2, phone_column)
        self.assertEqual(cell.data_type, "s")
        self.assertEqual(cell.number_format, "@")
        # Длительность разговора — по-прежнему число, а не текст.
        duration_column = headers.index("Время разговора в отчётном месяце") + 1
        self.assertEqual(sheet.cell(2, duration_column).number_format, "0")

    def test_phone_column_has_no_number_stored_as_text_warning(self):
        """Текстовый формат не должен тащить за собой значок ошибки Excel.

        Без <ignoredErrors> Excel рисует зелёный уголок «Число сохранено как
        текст» на каждой строке телефона — формат при этом правильный, но
        выгрузка выглядит битой.
        """
        _, content = self._export_bytes([self._detail_row(), self._detail_row()])
        with ZipFile(io.BytesIO(content)) as book:
            sheet_xml = book.read("xl/worksheets/sheet1.xml").decode("utf-8")

        self.assertIn('numberStoredAsText="1"', sheet_xml)
        # «Месяц базы» стоит первой колонкой, поэтому телефон — это C.
        self.assertIn('sqref="C2:C3"', sheet_xml)
        # Узел обязан стоять после pageMargins — иначе Excel считает книгу битой.
        self.assertLess(
            sheet_xml.index("<pageMargins"), sheet_xml.index("<ignoredErrors")
        )
        # Файл после правки zip остаётся читаемым книгой, а не только текстом.
        sheet = load_workbook(io.BytesIO(content), data_only=True).active
        self.assertEqual(sheet.cell(2, 3).value, "77010000001")

    def test_empty_export_is_not_patched(self):
        _, content = self._export_bytes([])
        with ZipFile(io.BytesIO(content)) as book:
            sheet_xml = book.read("xl/worksheets/sheet1.xml").decode("utf-8")
        self.assertNotIn("<ignoredErrors", sheet_xml)

    def _call_columns(self, sheet):
        headers = [cell.value for cell in sheet[1]]
        return {
            "prev_call": headers.index("Звонок в прошлом месяце") + 1,
            "prev_duration": headers.index("Время разговора в прошлом месяце") + 1,
            "month_call": headers.index("Звонок в отчётном месяце") + 1,
            "month_duration": headers.index("Время разговора в отчётном месяце") + 1,
        }

    def test_call_inside_reporting_month_fills_only_the_reporting_pair(self):
        _, sheet = self._export(
            [self._detail_row(call_at="2026-07-18T10:15:00", talk_duration_seconds=44)]
        )
        col = self._call_columns(sheet)

        self.assertEqual(sheet.cell(2, col["month_call"]).value, "2026-07-18 10:15:00")
        self.assertEqual(sheet.cell(2, col["month_duration"]).value, 44)
        self.assertIsNone(sheet.cell(2, col["prev_call"]).value)
        self.assertIsNone(sheet.cell(2, col["prev_duration"]).value)

    def test_call_in_last_seven_days_of_previous_month_fills_only_the_previous_pair(self):
        """Окно считается от конца месяца: в июне это 24–30, а не «после 24-го»."""
        _, sheet = self._export(
            [self._detail_row(call_at="2026-06-24T09:00:00", talk_duration_seconds=61)]
        )
        col = self._call_columns(sheet)

        self.assertEqual(sheet.cell(2, col["prev_call"]).value, "2026-06-24 09:00:00")
        self.assertEqual(sheet.cell(2, col["prev_duration"]).value, 61)
        self.assertIsNone(sheet.cell(2, col["month_call"]).value)
        self.assertIsNone(sheet.cell(2, col["month_duration"]).value)

    def test_call_before_the_window_lands_in_neither_pair(self):
        """23 июня — на день раньше окна, звонок успешку дать не может."""
        _, sheet = self._export(
            [
                self._detail_row(
                    call_at="2026-06-23T23:59:00",
                    status="not_counted",
                    status_rule="call_before_last7",
                    talk_duration_seconds=300,
                )
            ]
        )
        col = self._call_columns(sheet)

        for column in col.values():
            self.assertIsNone(sheet.cell(2, column).value)
        # Причина при этом остаётся видимой — иначе строка нечитаема.
        headers = [cell.value for cell in sheet[1]]
        rule_column = headers.index("Правило") + 1
        self.assertIn("раньше", sheet.cell(2, rule_column).value)

    def test_call_without_a_date_leaves_both_pairs_empty(self):
        _, sheet = self._export(
            [self._detail_row(call_at=None, talk_duration_seconds=None)]
        )
        col = self._call_columns(sheet)

        for column in col.values():
            self.assertIsNone(sheet.cell(2, column).value)

    def test_active_prev_month_exports_previous_reason_without_hiding_current_trip(self):
        _, sheet = self._export(
            [
                self._detail_row(
                    status="already_working",
                    status_rule="active_prev_month",
                    rule_code=None,
                    first_order_at="2026-07-03T22:47:00",
                    month_first_order_at="2026-07-03T22:47:00",
                    prev_month_first_order_at="2026-06-30T23:27:00",
                    call_at="2026-07-02T13:58:00",
                )
            ]
        )

        headers = [cell.value for cell in sheet[1]]
        previous_column = headers.index("Поездка в прошлом месяце") + 1
        current_column = headers.index("Поездка в отчётном месяце") + 1
        rule_column = headers.index("Правило") + 1

        self.assertEqual(sheet.cell(2, previous_column).value, "2026-06-30 23:27:00")
        self.assertEqual(sheet.cell(2, current_column).value, "2026-07-03 22:47:00")
        self.assertIn("прошлом месяце", sheet.cell(2, rule_column).value)


class TezLeadsCallBucketTests(unittest.TestCase):
    """Раскладка звонка по месяцам считается тем же окном, что и успешка."""

    def setUp(self):
        self.bucket = _load_top_level_function(
            "_tez_leads_call_bucket",
            namespace={"datetime": datetime, "dt_date": dt_date},
        )

    def test_reporting_month(self):
        self.assertEqual(self.bucket("2026-07-01T00:00:00", 2026, 7), "month")
        self.assertEqual(self.bucket("2026-07-31T23:59:00", 2026, 7), "month")

    def test_previous_month_window_is_counted_from_the_month_end(self):
        # Июнь (30 дней) -> окно 24–30, февраль-2026 (28 дней) -> 22–28.
        self.assertEqual(self.bucket("2026-06-24T00:00:00", 2026, 7), "prev")
        self.assertIsNone(self.bucket("2026-06-23T23:59:00", 2026, 7))
        self.assertEqual(self.bucket("2026-02-22T10:00:00", 2026, 3), "prev")
        self.assertIsNone(self.bucket("2026-02-21T10:00:00", 2026, 3))

    def test_january_looks_at_december_of_the_previous_year(self):
        self.assertEqual(self.bucket("2025-12-25T08:00:00", 2026, 1), "prev")
        self.assertIsNone(self.bucket("2025-12-24T08:00:00", 2026, 1))
        self.assertEqual(self.bucket("2026-01-05T08:00:00", 2026, 1), "month")

    def test_missing_or_broken_date_is_not_placed_anywhere(self):
        self.assertIsNone(self.bucket(None, 2026, 7))
        self.assertIsNone(self.bucket("", 2026, 7))
        self.assertIsNone(self.bucket("не дата", 2026, 7))

    def test_window_start_matches_the_success_rule(self):
        """Окно не должно разъехаться с расчётом успешки — источник правды один."""
        from tez_op_leads import call_window_for_period

        for year, month in ((2026, 1), (2026, 3), (2026, 7), (2026, 12)):
            start, _ = call_window_for_period(year, month)
            self.assertEqual(
                self.bucket(datetime.combine(start, datetime.min.time()).isoformat(),
                            year, month),
                "prev",
                f"первый день окна {start} обязан попасть в прошлый месяц",
            )
            before = start - timedelta(days=1)
            self.assertIsNone(
                self.bucket(datetime.combine(before, datetime.min.time()).isoformat(),
                            year, month),
                f"день до окна {before} не должен попадать никуда",
            )


class TezLeadsDisplayFilenameTests(unittest.TestCase):
    def test_unicode_basename_and_control_character_sanitizing(self):
        display_filename = _load_top_level_function(
            "_tez_leads_display_filename",
            namespace={"os": os, "re": re},
        )

        self.assertEqual(
            display_filename("Лиды Алматы июль.xlsx"),
            "Лиды Алматы июль.xlsx",
        )
        self.assertEqual(
            display_filename(r"C:\Users\operator\Downloads\Лиды Алматы.xlsx"),
            "Лиды Алматы.xlsx",
        )
        self.assertEqual(
            display_filename("/tmp/tez-imports/Лиды Алматы.csv"),
            "Лиды Алматы.csv",
        )
        self.assertEqual(
            display_filename("Лиды\x00 Алматы\r\nиюль\t.xlsx"),
            "Лиды Алматыиюль.xlsx",
        )


class TezLeadsDetailPayloadTests(unittest.TestCase):
    def test_detail_exposes_source_filename_and_binotel_talk_duration(self):
        detail_method, source = _load_database_method("get_tez_leads_detail")
        db = _FakeDatabase(
            (
                "8d0c0346-9563-46c3-b757-696414b19bf2",
                "77010000001",
                "Тестовый водитель",
                "success",
                "same_month",
                1,
                None,
                17,
                "Оператор ТЕЗ",
                None,
                None,
                "same_month",
                None,
                "Лиды Алматы июль.xlsx",
                187,
            )
        )

        result = detail_method(db, 2026, 7)

        self.assertEqual(result[0]["source_file_name"], "Лиды Алматы июль.xlsx")
        self.assertEqual(result[0]["talk_duration_seconds"], 187)
        self.assertIn("source_batch.file_name", db.cursor.sql)
        self.assertIn("success_call.billsec", db.cursor.sql)
        self.assertIn("ELSE lc.billsec", db.cursor.sql)
        self.assertIn(
            "success_call.general_call_id = s.call_general_id",
            db.cursor.sql,
        )
        self.assertIn("'source_file_name'", source)
        self.assertIn("'talk_duration_seconds'", source)

    def test_detail_keeps_previous_reason_and_current_trip_as_separate_fields(self):
        detail_method, source = _load_database_method("get_tez_leads_detail")
        current_trip = datetime(2026, 7, 3, 22, 47)
        previous_trip = datetime(2026, 6, 30, 23, 27)
        db = _FakeDatabase(
            (
                "8d0c0346-9563-46c3-b757-696414b19bf2",
                "77000000105",
                "Тестовый водитель",
                "already_working",
                "active_prev_month",
                1,
                current_trip,
                None,
                "",
                None,
                None,
                None,
                previous_trip,
                "Лиды ОП.xlsx",
                None,
            )
        )

        row = detail_method(db, 2026, 7)[0]

        # first_order_at остаётся для старых клиентов, но новое явное
        # поле не даёт спутать его с причиной active_prev_month.
        self.assertEqual(row["first_order_at"], "2026-07-03T22:47:00")
        self.assertEqual(row["month_first_order_at"], "2026-07-03T22:47:00")
        self.assertEqual(row["prev_month_first_order_at"], "2026-06-30T23:27:00")
        self.assertIn("'month_first_order_at'", source)


class TezLeadsReportPayloadTests(unittest.TestCase):
    """Контракт выборки отчёта: периметр периода и пересчёт колонок на него."""

    def setUp(self):
        self.method, self.source = _load_database_method("get_tez_leads_report")

    @staticmethod
    def _row(**overrides):
        values = {
            "lead_id": "8d0c0346-9563-46c3-b757-696414b19bf2",
            "phone_norm": "77010000001",
            "full_name": "Перенесённый водитель",
            "status": "success",
            "status_rule": "reactivated_30d",
            "upload_count": 2,
            "trip_at": datetime(2026, 8, 31, 21, 18, 56),
            "prev_trip_at": None,
            "operator_id": 17,
            "operator_name": "Оператор ТЕЗ",
            "call_at": datetime(2026, 7, 30, 15, 22, 49),
            "success_date": dt_date(2026, 8, 31),
            "rule_code": "reactivated_30d",
            "file_name": "Лиды Алматы июль.xlsx",
            "billsec": 39,
            "is_carried": True,
            "base_year": 2026,
            "base_month": 7,
        }
        values.update(overrides)
        return tuple(values.values())

    def test_carried_row_is_recomputed_onto_the_reporting_month(self):
        db = _FakeDatabase(self._row())
        row = self.method(db, 2026, 8)[0]

        self.assertTrue(row["is_carried"])
        self.assertEqual((row["base_year"], row["base_month"]), (2026, 7))
        # Статус, правило и дата успешки — августовские, хотя лид июльский.
        self.assertEqual(row["status"], "success")
        self.assertEqual(row["status_rule"], "reactivated_30d")
        self.assertEqual(row["success_date"], "2026-08-31")
        self.assertEqual(row["month_first_order_at"], "2026-08-31T21:18:56")
        self.assertEqual(row["talk_duration_seconds"], 39)
        self.assertEqual(row["upload_count"], 2)

    def test_period_perimeter_matches_the_funnel(self):
        """Строки отчёта = своя база месяца + перенос прошлого (status <> success),
        то есть ровно те лиды, на которых может лежать успешка периода."""
        db = _FakeDatabase(self._row())
        self.method(db, 2026, 8)
        sql = db.cursor.sql

        self.assertIn("WHERE l.year = %(year)s AND l.month = %(month)s", sql)
        self.assertIn("UNION ALL", sql)
        self.assertIn("l.year = %(prev_year)s AND l.month = %(prev_month)s", sql)
        self.assertIn("l.status <> 'success'", sql)
        self.assertEqual(
            db.cursor.params,
            {"year": 2026, "month": 8, "prev_year": 2026, "prev_month": 7,
             "limit": 50000},
        )

    def test_january_report_reaches_back_into_december(self):
        db = _FakeDatabase(self._row())
        self.method(db, 2026, 1)
        self.assertEqual(db.cursor.params["prev_year"], 2025)
        self.assertEqual(db.cursor.params["prev_month"], 12)

    def test_carried_columns_come_from_the_carry_pair(self):
        """Перенос читает carry_*, а не собственные колонки лида: иначе в отчёт
        уехали бы статус и поездка месяца базы."""
        db = _FakeDatabase(self._row())
        self.method(db, 2026, 8)
        query = db.cursor.sql

        self.assertIn("l.carry_status, l.carry_status_rule,", query)
        self.assertIn("l.carry_first_order_at,", query)
        # Отсечка «последний звонок до поездки» берётся из строки отчёта, то
        # есть у переноса — из его августовской поездки, а не из июльской.
        self.assertIn("(r.trip_at IS NULL OR c.started_at < r.trip_at)", query)

    def test_success_is_joined_only_for_the_reporting_period(self):
        """Успешка, забронированная на СЛЕДУЮЩИЙ месяц, в этот отчёт не входит —
        иначе счёт успешек разошёлся бы с разделом в другую сторону."""
        db = _FakeDatabase(self._row())
        self.method(db, 2026, 8)
        self.assertIn(
            "ON s.lead_id = r.id AND s.year = %(year)s AND s.month = %(month)s",
            db.cursor.sql,
        )

    def test_phone_present_in_both_bases_is_not_doubled_but_a_success_survives(self):
        db = _FakeDatabase(self._row())
        self.method(db, 2026, 8)
        sql = db.cursor.sql

        self.assertIn("FROM tez_leads own", sql)
        self.assertIn("own.phone_norm = l.phone_norm", sql)
        # Исключение: строку с успешкой отчётного месяца оставляем всегда.
        self.assertIn("FROM tez_lead_successes s", sql)
        self.assertIn("s.lead_id = l.id", sql)

    def test_own_base_goes_first_and_untouched_statuses_never_fall_out(self):
        db = _FakeDatabase(self._row())
        self.method(db, 2026, 8)
        self.assertIn("ORDER BY r.is_carried, r.trip_at DESC NULLS LAST", db.cursor.sql)
        # base_status у своей строки NOT NULL, у переноса до пересчёта NULL —
        # такое состояние называется своим именем, а не пустой ячейкой.
        self.assertIn("COALESCE(r.base_status, 'carry_pending')", db.cursor.sql)
        self.assertIn("'carry_pending'", self.source)

    def test_only_a_success_row_makes_a_row_a_success(self):
        """Равенство «успешек в файле = успешек в разделе» держится структурой
        запроса: статус 'success' без записи об успешке за период успешкой не
        называется."""
        db = _FakeDatabase(self._row())
        self.method(db, 2026, 8)
        self.assertIn("WHEN r.base_status = 'success' THEN 'stale_success'",
                      db.cursor.sql)


if __name__ == "__main__":
    unittest.main()
