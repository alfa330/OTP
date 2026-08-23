# -*- coding: utf-8 -*-
"""Статистика прохождений тренажёров: права, границы и выгрузка.

Что здесь важно проверить и почему.

1. ДВА РАЗНЫХ ГЕЙТА, И ПЕРЕПУТАТЬ ИХ ДОРОГО В ОБЕ СТОРОНЫ. Запись попытки
   открыта рядовому читателю — он и есть тот, кто проходит тренажёр; закрой её
   способностью, и прохождения молча пропадут по 403, а вкладка будет
   показывать нули. Статистику, наоборот, видит только редактор: там имена,
   отделы и число промахов каждого.

2. ГРАНИЦА ОТДЕЛА. Права редактора отделов не знают, и без отдельной границы
   супервайзер СЗоВ с правом публикации увидел бы поимённый состав отдела
   продаж. Проверяем, что в запрос уходит список отделов, а у супер-админа —
   NULL (без границы).

3. ЧУЖУЮ ПОПЫТКУ НЕ ЗАКРЫТЬ. id попытки приходит из браузера; в UPDATE обязано
   стоять условие по владельцу, иначе чужой урок можно пометить пройденным.

4. КНИГА СОБИРАЕТСЯ И НА ПУСТЫХ ДАННЫХ. Выгрузку просят в том числе у
   тренажёра, который ещё никто не проходил, и падать она там не должна.
"""

import json
import sys
import unittest
from contextlib import contextmanager
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import MagicMock

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from flask import Flask
except ImportError:  # pragma: no cover
    Flask = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

from io import BytesIO  # noqa: E402

from wiki import perimeter as wiki_perimeter  # noqa: E402
from wiki import queries  # noqa: E402
from wiki import trainer_report  # noqa: E402
from wiki import trainers as wiki_trainers  # noqa: E402
from wiki.access import collect_subjects  # noqa: E402
from wiki.routes import build_wiki_blueprint  # noqa: E402


def make_context(otp_role='operator', department_id=None, **caps):
    role = {'id': 5, 'code': 'wiki_role', 'can_read': True, 'can_create': False,
            'can_edit': False, 'can_delete': False, 'can_publish': False,
            'can_approve': False, 'can_manage_users': False,
            'can_manage_structure': False, 'can_manage_access': False}
    role.update(caps)
    return {
        'user_id': 42, 'otp_role': otp_role, 'department_id': department_id,
        'direction_id': None, 'headed_department_ids': [], 'group_ids': [],
        'wiki_roles': [role], 'access_mode': 'auto',
    }


class RecordingCursor(MagicMock):
    """Курсор, который помнит все запросы: по ним и проверяются границы."""


def build_client(test, context, *, fetchone=(1,), fetchall=()):
    cursor = MagicMock()
    cursor.calls = []

    def execute(sql, params=None):
        cursor.calls.append((sql, params))

    cursor.execute.side_effect = execute
    cursor.fetchone.return_value = fetchone
    cursor.fetchall.return_value = list(fetchall)
    db = MagicMock()

    @contextmanager
    def _get_cursor():
        yield cursor

    db._get_cursor = _get_cursor

    def fake_perimeter(_cursor, _ctx, **_kwargs):
        return collect_subjects(user_id=42, otp_role='operator'), {3}, {1, 2}

    patches = [
        (queries, 'load_access_context', lambda _c, _u: dict(context)),
        (queries, 'granted_rule_rights', lambda _c, _s, _u: ({}, [])),
        (queries, 'log_action', lambda *a, **k: None),
        (wiki_perimeter, 'read_perimeter', fake_perimeter),
    ]
    for module, name, replacement in patches:
        original = getattr(module, name)
        setattr(module, name, replacement)
        test.addCleanup(setattr, module, name, original)

    app = Flask(__name__)
    app.register_blueprint(build_wiki_blueprint(
        db=db, require_api_key=lambda f: f,
        build_cors_preflight_response=lambda: ('', 204),
        resolve_requester=lambda: (42, None, None),
        sensitive_access_granted=lambda _user_id, cursor=None: True,
        client_ip=lambda: '127.0.0.1',
        gcs={'signed_url': lambda *a, **k: 'https://x'},
    ))
    app.config['TESTING'] = True
    return app.test_client(), cursor


@unittest.skipIf(Flask is None, 'flask не установлен')
class GateTest(unittest.TestCase):
    """Читатель пишет попытки, но статистику не видит."""

    def test_reader_may_record_a_run(self):
        client, _ = build_client(self, make_context())
        response = client.post('/api/wiki/trainers/runs',
                               json={'key': 'sapar-site-avr', 'stages_total': 6})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['run_id'], 1)

    def test_run_without_key_is_refused(self):
        client, _ = build_client(self, make_context())
        self.assertEqual(client.post('/api/wiki/trainers/runs', json={}).status_code, 400)

    def test_reader_is_refused_stats(self):
        client, _ = build_client(self, make_context())
        for path in ('/api/wiki/trainers/stats',
                     '/api/wiki/trainers/sapar-site-avr/stats',
                     '/api/wiki/trainers/sapar-site-avr/export'):
            response = client.get(path)
            self.assertEqual(response.status_code, 403, path)
            self.assertEqual(response.get_json()['code'], 'WIKI_EDITOR_ONLY', path)

    def test_editor_sees_stats(self):
        client, _ = build_client(self, make_context(can_edit=True),
                                 fetchone=(0, 0, 0, 0, None, None, None, None, None, 0))
        self.assertEqual(client.get('/api/wiki/trainers/stats').status_code, 200)


@unittest.skipIf(Flask is None, 'flask не установлен')
class DepartmentBoundaryTest(unittest.TestCase):
    """Способность редактора не знает отделов — границу ставим отдельно."""

    def _departments_in_query(self, context):
        client, cursor = build_client(
            self, context, fetchone=(0, 0, 0, 0, None, None, None, None, None, 0))
        response = client.get('/api/wiki/trainers/sapar-site-avr/stats')
        self.assertEqual(response.status_code, 200)
        with_depts = [p for _sql, p in cursor.calls if isinstance(p, dict) and 'depts' in p]
        self.assertTrue(with_depts, 'ни один запрос статистики не получил границу отдела')
        return response.get_json(), with_depts[0]['depts']

    def test_supervisor_is_limited_to_own_department(self):
        context = make_context(otp_role='sv', department_id=7, can_publish=True)
        payload, depts = self._departments_in_query(context)
        self.assertEqual(depts, [7])
        self.assertTrue(payload['scoped'])

    def test_super_admin_has_no_boundary(self):
        context = make_context(otp_role='super_admin', department_id=7, can_edit=True)
        payload, depts = self._departments_in_query(context)
        self.assertIsNone(depts, 'у супер-админа границы отдела быть не должно')
        self.assertFalse(payload['scoped'])


class FinishRunTest(unittest.TestCase):
    """Дополняем ТОЛЬКО свою попытку и только пока она не закрыта."""

    def test_update_is_scoped_to_owner_and_open_run(self):
        cursor = MagicMock()
        cursor.fetchone.return_value = (1,)
        wiki_trainers.finish_run(cursor, run_id=10, user_id=42, status='finished',
                                 stages_done=6, errors=2, hints=1, restarts=0,
                                 duration_ms=120000)
        sql, params = cursor.execute.call_args[0]
        self.assertIn('user_id = %(user)s', sql)
        self.assertIn("status = 'started'", sql)
        self.assertEqual(params['user'], 42)
        self.assertEqual(params['id'], 10)

    def test_unknown_status_becomes_abandoned(self):
        """Мусор из браузера не должен превращаться в «прошёл»."""
        cursor = MagicMock()
        cursor.fetchone.return_value = None
        wiki_trainers.finish_run(cursor, run_id=1, user_id=42, status='ура-победа',
                                 stages_done=1, errors=0, hints=0, restarts=0,
                                 duration_ms=1000)
        self.assertEqual(cursor.execute.call_args[0][1]['status'], 'abandoned')

    def test_absurd_duration_is_dropped(self):
        """Вкладку оставили открытой на час — такое время портит медиану."""
        cursor = MagicMock()
        cursor.fetchone.return_value = (1,)
        wiki_trainers.finish_run(cursor, run_id=1, user_id=42, status='finished',
                                 stages_done=6, errors=0, hints=0, restarts=0,
                                 duration_ms=3 * 60 * 60 * 1000)
        self.assertEqual(cursor.execute.call_args[0][1]['duration'],
                         wiki_trainers._clamp(30 * 60 * 1000, 0, 30 * 60 * 1000))


class StaleRunTest(unittest.TestCase):
    """Попытка, о закрытии которой никто не сказал, — брошенная, а не «идёт»."""

    def _rows(self, started_at):
        # Последняя колонка — итог попытки (result). У тренажёров-прогулок его
        # нет, поэтому здесь None: ровно то, что отдаст база для них.
        return [(1, started_at, None, 'started', 'article', 'Иванов', 'СЗоВ', 'Группа 1',
                 'operator', 'Подписание', 'podpisanie', 3, 6, 1, 0, 0, None, None)]

    def _status(self, minutes_ago):
        now = datetime(2026, 8, 22, 12, 0, 0)
        cursor = MagicMock()
        cursor.fetchone.return_value = (1,)
        cursor.fetchall.return_value = self._rows(now - timedelta(minutes=minutes_ago))
        return wiki_trainers.runs(cursor, 'k', now=now)['items'][0]['status']

    def test_fresh_run_stays_started(self):
        self.assertEqual(self._status(5), 'started')

    def test_old_run_reads_as_abandoned(self):
        self.assertEqual(self._status(90), 'abandoned')


@unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
class ReportTest(unittest.TestCase):
    """Книга собирается и с данными, и на пустом тренажёре."""

    TRAINER = {'key': 'sapar-site-avr', 'title': 'Подписание на сайте Сапар',
               'app': 'Сайт Сапар'}

    def _book(self, **kwargs):
        stream = trainer_report.build_workbook(trainer=self.TRAINER, **kwargs)
        return load_workbook(BytesIO(stream.getvalue()))

    def test_empty_export_still_builds(self):
        book = self._book(totals={}, runs=[], people=[], articles=[])
        self.assertEqual(book.sheetnames,
                         ['Контекст', 'Прохождения', 'По людям', 'По статьям'])

    def test_rows_land_on_the_sheet(self):
        book = self._book(
            totals={'runs': 2, 'finished': 1, 'people': 1, 'people_done': 1,
                    'median_ms': 120000, 'avg_errors': 1.5, 'avg_hints': 0.5,
                    'restarts': 1, 'first_at': '2026-08-01T10:00:00',
                    'last_at': '2026-08-02T10:00:00'},
            runs=[{'started_at': '2026-08-02T10:00:00', 'name': 'Иванов Иван',
                   'department': 'СЗоВ', 'group': 'Группа 1', 'role': 'operator',
                   'status': 'finished', 'stages_done': 6, 'stages_total': 6,
                   'errors': 1, 'hints': 0, 'restarts': 0, 'duration_ms': 125000,
                   'source': 'article', 'article_title': 'Подписание',
                   'finished_at': '2026-08-02T10:02:05'}],
            people=[{'name': 'Иванов Иван', 'department': 'СЗоВ', 'group': 'Группа 1',
                     'role': 'operator', 'runs': 2, 'finished': 1, 'errors': 3,
                     'hints': 1, 'best_ms': 125000, 'first_at': '2026-08-01T10:00:00',
                     'last_at': '2026-08-02T10:00:00'}],
            articles=[{'title': 'Подписание', 'runs': 2, 'finished': 1, 'people': 1,
                       'last_at': '2026-08-02T10:00:00'}],
        )
        runs_sheet = book['Прохождения']
        self.assertEqual(runs_sheet['B2'].value, 'Иванов Иван')
        self.assertEqual(runs_sheet['F2'].value, 'Прошёл')
        self.assertEqual(runs_sheet['G2'].value, '6 из 6')
        # Время — числом секунд: по строке «2:05» не отсортировать и не усреднить.
        self.assertEqual(runs_sheet['K2'].value, 125)
        self.assertEqual(runs_sheet['L2'].value, '2:05')
        # Дата обязана быть датой, а не текстом: иначе фильтр Excel по ней не
        # работает, а это первое, что с выгрузкой делают.
        self.assertIsInstance(runs_sheet['A2'].value, datetime)
        self.assertEqual(book['По людям']['A2'].value, 'Иванов Иван')
        self.assertEqual(book['По статьям']['A2'].value, 'Подписание')

    def test_filename_is_safe(self):
        name = trainer_report.report_filename('sapar-site-avr',
                                              datetime(2026, 8, 22))
        self.assertEqual(name, 'trainer_sapar-site-avr_2026-08-22.xlsx')
        # Ключ приходит из адреса: слэши и точки в имени файла недопустимы.
        self.assertNotIn('/', trainer_report.report_filename('../../etc/passwd'))


if __name__ == '__main__':
    unittest.main()


class RunResultTest(unittest.TestCase):
    """Итог попытки: сама заведённая карточка, а не только счётчики.

    У тренажёров-прогулок итог один на всех — «дошёл до конца». У тренажёра
    «Обращение в CRM» итог это выбранная ветка категорий и текст комментария:
    ошибиться веткой можно и с нулём промахов, взяв подсказку.
    """

    def test_only_a_dict_survives(self):
        # Прилетает из браузера, поэтому мусор не пишем.
        for junk in (None, '', [], 'строка', 0, {}):
            self.assertIsNone(wiki_trainers._result_json(junk))

    def test_dict_becomes_json_with_russian_text(self):
        raw = wiki_trainers._result_json({'title': 'Обращение', 'correct': True})
        self.assertIn('Обращение', raw, 'русский текст ушёл в \\uXXXX')
        self.assertEqual(json.loads(raw)['correct'], True)

    def test_oversized_result_is_dropped_not_cut(self):
        # Резать JSON посередине нельзя: получилась бы неразбираемая строка,
        # и колонка стала бы мусорной. Лучше не записать ничего.
        self.assertIsNone(wiki_trainers._result_json({'x': 'я' * 5000}))

    def test_finish_writes_result_and_keeps_it_on_beacon(self):
        cursor = MagicMock()
        cursor.fetchone.return_value = (1,)
        wiki_trainers.finish_run(
            cursor, run_id=1, user_id=42, status='finished', stages_done=10,
            errors=0, hints=0, restarts=0, duration_ms=1000,
            result={'title': 'Обращение в CRM', 'correct': False},
        )
        sql, params = cursor.execute.call_args[0]
        self.assertIn('result', sql)
        self.assertIn('COALESCE', sql,
                      'без COALESCE досылка брошенной попытки затрёт карточку')
        self.assertIn('Обращение в CRM', params['result'])

    def test_finish_without_result_sends_null(self):
        cursor = MagicMock()
        cursor.fetchone.return_value = (1,)
        wiki_trainers.finish_run(
            cursor, run_id=1, user_id=42, status='abandoned', stages_done=2,
            errors=0, hints=0, restarts=0, duration_ms=None,
        )
        self.assertIsNone(cursor.execute.call_args[0][1]['result'])


class FinishedOnlyRouteTest(unittest.TestCase):
    """Попытка, записанная одним запросом: завести и сразу закрыть.

    Тренажёр, у которого итог — сделанная работа, до «Сохранить» молчит: строки
    со статусом «начал» у него не бывает вовсе.
    """

    def test_status_finished_in_start_closes_the_run(self):
        client, cursor = build_client(self, make_context())
        response = client.post('/api/wiki/trainers/runs', json={
            'key': 'crm-ticket-create',
            'source': 'article',
            'stages_total': 10,
            'status': 'finished',
            'stages_done': 10,
            'errors': 1,
            'hints': 0,
            'restarts': 0,
            'duration_ms': 9000,
            'result': {'title': 'Обращение в CRM', 'correct': True},
        })
        self.assertEqual(response.status_code, 200)
        statements = ' '.join(sql for sql, _ in cursor.calls)
        self.assertIn('INSERT INTO wiki_trainer_runs', statements)
        self.assertIn('UPDATE wiki_trainer_runs', statements,
                      'строка осталась в статусе «начал» — попытка не закрыта')

    def test_plain_start_does_not_close_the_run(self):
        client, cursor = build_client(self, make_context())
        client.post('/api/wiki/trainers/runs', json={'key': 'sapar-site-avr', 'stages_total': 6})
        statements = ' '.join(sql for sql, _ in cursor.calls)
        self.assertIn('INSERT INTO wiki_trainer_runs', statements)
        self.assertNotIn('UPDATE wiki_trainer_runs', statements,
                         'обычный тренажёр не должен закрывать попытку на старте')
