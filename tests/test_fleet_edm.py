# -*- coding: utf-8 -*-
"""Раздел «Провайдер ЭДО»: права, разбор файла, обход кабинета и сборка выгрузки.

Пакет fleet_edm импортируется напрямую: в нём нет ни Flask-контекста, ни пула к
боевой БД (в отличие от bot_schedule2.py). Сеть не трогаем — вместо кабинета
Fleet подставляется FakeClient, который ведёт себя ровно так, как измерено на
живом кабинете 20.08.2026:

* список контрагентов ПАРКО-ЗАВИСИМ — чужой парк отдаёт пустоту, а не ошибку;
* провайдера в полях списка нет, он выводится из того, ПО КАКОМУ фильтру строка
  нашлась;
* спросить можно про тысячи ID за раз, отдаются только совпавшие;
* архив — отдельный сегмент, по умолчанию список его не отдаёт;
* часть действующих профилей список молча не возвращает — их добирают карточкой.
"""
import sys
import threading
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from fleet_edm import access, engine, report  # noqa: E402
from fleet_edm.client import MAX_BATCH, FleetClient, FleetError, FleetSessionExpired  # noqa: E402
from fleet_edm.routes import _safe_name  # noqa: E402

PARK_A = 'a' * 32
PARK_B = 'b' * 32

PROVIDERS = [
    {'id': 'paperdo', 'name': 'Бумажный документооборот'},
    {'id': '2KZSP', 'name': 'Sapar'},
    {'id': '2KZVZ', 'name': 'Vezunchik.Pro'},
]


def _driver_id(seed):
    return '{:032x}'.format(seed)


class FakeClient:
    """Кабинет Fleet на столе: знает, кто в каком парке, у кого какой провайдер и
    кто лежит в архиве. Считает запросы — на них опираются проверки экономии."""

    def __init__(self, drivers, parks=(PARK_A, PARK_B), hidden_from_list=(),
                 missing_everywhere=(), concurrency=4):
        # drivers: {id: {'park': ..., 'provider': 'paperdo', 'archive': bool, ...}}
        self.drivers = drivers
        self._parks = list(parks)
        self.hidden_from_list = set(hidden_from_list)
        self.missing_everywhere = set(missing_everywhere)
        self.concurrency = concurrency
        self.requests_count = 0
        self.list_calls = []
        self.card_calls = []
        self._lock = threading.Lock()

    def parks(self, park_id=None):
        with self._lock:
            self.requests_count += 1
        return [{'id': park, 'name': 'Парк ' + park[:2], 'city': 'Алматы'}
                for park in self._parks]

    def edm_providers(self, park_id):
        with self._lock:
            self.requests_count += 1
        return list(PROVIDERS)

    def contractors_all(self, park_id, *, contractor_ids=None, edm_provider=None,
                        archive=False, projection=None, max_pages=400):
        ids = list(contractor_ids or [])
        with self._lock:
            self.requests_count += 1
            self.list_calls.append({'park': park_id, 'provider': edm_provider,
                                    'archive': archive, 'ids': ids})
        out = []
        for cid in ids:
            driver = self.drivers.get(cid)
            if not driver or cid in self.hidden_from_list:
                continue
            if driver['park'] != park_id:
                continue                                  # чужой парк — пустота
            if bool(driver.get('archive')) != bool(archive):
                continue                                  # архив отдельным проходом
            kind = driver.get('employment_type', 'individual_entrepreneur')
            if edm_provider:
                # У сотрудника парка провайдера нет вовсе — фильтр его не находит
                # (проверено карточками на живом кабинете 24.08.2026).
                if kind == 'park_employee' or driver.get('provider') != edm_provider:
                    continue
            out.append({
                'id': cid,
                'full_name': driver.get('full_name', 'Водитель ' + cid[:4]),
                'phone': driver.get('phone', '+77000000000'),
                'work_status': driver.get('work_status', 'working'),
                'employment_type': kind,
            })
        return out

    def driver_card(self, park_id, driver_id):
        with self._lock:
            self.requests_count += 1
            self.card_calls.append((park_id, driver_id))
        driver = self.drivers.get(driver_id)
        if not driver or driver_id in self.missing_everywhere:
            return None
        if driver['park'] != park_id:
            return None                                   # карточка привязана к парку
        kind = driver.get('employment_type', 'individual_entrepreneur')
        # У сотрудника парка поле ЭДО в карточке пусто — это не пропуск, а «поле
        # не про него» (проверено на живом кабинете 24.08.2026).
        if kind == 'park_employee':
            name = ''
        else:
            name = {'paperdo': 'Бумажный документооборот', '2KZSP': 'Sapar',
                    '2KZVZ': 'Vezunchik.Pro'}[driver['provider']]
        return {
            'edm_provider': name,
            'full_name': driver.get('full_name', 'Водитель ' + driver_id[:4]),
            'phone': driver.get('phone', '+77000000000'),
            'work_status': driver.get('work_status', 'working'),
            'employment_type': kind,
        }


def _xlsx(rows, headers=('Название парка', 'ID парка', 'Contractor ID', 'ФИО', 'телефон')):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(headers))
    for row in rows:
        sheet.append(list(row))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


# ── права ────────────────────────────────────────────────────────────────────

class AccessTest(unittest.TestCase):
    def test_global_admin_sees_section(self):
        user = {'role': 'admin', 'department_code': 'szov', 'is_department_head': False}
        self.assertTrue(access.can_view_section(user))
        self.assertTrue(access.can_run_job(user))
        self.assertTrue(access.can_manage_session(user))

    def test_szov_head_sees_section_but_not_session(self):
        head = {'role': 'admin', 'is_department_head': True, 'headed_department_code': 'szov'}
        self.assertTrue(access.can_view_section(head))
        self.assertTrue(access.can_run_job(head))
        # Куки кабинета — доступ ко всем 86 диспетчерским; их меняют только админы.
        self.assertFalse(access.can_manage_session(head))

    def test_head_of_other_department_is_not_admin(self):
        head = {'role': 'admin', 'is_department_head': True, 'headed_department_code': 'op'}
        self.assertFalse(access.can_view_section(head))

    def test_supervisor_and_operator_are_out(self):
        # Раздел отдаёт файл с ФИО и телефонами тысяч водителей — СВ его не видит,
        # хотя табло СЗоВ ему открыто.
        self.assertFalse(access.can_view_section({'role': 'sv', 'department_code': 'szov'}))
        self.assertFalse(access.can_view_section({'role': 'operator', 'department_code': 'szov'}))

    def test_super_admin_always_in(self):
        self.assertTrue(access.can_manage_session({'role': 'super_admin'}))


# ── разбор входного файла ────────────────────────────────────────────────────

class ParseInputTest(unittest.TestCase):
    def test_reads_park_and_driver_columns(self):
        content = _xlsx([
            ('Парк А', PARK_A, _driver_id(1), 'Иванов Иван', '77010000001'),
            ('Парк А', PARK_A, _driver_id(2), 'Петров Пётр', '77010000002'),
        ])
        rows, meta = engine.parse_input(content, 'список.xlsx')
        self.assertEqual(meta['rows_total'], 2)
        self.assertTrue(meta['has_park_column'])
        self.assertEqual(rows[0]['contractor_id'], _driver_id(1))
        self.assertEqual(rows[0]['park_id'], PARK_A)

    def test_file_with_only_ids(self):
        """ТЗ #176 просит именно такой файл — одна колонка с ID."""
        content = _xlsx([(_driver_id(3),), (_driver_id(4),)], headers=('ID водителя',))
        rows, meta = engine.parse_input(content, 'ids.xlsx')
        self.assertEqual(meta['rows_total'], 2)
        self.assertFalse(meta['has_park_column'])
        self.assertEqual(rows[1]['park_id'], '')

    def test_header_below_title_row(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Выгрузка из кабинета', None])
        sheet.append(['Contractor ID', 'ID парка'])
        sheet.append([_driver_id(5), PARK_B])
        stream = BytesIO()
        workbook.save(stream)
        rows, meta = engine.parse_input(stream.getvalue(), 'report.xlsx')
        self.assertEqual(meta['rows_total'], 1)
        self.assertEqual(rows[0]['park_id'], PARK_B)

    def test_bad_id_is_marked_not_dropped(self):
        content = _xlsx([('Парк', PARK_A, 'не-идентификатор', '', '')])
        rows, meta = engine.parse_input(content, 'x.xlsx')
        self.assertEqual(meta['rows_bad_id'], 1)
        self.assertIn('error', rows[0])

    def test_csv_is_accepted(self):
        text = 'Contractor ID;ID парка\n{};{}\n'.format(_driver_id(6), PARK_A)
        rows, meta = engine.parse_input(text.encode('utf-8'), 'ids.csv')
        self.assertEqual(meta['rows_total'], 1)
        self.assertEqual(rows[0]['contractor_id'], _driver_id(6))

    def test_file_without_id_column_is_rejected_with_words(self):
        content = _xlsx([('Парк', 'Алматы')], headers=('Название парка', 'Город'))
        with self.assertRaises(engine.InputError) as error:
            engine.parse_input(content, 'x.xlsx')
        self.assertIn('ID водителя', str(error.exception))


# ── обход ────────────────────────────────────────────────────────────────────

def _park_rows(drivers):
    return [{'contractor_id': cid, 'park_id': info['park']} for cid, info in drivers.items()]


class ResolveTest(unittest.TestCase):
    def test_provider_comes_from_filter_intersection(self):
        drivers = {
            _driver_id(1): {'park': PARK_A, 'provider': 'paperdo'},
            _driver_id(2): {'park': PARK_A, 'provider': '2KZSP'},
        }
        result = engine.resolve(_park_rows(drivers), FakeClient(drivers), control_sample=0)
        self.assertEqual(result['results'][_driver_id(1)]['provider_name'],
                         'Бумажный документооборот')
        self.assertEqual(result['results'][_driver_id(2)]['provider_name'], 'Sapar')
        # Поля ТЗ приходят тем же запросом, отдельных походов за ними нет.
        self.assertTrue(result['results'][_driver_id(1)]['full_name'])
        self.assertTrue(result['results'][_driver_id(1)]['phone'])

    def test_archive_segment_is_not_lost(self):
        """Без второго прохода по архиву в августе терялось 14 444 строки из 147 238."""
        drivers = {
            _driver_id(1): {'park': PARK_A, 'provider': 'paperdo'},
            _driver_id(2): {'park': PARK_A, 'provider': '2KZSP', 'archive': True},
            _driver_id(3): {'park': PARK_A, 'provider': 'paperdo'},
        }
        result = engine.resolve(_park_rows(drivers), FakeClient(drivers), control_sample=0)
        self.assertEqual(result['results'][_driver_id(2)]['provider_name'], 'Sapar')
        self.assertEqual(result['results'][_driver_id(2)]['source'], 'архив')

    def test_rows_hidden_from_list_are_taken_from_card(self):
        """Фильтр списка молча не отдаёт часть действующих профилей (0,05–0,2%)."""
        hidden = _driver_id(2)
        drivers = {
            _driver_id(1): {'park': PARK_A, 'provider': 'paperdo'},
            hidden: {'park': PARK_A, 'provider': '2KZVZ'},
        }
        # Четыре водителя, чтобы парк не ушёл в «мелкие» и провайдер спрашивался списком.
        drivers[_driver_id(3)] = {'park': PARK_A, 'provider': 'paperdo'}
        drivers[_driver_id(4)] = {'park': PARK_A, 'provider': 'paperdo'}
        client = FakeClient(drivers, hidden_from_list=[hidden])
        result = engine.resolve(_park_rows(drivers), client, control_sample=0)
        entry = result['results'][hidden]
        self.assertEqual(entry['provider_name'], 'Vezunchik.Pro')
        self.assertEqual(entry['source'], 'карточка')
        self.assertEqual(result['stats']['from_card'], 1)

    def test_park_is_probed_when_missing(self):
        drivers = {_driver_id(7): {'park': PARK_B, 'provider': '2KZSP'}}
        rows = [{'contractor_id': _driver_id(7), 'park_id': ''}]
        client = FakeClient(drivers)
        result = engine.resolve(rows, client, control_sample=0)
        self.assertEqual(result['results'][_driver_id(7)]['park_id'], PARK_B)
        self.assertEqual(result['results'][_driver_id(7)]['provider_name'], 'Sapar')
        self.assertGreater(result['park_probe_requests'], 0)

    def test_park_probe_asks_whole_list_at_once(self):
        """Главная экономия: диспетчерской отдают ВЕСЬ список сразу, а не сотнями.

        Пустой парк стоит один запрос независимо от того, 10 в файле строк или
        10 000 — иначе перебор 86 диспетчерских упирался бы в тысячи запросов.
        """
        drivers = {_driver_id(i): {'park': PARK_B, 'provider': 'paperdo'}
                   for i in range(1, 251)}
        rows = [{'contractor_id': cid, 'park_id': ''} for cid in drivers]
        client = FakeClient(drivers)
        engine.resolve(rows, client, control_sample=0)
        probe_calls = [call for call in client.list_calls if call['provider'] is None]
        self.assertTrue(probe_calls)
        self.assertEqual(max(len(call['ids']) for call in probe_calls), 250)
        # Два круга (обычный + архив) на два парка — а не 250/100 запросов на каждый.
        self.assertLessEqual(len(probe_calls), 4)

    def test_next_provider_round_asks_only_the_remainder(self):
        """После раунда найденные уходят из очереди — следующий провайдер
        спрашивается уже про остаток, иначе раунды стоили бы одинаково."""
        drivers = {_driver_id(i): {'park': PARK_A, 'provider': 'paperdo'} for i in range(1, 9)}
        drivers[_driver_id(9)] = {'park': PARK_A, 'provider': '2KZSP'}
        client = FakeClient(drivers)
        engine.resolve(_park_rows(drivers), client, control_sample=0)
        by_provider = {}
        for call in client.list_calls:
            if call['provider'] and not call['archive']:
                by_provider.setdefault(call['provider'], len(call['ids']))
        self.assertEqual(by_provider['paperdo'], 9)
        self.assertEqual(by_provider['2KZSP'], 1)

    def test_tiny_park_is_resolved_by_cards(self):
        """В парке один-два водителя — карточка дешевле семи проходов по провайдерам."""
        drivers = {_driver_id(1): {'park': PARK_A, 'provider': 'paperdo'}}
        client = FakeClient(drivers)
        result = engine.resolve(_park_rows(drivers), client, control_sample=0)
        self.assertEqual(result['results'][_driver_id(1)]['source'], 'карточка')
        self.assertFalse([call for call in client.list_calls if call['provider']])

    def test_missing_driver_is_reported_not_invented(self):
        ghost = _driver_id(9)
        drivers = {ghost: {'park': PARK_A, 'provider': 'paperdo'}}
        for index in range(10, 14):
            drivers[_driver_id(index)] = {'park': PARK_A, 'provider': 'paperdo'}
        rows = [{'contractor_id': ghost, 'park_id': PARK_A}]
        client = FakeClient(drivers, hidden_from_list=[ghost], missing_everywhere=[ghost])
        result = engine.resolve(rows, client, control_sample=0)
        self.assertNotIn(ghost, result['results'])
        self.assertEqual(result['stats']['not_found'], 1)

    def test_control_sample_catches_mismatch(self):
        # Подтверждение «бумажных» выключено намеренно: проверяем именно
        # контрольную сверку, а с включённым подтверждением расхождение до неё
        # уже не доживёт (см. VerifyByCardTest).
        drivers = {_driver_id(i): {'park': PARK_A, 'provider': '2KZVZ'} for i in range(1, 6)}

        class LyingClient(FakeClient):
            def driver_card(self, park_id, driver_id):
                card = super().driver_card(park_id, driver_id)
                if card:
                    card['edm_provider'] = 'Sapar'      # карточка говорит другое
                return card

        result = engine.resolve(_park_rows(drivers), LyingClient(drivers),
                                control_sample=5, verify_providers=())
        self.assertEqual(result['check']['checked'], 5)
        self.assertEqual(result['check']['matched'], 0)
        self.assertEqual(len(result['check']['mismatched']), 5)


class VerifyByCardTest(unittest.TestCase):
    """Подтверждение «Бумажного документооборота» карточками.

    Дефект, ради которого проход появился (измерено 01.09.2026): фильтр списка
    кабинета относит водителя к «Бумажному документообороту», а карточка того же
    кабинета говорит «Sapar». Список отстаёт, и наш обход честно повторял за ним.
    """

    class LyingClient(FakeClient):
        """Список говорит «бумажный», карточка — «Sapar». Ровно живой случай."""

        def driver_card(self, park_id, driver_id):
            card = super().driver_card(park_id, driver_id)
            if card:
                card['edm_provider'] = 'Sapar'
            return card

    def _drivers(self, count=4, provider='paperdo'):
        return {_driver_id(i): {'park': PARK_A, 'provider': provider}
                for i in range(1, count + 1)}

    def test_card_overrides_stale_list_value(self):
        drivers = self._drivers()
        result = engine.resolve(_park_rows(drivers), self.LyingClient(drivers),
                                control_sample=0)
        self.assertEqual(result['verify']['checked'], 4)
        self.assertEqual(len(result['verify']['fixed']), 4)
        for entry in result['results'].values():
            self.assertEqual(entry['provider_name'], 'Sapar')
            # Ярлык отдельный от 'карточка': там «список не отдал строку», а
            # здесь «отдал, но соврал». В отчёте это разные фразы.
            self.assertEqual(entry['source'], engine.SOURCE_CORRECTED)
            self.assertIn('отставал', entry['comment'])

    def test_confirmed_rows_get_their_own_source(self):
        # Карточка подтвердила список — строка тоже помечается, иначе перезапуск
        # спросит её заново, а контрольная сверка потратит на неё запросы.
        drivers = self._drivers()
        result = engine.resolve(_park_rows(drivers), FakeClient(drivers),
                                control_sample=0)
        self.assertEqual(result['verify']['checked'], 4)
        self.assertEqual(result['verify']['fixed'], [])
        for entry in result['results'].values():
            self.assertEqual(entry['provider_name'], 'Бумажный документооборот')
            self.assertEqual(entry['source'], engine.SOURCE_VERIFIED)

    def test_cache_replaces_requests(self):
        # Смысл кеша: цену подтверждения платит первый прогон. Люди в файлах
        # повторяются почти полностью, поэтому второй прогон не спрашивает никого.
        drivers = self._drivers()
        rows = _park_rows(drivers)

        class Cache:
            def __init__(self):
                self.store = {}

            def get(self, ids):
                return {cid: self.store[cid] for cid in ids if cid in self.store}

            def put(self, entries):
                for cid, _park, name in entries:
                    self.store[cid] = name

        cache = Cache()
        first = engine.resolve(rows, self.LyingClient(drivers), control_sample=0,
                               card_cache=cache)
        self.assertEqual(first['verify']['checked'], 4)
        self.assertEqual(first['verify']['from_cache'], 0)
        self.assertEqual(len(cache.store), 4)

        client = self.LyingClient(drivers)
        second = engine.resolve(rows, client, control_sample=0, card_cache=cache)
        self.assertEqual(second['verify']['checked'], 0)
        self.assertEqual(second['verify']['from_cache'], 4)
        self.assertEqual(second['verify']['requests'], 0)
        for entry in second['results'].values():
            self.assertEqual(entry['provider_name'], 'Sapar')

    def test_other_providers_are_not_re_asked(self):
        # Расхождение одностороннее: список держит человека в «бумажном» после
        # перехода. Остальные шесть корзин перепроверять незачем — это лишние
        # тысячи запросов в чужой кабинет.
        drivers = self._drivers(provider='2KZVZ')
        client = self.LyingClient(drivers)
        result = engine.resolve(_park_rows(drivers), client, control_sample=0)
        self.assertEqual(result['verify']['checked'], 0)
        self.assertEqual(result['verify']['requests'], 0)
        for entry in result['results'].values():
            self.assertEqual(entry['source'], 'список')

    def test_silent_card_keeps_list_value_and_no_mark(self):
        # Молчащая карточка — это «мы не спросили», а не «провайдера нет».
        # Значение списка остаётся, отметку не ставим: следующая попытка вернётся.
        # Четыре, а не два: парк меньше CARDS_CHEAPER_BELOW ушёл бы в добор
        # карточками ещё до подтверждения, и тест проверял бы не то.
        drivers = self._drivers(count=4)

        class SilentClient(FakeClient):
            def driver_card(self, park_id, driver_id):
                raise FleetError('кабинет молчит')

        result = engine.resolve(_park_rows(drivers), SilentClient(drivers),
                                control_sample=0)
        self.assertEqual(result['verify']['checked'], 0)
        self.assertEqual(result['verify']['silent'], 4)
        for entry in result['results'].values():
            self.assertEqual(entry['provider_name'], 'Бумажный документооборот')
            self.assertNotIn('card_checked', entry)

    def test_resume_does_not_re_ask_confirmed_rows(self):
        # Проход длинный, а деплой разрешения не спрашивает: подтверждённые
        # строки приезжают из контрольной точки с отметкой и второй раз не
        # спрашиваются.
        drivers = self._drivers(count=3)
        rows = _park_rows(drivers)
        first = engine.resolve(rows, self.LyingClient(drivers), control_sample=0)
        resume = {'results': {cid: dict(entry)
                              for cid, entry in first['results'].items()}}
        client = self.LyingClient(drivers)
        second = engine.resolve(rows, client, control_sample=0, resume=resume)
        self.assertEqual(second['verify']['checked'], 0)
        self.assertEqual(second['verify']['requests'], 0)


class ParkEmployeeTest(unittest.TestCase):
    """Сотрудник парка: провайдера ЭДО у него не бывает, и это ОТВЕТ, а не пропуск.

    Проверено на живом кабинете 24.08.2026 — у сотрудников парка edm_provider
    пуст, и ни один из 12 фильтров провайдера их не находит. На файле прогона №10
    таких было 955 из 15 738: каждый стоил 12 бесплодных раундов и запрос карточки.
    """

    def test_park_employee_needs_no_provider_rounds_or_card(self):
        # Один сотрудник парка среди ИП, но парк большой (>=3) — идёт списком.
        drivers = {
            _driver_id(1): {'park': PARK_A, 'provider': 'paperdo'},
            _driver_id(2): {'park': PARK_A, 'provider': 'paperdo'},
            _driver_id(3): {'park': PARK_A, 'provider': '2KZSP'},
            _driver_id(9): {'park': PARK_A, 'employment_type': 'park_employee',
                            'provider': 'paperdo'},   # provider в данных, но не отдаётся
        }
        rows = [{'contractor_id': cid, 'park_id': ''} for cid in drivers]  # нет ID парка
        client = FakeClient(drivers)
        result = engine.resolve(rows, client, control_sample=0)

        entry = result['results'][_driver_id(9)]
        self.assertEqual(entry['provider_name'], '')
        self.assertEqual(entry['source'], engine.SOURCE_NO_PROVIDER)
        self.assertEqual(result['stats'].get('no_provider_by_kind'), 1)
        # Сотрудника парка не искали карточкой — его сняли ещё на переборе.
        self.assertNotIn(_driver_id(9), [cid for _park, cid in client.card_calls])

    def test_park_employee_with_park_id_is_classified_not_carded(self):
        """ID парка в файле есть, перебора нет — тип занятости берём разбором остатка,
        а не карточкой на каждого. Иначе 955 сотрудников = 955 карточек."""
        drivers = {}
        for index in range(1, 26):
            drivers[_driver_id(index)] = {'park': PARK_A, 'employment_type': 'park_employee',
                                          'provider': 'paperdo'}
        drivers[_driver_id(99)] = {'park': PARK_A, 'provider': '2KZSP'}
        client = FakeClient(drivers)
        result = engine.resolve(_park_rows(drivers), client, control_sample=0)

        self.assertEqual(result['stats'].get('no_provider_by_kind'), 25)
        # Ни одного сотрудника парка не спрашивали карточкой.
        self.assertEqual(client.card_calls, [])
        self.assertGreater(result['stats'].get('classified', 0), 0)
        # ИП всё равно определился провайдером.
        self.assertEqual(result['results'][_driver_id(99)]['provider_name'], 'Sapar')

    def test_park_employee_is_checked_every_run(self):
        """Ярлык «провайдера не бывает» — наше утверждение о чужой системе, и
        контрольная сверка обязана трогать его каждым прогоном."""
        drivers = {_driver_id(i): {'park': PARK_A, 'employment_type': 'park_employee',
                                   'provider': 'paperdo'} for i in range(1, 6)}
        drivers[_driver_id(50)] = {'park': PARK_A, 'provider': 'paperdo'}
        rows = [{'contractor_id': cid, 'park_id': ''} for cid in drivers]
        client = FakeClient(drivers)
        result = engine.resolve(rows, client, control_sample=25)
        # Карточки сверки должны включать сотрудников парка (у честного клиента
        # карточка тоже отдаёт пусто — значит совпадает с нашим ярлыком).
        checked_ids = {cid for _park, cid in client.card_calls}
        employees = {_driver_id(i) for i in range(1, 6)}
        self.assertTrue(employees & checked_ids)
        self.assertEqual(result['check']['mismatched'], [])
        self.assertGreater(result['check']['checked'], 0)


class SilentParkTest(unittest.TestCase):
    """Диспетчерская, которая не ответила, — это НЕ «здесь никого нет».

    24.08.2026 на проде это стоило 1 250 ложных «не найден ни в одной
    диспетчерской»: кабинет придушил запрос по парку Jana Taxi, код поймал
    FleetError и пошёл дальше, а все 1 250 человек (из них 1 106 работающих ИП с
    провайдером) уехали в отчёт заказчицы как ненайденные. Список отдавал их за
    14 запросов.
    """

    def setUp(self):
        # Пауза между заходами настоящая (8 и 16 секунд) — в тестах ждать нечего.
        self._pause = engine.FAILED_PARK_PAUSE
        engine.FAILED_PARK_PAUSE = 0

    def tearDown(self):
        engine.FAILED_PARK_PAUSE = self._pause

    class _MutePark(FakeClient):
        """Кабинет, который по одному парку отвечает отказом ровно N первых раз."""

        def __init__(self, drivers, mute_park, times=99, **kw):
            super().__init__(drivers, **kw)
            self.mute_park = mute_park
            self.times = times
            self.refusals = 0

        def contractors_all(self, park_id, **kw):
            if park_id == self.mute_park and self.refusals < self.times:
                self.refusals += 1
                raise FleetError('429 Too Many Requests')
            return super().contractors_all(park_id, **kw)

    def test_silent_park_never_becomes_not_found(self):
        drivers = {_driver_id(i): {'park': PARK_B, 'provider': 'paperdo'}
                   for i in range(1, 6)}
        rows = [{'contractor_id': cid, 'park_id': ''} for cid in drivers]
        client = self._MutePark(drivers, mute_park=PARK_B)
        with self.assertRaises(FleetError):
            engine.resolve(rows, client, control_sample=0)
        # И ни одной строки не объявили ненайденной по дороге.
        self.assertGreaterEqual(client.refusals, engine.FAILED_PARK_RETRIES)

    def test_park_answering_on_retry_is_not_lost(self):
        drivers = {_driver_id(i): {'park': PARK_B, 'provider': 'paperdo'}
                   for i in range(1, 6)}
        rows = [{'contractor_id': cid, 'park_id': ''} for cid in drivers]
        client = self._MutePark(drivers, mute_park=PARK_B, times=1)
        result = engine.resolve(rows, client, control_sample=0)
        self.assertEqual(len(result['results']), 5)
        self.assertEqual(result['results'][_driver_id(1)]['provider_name'],
                         'Бумажный документооборот')

    def test_silent_card_is_not_absence(self):
        """Молчащая карточка не означает «такого водителя нет» — но и не роняет прогон.

        Раньше отказ карточки прерывал обход целиком: контрольная точка есть,
        подхват продолжит. На проде 01.09.2026 (выгрузка №33) это обернулось живым
        клинчем: у строки БЕЗ парка перебор идёт по всем 86 диспетчерским, под
        нагрузкой хоть одна молчит всегда, и «продолжим позже» не наступало
        никогда — шесть подхватов подряд умерли на одном и том же водителе.
        Теперь у такой строки свой ответ, и он по-прежнему НЕ «не найден».
        """
        drivers = {_driver_id(1): {'park': PARK_A, 'provider': 'paperdo'}}

        class MuteCard(FakeClient):
            def driver_card(self, park_id, driver_id):
                raise FleetError('429 Too Many Requests')

        rows = [{'contractor_id': _driver_id(1), 'park_id': PARK_A}]
        result = engine.resolve(rows, MuteCard(drivers), control_sample=0)
        entry = result['results'][_driver_id(1)]
        self.assertEqual(entry['source'], engine.SOURCE_UNVERIFIED)
        self.assertEqual(entry['provider_name'], '')
        self.assertEqual(result['stats'].get('unverified'), 1)
        # Главное: в «не найден» такая строка не попадает.
        self.assertEqual(result['stats'].get('not_found', 0), 0)

    def test_unverified_row_says_so_in_the_file(self):
        """В файле «не смогли спросить» обязано читаться иначе, чем «нет в кабинете»."""
        drivers = {_driver_id(1): {'park': PARK_A, 'provider': 'paperdo'}}

        class MuteCard(FakeClient):
            def driver_card(self, park_id, driver_id):
                raise FleetError('429 Too Many Requests')

        rows = [{'contractor_id': _driver_id(1), 'park_id': PARK_A}]
        result = engine.resolve(rows, MuteCard(drivers), control_sample=0)
        table = report._rows_for_sheet(rows, result['results'], {})
        self.assertIn('Не смогли проверить', table[0]['comment'])
        self.assertNotIn('не найден', table[0]['comment'].lower())


class SegmentTest(unittest.TestCase):
    """Архив спрашиваем там, где архивные есть.

    Перебор парков говорит не только парк, но и сегмент — тем же ответом. Значит
    шесть архивных раундов по диспетчерской без архивных водителей — это шесть
    заведомо пустых запросов. А вот когда сегмент неизвестен (ID парка пришёл в
    файле, перебора не было), спрашивать надо оба: в августе без второго прохода
    терялось 14 444 строки из 147 238.
    """

    def test_archive_is_not_asked_when_probe_saw_nobody_there(self):
        drivers = {_driver_id(i): {'park': PARK_A, 'provider': 'paperdo'}
                   for i in range(1, 6)}
        rows = [{'contractor_id': cid, 'park_id': ''} for cid in drivers]
        client = FakeClient(drivers)
        engine.resolve(rows, client, control_sample=0)
        archive_provider_calls = [call for call in client.list_calls
                                  if call['archive'] and call['provider']]
        self.assertEqual(archive_provider_calls, [])

    def test_archived_driver_is_still_found_after_probe(self):
        drivers = {_driver_id(1): {'park': PARK_A, 'provider': 'paperdo'},
                   _driver_id(2): {'park': PARK_A, 'provider': '2KZSP', 'archive': True},
                   _driver_id(3): {'park': PARK_A, 'provider': 'paperdo'}}
        rows = [{'contractor_id': cid, 'park_id': ''} for cid in drivers]
        result = engine.resolve(rows, FakeClient(drivers), control_sample=0)
        self.assertEqual(result['results'][_driver_id(2)]['provider_name'], 'Sapar')
        self.assertEqual(result['results'][_driver_id(2)]['source'], 'архив')

    def test_unknown_segment_is_asked_in_both(self):
        """ID парка в файле есть — сегмент неизвестен, архив обязан спрашиваться."""
        drivers = {_driver_id(1): {'park': PARK_A, 'provider': 'paperdo'},
                   _driver_id(2): {'park': PARK_A, 'provider': '2KZSP', 'archive': True},
                   _driver_id(3): {'park': PARK_A, 'provider': 'paperdo'}}
        client = FakeClient(drivers)
        result = engine.resolve(_park_rows(drivers), client, control_sample=0)
        self.assertEqual(result['results'][_driver_id(2)]['provider_name'], 'Sapar')
        self.assertTrue([call for call in client.list_calls
                         if call['archive'] and call['provider']])


class ResumeTest(unittest.TestCase):
    """Продолжение прерванной выгрузки — то, из-за чего сорвались оба прогона 21.08.

    Прерывание изображаем исключением из клиента посреди раундов, затем передаём
    накопленную контрольную точку во второй вызов resolve — как это сделает
    подхват после перезапуска процесса.
    """

    class _CheckpointSink:
        """То, что в бою делают queries.save_checkpoint/load_checkpoint, но в
        памяти: копит строки и этапы, отдаёт их обратно как resume."""

        def __init__(self):
            self.results = {}
            self.parks = {}
            self.stages = {}

        def save(self, rows=(), stages=None):
            for cid, park, payload in rows or ():
                if park:
                    self.parks[cid] = park
                if payload:
                    self.results[cid] = payload
            if stages is not None:
                self.stages = dict(stages)

        def resume(self):
            return {'results': dict(self.results), 'parks': dict(self.parks),
                    'stages': dict(self.stages)}

    def test_interrupted_run_resumes_without_redoing_rounds(self):
        drivers = {_driver_id(i): {'park': PARK_A, 'provider': 'paperdo'}
                   for i in range(1, 21)}
        drivers[_driver_id(90)] = {'park': PARK_A, 'provider': '2KZVZ'}
        rows = _park_rows(drivers)
        sink = self._CheckpointSink()

        # Первый прогон: клиент падает после того, как «Бумажный» уже прошёл.
        class DyingClient(FakeClient):
            def __init__(self, *a, **k):
                super().__init__(*a, **k)
                self.paperdo_seen = False

            def contractors_all(self, park_id, *, edm_provider=None, **kw):
                if edm_provider == '2KZVZ':
                    # Сессия «протухла» ровно на раунде Vezunchik — до него
                    # «Бумажный» успел записаться в контрольную точку.
                    raise FleetSessionExpired('обрыв посреди обхода')
                return super().contractors_all(park_id, edm_provider=edm_provider, **kw)

        with self.assertRaises(FleetSessionExpired):
            engine.resolve(rows, DyingClient(drivers), control_sample=0,
                           checkpoint=sink.save)
        # Контрольная точка успела снять «Бумажный».
        self.assertGreaterEqual(len(sink.results), 20)
        self.assertIn('paperdo|0', sink.stages.get('rounds', []))

        # Второй прогон на свежем клиенте продолжает с контрольной точки.
        client2 = FakeClient(drivers)
        result = engine.resolve(rows, client2, control_sample=0, resume=sink.resume())
        self.assertEqual(len(result['results']), 21)
        self.assertEqual(result['results'][_driver_id(90)]['provider_name'],
                         'Vezunchik.Pro')

        # Раунд «Бумажного» второй раз не запрашивался — его сняла контрольная точка.
        paperdo_calls = [c for c in client2.list_calls if c['provider'] == 'paperdo']
        self.assertEqual(paperdo_calls, [])

    def test_stopper_halts_the_walk(self):
        """Стоп-кран: если карточку подхватил другой процесс, обход обязан встать
        сам, а не жечь второй темп запросов к чужому кабинету."""
        drivers = {_driver_id(i): {'park': PARK_A, 'provider': 'paperdo'}
                   for i in range(1, 40)}
        with self.assertRaises(engine.Cancelled):
            engine.resolve(_park_rows(drivers), FakeClient(drivers), control_sample=0,
                           should_stop=lambda: True)


class NeedsClassificationTest(unittest.TestCase):
    def test_classify_only_pays_off_above_two_per_park(self):
        park = PARK_A
        few = [(park, _driver_id(i)) for i in range(3)]           # 3 строки, 1 парк
        many = [(park, _driver_id(i)) for i in range(30)]          # 30 строк, 1 парк
        self.assertFalse(engine._needs_classification(few))        # ниже пола
        self.assertTrue(engine._needs_classification(many))

    def test_orphans_without_park_are_not_classified(self):
        no_park = [('', _driver_id(i)) for i in range(30)]
        self.assertFalse(engine._needs_classification(no_park))


class ParallelismTest(unittest.TestCase):
    def test_one_park_is_sliced_so_threads_are_not_idle(self):
        """Файл из одной диспетчерской: без нарезки шесть потоков ждали бы один."""
        tasks = engine._split_tasks({PARK_A: [_driver_id(i) for i in range(4000)]},
                                    concurrency=6)
        self.assertGreater(len(tasks), 1)
        self.assertEqual(sum(len(ids) for _park, ids in tasks), 4000)

    def test_many_parks_are_not_sliced(self):
        pending = {'{:032x}'.format(index): [_driver_id(index)] for index in range(10)}
        tasks = engine._split_tasks(pending, concurrency=6)
        self.assertEqual(len(tasks), 10)

    def test_empty_parks_produce_no_tasks(self):
        self.assertEqual(engine._split_tasks({PARK_A: []}, concurrency=6), [])


class ThrottleTest(unittest.TestCase):
    """Просьба кабинета «помедленнее» — это ОДИН эпизод, а не отдельный ответ."""

    def _client(self):
        return FleetClient({'Session_id': 'x'}, 'UA', concurrency=6)

    def test_echo_refusals_do_not_multiply_the_pause(self):
        # При шести потоках отказ прилетает всем запросам, уже летящим по проводу.
        # Если наращивать паузу на каждый, она растёт в полтора раза шесть раз
        # подряд — так на проде и вышло 20 секунд простоя вместо секунды.
        client = self._client()
        first = client._note_throttled()
        echoes = [client._note_throttled() for _ in range(5)]
        self.assertEqual(first, 1.0)
        self.assertTrue(all(value <= first for value in echoes))
        self.assertEqual(client.concurrency, 5)     # поток убавили ровно один раз

    def test_threads_return_only_after_a_long_quiet_stretch(self):
        client = self._client()
        client._note_throttled()
        self.assertEqual(client.concurrency, 5)
        for _ in range(199):
            client._note_success()
        self.assertEqual(client.concurrency, 5)
        client._note_success()
        self.assertEqual(client.concurrency, 6)

    def test_growth_stops_at_the_measured_ceiling(self):
        client = FleetClient({'Session_id': 'x'}, 'UA', concurrency=6)
        for _ in range(1000):
            client._note_success()
        self.assertEqual(client.concurrency, 6)


class ClientPagingTest(unittest.TestCase):
    """Курсорная постраничность — на подменённом транспорте, без сети."""

    class _Response:
        def __init__(self, payload):
            self.status_code = 200
            self._payload = payload

        def json(self):
            return self._payload

    class _Session:
        def __init__(self, pages):
            self.pages = list(pages)
            self.bodies = []
            self.cookies = type('J', (), {'update': lambda self, value: None})()

        def request(self, method, url, headers=None, data=None, timeout=None):
            self.bodies.append(data)
            return ClientPagingTest._Response(self.pages.pop(0))

    def test_pages_until_cursor_runs_out(self):
        pages = [
            {'contractors': [{'id': _driver_id(i)} for i in range(100)], 'cursor': 'c1'},
            {'contractors': [{'id': _driver_id(i)} for i in range(100, 130)], 'cursor': None},
        ]
        session = self._Session(pages)
        client = FleetClient({'Session_id': 'x'}, 'UA', session=session)
        found = client.contractors_all(PARK_A, contractor_ids=[_driver_id(i) for i in range(500)])
        self.assertEqual(len(found), 130)
        self.assertEqual(len(session.bodies), 2)
        self.assertIn('c1', session.bodies[1])
        # Размер страницы остаётся сотней — это ограничение ответа, не запроса.
        self.assertIn('"limit": {}'.format(MAX_BATCH), session.bodies[0])

    def test_whole_list_goes_in_one_request(self):
        session = self._Session([{'contractors': [], 'cursor': None}])
        client = FleetClient({'Session_id': 'x'}, 'UA', session=session)
        client.contractors_all(PARK_A, contractor_ids=[_driver_id(i) for i in range(5000)])
        self.assertEqual(len(session.bodies), 1)
        self.assertIn(_driver_id(4999), session.bodies[0])


class NonJsonAnswerTest(unittest.TestCase):
    """Ответ без JSON — это ДВА разных случая, и путать их нельзя.

    Протухшая сессия отдаёт код 200 и HTML страницы входа. А неверно собранный
    запрос отвечает 400 с пустым телом (измерено на живом кабинете 01.09.2026:
    driver_ids вместо driver_id → 400, application/octet-stream, 0 байт).
    До этой правки оба случая говорили «сессия протухла, нужен новый вход», и
    человека отправляли логиниться заново из-за нашей же ошибки в запросе.
    """

    class _Response:
        def __init__(self, status_code, body=b''):
            self.status_code = status_code
            self.content = body
            self.text = body.decode('utf-8', 'replace')
            self.headers = {'content-type': 'application/octet-stream'}

        def json(self):
            raise ValueError('не JSON')

    class _Session:
        def __init__(self, response):
            self.response = response
            self.cookies = type('J', (), {'update': lambda self, value: None})()

        def request(self, method, url, headers=None, data=None, timeout=None):
            return self.response

    def _client(self, status_code, body=b''):
        session = self._Session(self._Response(status_code, body))
        return FleetClient({'Session_id': 'x'}, 'UA', session=session)

    def test_html_with_200_is_expired_session(self):
        client = self._client(200, b'<html><body>login</body></html>')
        with self.assertRaises(FleetSessionExpired):
            client.parks()

    def test_empty_400_is_our_broken_request(self):
        client = self._client(400)
        with self.assertRaises(FleetError) as caught:
            client.parks()
        self.assertNotIsInstance(caught.exception, FleetSessionExpired)
        self.assertIn('400', str(caught.exception))


# ── сборка файла ─────────────────────────────────────────────────────────────

class ReportTest(unittest.TestCase):
    def _build(self, **kwargs):
        drivers = {
            _driver_id(1): {'park': PARK_A, 'provider': 'paperdo', 'phone': '77010000001'},
            _driver_id(2): {'park': PARK_A, 'provider': '2KZSP', 'phone': '77010000002'},
            _driver_id(3): {'park': PARK_A, 'provider': 'paperdo', 'phone': '77010000003'},
            _driver_id(4): {'park': PARK_A, 'provider': 'paperdo', 'phone': '77010000004'},
        }
        rows = [{'contractor_id': cid, 'park_id': PARK_A,
                 'source_park_name': 'Парк А', 'row_number': index}
                for index, cid in enumerate(drivers, start=2)]
        resolution = engine.resolve(rows, FakeClient(drivers), control_sample=0)
        stream = report.build_workbook(rows, resolution, source_name='вход.xlsx', **kwargs)
        return rows, resolution, stream

    def test_sheets_and_columns(self):
        _rows, _resolution, stream = self._build()
        workbook = load_workbook(stream)
        self.assertEqual(workbook.sheetnames,
                         ['Контекст', 'Водители', 'Свод по провайдерам', 'Провайдеры по паркам'])
        sheet = workbook['Водители']
        headers = [cell.value for cell in sheet[1]]
        # Все поля из ТЗ #176 на месте.
        for title in ('Название парка', 'ID водителя', 'ФИО', 'Телефон', 'Провайдер ЭДО'):
            self.assertIn(title, headers)

    def test_phone_and_ids_stay_text(self):
        _rows, _resolution, stream = self._build()
        sheet = load_workbook(stream)['Водители']
        phone_column = [cell.value for cell in sheet[1]].index('Телефон') + 1
        cell = sheet.cell(row=2, column=phone_column)
        self.assertEqual(cell.number_format, '@')
        self.assertIsInstance(cell.value, str)

    def test_ignored_errors_patch_targets_drivers_sheet(self):
        """Зелёный уголок «Число сохранено как текст» гасится на листе «Водители»
        (второй в книге), а не на первом — иначе тег уедет в «Контекст»."""
        seen = {}

        def fake_patch(stream, sqref, sheet_path='xl/worksheets/sheet1.xml'):
            seen['sqref'] = sqref
            seen['sheet'] = sheet_path
            return stream

        self._build(text_warning_patch=fake_patch)
        self.assertEqual(seen['sheet'], 'xl/worksheets/sheet2.xml')
        self.assertTrue(seen['sqref'].startswith('B2:'))

    def test_context_sheet_states_the_date_and_caveats(self):
        _rows, _resolution, stream = self._build()
        sheet = load_workbook(stream)['Контекст']
        text = '\n'.join(str(cell.value or '') for row in sheet.iter_rows() for cell in row)
        self.assertIn('Дата и время сборки', text)
        self.assertIn('Бумажный документооборот', text)   # оговорка «это выбор, а не пропуск»


# ── мелочи, на которых уже обжигались ────────────────────────────────────────

class HelpersTest(unittest.TestCase):
    def test_upload_name_keeps_cyrillic(self):
        # secure_filename съедает кириллицу целиком, и «Нет провайдера.xlsx»
        # превращается в «.xlsx» — в списке выгрузок такое имя бесполезно.
        self.assertEqual(_safe_name('Нет провайдера 18.08.xlsx'), 'Нет провайдера 18.08.xlsx')
        self.assertNotIn('/', _safe_name('../../etc/passwd.xlsx'))

    def test_provider_name_is_stripped(self):
        # Справочник кабинета отдаёт «Partners Pay\n» — без strip сверка с
        # карточкой даёт ложные расхождения.
        self.assertEqual(FleetClient.card_provider({'edm_provider': 'Partners Pay\n'}),
                         'Partners Pay')
        self.assertEqual(FleetClient.card_provider({}), '')

    def test_cookies_accepted_in_both_shapes(self):
        as_list = FleetClient._normalize_cookies([{'name': 'Session_id', 'value': 'x'}])
        as_dict = FleetClient._normalize_cookies({'Session_id': 'x'})
        self.assertEqual(as_list, as_dict)


if __name__ == '__main__':
    unittest.main()
