# -*- coding: utf-8 -*-
"""Выгрузка реестра посылок в Excel (задача #257, поручил Ядигаров Руслан).

Проверяется не «собирается ли книга», а обещания, которые выгрузка даёт
человеку и которые молча ломаются при рефакторинге:

  * В ФАЙЛЕ ТО ЖЕ, ЧТО НА ЭКРАНЕ. Отбор списка и отбор выгрузки собираются одним
    `_filter_clause`, а страница выгрузку не сжимает: у списка потолок 200
    строк, и уйди выгрузка через него — файл обрезался бы молча.
  * ЧИСЛА СОВПАДАЮТ С РАЗДЕЛОМ. «Дней в офисе» и «Залежалась» считаются по тем
    же правилам, что `daysInOffice`/`isStale` в parcelMeta.js, а подписи
    статусов и типов — те же слова, что видит человек.
  * ФАЙЛ ОТКРЫВАЕТСЯ И РАБОТАЕТ. Дата — датой (иначе не сортируется), телефон и
    ID — текстом (иначе теряется ведущий ноль и округляется hex), описание с
    ведущим «=» — текстом, а не формулой.
  * ЧЕСТНОСТЬ. Сработавший потолок назван на листе «Контекст»: молча обрезанный
    файл читается как полный.
  * ДВЕРЬ. Выгрузка доступна читателю из СЗоВ — ради него реестр и заводили, —
    но не мягче гейта самого раздела (QR сторожит test_sensitive_section_qr_gate).

База здесь не нужна: курсор подменён, запросы разбираются текстом.
"""

import re
import sys
import unittest
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path
from unittest.mock import MagicMock
from zipfile import ZipFile

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from flask import Flask
except ImportError:  # pragma: no cover
    Flask = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

from parcels import queries as parcels_queries  # noqa: E402
from parcels import report  # noqa: E402
from parcels import schema as parcels_schema  # noqa: E402
from parcels.routes import build_parcels_blueprint  # noqa: E402


VIEW_PATH = ROOT / 'src' / 'components' / 'parcels' / 'ParcelsView.jsx'
META_PATH = ROOT / 'src' / 'components' / 'parcels' / 'parcelMeta.js'
TODAY = date(2026, 8, 31)


def parcel(**overrides):
    """Карточка в том же виде, в каком её отдаёт слой запросов.

    Телефоны учебные (7XX0000000): страж персональных данных считает такие
    отдельно от порога, а раздел весь про телефоны водителей.
    """
    base = {
        'id': 1,
        'received_on': '2026-08-01',
        'city': 'Алматы',
        'office_id': 17,
        'office_name': 'Абая',
        'office_address': 'ул. Абая, 1',
        'driver_account_id': 'a' * 32,
        'driver_name': 'Тестов Тест',
        'driver_phone': '+77000000001',
        'driver_park': 'Тестовый парк',
        'driver_callsign': '0451',
        'driver_car': 'Vento 123 ABC',
        'kind': 'parcel',
        'description': 'Синяя коробка',
        'sender': 'Отправитель',
        'recipient': 'Получатель',
        'order_url': 'https://fleet.example.kz/orders/abc',
        'order_number': '007',
        'status': 'in_office',
        'status_changed_at': None,
        'status_changed_by_name': '',
        'comment': '',
        'created_by_name': 'Менеджер',
        'created_at': '2026-08-01T10:20:30',
    }
    base.update(overrides)
    return base


def build(rows, **kwargs):
    kwargs.setdefault('today', TODAY)
    stream, written = report.build_workbook(rows, **kwargs)
    stream.seek(0)
    return load_workbook(stream), written


def header(sheet):
    return [cell.value for cell in sheet[1]]


def cell(sheet, row, title):
    return sheet.cell(row=row, column=header(sheet).index(title) + 1)


@unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
class WorkbookShapeTests(unittest.TestCase):
    def test_three_sheets_and_context_goes_first(self):
        """Контекст первым — как у «Касаний» и «Провайдера ЭДО».

        Реестр живой: без даты сборки и отбора файл через неделю уже нельзя
        сопоставить со вторым таким же.
        """
        workbook, _ = build([parcel()])
        self.assertEqual(workbook.sheetnames, ['Контекст', 'Посылки', 'По офисам'])

    def test_every_table_sheet_freezes_the_header_and_filters(self):
        """Правило портала: заморозка И автофильтр на каждом листе с таблицей."""
        workbook, _ = build([parcel()])
        for name in ('Посылки', 'По офисам'):
            sheet = workbook[name]
            self.assertEqual(sheet.freeze_panes, 'A2', name)
            self.assertTrue(sheet.auto_filter.ref, name)

    def test_empty_selection_still_gives_a_readable_book(self):
        """Пустой отбор — валидная книга с шапкой, а не отказ."""
        workbook, written = build([])
        self.assertEqual(written, 0)
        self.assertEqual(workbook['Посылки'].max_row, 1)
        self.assertIsNone(workbook['Посылки'].auto_filter.ref)


@unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
class ColumnContractTests(unittest.TestCase):
    def test_date_is_a_real_date_not_a_string(self):
        """Строка «01.08.2026» не сортируется и не попадает в фильтр по периоду."""
        workbook, _ = build([parcel()])
        received = cell(workbook['Посылки'], 2, 'Принята')
        self.assertEqual(received.value, datetime(2026, 8, 1))
        self.assertEqual(received.number_format, 'DD.MM.YYYY')

    def test_timestamps_are_naive(self):
        """Excel про зоны не знает: aware-объект превратился бы в текст."""
        workbook, _ = build([parcel(status='given_to_recipient',
                                    status_changed_at='2026-08-20T09:15:00')])
        changed = cell(workbook['Посылки'], 2, 'Статус изменён')
        self.assertEqual(changed.value, datetime(2026, 8, 20, 9, 15))
        self.assertIsNone(changed.value.tzinfo)

    def test_identifiers_stay_text(self):
        """Телефон, позывной, ID и номер заказа числом теряют ведущий ноль."""
        workbook, _ = build([parcel(driver_phone='+77000000001',
                                    driver_callsign='0451',
                                    order_number='007')])
        sheet = workbook['Посылки']
        for title, value in (('Телефон', '+77000000001'),
                             ('Позывной', '0451'),
                             ('Номер заказа', '007'),
                             ('ID во Флите', 'a' * 32)):
            found = cell(sheet, 2, title)
            self.assertEqual(found.value, value, title)
            self.assertEqual(found.data_type, 's', title)
            self.assertEqual(found.number_format, '@', title)

    def test_one_control_character_does_not_kill_the_whole_export(self):
        """Невидимый символ из чата не должен ломать выгрузку у всех.

        Управляющие символы в XML запрещены, и openpyxl роняет на них ВСЮ книгу
        (`IllegalCharacterError` в момент save). Одна такая карточка означала бы,
        что выгрузка перестала работать для всего раздела, а починить её мог бы
        только разработчик — при том что описание и комментарий сотрудники
        вставляют из переписки, откуда символ приезжает незаметно.
        """
        workbook, written = build([parcel(description='коробка\x07со\x0bзвонком',
                                          comment='\x00тихо')])
        self.assertEqual(written, 1)
        sheet = workbook['Посылки']
        self.assertEqual(cell(sheet, 2, 'Описание').value, 'коробка со звонком')
        # Пробелом, а не пустотой: иначе слова склеились бы в одно.
        self.assertNotIn('\x07', cell(sheet, 2, 'Описание').value)

    def test_context_sheet_survives_a_control_character_too(self):
        """В «Отобрано» приезжает поисковый запрос человека, в «Собрал» — его имя."""
        workbook, _ = build([parcel()], generated_by='Тест\x07ов',
                            filters_note='поиск: «\x0bчто-то»')
        lines = {row[0].value: (row[1].value if len(row) > 1 else None)
                 for row in workbook['Контекст'].iter_rows()}
        self.assertEqual(lines['Собрал'], 'Тест ов')
        self.assertNotIn('\x0b', lines['Отобрано'])
        # Числа на «Контексте» обязаны остаться числами.
        self.assertIsInstance(lines['Строк в файле'], int)

    def test_free_text_starting_with_equals_is_not_a_formula(self):
        """«=коробка» в описании — текст, а не вычисление.

        Описание и комментарий люди пишут руками, а openpyxl выводит тип из
        значения: без явного data_type такая книга открывалась бы с ошибкой
        вычисления, а то и с предупреждением безопасности.
        """
        workbook, _ = build([parcel(description='=коробка', comment='=1+1')])
        sheet = workbook['Посылки']
        for title in ('Описание', 'Комментарий'):
            self.assertEqual(cell(sheet, 2, title).data_type, 's', title)

    def test_only_http_links_become_clickable(self):
        """Ссылка, ведущая непонятно куда, хуже её отсутствия — то же правило,
        что в карточке."""
        workbook, _ = build([parcel(order_url='https://fleet.example.kz/orders/abc'),
                             parcel(id=2, order_url='javascript:alert(1)')])
        sheet = workbook['Посылки']
        self.assertTrue(cell(sheet, 2, 'Заказ').hyperlink)
        self.assertIsNone(cell(sheet, 3, 'Заказ').hyperlink)

    def test_order_link_keeps_the_address_as_the_value(self):
        """По id заказа в файле ищут поиском — подменённый текст сломал бы его."""
        workbook, _ = build([parcel(order_url='https://fleet.example.kz/orders/abc')])
        self.assertEqual(cell(workbook['Посылки'], 2, 'Заказ').value,
                         'https://fleet.example.kz/orders/abc')

    def test_licence_number_is_not_exported(self):
        """Номер прав раздел хранит, но нигде не показывает — в файл он не идёт.

        Выгрузка не должна выносить наружу больше, чем человек видит на экране.
        """
        titles = [title for _key, title, _width in report.COLUMNS]
        self.assertNotIn('driver_license', [key for key, _t, _w in report.COLUMNS])
        for word in ('Права', 'Удостоверение', 'ВУ'):
            self.assertNotIn(word, titles)


@unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
class NumbersMatchTheScreenTests(unittest.TestCase):
    def test_days_are_counted_only_for_parcels_still_in_the_office(self):
        """У переданной посылки в разделе прочерк — и в файле должен быть он же."""
        workbook, _ = build([
            parcel(id=1, received_on='2026-08-25', status='in_office'),
            parcel(id=2, received_on='2026-08-25', status='given_to_recipient'),
        ])
        sheet = workbook['Посылки']
        self.assertEqual(cell(sheet, 2, 'Дней в офисе').value, 6)
        self.assertIsNone(cell(sheet, 3, 'Дней в офисе').value)
        self.assertIsNone(cell(sheet, 3, 'Залежалась').value)

    def test_stale_starts_exactly_at_thirty_days(self):
        """Порог тот же, которым раздел подсвечивает строку янтарным."""
        workbook, _ = build([
            parcel(id=1, received_on='2026-08-02'),   # 29 дней
            parcel(id=2, received_on='2026-08-01'),   # 30 дней
        ])
        sheet = workbook['Посылки']
        self.assertEqual(cell(sheet, 2, 'Залежалась').value, 'нет')
        self.assertEqual(cell(sheet, 3, 'Залежалась').value, 'да')

    def test_stale_threshold_matches_the_frontend_constant(self):
        """Разъедься пороги — на экране «залежались 12», а в файле одиннадцать."""
        source = META_PATH.read_text(encoding='utf-8')
        found = re.search(r'STALE_AFTER_DAYS\s*=\s*(\d+)', source)
        self.assertIsNotNone(found, 'STALE_AFTER_DAYS пропал из parcelMeta.js')
        self.assertEqual(int(found.group(1)), report.STALE_AFTER_DAYS)

    def test_every_status_and_kind_of_the_registry_has_a_word(self):
        """Четвёртый статус не должен приехать в файл сырым кодом.

        Сверка с parcelMeta.js ловит расхождение подписей, но НЕ появление
        нового кода: словарь без него просто не проверялся бы. Поэтому набор
        кодов сверяется с DDL — с тем самым CHECK, который их и задаёт.
        """
        self.assertEqual(sorted(report.STATUS_LABELS), sorted(parcels_schema.PARCEL_STATUSES),
                         'в report.py нет слова для статуса из схемы')
        self.assertEqual(sorted(report.KIND_LABELS), sorted(parcels_schema.PARCEL_KINDS),
                         'в report.py нет слова для типа посылки из схемы')

    def test_labels_are_the_same_words_the_person_sees(self):
        """Подписи статусов и типов сверяются с parcelMeta.js буквально."""
        source = META_PATH.read_text(encoding='utf-8')
        for code, label in report.STATUS_LABELS.items():
            self.assertRegex(source, r"%s:\s*\{\s*\n\s*label:\s*'%s'" % (code, label),
                             'подпись статуса %s разошлась' % code)
        for code, label in report.KIND_LABELS.items():
            self.assertIn("%s: { label: '%s' }" % (code, label), source,
                          'подпись типа %s разошлась' % code)

    def test_office_summary_is_counted_from_the_very_same_rows(self):
        """Сводка, не сходящаяся с данными под ней, хуже её отсутствия."""
        workbook, _ = build([
            parcel(id=1, city='Алматы', office_name='Абая', received_on='2026-08-01'),
            parcel(id=2, city='Алматы', office_name='Абая', status='given_to_sender'),
            parcel(id=3, city='Астана', office_name='Сарыарка', received_on='2026-08-30'),
        ])
        sheet = workbook['По офисам']
        rows = {row[1]: row for row in sheet.iter_rows(min_row=2, values_only=True)}
        self.assertEqual(rows['Абая'][2:], (2, 1, 1, 0, 1))
        self.assertEqual(rows['Сарыарка'][2:], (1, 1, 0, 0, 0))

    def test_offices_with_something_lying_come_first(self):
        """Раздел про невостребованное: первым — тот, кому пора разбирать полку."""
        workbook, _ = build([
            parcel(id=1, office_name='Тихий', status='given_to_recipient'),
            parcel(id=2, office_name='Тихий', status='given_to_recipient'),
            parcel(id=3, office_name='Занятой', status='in_office'),
        ])
        names = [row[1] for row in workbook['По офисам'].iter_rows(min_row=2, values_only=True)]
        self.assertEqual(names[0], 'Занятой')


@unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
class ContextSheetTests(unittest.TestCase):
    def context(self, workbook):
        return {row[0].value: (row[1].value if len(row) > 1 else None)
                for row in workbook['Контекст'].iter_rows()}

    def test_context_names_who_collected_and_what_was_selected(self):
        workbook, _ = build([parcel()], generated_by='Хайрихан Шерзад',
                            filters_note='город: Алматы')
        lines = self.context(workbook)
        self.assertEqual(lines['Собрал'], 'Хайрихан Шерзад')
        self.assertEqual(lines['Отобрано'], 'город: Алматы')

    def test_without_filters_the_context_says_so_instead_of_leaving_a_blank(self):
        workbook, _ = build([parcel()], filters_note='')
        self.assertEqual(self.context(workbook)['Отобрано'],
                         'без фильтров — весь реестр целиком')

    def test_truncated_export_says_so_out_loud(self):
        """Молча обрезанный файл читается как полный."""
        workbook, _ = build([parcel()], total=99999, truncated=True)
        lines = self.context(workbook)
        self.assertIn('В ФАЙЛ ПОПАЛО НЕ ВСЁ', lines)
        self.assertIn('99999', lines['В ФАЙЛ ПОПАЛО НЕ ВСЁ'])

    def test_complete_export_does_not_scare_with_the_warning(self):
        workbook, _ = build([parcel()], total=1, truncated=False)
        self.assertNotIn('В ФАЙЛ ПОПАЛО НЕ ВСЁ', self.context(workbook))

    def test_counters_are_counted_over_the_rows_of_the_file(self):
        workbook, _ = build([
            parcel(id=1, received_on='2026-08-01'),
            parcel(id=2, status='given_to_recipient'),
            parcel(id=3, status='given_to_sender'),
        ])
        lines = self.context(workbook)
        self.assertEqual(lines['Строк в файле'], 3)
        self.assertEqual(lines['В офисе'], 1)
        self.assertEqual(lines['Из них залежались (30 дней и больше)'], 1)
        self.assertEqual(lines['Передали получателю'], 1)
        self.assertEqual(lines['Вернули отправителю'], 1)


class TextWarningPatchTests(unittest.TestCase):
    """Зелёный уголок «Число сохранено как текст» гасится, но не любой ценой."""

    def calls(self):
        recorded = []

        def patch(stream, sqref, sheet_path=None):
            recorded.append({'sqref': sqref, 'sheet_path': sheet_path})
            return stream

        return recorded, patch

    @unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
    def test_patch_targets_the_sheet_with_the_table(self):
        """Хелпер адресует лист ФИЗИЧЕСКИМ номером (sheetN.xml), а не именем.

        Поэтому проверяется не «sheet2» наизусть, а совпадение с фактическим
        местом листа «Посылки» в книге: переставь листы — и патч молча гасил бы
        уголок на «Контексте», где гасить нечего.
        """
        recorded, patch = self.calls()
        report.build_workbook([parcel()], today=TODAY, text_warning_patch=patch)
        self.assertEqual(len(recorded), 1)

        workbook, _ = build([parcel()])
        expected = workbook.sheetnames.index(report.SHEET_PARCELS) + 1
        self.assertEqual(recorded[0]['sheet_path'],
                         'xl/worksheets/sheet%d.xml' % expected)

    @unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
    def test_patch_really_lands_in_the_saved_book(self):
        """Сквозная проверка: тег доезжает до листа с таблицей и не ломает книгу.

        Двойник хелпера здесь не подошёл бы — узел обязан встать ПОСЛЕ
        <pageMargins>, и это свойство настоящей перепаковки, а не наша.
        """
        from tests import source_cache
        import ast
        import io as _io
        from zipfile import ZipFile as _Zip, ZIP_DEFLATED as _DEFLATED

        node = next(n for n in source_cache.tree(ROOT / 'bot_schedule2.py').body
                    if isinstance(n, ast.FunctionDef)
                    and n.name == '_excel_suppress_number_as_text_warning')
        namespace = {}
        exec(compile(ast.Module(body=[node], type_ignores=[]), 'bot', 'exec'),
             {'BytesIO': _io.BytesIO, 'ZipFile': _Zip, 'ZIP_DEFLATED': _DEFLATED},
             namespace)

        stream, _ = report.build_workbook(
            [parcel()], today=TODAY,
            text_warning_patch=namespace['_excel_suppress_number_as_text_warning'])
        stream.seek(0)
        xml = ZipFile(stream).read('xl/worksheets/sheet2.xml').decode('utf-8')
        self.assertIn('<ignoredErrors>', xml)
        self.assertLess(xml.index('<pageMargins'), xml.index('<ignoredErrors'))

        stream.seek(0)
        self.assertEqual(load_workbook(stream).sheetnames[1], report.SHEET_PARCELS)

    @unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
    def test_every_text_column_is_covered_and_no_other(self):
        """Гасим уголок ровно у текстовых колонок — не меньше и не больше.

        Считать количество диапазонов мало: опечатка в TEXT_COLUMNS даёт то же
        число, но `_column_index` для неизвестного ключа возвращает 1, и уголок
        гасился бы у колонки «№», а у телефона оставался. Поэтому сверяются
        сами буквы колонок.
        """
        from openpyxl.utils import get_column_letter
        recorded, patch = self.calls()
        report.build_workbook([parcel()], today=TODAY, text_warning_patch=patch)

        titles = [title for _key, title, _width in report.COLUMNS]
        keys = [key for key, _title, _width in report.COLUMNS]
        expected = {get_column_letter(keys.index(key) + 1) for key in report.TEXT_COLUMNS}
        got = {part.split('2')[0] for part in recorded[0]['sqref'].split(' ')}
        self.assertEqual(got, expected, recorded[0]['sqref'])
        # И ни одна из них не «№» — эту колонку гасить незачем, она и есть число.
        self.assertNotIn(get_column_letter(titles.index('№') + 1), got)

    def test_empty_export_is_not_patched(self):
        """sqref вида 'B2:B1' — битый диапазон, Excel объявит книгу повреждённой."""
        recorded, patch = self.calls()
        report.build_workbook([], today=TODAY, text_warning_patch=patch)
        self.assertEqual(recorded, [])

    @unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
    def test_a_broken_patch_does_not_swallow_the_file(self):
        """Значок в углу ячейки — досадно, но не повод не отдать файл."""
        def explode(_stream, _sqref, sheet_path=None):
            raise RuntimeError('перепаковка не удалась')

        stream, written = report.build_workbook([parcel()], today=TODAY,
                                                text_warning_patch=explode)
        self.assertEqual(written, 1)
        stream.seek(0)
        self.assertIn('xl/worksheets/sheet2.xml', ZipFile(stream).namelist())


class TextWarningWiringTests(unittest.TestCase):
    """Проводка хелпера от монолита до книги.

    Три звена, и каждое рвётся молча. Особенно последнее: подключение раздела в
    `bot_schedule2` обёрнуто в `try/except`, поэтому переименованный аргумент не
    падает с ошибкой — блюпринт просто не регистрируется, и раздел «Посылки»
    исчезает целиком, а в логах остаётся одна строка.
    """

    def setUp(self):
        self.bot = (ROOT / 'bot_schedule2.py').read_text(encoding='utf-8-sig')
        self.routes = (ROOT / 'parcels' / 'routes.py').read_text(encoding='utf-8')

    def test_monolith_hands_the_helper_to_the_section(self):
        block = self.bot.split('build_parcels_blueprint(')[1].split('))')[0]
        self.assertIn('excel_text_warning=_excel_suppress_number_as_text_warning', block)

    def test_the_factory_accepts_it_under_that_very_name(self):
        signature = self.routes.split('def build_parcels_blueprint(')[1].split('):')[0]
        self.assertIn('excel_text_warning', signature)

    def test_the_factory_survives_without_it(self):
        """У аргумента есть значение по умолчанию: без хелпера книга соберётся,
        просто с зелёным уголком. Ронять раздел из-за косметики нельзя."""
        signature = self.routes.split('def build_parcels_blueprint(')[1].split('):')[0]
        self.assertIn('excel_text_warning=None', signature)

    def test_the_route_passes_it_into_the_book(self):
        self.assertIn('text_warning_patch=excel_text_warning', self.routes)


class FilenameTests(unittest.TestCase):
    def test_filename_carries_the_date_of_collection(self):
        self.assertEqual(report.report_filename(datetime(2026, 8, 31, 14, 0)),
                         'Посылки 31.08.2026.xlsx')

    def test_frontend_builds_the_same_name(self):
        """Content-Disposition до фронта не доходит (его нет в expose-headers),
        поэтому имя собирается на клиенте и обязано совпадать с серверным."""
        source = VIEW_PATH.read_text(encoding='utf-8')
        self.assertIn("link.download = `Посылки ${todayISO().split('-').reverse().join('.')}.xlsx`",
                      source)


class _Cursor:
    """Курсор-двойник: помнит запросы и отдаёт заготовленные строки."""

    def __init__(self, rows=(), total=0):
        self.statements = []
        self._rows = list(rows)
        self._total = total
        self._last = ''

    def execute(self, statement, params=None):
        self._last = ' '.join(str(statement).split())
        self.statements.append((self._last, params))

    def fetchall(self):
        return list(self._rows) if 'SELECT p.id' in self._last else []

    def fetchone(self):
        return [self._total] if 'COUNT(*)' in self._last else None

    def selects(self):
        return [statement for statement, _params in self.statements
                if statement.startswith('SELECT')]


class SelectionTests(unittest.TestCase):
    """Файл обязан содержать то же, что экран, — и весь отбор, а не страницу."""

    def test_page_size_of_the_list_does_not_shrink_the_file(self):
        """У списка потолок 200 строк, и он там намеренный. Уйди выгрузка через
        него — файл обрезался бы молча."""
        cursor = _Cursor()
        parcels_queries.parcels_for_export(cursor, limit=20000)
        select = next(s for s in cursor.selects() if 'SELECT p.id' in s)
        self.assertIn('LIMIT %(limit)s', select)
        self.assertNotIn('OFFSET', select)
        params = cursor.statements[0][1]
        self.assertEqual(params['limit'], 20000)

    def test_export_and_list_filter_by_the_very_same_condition(self):
        """Один `_filter_clause` на список, счётчики и выгрузку."""
        filters = dict(query='771', status=['in_office'], city='Алматы',
                       office_id=17, manager_id=5,
                       date_from=date(2026, 8, 1), date_to=date(2026, 8, 31))

        listing = _Cursor()
        parcels_queries.list_parcels(listing, **filters)
        export = _Cursor()
        parcels_queries.parcels_for_export(export, **filters)

        def condition(statement):
            return statement.split('WHERE', 1)[1].split('ORDER BY', 1)[0].strip()

        self.assertEqual(
            condition(next(s for s in listing.selects() if 'SELECT p.id' in s)),
            condition(next(s for s in export.selects() if 'SELECT p.id' in s)))

    def test_export_keeps_the_order_of_the_screen(self):
        """Человек ждёт в файле те же строки в том же порядке, что видел.

        Порядок сверяется СО СПИСКОМ, а не с литералом: перепиши сортировку в
        одном месте — и файл начал бы приезжать в другом порядке, а тест на
        литерал этого бы не заметил (он бы просто устарел вместе с кодом,
        которого не сторожит).
        """
        def order(cursor):
            select = next(s for s in cursor.selects() if 'SELECT p.id' in s)
            return select.split('ORDER BY', 1)[1].split('LIMIT', 1)[0].strip()

        listing = _Cursor()
        parcels_queries.list_parcels(listing)
        export = _Cursor()
        parcels_queries.parcels_for_export(export)
        self.assertEqual(order(listing), order(export))

    def test_export_reports_how_many_rows_the_registry_really_has(self):
        """Без этого числа нельзя честно сказать, что потолок сработал."""
        cursor = _Cursor(total=4242)
        _items, total = parcels_queries.parcels_for_export(cursor)
        self.assertEqual(total, 4242)


@unittest.skipIf(Flask is None or load_workbook is None, 'flask/openpyxl не установлены')
class RouteTests(unittest.TestCase):
    """Дверь в выгрузку: читатель проходит, отбор разбирается как у списка."""

    def build(self, context=None, rows=(), total=0):
        context = context or {'user_id': 10, 'name': 'Оператор СЗоВ', 'role': 'operator',
                              'department_id': 3, 'department_code': 'szov',
                              'city': 'Алматы', 'headed_department_ids': [],
                              'headed_department_codes': []}
        captured = {}

        def _export(cursor, **filters):
            captured.update(filters)
            return list(rows), total or len(rows)

        cursor = MagicMock()
        db = MagicMock()

        @contextmanager
        def _get_cursor():
            yield cursor

        db._get_cursor = _get_cursor

        def _patch(module, name, value):
            original = getattr(module, name)
            setattr(module, name, value)
            self.addCleanup(setattr, module, name, original)

        _patch(parcels_queries, 'load_access_context', lambda _c, _uid: dict(context))
        _patch(parcels_queries, 'parcels_for_export', _export)
        _patch(parcels_queries, 'read_office', lambda _c, oid, **_kw: {'name': 'Абая'})
        _patch(parcels_queries, 'section_space_ids', lambda _c: [1])
        _patch(parcels_queries, 'list_managers',
               lambda _c: [{'id': 5, 'name': 'Менеджер Тест', 'parcels': 3}])
        _patch(parcels_schema, 'schema_is_ready', lambda _c: True)

        app = Flask(__name__)
        app.register_blueprint(build_parcels_blueprint(
            db=db,
            require_api_key=lambda f: f,
            build_cors_preflight_response=lambda: ('', 204),
            resolve_requester=lambda: (context['user_id'], None, None),
            sensitive_access_granted=lambda _uid: True,
        ))
        app.config['TESTING'] = True
        return app.test_client(), captured

    def book(self, response):
        from io import BytesIO
        return load_workbook(BytesIO(response.data))

    def test_reader_from_szov_may_download(self):
        """Ради оператора СЗоВ реестр и заводили: «сохранить то, что вижу» —
        то же самое, что «посмотреть»."""
        client, _ = self.build(rows=[parcel()])
        response = client.get('/api/parcels/export')
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response.mimetype)

    def test_the_answer_really_carries_the_rows(self):
        """Книга с одной шапкой — тоже 200 и тоже xlsx.

        Поэтому проверяется не код ответа, а содержимое: доехали ли строки до
        листа «Посылки» и те ли это строки.
        """
        client, _ = self.build(rows=[parcel(id=7, description='Синяя коробка'),
                                     parcel(id=8, description='Пакет')])
        sheet = self.book(client.get('/api/parcels/export'))[report.SHEET_PARCELS]
        self.assertEqual(sheet.max_row, 3, 'строки не доехали до листа')
        self.assertEqual([cell(sheet, row, '№').value for row in (2, 3)], [7, 8])
        self.assertEqual(cell(sheet, 2, 'Описание').value, 'Синяя коробка')

    def test_who_collected_the_file_comes_from_the_requester(self):
        """«Собрал» — не декорация: файл с ФИО и телефонами уносят с собой, и по
        нему потом спрашивают, кто его сделал."""
        client, _ = self.build(rows=[parcel()])
        lines = {row[0].value: (row[1].value if len(row) > 1 else None)
                 for row in self.book(client.get('/api/parcels/export'))['Контекст'].iter_rows()}
        self.assertEqual(lines['Собрал'], 'Оператор СЗоВ')

    def test_route_itself_decides_the_file_is_incomplete(self):
        """Признак «попало не всё» считает КОД роута, а не тест.

        Отдаём меньше строк, чем есть в реестре, и ждём предупреждение в книге:
        иначе честность про потолок держалась бы на константе в тесте.
        """
        client, _ = self.build(rows=[parcel()], total=12345)
        lines = {row[0].value: (row[1].value if len(row) > 1 else None)
                 for row in self.book(client.get('/api/parcels/export'))['Контекст'].iter_rows()}
        self.assertIn('В ФАЙЛ ПОПАЛО НЕ ВСЁ', lines)
        self.assertIn('12345', lines['В ФАЙЛ ПОПАЛО НЕ ВСЁ'])
        # И счётчики честно названы «в этом файле», а не выданы за весь реестр.
        self.assertIn('В офисе (в этом файле)', lines)

    def test_a_complete_file_does_not_hedge(self):
        client, _ = self.build(rows=[parcel()], total=1)
        lines = {row[0].value: (row[1].value if len(row) > 1 else None)
                 for row in self.book(client.get('/api/parcels/export'))['Контекст'].iter_rows()}
        self.assertNotIn('В ФАЙЛ ПОПАЛО НЕ ВСЁ', lines)
        self.assertIn('В офисе', lines)

    def test_filename_reaches_the_browser(self):
        client, _ = self.build(rows=[parcel()])
        response = client.get('/api/parcels/export')
        self.assertIn('attachment', response.headers.get('Content-Disposition', ''))

    def test_all_seven_filters_reach_the_query(self):
        """Отбор разбирается тем же кодом, что у списка, — все семь условий."""
        client, captured = self.build(rows=[parcel()])
        client.get('/api/parcels/export?status=in_office&q=771&city=Алматы'
                   '&office_id=17&manager_id=5&date_from=2026-08-01&date_to=2026-08-31')
        self.assertEqual(captured['status'], ['in_office'])
        self.assertEqual(captured['query'], '771')
        self.assertEqual(captured['city'], 'Алматы')
        self.assertEqual(captured['office_id'], 17)
        self.assertEqual(captured['manager_id'], 5)
        self.assertEqual(captured['date_from'], date(2026, 8, 1))
        self.assertEqual(captured['date_to'], date(2026, 8, 31))

    def test_screen_page_does_not_shrink_the_file(self):
        """limit и offset — это страница, а не отбор.

        Пролезь экранный limit в выгрузку — человек, у которого на экране
        полсотни строк, получил бы файл на полсотни строк и считал бы его
        полным. Потолок у выгрузки свой и всегда один.
        """
        client, captured = self.build(rows=[parcel()])
        client.get('/api/parcels/export?limit=5&offset=100')
        self.assertEqual(captured['limit'], report.EXPORT_LIMIT)
        self.assertNotIn('offset', captured)

    def test_unknown_status_is_refused_the_same_way_as_in_the_list(self):
        client, _ = self.build(rows=[parcel()])
        response = client.get('/api/parcels/export?status=выдумка')
        self.assertEqual(response.status_code, 400)
        self.assertIn('Неизвестный статус', response.get_json()['error'])

    def test_section_that_is_not_rolled_out_yet_answers_plainly(self):
        """«Раздел разворачивается» — это не «внутренняя ошибка»."""
        client, _ = self.build(rows=[])
        parcels_schema.schema_is_ready = lambda _c: False
        response = client.get('/api/parcels/export')
        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.get_json()['code'], 'PARCELS_SCHEMA_NOT_READY')

    def test_selection_is_spelled_out_in_the_context_sheet(self):
        """Офис и менеджер разворачиваются в имена: «офис 17» не объясняет ничего."""
        client, _ = self.build(rows=[parcel()])
        response = client.get('/api/parcels/export?city=Алматы&office_id=17'
                              '&manager_id=5&status=in_office&q=771'
                              '&date_from=2026-08-01&date_to=2026-08-31')
        from io import BytesIO
        workbook = load_workbook(BytesIO(response.data))
        note = {row[0].value: (row[1].value if len(row) > 1 else None)
                for row in workbook['Контекст'].iter_rows()}['Отобрано']
        self.assertIn('статус: В офисе', note)
        self.assertIn('город: Алматы', note)
        self.assertIn('офис: Абая', note)
        self.assertIn('менеджер: Менеджер Тест', note)
        self.assertIn('принята с 01.08.2026 по 31.08.2026', note)
        self.assertIn('поиск: «771»', note)


class FrontendTests(unittest.TestCase):
    """Интерфейсные решения выгрузки. Сторожит их python, читая .jsx текстом —
    по коду компонента о них не догадаться, сборка проходит и без них."""

    def setUp(self):
        self.source = VIEW_PATH.read_text(encoding='utf-8')

    def test_button_is_outside_the_write_gate(self):
        """Внутри {canEdit && …} кнопку видели бы одни фронт-офисы, а выгрузка
        нужна прежде всего читателю из СЗоВ."""
        button = self.source.index('Выгрузить')
        gate = self.source.index('{canEdit && (')
        self.assertLess(button, gate,
                        'кнопка выгрузки уехала внутрь гейта записи')

    def test_file_is_fetched_with_the_authorization_header(self):
        """Ссылкой <a href> файл качать нельзя: портал авторизуется заголовком,
        и вместо книги приехала бы страница входа."""
        self.assertIn('/api/parcels/export', self.source)
        self.assertRegex(self.source,
                         r"axios\.get\(`\$\{apiBaseUrl\}/api/parcels/export[^`]*`,\s*\n"
                         r"\s*\{ headers: headers\(\), responseType: 'blob' \}")

    def selection_body(self):
        self.assertIn('const selection = useMemo(', self.source,
                      'отбор перестал собираться одним местом')
        return self.source.split('const selection = useMemo(')[1].split('}, [')[0]

    def test_export_sends_the_same_selection_as_the_list(self):
        """Один источник отбора на список и на выгрузку: собери выгрузка свою
        строку параметров — файл разошёлся бы с экраном."""
        self.assertIn('/api/parcels/export?${selection.toString()}', self.source)
        self.assertIn('const params = new URLSearchParams(selection);', self.source)

    def test_all_seven_conditions_are_inside_the_shared_selection(self):
        """Выпади хоть одно условие из общего отбора — список продолжит работать
        (у него свои `filters` в состоянии), а файл начнёт приезжать шире, чем
        экран, и заметят это не сразу. Поэтому проверяется каждое поимённо."""
        body = self.selection_body()
        for name in ('status', 'q', 'city', 'office_id', 'manager_id',
                     'date_from', 'date_to'):
            self.assertIn("params.set('%s'" % name, body,
                          'условие %s пропало из общего отбора' % name)

    def test_page_is_added_only_for_the_list(self):
        """limit/offset остаются у списка и не попадают в общий отбор."""
        body = self.selection_body()
        self.assertNotIn('limit', body)
        self.assertNotIn('offset', body)

    def test_blob_error_is_unwrapped(self):
        """Ошибку сервер шлёт JSON-ом, а мы просили blob: без разворота в тост
        уехало бы «[object Blob]»."""
        self.assertIn('error?.response?.data?.text?.()', self.source)
        self.assertIn('Не удалось собрать выгрузку', self.source)

    def test_toast_is_held_in_a_ref(self):
        """showToast пересоздаётся на каждом рендере App — известная ловушка."""
        self.assertIn('const toastRef = useRef(showToast);', self.source)
        self.assertIn('toastRef.current?.(', self.source)

    def test_button_is_off_while_the_file_is_being_prepared(self):
        self.assertIn('disabled={downloading || !total}', self.source)
        self.assertIn('Готовим файл…', self.source)

    def test_header_row_wraps_instead_of_breaking_a_button_in_half(self):
        """Третья кнопка не помещается в строку на узком телефоне.

        Замерено 31.08.2026: на 375 px и уже «Добавить посылку» ломалась внутри
        себя на две строки, и ряд получался рваным (40/40/61). Лечится
        переносом РЯДА (`flex-wrap`) плюс запретом переноса ВНУТРИ подписи
        (`whitespace-nowrap`) — по одному ни то, ни другое не работает: без
        wrap кнопка всё равно рвётся, без nowrap перенос ряда не наступает,
        потому что кнопка соглашается сжаться и порвать подпись.
        """
        row = self.source.split('<div className="flex shrink-0')[1].split('>')[0]
        self.assertIn('flex-wrap', row, 'ряд действий перестал переноситься')
        self.assertIn('whitespace-nowrap sm:flex-none', self.source,
                      'подпись главной кнопки снова может порваться пополам')


if __name__ == '__main__':
    unittest.main()
