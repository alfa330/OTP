# -*- coding: utf-8 -*-
"""Выгрузка раздела «Задачи» в Excel: лист на каждую колонку доски.

Книгу собираем настоящую — функцию роута вытаскиваем из bot_schedule2.py через
AST, чтобы не поднимать Flask и БД, и читаем результат openpyxl'ом.
"""

import ast
import collections
import copy
import io
import sys
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tests import source_cache


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
BOT_PATH = ROOT / "bot_schedule2.py"
BOT_SOURCE = BOT_PATH.read_text(encoding="utf-8-sig")
BOT_TREE = source_cache.parse(BOT_SOURCE)
TOP_LEVEL_FUNCTIONS = {
    node.name: node
    for node in BOT_TREE.body
    if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
}
DATABASE_SOURCE = (ROOT / "database.py").read_text(encoding="utf-8-sig")
WORKSPACE_SOURCE = (
    ROOT / "src" / "components" / "tasks" / "TaskBoardWorkspace.jsx"
).read_text(encoding="utf-8-sig")
TASKS_VIEW_SOURCE = (
    ROOT / "src" / "components" / "tasks" / "TasksView.jsx"
).read_text(encoding="utf-8-sig")
BOARD_QUERY_SOURCE = (
    ROOT / "src" / "components" / "tasks" / "boardQuery.js"
).read_text(encoding="utf-8-sig")

NOW = datetime(2026, 8, 12, 15, 0, 0)


def _load_export_function(namespace):
    """Роут выгрузки + его helper'ы и константы, без импорта bot_schedule2."""
    pending = ["export_tasks_excel"]
    selected = {}
    while pending:
        name = pending.pop()
        if name in selected or name in namespace:
            continue
        node = TOP_LEVEL_FUNCTIONS.get(name)
        if node is None:
            continue
        selected[name] = node
        for child in ast.walk(node):
            if isinstance(child, ast.Call) and isinstance(child.func, ast.Name):
                pending.append(child.func.id)

    nodes = []
    for original in BOT_TREE.body:
        # Константы выгрузки берём из исходника, а не переписываем в тест: иначе
        # тест перестал бы замечать, что состав колонок разъехался с кодом.
        if isinstance(original, ast.Assign) and all(
            isinstance(target, ast.Name)
            and (
                target.id.lstrip('_').startswith('TASK_EXPORT')
                or target.id in ('TASK_TAG_LABELS', 'TASK_PRIORITY_LABELS')
            )
            for target in original.targets
        ):
            nodes.append(copy.deepcopy(original))
        elif (
            isinstance(original, (ast.FunctionDef, ast.AsyncFunctionDef))
            and original.name in selected
        ):
            cloned = copy.deepcopy(original)
            cloned.decorator_list = []
            nodes.append(cloned)

    module = ast.Module(body=nodes, type_ignores=[])
    ast.fix_missing_locations(module)
    result = dict(namespace)
    exec(compile(module, str(BOT_PATH), "exec"), result)
    return result


class _FakeDb:
    def __init__(self, tasks):
        self.tasks = tasks
        self.calls = []

    def get_tasks_for_export(self, **kwargs):
        self.calls.append(kwargs)
        return self.tasks


class _FakeArgs:
    def __init__(self, values):
        self.values = values

    def get(self, key, default=None):
        return self.values.get(key, default)


class _FakeRequest:
    def __init__(self, values=None):
        self.args = _FakeArgs(values or {})


class _FakeG:
    effective_task_role = 'admin'


class _FakeDatetime(datetime):
    """Фиксируем «сейчас»: от него зависит подсветка просрочки и имя файла."""

    @classmethod
    def now(cls, tz=None):
        return NOW


class _QuietLogging:
    @staticmethod
    def exception(*_args, **_kwargs):
        return None


def _task(**overrides):
    task = {
        "id": 1,
        "subject": "Тема",
        "description": "Описание",
        "tag": "task",
        "status": "assigned",
        "priority": "normal",
        "is_backlog": False,
        "created_at": datetime(2026, 8, 1, 10, 0),
        "planned_start_at": None,
        "started_at": None,
        "due_at": None,
        "completed_at": None,
        "completion_summary": None,
        "estimate_minutes": None,
        "spent_minutes": None,
        "assignee_name": "Ядигаров Руслан",
        "creator_name": "Алчинбаева Анель",
        "requested_by_name": "Алчинбаева Анель",
    }
    task.update(overrides)
    return task


class TaskExcelExportTests(unittest.TestCase):
    def _export(self, tasks, params=None, guard=None):
        db = _FakeDb(tasks)
        sent = {}

        def send_file(stream, **kwargs):
            sent["content"] = stream.getvalue()
            sent["kwargs"] = kwargs
            return sent

        namespace = _load_export_function({
            "_task_route_guard": lambda: (guard or (2, (2, '', '', 'admin'), None, None)),
            "db": db,
            "g": _FakeG(),
            "request": _FakeRequest(params),
            "jsonify": lambda value: value,
            "logging": _QuietLogging,
            "collections": collections,
            "Workbook": Workbook,
            "Font": Font,
            "PatternFill": PatternFill,
            "Alignment": Alignment,
            "get_column_letter": get_column_letter,
            "BytesIO": io.BytesIO,
            "datetime": _FakeDatetime,
            "send_file": send_file,
        })
        response = namespace["export_tasks_excel"]()
        if response is not sent:
            return db, response, None
        return db, sent, load_workbook(io.BytesIO(sent["content"]))

    @staticmethod
    def _headers(sheet):
        return [cell.value for cell in sheet[1]]

    def test_sheet_per_board_column_in_board_order(self):
        _db, _sent, book = self._export([
            _task(id=1, status='accepted'),
            _task(id=2, status='in_progress'),
            _task(id=3, is_backlog=True),
            _task(id=4, status='completed'),
            _task(id=5, status='assigned'),
            # returned живёт в той же колонке, что и in_progress.
            _task(id=6, status='returned'),
        ])
        self.assertEqual(
            book.sheetnames,
            ['Бэклог', 'К выполнению', 'В работе', 'На проверке', 'Готово'],
        )
        self.assertEqual(
            [row[0] for row in book['В работе'].iter_rows(min_row=2, max_col=1, values_only=True)],
            [2, 6],
        )

    def test_backlog_flag_wins_over_status(self):
        """Бэклог на доске — это backlog=only поверх любого статуса."""
        _db, _sent, book = self._export([_task(id=7, is_backlog=True, status='in_progress')])
        self.assertEqual(book.sheetnames, ['Бэклог'])

    def test_empty_columns_do_not_become_sheets(self):
        _db, _sent, book = self._export([_task(id=1, status='accepted')])
        self.assertEqual(book.sheetnames, ['Готово'])

    def test_no_tasks_still_produces_one_sheet(self):
        _db, _sent, book = self._export([])
        self.assertEqual(book.sheetnames, ['Задачи'])
        self.assertEqual(book['Задачи'].max_row, 1)

    def test_columns_are_scoped_to_the_sheet(self):
        closed = dict(
            started_at=datetime(2026, 8, 2, 9, 0),
            completed_at=datetime(2026, 8, 3, 9, 0),
            completion_summary='Готово',
            planned_start_at=datetime(2026, 8, 1, 9, 0),
            due_at=datetime(2026, 8, 4, 9, 0),
            estimate_minutes=60,
            spent_minutes=60,
        )
        _db, _sent, book = self._export([
            _task(id=1, status='in_progress', **closed),
            _task(id=2, status='accepted', **closed),
            _task(id=3, is_backlog=True, **closed),
        ])
        progress = self._headers(book['В работе'])
        done = self._headers(book['Готово'])
        backlog = self._headers(book['Бэклог'])
        # Пустая в каждой строке колонка — это шум, поэтому «Завершена»/«Итог»
        # есть только там, где задача уже закрыта, а «Факт» — где работа шла.
        self.assertNotIn('Завершена', progress)
        self.assertNotIn('Итог', progress)
        self.assertIn('Завершена', done)
        self.assertIn('Итог', done)
        self.assertNotIn('Факт, ч', backlog)
        self.assertNotIn('Плановый старт', backlog)
        for headers in (progress, done, backlog):
            self.assertEqual(headers[:2], ['№', 'Тема'])
            self.assertIn('Исполнители', headers)
            self.assertIn('Поручил', headers)

    def test_column_empty_in_every_row_is_dropped(self):
        """Заголовок без единого значения — шум: лист приходится листать вбок."""
        _db, _sent, book = self._export([
            _task(id=1, status='in_progress', estimate_minutes=60),
            _task(id=2, status='in_progress'),
        ])
        headers = self._headers(book['В работе'])
        self.assertIn('Оценка, ч', headers)
        # Плановый старт и факт не проставлены ни у одной задачи листа.
        self.assertNotIn('Плановый старт', headers)
        self.assertNotIn('Факт, ч', headers)
        self.assertNotIn('Дедлайн', headers)

    def test_dates_are_real_dates_and_hours_are_numbers(self):
        _db, _sent, book = self._export([_task(
            id=1,
            status='in_progress',
            created_at=datetime(2026, 8, 1, 10, 0),
            started_at=datetime(2026, 8, 10, 14, 59),
            due_at=datetime(2026, 8, 20, 16, 30),
            estimate_minutes=90,
            spent_minutes=210,
        )])
        sheet = book['В работе']
        headers = self._headers(sheet)
        row = {title: sheet.cell(row=2, column=idx + 1) for idx, title in enumerate(headers)}
        self.assertEqual(row['Дедлайн'].value, datetime(2026, 8, 20, 16, 30))
        self.assertEqual(row['Начата'].value, datetime(2026, 8, 10, 14, 59))
        self.assertIn('DD.MM.YYYY', row['Создана'].number_format)
        # Часы, а не «1ч 30м»: по колонке в Excel считают суммы.
        self.assertEqual(row['Оценка, ч'].value, 1.5)
        self.assertEqual(row['Факт, ч'].value, 3.5)

    def test_labels_are_russian(self):
        _db, _sent, book = self._export([
            _task(id=1, status='assigned', tag='problem', priority='critical'),
        ])
        sheet = book['К выполнению']
        values = [cell.value for cell in sheet[2]]
        self.assertIn('Проблема', values)
        self.assertIn('Критичная', values)

    @staticmethod
    def _font_hex(cell):
        color = cell.font.color
        rgb = getattr(color, 'rgb', None) if color is not None else None
        return str(rgb)[-6:].upper() if isinstance(rgb, str) else None

    def _due_cell(self, sheet, row):
        return sheet.cell(row=row, column=self._headers(sheet).index('Дедлайн') + 1)

    def test_only_overdue_deadline_is_coloured(self):
        overdue = NOW - timedelta(days=2)
        _db, _sent, book = self._export([
            _task(id=1, status='in_progress', due_at=overdue),
            _task(id=2, status='in_progress', due_at=NOW + timedelta(days=2)),
            # Закрытую задачу просрочка уже не касается — так же считает сводка.
            _task(id=3, status='accepted', due_at=overdue),
        ])
        progress = book['В работе']
        self.assertEqual(self._font_hex(self._due_cell(progress, 2)), 'B91C1C')
        self.assertNotEqual(self._font_hex(self._due_cell(progress, 3)), 'B91C1C')
        self.assertNotEqual(self._font_hex(self._due_cell(book['Готово'], 2)), 'B91C1C')

    def test_header_is_frozen_and_filterable(self):
        _db, _sent, book = self._export([_task(id=1, status='accepted')])
        sheet = book['Готово']
        self.assertEqual(sheet.freeze_panes, 'A2')
        self.assertTrue(sheet.auto_filter.ref.startswith('A1:'))

    def test_scope_goes_to_the_database_untouched(self):
        db, _sent, _book = self._export(
            [_task(id=1)],
            params={'mine': 'assignee', 'person_id': '300', 'person_scope': 'any'},
        )
        self.assertEqual(len(db.calls), 1)
        call = db.calls[0]
        self.assertEqual(call['mine'], 'assignee')
        self.assertEqual(call['person_id'], '300')
        self.assertEqual(call['person_scope'], 'any')
        self.assertEqual(call['requester_id'], 2)

    def test_guard_rejection_is_returned_as_is(self):
        _db, response, book = self._export(
            [_task(id=1)],
            guard=(None, None, {"error": "Only admin, sv, trainer and department heads can access tasks"}, 403),
        )
        self.assertIsNone(book)
        self.assertEqual(response[1], 403)

    def test_download_name_is_dated_xlsx(self):
        _db, sent, _book = self._export([_task(id=1)])
        self.assertEqual(sent["kwargs"]["download_name"], 'Задачи_2026-08-12.xlsx')
        self.assertTrue(sent["kwargs"]["as_attachment"])


class TaskExportDbLayerTests(unittest.TestCase):
    def setUp(self):
        self.src = DATABASE_SOURCE

    def test_export_reuses_the_same_visibility_filter_as_the_list(self):
        # Разъедься эти два места — в файл уехало бы больше, чем видно на доске.
        self.assertIn("def _task_scope_filters(self", self.src)
        export = self.src[self.src.index("def get_tasks_for_export("):]
        export = export[:export.index("def update_task_status(")]
        self.assertIn("self._task_scope_filters(", export)
        listing = self.src[self.src.index("def get_tasks_for_requester("):]
        listing = listing[:listing.index("def get_tasks_for_export(")]
        self.assertIn("self._task_scope_filters(", listing)

    def test_export_has_its_own_higher_cap(self):
        self.assertIn("TASKS_EXPORT_MAX_ROWS = 5000", self.src)
        export = self.src[self.src.index("def get_tasks_for_export("):]
        export = export[:export.index("def update_task_status(")]
        self.assertIn("self.TASKS_EXPORT_MAX_ROWS", export)
        self.assertIn("LIMIT %s", export)

    def test_export_query_stays_flat(self):
        """Ни истории, ни вложений, ни чек-листа — иначе это четыре лишних запроса."""
        export = self.src[self.src.index("def get_tasks_for_export("):]
        export = export[:export.index("def update_task_status(")]
        for table in ("task_status_history", "task_attachments", "task_checklist_items"):
            self.assertNotIn(table, export)
        # Трудозатраты — агрегатом в том же запросе, а не журналом отчётов.
        self.assertIn("SUM(r.spent_minutes)", export)

    def test_export_order_matches_the_board(self):
        export = self.src[self.src.index("def get_tasks_for_export("):]
        export = export[:export.index("def update_task_status(")]
        self.assertIn("t.is_backlog DESC", export)
        self.assertIn("t.backlog_rank END ASC NULLS LAST", export)
        self.assertIn("t.created_at DESC, t.id DESC", export)


class TaskExportApiTests(unittest.TestCase):
    def test_route_registered(self):
        self.assertIn("@app.route('/api/tasks/export', methods=['GET', 'OPTIONS'])", BOT_SOURCE)
        self.assertIn("def export_tasks_excel():", BOT_SOURCE)

    def test_route_is_behind_the_tasks_guard(self):
        route = BOT_SOURCE[BOT_SOURCE.index("def export_tasks_excel():"):]
        route = route[:route.index("\n@app.route")]
        self.assertIn("_task_route_guard()", route)
        self.assertIn("effective_task_role", route)


class TaskExportFrontendTests(unittest.TestCase):
    def test_scope_is_described_once_for_board_and_export(self):
        self.assertIn("export const scopeQueryParams = (scope, departmentId = null)", BOARD_QUERY_SOURCE)
        board_params = BOARD_QUERY_SOURCE[BOARD_QUERY_SOURCE.index("export const boardQueryParams"):]
        self.assertIn("scopeQueryParams(scope, departmentId)", board_params)
        self.assertIn("scopeQueryParams(scope, departmentId)", TASKS_VIEW_SOURCE)

    def test_download_goes_through_axios_with_headers(self):
        # Ссылка <a href> не несёт bearer-токен и упёрлась бы в 401.
        handler = TASKS_VIEW_SOURCE[TASKS_VIEW_SOURCE.index("const exportBoardTasks"):]
        handler = handler[:handler.index("useEffect(")]
        self.assertIn("/api/tasks/export", handler)
        self.assertIn("responseType: 'blob'", handler)
        self.assertIn("buildHeaders()", handler)
        self.assertIn(".xlsx", handler)
        self.assertIn("onExport={exportBoardTasks}", TASKS_VIEW_SOURCE)

    def test_button_exports_the_selected_scope(self):
        handler = WORKSPACE_SOURCE[WORKSPACE_SOURCE.index("const handleExport"):]
        handler = handler[:handler.index("}, [onExport")]
        self.assertIn("onExport({ scope, departmentId })", handler)
        self.assertIn("setIsExporting(true)", handler)
        self.assertIn("Выгрузка", WORKSPACE_SOURCE)
        self.assertIn("disabled={isExporting}", WORKSPACE_SOURCE)


if __name__ == "__main__":
    unittest.main()
