# -*- coding: utf-8 -*-
"""Вкладка «Аналитика»: гейт, честный счёт и запись поисковых запросов.

Что здесь важно проверить и почему.

1. ГЕЙТ. Вкладку открывает тот, кто ВЕДЁТ статьи: can_create | can_edit |
   can_publish (плюс отдельно администратор доступов). Читателю она закрыта —
   там видно, кто что читал и кто просрочил ознакомление, поимённо. Гард во
   фронте вкладку прячет, но не защищает запрос: проверка на сервере.

2. ГРАНИЦА ОТДЕЛА — ОТДЕЛЬНАЯ ВЕЩЬ, И ЕЁ ЛЕГКО ПОТЕРЯТЬ. Способность редактора
   отделов не знает: без границы супервайзер СЗоВ с правом публикации увидел бы
   поимённо, кто в отделе продаж просрочил ознакомление. Проверяем, что в
   поимённый запрос уходит список отделов, а у супер-админа и администратора
   доступов — None (границы нет по построению).

   Отдельно закреплено то, ради чего гейт и меняли: пока вкладка гейтилась одной
   способностью can_manage_access, обе ветки _departments возвращали None и
   границы не было вовсе — она недостижима правилом раздела (нет в
   PERMISSION_COLUMNS) и снята у роли admin. Тест на границу тогда проходил бы
   вхолостую, сертифицируя несуществующую защиту.

3. ПРОЧТЕНИЕ ≠ СТРОКА ЖУРНАЛА. Просмотр пишется на каждый GET статьи, включая
   обновление страницы. Дедупликация проверяется НАСТОЯЩИМ запросом на
   синтетических данных: в PostgreSQL CTE перекрывает одноимённую таблицу,
   поэтому заглушки подставляются перед боевым текстом, и исполняется он сам.

4. ОТМЕНЁННОЕ НАЗНАЧЕНИЕ — НЕ ПРОСРОЧКА. У отменённого ознакомления остаются и
   срок, и отсутствие подписи. Без фильтра по 'cancelled' оно вечно висело бы в
   письме супервайзеру. Остальной раздел фильтрует именно так (wiki/ack.py),
   и разъезжаться этим правилам нельзя.

5. ВЫГРУЗКА — ТА ЖЕ ДВЕРЬ, ЧТО И ЭКРАН. Гейт и граница отдела у /analytics и
   у /analytics/export обязаны совпадать: файл уносят с собой, и поимённый
   список чужого отдела в нём хуже, чем на экране. Отличаться выгрузка вправе
   ровно потолком строк — и проверяется, что экранный `limit` из строки запроса
   на неё не действует, иначе файл обрезался бы ровно там, где начинается то,
   ради чего его просили.

6. ЗАПРОСЫ-ОГРЫЗКИ. Поле поиска ищет по мере набора, поэтому одна фраза
   приезжает шестью запросами. Свёртка проверяется на уровне питона: какие
   параметры уходят в запрос и какой запрос выбирается.
"""

import sys
import unittest
from contextlib import contextmanager
from datetime import datetime
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from flask import Flask
except ImportError:  # pragma: no cover
    Flask = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

from tests import prod_db  # noqa: E402

from wiki import analytics as wiki_analytics  # noqa: E402
from wiki import analytics_report  # noqa: E402
from wiki import routes_analytics  # noqa: E402
from wiki import perimeter as wiki_perimeter  # noqa: E402
from wiki import queries  # noqa: E402
from wiki import search as wiki_search  # noqa: E402
from wiki.access import collect_subjects  # noqa: E402
from wiki.routes import build_wiki_blueprint  # noqa: E402


def make_context(otp_role='operator', department_id=None, wiki_role=True, **caps):
    """Контекст доступа.

    wiki_role=False — способности выводятся из ДОЛЖНОСТИ. Это не украшение:
    непустой список ролей вики ЗАМЕЩАЕТ роль OTP, а не дополняет её, поэтому
    супер-админ с ролью «читатель» теряет мастер-ключ. Проверять права
    супер-админа с приклеенной ролью значило бы проверять не то.
    """
    role = {'id': 5, 'code': 'wiki_role', 'can_read': True, 'can_create': False,
            'can_edit': False, 'can_delete': False, 'can_publish': False,
            'can_approve': False, 'can_manage_users': False,
            'can_manage_structure': False, 'can_manage_access': False}
    role.update(caps)
    return {
        'user_id': 42, 'otp_role': otp_role, 'department_id': department_id,
        'direction_id': None, 'headed_department_ids': [], 'group_ids': [],
        'wiki_roles': [role] if wiki_role else [], 'access_mode': 'auto',
    }


# Сколько значений отдаёт fetchone на каждый запрос отчёта, в порядке вызова:
# итоги чтения (6), счётчик устаревших (1), итоги ознакомлений (6), итоги
# поиска (4), дата начала журнала (1), итоги помощника (5). Считать их «как
# получится» нельзя — распаковка кортежа не той длины падает, и падает она уже
# внутри роута.
REPORT_FETCHONE = [(0,) * 6, (0,), (0,) * 6, (0,) * 4, (None,), (0,) * 5]


def build_client(test, context, *, fetchone=None, fetchall=()):
    cursor = MagicMock()
    cursor.calls = []

    def execute(sql, params=None):
        cursor.calls.append((sql, params))

    cursor.execute.side_effect = execute
    if fetchone is None:
        fetchone = list(REPORT_FETCHONE)
    if isinstance(fetchone, list):
        # Кортежи разной длины по порядку вызова; когда список кончился,
        # отдаём последний — так тест не разваливается от лишнего запроса.
        queue = list(fetchone)
        cursor.fetchone.side_effect = lambda: queue.pop(0) if len(queue) > 1 else queue[0]
    else:
        cursor.fetchone.return_value = fetchone
    cursor.fetchall.return_value = list(fetchall)
    db = MagicMock()

    @contextmanager
    def _get_cursor():
        yield cursor

    db._get_cursor = _get_cursor

    def fake_perimeter(_cursor, _ctx, **_kwargs):
        return collect_subjects(user_id=42, otp_role='operator'), {3}, {1, 2}

    patches = [
        (queries, 'load_access_context', lambda _c, _u: dict(context)),
        (queries, 'granted_rule_rights', lambda _c, _s, _u: ({}, [])),
        (queries, 'log_action', lambda *a, **k: None),
        (wiki_perimeter, 'read_perimeter', fake_perimeter),
    ]
    for module, name, replacement in patches:
        original = getattr(module, name)
        setattr(module, name, replacement)
        test.addCleanup(setattr, module, name, original)

    app = Flask(__name__)
    app.register_blueprint(build_wiki_blueprint(
        db=db, require_api_key=lambda f: f,
        build_cors_preflight_response=lambda: ('', 204),
        resolve_requester=lambda: (42, None, None),
        sensitive_access_granted=lambda _user_id, cursor=None: True,
        client_ip=lambda: '127.0.0.1',
        gcs={'signed_url': lambda *a, **k: 'https://x'},
    ))
    app.config['TESTING'] = True
    return app.test_client(), cursor


@unittest.skipIf(Flask is None, 'flask не установлен')
class GateTest(unittest.TestCase):

    def test_reader_is_refused(self):
        """Читателю вкладка закрыта: там люди поимённо."""
        client, _ = build_client(self, make_context())
        response = client.get('/api/wiki/analytics')
        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.get_json()['code'], 'WIKI_EDITOR_ONLY')

    def test_every_editing_right_opens_the_tab(self):
        """Любое из трёх прав правки открывает вкладку — по отдельности."""
        for right in ('can_create', 'can_edit', 'can_publish'):
            client, _ = build_client(self, make_context(**{right: True}))
            response = client.get('/api/wiki/analytics')
            self.assertEqual(response.status_code, 200, right)

    def test_access_manager_without_editing_still_sees_report(self):
        """У администратора доступов вкладка была и остаётся."""
        client, _ = build_client(self, make_context(can_manage_access=True))
        response = client.get('/api/wiki/analytics')
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        for block in ('reading', 'content', 'acknowledgements', 'demand',
                      'notes', 'period'):
            self.assertIn(block, payload)

    def test_period_reaches_every_query(self):
        """Пустые границы уходят как None, заполненные — как есть."""
        client, cursor = build_client(self, make_context(can_manage_access=True))
        client.get('/api/wiki/analytics?since=2026-08-01&until=2026-08-24')
        dated = [p for _sql, p in cursor.calls
                 if isinstance(p, dict) and 'since' in p]
        self.assertTrue(dated, 'ни один запрос не получил период')
        for params in dated:
            self.assertEqual(params['since'], '2026-08-01')
            self.assertEqual(params['until'], '2026-08-24')


@unittest.skipIf(Flask is None, 'flask не установлен')
class DepartmentBoundaryTest(unittest.TestCase):
    """Способность редактора отделов не знает — границу ставим отдельно."""

    def _depts(self, context):
        client, cursor = build_client(self, context)
        response = client.get('/api/wiki/analytics')
        self.assertEqual(response.status_code, 200)
        with_depts = [p for _sql, p in cursor.calls
                      if isinstance(p, dict) and 'depts' in p]
        self.assertTrue(with_depts, 'поимённый список не получил границу отдела')
        return response.get_json(), with_depts[0]['depts']

    def test_supervisor_editor_is_limited_to_own_department(self):
        payload, depts = self._depts(
            make_context(otp_role='sv', department_id=7, can_publish=True))
        self.assertEqual(depts, [7])
        self.assertTrue(payload['scoped'])
        # Оговорка приходит ключом, а не строкой в общем списке: фронт ставит
        # её к поимённому списку, а не подвалом страницы.
        self.assertIn('только по вашим отделам', payload['notes'].get('scoped', ''),
                      'человеку не сказали, что список сужен')

    def test_department_head_sees_headed_departments(self):
        context = make_context(otp_role='admin', department_id=7, can_edit=True)
        context['headed_department_ids'] = [7, 9]
        _payload, depts = self._depts(context)
        self.assertEqual(depts, [7, 9])

    def test_super_admin_has_no_boundary(self):
        payload, depts = self._depts(
            make_context(otp_role='super_admin', department_id=7, wiki_role=False))
        self.assertIsNone(depts)
        self.assertFalse(payload['scoped'])
        self.assertNotIn('scoped', payload['notes'],
                         'без границы оговорки о сужении быть не должно')

    def test_access_manager_has_no_boundary_either(self):
        _payload, depts = self._depts(
            make_context(otp_role='sv', department_id=7, can_manage_access=True))
        self.assertIsNone(depts)


class SearchLogTest(unittest.TestCase):
    """Запись поискового запроса: маска, нормализация, свёртка, короткий запрос."""

    def _capture(self, **kwargs):
        cursor = MagicMock()
        cursor.calls = []
        cursor.execute.side_effect = lambda sql, params=None: cursor.calls.append((sql, params))
        cursor.fetchone.return_value = None      # свернуть не во что
        written = wiki_search.log_query(cursor, **kwargs)
        return written, cursor.calls

    def test_short_query_is_not_written(self):
        written, calls = self._capture(user_id=1, query='а', results_count=0,
                                       perimeter_size=10)
        self.assertFalse(written)
        self.assertEqual(calls, [])

    def test_digits_are_masked_before_writing(self):
        _written, calls = self._capture(user_id=1, query='телефон 87051234567 водителя',
                                        results_count=0, perimeter_size=10)
        params = calls[-1][1]
        self.assertEqual(params['query'], 'телефон # водителя')
        self.assertNotIn('87051234567', params['query'])

    def test_short_numbers_survive(self):
        """Маска бьёт по телефонам и ИИН, а не по любому числу: «форма 2» нужна."""
        _written, calls = self._capture(user_id=1, query='приказ 2024 года',
                                        results_count=3, perimeter_size=10)
        self.assertEqual(calls[-1][1]['query'], 'приказ 2024 года')

    def test_kazakh_letters_are_folded_for_the_report(self):
        _written, calls = self._capture(user_id=1, query='Қазына', results_count=0,
                                        perimeter_size=10)
        params = calls[-1][1]
        self.assertEqual(params['query'], 'Қазына')
        self.assertEqual(params['norm'], 'казына')

    def test_prefix_is_collapsed_instead_of_a_new_row(self):
        """Нашлась строка-огрызок — переписываем её, а не плодим вторую."""
        cursor = MagicMock()
        cursor.calls = []
        cursor.execute.side_effect = lambda sql, params=None: cursor.calls.append((sql, params))
        cursor.fetchone.return_value = (77,)     # UPDATE нашёл, что свернуть
        self.assertTrue(wiki_search.log_query(
            cursor, user_id=1, query='как оформить самозанятость',
            results_count=0, perimeter_size=10))
        self.assertEqual(len(cursor.calls), 1, 'после свёртки INSERT не нужен')
        self.assertIn('UPDATE wiki_search_log', cursor.calls[0][0])

    def test_anonymous_query_is_never_collapsed(self):
        """Без владельца «предыдущая строка за 30 секунд» склеила бы разных людей."""
        _written, calls = self._capture(user_id=None, query='отпуск',
                                        results_count=1, perimeter_size=10)
        self.assertEqual(len(calls), 1)
        self.assertIn('INSERT INTO wiki_search_log', calls[0][0])

    def test_perimeter_is_written_next_to_the_result_count(self):
        """Ноль находок без размера периметра неинтерпретируем."""
        _written, calls = self._capture(user_id=1, query='самозанятость',
                                        results_count=0, perimeter_size=16)
        params = calls[-1][1]
        self.assertEqual(params['found'], 0)
        self.assertEqual(params['perimeter'], 16)


# ── Настоящий SQL на синтетических данных ────────────────────────────────────
#
# Приём тот же, что в tests/test_wiki_ai_perimeter.py: в PostgreSQL CTE
# перекрывает одноимённую таблицу, поэтому заглушки подставляются ПЕРЕД боевым
# текстом запроса, и исполняется он сам, без правок. Соединение read only.

def _stub(sql, stubs):
    """Вклеить заглушки в начало боевого запроса.

    Боевые запросы блока чтения уже начинаются с «WITH reads AS», поэтому
    заглушки дописываются в тот же WITH, а не вторым.
    """
    body = sql.lstrip()
    if body.startswith('WITH '):
        return 'WITH ' + stubs + ',\n' + body[len('WITH '):]
    return 'WITH ' + stubs + '\n' + body


_VIEWS_STUB = """
wiki_article_views_log AS (
    SELECT article_id::int, user_id::int, viewed_at::timestamp,
           snapshot_department_id::int
      FROM (VALUES {rows}) AS t(article_id, user_id, viewed_at, snapshot_department_id)
),
users AS (
    SELECT id::int, department_id::int, status::text
      FROM (VALUES {people}) AS t(id, department_id, status)
),
departments AS (
    SELECT id::int, name::text FROM (VALUES (1, 'СЗоВ')) AS t(id, name)
),
wiki_articles AS (
    SELECT id::int, slug::text, title::text, status::text, updated_at::timestamp
      FROM (VALUES {articles}) AS t(id, slug, title, status, updated_at)
)
"""


@unittest.skipIf(not prod_db.available(), 'нет доступа к базе для прогона SQL')
class ReadDeduplicationTest(unittest.TestCase):
    """Обновление страницы не должно превращаться в три прочтения."""

    @classmethod
    def setUpClass(cls):
        reason = prod_db.skip_reason()
        if reason:
            raise unittest.SkipTest(reason)
        cls.conn = prod_db.connection()

    def tearDown(self):
        prod_db.rollback()

    def _totals(self, rows, articles=None):
        stubs = _VIEWS_STUB.format(
            rows=', '.join(rows),
            people="(5, 1, 'working'), (6, 1, 'fired')",
            articles=articles
            or "(1, 'a', 'Статья', 'published', '2026-08-01'::timestamp)")
        sql = _stub(wiki_analytics._TOTALS_SQL, stubs)
        with self.conn.cursor() as cursor:
            cursor.execute(sql, {'visible': [1], 'since': None, 'until': None})
            return cursor.fetchone()

    def test_untouched_published_counts_whole(self):
        """Нетронутое считается ЦЕЛИКОМ, а не по длине урезанного списка.

        Плитка «Статей без чтений» брала длину списка, а список режется
        потолком строк: при полусотне нетронутых статей она показывала бы
        ровно столько, сколько влезло в таблицу.
        """
        rows = ["(1, 5, '2026-08-10 10:00:05'::timestamp, 1)"]
        articles = ("(1, 'a', 'Читали', 'published', '2026-08-01'::timestamp), "
                    "(2, 'b', 'Не читали', 'published', '2026-08-01'::timestamp), "
                    "(3, 'c', 'Черновик', 'draft', '2026-08-01'::timestamp)")
        with self.conn.cursor() as cursor:
            cursor.execute(_stub(wiki_analytics._TOTALS_SQL, _VIEWS_STUB.format(
                rows=', '.join(rows), people="(5, 1, 'working')",
                articles=articles)),
                {'visible': [1, 2, 3], 'since': None, 'until': None})
            row = cursor.fetchone()
        published, published_read = row[4], row[5]
        self.assertEqual(published, 2, 'черновик не опубликован')
        self.assertEqual(published_read, 1)
        # Ровно эту разность роут кладёт в totals['unread'].
        self.assertEqual(published - published_read, 1,
                         'нетронутая опубликованная статья ровно одна')

    def test_three_opens_in_one_minute_are_one_read(self):
        rows = ["(1, 5, '2026-08-10 10:00:05'::timestamp, 1)",
                "(1, 5, '2026-08-10 10:00:31'::timestamp, 1)",
                "(1, 5, '2026-08-10 10:00:59'::timestamp, 1)"]
        reads, readers, articles_read, opens, published, _read_pub = self._totals(rows)
        self.assertEqual(opens, 3, 'сырые открытия считаются как есть')
        self.assertEqual(reads, 1, 'обновление страницы — не второе прочтение')
        self.assertEqual(readers, 1)
        self.assertEqual(articles_read, 1)
        self.assertEqual(published, 1)

    def test_next_minute_is_a_second_read(self):
        rows = ["(1, 5, '2026-08-10 10:00:05'::timestamp, 1)",
                "(1, 5, '2026-08-10 10:01:05'::timestamp, 1)"]
        reads, _readers, _articles, opens, _published, _read_pub = self._totals(rows)
        self.assertEqual((opens, reads), (2, 2))

    def test_different_people_are_different_reads(self):
        rows = ["(1, 5, '2026-08-10 10:00:05'::timestamp, 1)",
                "(1, 6, '2026-08-10 10:00:07'::timestamp, 1)"]
        reads, readers, _articles, _opens, _published, _read_pub = self._totals(rows)
        self.assertEqual((reads, readers), (2, 2))

    def test_period_cuts_by_almaty_day_inclusive(self):
        """Верхняя граница включает ВЕСЬ день «по», а не полночь."""
        rows = ["(1, 5, '2026-08-10 23:50:00'::timestamp, 1)"]
        stubs = _VIEWS_STUB.format(
            rows=', '.join(rows),
            people="(5, 1, 'working')",
            articles="(1, 'a', 'Статья', 'published', '2026-08-01'::timestamp)")
        sql = _stub(wiki_analytics._TOTALS_SQL, stubs)
        with self.conn.cursor() as cursor:
            cursor.execute(sql, {'visible': [1], 'since': '2026-08-10',
                                 'until': '2026-08-10'})
            self.assertEqual(cursor.fetchone()[0], 1)


@unittest.skipIf(not prod_db.available(), 'нет доступа к базе для прогона SQL')
class HeadcountTest(unittest.TestCase):
    """Уволенные не должны занижать охват отдела."""

    @classmethod
    def setUpClass(cls):
        reason = prod_db.skip_reason()
        if reason:
            raise unittest.SkipTest(reason)
        cls.conn = prod_db.connection()

    def tearDown(self):
        prod_db.rollback()

    def test_fired_are_out_of_the_denominator(self):
        stubs = _VIEWS_STUB.format(
            rows="(1, 5, '2026-08-10 10:00:00'::timestamp, 1)",
            people=("(5, 1, 'working'), (6, 1, 'fired'), (7, 1, 'annual_leave'), "
                    "(8, 1, 'dismissal')"),
            articles="(1, 'a', 'Статья', 'published', '2026-08-01'::timestamp)")
        sql = _stub(wiki_analytics._BY_DEPARTMENT_SQL, stubs)
        with self.conn.cursor() as cursor:
            cursor.execute(sql, {'visible': [1], 'since': None, 'until': None})
            row = cursor.fetchone()
        # Штат: работающий и отпускник. Уволенный и уволившийся — нет.
        self.assertEqual(row[5], 2)
        self.assertEqual(row[3], 1, 'читатель один')


_ROSTER_STUB = """
wiki_article_views_log AS (
    SELECT article_id::int, user_id::int, viewed_at::timestamp,
           snapshot_department_id::int
      FROM (VALUES {rows}) AS t(article_id, user_id, viewed_at,
                                snapshot_department_id)
),
users AS (
    SELECT id::int, department_id::int, name::text
      FROM (VALUES (5, 1, 'Ахметова Асель'), (6, 2, 'Ким Владислав')) AS t(id, department_id, name)
),
departments AS (
    SELECT id::int, name::text
      FROM (VALUES (1, 'СЗоВ'), (2, 'Отдел продаж')) AS t(id, name)
)
"""


@unittest.skipIf(not prod_db.available(), 'нет доступа к базе для прогона SQL')
class ReadersRosterTest(unittest.TestCase):
    """Перепись читателей: человек одной строкой, отдел — последний из его чтений."""

    @classmethod
    def setUpClass(cls):
        reason = prod_db.skip_reason()
        if reason:
            raise unittest.SkipTest(reason)
        cls.conn = prod_db.connection()

    def tearDown(self):
        prod_db.rollback()

    def _rows(self, views, depts=None):
        # Заглушка своя, а не общая: переписи нужны ИМЕНА людей, а в общей
        # users только отдел и статус. Дописывать колонку в общую значило бы
        # править арность фикстур во всех остальных тестах файла.
        stubs = _ROSTER_STUB.format(rows=', '.join(views))
        sql = _stub(wiki_analytics._READERS_SQL, stubs)
        with self.conn.cursor() as cursor:
            cursor.execute(sql, {'visible': [1], 'since': None, 'until': None,
                                 'roster': 100, 'depts': depts})
            return cursor.fetchall()

    def test_mover_is_one_row_with_the_latest_department(self):
        """Перешедший между отделами не должен рваться на две строки."""
        rows = self._rows([
            "(1, 5, '2026-08-01 10:00:00'::timestamp, 2)",   # тогда — второй отдел
            "(1, 5, '2026-08-10 10:00:00'::timestamp, 1)",   # теперь — первый
            "(1, 5, '2026-08-11 10:00:00'::timestamp, 1)",
        ])
        self.assertEqual(len(rows), 1, 'человек обязан быть одной строкой')
        _uid, _name, department, reads, articles, _last = rows[0]
        self.assertEqual(reads, 3)
        self.assertEqual(articles, 1)
        self.assertEqual(department, 'СЗоВ', 'отдел берётся из последнего чтения')

    def test_department_boundary_cuts_the_roster(self):
        """Супервайзеру видны читатели только своих отделов."""
        views = ["(1, 5, '2026-08-10 10:00:00'::timestamp, 1)",
                 "(1, 6, '2026-08-10 10:00:00'::timestamp, 2)"]
        self.assertEqual(len(self._rows(views)), 2, 'без границы видны оба')
        limited = self._rows(views, depts=[1])
        self.assertEqual(len(limited), 1)
        self.assertEqual(limited[0][0], 5)


_SECTIONS_STUB = """
wiki_sections AS (
    SELECT id::int, parent_section_id::int, name::text, status::text,
           position::int
      FROM (VALUES {sections}) AS t(id, parent_section_id, name, status, position)
),
wiki_article_sections AS (
    SELECT article_id::int, section_id::int
      FROM (VALUES {links}) AS t(article_id, section_id)
),
wiki_articles AS (
    SELECT id::int, slug::text, title::text, status::text,
           updated_at::timestamp, review_due_at::timestamp,
           updated_by::int, author_id::int
      FROM (VALUES {articles}) AS t(id, slug, title, status, updated_at,
                                    review_due_at, updated_by, author_id)
),
users AS (
    SELECT id::int, name::text FROM (VALUES (5, 'Иванов'), (6, 'Петров')) AS t(id, name)
)
"""


@unittest.skipIf(not prod_db.available(), 'нет доступа к базе для прогона SQL')
class SectionsAndStaleTest(unittest.TestCase):
    """Разделы считают только видимые статьи, устаревшее — только опубликованное."""

    SECTIONS = ("(10, NULL, 'Регламенты', 'active', 0), "
                "(11, 10, 'Фотоконтроль', 'active', 1), "
                "(12, NULL, 'Пустой', 'active', 2), "
                "(13, NULL, 'В архиве', 'archived', 3)")
    LINKS = "(1, 10), (2, 10), (3, 11), (99, 12)"
    ARTICLES = (
        "(1, 'a', 'Первая', 'published', '2026-08-20'::timestamp, NULL, 5, 5), "
        "(2, 'b', 'Вторая', 'draft', '2026-08-24'::timestamp, NULL, 6, 6), "
        "(3, 'c', 'Третья', 'published', '2020-01-01'::timestamp, "
        "'2020-06-01'::timestamp, 5, 5), "
        "(99, 'x', 'Чужая', 'published', '2026-08-24'::timestamp, NULL, 5, 5)")

    @classmethod
    def setUpClass(cls):
        reason = prod_db.skip_reason()
        if reason:
            raise unittest.SkipTest(reason)
        cls.conn = prod_db.connection()

    def tearDown(self):
        prod_db.rollback()

    def _stubs(self):
        return _SECTIONS_STUB.format(sections=self.SECTIONS, links=self.LINKS,
                                     articles=self.ARTICLES)

    def _run(self, sql, params):
        with self.conn.cursor() as cursor:
            cursor.execute(_stub(sql, self._stubs()), params)
            return cursor.fetchall()

    def test_section_counts_only_visible_articles(self):
        """Статья 99 не в периметре — раздел «Пустой» обязан остаться пустым."""
        rows = self._run(wiki_analytics._SECTIONS_SQL,
                         {'visible': [1, 2, 3], 'sections': [10, 11, 12, 13],
                          'roster': 100})
        by_name = {r[1]: r for r in rows}
        self.assertNotIn('В архиве', by_name, 'архивный раздел не показываем')
        self.assertEqual(by_name['Регламенты'][3], 2, 'две статьи')
        self.assertEqual(by_name['Регламенты'][4], 1, 'опубликована одна')
        self.assertEqual(by_name['Пустой'][3], 0)
        self.assertIsNone(by_name['Пустой'][5], 'у пустого нет даты правки')
        self.assertEqual(by_name['Фотоконтроль'][2], 'Регламенты',
                         'у вложенного раздела виден родитель')

    def test_stale_takes_published_older_than_threshold(self):
        """Черновик и свежая статья в устаревшие не попадают."""
        params = {'visible': [1, 2, 3, 99], 'days': 180, 'limit': 10}
        rows = self._run(wiki_analytics._STALE_SQL, params)
        self.assertEqual([r[2] for r in rows], ['Третья'])
        row = rows[0]
        self.assertGreater(row[4], 180, 'дней без правки больше порога')
        self.assertTrue(row[5], 'срок пересмотра прошёл — признак поднят')
        self.assertEqual(row[7], 'Фотоконтроль', 'раздел статьи подписан')
        count = self._run(wiki_analytics._STALE_COUNT_SQL, params)
        self.assertEqual(count[0][0], 1)


_ACK_STUB = """
wiki_ack_assignments AS (
    SELECT article_id::int, user_id::int, status::text, due_at::timestamp,
           acknowledged_at::timestamp, snapshot_department_id::int,
           snapshot_department_name::text
      FROM (VALUES {rows}) AS t(article_id, user_id, status, due_at,
                                acknowledged_at, snapshot_department_id,
                                snapshot_department_name)
)
"""


@unittest.skipIf(not prod_db.available(), 'нет доступа к базе для прогона SQL')
class AckOverdueTest(unittest.TestCase):
    """Отменённое и перевыпущенное назначение не считаются просрочкой."""

    @classmethod
    def setUpClass(cls):
        reason = prod_db.skip_reason()
        if reason:
            raise unittest.SkipTest(reason)
        cls.conn = prod_db.connection()

    def tearDown(self):
        prod_db.rollback()

    def test_cancelled_and_superseded_are_not_overdue(self):
        rows = [
            # Живая просрочка: срок прошёл, подписи нет.
            "(1, 5, 'not_open', '2020-01-01'::timestamp, NULL, 1, 'СЗоВ')",
            # Отменённое назначение с тем же прошедшим сроком.
            "(1, 6, 'cancelled', '2020-01-01'::timestamp, NULL, 1, 'СЗоВ')",
            # Перевыпущенное — снято выходом новой версии статьи.
            "(1, 7, 'superseded', '2020-01-01'::timestamp, NULL, 1, 'СЗоВ')",
            # Подтверждённое вовремя.
            "(1, 8, 'acknowledged', '2020-01-01'::timestamp, "
            "'2019-12-01'::timestamp, 1, 'СЗоВ')",
        ]
        sql = _stub(wiki_analytics._ACK_TOTALS_SQL,
                    _ACK_STUB.format(rows=', '.join(rows)))
        with self.conn.cursor() as cursor:
            cursor.execute(sql, {'visible': [1]})
            total, done, not_open, overdue, people, articles = cursor.fetchone()
        self.assertEqual(total, 2, 'живых назначений два')
        self.assertEqual(done, 1)
        self.assertEqual(not_open, 1)
        self.assertEqual(overdue, 1, 'просрочка ровно одна — отменённое не в счёт')
        self.assertEqual((people, articles), (2, 1))


# ── Выгрузка ─────────────────────────────────────────────────────────────────

# Порядок fetchone у выгрузки — тот же, что у экрана, плюс один запрос на
# титульный лист (имя автора и название пространства одной строкой).
EXPORT_FETCHONE = list(REPORT_FETCHONE) + [('Иванов Иван', 'Тез')]


@unittest.skipIf(Flask is None, 'flask не установлен')
@unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
class ExportGateTest(unittest.TestCase):
    """Дверь в файл — та же, что дверь на экран."""

    def test_reader_is_refused(self):
        """Читателю выгрузка закрыта так же, как вкладка: в книге люди поимённо
        и она уходит из портала вместе с человеком."""
        client, _ = build_client(self, make_context(), fetchone=EXPORT_FETCHONE)
        response = client.get('/api/wiki/analytics/export')
        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.get_json()['code'], 'WIKI_EDITOR_ONLY')

    def test_editor_gets_a_workbook(self):
        client, _ = build_client(self, make_context(can_edit=True),
                                 fetchone=EXPORT_FETCHONE)
        response = client.get('/api/wiki/analytics/export')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, routes_analytics.XLSX_MIME)
        self.assertIn('attachment', response.headers.get('Content-Disposition', ''))
        book = load_workbook(BytesIO(response.data))
        self.assertEqual(book.sheetnames[0], 'Контекст')

    def test_department_boundary_reaches_the_file(self):
        """Граница отдела у файла своя не бывает: тот же _departments."""
        client, cursor = build_client(
            self, make_context(otp_role='sv', department_id=7, can_publish=True),
            fetchone=EXPORT_FETCHONE)
        self.assertEqual(client.get('/api/wiki/analytics/export').status_code, 200)
        with_depts = [p for _sql, p in cursor.calls
                      if isinstance(p, dict) and 'depts' in p]
        self.assertTrue(with_depts, 'поимённый список в файле не получил границу')
        for params in with_depts:
            self.assertEqual(params['depts'], [7])

    def test_screen_limit_does_not_shrink_the_file(self):
        """`limit` из строки запроса — потолок ЭКРАНА.

        Он приходит с фронта вместе с остальными параметрами периода, и если бы
        выгрузка его слушалась, файл обрезался бы сотней строк — ровно там, где
        начинается то, ради чего его и просили.
        """
        client, cursor = build_client(self, make_context(can_edit=True),
                                      fetchone=EXPORT_FETCHONE)
        self.assertEqual(
            client.get('/api/wiki/analytics/export?limit=5').status_code, 200)
        limits = {p['limit'] for _sql, p in cursor.calls
                  if isinstance(p, dict) and 'limit' in p}
        self.assertEqual(limits, {routes_analytics.EXPORT_ROWS})

    def test_period_and_space_reach_the_file(self):
        """Файл повторяет то, на что человек смотрит: тот же период и та же вики."""
        client, cursor = build_client(self, make_context(can_edit=True),
                                      fetchone=EXPORT_FETCHONE)
        client.get('/api/wiki/analytics/export?since=2026-08-01&until=2026-08-24'
                   '&space_id=3')
        dated = [p for _sql, p in cursor.calls
                 if isinstance(p, dict) and 'since' in p]
        self.assertTrue(dated)
        for params in dated:
            self.assertEqual((params['since'], params['until']),
                             ('2026-08-01', '2026-08-24'))
        # Название пространства для титульного листа сервер берёт из базы, а не
        # из строки запроса: подставлять в файл присланное браузером незачем.
        titles = [p for _sql, p in cursor.calls
                  if isinstance(p, dict) and 'space' in p and 'user' in p]
        self.assertEqual(titles, [{'user': 42, 'space': 3}])


@unittest.skipIf(load_workbook is None, 'openpyxl не установлен')
class ExportWorkbookTest(unittest.TestCase):
    """Книга: листы, форматы ячеек и оговорки.

    Проверяется не «собралось без ошибки», а то, ради чего файл и берут: дата
    обязана быть датой (иначе не работает фильтр по периоду), доля — долей
    (иначе по колонке не считается среднее), а обрез списка — подписанным.
    """

    SHEETS = ['Контекст', 'Прочтения по дням', 'Кто читает по отделам',
              'Что читают чаще всего', 'Не открывали ни разу',
              'Кто пользовался вики', 'Разделы', 'Давно не обновляли',
              'Ознакомления по отделам', 'Просрочено поимённо',
              'Темы без ответа']

    def _book(self, report, **kwargs):
        kwargs.setdefault('generated_at', datetime(2026, 8, 26, 12, 0))
        stream = analytics_report.build_workbook(report=report, **kwargs)
        return load_workbook(BytesIO(stream.getvalue()))

    def test_empty_report_still_builds(self):
        """Пустой периметр не должен ронять выгрузку: у нового редактора он
        именно такой, и вместо файла он получил бы 500."""
        book = self._book({})
        self.assertEqual(book.sheetnames, self.SHEETS)

    def test_rows_land_with_the_right_types(self):
        book = self._book(FULL_REPORT, space_name='Тез',
                          requested_by='Иванов Иван',
                          since='2026-08-01', until='2026-08-26')

        top = book['Что читают чаще всего']
        self.assertEqual(top['A2'].value, 'Как принять заказ')
        self.assertEqual(top['B2'].value, 'опубликована')
        # Дата — датой: по тексту «22.08.2026» фильтр Excel не работает.
        self.assertIsInstance(top['E2'].value, datetime)

        # Доля — числом от нуля до единицы плюс процентный формат: строка «50%»
        # не усредняется и не ложится в сводную.
        depts = book['Кто читает по отделам']
        self.assertAlmostEqual(depts['D2'].value, 0.5)
        self.assertEqual(depts['D2'].number_format, '0%')

        unread = book['Не открывали ни разу']
        self.assertEqual(unread['B2'].value, 'нет', 'статью не открывали никогда')
        self.assertIsNone(unread['C2'].value)
        self.assertEqual(unread['B3'].value, 'да')
        self.assertEqual(unread['D3'].value, 25, 'дней с последнего чтения')

        # Коды превращаются в те же слова, что на экране.
        self.assertEqual(book['Просрочено поимённо']['F2'].value, 'не открывал')
        self.assertEqual(book['Темы без ответа']['B2'].value, 'Поиск')
        self.assertEqual(book['Темы без ответа']['C2'].value, 'Нет статьи')

        sections = book['Разделы']
        self.assertEqual(sections['F2'].value, 'Пётр · 3, Анна · 1')

    def test_context_carries_period_author_and_notes(self):
        book = self._book(FULL_REPORT, space_name='Тез',
                          requested_by='Иванов Иван',
                          since='2026-08-01', until='2026-08-26')
        text = '\n'.join(str(cell.value) for row in book['Контекст'].iter_rows()
                          for cell in row if cell.value is not None)
        self.assertIn('с 2026-08-01 по 2026-08-26', text)
        self.assertIn('Иванов Иван', text)
        self.assertIn('Тез', text)
        # Оговорки приезжают с сервера теми же словами, что на экране: иначе
        # файл и вкладка объясняют одно число по-разному.
        self.assertIn('Прочтение — это человек, статья и минута', text)
        self.assertIn('только по вашим отделам', text)
        self.assertIn('только ваши отделы', text, 'охват выгрузки не подписан')
        # Легенда причин — только те, что встретились в выборке.
        self.assertIn('«Нет статьи»', text)
        self.assertNotIn('«Помощник отказал»', text)

    def test_period_without_borders_is_named(self):
        """Пустые границы — «за всё время», а не пустая строка: через месяц по
        одному числу уже не сказать, за какие дни оно посчитано."""
        book = self._book({})
        values = [cell.value for row in book['Контекст'].iter_rows() for cell in row]
        self.assertIn('за всё время', values)

    def test_truncated_sheet_is_named(self):
        """Молчаливый обрез читается как «больше и нет»."""
        report = {'reading': {'top': [{'title': 'Статья', 'reads': 1}] * 3}}
        book = self._book(report, row_cap=3)
        text = '\n'.join(str(cell.value) for row in book['Контекст'].iter_rows()
                          for cell in row if cell.value is not None)
        self.assertIn('«Что читают чаще всего» упёрся в потолок', text)

    def test_filename_has_no_slashes(self):
        self.assertEqual(analytics_report.report_filename(datetime(2026, 8, 26)),
                         'wiki_analytics_2026-08-26.xlsx')


# Один отчёт на все проверки книги: он повторяет форму ответа /analytics, и
# собирать его заново в каждом тесте значило бы завести пять слегка разных
# форм ответа вместо одной.
FULL_REPORT = {
    'scoped': True,
    'notes': {
        'read': 'Прочтение — это человек, статья и минута: обновление страницы…',
        'ack_now': 'Выбранный период на этот блок не действует.',
        'scoped': 'Люди показаны поимённо только по вашим отделам.',
    },
    'reading': {
        'totals': {'reads': 47, 'opens': 56, 'readers': 12, 'articles_read': 9,
                   'published': 20, 'unread': 11, 'coverage': 45.0},
        'days': [{'day': '2026-08-25', 'reads': 4, 'readers': 2}],
        'top': [{'id': 1, 'slug': 'order', 'title': 'Как принять заказ',
                 'status': 'published', 'reads': 12, 'readers': 5,
                 'updated_at': '2026-08-20T10:00:00'}],
        'unread': [
            {'id': 2, 'slug': 'never', 'title': 'Регламент выдачи',
             'updated_at': '2026-01-10T10:00:00', 'last_at': None},
            {'id': 3, 'slug': 'old', 'title': 'Памятка по кассе',
             'updated_at': '2026-02-10T10:00:00', 'last_at': '2026-08-01T12:00:00'},
        ],
        'departments': [{'department_id': 1, 'name': 'СЗоВ', 'reads': 30,
                         'readers': 5, 'articles_read': 7, 'headcount': 10}],
        'people': [{'user_id': 5, 'name': 'Пётр Петров', 'department': 'СЗоВ',
                    'reads': 9, 'articles': 4, 'last_at': '2026-08-25T09:30:00'}],
    },
    'content': {
        'sections': [{'id': 1, 'name': 'Регламенты', 'parent': 'СЗоВ',
                      'articles': 12, 'published': 10,
                      'last_update': '2026-08-19T10:00:00',
                      'editors': [{'name': 'Пётр', 'edits': 3},
                                  {'name': 'Анна', 'edits': 1}]}],
        'stale': [{'id': 4, 'slug': 'stale', 'title': 'Старый регламент',
                   'updated_at': '2025-06-01T10:00:00', 'days': 451,
                   'review_overdue': True, 'editor': 'Анна', 'section': 'Регламенты'}],
        'stale_total': 1,
        'stale_days': 180,
    },
    'acknowledgements': {
        'totals': {'total': 8, 'done': 5, 'not_open': 2, 'overdue': 1,
                   'people': 4, 'articles': 3},
        'departments': [{'department_id': 1, 'name': 'СЗоВ', 'total': 8,
                         'done': 4, 'overdue': 1}],
        'overdue': [{'user_id': 6, 'name': 'Сидоров Сидор', 'department': 'СЗоВ',
                     'team': 'Группа 1', 'supervisor': 'Иванов Иван',
                     'article_id': 1, 'slug': 'order', 'title': 'Как принять заказ',
                     'due_at': '2026-08-10T00:00:00', 'status': 'not_open',
                     'days': 16}],
    },
    'demand': {
        'search': {'total': 120, 'empty': 14, 'people': 9, 'steps': 340,
                   'empty_share': 11.7, 'logging_since': '2026-08-01T00:00:00'},
        'assistant': {'total': 40, 'answered': 30, 'no_answer': 8, 'clarify': 2,
                      'people': 6},
        'items': [{'channel': 'search', 'key': 'справка ндс', 'text': 'справка ндс',
                   'times': 6, 'people': 3, 'last_at': '2026-08-24T15:00:00',
                   'reason': 'missing'}],
    },
}


class ExportButtonTest(unittest.TestCase):
    """Кнопка выгрузки во вкладке. Тест читает WikiAnalytics.jsx текстом.

    Решения здесь интерфейсные, но ломаются они молча: с экранным `limit` файл
    приедет обрезанным сотней строк и будет выглядеть исправным, а без blob и
    заголовков вместо книги скачается страница входа.
    """

    @classmethod
    def setUpClass(cls):
        cls.src = (ROOT / 'src' / 'components' / 'wiki'
                   / 'WikiAnalytics.jsx').read_text(encoding='utf-8')

    def test_button_calls_the_export_route(self):
        self.assertIn('/analytics/export', self.src)
        self.assertIn('Выгрузить в Excel', self.src)

    def test_file_is_fetched_with_headers_and_as_blob(self):
        """Раздел авторизуется заголовком: обычная ссылка их не несёт."""
        self.assertIn("responseType: 'blob'", self.src)
        self.assertRegex(self.src, r'analytics/export`,\s*\{\s*\n?\s*headers')

    def test_screen_row_cap_is_dropped(self):
        """`limit` — потолок экрана; в файле он обрезал бы ровно подробность."""
        self.assertIn('limit: undefined', self.src)

    def test_period_and_space_go_with_the_file(self):
        """Выгрузка повторяет то, на что человек смотрит."""
        self.assertIn('...params', self.src)


if __name__ == '__main__':
    unittest.main()
