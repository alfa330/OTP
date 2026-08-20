"""Табло СЗоВ, направление «Чат»: статусы чатников, почасовой срез и кэш снимка.

Функции бэкенда вытаскиваем из bot_schedule2.py тем же загрузчиком, что и тесты «Линии»
(`_load_names`): проверяется настоящая логика, а не строковое совпадение.
"""
import logging
import math
import os
import re
import threading
import time
import unittest
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

from tests.test_szov_wallboard import _FakeDb, _load_names

ROOT = Path(__file__).resolve().parents[1]

# Имена вымышленные, но в тех же двух формах, что расходятся в жизни: Chat2Desk отдаёт
# «Имя Фамилия», в OTP человек записан как «Фамилия Имя Отчество».
CHAT_MANAGERS = {
    'Алия Тестова': {'id': 235, 'name': 'Тестова Алия Тестовна',
                     'calculation_model_code': 'chat_manager', 'direction_name': 'Чат менеджер'},
    'Бекзат Примеров': {'id': 149, 'name': 'Примеров Бекзат Примерулы',
                        'calculation_model_code': 'chat_manager', 'direction_name': 'Чат менеджер'},
    'Дана Ночная': {'id': 18, 'name': 'Ночная Дана Сменовна',
                    'calculation_model_code': 'chat_manager', 'direction_name': 'Чат менеджер'},
    'Ерлан Учебный': {'id': 37, 'name': 'Учебный Ерлан Тренингулы',
                      'calculation_model_code': 'chat_manager', 'direction_name': 'Чат менеджер'},
}

NAMES = {
    'SZOV_WALLBOARD_DEPARTMENT_CODE',
    'SZOV_CHAT_WALLBOARD_TARGET_SECONDS',
    'SZOV_CHAT_WALLBOARD_ONLINE_MIN_SECONDS',
    'SZOV_CHAT_WALLBOARD_CACHE_TTL_SECONDS',
    'SZOV_CHAT_WALLBOARD_REQUESTS_TTL_SECONDS',
    'SZOV_CHAT_WALLBOARD_STALE_MAX_SECONDS',
    'SZOV_CHAT_WALLBOARD_RETRY_AFTER_FAIL_SECONDS',
    'SZOV_CHAT_WALLBOARD_LOCK_WAIT_SECONDS',
    'SZOV_CHAT_WALLBOARD_EVENT_MAX_PAGES',
    '_SZOV_CHAT_WALLBOARD_STATUSES',
    '_SZOV_CHAT_WALLBOARD_OTHER_KEY',
    '_SZOV_CHAT_WALLBOARD_OFFLINE',
    '_SZOV_CHAT_WALLBOARD_STATUS_ORDER',
    '_SZOV_CHAT_WALLBOARD_STATUS_RANK',
    '_SZOV_WALLBOARD_DEPARTMENT_CACHE',
    '_SZOV_WALLBOARD_DEPARTMENT_CACHE_TTL',
    '_szov_wallboard_department_id',
    '_szov_chat_wallboard_events_cache',
    '_szov_chat_wallboard_requests_cache',
    '_szov_chat_wallboard_cache',
    '_szov_chat_wallboard_lock',
    '_operator_info_is_chat_manager',
    '_szov_chat_wallboard_status',
    '_szov_chat_wallboard_operator_lookup',
    '_szov_chat_wallboard_resolve',
    '_szov_chat_wallboard_day_seconds',
    '_szov_chat_wallboard_timelines',
    '_szov_chat_wallboard_online_seconds',
    '_szov_chat_wallboard_reply_sums',
    '_szov_chat_wallboard_reply_average',
    '_szov_chat_wallboard_hourly',
    '_szov_chat_wallboard_display_names',
    '_szov_chat_wallboard_now',
    '_szov_chat_wallboard_fetch_events',
    '_szov_chat_wallboard_fetch_snapshot',
    '_wallboard_snapshot_with_cache',
    '_szov_chat_wallboard_snapshot',
}


def _event(name, event, created_at, role='Супервайзер', operator_id=1):
    return {'operator_id': operator_id, 'operator_name': name, 'operator_role': role,
            'event': event, 'dialog_id': '', 'created_at': created_at}


def _operator_row(name, online=1, offline_type=None, dialogs=0, status='enabled'):
    first, last = name.split(' ', 1)
    return {'first_name': first, 'last_name': last, 'online': online, 'status': status,
            'offline_type': offline_type, 'opened_dialogs': dialogs}


def _request(start, reaction, replies, total, operator='Алия Тестова', end='x'):
    # average_replies_time = total_replies_time / replies — так его отдаёт Chat2Desk.
    return {'request_start': start, 'request_end': end, 'request_type': 'common',
            'operator_name': operator, 'reaction_time': reaction,
            'replies': replies, 'total_replies_time': total,
            'average_replies_time': (total / replies) if (total is not None and replies) else ''}


class _Harness:
    """Namespace с настоящими функциями табло и поддельными Chat2Desk/БД."""

    def _namespace(self, *, members=None, operators=None, requests_get=None):
        source = (ROOT / "bot_schedule2.py").read_text(encoding="utf-8-sig")
        matched = operators if operators is not None else CHAT_MANAGERS

        class _FakeRequests:
            def __init__(self, handler):
                self.handler = handler
                self.calls = []

            def get(self, url, headers=None, timeout=None, params=None):
                self.calls.append(params or {})
                return self.handler(params or {})

        fake_requests = _FakeRequests(requests_get or (lambda params: None))
        ns = {
            'time': time,
            'math': math,
            'logging': logging,
            'threading': threading,
            'datetime': datetime,
            'ZoneInfo': ZoneInfo,
            'requests': fake_requests,
            're': __import__('re'),
            '_env_int': lambda name, default, minimum=None, maximum=None: default,
            'CHAT2DESK_API_PAGE_LIMIT': 200,
            'CHAT2DESK_API_TIMEOUT_SECONDS': 45,
            'CHAT2DESK_STATISTICS_REPORT_OPERATOR_EVENTS': 'operator_events',
            'CHAT_HOURLY_TIMEZONE': 'Asia/Almaty',
            'CHAT_HOURLY_REQUEST_TYPE': 'common',
            'CHAT2DESK_STATISTICS_REPORT_REQUEST_STATS': 'request_stats',
            'db': _FakeDb(members={1: members if members is not None else {18, 37, 149, 235}}),
            '_chat2desk_authorization_header': lambda: 'token',
            '_chat2desk_api_base_url': lambda: 'https://api.example',
            '_chat2desk_api_error_message': lambda response, report, day: f'HTTP {response.status_code}',
            '_chat2desk_extract_response_rows': lambda payload: (payload or {}).get('data') or [],
            '_chat2desk_extract_total': lambda payload: ((payload or {}).get('meta') or {}).get('total'),
            '_chat2desk_row_first': lambda row, *keys: next(
                (row[key] for key in keys if row.get(key) not in (None, '')), None),
            '_chat2desk_operator_display_name': lambda row: (
                f"{row.get('first_name') or ''} {row.get('last_name') or ''}".strip()),
            '_chat_hourly_number': lambda value: (
                None if str(value if value is not None else '').strip() == ''
                else float(str(value).strip())),
            '_chat_hourly_is_open': lambda row: str(row.get('request_end') or '').strip() in ('', 'None', 'null'),
            '_chat_hourly_request_start': lambda row: str(row.get('request_start') or '').strip(),
            '_chat_hourly_operator_name': lambda row: str(row.get('operator_name') or '').strip(),
            '_chat_hourly_lock': threading.Lock(),
            '_chat_hourly_requests_cache': {'day': None, 'rows': {}},
            '_status_import_build_operator_lookup': lambda restrict_to_ids=None: {
                'lookup': [dict(info, id=info['id']) for info in matched.values()
                           if restrict_to_ids is None or info['id'] in restrict_to_ids]
            },
            '_status_import_resolve_operator_matches': lambda name, lookup: (
                [matched[name]] if name in matched and any(
                    info['id'] == matched[name]['id'] for info in lookup.get('lookup', [])) else []
            ),
        }
        # Ответ внутри чата считает боевая функция отчёта по чатам — берём её как есть.
        _load_names(source, {'_chat_hourly_response_times', '_chat_hourly_response_sums'}, ns)
        _load_names(source, NAMES, ns)
        ns['_SZOV_WALLBOARD_DEPARTMENT_CACHE'].update(ts=0.0, id=None)
        ns['_szov_chat_wallboard_events_cache'].update(day=None, rows={}, newest=None, truncated=False)
        ns['_szov_chat_wallboard_requests_cache'].update(day=None, ts=0.0)
        ns['_szov_chat_wallboard_cache'].update(ts=0.0, payload=None, failed_at=0.0, error=None)
        ns['_fake_requests'] = fake_requests
        return ns


class ChatWallboardTimelineTests(_Harness, unittest.TestCase):
    """Ленты статусов: что считается «на линии» и как достраивается ночная смена."""

    def test_only_status_events_of_szov_chat_managers_get_in(self):
        ns = self._namespace()
        events = [
            _event('Алия Тестова', 'online', '2026-08-18 09:00:00'),
            _event('Алия Тестова', 'take_chat', '2026-08-18 09:05:00'),
            _event('Посторонний Оператор', 'online', '2026-08-18 09:10:00'),
            _event('Администратор компании', 'offline', '2026-08-18 09:15:00', role='admin'),
        ]
        lookup, _ = ns['_szov_chat_wallboard_operator_lookup']()
        timelines, unmatched = ns['_szov_chat_wallboard_timelines'](events, lookup)
        self.assertEqual(list(timelines), ['Алия Тестова'])
        self.assertEqual([entry[1] for entry in timelines['Алия Тестова']], ['online'])
        # Админская учётка чатов не ведёт и в «потерянные имена» не попадает.
        self.assertEqual(unmatched, ['Посторонний Оператор'])

    def test_night_shift_is_backfilled_from_midnight(self):
        """Первое событие суток — выход: человек работал с полуночи, а не появился в 02:00."""
        ns = self._namespace()
        events = [_event('Дана Ночная', 'logout', '2026-08-18 02:00:45')]
        lookup, _ = ns['_szov_chat_wallboard_operator_lookup']()
        timelines, _ = ns['_szov_chat_wallboard_timelines'](events, lookup)
        entries = timelines['Дана Ночная']
        self.assertEqual(entries[0][0], 0)
        self.assertEqual(entries[0][1], 'online')
        self.assertEqual(entries[1][1], 'offline')

    def test_morning_login_is_not_backfilled(self):
        ns = self._namespace()
        events = [_event('Алия Тестова', 'login', '2026-08-18 09:00:00'),
                  _event('Алия Тестова', 'online', '2026-08-18 09:00:00')]
        lookup, _ = ns['_szov_chat_wallboard_operator_lookup']()
        timelines, _ = ns['_szov_chat_wallboard_timelines'](events, lookup)
        self.assertEqual(timelines['Алия Тестова'][0][0], 9 * 3600)

    def test_online_seconds_land_in_every_hour_they_touched_and_stop_at_now(self):
        """Смена через границу часа считается в каждом часе своей частью и не идёт в будущее."""
        ns = self._namespace()
        timelines = {'Алия Тестова': [(8 * 3600 + 1800, 'online', 'Онлайн', '')]}
        per_hour = ns['_szov_chat_wallboard_online_seconds'](timelines, 10 * 3600 + 900)
        # Разбивка по людям: из неё считаются и голова часа, и ФИО в подсказке графика.
        self.assertEqual(per_hour, {8: {'Алия Тестова': 1800}, 9: {'Алия Тестова': 3600},
                                    10: {'Алия Тестова': 900}})

    def test_busy_and_training_do_not_hold_the_line(self):
        """Занят и тренинг — это НЕ на линии; отрезки одного человека внутри часа складываются."""
        ns = self._namespace()
        timelines = {'Алия Тестова': [
            (9 * 3600, 'online', 'Онлайн', ''),
            (9 * 3600 + 600, 'busy', 'Занят', ''),
            (9 * 3600 + 1200, 'break', 'Перерыв', ''),
            (9 * 3600 + 1800, 'online', 'Онлайн', ''),
        ],
            # Весь час занят, весь час на тренинге: линию не держали, в счёт не идут.
            'Бекзат Примеров': [(9 * 3600, 'busy', 'Занят', '')],
            'Ерлан Учебный': [(9 * 3600, 'training', 'Тренинг', '')]}
        per_hour = ns['_szov_chat_wallboard_online_seconds'](timelines, 10 * 3600)
        self.assertEqual(per_hour, {9: {'Алия Тестова': 600 + 1800}})

    def test_status_spellings_are_normalized(self):
        """Chat2Desk пишет один статус четырьмя способами — ключ должен получиться один."""
        ns = self._namespace()
        for raw in ('tech break', 'tech_break', 'tech.break', 'status.tech.break', 'TECH BRAKE'):
            self.assertEqual(ns['_szov_chat_wallboard_status'](raw), ('tech', 'Тех. перерыв'), raw)
        self.assertEqual(ns['_szov_chat_wallboard_status']('status.online'), ('online', 'Онлайн'))
        self.assertIsNone(ns['_szov_chat_wallboard_status']('take_chat'))

    def test_status_labels_come_from_the_chat2desk_dictionary(self):
        """Подписи сверены со справочником Chat2Desk (GET /v1/operators/statuses).

        Главное здесь — `holiday`: он был подписан «Отпуск», а Chat2Desk называет его
        «Закрытие чатов» (человек в системе, дорабатывает открытые, новых не берёт).
        Супервайзеры читали со стены отпуск там, где смена работает."""
        ns = self._namespace()
        for raw, expected in (
            ('online', ('online', 'Онлайн')),
            ('busy', ('busy', 'Занят')),
            ('break', ('break', 'Перерыв')),
            ('study', ('training', 'Тренинг')),
            ('tech_break', ('tech', 'Тех. перерыв')),
            ('holiday', ('holiday', 'Закрытие чатов')),
        ):
            self.assertEqual(ns['_szov_chat_wallboard_status'](raw), expected, raw)
        # Отпуска среди статусов Chat2Desk нет вовсе — ни в одной подписи справочника.
        self.assertNotIn('Отпуск', [label for _key, label in
                                    ns['_SZOV_CHAT_WALLBOARD_STATUSES'].values()])

    def test_unknown_live_status_goes_to_the_other_bucket(self):
        """Незнакомый статус не выдумываем ключом, иначе человек выпадет из всех счётчиков."""
        ns = self._namespace()
        self.assertEqual(ns['_szov_chat_wallboard_status']('dinner', unknown_as_other=True),
                         ('other', 'dinner'))

    def test_status_event_with_a_prefixed_spelling_breaks_the_online_span(self):
        """Раньше `tech.break` не распознавался, и человек «держал линию» весь перерыв."""
        ns = self._namespace()
        events = [_event('Алия Тестова', 'online', '2026-08-18 09:00:00'),
                  _event('Алия Тестова', 'status.tech.break', '2026-08-18 09:30:00')]
        lookup, _ = ns['_szov_chat_wallboard_operator_lookup']()
        timelines, _ = ns['_szov_chat_wallboard_timelines'](events, lookup)
        # Нераспознанный `tech.break` тянул бы «Онлайн» до самого «сейчас» — человек попал бы
        # ещё и в 10-й час.
        per_hour = ns['_szov_chat_wallboard_online_seconds'](timelines, 11 * 3600)
        self.assertEqual(per_hour, {9: {'Алия Тестова': 1800}})

    def test_truncated_events_do_not_backfill_the_night(self):
        """Выгрузка обрезалась — самое раннее событие уже не начало смены, достраивать нельзя."""
        ns = self._namespace()
        ns['_szov_chat_wallboard_events_cache']['truncated'] = True
        events = [_event('Дана Ночная', 'break', '2026-08-18 14:10:00')]
        lookup, _ = ns['_szov_chat_wallboard_operator_lookup']()
        timelines, _ = ns['_szov_chat_wallboard_timelines'](events, lookup)
        self.assertEqual([entry[1] for entry in timelines['Дана Ночная']], ['break'])
        self.assertEqual(ns['_szov_chat_wallboard_online_seconds'](timelines, 15 * 3600), {})


class ChatWallboardHourlyTests(_Harness, unittest.TestCase):
    """Почасовые строки графика: чаты, время ответа и КТО держал линию."""

    def _rows(self, ns, timelines, requests, now, **kwargs):
        return {row['hour']: row
                for row in ns['_szov_chat_wallboard_hourly'](requests, timelines, now, **kwargs)}

    def test_hours_stop_at_the_current_hour(self):
        ns = self._namespace()
        now = datetime(2026, 8, 18, 13, 36, 0)
        rows = ns['_szov_chat_wallboard_hourly']([], {}, now)
        self.assertEqual([row['hour'] for row in rows], list(range(0, 14)))
        self.assertTrue(rows[-1]['partial'])
        self.assertFalse(rows[-2]['partial'])

    def test_closed_day_covers_all_24_hours(self):
        """Выгрузка за прошедший день: сутки закрыты, «час идёт» там некому идти."""
        ns = self._namespace()
        rows = ns['_szov_chat_wallboard_hourly']([], {}, None, full_day=True)
        self.assertEqual([row['hour'] for row in rows], list(range(0, 24)))
        self.assertFalse(any(row['partial'] for row in rows))

    def test_people_are_counted_by_heads(self):
        """Решение владельца 19.08.2026 (вечер): двое на линии — это 2, а не доля часа."""
        ns = self._namespace()
        now = datetime(2026, 8, 18, 13, 20, 0)
        timelines = {
            'Алия Тестова': [(13 * 3600, 'online', 'Онлайн', '')],
            'Бекзат Примеров': [(13 * 3600, 'online', 'Онлайн', '')],
        }
        rows = self._rows(ns, timelines, [], now)
        self.assertEqual(rows[13]['operators_online'], 2)
        self.assertIsInstance(rows[13]['operators_online'], int)

    def test_two_half_shifts_count_as_two_people(self):
        """Сменились по получасу — на линии в этом часу было двое (прежняя FTE давала 1,0)."""
        ns = self._namespace()
        now = datetime(2026, 8, 18, 10, 0, 0)
        timelines = {
            'Алия Тестова': [(9 * 3600, 'online', 'Онлайн', ''),
                             (9 * 3600 + 1800, 'logout', 'Не в системе', '')],
            'Бекзат Примеров': [(9 * 3600 + 1800, 'online', 'Онлайн', ''),
                                (10 * 3600, 'logout', 'Не в системе', '')],
        }
        rows = self._rows(ns, timelines, [], now)
        self.assertEqual(rows[9]['operators_online'], 2)
        # Минуты смены остались рядом: по ним видно, что часа на двоих было ровно один.
        self.assertEqual(rows[9]['online_seconds'], 3600)

    def test_visit_under_a_minute_is_not_a_person_but_is_not_lost(self):
        """Заглянул на 40 секунд — не человек часа; в счёт не идёт, но и не исчезает."""
        ns = self._namespace()
        now = datetime(2026, 8, 18, 10, 0, 0)
        timelines = {'Алия Тестова': [(9 * 3600, 'online', 'Онлайн', ''),
                                      (9 * 3600 + 40, 'logout', 'Не в системе', '')]}
        rows = self._rows(ns, timelines, [], now)
        self.assertEqual(rows[9]['operators_online'], 0)
        self.assertEqual(rows[9]['operators'], [])
        self.assertEqual(rows[9]['operators_under_minute'], 1)
        self.assertEqual(rows[9]['online_seconds'], 40)

    def test_a_full_minute_already_counts(self):
        """Порог — минута и больше: ровно минута человека уже делает человеком часа."""
        ns = self._namespace()
        now = datetime(2026, 8, 18, 10, 0, 0)
        timelines = {'Алия Тестова': [(9 * 3600, 'online', 'Онлайн', ''),
                                      (9 * 3600 + 60, 'logout', 'Не в системе', '')]}
        rows = self._rows(ns, timelines, [], now)
        self.assertEqual(rows[9]['operators_online'], 1)
        self.assertEqual(rows[9]['operators_under_minute'], 0)

    def test_row_carries_the_people_by_name_and_minutes(self):
        """Подсказка графика показывает ФИО: сверху те, кто держал линию дольше."""
        ns = self._namespace()
        now = datetime(2026, 8, 18, 10, 0, 0)
        timelines = {
            'Алия Тестова': [(9 * 3600, 'online', 'Онлайн', ''),
                             (9 * 3600 + 1200, 'break', 'Перерыв', '')],
            'Бекзат Примеров': [(9 * 3600, 'online', 'Онлайн', '')],
        }
        names = {'Алия Тестова': 'Тестова Алия Тестовна',
                 'Бекзат Примеров': 'Примеров Бекзат Примерулы'}
        rows = self._rows(ns, timelines, [], now, names=names)
        self.assertEqual([person['name'] for person in rows[9]['operators']],
                         ['Примеров Бекзат Примерулы', 'Тестова Алия Тестовна'])
        self.assertEqual([person['seconds'] for person in rows[9]['operators']], [3600, 1200])

    def test_name_without_a_matched_employee_stays_itself(self):
        """ФИО не нашли — оставляем имя учётки: минуты есть, а человека под ними быть обязано."""
        ns = self._namespace()
        now = datetime(2026, 8, 18, 10, 0, 0)
        timelines = {'Алия Тестова': [(9 * 3600, 'online', 'Онлайн', '')]}
        rows = self._rows(ns, timelines, [], now, names={})
        self.assertEqual([person['name'] for person in rows[9]['operators']], ['Алия Тестова'])

    def test_row_carries_the_minutes_the_shift_stood_on_the_line(self):
        """Рядом с головами — сколько всего минут смена простояла на линии в этом часу."""
        ns = self._namespace()
        now = datetime(2026, 8, 18, 10, 0, 0)
        timelines = {
            'Алия Тестова': [(9 * 3600, 'online', 'Онлайн', '')],
            'Бекзат Примеров': [(9 * 3600, 'online', 'Онлайн', ''),
                                (9 * 3600 + 1800, 'break', 'Перерыв', '')],
        }
        rows = self._rows(ns, timelines, [], now)
        self.assertEqual(rows[9]['online_seconds'], 3600 + 1800)
        self.assertEqual(rows[9]['operators_online'], 2)

    def test_row_carries_chats_response_and_required(self):
        ns = self._namespace()
        now = datetime(2026, 8, 18, 12, 0, 0)
        timelines = {'Алия Тестова': [(11 * 3600, 'online', 'Онлайн', '')],
                     'Бекзат Примеров': [(11 * 3600, 'online', 'Онлайн', '')]}
        requests = [
            # Ответ внутри чата = average_replies_time = total/replies = 720/3 = 240 c.
            _request('2026-08-18 11:05:00', 20, 3, 720),
            _request('2026-08-18 12:30:00', 10, 2, 200),
        ]
        rows = self._rows(ns, timelines, requests, now)
        # Час — промежуток: заявка из 12:30 в строку 11 часа не попадает.
        self.assertEqual(rows[11]['chats'], 1)
        self.assertEqual(rows[12]['chats'], 1)
        self.assertEqual(rows[11]['inner_reply_seconds'], 240)
        self.assertEqual(rows[11]['operators_online'], 2)
        self.assertNotIn('operators_required', rows[11])

    def test_hour_carries_the_reply_time_of_every_chatter(self):
        """Из этого разреза выгрузка строит два листа «оператор × час»."""
        ns = self._namespace()
        now = datetime(2026, 8, 18, 12, 0, 0)
        requests = [
            _request('2026-08-18 11:05:00', 30, 2, 200, operator='Алия Тестова'),
            _request('2026-08-18 11:40:00', 90, 1, 100, operator='Алия Тестова'),
            _request('2026-08-18 11:50:00', 60, 4, 800, operator='Бекзат Примеров'),
        ]
        rows = self._rows(ns, {}, requests, now, names={'Алия Тестова': 'Тестова Алия Тестовна'})
        replies = rows[11]['operators_reply']
        # Сверху тот, кто взял больше чатов в этом часу; ФИО — из OTP, если человек сопоставлен.
        self.assertEqual([item['name'] for item in replies],
                         ['Тестова Алия Тестовна', 'Бекзат Примеров'])
        self.assertEqual([item['chats'] for item in replies], [2, 1])
        # Суммы, а не средние: складывать часы разных дней можно только ими.
        self.assertEqual(replies[0]['first_sum'], 120.0)
        self.assertEqual(replies[0]['first_count'], 2)
        self.assertEqual(ns['_szov_chat_wallboard_reply_average'](replies[0], 'first'), 60.0)
        # Ответ внутри чата = average_replies_time: 200/2 = 100 и 100/1 = 100.
        self.assertEqual(ns['_szov_chat_wallboard_reply_average'](replies[0], 'inner'), 100.0)
        self.assertEqual(ns['_szov_chat_wallboard_reply_average'](replies[1], 'inner'), 200.0)

    def test_reply_cut_of_an_hour_adds_up_to_the_hour_itself(self):
        """Разрез по людям и итог часа считаются одними функциями и обязаны сходиться."""
        ns = self._namespace()
        now = datetime(2026, 8, 18, 12, 0, 0)
        requests = [
            _request('2026-08-18 11:05:00', 30, 2, 200, operator='Алия Тестова'),
            _request('2026-08-18 11:50:00', 90, 1, 400, operator='Бекзат Примеров'),
            # Чат, который никто не взял: времени ответа у него нет вовсе, и он не мешает.
            {'request_start': '2026-08-18 11:20:00', 'request_end': '', 'request_type': 'common',
             'operator_name': '', 'reaction_time': '', 'replies': 0,
             'total_replies_time': '', 'average_replies_time': ''},
        ]
        row = self._rows(ns, {}, requests, now)[11]
        self.assertEqual(row['chats'], 3)  # в итог часа чат без оператора входит
        merged = {}
        for item in row['operators_reply']:
            for key in ('first_sum', 'first_count', 'inner_sum', 'inner_count'):
                merged[key] = (merged.get(key) or 0) + item[key]
        self.assertEqual(ns['_szov_chat_wallboard_reply_average'](merged, 'first'),
                         row['first_reply_seconds'])
        self.assertEqual(ns['_szov_chat_wallboard_reply_average'](merged, 'inner'),
                         row['inner_reply_seconds'])

    def test_reply_average_keeps_the_gap_between_no_data_and_zero(self):
        """None — «мерить было не по чему», а не «ответили мгновенно»."""
        ns = self._namespace()
        self.assertIsNone(ns['_szov_chat_wallboard_reply_average']({}, 'first'))
        self.assertIsNone(ns['_szov_chat_wallboard_reply_average'](None, 'inner'))
        self.assertEqual(ns['_szov_chat_wallboard_reply_average'](
            {'first_sum': 0.0, 'first_count': 3}, 'first'), 0.0)

    def test_hour_without_people_still_carries_the_chats(self):
        ns = self._namespace()
        now = datetime(2026, 8, 18, 3, 0, 0)
        rows = self._rows(ns, {}, [_request('2026-08-18 02:10:00', 20, 3, 740)], now)
        self.assertEqual(rows[2]['chats'], 1)
        self.assertEqual(rows[2]['operators_online'], 0)
        self.assertEqual(rows[2]['operators'], [])
        self.assertEqual(rows[2]['online_seconds'], 0)


class ChatWallboardNowTests(_Harness, unittest.TestCase):
    """Счётчики «сейчас»: онлайн, занят, тренинг."""

    def _now(self, ns, events, operator_rows, now_seconds=13 * 3600 + 36 * 60, requests=None):
        lookup, _ = ns['_szov_chat_wallboard_operator_lookup']()
        timelines, _ = ns['_szov_chat_wallboard_timelines'](events, lookup)
        return ns['_szov_chat_wallboard_now'](
            timelines, operator_rows, lookup, now_seconds,
            replies=ns['_szov_chat_wallboard_reply_sums'](requests or []))

    def test_busy_is_not_confused_with_a_finished_shift(self):
        """Ключевое: у вышедшего offline_type остаётся прежним, поэтому статус берём из событий."""
        ns = self._namespace()
        events = [
            _event('Бекзат Примеров', 'busy', '2026-08-18 10:08:34'),
            _event('Дана Ночная', 'logout', '2026-08-18 02:00:45'),
        ]
        rows = [
            # Оба висят в живом списке одинаково: online=0, offline_type='busy'.
            _operator_row('Бекзат Примеров', online=0, offline_type='busy'),
            _operator_row('Дана Ночная', online=0, offline_type='busy'),
        ]
        now = self._now(ns, events, rows)
        self.assertEqual(now['operators_busy'], 1)
        self.assertEqual(now['operators_offline'], 1)
        self.assertEqual([person['name'] for person in now['operators']],
                         ['Примеров Бекзат Примерулы'])

    def test_counts_by_status_and_open_chats(self):
        ns = self._namespace()
        events = [
            _event('Алия Тестова', 'online', '2026-08-18 13:10:46'),
            _event('Ерлан Учебный', 'study', '2026-08-18 13:30:00'),
            _event('Бекзат Примеров', 'break', '2026-08-18 13:20:00'),
        ]
        rows = [
            _operator_row('Алия Тестова', dialogs=75),
            _operator_row('Ерлан Учебный', dialogs=3),
            _operator_row('Бекзат Примеров', dialogs=1),
        ]
        now = self._now(ns, events, rows)
        self.assertEqual(now['operators_online'], 1)
        self.assertEqual(now['operators_on_training'], 1)
        self.assertEqual(now['operators_on_break'], 1)
        self.assertEqual(now['open_chats'], 79)
        # Первым в списке тот, кто держит линию, дальше — по убыванию доступности.
        self.assertEqual([person['status_key'] for person in now['operators']],
                         ['online', 'training', 'break'])

    def test_every_person_carries_his_own_reply_time_of_the_day(self):
        """Запрос владельца: рядом с чатником в правой колонке — как он отвечает.

        Занятость без качества читается неверно: «7 в работе» у того, кто отвечает за минуту,
        и у того, кто держит клиента десять, — это разные семь чатов. Время ответа тут за
        СУТКИ, а не «сейчас»: по одному чату оно скачет от минуты до получаса."""
        ns = self._namespace()
        events = [_event('Алия Тестова', 'online', '2026-08-18 09:00:00'),
                  _event('Бекзат Примеров', 'online', '2026-08-18 09:00:00')]
        rows = [_operator_row('Алия Тестова', dialogs=7), _operator_row('Бекзат Примеров', dialogs=2)]
        requests = [
            _request('2026-08-18 09:05:00', 30, 2, 200, operator='Алия Тестова'),
            _request('2026-08-18 11:40:00', 90, 1, 100, operator='Алия Тестова'),
            _request('2026-08-18 12:10:00', 60, 4, 800, operator='Бекзат Примеров'),
        ]
        now = self._now(ns, events, rows, requests=requests)
        people = {person['name']: person for person in now['operators']}
        alia = people['Тестова Алия Тестовна']
        self.assertEqual(alia['open_chats'], 7)   # сколько чатов у него в работе сейчас
        self.assertEqual(alia['chats'], 2)        # и сколько он взял за сутки
        self.assertEqual(alia['first_reply_seconds'], 60.0)
        self.assertEqual(alia['inner_reply_seconds'], 100.0)
        self.assertEqual(people['Примеров Бекзат Примерулы']['inner_reply_seconds'], 200.0)

    def test_person_without_chats_today_shows_no_reply_time(self):
        """Прочерк, а не ноль: ноль прочитался бы как «отвечает мгновенно»."""
        ns = self._namespace()
        now = self._now(ns, [_event('Алия Тестова', 'online', '2026-08-18 09:00:00')],
                        [_operator_row('Алия Тестова')])
        person = now['operators'][0]
        self.assertEqual(person['chats'], 0)
        self.assertIsNone(person['first_reply_seconds'])
        self.assertIsNone(person['inner_reply_seconds'])

    def test_time_in_status_counts_from_the_last_event(self):
        ns = self._namespace()
        events = [_event('Алия Тестова', 'online', '2026-08-18 13:00:00')]
        now = self._now(ns, events, [_operator_row('Алия Тестова')],
                        now_seconds=13 * 3600 + 36 * 60)
        self.assertEqual(now['operators'][0]['seconds'], 36 * 60)

    def test_operator_without_events_falls_back_to_the_live_flag(self):
        """Смена началась вчера: событий сегодня нет, но человек на линии."""
        ns = self._namespace()
        now = self._now(ns, [], [_operator_row('Алия Тестова', online=1, dialogs=12)])
        self.assertEqual(now['operators_online'], 1)
        self.assertIsNone(now['operators'][0]['seconds'])

    def test_disabled_accounts_are_ignored(self):
        ns = self._namespace()
        now = self._now(ns, [], [_operator_row('Алия Тестова', online=1, status='deleted')])
        self.assertEqual(now['operators_online'], 0)
        self.assertEqual(now['operators'], [])

    def test_only_szov_chat_managers_are_counted(self):
        ns = self._namespace(members=set())
        now = self._now(ns, [], [_operator_row('Алия Тестова', online=1)])
        self.assertEqual(now['operators_online'], 0)

    def test_unknown_status_is_counted_in_other_not_lost(self):
        ns = self._namespace()
        now = self._now(ns, [], [_operator_row('Алия Тестова', online=1, offline_type='dinner')])
        self.assertEqual(now['operators_other'], 1)
        self.assertEqual(now['operators_online'], 0)
        self.assertEqual(now['operators'][0]['status'], 'dinner')

    def test_account_disabled_midday_counts_as_offline(self):
        """Учётки в живом списке уже нет: на линии человека нет, а плитка «Онлайн» не должна врать."""
        ns = self._namespace()
        events = [_event('Алия Тестова', 'online', '2026-08-18 09:00:00')]
        now = self._now(ns, events, [])
        self.assertEqual(now['operators_online'], 0)
        self.assertEqual(now['operators_offline'], 1)
        self.assertEqual(now['operators'], [])

    def test_strangers_do_not_pollute_the_rename_diagnostic(self):
        """В диагностику идут только потерянные чат-менеджеры, а не операторы других отделов."""
        ns = self._namespace()
        lookup, _ = ns['_szov_chat_wallboard_operator_lookup']()
        timelines, unmatched = ns['_szov_chat_wallboard_timelines']([], lookup)
        ns['_szov_chat_wallboard_now'](timelines, [_operator_row('Чужой Оператор', online=1)],
                                       lookup, 10 * 3600)
        self.assertEqual(unmatched, [])


class ChatWallboardEventsFetchTests(_Harness, unittest.TestCase):
    """Догрузка событий: страницы качаем сверху и до уже известных."""

    def _paged(self, pages):
        def handler(params):
            offset = int(params.get('offset') or 0)
            index = offset // 200
            rows = pages[index] if index < len(pages) else []

            class _Response:
                status_code = 200

                @staticmethod
                def json():
                    return {'data': rows, 'meta': {'total': sum(len(page) for page in pages)}}
            return _Response()
        return handler

    def test_first_pass_walks_the_whole_day(self):
        first = [_event('Алия Тестова', 'online', f'2026-08-18 09:{index:02d}:00') for index in range(60)]
        second = [_event('Алия Тестова', 'busy', f'2026-08-18 08:{index:02d}:00') for index in range(60)]
        ns = self._namespace(requests_get=self._paged([first[:200] + second[:200]]))
        rows = ns['_szov_chat_wallboard_fetch_events']('2026-08-18')
        self.assertEqual(len(rows), 120)
        self.assertEqual(len(ns['_fake_requests'].calls), 1)

    def test_second_pass_stops_at_known_events(self):
        page_one = [_event('Алия Тестова', 'online', f'2026-08-18 12:{index:02d}:00', operator_id=index)
                    for index in range(59, -1, -1)]
        page_two = [_event('Алия Тестова', 'busy', f'2026-08-18 11:{index:02d}:00', operator_id=index)
                    for index in range(59, -1, -1)]
        pages = [(page_one + page_two)[:200]]
        ns = self._namespace(requests_get=self._paged(pages))
        ns['_szov_chat_wallboard_fetch_events']('2026-08-18')
        calls_after_first = len(ns['_fake_requests'].calls)
        rows = ns['_szov_chat_wallboard_fetch_events']('2026-08-18')
        # Ни одного нового события: страниц столько же, дублей в кэше нет.
        self.assertEqual(len(rows), 120)
        self.assertEqual(len(ns['_fake_requests'].calls), calls_after_first + 1)

    def test_new_day_drops_the_cache(self):
        ns = self._namespace(requests_get=self._paged([[_event('Алия Тестова', 'online', '2026-08-18 09:00:00')]]))
        ns['_szov_chat_wallboard_fetch_events']('2026-08-18')
        ns['_szov_chat_wallboard_fetch_events']('2026-08-19')
        self.assertEqual(ns['_szov_chat_wallboard_events_cache']['day'], '2026-08-19')


class ChatWallboardSnapshotCacheTests(_Harness, unittest.TestCase):
    """Общий кэш снимка: квота Chat2Desk одна на компанию, N открытых табло — один запрос."""

    def _namespace_with_fetch(self, payloads):
        ns = self._namespace()
        state = {'calls': 0}

        def fake_fetch():
            state['calls'] += 1
            result = payloads[min(state['calls'] - 1, len(payloads) - 1)]
            if isinstance(result, Exception):
                raise result
            return dict(result)

        ns['_szov_chat_wallboard_fetch_snapshot'] = fake_fetch
        ns['_fetch_state'] = state
        return ns

    def test_second_call_within_ttl_reuses_the_snapshot(self):
        ns = self._namespace_with_fetch([{'day': '2026-08-18'}])
        first = ns['_szov_chat_wallboard_snapshot']()
        second = ns['_szov_chat_wallboard_snapshot']()
        self.assertEqual(ns['_fetch_state']['calls'], 1)
        self.assertFalse(first['stale'])
        self.assertFalse(second['stale'])

    def test_failure_returns_the_last_snapshot_marked_stale(self):
        ns = self._namespace_with_fetch([{'day': '2026-08-18'}, RuntimeError('Chat2Desk HTTP 500')])
        ns['_szov_chat_wallboard_snapshot']()
        ns['_szov_chat_wallboard_cache']['ts'] -= ns['SZOV_CHAT_WALLBOARD_CACHE_TTL_SECONDS'] + 1
        stale = ns['_szov_chat_wallboard_snapshot']()
        self.assertTrue(stale['stale'])
        self.assertIn('Chat2Desk', stale['error'])
        self.assertEqual(stale['day'], '2026-08-18')

    def test_failure_without_any_snapshot_raises(self):
        ns = self._namespace_with_fetch([RuntimeError('Chat2Desk HTTP 500')])
        with self.assertRaises(RuntimeError):
            ns['_szov_chat_wallboard_snapshot']()


class ChatWallboardWiringTests(unittest.TestCase):
    """Разводка на фронте: переключатель направления, отдельный опрос, состав экрана."""

    @classmethod
    def setUpClass(cls):
        cls.view = (ROOT / "src" / "components" / "monitoring" / "SzovWallboardView.jsx").read_text(encoding="utf-8-sig")
        cls.board = (ROOT / "src" / "components" / "monitoring" / "SzovChatWallboard.jsx").read_text(encoding="utf-8-sig")
        cls.shared = (ROOT / "src" / "components" / "monitoring" / "szovWallboardShared.js").read_text(encoding="utf-8-sig")

    def test_header_offers_both_directions(self):
        """Подписи направлений живут в общем каталоге: раздел и виджет читают одно описание,
        и разъехаться названия одного направления не могут."""
        self.assertIn("const DIRECTIONS = WALLBOARD_DIRECTION_LIST;", self.view)
        self.assertIn("<SegmentedSwitch value={direction} options={DIRECTIONS}", self.view)
        self.assertIn("label: 'Линия'", self.shared)
        self.assertIn("label: 'Чат'", self.shared)

    def test_line_direction_keeps_its_historic_key(self):
        """Подпись сменили на «Линию», а ключ 'osnova' трогать нельзя: по нему лежат выбор
        направления и набор показателей виджета в localStorage, строки получателей отбивки в
        БД и параметр `direction` в запросах."""
        self.assertIn("key: 'osnova'", self.shared)
        self.assertNotIn("Основа", self.view)

    def test_only_the_open_direction_is_polled(self):
        """Закрытое направление не должно тратить квоту Chat2Desk: хук живёт в своём компоненте."""
        self.assertIn("if (direction === 'chat') {", self.view)
        line_board = self.view[self.view.index("const LineWallboard ="):self.view.index("const ChatWallboard =")]
        self.assertNotIn("useSzovChatWallboardSnapshot", line_board)
        chat_board = self.view[self.view.index("const ChatWallboard ="):
                               self.view.index("export default function SzovWallboardView")]
        self.assertNotIn("useSzovWallboardSnapshot", chat_board)

    def test_chat_direction_polls_its_own_endpoint_slower(self):
        self.assertIn("const CHAT_POLL_INTERVAL_MS = 60000;", self.shared)
        self.assertIn("path: '/api/szov_wallboard/chat_snapshot'", self.shared)
        self.assertIn("pollIntervalMs: CHAT_POLL_INTERVAL_MS", self.shared)

    def test_broadcast_is_offered_on_both_directions(self):
        """Владелец попросил у чатов такую же настройку отбивки, как у линии (20.08.2026).

        Кнопка, форма и права — общие; различается направление, которое уезжает параметром."""
        chat_board = self.view[self.view.index("const ChatWallboard ="):
                               self.view.index("export default function SzovWallboardView")]
        self.assertIn("<BroadcastControls", chat_board)
        self.assertIn("canManageBroadcast", chat_board)
        # Цель времени ответа берём из снимка: иначе подсказка о норме назвала бы значение
        # по умолчанию вместо настоящего.
        self.assertIn("targetSeconds={snapshot?.target_seconds}", chat_board)
        self.assertIn("direction=${encodeURIComponent(direction)}", self.view)

    def test_chat_direction_offers_the_widget_too(self):
        """Виджет «поверх окон» есть у обоих направлений, у каждого со своим набором."""
        chat_board = self.view[self.view.index("const ChatWallboard ="):
                               self.view.index("export default function SzovWallboardView")]
        self.assertIn("onToggleWidget={onToggleWidget}", chat_board)
        self.assertIn("<WidgetButton direction={direction}", self.view)
        widget = (ROOT / "src" / "components" / "monitoring" / "SzovWallboardWidget.jsx").read_text(encoding="utf-8-sig")
        self.assertIn("direction = 'osnova'", widget)
        self.assertIn("wallboardDirection(direction)", widget)
        for key in ('chat_online', 'chat_busy', 'chat_training', 'chat_inner', 'chat_shift_list'):
            self.assertIn(f"key: '{key}'", self.shared, key)

    def test_chosen_direction_is_remembered(self):
        self.assertIn("otp:szov-wallboard-direction", self.view)
        self.assertIn("readStoredDirection(userId)", self.view)

    def test_every_person_on_shift_carries_a_status(self):
        """Запрос владельца: статус рядом с каждым чатником, включая «Онлайн»."""
        column = self.board[self.board.index('const ChatPeopleColumn'):self.board.index('const formatPeople')]
        self.assertIn('{item.status}', column)
        self.assertNotIn("item.status_key !== 'online'", column)

    def test_chat_board_shows_the_three_asked_counters(self):
        for label in ('label="Онлайн"', 'label="Занят"', 'label="Тренинг"'):
            self.assertIn(label, self.board, label)

    def test_shift_column_pairs_the_workload_with_the_reply_time(self):
        """Запрос владельца 20.08.2026: у каждого чатника видно и сколько чатов в работе,
        и как он на них отвечает — первый ответ и ответ внутри чата за сутки."""
        column = self.board[self.board.index('const ChatPeopleColumn'):
                            self.board.index('const formatPeople')]
        # Со стены должно быть видно, ЧТО человек обрабатывает: «12 чатов в работе», а не
        # «12 в работе» (запрос владельца 20.08.2026). Склонение обязательно — иначе «1 чатов».
        self.assertIn("{pluralRu(item.open_chats, 'чат', 'чата', 'чатов')} в работе", column)
        self.assertIn('{formatMinutes(item.first_reply_seconds, 1, false)}', column)
        self.assertIn('{formatMinutes(item.inner_reply_seconds, 1, false)}', column)
        # Оценочный цвет — только у ответа внутри чата: цель задана ему одному.
        self.assertIn('chatReplyTone(item.inner_reply_seconds, targetSeconds)', column)
        # Строка появляется только у того, кто сегодня брал чаты: иначе это пара прочерков.
        self.assertIn('{item.chats ? (', column)

    def test_hours_in_target_tile_is_gone(self):
        """«Часов в цели» убрана по решению владельца 20.08.2026: оценка дня задним числом,
        по которой со стены ничего не сделаешь. Сетка при этом становится на три плитки —
        пустая четвёртая ячейка читалась бы как «сюда что-то не приехало»."""
        # Ищем именно плитку, а не слова: почему её убрали, в файле остаётся комментарием.
        self.assertNotIn('label="Часов в цели"', self.board)
        self.assertNotIn('hoursInTarget', self.board)
        self.assertIn('<Grid cols={3}>', self.board)
        tiles = (ROOT / "src" / "components" / "monitoring" / "SzovWallboardTiles.jsx").read_text(
            encoding="utf-8-sig")
        # Классы перечислены целиком: собранное в рантайме имя Tailwind в бандл не положит.
        self.assertIn("const GRID_COLS = { 3: 'lg:grid-cols-3', 4: 'lg:grid-cols-4' };", tiles)

    def test_holiday_is_closing_chats_not_a_vacation(self):
        """Подпись из справочника Chat2Desk (GET /v1/operators/statuses): `holiday` — это
        «Закрытие чатов», человек в системе и дорабатывает открытые. Отпуска там нет."""
        self.assertIn("holiday: { label: 'Закрытие чатов'", self.shared)
        self.assertNotIn('Отпуск', self.shared)
        self.assertNotIn('в отпуске', self.board)
        self.assertIn("'на закрытии чатов'", self.board)

    def test_chart_pairs_minutes_with_people_on_the_line(self):
        self.assertIn('yAxisId="left"', self.board)
        self.assertIn('yAxisId="right"', self.board)
        self.assertIn('dataKey="innerMinutes"', self.board)
        self.assertIn('dataKey="online"', self.board)
        self.assertIn('<ReferenceLine yAxisId="left" y={targetSeconds / 60}', self.board)
        # У каждого ряда подпись: две оси без легенды прочитать невозможно.
        self.assertIn('<Legend', self.board)
        for name in ('name="Чатников на линии"', 'name="Ответ внутри чата"'):
            self.assertIn(name, self.board, name)

    def test_the_hour_hint_pairs_people_with_minutes_on_the_line(self):
        """Головы без минут врут: двое по часу и двое по десять минут — разные вещи."""
        self.assertIn("label={{ value: 'чатники'", self.board)
        self.assertIn("${formatPeople(row.online) || '0'} чел.", self.board)
        self.assertIn("{formatMinutes(row.onlineSeconds, 0)} на линии", self.board)
        self.assertIn('onlineSeconds: row.online_seconds ?? null', self.board)
        # Не добравшие до минуты видны отдельной пометкой, а не растворяются в счётчике.
        self.assertIn("{formatInt(row.underMinute)} меньше минуты", self.board)

    def test_required_staffing_is_gone_from_the_board(self):
        """Расчёт «сколько нужно под 2 минуты» убран: пропорция от факта завышала в разы."""
        self.assertNotIn('dataKey="required"', self.board)
        self.assertNotIn('Нужно под цель', self.board)
        api = (ROOT / "bot_schedule2.py").read_text(encoding="utf-8-sig")
        self.assertNotIn('operators_required', api)

    def test_chat_direction_offers_the_excel_export(self):
        """Выгрузка есть только у «Чата» и только когда снимок уже пришёл."""
        self.assertIn("/api/szov_wallboard/chat_export", self.view)
        self.assertIn("<ChatExportControls", self.view)
        line_board = self.view[self.view.index("const LineWallboard ="):
                               self.view.index("const chatExportFileName =")]
        self.assertNotIn("ChatExportControls", line_board)
        chat_board = self.view[self.view.index("const ChatWallboard ="):
                               self.view.index("export default function SzovWallboardView")]
        self.assertIn("{snapshot ? (\n                <ChatExportControls", chat_board)

    def test_export_period_is_chosen_inside_the_button(self):
        """Запрос владельца: период выбирается по нажатию «Выгрузить», под ним «Подтвердить».

        Календарь — общий кирпич из «Чатов ChatApp», а не копия; чипа с датой в шапке табло нет,
        иначе на стене висел бы элемент управления, который смотрящим не нужен."""
        self.assertIn("import { IosDateRangeCalendar, isoDate, rangeLabel } from '../ui/DateRangePicker';",
                      self.view)
        self.assertIn("<IosDateRangeCalendar", self.view)
        self.assertNotIn("<IosDateRangePicker", self.view)
        self.assertIn("Подтвердить", self.view)
        self.assertIn("onClick={() => download(range.from, range.to)}", self.view)
        self.assertIn("query.set('date_from', from)", self.view)
        self.assertIn("query.set('date_to', to)", self.view)
        # «Весь период» тут не годится: выгрузка качает Chat2Desk по дню на день.
        self.assertIn("chatExportPresets", self.view)
        # Пресеты ищем в САМОМ списке выгрузки, а не по всему файлу: в нём живут и другие
        # наборы периодов (журнал перерывов, задача #114), и поиск по тексту принимал бы
        # их подписи за пресеты выгрузки.
        presets = re.search(r"const chatExportPresets = \[.*?\];", self.view, flags=re.DOTALL)
        self.assertIsNotNone(presets, "не нашёл список пресетов выгрузки")
        presets = presets.group(0)
        self.assertNotIn("'Весь период'", presets)
        # Потолок периода гасит «Подтвердить» ДО запроса, а не после минуты ожидания.
        self.assertIn("const CHAT_EXPORT_MAX_DAYS = 7;", self.view)
        self.assertIn("disabled={tooLong}", self.view)
        # Пресета длиннее потолка быть не должно: он всегда гасил бы «Подтвердить».
        self.assertNotIn("'30 дней'", presets)
        for preset in ("{ label: 'Сегодня'", "{ label: '3 дня'", "{ label: '7 дней'"):
            self.assertIn(preset, presets, preset)

    def test_calendar_is_one_brick_for_every_section(self):
        """Календарь вынесен из чипа-пикера, а сам пикер построен на нём — копий быть не должно."""
        picker = (ROOT / "src" / "components" / "ui" / "DateRangePicker.jsx").read_text(encoding="utf-8-sig")
        self.assertIn("export function IosDateRangeCalendar(", picker)
        self.assertIn("export function IosDateRangePicker(", picker)
        self.assertIn("<IosDateRangeCalendar", picker)
        # Сетка месяца существует в одном экземпляре: в чипе её больше нет.
        self.assertEqual(picker.count("daysInMonth(view.y, view.m)"), 1)

    def test_export_file_name_carries_the_period(self):
        """Имя собирает фронт: Content-Disposition через CORS сюда не доходит."""
        self.assertIn("const chatExportFileName = (snapshot, from, to) => {", self.view)
        self.assertIn("if (from && to && from !== to) return `szov_wallboard_chat_${day(from)}_${day(to)}.xlsx`;",
                      self.view)
        self.assertIn("String(snapshot?.chat2desk_now || '').slice(11, 16)", self.view)

    def test_export_failure_is_reported_by_toast(self):
        """showToast держим в ref: он новый на каждом рендере, а табло рендерится по опросу."""
        self.assertIn("Не удалось выгрузить показатели", self.view)
        self.assertIn("const toastRef = useRef(showToast);", self.view)
        element = self.view[self.view.index("<ChatExportControls"):]
        element = element[:element.index("/>")]
        for prop in ('apiBaseUrl={apiBaseUrl}', 'withAccessTokenHeader={withAccessTokenHeader}',
                     'showToast={showToast}', 'snapshot={snapshot}'):
            self.assertIn(prop, element, prop)

    def test_hourly_chart_can_show_who_was_on_the_line(self):
        """Переключатель у «По часам»: график тот же, меняется только подсказка (запрос владельца)."""
        self.assertIn("export const HOURLY_TOOLTIP_MODES = [", self.board)
        self.assertIn("{ key: 'people', label: 'Кто на линии'", self.board)
        self.assertIn("<SegmentedSwitch", self.board)
        self.assertIn("tooltipMode === 'people' ? <PeopleTooltip /> : <ChartTooltip />", self.board)
        self.assertIn("people: Array.isArray(row.operators) ? row.operators : []", self.board)
        self.assertIn("underMinute: Number(row.operators_under_minute) || 0", self.board)

    def test_people_on_the_line_are_whole_people(self):
        """Дробей на графике больше нет: считаем головами (решение владельца 19.08.2026, вечер)."""
        self.assertIn("return String(Math.round(number));", self.board)
        self.assertNotIn("rounded.toFixed(1)", self.board)
        self.assertIn('allowDecimals={false}', self.board)

    def test_chart_hours_are_labelled_as_intervals(self):
        """«12–13», как в почасовом отчёте: иначе непонятно, промежуток это или накопление."""
        self.assertIn("const hourLabel = (hour) => `${String(hour).padStart(2, '0')}–"
                      "${String((hour + 1) % 24).padStart(2, '0')}`;", self.board)
        self.assertIn("rows || []", self.board)


# Имена сборщика выгрузки. Отдельным набором: файл собирается из готовых дней, источники
# Chat2Desk ему не нужны, и тянуть в namespace весь их обвес было бы шумом.
EXPORT_NAMES = {
    'SZOV_CHAT_WALLBOARD_TARGET_SECONDS',
    'SZOV_CHAT_EXPORT_MAX_DAYS',
    'SZOV_CHAT_EXPORT_WORKERS',
    '_SZOV_CHAT_EXPORT_STATUS_FILL',
    '_szov_chat_export_minutes',
    '_szov_chat_export_hour_label',
    '_szov_chat_export_parse_day',
    '_szov_chat_export_period',
    '_szov_chat_export_day_people',
    '_szov_chat_export_finish_day',
    '_szov_chat_export_today_block',
    '_szov_chat_wallboard_reply_average',
    '_szov_chat_export_file_name',
    '_szov_chat_wallboard_workbook',
}


def _export_namespace():
    """namespace для сборщика выгрузки: openpyxl настоящий, дни подаём руками."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    source = (ROOT / "bot_schedule2.py").read_text(encoding="utf-8-sig")
    ns = {
        'math': math,
        'datetime': datetime,
        'timedelta': timedelta,
        'ZoneInfo': ZoneInfo,
        'BytesIO': BytesIO,
        'Workbook': Workbook,
        'Font': Font,
        'PatternFill': PatternFill,
        'Alignment': Alignment,
        'Border': Border,
        'Side': Side,
        'get_column_letter': get_column_letter,
        'CHAT_HOURLY_TIMEZONE': 'Asia/Almaty',
        '_env_int': lambda name, default, minimum=None, maximum=None: default,
    }
    _load_names(source, EXPORT_NAMES, ns)
    return ns


# Час выгрузки: люди уже посчитаны головами, как их отдаёт _szov_chat_wallboard_hourly.
# `replies` — разрез времени ответа по отвечавшим, кортежами (ФИО, чатов, первый, внутри).
# В строку он едет СУММАМИ, как его отдаёт сервер: складывать часы разных дней можно только
# суммами, среднее средних дало бы третью цифру.
def _export_hour(hour, chats, inner, first, people, *, under=0, partial=False, replies=()):
    return {
        'hour': hour, 'chats': chats, 'inner_reply_seconds': inner, 'first_reply_seconds': first,
        'operators_online': len(people),
        'operators': [{'name': name, 'seconds': seconds} for name, seconds in people],
        'operators_reply': [
            {'name': name, 'chats': count,
             'first_sum': float(first_seconds) * count, 'first_count': count,
             'inner_sum': float(inner_seconds) * count, 'inner_count': count}
            for name, count, first_seconds, inner_seconds in replies
        ],
        'operators_under_minute': under,
        'online_seconds': sum(seconds for _name, seconds in people),
        'partial': partial,
    }


ALIA, BEK, DANA = 'Тестова Алия Тестовна', 'Примеров Бекзат Примерулы', 'Ночная Дана Сменовна'
# Кто попадает в разрез «оператор × час»: те, кто в дне выгрузки отвечал хотя бы в одном часу.
BOARD_CHATTERS = (ALIA, BEK, DANA)


def _export_day(ns, day, *, chats_open=None, error=None, truncated=False, unmatched=(),
                partial_last=False):
    # Разрез по отвечавшим подобран так, чтобы его среднее совпадало с итогом часа: в жизни
    # это выполняется само (обращение без оператора не несёт времени ответа вовсе), и файл
    # обязан сходиться сам с собой.
    hourly = [
        _export_hour(9, 10, 90.0, 60.0, [(ALIA, 3600), (DANA, 1800)], under=1,
                     replies=[(ALIA, 6, 48.0, 78.0), (DANA, 4, 78.0, 108.0)]),
        _export_hour(10, 30, 210.0, 120.0, [(ALIA, 3600), (BEK, 3600), (DANA, 900)],
                     replies=[(ALIA, 15, 96.0, 180.0), (BEK, 10, 156.0, 240.0),
                              (DANA, 5, 120.0, 240.0)]),
        _export_hour(11, 0, None, None, []),
        _export_hour(15, 6, 105.0, 75.0, [(ALIA, 1800)], partial=partial_last,
                     replies=[(ALIA, 5, 72.0, 96.0), (BEK, 1, 90.0, 150.0)]),
    ]
    return ns['_szov_chat_export_finish_day'](day, hourly, {
        'chats': 46, 'chats_open': chats_open,
        'first_reply_seconds': 90.0, 'inner_reply_seconds': 150.0,
        'reply_sums': {'first_sum': 4140.0, 'first_count': 46,
                       'inner_sum': 6900.0, 'inner_count': 46},
    }, events_truncated=truncated, unmatched_names=unmatched, error=error)


NOW_BLOCK = {
    'operators_online': 3, 'operators_busy': 1, 'operators_on_training': 1,
    'operators_on_break': 2, 'operators_on_holiday': 1, 'operators_other': 0,
    'operators_offline': 4, 'open_chats': 12,
    'operators': [
        {'name': ALIA, 'status': 'Онлайн', 'status_key': 'online', 'seconds': 1800, 'open_chats': 7,
         'chats': 26, 'first_reply_seconds': 78.0, 'inner_reply_seconds': 108.0},
        {'name': DANA, 'status': 'Онлайн', 'status_key': 'online', 'seconds': None, 'open_chats': 5,
         'chats': 9, 'first_reply_seconds': 96.0, 'inner_reply_seconds': 180.0},
        {'name': 'Учебный Ерлан Тренингулы', 'status': 'Тренинг', 'status_key': 'training',
         'seconds': 600, 'open_chats': 0,
         'chats': 0, 'first_reply_seconds': None, 'inner_reply_seconds': None},
    ],
}


class ChatWallboardExportPeriodTests(unittest.TestCase):
    """Период выгрузки: что разрешено просить и как это превращается в дни."""

    @classmethod
    def setUpClass(cls):
        cls.ns = _export_namespace()

    def _period(self, date_from, date_to, today=date(2026, 8, 19)):
        return self.ns['_szov_chat_export_period'](date_from, date_to, today)

    def test_no_dates_mean_today(self):
        """Самый частый случай — глянул на табло и забрал цифры: он не стоит ни одного запроса."""
        self.assertEqual(self._period(None, None), ['2026-08-19'])

    def test_period_is_expanded_day_by_day(self):
        self.assertEqual(self._period('2026-08-17', '2026-08-19'),
                         ['2026-08-17', '2026-08-18', '2026-08-19'])

    def test_reversed_period_is_straightened(self):
        self.assertEqual(self._period('2026-08-19', '2026-08-17'),
                         ['2026-08-17', '2026-08-18', '2026-08-19'])

    def test_one_date_is_a_single_day(self):
        self.assertEqual(self._period('2026-08-17', None), ['2026-08-17'])
        self.assertEqual(self._period(None, '2026-08-17'), ['2026-08-17'])

    def test_tail_in_the_future_is_cut_off(self):
        """Chat2Desk отдал бы пустые сутки, а в файле они читались бы как день простоя."""
        self.assertEqual(self._period('2026-08-18', '2026-08-25'), ['2026-08-18', '2026-08-19'])

    def test_broken_and_oversized_periods_are_refused(self):
        for date_from, date_to in (('2026-13-01', None), (None, 'вчера')):
            with self.assertRaises(ValueError):
                self._period(date_from, date_to)
        with self.assertRaises(ValueError):
            self._period('2026-08-25', '2026-08-26')  # весь период в будущем
        with self.assertRaises(ValueError):
            self._period('2026-01-01', '2026-08-19')  # больше потолка суток

    def test_week_is_the_ceiling(self):
        """Потолок — неделя (решение владельца): каждый прошедший день качается отдельно."""
        self.assertEqual(self.ns['SZOV_CHAT_EXPORT_MAX_DAYS'], 7)
        self.assertEqual(len(self._period('2026-08-13', '2026-08-19')), 7)
        with self.assertRaises(ValueError) as refused:
            self._period('2026-08-12', '2026-08-19')  # восемь суток
        self.assertIn('7 суток', str(refused.exception))

    def test_file_name_says_what_is_inside(self):
        name = self.ns['_szov_chat_export_file_name']
        self.assertEqual(name({'from': '2026-08-17', 'to': '2026-08-19'}),
                         'szov_wallboard_chat_20260817_20260819.xlsx')
        # У сегодняшнего дня в имени стоит момент данных, а не скачивания.
        self.assertEqual(name({'from': '2026-08-19', 'to': '2026-08-19', 'day': '2026-08-19',
                               'chat2desk_now': '2026-08-19 15:20:30'}),
                         'szov_wallboard_chat_20260819_1520.xlsx')
        self.assertEqual(name({'from': '2026-08-17', 'to': '2026-08-17'}),
                         'szov_wallboard_chat_20260817.xlsx')


class ChatWallboardExcelExportTests(unittest.TestCase):
    """Выгрузка показателей табло по чатам в .xlsx: те же цифры, что на экране."""

    @classmethod
    def setUpClass(cls):
        cls.ns = _export_namespace()

    def _payload(self, days, **overrides):
        payload = {
            'from': days[0]['day'], 'to': days[-1]['day'], 'today': '2026-08-19',
            'days': days, 'target_seconds': 120, 'now': NOW_BLOCK,
            'chat2desk_now': '2026-08-19 15:20:30', 'stale': False,
            'generated_at': '2026-08-19 15:21',
        }
        payload.update(overrides)
        return payload

    def _workbook(self, payload):
        from openpyxl import load_workbook
        return load_workbook(BytesIO(self.ns['_szov_chat_wallboard_workbook'](payload)))

    def _today(self):
        return self._payload([_export_day(self.ns, '2026-08-19', chats_open=12, partial_last=True)])

    def _period(self):
        return self._payload([
            _export_day(self.ns, '2026-08-17', truncated=True, unmatched=['Посторонний Оператор']),
            _export_day(self.ns, '2026-08-18'),
            _export_day(self.ns, '2026-08-19', chats_open=12, partial_last=True),
        ], stale=True, age_seconds=420, error='Chat2Desk не отвечает')

    def _find_row(self, sheet, label):
        for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
            if row and row[0] == label:
                return row
        raise AssertionError(f'строка «{label}» не найдена')

    def test_minutes_keep_the_gap_between_no_data_and_zero(self):
        """None -> пустая ячейка, а не ноль: иначе «мерить было нечего» станет идеалом."""
        minutes = self.ns['_szov_chat_export_minutes']
        self.assertIsNone(minutes(None))
        self.assertIsNone(minutes('—'))
        self.assertEqual(minutes(0), 0.0)
        self.assertEqual(minutes(150), 2.5)
        self.assertEqual(minutes(13824, 0), 230.0)

    def test_hour_is_labelled_as_an_interval(self):
        """Час — промежуток, как в почасовом отчёте, а полночь замыкает сутки."""
        self.assertEqual(self.ns['_szov_chat_export_hour_label'](12), '12:00–13:00')
        self.assertEqual(self.ns['_szov_chat_export_hour_label'](23), '23:00–00:00')

    def test_single_day_has_no_day_sheet(self):
        """На однодневной выгрузке лист «По дням» повторял бы сводку — его нет."""
        self.assertEqual(self._workbook(self._today()).sheetnames,
                         ['Показатели', 'По часам', 'Первый ответ по часам', 'Ответ внутри чата по часам', 'Чатники'])

    def test_period_adds_the_day_sheet(self):
        self.assertEqual(self._workbook(self._period()).sheetnames,
                         ['Показатели', 'По дням', 'По часам', 'Первый ответ по часам', 'Ответ внутри чата по часам', 'Чатники'])

    def test_shift_sheet_appears_only_with_today_inside(self):
        """Состава смены «сейчас» у прошедшего дня не существует: живой список — только на сейчас."""
        payload = self._payload([_export_day(self.ns, '2026-08-17')], now=None, chat2desk_now=None)
        wb = self._workbook(payload)
        self.assertNotIn('Чатники', wb.sheetnames)
        with self.assertRaises(AssertionError):
            self._find_row(wb['Показатели'], 'Открыто чатов')
        self.assertEqual(self._find_row(wb['Показатели'], 'Дата')[1], '2026-08-17')

    def test_summary_repeats_the_tiles_of_the_board(self):
        sheet = self._workbook(self._today())['Показатели']
        self.assertEqual(self._find_row(sheet, 'Дата')[1], '2026-08-19')
        self.assertEqual(self._find_row(sheet, 'Данные Chat2Desk на')[1], '2026-08-19 15:20:30')
        self.assertEqual(self._find_row(sheet, 'Цель по ответу внутри чата, мин')[1], 2.0)
        self.assertEqual(self._find_row(sheet, 'Онлайн')[1], 3)
        self.assertEqual(self._find_row(sheet, 'Не в системе')[1], 4)
        self.assertEqual(self._find_row(sheet, 'Открыто чатов')[1], 12)
        self.assertEqual(self._find_row(sheet, 'Чатов')[1], 46)
        self.assertEqual(self._find_row(sheet, 'Открыто сейчас')[1], 12)
        self.assertEqual(self._find_row(sheet, 'Ответ внутри чата, мин')[1], 2.5)
        self.assertEqual(self._find_row(sheet, 'Первый ответ, мин')[1], 1.5)

    def test_summary_has_no_column_of_explanations(self):
        """Пояснительных слов в выгрузке нет (решение владельца 20.08.2026): подписи
        показателей говорят сами за себя, а колонка текста рядом с числами удлиняла лист."""
        sheet = self._workbook(self._today())['Показатели']
        headers = [row for row in sheet.iter_rows(min_col=1, max_col=3, values_only=True)
                   if row and row[0] == 'Показатель']
        self.assertTrue(headers, 'не нашёл шапку блока')
        for header in headers:
            self.assertEqual(header[1], 'Значение')
            self.assertIsNone(header[2], 'колонка пояснений вернулась')

    def test_vacation_counter_is_gone_from_the_file(self):
        """Отпуска в Chat2Desk не существует вовсе (`holiday` — это «Закрытие чатов»),
        отдельной строкой он не нужен — решение владельца 20.08.2026. Сами люди не теряются:
        они стоят на листе «Чатники» со своим статусом."""
        wb = self._workbook(self._today())
        with self.assertRaises(AssertionError):
            self._find_row(wb['Показатели'], 'В отпуске')
        self.assertNotIn('Отпуск', str([cell.value for cell in wb['Показатели']['A']]))

    def test_hours_in_target_is_gone_from_the_file(self):
        """«Часов в цели» убран по решению владельца 20.08.2026 — и со стены, и из файла.

        Оценка дня задним числом: по ней ничего не сделаешь, а место она занимала наравне с
        показателями, на которые смена реагирует прямо сейчас."""
        wb = self._workbook(self._period())
        for label in ('Часов в цели', 'Часов измерено'):
            with self.assertRaises(AssertionError):
                self._find_row(wb['Показатели'], label)
        self.assertNotIn('Часов в цели', [cell.value for cell in wb['По дням'][1]])
        # Колонка «В цели» (да/нет) с листа «По часам» тоже убрана 20.08.2026: она повторяла
        # словами то, что уже видно цветом самой минуты.
        self.assertNotIn('В цели', [cell.value for cell in wb['По часам'][1]])

    def test_summary_counts_people_as_heads(self):
        """Людей — головами: трое разных за день и трое в самом плотном часу."""
        sheet = self._workbook(self._today())['Показатели']
        self.assertEqual(self._find_row(sheet, 'Людей на линии')[1], 3)
        self.assertEqual(self._find_row(sheet, 'Максимум людей в часе')[1], 3)
        self.assertEqual(self._find_row(sheet, 'На линии, мин')[1], 255.0)

    def test_period_averages_go_by_requests_not_by_days(self):
        """Среднее за период считается по обращениям всех дней, а не как среднее средних."""
        days = [
            _export_day(self.ns, '2026-08-17'),
            _export_day(self.ns, '2026-08-18'),
        ]
        # Второй день: вдвое больше обращений и ответ по 4 минуты. Среднее средних дало бы
        # 3,25 мин, среднее по обращениям — 3,5.
        days[1]['reply_sums'] = {'first_sum': 100.0, 'first_count': 1,
                                 'inner_sum': 240.0 * 92, 'inner_count': 92}
        days[1]['inner_reply_seconds'] = 240.0
        sheet = self._workbook(self._payload(days))['Показатели']
        self.assertEqual(self._find_row(sheet, 'Ответ внутри чата, мин')[1], 3.5)
        self.assertEqual(self._find_row(sheet, 'Дней в периоде')[1], 2)

    def test_hours_carry_numbers_and_names(self):
        """По файлу считают дальше: минуты — числа, люди — целые, состав часа — ФИО."""
        sheet = self._workbook(self._today())['По часам']
        header = [cell.value for cell in sheet[1]]
        self.assertEqual(header, ['Час', 'Чатов', 'Ответ внутри чата, мин', 'Первый ответ, мин',
                                  'Людей на линии', 'На линии, мин', 'Кто на линии',
                                  'Примечание'])
        first = [cell.value for cell in sheet[2]]
        self.assertEqual(first[0], '09:00–10:00')
        self.assertEqual(first[1:6], [10, 1.5, 1.0, 2, 90.0])
        self.assertEqual(first[6], f'{ALIA}, {DANA}')
        self.assertEqual(first[7], 'ещё 1 меньше минуты')

    def test_hour_without_answers_stays_empty_and_current_hour_is_marked(self):
        sheet = self._workbook(self._today())['По часам']
        empty = [cell.value for cell in sheet[4]]
        self.assertIsNone(empty[2])
        self.assertIsNone(empty[3])
        self.assertEqual(empty[4], 0)
        current = [cell.value for cell in sheet[5]]
        # Уложились в цель (105 c против 120 c) — теперь это видно цветом самой минуты,
        # а не отдельной колонкой «В цели».
        self.assertEqual(sheet.cell(row=5, column=3).font.color.rgb[-6:], '15803D')
        self.assertEqual(current[7], 'час идёт')

    def test_daily_total_row_keeps_heads_out_of_the_sum(self):
        """В итоге по людям стоит максимум в часе: суммировать головы разных часов нельзя."""
        sheet = self._workbook(self._today())['По часам']
        total = [cell.value for cell in sheet[6]]
        self.assertEqual(total[0], 'За сутки')
        self.assertEqual(total[1], 46)
        self.assertEqual(total[2], 2.5)
        self.assertEqual(total[4], 3)
        self.assertEqual(total[6], 'разных людей: 3')
        # Фильтр стоит только по часам: строка итога под ним.
        self.assertEqual(sheet.auto_filter.ref, 'A1:H5')
        self.assertEqual(len([cell.value for cell in sheet[1]]), 8)

    def test_period_hours_carry_the_day(self):
        sheet = self._workbook(self._period())['По часам']
        self.assertEqual([cell.value for cell in sheet[1]][0], 'День')
        self.assertEqual([cell.value for cell in sheet[2]][0], '2026-08-17')
        self.assertEqual([cell.value for cell in sheet[6]][0], '2026-08-18')
        total = [cell.value for cell in sheet[14]]
        self.assertEqual(total[0], '2026-08-17 — 2026-08-19')
        self.assertEqual(total[1], 'За период')

    def test_day_sheet_lists_every_day_with_its_note(self):
        sheet = self._workbook(self._period())['По дням']
        rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_col=8)]
        self.assertEqual([row[0] for row in rows], ['2026-08-17', '2026-08-18', '2026-08-19'])
        self.assertEqual(rows[0][1:7], [46, 2.5, 1.5, 3, 3, 255.0])
        self.assertEqual(rows[0][7], 'события обрезаны')
        self.assertFalse(rows[1][7])

    def test_failed_day_says_why_instead_of_showing_zeros(self):
        """День не дался — в файле причина, а не нули: нули прочитались бы как простой смены."""
        payload = self._payload([
            self.ns['_szov_chat_export_finish_day']('2026-08-17', [], {},
                                                    error='Chat2Desk HTTP 500'),
            _export_day(self.ns, '2026-08-19', chats_open=12),
        ])
        wb = self._workbook(payload)
        notes = [row[1] for row in wb['Показатели'].iter_rows(min_col=1, max_col=2, values_only=True)
                 if row and row[0] == 'Внимание']
        self.assertTrue(any('2026-08-17' in str(note) and 'HTTP 500' in str(note) for note in notes))
        day_rows = [[cell.value for cell in row] for row in wb['По дням'].iter_rows(min_row=2, max_col=8)]
        self.assertEqual(day_rows[0][1], 0)
        self.assertIn('HTTP 500', day_rows[0][7])

    def test_line_time_sheet_is_replaced_by_the_reply_cut(self):
        """Лист «Люди на линии» снят по решению владельца 20.08.2026.

        Минуты на линии никуда не делись — они в «По часам» и в сводке; а времени ответа по
        людям в файле не было нигде, и именно за ним в выгрузку и ходят."""
        self.assertNotIn('Люди на линии', self._workbook(self._period()).sheetnames)

    def test_reply_sheets_lay_operators_against_hours(self):
        """Строка — чатник, колонка — час, в клетке минуты: разрез, которого в файле не было."""
        sheet = self._workbook(self._today())['Первый ответ по часам']
        # Колонки — только часы, в которые чаты приходили: у часа 11 их не было, и колонки нет.
        self.assertEqual([cell.value for cell in sheet[1]],
                         ['Сотрудник', 'Чатов', 'Среднее, мин',
                          '09:00–10:00', '10:00–11:00', '15:00–16:00'])
        rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_col=6)]
        # Сверху тот, кто взял больше чатов за период.
        self.assertEqual(rows[0], [ALIA, 26, 1.3, 0.8, 1.6, 1.2])
        self.assertEqual(rows[1], [BEK, 11, 2.5, None, 2.6, 1.5])
        self.assertEqual(rows[2], [DANA, 9, 1.7, 1.3, 2.0, None])

    def test_inner_reply_sheet_is_the_same_shape_by_its_own_metric(self):
        """Два показателя — два листа: в одной таблице их пришлось бы чередовать колонками."""
        sheet = self._workbook(self._today())['Ответ внутри чата по часам']
        self.assertEqual([cell.value for cell in sheet[1]][:3], ['Сотрудник', 'Чатов', 'Среднее, мин'])
        rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_col=6)]
        self.assertEqual(rows[0], [ALIA, 26, 2.3, 1.3, 3.0, 1.6])
        self.assertEqual(rows[1], [BEK, 11, 3.9, None, 4.0, 2.5])
        self.assertEqual(rows[2], [DANA, 9, 3.0, 1.8, 4.0, None])

    def test_reply_cut_adds_up_to_the_hour_on_the_hours_sheet(self):
        """Итоговая строка — сумма колонки, и она обязана сойтись с листом «По часам».

        В жизни это выполняется само: обращение, которое никто не взял, не несёт времени
        ответа вовсе, поэтому разрез по людям складывается ровно в итог часа."""
        wb = self._workbook(self._today())
        hours = {row[0]: row for row in
                 ([cell.value for cell in line] for line in wb['По часам'].iter_rows(min_row=2))}
        for title, column in (('Первый ответ по часам', 3), ('Ответ внутри чата по часам', 2)):
            sheet = wb[title]
            total = [cell.value for cell in sheet[len(BOARD_CHATTERS) + 2]]
            self.assertEqual(total[0], 'Все чатники')
            self.assertEqual(total[1], 46)  # все 46 чатов дня разошлись по людям
            for index, label in enumerate(('09:00–10:00', '10:00–11:00', '15:00–16:00'), start=3):
                self.assertEqual(total[index], hours[label][column], f'{title} · {label}')

    def test_reply_sheet_says_when_chats_were_left_without_an_operator(self):
        """Разница между итогом дня и суммой колонки «Чатов» подписана, а не оставлена загадкой."""
        day = _export_day(self.ns, '2026-08-19')
        day['chats'] = 50  # четыре чата закрылись сами, оператора у них не было
        sheet = self._workbook(self._payload([day]))['Первый ответ по часам']
        notes = [cell.value for cell in sheet['A'] if isinstance(cell.value, str)]
        self.assertIn('Чатов без оператора: 4', notes)

    def test_reply_sheet_survives_a_day_nobody_answered(self):
        """Ночь без чатов: лист собирается и говорит об этом словами, а не пустой сеткой."""
        payload = self._payload(
            [self.ns['_szov_chat_export_finish_day']('2026-08-19', [], {'chats': 0})],
            now={'operators_online': 0, 'operators_offline': 0, 'open_chats': 0, 'operators': []})
        sheet = self._workbook(payload)['Ответ внутри чата по часам']
        self.assertEqual([cell.value for cell in sheet[1]], ['Сотрудник', 'Чатов', 'Среднее, мин'])
        self.assertIn('За период никто не отвечал в чатах',
                      [cell.value for cell in sheet['A'] if isinstance(cell.value, str)])

    def test_reply_cut_of_a_period_sums_the_same_hour_of_every_day(self):
        """Клетка периода — один и тот же час РАЗНЫХ дней, сложенный суммами, а не средними."""
        sheet = self._workbook(self._period())['Первый ответ по часам']
        rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_col=6)]
        # Три одинаковых дня: чатов втрое больше, а средние — те же самые.
        self.assertEqual(rows[0], [ALIA, 78, 1.3, 0.8, 1.6, 1.2])
        self.assertEqual(rows[1], [BEK, 33, 2.5, None, 2.6, 1.5])

    def test_inner_reply_cells_are_coloured_against_the_goal(self):
        """Цвет — только у ответа внутри чата: цель задана ему, у первого ответа её нет."""
        wb = self._workbook(self._today())
        inner, first = wb['Ответ внутри чата по часам'], wb['Первый ответ по часам']
        # Час 09 у ALIA — 1,3 мин при цели 2 мин: зелёный; час 10 — 3,0 мин: красный.
        self.assertEqual(inner.cell(row=2, column=4).font.color.rgb[-6:], '15803D')
        self.assertEqual(inner.cell(row=2, column=5).font.color.rgb[-6:], 'B91C1C')
        # У первого ответа нормы нет — красить его той же меркой значило бы её выдумать:
        # цвет шрифта остаётся темой книги, а не зелёным/красным вердиктом.
        self.assertEqual(first.cell(row=2, column=5).font.color.type, 'theme')

    def test_day_people_follow_the_same_threshold_as_hours(self):
        """Атом счёта один: день складывается из тех же людей, что стоят в подсказке часа."""
        people = self.ns['_szov_chat_export_day_people']([
            _export_hour(9, 0, None, None, [(ALIA, 3600)], under=2),
            _export_hour(10, 0, None, None, [(ALIA, 600), (DANA, 1800)]),
        ])
        self.assertEqual(people, {ALIA: 4200, DANA: 1800})

    def test_shift_sheet_keeps_the_since_midnight_wording(self):
        """seconds=None — человек в статусе с начала суток; «0 мин» тут соврало бы."""
        sheet = self._workbook(self._today())['Чатники']
        rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2, max_col=4)]
        self.assertEqual(rows[0], [ALIA, 'Онлайн', 30.0, 7])
        self.assertEqual(rows[1], [DANA, 'Онлайн', 'с начала суток', 5])
        self.assertEqual(rows[2], ['Учебный Ерлан Тренингулы', 'Тренинг', 10.0, 0])
        # Вышедшие в список не идут, но их число видно — иначе смена выглядит меньше.
        self.assertIn('Не в системе: 4',
                      [cell.value for cell in sheet['A'] if isinstance(cell.value, str)])

    def test_status_colour_matches_the_chip_on_the_board(self):
        sheet = self._workbook(self._today())['Чатники']
        self.assertIn(self.ns['_SZOV_CHAT_EXPORT_STATUS_FILL']['online'],
                      str(sheet.cell(row=2, column=2).fill.fgColor.rgb))
        self.assertIn(self.ns['_SZOV_CHAT_EXPORT_STATUS_FILL']['training'],
                      str(sheet.cell(row=4, column=2).fill.fgColor.rgb))

    def test_file_warns_about_stale_and_truncated_data(self):
        """Оговорки живут В ФАЙЛЕ: он уходит в переписку без экрана, с которого его сняли."""
        sheet = self._workbook(self._period())['Показатели']
        notes = [row[1] for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True)
                 if row and row[0] == 'Внимание']
        self.assertEqual(len(notes), 3)
        self.assertIn('7 мин', notes[0])
        self.assertIn('Chat2Desk не отвечает', notes[0])
        self.assertIn('обрезаны', notes[1])
        self.assertIn('2026-08-17', notes[1])
        self.assertIn('Посторонний Оператор', notes[2])

    def test_clean_snapshot_has_no_warnings(self):
        """Шум = брак: пока всё в порядке, строк «Внимание» в файле нет."""
        sheet = self._workbook(self._today())['Показатели']
        self.assertEqual([row[0] for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True)
                          if row and row[0] == 'Внимание'], [])

    def test_empty_day_does_not_break_the_file(self):
        """Ночь, чатов ещё нет: файл собирается, а не падает на пустых списках."""
        payload = self._payload(
            [self.ns['_szov_chat_export_finish_day']('2026-08-19', [], {'chats': 0})],
            now={'operators_online': 0, 'operators_offline': 0, 'open_chats': 0, 'operators': []})
        wb = self._workbook(payload)
        self.assertEqual([cell.value for cell in wb['По часам'][2]][0], 'За сутки')
        self.assertIsNone(self._find_row(wb['Показатели'], 'Ответ внутри чата, мин')[1])


class ChatBroadcastTests(unittest.TestCase):
    """Отбивка направления «Чат»: что считаем отклонением, что пишем в сообщении.

    Снимок подсовываем готовым — сбор данных проверен тестами самого табло, а здесь важно
    ровно то, что видит получатель в Telegram."""

    SNAPSHOT = {
        'day': '2026-08-20',
        'chat2desk_now': '2026-08-20 14:05:00',
        'target_seconds': 120,
        'now': {'operators_online': 6, 'operators_busy': 2, 'operators_on_training': 1,
                'operators_on_break': 3, 'operators_offline': 8, 'open_chats': 41,
                'operators': []},
        'today': {'chats': 512, 'chats_open': 41,
                  'first_reply_seconds': 74.2, 'inner_reply_seconds': 187.5},
        'stale': False,
        'age_seconds': 30,
    }

    def _namespace(self, snapshot=None, break_violations=None):
        source = (ROOT / "bot_schedule2.py").read_text(encoding="utf-8-sig")
        ns = {
            'os': os, 're': re, 'logging': logging,
            'datetime': datetime, 'ZoneInfo': ZoneInfo,
            'CHAT_HOURLY_TIMEZONE': 'Asia/Almaty',
            'SZOV_CHAT_WALLBOARD_TARGET_SECONDS': 120,
            '_env_int': lambda name, default, minimum=None, maximum=None: default,
            '_szov_chat_wallboard_snapshot': lambda: (self.SNAPSHOT if snapshot is None
                                                      else snapshot),
            # Журнал перерывов лежит в БД, его наполняет отдельный разбор (задача #114):
            # здесь подставляем готовые строки, а сборку текста проверяем настоящую.
            '_szov_chat_broadcast_break_violations': (
                lambda scheduled=False, now=None: list(break_violations or [])
            ),
        }
        _load_names(source, {
            'SZOV_BROADCAST_SEND_TIMES', '_SZOV_BROADCAST_HOURLY',
            'SZOV_CHAT_BROADCAST_SEND_TIMES', 'SZOV_CHAT_BROADCAST_MIN_CHATS',
            '_szov_wallboard_int', '_szov_plural', '_szov_format_seconds_mmss',
            '_szov_format_age_ru', '_szov_broadcast_parse_times', '_szov_broadcast_stale_note',
            '_szov_chat_broadcast_send_times', '_szov_chat_broadcast_duration',
            '_szov_chat_broadcast_collect', '_szov_chat_broadcast_deviations',
            '_szov_chat_broadcast_notes', '_szov_chat_broadcast_text',
            'SZOV_BREAK_NOTE_LIMIT', 'SZOV_BREAK_KIND_OFF_SCHEDULE',
            'SZOV_BREAK_KIND_NOT_PLANNED', 'SZOV_BREAK_KIND_NO_SHIFT',
            '_szov_break_violation_detail', '_szov_break_violation_notes',
        }, ns)
        return ns

    def _data(self, snapshot=None, break_violations=None):
        ns = self._namespace(snapshot, break_violations)
        return ns, ns['_szov_chat_broadcast_collect']()

    def test_break_violations_reach_the_chat_broadcast(self):
        """Тот же контроль, что у «Линии»: строка приходит почасовым отчётом чатов."""
        ns, data = self._data(break_violations=[{
            'operator_name': 'Иванов Иван', 'started_at': '2026-08-19T16:30:00',
            'kind': 'off_schedule', 'planned_start_minutes': 14 * 60,
        }])
        text = ns['_szov_chat_broadcast_text'](data)
        self.assertIn('Обратите внимание: Иванов Иван вышел(а) на перерыв не по графику', text)
        self.assertIn('по графику в 14:00', text)

    def test_break_violation_is_a_routine_line_for_chat_too(self):
        """Решение владельца: такой строкой чат «только при отклонениях» не будим."""
        # Снимок берём заведомо спокойный: иначе отклонением стал бы медленный ответ,
        # и тест перестал бы проверять то, ради чего написан.
        ns, data = self._data(snapshot=self._with(inner_reply_seconds=60.0), break_violations=[{
            'operator_name': 'Иванов Иван', 'started_at': '2026-08-19T16:30:00',
            'kind': 'not_planned', 'planned_start_minutes': None,
        }])
        self.assertEqual(ns['_szov_chat_broadcast_deviations'](data), [])
        self.assertIn('перерывов в графике на этот день нет',
                      ' '.join(ns['_szov_chat_broadcast_notes'](data)))

    def test_quiet_hour_adds_no_break_line(self):
        ns, data = self._data(break_violations=[])
        self.assertNotIn('не по графику', ns['_szov_chat_broadcast_text'](data))

    @staticmethod
    def _with(**today):
        snapshot = dict(ChatBroadcastTests.SNAPSHOT)
        snapshot['today'] = dict(snapshot['today'], **today)
        return snapshot

    # --- отклонения ---

    def test_slow_inner_reply_is_a_deviation(self):
        """Единственный оценочный показатель направления — тот же, что красит плитку."""
        ns, data = self._data()
        notes = ns['_szov_chat_broadcast_deviations'](data)
        self.assertEqual(len(notes), 1)
        self.assertIn('3:07', notes[0])
        self.assertIn('2:00', notes[0])

    def test_reply_within_the_goal_is_not_a_deviation(self):
        ns, data = self._data(self._with(inner_reply_seconds=90.0))
        self.assertEqual(ns['_szov_chat_broadcast_deviations'](data), [])

    def test_exactly_on_the_goal_is_still_within_norm(self):
        """Граница включительная — как chatReplyTone на фронте (value <= target)."""
        ns, data = self._data(self._with(inner_reply_seconds=120))
        self.assertEqual(ns['_szov_chat_broadcast_deviations'](data), [])

    def test_a_handful_of_night_chats_does_not_raise_an_alarm(self):
        """«Ответ внутри чата» — среднее по суткам: в 00:30 оно стоит на паре обращений, и
        один медленный ответ поднял бы тревогу там, где нет ни очереди, ни проблемы."""
        ns, data = self._data(self._with(chats=2, inner_reply_seconds=600.0))
        self.assertEqual(ns['_szov_chat_broadcast_deviations'](data), [])

    def test_the_same_delay_on_a_full_day_does_raise_it(self):
        """Порог выборки не должен глушить настоящую проблему."""
        ns, data = self._data(self._with(chats=200, inner_reply_seconds=600.0))
        self.assertEqual(len(ns['_szov_chat_broadcast_deviations'](data)), 1)

    def test_missing_reply_time_is_not_a_deviation(self):
        """Обращений не было — сравнивать нечего; ноль здесь означал бы «отвечаем мгновенно»."""
        ns, data = self._data(self._with(chats=0, first_reply_seconds=None,
                                         inner_reply_seconds=None))
        self.assertEqual(ns['_szov_chat_broadcast_deviations'](data), [])

    def test_frozen_snapshot_names_chat2desk_not_oktell(self):
        """Получателю важно, КТО молчит: от этого зависит, к кому идти."""
        stale = dict(self.SNAPSHOT, stale=True, age_seconds=2400)
        ns, data = self._data(stale)
        notes = ns['_szov_chat_broadcast_deviations'](data)
        self.assertTrue(any('Chat2Desk' in note for note in notes), notes)
        self.assertFalse(any('Oktell' in note for note in notes), notes)
        self.assertTrue(any('40 минут' in note for note in notes), notes)

    # --- примечания и текст ---

    def test_duty_lines_are_not_deviations(self):
        """Иначе чат в режиме «только при отклонениях» получал бы сообщение каждый раз."""
        ns, data = self._data(self._with(inner_reply_seconds=90.0))
        notes = ns['_szov_chat_broadcast_notes'](data)
        self.assertEqual(ns['_szov_chat_broadcast_deviations'](data), [])
        self.assertTrue(notes)
        self.assertIn('6 чатников на линии', notes[0])
        self.assertIn('3 на перерыве', notes[0])

    def test_empty_statuses_are_left_out_of_the_line(self):
        """Ноль в строке — шум: перечисляем только то, что есть."""
        snapshot = dict(self.SNAPSHOT)
        snapshot['now'] = dict(snapshot['now'], operators_busy=0, operators_on_training=0)
        ns, data = self._data(snapshot)
        line = ns['_szov_chat_broadcast_notes'](data)[1]
        self.assertIn('на перерыве', line)
        self.assertNotIn('заняты', line)
        self.assertNotIn('на тренинге', line)

    def test_an_empty_day_says_so_in_words(self):
        """«Первый ответ — —, внутри чата —» читается как сбой, хотя пустая ночь — норма."""
        ns, data = self._data(self._with(chats=0, first_reply_seconds=None,
                                         inner_reply_seconds=None))
        notes = ns['_szov_chat_broadcast_notes'](data)
        self.assertIn('Обращений за сутки пока не было.', notes)
        self.assertFalse(any('Первый ответ' in note for note in notes), notes)

    def test_text_leads_with_the_deviation(self):
        ns, data = self._data()
        text = ns['_szov_chat_broadcast_text'](data)
        self.assertTrue(text.startswith('<b>Чаты сейчас</b>'))
        self.assertIn('ответ внутри чата дольше цели', text)
        self.assertIn('Первый ответ — 1:14', text)

    def test_frozen_note_goes_last(self):
        """Замершие данные упоминаем после дежурных строк, а не вместо них."""
        stale = dict(self.SNAPSHOT, stale=True, age_seconds=600)
        ns, data = self._data(stale)
        notes = ns['_szov_chat_broadcast_notes'](data)
        self.assertIn('Chat2Desk', notes[-1])
        self.assertEqual([note for note in notes if 'Chat2Desk' in note], [notes[-1]])

    # --- расписание ---

    def test_chat_schedule_follows_the_line_unless_set_apart(self):
        """Владелец просил «такую же»: по умолчанию расписание совпадает с «Линией»."""
        ns = self._namespace()
        self.assertEqual(ns['_szov_chat_broadcast_send_times'](),
                         [(hour, 0) for hour in range(24)])

    def test_broken_chat_send_time_is_skipped_not_fatal(self):
        ns = self._namespace()
        ns['SZOV_CHAT_BROADCAST_SEND_TIMES'] = '09:00, ерунда, 24:00, 21:30'
        self.assertEqual(ns['_szov_chat_broadcast_send_times'](), [(9, 0), (21, 30)])


class ChatBroadcastWiringTests(unittest.TestCase):
    """Отбивка чатов подключена: снимок из кэша, картинка общим рисовальщиком, отправка."""

    @classmethod
    def setUpClass(cls):
        cls.api = (ROOT / "bot_schedule2.py").read_text(encoding="utf-8-sig")

    def _section(self, start, end):
        return self.api[self.api.index(start):self.api.index(end)]

    def test_broadcast_never_spends_chat2desk_quota_of_its_own(self):
        """Снимок направления уже собран для экрана: второй сбор жёг бы квоту компании и
        разводил цифры в Telegram с цифрами на стене."""
        block = self._section("def _szov_chat_broadcast_collect(",
                              "def _szov_chat_broadcast_deviations(")
        self.assertIn("_szov_chat_wallboard_snapshot()", block)
        self.assertNotIn("_chat2desk_", block)
        self.assertNotIn("requests.", block)

    def test_one_picture_goes_as_a_photo_not_an_album(self):
        """sendMediaGroup у Telegram принимает от ДВУХ вложений и на одном честно падает."""
        block = self._section("async def _szov_broadcast_deliver(",
                              "async def _szov_broadcast_send(")
        self.assertIn("if len(media) == 1:", block)
        self.assertIn("await bot.send_photo(", block)
        self.assertIn("group = types.MediaGroup()", block)

    def test_tiles_are_drawn_by_one_shared_renderer(self):
        """Картинки «Линии» и «Чата» лежат в одном чате рядом — оформление разъезжаться не должно."""
        self.assertEqual(self.api.count("def _szov_render_tiles_png("), 1)
        for renderer in ("def _szov_render_wallboard_png(", "def _szov_render_chat_wallboard_png("):
            block = self.api[self.api.index(renderer):]
            self.assertIn("_szov_render_tiles_png(", block[:2500], renderer)
        # Второй копии раскладки плиток в файле нет.
        self.assertEqual(self.api.count("draw.rounded_rectangle([x, y, x + w, y + tile_h]"), 1)

    def test_picture_failure_never_swallows_the_numbers(self):
        block = self._section("async def _szov_chat_broadcast_prepare(",
                              "async def _szov_chat_broadcast_send(")
        self.assertIn("media = []", block)
        self.assertIn("не удалось собрать картинку", block)

    def test_snapshot_is_collected_off_the_bot_loop(self):
        """Chat2Desk — синхронный requests: держать на нём event loop бота недопустимо."""
        block = self._section("async def _szov_chat_broadcast_prepare(",
                              "async def _szov_chat_broadcast_send(")
        self.assertIn("run_in_executor(executor_pool, _szov_chat_broadcast_collect", block)

    def test_preview_shows_the_chat_direction_without_sending(self):
        block = self._section("def _szov_chat_broadcast_preview(",
                              "@app.route('/api/szov_wallboard/broadcast_test'")
        self.assertIn("font_path", block)
        self.assertNotIn("send_media_group", block)
        self.assertNotIn("send_photo", block)
        self.assertNotIn("send_message", block)

    def test_endpoints_default_to_the_line_for_an_older_bundle(self):
        """Фронт едет отдельным деплоем (GitHub Pages против Render): бандл, не знающий о
        втором направлении, обязан продолжать настраивать «Линию», а не получать 400."""
        block = self._section("def _szov_broadcast_direction_arg(",
                              "def _szov_broadcast_direction_times(")
        self.assertIn("or SZOV_BROADCAST_DIRECTION_LINE", block)
        self.assertIn("raise ValueError", block)


if __name__ == '__main__':
    unittest.main()
