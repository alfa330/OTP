"""Тесты «Биллинг Oktell» (Расчет ресурсов).

Проверяют SQL-билдер, разбор времени HH:MM, агрегацию день×таксопарк и
чанковую выгрузку с фолбэком по одному дню при упоре в лимит прокси (1000 строк).

Приём тот же, что в test_oktell_operator_day_gate.py: функции извлекаются из
bot_schedule2.py через AST и исполняются в изолированном namespace со стабами
(модуль целиком импортировать нельзя — сетевые/БД сайд-эффекты на импорте).
"""

import ast
import re
import unittest
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BOT_PATH = Path(__file__).resolve().parents[1] / "bot_schedule2.py"

FUNCTION_NAMES = [
    "_oktell_billing_range_exceeds_limit",
    "_oktell_billing_parse_time",
    "_oktell_billing_minute_filter",
    "_oktell_billing_sql",
    "_oktell_fetch_billing_rows",
    "_oktell_billing_int",
    "_oktell_billing_sec",
    "_oktell_billing_build_report",
    "_oktell_billing_detail_source_sql",
    "_oktell_billing_detail_select_sql",
    "_oktell_billing_detail_line_apply_sql",
    "_oktell_billing_detail_page_sql",
    "_oktell_billing_detail_export_page_sql",
    "_oktell_billing_detail_row",
    "_oktell_billing_build_detail_page",
    "_oktell_fetch_billing_detail_page",
    "_oktell_fetch_billing_detail_export_rows",
    "_oktell_billing_operator_calls_sql",
    "_oktell_billing_operator_states_sql",
    "_oktell_fetch_billing_operator_rows",
    "_oktell_billing_build_operator_report",
    "_oktell_billing_park_label",
    "_oktell_billing_line_digits",
    "_oktell_billing_phone_display",
    "_oktell_billing_line_label",
    "_oktell_billing_ratio",
    "_oktell_billing_export_columns",
    "_oktell_billing_export_values",
    "_oktell_billing_export_workbook",
]

CONST_NAMES = (
    "OKTELL_BILLING_MAX_RANGE_DAYS",
    "_OKTELL_BILLING_PARK_LABELS",
    "_OKTELL_BILLING_LINE_LABELS",
    "_OKTELL_BILLING_EXPORT_DUR_FMT",
    "_OKTELL_BILLING_EXPORT_PCT_FMT",
)


def _extract_namespace(oktell_query=None, page_size=1000, chunk_days=7):
    module = ast.parse(BOT_PATH.read_text(encoding="utf-8"))
    wanted = set(FUNCTION_NAMES)
    selected = [
        node for node in module.body
        if isinstance(node, ast.FunctionDef) and node.name in wanted
    ]
    found = {node.name for node in selected}
    missing = wanted - found
    if missing:
        raise AssertionError("Не найдены функции в bot_schedule2.py: %s" % sorted(missing))
    consts = [
        node for node in module.body
        if isinstance(node, ast.Assign) and any(getattr(t, "id", "") in CONST_NAMES for t in node.targets)
    ]
    ns = {
        "re": re,
        "datetime": datetime,
        "timedelta": timedelta,
        "BytesIO": BytesIO,
        "Workbook": Workbook,
        "Font": Font,
        "PatternFill": PatternFill,
        "get_column_letter": get_column_letter,
        "_OKTELL_GREETING_ABANDON": "Бросили трубку на приветствии",
        "_OKTELL_FAILED_CALL": "Неудачный звонок",
        "_OKTELL_BILLING_METRICS": (
            "arrived", "served", "lost", "served_sl", "greet_drop",
            "talk_seconds", "wait_ok_seconds", "wait_lost_seconds", "total_seconds",
        ),
        "_OKTELL_BILLING_OPERATOR_METRICS": (
            "served", "talk_seconds", "talk_in_seconds", "talk_out_seconds", "postproc_seconds",
            "hold_seconds", "wait_seconds", "pause_seconds", "dial_seconds",
        ),
        "OKTELL_API_PAGE_SIZE": page_size,
        "OKTELL_RESOURCE_CHUNK_DAYS": chunk_days,
        "_oktell_query": oktell_query or (lambda sql: []),
    }
    exec(compile(ast.Module(body=consts + selected, type_ignores=[]), str(BOT_PATH), "exec"), ns)
    return ns


class ParseTimeTests(unittest.TestCase):
    def setUp(self):
        self.ns = _extract_namespace()

    def test_valid_times(self):
        parse = self.ns["_oktell_billing_parse_time"]
        self.assertEqual(parse("00:00"), 0)
        self.assertEqual(parse("08:30"), 510)
        self.assertEqual(parse("23:59"), 1439)
        self.assertEqual(parse("9:05"), 545)
        self.assertEqual(parse(" 10:00 "), 600)

    def test_invalid_times(self):
        parse = self.ns["_oktell_billing_parse_time"]
        for bad in ("24:00", "12:60", "1200", "12", "", None, "ab:cd", "12:5"):
            self.assertIsNone(parse(bad), bad)


class BillingRangeLimitTests(unittest.TestCase):
    def setUp(self):
        self.ns = _extract_namespace()

    def test_exactly_31_days_are_allowed(self):
        exceeds = self.ns["_oktell_billing_range_exceeds_limit"]
        self.assertEqual(self.ns["OKTELL_BILLING_MAX_RANGE_DAYS"], 31)
        self.assertFalse(exceeds(date(2026, 7, 1), date(2026, 7, 31)))

    def test_32_days_are_rejected(self):
        exceeds = self.ns["_oktell_billing_range_exceeds_limit"]
        self.assertTrue(exceeds(date(2026, 7, 1), date(2026, 8, 1)))


class SqlBuilderTests(unittest.TestCase):
    def setUp(self):
        self.ns = _extract_namespace()

    def test_full_day_has_no_minute_filter(self):
        sql = self.ns["_oktell_billing_sql"]("20260701", "20260702", 0, 1439, 20)
        self.assertNotIn("DATEPART(MINUTE", sql)
        self.assertIn("t.dt_insert >= '20260701'", sql)
        self.assertIn("t.dt_insert < '20260702'", sql)
        self.assertIn("GROUP BY CONVERT(varchar(10), t.dt_insert, 23), t.taxi_park", sql)

    def test_time_window_adds_minute_filter(self):
        sql = self.ns["_oktell_billing_sql"]("20260701", "20260702", 480, 1200, 20)
        self.assertIn("BETWEEN 480 AND 1200", sql)
        self.assertIn("DATEPART(HOUR, t.dt_insert) * 60 + DATEPART(MINUTE, t.dt_insert)", sql)

    def test_base_filters_and_sl(self):
        sql = self.ns["_oktell_billing_sql"]("20260701", "20260702", 0, 1439, 25)
        self.assertIn("t.route = 'incoming'", sql)
        self.assertIn("t.taxi_park <> ''", sql)
        self.assertIn("N'Неудачный звонок'", sql)
        self.assertIn("t.LenQueue <= 25", sql)
        # прокси не принимает несколько statement'ов
        self.assertNotIn(";", sql)

    def test_park_mode_has_no_line_join(self):
        sql = self.ns["_oktell_billing_sql"]("20260701", "20260702", 0, 1439, 20, "park")
        self.assertNotIn("A_Stat_Connections_1x1", sql)
        self.assertNotIn("line_number", sql)

    def test_line_mode_joins_dialed_number(self):
        sql = self.ns["_oktell_billing_sql"]("20260701", "20260708", 0, 1439, 20, "line")
        self.assertIn("MIN(ANumberDialed) AS line_number", sql)
        self.assertIn("ConnectionType = 4", sql)
        # окно по TimeStart расширено на день назад (звонок может начаться до полуночи)
        self.assertIn("TimeStart >= '20260630'", sql)
        self.assertIn("TimeStart < '20260708'", sql)
        self.assertIn("TRY_CAST(t.chainid AS uniqueidentifier)", sql)
        self.assertIn("GROUP BY CONVERT(varchar(10), t.dt_insert, 23), t.taxi_park, COALESCE(c.line_number, N'')", sql)
        self.assertNotIn(";", sql)

    def test_operator_sql_builders(self):
        calls_sql = self.ns["_oktell_billing_operator_calls_sql"]("20260701", "20260702", 480, 1200)
        self.assertIn("A_Cube_CC_Cat_OperatorInfo", calls_sql)
        self.assertIn("TRY_CAST(t.id_operator AS uniqueidentifier)", calls_sql)
        self.assertIn("t.call_result IN (5)", calls_sql)
        self.assertIn("BETWEEN 480 AND 1200", calls_sql)
        self.assertNotIn(";", calls_sql)
        states_sql = self.ns["_oktell_billing_operator_states_sql"]("20260701", "20260702", 480, 1200)
        self.assertIn("A_Cube_CC_OperatorStates", states_sql)
        for state_alias in ("talk_in_seconds", "talk_out_seconds", "postproc_seconds",
                            "hold_seconds", "wait_seconds", "pause_seconds", "dial_seconds"):
            self.assertIn(state_alias, states_sql)
        self.assertIn("DATEPART(HOUR, s.DateTimeStart)", states_sql)
        self.assertNotIn(";", states_sql)


class BillingDetailTests(unittest.TestCase):
    def test_page_sql_contains_driver_outcomes_snapshot_and_stable_order(self):
        ns = _extract_namespace()
        sql = ns["_oktell_billing_detail_page_sql"](
            "20260701", "20260703", 480, 1200, page=2, per_page=25, snapshot_id=500)
        self.assertIn("COALESCE(t.[number], N'') AS driver_number", sql)
        self.assertIn("s.ANumberDialed AS line_number", sql)
        self.assertIn("s.AOutNumber AS caller_number", sql)
        self.assertIn("s.ConnectionType = 4", sql)
        self.assertIn("OUTER APPLY", sql)
        self.assertIn("t.call_result IN (5,13,19)", sql)
        self.assertIn("AS ivr_drop", sql)
        self.assertIn("AS queue_drop", sql)
        self.assertIn("AS talk_seconds", sql)
        self.assertIn("COUNT(*) OVER() AS total_count", sql)
        self.assertIn("MAX(t.Id) OVER() AS snapshot_id", sql)
        self.assertIn("ORDER BY t.dt_insert DESC, t.Id DESC", sql)
        self.assertIn("t.Id <= 500", sql)
        self.assertIn("ranked.row_num BETWEEN 26 AND 50", sql)
        self.assertIn("BETWEEN 480 AND 1200", sql)
        self.assertNotIn(";", sql)

    def test_export_sql_uses_keyset_cursor_and_proxy_page_size(self):
        sql = _extract_namespace(page_size=1000)["_oktell_billing_detail_export_page_sql"](
            "20260701", "20260702", 0, 1439, max_call_id=777, page_size=5000)
        self.assertIn("SELECT TOP 1000", sql)
        self.assertIn("t.Id <= 777", sql)
        self.assertIn("ORDER BY t.Id DESC", sql)
        self.assertIn("ORDER BY p.call_id DESC", sql)
        self.assertNotIn("ROW_NUMBER()", sql)
        self.assertNotIn(";", sql)

    def test_build_page_normalizes_rows_and_preserves_snapshot(self):
        ns = _extract_namespace()
        page = ns["_oktell_billing_build_detail_page"]([{
            "call_id": "42",
            "occurred_at": "2026-07-15 10:20:30",
            "taxi_park": " iTaxi ",
            "line_number": "+77075050880",
            "driver_number": "+77071234567",
            "ivr_drop": 0,
            "queue_drop": 1,
            "talk_seconds": "0",
            "total_count": 76,
            "snapshot_id": 99,
        }], page=2, per_page=25, snapshot_id=100)
        self.assertEqual(page["snapshot_id"], 100)
        self.assertEqual(page["pagination"], {
            "page": 2, "per_page": 25, "total": 76, "total_pages": 4,
        })
        self.assertEqual(page["rows"], [{
            "id": 42,
            "occurred_at": "2026-07-15 10:20:30",
            "park": "iTaxi",
            "line": "+77075050880",
            "driver_number": "+77071234567",
            "ivr_drop": 0,
            "queue_drop": 1,
            "talk_seconds": 0,
        }])

    def test_full_export_fetches_every_keyset_page(self):
        calls = []
        pages = [
            [
                {"call_id": 5, "occurred_at": "2026-07-15 10:05:00", "taxi_park": "iTaxi"},
                {"call_id": 4, "occurred_at": "2026-07-15 10:04:00", "taxi_park": "iTaxi"},
            ],
            [
                {"call_id": 3, "occurred_at": "2026-07-15 10:03:00", "taxi_park": "iTaxi"},
                {"call_id": 2, "occurred_at": "2026-07-15 10:02:00", "taxi_park": "iTaxi"},
            ],
            [
                {"call_id": 1, "occurred_at": "2026-07-15 10:01:00", "taxi_park": "iTaxi"},
            ],
        ]

        def fake_query(sql):
            calls.append(sql)
            return pages[len(calls) - 1]

        ns = _extract_namespace(oktell_query=fake_query, page_size=2)
        rows = ns["_oktell_fetch_billing_detail_export_rows"](
            date(2026, 7, 15), date(2026, 7, 15), 0, 1439)
        self.assertEqual([row["id"] for row in rows], [5, 4, 3, 2, 1])
        self.assertEqual(len(calls), 3)
        self.assertNotIn("t.Id <=", calls[0])
        self.assertIn("t.Id <= 3", calls[1])
        self.assertIn("t.Id <= 1", calls[2])


class BuildReportTests(unittest.TestCase):
    def setUp(self):
        self.ns = _extract_namespace()

    @staticmethod
    def _row(report_date, park, arrived=0, served=0, **extra):
        row = {
            "report_date": report_date,
            "taxi_park": park,
            "arrived": arrived,
            "served": served,
            "served_sl": 0,
            "greet_drop": 0,
            "talk_seconds": 0,
            "wait_ok_seconds": 0,
            "wait_lost_seconds": 0,
            "total_seconds": 0,
        }
        row.update(extra)
        return row

    def test_grouping_totals_and_sorting(self):
        build = self.ns["_oktell_billing_build_report"]
        report = build([
            self._row("2026-07-02", "Amanat", arrived=5, served=4, talk_seconds=100.6),
            self._row("2026-07-01", "iTaxi", arrived=50, served=45, served_sl=40, talk_seconds=1000),
            self._row("2026-07-01", "Amanat", arrived=10, served=9, talk_seconds=200),
        ])
        self.assertEqual([day["date"] for day in report["days"]], ["2026-07-01", "2026-07-02"])
        first_day = report["days"][0]
        # сортировка по «поступило» убыванием
        self.assertEqual([p["park"] for p in first_day["parks"]], ["iTaxi", "Amanat"])
        self.assertEqual(first_day["totals"]["arrived"], 60)
        self.assertEqual(first_day["totals"]["served"], 54)
        self.assertEqual(first_day["totals"]["lost"], 6)
        # итоги за период по паркам
        parks = {p["park"]: p for p in report["parks"]}
        self.assertEqual(parks["Amanat"]["arrived"], 15)
        self.assertEqual(parks["Amanat"]["talk_seconds"], 301)  # 200 + round(100.6)
        self.assertEqual(report["totals"]["arrived"], 65)
        self.assertEqual(report["totals"]["served_sl"], 40)

    def test_lost_clamped_and_bad_rows_skipped(self):
        build = self.ns["_oktell_billing_build_report"]
        report = build([
            self._row("2026-07-01", "Amanat", arrived=3, served=5),  # served > arrived
            self._row("", "iTaxi", arrived=10, served=10),           # нет даты
            self._row("2026-07-01", "", arrived=10, served=10),      # пустой парк
            self._row("2026-07-01", "Global", arrived=None, served=None, talk_seconds="oops"),
        ])
        self.assertEqual(len(report["days"]), 1)
        parks = {p["park"]: p for p in report["days"][0]["parks"]}
        self.assertEqual(set(parks), {"Amanat", "Global"})
        self.assertEqual(parks["Amanat"]["lost"], 0)
        self.assertEqual(parks["Global"]["talk_seconds"], 0)
        self.assertEqual(report["totals"]["arrived"], 3)

    def test_empty_input(self):
        report = self.ns["_oktell_billing_build_report"]([])
        self.assertEqual(report, {"days": [], "parks": [], "totals": {
            "arrived": 0, "served": 0, "lost": 0, "served_sl": 0, "greet_drop": 0,
            "talk_seconds": 0, "wait_ok_seconds": 0, "wait_lost_seconds": 0, "total_seconds": 0,
        }})

    def test_line_grouping_splits_park(self):
        build = self.ns["_oktell_billing_build_report"]
        report = build([
            self._row("2026-07-01", "Честный", arrived=10, served=9, line_number="+77003330402"),
            self._row("2026-07-01", "Честный", arrived=3, served=3, line_number="+77001110200"),
            self._row("2026-07-01", "iTaxi", arrived=50, served=45, line_number="+77075050880"),
        ], include_line=True)
        day_rows = report["days"][0]["parks"]
        self.assertEqual(
            [(r["park"], r["line"]) for r in day_rows],
            [("iTaxi", "+77075050880"), ("Честный", "+77003330402"), ("Честный", "+77001110200")],
        )
        # итог дня объединяет обе линии парка
        self.assertEqual(report["days"][0]["totals"]["arrived"], 63)
        parks = {(p["park"], p["line"]) for p in report["parks"]}
        self.assertEqual(len(parks), 3)

    def test_without_line_flag_lines_merge(self):
        build = self.ns["_oktell_billing_build_report"]
        report = build([
            self._row("2026-07-01", "Честный", arrived=10, served=9, line_number="+77003330402"),
            self._row("2026-07-01", "Честный", arrived=3, served=3, line_number="+77001110200"),
        ], include_line=False)
        rows = report["days"][0]["parks"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["arrived"], 13)
        self.assertNotIn("line", rows[0])


class BuildOperatorReportTests(unittest.TestCase):
    def setUp(self):
        self.ns = _extract_namespace()

    def test_merge_calls_and_states(self):
        build = self.ns["_oktell_billing_build_operator_report"]
        calls = [
            {"report_date": "2026-07-01", "operator_name": "Иванова А.", "served": 30, "talk_seconds": 6000.4},
            {"report_date": "2026-07-02", "operator_name": "Иванова А.", "served": 20, "talk_seconds": 4000},
        ]
        states = [
            {"report_date": "2026-07-01", "operator_name": "Иванова А.", "talk_in_seconds": 5900,
             "talk_out_seconds": 0, "postproc_seconds": 100, "hold_seconds": 50,
             "wait_seconds": 3000, "pause_seconds": 1000, "dial_seconds": 20},
            {"report_date": "2026-07-01", "operator_name": "Петров Б.", "talk_in_seconds": 0,
             "talk_out_seconds": 2000, "postproc_seconds": 40, "hold_seconds": 0,
             "wait_seconds": 500, "pause_seconds": 0, "dial_seconds": 100},
        ]
        report = build(calls, states)
        self.assertEqual([d["date"] for d in report["days"]], ["2026-07-01", "2026-07-02"])
        day1 = {r["operator"]: r for r in report["days"][0]["operators"]}
        self.assertEqual(day1["Иванова А."]["served"], 30)
        self.assertEqual(day1["Иванова А."]["talk_seconds"], 6000)
        self.assertEqual(day1["Иванова А."]["hold_seconds"], 50)
        # Петров без входящих звонков, но с исходящими разговорами — остаётся
        self.assertEqual(day1["Петров Б."]["served"], 0)
        self.assertEqual(day1["Петров Б."]["talk_out_seconds"], 2000)
        period = {r["operator"]: r for r in report["operators"]}
        self.assertEqual(period["Иванова А."]["served"], 50)
        self.assertEqual(report["totals"]["served"], 50)

    def test_technical_accounts_dropped(self):
        build = self.ns["_oktell_billing_build_operator_report"]
        states = [
            {"report_date": "2026-07-01", "operator_name": "admin", "talk_in_seconds": 0,
             "talk_out_seconds": 0, "postproc_seconds": 0, "hold_seconds": 0,
             "wait_seconds": 65000, "pause_seconds": 5000, "dial_seconds": 0},
        ]
        report = build([], states)
        self.assertEqual(report["days"], [])
        self.assertEqual(report["operators"], [])

    def test_export_values_park_row(self):
        ns = _extract_namespace()
        values = ns["_oktell_billing_export_values"]("park", {
            "park": "Tenge_taxi", "arrived": 100, "served": 80, "lost": 20, "served_sl": 60,
            "greet_drop": 2, "talk_seconds": 8000, "wait_ok_seconds": 400,
            "wait_lost_seconds": 100, "total_seconds": 8500,
        })
        self.assertEqual(values[0], "Тенге Такси")
        self.assertEqual(values[1:4], [100, 80, 20])
        self.assertAlmostEqual(values[4], 0.2)   # AR
        self.assertAlmostEqual(values[5], 0.75)  # SL
        self.assertAlmostEqual(values[6], (8000 / 80) / 86400.0)  # АТТ как доля суток
        self.assertAlmostEqual(values[8], 8000 / 86400.0)
        self.assertEqual(values[10], 2)

    def test_export_values_line_and_zero_served(self):
        ns = _extract_namespace()
        values = ns["_oktell_billing_export_values"]("line", {
            "park": "Салам Такси", "line": "+77001551198", "arrived": 0, "served": 0, "lost": 0,
            "served_sl": 0, "greet_drop": 1, "talk_seconds": 0, "wait_ok_seconds": 0,
            "wait_lost_seconds": 0, "total_seconds": 0,
        })
        self.assertEqual(values[0], "87001551198")
        self.assertEqual(values[1], "СТ 1")
        self.assertEqual(values[2], "Салам Такси")
        # нет обслуженных -> проценты и средние '—'
        self.assertEqual(values[6], "—")
        self.assertEqual(values[7], "—")
        self.assertEqual(values[8], "—")

    def test_detail_export_columns_and_values(self):
        ns = _extract_namespace()
        headers, widths, formats = ns["_oktell_billing_export_columns"]("detail")
        self.assertEqual(headers, [
            "Дата",
            "Парк на который звонят",
            "Номер на который звонят",
            "Номер водителя",
            "Сброс на IVR",
            "Сброс в очереди/пропущенные",
            "Время разговора",
        ])
        self.assertEqual(len(widths), 7)
        self.assertEqual(formats, {1: "dd.mm.yyyy hh:mm:ss", 7: "[h]:mm:ss"})

        values = ns["_oktell_billing_export_values"]("detail", {
            "occurred_at": "2026-07-15 11:22:33",
            "park": "Tenge_taxi",
            "line": "+7 (700) 155-1198",
            "driver_number": "+7 707 123 45 67",
            "ivr_drop": 0,
            "queue_drop": 0,
            "talk_seconds": 901,
        })
        self.assertEqual(values[:6], [
            datetime(2026, 7, 15, 11, 22, 33), "Тенге Такси", "87001551198",
            "87071234567", 0, 0,
        ])
        self.assertAlmostEqual(values[6], 901 / 86400.0)

    def test_detail_export_uses_dash_for_missing_numbers(self):
        values = _extract_namespace()["_oktell_billing_export_values"]("detail", {
            "occurred_at": "2026-07-15 10:00:00", "park": "iTaxi", "line": "",
            "driver_number": "", "ivr_drop": 1, "queue_drop": 0, "talk_seconds": 0,
        })
        self.assertEqual(values[2], "—")
        self.assertEqual(values[3], "—")
        self.assertEqual(values[4:6], [1, 0])

    def test_export_workbook_structure(self):
        ns = _extract_namespace()
        report = ns["_oktell_billing_build_report"]([
            {"report_date": "2026-07-13", "taxi_park": "iTaxi", "arrived": 10, "served": 9,
             "served_sl": 8, "greet_drop": 0, "talk_seconds": 900, "wait_ok_seconds": 30,
             "wait_lost_seconds": 5, "total_seconds": 940},
        ])
        params = {"start_day": date(2026, 7, 13), "end_day": date(2026, 7, 13),
                  "minute_from": 0, "minute_to": 1439}
        output = ns["_oktell_billing_export_workbook"]("park", params, report, 20)
        wb = load_workbook(output)
        self.assertEqual(wb.sheetnames, ["Итого за период", "По дням"])
        ws = wb["Итого за период"]
        header_row = next(
            idx for idx in range(1, ws.max_row + 1)
            if ws.cell(row=idx, column=1).value == "Таксопарк"
        )
        self.assertEqual(ws.cell(row=header_row + 1, column=1).value, "iTaxi")
        self.assertEqual(ws.cell(row=header_row + 2, column=1).value, "Итого за период")
        self.assertEqual(ws.cell(row=header_row + 1, column=5).number_format, "0.0%")
        ws_days = wb["По дням"]
        self.assertEqual(ws_days.cell(row=1, column=1).value, "Дата")
        self.assertEqual(ws_days.cell(row=2, column=1).value, "13.07.2026")
        self.assertEqual(ws_days.cell(row=3, column=2).value, "Итого за день")
        # формат длительности в дневном листе со сдвигом на колонку даты
        self.assertEqual(ws_days.cell(row=2, column=8).number_format, "[h]:mm:ss")

    def test_detail_workbook_is_one_flat_filtered_sheet(self):
        ns = _extract_namespace()
        report = {"rows": [
            {"id": 103, "occurred_at": "2026-07-14 12:15:00", "park": "iTaxi",
             "line": "+77075050880", "driver_number": "+77071234567",
             "ivr_drop": 0, "queue_drop": 0, "talk_seconds": 960},
            {"id": 102, "occurred_at": "2026-07-14 12:10:00", "park": "iTaxi",
             "line": "+77075050880", "driver_number": "87079876543",
             "ivr_drop": 0, "queue_drop": 1, "talk_seconds": 0},
            {"id": 101, "occurred_at": "2026-07-13 09:00:00", "park": "Tenge_taxi",
             "line": "", "driver_number": "", "ivr_drop": 1, "queue_drop": 0,
             "talk_seconds": 0},
        ]}
        params = {"start_day": date(2026, 7, 13), "end_day": date(2026, 7, 14),
                  "minute_from": 0, "minute_to": 1439}

        output = ns["_oktell_billing_export_workbook"]("detail", params, report, 20)
        wb = load_workbook(output)
        self.assertEqual(wb.sheetnames, ["Детализация"])
        ws = wb.active
        self.assertEqual([cell.value for cell in ws[1]], [
            "Дата", "Парк на который звонят", "Номер на который звонят", "Номер водителя",
            "Сброс на IVR", "Сброс в очереди/пропущенные", "Время разговора",
        ])
        self.assertEqual(ws.max_row, 4)
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertEqual(ws.auto_filter.ref, "A1:G4")
        self.assertTrue(ws["A1"].font.bold)
        self.assertEqual(ws["A1"].font.color.rgb, "00FFFFFF")

        rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertEqual(rows[0][:6], [
            datetime(2026, 7, 14, 12, 15), "iTaxi", "87075050880", "87071234567", 0, 0,
        ])
        self.assertEqual(rows[0][6], timedelta(seconds=960))
        self.assertEqual(ws["A2"].number_format, "dd.mm.yyyy hh:mm:ss")
        self.assertEqual(ws["G2"].number_format, "[h]:mm:ss")
        self.assertEqual(rows[1][3:6], ["87079876543", 0, 1])
        self.assertEqual(rows[2][1:4], ["Тенге Такси", "—", "—"])

    def test_operator_fetch_runs_two_queries_per_chunk(self):
        calls = []

        def fake_query(sql):
            calls.append(sql)
            return [{"n": len(calls)}]

        ns = _extract_namespace(oktell_query=fake_query, chunk_days=7)
        calls_rows, states_rows = ns["_oktell_fetch_billing_operator_rows"](
            date(2026, 7, 1), date(2026, 7, 10), 0, 1439)
        # 10 дней = 2 чанка, по 2 запроса на чанк
        self.assertEqual(len(calls), 4)
        self.assertEqual(len(calls_rows), 2)
        self.assertEqual(len(states_rows), 2)
        self.assertIn("Call_Systems_hst", calls[0])
        self.assertIn("A_Cube_CC_OperatorStates", calls[1])


class FetchChunkingTests(unittest.TestCase):
    def test_chunks_are_serial_windows(self):
        calls = []

        def fake_query(sql):
            calls.append(sql)
            return [{"report_date": "x"}]

        ns = _extract_namespace(oktell_query=fake_query, chunk_days=7)
        rows = ns["_oktell_fetch_billing_rows"](date(2026, 7, 1), date(2026, 7, 16), 0, 1439, 20)
        # 16 дней при чанке 7 -> окна 01-07, 08-14, 15-16
        self.assertEqual(len(calls), 3)
        self.assertIn("'20260701'", calls[0])
        self.assertIn("'20260708'", calls[0])
        self.assertIn("'20260708'", calls[1])
        self.assertIn("'20260715'", calls[1])
        self.assertIn("'20260715'", calls[2])
        self.assertIn("'20260717'", calls[2])
        self.assertEqual(len(rows), 3)

    def test_truncated_window_falls_back_to_days(self):
        calls = []

        def fake_query(sql):
            calls.append(sql)
            # первое (оконное) обращение упирается в лимит, дневные — нет
            if len(calls) == 1:
                return [{"n": i} for i in range(5)]
            return [{"n": "day"}]

        ns = _extract_namespace(oktell_query=fake_query, page_size=5, chunk_days=7)
        rows = ns["_oktell_fetch_billing_rows"](date(2026, 7, 1), date(2026, 7, 3), 0, 1439, 20)
        # 1 оконный + 3 дневных запроса; строки берутся только из дневных
        self.assertEqual(len(calls), 4)
        self.assertEqual(rows, [{"n": "day"}] * 3)

    def test_single_day_truncated_keeps_page(self):
        calls = []

        def fake_query(sql):
            calls.append(sql)
            return [{"n": i} for i in range(5)]

        ns = _extract_namespace(oktell_query=fake_query, page_size=5, chunk_days=7)
        rows = ns["_oktell_fetch_billing_rows"](date(2026, 7, 1), date(2026, 7, 1), 0, 1439, 20)
        # окно из одного дня не дробится (a == b) — страница остаётся как есть
        self.assertEqual(len(calls), 1)
        self.assertEqual(len(rows), 5)


if __name__ == "__main__":
    unittest.main()
