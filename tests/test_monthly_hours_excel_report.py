# -*- coding: utf-8 -*-
"""Регрессии XLSX-отчёта «Учёт часов»: штрафы и бонусы.

Большой модуль ``database.py`` нельзя импортировать в unit-тесте: при импорте он
настраивает инфраструктуру приложения. Поэтому извлекаем только метод
``Database.generate_excel_report_from_view`` через AST и исполняем его с
минимальным набором зависимостей, как в остальных Excel-тестах проекта.
"""

import ast
import calendar
import logging
import re
import textwrap
import unittest
from collections import defaultdict
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATABASE_PATH = ROOT / "database.py"


def _load_report_builder():
    source = DATABASE_PATH.read_text(encoding="utf-8-sig")
    module = ast.parse(source)
    database_class = next(
        node
        for node in module.body
        if isinstance(node, ast.ClassDef) and node.name == "Database"
    )
    method = next(
        node
        for node in database_class.body
        if isinstance(node, ast.FunctionDef)
        and node.name == "generate_excel_report_from_view"
    )
    method_source = textwrap.dedent(ast.get_source_segment(source, method))
    namespace = {
        "Any": Any,
        "Dict": Dict,
        "List": List,
        "Tuple": Tuple,
        "Alignment": Alignment,
        "Border": Border,
        "BytesIO": BytesIO,
        "CellRichText": CellRichText,
        "Font": Font,
        "InlineFont": InlineFont,
        "PatternFill": PatternFill,
        "Side": Side,
        "TextBlock": TextBlock,
        "Workbook": Workbook,
        "CALCULATION_MODEL_CHAT_MANAGER": "chat_manager",
        "SCHEDULE_AUTO_FINE_RATE_PER_MINUTE": 50,
        "calendar": calendar,
        "date": date,
        "defaultdict": defaultdict,
        "get_column_letter": get_column_letter,
        "logging": logging,
        "re": re,
    }
    exec(compile(method_source, str(DATABASE_PATH), "exec"), namespace)
    return namespace["generate_excel_report_from_view"]


class _FakeDatabase:
    generate_excel_report_from_view = _load_report_builder()

    def _get_cursor(self):
        raise AssertionError("Активные тестовые операторы не должны обращаться к БД")


def _operator(operator_id, name, daily=None, **overrides):
    value = {
        "operator_id": operator_id,
        "name": name,
        "rate": 1.0,
        "norm_hours": 160.0,
        "status": "working",
        "calculation_model_code": "operator",
        "aggregates": {},
        "daily": daily or {},
    }
    value.update(overrides)
    return value


def _find_row(sheet, value, start=1):
    for row in range(start, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == value:
            return row
    raise AssertionError(f"Не найдена строка {value!r} на листе {sheet.title!r}")


def _headers(sheet, row):
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[row]
        if cell.value not in (None, "")
    }


def _metric_column(headers, label, metric):
    """Находит колонку по смыслу, не привязываясь к ``шт``/``шт.``."""
    aliases = {
        "count": ("шт", "колич", "кол-"),
        "minutes": ("мин",),
        "hours": ("час", "(ч", " ч"),
        "amount": ("сумм", "₸", "тенге"),
    }
    label_norm = label.casefold()
    matches = []
    for header, column in headers.items():
        normalized = header.casefold()
        if label_norm not in normalized:
            continue
        if any(alias in normalized for alias in aliases[metric]):
            matches.append((header, column))
    if len(matches) != 1:
        raise AssertionError(
            f"Ожидалась одна колонка {label!r}/{metric}, найдены: {matches}; "
            f"все заголовки: {list(headers)}"
        )
    return matches[0][1]


def _summary(sheet):
    title_row = _find_row(sheet, "Сводка по операторам")
    header_row = title_row + 1
    headers = _headers(sheet, header_row)
    rows = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        if name not in (None, ""):
            rows[str(name)] = row
    return title_row, header_row, headers, rows


class MonthlyHoursExcelReportTests(unittest.TestCase):
    month = "2026-02"

    def _workbook(self, operators, *, include_supervisor=False, report_group_id=None):
        filename, content = _FakeDatabase().generate_excel_report_from_view(
            {"operators": operators},
            {},
            {},
            self.month,
            offline_activities_map={},
            include_supervisor=include_supervisor,
            report_group_id=report_group_id,
        )
        self.assertEqual(filename, f"report_{self.month}.xlsx")
        self.assertTrue(content.startswith(b"PK"))
        return load_workbook(BytesIO(content), data_only=False, rich_text=True)

    def test_chat_score_total_uses_weighted_month_average(self):
        weighted_average = 501 / 101
        operator = _operator(
            99,
            "Чат Оператор",
            calculation_model_code="chat_manager",
            aggregates={},
            daily={
                "1": {
                    "chat_metrics": {
                        "avg_score": 1.0,
                        "score_sum": 1.0,
                        "score_count": 1,
                    },
                },
                "2": {
                    "chat_metrics": {
                        "avg_score": 5.0,
                        "score_sum": 500.0,
                        "score_count": 100,
                    },
                },
            },
        )

        workbook = self._workbook([operator])
        score_sheet = workbook["Средняя оценка"]
        headers = _headers(score_sheet, 1)

        self.assertEqual(score_sheet.cell(2, headers["01.02"]).value, 1.0)
        self.assertEqual(score_sheet.cell(2, headers["02.02"]).value, 5.0)
        self.assertAlmostEqual(
            score_sheet.cell(2, headers["Итого"]).value,
            weighted_average,
            places=12,
        )
        self.assertNotEqual(score_sheet.cell(2, headers["Итого"]).value, 3.0)

    def test_chat_score_fallback_excludes_days_from_another_group(self):
        operator = _operator(
            100,
            "Переведённый чат-оператор",
            calculation_model_code="chat_manager",
            aggregates={},
            group_segments=[
                {"group_id": 10, "start_day": 1, "end_day": 1},
                {"group_id": 20, "start_day": 2, "end_day": 28},
            ],
            daily={
                "1": {
                    "chat_metrics": {
                        "avg_score": 1.0,
                        "score_sum": 1.0,
                        "score_count": 1,
                    },
                },
                "2": {
                    "chat_metrics": {
                        "avg_score": 5.0,
                        "score_sum": 45.0,
                        "score_count": 9,
                    },
                },
            },
        )

        workbook = self._workbook([operator], report_group_id=10)
        score_sheet = workbook["Средняя оценка"]
        headers = _headers(score_sheet, 1)

        self.assertEqual(score_sheet.cell(2, headers["01.02"]).value, 1.0)
        self.assertEqual(score_sheet.cell(2, headers["02.02"]).value, "др.")
        self.assertEqual(score_sheet.cell(2, headers["Итого"]).value, 1.0)

    def test_summaries_aggregate_multiple_records_and_keep_supervisor_layout(self):
        operators = [
            _operator(
                101,
                "Анна Оператор",
                supervisor_name="СВ Север",
                rate=0.75,
                daily={
                    "1": {
                        "fines": [
                            {"reason": "Опоздание", "minutes": 7, "amount": 350},
                            {"reason": "Не выход", "amount": 10000},
                            {"reason": "Нарушение регламента", "amount": 1200},
                        ],
                        "bonuses": [
                            {
                                "type": "Обучение",
                                "training_hours": 1.5,
                                "quantity": 1,
                                "amount": 750,
                            },
                            {"type": "Приведи друга", "quantity": 2, "amount": 10000},
                            {"type": "Премия", "quantity": 1, "amount": 900},
                        ],
                    },
                    "2": {
                        "fines": [
                            {"reason": "Опоздание", "minutes": 3, "amount": 150},
                            {"reason": "Корп такси", "amount": 700},
                            {"reason": "Прокси карта", "amount": 5000},
                        ],
                        "bonuses": [
                            {"type": "Съемки", "quantity": 2, "amount": 10000},
                            {"type": "Приведи друга", "quantity": 1, "amount": 5000},
                        ],
                    },
                },
            ),
            _operator(
                102,
                "Борис Оператор",
                supervisor_name="СВ Юг",
                rate=1.0,
                daily={
                    # Legacy-строка: списка fines нет, штраф должен попасть в сводку.
                    "1": {
                        "fine_amount": 250,
                        "fine_reason": "Опоздание",
                        "fine_comment": "legacy",
                        "bonuses": [],
                    },
                    # Новый список авторитетен: агрегированные legacy-поля не удваиваются.
                    "2": {
                        "fine_amount": 100,
                        "fine_reason": "Опоздание",
                        "fines": [
                            {"reason": "Опоздание", "minutes": 2, "amount": 100},
                        ],
                        "bonuses": [],
                    },
                },
            ),
        ]

        workbook = self._workbook(operators, include_supervisor=True)
        self.assertIn("Штрафы", workbook.sheetnames)
        self.assertIn("Бонусы", workbook.sheetnames)
        self.assertEqual(
            workbook.sheetnames.index("Бонусы"),
            workbook.sheetnames.index("Штрафы") + 1,
        )

        fines = workbook["Штрафы"]
        bonuses = workbook["Бонусы"]
        self.assertEqual([fines.cell(1, col).value for col in range(1, 4)], [
            "ФИО", "Супервайзер", "Ставка",
        ])
        self.assertEqual([bonuses.cell(1, col).value for col in range(1, 4)], [
            "ФИО", "Супервайзер", "Ставка",
        ])
        self.assertEqual(fines.freeze_panes, "D2")
        self.assertEqual(bonuses.freeze_panes, "D2")

        fines_title, fines_header_row, fine_headers, fine_rows = _summary(fines)
        self.assertEqual(fine_headers.get("ФИО"), 1)
        self.assertEqual(fine_headers.get("Супервайзер"), 2)
        self.assertEqual(fine_headers.get("Ставка"), 3)
        self.assertTrue(fines.cell(fines_header_row, 1).font.bold)
        self.assertEqual(fines.cell(fines_header_row, 1).border.left.style, "thin")
        self.assertGreater(fines_title, _find_row(fines, "Легенда причин"))

        anna = fine_rows["Анна Оператор"]
        self.assertEqual(fines.cell(anna, fine_headers["Супервайзер"]).value, "СВ Север")
        self.assertEqual(fines.cell(anna, fine_headers["Ставка"]).value, 0.75)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "опоздани", "count")).value, 2)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "опоздани", "minutes")).value, 10)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "опоздани", "amount")).value, 500)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "такси", "count")).value, 1)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "такси", "amount")).value, 700)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "Не выход", "count")).value, 1)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "Не выход", "amount")).value, 10000)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "прокси", "count")).value, 1)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "прокси", "amount")).value, 5000)
        # Неизвестная причина нормализуется в «Другое».
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "друг", "count")).value, 1)
        self.assertEqual(fines.cell(anna, _metric_column(fine_headers, "друг", "amount")).value, 1200)
        self.assertEqual(fines.cell(anna, fine_headers["Всего штрафов"]).value, 6)
        self.assertEqual(fines.cell(anna, fine_headers["Итого"]).value, 17400)

        boris = fine_rows["Борис Оператор"]
        self.assertEqual(fines.cell(boris, _metric_column(fine_headers, "опоздани", "count")).value, 2)
        self.assertEqual(fines.cell(boris, _metric_column(fine_headers, "опоздани", "minutes")).value, 7)
        self.assertEqual(fines.cell(boris, _metric_column(fine_headers, "опоздани", "amount")).value, 350)
        self.assertEqual(fines.cell(boris, fine_headers["Всего штрафов"]).value, 2)
        self.assertEqual(fines.cell(boris, fine_headers["Итого"]).value, 350)

        bonus_title, bonus_header_row, bonus_headers, bonus_rows = _summary(bonuses)
        self.assertEqual(bonus_headers.get("ФИО"), 1)
        self.assertEqual(bonus_headers.get("Супервайзер"), 2)
        self.assertEqual(bonus_headers.get("Ставка"), 3)
        self.assertTrue(bonuses.cell(bonus_header_row, 1).font.bold)
        self.assertEqual(bonuses.cell(bonus_header_row, 1).border.left.style, "thin")
        legend_rows = [
            row
            for row in range(1, bonuses.max_row + 1)
            if str(bonuses.cell(row, 1).value or "").startswith("Легенда")
        ]
        self.assertTrue(legend_rows)
        self.assertGreater(bonus_title, max(legend_rows))

        anna_bonus = bonus_rows["Анна Оператор"]
        self.assertEqual(bonuses.cell(anna_bonus, bonus_headers["Супервайзер"]).value, "СВ Север")
        self.assertEqual(bonuses.cell(anna_bonus, _metric_column(bonus_headers, "Приведи друга", "count")).value, 3)
        self.assertEqual(bonuses.cell(anna_bonus, _metric_column(bonus_headers, "Приведи друга", "amount")).value, 15000)
        self.assertEqual(bonuses.cell(anna_bonus, _metric_column(bonus_headers, "Обучение", "hours")).value, 1.5)
        self.assertEqual(bonuses.cell(anna_bonus, _metric_column(bonus_headers, "Обучение", "amount")).value, 750)
        self.assertEqual(bonuses.cell(anna_bonus, _metric_column(bonus_headers, "Съемки", "count")).value, 2)
        self.assertEqual(bonuses.cell(anna_bonus, _metric_column(bonus_headers, "Съемки", "amount")).value, 10000)
        self.assertEqual(bonuses.cell(anna_bonus, _metric_column(bonus_headers, "Проч", "count")).value, 1)
        self.assertEqual(bonuses.cell(anna_bonus, _metric_column(bonus_headers, "Проч", "amount")).value, 900)
        self.assertEqual(bonuses.cell(anna_bonus, bonus_headers["Всего бонусов"]).value, 5)
        self.assertEqual(bonuses.cell(anna_bonus, bonus_headers["Итого"]).value, 26650)

        # Дневная матрица остаётся числовой по итогу и не заменяется сводкой.
        fine_matrix_headers = _headers(fines, 1)
        bonus_matrix_headers = _headers(bonuses, 1)
        self.assertEqual(fines.cell(2, fine_matrix_headers["Итого"]).value, 17400)
        self.assertEqual(bonuses.cell(2, bonus_matrix_headers["Итого"]).value, 26650)

    def test_group_report_excludes_foreign_days_from_matrix_and_summaries(self):
        operator = _operator(
            201,
            "Переведённый оператор",
            daily={
                "1": {
                    "fines": [{"reason": "Опоздание", "minutes": 1, "amount": 50}],
                    "bonuses": [{"type": "Приведи друга", "quantity": 1, "amount": 5000}],
                },
                "2": {
                    "fines": [{"reason": "Не выход", "amount": 10000}],
                    "bonuses": [{"type": "Съемки", "quantity": 1, "amount": 5000}],
                },
            },
            group_segments=[
                {"group_id": 10, "start_day": 1, "end_day": 1},
                {"group_id": 20, "start_day": 2, "end_day": 28},
            ],
        )

        workbook = self._workbook([operator], report_group_id=10)
        fines = workbook["Штрафы"]
        bonuses = workbook["Бонусы"]
        fine_matrix_headers = _headers(fines, 1)
        bonus_matrix_headers = _headers(bonuses, 1)
        self.assertEqual(fines.cell(2, fine_matrix_headers["02.02"]).value, "др.")
        self.assertEqual(bonuses.cell(2, bonus_matrix_headers["02.02"]).value, "др.")
        self.assertEqual(fines.cell(2, fine_matrix_headers["Итого"]).value, 50)
        self.assertEqual(bonuses.cell(2, bonus_matrix_headers["Итого"]).value, 5000)

        _, _, fine_headers, fine_rows = _summary(fines)
        fine_row = fine_rows["Переведённый оператор"]
        self.assertEqual(fines.cell(fine_row, _metric_column(fine_headers, "опоздани", "count")).value, 1)
        self.assertEqual(fines.cell(fine_row, _metric_column(fine_headers, "Не выход", "count")).value, 0)
        self.assertEqual(fines.cell(fine_row, fine_headers["Всего штрафов"]).value, 1)
        self.assertEqual(fines.cell(fine_row, fine_headers["Итого"]).value, 50)

        _, _, bonus_headers, bonus_rows = _summary(bonuses)
        bonus_row = bonus_rows["Переведённый оператор"]
        self.assertEqual(bonuses.cell(bonus_row, _metric_column(bonus_headers, "Приведи друга", "count")).value, 1)
        self.assertEqual(bonuses.cell(bonus_row, _metric_column(bonus_headers, "Съемки", "count")).value, 0)
        self.assertEqual(bonuses.cell(bonus_row, bonus_headers["Итого"]).value, 5000)

    def test_empty_records_keep_zero_rows_in_both_summaries(self):
        workbook = self._workbook([_operator(301, "Оператор без начислений")])

        for sheet_name, total_header in (("Штрафы", "Всего штрафов"), ("Бонусы", "Всего бонусов")):
            sheet = workbook[sheet_name]
            _, _, headers, rows = _summary(sheet)
            row = rows["Оператор без начислений"]
            self.assertEqual(sheet.cell(row, headers[total_header]).value, 0)
            self.assertEqual(sheet.cell(row, headers["Итого"]).value, 0)
            self.assertTrue(sheet.cell(row, headers["Итого"]).border.left.style)


if __name__ == "__main__":
    unittest.main()
