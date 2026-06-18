import ast
import unittest
from io import BytesIO
from pathlib import Path

import openpyxl

try:
    import xlsxwriter
except ImportError:  # Локальный облегчённый runtime; в приложении пакет закреплён в requirements.txt.
    xlsxwriter = None


ROOT = Path(__file__).resolve().parents[1]
BOT_PATH = ROOT / "bot_schedule2.py"
DATABASE_PATH = ROOT / "database.py"
MAIN_JSX_PATH = ROOT / "src" / "call_evaluation" / "main.jsx"


def _read(path):
    return path.read_text(encoding="utf-8-sig")


def _load_report_builder():
    source = _read(BOT_PATH)
    module = ast.parse(source)
    function = next(
        node for node in module.body
        if isinstance(node, ast.FunctionDef) and node.name == "_build_call_feedback_report_xlsx"
    )
    namespace = {"BytesIO": BytesIO, "xlsxwriter": xlsxwriter}
    exec(compile(ast.Module(body=[function], type_ignores=[]), str(BOT_PATH), "exec"), namespace)
    return namespace["_build_call_feedback_report_xlsx"]


class CallFeedbackExcelReportTests(unittest.TestCase):
    def setUp(self):
        self.report_data = {
            "overview": {
                "total_evaluated": 4,
                "feedback_provided": 3,
                "feedback_on_time": 2,
                "feedback_overdue": 1,
                "pending": 1,
                "pending_overdue": 1,
                "overdue_total": 2,
            },
            "supervisors": [
                {
                    "supervisor_name": "Анна Соколова",
                    "total_evaluated": 4,
                    "feedback_provided": 3,
                    "feedback_on_time": 2,
                    "feedback_overdue": 1,
                    "pending": 1,
                    "pending_overdue": 1,
                    "overdue_total": 2,
                    "items": [
                        {
                            "call_id": 101,
                            "operator_name": "Иван Петров",
                            "score": 95,
                            "evaluation_date": "2026-06-02",
                            "due_date": "2026-06-04",
                            "feedback_date": "2026-06-03",
                            "feedback_provided": True,
                            "feedback_supervisor_name": "Анна Соколова",
                            "status": "on_time",
                            "overdue_days": 0,
                            "has_critical_error": False,
                        },
                        {
                            "call_id": 102,
                            "operator_name": "Иван Петров",
                            "score": 72,
                            "evaluation_date": "2026-06-05",
                            "due_date": "2026-06-07",
                            "feedback_date": "2026-06-09",
                            "feedback_provided": True,
                            "feedback_supervisor_name": "Анна Соколова",
                            "status": "overdue",
                            "overdue_days": 2,
                            "has_critical_error": False,
                        },
                        {
                            "call_id": 103,
                            "operator_name": "Мария Орлова",
                            "score": 45,
                            "evaluation_date": "2026-06-08",
                            "due_date": "2026-06-09",
                            "feedback_date": None,
                            "feedback_provided": False,
                            "feedback_supervisor_name": None,
                            "status": "overdue",
                            "overdue_days": 4,
                            "has_critical_error": True,
                        },
                        {
                            "call_id": 104,
                            "operator_name": "Мария Орлова",
                            "score": 100,
                            "evaluation_date": "2026-06-17",
                            "due_date": "2026-06-19",
                            "feedback_date": None,
                            "feedback_provided": False,
                            "feedback_supervisor_name": None,
                            "status": "pending",
                            "overdue_days": 0,
                            "has_critical_error": False,
                        },
                    ],
                }
            ],
        }

    @unittest.skipIf(xlsxwriter is None, "xlsxwriter is not installed in this test runtime")
    def test_builder_creates_styled_summary_and_detail_workbook(self):
        builder = _load_report_builder()
        content = builder(
            self.report_data,
            "2026-06",
            "18.06.2026 09:05",
            scope_label="Отдел поддержки",
        )

        self.assertTrue(content.startswith(b"PK"))
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Сводка", "Детализация"])
        self.assertIn("Отчёт по обратной связи", workbook["Сводка"]["A1"].value)
        self.assertEqual(workbook["Сводка"]["A6"].value, 4)
        self.assertIn("Отдел поддержки", workbook["Сводка"]["A3"].value)
        self.assertTrue(workbook["Сводка"]._charts)
        self.assertEqual(workbook["Детализация"].max_row, 9)
        statuses = {workbook["Детализация"].cell(row=row, column=9).value for row in range(6, 10)}
        self.assertEqual(
            statuses,
            {"В срок", "Просрочено с ОС", "Просрочено без ОС", "Ожидает ОС"},
        )
        self.assertTrue(workbook["Сводка"].freeze_panes)
        self.assertTrue(workbook["Детализация"].freeze_panes)

    def test_scheduler_sends_one_excel_document(self):
        source = _read(BOT_PATH)
        module = ast.parse(source)
        sender = next(
            node for node in module.body
            if isinstance(node, ast.FunctionDef) and node.name == "sync_send_weekly_call_feedback_report"
        )
        body = ast.get_source_segment(source, sender)
        self.assertIn("_build_call_feedback_report_xlsx", body)
        self.assertIn("_send_call_feedback_report_document", body)
        self.assertNotIn("_send_telegram_text_message", body)
        self.assertIn("sendDocument", source)

    def test_preview_endpoint_and_frontend_action_exist(self):
        backend = _read(BOT_PATH)
        frontend = _read(MAIN_JSX_PATH)
        self.assertIn("@app.route('/api/admin/call_feedback_report_preview', methods=['POST'])", backend)
        self.assertIn("def admin_call_feedback_report_preview():", backend)
        self.assertIn("/api/admin/call_feedback_report_preview", frontend)
        self.assertIn("Отправить Excel", frontend)
        self.assertIn("Разовая отправка не включает еженедельную подписку", frontend)

    def test_report_payload_contains_all_evaluations(self):
        database_source = _read(DATABASE_PATH)
        self.assertIn('"items": []', database_source)
        self.assertIn('bucket["items"].append(call_item)', database_source)
        self.assertIn('"evaluation_date": item.get("evaluation_date")', database_source)

    def test_department_heads_can_subscribe_and_receive_scoped_report(self):
        backend = _read(BOT_PATH)
        database_source = _read(DATABASE_PATH)
        frontend = _read(MAIN_JSX_PATH)

        self.assertIn("Only admins and department heads can manage this setting", backend)
        self.assertIn("get_headed_departments_for_user(requester_id)", backend)
        self.assertIn("department_ids=department_ids", backend)
        self.assertIn("scope_label=scope_label", backend)
        self.assertIn("d.head_user_id = u.id", database_source)
        self.assertIn('"department_ids": None if is_admin else department_ids', database_source)
        self.assertIn("op.department_id = ANY(%s)", database_source)
        self.assertIn("isGlobalAdminRole || isDepartmentHead", frontend)
        self.assertIn("feedbackReportSetting.departmentName", frontend)


if __name__ == "__main__":
    unittest.main()
