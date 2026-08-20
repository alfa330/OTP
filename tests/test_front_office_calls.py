# -*- coding: utf-8 -*-
"""Обзвон фронт-офиса: разбор ответа CRM, сведение с реестром и текст отбивки.

front_office_calls.py — чистая логика без БД и Flask, поэтому импортируется
напрямую (в отличие от bot_schedule2.py, который на старте поднимает пул к
боевой БД). Сеть в тестах не трогаем: клиент проверяем на подменённом _post.

Здесь же разрез «город × сотрудник × дни» и сборка .xlsx: город приходит из
кадровой карточки, дни — по одному запросу CRM на день, и оба измерения
проверяются на готовом файле, а не на моках.
"""
import ast
import sys
import unittest
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import front_office_calls as foc


def _user(uid, name, email=None):
    return {"id": uid, "name": name, "email": email}


def _crm(manager_id, name, login, calls):
    return {"manager_id": manager_id, "manager_name": name,
            "manager_login": login, "total_calls": calls,
            "by_status": [{"status_id": 23, "status_code": "agreement",
                           "status_title": "Согласие", "count": calls}]}


DAY = date(2026, 8, 16)


class FetchManagersTests(unittest.TestCase):
    """Разбор строгий: незнакомый формат = ошибка, а не пустой список."""

    def _client(self, payload):
        client = foc.RegionCallStatsClient("http://crm.test", "token")
        client._post = lambda body: payload
        return client

    def test_returns_managers_as_is(self):
        rows = [_crm(1, "Иванов Иван", "ivanov@x.kz", 7)]
        client = self._client({"total": 7, "dialog_statuses": [], "managers": rows})
        self.assertEqual(client.fetch_managers(DAY, DAY), rows)

    def test_empty_day_is_not_an_error(self):
        # Будущая дата — законный пустой ответ живого API, а не поломка.
        client = self._client({"total": 0, "dialog_statuses": [], "managers": []})
        self.assertEqual(client.fetch_managers(DAY, DAY), [])

    def test_missing_managers_key_raises(self):
        # Так CRM уже ломала конкурс регистраций: ключ пропал, а синк
        # отрапортовал «ok» и записал пустой срез.
        client = self._client({"total": 12, "rows": []})
        with self.assertRaises(RuntimeError) as ctx:
            client.fetch_managers(DAY, DAY)
        self.assertIn("managers", str(ctx.exception))

    def test_manager_without_total_calls_raises(self):
        client = self._client({"managers": [{"manager_id": 1, "manager_name": "И"}]})
        with self.assertRaises(RuntimeError):
            client.fetch_managers(DAY, DAY)

    def test_html_answer_raises(self):
        # На кривой период CRM отдаёт HTML-страницу с кодом 200.
        client = self._client("<!doctype html><html></html>")
        with self.assertRaises(RuntimeError):
            client.fetch_managers(DAY, DAY)

    def test_period_is_sent_inclusive(self):
        sent = {}
        client = foc.RegionCallStatsClient("http://crm.test", "token")
        client._post = lambda body: sent.update(body) or {"managers": []}
        client.fetch_managers(date(2026, 8, 10), date(2026, 8, 12))
        self.assertEqual(sent, {"from": "2026-08-10", "to": "2026-08-12"})


class BuildReportTests(unittest.TestCase):
    def test_manager_without_calls_stays_in_report(self):
        # Главное правило: CRM присылает только звонивших, а в отбивке нужен
        # именно молчун — иначе «план не выполнен» никогда не сработает.
        roster = [_user(1, "Иванов Иван", "ivanov@x.kz"),
                  _user(2, "Петров Пётр", "petrov@x.kz")]
        report = foc.build_report(roster, [_crm(10, "Иванов Иван", "ivanov@x.kz", 5)],
                                  DAY, plan_per_day=3)
        by_name = {r["name"]: r for r in report["rows"]}
        self.assertEqual(by_name["Петров Пётр"]["calls"], 0)
        self.assertIs(by_name["Петров Пётр"]["met"], False)
        self.assertIs(by_name["Иванов Иван"]["met"], True)
        self.assertEqual([r["name"] for r in report["missing"]], ["Петров Пётр"])
        self.assertEqual(report["roster_size"], 2)
        self.assertEqual(report["called_count"], 1)

    def test_matched_by_email_case_insensitive(self):
        roster = [_user(1, "Иванов Иван", "Ivanov@X.kz")]
        report = foc.build_report(roster, [_crm(10, "кто-то другой", "ivanov@x.kz", 4)], DAY)
        self.assertEqual(report["rows"][0]["calls"], 4)
        self.assertEqual(report["unmatched_calls"], 0)

    def test_matched_by_name_when_logins_differ(self):
        # Живой случай: у CRM kairzhanova_diana_fo@, у нас diana_kairzhanova_fo@.
        roster = [_user(426, "Каиржанова Диана", "diana_kairzhanova_fo@yandextaxi.kz")]
        crm = [_crm(468, "Каиржанова Диана", "kairzhanova_diana_fo@yandextaxi.kz", 5)]
        report = foc.build_report(roster, crm, DAY)
        self.assertEqual(report["rows"][0]["calls"], 5)
        self.assertEqual(report["unmatched_calls"], 0)

    def test_two_crm_accounts_of_one_person_are_summed(self):
        # У отдела есть люди с парой учёток CRM из разных наборов id.
        roster = [_user(1, "Иванов Иван", "ivanov@x.kz")]
        crm = [_crm(10, "Иванов Иван", "ivanov@x.kz", 4),
               _crm(99, "Иванов Иван", None, 3)]
        report = foc.build_report(roster, crm, DAY)
        self.assertEqual(report["rows"][0]["calls"], 7)
        self.assertEqual(report["total_calls"], 7)

    def test_unknown_manager_with_calls_is_shown_separately(self):
        # Строки без имени и логина живой API отдаёт: терять их звонки молча нельзя.
        roster = [_user(1, "Иванов Иван", "ivanov@x.kz")]
        crm = [_crm(10, "Иванов Иван", "ivanov@x.kz", 4), _crm(587, None, None, 3)]
        report = foc.build_report(roster, crm, DAY)
        self.assertEqual(report["total_calls"], 4)
        self.assertEqual(report["unmatched_calls"], 3)
        self.assertEqual(report["unmatched"][0]["crm_manager_id"], 587)

    def test_unknown_manager_without_calls_is_ignored(self):
        report = foc.build_report([_user(1, "Иванов Иван")], [_crm(587, None, None, 0)], DAY)
        self.assertEqual(report["unmatched"], [])

    def test_rows_sorted_by_calls_then_name(self):
        roster = [_user(1, "Борисов Борис"), _user(2, "Абдулов Абдул"), _user(3, "Ким Ким")]
        crm = [_crm(1, "Борисов Борис", None, 2), _crm(2, "Абдулов Абдул", None, 2),
               _crm(3, "Ким Ким", None, 9)]
        report = foc.build_report(roster, crm, DAY)
        self.assertEqual([r["name"] for r in report["rows"]],
                         ["Ким Ким", "Абдулов Абдул", "Борисов Борис"])

    def test_without_plan_nobody_is_missing(self):
        report = foc.build_report([_user(1, "Иванов Иван")], [], DAY)
        self.assertIsNone(report["plan_total"])
        self.assertEqual(report["missing"], [])
        self.assertIsNone(report["rows"][0]["met"])

    def test_plan_multiplies_over_period(self):
        # За три дня норма — тройная, иначе период всегда «выполнен».
        roster = [_user(1, "Иванов Иван"), _user(2, "Петров Пётр")]
        crm = [_crm(1, "Иванов Иван", None, 30), _crm(2, "Петров Пётр", None, 20)]
        report = foc.build_report(roster, crm, date(2026, 8, 10), date(2026, 8, 12),
                                  plan_per_day=10)
        self.assertEqual(report["days"], 3)
        self.assertEqual(report["plan_total"], 30)
        self.assertEqual([r["name"] for r in report["missing"]], ["Петров Пётр"])

    def test_zero_plan_treated_as_unset(self):
        report = foc.build_report([_user(1, "Иванов Иван")], [], DAY, plan_per_day=0)
        self.assertIsNone(report["plan_total"])

    def test_has_data_false_only_when_nothing_at_all(self):
        roster = [_user(1, "Иванов Иван")]
        self.assertFalse(foc.has_data(foc.build_report(roster, [], DAY)))
        self.assertTrue(foc.has_data(
            foc.build_report(roster, [_crm(1, "Иванов Иван", None, 1)], DAY)))
        # Даже если звонили только «чужие», день рабочий — молчать нельзя.
        self.assertTrue(foc.has_data(
            foc.build_report(roster, [_crm(587, None, None, 2)], DAY)))


class RenderReportTests(unittest.TestCase):
    def _report(self, plan=10):
        roster = [_user(1, "Иванов Иван"), _user(2, "Петров Пётр"),
                  _user(3, "Сидоров Сидор")]
        crm = [_crm(1, "Иванов Иван", None, 12), _crm(2, "Петров Пётр", None, 3)]
        return foc.build_report(roster, crm, DAY, plan_per_day=plan)

    def test_alert_lists_only_those_below_plan(self):
        text = foc.render_report(self._report(), only_missing=True)
        self.assertIn("Петров Пётр — 3 из 10", text)
        self.assertIn("Сидоров Сидор — 0 из 10", text)
        # Выполнивший план в утреннюю отбивку поимённо не попадает.
        self.assertNotIn("Иванов Иван", text)
        self.assertIn("Не выполнили: 2 из 3.", text)

    def test_full_report_lists_everyone_with_one_divider(self):
        text = foc.render_report(self._report())
        for name in ("Иванов Иван", "Петров Пётр", "Сидоров Сидор"):
            self.assertIn(name, text)
        self.assertEqual(text.count("— ниже плана —"), 1)
        # Черта стоит ровно между выполнившим и первым отставшим.
        lines = text.splitlines()
        self.assertLess(lines.index("Иванов Иван — 12"), lines.index("— ниже плана —"))
        self.assertLess(lines.index("— ниже плана —"), lines.index("Петров Пётр — 3"))

    def test_no_divider_when_nobody_reached_the_plan(self):
        # Живой случай 16.08: план не вытянул никто — черте не место в начале.
        roster = [_user(1, "Иванов Иван"), _user(2, "Петров Пётр")]
        crm = [_crm(1, "Иванов Иван", None, 2), _crm(2, "Петров Пётр", None, 1)]
        text = foc.render_report(foc.build_report(roster, crm, DAY, plan_per_day=10))
        self.assertNotIn("— ниже плана —", text)

    def test_unmatched_ids_are_grouped_under_one_prefix(self):
        roster = [_user(1, "Иванов Иван")]
        crm = [_crm(583, None, None, 2), _crm(476, None, None, 1),
               _crm(608, "Новенький Новичок", None, 3)]
        text = foc.render_report(foc.build_report(roster, crm, DAY))
        self.assertIn("Новенький Новичок, id 583, 476", text)

    def test_full_report_without_plan_has_no_divider(self):
        text = foc.render_report(self._report(plan=None))
        self.assertNotIn("— ниже плана —", text)
        self.assertNotIn("Не выполнили", text)

    def test_period_header_shows_both_dates(self):
        report = foc.build_report([_user(1, "Иванов Иван")], [],
                                  date(2026, 8, 10), date(2026, 8, 12), plan_per_day=10)
        text = foc.render_report(report)
        self.assertIn("10.08.2026 — 12.08.2026", text)
        self.assertIn("План: 10 в день, за 3 дня — 30.", text)

    def test_single_day_header(self):
        text = foc.render_report(self._report())
        self.assertIn("Обзвон фронт-офиса за 16.08.2026", text)
        self.assertIn("План: 10 звонков в день.", text)

    def test_unmatched_calls_are_mentioned(self):
        roster = [_user(1, "Иванов Иван")]
        crm = [_crm(1, "Иванов Иван", None, 4), _crm(587, None, None, 3)]
        text = foc.render_report(foc.build_report(roster, crm, DAY))
        self.assertIn("Ещё 3 звонка", text)
        self.assertIn("id 587", text)

    def test_html_in_names_is_escaped(self):
        roster = [_user(1, "Иванов <b>Иван</b>")]
        text = foc.render_report(foc.build_report(roster, [], DAY))
        self.assertIn("Иванов &lt;b&gt;Иван&lt;/b&gt;", text)

    def test_plural_forms(self):
        self.assertIn("Всего 1 звонок", foc.render_report(
            foc.build_report([_user(1, "И")], [_crm(1, "И", None, 1)], DAY)))
        self.assertIn("Всего 2 звонка", foc.render_report(
            foc.build_report([_user(1, "И")], [_crm(1, "И", None, 2)], DAY)))
        self.assertIn("Всего 11 звонков", foc.render_report(
            foc.build_report([_user(1, "И")], [_crm(1, "И", None, 11)], DAY)))


class ConfigTests(unittest.TestCase):
    def test_url_defaults_without_env(self):
        with patch.dict("os.environ", {}, clear=False):
            with patch.object(foc.reg_contest, "get_config", return_value={"token": "t"}):
                cfg = foc.get_config()
        self.assertEqual(cfg["url"], foc.DEFAULT_API_URL)
        self.assertTrue(foc.is_configured(cfg))

    def test_not_configured_without_token(self):
        with patch.object(foc.reg_contest, "get_config", return_value={"token": None}):
            self.assertFalse(foc.is_configured(foc.get_config()))


class YesterdayTests(unittest.TestCase):
    def test_yesterday_crosses_month(self):
        self.assertEqual(foc.yesterday(date(2026, 9, 1)), date(2026, 8, 31))


class ParsePeriodTests(unittest.TestCase):
    TODAY = date(2026, 8, 17)

    def test_empty_argument_is_yesterday(self):
        self.assertEqual(foc.parse_period("", self.TODAY),
                         (date(2026, 8, 16), date(2026, 8, 16)))
        self.assertEqual(foc.parse_period(None, self.TODAY),
                         (date(2026, 8, 16), date(2026, 8, 16)))

    def test_single_day_formats(self):
        expected = (date(2026, 8, 12), date(2026, 8, 12))
        for text in ("12.08.2026", "12.08", "2026-08-12"):
            self.assertEqual(foc.parse_period(text, self.TODAY), expected, text)

    def test_range(self):
        self.assertEqual(foc.parse_period("10.08 12.08", self.TODAY),
                         (date(2026, 8, 10), date(2026, 8, 12)))
        self.assertEqual(foc.parse_period("10.08, 12.08", self.TODAY),
                         (date(2026, 8, 10), date(2026, 8, 12)))

    def test_reversed_range_is_straightened(self):
        # CRM на перевёрнутом периоде отдаёт HTML с кодом 200 — не доводим до неё.
        self.assertEqual(foc.parse_period("12.08 10.08", self.TODAY),
                         (date(2026, 8, 10), date(2026, 8, 12)))

    def test_garbage_returns_none(self):
        for text in ("вчера", "32.08", "10.08 11.08 12.08", "08/10"):
            self.assertIsNone(foc.parse_period(text, self.TODAY), text)


if __name__ == "__main__":
    unittest.main()


# ---------------------------------------------------------------------------
# Город, дни в штате и поденная сетка
# ---------------------------------------------------------------------------

def _fo(uid, name, email=None, city=None, hire_date=None):
    """Строка реестра в том виде, в котором её отдаёт get_front_office_call_roster."""
    return {"id": uid, "name": name, "email": email, "city": city,
            "hire_date": hire_date}


PERIOD_FROM = date(2026, 8, 1)
PERIOD_TO = date(2026, 8, 5)


def _days(date_from=PERIOD_FROM, date_to=PERIOD_TO):
    day = date_from
    while day <= date_to:
        yield day
        day += timedelta(days=1)


class CityTests(unittest.TestCase):
    """Город берётся из карточки: CRM его не присылает вовсе."""

    def test_city_from_roster_lands_in_rows(self):
        report = foc.build_report([_fo(1, "Иванов Иван", "i@x.kz", city="Актау")],
                                  [_crm(1, "Иванов Иван", "i@x.kz", 5)], DAY)
        self.assertEqual(report["rows"][0]["city"], "Актау")

    def test_empty_city_becomes_placeholder(self):
        report = foc.build_report([_fo(1, "Иванов Иван", "i@x.kz", city="  ")],
                                  [], DAY)
        self.assertEqual(report["rows"][0]["city"], foc.CITY_UNKNOWN)

    def test_cities_sum_calls_and_headcount(self):
        roster = [_fo(1, "А А", "a@x.kz", city="Алматы"),
                  _fo(2, "Б Б", "b@x.kz", city="Алматы"),
                  _fo(3, "В В", "c@x.kz", city="Астана")]
        report = foc.build_report(roster, [_crm(1, "А А", "a@x.kz", 4),
                                           _crm(2, "Б Б", "b@x.kz", 6)], DAY)
        by_city = {city["city"]: city for city in report["cities"]}
        self.assertEqual(by_city["Алматы"]["calls"], 10)
        self.assertEqual(by_city["Алматы"]["roster_size"], 2)
        self.assertEqual(by_city["Астана"]["calls"], 0)

    def test_worst_city_first_and_unknown_last(self):
        roster = [_fo(1, "А А", "a@x.kz", city="Алматы"),
                  _fo(2, "Б Б", "b@x.kz", city="Астана"),
                  _fo(3, "В В", "c@x.kz", city=None)]
        report = foc.build_report(roster, [_crm(1, "А А", "a@x.kz", 1),
                                           _crm(2, "Б Б", "b@x.kz", 9),
                                           _crm(3, "В В", "c@x.kz", 5)],
                                  DAY, plan_per_day=10)
        self.assertEqual([city["city"] for city in report["cities"]],
                         ["Алматы", "Астана", foc.CITY_UNKNOWN])


class StaffDaysTests(unittest.TestCase):
    """План умножается на личные дни в штате, а не на длину периода."""

    def test_hire_inside_period_shrinks_plan(self):
        roster = [_fo(1, "А А", "a@x.kz", city="Алматы",
                      hire_date=date(2026, 8, 4))]
        report = foc.build_report(roster, [], PERIOD_FROM, PERIOD_TO,
                                  plan_per_day=10)
        row = report["rows"][0]
        self.assertEqual(row["staff_days"], 2)
        self.assertEqual(row["plan_total"], 20)
        # Общий ориентир отдела при этом остаётся за весь период.
        self.assertEqual(report["plan_total"], 50)

    def test_hire_before_period_gives_full_plan(self):
        roster = [_fo(1, "А А", "a@x.kz", hire_date=date(2020, 1, 1))]
        report = foc.build_report(roster, [], PERIOD_FROM, PERIOD_TO,
                                  plan_per_day=10)
        self.assertEqual(report["rows"][0]["plan_total"], 50)

    def test_broken_hire_year_is_ignored(self):
        # На проде есть карточка с годом 0024 вместо 2024.
        roster = [_fo(1, "А А", "a@x.kz", hire_date=date(24, 9, 11))]
        report = foc.build_report(roster, [], PERIOD_FROM, PERIOD_TO,
                                  plan_per_day=10)
        self.assertIsNone(report["rows"][0]["hire_date"])
        self.assertEqual(report["rows"][0]["staff_days"], 5)

    def test_hire_after_period_is_not_a_failure(self):
        roster = [_fo(1, "А А", "a@x.kz", hire_date=date(2026, 9, 1))]
        report = foc.build_report(roster, [], PERIOD_FROM, PERIOD_TO,
                                  plan_per_day=10)
        self.assertEqual(report["rows"][0]["staff_days"], 0)
        self.assertIsNone(report["rows"][0]["met"])
        self.assertEqual(report["missing"], [])

    def test_day_without_crm_answer_leaves_plan(self):
        roster = [_fo(1, "А А", "a@x.kz")]
        report = foc.build_report(roster, [], PERIOD_FROM, PERIOD_TO,
                                  plan_per_day=10,
                                  no_data_days=[date(2026, 8, 3)])
        self.assertEqual(report["rows"][0]["staff_days"], 4)
        self.assertEqual(report["rows"][0]["plan_total"], 40)


class DailyReportTests(unittest.TestCase):
    """Поденная сетка: один запрос CRM на день, итоги — той же формулой."""

    def _roster(self):
        return [_fo(1, "А А", "a@x.kz", city="Алматы"),
                _fo(2, "Б Б", "b@x.kz", city="Астана")]

    def _by_day(self):
        return {day: [_crm(1, "А А", "a@x.kz", day.day),
                      _crm(2, "Б Б", "b@x.kz", 1)]
                for day in _days()}

    def test_per_day_holds_every_day(self):
        report = foc.build_daily_report(self._roster(), self._by_day())
        row = next(r for r in report["rows"] if r["user_id"] == 1)
        self.assertEqual(row["per_day"], {day: day.day for day in _days()})

    def test_totals_match_single_shot_report(self):
        by_day = self._by_day()
        daily = foc.build_daily_report(self._roster(), by_day, plan_per_day=3)
        flat = [row for day in sorted(by_day) for row in by_day[day]]
        single = foc.build_report(self._roster(), flat, PERIOD_FROM, PERIOD_TO,
                                  plan_per_day=3)
        self.assertEqual(daily["total_calls"], single["total_calls"])
        self.assertEqual([r["calls"] for r in daily["rows"]],
                         [r["calls"] for r in single["rows"]])

    def test_days_met_counts_cells_not_totals(self):
        report = foc.build_daily_report(self._roster(), self._by_day(),
                                        plan_per_day=3)
        row = next(r for r in report["rows"] if r["user_id"] == 1)
        # Звонки по дням 1..5, норма 3 — взята в дни 3, 4, 5.
        self.assertEqual(row["days_met"], 3)

    def test_per_day_totals_skip_days_without_answer(self):
        by_day = self._by_day()
        missing = date(2026, 8, 3)
        by_day.pop(missing)
        report = foc.build_daily_report(self._roster(), by_day,
                                        no_data_days=[missing])
        self.assertNotIn(missing, report["per_day_totals"])
        self.assertEqual(len(report["day_list"]), 5)

    def test_unmatched_counts_days_with_calls(self):
        by_day = self._by_day()
        for day in list(by_day)[:2]:
            by_day[day].append(_crm(583, None, None, 7))
        report = foc.build_daily_report(self._roster(), by_day)
        entry = next(e for e in report["unmatched"] if e["crm_manager_id"] == 583)
        self.assertEqual(entry["calls"], 14)
        self.assertEqual(entry["days"], 2)

    def test_empty_input_is_an_error_not_an_empty_report(self):
        with self.assertRaises(ValueError):
            foc.build_daily_report(self._roster(), {})

    def test_control_total_is_carried_into_report(self):
        report = foc.build_daily_report(self._roster(), self._by_day(),
                                        crm_period_total=999)
        self.assertEqual(report["crm_period_total"], 999)


class DayCellTests(unittest.TestCase):
    """Три состояния клетки: ноль, «не было выгрузки» и «ещё не принят»."""

    def _report(self, hire_date=None, no_data=()):
        roster = [_fo(1, "А А", "a@x.kz", city="Алматы", hire_date=hire_date)]
        by_day = {day: [] for day in _days() if day not in no_data}
        return foc.build_daily_report(roster, by_day, no_data_days=no_data)

    def test_zero_is_data(self):
        report = self._report()
        self.assertEqual(foc.day_cell(report, report["rows"][0], PERIOD_FROM),
                         ("data", 0))

    def test_missing_export_is_not_zero(self):
        missing = date(2026, 8, 3)
        report = self._report(no_data=[missing])
        self.assertEqual(foc.day_cell(report, report["rows"][0], missing),
                         ("no_data", None))

    def test_before_hire_is_off_staff(self):
        report = self._report(hire_date=date(2026, 8, 4))
        self.assertEqual(foc.day_cell(report, report["rows"][0], PERIOD_FROM),
                         ("off_staff", None))
        self.assertEqual(foc.day_cell(report, report["rows"][0], date(2026, 8, 4)),
                         ("data", 0))


class RenderPeriodTests(unittest.TestCase):
    """За период в сообщение идут итоги и города, а не 22 человека на 31 день."""

    def _report(self):
        roster = [_fo(1, "Асанов Асан", "a@x.kz", city="Алматы"),
                  _fo(2, "Бекова Бота", "b@x.kz", city="Астана")]
        by_day = {day: [_crm(1, "Асанов Асан", "a@x.kz", 1),
                        _crm(2, "Бекова Бота", "b@x.kz", 9)]
                  for day in _days()}
        return foc.build_daily_report(roster, by_day, plan_per_day=10)

    def test_period_text_shows_city_table(self):
        text = foc.render_report(self._report())
        self.assertIn("<pre>", text)
        self.assertIn("Алматы", text)
        self.assertIn("Астана", text)

    def test_period_text_has_no_person_list(self):
        text = foc.render_report(self._report())
        self.assertNotIn("Бекова Бота —", text)
        self.assertIn("Ниже всех:", text)
        self.assertIn("Асанов Асан", text)

    def test_period_text_reports_average_and_percent(self):
        text = foc.render_report(self._report())
        self.assertIn("В среднем 5,0 на человека в день — 50% нормы.", text)

    def test_single_day_still_lists_everyone(self):
        roster = [_fo(1, "Асанов Асан", "a@x.kz", city="Алматы")]
        report = foc.build_daily_report(roster,
                                        {DAY: [_crm(1, "Асанов Асан", "a@x.kz", 3)]},
                                        plan_per_day=10)
        text = foc.render_report(report)
        self.assertIn("Асанов Асан — 3", text)
        self.assertNotIn("<pre>", text)

    def test_alert_uses_personal_plan(self):
        roster = [_fo(1, "Асанов Асан", "a@x.kz", city="Алматы",
                      hire_date=date(2026, 8, 4))]
        by_day = {day: [] for day in _days()}
        report = foc.build_daily_report(roster, by_day, plan_per_day=10)
        self.assertIn("Асанов Асан — 0 из 20",
                      foc.render_report(report, only_missing=True))

    def test_unmatched_tail_has_single_dot(self):
        roster = [_fo(index, "Имя %d" % index, "m%d@x.kz" % index, city="Алматы")
                  for index in range(1, 3)]
        rows = [_crm(600 + index, None, None, index) for index in range(1, 8)]
        report = foc.build_daily_report(roster, {DAY: rows})
        text = foc.render_report(report)
        self.assertIn("и др.", text)
        self.assertNotIn("и др..", text)


class WorkbookTests(unittest.TestCase):
    """Проверяем сам файл: листы, шапки, итоги, заливки и форматы."""

    def _report(self, plan_per_day=10, no_data=(), hire_date=None):
        roster = [_fo(1, "Асанов Асан", "a@x.kz", city="Алматы"),
                  _fo(2, "Бекова Бота", "b@x.kz", city="Астана",
                      hire_date=hire_date),
                  _fo(3, "Вагапов Вали", "c@x.kz", city=None)]
        by_day = {day: [_crm(1, "Асанов Асан", "a@x.kz", 11),
                        _crm(2, "Бекова Бота", "b@x.kz", 0)]
                  for day in _days() if day not in no_data}
        return foc.build_daily_report(roster, by_day, plan_per_day=plan_per_day,
                                      no_data_days=no_data)

    def _book(self, report):
        return load_workbook(BytesIO(foc.build_workbook(report)))

    def test_sheets_are_named_and_ordered(self):
        self.assertEqual(self._book(self._report()).sheetnames,
                         list(foc.SHEET_NAMES))

    def test_matrix_header_has_city_person_and_every_day(self):
        ws = self._book(self._report())["Обзвон"]
        header = [cell.value for cell in ws[4]]
        self.assertEqual(header[:3], ["Город", "Сотрудник", "Факт"])
        self.assertEqual([value.date() for value in header[-5:]], list(_days()))

    def test_single_day_matrix_has_no_day_columns(self):
        roster = [_fo(1, "Асанов Асан", "a@x.kz", city="Алматы")]
        report = foc.build_daily_report(roster,
                                        {DAY: [_crm(1, "Асанов Асан", "a@x.kz", 3)]},
                                        plan_per_day=10)
        header = [cell.value for cell in self._book(report)["Обзвон"][4]]
        self.assertEqual(header, ["Город", "Сотрудник", "Факт", "Норма",
                                  "Отклонение"])

    def test_without_plan_percent_columns_disappear(self):
        header = [cell.value
                  for cell in self._book(self._report(plan_per_day=None))["Обзвон"][4]]
        self.assertNotIn("% нормы", header)
        self.assertNotIn("Дней с нормой", header)

    def test_rows_grouped_by_city_with_unknown_last(self):
        ws = self._book(self._report())["Обзвон"]
        cities = [ws.cell(row=row, column=1).value
                  for row in range(5, ws.max_row)]
        self.assertEqual(cities, ["Алматы", "Астана", foc.CITY_UNKNOWN])

    def test_totals_row_is_outside_the_filter(self):
        ws = self._book(self._report())["Обзвон"]
        self.assertEqual(ws.cell(row=ws.max_row, column=1).value, "ИТОГО")
        self.assertTrue(ws.auto_filter.ref.endswith(str(ws.max_row - 1)))
        self.assertEqual(ws.freeze_panes, "C5")

    def test_zero_day_cell_is_a_number_shown_as_dash(self):
        ws = self._book(self._report())["Обзвон"]
        row = next(index for index in range(5, ws.max_row)
                   if ws.cell(row=index, column=2).value == "Бекова Бота")
        cell = ws.cell(row=row, column=ws.max_column)
        self.assertEqual(cell.value, 0)
        self.assertEqual(cell.number_format, '0;-0;"—"')

    def test_day_before_hire_is_empty_and_grey(self):
        report = self._report(hire_date=date(2026, 8, 4))
        ws = self._book(report)["Обзвон"]
        row = next(index for index in range(5, ws.max_row)
                   if ws.cell(row=index, column=2).value == "Бекова Бота")
        first_day = ws.cell(row=row, column=ws.max_column - 4)
        self.assertIsNone(first_day.value)
        self.assertEqual(first_day.fill.fgColor.rgb[-6:], foc._OFF_STAFF_FILL)

    def test_day_without_export_is_marked_in_the_header(self):
        missing = date(2026, 8, 3)
        ws = self._book(self._report(no_data=[missing]))["Обзвон"]
        header = [cell for cell in ws[4] if isinstance(cell.value, object)]
        marked = [cell for cell in header
                  if getattr(cell.value, "date", None)
                  and cell.value.date() == missing]
        self.assertEqual(len(marked), 1)
        self.assertEqual(marked[0].fill.fgColor.rgb[-6:], foc._NO_DATA_FILL)

    def test_met_day_cell_is_highlighted(self):
        ws = self._book(self._report())["Обзвон"]
        row = next(index for index in range(5, ws.max_row)
                   if ws.cell(row=index, column=2).value == "Асанов Асан")
        cell = ws.cell(row=row, column=ws.max_column)
        self.assertEqual(cell.value, 11)
        self.assertEqual(cell.fill.fgColor.rgb[-6:], foc._MET_FILL)

    def test_cities_sheet_lists_every_city_with_totals(self):
        ws = self._book(self._report())["Города"]
        cities = [ws.cell(row=row, column=1).value
                  for row in range(5, ws.max_row)]
        self.assertEqual(sorted(cities),
                         sorted(["Алматы", "Астана", foc.CITY_UNKNOWN]))
        self.assertEqual(ws.cell(row=ws.max_row, column=1).value, "ИТОГО")

    def test_notes_sheet_warns_about_missing_shifts(self):
        ws = self._book(self._report())["Как читать"]
        text = " ".join(str(cell.value or "")
                        for row in ws.iter_rows(min_row=4) for cell in row)
        self.assertIn("графиков смен", text)
        self.assertIn("ещё не был принят", text)

    def test_filename_and_caption(self):
        report = self._report()
        self.assertEqual(foc.report_filename(report),
                         "obzvon_front_office_2026-08-01_2026-08-05.xlsx")
        day_report = foc.build_daily_report([_fo(1, "А А", "a@x.kz")],
                                           {DAY: []})
        self.assertEqual(foc.report_filename(day_report),
                         "obzvon_front_office_2026-08-16.xlsx")
        self.assertIn("Сводка", foc.document_caption(report))


class TotalCallsTests(unittest.TestCase):
    def test_sums_raw_crm_rows(self):
        self.assertEqual(
            foc.total_calls([_crm(1, "А", "a", 3), _crm(2, "Б", "b", 4)]), 7)

    def test_empty_and_broken_rows_are_zero(self):
        self.assertEqual(foc.total_calls(None), 0)
        self.assertEqual(foc.total_calls([{"total_calls": None}]), 0)


class BotWiringTests(unittest.TestCase):
    """Проводка в bot_schedule2.py: импортировать его нельзя, читаем исходник."""

    @classmethod
    def setUpClass(cls):
        from tests import source_cache
        cls.bot_path = Path(__file__).resolve().parents[1] / "bot_schedule2.py"
        cls.source_cache = source_cache

    @staticmethod
    def _names(node):
        """Всё, что функция упоминает: и вызовы, и имена.

        Проверять только ast.Call нельзя: блокирующие функции уходят в
        run_in_executor через functools.partial, то есть передаются ИМЕНЕМ,
        и в дереве это ast.Name, а не вызов.
        """
        found = set()
        for inner in ast.walk(node):
            if isinstance(inner, ast.Name):
                found.add(inner.id)
            elif isinstance(inner, ast.Attribute):
                found.add(inner.attr)
        return found

    def _calls(self, name):
        return self._names(
            self.source_cache.function_node(str(self.bot_path), name))

    def test_report_is_collected_day_by_day(self):
        calls = self._calls("_front_office_calls_report")
        self.assertIn("fetch_managers", calls)
        self.assertIn("build_daily_report", calls)
        self.assertIn("get_front_office_call_roster", calls)

    def test_command_sends_the_table(self):
        calls = self._calls("front_office_calls_command")
        self.assertIn("_front_office_calls_document", calls)
        self.assertIn("_front_office_calls_send", calls)

    def test_broadcast_sends_the_table(self):
        calls = self._calls("front_office_calls_broadcast_job")
        self.assertIn("_front_office_calls_document", calls)
        self.assertIn("_front_office_calls_send", calls)

    def test_broadcast_builds_the_file_once_for_all_chats(self):
        node = self.source_cache.function_node(str(self.bot_path),
                                               "front_office_calls_broadcast_job")
        inside_loop = set()
        for loop in ast.walk(node):
            if isinstance(loop, (ast.For, ast.AsyncFor)):
                inside_loop |= self._names(loop)
        self.assertNotIn("_front_office_calls_document", inside_loop)
        self.assertIn("_front_office_calls_send", inside_loop)

    def test_sender_falls_back_to_a_message_without_the_file(self):
        calls = self._calls("_front_office_calls_send")
        self.assertIn("send_document", calls)
        self.assertIn("send_message", calls)
