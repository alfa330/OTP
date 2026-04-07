import os
import logging
import psycopg2
from contextlib import contextmanager
from passlib.hash import pbkdf2_sha256
from datetime import datetime, timedelta, date, time as dt_time
import time
import uuid
import requests
import openpyxl
import re
import json
import csv
import io
import math
from io import BytesIO
import asyncio
from concurrent.futures import ThreadPoolExecutor
import gc
from psycopg2.pool import ThreadedConnectionPool  # Added for connection pooling
from psycopg2.extras import execute_values, Json
import calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles.fills import Fill
from collections import defaultdict
import pandas as pd
from typing import List, Dict, Any, Tuple, Optional

logging.basicConfig(level=logging.INFO)

os.environ['TZ'] = 'Asia/Almaty'
time.tzset()

# Global connection pool
MIN_CONN = 1
MAX_CONN = 20  # Adjust based on expected load
POOL = None
STATUS_IMPORT_INSERT_PAGE_SIZE = max(200, int(os.getenv('STATUS_IMPORT_INSERT_PAGE_SIZE', '2000')))

ROLE_ALIASES = {
    'supervisor': 'sv',
    'superadmin': 'super_admin',
    'super-admin': 'super_admin',
    'super admin': 'super_admin'
}

ROLE_HIERARCHY = {
    'operator': 10,
    'trainee': 10,
    'trainer': 20,
    'sv': 30,
    'admin': 40,
    'super_admin': 50
}


def normalize_role_value(role_value: Optional[str]) -> str:
    role_norm = str(role_value or '').strip().lower()
    return ROLE_ALIASES.get(role_norm, role_norm)


def role_has_min(role_value: Optional[str], required_role: str) -> bool:
    role_level = int(ROLE_HIERARCHY.get(normalize_role_value(role_value), 0))
    required_level = int(ROLE_HIERARCHY.get(normalize_role_value(required_role), 0))
    if required_level <= 0:
        return False
    return role_level >= required_level


def role_is_any(role_value: Optional[str], allowed_roles: List[str]) -> bool:
    role_norm = normalize_role_value(role_value)
    if not role_norm:
        return False
    allowed = {normalize_role_value(item) for item in (allowed_roles or [])}
    return role_norm in allowed

# Вставьте/адаптируйте этот helper в ваш модуль
def _normalize_phone(phone: Optional[str]) -> Optional[str]:
    if not phone:
        return None
    digits = re.sub(r'\D', '', str(phone))
    if not digits:
        return None
    # Оставим, например, последние 10-11 цифр (зависит от вашей логики)
    return digits

def _parse_datetime_raw(dt_str: Optional[str]) -> Optional[datetime]:
    if not dt_str:
        return None
    # ожидаемый формат 'dd.mm.yyyy hh:mm:ss'
    formats = ['%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M']  # запасные форматы
    for fmt in formats:
        try:
            # возвращаем timezone-naive datetime; при вставке используем AT TIME ZONE 'Asia/Almaty' или локаль
            return datetime.strptime(dt_str, fmt)
        except Exception:
            continue
    return None

def _time_to_minutes(time_str: str) -> int:
    """Преобразует время в формате 'HH:MM' в минуты от начала дня."""
    try:
        parts = time_str.split(':')
        hours = int(parts[0])
        minutes = int(parts[1]) if len(parts) > 1 else 0
        return hours * 60 + minutes
    except (ValueError, AttributeError):
        return 0

def _minutes_to_time(minutes: int) -> str:
    """Преобразует минуты от начала дня в формат 'HH:MM'."""
    minutes = minutes % (24 * 60)  # нормализация для смен через полночь
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours:02d}:{mins:02d}"

def _merge_shifts_for_date(shifts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Объединяет перекрывающиеся смены на одну дату.
    shifts: список словарей с ключами 'start', 'end', 'breaks', 'id'
    Возвращает список объединенных смен.
    """
    if not shifts or len(shifts) == 0:
        return []
    
    # Преобразуем смены в интервалы в минутах
    intervals = []
    for shift in shifts:
        start_min = _time_to_minutes(shift.get('start', '00:00'))
        end_min = _time_to_minutes(shift.get('end', '00:00'))
        
        # Обработка смен через полночь
        if end_min <= start_min:
            end_min += 24 * 60
        
        intervals.append({
            'start': start_min,
            'end': end_min,
            'original': shift,
            'breaks': shift.get('breaks', [])
        })
    
    # Сортируем по времени начала
    intervals.sort(key=lambda x: x['start'])
    
    # Объединяем перекрывающиеся интервалы
    merged = []
    for interval in intervals:
        if not merged:
            merged.append(interval.copy())
        else:
            last = merged[-1]
            # Проверяем перекрытие: если начало новой смены <= конца последней
            if interval['start'] <= last['end']:
                # Объединяем: расширяем конец до максимума
                last['end'] = max(last['end'], interval['end'])
                # Объединяем перерывы
                last['breaks'].extend(interval['breaks'])
                # Сохраняем ID оригинальных смен для удаления старых
                if 'original_ids' not in last:
                    last['original_ids'] = [last['original'].get('id')]
                if interval['original'].get('id'):
                    last['original_ids'].append(interval['original'].get('id'))
            else:
                merged.append(interval.copy())
    
    # Преобразуем обратно в формат смен
    result = []
    for m in merged:
        start_time = _minutes_to_time(m['start'])
        end_time = _minutes_to_time(m['end'] % (24 * 60))
        
        # Объединяем и нормализуем перерывы (убираем дубликаты, сортируем)
        breaks = m.get('breaks', [])
        # Если перерывы в формате {'start': minutes, 'end': minutes}, оставляем как есть
        # Если в формате {'start': 'HH:MM', 'end': 'HH:MM'}, преобразуем
        normalized_breaks = []
        seen_breaks = set()
        for b in breaks:
            if isinstance(b, dict):
                start_br = b.get('start')
                end_br = b.get('end')
                # Если это строки времени, преобразуем в минуты
                if isinstance(start_br, str):
                    start_br_min = _time_to_minutes(start_br)
                else:
                    start_br_min = int(start_br) if start_br is not None else 0
                
                if isinstance(end_br, str):
                    end_br_min = _time_to_minutes(end_br)
                else:
                    end_br_min = int(end_br) if end_br is not None else 0
                
                break_key = (start_br_min, end_br_min)
                if break_key not in seen_breaks:
                    seen_breaks.add(break_key)
                    normalized_breaks.append({
                        'start': start_br_min,
                        'end': end_br_min
                    })
        
        # Сортируем перерывы по началу
        normalized_breaks.sort(key=lambda x: x['start'])
        
        result.append({
            'start': start_time,
            'end': end_time,
            'breaks': normalized_breaks
        })
    
    return result

SCHEDULE_SPECIAL_STATUS_META: Dict[str, Dict[str, str]] = {
    'bs': {
        'label': 'Б/С',
        'kind': 'absence'
    },
    'sick_leave': {
        'label': 'Больничный',
        'kind': 'absence'
    },
    'annual_leave': {
        'label': 'Ежегодный отпуск',
        'kind': 'absence'
    },
    'dismissal': {
        'label': 'Увольнение',
        'kind': 'dismissal'
    }
}

SCHEDULE_DISMISSAL_REASONS: List[str] = [
    'Б/С на летний период',
    'Мошенничество',
    'Нарушение дисциплины',
    'Не может совмещать с учебой',
    'Не может совмещать с работой',
    'Не нравится работа',
    'Выгорание',
    'Не устраивает доход',
    'Перевод в другой отдел',
    'Переезд',
    'По состоянию здоровья',
    'Пропал',
    'Слабый/не выполняет kpi',
    'Забрали в армию',
    'Нашел работу по профессии',
    'По семейным обстоятельствам'
]

USER_STATUS_ALLOWED_VALUES: List[str] = [
    'working',
    'fired',
    'unpaid_leave',
    'bs',
    'sick_leave',
    'annual_leave',
    'dismissal'
]

SCHEDULE_STATUS_TO_USER_STATUS: Dict[str, str] = {
    'bs': 'bs',
    'sick_leave': 'sick_leave',
    'annual_leave': 'annual_leave',
    # Для совместимости с остальными экранами "увольнение" продолжаем отражать как fired
    'dismissal': 'fired'
}

SCHEDULE_AUTO_AGGREGATION_START_DATE = date(2026, 3, 1)
SCHEDULE_AUTO_FLAG_PENDING = 'pending'
SCHEDULE_AUTO_FLAG_CONFIRMED = 'confirmed'
SCHEDULE_AUTO_FLAG_REJECTED = 'rejected'
SCHEDULE_AUTO_FLAG_ALLOWED = {
    SCHEDULE_AUTO_FLAG_PENDING,
    SCHEDULE_AUTO_FLAG_CONFIRMED,
    SCHEDULE_AUTO_FLAG_REJECTED
}
SCHEDULE_AUTO_WORK_STATUS_KEYS = {'готов', 'занят', 'занята', 'перезвон'}
SCHEDULE_AUTO_TALK_STATUS_KEYS = {'занят', 'занята'}
SCHEDULE_AUTO_BREAK_STATUS_KEYS = {'перерыв', 'авто'}
SCHEDULE_AUTO_NO_PHONE_STATUS_KEY = 'без телефона'
SCHEDULE_AUTO_TRAINING_STATUS_KEY = 'тренинг'
SCHEDULE_STATUS_KEY_LABELS = {
    'готов': 'Готов',
    'занят': 'Занят',
    'занята': 'Занята',
    'перезвон': 'Перезвон',
    'перерыв': 'Перерыв',
    'авто': 'Авто',
    'вышел': 'Вышел',
    'тренинг': 'Тренинг',
    'training': 'Training',
    'без телефона': 'Без телефона',
    'нет статуса': 'Нет статуса',
    'отключен': 'Отключен',
    'отключена': 'Отключена',
    'отключено': 'Отключено'
}
SCHEDULE_AUTO_FINE_RATE_PER_MINUTE = float(os.getenv('SCHEDULE_AUTO_FINE_RATE_PER_MINUTE', '50'))

TECHNICAL_ISSUE_REASONS: List[str] = [
    'Не работает интернет',
    'Замена мыши',
    'Не работает микрофон',
    'Не работает Oktell',
    'Проблема с маршрутизацией Oktell (не идут исходящие звонки), переключение в ручной режим',
    'Замена клавиатуры',
    'Не заходит в корпоративный чат',
    'Не включается компьютер',
    'Переполнена память',
    'Кнопка "Войти в колл-центр" в Oktell не реагирует на действия',
    'Виснет компьютер',
    'Не работают программы на ПК (ошибка "Меню \"Пуск\" не работает")',
    'Проблема с подключением к сайту Oktell',
    'Не может войти в учетную запись ПК',
    'Не поступают звонки',
    'Не может войти в учетную запись Oktell',
    'Отключение света',
    'Массовая проблема с Октелл',
    'Массовая проблема с интернетом',
    'Массовая проблема с телефонией'
]
TECHNICAL_ISSUE_REASONS_SET = set(TECHNICAL_ISSUE_REASONS)

def get_pool():
    global POOL
    if POOL is None:
        conn_params = {
            'dbname': os.getenv('POSTGRES_DB'),
            'user': os.getenv('POSTGRES_USER'),
            'password': os.getenv('POSTGRES_PASSWORD'),
            'host': os.getenv('POSTGRES_HOST'),
            'port': os.getenv('POSTGRES_PORT', 5432)
        }
        POOL = ThreadedConnectionPool(MIN_CONN, MAX_CONN, **conn_params)
    return POOL

class Database:
    SURVEY_OTHER_ANSWER_MAX_LENGTH = 500
    SCHEMA_INIT_LOCK_KEY = 915904137
    SCHEMA_INIT_LOCK_TIMEOUT_SEC = 120
    SCHEMA_INIT_RETRY_ATTEMPTS = 4

    def __init__(self):
        self._init_db_with_retry()

    @contextmanager
    def _get_connection(self):
        pool = get_pool()
        conn = pool.getconn()
        try:
            yield conn
        finally:
            pool.putconn(conn)

    @contextmanager
    def _get_cursor(self):
        with self._get_connection() as conn:
            cursor = conn.cursor()
            try:
                yield cursor
                conn.commit()
            except Exception as e:
                conn.rollback()
                raise e
            finally:
                cursor.close()

    @contextmanager
    def _schema_init_lock(self):
        with self._get_connection() as conn:
            cursor = conn.cursor()
            acquired = False
            started_at = time.time()
            try:
                while True:
                    cursor.execute("SELECT pg_try_advisory_lock(%s)", (self.SCHEMA_INIT_LOCK_KEY,))
                    acquired = bool(cursor.fetchone()[0])
                    if acquired:
                        break
                    if (time.time() - started_at) >= float(self.SCHEMA_INIT_LOCK_TIMEOUT_SEC):
                        raise TimeoutError("Timeout waiting for schema init advisory lock")
                    time.sleep(0.25)
                yield
            finally:
                if acquired:
                    try:
                        cursor.execute("SELECT pg_advisory_unlock(%s)", (self.SCHEMA_INIT_LOCK_KEY,))
                        conn.commit()
                    except Exception:
                        conn.rollback()
                        logging.exception("Failed to release schema init advisory lock")
                cursor.close()

    def _init_db_with_retry(self):
        last_error = None
        for attempt in range(1, int(self.SCHEMA_INIT_RETRY_ATTEMPTS) + 1):
            try:
                with self._schema_init_lock():
                    self._init_db()
                return
            except psycopg2.Error as exc:
                last_error = exc
                # 40P01 = deadlock_detected
                if getattr(exc, "pgcode", None) == '40P01' and attempt < int(self.SCHEMA_INIT_RETRY_ATTEMPTS):
                    delay = min(3.0, 0.4 * attempt)
                    logging.warning(
                        "Deadlock during DB init (attempt %s/%s). Retrying in %.1fs",
                        attempt,
                        self.SCHEMA_INIT_RETRY_ATTEMPTS,
                        delay
                    )
                    time.sleep(delay)
                    continue
                raise
        if last_error:
            raise last_error

    def _init_db(self):
        with self._get_cursor() as cursor:
            # Users table (без direction_id на этом этапе)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    telegram_id BIGINT UNIQUE,
                    name VARCHAR(255) NOT NULL,
                    role VARCHAR(20) NOT NULL CHECK(role IN ('super_admin', 'admin', 'sv', 'supervisor', 'trainer', 'operator', 'trainee')),
                    hire_date DATE,
                    login VARCHAR(255) UNIQUE,
                    password_hash VARCHAR(255),
                    supervisor_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    hours_table_url TEXT,
                    scores_table_url TEXT,
                    is_active BOOLEAN NOT NULL DEFAULT FALSE,
                    status VARCHAR(20) NOT NULL DEFAULT 'working' CHECK(status IN ('working', 'fired', 'unpaid_leave', 'bs', 'sick_leave', 'annual_leave', 'dismissal')),
                    rate DECIMAL(3,2) NOT NULL DEFAULT 1.00 CHECK(rate IN (1.00, 0.75, 0.50)),
                    CONSTRAINT unique_name_role UNIQUE (name, role)
                );
            """)
            cursor.execute("""
                DO $$
                DECLARE role_constraint RECORD;
                BEGIN
                    FOR role_constraint IN
                        SELECT conname
                        FROM pg_constraint
                        WHERE conrelid = 'users'::regclass
                          AND contype = 'c'
                          AND pg_get_constraintdef(oid) ILIKE '%role%'
                    LOOP
                        EXECUTE format('ALTER TABLE users DROP CONSTRAINT IF EXISTS %I', role_constraint.conname);
                    END LOOP;

                    BEGIN
                        ALTER TABLE users
                            ADD CONSTRAINT users_role_check
                            CHECK (role IN ('super_admin', 'admin', 'sv', 'supervisor', 'trainer', 'operator', 'trainee'));
                    EXCEPTION
                        WHEN duplicate_object THEN
                            NULL;
                    END;
                END $$;
            """)
            # Directions table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS directions (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    has_file_upload BOOLEAN NOT NULL DEFAULT TRUE,
                    criteria JSONB NOT NULL DEFAULT '[]',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    is_active BOOLEAN NOT NULL DEFAULT TRUE,
                    version INTEGER NOT NULL DEFAULT 1,
                    previous_version_id INTEGER REFERENCES directions(id)
                );
            """)

            # Добавляем direction_id в users после создания directions
            cursor.execute("""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns 
                        WHERE table_name = 'users' AND column_name = 'direction_id'
                    ) THEN
                        ALTER TABLE users ADD COLUMN direction_id INTEGER REFERENCES directions(id) ON DELETE SET NULL;
                    END IF;
                END $$;
            """)
            cursor.execute("""
                ALTER TABLE users
                    ADD COLUMN IF NOT EXISTS gender VARCHAR(20) CHECK (gender IN ('male', 'female')),
                    ADD COLUMN IF NOT EXISTS birth_date DATE;
            """)
            cursor.execute("""
                ALTER TABLE users
                    ADD COLUMN IF NOT EXISTS avatar_bucket VARCHAR(255),
                    ADD COLUMN IF NOT EXISTS avatar_blob_path TEXT,
                    ADD COLUMN IF NOT EXISTS avatar_original_blob_path TEXT,
                    ADD COLUMN IF NOT EXISTS avatar_content_type VARCHAR(128),
                    ADD COLUMN IF NOT EXISTS avatar_file_size INTEGER,
                    ADD COLUMN IF NOT EXISTS avatar_updated_at TIMESTAMP;
            """)
            cursor.execute("""
                ALTER TABLE users
                    ADD COLUMN IF NOT EXISTS phone VARCHAR(50),
                    ADD COLUMN IF NOT EXISTS email VARCHAR(255),
                    ADD COLUMN IF NOT EXISTS instagram VARCHAR(255),
                    ADD COLUMN IF NOT EXISTS telegram_nick VARCHAR(255),
                    ADD COLUMN IF NOT EXISTS company_name VARCHAR(255),
                    ADD COLUMN IF NOT EXISTS employment_type VARCHAR(10),
                    ADD COLUMN IF NOT EXISTS has_proxy BOOLEAN NOT NULL DEFAULT FALSE,
                    ADD COLUMN IF NOT EXISTS proxy_card_number VARCHAR(64),
                    ADD COLUMN IF NOT EXISTS has_driver_license BOOLEAN NOT NULL DEFAULT FALSE,
                    ADD COLUMN IF NOT EXISTS sip_number VARCHAR(64),
                    ADD COLUMN IF NOT EXISTS study_place VARCHAR(255),
                    ADD COLUMN IF NOT EXISTS study_course VARCHAR(100),
                    ADD COLUMN IF NOT EXISTS close_contact_1_relation VARCHAR(100),
                    ADD COLUMN IF NOT EXISTS close_contact_1_full_name VARCHAR(255),
                    ADD COLUMN IF NOT EXISTS close_contact_1_phone VARCHAR(50),
                    ADD COLUMN IF NOT EXISTS close_contact_2_relation VARCHAR(100),
                    ADD COLUMN IF NOT EXISTS close_contact_2_full_name VARCHAR(255),
                    ADD COLUMN IF NOT EXISTS close_contact_2_phone VARCHAR(50),
                    ADD COLUMN IF NOT EXISTS card_number VARCHAR(64),
                    ADD COLUMN IF NOT EXISTS internship_in_company BOOLEAN NOT NULL DEFAULT FALSE,
                    ADD COLUMN IF NOT EXISTS front_office_training BOOLEAN NOT NULL DEFAULT FALSE,
                    ADD COLUMN IF NOT EXISTS front_office_training_date DATE,
                    ADD COLUMN IF NOT EXISTS taxipro_id VARCHAR(128);
            """)
            cursor.execute("""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1
                        FROM pg_constraint
                        WHERE conname = 'users_employment_type_check'
                          AND conrelid = 'users'::regclass
                    ) THEN
                        ALTER TABLE users
                            ADD CONSTRAINT users_employment_type_check
                            CHECK (employment_type IN ('gph', 'of') OR employment_type IS NULL);
                    END IF;
                END $$;
            """)
            # Calls table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operator_activity_logs (
                    id SERIAL PRIMARY KEY,
                    operator_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                    is_active VARCHAR(20) NOT NULL,
                    change_time TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS calls (
                    id SERIAL PRIMARY KEY,
                    evaluator_id INTEGER NOT NULL REFERENCES users(id),
                    operator_id INTEGER NOT NULL REFERENCES users(id),
                    month VARCHAR(7) NOT NULL DEFAULT TO_CHAR(CURRENT_DATE, 'YYYY-MM'),
                    phone_number VARCHAR(255) NOT NULL,
                    appeal_date TIMESTAMP,
                    score FLOAT NOT NULL,
                    comment TEXT,
                    audio_path TEXT,
                    is_draft BOOLEAN NOT NULL DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    is_correction BOOLEAN DEFAULT FALSE,
                    previous_version_id INTEGER REFERENCES calls(id),
                    scores JSONB,
                    criterion_comments JSONB,
                    direction_id INTEGER REFERENCES directions(id) ON DELETE SET NULL,
                    UNIQUE(evaluator_id, operator_id, month, phone_number, score, comment, is_draft)
                );
                ALTER TABLE calls
                    ADD COLUMN IF NOT EXISTS comment_visible_to_operator BOOLEAN NOT NULL DEFAULT TRUE,
                    ADD COLUMN IF NOT EXISTS sv_request BOOLEAN NOT NULL DEFAULT FALSE,
                    ADD COLUMN IF NOT EXISTS sv_request_comment TEXT,
                    ADD COLUMN IF NOT EXISTS sv_request_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    ADD COLUMN IF NOT EXISTS sv_request_at TIMESTAMP,
                    ADD COLUMN IF NOT EXISTS sv_request_approved BOOLEAN NOT NULL DEFAULT FALSE,
                    ADD COLUMN IF NOT EXISTS sv_request_approved_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    ADD COLUMN IF NOT EXISTS sv_request_approved_at TIMESTAMP;
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS imported_calls (
                    id SERIAL PRIMARY KEY,
                    external_id TEXT,                      -- id из JSON (можно UUID или любой string)
                    operator_name TEXT NOT NULL,           -- ФИО из JSON
                    operator_id INTEGER,                   -- resolved users.id (NULL допустим)
                    month VARCHAR(7) NOT NULL,             -- format 'YYYY-MM'
                    datetime_raw TIMESTAMP WITH TIME ZONE, -- parsed datetime (store with timezone)
                    phone_number TEXT,
                    phone_normalized TEXT,
                    duration_sec DOUBLE PRECISION,
                    desired INTEGER,
                    available INTEGER,
                    status VARCHAR(20) NOT NULL DEFAULT 'not_evaluated', -- not_evaluated / evaluated / skipped
                    imported_at TIMESTAMP WITH TIME ZONE DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    evaluated_at TIMESTAMP WITH TIME ZONE,
                    evaluated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    notes TEXT
                );

                ALTER TABLE imported_calls ADD COLUMN id_int SERIAL;
                ALTER TABLE imported_calls DROP CONSTRAINT imported_calls_pkey;
                ALTER TABLE imported_calls DROP COLUMN id;
                ALTER TABLE imported_calls RENAME COLUMN id_int TO id;
                ALTER TABLE imported_calls ADD PRIMARY KEY (id);
                CREATE UNIQUE INDEX IF NOT EXISTS uq_imported_calls_external_id_month ON imported_calls (external_id, month);
                CREATE INDEX IF NOT EXISTS idx_imported_calls_month ON imported_calls(month);
                CREATE INDEX IF NOT EXISTS idx_imported_calls_operator_id ON imported_calls(operator_id);
                CREATE INDEX IF NOT EXISTS idx_imported_calls_status ON imported_calls(status);
                CREATE INDEX IF NOT EXISTS idx_imported_calls_phone_normalized ON imported_calls(phone_normalized);
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS calibration_rooms (
                    id SERIAL PRIMARY KEY,
                    created_by_admin_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    month VARCHAR(7) NOT NULL,
                    phone_number VARCHAR(255) NOT NULL,
                    appeal_date TIMESTAMP,
                    direction_id INTEGER REFERENCES directions(id) ON DELETE SET NULL,
                    score FLOAT NOT NULL DEFAULT 0,
                    comment TEXT,
                    audio_path TEXT,
                    scores JSONB NOT NULL DEFAULT '[]',
                    criterion_comments JSONB NOT NULL DEFAULT '[]',
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor.execute("""
                ALTER TABLE calibration_rooms
                ADD COLUMN IF NOT EXISTS room_title TEXT;
            """) 
            cursor.execute("""
                ALTER TABLE calibration_rooms
                ALTER COLUMN operator_id DROP NOT NULL;
            """)
            cursor.execute("""
                ALTER TABLE calibration_rooms
                ALTER COLUMN phone_number DROP NOT NULL;
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS calibration_room_members (
                    id SERIAL PRIMARY KEY,
                    room_id INTEGER NOT NULL REFERENCES calibration_rooms(id) ON DELETE CASCADE,
                    supervisor_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    joined_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(room_id, supervisor_id)
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS calibration_room_calls (
                    id SERIAL PRIMARY KEY,
                    room_id INTEGER NOT NULL REFERENCES calibration_rooms(id) ON DELETE CASCADE,
                    created_by_admin_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    month VARCHAR(7) NOT NULL,
                    phone_number VARCHAR(255) NOT NULL,
                    appeal_date TIMESTAMP,
                    direction_id INTEGER REFERENCES directions(id) ON DELETE SET NULL,
                    score FLOAT NOT NULL DEFAULT 0,
                    comment TEXT,
                    audio_path TEXT,
                    scores JSONB NOT NULL DEFAULT '[]',
                    criterion_comments JSONB NOT NULL DEFAULT '[]',
                    etalon_scores JSONB NOT NULL DEFAULT '[]',
                    etalon_criterion_comments JSONB NOT NULL DEFAULT '[]',
                    etalon_updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    etalon_updated_at TIMESTAMP,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor.execute("""
                ALTER TABLE calibration_room_calls
                    ADD COLUMN IF NOT EXISTS etalon_scores JSONB NOT NULL DEFAULT '[]',
                    ADD COLUMN IF NOT EXISTS etalon_criterion_comments JSONB NOT NULL DEFAULT '[]',
                    ADD COLUMN IF NOT EXISTS etalon_updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    ADD COLUMN IF NOT EXISTS etalon_updated_at TIMESTAMP,
                    ADD COLUMN IF NOT EXISTS general_comment TEXT,
                    ADD COLUMN IF NOT EXISTS general_comment_updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    ADD COLUMN IF NOT EXISTS general_comment_updated_at TIMESTAMP;
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS calibration_room_call_evaluations (
                    id SERIAL PRIMARY KEY,
                    room_call_id INTEGER NOT NULL REFERENCES calibration_room_calls(id) ON DELETE CASCADE,
                    evaluator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    score FLOAT NOT NULL DEFAULT 0,
                    comment TEXT,
                    scores JSONB NOT NULL DEFAULT '[]',
                    criterion_comments JSONB NOT NULL DEFAULT '[]',
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(room_call_id, evaluator_id)
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS calibration_room_evaluations (
                    id SERIAL PRIMARY KEY,
                    room_id INTEGER NOT NULL REFERENCES calibration_rooms(id) ON DELETE CASCADE,
                    evaluator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    score FLOAT NOT NULL DEFAULT 0,
                    comment TEXT,
                    scores JSONB NOT NULL DEFAULT '[]',
                    criterion_comments JSONB NOT NULL DEFAULT '[]',
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(room_id, evaluator_id)
                );
            """)


            #Trainings table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS trainings (
                    id SERIAL PRIMARY KEY,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    training_date DATE NOT NULL,
                    start_time TIME NOT NULL,
                    end_time TIME NOT NULL,
                    reason VARCHAR(255) NOT NULL CHECK(reason IN (
                        'Обратная связь', 'Собрание', 'Тех. сбой', 'Мотивационная беседа', 
                        'Дисциплинарный тренинг', 'Тренинг по качеству. Разбор ошибок', 
                        'Тренинг по качеству. Объяснение МШ', 'Тренинг по продукту', 
                        'Мониторинг', 'Практика в офисе таксопарка', 'Другое'
                    )),
                    comment TEXT,
                    created_by INTEGER REFERENCES users(id),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    count_in_hours BOOLEAN NOT NULL DEFAULT TRUE,
                    UNIQUE(operator_id, training_date, start_time, end_time)
                );
            """)
            
            # Work hours table with fines column
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS work_hours (
                    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                    operator_id INTEGER NOT NULL REFERENCES users(id),
                    month VARCHAR(7) NOT NULL DEFAULT TO_CHAR(CURRENT_DATE, 'YYYY-MM'),
                    regular_hours FLOAT NOT NULL DEFAULT 0,    
                    training_hours FLOAT NOT NULL DEFAULT 0,
                    fines FLOAT NOT NULL DEFAULT 0,
                    norm_hours FLOAT NOT NULL DEFAULT 0,
                    total_break_time FLOAT NOT NULL DEFAULT 0,
                    total_talk_time FLOAT NOT NULL DEFAULT 0,
                    total_calls INTEGER NOT NULL DEFAULT 0,
                    total_efficiency_hours FLOAT NOT NULL DEFAULT 0, 
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    calls_per_hour FLOAT NOT NULL DEFAULT 0.0,
                    UNIQUE(operator_id, month)
                );
            """)

            cursor.execute("""
                CREATE EXTENSION IF NOT EXISTS pgcrypto;
                CREATE TABLE IF NOT EXISTS daily_hours (
                    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    day DATE NOT NULL,
                    work_time FLOAT NOT NULL DEFAULT 0,   -- часы работы
                    break_time FLOAT NOT NULL DEFAULT 0,
                    talk_time FLOAT NOT NULL DEFAULT 0,
                    calls INTEGER NOT NULL DEFAULT 0,
                    efficiency FLOAT NOT NULL DEFAULT 0,  -- часы
                    fine_amount FLOAT NOT NULL DEFAULT 0,
                    fine_reason VARCHAR(64),
                    fine_comment TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(operator_id, day)
                );
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS training_time FLOAT NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS late_minutes INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS late_seconds INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS early_leave_minutes INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS early_leave_seconds INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS overtime_minutes INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS overtime_seconds INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS training_minutes INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS training_seconds INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS technical_reason_minutes INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS technical_reason_seconds INTEGER NOT NULL DEFAULT 0;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS late_status VARCHAR(16);
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS early_leave_status VARCHAR(16);
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS training_status VARCHAR(16);
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS technical_reason_status VARCHAR(16);
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS late_fine_id INTEGER;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS early_leave_fine_id INTEGER;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS training_fine_id INTEGER;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS technical_reason_fine_id INTEGER;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS auto_aggregated BOOLEAN NOT NULL DEFAULT FALSE;
            """)
            cursor.execute("""
                ALTER TABLE daily_hours
                ADD COLUMN IF NOT EXISTS auto_aggregated_at TIMESTAMP;
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_daily_hours_operator_day
                ON daily_hours(operator_id, day);
            """)

            # Table for multiple fines per day (new schema)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS daily_fines (
                    id SERIAL PRIMARY KEY,
                    daily_hours_id UUID REFERENCES daily_hours(id) ON DELETE CASCADE,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    day DATE NOT NULL,
                    amount FLOAT NOT NULL DEFAULT 0,
                    reason VARCHAR(64),
                    comment TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operator_technical_issues (
                    id SERIAL PRIMARY KEY,
                    batch_id UUID NOT NULL,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    issue_date DATE NOT NULL,
                    start_time TIME NOT NULL DEFAULT TIME '00:00:00',
                    end_time TIME NOT NULL DEFAULT TIME '23:59:59',
                    reason VARCHAR(255) NOT NULL,
                    comment TEXT,
                    direction_ids JSONB NOT NULL DEFAULT '[]'::jsonb,
                    created_by INTEGER NOT NULL REFERENCES users(id) ON DELETE RESTRICT,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor.execute("""
                ALTER TABLE operator_technical_issues
                ADD COLUMN IF NOT EXISTS start_time TIME NOT NULL DEFAULT TIME '00:00:00';
            """)
            cursor.execute("""
                ALTER TABLE operator_technical_issues
                ADD COLUMN IF NOT EXISTS end_time TIME NOT NULL DEFAULT TIME '23:59:59';
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operator_offline_activities (
                    id SERIAL PRIMARY KEY,
                    batch_id UUID NOT NULL,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    activity_date DATE NOT NULL,
                    start_time TIME NOT NULL DEFAULT TIME '00:00:00',
                    end_time TIME NOT NULL DEFAULT TIME '23:59:59',
                    comment TEXT,
                    created_by INTEGER NOT NULL REFERENCES users(id) ON DELETE RESTRICT,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor.execute("""
                ALTER TABLE operator_offline_activities
                ADD COLUMN IF NOT EXISTS start_time TIME NOT NULL DEFAULT TIME '00:00:00';
            """)
            cursor.execute("""
                ALTER TABLE operator_offline_activities
                ADD COLUMN IF NOT EXISTS end_time TIME NOT NULL DEFAULT TIME '23:59:59';
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS user_history (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    changed_by INTEGER REFERENCES users(id),  -- Who made the change (e.g., admin or sv ID)
                    field_changed VARCHAR(50) NOT NULL,
                    old_value TEXT,
                    new_value TEXT,
                    changed_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS user_sessions (
                    session_id UUID PRIMARY KEY,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    refresh_token_hash TEXT NOT NULL,
                    user_agent TEXT,
                    ip_address VARCHAR(64),
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    last_seen_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    expires_at TIMESTAMP NOT NULL,
                    revoked_at TIMESTAMP,
                    sensitive_data_unlocked BOOLEAN NOT NULL DEFAULT FALSE,
                    sensitive_data_unlocked_at TIMESTAMP
                );
            """)
            cursor.execute("""
                ALTER TABLE user_sessions
                ADD COLUMN IF NOT EXISTS sensitive_data_unlocked BOOLEAN NOT NULL DEFAULT FALSE;
            """)
            cursor.execute("""
                ALTER TABLE user_sessions
                ADD COLUMN IF NOT EXISTS sensitive_data_unlocked_at TIMESTAMP;
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_user_sessions_user_id
                ON user_sessions(user_id);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_user_sessions_expires_at
                ON user_sessions(expires_at);
            """)
            cursor.execute("""
                CREATE OR REPLACE FUNCTION revoke_sessions_on_inactive_status()
                RETURNS TRIGGER
                LANGUAGE plpgsql
                AS $$
                BEGIN
                    IF NEW.status IN ('fired', 'dismissal')
                       AND (OLD.status IS DISTINCT FROM NEW.status) THEN
                        UPDATE user_sessions
                        SET revoked_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                        WHERE user_id = NEW.id
                          AND revoked_at IS NULL;
                    END IF;

                    RETURN NEW;
                END;
                $$;
            """)
            cursor.execute("""
                DROP TRIGGER IF EXISTS trg_revoke_sessions_on_inactive_status ON users;
            """)
            cursor.execute("""
                CREATE TRIGGER trg_revoke_sessions_on_inactive_status
                AFTER UPDATE OF status ON users
                FOR EACH ROW
                EXECUTE FUNCTION revoke_sessions_on_inactive_status();
            """)
            # Work schedules (shifts) table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS work_shifts (
                    id SERIAL PRIMARY KEY,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    shift_date DATE NOT NULL,
                    start_time TIME NOT NULL,
                    end_time TIME NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(operator_id, shift_date, start_time, end_time)
                );
            """)

            # Break periods within shifts
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS shift_breaks (
                    id SERIAL PRIMARY KEY,
                    shift_id INTEGER NOT NULL REFERENCES work_shifts(id) ON DELETE CASCADE,
                    start_minutes INTEGER NOT NULL,
                    end_minutes INTEGER NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS work_schedule_break_rules (
                    id SERIAL PRIMARY KEY,
                    direction_name VARCHAR(255) NOT NULL,
                    min_shift_minutes INTEGER NOT NULL,
                    max_shift_minutes INTEGER NOT NULL,
                    break_durations_json JSONB NOT NULL DEFAULT '[]'::jsonb,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    CHECK (min_shift_minutes >= 0),
                    CHECK (max_shift_minutes > min_shift_minutes)
                );
            """)

            # Days off table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS days_off (
                    id SERIAL PRIMARY KEY,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    day_off_date DATE NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(operator_id, day_off_date)
                );
            """)
            # Shift swap requests between operators (operator -> operator)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS work_shift_swap_requests (
                    id SERIAL PRIMARY KEY,
                    requester_operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    target_operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    direction_id INTEGER NULL REFERENCES directions(id) ON DELETE SET NULL,
                    start_date DATE NOT NULL,
                    end_date DATE NOT NULL,
                    status VARCHAR(20) NOT NULL DEFAULT 'pending',
                    request_comment TEXT NULL,
                    response_comment TEXT NULL,
                    requested_shifts_json JSONB NOT NULL DEFAULT '{}'::jsonb,
                    responded_by INTEGER NULL REFERENCES users(id) ON DELETE SET NULL,
                    responded_at TIMESTAMP NULL,
                    accepted_at TIMESTAMP NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    CONSTRAINT work_shift_swap_requests_status_check CHECK (status IN ('pending', 'accepted', 'rejected', 'cancelled')),
                    CONSTRAINT work_shift_swap_requests_period_check CHECK (end_date >= start_date),
                    CONSTRAINT work_shift_swap_requests_participants_check CHECK (requester_operator_id <> target_operator_id)
                );
            """)
            cursor.execute("""
                ALTER TABLE work_shift_swap_requests
                ADD COLUMN IF NOT EXISTS direction_id INTEGER NULL REFERENCES directions(id) ON DELETE SET NULL;
            """)
            cursor.execute("""
                ALTER TABLE work_shift_swap_requests
                ADD COLUMN IF NOT EXISTS request_comment TEXT;
            """)
            cursor.execute("""
                ALTER TABLE work_shift_swap_requests
                ADD COLUMN IF NOT EXISTS response_comment TEXT;
            """)
            cursor.execute("""
                ALTER TABLE work_shift_swap_requests
                ADD COLUMN IF NOT EXISTS requested_shifts_json JSONB NOT NULL DEFAULT '{}'::jsonb;
            """)
            cursor.execute("""
                ALTER TABLE work_shift_swap_requests
                ADD COLUMN IF NOT EXISTS responded_by INTEGER NULL REFERENCES users(id) ON DELETE SET NULL;
            """)
            cursor.execute("""
                ALTER TABLE work_shift_swap_requests
                ADD COLUMN IF NOT EXISTS responded_at TIMESTAMP;
            """)
            cursor.execute("""
                ALTER TABLE work_shift_swap_requests
                ADD COLUMN IF NOT EXISTS accepted_at TIMESTAMP;
            """)
            cursor.execute("""
                ALTER TABLE work_shift_swap_requests
                ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;
            """)
            cursor.execute("""
                ALTER TABLE users
                DROP CONSTRAINT IF EXISTS users_status_check;
            """)
            cursor.execute("""
                ALTER TABLE users
                ADD CONSTRAINT users_status_check
                CHECK (status IN ('working', 'fired', 'unpaid_leave', 'bs', 'sick_leave', 'annual_leave', 'dismissal'));
            """)
            # Special statuses by period (vacation/sick leave/unpaid leave/dismissal)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operator_schedule_status_periods (
                    id SERIAL PRIMARY KEY,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    status_code VARCHAR(32) NOT NULL,
                    start_date DATE NOT NULL,
                    end_date DATE NULL,
                    dismissal_reason TEXT NULL,
                    is_blacklist BOOLEAN NOT NULL DEFAULT FALSE,
                    comment TEXT NULL,
                    created_by INTEGER NULL REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor.execute("""
                ALTER TABLE operator_schedule_status_periods
                ADD COLUMN IF NOT EXISTS is_blacklist BOOLEAN;
            """)
            cursor.execute("""
                UPDATE operator_schedule_status_periods
                SET is_blacklist = FALSE
                WHERE is_blacklist IS NULL;
            """)
            cursor.execute("""
                ALTER TABLE operator_schedule_status_periods
                ALTER COLUMN is_blacklist SET DEFAULT FALSE;
            """)
            cursor.execute("""
                ALTER TABLE operator_schedule_status_periods
                ALTER COLUMN is_blacklist SET NOT NULL;
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_shift_breaks_shift_id
                ON shift_breaks(shift_id);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_work_schedule_break_rules_direction
                ON work_schedule_break_rules(LOWER(direction_name));
            """)
            cursor.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS uq_work_schedule_break_rules_direction_range
                ON work_schedule_break_rules(LOWER(direction_name), min_shift_minutes, max_shift_minutes);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_op_sched_status_periods_operator_start
                ON operator_schedule_status_periods(operator_id, start_date);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_op_sched_status_periods_operator_end
                ON operator_schedule_status_periods(operator_id, end_date);
            """)
            # Status timeline imports (CSV transitions -> events/segments for analytics and KPI calculations)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operator_status_import_batches (
                    id UUID PRIMARY KEY,
                    imported_by INTEGER NULL REFERENCES users(id) ON DELETE SET NULL,
                    source_rows INTEGER NOT NULL DEFAULT 0,
                    valid_events INTEGER NOT NULL DEFAULT 0,
                    matched_events INTEGER NOT NULL DEFAULT 0,
                    segments_saved INTEGER NOT NULL DEFAULT 0,
                    deleted_events INTEGER NOT NULL DEFAULT 0,
                    deleted_segments INTEGER NOT NULL DEFAULT 0,
                    invalid_rows_count INTEGER NOT NULL DEFAULT 0,
                    parse_errors_count INTEGER NOT NULL DEFAULT 0,
                    operators_count INTEGER NOT NULL DEFAULT 0,
                    open_tail_events INTEGER NOT NULL DEFAULT 0,
                    zero_or_negative_transitions INTEGER NOT NULL DEFAULT 0,
                    date_from DATE NULL,
                    date_to DATE NULL,
                    meta_json JSONB NOT NULL DEFAULT '{}'::jsonb
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operator_status_events (
                    id BIGSERIAL PRIMARY KEY,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    event_at TIMESTAMP NOT NULL,
                    event_date DATE NOT NULL,
                    status_key VARCHAR(128) NOT NULL,
                    state_note VARCHAR(255) NULL,
                    imported_by INTEGER NULL REFERENCES users(id) ON DELETE SET NULL
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operator_status_segments (
                    id BIGSERIAL PRIMARY KEY,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    status_date DATE NOT NULL,
                    start_at TIMESTAMP NOT NULL,
                    end_at TIMESTAMP NOT NULL,
                    duration_sec INTEGER NOT NULL,
                    status_key VARCHAR(128) NOT NULL,
                    state_note VARCHAR(255) NULL,
                    imported_by INTEGER NULL REFERENCES users(id) ON DELETE SET NULL,
                    CHECK (end_at > start_at),
                    CHECK (duration_sec > 0)
                );
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_operator_status_events_operator_time
                ON operator_status_events(operator_id, event_at);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_operator_status_events_operator_date
                ON operator_status_events(operator_id, event_date);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_operator_status_events_imported_by
                ON operator_status_events(imported_by);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_operator_status_segments_operator_date
                ON operator_status_segments(operator_id, status_date);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_operator_status_segments_operator_start
                ON operator_status_segments(operator_id, start_at);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_operator_status_segments_status_key
                ON operator_status_segments(status_key);
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_operator_status_segments_imported_by
                ON operator_status_segments(imported_by);
            """)

            # AI feedback cache table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS ai_feedback_cache (
                    id SERIAL PRIMARY KEY,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    month VARCHAR(7) NOT NULL,
                    feedback_data JSONB NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(operator_id, month)
                );
            """)

            # AI birthday greeting cache table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS ai_birthday_greeting_cache (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    greeting_date DATE NOT NULL,
                    greeting_data JSONB NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(user_id, greeting_date)
                );
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_ai_birthday_greeting_cache_user_date
                ON ai_birthday_greeting_cache(user_id, greeting_date);
            """)

            # Recruiting parser runs + snapshots
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS recruiting_parse_runs (
                    id SERIAL PRIMARY KEY,
                    run_uuid UUID UNIQUE NOT NULL,
                    source VARCHAR(64) NOT NULL DEFAULT 'enbek',
                    triggered_by VARCHAR(32) NOT NULL DEFAULT 'scheduler',
                    status VARCHAR(16) NOT NULL CHECK (status IN ('success', 'failed')),
                    total_items INTEGER NOT NULL DEFAULT 0,
                    error_message TEXT,
                    started_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    finished_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS recruiting_resumes (
                    id SERIAL PRIMARY KEY,
                    run_uuid UUID NOT NULL REFERENCES recruiting_parse_runs(run_uuid) ON DELETE CASCADE,
                    keyword_group VARCHAR(128),
                    keyword_query TEXT,
                    page_found INTEGER,
                    title TEXT,
                    category TEXT,
                    experience TEXT,
                    location TEXT,
                    salary TEXT,
                    education TEXT,
                    published_at TEXT,
                    detail_url TEXT,
                    scraped_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );
            """)
            cursor.execute("""
                DELETE FROM recruiting_resumes newer
                USING recruiting_resumes older
                WHERE newer.id > older.id
                  AND newer.run_uuid = older.run_uuid
                  AND newer.detail_url IS NOT NULL
                  AND newer.detail_url = older.detail_url;
            """)
            cursor.execute("""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1
                        FROM pg_constraint
                        WHERE conname = 'uq_recruiting_resumes_run_detail_url'
                          AND conrelid = 'recruiting_resumes'::regclass
                    ) THEN
                        ALTER TABLE recruiting_resumes
                            ADD CONSTRAINT uq_recruiting_resumes_run_detail_url
                            UNIQUE (run_uuid, detail_url);
                    END IF;
                END $$;
            """)

            # Surveys tables
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS surveys (
                    id SERIAL PRIMARY KEY,
                    title VARCHAR(255) NOT NULL,
                    description TEXT,
                    created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    direction_ids JSONB NOT NULL DEFAULT '[]'::jsonb,
                    tenure_weeks_min INTEGER,
                    tenure_weeks_max INTEGER,
                    repeat_root_id INTEGER REFERENCES surveys(id) ON DELETE SET NULL,
                    repeat_iteration INTEGER NOT NULL DEFAULT 1,
                    is_test BOOLEAN NOT NULL DEFAULT FALSE,
                    is_active BOOLEAN NOT NULL DEFAULT TRUE,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    updated_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    CHECK (tenure_weeks_min IS NULL OR tenure_weeks_min >= 0),
                    CHECK (tenure_weeks_max IS NULL OR tenure_weeks_max >= 0),
                    CHECK (tenure_weeks_min IS NULL OR tenure_weeks_max IS NULL OR tenure_weeks_min <= tenure_weeks_max),
                    CHECK (repeat_iteration >= 1)
                );
            """)
            cursor.execute("""
                ALTER TABLE surveys
                ADD COLUMN IF NOT EXISTS repeat_root_id INTEGER REFERENCES surveys(id) ON DELETE SET NULL;
            """)
            cursor.execute("""
                ALTER TABLE surveys
                ADD COLUMN IF NOT EXISTS repeat_iteration INTEGER;
            """)
            cursor.execute("""
                UPDATE surveys
                SET repeat_iteration = 1
                WHERE repeat_iteration IS NULL OR repeat_iteration < 1;
            """)
            cursor.execute("""
                ALTER TABLE surveys
                ALTER COLUMN repeat_iteration SET DEFAULT 1;
            """)
            cursor.execute("""
                ALTER TABLE surveys
                ALTER COLUMN repeat_iteration SET NOT NULL;
            """)
            cursor.execute("""
                UPDATE surveys
                SET repeat_root_id = id
                WHERE repeat_root_id IS NULL;
            """)
            cursor.execute("""
                ALTER TABLE surveys
                ADD COLUMN IF NOT EXISTS is_test BOOLEAN NOT NULL DEFAULT FALSE;
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS survey_questions (
                    id SERIAL PRIMARY KEY,
                    survey_id INTEGER NOT NULL REFERENCES surveys(id) ON DELETE CASCADE,
                    position INTEGER NOT NULL,
                    question_text TEXT NOT NULL,
                    question_type VARCHAR(16) NOT NULL CHECK (question_type IN ('single', 'multiple', 'rating')),
                    is_required BOOLEAN NOT NULL DEFAULT TRUE,
                    allow_other BOOLEAN NOT NULL DEFAULT FALSE,
                    options_json JSONB NOT NULL DEFAULT '[]'::jsonb,
                    correct_options_json JSONB NOT NULL DEFAULT '[]'::jsonb,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    UNIQUE (survey_id, position)
                );
            """)
            cursor.execute("""
                ALTER TABLE survey_questions
                ADD COLUMN IF NOT EXISTS correct_options_json JSONB;
            """)
            cursor.execute("""
                UPDATE survey_questions
                SET correct_options_json = '[]'::jsonb
                WHERE correct_options_json IS NULL;
            """)
            cursor.execute("""
                ALTER TABLE survey_questions
                ALTER COLUMN correct_options_json SET DEFAULT '[]'::jsonb;
            """)
            cursor.execute("""
                ALTER TABLE survey_questions
                ALTER COLUMN correct_options_json SET NOT NULL;
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS survey_assignments (
                    id SERIAL PRIMARY KEY,
                    survey_id INTEGER NOT NULL REFERENCES surveys(id) ON DELETE CASCADE,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    assigned_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    status VARCHAR(16) NOT NULL DEFAULT 'assigned' CHECK (status IN ('assigned', 'in_progress', 'completed')),
                    assigned_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    completed_at TIMESTAMP,
                    UNIQUE (survey_id, operator_id)
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS survey_responses (
                    id SERIAL PRIMARY KEY,
                    survey_id INTEGER NOT NULL REFERENCES surveys(id) ON DELETE CASCADE,
                    operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    assignment_id INTEGER REFERENCES survey_assignments(id) ON DELETE SET NULL,
                    submitted_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    UNIQUE (survey_id, operator_id)
                );
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS survey_answers (
                    id SERIAL PRIMARY KEY,
                    response_id INTEGER NOT NULL REFERENCES survey_responses(id) ON DELETE CASCADE,
                    question_id INTEGER NOT NULL REFERENCES survey_questions(id) ON DELETE CASCADE,
                    answer_text TEXT,
                    selected_options_json JSONB NOT NULL DEFAULT '[]'::jsonb,
                    rating_value INTEGER,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    CHECK (rating_value IS NULL OR (rating_value >= 1 AND rating_value <= 5))
                );
            """)

            # Tasks table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS tasks (
                    id SERIAL PRIMARY KEY,
                    subject VARCHAR(255) NOT NULL,
                    description TEXT,
                    tag VARCHAR(32) NOT NULL DEFAULT 'task' CHECK (tag IN ('task', 'problem', 'suggestion')),
                    status VARCHAR(32) NOT NULL DEFAULT 'assigned' CHECK (status IN ('assigned', 'in_progress', 'completed', 'accepted', 'returned')),
                    assigned_to INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    completion_summary TEXT,
                    completed_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    completed_at TIMESTAMP,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    updated_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );
            """)
            cursor.execute("""
                ALTER TABLE tasks
                ADD COLUMN IF NOT EXISTS completion_summary TEXT;
            """)
            cursor.execute("""
                ALTER TABLE tasks
                ADD COLUMN IF NOT EXISTS completed_by INTEGER REFERENCES users(id) ON DELETE SET NULL;
            """)
            cursor.execute("""
                ALTER TABLE tasks
                ADD COLUMN IF NOT EXISTS completed_at TIMESTAMP;
            """)

            # Tasks status history
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS task_status_history (
                    id SERIAL PRIMARY KEY,
                    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
                    status_code VARCHAR(32) NOT NULL CHECK (status_code IN ('assigned', 'in_progress', 'completed', 'accepted', 'returned', 'reopened')),
                    changed_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    comment TEXT,
                    changed_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );
            """)

            # Task attachments (GCS storage; DB `file_data` kept for legacy reads)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS task_attachments (
                    id SERIAL PRIMARY KEY,
                    task_id INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
                    file_name TEXT NOT NULL,
                    content_type VARCHAR(255),
                    file_size INTEGER NOT NULL DEFAULT 0,
                    file_data BYTEA,
                    storage_type VARCHAR(16) NOT NULL DEFAULT 'gcs' CHECK (storage_type IN ('db', 'gcs')),
                    gcs_bucket VARCHAR(255),
                    gcs_blob_path TEXT,
                    attachment_kind VARCHAR(16) NOT NULL DEFAULT 'initial' CHECK (attachment_kind IN ('initial', 'result')),
                    uploaded_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                ADD COLUMN IF NOT EXISTS storage_type VARCHAR(16) NOT NULL DEFAULT 'gcs';
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                ALTER COLUMN storage_type SET DEFAULT 'gcs';
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                ADD COLUMN IF NOT EXISTS gcs_bucket VARCHAR(255);
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                ADD COLUMN IF NOT EXISTS gcs_blob_path TEXT;
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                ADD COLUMN IF NOT EXISTS attachment_kind VARCHAR(16) NOT NULL DEFAULT 'initial';
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                DROP CONSTRAINT IF EXISTS task_attachments_storage_type_check;
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                ADD CONSTRAINT task_attachments_storage_type_check
                CHECK (storage_type IN ('db', 'gcs'));
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                DROP CONSTRAINT IF EXISTS task_attachments_attachment_kind_check;
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                ADD CONSTRAINT task_attachments_attachment_kind_check
                CHECK (attachment_kind IN ('initial', 'result'));
            """)
            cursor.execute("""
                ALTER TABLE task_attachments
                ALTER COLUMN file_data DROP NOT NULL;
            """)

            # Optimized Indexes (added more based on query patterns)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_calls_month ON calls(month);
                CREATE INDEX IF NOT EXISTS idx_calls_operator_id ON calls(operator_id);
                CREATE INDEX IF NOT EXISTS idx_calls_operator_month ON calls(operator_id, month);
                CREATE INDEX IF NOT EXISTS idx_calls_op_month_phone_created ON calls(operator_id, month, phone_number, created_at DESC);
                CREATE INDEX IF NOT EXISTS idx_calls_phone_number ON calls(phone_number);
                CREATE INDEX IF NOT EXISTS idx_calls_created_at ON calls(created_at);
                CREATE INDEX IF NOT EXISTS idx_calls_evaluator_id ON calls(evaluator_id);
                CREATE INDEX IF NOT EXISTS idx_calls_is_draft ON calls(is_draft);
                CREATE INDEX IF NOT EXISTS idx_work_hours_operator_id ON work_hours(operator_id);
                CREATE INDEX IF NOT EXISTS idx_work_hours_month ON work_hours(month);
                CREATE INDEX IF NOT EXISTS idx_directions_name ON directions(name);
                CREATE INDEX IF NOT EXISTS idx_directions_is_active ON directions(is_active);
                CREATE INDEX IF NOT EXISTS idx_users_role ON users(role);
                CREATE INDEX IF NOT EXISTS idx_users_supervisor_id ON users(supervisor_id);
                CREATE INDEX IF NOT EXISTS idx_users_login ON users(login);
                CREATE INDEX IF NOT EXISTS idx_users_is_active ON users(is_active);
                CREATE INDEX IF NOT EXISTS idx_users_status ON users(status);
                CREATE INDEX IF NOT EXISTS idx_users_rate ON users(rate);
                CREATE INDEX IF NOT EXISTS idx_user_history_user_id ON user_history(user_id);
                CREATE INDEX IF NOT EXISTS idx_user_history_changed_at ON user_history(changed_at);
                CREATE INDEX IF NOT EXISTS idx_work_hours_calls_per_hour ON work_hours(calls_per_hour);
                CREATE INDEX IF NOT EXISTS idx_trainings_operator_id ON trainings(operator_id);
                CREATE INDEX IF NOT EXISTS idx_trainings_training_date ON trainings(training_date);
                CREATE INDEX IF NOT EXISTS idx_operator_technical_issues_issue_date
                ON operator_technical_issues(issue_date);
                CREATE INDEX IF NOT EXISTS idx_operator_technical_issues_issue_date_time
                ON operator_technical_issues(issue_date, start_time, end_time);
                CREATE INDEX IF NOT EXISTS idx_operator_technical_issues_operator_id
                ON operator_technical_issues(operator_id);
                CREATE INDEX IF NOT EXISTS idx_operator_technical_issues_operator_date
                ON operator_technical_issues(operator_id, issue_date);
                CREATE INDEX IF NOT EXISTS idx_operator_technical_issues_reason
                ON operator_technical_issues(reason);
                CREATE INDEX IF NOT EXISTS idx_operator_technical_issues_created_by
                ON operator_technical_issues(created_by);
                CREATE INDEX IF NOT EXISTS idx_operator_technical_issues_created_at
                ON operator_technical_issues(created_at);
                CREATE INDEX IF NOT EXISTS idx_operator_offline_activities_activity_date
                ON operator_offline_activities(activity_date);
                CREATE INDEX IF NOT EXISTS idx_operator_offline_activities_activity_date_time
                ON operator_offline_activities(activity_date, start_time, end_time);
                CREATE INDEX IF NOT EXISTS idx_operator_offline_activities_operator_id
                ON operator_offline_activities(operator_id);
                CREATE INDEX IF NOT EXISTS idx_operator_offline_activities_operator_date
                ON operator_offline_activities(operator_id, activity_date);
                CREATE INDEX IF NOT EXISTS idx_operator_offline_activities_created_by
                ON operator_offline_activities(created_by);
                CREATE INDEX IF NOT EXISTS idx_operator_offline_activities_created_at
                ON operator_offline_activities(created_at);
                CREATE INDEX IF NOT EXISTS idx_work_shifts_operator_date ON work_shifts(operator_id, shift_date);
                CREATE INDEX IF NOT EXISTS idx_work_shifts_date ON work_shifts(shift_date);
                CREATE INDEX IF NOT EXISTS idx_days_off_operator_date ON days_off(operator_id, day_off_date);
                CREATE INDEX IF NOT EXISTS idx_work_shift_swap_requests_target_status
                ON work_shift_swap_requests(target_operator_id, status, created_at DESC);
                CREATE INDEX IF NOT EXISTS idx_work_shift_swap_requests_requester_status
                ON work_shift_swap_requests(requester_operator_id, status, created_at DESC);
                CREATE INDEX IF NOT EXISTS idx_work_shift_swap_requests_period
                ON work_shift_swap_requests(start_date, end_date);
                CREATE INDEX IF NOT EXISTS idx_ai_feedback_cache_operator_month ON ai_feedback_cache(operator_id, month);
                CREATE INDEX IF NOT EXISTS idx_recruiting_parse_runs_status_finished
                ON recruiting_parse_runs(status, finished_at DESC);
                CREATE INDEX IF NOT EXISTS idx_recruiting_resumes_run_uuid ON recruiting_resumes(run_uuid);
                CREATE INDEX IF NOT EXISTS idx_recruiting_resumes_group ON recruiting_resumes(keyword_group);
                CREATE INDEX IF NOT EXISTS idx_recruiting_resumes_page ON recruiting_resumes(page_found);
                CREATE INDEX IF NOT EXISTS idx_recruiting_resumes_detail_url ON recruiting_resumes(detail_url);
                CREATE INDEX IF NOT EXISTS idx_surveys_created_by ON surveys(created_by, created_at DESC);
                CREATE INDEX IF NOT EXISTS idx_surveys_repeat_root_iteration ON surveys(repeat_root_id, repeat_iteration);
                CREATE INDEX IF NOT EXISTS idx_survey_questions_survey_position ON survey_questions(survey_id, position);
                CREATE INDEX IF NOT EXISTS idx_survey_assignments_operator ON survey_assignments(operator_id, status);
                CREATE INDEX IF NOT EXISTS idx_survey_assignments_survey ON survey_assignments(survey_id);
                CREATE INDEX IF NOT EXISTS idx_survey_responses_survey ON survey_responses(survey_id);
                CREATE INDEX IF NOT EXISTS idx_survey_answers_response ON survey_answers(response_id);
                CREATE INDEX IF NOT EXISTS idx_tasks_assigned_to ON tasks(assigned_to);
                CREATE INDEX IF NOT EXISTS idx_tasks_created_by ON tasks(created_by);
                CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status);
                CREATE INDEX IF NOT EXISTS idx_task_status_history_task_id ON task_status_history(task_id);
                CREATE INDEX IF NOT EXISTS idx_task_attachments_task_id ON task_attachments(task_id);
                CREATE INDEX IF NOT EXISTS idx_calibration_rooms_month ON calibration_rooms(month);
                CREATE INDEX IF NOT EXISTS idx_calibration_rooms_operator_id ON calibration_rooms(operator_id);
                CREATE INDEX IF NOT EXISTS idx_calibration_rooms_admin_id ON calibration_rooms(created_by_admin_id);
                CREATE INDEX IF NOT EXISTS idx_calibration_rooms_created_at ON calibration_rooms(created_at);
                CREATE INDEX IF NOT EXISTS idx_calibration_members_room_sv ON calibration_room_members(room_id, supervisor_id);
                CREATE INDEX IF NOT EXISTS idx_calibration_room_calls_room_id ON calibration_room_calls(room_id);
                CREATE INDEX IF NOT EXISTS idx_calibration_room_calls_operator_id ON calibration_room_calls(operator_id);
                CREATE INDEX IF NOT EXISTS idx_calibration_room_calls_month ON calibration_room_calls(month);
                CREATE INDEX IF NOT EXISTS idx_calibration_call_evals_room_call_id ON calibration_room_call_evaluations(room_call_id);
                CREATE INDEX IF NOT EXISTS idx_calibration_call_evals_evaluator_id ON calibration_room_call_evaluations(evaluator_id);
                CREATE INDEX IF NOT EXISTS idx_calibration_eval_room_id ON calibration_room_evaluations(room_id);
                CREATE INDEX IF NOT EXISTS idx_calibration_eval_evaluator_id ON calibration_room_evaluations(evaluator_id);
                
                    -- Table for Baiga daily scores (points per operator per day)
                    CREATE TABLE IF NOT EXISTS baiga_daily_scores (
                        id SERIAL PRIMARY KEY,
                        operator_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                        day DATE NOT NULL,
                        points INTEGER NOT NULL DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE(operator_id, day)
                    );

                CREATE INDEX IF NOT EXISTS idx_baiga_daily_scores_day ON baiga_daily_scores(day);
                CREATE INDEX IF NOT EXISTS idx_baiga_daily_scores_operator ON baiga_daily_scores(operator_id);

                -- LMS core entities
                CREATE TABLE IF NOT EXISTS lms_courses (
                    id SERIAL PRIMARY KEY,
                    slug VARCHAR(255) UNIQUE,
                    title VARCHAR(255) NOT NULL,
                    description TEXT,
                    category VARCHAR(100),
                    status VARCHAR(20) NOT NULL DEFAULT 'draft' CHECK(status IN ('draft', 'published', 'archived')),
                    default_pass_threshold NUMERIC(5,2) NOT NULL DEFAULT 80.00 CHECK(default_pass_threshold >= 0 AND default_pass_threshold <= 100),
                    default_attempt_limit INTEGER NOT NULL DEFAULT 3 CHECK(default_attempt_limit >= 1),
                    current_version_id INTEGER,
                    created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    updated_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );

                CREATE TABLE IF NOT EXISTS lms_course_versions (
                    id SERIAL PRIMARY KEY,
                    course_id INTEGER NOT NULL REFERENCES lms_courses(id) ON DELETE CASCADE,
                    version_number INTEGER NOT NULL CHECK(version_number >= 1),
                    title VARCHAR(255),
                    description TEXT,
                    status VARCHAR(20) NOT NULL DEFAULT 'draft' CHECK(status IN ('draft', 'published', 'archived')),
                    pass_threshold NUMERIC(5,2) NOT NULL DEFAULT 80.00 CHECK(pass_threshold >= 0 AND pass_threshold <= 100),
                    attempt_limit INTEGER NOT NULL DEFAULT 3 CHECK(attempt_limit >= 1),
                    anti_cheat_settings JSONB NOT NULL DEFAULT '{}'::jsonb,
                    created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    published_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    published_at TIMESTAMP,
                    UNIQUE(course_id, version_number)
                );

                ALTER TABLE lms_courses
                ADD COLUMN IF NOT EXISTS current_version_id INTEGER;

                DO $$
                BEGIN
                    ALTER TABLE lms_courses
                    ADD CONSTRAINT lms_courses_current_version_fk
                    FOREIGN KEY (current_version_id)
                    REFERENCES lms_course_versions(id)
                    ON DELETE SET NULL;
                EXCEPTION
                    WHEN duplicate_object THEN
                        NULL;
                END $$;

                CREATE TABLE IF NOT EXISTS lms_modules (
                    id SERIAL PRIMARY KEY,
                    course_version_id INTEGER NOT NULL REFERENCES lms_course_versions(id) ON DELETE CASCADE,
                    title VARCHAR(255) NOT NULL,
                    description TEXT,
                    position INTEGER NOT NULL DEFAULT 1 CHECK(position >= 1),
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );

                CREATE TABLE IF NOT EXISTS lms_lessons (
                    id SERIAL PRIMARY KEY,
                    module_id INTEGER NOT NULL REFERENCES lms_modules(id) ON DELETE CASCADE,
                    title VARCHAR(255) NOT NULL,
                    description TEXT,
                    position INTEGER NOT NULL DEFAULT 1 CHECK(position >= 1),
                    duration_seconds INTEGER NOT NULL DEFAULT 0 CHECK(duration_seconds >= 0),
                    allow_fast_forward BOOLEAN NOT NULL DEFAULT FALSE,
                    completion_threshold NUMERIC(5,2) NOT NULL DEFAULT 95.00 CHECK(completion_threshold >= 0 AND completion_threshold <= 100),
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );

                CREATE TABLE IF NOT EXISTS lms_lesson_materials (
                    id SERIAL PRIMARY KEY,
                    lesson_id INTEGER NOT NULL REFERENCES lms_lessons(id) ON DELETE CASCADE,
                    title VARCHAR(255) NOT NULL,
                    material_type VARCHAR(20) NOT NULL CHECK(material_type IN ('video', 'pdf', 'link', 'text', 'file')),
                    content_text TEXT,
                    content_url TEXT,
                    gcs_bucket VARCHAR(255),
                    gcs_blob_path TEXT,
                    mime_type VARCHAR(255),
                    metadata JSONB NOT NULL DEFAULT '{}'::jsonb,
                    position INTEGER NOT NULL DEFAULT 1 CHECK(position >= 1),
                    created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );

                CREATE TABLE IF NOT EXISTS lms_tests (
                    id SERIAL PRIMARY KEY,
                    course_version_id INTEGER NOT NULL REFERENCES lms_course_versions(id) ON DELETE CASCADE,
                    module_id INTEGER REFERENCES lms_modules(id) ON DELETE SET NULL,
                    title VARCHAR(255) NOT NULL,
                    description TEXT,
                    pass_threshold NUMERIC(5,2) NOT NULL DEFAULT 80.00 CHECK(pass_threshold >= 0 AND pass_threshold <= 100),
                    attempt_limit INTEGER NOT NULL DEFAULT 3 CHECK(attempt_limit >= 1),
                    time_limit_minutes INTEGER CHECK(time_limit_minutes IS NULL OR time_limit_minutes > 0),
                    is_final BOOLEAN NOT NULL DEFAULT TRUE,
                    status VARCHAR(20) NOT NULL DEFAULT 'draft' CHECK(status IN ('draft', 'published', 'archived')),
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );

                CREATE TABLE IF NOT EXISTS lms_questions (
                    id SERIAL PRIMARY KEY,
                    test_id INTEGER NOT NULL REFERENCES lms_tests(id) ON DELETE CASCADE,
                    question_type VARCHAR(20) NOT NULL CHECK(question_type IN ('single', 'multiple', 'true_false', 'matching', 'text')),
                    prompt TEXT NOT NULL,
                    points NUMERIC(8,2) NOT NULL DEFAULT 1 CHECK(points > 0),
                    position INTEGER NOT NULL DEFAULT 1 CHECK(position >= 1),
                    required BOOLEAN NOT NULL DEFAULT TRUE,
                    metadata JSONB NOT NULL DEFAULT '{}'::jsonb,
                    correct_text_answers JSONB NOT NULL DEFAULT '[]'::jsonb,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );

                CREATE TABLE IF NOT EXISTS lms_question_options (
                    id SERIAL PRIMARY KEY,
                    question_id INTEGER NOT NULL REFERENCES lms_questions(id) ON DELETE CASCADE,
                    option_key VARCHAR(255),
                    option_text TEXT NOT NULL,
                    position INTEGER NOT NULL DEFAULT 1 CHECK(position >= 1),
                    is_correct BOOLEAN NOT NULL DEFAULT FALSE,
                    match_key VARCHAR(255),
                    metadata JSONB NOT NULL DEFAULT '{}'::jsonb
                );

                CREATE TABLE IF NOT EXISTS lms_course_assignments (
                    id SERIAL PRIMARY KEY,
                    course_id INTEGER NOT NULL REFERENCES lms_courses(id) ON DELETE CASCADE,
                    course_version_id INTEGER REFERENCES lms_course_versions(id) ON DELETE SET NULL,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    assigned_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    due_at TIMESTAMP,
                    started_at TIMESTAMP,
                    completed_at TIMESTAMP,
                    status VARCHAR(20) NOT NULL DEFAULT 'assigned' CHECK(status IN ('assigned', 'in_progress', 'completed')),
                    completion_color_status VARCHAR(20) CHECK(completion_color_status IN ('green', 'orange', 'red')),
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    updated_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    UNIQUE(course_id, user_id)
                );

                CREATE TABLE IF NOT EXISTS lms_lesson_progress (
                    id SERIAL PRIMARY KEY,
                    assignment_id INTEGER NOT NULL REFERENCES lms_course_assignments(id) ON DELETE CASCADE,
                    lesson_id INTEGER NOT NULL REFERENCES lms_lessons(id) ON DELETE CASCADE,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    status VARCHAR(20) NOT NULL DEFAULT 'not_started' CHECK(status IN ('not_started', 'in_progress', 'completed')),
                    max_position_seconds NUMERIC(10,2) NOT NULL DEFAULT 0,
                    confirmed_seconds NUMERIC(10,2) NOT NULL DEFAULT 0,
                    completion_ratio NUMERIC(5,2) NOT NULL DEFAULT 0,
                    active_seconds INTEGER NOT NULL DEFAULT 0,
                    last_heartbeat_at TIMESTAMP,
                    last_event_at TIMESTAMP,
                    tab_hidden_count INTEGER NOT NULL DEFAULT 0,
                    stale_gap_count INTEGER NOT NULL DEFAULT 0,
                    anti_cheat_flags JSONB NOT NULL DEFAULT '[]'::jsonb,
                    started_at TIMESTAMP,
                    completed_at TIMESTAMP,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    updated_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    UNIQUE(assignment_id, lesson_id, user_id)
                );

                CREATE TABLE IF NOT EXISTS lms_learning_sessions (
                    id SERIAL PRIMARY KEY,
                    assignment_id INTEGER NOT NULL REFERENCES lms_course_assignments(id) ON DELETE CASCADE,
                    lesson_id INTEGER NOT NULL REFERENCES lms_lessons(id) ON DELETE CASCADE,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    started_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    ended_at TIMESTAMP,
                    last_heartbeat_at TIMESTAMP,
                    last_visible_at TIMESTAMP,
                    is_active BOOLEAN NOT NULL DEFAULT TRUE,
                    max_position_seconds NUMERIC(10,2) NOT NULL DEFAULT 0,
                    confirmed_seconds NUMERIC(10,2) NOT NULL DEFAULT 0,
                    active_seconds INTEGER NOT NULL DEFAULT 0,
                    tab_hidden_count INTEGER NOT NULL DEFAULT 0,
                    stale_gap_count INTEGER NOT NULL DEFAULT 0,
                    client_fingerprint VARCHAR(255)
                );

                CREATE TABLE IF NOT EXISTS lms_learning_events (
                    id SERIAL PRIMARY KEY,
                    session_id INTEGER REFERENCES lms_learning_sessions(id) ON DELETE CASCADE,
                    lesson_id INTEGER NOT NULL REFERENCES lms_lessons(id) ON DELETE CASCADE,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    event_type VARCHAR(50) NOT NULL,
                    payload JSONB NOT NULL DEFAULT '{}'::jsonb,
                    client_ts TIMESTAMP,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );

                CREATE TABLE IF NOT EXISTS lms_test_attempts (
                    id SERIAL PRIMARY KEY,
                    assignment_id INTEGER NOT NULL REFERENCES lms_course_assignments(id) ON DELETE CASCADE,
                    test_id INTEGER NOT NULL REFERENCES lms_tests(id) ON DELETE CASCADE,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    attempt_no INTEGER NOT NULL CHECK(attempt_no >= 1),
                    status VARCHAR(20) NOT NULL DEFAULT 'in_progress' CHECK(status IN ('in_progress', 'finished', 'expired')),
                    score_percent NUMERIC(6,2),
                    passed BOOLEAN,
                    started_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    finished_at TIMESTAMP,
                    duration_seconds INTEGER,
                    proctor_flags JSONB NOT NULL DEFAULT '[]'::jsonb,
                    UNIQUE(assignment_id, test_id, attempt_no)
                );

                CREATE TABLE IF NOT EXISTS lms_test_attempt_answers (
                    id SERIAL PRIMARY KEY,
                    attempt_id INTEGER NOT NULL REFERENCES lms_test_attempts(id) ON DELETE CASCADE,
                    question_id INTEGER NOT NULL REFERENCES lms_questions(id) ON DELETE CASCADE,
                    answer_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
                    is_correct BOOLEAN,
                    points_awarded NUMERIC(8,2),
                    answered_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    UNIQUE(attempt_id, question_id)
                );

                CREATE TABLE IF NOT EXISTS lms_certificates (
                    id SERIAL PRIMARY KEY,
                    assignment_id INTEGER REFERENCES lms_course_assignments(id) ON DELETE SET NULL,
                    course_id INTEGER NOT NULL REFERENCES lms_courses(id) ON DELETE CASCADE,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    test_attempt_id INTEGER REFERENCES lms_test_attempts(id) ON DELETE SET NULL,
                    certificate_number VARCHAR(64) NOT NULL UNIQUE,
                    verify_token VARCHAR(128) NOT NULL UNIQUE,
                    score_percent NUMERIC(6,2),
                    status VARCHAR(20) NOT NULL DEFAULT 'active' CHECK(status IN ('active', 'revoked')),
                    issued_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    revoked_at TIMESTAMP,
                    revoked_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    revoke_reason TEXT,
                    pdf_storage_type VARCHAR(16) NOT NULL DEFAULT 'db' CHECK(pdf_storage_type IN ('db', 'gcs')),
                    pdf_data BYTEA,
                    gcs_bucket VARCHAR(255),
                    gcs_blob_path TEXT,
                    metadata JSONB NOT NULL DEFAULT '{}'::jsonb
                );

                CREATE TABLE IF NOT EXISTS lms_notifications (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    notification_type VARCHAR(50) NOT NULL,
                    title VARCHAR(255) NOT NULL,
                    message TEXT,
                    payload JSONB NOT NULL DEFAULT '{}'::jsonb,
                    is_read BOOLEAN NOT NULL DEFAULT FALSE,
                    read_at TIMESTAMP,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );

                CREATE TABLE IF NOT EXISTS lms_admin_audit_log (
                    id SERIAL PRIMARY KEY,
                    actor_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    actor_role VARCHAR(20),
                    action VARCHAR(100) NOT NULL,
                    entity_type VARCHAR(50) NOT NULL,
                    entity_id INTEGER,
                    details JSONB NOT NULL DEFAULT '{}'::jsonb,
                    created_at TIMESTAMP NOT NULL DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                );

                -- LMS indexes
                CREATE INDEX IF NOT EXISTS idx_lms_courses_status ON lms_courses(status);
                CREATE INDEX IF NOT EXISTS idx_lms_course_versions_course_status ON lms_course_versions(course_id, status);
                CREATE INDEX IF NOT EXISTS idx_lms_modules_course_version_position ON lms_modules(course_version_id, position);
                CREATE INDEX IF NOT EXISTS idx_lms_lessons_module_position ON lms_lessons(module_id, position);
                CREATE INDEX IF NOT EXISTS idx_lms_lesson_materials_lesson_position ON lms_lesson_materials(lesson_id, position);
                CREATE INDEX IF NOT EXISTS idx_lms_tests_course_version ON lms_tests(course_version_id);
                CREATE INDEX IF NOT EXISTS idx_lms_questions_test_position ON lms_questions(test_id, position);
                CREATE INDEX IF NOT EXISTS idx_lms_question_options_question_position ON lms_question_options(question_id, position);
                CREATE INDEX IF NOT EXISTS idx_lms_assignments_user_status ON lms_course_assignments(user_id, status);
                CREATE INDEX IF NOT EXISTS idx_lms_assignments_course ON lms_course_assignments(course_id);
                CREATE INDEX IF NOT EXISTS idx_lms_assignments_due_at ON lms_course_assignments(due_at);
                CREATE INDEX IF NOT EXISTS idx_lms_lesson_progress_assignment ON lms_lesson_progress(assignment_id);
                CREATE INDEX IF NOT EXISTS idx_lms_lesson_progress_user_lesson ON lms_lesson_progress(user_id, lesson_id);
                CREATE INDEX IF NOT EXISTS idx_lms_lesson_progress_assignment_user_lesson ON lms_lesson_progress(assignment_id, user_id, lesson_id);
                CREATE INDEX IF NOT EXISTS idx_lms_learning_sessions_lesson_user_active ON lms_learning_sessions(lesson_id, user_id, is_active);
                CREATE INDEX IF NOT EXISTS idx_lms_learning_events_session_created ON lms_learning_events(session_id, created_at);
                CREATE INDEX IF NOT EXISTS idx_lms_test_attempts_assignment_test ON lms_test_attempts(assignment_id, test_id);
                CREATE INDEX IF NOT EXISTS idx_lms_test_attempts_assignment_user_test_status_passed ON lms_test_attempts(assignment_id, user_id, test_id, status, passed);
                CREATE INDEX IF NOT EXISTS idx_lms_test_attempts_user_started ON lms_test_attempts(user_id, started_at DESC);
                CREATE INDEX IF NOT EXISTS idx_lms_attempt_answers_attempt ON lms_test_attempt_answers(attempt_id);
                CREATE INDEX IF NOT EXISTS idx_lms_certificates_user_status ON lms_certificates(user_id, status);
                CREATE INDEX IF NOT EXISTS idx_lms_certificates_token ON lms_certificates(verify_token);
                CREATE INDEX IF NOT EXISTS idx_lms_notifications_user_read ON lms_notifications(user_id, is_read, created_at DESC);
                CREATE INDEX IF NOT EXISTS idx_lms_admin_audit_actor_created ON lms_admin_audit_log(actor_id, created_at DESC);
            """)

    def create_user(
        self,
        telegram_id,
        name,
        role,
        direction_id=None,
        rate=None,
        hire_date=None,
        supervisor_id=None,
        login=None,
        password=None,
        hours_table_url=None,
        gender=None,
        birth_date=None,
        phone=None,
        email=None,
        instagram=None,
        telegram_nick=None,
        company_name=None,
        employment_type=None,
        has_proxy=None,
        proxy_card_number=None,
        has_driver_license=None,
        sip_number=None,
        study_place=None,
        study_course=None,
        close_contact_1_relation=None,
        close_contact_1_full_name=None,
        close_contact_1_phone=None,
        close_contact_2_relation=None,
        close_contact_2_full_name=None,
        close_contact_2_phone=None,
        card_number=None,
        internship_in_company=None,
        front_office_training=None,
        front_office_training_date=None,
        taxipro_id=None
    ):
        if login is None:
            base_login = f"user_{str(uuid.uuid4())[:8]}"
            with self._get_cursor() as cursor:
                while True:
                    cursor.execute("SELECT id FROM users WHERE login = %s", (base_login,))
                    if not cursor.fetchone():
                        login = base_login
                        break
                    base_login = f"user_{str(uuid.uuid4())[:8]}"
        else:
            login = login or str(telegram_id)

        password = password or "123456"
        password_hash = pbkdf2_sha256.hash(password)
        role_norm = str(role or '').strip().lower()
        phone = str(phone).strip() if phone is not None else ""
        email = str(email).strip() if email is not None else ""
        instagram = str(instagram).strip() if instagram is not None else ""
        telegram_nick = str(telegram_nick).strip() if telegram_nick is not None else ""
        company_name = str(company_name).strip() if company_name is not None else ""
        employment_type = str(employment_type).strip().lower() if employment_type is not None else ""
        proxy_card_number = str(proxy_card_number).strip() if proxy_card_number is not None else ""
        sip_number = str(sip_number).strip() if sip_number is not None else ""
        study_place = str(study_place).strip() if study_place is not None else ""
        study_course = str(study_course).strip() if study_course is not None else ""
        close_contact_1_relation = str(close_contact_1_relation).strip() if close_contact_1_relation is not None else ""
        close_contact_1_full_name = str(close_contact_1_full_name).strip() if close_contact_1_full_name is not None else ""
        close_contact_1_phone = str(close_contact_1_phone).strip() if close_contact_1_phone is not None else ""
        close_contact_2_relation = str(close_contact_2_relation).strip() if close_contact_2_relation is not None else ""
        close_contact_2_full_name = str(close_contact_2_full_name).strip() if close_contact_2_full_name is not None else ""
        close_contact_2_phone = str(close_contact_2_phone).strip() if close_contact_2_phone is not None else ""
        card_number = str(card_number).strip() if card_number is not None else ""
        taxipro_id = str(taxipro_id).strip() if taxipro_id is not None else ""

        phone = phone or None
        email = email or None
        instagram = instagram or None
        telegram_nick = telegram_nick or None
        company_name = company_name or None
        employment_type = employment_type or None
        proxy_card_number = proxy_card_number or None
        sip_number = sip_number or None
        study_place = study_place or None
        study_course = study_course or None
        close_contact_1_relation = close_contact_1_relation or None
        close_contact_1_full_name = close_contact_1_full_name or None
        close_contact_1_phone = close_contact_1_phone or None
        close_contact_2_relation = close_contact_2_relation or None
        close_contact_2_full_name = close_contact_2_full_name or None
        close_contact_2_phone = close_contact_2_phone or None
        card_number = card_number or None
        taxipro_id = taxipro_id or None
        if employment_type not in (None, 'gph', 'of'):
            raise ValueError("Invalid employment_type")
        has_proxy_value = None if has_proxy is None else bool(has_proxy)
        if not has_proxy_value:
            proxy_card_number = None
        has_driver_license_value = None if has_driver_license is None else bool(has_driver_license)
        internship_in_company_value = None if internship_in_company is None else bool(internship_in_company)
        front_office_training_value = None if front_office_training is None else bool(front_office_training)
        if not front_office_training_value:
            front_office_training_date = None
        if role_norm != 'operator':
            proxy_card_number = None
        if role_norm == 'trainer':
            direction_id = None
            supervisor_id = None
            sip_number = None

        with self._get_cursor() as cursor:
            def _clear_trainer_links(target_user_id):
                if role_norm == 'trainer' and target_user_id is not None:
                    cursor.execute(
                        "UPDATE users SET direction_id = NULL, supervisor_id = NULL WHERE id = %s",
                        (target_user_id,)
                    )

            cursor.execute("SAVEPOINT before_insert")
            try:
                cursor.execute("""
                    INSERT INTO users (
                        telegram_id, name, role, direction_id, rate, hire_date, supervisor_id,
                        login, password_hash, hours_table_url, gender, birth_date, phone, email,
                        instagram, telegram_nick, company_name, employment_type, has_proxy, proxy_card_number, has_driver_license, sip_number,
                        study_place, study_course,
                        close_contact_1_relation, close_contact_1_full_name, close_contact_1_phone,
                        close_contact_2_relation, close_contact_2_full_name, close_contact_2_phone,
                        card_number, internship_in_company, front_office_training, front_office_training_date, taxipro_id
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    telegram_id, name, role, direction_id, rate, hire_date, supervisor_id,
                    login, password_hash, hours_table_url, gender, birth_date, phone, email,
                    instagram, telegram_nick, company_name, employment_type,
                    (has_proxy_value if has_proxy_value is not None else False),
                    proxy_card_number,
                    (has_driver_license_value if has_driver_license_value is not None else False),
                    sip_number,
                    study_place, study_course,
                    close_contact_1_relation, close_contact_1_full_name, close_contact_1_phone,
                    close_contact_2_relation, close_contact_2_full_name, close_contact_2_phone,
                    card_number,
                    (internship_in_company_value if internship_in_company_value is not None else False),
                    (front_office_training_value if front_office_training_value is not None else False),
                    front_office_training_date,
                    taxipro_id
                ))
                created_user_id = cursor.fetchone()[0]
                _clear_trainer_links(created_user_id)
                return created_user_id
            except psycopg2.IntegrityError as e:
                cursor.execute("ROLLBACK TO SAVEPOINT before_insert")
                if 'unique_name_role' in str(e):
                    cursor.execute("""
                        UPDATE users
                        SET direction_id = COALESCE(%s, direction_id),
                            supervisor_id = COALESCE(%s, supervisor_id),
                            hours_table_url = COALESCE(%s, hours_table_url),
                            gender = COALESCE(%s, gender),
                            birth_date = COALESCE(%s, birth_date),
                            phone = COALESCE(%s, phone),
                            email = COALESCE(%s, email),
                            instagram = COALESCE(%s, instagram),
                            telegram_nick = COALESCE(%s, telegram_nick),
                            company_name = COALESCE(%s, company_name),
                            employment_type = COALESCE(%s, employment_type),
                            has_proxy = COALESCE(%s, has_proxy),
                            proxy_card_number = COALESCE(%s, proxy_card_number),
                            has_driver_license = COALESCE(%s, has_driver_license),
                            sip_number = COALESCE(%s, sip_number),
                            study_place = COALESCE(%s, study_place),
                            study_course = COALESCE(%s, study_course),
                            close_contact_1_relation = COALESCE(%s, close_contact_1_relation),
                            close_contact_1_full_name = COALESCE(%s, close_contact_1_full_name),
                            close_contact_1_phone = COALESCE(%s, close_contact_1_phone),
                            close_contact_2_relation = COALESCE(%s, close_contact_2_relation),
                            close_contact_2_full_name = COALESCE(%s, close_contact_2_full_name),
                            close_contact_2_phone = COALESCE(%s, close_contact_2_phone),
                            card_number = COALESCE(%s, card_number),
                            internship_in_company = COALESCE(%s, internship_in_company),
                            front_office_training = COALESCE(%s, front_office_training),
                            front_office_training_date = COALESCE(%s, front_office_training_date),
                            taxipro_id = COALESCE(%s, taxipro_id)
                        WHERE name = %s AND role = %s
                        RETURNING id
                    """, (
                        direction_id, supervisor_id, hours_table_url, gender, birth_date,
                        phone, email, instagram, telegram_nick, company_name, employment_type, has_proxy_value, proxy_card_number, has_driver_license_value, sip_number,
                        study_place, study_course,
                        close_contact_1_relation, close_contact_1_full_name, close_contact_1_phone,
                        close_contact_2_relation, close_contact_2_full_name, close_contact_2_phone,
                        card_number, internship_in_company_value, front_office_training_value, front_office_training_date, taxipro_id,
                        name, role
                    ))
                    result = cursor.fetchone()
                    if result:
                        updated_user_id = result[0]
                        _clear_trainer_links(updated_user_id)
                        return updated_user_id
                    else:
                        raise ValueError("User with this name and role not found")
                elif 'telegram_id' in str(e):
                    cursor.execute("""
                        UPDATE users
                        SET name = %s,
                            role = %s,
                            direction_id = COALESCE(%s, direction_id),
                            hire_date = COALESCE(%s, hire_date),
                            supervisor_id = COALESCE(%s, supervisor_id),
                            login = %s,
                            password_hash = %s,
                            hours_table_url = COALESCE(%s, hours_table_url),
                            gender = COALESCE(%s, gender),
                            birth_date = COALESCE(%s, birth_date),
                            phone = COALESCE(%s, phone),
                            email = COALESCE(%s, email),
                            instagram = COALESCE(%s, instagram),
                            telegram_nick = COALESCE(%s, telegram_nick),
                            company_name = COALESCE(%s, company_name),
                            employment_type = COALESCE(%s, employment_type),
                            has_proxy = COALESCE(%s, has_proxy),
                            proxy_card_number = COALESCE(%s, proxy_card_number),
                            has_driver_license = COALESCE(%s, has_driver_license),
                            sip_number = COALESCE(%s, sip_number),
                            study_place = COALESCE(%s, study_place),
                            study_course = COALESCE(%s, study_course),
                            close_contact_1_relation = COALESCE(%s, close_contact_1_relation),
                            close_contact_1_full_name = COALESCE(%s, close_contact_1_full_name),
                            close_contact_1_phone = COALESCE(%s, close_contact_1_phone),
                            close_contact_2_relation = COALESCE(%s, close_contact_2_relation),
                            close_contact_2_full_name = COALESCE(%s, close_contact_2_full_name),
                            close_contact_2_phone = COALESCE(%s, close_contact_2_phone),
                            card_number = COALESCE(%s, card_number),
                            internship_in_company = COALESCE(%s, internship_in_company),
                            front_office_training = COALESCE(%s, front_office_training),
                            front_office_training_date = COALESCE(%s, front_office_training_date),
                            taxipro_id = COALESCE(%s, taxipro_id)
                        WHERE telegram_id = %s
                        RETURNING id
                    """, (
                        name, role, direction_id, hire_date, supervisor_id, login, password_hash, hours_table_url,
                        gender, birth_date, phone, email, instagram, telegram_nick, company_name, employment_type,
                        has_proxy_value, proxy_card_number, has_driver_license_value, sip_number,
                        study_place, study_course,
                        close_contact_1_relation, close_contact_1_full_name, close_contact_1_phone,
                        close_contact_2_relation, close_contact_2_full_name, close_contact_2_phone,
                        card_number, internship_in_company_value, front_office_training_value, front_office_training_date, taxipro_id,
                        telegram_id
                    ))
                    updated_user_id = cursor.fetchone()[0]
                    _clear_trainer_links(updated_user_id)
                    return updated_user_id
                else:
                    raise

    def update_telegram_id(self, user_id, telegram_id):
        with self._get_cursor() as cursor:
            cursor.execute("""
                UPDATE users 
                SET telegram_id = %s
                WHERE id = %s
                RETURNING telegram_id
            """, (telegram_id, user_id))
            result = cursor.fetchone()
            if result:
                return result[0]
            return None
            
    def get_directions(self):
        """Получить все направления из таблицы directions."""
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, name, has_file_upload, criteria, is_active
                FROM directions
                WHERE is_active = TRUE  -- Added filter for performance
                ORDER BY name
            """)
            return [
                {
                    "id": row[0],
                    "name": row[1],
                    "hasFileUpload": row[2],
                    "criteria": row[3],
                    "isActive": row[4]
                } for row in cursor.fetchall()
            ]

    def insert_or_update_daily_hours(self, operator_id, day, work_time=0.0, break_time=0.0,
                                    talk_time=0.0, calls=0, efficiency=0.0,
                                    fine_amount=0.0, fine_reason=None, fine_comment=None,
                                    fines: List[Dict[str, Any]] = None):
        """
        Вставляет/обновляет запись daily_hours (operator_id + day).
        efficiency и fine_amount ожидаются в часах/суммах соответственно.
        """
        if isinstance(day, str):
            day = datetime.strptime(day, "%Y-%m-%d").date()
        with self._get_cursor() as cursor:
            # Insert or update daily_hours and return its id for managing daily_fines
            cursor.execute("""
                INSERT INTO daily_hours (
                    operator_id, day, work_time, break_time, talk_time, calls, efficiency,
                    fine_amount, fine_reason, fine_comment
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (operator_id, day)
                DO UPDATE SET
                    work_time = EXCLUDED.work_time,
                    break_time = EXCLUDED.break_time,
                    talk_time = EXCLUDED.talk_time,
                    calls = EXCLUDED.calls,
                    efficiency = EXCLUDED.efficiency,
                    fine_amount = EXCLUDED.fine_amount,
                    fine_reason = EXCLUDED.fine_reason,
                    fine_comment = EXCLUDED.fine_comment,
                    created_at = CURRENT_TIMESTAMP
                RETURNING id
            """, (operator_id, day, work_time, break_time, talk_time, calls, efficiency,
                float(fine_amount) if fine_amount is not None else 0.0,
                fine_reason, fine_comment))
            res = cursor.fetchone()
            daily_id = res[0] if res else None

            # If fines provided, replace existing fines for this daily_hours record
            if daily_id and isinstance(fines, list):
                # remove previous fines for this daily row
                cursor.execute("DELETE FROM daily_fines WHERE daily_hours_id = %s", (daily_id,))
                insert_vals = []
                total_amount = 0.0
                first_reason = None
                comments = []
                for f in fines:
                    amt = float(f.get('amount') or 0)
                    reason = f.get('reason')
                    comment = f.get('comment')
                    insert_vals.append((daily_id, operator_id, day, amt, reason, comment))
                    total_amount += amt
                    if not first_reason and reason:
                        first_reason = reason
                    if comment:
                        comments.append(str(comment))

                if insert_vals:
                    cursor.executemany(
                        "INSERT INTO daily_fines (daily_hours_id, operator_id, day, amount, reason, comment) VALUES (%s, %s, %s, %s, %s, %s)",
                        insert_vals
                    )

                # update aggregated fields in daily_hours for backward compatibility
                agg_comment = '; '.join(comments) if comments else None
                cursor.execute("""
                    UPDATE daily_hours SET fine_amount = %s, fine_reason = %s, fine_comment = %s
                    WHERE id = %s
                """, (float(total_amount), first_reason, agg_comment, daily_id))

    def _load_technical_issues_by_operator_day_tx(self, cursor, operator_ids, start_date, end_date):
        """
        Возвращает:
        - map: {operator_id: {day_num_str: [issues...]}}
        - totals: {operator_id: total_hours}
        """
        op_ids = sorted({int(v) for v in (operator_ids or []) if v is not None})
        result = {op_id: {} for op_id in op_ids}
        totals = {op_id: 0.0 for op_id in op_ids}
        if not op_ids:
            return result, totals

        cursor.execute(
            """
            SELECT
                ti.id,
                ti.operator_id,
                ti.issue_date,
                ti.start_time,
                ti.end_time,
                ti.reason,
                ti.comment,
                cb.name,
                ti.created_at
            FROM operator_technical_issues ti
            LEFT JOIN users cb ON cb.id = ti.created_by
            WHERE ti.operator_id = ANY(%s)
              AND ti.issue_date >= %s
              AND ti.issue_date <= %s
            ORDER BY ti.operator_id, ti.issue_date, ti.start_time, ti.end_time, ti.id
            """,
            (op_ids, start_date, end_date),
        )
        for issue_id, op_id, issue_date, start_time, end_time, reason, comment, created_by_name, created_at in cursor.fetchall() or []:
            try:
                op_id_int = int(op_id)
            except Exception:
                continue
            if op_id_int not in result or issue_date is None:
                continue

            start_text = start_time.strftime('%H:%M') if start_time else None
            end_text = end_time.strftime('%H:%M') if end_time else None
            duration_minutes = 0
            try:
                if start_time and end_time:
                    start_min, end_min = self._schedule_interval_minutes(start_time, end_time)
                    duration_minutes = max(0, int(end_min - start_min))
            except Exception:
                duration_minutes = 0
            duration_hours = round(float(duration_minutes) / 60.0, 2)

            day_key = str(int(issue_date.day))
            result.setdefault(op_id_int, {}).setdefault(day_key, []).append({
                "id": int(issue_id),
                "date": issue_date.strftime('%Y-%m-%d'),
                "start_time": start_text,
                "end_time": end_text,
                "time_range": f"{start_text} - {end_text}" if start_text and end_text else None,
                "reason": reason,
                "comment": comment,
                "created_by_name": created_by_name,
                "created_at": created_at.strftime('%Y-%m-%d %H:%M:%S') if created_at else None,
                "duration_minutes": int(duration_minutes),
                "duration_hours": float(duration_hours),
            })
            totals[op_id_int] = round(float(totals.get(op_id_int, 0.0)) + float(duration_hours), 2)

        return result, totals

    def _load_offline_activities_by_operator_day_tx(self, cursor, operator_ids, start_date, end_date):
        """
        Возвращает:
        - map: {operator_id: {day_num_str: [activities...]}}
        - totals: {operator_id: total_hours}
        """
        op_ids = sorted({int(v) for v in (operator_ids or []) if v is not None})
        result = {op_id: {} for op_id in op_ids}
        totals = {op_id: 0.0 for op_id in op_ids}
        if not op_ids:
            return result, totals

        cursor.execute(
            """
            SELECT
                oa.id,
                oa.operator_id,
                oa.activity_date,
                oa.start_time,
                oa.end_time,
                oa.comment,
                cb.name,
                oa.created_at
            FROM operator_offline_activities oa
            LEFT JOIN users cb ON cb.id = oa.created_by
            WHERE oa.operator_id = ANY(%s)
              AND oa.activity_date >= %s
              AND oa.activity_date <= %s
            ORDER BY oa.operator_id, oa.activity_date, oa.start_time, oa.end_time, oa.id
            """,
            (op_ids, start_date, end_date),
        )
        for activity_id, op_id, activity_date, start_time, end_time, comment, created_by_name, created_at in cursor.fetchall() or []:
            try:
                op_id_int = int(op_id)
            except Exception:
                continue
            if op_id_int not in result or activity_date is None:
                continue

            start_text = start_time.strftime('%H:%M') if start_time else None
            end_text = end_time.strftime('%H:%M') if end_time else None
            duration_minutes = 0
            try:
                if start_time and end_time:
                    start_min, end_min = self._schedule_interval_minutes(start_time, end_time)
                    duration_minutes = max(0, int(end_min - start_min))
            except Exception:
                duration_minutes = 0
            duration_hours = round(float(duration_minutes) / 60.0, 2)

            day_key = str(int(activity_date.day))
            comment_text = str(comment or '').strip()
            result.setdefault(op_id_int, {}).setdefault(day_key, []).append({
                "id": int(activity_id),
                "date": activity_date.strftime('%Y-%m-%d'),
                "start_time": start_text,
                "end_time": end_text,
                "time_range": f"{start_text} - {end_text}" if start_text and end_text else None,
                "comment": comment_text,
                "created_by_name": created_by_name,
                "created_at": created_at.strftime('%Y-%m-%d %H:%M:%S') if created_at else None,
                "duration_minutes": int(duration_minutes),
                "duration_hours": float(duration_hours),
            })
            totals[op_id_int] = round(float(totals.get(op_id_int, 0.0)) + float(duration_hours), 2)

        return result, totals

    def get_daily_hours_for_operator_month(self, operator_id: int, month: str):
        """
        Возвращает daily_hours для одного оператора за месяц YYYY-MM,
        а также norm_hours, aggregates (из work_hours) и rate/name из users.

        Включает поля штрафов из daily_hours: fine_amount, fine_reason, fine_comment
        и агрегированное поле fines из work_hours.

        Результат:
        {
            "month": month,
            "days_in_month": <int>,
            "operator": { ... }
        }
        """
        import calendar as _py_calendar
        from datetime import date as _date

        # validate month format YYYY-MM and compute date range
        try:
            year, mon = map(int, month.split('-'))
            days_in_month = _py_calendar.monthrange(year, mon)[1]
            start = _date(year, mon, 1)
            end = _date(year, mon, days_in_month)
        except Exception as e:
            raise ValueError("Invalid month format, expected YYYY-MM") from e

        with self._get_cursor() as cursor:
            # 1) Получаем daily_hours (одним запросом), теперь с информацией о штрафах
            cursor.execute(
                """
                SELECT day, work_time, break_time, talk_time, calls, efficiency,
                    fine_amount, fine_reason, fine_comment
                FROM daily_hours
                WHERE operator_id = %s AND day >= %s AND day <= %s
                ORDER BY day
                """,
                (operator_id, start, end),
            )
            daily_rows = cursor.fetchall()

            # build daily map: { "1": {...}, "2": {...}, ... }
            daily_map = {}
            for row in daily_rows:
                # row: (day, work_time, break_time, talk_time, calls, efficiency, fine_amount, fine_reason, fine_comment)
                day_obj = row[0]
                day_key = str(int(day_obj.day))
                daily_map[day_key] = {
                    "work_time": float(row[1]) if row[1] is not None else 0.0,
                    "break_time": float(row[2]) if row[2] is not None else 0.0,
                    "talk_time": float(row[3]) if row[3] is not None else 0.0,
                    "calls": int(row[4]) if row[4] is not None else 0,
                    "efficiency": float(row[5]) if row[5] is not None else 0.0,
                    "fine_amount": float(row[6]) if row[6] is not None else 0.0,
                    "fine_reason": row[7],
                    "fine_comment": row[8],
                    "fines": []  # will populate from daily_fines table if present
                }

            # fetch fines for this operator in the date range
            cursor.execute(
                """
                SELECT df.amount, df.reason, df.comment, dh.day
                FROM daily_fines df
                JOIN daily_hours dh ON df.daily_hours_id = dh.id
                WHERE dh.operator_id = %s AND dh.day >= %s AND dh.day <= %s
                ORDER BY dh.day, df.id
                """,
                (operator_id, start, end)
            )
            fines_rows = cursor.fetchall()
            for amt, reason, comment, day_obj in fines_rows:
                day_key = str(int(day_obj.day))
                entry = daily_map.get(day_key)
                fine_obj = {
                    "amount": float(amt) if amt is not None else 0.0,
                    "reason": reason,
                    "comment": comment
                }
                # include minutes for Опоздание for client convenience
                try:
                    if str(reason) == 'Опоздание':
                        fine_obj['minutes'] = int(round(float(amt) / 50)) if amt else 0
                except Exception:
                    pass
                if entry is not None:
                    entry.setdefault('fines', []).append(fine_obj)

            technical_map_by_operator, technical_totals_by_operator = self._load_technical_issues_by_operator_day_tx(
                cursor=cursor,
                operator_ids=[operator_id],
                start_date=start,
                end_date=end
            )
            offline_map_by_operator, offline_totals_by_operator = self._load_offline_activities_by_operator_day_tx(
                cursor=cursor,
                operator_ids=[operator_id],
                start_date=start,
                end_date=end
            )

            # 2) Получаем имя/ставку + данные work_hours одним LEFT JOIN запросом (включая fines)
            cursor.execute(
                """
                SELECT u.name, u.rate,
                    d.name as direction_name,
                    COALESCE(w.norm_hours, 0) AS norm_hours,
                    COALESCE(w.regular_hours, 0) AS regular_hours,
                    COALESCE(w.total_break_time, 0) AS total_break_time,
                    COALESCE(w.total_talk_time, 0) AS total_talk_time,
                    COALESCE(w.total_calls, 0) AS total_calls,
                    COALESCE(w.total_efficiency_hours, 0) AS total_efficiency_hours,
                    COALESCE(w.calls_per_hour, 0) AS calls_per_hour,
                    COALESCE(w.fines, 0) AS fines
                FROM users u
                LEFT JOIN work_hours w
                ON w.operator_id = u.id AND w.month = %s
                LEFT JOIN directions d ON u.direction_id = d.id
                WHERE u.id = %s
                LIMIT 1
                """,
                (month, operator_id),
            )
            row = cursor.fetchone()

        technical_issues_by_day = technical_map_by_operator.get(int(operator_id), {}) if isinstance(technical_map_by_operator, dict) else {}
        technical_issue_hours = float(technical_totals_by_operator.get(int(operator_id), 0.0)) if isinstance(technical_totals_by_operator, dict) else 0.0
        offline_activities_by_day = offline_map_by_operator.get(int(operator_id), {}) if isinstance(offline_map_by_operator, dict) else {}
        offline_activity_hours = float(offline_totals_by_operator.get(int(operator_id), 0.0)) if isinstance(offline_totals_by_operator, dict) else 0.0

        # default values if user / work_hours not found
        if row:
            name, rate, direction_name, norm_hours, regular_hours, total_break_time, total_talk_time, \
                total_calls, total_efficiency_hours, calls_per_hour, fines = row
            rate = float(rate) if rate is not None else 0.0
        else:
            name = None
            rate = 0.0
            direction_name = None
            norm_hours = regular_hours = total_break_time = total_talk_time = 0.0
            total_calls = 0
            total_efficiency_hours = calls_per_hour = fines = 0.0

        operator_obj = {
            "operator_id": operator_id,
            "name": name,
            "direction": direction_name,
            "rate": rate,
            "norm_hours": float(norm_hours),
            "fines": float(fines),
            "daily": daily_map,
            "technical_issues_by_day": technical_issues_by_day,
            "technical_issue_hours": round(float(technical_issue_hours), 2),
            "offline_activities_by_day": offline_activities_by_day,
            "offline_activity_hours": round(float(offline_activity_hours), 2),
            "aggregates": {
                "regular_hours": float(regular_hours),
                "total_break_time": float(total_break_time),
                "total_talk_time": float(total_talk_time),
                "total_calls": int(total_calls),
                "total_efficiency_hours": float(total_efficiency_hours),
                "calls_per_hour": float(calls_per_hour),
            },
        }

        return {"month": month, "days_in_month": days_in_month, "operator": operator_obj}

    def get_daily_hours_by_supervisor_month(self, supervisor_id, month):
        """
        Возвращает все daily_hours и агрегаты work_hours для всех операторов указанного супервайзера за месяц YYYY-MM.
        Включает также ставку (rate) из users, norm_hours и fines из work_hours.
        """
        import calendar as _py_calendar
        from datetime import date as _date

        # validate month format YYYY-MM
        try:
            year, mon = map(int, month.split('-'))
            days = _py_calendar.monthrange(year, mon)[1]
            start = _date(year, mon, 1)
            end = _date(year, mon, days)
        except Exception as e:
            raise ValueError("Invalid month format, expected YYYY-MM") from e

        with self._get_cursor() as cursor:
            # Получаем операторов + ставка + norm_hours + агрегаты work_hours (включая fines)
            cursor.execute("""
                SELECT u.id, u.name, u.rate, u.status, u.supervisor_id,
                    d.name as direction_name,
                    COALESCE(w.norm_hours, 0) as norm_hours,
                    COALESCE(w.regular_hours, 0) as regular_hours,
                    COALESCE(w.total_break_time, 0) as total_break_time,
                    COALESCE(w.total_talk_time, 0) as total_talk_time,
                    COALESCE(w.total_calls, 0) as total_calls,
                    COALESCE(w.total_efficiency_hours, 0) as total_efficiency_hours,
                    COALESCE(w.calls_per_hour, 0) as calls_per_hour,
                    COALESCE(w.fines, 0) as fines
                FROM users u
                LEFT JOIN work_hours w
                ON w.operator_id = u.id AND w.month = %s
                LEFT JOIN directions d ON u.direction_id = d.id
                WHERE u.role = 'operator' AND u.supervisor_id = %s
                ORDER BY u.name
            """, (month, supervisor_id))
            operator_rows = cursor.fetchall()  # list of tuples

            if not operator_rows:
                return {"month": month, "days_in_month": days, "operators": []}

            op_ids = [row[0] for row in operator_rows]

            # Получаем daily_hours для этих операторов за месяц (с информацией о штрафах)
            cursor.execute("""
                SELECT d.operator_id, d.day, d.work_time, d.break_time, d.talk_time, d.calls, d.efficiency,
                    d.fine_amount, d.fine_reason, d.fine_comment
                FROM daily_hours d
                WHERE d.operator_id = ANY(%s)
                AND d.day >= %s AND d.day <= %s
                ORDER BY d.operator_id, d.day
            """, (op_ids, start, end))
            daily_rows = cursor.fetchall()

            # Готовим словарь daily: operator_id -> {day_number: {...}}
            daily_map = {}
            for (op_id, day, work_time, break_time, talk_time, calls, eff,
                fine_amount, fine_reason, fine_comment) in daily_rows:
                day_num = str(int(day.day))
                d = {
                    "work_time": float(work_time) if work_time is not None else 0.0,
                    "break_time": float(break_time) if break_time is not None else 0.0,
                    "talk_time": float(talk_time) if talk_time is not None else 0.0,
                    "calls": int(calls) if calls is not None else 0,
                    "efficiency": float(eff) if eff is not None else 0.0,
                    "fine_amount": float(fine_amount) if fine_amount is not None else 0.0,
                    "fine_reason": fine_reason,
                    "fine_comment": fine_comment,
                    "fines": []
                }
                daily_map.setdefault(op_id, {})[day_num] = d

            # fetch fines for all operators in range
            cursor.execute("""
                SELECT df.operator_id, df.amount, df.reason, df.comment, df.day
                FROM daily_fines df
                WHERE df.operator_id = ANY(%s)
                AND df.day >= %s AND df.day <= %s
                ORDER BY df.operator_id, df.day, df.id
            """, (op_ids, start, end))
            fines_rows = cursor.fetchall()
            for op_id, amt, reason, comment, day_obj in fines_rows:
                day_num = str(int(day_obj.day))
                fine_obj = {"amount": float(amt) if amt is not None else 0.0, "reason": reason, "comment": comment}
                try:
                    if str(reason) == 'Опоздание':
                        fine_obj['minutes'] = int(round(float(amt) / 50)) if amt else 0
                except Exception:
                    pass
                if op_id in daily_map and day_num in daily_map[op_id]:
                    daily_map[op_id][day_num].setdefault('fines', []).append(fine_obj)

            technical_map_by_operator, technical_totals_by_operator = self._load_technical_issues_by_operator_day_tx(
                cursor=cursor,
                operator_ids=op_ids,
                start_date=start,
                end_date=end
            )
            offline_map_by_operator, offline_totals_by_operator = self._load_offline_activities_by_operator_day_tx(
                cursor=cursor,
                operator_ids=op_ids,
                start_date=start,
                end_date=end
            )

            # Сбор финального списка операторов
            operators = []
            for row in operator_rows:
                (op_id, op_name, rate, status, sup_id, direction_name, norm_hours,
                regular_hours, total_break_time, total_talk_time,
                total_calls, total_efficiency_hours, calls_per_hour, fines) = row

                operators.append({
                    "operator_id": op_id,
                    "name": op_name,
                    "direction": direction_name,
                    "supervisor_id": sup_id,
                    "rate": float(rate) if rate is not None else 0.0,
                    "status": status,
                    "norm_hours": float(norm_hours) if norm_hours is not None else 0.0,
                    "daily": daily_map.get(op_id, {}),
                    "technical_issues_by_day": technical_map_by_operator.get(op_id, {}) if isinstance(technical_map_by_operator, dict) else {},
                    "technical_issue_hours": round(float(technical_totals_by_operator.get(op_id, 0.0)), 2) if isinstance(technical_totals_by_operator, dict) else 0.0,
                    "offline_activities_by_day": offline_map_by_operator.get(op_id, {}) if isinstance(offline_map_by_operator, dict) else {},
                    "offline_activity_hours": round(float(offline_totals_by_operator.get(op_id, 0.0)), 2) if isinstance(offline_totals_by_operator, dict) else 0.0,
                    "aggregates": {
                        "regular_hours": float(regular_hours),
                        "total_break_time": float(total_break_time),
                        "total_talk_time": float(total_talk_time),
                        "total_calls": int(total_calls),
                        "total_efficiency_hours": float(total_efficiency_hours),
                        "calls_per_hour": float(calls_per_hour),
                        "fines": float(fines)
                    }
                })

        return {"month": month, "days_in_month": days, "operators": operators}

    def get_daily_hours_for_all_month(self, month):
        """
        Возвращает все daily_hours и агрегаты work_hours для всех операторов за месяц YYYY-MM.
        Аналогично get_daily_hours_by_supervisor_month, но без фильтра по супервайзеру.
        """
        import calendar as _py_calendar
        from datetime import date as _date

        try:
            year, mon = map(int, month.split('-'))
            days = _py_calendar.monthrange(year, mon)[1]
            start = _date(year, mon, 1)
            end = _date(year, mon, days)
        except Exception as e:
            raise ValueError("Invalid month format, expected YYYY-MM") from e

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT u.id, u.name, u.rate, u.status, u.supervisor_id,
                    d.name as direction_name,
                    COALESCE(w.norm_hours, 0) as norm_hours,
                    COALESCE(w.regular_hours, 0) as regular_hours,
                    COALESCE(w.total_break_time, 0) as total_break_time,
                    COALESCE(w.total_talk_time, 0) as total_talk_time,
                    COALESCE(w.total_calls, 0) as total_calls,
                    COALESCE(w.total_efficiency_hours, 0) as total_efficiency_hours,
                    COALESCE(w.calls_per_hour, 0) as calls_per_hour,
                    COALESCE(w.fines, 0) as fines
                FROM users u
                LEFT JOIN work_hours w
                ON w.operator_id = u.id AND w.month = %s
                LEFT JOIN directions d ON u.direction_id = d.id
                WHERE u.role = 'operator'
                ORDER BY u.name
            """, (month,))
            operator_rows = cursor.fetchall()

            if not operator_rows:
                return {"month": month, "days_in_month": days, "operators": []}

            op_ids = [row[0] for row in operator_rows]

            cursor.execute("""
                SELECT d.operator_id, d.day, d.work_time, d.break_time, d.talk_time, d.calls, d.efficiency,
                    d.fine_amount, d.fine_reason, d.fine_comment
                FROM daily_hours d
                WHERE d.operator_id = ANY(%s)
                AND d.day >= %s AND d.day <= %s
                ORDER BY d.operator_id, d.day
            """, (op_ids, start, end))
            daily_rows = cursor.fetchall()

            daily_map = {}
            for (op_id, day, work_time, break_time, talk_time, calls, eff,
                fine_amount, fine_reason, fine_comment) in daily_rows:
                day_num = str(int(day.day))
                d = {
                    "work_time": float(work_time) if work_time is not None else 0.0,
                    "break_time": float(break_time) if break_time is not None else 0.0,
                    "talk_time": float(talk_time) if talk_time is not None else 0.0,
                    "calls": int(calls) if calls is not None else 0,
                    "efficiency": float(eff) if eff is not None else 0.0,
                    "fine_amount": float(fine_amount) if fine_amount is not None else 0.0,
                    "fine_reason": fine_reason,
                    "fine_comment": fine_comment,
                    "fines": []
                }
                daily_map.setdefault(op_id, {})[day_num] = d

            cursor.execute("""
                SELECT df.operator_id, df.amount, df.reason, df.comment, df.day
                FROM daily_fines df
                WHERE df.operator_id = ANY(%s)
                AND df.day >= %s AND df.day <= %s
                ORDER BY df.operator_id, df.day, df.id
            """, (op_ids, start, end))
            fines_rows = cursor.fetchall()
            for op_id, amt, reason, comment, day_obj in fines_rows:
                day_num = str(int(day_obj.day))
                fine_obj = {"amount": float(amt) if amt is not None else 0.0, "reason": reason, "comment": comment}
                try:
                    if str(reason) == 'Опоздание':
                        fine_obj['minutes'] = int(round(float(amt) / 50)) if amt else 0
                except Exception:
                    pass
                if op_id in daily_map and day_num in daily_map[op_id]:
                    daily_map[op_id][day_num].setdefault('fines', []).append(fine_obj)

            technical_map_by_operator, technical_totals_by_operator = self._load_technical_issues_by_operator_day_tx(
                cursor=cursor,
                operator_ids=op_ids,
                start_date=start,
                end_date=end
            )
            offline_map_by_operator, offline_totals_by_operator = self._load_offline_activities_by_operator_day_tx(
                cursor=cursor,
                operator_ids=op_ids,
                start_date=start,
                end_date=end
            )

            operators = []
            for row in operator_rows:
                (op_id, op_name, rate, status, sup_id, direction_name, norm_hours,
                regular_hours, total_break_time, total_talk_time,
                total_calls, total_efficiency_hours, calls_per_hour, fines) = row

                operators.append({
                    "operator_id": op_id,
                    "name": op_name,
                    "direction": direction_name,
                    "supervisor_id": sup_id,
                    "rate": float(rate) if rate is not None else 0.0,
                    "status": status,
                    "norm_hours": float(norm_hours) if norm_hours is not None else 0.0,
                    "daily": daily_map.get(op_id, {}),
                    "technical_issues_by_day": technical_map_by_operator.get(op_id, {}) if isinstance(technical_map_by_operator, dict) else {},
                    "technical_issue_hours": round(float(technical_totals_by_operator.get(op_id, 0.0)), 2) if isinstance(technical_totals_by_operator, dict) else 0.0,
                    "offline_activities_by_day": offline_map_by_operator.get(op_id, {}) if isinstance(offline_map_by_operator, dict) else {},
                    "offline_activity_hours": round(float(offline_totals_by_operator.get(op_id, 0.0)), 2) if isinstance(offline_totals_by_operator, dict) else 0.0,
                    "aggregates": {
                        "regular_hours": float(regular_hours),
                        "total_break_time": float(total_break_time),
                        "total_talk_time": float(total_talk_time),
                        "total_calls": int(total_calls),
                        "total_efficiency_hours": float(total_efficiency_hours),
                        "calls_per_hour": float(calls_per_hour),
                        "fines": float(fines)
                    }
                })

        return {"month": month, "days_in_month": days, "operators": operators}

    def aggregate_month_from_daily(self, operator_id, month):
        """
        Публичный wrapper месячной агрегации daily_hours -> work_hours.
        """
        with self._get_cursor() as cursor:
            return self._aggregate_month_from_daily_tx(
                cursor=cursor,
                operator_id=int(operator_id),
                month=str(month)
            )

    def auto_fill_norm_hours(self, month):
        """
        Авто-подсчет `norm_hours` для всех операторов за указанный месяц (формат 'YYYY-MM'),
        где текущая `norm_hours` равна 0. Формула: рабочие дни * 8 * rate.

        Возвращает словарь: {"month": month, "work_days": int, "processed": int}
        processed — число строк, на которые сработал INSERT/UPDATE (оценочно).
        """
        try:
            year, mon = map(int, month.split('-'))
        except Exception as e:
            raise ValueError("Invalid month format, expected YYYY-MM") from e

        work_days = self._get_month_work_days(f"{year}-{mon:02d}")

        with self._get_cursor() as cursor:
            # Вставляем/обновляем norm_hours для всех операторов.
            # Для операторов без строки в work_hours будет INSERT, для существующих с norm_hours=0 — UPDATE.
            cursor.execute("""
                INSERT INTO work_hours (operator_id, month, norm_hours)
                SELECT u.id, %s as month, (%s::float * 8.0 * COALESCE(u.rate, 1.0))::float
                FROM users u
                WHERE u.role = 'operator'
                ON CONFLICT (operator_id, month) DO UPDATE
                  SET norm_hours = EXCLUDED.norm_hours
                  WHERE COALESCE(work_hours.norm_hours, 0) = 0
            """, (month, work_days))

            # rowcount может быть не точен при ON CONFLICT, но даёт оценку затронутых строк
            processed = cursor.rowcount if cursor.rowcount is not None else 0

        return {"month": month, "work_days": work_days, "processed": int(processed)}

    def save_directions(self, directions):
        """Сохранить направления в таблицу directions, создавая новые версии при изменениях."""
        with self._get_cursor() as cursor:
            # 1. Получаем текущие активные направления (только нужные поля)
            cursor.execute("""
                SELECT id, name, has_file_upload, criteria, version
                FROM directions
                WHERE is_active = TRUE
            """)
            existing_directions = {
                row[1]: {
                    "id": row[0],
                    "has_file_upload": row[2],
                    "criteria": row[3],
                    "version": row[4]
                } for row in cursor.fetchall()
            }
            
            new_direction_names = {d['name'] for d in directions}
            directions_to_deactivate = []
            insert_values = []
            
            # 2. Подготовка данных для пакетной обработки
            for direction in directions:
                name = direction['name']
                has_file_upload = direction['hasFileUpload']
                criteria = json.dumps(direction['criteria'])
                
                if name in existing_directions:
                    existing = existing_directions[name]
                    if (existing['has_file_upload'] == has_file_upload and
                        existing['criteria'] == criteria):
                        continue  # Ничего не изменилось, пропускаем
                    
                    directions_to_deactivate.append(existing['id'])
                    insert_values.append((
                        name, has_file_upload, criteria,
                        existing['version'] + 1, existing['id']
                    ))
                else:
                    insert_values.append((
                        name, has_file_upload, criteria,
                        1, None  # version=1, no previous version
                    ))
            
            # 3. Пакетное деактивирование старых версий
            if directions_to_deactivate:
                cursor.execute("""
                    UPDATE directions
                    SET is_active = FALSE
                    WHERE id = ANY(%s)
                """, (directions_to_deactivate,))
            
            # 4. Пакетное добавление новых версий
            if insert_values:
                cursor.executemany("""
                    INSERT INTO directions (
                        name, has_file_upload, criteria,
                        version, previous_version_id
                    )
                    VALUES (%s, %s, %s, %s, %s)
                """, insert_values)
            
            # 5. Деактивируем направления, которых нет в новом списке
            if new_direction_names:
                cursor.execute("""
                    UPDATE directions
                    SET is_active = FALSE
                    WHERE is_active = TRUE AND name NOT IN %s
                """, (tuple(new_direction_names),))
            else:
                cursor.execute("""
                    UPDATE directions
                    SET is_active = FALSE
                    WHERE is_active = TRUE
                """)
            
            # 6. Оптимизированное обновление direction_id в users
            cursor.execute("""
                -- Обновляем direction_id на последнюю активную версию с тем же именем
                UPDATE users u
                SET direction_id = latest.id
                FROM directions current
                JOIN (
                    SELECT name, id,
                        ROW_NUMBER() OVER (PARTITION BY name ORDER BY version DESC) as rn
                    FROM directions
                    WHERE is_active = TRUE
                ) latest ON current.name = latest.name AND latest.rn = 1
                WHERE u.direction_id = current.id AND current.is_active = FALSE;
                
                -- Обнуляем direction_id где нет активных направлений
                UPDATE users
                SET direction_id = NULL
                WHERE direction_id IS NOT NULL
                AND direction_id NOT IN (
                    SELECT id FROM directions WHERE is_active = TRUE
                );
            """)

    def get_operator_credentials(self, operator_id, supervisor_id):
        """Получить логин и хеш пароля оператора с проверкой принадлежности супервайзеру"""
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT u.id, u.login 
                FROM users u
                WHERE u.id = %s AND u.role = 'operator' AND u.supervisor_id = %s
            """, (operator_id, supervisor_id))
            return cursor.fetchone()

    def migrate_daily_fines_from_daily_hours(self):
        """
        Migration helper: convert existing aggregated fine_amount/fine_reason/fine_comment
        from daily_hours into individual rows in daily_fines when no daily_fines exist for that day.
        This is idempotent: it will skip days that already have entries in daily_fines.
        """
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, operator_id, day, fine_amount, fine_reason, fine_comment
                FROM daily_hours
                WHERE COALESCE(fine_amount, 0) > 0
            """)
            rows = cursor.fetchall()
            for dh_id, op_id, day_obj, fine_amount, fine_reason, fine_comment in rows:
                # skip if fines already exist for this daily_hours
                cursor.execute("SELECT 1 FROM daily_fines WHERE daily_hours_id = %s LIMIT 1", (dh_id,))
                if cursor.fetchone():
                    continue
                # insert single aggregated fine as one row
                cursor.execute(
                    "INSERT INTO daily_fines (daily_hours_id, operator_id, day, amount, reason, comment) VALUES (%s, %s, %s, %s, %s, %s)",
                    (dh_id, op_id, day_obj, float(fine_amount or 0.0), fine_reason, fine_comment)
                )

    def update_user_password(self, user_id, new_password):
        password_hash = pbkdf2_sha256.hash(new_password)
        with self._get_cursor() as cursor:
            cursor.execute("""
                UPDATE users SET password_hash = %s
                WHERE id = %s
                RETURNING id
            """, (password_hash, user_id))
            return cursor.fetchone() is not None
    
    def update_operator_login(self, operator_id, supervisor_id, new_login):
        """Обновить логин оператора с проверкой принадлежности"""
        if supervisor_id:
            with self._get_cursor() as cursor:
                cursor.execute("""
                    UPDATE users SET login = %s
                    WHERE id = %s AND supervisor_id = %s
                    RETURNING id
                """, (new_login, operator_id, supervisor_id))
                return cursor.fetchone() is not None
        else:
            with self._get_cursor() as cursor:
                cursor.execute("""
                    UPDATE users SET login = %s
                    WHERE id = %s
                    RETURNING id
                """, (new_login, operator_id))
                return cursor.fetchone() is not None
    
    def update_operator_password(self, operator_id, supervisor_id, new_password):
        """Обновить пароль оператора с проверкой принадлежности"""
        password_hash = pbkdf2_sha256.hash(new_password)
        with self._get_cursor() as cursor:
            cursor.execute("""
                UPDATE users SET password_hash = %s
                WHERE id = %s AND role = 'operator' AND supervisor_id = %s
                RETURNING id
            """, (password_hash, operator_id, supervisor_id))
            return cursor.fetchone() is not None
            
    def get_user(self, **kwargs):
        conditions = []
        params = []
        for key, value in kwargs.items():
            if key == 'direction':
                conditions.append("d.name = %s")
                params.append(value)
            else:
                conditions.append(f"u.{key} = %s")
                params.append(value)
        
        query = f"""
            SELECT u.id, u.telegram_id, u.name, u.role, d.name, u.hire_date, u.supervisor_id, u.login, u.hours_table_url, u.scores_table_url, u.is_active, u.status, u.rate, u.gender, u.birth_date, u.avatar_bucket, u.avatar_blob_path, u.avatar_content_type, u.avatar_file_size, u.avatar_updated_at
            FROM users u
            LEFT JOIN directions d ON u.direction_id = d.id
            WHERE {' AND '.join(conditions)}
        """
        
        with self._get_cursor() as cursor:
            cursor.execute(query, params)
            return cursor.fetchone()

    def get_birthdays_for_date(self, day_value):
        """Return list of active users with birthdays matching the given date (month/day)."""
        day = None
        if isinstance(day_value, str):
            try:
                day = datetime.strptime(day_value, "%Y-%m-%d").date()
            except ValueError:
                raise ValueError("Invalid date format. Use YYYY-MM-DD")
        elif isinstance(day_value, datetime):
            day = day_value.date()
        else:
            day = day_value

        if not day:
            return []

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    u.id,
                    u.name,
                    u.role,
                    d.name AS direction,
                    u.status,
                    u.birth_date
                FROM users u
                LEFT JOIN directions d ON u.direction_id = d.id
                WHERE u.birth_date IS NOT NULL
                  AND EXTRACT(MONTH FROM u.birth_date) = %s
                  AND EXTRACT(DAY FROM u.birth_date) = %s
                  AND (u.status IS NULL OR u.status NOT IN ('fired', 'dismissal'))
            """, (day.month, day.day))
            rows = cursor.fetchall()

        return [
            {
                "id": row[0],
                "name": row[1],
                "role": row[2],
                "direction": row[3],
                "status": row[4],
                "birth_date": row[5].strftime('%Y-%m-%d') if row[5] else None
            }
            for row in rows
        ]

    def get_call_by_id(self, call_id):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, operator_id, phone_number, audio_path
                FROM calls
                WHERE id = %s
            """, (call_id,))
            row = cursor.fetchone()
            if row:
                return {
                    "id": row[0],
                    "operator_id": row[1],
                    "phone_number": row[2],
                    "audio_path": row[3]
                }
            return None
        
    def update_user_table(self, user_id, hours_table_url=None, scores_table_url=None):
        with self._get_cursor() as cursor:
            cursor.execute("""
                UPDATE users 
                SET hours_table_url = COALESCE(%s, hours_table_url),
                    scores_table_url = COALESCE(%s, scores_table_url)
                WHERE id = %s
            """, (hours_table_url, scores_table_url, user_id))

    def add_call_evaluation(self,
                            evaluator_id,
                            operator_id,
                            phone_number,
                            score,
                            comment=None,
                            comment_visible_to_operator=True,
                            month=None,
                            audio_path=None,
                            is_draft=False,
                            scores=None,
                            criterion_comments=None,
                            direction_id=None,
                            is_correction=False,
                            previous_version_id=None,
                            appeal_date=None,
                            external_id=None):   # <-- NEW optional param
        """
        Создаёт/обновляет запись в calls и, если возможно, помечает связанную запись в imported_calls как evaluated.
        """
        month = month or datetime.now().strftime('%Y-%m')

        # Подготовка JSON данных один раз
        scores_json = json.dumps(scores) if scores else None
        criterion_comments_json = json.dumps(criterion_comments) if criterion_comments else None

        with self._get_cursor() as cursor:
            # Проверка существования direction_id одним запросом
            if direction_id:
                cursor.execute("SELECT 1 FROM directions WHERE id = %s AND is_active = TRUE", (direction_id,))
                if not cursor.fetchone():
                    direction_id = None

            if is_correction and previous_version_id and not audio_path:
                cursor.execute("SELECT audio_path FROM calls WHERE id = %s", (previous_version_id,))
                result = cursor.fetchone()
                if result and result[0]:
                    audio_path = result[0]

            # Объединенная проверка черновика и существующей оценки
            cursor.execute("""
                WITH existing_data AS (
                    -- Проверка на существующий черновик
                    SELECT id, audio_path, FALSE AS is_existing_eval, created_at
                    FROM calls 
                    WHERE evaluator_id = %s 
                    AND operator_id = %s 
                    AND month = %s 
                    AND is_draft = TRUE
                    
                    UNION ALL
                    
                    -- Проверка на существующую оценку (не черновик)
                    SELECT id, NULL AS audio_path, TRUE AS is_existing_eval, created_at
                    FROM calls 
                    WHERE operator_id = %s 
                    AND month = %s 
                    AND phone_number = %s 
                    AND is_draft = FALSE
                )
                SELECT id, audio_path, is_existing_eval 
                FROM existing_data
                ORDER BY created_at DESC
                LIMIT 1
            """, (evaluator_id, operator_id, month, operator_id, month, phone_number))

            existing_record = cursor.fetchone()
            old_audio_path = None
            call_id = None

            if existing_record and not is_correction:
                call_id, old_audio_path, is_existing_eval = existing_record

                # Обновление существующей записи
                if is_draft:
                    # Для черновика обновляем все поля
                    cursor.execute("""
                        UPDATE calls
                        SET phone_number = %s,
                            score = %s,
                            comment = %s,
                            comment_visible_to_operator = %s,
                            audio_path = COALESCE(%s, audio_path),
                            created_at = CURRENT_TIMESTAMP,
                            scores = %s,
                            criterion_comments = %s,
                            is_correction = %s,
                            direction_id = %s
                        WHERE id = %s
                        RETURNING id
                    """, (
                        phone_number, score, comment, bool(comment_visible_to_operator), audio_path,
                        scores_json, criterion_comments_json,
                        False,
                        direction_id,
                        call_id
                    ))
                    call_id = cursor.fetchone()[0]

                    # Удаляем старый аудиофайл если он был заменен
                    if old_audio_path and audio_path and old_audio_path != audio_path:
                        try:
                            bucket_name = os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET')
                            client = get_gcs_client()
                            bucket = client.bucket(bucket_name)
                            blob = bucket.blob(old_audio_path.replace("my-app-audio-uploads/", ""))
                            blob.delete()
                        except Exception as e:
                            logging.error(f"Error removing old audio file: {str(e)}")

                    # Перед возвратом — пометить imported_calls (если найдено)
                    try:
                        if external_id:
                            cursor.execute("""
                                UPDATE imported_calls
                                SET status = 'evaluated',
                                    evaluated_at = CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty',
                                    evaluated_by = %s
                                WHERE external_id = %s AND month = %s AND status != 'evaluated'
                            """, (evaluator_id, external_id, month))
                        else:
                            # попытаться найти по телефону + дате (точное совпадение)
                            if phone_number and appeal_date:
                                phone_norm = _normalize_phone(phone_number)
                                cursor.execute("""
                                    UPDATE imported_calls
                                    SET status = 'evaluated',
                                        evaluated_at = CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty',
                                        evaluated_by = %s
                                    WHERE phone_normalized = %s AND month = %s AND datetime_raw = %s AND status != 'evaluated'
                                """, (evaluator_id, phone_norm, month, appeal_date))
                    except Exception as e:
                        logging.exception("Error marking imported_call as evaluated: %s", e)

                    return call_id

            # Создаем новую запись (для новой оценки, переоценки или если нет черновика)
            cursor.execute("""
                INSERT INTO calls (
                    evaluator_id, operator_id, month, phone_number, score, comment,
                    comment_visible_to_operator, audio_path, is_draft, is_correction, previous_version_id,
                    scores, criterion_comments, direction_id, appeal_date
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                evaluator_id, operator_id, month, phone_number, score, comment,
                bool(comment_visible_to_operator), audio_path, is_draft, is_correction, previous_version_id,
                scores_json, criterion_comments_json, direction_id, appeal_date
            ))
            call_id = cursor.fetchone()[0]

            # Удаляем старый аудиофайл, если он был заменен
            if old_audio_path and audio_path and old_audio_path != audio_path:
                try:
                    bucket_name = os.getenv('GOOGLE_CLOUD_STORAGE_BUCKET')
                    client = get_gcs_client()
                    bucket = client.bucket(bucket_name)
                    blob = bucket.blob(old_audio_path.replace("my-app-audio-uploads/", ""))
                    blob.delete()
                except Exception as e:
                    logging.error(f"Error removing old audio file: {str(e)}")

            # После успешного создания — помечаем imported_calls как evaluated (если можем сопоставить)
            try:
                if external_id:
                    cursor.execute("""
                        UPDATE imported_calls
                        SET status = 'evaluated',
                            evaluated_at = CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty',
                            evaluated_by = %s
                        WHERE external_id = %s AND month = %s AND status != 'evaluated'
                    """, (evaluator_id, external_id, month))
                else:
                    if phone_number and appeal_date:
                        phone_norm = _normalize_phone(phone_number)
                        cursor.execute("""
                            UPDATE imported_calls
                            SET status = 'evaluated',
                                evaluated_at = CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty',
                                evaluated_by = %s
                            WHERE phone_normalized = %s AND month = %s AND datetime_raw = %s AND status != 'evaluated'
                        """, (evaluator_id, phone_norm, month, appeal_date))
            except Exception as e:
                logging.exception("Error marking imported_call as evaluated: %s", e)

            return call_id
    
    def import_calls_from_distribution(self, payload: dict, importer_id: Optional[int] = None) -> dict:
        """
        Импортирует JSON payload в imported_calls.
        Возвращает отчёт: {'imported': n, 'updated': m, 'skipped': k, 'errors': [...], 'missing_operators': [...]}
        """
        month = payload.get('month')
        if not month:
            raise ValueError("Payload must include 'month' (YYYY-MM)")

        imported = 0
        updated = 0
        skipped = 0
        errors = []
        missing_operators = set()

        with self._get_cursor() as cur:
            for op in payload.get('distribution', []):
                op_name = op.get('operator')
                desired = op.get('desired')
                available = op.get('available')

                cur.execute("SELECT id FROM users WHERE name = %s LIMIT 1", (op_name,))
                r = cur.fetchone()
                operator_id = r[0] if r else None
                if not operator_id:
                    missing_operators.add(op_name)

                for c in op.get('calls', []):
                    external_id = c.get('id')
                    dt_raw_str = c.get('datetimeRaw')
                    parsed_dt = _parse_datetime_raw(dt_raw_str)
                    phone = c.get('phone')
                    phone_norm = _normalize_phone(phone)
                    duration = c.get('durationSec')

                    try:
                        cur.execute("""
                            INSERT INTO imported_calls
                            (external_id, operator_name, operator_id, month, datetime_raw,
                            phone_number, phone_normalized, duration_sec, desired, available, status, imported_at)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'not_evaluated', CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                            ON CONFLICT (external_id, month) DO UPDATE
                            SET operator_name = EXCLUDED.operator_name,
                                operator_id = COALESCE(EXCLUDED.operator_id, imported_calls.operator_id),
                                datetime_raw = COALESCE(EXCLUDED.datetime_raw, imported_calls.datetime_raw),
                                phone_number = COALESCE(EXCLUDED.phone_number, imported_calls.phone_number),
                                phone_normalized = COALESCE(EXCLUDED.phone_normalized, imported_calls.phone_normalized),
                                duration_sec = COALESCE(EXCLUDED.duration_sec, imported_calls.duration_sec),
                                desired = COALESCE(EXCLUDED.desired, imported_calls.desired),
                                available = COALESCE(EXCLUDED.available, imported_calls.available),
                                status = COALESCE(imported_calls.status, 'not_evaluated'),
                                imported_at = CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'
                            """,
                            (external_id, op_name, operator_id, month, parsed_dt, phone, phone_norm, duration, desired, available)
                        )
                        # cur.rowcount может быть ненадёжен для ON CONFLICT; просто учитываем как imported/updated логически
                        imported += 1
                    except Exception as exc:
                        errors.append({"external_id": external_id, "error": str(exc)})
                        skipped += 1

        return {
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
            "missing_operators": list(missing_operators)
        }

    def list_imported_calls(self, month: str, status: Optional[str] = None, operator_name: Optional[str] = None, limit: int = 200):
        q = "SELECT id, external_id, operator_name, operator_id, datetime_raw, phone_number, phone_normalized, duration_sec, desired, available, status, evaluated_at, evaluated_by, notes FROM imported_calls WHERE month = %s"
        params = [month]
        if status:
            q += " AND status = %s"
            params.append(status)
        if operator_name:
            q += " AND operator_name = %s"
            params.append(operator_name)
        q += " ORDER BY datetime_raw NULLS LAST LIMIT %s"
        params.append(limit)
        with self._get_cursor() as cur:
            cur.execute(q, tuple(params))
            rows = cur.fetchall()
        keys = ["id","external_id","operator_name","operator_id","datetime_raw","phone_number","phone_normalized","duration_sec","desired","available","status","evaluated_at","evaluated_by","notes"]
        return [dict(zip(keys, r)) for r in rows]

    def mark_imported_call_skipped(self, imported_call_id: str, reason: Optional[str] = None):
        with self._get_cursor() as cur:
            cur.execute("UPDATE imported_calls SET status = 'skipped', notes = COALESCE(notes,'') || %s WHERE id = %s",
                        (f"\nskipped: {reason}" if reason else "\nskipped", imported_call_id))
            return cur.rowcount == 1

    def parse_calls_file(self, file, target_date: Optional[str] = None):
        try:
            filename = str(file.filename or "").lower()

            if filename.endswith(".csv"):
                try:
                    df = pd.read_csv(file, sep=None, engine="python")
                except Exception:
                    file.stream.seek(0)
                    df = pd.read_csv(file, sep=";")
                sheet_name = "CSV"
            elif filename.endswith((".xls", ".xlsx")):
                excel = pd.ExcelFile(file)
                sheet_name = excel.sheet_names[0]
                df = pd.read_excel(excel, sheet_name=sheet_name)
            else:
                return None, None, "Неверный формат файла. Поддерживаются только CSV, XLS, XLSX"

            if df is None or df.empty:
                return None, None, "Файл пустой или не содержит данных"

            df.columns = [str(col).strip() for col in df.columns]

            def _norm_header(value):
                txt = str(value or "").strip().lower()
                txt = txt.replace("ё", "е")
                txt = re.sub(r"\s+", " ", txt)
                return txt

            normalized_to_original = {_norm_header(col): col for col in df.columns}

            def _pick_column(candidates):
                # exact normalized match
                for cand in candidates:
                    found = normalized_to_original.get(_norm_header(cand))
                    if found:
                        return found
                # soft contains match
                candidate_norms = [_norm_header(c) for c in candidates]
                for col in df.columns:
                    col_norm = _norm_header(col)
                    for cand_norm in candidate_norms:
                        if cand_norm and (cand_norm in col_norm or col_norm in cand_norm):
                            return col
                return None

            name_col = _pick_column(["ФИО", "ФИО оператора", "Оператор", "Name"])
            date_col = _pick_column(["Дата", "Date", "День", "Дата звонка"])
            calls_col = _pick_column([
                "Кол-во поступивших",
                "Количество поступивших"
            ])

            if not name_col or not date_col or not calls_col:
                return None, None, (
                    "Ожидаются колонки: ФИО, Дата, Кол-во поступивших "
                    f"(найдены: {', '.join(map(str, df.columns))})"
                )

            target_date_obj = None
            if target_date:
                try:
                    target_date_obj = datetime.strptime(str(target_date), "%Y-%m-%d").date()
                except Exception:
                    return None, None, "Некорректный формат выбранной даты. Ожидается YYYY-MM-DD"

            def _parse_row_date(raw_value):
                if raw_value is None or (isinstance(raw_value, float) and pd.isna(raw_value)):
                    return None
                if isinstance(raw_value, pd.Timestamp):
                    return raw_value.date()
                if isinstance(raw_value, datetime):
                    return raw_value.date()
                if isinstance(raw_value, date):
                    return raw_value

                text = str(raw_value).strip()
                if not text or text.lower() == "nan":
                    return None

                for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
                    try:
                        return datetime.strptime(text, fmt).date()
                    except Exception:
                        pass

                parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
                if pd.isna(parsed):
                    return None
                if isinstance(parsed, pd.Timestamp):
                    return parsed.date()
                return None

            def _parse_calls(raw_value):
                if raw_value is None or (isinstance(raw_value, float) and pd.isna(raw_value)):
                    return 0
                if isinstance(raw_value, (int, float)) and not pd.isna(raw_value):
                    try:
                        return max(int(round(float(raw_value))), 0)
                    except Exception:
                        return 0
                text = str(raw_value).strip().replace(" ", "").replace("\xa0", "")
                if not text or text.lower() == "nan":
                    return 0
                text = text.replace(",", ".")
                try:
                    return max(int(round(float(text))), 0)
                except Exception:
                    digits = re.sub(r"[^\d\-]", "", text)
                    try:
                        return max(int(digits), 0)
                    except Exception:
                        return 0

            aggregated = {}
            skipped_other_date = 0

            for _, row in df.iterrows():
                name_val = row.get(name_col)
                name = str(name_val or "").strip()
                if not name or name.lower() == "nan":
                    continue

                row_date_obj = _parse_row_date(row.get(date_col))
                if not row_date_obj:
                    continue

                if target_date_obj and row_date_obj != target_date_obj:
                    skipped_other_date += 1
                    continue

                calls_value = _parse_calls(row.get(calls_col))
                date_str = row_date_obj.strftime("%Y-%m-%d")
                key = (name, date_str)
                if key not in aggregated:
                    aggregated[key] = {
                        "name": name,
                        "date": date_str,
                        "calls": int(calls_value),
                        "month": date_str[:7]
                    }
                else:
                    aggregated[key]["calls"] = int(aggregated[key]["calls"]) + int(calls_value)

            operators = sorted(
                aggregated.values(),
                key=lambda item: (item.get("date") or "", str(item.get("name") or "").lower())
            )

            if not operators:
                if target_date_obj and skipped_other_date > 0:
                    return None, None, (
                        f"В файле нет строк за выбранную дату {target_date_obj.strftime('%Y-%m-%d')}"
                    )
                return None, None, "Не удалось распознать строки с данными ФИО/Дата/Кол-во звонков"

            return sheet_name, operators, None

        except Exception as e:
            return None, None, str(e)


    def _get_month_end_date(self, month: Optional[str]) -> date:
        month_str = str(month or '').strip()
        try:
            year_str, month_part = month_str.split('-', 1)
            year_value = int(year_str)
            month_value = int(month_part)
            last_day = calendar.monthrange(year_value, month_value)[1]
            return date(year_value, month_value, last_day)
        except Exception:
            today = datetime.now().date()
            last_day = calendar.monthrange(today.year, today.month)[1]
            return date(today.year, today.month, last_day)

    def _get_month_date_range(self, month: Optional[str]) -> Tuple[date, date]:
        month_end_date = self._get_month_end_date(month)
        month_start_date = date(month_end_date.year, month_end_date.month, 1)
        if month_end_date.month == 12:
            next_month_start = date(month_end_date.year + 1, 1, 1)
        else:
            next_month_start = date(month_end_date.year, month_end_date.month + 1, 1)
        return month_start_date, next_month_start

    def _get_month_work_days(self, month: Optional[str]) -> int:
        month_start_date, next_month_start = self._get_month_date_range(month)
        work_days = 0
        current_date = month_start_date
        while current_date < next_month_start:
            if current_date.weekday() < 5:
                work_days += 1
            current_date += timedelta(days=1)
        return int(work_days)

    def _get_full_rate_norm_hours(self, month: Optional[str]) -> float:
        return float(self._get_month_work_days(month) * 8.0)

    def _get_tenure_months_for_reference(self, hire_date_value, reference_date: Optional[date]) -> Optional[int]:
        if not hire_date_value or not reference_date:
            return None

        if isinstance(hire_date_value, datetime):
            hire_date_obj = hire_date_value.date()
        elif isinstance(hire_date_value, date):
            hire_date_obj = hire_date_value
        elif isinstance(hire_date_value, str):
            hire_date_obj = None
            for date_format in ('%Y-%m-%d', '%d-%m-%Y'):
                try:
                    hire_date_obj = datetime.strptime(hire_date_value, date_format).date()
                    break
                except ValueError:
                    continue
            if hire_date_obj is None:
                return None
        else:
            return None

        months = (reference_date.year - hire_date_obj.year) * 12 + (reference_date.month - hire_date_obj.month)
        if reference_date.day < hire_date_obj.day:
            months -= 1
        return max(int(months), 0)

    def _get_base_call_target_by_tenure(self, tenure_months: Optional[int]) -> int:
        if tenure_months is None:
            return 20
        if tenure_months <= 2:
            return 20
        if tenure_months <= 5:
            return 10
        return 5

    def _fetch_operator_call_evaluation_target_source(self, cursor, operator_id: int, target_month: str):
        month_start_date, next_month_start = self._get_month_date_range(target_month)
        cursor.execute("""
            SELECT
                u.hire_date,
                COALESCE(wh.regular_hours, 0) AS regular_hours,
                COALESCE(wh.norm_hours, 0) AS norm_hours,
                COALESCE((
                    SELECT SUM(
                        CASE
                            WHEN t.end_time <= t.start_time
                                THEN EXTRACT(EPOCH FROM (t.end_time + INTERVAL '24 hours' - t.start_time))
                            ELSE EXTRACT(EPOCH FROM (t.end_time - t.start_time))
                        END
                    ) / 3600.0
                    FROM trainings t
                    WHERE t.operator_id = u.id
                      AND t.count_in_hours = TRUE
                      AND t.training_date >= %s
                      AND t.training_date < %s
                ), 0) AS training_hours,
                COALESCE((
                    SELECT SUM(
                        CASE
                            WHEN ti.end_time <= ti.start_time
                                THEN EXTRACT(EPOCH FROM (ti.end_time + INTERVAL '24 hours' - ti.start_time))
                            ELSE EXTRACT(EPOCH FROM (ti.end_time - ti.start_time))
                        END
                    ) / 3600.0
                    FROM operator_technical_issues ti
                    WHERE ti.operator_id = u.id
                      AND ti.issue_date >= %s
                      AND ti.issue_date < %s
                ), 0) AS technical_issue_hours,
                COALESCE((
                    SELECT SUM(
                        CASE
                            WHEN oa.end_time <= oa.start_time
                                THEN EXTRACT(EPOCH FROM (oa.end_time + INTERVAL '24 hours' - oa.start_time))
                            ELSE EXTRACT(EPOCH FROM (oa.end_time - oa.start_time))
                        END
                    ) / 3600.0
                    FROM operator_offline_activities oa
                    WHERE oa.operator_id = u.id
                      AND oa.activity_date >= %s
                      AND oa.activity_date < %s
                ), 0) AS offline_activity_hours
            FROM users u
            LEFT JOIN work_hours wh
                ON wh.operator_id = u.id
               AND wh.month = %s
            WHERE u.id = %s
            LIMIT 1
        """, (
            month_start_date,
            next_month_start,
            month_start_date,
            next_month_start,
            month_start_date,
            next_month_start,
            target_month,
            operator_id
        ))
        row = cursor.fetchone()
        if not row:
            return {
                'hire_date_value': None,
                'regular_hours': 0.0,
                'operator_norm_hours': 0.0,
                'training_hours': 0.0,
                'technical_issue_hours': 0.0,
                'offline_activity_hours': 0.0,
            }

        return {
            'hire_date_value': row[0],
            'regular_hours': float(row[1] or 0.0),
            'operator_norm_hours': float(row[2] or 0.0),
            'training_hours': float(row[3] or 0.0),
            'technical_issue_hours': float(row[4] or 0.0),
            'offline_activity_hours': float(row[5] or 0.0),
        }

    def _build_operator_call_evaluation_target(
        self,
        operator_id: int,
        target_month: str,
        hire_date_value=None,
        regular_hours: float = 0.0,
        training_hours: float = 0.0,
        technical_issue_hours: float = 0.0,
        offline_activity_hours: float = 0.0,
        operator_norm_hours: float = 0.0
    ):
        month_end_date = self._get_month_end_date(target_month)
        regular_hours_value = float(regular_hours or 0.0)
        training_hours_value = float(training_hours or 0.0)
        technical_issue_hours_value = float(technical_issue_hours or 0.0)
        offline_activity_hours_value = float(offline_activity_hours or 0.0)
        accounted_hours_value = (
            regular_hours_value
            + training_hours_value
            + technical_issue_hours_value
            + offline_activity_hours_value
        )
        operator_norm_hours_value = float(operator_norm_hours or 0.0)
        work_days_value = int(self._get_month_work_days(target_month))
        full_rate_norm_hours_value = float(self._get_full_rate_norm_hours(target_month) or 0.0)

        tenure_months = self._get_tenure_months_for_reference(hire_date_value, month_end_date)
        base_call_target = self._get_base_call_target_by_tenure(tenure_months)
        worked_hours_ratio = (
            accounted_hours_value / full_rate_norm_hours_value
        ) if full_rate_norm_hours_value > 0 else 0.0
        required_calls_raw = float(worked_hours_ratio) * float(base_call_target)
        required_calls = int(math.ceil(required_calls_raw - 1e-9)) if required_calls_raw > 0 else 0

        if tenure_months is None:
            tenure_bucket = 'unknown'
        elif tenure_months <= 2:
            tenure_bucket = '0_2_months'
        elif tenure_months <= 5:
            tenure_bucket = '3_5_months'
        else:
            tenure_bucket = '6_plus_months'

        return {
            'operator_id': int(operator_id),
            'month': target_month,
            'reference_date': month_end_date.isoformat(),
            'tenure_months': tenure_months,
            'tenure_bucket': tenure_bucket,
            'base_call_target': int(base_call_target),
            'worked_hours_for_calls': round(float(regular_hours_value), 2),
            'training_hours': round(float(training_hours_value), 2),
            'technical_issue_hours': round(float(technical_issue_hours_value), 2),
            'offline_activity_hours': round(float(offline_activity_hours_value), 2),
            'accounted_hours': round(float(accounted_hours_value), 2),
            'worked_hours_used': round(float(accounted_hours_value), 2),
            'work_days': int(work_days_value),
            'operator_norm_hours': round(float(operator_norm_hours_value), 2),
            'full_rate_norm_hours': round(float(full_rate_norm_hours_value), 2),
            'norm_hours': round(float(full_rate_norm_hours_value), 2),
            'worked_hours_ratio': round(float(worked_hours_ratio), 4),
            'required_calls_raw': round(float(required_calls_raw), 2),
            'required_calls': int(required_calls),
        }

    def get_operator_call_evaluation_target(self, operator_id: int, month: Optional[str] = None):
        target_month = str(month or datetime.now().strftime('%Y-%m'))

        with self._get_cursor() as cursor:
            source = self._fetch_operator_call_evaluation_target_source(cursor, operator_id, target_month)

        return self._build_operator_call_evaluation_target(
            operator_id=operator_id,
            target_month=target_month,
            hire_date_value=source.get('hire_date_value'),
            regular_hours=source.get('regular_hours'),
            training_hours=source.get('training_hours'),
            technical_issue_hours=source.get('technical_issue_hours'),
            offline_activity_hours=source.get('offline_activity_hours'),
            operator_norm_hours=source.get('operator_norm_hours')
        )

    def get_operator_call_evaluation_targets_for_month(self, operator_ids, month: Optional[str] = None):
        target_month = str(month or datetime.now().strftime('%Y-%m'))

        normalized_ids = []
        seen_ids = set()
        for raw_id in (operator_ids or []):
            try:
                op_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if op_id in seen_ids:
                continue
            seen_ids.add(op_id)
            normalized_ids.append(op_id)

        if not normalized_ids:
            return {}

        result = {}
        with self._get_cursor() as cursor:
            for op_id in normalized_ids:
                source = self._fetch_operator_call_evaluation_target_source(cursor, op_id, target_month)
                result[op_id] = self._build_operator_call_evaluation_target(
                    operator_id=op_id,
                    target_month=target_month,
                    hire_date_value=source.get('hire_date_value'),
                    regular_hours=source.get('regular_hours'),
                    training_hours=source.get('training_hours'),
                    technical_issue_hours=source.get('technical_issue_hours'),
                    offline_activity_hours=source.get('offline_activity_hours'),
                    operator_norm_hours=source.get('operator_norm_hours')
                )

        return result

    def get_operator_stats(self, operator_id):
        """Получить статистику оператора (часы, оценки, звонки в час, тренинги)"""
        with self._get_cursor() as cursor:
            # Текущий месяц в формате YYYY-MM
            current_month = datetime.now().strftime('%Y-%m')

            # 1) Берём агрегаты из work_hours (regular_hours, total_calls, fines, norm_hours)
            cursor.execute("""
                SELECT 
                    COALESCE(wh.regular_hours, 0) AS regular_hours,
                    COALESCE(wh.fines, 0) AS fines,
                    COALESCE(wh.norm_hours, 0) AS norm_hours,
                    COALESCE(wh.total_calls, 0) AS total_calls
                FROM work_hours wh
                WHERE wh.operator_id = %s AND wh.month = %s
                LIMIT 1
            """, (operator_id, current_month))
            wh_row = cursor.fetchone()

            if wh_row:
                regular_hours, fines, norm_hours, total_calls = wh_row
                regular_hours = float(regular_hours or 0.0)
                fines = float(fines or 0.0)
                norm_hours = float(norm_hours or 0.0)
                total_calls = int(total_calls or 0)
            else:
                regular_hours = 0.0
                fines = 0.0
                norm_hours = 0.0
                total_calls = 0

            # 2) Рассчитываем training_hours прямо по таблице trainings (учитываем count_in_hours = TRUE)
            cursor.execute("""
                SELECT COALESCE(SUM(
                    CASE 
                    WHEN t.end_time <= t.start_time 
                        THEN EXTRACT(EPOCH FROM (t.end_time + INTERVAL '24 hours' - t.start_time))
                    ELSE EXTRACT(EPOCH FROM (t.end_time - t.start_time))
                    END
                ) / 3600.0, 0) AS training_hours_seconds
                FROM trainings t
                WHERE t.operator_id = %s
                AND TO_CHAR(t.training_date, 'YYYY-MM') = %s
                AND t.count_in_hours = TRUE
            """, (operator_id, current_month))
            tr_row = cursor.fetchone()
            training_hours = float(tr_row[0] or 0.0)

            # 3) Рассчитываем техсбои за месяц (учитываются в часах выполнения нормы)
            cursor.execute("""
                SELECT COALESCE(SUM(
                    CASE
                    WHEN ti.end_time <= ti.start_time
                        THEN EXTRACT(EPOCH FROM (ti.end_time + INTERVAL '24 hours' - ti.start_time))
                    ELSE EXTRACT(EPOCH FROM (ti.end_time - ti.start_time))
                    END
                ) / 3600.0, 0) AS technical_issue_hours_seconds
                FROM operator_technical_issues ti
                WHERE ti.operator_id = %s
                AND TO_CHAR(ti.issue_date, 'YYYY-MM') = %s
            """, (operator_id, current_month))
            tech_row = cursor.fetchone()
            technical_issue_hours = float(tech_row[0] or 0.0)

            # 3.1) Рассчитываем офлайн-активность за месяц
            cursor.execute("""
                SELECT COALESCE(SUM(
                    CASE
                    WHEN oa.end_time <= oa.start_time
                        THEN EXTRACT(EPOCH FROM (oa.end_time + INTERVAL '24 hours' - oa.start_time))
                    ELSE EXTRACT(EPOCH FROM (oa.end_time - oa.start_time))
                    END
                ) / 3600.0, 0) AS offline_activity_hours_seconds
                FROM operator_offline_activities oa
                WHERE oa.operator_id = %s
                AND TO_CHAR(oa.activity_date, 'YYYY-MM') = %s
            """, (operator_id, current_month))
            offline_row = cursor.fetchone()
            offline_activity_hours = float(offline_row[0] or 0.0)

            # 4) Количество оценённых звонков и средняя оценка (как раньше)
            cursor.execute("""
                SELECT 
                    (SELECT COUNT(*) FROM calls WHERE operator_id = %s AND month = %s AND is_draft = FALSE) AS call_count,
                    (SELECT AVG(score) FROM calls WHERE operator_id = %s AND month = %s AND is_draft = FALSE) AS avg_score
            """, (operator_id, current_month, operator_id, current_month))
            calls_row = cursor.fetchone()
            call_count = int(calls_row[0] or 0)
            avg_score = float(calls_row[1]) if calls_row[1] is not None else 0.0

            # 5) calls_per_hour: total_calls / regular_hours (время в работе из статусов).
            effective_call_hours = max(0.0, float(regular_hours))
            if effective_call_hours > 0:
                calls_per_hour = float(total_calls) / float(effective_call_hours)
            else:
                calls_per_hour = 0.0

            # 6) percent_complete: считаем с учётом зачётных тренингов, техсбоев и офлайн-активности.
            accounted_hours = regular_hours + training_hours + technical_issue_hours + offline_activity_hours
            if norm_hours and norm_hours > 0:
                percent_complete = (accounted_hours / norm_hours) * 100.0
            else:
                percent_complete = 0.0

            return {
                'operator_id': operator_id,
                'month': current_month,
                'regular_hours': round(float(regular_hours), 2),
                'training_hours': round(float(training_hours), 2),
                'technical_issue_hours': round(float(technical_issue_hours), 2),
                'offline_activity_hours': round(float(offline_activity_hours), 2),
                'accounted_hours': round(float(accounted_hours), 2),
                'fines': round(float(fines), 2),
                'norm_hours': float(norm_hours),
                'percent_complete': round(percent_complete, 2),
                'call_count': int(call_count),
                'avg_score': round(float(avg_score), 2),
                'total_calls': int(total_calls),
                'calls_per_hour': round(float(calls_per_hour), 2)
            }


    def get_operators_by_supervisor(self, supervisor_id):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    u.id,
                    u.name,
                    u.direction_id,
                    u.supervisor_id,
                    u.hire_date,
                    u.hours_table_url,
                    u.scores_table_url,
                    s.name as supervisor_name,
                    u.status,
                    u.rate,
                    u.gender,
                    u.birth_date,
                    u.avatar_bucket,
                    u.avatar_blob_path,
                    u.avatar_updated_at,
                    sp.status_code as status_period_status_code,
                    sp.start_date as status_period_start_date,
                    sp.end_date as status_period_end_date,
                    sp.dismissal_reason as status_period_dismissal_reason,
                    COALESCE(sp.is_blacklist, FALSE) as status_period_is_blacklist,
                    sp.comment as status_period_comment
                FROM users u
                LEFT JOIN directions d ON u.direction_id = d.id
                LEFT JOIN users s ON u.supervisor_id = s.id
                LEFT JOIN LATERAL (
                    SELECT
                        p.status_code,
                        p.start_date,
                        p.end_date,
                        p.dismissal_reason,
                        p.is_blacklist,
                        p.comment
                    FROM operator_schedule_status_periods p
                    WHERE p.operator_id = u.id
                      AND p.status_code = (
                          CASE
                              WHEN u.status = 'fired' THEN 'dismissal'
                              WHEN u.status = 'dismissal' THEN 'dismissal'
                              WHEN u.status = 'unpaid_leave' THEN 'bs'
                              ELSE u.status
                          END
                      )
                    ORDER BY
                        CASE
                            WHEN p.start_date <= CURRENT_DATE
                             AND COALESCE(p.end_date, DATE '9999-12-31') >= CURRENT_DATE
                            THEN 0
                            ELSE 1
                        END,
                        p.start_date DESC,
                        p.id DESC
                    LIMIT 1
                ) sp ON TRUE
                WHERE u.supervisor_id = %s AND u.role = 'operator'
            """, (supervisor_id,))
            return [
                {
                    'id': row[0],
                    'name': row[1],
                    'direction_id': row[2],
                    'supervisor_id': row[3],
                    'hire_date': row[4].strftime('%d-%m-%Y') if row[4] else None,
                    'hours_table_url': row[5],
                    'scores_table_url': row[6],
                    'supervisor_name': row[7],
                    'status': row[8],
                    'rate': row[9],
                    'gender': row[10],
                    'birth_date': row[11].strftime('%d-%m-%Y') if row[11] else None,
                    'avatar_bucket': row[12],
                    'avatar_blob_path': row[13],
                    'avatar_updated_at': row[14].isoformat() if row[14] else None,
                    'status_period_status_code': row[15],
                    'status_period_start_date': row[16].strftime('%Y-%m-%d') if row[16] else None,
                    'status_period_end_date': row[17].strftime('%Y-%m-%d') if row[17] else None,
                    'status_period_dismissal_reason': row[18] or '',
                    'status_period_is_blacklist': bool(row[19]) if row[19] is not None else False,
                    'status_period_comment': row[20] or ''
                } for row in cursor.fetchall()
            ]

    def get_all_operators_with_details(self):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    u.id,
                    u.name,
                    u.direction_id,
                    u.supervisor_id,
                    u.hire_date,
                    u.hours_table_url,
                    u.scores_table_url,
                    s.name as supervisor_name,
                    u.status,
                    u.rate,
                    u.gender,
                    u.birth_date,
                    u.avatar_bucket,
                    u.avatar_blob_path,
                    u.avatar_updated_at,
                    sp.status_code as status_period_status_code,
                    sp.start_date as status_period_start_date,
                    sp.end_date as status_period_end_date,
                    sp.dismissal_reason as status_period_dismissal_reason,
                    COALESCE(sp.is_blacklist, FALSE) as status_period_is_blacklist,
                    sp.comment as status_period_comment
                FROM users u
                LEFT JOIN directions d ON u.direction_id = d.id
                LEFT JOIN users s ON u.supervisor_id = s.id
                LEFT JOIN LATERAL (
                    SELECT
                        p.status_code,
                        p.start_date,
                        p.end_date,
                        p.dismissal_reason,
                        p.is_blacklist,
                        p.comment
                    FROM operator_schedule_status_periods p
                    WHERE p.operator_id = u.id
                      AND p.status_code = (
                          CASE
                              WHEN u.status = 'fired' THEN 'dismissal'
                              WHEN u.status = 'dismissal' THEN 'dismissal'
                              WHEN u.status = 'unpaid_leave' THEN 'bs'
                              ELSE u.status
                          END
                      )
                    ORDER BY
                        CASE
                            WHEN p.start_date <= CURRENT_DATE
                             AND COALESCE(p.end_date, DATE '9999-12-31') >= CURRENT_DATE
                            THEN 0
                            ELSE 1
                        END,
                        p.start_date DESC,
                        p.id DESC
                    LIMIT 1
                ) sp ON TRUE
                WHERE u.role = 'operator'
            """)
            return [
                {
                    'id': row[0],
                    'name': row[1],
                    'direction_id': row[2],
                    'supervisor_id': row[3],
                    'hire_date': row[4].strftime('%d-%m-%Y') if row[4] else None,
                    'hours_table_url': row[5],
                    'scores_table_url': row[6],
                    'supervisor_name': row[7],
                    'status': row[8],
                    'rate': row[9],
                    'gender': row[10],
                    'birth_date': row[11].strftime('%d-%m-%Y') if row[11] else None,
                    'avatar_bucket': row[12],
                    'avatar_blob_path': row[13],
                    'avatar_updated_at': row[14].isoformat() if row[14] else None,
                    'status_period_status_code': row[15],
                    'status_period_start_date': row[16].strftime('%Y-%m-%d') if row[16] else None,
                    'status_period_end_date': row[17].strftime('%Y-%m-%d') if row[17] else None,
                    'status_period_dismissal_reason': row[18] or '',
                    'status_period_is_blacklist': bool(row[19]) if row[19] is not None else False,
                    'status_period_comment': row[20] or ''
                } for row in cursor.fetchall()
            ]

    def get_activity_logs(self, operator_id, date_str=None):
        # Если date_str не указан, используем текущую дату
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        try:
            log_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError("Invalid date format. Use YYYY-MM-DD")
        
        with self._get_cursor() as cursor:
            # Берём последний статус до указанной даты
            cursor.execute("""
                SELECT change_time, is_active
                FROM operator_activity_logs
                WHERE operator_id = %s
                  AND change_time < %s
                ORDER BY change_time DESC
                LIMIT 1
            """, (operator_id, log_date))
            prev_log = cursor.fetchone()
    
            # Берём все записи за указанную дату
            cursor.execute("""
                SELECT change_time, is_active 
                FROM operator_activity_logs 
                WHERE operator_id = %s 
                  AND change_time::date = %s 
                ORDER BY change_time ASC
            """, (operator_id, log_date))
            logs = cursor.fetchall()
    
        result = []
        if prev_log:
            result.append({"change_time": prev_log[0].isoformat(), "is_active": prev_log[1]})
        result.extend([
            {"change_time": row[0].isoformat(), "is_active": row[1]}
            for row in logs
        ])
        return result
    
    def get_hours_summary(self, operator_id=None, month=None):
        query = """
            SELECT 
                u.id AS operator_id,
                u.name AS operator_name,
                u.direction_id,

                COALESCE(wh.regular_hours, 0) AS regular_hours,
                COALESCE(wh.fines, 0) AS fines,
                COALESCE(wh.norm_hours, 0) AS norm_hours,
                COALESCE(wh.total_calls, 0) AS total_calls,

                -- training hours считаем из trainings
                COALESCE((
                    SELECT SUM(
                        CASE
                            WHEN t.end_time <= t.start_time
                                THEN EXTRACT(EPOCH FROM (t.end_time + INTERVAL '24 hours' - t.start_time))
                            ELSE EXTRACT(EPOCH FROM (t.end_time - t.start_time))
                        END
                    ) / 3600.0
                    FROM trainings t
                    WHERE t.operator_id = u.id
                    AND t.count_in_hours = TRUE
                    AND (%s IS NULL OR TO_CHAR(t.training_date, 'YYYY-MM') = %s)
                ), 0) AS training_hours,

                -- technical issue hours считаем отдельно и учитываем в часах выполнения
                COALESCE((
                    SELECT SUM(
                        CASE
                            WHEN ti.end_time <= ti.start_time
                                THEN EXTRACT(EPOCH FROM (ti.end_time + INTERVAL '24 hours' - ti.start_time))
                            ELSE EXTRACT(EPOCH FROM (ti.end_time - ti.start_time))
                        END
                    ) / 3600.0
                    FROM operator_technical_issues ti
                    WHERE ti.operator_id = u.id
                    AND (%s IS NULL OR TO_CHAR(ti.issue_date, 'YYYY-MM') = %s)
                ), 0) AS technical_issue_hours,

                COALESCE((
                    SELECT SUM(
                        CASE
                            WHEN oa.end_time <= oa.start_time
                                THEN EXTRACT(EPOCH FROM (oa.end_time + INTERVAL '24 hours' - oa.start_time))
                            ELSE EXTRACT(EPOCH FROM (oa.end_time - oa.start_time))
                        END
                    ) / 3600.0
                    FROM operator_offline_activities oa
                    WHERE oa.operator_id = u.id
                    AND (%s IS NULL OR TO_CHAR(oa.activity_date, 'YYYY-MM') = %s)
                ), 0) AS offline_activity_hours

            FROM users u
            LEFT JOIN work_hours wh
                ON u.id = wh.operator_id
            AND (%s IS NULL OR wh.month = %s)
            WHERE u.role = 'operator'
        """

        params = [month, month, month, month, month, month, month, month]

        if operator_id:
            query += " AND u.id = %s"
            params.append(operator_id)

        with self._get_cursor() as cursor:
            cursor.execute(query, params)
            rows = cursor.fetchall()

            result = []
            for row in rows:
                regular_hours = float(row[3])
                total_calls = int(row[6])
                training_hours = round(float(row[7]), 2)
                technical_issue_hours = round(float(row[8]), 2)
                offline_activity_hours = round(float(row[9]), 2)
                accounted_hours = round(regular_hours + training_hours + technical_issue_hours + offline_activity_hours, 2)

                effective_call_hours = max(0.0, float(regular_hours))
                calls_per_hour = (
                    round(total_calls / effective_call_hours, 2)
                    if effective_call_hours > 0 else 0.0
                )

                result.append({
                    "operator_id": row[0],
                    "operator_name": row[1],
                    "direction": row[2],
                    "regular_hours": regular_hours,
                    "training_hours": training_hours,
                    "technical_issue_hours": technical_issue_hours,
                    "offline_activity_hours": offline_activity_hours,
                    "accounted_hours": accounted_hours,
                    "fines": float(row[4]),
                    "norm_hours": float(row[5]),
                    "calls_per_hour": calls_per_hour
                })

            return result

    def get_supervisors(self):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, name, hours_table_url, role, hire_date, status
                FROM users 
                WHERE role = 'sv'
            """)
            return cursor.fetchall()

    def get_user_by_login(self, login):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, telegram_id, name, role, direction_id, hire_date, supervisor_id, login, password_hash, hours_table_url, scores_table_url, avatar_bucket, avatar_blob_path, avatar_content_type, avatar_file_size, avatar_updated_at, gender
                FROM users WHERE login = %s
            """, (login,))
            return cursor.fetchone()

    def get_user_by_name(self, name):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, telegram_id, name, role, direction_id, hire_date, supervisor_id, login
                FROM users WHERE LOWER(name) = LOWER(%s) LIMIT 1
            """, (name,))
            return cursor.fetchone()

    def verify_password(self, user_id, password):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT password_hash FROM users WHERE id = %s
            """, (user_id,))
            result = cursor.fetchone()
            if result and result[0]:
                return pbkdf2_sha256.verify(password, result[0])
            return False

    def create_user_session(
        self,
        session_id,
        user_id,
        refresh_token_hash,
        expires_at,
        user_agent=None,
        ip_address=None
    ):
        with self._get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO user_sessions (
                    session_id,
                    user_id,
                    refresh_token_hash,
                    user_agent,
                    ip_address,
                    expires_at
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (session_id, user_id, refresh_token_hash, user_agent, ip_address, expires_at))

    def get_user_session(self, session_id, user_id=None):
        with self._get_cursor() as cursor:
            if user_id is None:
                cursor.execute("""
                    SELECT
                        session_id::text,
                        user_id,
                        refresh_token_hash,
                        user_agent,
                        ip_address,
                        created_at,
                        last_seen_at,
                        expires_at,
                        revoked_at,
                        sensitive_data_unlocked,
                        sensitive_data_unlocked_at
                    FROM user_sessions
                    WHERE session_id = %s
                """, (session_id,))
            else:
                cursor.execute("""
                    SELECT
                        session_id::text,
                        user_id,
                        refresh_token_hash,
                        user_agent,
                        ip_address,
                        created_at,
                        last_seen_at,
                        expires_at,
                        revoked_at,
                        sensitive_data_unlocked,
                        sensitive_data_unlocked_at
                    FROM user_sessions
                    WHERE session_id = %s AND user_id = %s
                """, (session_id, user_id))

            row = cursor.fetchone()
            if not row:
                return None
            return {
                "session_id": row[0],
                "user_id": row[1],
                "refresh_token_hash": row[2],
                "user_agent": row[3],
                "ip_address": row[4],
                "created_at": row[5],
                "last_seen_at": row[6],
                "expires_at": row[7],
                "revoked_at": row[8],
                "sensitive_data_unlocked": bool(row[9]),
                "sensitive_data_unlocked_at": row[10]
            }

    def rotate_user_session_token(self, session_id, user_id, refresh_token_hash, expires_at):
        with self._get_cursor() as cursor:
            cursor.execute("""
                UPDATE user_sessions
                SET refresh_token_hash = %s,
                    expires_at = %s,
                    last_seen_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                WHERE session_id = %s
                  AND user_id = %s
                  AND revoked_at IS NULL
                RETURNING session_id
            """, (refresh_token_hash, expires_at, session_id, user_id))
            return cursor.fetchone() is not None

    def touch_user_session(self, session_id, user_id=None, ip_address=None, user_agent=None):
        with self._get_cursor() as cursor:
            if user_id is None:
                cursor.execute("""
                    UPDATE user_sessions
                    SET last_seen_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                        ip_address = COALESCE(%s, ip_address),
                        user_agent = COALESCE(%s, user_agent)
                    WHERE session_id = %s
                      AND revoked_at IS NULL
                    RETURNING session_id
                """, (ip_address, user_agent, session_id))
            else:
                cursor.execute("""
                    UPDATE user_sessions
                    SET last_seen_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                        ip_address = COALESCE(%s, ip_address),
                        user_agent = COALESCE(%s, user_agent)
                    WHERE session_id = %s
                      AND user_id = %s
                      AND revoked_at IS NULL
                    RETURNING session_id
                """, (ip_address, user_agent, session_id, user_id))
            return cursor.fetchone() is not None

    def revoke_user_session(self, session_id, user_id=None):
        with self._get_cursor() as cursor:
            if user_id is None:
                cursor.execute("""
                    UPDATE user_sessions
                    SET revoked_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    WHERE session_id = %s
                      AND revoked_at IS NULL
                    RETURNING session_id
                """, (session_id,))
            else:
                cursor.execute("""
                    UPDATE user_sessions
                    SET revoked_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    WHERE session_id = %s
                      AND user_id = %s
                      AND revoked_at IS NULL
                    RETURNING session_id
                """, (session_id, user_id))
            return cursor.fetchone() is not None

    def revoke_all_user_sessions(self, user_id, except_session_id=None):
        with self._get_cursor() as cursor:
            if except_session_id:
                cursor.execute("""
                    UPDATE user_sessions
                    SET revoked_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    WHERE user_id = %s
                      AND session_id <> %s
                      AND revoked_at IS NULL
                """, (user_id, except_session_id))
            else:
                cursor.execute("""
                    UPDATE user_sessions
                    SET revoked_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    WHERE user_id = %s
                      AND revoked_at IS NULL
                """, (user_id,))
            return cursor.rowcount

    def set_session_sensitive_access(self, session_id, user_id, unlocked=True):
        with self._get_cursor() as cursor:
            cursor.execute("""
                UPDATE user_sessions
                SET sensitive_data_unlocked = %s,
                    sensitive_data_unlocked_at = CASE
                        WHEN %s THEN (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                        ELSE NULL
                    END
                WHERE session_id = %s
                  AND user_id = %s
                  AND revoked_at IS NULL
                RETURNING session_id
            """, (bool(unlocked), bool(unlocked), session_id, user_id))
            return cursor.fetchone() is not None

    def is_session_sensitive_access_unlocked(self, session_id, user_id):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT sensitive_data_unlocked
                FROM user_sessions
                WHERE session_id = %s
                  AND user_id = %s
                  AND revoked_at IS NULL
                  AND expires_at > (CURRENT_TIMESTAMP AT TIME ZONE 'UTC')
                LIMIT 1
            """, (session_id, user_id))
            row = cursor.fetchone()
            return bool(row[0]) if row else False

    def list_user_sessions(self, user_id):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    session_id::text,
                    user_agent,
                    ip_address,
                    created_at,
                    last_seen_at,
                    expires_at,
                    revoked_at
                FROM user_sessions
                WHERE user_id = %s
                ORDER BY created_at DESC
            """, (user_id,))
            rows = cursor.fetchall()
            return [
                {
                    "session_id": row[0],
                    "user_agent": row[1],
                    "ip_address": row[2],
                    "created_at": row[3],
                    "last_seen_at": row[4],
                    "expires_at": row[5],
                    "revoked_at": row[6]
                }
                for row in rows
            ]

    def _build_active_sessions_where_clause(self, search: Optional[str] = None):
        where_clauses = [
            "us.revoked_at IS NULL",
            "us.expires_at > (CURRENT_TIMESTAMP AT TIME ZONE 'UTC')"
        ]
        params: List[Any] = []

        search_text = (search or '').strip()
        if search_text:
            pattern = f"%{search_text}%"
            where_clauses.append("""
                (
                    us.session_id::text ILIKE %s
                    OR COALESCE(us.ip_address, '') ILIKE %s
                    OR COALESCE(us.user_agent, '') ILIKE %s
                    OR COALESCE(u.name, '') ILIKE %s
                    OR COALESCE(u.login, '') ILIKE %s
                    OR COALESCE(sv.name, '') ILIKE %s
                )
            """)
            params.extend([pattern, pattern, pattern, pattern, pattern, pattern])

        return " AND ".join(where_clauses), params

    def list_all_active_sessions(self, limit: Optional[int] = None, offset: int = 0, search: Optional[str] = None):
        where_sql, params = self._build_active_sessions_where_clause(search=search)

        query = f"""
            SELECT
                us.session_id::text,
                us.user_id,
                u.name,
                u.role,
                u.login,
                u.supervisor_id,
                sv.name AS supervisor_name,
                u.avatar_bucket,
                u.avatar_blob_path,
                us.user_agent,
                us.ip_address,
                us.created_at,
                us.last_seen_at,
                us.expires_at,
                us.sensitive_data_unlocked,
                us.sensitive_data_unlocked_at
            FROM user_sessions us
            JOIN users u ON u.id = us.user_id
            LEFT JOIN users sv ON sv.id = u.supervisor_id
            WHERE {where_sql}
            ORDER BY u.name ASC, us.last_seen_at DESC, us.created_at DESC
        """

        final_params = list(params)
        if limit is not None:
            query += " LIMIT %s"
            final_params.append(max(1, int(limit)))
        if offset:
            query += " OFFSET %s"
            final_params.append(max(0, int(offset)))

        with self._get_cursor() as cursor:
            cursor.execute(query, tuple(final_params))
            rows = cursor.fetchall()
            return [
                {
                    "session_id": row[0],
                    "user_id": row[1],
                    "user_name": row[2],
                    "user_role": row[3],
                    "user_login": row[4],
                    "supervisor_id": row[5],
                    "supervisor_name": row[6],
                    "avatar_bucket": row[7],
                    "avatar_blob_path": row[8],
                    "user_agent": row[9],
                    "ip_address": row[10],
                    "created_at": row[11],
                    "last_seen_at": row[12],
                    "expires_at": row[13],
                    "sensitive_data_unlocked": bool(row[14]),
                    "sensitive_data_unlocked_at": row[15]
                }
                for row in rows
            ]

    def get_all_active_sessions_summary(self, search: Optional[str] = None):
        where_sql, params = self._build_active_sessions_where_clause(search=search)

        bot_expr = "ua ~* 'bot|crawl|spider|slurp|bingpreview|facebookexternalhit|linkedinbot|twitterbot'"
        tablet_expr = (
            "("
            "ua ~* 'ipad|tablet|kindle|playbook|silk' "
            "OR (ua ~* 'android' AND ua !~* 'mobile') "
            "OR (ua ~* 'windows' AND ua !~* 'phone' AND ua ~* 'touch') "
            "OR (ua ~* 'puffin' AND ua !~* 'ip|ap|wp')"
            ")"
        )
        mobile_expr = "ua ~* 'mobi|android|iphone|ipod|blackberry|iemobile|opera mini|windows phone'"

        query = f"""
            WITH filtered AS (
                SELECT
                    us.user_id,
                    LOWER(COALESCE(u.role, '')) AS user_role,
                    LOWER(COALESCE(us.user_agent, '')) AS ua
                FROM user_sessions us
                JOIN users u ON u.id = us.user_id
                LEFT JOIN users sv ON sv.id = u.supervisor_id
                WHERE {where_sql}
            )
            SELECT
                COUNT(*) AS total_sessions,
                COUNT(DISTINCT user_id) AS total_users,
                COUNT(*) FILTER (WHERE user_role IN ('admin', 'super_admin')) AS admin_sessions,
                COUNT(*) FILTER (WHERE user_role IN ('sv', 'supervisor')) AS sv_sessions,
                COUNT(*) FILTER (WHERE user_role = 'operator') AS operator_sessions,
                COUNT(*) FILTER (WHERE ua = '') AS unknown_sessions,
                COUNT(*) FILTER (WHERE {bot_expr}) AS bot_sessions,
                COUNT(*) FILTER (
                    WHERE ua <> ''
                      AND NOT ({bot_expr})
                      AND {tablet_expr}
                ) AS tablet_sessions,
                COUNT(*) FILTER (
                    WHERE ua <> ''
                      AND NOT ({bot_expr})
                      AND NOT ({tablet_expr})
                      AND {mobile_expr}
                ) AS mobile_sessions,
                COUNT(*) FILTER (
                    WHERE ua <> ''
                      AND NOT ({bot_expr})
                      AND NOT ({tablet_expr})
                      AND NOT ({mobile_expr})
                ) AS desktop_sessions
            FROM filtered
        """

        with self._get_cursor() as cursor:
            cursor.execute(query, tuple(params))
            row = cursor.fetchone() or (0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
            return {
                "total_sessions": int(row[0] or 0),
                "total_users": int(row[1] or 0),
                "role_counts": {
                    "admin": int(row[2] or 0),
                    "sv": int(row[3] or 0),
                    "operator": int(row[4] or 0)
                },
                "device_counts": {
                    "unknown": int(row[5] or 0),
                    "bot": int(row[6] or 0),
                    "tablet": int(row[7] or 0),
                    "mobile": int(row[8] or 0),
                    "desktop": int(row[9] or 0)
                }
            }

    def get_call_evaluations(self, operator_id, month=None, include_target: bool = False):
        """
        Возвращает оценки звонков (calls) + неоценённые звонки (imported_calls).
        Последняя версия оценки определяется по ключу
        (operator_id, phone_number, month, appeal_date).
        Для calls.duration -> NULL, для imported_calls.duration -> ic.duration_sec.
        """
        query = """
            WITH latest_versions AS (
                SELECT 
                    operator_id,
                    phone_number,
                    month,
                    appeal_date,
                    MAX(created_at) AS latest_date
                FROM calls
                WHERE operator_id = %s
                GROUP BY operator_id, phone_number, month, appeal_date
            ),
            latest_calls AS (
                SELECT 
                    c.id::text AS id_text,
                    c.operator_id,
                    c.phone_number,
                    c.month,
                    c.appeal_date,
                    c.created_at
                FROM calls c
                JOIN latest_versions lv ON 
                    c.operator_id = lv.operator_id
                    AND (
                        (c.appeal_date IS NULL AND lv.appeal_date IS NULL)
                        OR c.appeal_date = lv.appeal_date
                    )
                    AND c.phone_number = lv.phone_number 
                    AND c.month = lv.month 
                    AND c.created_at = lv.latest_date
                WHERE c.operator_id = %s
            )
            SELECT 
                c.id::text AS id,
                c.month, 
                c.phone_number, 
                c.appeal_date,
                c.score, 
                c.comment,
                c.audio_path,
                c.is_draft,
                c.is_correction,
                TO_CHAR(c.created_at, 'YYYY-MM-DD HH24:MI') AS evaluation_date,
                c.scores,
                c.criterion_comments,
                d.id AS direction_id,
                d.name AS direction_name,
                d.criteria AS direction_criteria,
                d.has_file_upload AS direction_has_file_upload,
                u.name AS evaluator_name,
                c.created_at,
                c.sv_request,
                c.sv_request_comment,
                c.sv_request_by,
                su.name AS sv_request_by_name,
                TO_CHAR(c.sv_request_at, 'YYYY-MM-DD HH24:MI') AS sv_request_at,
                c.sv_request_approved,
                c.sv_request_approved_by,
                apu.name AS sv_request_approved_by_name,
                TO_CHAR(c.sv_request_approved_at, 'YYYY-MM-DD HH24:MI') AS sv_request_approved_at,
                NULL::numeric AS duration, -- у оценённых нет длительности
                FALSE AS is_imported,
                COALESCE(c.comment_visible_to_operator, TRUE) AS comment_visible_to_operator
            FROM calls c
            JOIN latest_calls lc ON c.id::text = lc.id_text
            LEFT JOIN directions d ON c.direction_id = d.id  
            LEFT JOIN users u ON c.evaluator_id = u.id
            LEFT JOIN users su ON c.sv_request_by = su.id
            LEFT JOIN users apu ON c.sv_request_approved_by = apu.id
            WHERE c.operator_id = %s
        """
        params = [operator_id, operator_id, operator_id]
        if month:
            query += " AND c.month = %s"
            params.append(month)

        # добавляем неоценённые звонки из imported_calls
        query += """
            UNION ALL
            SELECT 
                ic.id::text AS id,
                ic.month,
                ic.phone_number,
                ic.datetime_raw AS appeal_date,
                NULL::numeric AS score,
                NULL::text AS comment,
                NULL::text AS audio_path,
                FALSE AS is_draft,
                FALSE AS is_correction,
                NULL::text AS evaluation_date,
                NULL::jsonb AS scores,
                NULL::jsonb AS criterion_comments,
                NULL::integer AS direction_id,
                NULL::text AS direction_name,
                NULL::jsonb AS direction_criteria,
                NULL::boolean AS direction_has_file_upload,
                NULL::text AS evaluator_name,
                ic.imported_at AS created_at,
                FALSE::boolean AS sv_request,
                NULL::text AS sv_request_comment,
                NULL::integer AS sv_request_by,
                NULL::text AS sv_request_by_name,
                NULL::text AS sv_request_at,
                FALSE::boolean AS sv_request_approved,
                NULL::integer AS sv_request_approved_by,
                NULL::text AS sv_request_approved_by_name,
                NULL::text AS sv_request_approved_at,
                ic.duration_sec::numeric AS duration, -- берем из imported_calls
                TRUE AS is_imported,
                TRUE AS comment_visible_to_operator
            FROM imported_calls ic
            WHERE ic.operator_id = %s AND ic.status = 'not_evaluated'
        """
        params.append(operator_id)

        if month:
            query += " AND ic.month = %s"
            params.append(month)

        query += " ORDER BY created_at DESC"

        target_month = str(month or datetime.now().strftime('%Y-%m'))
        evaluation_target = None

        with self._get_cursor() as cursor:
            if include_target:
                source = self._fetch_operator_call_evaluation_target_source(cursor, operator_id, target_month)
                evaluation_target = self._build_operator_call_evaluation_target(
                    operator_id=operator_id,
                    target_month=target_month,
                    hire_date_value=source.get('hire_date_value'),
                    regular_hours=source.get('regular_hours'),
                    training_hours=source.get('training_hours'),
                    technical_issue_hours=source.get('technical_issue_hours'),
                    offline_activity_hours=source.get('offline_activity_hours'),
                    operator_norm_hours=source.get('operator_norm_hours')
                )

            cursor.execute(query, params)
            rows = cursor.fetchall()

            evaluations = [
                {
                    "id": row[0],
                    "month": row[1],
                    "phone_number": row[2],
                    "appeal_date": (
                        row[3].strftime('%Y-%m-%d %H:%M:%S')
                        if row[3] and hasattr(row[3], "strftime")
                        else (row[3] if row[3] else None)
                    ),
                    "score": float(row[4]) if row[4] is not None else None,
                    "comment": row[5],
                    "audio_path": row[6],
                    "is_draft": bool(row[7]),
                    "is_correction": bool(row[8]),
                    "evaluation_date": row[9],
                    "scores": row[10] if row[10] else [],
                    "criterion_comments": row[11] if row[11] else [],
                    "direction": {
                        "id": row[12],
                        "name": row[13],
                        "criteria": row[14] if row[14] else [],
                        "hasFileUpload": row[15] if row[15] is not None else True,
                    }
                    if row[12]
                    else None,
                    "evaluator": row[16] if row[16] else None,
                    "created_at": row[17],
                    "sv_request": bool(row[18]),
                    "sv_request_comment": row[19],
                    "sv_request_by": row[20],
                    "sv_request_by_name": row[21],
                    "sv_request_at": row[22],
                    "sv_request_approved": bool(row[23]),
                    "sv_request_approved_by": row[24],
                    "sv_request_approved_by_name": row[25],
                    "sv_request_approved_at": row[26],
                    "duration": float(row[27]) if row[27] is not None else None,
                    "is_imported": bool(row[28]),
                    "comment_visible_to_operator": bool(row[29]) if row[29] is not None else True,
                }
                for row in rows
            ]

            if include_target:
                return {
                    "evaluations": evaluations,
                    "evaluation_target": evaluation_target
                }

            return evaluations

    def get_operator_score_aggregates_for_month(self, month, operator_ids: Optional[List[int]] = None):
        """
        Быстрые агрегаты по операторам за месяц:
        - call_count: количество оцененных звонков (score IS NOT NULL)
        - avg_score: средний балл

        Логика "последней версии" соответствует get_call_evaluations:
        берем MAX(created_at) для пары
        (operator_id, phone_number, month, appeal_date).
        """
        if not month:
            return {}

        normalized_ids = None
        if operator_ids is not None:
            normalized_ids = []
            seen_ids = set()
            for raw_id in operator_ids:
                try:
                    op_id = int(raw_id)
                except (TypeError, ValueError):
                    continue
                if op_id in seen_ids:
                    continue
                seen_ids.add(op_id)
                normalized_ids.append(op_id)
            if not normalized_ids:
                return {}

        params = [month]
        filter_clause = "month = %s"
        if normalized_ids is not None:
            filter_clause += " AND operator_id = ANY(%s)"
            params.append(normalized_ids)

        query = f"""
            WITH latest_versions AS (
                SELECT
                    operator_id,
                    phone_number,
                    month,
                    appeal_date,
                    MAX(created_at) AS latest_date
                FROM calls
                WHERE {filter_clause}
                GROUP BY operator_id, phone_number, month, appeal_date
            ),
            latest_calls AS (
                SELECT
                    c.operator_id,
                    c.score
                FROM calls c
                JOIN latest_versions lv
                  ON c.operator_id = lv.operator_id
                 AND c.phone_number = lv.phone_number
                 AND c.month = lv.month
                 AND (
                     (c.appeal_date IS NULL AND lv.appeal_date IS NULL)
                     OR c.appeal_date = lv.appeal_date
                 )
                 AND c.created_at = lv.latest_date
            )
            SELECT
                operator_id,
                COUNT(*) FILTER (WHERE score IS NOT NULL) AS call_count,
                AVG(score)::float AS avg_score
            FROM latest_calls
            GROUP BY operator_id
        """

        with self._get_cursor() as cursor:
            cursor.execute(query, tuple(params))
            rows = cursor.fetchall()

        result = {}
        for operator_id, call_count, avg_score in rows:
            op_id = int(operator_id)
            result[op_id] = {
                "call_count": int(call_count or 0),
                "avg_score": round(float(avg_score), 2) if avg_score is not None else None
            }
        return result

        
    def get_operators_summary_for_month(self, month, supervisor_id=None):
        """
        Возвращает список сотрудников с количеством последних версий звонков за месяц.
        Включает:
        - операторов;
        - супервайзеров, если у них есть оценки за месяц.

        Последние версии определяются как MAX(created_at) для каждой комбинации:
        (phone_number, operator_id, month, appeal_date)
        Если supervisor_id задан — фильтрует операторов по этому SV и включает самого SV.
        """
        query = """
        WITH latest_versions AS (
            -- для каждой пары (phone_number, operator_id, month, appeal_date) берем последний created_at
            SELECT
                phone_number,
                operator_id,
                month,
                appeal_date,
                MAX(created_at) AS latest_date
            FROM calls
            WHERE month = %s
            GROUP BY phone_number, operator_id, month, appeal_date
        ),
        latest_calls AS (
            -- берем сами записи (последние версии) для указанного месяца и appeal_date
            SELECT c.*
            FROM calls c
            JOIN latest_versions lv
            ON c.phone_number = lv.phone_number
            AND c.operator_id = lv.operator_id
            AND c.month = lv.month
            AND ( (c.appeal_date IS NULL AND lv.appeal_date IS NULL) OR (c.appeal_date = lv.appeal_date) )
            AND c.created_at = lv.latest_date
            WHERE c.month = %s
        ),
        counts AS (
            -- считаем количество последних версий звонков по оператору
            SELECT
                operator_id,
                COUNT(*) FILTER (WHERE score IS NOT NULL) AS call_count,
                AVG(score)::float AS avg_score
            FROM latest_calls
            GROUP BY operator_id
        )
        SELECT
            u.id AS operator_id,
            u.name AS operator_name,
            u.status,
            u.direction_id,
            d.name AS direction_name,
            u.supervisor_id,
            su.name AS supervisor_name,
            u.hire_date,
            COALESCE(c.call_count, 0) AS call_count,
            c.avg_score,
            LOWER(COALESCE(u.role, '')) AS role_norm
        FROM users u
        LEFT JOIN directions d ON u.direction_id = d.id
        LEFT JOIN users su ON u.supervisor_id = su.id
        LEFT JOIN counts c ON c.operator_id = u.id
        WHERE (
            LOWER(COALESCE(u.role, '')) = 'operator'
            OR (
                LOWER(COALESCE(u.role, '')) IN ('sv', 'supervisor')
                AND COALESCE(c.call_count, 0) > 0
            )
        )
        """
        params = [month, month]

        if supervisor_id is not None:
            query += """
                AND (
                    u.supervisor_id = %s
                    OR (
                        LOWER(COALESCE(u.role, '')) IN ('sv', 'supervisor')
                        AND u.id = %s
                    )
                )
            """
            params.extend([supervisor_id, supervisor_id])

        query += " ORDER BY u.name"

        with self._get_cursor() as cursor:
            cursor.execute(query, params)
            rows = cursor.fetchall()
            result = []
            for row in rows:
                role_norm = str(row[10] or '').strip().lower()
                is_supervisor_row = role_norm in ('sv', 'supervisor')
                effective_supervisor_id = row[5]
                effective_supervisor_name = row[6]
                if is_supervisor_row:
                    effective_supervisor_id = row[0]
                    effective_supervisor_name = row[1]

                result.append({
                    "id": row[0],
                    "name": row[1],
                    "status": row[2],
                    "direction_id": row[3],
                    "direction_name": row[4],
                    "supervisor_id": effective_supervisor_id,
                    "supervisor_name": effective_supervisor_name,
                    "hire_date": row[7].strftime('%d-%m-%Y') if row[7] else None,
                    "call_count": int(row[8] or 0),
                    "avg_score": round(float(row[9]), 2) if row[9] is not None else None,
                    "role": role_norm
                })
            return result

    def update_user(self, user_id, field, value, changed_by=None):
        allowed_fields = [
            'direction_id',
            'supervisor_id',
            'status',
            'rate',
            'hire_date',
            'name',
            'gender',
            'birth_date',
            'phone',
            'email',
            'instagram',
            'telegram_nick',
            'company_name',
            'employment_type',
            'has_proxy',
            'proxy_card_number',
            'has_driver_license',
            'sip_number',
            'study_place',
            'study_course',
            'close_contact_1_relation',
            'close_contact_1_full_name',
            'close_contact_1_phone',
            'close_contact_2_relation',
            'close_contact_2_full_name',
            'close_contact_2_phone',
            'card_number',
            'internship_in_company',
            'front_office_training',
            'front_office_training_date',
            'taxipro_id'
        ]
        if field not in allowed_fields:
            raise ValueError("Invalid field to update")
        
        with self._get_cursor() as cursor:
            # Fetch old value
            cursor.execute(f"SELECT role, {field} FROM users WHERE id = %s", (user_id,))
            row = cursor.fetchone()
            if not row:
                return False
            current_role = str(row[0] or '').strip().lower()
            old_field_value = row[1]
            if current_role == 'trainer' and field in ('direction_id', 'supervisor_id'):
                value = None
            old_value = str(old_field_value) if old_field_value is not None else None
            
            # Update
            cursor.execute(f"UPDATE users SET {field} = %s WHERE id = %s RETURNING id", (value, user_id))
            updated = cursor.fetchone() is not None
            
            # Log history if updated
            if updated:
                cursor.execute("""
                    INSERT INTO user_history (user_id, changed_by, field_changed, old_value, new_value)
                    VALUES (%s, %s, %s, %s, %s)
                """, (user_id, changed_by, field, old_value, str(value)))
            
            return updated

    def promote_operator_to_supervisor(self, user_id, changed_by=None):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, name, role, supervisor_id
                FROM users
                WHERE id = %s
                FOR UPDATE
            """, (user_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("User not found")

            target_id, target_name, current_role, current_supervisor_id = row
            role_value = str(current_role or '').lower()

            if role_value == 'sv':
                raise ValueError("User is already a supervisor")
            if role_value != 'operator':
                raise ValueError("Only operators can be promoted to supervisor")

            cursor.execute("""
                UPDATE users
                SET role = 'sv',
                    supervisor_id = NULL
                WHERE id = %s
                RETURNING id, name
            """, (target_id,))
            updated = cursor.fetchone()
            if not updated:
                raise ValueError("Failed to promote user")

            cursor.execute("""
                INSERT INTO user_history (user_id, changed_by, field_changed, old_value, new_value)
                VALUES (%s, %s, %s, %s, %s)
            """, (target_id, changed_by, 'role', role_value, 'sv'))

            if current_supervisor_id is not None:
                cursor.execute("""
                    INSERT INTO user_history (user_id, changed_by, field_changed, old_value, new_value)
                    VALUES (%s, %s, %s, %s, %s)
                """, (target_id, changed_by, 'supervisor_id', str(current_supervisor_id), None))

            return {
                "id": int(updated[0]),
                "name": updated[1] or target_name
            }

    def get_user_avatar_storage(self, user_id):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT avatar_bucket, avatar_blob_path, avatar_original_blob_path, avatar_content_type, avatar_file_size, avatar_updated_at
                FROM users
                WHERE id = %s
            """, (user_id,))
            row = cursor.fetchone()
            if not row:
                return None
            return {
                "avatar_bucket": row[0],
                "avatar_blob_path": row[1],
                "avatar_original_blob_path": row[2],
                "avatar_content_type": row[3],
                "avatar_file_size": int(row[4]) if row[4] is not None else None,
                "avatar_updated_at": row[5].isoformat() if row[5] else None
            }

    def set_user_avatar(self, user_id, bucket_name, blob_path, original_blob_path=None, content_type=None, file_size=None):
        with self._get_cursor() as cursor:
            cursor.execute("""
                UPDATE users
                SET avatar_bucket = %s,
                    avatar_blob_path = %s,
                    avatar_original_blob_path = %s,
                    avatar_content_type = %s,
                    avatar_file_size = %s,
                    avatar_updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                RETURNING avatar_bucket, avatar_blob_path, avatar_original_blob_path, avatar_content_type, avatar_file_size, avatar_updated_at
            """, (bucket_name, blob_path, original_blob_path, content_type, file_size, user_id))
            row = cursor.fetchone()
            if not row:
                return None
            return {
                "avatar_bucket": row[0],
                "avatar_blob_path": row[1],
                "avatar_original_blob_path": row[2],
                "avatar_content_type": row[3],
                "avatar_file_size": int(row[4]) if row[4] is not None else None,
                "avatar_updated_at": row[5].isoformat() if row[5] else None
            }

    def clear_user_avatar(self, user_id):
        with self._get_cursor() as cursor:
            cursor.execute("""
                UPDATE users
                SET avatar_bucket = NULL,
                    avatar_blob_path = NULL,
                    avatar_original_blob_path = NULL,
                    avatar_content_type = NULL,
                    avatar_file_size = NULL,
                    avatar_updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                RETURNING id
            """, (user_id,))
            row = cursor.fetchone()
            return bool(row)

    def get_all_operators(self):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, name, direction_id, hire_date, supervisor_id, scores_table_url, hours_table_url
                FROM users 
                WHERE role = 'operator'
            """)
            return cursor.fetchall()

    def get_week_call_stats(self, operator_id, start_date, end_date):
        with self._get_cursor() as cursor:
            query = """
                WITH latest_versions AS (
                    SELECT 
                        operator_id,
                        phone_number,
                        month,
                        appeal_date,
                        MAX(created_at) as latest_date
                    FROM calls
                    WHERE operator_id = %s
                    AND created_at >= %s
                    AND created_at <= %s
                    AND is_draft = FALSE
                    GROUP BY operator_id, phone_number, month, appeal_date
                ),
                latest_calls AS (
                    SELECT 
                        c.id,
                        c.score
                    FROM calls c
                    JOIN latest_versions lv ON 
                        c.operator_id = lv.operator_id AND
                        c.phone_number = lv.phone_number AND 
                        c.month = lv.month AND 
                        (
                            (c.appeal_date IS NULL AND lv.appeal_date IS NULL)
                            OR c.appeal_date = lv.appeal_date
                        ) AND
                        c.created_at = lv.latest_date
                    WHERE c.operator_id = %s
                    AND c.is_draft = FALSE
                )
                SELECT COUNT(*), AVG(score)::float
                FROM latest_calls
            """
            cursor.execute(query, (operator_id, start_date, end_date, operator_id))
            result = cursor.fetchone()
            count = result[0] or 0
            avg = result[1] or 0.0
            return count, avg

    def get_average_scores_for_period(self, start_date, end_date, operator_ids: Optional[List[int]] = None):
        """
        Возвращает среднюю оценку и количество оценок по операторам за выбранный период.

        Параметры:
        - start_date: начало периода (datetime или строка в формате SQL-совместимом, напр. 'YYYY-MM-DD' или full timestamp)
        - end_date: конец периода (inclusive)
        - operator_ids: опциональный список id операторов для фильтрации (если None — по всем операторам)

        Возвращает словарь:
        {
            "operators": { operator_id: {"count": int, "avg_score": float}, ... },
            "overall": {"count": total_count, "avg_score": overall_avg}
        }

        Реализация основана на логике последних версий оценок
        (по phone_number/operator_id/month/appeal_date)
        аналогично `get_week_call_stats`, но аггрегирует по операторам за период.
        """
        # normalize dates if strings were provided (leave as-is otherwise)
        sd = start_date
        ed = end_date
        params = [sd, ed]

        query = """
            WITH latest_versions AS (
                SELECT phone_number, operator_id, month, appeal_date, MAX(created_at) AS latest_date
                FROM calls
                WHERE is_draft = FALSE
                  AND created_at >= %s AND created_at <= %s
                GROUP BY phone_number, operator_id, month, appeal_date
            )
            SELECT c.operator_id, COUNT(*) AS cnt, AVG(c.score)::float AS avg_score
            FROM calls c
            JOIN latest_versions lv
              ON c.phone_number = lv.phone_number
             AND c.operator_id = lv.operator_id
             AND c.month = lv.month
             AND (
                 (c.appeal_date IS NULL AND lv.appeal_date IS NULL)
                 OR c.appeal_date = lv.appeal_date
             )
             AND c.created_at = lv.latest_date
        """

        if operator_ids:
            query += "\nWHERE c.operator_id = ANY(%s)"
            params.append(operator_ids)

        query += "\nGROUP BY c.operator_id"

        with self._get_cursor() as cursor:
            cursor.execute(query, tuple(params))
            rows = cursor.fetchall()

        operators = {}
        total_count = 0
        total_weighted = 0.0
        for op_id, cnt, avg in rows:
            c = int(cnt or 0)
            a = float(avg) if avg is not None else 0.0
            operators[int(op_id)] = {"count": c, "avg_score": round(a, 2)}
            total_count += c
            total_weighted += a * c

        overall_avg = round((total_weighted / total_count), 2) if total_count > 0 else 0.0

        return {"operators": operators, "overall": {"count": total_count, "avg_score": overall_avg}}

    def set_user_active(self, user_id, status):
        # допустимые статусы
        allowed_statuses = {"active", "break", "training", "inactive", "tech", "iesigning"}
        if status not in allowed_statuses:
            return False  # неверный статус
    
        # преобразуем статус в boolean
        is_active = True if status == "active" else False
    
        with self._get_cursor() as cursor:
            cursor.execute("""
                UPDATE users 
                SET is_active = %s
                WHERE id = %s AND role = 'operator'
                RETURNING id
            """, (is_active, user_id))
            return cursor.fetchone() is not None
    
    def get_active_operators(self, direction_name=None):
        with self._get_cursor() as cursor:
            if direction_name:
                cursor.execute("""
                    SELECT u.id, u.name, u.direction_id
                    FROM users u
                    JOIN directions d ON u.direction_id = d.id
                    WHERE u.role = 'operator' AND u.is_active = TRUE
                      AND d.name = %s AND d.is_active = TRUE
                    ORDER BY u.name
                """, (direction_name,))
            else:
                cursor.execute("""
                    SELECT id, name, direction_id 
                    FROM users 
                    WHERE role = 'operator' AND is_active = TRUE
                    ORDER BY name
                """)
            return [{"id": row[0], "name": row[1], "direction_id": row[2]} for row in cursor.fetchall()]

    
    def log_activity(self, operator_id, status):
        allowed_statuses = {"active", "break", "training", "inactive","tech","iesigning"}
        if status not in allowed_statuses:
            return False
    
        try:
            with self._get_cursor() as cursor:
                cursor.execute("""
                    INSERT INTO operator_activity_logs (operator_id, is_active)
                    VALUES (%s, %s)
                """, (operator_id, status))
            return True
        except Exception as e:
            logging.error(f"Error logging activity: {e}")
            return False
            
    def get_last_activity_status(self, operator_id):
        try:
            with self._get_cursor() as cursor:
                cursor.execute("""
                    SELECT is_active
                    FROM operator_activity_logs
                    WHERE operator_id = %s
                    ORDER BY change_time DESC
                    LIMIT 1
                """, (operator_id,))
                row = cursor.fetchone()
                return row[0] if row else None
        except Exception as e:
            logging.error(f"Error fetching last activity status: {e}")
            return None

        
    def generate_monthly_report(self, supervisor_id, month=None, current_date=None):
        """
        Генерация месячного отчёта (xlsx).
        Включает:
        - Summary (как раньше);
        - All active (даты dd.mm.yyyy, ФИО с рамкой, ячейки часов >0 — серый B3B3B3);
        - Листы операторов: дата в формате dd.mm.yyyy в объединённой ячейке на день,
        длительности в числовом формате (часы, 2 знака), длительности окрашены в цвет статуса.
        """
        def sanitize_table_name(name):
            sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', (name or ""))
            if not sanitized or (not sanitized[0].isalpha() and sanitized[0] != '_'):
                sanitized = f"_{sanitized}"
            return sanitized[:255]

        def format_hms(duration):
            """Оставлю для совместимости — не используется для числовых полей."""
            if duration is None:
                return "N/A"
            total = int(duration.total_seconds())
            if total < 0:
                return "N/A"
            h = total // 3600
            m = (total % 3600) // 60
            s = total % 60
            return f"{h:02d}:{m:02d}:{s:02d}"

        try:
            if current_date is None:
                current_date = date.today()
            else:
                current_date = datetime.strptime(current_date, "%Y-%m-%d").date() if isinstance(current_date, str) else current_date

            if month is None:
                year = current_date.year
                mon = current_date.month
            else:
                try:
                    year, mon = map(int, month.split('-'))
                    if not (1 <= mon <= 12):
                        raise ValueError("Invalid month")
                except:
                    raise ValueError("Invalid month format. Use YYYY-MM")

            month_str = f"{year}-{mon:02d}"
            filename = f"monthly_report_supervisor_{supervisor_id}_{month_str}.xlsx"

            month_start_date = date(year, mon, 1)
            days_in_month = calendar.monthrange(year, mon)[1]
            month_end = date(year, mon, days_in_month)
            end_date = min(month_end, current_date)

            end_time = datetime.combine(end_date, dt_time(23, 59, 59))

            # Получаем данные
            with self._get_cursor() as cursor:
                cursor.execute("""
                    SELECT id, name
                    FROM users
                    WHERE supervisor_id = %s AND role = 'operator'
                """, (supervisor_id,))
                operators = cursor.fetchall()
                if not operators:
                    raise ValueError("No operators found for the given supervisor")
                operators_dict = {op[0]: op[1] for op in operators}

                cursor.execute("""
                    SELECT o.operator_id, o.change_time, o.is_active
                    FROM operator_activity_logs o
                    JOIN users u ON o.operator_id = u.id
                    WHERE u.supervisor_id = %s AND u.role = 'operator'
                    AND o.change_time >= %s
                    AND o.change_time < %s + INTERVAL '1 DAY'
                    ORDER BY o.operator_id, o.change_time
                """, (supervisor_id, month_start_date, end_date))
                all_logs = cursor.fetchall()

            # Группируем логи и считаем активации/деактивации для summary
            logs_per_op = defaultdict(list)
            counts_per_op = defaultdict(lambda: defaultdict(lambda: {'act': 0, 'deact': 0}))
            total_counts_per_op = defaultdict(lambda: {'act': 0, 'deact': 0})

            for log in all_logs:
                op_id, change_time, is_active = log
                dt = change_time.date()
                logs_per_op[op_id].append({'change_time': change_time, 'is_active': is_active, 'date': dt})
                if is_active == 'active':
                    counts_per_op[op_id][dt]['act'] += 1
                    total_counts_per_op[op_id]['act'] += 1
                else:
                    counts_per_op[op_id][dt]['deact'] += 1
                    total_counts_per_op[op_id]['deact'] += 1

            # Список дат для summary (от 1 числа до end_date)
            dates = [month_start_date + timedelta(days=i) for i in range((end_date - month_start_date).days + 1)]

            # Подготовим переводы статусов и заливки
            # перевод статусов, согласно вашему запросу:
            status_display = {
                'active': 'Активен',
                'tech': 'Перерыв',
                'break': 'Перерыв',
                'training': 'Тренинг',
                'iesigning': 'Подписание',
                'inactive': 'Завершил смену'
            }

            # стили и заливки
            total_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
            status_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
            green_fill = PatternFill(start_color="CFF5D0", end_color="CFF5D0", fill_type="solid")   # active
            orange_fill = PatternFill(start_color="FFE8C0", end_color="FFE8C0", fill_type="solid")  # break/tech
            yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # training
            purple_fill = PatternFill(start_color="E8D7F7", end_color="E8D7F7", fill_type="solid")  # tech/alt
            blue_fill = PatternFill(start_color="DCEEFF", end_color="DCEEFF", fill_type="solid")    # iesigning
            inactive_fill = PatternFill(start_color="FFDADA", end_color="FFDADA", fill_type="solid")# inactive
            grey_fill = PatternFill(start_color="B3B3B3", end_color="B3B3B3", fill_type="solid")     # All active >0

            # Соответствие статуса -> заливка (используем явное сопоставление)
            status_fill_map = {
                'active': green_fill,
                'iesigning': blue_fill,
                'break': orange_fill,
                'training': yellow_fill,
                'tech': orange_fill,      # вы просили tech = "Перерыв", поэтому даём тот же цвет что и break
                'inactive': inactive_fill
            }

            # порядок статусов для колонок итогов E..I (исключаем 'inactive')
            status_order = [
                ('active', status_display.get('active', 'active'), status_fill_map.get('active')),
                ('iesigning', status_display.get('iesigning', 'iesigning'), status_fill_map.get('iesigning')),
                ('break', status_display.get('break', 'break'), status_fill_map.get('break')),
                ('training', status_display.get('training', 'training'), status_fill_map.get('training')),
                ('tech', status_display.get('tech', 'tech'), status_fill_map.get('tech')),
            ]

            # --------- НОВЫЙ БЛОК: подсчёт суммарных секунд активности (active + iesigning) по дням для каждого оператора
            active_seconds_per_op_per_day = defaultdict(lambda: defaultdict(int))
            for op_id, logs in logs_per_op.items():
                logs_sorted = sorted(logs, key=lambda x: x['change_time'])
                for i, log in enumerate(logs_sorted):
                    dt = log['date']
                    if i < len(logs_sorted) - 1:
                        next_time = logs_sorted[i + 1]['change_time']
                    else:
                        next_time = end_time
                    duration = next_time - log['change_time']
                    # пропускаем дубликаты состояния
                    if i > 0 and log['is_active'] == logs_sorted[i - 1]['is_active']:
                        continue
                    if log['is_active'] in ('active', 'iesigning'):
                        active_seconds_per_op_per_day[op_id][dt] += int(duration.total_seconds())
            # --------- КОНЕЦ НОВОГО БЛОКА

            wb = Workbook()
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # -------------------------
            # 1) Summary: заголовки и данные
            # -------------------------
            ws_summary.cell(1, 1).value = "ФИО"
            for col, dt in enumerate(dates, start=2):
                ws_summary.cell(1, col).value = dt.strftime("%Y-%m-%d")
                ws_summary.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
            last_date_col = 1 + len(dates)
            ws_summary.cell(1, last_date_col + 1).value = "Итого активаций"
            ws_summary.cell(1, last_date_col + 2).value = "Итого деактиваций"

            # шрифты для rich text (если доступны)
            green_font = InlineFont(color="00AA00") if InlineFont else None
            red_font = InlineFont(color="AA0000") if InlineFont else None
            default_font = InlineFont() if InlineFont else None

            # Заполняем summary
            row = 2
            for op_id, name in operators_dict.items():
                ws_summary.cell(row, 1).value = name
                for col, dt in enumerate(dates, start=2):
                    activations = counts_per_op[op_id][dt]['act']
                    deactivations = counts_per_op[op_id][dt]['deact']
                    cell = ws_summary.cell(row, col)
                    if CellRichText and TextBlock and green_font and red_font:
                        rt = CellRichText([
                            TextBlock(green_font, str(activations)),
                            TextBlock(default_font, " | "),
                            TextBlock(red_font, str(deactivations))
                        ])
                        cell.value = rt
                    else:
                        cell.value = f"{activations} | {deactivations}"
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                # итоговые колонки
                act_cell = ws_summary.cell(row, last_date_col + 1)
                deact_cell = ws_summary.cell(row, last_date_col + 2)
                if CellRichText and TextBlock and green_font and red_font:
                    act_cell.value = CellRichText([TextBlock(green_font, str(total_counts_per_op[op_id]['act']))])
                    deact_cell.value = CellRichText([TextBlock(red_font, str(total_counts_per_op[op_id]['deact']))])
                else:
                    act_cell.value = total_counts_per_op[op_id]['act']
                    deact_cell.value = total_counts_per_op[op_id]['deact']
                act_cell.alignment = Alignment(horizontal='center', vertical='center')
                deact_cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1

            # легенда
            ws_summary.cell(row + 1, 1).value = "Легенда:"
            ws_summary.cell(row + 2, 1).value = "Зелёный — активации"
            ws_summary.cell(row + 3, 1).value = "Красный — деактивации"

            # стили summary: жирная шапка, границы, автоширины
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for c in range(1, last_date_col + 3):
                cell = ws_summary.cell(1, c)
                cell.font = Font(bold=True)
                cell.border = thin_border
            for r in range(1, row):
                for c in range(1, last_date_col + 3):
                    ws_summary.cell(r, c).border = thin_border
            ws_summary.column_dimensions['A'].width = 24
            for idx in range(2, last_date_col + 3):
                col_letter = ws_summary.cell(1, idx).column_letter
                ws_summary.column_dimensions[col_letter].width = 12
            ws_summary.freeze_panes = "A2"

            # -------------------------
            # НОВЫЙ ЛИСТ: All active (суммы часов активности по дням для каждого оператора)
            # -------------------------
            ws_all_active = wb.create_sheet(title="All active")
            # Заголовок: даты в формате dd.mm.yyyy
            ws_all_active.cell(1, 1).value = "ФИО"
            for col, dt in enumerate(dates, start=2):
                ws_all_active.cell(1, col).value = dt.strftime("%d.%m.%Y")
                ws_all_active.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
            total_col_idx = 1 + len(dates) + 1  # колонка для Итого (ч)
            ws_all_active.cell(1, total_col_idx).value = "Итого (ч)"

            # стиль шапки
            for c in range(1, total_col_idx + 1):
                cell = ws_all_active.cell(1, c)
                cell.font = Font(bold=True)
                cell.border = thin_border

            row = 2
            daily_totals_seconds = defaultdict(int)
            grand_total_seconds = 0

            for op_id in operators_dict.keys():
                name = operators_dict[op_id]
                name_cell = ws_all_active.cell(row, 1)
                name_cell.value = name
                name_cell.border = thin_border  # обводка ФИО
                name_cell.alignment = Alignment(horizontal='left', vertical='center')

                row_total_seconds = 0
                for col_idx, dt in enumerate(dates, start=2):
                    secs = active_seconds_per_op_per_day[op_id].get(dt, 0)
                    hours = secs / 3600.0
                    cell = ws_all_active.cell(row, col_idx)
                    cell.value = round(hours, 2)
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    # если сумма больше 0 — красим в серый
                    if hours > 0:
                        cell.fill = grey_fill

                    row_total_seconds += secs
                    daily_totals_seconds[dt] += secs

                # итого по строке (оператору) в часах
                row_total_hours = row_total_seconds / 3600.0
                total_cell = ws_all_active.cell(row, total_col_idx)
                total_cell.value = round(row_total_hours, 2)
                total_cell.number_format = '0.00'
                total_cell.alignment = Alignment(horizontal='center', vertical='center')
                total_cell.border = thin_border
                if row_total_hours > 0:
                    total_cell.fill = grey_fill

                grand_total_seconds += row_total_seconds
                row += 1

            # Добавим строку итогов по дням (и общий итог)
            total_row = row
            tot_label_cell = ws_all_active.cell(total_row, 1)
            tot_label_cell.value = "Итого"
            tot_label_cell.font = Font(bold=True)
            tot_label_cell.border = thin_border
            tot_label_cell.alignment = Alignment(horizontal='left', vertical='center')

            for col_idx, dt in enumerate(dates, start=2):
                secs = daily_totals_seconds.get(dt, 0)
                hours = secs / 3600.0
                cell = ws_all_active.cell(total_row, col_idx)
                cell.value = round(hours, 2)
                cell.number_format = '0.00'
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                if hours > 0:
                    cell.fill = grey_fill

            # общий итог всех операторов в часах
            grand_cell = ws_all_active.cell(total_row, total_col_idx)
            grand_cell.value = round(grand_total_seconds / 3600.0, 2)
            grand_cell.font = Font(bold=True)
            grand_cell.number_format = '0.00'
            grand_cell.alignment = Alignment(horizontal='center', vertical='center')
            grand_cell.border = thin_border
            if grand_total_seconds > 0:
                grand_cell.fill = grey_fill

            # автоширины и фиксация панелей
            ws_all_active.column_dimensions['A'].width = 24
            for idx in range(2, total_col_idx + 1):
                col_letter = ws_all_active.cell(1, idx).column_letter
                ws_all_active.column_dimensions[col_letter].width = 12
            ws_all_active.freeze_panes = "A2"

            # -------------------------
            # 2) Листы по операторам: события + одна строка итого на день с колонками итогов
            #    (E..I — ЧАСЫ с 2 знаками, J..N — секунды скрытые)
            # -------------------------
            for op_id, name in operators_dict.items():
                ws = wb.create_sheet(title=(name[:31] or f"op_{op_id}"))

                headers = [
                    "Дата", "Время / Описание", "Статус", "Длительность (ч)",
                    "Итого: Активен", "Итого: Подписание", "Итого: Перерыв", "Итого: Тренинг", "Итого: Тех",
                    "Итого Активен (сек)", "Итого Подписание (сек)", "Итого Перерыв (сек)", "Итого Тренинг (сек)", "Итого Тех (сек)"
                ]
                for col_idx, h in enumerate(headers, start=1):
                    cell = ws.cell(1, col_idx)
                    cell.value = h
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                for col_letter in ['J', 'K', 'L', 'M', 'N']:
                    ws.column_dimensions[col_letter].hidden = True

                ws.freeze_panes = "A2"
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 36
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 14
                ws.column_dimensions['E'].width = 16
                ws.column_dimensions['F'].width = 16
                ws.column_dimensions['G'].width = 14
                ws.column_dimensions['H'].width = 14
                ws.column_dimensions['I'].width = 14

                logs = sorted(logs_per_op[op_id], key=lambda x: x['change_time'])
                current_row = 2
                current_day = None
                day_status_times = defaultdict(timedelta)  # суммарное timedelta по статусам для дня
                day_start_row = None  # для объединения дат

                def write_day_totals_row(r, day_dt, counts_dict, day_status_times_dict):
                    """
                    Строка 'Итого' для оператора за день.
                    A — пустая объединяемая ячейка
                    B — "Итого"
                    C — "В работе" (ярко-зелёная заливка)
                    D — итог рабочих часов (active + iesigning) в часах, 2 знака
                    E..I — часы по статусам (число часов, 2 знака), J..N — секунды
                    Возвращает следующий свободный row (r + 1).

                    ВАЖНО: НЕ присваивает cell.fill = None (openpyxl ожидает объект Fill).
                            Вместо этого fill задаётся только когда нужно.
                    """
                    # fallback на случай отсутствия глобальных стилей
                    try:
                        _ = total_fill
                    except NameError:
                        total_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
                    try:
                        _ = thin_border
                    except NameError:
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                            top=Side(style='thin'), bottom=Side(style='thin'))
                    try:
                        _ = status_order
                    except NameError:
                        # default minimal status_order: (key, name, fill)
                        status_order = [
                            ('active', 'Активен', PatternFill(start_color="CFF5D0", end_color="CFF5D0", fill_type="solid")),
                            ('iesigning', 'Подписание', PatternFill(start_color="DCEEFF", end_color="DCEEFF", fill_type="solid")),
                            ('break', 'Перерыв', PatternFill(start_color="FFE8C0", end_color="FFE8C0", fill_type="solid")),
                            ('training', 'Тренинг', PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")),
                            ('tech', 'Перерыв', PatternFill(start_color="FFE8C0", end_color="FFE8C0", fill_type="solid")),
                        ]

                    # жирная нижняя граница (для визуального раздела)
                    bottom_bold_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='medium')
                    )

                    # насыщённый зелёный цвет для "В работе" и часов
                    strong_green_fill = PatternFill(start_color="66CC66", end_color="66CC66", fill_type="solid")

                    # A — пустая объединяемая ячейка (обычно объединена выше)
                    a_cell = ws.cell(r, 1)
                    a_cell.fill = total_fill
                    a_cell.border = bottom_bold_border
                    a_cell.alignment = Alignment(horizontal='center', vertical='center')

                    # B — "Итого"
                    b_cell = ws.cell(r, 2)
                    b_cell.value = "Итого"
                    b_cell.font = Font(bold=True)
                    b_cell.alignment = Alignment(horizontal='center', vertical='center')
                    b_cell.fill = total_fill
                    b_cell.border = bottom_bold_border

                    # C — "В работе" (ярко-зелёный)
                    c_cell = ws.cell(r, 3)
                    c_cell.value = "В работе"
                    c_cell.font = Font(bold=True)
                    c_cell.alignment = Alignment(horizontal='center', vertical='center')
                    c_cell.fill = strong_green_fill
                    c_cell.border = bottom_bold_border

                    # D — итог рабочих часов = active + iesigning (в часах)
                    active_td = day_status_times_dict.get('active', timedelta(0))
                    iesigning_td = day_status_times_dict.get('iesigning', timedelta(0))
                    total_work_sec = int((active_td + iesigning_td).total_seconds())
                    total_work_hours = round(total_work_sec / 3600.0, 2)

                    d_cell = ws.cell(r, 4)
                    d_cell.value = total_work_hours
                    d_cell.number_format = '0.00'
                    d_cell.alignment = Alignment(horizontal='center', vertical='center')
                    # ставим заливку ТОЛЬКО если > 0
                    if total_work_hours > 0:
                        d_cell.fill = strong_green_fill
                    d_cell.border = bottom_bold_border

                    # E..I (часы по статусам) и J..N (секунды) — заполняем и даём заливку по статусу, если >0
                    for idx, (status_key, status_label, fill) in enumerate(status_order):
                        dur_td = day_status_times_dict.get(status_key, timedelta(0))
                        dur_sec = int(dur_td.total_seconds())
                        dur_hours = round(dur_sec / 3600.0, 2)

                        col_read = 5 + idx   # E..I
                        col_sec = 10 + idx   # J..N

                        # часы (число)
                        rcell = ws.cell(r, col_read)
                        rcell.value = dur_hours
                        rcell.number_format = '0.00'
                        rcell.alignment = Alignment(horizontal='center', vertical='center')
                        # заливаем ТОЛЬКО если >0 и у нас есть валидная заливка
                        if dur_hours > 0 and isinstance(fill, PatternFill):
                            rcell.fill = fill
                        rcell.border = bottom_bold_border

                        # секунды (скрытые колонки) — всегда записываем число
                        scell = ws.cell(r, col_sec)
                        scell.value = dur_sec
                        scell.alignment = Alignment(horizontal='center', vertical='center')
                        scell.border = bottom_bold_border

                    return r + 1


                for i, log in enumerate(logs):
                    dt = log['date']
                    if current_day is None:
                        current_day = dt
                        day_start_row = current_row  # первый ряд для этого дня (включая будущий "Итого")
                    if dt != current_day:
                        # Перед записью итоговой строки — объединяем столбец A для предыдущего дня:
                        if day_start_row is not None:
                            merge_end = current_row  # текущ_row — место где будет написан "Итого"
                            ws.merge_cells(start_row=day_start_row, start_column=1, end_row=merge_end, end_column=1)
                            merged_cell = ws.cell(day_start_row, 1)
                            merged_cell.value = current_day.strftime("%d.%m.%Y")
                            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

                        # записать строку итогов предыдущего дня
                        counts = counts_per_op[op_id][current_day]
                        current_row = write_day_totals_row(current_row, current_day, counts, day_status_times)
                        # сброс
                        day_status_times = defaultdict(timedelta)
                        # переходим к новому дню
                        current_day = dt
                        day_start_row = current_row  # первый ряд нового дня (включая будущий Итого)

                    # next_time
                    if i < len(logs) - 1:
                        next_time = logs[i + 1]['change_time']
                    else:
                        next_time = end_time

                    duration = next_time - log['change_time']

                    # дубликат состояния?
                    is_duplicate = (i > 0 and log['is_active'] == logs[i - 1]['is_active'])
                    if is_duplicate:
                        dur_hours = None
                    else:
                        dur_hours = round(duration.total_seconds() / 3600.0, 2)
                        if log['is_active'] != 'inactive':
                            day_status_times[log['is_active']] += duration

                    # пишем событие (A..D)
                    # A оставляем пустой — объединится позже
                    ws.cell(current_row, 2).value = log['change_time'].strftime("%H:%M:%S")
                    # статус — русская версия
                    status_key = log['is_active']
                    status_rus = status_display.get(status_key, status_key)
                    ws.cell(current_row, 3).value = status_rus

                    # длительность — ЧАСЫ числом с 2 знаками или пусто для дубликата
                    dur_cell = ws.cell(current_row, 4)
                    if dur_hours is None:
                        dur_cell.value = None
                    else:
                        dur_cell.value = dur_hours
                        dur_cell.number_format = '0.00'
                        # окрашиваем длительность в цвет статуса (если сопоставление есть)
                        fill = status_fill_map.get(status_key)
                        if fill is not None:
                            dur_cell.fill = fill

                    # заливка статуса в колонке C (сам статус)
                    cell_stat = ws.cell(current_row, 3)
                    fill_for_cell = status_fill_map.get(status_key)
                    if fill_for_cell is not None:
                        cell_stat.fill = fill_for_cell
                    else:
                        # fallback
                        if status_key == 'inactive':
                            cell_stat.fill = inactive_fill

                    # границы и выравнивание для строки события
                    for col_idx in range(1, 15):
                        ccell = ws.cell(current_row, col_idx)
                        ccell.border = thin_border
                        if col_idx == 4:
                            ccell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            ccell.alignment = Alignment(horizontal='left', vertical='center')

                    current_row += 1

                # итоги для последнего дня
                if current_day is not None:
                    if day_start_row is not None:
                        merge_end = current_row  # сюда запишется "Итого"
                        ws.merge_cells(start_row=day_start_row, start_column=1, end_row=merge_end, end_column=1)
                        merged_cell = ws.cell(day_start_row, 1)
                        merged_cell.value = current_day.strftime("%d.%m.%Y")
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

                    counts = counts_per_op[op_id][current_day]
                    current_row = write_day_totals_row(current_row, current_day, counts, day_status_times)

                # добавляем таблицу, охватывающую A..N
                if current_row > 2:
                    last_row = current_row - 1
                    sanitized_name = sanitize_table_name(f"{name}_{op_id}")
                    tab_ref = f"A1:N{last_row}"
                    tab_op = Table(displayName=f"Table_{sanitized_name}", ref=tab_ref)
                    style_op = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showRowStripes=True)
                    tab_op.tableStyleInfo = style_op
                    ws.add_table(tab_op)

            # Сохранение в BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            content = output.getvalue()
            return filename, content

        except Exception as e:
            logging.error(f"Error generating report: {e}")
            return None, None



    def add_training(self, operator_id, training_date, start_time, end_time, reason, comment, created_by, count_in_hours=True):
        with self._get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO trainings (operator_id, training_date, start_time, end_time, reason, comment, created_by, count_in_hours)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (operator_id, training_date, start_time, end_time, reason, comment, created_by, count_in_hours))
            return cursor.fetchone()[0]


    def get_trainings(self, requester_id=None, month=None):
        """
        Получить тренинги для оператора/всех, в зависимости от роли requester_id.
        Если operator_id указан — только для него.
        Если requester_id — определяет роль и фильтрует:
            - admin, sv: все тренинги (с фильтром по month, если указан)
            - operator: только свои тренинги (с фильтром по month, если указан)
        """
        # Получаем роль requester_id
        role = None
        if requester_id:
            with self._get_cursor() as cursor:
                cursor.execute("SELECT role FROM users WHERE id = %s", (requester_id,))
                res = cursor.fetchone()
                if res:
                    role = res[0]
        query = """
            SELECT t.id, t.operator_id, t.training_date, t.start_time, t.end_time, t.reason, t.comment, t.created_at, cb.name as created_by_name, t.count_in_hours
            FROM trainings t
            JOIN users u ON t.operator_id = u.id
            LEFT JOIN users cb ON t.created_by = cb.id
        """
        params = []
        where_clauses = []
        if month:
            where_clauses.append("TO_CHAR(t.training_date, 'YYYY-MM') = %s")
            params.append(month)
        if role == "operator":
            where_clauses.append("t.operator_id = %s")
            params.append(requester_id)
        # Для admin — без ограничений, кроме month
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
        query += " ORDER BY t.training_date DESC, t.start_time DESC"
        with self._get_cursor() as cursor:
            cursor.execute(query, params)
            return [
                {
                    "id": row[0],
                    "operator_id": row[1],
                    "date": row[2].strftime('%Y-%m-%d'),
                    "start_time": row[3].strftime('%H:%M'),
                    "end_time": row[4].strftime('%H:%M'),
                    "reason": row[5],
                    "comment": row[6],
                    "created_at": row[7].strftime('%Y-%m-%d %H:%M'),
                    "created_by_name": row[8] if row[8] else "System",
                    "count_in_hours": bool(row[9])
                } for row in cursor.fetchall()
            ]

    def _normalize_technical_issue_role(self, role_value):
        return normalize_role_value(role_value)

    def _coerce_int_list(self, values):
        if values is None:
            return []
        if isinstance(values, (list, tuple, set)):
            raw_values = list(values)
        else:
            raw_values = [values]

        normalized = []
        seen = set()
        for item in raw_values:
            if item is None:
                continue
            if isinstance(item, str) and not item.strip():
                continue
            try:
                ivalue = int(item)
            except (TypeError, ValueError):
                continue
            if ivalue <= 0 or ivalue in seen:
                continue
            seen.add(ivalue)
            normalized.append(ivalue)
        return normalized

    def _parse_technical_issue_date(self, value, field_name='date'):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if not value:
            raise ValueError(f"Field '{field_name}' is required (YYYY-MM-DD)")
        try:
            return datetime.strptime(str(value), '%Y-%m-%d').date()
        except (TypeError, ValueError):
            raise ValueError(f"Invalid '{field_name}' format. Use YYYY-MM-DD")

    def _parse_technical_issue_time(self, value, field_name='start_time', default_value=None):
        raw_value = value
        if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
            if default_value is None:
                raise ValueError(f"Field '{field_name}' is required (HH:MM)")
            raw_value = default_value

        if isinstance(raw_value, dt_time):
            return raw_value.replace(second=0, microsecond=0)

        raw_text = str(raw_value).strip()
        for fmt in ('%H:%M', '%H:%M:%S'):
            try:
                parsed = datetime.strptime(raw_text, fmt).time()
                return parsed.replace(second=0, microsecond=0)
            except (TypeError, ValueError):
                continue
        raise ValueError(f"Invalid '{field_name}' format. Use HH:MM")

    def _technical_issue_time_to_hhmm(self, time_value):
        if isinstance(time_value, dt_time):
            return time_value.strftime('%H:%M')
        if isinstance(time_value, str):
            try:
                parsed = self._parse_technical_issue_time(time_value, field_name='time')
                return parsed.strftime('%H:%M')
            except ValueError:
                return str(time_value)
        return str(time_value or '')

    def _split_absolute_minutes_intervals_by_day(self, intervals, anchor_date_obj):
        """
        Splits absolute-minute intervals (relative to anchor_date midnight) into day chunks.
        Returns list of dicts: {'date': date, 'start_min': int, 'end_min': int, 'start_time': 'HH:MM', 'end_time': 'HH:MM'}.
        """
        anchor_date_obj = self._parse_technical_issue_date(anchor_date_obj, field_name='date')
        merged_intervals = self._merge_break_intervals(intervals or [])
        result = []
        for interval in merged_intervals:
            seg_start = int(interval.get('start', 0))
            seg_end = int(interval.get('end', 0))
            if seg_end <= seg_start:
                continue

            cursor_min = seg_start
            while cursor_min < seg_end:
                day_offset = int(cursor_min // 1440)
                day_floor = int(day_offset) * 1440
                day_ceil = day_floor + 1440
                chunk_end = min(seg_end, day_ceil)
                if chunk_end <= cursor_min:
                    break

                local_start = int(cursor_min - day_floor)
                local_end = int(chunk_end - day_floor)
                if local_end > local_start:
                    issue_day = anchor_date_obj + timedelta(days=day_offset)
                    result.append({
                        'date': issue_day,
                        'start_min': local_start,
                        'end_min': local_end,
                        'start_time': _minutes_to_time(local_start),
                        'end_time': _minutes_to_time(local_end)
                    })
                cursor_min = chunk_end
        return result

    def _find_shift_overlap_intervals_for_technical_issue_tx(
        self,
        cursor,
        operator_ids,
        issue_date_obj,
        start_time_obj,
        end_time_obj
    ):
        operator_ids_norm = self._coerce_int_list(operator_ids)
        if not operator_ids_norm:
            return {}

        issue_date_obj = self._parse_technical_issue_date(issue_date_obj, field_name='date')
        issue_start_min, issue_end_min = self._schedule_interval_minutes(start_time_obj, end_time_obj)
        issue_start_abs = int(issue_start_min)
        issue_end_abs = int(issue_end_min)
        if issue_end_abs <= issue_start_abs:
            return {}

        shifts_source_start = issue_date_obj - timedelta(days=1)
        shifts_source_end = issue_date_obj + timedelta(days=1)
        cursor.execute(
            """
            SELECT operator_id, shift_date, start_time, end_time
            FROM work_shifts
            WHERE operator_id = ANY(%s)
              AND shift_date >= %s
              AND shift_date <= %s
            ORDER BY operator_id, shift_date, start_time, end_time
            """,
            (operator_ids_norm, shifts_source_start, shifts_source_end)
        )
        rows = cursor.fetchall() or []

        overlaps_by_operator = {}
        for operator_id, shift_date_value, shift_start_value, shift_end_value in rows:
            if shift_date_value is None:
                continue
            day_offset = int((shift_date_value - issue_date_obj).days)
            shift_start_min, shift_end_min = self._schedule_interval_minutes(shift_start_value, shift_end_value)
            shift_start_abs = day_offset * 1440 + int(shift_start_min)
            shift_end_abs = day_offset * 1440 + int(shift_end_min)

            overlap_start = max(issue_start_abs, shift_start_abs)
            overlap_end = min(issue_end_abs, shift_end_abs)
            if overlap_end <= overlap_start:
                continue
            overlaps_by_operator.setdefault(int(operator_id), []).append({
                'start': int(overlap_start),
                'end': int(overlap_end)
            })

        for op_id in list(overlaps_by_operator.keys()):
            overlaps_by_operator[op_id] = self._merge_break_intervals(overlaps_by_operator.get(op_id) or [])
            if not overlaps_by_operator[op_id]:
                overlaps_by_operator.pop(op_id, None)

        return overlaps_by_operator

    def _normalize_direction_ids_json_payload(self, payload):
        if payload is None:
            return []
        parsed = payload
        if isinstance(payload, str):
            payload = payload.strip()
            if not payload:
                return []
            try:
                parsed = json.loads(payload)
            except Exception:
                return []
        if not isinstance(parsed, list):
            return []
        result = []
        seen = set()
        for item in parsed:
            try:
                direction_id = int(item)
            except (TypeError, ValueError):
                continue
            if direction_id <= 0 or direction_id in seen:
                continue
            seen.add(direction_id)
            result.append(direction_id)
        return result

    def _get_direction_name_map_tx(self, cursor, direction_ids):
        normalized_ids = self._coerce_int_list(direction_ids)
        if not normalized_ids:
            return {}
        cursor.execute("""
            SELECT id, name
            FROM directions
            WHERE id = ANY(%s)
        """, (normalized_ids,))
        return {int(row[0]): row[1] for row in cursor.fetchall()}

    def _resolve_technical_issue_operator_ids_tx(
        self,
        cursor,
        requester_id,
        requester_role,
        operator_ids=None,
        direction_ids=None,
        enforce_supervisor_scope=True
    ):
        role_norm = self._normalize_technical_issue_role(requester_role)
        if not (role_has_min(role_norm, 'admin') or role_norm == 'sv'):
            raise ValueError("Only admin and sv can register technical issues")

        operator_ids_norm = self._coerce_int_list(operator_ids)
        direction_ids_norm = self._coerce_int_list(direction_ids)
        target_operator_ids = set()

        if operator_ids_norm:
            cursor.execute("""
                SELECT id, supervisor_id
                FROM users
                WHERE role = 'operator' AND id = ANY(%s)
            """, (operator_ids_norm,))
            rows = cursor.fetchall()
            found_ids = {int(row[0]) for row in rows}
            missing_ids = [op_id for op_id in operator_ids_norm if op_id not in found_ids]
            if missing_ids:
                raise ValueError(f"Operators not found: {', '.join(map(str, missing_ids))}")

            if role_norm == 'sv' and enforce_supervisor_scope:
                forbidden_ids = [int(op_id) for op_id, supervisor_id in rows if int(supervisor_id or 0) != int(requester_id)]
                if forbidden_ids:
                    raise ValueError(f"Forbidden operators for sv: {', '.join(map(str, forbidden_ids))}")
            target_operator_ids.update(found_ids)

        if direction_ids_norm:
            cursor.execute("""
                SELECT id
                FROM directions
                WHERE id = ANY(%s) AND is_active = TRUE
            """, (direction_ids_norm,))
            valid_direction_ids = {int(row[0]) for row in cursor.fetchall()}
            missing_directions = [dir_id for dir_id in direction_ids_norm if dir_id not in valid_direction_ids]
            if missing_directions:
                raise ValueError(f"Directions not found or inactive: {', '.join(map(str, missing_directions))}")

            if role_norm == 'sv' and enforce_supervisor_scope:
                cursor.execute("""
                    SELECT id
                    FROM users
                    WHERE role = 'operator'
                      AND direction_id = ANY(%s)
                      AND supervisor_id = %s
                      AND COALESCE(status, 'working') <> 'fired'
                """, (direction_ids_norm, int(requester_id)))
            else:
                cursor.execute("""
                    SELECT id
                    FROM users
                    WHERE role = 'operator'
                      AND direction_id = ANY(%s)
                      AND COALESCE(status, 'working') <> 'fired'
                """, (direction_ids_norm,))
            target_operator_ids.update(int(row[0]) for row in cursor.fetchall())

        if not target_operator_ids:
            raise ValueError("Select at least one operator or direction with active operators")

        return sorted(target_operator_ids), direction_ids_norm

    def get_technical_issue_reasons(self):
        return list(TECHNICAL_ISSUE_REASONS)

    def create_operator_technical_issues(
        self,
        requester_id,
        requester_role,
        issue_date,
        reason,
        start_time=None,
        end_time=None,
        comment=None,
        operator_ids=None,
        direction_ids=None
    ):
        role_norm = self._normalize_technical_issue_role(requester_role)
        if not (role_has_min(role_norm, 'admin') or role_norm == 'sv'):
            raise ValueError("Only admin and sv can create technical issues")

        reason_text = str(reason or '').strip()
        if reason_text not in TECHNICAL_ISSUE_REASONS_SET:
            raise ValueError("Invalid technical issue reason")

        issue_date_obj = self._parse_technical_issue_date(issue_date, field_name='date')
        start_time_obj = self._parse_technical_issue_time(start_time, field_name='start_time', default_value='00:00')
        end_time_obj = self._parse_technical_issue_time(end_time, field_name='end_time', default_value='23:59')
        if start_time_obj == end_time_obj:
            raise ValueError("start_time and end_time cannot be equal")
        comment_text = str(comment or '').strip() or None
        requester_id_int = int(requester_id)

        with self._get_cursor() as cursor:
            target_operator_ids, selected_direction_ids = self._resolve_technical_issue_operator_ids_tx(
                cursor=cursor,
                requester_id=requester_id_int,
                requester_role=role_norm,
                operator_ids=operator_ids,
                direction_ids=direction_ids,
                enforce_supervisor_scope=False
            )

            overlaps_by_operator = self._find_shift_overlap_intervals_for_technical_issue_tx(
                cursor=cursor,
                operator_ids=target_operator_ids,
                issue_date_obj=issue_date_obj,
                start_time_obj=start_time_obj,
                end_time_obj=end_time_obj
            )
            eligible_operator_ids = sorted(overlaps_by_operator.keys())
            if not eligible_operator_ids:
                raise ValueError("No selected operators have shifts in the specified time range")

            batch_id = str(uuid.uuid4())
            issue_values = []

            for operator_id in eligible_operator_ids:
                day_chunks = self._split_absolute_minutes_intervals_by_day(
                    overlaps_by_operator.get(operator_id) or [],
                    issue_date_obj
                )
                for chunk in day_chunks:
                    issue_values.append((
                        batch_id,
                        int(operator_id),
                        chunk['date'],
                        chunk['start_time'],
                        chunk['end_time'],
                        reason_text,
                        comment_text,
                        Json(selected_direction_ids),
                        requester_id_int
                    ))

            if not issue_values:
                raise ValueError("No overlapping shift intervals were found")

            execute_values(
                cursor,
                """
                    INSERT INTO operator_technical_issues (
                        batch_id,
                        operator_id,
                        issue_date,
                        start_time,
                        end_time,
                        reason,
                        comment,
                        direction_ids,
                        created_by
                    )
                    VALUES %s
                """,
                issue_values
            )

            direction_name_map = self._get_direction_name_map_tx(cursor, selected_direction_ids)
            skipped_operator_ids = [op_id for op_id in target_operator_ids if op_id not in set(eligible_operator_ids)]
            return {
                'batch_id': batch_id,
                'created_count': len(issue_values),
                'created_operator_count': len(eligible_operator_ids),
                'operator_ids': eligible_operator_ids,
                'requested_operator_ids': target_operator_ids,
                'requested_operator_count': len(target_operator_ids),
                'skipped_operator_ids': skipped_operator_ids,
                'skipped_operator_count': len(skipped_operator_ids),
                'selected_direction_ids': selected_direction_ids,
                'selected_direction_names': [direction_name_map.get(dir_id) for dir_id in selected_direction_ids if direction_name_map.get(dir_id)],
                'date': issue_date_obj.strftime('%Y-%m-%d'),
                'start_time': start_time_obj.strftime('%H:%M'),
                'end_time': end_time_obj.strftime('%H:%M'),
                'reason': reason_text
            }

    def get_operator_technical_issues(
        self,
        requester_id,
        requester_role,
        issue_date=None,
        date_from=None,
        date_to=None,
        operator_id=None,
        reason=None,
        limit=500,
        offset=0
    ):
        role_norm = self._normalize_technical_issue_role(requester_role)
        if not (role_has_min(role_norm, 'admin') or role_norm == 'sv'):
            raise ValueError("Only admin and sv can view technical issues")

        limit_int = max(1, min(int(limit or 500), 5000))
        offset_int = max(0, int(offset or 0))
        where_parts = []
        params = []

        if issue_date:
            issue_date_obj = self._parse_technical_issue_date(issue_date, field_name='date')
            where_parts.append("ti.issue_date = %s")
            params.append(issue_date_obj)
        else:
            if date_from:
                from_obj = self._parse_technical_issue_date(date_from, field_name='date_from')
                where_parts.append("ti.issue_date >= %s")
                params.append(from_obj)
            if date_to:
                to_obj = self._parse_technical_issue_date(date_to, field_name='date_to')
                where_parts.append("ti.issue_date <= %s")
                params.append(to_obj)

        operator_id_int = None
        if operator_id is not None and str(operator_id).strip() != '':
            try:
                operator_id_int = int(operator_id)
            except (TypeError, ValueError):
                raise ValueError("Invalid operator_id")
            where_parts.append("ti.operator_id = %s")
            params.append(operator_id_int)

        reason_text = str(reason or '').strip()
        if reason_text:
            if reason_text not in TECHNICAL_ISSUE_REASONS_SET:
                raise ValueError("Invalid reason filter")
            where_parts.append("ti.reason = %s")
            params.append(reason_text)

        where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
        base_sql = """
            FROM operator_technical_issues ti
            JOIN users op ON op.id = ti.operator_id
            LEFT JOIN users sv ON sv.id = op.supervisor_id
            LEFT JOIN users cb ON cb.id = ti.created_by
            LEFT JOIN directions d ON d.id = op.direction_id
        """

        with self._get_cursor() as cursor:
            cursor.execute(f"SELECT COUNT(*) {base_sql} {where_sql}", params)
            total = int(cursor.fetchone()[0] or 0)

            cursor.execute(
                f"""
                    SELECT
                        ti.id,
                        ti.batch_id::text,
                        ti.operator_id,
                        op.name,
                        op.supervisor_id,
                        sv.name,
                        op.direction_id,
                        d.name,
                        ti.issue_date,
                        ti.start_time,
                        ti.end_time,
                        ti.reason,
                        ti.comment,
                        ti.created_by,
                        cb.name,
                        ti.created_at,
                        ti.direction_ids
                    {base_sql}
                    {where_sql}
                    ORDER BY ti.issue_date DESC, ti.start_time DESC, ti.created_at DESC, ti.id DESC
                    LIMIT %s OFFSET %s
                """,
                params + [limit_int, offset_int]
            )
            rows = cursor.fetchall()

            all_selected_direction_ids = set()
            prepared_rows = []
            for row in rows:
                selected_direction_ids = self._normalize_direction_ids_json_payload(row[16])
                all_selected_direction_ids.update(selected_direction_ids)
                prepared_rows.append((row, selected_direction_ids))

            direction_name_map = self._get_direction_name_map_tx(cursor, list(all_selected_direction_ids))
            items = []
            for row, selected_direction_ids in prepared_rows:
                start_time_text = row[9].strftime('%H:%M') if row[9] else None
                end_time_text = row[10].strftime('%H:%M') if row[10] else None
                items.append({
                    'id': int(row[0]),
                    'batch_id': row[1],
                    'operator_id': int(row[2]),
                    'operator_name': row[3],
                    'supervisor_id': int(row[4]) if row[4] is not None else None,
                    'supervisor_name': row[5],
                    'direction_id': int(row[6]) if row[6] is not None else None,
                    'direction_name': row[7],
                    'date': row[8].strftime('%Y-%m-%d') if row[8] else None,
                    'start_time': start_time_text,
                    'end_time': end_time_text,
                    'time_range': f"{start_time_text} - {end_time_text}" if start_time_text and end_time_text else None,
                    'reason': row[11],
                    'comment': row[12],
                    'created_by_id': int(row[13]) if row[13] is not None else None,
                    'created_by_name': row[14],
                    'created_at': row[15].strftime('%Y-%m-%d %H:%M:%S') if row[15] else None,
                    'selected_direction_ids': selected_direction_ids,
                    'selected_direction_names': [direction_name_map.get(dir_id) for dir_id in selected_direction_ids if direction_name_map.get(dir_id)]
                })

            return {
                'total': total,
                'items': items
            }

    def delete_operator_technical_issue(self, requester_id, requester_role, issue_id):
        role_norm = self._normalize_technical_issue_role(requester_role)
        if not (role_has_min(role_norm, 'admin') or role_norm == 'sv'):
            raise ValueError("Only admin and sv can delete technical issues")

        requester_id_int = int(requester_id)
        try:
            issue_id_int = int(issue_id)
        except (TypeError, ValueError):
            raise ValueError("Invalid technical issue id")
        if issue_id_int <= 0:
            raise ValueError("Invalid technical issue id")

        with self._get_cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    ti.id,
                    ti.batch_id::text,
                    ti.operator_id,
                    op.supervisor_id,
                    ti.created_by,
                    ti.issue_date,
                    ti.start_time,
                    ti.end_time,
                    ti.reason
                FROM operator_technical_issues ti
                JOIN users op ON op.id = ti.operator_id
                WHERE ti.id = %s
                """,
                (issue_id_int,)
            )
            row = cursor.fetchone()
            if not row:
                raise ValueError("Technical issue not found")

            supervisor_id = int(row[3]) if row[3] is not None else None
            issue_created_by = int(row[4]) if row[4] is not None else None
            if role_norm == 'sv':
                can_manage_by_scope = supervisor_id == requester_id_int
                can_manage_by_creator = issue_created_by == requester_id_int
                if not (can_manage_by_scope or can_manage_by_creator):
                    raise PermissionError("Forbidden")

            cursor.execute(
                """
                DELETE FROM operator_technical_issues
                WHERE id = %s
                RETURNING id
                """,
                (issue_id_int,)
            )
            deleted_row = cursor.fetchone()
            if not deleted_row:
                raise ValueError("Technical issue not found")

            return {
                'id': int(row[0]),
                'batch_id': row[1],
                'operator_id': int(row[2]),
                'date': row[5].strftime('%Y-%m-%d') if row[5] else None,
                'start_time': row[6].strftime('%H:%M') if row[6] else None,
                'end_time': row[7].strftime('%H:%M') if row[7] else None,
                'reason': row[8]
            }

    def create_operator_offline_activity(
        self,
        requester_id,
        requester_role,
        activity_date,
        start_time=None,
        end_time=None,
        comment=None,
        operator_id=None
    ):
        role_norm = self._normalize_technical_issue_role(requester_role)
        if not (role_has_min(role_norm, 'admin') or role_norm == 'sv'):
            raise ValueError("Only admin and sv can create offline activity")

        issue_date_obj = self._parse_technical_issue_date(activity_date, field_name='date')
        start_time_obj = self._parse_technical_issue_time(start_time, field_name='start_time', default_value='00:00')
        end_time_obj = self._parse_technical_issue_time(end_time, field_name='end_time', default_value='23:59')
        if start_time_obj == end_time_obj:
            raise ValueError("start_time and end_time cannot be equal")

        operator_ids_norm = self._coerce_int_list([operator_id])
        if not operator_ids_norm:
            raise ValueError("Field 'operator_id' is required")

        comment_text = str(comment or '').strip() or None
        requester_id_int = int(requester_id)

        with self._get_cursor() as cursor:
            target_operator_ids, _ = self._resolve_technical_issue_operator_ids_tx(
                cursor=cursor,
                requester_id=requester_id_int,
                requester_role=role_norm,
                operator_ids=operator_ids_norm,
                direction_ids=None
            )
            target_operator_id = int(target_operator_ids[0])

            overlaps_by_operator = self._find_shift_overlap_intervals_for_technical_issue_tx(
                cursor=cursor,
                operator_ids=[target_operator_id],
                issue_date_obj=issue_date_obj,
                start_time_obj=start_time_obj,
                end_time_obj=end_time_obj
            )
            eligible_operator_ids = sorted(overlaps_by_operator.keys())
            if not eligible_operator_ids:
                raise ValueError("Selected operator has no shifts in the specified time range")

            batch_id = str(uuid.uuid4())
            activity_values = []
            for resolved_operator_id in eligible_operator_ids:
                day_chunks = self._split_absolute_minutes_intervals_by_day(
                    overlaps_by_operator.get(resolved_operator_id) or [],
                    issue_date_obj
                )
                for chunk in day_chunks:
                    activity_values.append((
                        batch_id,
                        int(resolved_operator_id),
                        chunk['date'],
                        chunk['start_time'],
                        chunk['end_time'],
                        comment_text,
                        requester_id_int
                    ))

            if not activity_values:
                raise ValueError("No overlapping shift intervals were found")

            execute_values(
                cursor,
                """
                    INSERT INTO operator_offline_activities (
                        batch_id,
                        operator_id,
                        activity_date,
                        start_time,
                        end_time,
                        comment,
                        created_by
                    )
                    VALUES %s
                """,
                activity_values
            )

            return {
                'batch_id': batch_id,
                'created_count': len(activity_values),
                'created_operator_count': len(eligible_operator_ids),
                'operator_ids': eligible_operator_ids,
                'date': issue_date_obj.strftime('%Y-%m-%d'),
                'start_time': start_time_obj.strftime('%H:%M'),
                'end_time': end_time_obj.strftime('%H:%M'),
                'comment': comment_text
            }

    def get_operator_offline_activities(
        self,
        requester_id,
        requester_role,
        activity_date=None,
        date_from=None,
        date_to=None,
        operator_id=None,
        limit=500,
        offset=0
    ):
        role_norm = self._normalize_technical_issue_role(requester_role)
        if not (role_has_min(role_norm, 'admin') or role_norm == 'sv'):
            raise ValueError("Only admin and sv can view offline activities")

        limit_int = max(1, min(int(limit or 500), 5000))
        offset_int = max(0, int(offset or 0))
        requester_id_int = int(requester_id)

        where_parts = []
        params = []
        if role_norm == 'sv':
            where_parts.append("op.supervisor_id = %s")
            params.append(requester_id_int)

        if activity_date:
            activity_date_obj = self._parse_technical_issue_date(activity_date, field_name='date')
            where_parts.append("oa.activity_date = %s")
            params.append(activity_date_obj)
        else:
            if date_from:
                from_obj = self._parse_technical_issue_date(date_from, field_name='date_from')
                where_parts.append("oa.activity_date >= %s")
                params.append(from_obj)
            if date_to:
                to_obj = self._parse_technical_issue_date(date_to, field_name='date_to')
                where_parts.append("oa.activity_date <= %s")
                params.append(to_obj)

        operator_id_int = None
        if operator_id is not None and str(operator_id).strip() != '':
            try:
                operator_id_int = int(operator_id)
            except (TypeError, ValueError):
                raise ValueError("Invalid operator_id")
            where_parts.append("oa.operator_id = %s")
            params.append(operator_id_int)

        where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
        base_sql = """
            FROM operator_offline_activities oa
            JOIN users op ON op.id = oa.operator_id
            LEFT JOIN users sv ON sv.id = op.supervisor_id
            LEFT JOIN users cb ON cb.id = oa.created_by
            LEFT JOIN directions d ON d.id = op.direction_id
        """

        with self._get_cursor() as cursor:
            cursor.execute(f"SELECT COUNT(*) {base_sql} {where_sql}", params)
            total = int(cursor.fetchone()[0] or 0)

            cursor.execute(
                f"""
                    SELECT
                        oa.id,
                        oa.batch_id::text,
                        oa.operator_id,
                        op.name,
                        op.supervisor_id,
                        sv.name,
                        op.direction_id,
                        d.name,
                        oa.activity_date,
                        oa.start_time,
                        oa.end_time,
                        oa.comment,
                        oa.created_by,
                        cb.name,
                        oa.created_at
                    {base_sql}
                    {where_sql}
                    ORDER BY oa.activity_date DESC, oa.start_time DESC, oa.created_at DESC, oa.id DESC
                    LIMIT %s OFFSET %s
                """,
                params + [limit_int, offset_int]
            )
            rows = cursor.fetchall() or []

            items = []
            for row in rows:
                start_time_text = row[9].strftime('%H:%M') if row[9] else None
                end_time_text = row[10].strftime('%H:%M') if row[10] else None
                duration_minutes = 0
                try:
                    if row[9] and row[10]:
                        start_min, end_min = self._schedule_interval_minutes(row[9], row[10])
                        duration_minutes = max(0, int(end_min - start_min))
                except Exception:
                    duration_minutes = 0
                items.append({
                    'id': int(row[0]),
                    'batch_id': row[1],
                    'operator_id': int(row[2]),
                    'operator_name': row[3],
                    'supervisor_id': int(row[4]) if row[4] is not None else None,
                    'supervisor_name': row[5],
                    'direction_id': int(row[6]) if row[6] is not None else None,
                    'direction_name': row[7],
                    'date': row[8].strftime('%Y-%m-%d') if row[8] else None,
                    'start_time': start_time_text,
                    'end_time': end_time_text,
                    'time_range': f"{start_time_text} - {end_time_text}" if start_time_text and end_time_text else None,
                    'comment': row[11],
                    'created_by_id': int(row[12]) if row[12] is not None else None,
                    'created_by_name': row[13],
                    'created_at': row[14].strftime('%Y-%m-%d %H:%M:%S') if row[14] else None,
                    'duration_minutes': int(duration_minutes),
                    'duration_hours': round(float(duration_minutes) / 60.0, 2)
                })

            return {
                'total': total,
                'items': items
            }

    def delete_operator_offline_activity(self, requester_id, requester_role, activity_id):
        role_norm = self._normalize_technical_issue_role(requester_role)
        if not (role_has_min(role_norm, 'admin') or role_norm == 'sv'):
            raise ValueError("Only admin and sv can delete offline activities")

        requester_id_int = int(requester_id)
        try:
            activity_id_int = int(activity_id)
        except (TypeError, ValueError):
            raise ValueError("Invalid offline activity id")
        if activity_id_int <= 0:
            raise ValueError("Invalid offline activity id")

        with self._get_cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    oa.id,
                    oa.batch_id::text,
                    oa.operator_id,
                    op.supervisor_id,
                    oa.activity_date,
                    oa.start_time,
                    oa.end_time,
                    oa.comment
                FROM operator_offline_activities oa
                JOIN users op ON op.id = oa.operator_id
                WHERE oa.id = %s
                """,
                (activity_id_int,)
            )
            row = cursor.fetchone()
            if not row:
                raise ValueError("Offline activity not found")

            supervisor_id = int(row[3]) if row[3] is not None else None
            if role_norm == 'sv' and supervisor_id != requester_id_int:
                raise PermissionError("Forbidden")

            cursor.execute(
                """
                DELETE FROM operator_offline_activities
                WHERE id = %s
                RETURNING id
                """,
                (activity_id_int,)
            )
            deleted_row = cursor.fetchone()
            if not deleted_row:
                raise ValueError("Offline activity not found")

            return {
                'id': int(row[0]),
                'batch_id': row[1],
                'operator_id': int(row[2]),
                'date': row[4].strftime('%Y-%m-%d') if row[4] else None,
                'start_time': row[5].strftime('%H:%M') if row[5] else None,
                'end_time': row[6].strftime('%H:%M') if row[6] else None,
                'comment': row[7]
            }
    
    def get_user_history(self, user_id):
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT uh.id, uh.field_changed, uh.old_value, uh.new_value, uh.changed_at, u.name AS changed_by_name
                FROM user_history uh
                LEFT JOIN users u ON uh.changed_by = u.id
                WHERE uh.user_id = %s
                ORDER BY uh.changed_at DESC
            """, (user_id,))
            return [
                {
                    "id": row[0],
                    "field": row[1],
                    "old_value": row[2],
                    "new_value": row[3],
                    "changed_at": row[4].strftime('%Y-%m-%d %H:%M:%S'),
                    "changed_by": row[5] or "System"
                } for row in cursor.fetchall()
            ]

    def update_training(self, training_id, training_date=None, start_time=None, end_time=None, reason=None, comment=None, count_in_hours=None):
        updates = []
        params = []
        if training_date:
            updates.append("training_date = %s"); params.append(training_date)
        if start_time:
            updates.append("start_time = %s"); params.append(start_time)
        if end_time:
            updates.append("end_time = %s"); params.append(end_time)
        if reason:
            updates.append("reason = %s"); params.append(reason)
        if comment is not None:
            updates.append("comment = %s"); params.append(comment)
        if count_in_hours is not None:
            updates.append("count_in_hours = %s"); params.append(count_in_hours)

        if not updates:
            return False

        query = f"UPDATE trainings SET {', '.join(updates)} WHERE id = %s RETURNING id"
        params.append(training_id)

        with self._get_cursor() as cursor:
            cursor.execute(query, params)
            return cursor.fetchone() is not None

    def delete_training(self, training_id):
        with self._get_cursor() as cursor:
            cursor.execute("DELETE FROM trainings WHERE id = %s RETURNING id", (training_id,))
            return cursor.fetchone() is not None

    def generate_users_report(self, current_date=None):
        """
        Generates an Excel report of operators with extended profile fields:
        base data, contacts, study, corporate data, proxy/SIP, and two close contacts.
        
        :param current_date: Optional, current date for filename (defaults to today).
        :return: (filename, content) or (None, None) on error.
        """
        try:
            if current_date is None:
                current_date = date.today()
            else:
                current_date = datetime.strptime(current_date, "%Y-%m-%d").date() if isinstance(current_date, str) else current_date
            
            filename = f"users_report_{current_date.strftime('%Y-%m-%d')}.xlsx"
            
            # Fetch operators with all export fields
            with self._get_cursor() as cursor:
                cursor.execute("""
                    SELECT
                        u.name,
                        u.login,
                        u.role,
                        COALESCE(d.name, 'N/A') as direction,
                        COALESCE(s.name, 'N/A') as supervisor,
                        u.status,
                        u.rate,
                        u.hire_date,
                        u.phone,
                        u.email,
                        u.instagram,
                        u.telegram_nick,
                        u.study_place,
                        u.study_course,
                        u.company_name,
                        u.employment_type,
                        COALESCE(u.has_proxy, FALSE) as has_proxy,
                        u.proxy_card_number,
                        COALESCE(u.has_driver_license, FALSE) as has_driver_license,
                        u.sip_number,
                        u.close_contact_1_relation,
                        u.close_contact_1_full_name,
                        u.close_contact_1_phone,
                        u.close_contact_2_relation,
                        u.close_contact_2_full_name,
                        u.close_contact_2_phone,
                        u.card_number,
                        COALESCE(u.internship_in_company, FALSE) as internship_in_company,
                        COALESCE(u.front_office_training, FALSE) as front_office_training,
                        u.front_office_training_date,
                        u.taxipro_id,
                        u.supervisor_id
                    FROM users u
                    LEFT JOIN directions d ON u.direction_id = d.id
                    LEFT JOIN users s ON u.supervisor_id = s.id
                    WHERE u.role = 'operator'
                    ORDER BY s.name, u.name
                """)
                all_operators = cursor.fetchall()
            
            if not all_operators:
                logging.warning("No operators found for users report")
                return None, None
            
            def _format_date(value):
                return value.strftime('%Y-%m-%d') if value else 'N/A'

            def _format_employment(value):
                normalized = str(value or '').strip().lower()
                if normalized == 'gph':
                    return 'ГПХ'
                if normalized == 'of':
                    return 'ОФ'
                return ''

            def _format_proxy(value):
                return 'Да' if bool(value) else 'Нет'

            headers = [
                "ФИО",
                "Логин",
                "Роль",
                "Направление",
                "Супервайзер",
                "Статус",
                "Ставка",
                "Дата принятия",
                "Номер телефона",
                "Почта",
                "Инстаграм",
                "Ник Telegram",
                "Место учебы",
                "Курс",
                "Номер карты",
                "Наименование ТОО/ИП",
                "Оформлен ГПХ/ОФ",
                "Практика в компании",
                "Обучение во фронт офисе",
                "Дата обучения во фронт офисе",
                "ID таксипро",
                "Наличие прокси",
                "Номер прокси карты",
                "Наличие водительских прав",
                "SIP номер",
                "Близкий 1: Кем приходится",
                "Близкий 1: ФИО",
                "Близкий 1: Номер",
                "Близкий 2: Кем приходится",
                "Близкий 2: ФИО",
                "Близкий 2: Номер"
            ]

            # Group by supervisor and normalize rows for export
            operators_by_supervisor = defaultdict(list)
            supervisors = {}  # supervisor_id -> name
            summary_rows = []
            for row in all_operators:
                (
                    name, login, role, direction, supervisor, status, rate, hire_date,
                    phone, email, instagram, telegram_nick,
                    study_place, study_course, card_number,
                    company_name, employment_type, internship_in_company, front_office_training, front_office_training_date, taxipro_id, has_proxy, proxy_card_number, has_driver_license, sip_number,
                    close_contact_1_relation, close_contact_1_full_name, close_contact_1_phone,
                    close_contact_2_relation, close_contact_2_full_name, close_contact_2_phone,
                    sup_id
                ) = row

                row_values = [
                    name or "",
                    login or "",
                    role or "",
                    direction or "N/A",
                    supervisor or "N/A",
                    status or "",
                    float(rate) if rate else 1.0,
                    _format_date(hire_date),
                    phone or "",
                    email or "",
                    instagram or "",
                    telegram_nick or "",
                    study_place or "",
                    study_course or "",
                    card_number or "",
                    company_name or "",
                    _format_employment(employment_type),
                    _format_proxy(internship_in_company),
                    _format_proxy(front_office_training),
                    _format_date(front_office_training_date),
                    taxipro_id or "",
                    _format_proxy(has_proxy),
                    (proxy_card_number or "") if bool(has_proxy) else "",
                    _format_proxy(has_driver_license),
                    sip_number or "",
                    close_contact_1_relation or "",
                    close_contact_1_full_name or "",
                    close_contact_1_phone or "",
                    close_contact_2_relation or "",
                    close_contact_2_full_name or "",
                    close_contact_2_phone or ""
                ]

                summary_rows.append(row_values)
                operators_by_supervisor[sup_id].append(row_values)
                if sup_id and supervisor != 'N/A':
                    supervisors[sup_id] = supervisor
            
            # Create workbook
            wb = Workbook()
            ws_summary = wb.active
            ws_summary.title = "Summary"

            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=True,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            def _write_rows_to_sheet(ws, rows, table_name):
                # headers
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(1, col)
                    cell.value = header
                    cell.font = Font(bold=True)

                # data
                for row_idx, row_values in enumerate(rows, start=2):
                    for col_idx, value in enumerate(row_values, start=1):
                        ws.cell(row_idx, col_idx).value = value

                # table
                if rows:
                    last_col = get_column_letter(len(headers))
                    tab = Table(displayName=table_name, ref=f"A1:{last_col}{len(rows) + 1}")
                    tab.tableStyleInfo = style
                    ws.add_table(tab)

                # column widths
                for col_idx in range(1, len(headers) + 1):
                    max_len = len(str(headers[col_idx - 1]))
                    for row_values in rows:
                        value = row_values[col_idx - 1]
                        value_len = len(str(value)) if value is not None else 0
                        if value_len > max_len:
                            max_len = value_len
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))

            _write_rows_to_sheet(ws_summary, summary_rows, "SummaryTable")
            
            # Create per-supervisor sheets
            for sup_id, ops in operators_by_supervisor.items():
                sup_name = supervisors.get(sup_id, "No Supervisor")
                base_sheet_title = (sup_name or "No Supervisor")[:31]
                sheet_title = base_sheet_title
                if sheet_title in wb.sheetnames:
                    suffix_counter = 2
                    while sheet_title in wb.sheetnames:
                        suffix = f"_{suffix_counter}"
                        sheet_title = f"{base_sheet_title[:31 - len(suffix)]}{suffix}"
                        suffix_counter += 1
                ws = wb.create_sheet(title=sheet_title)
                table_name = f"Table_sup_{sup_id if sup_id is not None else 'none'}"
                table_name = re.sub(r'[^A-Za-z0-9_]', '_', table_name)
                _write_rows_to_sheet(ws, ops, table_name)
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            content = output.getvalue()
            
            return filename, content
        
        except Exception as e:
            logging.error(f"Error generating users report: {e}")
            return None, None

    def generate_excel_report_from_view(self,
        operators: List[Dict[str, Any]],
        trainings_map: Dict[int, Dict[int, List[Dict[str, Any]]]],
        technical_issues_map: Dict[int, Dict[int, List[Dict[str, Any]]]],
        month: str,  # 'YYYY-MM'
        offline_activities_map: Dict[int, Dict[int, List[Dict[str, Any]]]] = None,
        filename: str = None,
        include_supervisor: bool = False
    ) -> Tuple[str, bytes]:
        """
        Генерирует xlsx с листами: Отработанные часы, Перерыв, Звонки, Эффективность,
        Штрафы, Тренинги, Тех. сбои, Офлайн активность.

        Правила форматирования (по заданию пользователя):
        - Все числовые данные пишем без дополнительного округления.
        - В Excel визуально показываем максимум 2 знака после запятой (через number format), без изменения фактического значения.
        - Внутри таблицы по дням значение 0 отображается как целое 0 (не 0.0).
        - Ячейки по дням, где значение > 0, закрашивать светло-серым (теперь чуть темнее).
        - Calls (количество звонков) остаются целыми числами.
        - Добавить границы ко всем ячейкам таблиц.
        """

        def _normalize_day_items_map(source_map: Any) -> Dict[int, Dict[int, List[Dict[str, Any]]]]:
            normalized: Dict[int, Dict[int, List[Dict[str, Any]]]] = {}
            if not isinstance(source_map, dict):
                return normalized
            for op_key, by_day in source_map.items():
                try:
                    op_id_int = int(op_key)
                except Exception:
                    continue
                if not isinstance(by_day, dict):
                    continue
                op_bucket = normalized.setdefault(op_id_int, {})
                for day_key, items in by_day.items():
                    try:
                        day_int = int(day_key)
                    except Exception:
                        continue
                    if day_int <= 0:
                        continue
                    if isinstance(items, list):
                        day_items = items
                    elif isinstance(items, dict):
                        day_items = [items]
                    else:
                        day_items = []
                    op_bucket.setdefault(day_int, []).extend(day_items)
            return normalized

        trainings_map = _normalize_day_items_map(trainings_map or {})
        technical_issues_map = _normalize_day_items_map(technical_issues_map or {})
        offline_activities_map = _normalize_day_items_map(offline_activities_map or {})
        operators = operators["operators"]

        # If requested, ensure each operator has `supervisor_name` populated.
        if include_supervisor:
            # collect supervisor ids that don't already have a name
            sup_ids = []
            for op in operators:
                if not op.get('supervisor_name') and op.get('supervisor_id'):
                    try:
                        sup_ids.append(int(op.get('supervisor_id')))
                    except Exception:
                        continue
            sup_ids = list(set(sup_ids))
            sup_map = {}
            if sup_ids:
                try:
                    with self._get_cursor() as cursor:
                        cursor.execute("SELECT id, name FROM users WHERE id = ANY(%s)", (sup_ids,))
                        sup_map = {r[0]: r[1] for r in cursor.fetchall()}
                except Exception:
                    logging.exception("Error fetching supervisors for report")
            # fill supervisor_name where missing
            for op in operators:
                if not op.get('supervisor_name') and op.get('supervisor_id'):
                    try:
                        op['supervisor_name'] = sup_map.get(int(op.get('supervisor_id')))
                    except Exception:
                        op['supervisor_name'] = None

        def _num_has_value(value: Any) -> bool:
            try:
                return abs(float(value)) > 0
            except Exception:
                return False

        def _has_any_hours_indicators(op: Dict[str, Any]) -> bool:
            if not isinstance(op, dict):
                return False

            aggregates = op.get('aggregates') if isinstance(op.get('aggregates'), dict) else {}
            for key in (
                'regular_hours',
                'total_break_time',
                'total_talk_time',
                'total_calls',
                'total_efficiency_hours',
                'calls_per_hour',
                'fines'
            ):
                if _num_has_value(aggregates.get(key)):
                    return True

            if (
                _num_has_value(op.get('fines')) or
                _num_has_value(op.get('training_hours')) or
                _num_has_value(op.get('technical_issue_hours')) or
                _num_has_value(op.get('offline_activity_hours'))
            ):
                return True

            daily = op.get('daily') if isinstance(op.get('daily'), dict) else {}
            for day_payload in daily.values():
                if not isinstance(day_payload, dict):
                    continue
                for key in ('work_time', 'break_time', 'talk_time', 'calls', 'efficiency', 'fine_amount'):
                    if _num_has_value(day_payload.get(key)):
                        return True
                if str(day_payload.get('fine_reason') or '').strip():
                    return True
                if str(day_payload.get('fine_comment') or '').strip():
                    return True
                if isinstance(day_payload.get('fines'), list) and len(day_payload.get('fines')) > 0:
                    return True

            op_id_int = None
            try:
                op_id_int = int(op.get('operator_id'))
            except Exception:
                op_id_int = None

            if op_id_int is not None:
                for source_map in (trainings_map, technical_issues_map, offline_activities_map):
                    if not isinstance(source_map, dict):
                        continue
                    by_day = source_map.get(op_id_int)
                    if by_day is None:
                        by_day = source_map.get(str(op_id_int))
                    if not isinstance(by_day, dict):
                        continue
                    if any(isinstance(items, list) and len(items) > 0 for items in by_day.values()):
                        return True

            return False

        # Фильтруем уволенных операторов:
        # - оставляем, если увольнение попадает в выбранный месяц
        # - либо если у оператора есть любые показатели за период отчета.
        try:
            year, mon = map(int, month.split('-'))
            month_start = date(year, mon, 1)
            fired_candidates = set()
            for op in operators:
                status_norm = str(op.get('status') or '').strip().lower()
                if status_norm in ('fired', 'dismissal'):
                    try:
                        fired_candidates.add(int(op.get('operator_id')))
                    except Exception:
                        continue

            if fired_candidates:
                with self._get_cursor() as cursor:
                    cursor.execute(
                        """
                        SELECT user_id, changed_at
                        FROM user_history
                        WHERE user_id = ANY(%s)
                          AND field_changed = 'status'
                          AND lower(new_value) IN ('fired', 'dismissal')
                        """,
                        (list(fired_candidates),)
                    )
                    rows = cursor.fetchall()
                allowed_fired_ids = set()
                for r in rows:
                    try:
                        uid = r[0]
                        changed_at = r[1]
                        if changed_at is None:
                            continue
                        ch_date = changed_at.date() if hasattr(changed_at, 'date') else changed_at
                        if ch_date >= month_start:
                            allowed_fired_ids.add(uid)
                    except Exception:
                        continue
            else:
                allowed_fired_ids = set()

            filtered_operators = []
            for op in operators:
                status_norm = str(op.get('status') or '').strip().lower()
                if status_norm not in ('fired', 'dismissal'):
                    filtered_operators.append(op)
                    continue

                op_id_int = None
                try:
                    op_id_int = int(op.get('operator_id'))
                except Exception:
                    op_id_int = None

                keep_by_dismissal_month = bool(op_id_int in allowed_fired_ids) if op_id_int is not None else False
                keep_by_metrics = _has_any_hours_indicators(op)
                if keep_by_dismissal_month or keep_by_metrics:
                    filtered_operators.append(op)

            operators = filtered_operators
        except Exception:
            logging.exception("Error filtering fired operators for report; proceeding without filter")

        FILL_POS = PatternFill(fill_type='solid', start_color='b3b3b3')  # чуть темнее серый
        THIN = Side(style='thin')
        BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        def _make_header(ws, headers: List[str]):
            thin = Side(style='thin')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            bold = Font(bold=True)
            for i, h in enumerate(headers, start=1):
                cell = ws.cell(1, i)
                cell.value = h
                cell.font = bold
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        def parse_time_to_minutes(t: str):
            if not t:
                return None
            try:
                parts = t.strip().split(':')
                hh = int(parts[0])
                mm = int(parts[1]) if len(parts) > 1 else 0
                return hh * 60 + mm
            except Exception:
                return None

        def compute_training_duration_hours(t: Dict[str, Any]) -> float:
            try:
                if t.get('duration_minutes') is not None:
                    return float(t['duration_minutes']) / 60.0
                if t.get('duration_hours') is not None:
                    return float(t['duration_hours'])
                s = parse_time_to_minutes(t.get('start_time'))
                e = parse_time_to_minutes(t.get('end_time'))
                if s is None or e is None:
                    return 0.0
                diff = e - s
                if diff < 0:
                    diff += 24 * 60
                return diff / 60.0
            except Exception:
                return 0.0

        def compute_technical_issue_duration_hours(item: Dict[str, Any]) -> float:
            try:
                if item.get('duration_minutes') is not None:
                    return float(item['duration_minutes']) / 60.0
                if item.get('duration_hours') is not None:
                    return float(item['duration_hours'])
                s = parse_time_to_minutes(item.get('start_time'))
                e = parse_time_to_minutes(item.get('end_time'))
                if s is None or e is None:
                    return 0.0
                diff = e - s
                if diff < 0:
                    diff += 24 * 60
                return diff / 60.0
            except Exception:
                return 0.0

        def compute_offline_activity_duration_hours(item: Dict[str, Any]) -> float:
            try:
                if item.get('duration_minutes') is not None:
                    return float(item['duration_minutes']) / 60.0
                if item.get('duration_hours') is not None:
                    return float(item['duration_hours'])
                s = parse_time_to_minutes(item.get('start_time'))
                e = parse_time_to_minutes(item.get('end_time'))
                if s is None or e is None:
                    return 0.0
                diff = e - s
                if diff < 0:
                    diff += 24 * 60
                return diff / 60.0
            except Exception:
                return 0.0

        def fmt_day_value(metric_key: str, value: Any):
            """Формат для значений в столбцах по дням.
            - calls -> int
            - for hours/perc/eff -> без дополнительного округления; 0 как integer 0
            """
            if metric_key == 'calls':
                try:
                    return int(value or 0)
                except Exception:
                    return 0
            try:
                num = float(value or 0)
            except Exception:
                return 0
            # represent exactly zero as int 0
            if abs(num) < 1e-9:
                return 0
            return num

        def fmt_total_value(metric_key: str, value: Any):
            """Формат для итоговых/доп. колонок (итого, проценты и т.п.).
            - для calls оставляем int
            - для остальных без дополнительного округления
            """
            if value is None:
                return None
            if metric_key == 'calls':
                try:
                    return int(value or 0)
                except Exception:
                    return 0
            try:
                num = float(value)
            except Exception:
                return None
            return num

        logging.info(month)
        year, mon = map(int, month.split('-'))
        days_in_month = calendar.monthrange(year, mon)[1]
        days = list(range(1, days_in_month + 1))

        wb = Workbook()
        default = wb.active
        wb.remove(default)

        FLOAT_DISPLAY_FORMAT = '0.##'
        INTEGER_DISPLAY_FORMAT = '0'

        def _normalize_excel_number(value: Any):
            """
            Avoid showing a dangling decimal separator in Excel for values like 10.0.
            We keep non-integer floats as-is (shown with max 2 fraction digits),
            but convert integer-like floats to int for cleaner display.
            """
            if isinstance(value, float):
                rounded = round(value)
                if abs(value - rounded) < 1e-9:
                    return int(rounded)
            return value

        def set_cell(ws, r, c, value, align_center=True, fill=None):
            cell = ws.cell(r, c)
            normalized_value = _normalize_excel_number(value)
            cell.value = normalized_value
            # Визуально показываем максимум 2 знака после запятой без фактического округления значения.
            if isinstance(normalized_value, float):
                cell.number_format = FLOAT_DISPLAY_FORMAT
            elif isinstance(normalized_value, int) and not isinstance(normalized_value, bool):
                cell.number_format = INTEGER_DISPLAY_FORMAT
            if align_center:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_ALL
            if fill:
                cell.fill = fill
            return cell

        def build_generic_sheet(key: str, label: str, metric_key: str, is_hour=True, format_fn=None, extra_cols=None):
            ws = wb.create_sheet(title=label[:31])
            # Build headers with optional supervisor column
            headers = ["Оператор"]
            if include_supervisor:
                headers.append("Супервайзер")
            headers += ["Ставка", "Норма часов (ч)"] + [f"{d:02d}.{mon:02d}" for d in days] + ["Итого"]
            if extra_cols:
                headers += [c[0] for c in extra_cols]
            _make_header(ws, headers)

            # column indexes depending on presence of supervisor column
            rate_col = 2 + (1 if include_supervisor else 0)
            norm_col = rate_col + 1
            day_start_col = norm_col + 1

            row = 2
            for op in operators:
                daily = op.get('daily', {})
                name = op.get('name') or f"op_{op.get('operator_id')}"
                op_id_raw = op.get('operator_id')
                try:
                    op_id_int = int(op_id_raw)
                except Exception:
                    op_id_int = None
                trainings_by_day = trainings_map.get(op_id_int, {}) if op_id_int is not None else {}
                technical_by_day = technical_issues_map.get(op_id_int, {}) if op_id_int is not None else {}
                offline_by_day = offline_activities_map.get(op_id_int, {}) if op_id_int is not None else {}
                set_cell(ws, row, 1, name, align_center=False)
                if include_supervisor:
                    sup_name = op.get('supervisor_name') or ""
                    set_cell(ws, row, 2, sup_name, align_center=False)
                # Ставка и Норма оставляем без изменения/округления
                set_cell(ws, row, rate_col, float(op.get('rate') or 0), align_center=False)
                set_cell(ws, row, norm_col, float(op.get('norm_hours') or 0), align_center=False)
                total = 0.0
                totals = { 'work_time': 0.0, 'calls': 0, 'efficiency': 0.0 }
                for c_idx, day in enumerate(days, start=day_start_col):
                    dkey = str(day)
                    if metric_key == 'trainings':
                        set_cell(ws, row, c_idx, "")
                    else:
                        d = daily.get(dkey)
                        raw_v = None
                        if d:
                            raw_v = d.get(metric_key, 0)
                        cell_val = fmt_day_value(metric_key, raw_v)
                        # apply fill if >0
                        fill = FILL_POS if (isinstance(cell_val, (int, float)) and cell_val > 0) else None
                        set_cell(ws, row, c_idx, cell_val, fill=fill)
                        if metric_key == 'calls':
                            totals['calls'] += int(cell_val or 0)
                            total += int(cell_val or 0)
                        else:
                            num_for_tot = float(raw_v or 0)
                            total += num_for_tot
                            if metric_key == 'work_time':
                                totals['work_time'] += num_for_tot
                            if metric_key == 'efficiency':
                                totals['efficiency'] += num_for_tot
                # total cell
                total_col = day_start_col + len(days)
                set_cell(ws, row, total_col, fmt_total_value(metric_key, total))

                # extra cols
                if extra_cols:
                    for i, (_, fn) in enumerate(extra_cols, start=1):
                        val = fn(op, totals)
                        # if fn returns percent-string already, keep
                        if isinstance(val, str):
                            set_cell(ws, row, total_col + i, val)
                        elif metric_key == 'calls':
                            if val is None:
                                set_cell(ws, row, total_col + i, None)
                            else:
                                try:
                                    set_cell(ws, row, total_col + i, float(val))
                                except Exception:
                                    set_cell(ws, row, total_col + i, None)
                        else:
                            set_cell(ws, row, total_col + i, fmt_total_value(metric_key, val))

                row += 1

            ws.column_dimensions['A'].width = 24
            for i in range(2, len(headers) + 1):
                col = ws.cell(1, i).column_letter
                ws.column_dimensions[col].width = 12

        def build_work_time_sheet():
            ws = wb.create_sheet(title='Отработанные часы'[:31])
            headers = ["Оператор"]
            if include_supervisor:
                headers.append("Супервайзер")
            headers += ["Ставка", "Норма часов (ч)"] + [f"{d:02d}.{mon:02d}" for d in days] + ["Итого часов", "База часов", "Тех. сбои (ч)", "Тренинги (ч)", "Офлайн активность (ч)", "Вып нормы (%)", "Выработка"]
            _make_header(ws, headers)
            row = 2
            for op in operators:
                daily = op.get('daily', {})
                name = op.get('name') or f"op_{op.get('operator_id')}"
                op_id_raw = op.get('operator_id')
                try:
                    op_id_int = int(op_id_raw)
                except Exception:
                    op_id_int = None
                trainings_by_day = trainings_map.get(op_id_int, {}) if op_id_int is not None else {}
                technical_by_day = technical_issues_map.get(op_id_int, {}) if op_id_int is not None else {}
                offline_by_day = offline_activities_map.get(op_id_int, {}) if op_id_int is not None else {}
                set_cell(ws, row, 1, name, align_center=False)
                if include_supervisor:
                    sup_name = op.get('supervisor_name') or ""
                    set_cell(ws, row, 2, sup_name, align_center=False)
                rate_col = 2 + (1 if include_supervisor else 0)
                norm_col = rate_col + 1
                set_cell(ws, row, rate_col, float(op.get('rate') or 0), align_center=False)
                norm = float(op.get('norm_hours') or 0)
                set_cell(ws, row, norm_col, norm, align_center=False)

                total_work = 0.0
                total_counted_trainings = 0.0
                total_technical_issues = 0.0
                total_offline_activities = 0.0

                day_start = norm_col + 1
                for c_idx, day in enumerate(days, start=day_start):
                    dkey = str(day)
                    work_val = 0.0
                    d = daily.get(dkey)
                    if d:
                        work_val = float(d.get('work_time') or 0.0)
                    # Рассчитываем зачётные часы тренинга для дня и добавляем их к дневному показателю
                    trainings_for_day = trainings_by_day.get(day, []) if trainings_by_day else []
                    counted_for_day = 0.0
                    for t in trainings_for_day:
                        dur = compute_training_duration_hours(t)
                        if t.get('count_in_hours'):
                            counted_for_day += dur

                    technical_for_day = 0.0
                    technical_items_for_day = technical_by_day.get(day, []) if technical_by_day else []
                    for item in technical_items_for_day:
                        technical_for_day += compute_technical_issue_duration_hours(item)

                    offline_for_day = 0.0
                    offline_items_for_day = offline_by_day.get(day, []) if offline_by_day else []
                    for item in offline_items_for_day:
                        offline_for_day += compute_offline_activity_duration_hours(item)

                    # Сохраняем в суммарные показатели отдельно, итоговая ячейка по дню — work + trainings + technical + offline
                    total_work += work_val
                    total_counted_trainings += counted_for_day
                    total_technical_issues += technical_for_day
                    total_offline_activities += offline_for_day
                    combined = work_val + counted_for_day + technical_for_day + offline_for_day
                    cell_val = fmt_day_value('work_time', combined)
                    fill = FILL_POS if (isinstance(cell_val, (int, float)) and cell_val > 0) else None
                    set_cell(ws, row, c_idx, cell_val, fill=fill)

                itogo_chasov = total_work + total_counted_trainings + total_technical_issues + total_offline_activities
                base_total_col = day_start + len(days)
                set_cell(ws, row, base_total_col, fmt_total_value('work_time', itogo_chasov))
                set_cell(ws, row, base_total_col + 1, fmt_total_value('work_time', total_work))
                set_cell(ws, row, base_total_col + 2, fmt_total_value('work_time', total_technical_issues))
                set_cell(ws, row, base_total_col + 3, fmt_total_value('work_time', total_counted_trainings))
                set_cell(ws, row, base_total_col + 4, fmt_total_value('work_time', total_offline_activities))
                if norm and norm != 0:
                    percent_display = (itogo_chasov / norm) * 100
                else:
                    percent_display = None
                set_cell(ws, row, base_total_col + 5, percent_display)
                set_cell(ws, row, base_total_col + 6, fmt_total_value('work_time', (norm - itogo_chasov) if norm is not None else None))

                row += 1

            ws.column_dimensions['A'].width = 24
            for i in range(2, len(headers) + 1):
                col = ws.cell(1, i).column_letter
                ws.column_dimensions[col].width = 14

        def build_calls_sheet():
            def kvz_fn(op, totals):
                work_hours = 0.0
                daily = op.get('daily', {})
                for dnum in daily.values():
                    work_hours += float(dnum.get('work_time') or 0.0)
                calls = totals.get('calls', 0)
                effective_hours = max(0.0, float(work_hours))
                if effective_hours > 0:
                    return calls / effective_hours
                return None

            build_generic_sheet('calls', 'Звонки', 'calls', is_hour=False, extra_cols=[('КВЗ', kvz_fn)])

        def build_efficiency_sheet():
            def otn_fn(op, totals):
                sum_work = 0.0
                sum_eff = totals.get('efficiency', 0.0)
                daily = op.get('daily', {})
                for dnum in daily.values():
                    sum_work += float(dnum.get('work_time') or 0.0)
                if sum_work and sum_work != 0:
                    return (sum_eff / sum_work) * 100
                return None

            build_generic_sheet('efficiency', 'Эффективность', 'efficiency', is_hour=True, extra_cols=[('Отн.', otn_fn)])

        def build_fines_sheet():
            ws_f = wb.create_sheet(title='Штрафы'[:31])
            headers = ["ФИО"]
            if include_supervisor:
                headers.append("Супервайзер")
            headers += [
                "Ставка",
                "Кол-во Опозданий",
                "Минуты",
                "Сумма Опозданий",
                "Корп такси",
                "Кол-во Не выход",
                "Сумма Не выход",
                "Прокси Карта",
                "Другое",
                "Итого"   # <--- НОВАЯ КОЛОНКА
            ]
            _make_header(ws_f, headers)

            row = 2
            def fmt_amt(x):
                try:
                    return float(x or 0.0)
                except Exception:
                    return 0.0

            for op in operators:
                name = op.get('name') or f"op_{op.get('operator_id')}"
                set_cell(ws_f, row, 1, name, align_center=False)

                col_idx = 2
                if include_supervisor:
                    sup_name = op.get('supervisor_name') or ""
                    set_cell(ws_f, row, col_idx, sup_name, align_center=False)
                    col_idx += 1

                rate = op.get('rate') or 0
                set_cell(ws_f, row, col_idx, float(rate), align_center=False)
                col_idx += 1

                fines_map = op.get('daily', {})

                count_late = 0
                minutes_late = 0
                sum_late = 0.0
                sum_korp = 0.0
                count_no_show = 0
                sum_no_show = 0.0
                sum_proxy = 0.0
                sum_other = 0.0

                for day_entry in fines_map.values():
                    fines_list = day_entry.get('fines', []) if isinstance(day_entry, dict) else []
                    for f in fines_list:
                        reason = (f.get('reason') or '').strip()
                        amt = float(f.get('amount') or 0.0)
                        rl = reason.lower()

                        if rl == 'опоздание':
                            count_late += 1
                            minutes = float(f.get('minutes')) if f.get('minutes') is not None else (amt / 50.0 if amt else 0.0)
                            minutes_late += minutes
                            sum_late += amt
                        elif 'корп' in rl and 'такси' in rl:
                            sum_korp += amt
                        elif rl == 'не выход' or 'не выход' in rl:
                            count_no_show += 1
                            sum_no_show += amt
                        elif 'прокси' in rl:
                            sum_proxy += amt
                        else:
                            sum_other += amt

                # Итоговая сумма штрафов
                total_fines = sum_late + sum_korp + sum_no_show + sum_proxy + sum_other

                # Записываем данные
                set_cell(ws_f, row, col_idx, int(count_late)); col_idx += 1
                set_cell(ws_f, row, col_idx, minutes_late); col_idx += 1
                set_cell(ws_f, row, col_idx, fmt_amt(sum_late)); col_idx += 1
                set_cell(ws_f, row, col_idx, fmt_amt(sum_korp)); col_idx += 1
                set_cell(ws_f, row, col_idx, int(count_no_show)); col_idx += 1
                set_cell(ws_f, row, col_idx, fmt_amt(sum_no_show)); col_idx += 1
                set_cell(ws_f, row, col_idx, fmt_amt(sum_proxy)); col_idx += 1
                set_cell(ws_f, row, col_idx, fmt_amt(sum_other)); col_idx += 1

                # Новая колонка — ИТОГО
                set_cell(ws_f, row, col_idx, fmt_amt(total_fines)); col_idx += 1

                row += 1

            ws_f.column_dimensions['A'].width = 28
            for i in range(2, len(headers) + 1):
                col = ws_f.cell(1, i).column_letter
                ws_f.column_dimensions[col].width = 14

        build_work_time_sheet()
        build_generic_sheet('break_time', 'Перерыв', 'break_time', is_hour=True)
        build_calls_sheet()
        build_efficiency_sheet()
        build_fines_sheet()

        ws_t = wb.create_sheet(title='Тренинги'[:31])

        headers = ["Оператор"]
        if include_supervisor:
            headers.append("Супервайзер")
        headers += [f"{d:02d}.{mon:02d}" for d in days] + ["Всего (ч)"]
        _make_header(ws_t, headers)

        def fmt_num(n):
            """Возвращаем число без дополнительного округления."""
            try:
                return float(n or 0.0)
            except Exception:
                return 0.0

        # Заполняем строки — показываем по дням только зачётные часы и одну итоговую колонку "Всего (ч)".
        row_counted = 2

        for op in operators:
            name = op.get('name') or f"op_{op.get('operator_id')}"
            op_id = op.get('operator_id')
            try:
                op_id_int = int(op_id)
            except Exception:
                op_id_int = None

            # Берём все тренинги оператора (словарь day -> list)
            op_trainings = trainings_map.get(op_id_int) or {}

            # Инициализация итогов
            total_all = 0.0

            # Сначала пройдем все дни, чтобы посчитать общие итоги
            for day in days:
                arr = op_trainings.get(day, []) if isinstance(op_trainings, dict) else []
                for t in arr:
                    dur = compute_training_duration_hours(t)
                    total_all += dur

            # --- Заполнение вкладки "Тренинги" ---
            set_cell(ws_t, row_counted, 1, name, align_center=False)
            if include_supervisor:
                sup_name = op.get('supervisor_name') or ""
                set_cell(ws_t, row_counted, 2, sup_name, align_center=False)
            day_start = 2 + (1 if include_supervisor else 0)
            for c_idx, day in enumerate(days, start=day_start):
                arr = op_trainings.get(day, []) if isinstance(op_trainings, dict) else []
                counted = 0.0
                for t in arr:
                    if t.get('count_in_hours'):
                        counted += compute_training_duration_hours(t)
                if counted == 0:
                    set_cell(ws_t, row_counted, c_idx, "")
                else:
                    set_cell(ws_t, row_counted, c_idx, fmt_num(counted), fill=FILL_POS)

            # Итоги по строке — только Всего (ч)
            total_col = len(headers)
            set_cell(ws_t, row_counted, total_col, fmt_num(total_all))
            row_counted += 1

        # Настройка ширины колонок для вкладки Тренинги
        ws_t.column_dimensions['A'].width = 24
        for i in range(2, 3 + len(days)):
            col = ws_t.cell(1, i).column_letter
            ws_t.column_dimensions[col].width = 14

        ws_tech = wb.create_sheet(title='Тех. сбои'[:31])
        tech_headers = ["Оператор"]
        if include_supervisor:
            tech_headers.append("Супервайзер")
        tech_headers += [f"{d:02d}.{mon:02d}" for d in days] + ["Всего (ч)"]
        _make_header(ws_tech, tech_headers)

        row_tech = 2
        for op in operators:
            name = op.get('name') or f"op_{op.get('operator_id')}"
            op_id = op.get('operator_id')
            try:
                op_id_int = int(op_id)
            except Exception:
                op_id_int = None
            op_issues = technical_issues_map.get(op_id_int) or {}

            total_issue_hours = 0.0
            for day in days:
                arr = op_issues.get(day, []) if isinstance(op_issues, dict) else []
                for item in arr:
                    total_issue_hours += compute_technical_issue_duration_hours(item)

            set_cell(ws_tech, row_tech, 1, name, align_center=False)
            if include_supervisor:
                sup_name = op.get('supervisor_name') or ""
                set_cell(ws_tech, row_tech, 2, sup_name, align_center=False)

            day_start = 2 + (1 if include_supervisor else 0)
            for c_idx, day in enumerate(days, start=day_start):
                arr = op_issues.get(day, []) if isinstance(op_issues, dict) else []
                day_sum = 0.0
                for item in arr:
                    day_sum += compute_technical_issue_duration_hours(item)
                if day_sum == 0:
                    set_cell(ws_tech, row_tech, c_idx, "")
                else:
                    set_cell(ws_tech, row_tech, c_idx, fmt_num(day_sum), fill=FILL_POS)

            total_col = len(tech_headers)
            set_cell(ws_tech, row_tech, total_col, fmt_num(total_issue_hours))
            row_tech += 1

        ws_tech.column_dimensions['A'].width = 24
        for i in range(2, 3 + len(days)):
            col = ws_tech.cell(1, i).column_letter
            ws_tech.column_dimensions[col].width = 14

        ws_offline = wb.create_sheet(title='Офлайн активность'[:31])
        offline_headers = ["Оператор"]
        if include_supervisor:
            offline_headers.append("Супервайзер")
        offline_headers += [f"{d:02d}.{mon:02d}" for d in days] + ["Всего (ч)"]
        _make_header(ws_offline, offline_headers)

        row_offline = 2
        for op in operators:
            name = op.get('name') or f"op_{op.get('operator_id')}"
            op_id = op.get('operator_id')
            try:
                op_id_int = int(op_id)
            except Exception:
                op_id_int = None
            op_activities = offline_activities_map.get(op_id_int) or {}

            total_activity_hours = 0.0
            for day in days:
                arr = op_activities.get(day, []) if isinstance(op_activities, dict) else []
                for item in arr:
                    total_activity_hours += compute_offline_activity_duration_hours(item)

            set_cell(ws_offline, row_offline, 1, name, align_center=False)
            if include_supervisor:
                sup_name = op.get('supervisor_name') or ""
                set_cell(ws_offline, row_offline, 2, sup_name, align_center=False)

            day_start = 2 + (1 if include_supervisor else 0)
            for c_idx, day in enumerate(days, start=day_start):
                arr = op_activities.get(day, []) if isinstance(op_activities, dict) else []
                day_sum = 0.0
                for item in arr:
                    day_sum += compute_offline_activity_duration_hours(item)
                if day_sum == 0:
                    set_cell(ws_offline, row_offline, c_idx, "")
                else:
                    set_cell(ws_offline, row_offline, c_idx, fmt_num(day_sum), fill=FILL_POS)

            total_col = len(offline_headers)
            set_cell(ws_offline, row_offline, total_col, fmt_num(total_activity_hours))
            row_offline += 1

        ws_offline.column_dimensions['A'].width = 24
        for i in range(2, 3 + len(days)):
            col = ws_offline.cell(1, i).column_letter
            ws_offline.column_dimensions[col].width = 14

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        content = out.getvalue()
        if filename is None:
            filename = f"report_{month}.xlsx"
        return filename, content

    def generate_excel_report_all_operators_from_view(self,
        operators: List[Dict[str, Any]],
        trainings_map: Dict[int, Dict[int, List[Dict[str, Any]]]],
        technical_issues_map: Dict[int, Dict[int, List[Dict[str, Any]]]],
        month: str,  # 'YYYY-MM'
        offline_activities_map: Dict[int, Dict[int, List[Dict[str, Any]]]] = None,
        filename: str = None
        ) -> Tuple[str, bytes]:
        """
        Быстрая версия: дополняет операторов supervisor_name и вызывает
        generate_excel_report_from_view(include_supervisor=True).
        """
        # operators может быть {"operators": [...]} или список
        ops = operators["operators"] if isinstance(operators, dict) and "operators" in operators else operators
        logging.info("First operator for all-operators report: %s", ops[0])
        # Deduplicate ids и получить map одним запросом
        sup_ids = list({int(op.get('supervisor_id')) for op in ops if op.get('supervisor_id')})
        logging.info("Fetching supervisors for all-operators report: %s", sup_ids)
        sup_map = {}
        if sup_ids:
            try:
                with self._get_cursor() as cursor:
                    cursor.execute("SELECT id, name FROM users WHERE id = ANY(%s)", (sup_ids,))
                    logging.info("Fetching supervisors for all-operators report: %s", sup_ids)
                    sup_map = {r[0]: r[1] for r in cursor.fetchall()}
            except Exception:
                logging.exception("Error fetching supervisors for all-operators report")

        # Добавляем supervisor_name к каждому оператору (O(N))
        for op in ops:
            sid = op.get('supervisor_id')
            op['supervisor_name'] = sup_map.get(int(sid)) if sid else None

        # Вызов генератора — один проход, файл формируется сразу с колонкой "Супервайзер"
        return self.generate_excel_report_from_view(
            {"operators": ops},
            trainings_map,
            technical_issues_map,
            month,
            offline_activities_map=offline_activities_map,
            filename=filename,
            include_supervisor=True
        )

    def _normalize_break_durations_list(self, value):
        if value is None:
            return []
        if isinstance(value, str):
            parts = [p.strip() for p in value.split(',')]
            raw_items = [p for p in parts if p]
        elif isinstance(value, list):
            raw_items = value
        else:
            raise ValueError("break_durations must be a list of minutes")

        result = []
        for item in raw_items:
            try:
                minutes = int(item)
            except Exception:
                raise ValueError("break_durations must contain integer minutes")
            if minutes <= 0:
                raise ValueError("break_durations values must be > 0")
            if minutes > 240:
                raise ValueError("break_durations values must be <= 240")
            result.append(minutes)
        return result

    def _normalize_break_rule_ranges(self, rules):
        if rules is None:
            return []
        if not isinstance(rules, list):
            raise ValueError("rules must be a list")

        normalized = []

        def _to_minutes(item, minutes_key, minutes_key_alt, hours_key, hours_key_alt):
            if item.get(minutes_key) is not None:
                return int(item.get(minutes_key))
            if item.get(minutes_key_alt) is not None:
                return int(item.get(minutes_key_alt))
            if item.get(hours_key) is not None:
                return int(round(float(item.get(hours_key)) * 60))
            if item.get(hours_key_alt) is not None:
                return int(round(float(item.get(hours_key_alt)) * 60))
            return None

        for rule in rules:
            if not isinstance(rule, dict):
                raise ValueError("Each rule must be an object")
            min_minutes = _to_minutes(rule, 'min_minutes', 'minMinutes', 'min_hours', 'minHours')
            max_minutes = _to_minutes(rule, 'max_minutes', 'maxMinutes', 'max_hours', 'maxHours')
            if min_minutes is None or max_minutes is None:
                raise ValueError("Each rule must contain min/max hours or minutes")
            if min_minutes < 0:
                raise ValueError("min_minutes must be >= 0")
            if max_minutes <= min_minutes:
                raise ValueError("max_minutes must be > min_minutes")
            break_durations = self._normalize_break_durations_list(
                rule.get('break_durations')
                if 'break_durations' in rule else (
                    rule.get('breakDurations')
                    if 'breakDurations' in rule else rule.get('durations')
                )
            )
            normalized.append({
                'minMinutes': int(min_minutes),
                'maxMinutes': int(max_minutes),
                'breakDurations': break_durations
            })

        normalized.sort(key=lambda x: (x['minMinutes'], x['maxMinutes']))
        prev = None
        for cur in normalized:
            if prev is not None and cur['minMinutes'] < prev['maxMinutes']:
                raise ValueError("Break rules ranges must not overlap within one direction")
            prev = cur
        return normalized

    def get_work_schedule_break_rules(self):
        """
        Возвращает правила автоперерывов по направлениям.
        Формат:
        [
          {
            "direction": "Название",
            "rules": [
              {"minMinutes": 360, "maxMinutes": 540, "breakDurations": [30]}
            ]
          }
        ]
        """
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT direction_name, min_shift_minutes, max_shift_minutes, break_durations_json
                FROM work_schedule_break_rules
                ORDER BY LOWER(direction_name), min_shift_minutes, max_shift_minutes, id
            """)
            rows = cursor.fetchall() or []

        by_direction = {}
        for direction_name, min_shift_minutes, max_shift_minutes, break_durations_json in rows:
            direction_label = str(direction_name or '').strip()
            if not direction_label:
                continue
            item = by_direction.setdefault(direction_label, {"direction": direction_label, "rules": []})
            item["rules"].append({
                "minMinutes": int(min_shift_minutes),
                "maxMinutes": int(max_shift_minutes),
                "breakDurations": self._normalize_break_durations_list(break_durations_json)
            })

        result = list(by_direction.values())
        result.sort(key=lambda x: str(x.get('direction') or '').lower())
        return result

    def get_work_schedule_break_rules_map(self):
        """
        Возвращает map по normalized direction key -> rules list.
        """
        result = {}
        for item in (self.get_work_schedule_break_rules() or []):
            direction = str(item.get('direction') or '').strip()
            key = self._normalize_direction_key(direction)
            if not key:
                continue
            result[key] = [dict(rule) for rule in (item.get('rules') or []) if isinstance(rule, dict)]
        return result

    def _get_work_schedule_break_rules_for_direction_tx(self, cursor, direction_name):
        direction_label = str(direction_name or '').strip()
        if not direction_label:
            return []
        cursor.execute("""
            SELECT min_shift_minutes, max_shift_minutes, break_durations_json
            FROM work_schedule_break_rules
            WHERE LOWER(direction_name) = LOWER(%s)
            ORDER BY min_shift_minutes, max_shift_minutes, id
        """, (direction_label,))
        rows = cursor.fetchall() or []
        return [
            {
                "minMinutes": int(min_shift_minutes),
                "maxMinutes": int(max_shift_minutes),
                "breakDurations": self._normalize_break_durations_list(break_durations_json)
            }
            for min_shift_minutes, max_shift_minutes, break_durations_json in rows
        ]

    def save_work_schedule_break_rules(self, direction_rules):
        """
        Сохранить правила автоперерывов по направлениям.
        direction_rules:
        [
          {"direction": "Название", "rules": [{"minMinutes": 360, "maxMinutes": 540, "breakDurations": [30]}]}
        ]
        Для направления переданные rules полностью заменяют старые.
        """
        if not isinstance(direction_rules, list):
            raise ValueError("direction_rules must be a list")

        normalized_payload = []
        seen = set()
        for item in direction_rules:
            if not isinstance(item, dict):
                raise ValueError("Each direction_rules item must be an object")
            direction = str(item.get('direction') or item.get('direction_name') or '').strip()
            if not direction:
                raise ValueError("direction is required")
            direction_key = self._normalize_direction_key(direction)
            if not direction_key:
                raise ValueError("direction is required")
            if direction_key in seen:
                raise ValueError("Duplicate direction in direction_rules")
            seen.add(direction_key)
            rules = self._normalize_break_rule_ranges(item.get('rules'))
            normalized_payload.append({
                'direction': direction,
                'rules': rules
            })

        with self._get_cursor() as cursor:
            for item in normalized_payload:
                direction = item['direction']
                cursor.execute("""
                    DELETE FROM work_schedule_break_rules
                    WHERE LOWER(direction_name) = LOWER(%s)
                """, (direction,))

                rows_to_insert = item['rules']
                if not rows_to_insert:
                    continue

                cursor.executemany("""
                    INSERT INTO work_schedule_break_rules (
                        direction_name,
                        min_shift_minutes,
                        max_shift_minutes,
                        break_durations_json,
                        updated_at
                    )
                    VALUES (%s, %s, %s, %s::jsonb, CURRENT_TIMESTAMP)
                """, [
                    (
                        direction,
                        int(rule['minMinutes']),
                        int(rule['maxMinutes']),
                        json.dumps(rule['breakDurations'], ensure_ascii=False)
                    )
                    for rule in rows_to_insert
                ])

        return self.get_work_schedule_break_rules()

    # ==================== Work Shifts Methods ====================
    
    def _normalize_schedule_date(self, value):
        if value is None:
            raise ValueError("shift_date is required")
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            return datetime.strptime(value, '%Y-%m-%d').date()
        raise ValueError("Invalid date format, expected YYYY-MM-DD")

    def _normalize_schedule_time(self, value, field_name):
        if value is None:
            raise ValueError(f"{field_name} is required")
        if isinstance(value, dt_time):
            return value
        if isinstance(value, str):
            return datetime.strptime(value, '%H:%M').time()
        raise ValueError(f"Invalid {field_name} format, expected HH:MM")

    def _normalize_shift_breaks(self, breaks):
        if breaks is None:
            return []
        if not isinstance(breaks, list):
            raise ValueError("breaks must be a list")

        normalized = []
        seen = set()
        for item in breaks:
            if not isinstance(item, dict):
                raise ValueError("Each break must be an object")
            start_raw = item.get('start')
            end_raw = item.get('end')
            if start_raw is None or end_raw is None:
                raise ValueError("Each break must contain start and end")

            start_val = _time_to_minutes(start_raw) if isinstance(start_raw, str) else int(start_raw)
            end_val = _time_to_minutes(end_raw) if isinstance(end_raw, str) else int(end_raw)
            if end_val <= start_val:
                raise ValueError("Break end must be greater than break start")

            key = (start_val, end_val)
            if key in seen:
                continue
            seen.add(key)
            normalized.append({'start': start_val, 'end': end_val})

        normalized.sort(key=lambda b: (b['start'], b['end']))
        return normalized

    def _schedule_interval_minutes(self, start_time_value, end_time_value):
        if isinstance(start_time_value, dt_time):
            start_str = start_time_value.strftime('%H:%M')
        else:
            start_str = str(start_time_value)
        if isinstance(end_time_value, dt_time):
            end_str = end_time_value.strftime('%H:%M')
        else:
            end_str = str(end_time_value)

        start_min = _time_to_minutes(start_str)
        end_min = _time_to_minutes(end_str)
        if end_min <= start_min:
            end_min += 24 * 60
        return start_min, end_min

    def _normalize_direction_key(self, direction_name):
        return ' '.join(str(direction_name or '').strip().lower().split())

    def _is_chat_manager_direction(self, direction_name):
        key = self._normalize_direction_key(direction_name)
        return key in ('чат менеджер', 'chat manager')

    def _is_smz_direction(self, direction_name):
        key = self._normalize_direction_key(direction_name)
        if not key:
            return False
        # Поддерживаем варианты вида "СМЗ 1", "smz team" и смешанную запись "sмz".
        return ('смз' in key) or ('smz' in key) or ('sмz' in key)

    def _is_line_direction(self, direction_name):
        key = self._normalize_direction_key(direction_name)
        if not key:
            return False
        # Поддерживаем варианты вида "Основа 1", "Линия 2", "osnova team", "line 2".
        return ('основа' in key) or ('osnova' in key) or ('линия' in key) or ('line' in key)

    def _are_swap_directions_compatible(self, requester_direction_name, target_direction_name):
        requester_key = self._normalize_direction_key(requester_direction_name)
        target_key = self._normalize_direction_key(target_direction_name)
        if not requester_key or not target_key:
            return False
        if requester_key == target_key:
            return True
        requester_is_smz_or_line = self._is_smz_direction(requester_direction_name) or self._is_line_direction(requester_direction_name)
        target_is_smz_or_line = self._is_smz_direction(target_direction_name) or self._is_line_direction(target_direction_name)
        if requester_is_smz_or_line and target_is_smz_or_line:
            return True
        return False

    def _get_operator_direction_name_tx(self, cursor, operator_id):
        cursor.execute(
            """
            SELECT d.name
            FROM users u
            LEFT JOIN directions d ON d.id = u.direction_id
            WHERE u.id = %s
            LIMIT 1
            """,
            (int(operator_id),)
        )
        row = cursor.fetchone()
        return row[0] if row else None

    def _pick_break_durations_for_shift(self, duration_minutes, direction_name=None, direction_rules=None):
        dur = int(duration_minutes)
        rules = direction_rules if isinstance(direction_rules, list) else []
        selected_custom = None
        for rule in sorted(
            [r for r in rules if isinstance(r, dict)],
            key=lambda x: (int(x.get('minMinutes', 0)), int(x.get('maxMinutes', 0)))
        ):
            try:
                min_minutes = int(rule.get('minMinutes'))
                max_minutes = int(rule.get('maxMinutes'))
            except Exception:
                continue
            if max_minutes <= min_minutes:
                continue
            if dur >= min_minutes and dur <= max_minutes:
                # При совпадении нескольких диапазонов берем с максимальным minMinutes
                # (например, 6-9 и 9-12 -> для 9ч сработает 9-12).
                if selected_custom is None or min_minutes >= int(selected_custom.get('minMinutes', -1)):
                    selected_custom = rule

        if selected_custom is not None:
            return self._normalize_break_durations_list(selected_custom.get('breakDurations'))

        # Fallback: действующий базовый профиль + чат-менеджер по умолчанию.
        if self._is_chat_manager_direction(direction_name):
            if dur >= 6 * 60 and dur < 9 * 60:
                return [30]
            if dur >= 9 * 60 and dur <= 12 * 60:
                return [30, 30]

        if dur >= 5 * 60 and dur < 6 * 60:
            return [15]
        if dur >= 6 * 60 and dur < 8 * 60:
            return [15, 15]
        if dur >= 8 * 60 and dur < 11 * 60:
            return [15, 30, 15]
        if dur >= 11 * 60:
            return [15, 30, 15, 15]
        return []

    def _compute_auto_shift_breaks_minutes(self, start_min, end_min, direction_name=None, direction_rules=None):
        """
        Серверная автогенерация перерывов (зеркалит фронтенд-базовую логику по длительности смены).
        Возвращает список интервалов в минутах от начала дня; для ночных смен end_min может быть > 1440.
        """
        start_min = int(start_min)
        end_min = int(end_min)
        dur = end_min - start_min
        if dur <= 0:
            return []

        breaks = []

        def snap5(x):
            return int(round(float(x) / 5.0) * 5)

        def push_centered(center, size):
            center_snapped = snap5(center)
            s = snap5(center_snapped - (size / 2))
            e = s + int(size)
            s = max(start_min, min(end_min, s))
            e = max(start_min, min(end_min, e))
            if e > s:
                breaks.append({'start': s, 'end': e})

        break_durations = self._pick_break_durations_for_shift(
            duration_minutes=dur,
            direction_name=direction_name,
            direction_rules=direction_rules
        )
        if break_durations:
            count = len(break_durations)
            for idx, size in enumerate(break_durations):
                center = start_min + (dur * ((idx + 1) / (count + 1)))
                push_centered(center, int(size))

        normalized = []
        seen = set()
        for b in sorted(breaks, key=lambda x: (x['start'], x['end'])):
            key = (int(b['start']), int(b['end']))
            if key[1] <= key[0] or key in seen:
                continue
            seen.add(key)
            normalized.append({'start': key[0], 'end': key[1]})
        return normalized

    def _insert_shift_breaks(self, cursor, shift_id, breaks):
        cursor.execute("DELETE FROM shift_breaks WHERE shift_id = %s", (shift_id,))
        if breaks:
            cursor.executemany(
                """
                INSERT INTO shift_breaks (shift_id, start_minutes, end_minutes)
                VALUES (%s, %s, %s)
                """,
                [(shift_id, b['start'], b['end']) for b in breaks]
            )

    def _load_shift_breaks_tx(self, cursor, shift_id):
        cursor.execute(
            """
            SELECT start_minutes, end_minutes
            FROM shift_breaks
            WHERE shift_id = %s
            ORDER BY start_minutes, end_minutes
            """,
            (int(shift_id),)
        )
        result = []
        for start_minutes, end_minutes in cursor.fetchall() or []:
            result.append({
                'start': int(start_minutes),
                'end': int(end_minutes)
            })
        return result

    def _break_intervals_overlap(self, a, b):
        return int(a['start']) < int(b['end']) and int(b['start']) < int(a['end'])

    def _merge_break_intervals(self, items):
        if not items:
            return []
        normalized = sorted(
            [
                {'start': int(x.get('start', 0)), 'end': int(x.get('end', 0))}
                for x in items
                if int(x.get('end', 0)) > int(x.get('start', 0))
            ],
            key=lambda x: (x['start'], x['end'])
        )
        if not normalized:
            return []
        result = [dict(normalized[0])]
        for cur in normalized[1:]:
            last = result[-1]
            if cur['start'] <= last['end']:
                last['end'] = max(last['end'], cur['end'])
            else:
                result.append(dict(cur))
        return result

    def _find_non_overlapping_break_start(self, desired_start, length, seg_start, seg_end, occupied_intervals):
        step = 5

        def snap(v):
            return int(round(float(v) / step) * step)

        desired_start = snap(desired_start)
        candidates = [0]
        max_shift = max(int(seg_end) - int(seg_start), 60)
        for delta in range(step, max_shift + step, step):
            candidates.append(delta)
            candidates.append(-delta)

        for delta in candidates:
            s = desired_start + delta
            start = max(int(seg_start), min(int(seg_end) - int(length), s))
            end = start + int(length)
            if start < int(seg_start) or end > int(seg_end):
                continue
            interval = {'start': start, 'end': end}
            if any(self._break_intervals_overlap(interval, occ) for occ in (occupied_intervals or [])):
                continue
            return start
        return None

    def _load_occupied_break_intervals_for_operator_date_tx(self, cursor, operator_id, shift_date):
        """
        Возвращает занятые интервалы перерывов других операторов того же направления
        для shift_date с учетом переходов через полночь (пред/след день со сдвигом).
        """
        operator_id = int(operator_id)
        date_obj = self._normalize_schedule_date(shift_date)

        cursor.execute(
            """
            SELECT direction_id
            FROM users
            WHERE id = %s
            LIMIT 1
            """,
            (operator_id,)
        )
        row = cursor.fetchone()
        direction_id = row[0] if row else None
        if direction_id is None:
            return []

        prev_date = date_obj - timedelta(days=1)
        next_date = date_obj + timedelta(days=1)
        occupied = []

        cursor.execute(
            """
            SELECT ws.shift_date, sb.start_minutes, sb.end_minutes
            FROM work_shifts ws
            JOIN users u ON u.id = ws.operator_id
            JOIN shift_breaks sb ON sb.shift_id = ws.id
            WHERE ws.operator_id <> %s
              AND u.role = 'operator'
              AND u.direction_id = %s
              AND ws.shift_date IN (%s, %s, %s)
            """,
            (operator_id, direction_id, prev_date, date_obj, next_date)
        )

        for break_date, start_minutes, end_minutes in cursor.fetchall() or []:
            try:
                s = int(start_minutes)
                e = int(end_minutes)
            except Exception:
                continue
            if e <= s:
                continue

            if break_date == prev_date:
                s -= 1440
                e -= 1440
            elif break_date == next_date:
                s += 1440
                e += 1440

            ns = max(0, min(2880, int(s)))
            ne = max(0, min(2880, int(e)))
            if ne > ns:
                occupied.append({'start': ns, 'end': ne})

        return self._merge_break_intervals(occupied)

    def _adjust_shift_breaks_against_occupied_tx(self, cursor, operator_id, shift_date, start_time, end_time, breaks):
        if not isinstance(breaks, list) or not breaks:
            return []

        seg_start, seg_end = self._schedule_interval_minutes(start_time, end_time)
        seg_start = max(0, seg_start)
        seg_end = max(seg_start, min(2880, seg_end))
        if seg_end <= seg_start:
            return []

        occupied = self._load_occupied_break_intervals_for_operator_date_tx(cursor, operator_id, shift_date)
        new_breaks = []
        for b in breaks:
            b_start = int(b.get('start', 0))
            b_end = int(b.get('end', 0))
            length = b_end - b_start
            if length <= 0 or (seg_end - seg_start) <= 0:
                continue

            desired_start = max(seg_start, min(seg_end - length, b_start))
            found = self._find_non_overlapping_break_start(
                desired_start=desired_start,
                length=length,
                seg_start=seg_start,
                seg_end=seg_end,
                occupied_intervals=occupied
            )
            if found is not None:
                nb = {'start': int(found), 'end': int(found) + int(length)}
            else:
                clamped_start = max(seg_start, min(seg_end - length, b_start))
                clamped_end = max(seg_start, min(seg_end, b_end))
                if clamped_end <= clamped_start:
                    continue
                nb = {'start': int(clamped_start), 'end': int(clamped_end)}

            new_breaks.append(nb)
            occupied.append(nb)
            occupied = self._merge_break_intervals(occupied)

        return new_breaks

    def _schedule_auto_normalize_flag_status(self, value):
        status = str(value or '').strip().lower()
        if status in SCHEDULE_AUTO_FLAG_ALLOWED:
            return status
        return None

    def _schedule_auto_clamp_range(self, start_date_obj, end_date_obj):
        if start_date_obj is None or end_date_obj is None:
            return None, None
        if end_date_obj < SCHEDULE_AUTO_AGGREGATION_START_DATE:
            return None, None
        return max(start_date_obj, SCHEDULE_AUTO_AGGREGATION_START_DATE), end_date_obj

    def _schedule_auto_total_minutes(self, intervals):
        total = 0.0
        for item in (intervals or []):
            start = float(item.get('start', 0))
            end = float(item.get('end', 0))
            if end > start:
                total += (end - start)
        return float(total)

    def _schedule_auto_seconds_to_display_minutes(self, seconds_value):
        seconds_int = max(0, int(seconds_value or 0))
        if seconds_int <= 0:
            return 0
        return int((seconds_int + 59) // 60)

    def _schedule_auto_subtract_intervals(self, source_intervals, blocked_intervals):
        source = self._merge_break_intervals(source_intervals or [])
        blocked = self._merge_break_intervals(blocked_intervals or [])
        if not source:
            return []
        if not blocked:
            return source

        current = source
        for block in blocked:
            b_start = int(block.get('start', 0))
            b_end = int(block.get('end', 0))
            if b_end <= b_start:
                continue
            next_items = []
            for seg in current:
                s_start = int(seg.get('start', 0))
                s_end = int(seg.get('end', 0))
                if s_end <= s_start:
                    continue
                if s_end <= b_start or s_start >= b_end:
                    next_items.append({'start': s_start, 'end': s_end})
                    continue
                if s_start < b_start:
                    next_items.append({'start': s_start, 'end': b_start})
                if s_end > b_end:
                    next_items.append({'start': b_end, 'end': s_end})
            current = self._merge_break_intervals(next_items)
            if not current:
                break
        return current

    def _schedule_auto_overlap_minutes(self, a_intervals, b_intervals):
        a_list = self._merge_break_intervals(a_intervals or [])
        b_list = self._merge_break_intervals(b_intervals or [])
        if not a_list or not b_list:
            return 0.0

        total = 0.0
        i = 0
        j = 0
        while i < len(a_list) and j < len(b_list):
            a = a_list[i]
            b = b_list[j]
            start = max(int(a['start']), int(b['start']))
            end = min(int(a['end']), int(b['end']))
            if end > start:
                total += (end - start)
            if int(a['end']) <= int(b['end']):
                i += 1
            else:
                j += 1
        return float(total)

    def _schedule_auto_intersect_interval_with_list(self, interval, intervals):
        start = int(interval.get('start', 0))
        end = int(interval.get('end', 0))
        if end <= start:
            return []
        result = []
        for seg in self._merge_break_intervals(intervals or []):
            seg_start = int(seg.get('start', 0))
            seg_end = int(seg.get('end', 0))
            if seg_end <= seg_start:
                continue
            overlap_start = max(start, seg_start)
            overlap_end = min(end, seg_end)
            if overlap_end > overlap_start:
                result.append({'start': overlap_start, 'end': overlap_end})
        return self._merge_break_intervals(result)

    def _schedule_auto_split_status_segment_by_day(self, status_date_value, start_at_value, end_at_value):
        """
        Разбивает статус-сегмент на дневные куски в секундах [0..86400].
        Корректно обрабатывает переход через полночь и интервалы, где end_at
        лежит в следующем календарном дне.
        """
        chunks = []

        if isinstance(start_at_value, datetime) and isinstance(end_at_value, datetime):
            if end_at_value <= start_at_value:
                return []

            cursor = start_at_value
            while cursor < end_at_value:
                day_start = cursor.replace(hour=0, minute=0, second=0, microsecond=0)
                day_end = day_start + timedelta(days=1)
                seg_end = end_at_value if end_at_value <= day_end else day_end
                if seg_end <= cursor:
                    break

                start_delta = (cursor - day_start).total_seconds()
                end_delta = (seg_end - day_start).total_seconds()
                start_sec = int(start_delta)
                end_sec = int(end_delta)
                if end_delta > float(end_sec):
                    end_sec += 1
                if end_sec <= start_sec and seg_end > cursor:
                    end_sec = min(86400, start_sec + 1)

                start_sec = max(0, min(86399, int(start_sec)))
                end_sec = max(1, min(86400, int(end_sec)))
                if end_sec > start_sec:
                    chunks.append({
                        'day': cursor.date(),
                        'start': int(start_sec),
                        'end': int(end_sec)
                    })
                cursor = seg_end

            if chunks:
                return chunks

        # Fallback для старых/нестандартных данных: якоримся на status_date.
        if not isinstance(status_date_value, date):
            return []

        if isinstance(start_at_value, datetime):
            start_sec = int((start_at_value.hour * 3600) + (start_at_value.minute * 60) + start_at_value.second)
        elif isinstance(start_at_value, dt_time):
            start_sec = int((start_at_value.hour * 3600) + (start_at_value.minute * 60) + start_at_value.second)
        else:
            return []

        if isinstance(end_at_value, datetime):
            end_sec = int((end_at_value.hour * 3600) + (end_at_value.minute * 60) + end_at_value.second)
            if end_at_value.microsecond > 0:
                end_sec += 1
        elif isinstance(end_at_value, dt_time):
            end_sec = int((end_at_value.hour * 3600) + (end_at_value.minute * 60) + end_at_value.second)
            if getattr(end_at_value, 'microsecond', 0) > 0:
                end_sec += 1
        else:
            return []

        start_sec = max(0, min(86399, start_sec))
        end_sec = max(1, min(86400, end_sec))
        if end_sec <= start_sec:
            end_sec = min(86400, start_sec + 1)

        if end_sec > start_sec:
            chunks.append({
                'day': status_date_value,
                'start': int(start_sec),
                'end': int(end_sec)
            })
        return chunks

    def _normalize_import_status_key(self, status_key_value):
        return ' '.join(str(status_key_value or '').strip().lower().split())

    def _status_label_from_key(self, status_key_value, fallback_name=None):
        key = self._normalize_import_status_key(status_key_value)
        if not key:
            fallback_text = str(fallback_name or '').strip()
            return fallback_text or 'Статус'
        if self._schedule_auto_is_tech_reason_status_key(key):
            return 'Тех причина'
        direct = SCHEDULE_STATUS_KEY_LABELS.get(key)
        if direct:
            return direct
        return ' '.join(part[:1].upper() + part[1:] for part in key.split(' ') if part)

    def _schedule_auto_compact_status_key(self, status_key_value):
        key = str(status_key_value or '').strip().lower()
        if not key:
            return ''
        return re.sub(r'[\s._-]+', '', key)

    def _schedule_auto_is_tech_reason_status_key(self, status_key_value):
        compact_key = self._schedule_auto_compact_status_key(status_key_value)
        return compact_key == 'техпричина'

    def _schedule_auto_is_late_excused_status_key(self, status_key_value):
        key = self._normalize_import_status_key(status_key_value)
        if not key:
            return False
        if key in (SCHEDULE_AUTO_TRAINING_STATUS_KEY, 'training'):
            return True
        return self._schedule_auto_is_tech_reason_status_key(key)

    def _schedule_auto_shift_start_excused_boundary(self, shift_interval, excused_intervals):
        shift_start = int(shift_interval.get('start', 0))
        shift_end = int(shift_interval.get('end', 0))
        if shift_end <= shift_start:
            return shift_start

        covered_until = shift_start
        for seg in self._merge_break_intervals(excused_intervals or []):
            seg_start = int(seg.get('start', 0))
            seg_end = int(seg.get('end', 0))
            if seg_end <= seg_start or seg_end <= covered_until:
                continue
            if seg_start > covered_until:
                break
            covered_until = min(shift_end, max(covered_until, seg_end))
            if covered_until >= shift_end:
                break
        return covered_until

    def _schedule_auto_shift_end_excused_boundary(self, shift_interval, excused_intervals):
        shift_start = int(shift_interval.get('start', 0))
        shift_end = int(shift_interval.get('end', 0))
        if shift_end <= shift_start:
            return shift_end

        covered_from = shift_end
        merged = self._merge_break_intervals(excused_intervals or [])
        for seg in reversed(merged):
            seg_start = int(seg.get('start', 0))
            seg_end = int(seg.get('end', 0))
            if seg_end <= seg_start:
                continue
            seg_start = max(shift_start, seg_start)
            seg_end = min(shift_end, seg_end)
            if seg_end <= seg_start:
                continue

            # Нет непрерывного покрытия до конца смены — останавливаемся.
            if seg_end < covered_from:
                break
            covered_from = max(shift_start, min(covered_from, seg_start))
            if covered_from <= shift_start:
                break
        return covered_from

    def _schedule_auto_flag_status_for_minutes(self, status_value, minutes_value):
        minutes_int = max(0, int(minutes_value or 0))
        if minutes_int <= 0:
            return None
        normalized = self._schedule_auto_normalize_flag_status(status_value)
        return normalized or SCHEDULE_AUTO_FLAG_PENDING

    def _schedule_auto_flag_meta(self, flag_type):
        key = str(flag_type or '').strip().lower()
        if key == 'late':
            return {
                'reason': 'Опоздание',
                'label': 'Опоздание'
            }
        if key == 'early_leave':
            return {
                'reason': 'Опоздание',
                'label': 'Ранний уход'
            }
        if key == 'training':
            return {
                'reason': 'Тренинг',
                'label': 'Тренинг'
            }
        if key in ('technical_reason', 'tech_reason', 'tech', 'technical'):
            return {
                'reason': 'Тех.причина',
                'label': 'Тех.причина'
            }
        raise ValueError("Unsupported auto flag type")

    def _schedule_auto_build_fine_payload(self, flag_type, minutes_value):
        flag_key = str(flag_type or '').strip().lower()
        if flag_key in ('training', 'technical_reason', 'tech_reason', 'tech', 'technical'):
            raise ValueError("This auto flag must not be stored in daily_fines")
        meta = self._schedule_auto_flag_meta(flag_type)
        minutes_int = max(0, int(minutes_value or 0))
        rate_per_minute = float(SCHEDULE_AUTO_FINE_RATE_PER_MINUTE or 0.0)
        amount = round(float(minutes_int) * rate_per_minute, 2)
        comment = f"Автоагрегация: {meta['label']} {minutes_int} мин"
        return {
            'amount': float(amount),
            'reason': meta['reason'],
            'comment': comment
        }

    def _schedule_auto_sync_flag_fines_for_range_tx(self, cursor, operator_ids, start_date, end_date):
        op_ids = sorted({int(v) for v in (operator_ids or []) if v is not None})
        if not op_ids:
            return {'inserted': 0, 'updated': 0, 'deleted': 0}

        start_date_obj = self._normalize_schedule_date(start_date)
        end_date_obj = self._normalize_schedule_date(end_date)
        if end_date_obj < start_date_obj:
            start_date_obj, end_date_obj = end_date_obj, start_date_obj

        cursor.execute("""
            SELECT
                id,
                operator_id,
                day,
                COALESCE(late_minutes, 0),
                COALESCE(early_leave_minutes, 0),
                COALESCE(training_minutes, 0),
                COALESCE(technical_reason_minutes, 0),
                late_status,
                early_leave_status,
                training_status,
                technical_reason_status,
                late_fine_id,
                early_leave_fine_id,
                training_fine_id,
                technical_reason_fine_id
            FROM daily_hours
            WHERE operator_id = ANY(%s)
              AND day >= %s
              AND day <= %s
        """, (op_ids, start_date_obj, end_date_obj))
        rows = cursor.fetchall() or []
        if not rows:
            return {'inserted': 0, 'updated': 0, 'deleted': 0}

        inserted_count = 0
        updated_count = 0
        deleted_count = 0

        def sync_one_flag(daily_id, operator_id, day_value, flag_type, minutes_value, status_value, fine_id_value):
            nonlocal inserted_count, updated_count, deleted_count

            minutes_int = max(0, int(minutes_value or 0))
            status_norm = self._schedule_auto_flag_status_for_minutes(status_value, minutes_int)
            fine_id_norm = int(fine_id_value) if fine_id_value is not None else None

            flag_key = str(flag_type or '').strip().lower()
            # Штрафы из авто-флагов создаем только для нарушений дисциплины.
            # Тренинг подтверждается/отклоняется, но не должен попадать в daily_fines.
            supports_fine = flag_key in ('late', 'early', 'early_leave')
            should_keep_fine = supports_fine and (minutes_int > 0 and status_norm == SCHEDULE_AUTO_FLAG_CONFIRMED)
            if should_keep_fine:
                payload = self._schedule_auto_build_fine_payload(flag_type, minutes_int)
                if fine_id_norm is None:
                    cursor.execute("""
                        INSERT INTO daily_fines (daily_hours_id, operator_id, day, amount, reason, comment)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        RETURNING id
                    """, (
                        daily_id,
                        int(operator_id),
                        day_value,
                        float(payload['amount']),
                        payload['reason'],
                        payload['comment']
                    ))
                    fine_id_norm = int(cursor.fetchone()[0])
                    inserted_count += 1
                else:
                    cursor.execute("""
                        UPDATE daily_fines
                        SET daily_hours_id = %s,
                            operator_id = %s,
                            day = %s,
                            amount = %s,
                            reason = %s,
                            comment = %s
                        WHERE id = %s
                    """, (
                        daily_id,
                        int(operator_id),
                        day_value,
                        float(payload['amount']),
                        payload['reason'],
                        payload['comment'],
                        fine_id_norm
                    ))
                    if int(cursor.rowcount or 0) > 0:
                        updated_count += 1
                    else:
                        cursor.execute("""
                            INSERT INTO daily_fines (daily_hours_id, operator_id, day, amount, reason, comment)
                            VALUES (%s, %s, %s, %s, %s, %s)
                            RETURNING id
                        """, (
                            daily_id,
                            int(operator_id),
                            day_value,
                            float(payload['amount']),
                            payload['reason'],
                            payload['comment']
                        ))
                        fine_id_norm = int(cursor.fetchone()[0])
                        inserted_count += 1
            else:
                if fine_id_norm is not None:
                    cursor.execute("DELETE FROM daily_fines WHERE id = %s", (fine_id_norm,))
                    if int(cursor.rowcount or 0) > 0:
                        deleted_count += 1
                fine_id_norm = None

            return status_norm, fine_id_norm

        for row in rows:
            (
                daily_id,
                operator_id,
                day_value,
                late_minutes,
                early_leave_minutes,
                training_minutes,
                technical_reason_minutes,
                late_status,
                early_status,
                training_status,
                technical_reason_status,
                late_fine_id,
                early_fine_id,
                training_fine_id,
                technical_reason_fine_id
            ) = row

            late_status_norm, late_fine_id_norm = sync_one_flag(
                daily_id=daily_id,
                operator_id=operator_id,
                day_value=day_value,
                flag_type='late',
                minutes_value=late_minutes,
                status_value=late_status,
                fine_id_value=late_fine_id
            )
            early_status_norm, early_fine_id_norm = sync_one_flag(
                daily_id=daily_id,
                operator_id=operator_id,
                day_value=day_value,
                flag_type='early_leave',
                minutes_value=early_leave_minutes,
                status_value=early_status,
                fine_id_value=early_fine_id
            )
            training_status_norm, training_fine_id_norm = sync_one_flag(
                daily_id=daily_id,
                operator_id=operator_id,
                day_value=day_value,
                flag_type='training',
                minutes_value=training_minutes,
                status_value=training_status,
                fine_id_value=training_fine_id
            )
            technical_reason_status_norm, technical_reason_fine_id_norm = sync_one_flag(
                daily_id=daily_id,
                operator_id=operator_id,
                day_value=day_value,
                flag_type='technical_reason',
                minutes_value=technical_reason_minutes,
                status_value=technical_reason_status,
                fine_id_value=technical_reason_fine_id
            )

            if (
                self._schedule_auto_normalize_flag_status(late_status) != late_status_norm
                or self._schedule_auto_normalize_flag_status(early_status) != early_status_norm
                or self._schedule_auto_normalize_flag_status(training_status) != training_status_norm
                or self._schedule_auto_normalize_flag_status(technical_reason_status) != technical_reason_status_norm
                or (int(late_fine_id) if late_fine_id is not None else None) != late_fine_id_norm
                or (int(early_fine_id) if early_fine_id is not None else None) != early_fine_id_norm
                or (int(training_fine_id) if training_fine_id is not None else None) != training_fine_id_norm
                or (int(technical_reason_fine_id) if technical_reason_fine_id is not None else None) != technical_reason_fine_id_norm
            ):
                cursor.execute("""
                    UPDATE daily_hours
                    SET late_status = %s,
                        early_leave_status = %s,
                        training_status = %s,
                        technical_reason_status = %s,
                        late_fine_id = %s,
                        early_leave_fine_id = %s,
                        training_fine_id = %s,
                        technical_reason_fine_id = %s,
                        created_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                """, (
                    late_status_norm,
                    early_status_norm,
                    training_status_norm,
                    technical_reason_status_norm,
                    late_fine_id_norm,
                    early_fine_id_norm,
                    training_fine_id_norm,
                    technical_reason_fine_id_norm,
                    daily_id
                ))

        return {
            'inserted': int(inserted_count),
            'updated': int(updated_count),
            'deleted': int(deleted_count)
        }

    def _aggregate_month_from_daily_tx(self, cursor, operator_id, month):
        operator_id = int(operator_id)
        year, mon = map(int, str(month).split('-'))
        start = date(year, mon, 1)
        end = date(year, mon, calendar.monthrange(year, mon)[1])

        cursor.execute("""
            SELECT
                COALESCE(SUM(work_time),0),
                COALESCE(SUM(training_time),0),
                COALESCE(SUM(break_time),0),
                COALESCE(SUM(talk_time),0),
                COALESCE(SUM(calls),0),
                COALESCE(SUM(efficiency),0)
            FROM daily_hours
            WHERE operator_id = %s AND day >= %s AND day <= %s
        """, (operator_id, start, end))
        row = cursor.fetchone()
        total_work_time, total_training_time, total_break_time, total_talk_time, total_calls, total_efficiency_hours = row

        cursor.execute("""
            SELECT COALESCE(SUM(df.amount), 0)
            FROM daily_fines df
            JOIN daily_hours dh ON df.daily_hours_id = dh.id
            WHERE dh.operator_id = %s AND dh.day >= %s AND dh.day <= %s
        """, (operator_id, start, end))
        total_fines = cursor.fetchone()[0] or 0.0

        cursor.execute("""
            SELECT COALESCE(SUM(
                CASE
                    WHEN oa.end_time <= oa.start_time
                        THEN EXTRACT(EPOCH FROM (oa.end_time + INTERVAL '24 hours' - oa.start_time))
                    ELSE EXTRACT(EPOCH FROM (oa.end_time - oa.start_time))
                END
            ) / 3600.0, 0)
            FROM operator_offline_activities oa
            WHERE oa.operator_id = %s
              AND oa.activity_date >= %s
              AND oa.activity_date <= %s
        """, (operator_id, start, end))
        total_offline_hours = float(cursor.fetchone()[0] or 0.0)

        effective_hours_for_calls = max(0.0, float(total_work_time or 0.0))

        if effective_hours_for_calls > 0:
            calls_per_hour = float(total_calls) / float(effective_hours_for_calls)
        else:
            calls_per_hour = 0.0

        cursor.execute("""
            INSERT INTO work_hours (
                operator_id, month, regular_hours, training_hours, fines, total_break_time,
                total_talk_time, total_calls, total_efficiency_hours, calls_per_hour
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (operator_id, month)
            DO UPDATE SET
                regular_hours = EXCLUDED.regular_hours,
                training_hours = EXCLUDED.training_hours,
                fines = EXCLUDED.fines,
                total_break_time = EXCLUDED.total_break_time,
                total_talk_time = EXCLUDED.total_talk_time,
                total_calls = EXCLUDED.total_calls,
                total_efficiency_hours = EXCLUDED.total_efficiency_hours,
                calls_per_hour = EXCLUDED.calls_per_hour,
                created_at = CURRENT_TIMESTAMP
        """, (
            operator_id,
            month,
            float(total_work_time or 0.0),
            float(total_training_time or 0.0),
            float(total_fines or 0.0),
            float(total_break_time or 0.0),
            float(total_talk_time or 0.0),
            int(total_calls or 0),
            float(total_efficiency_hours or 0.0),
            float(calls_per_hour or 0.0)
        ))

        return {
            "regular_hours": float(total_work_time or 0.0),
            "training_hours": float(total_training_time or 0.0),
            "total_break_time": float(total_break_time or 0.0),
            "total_talk_time": float(total_talk_time or 0.0),
            "total_calls": int(total_calls or 0),
            "total_efficiency_hours": float(total_efficiency_hours or 0.0),
            "calls_per_hour": float(calls_per_hour or 0.0),
            "fines": float(total_fines or 0.0),
            "offline_activity_hours": float(total_offline_hours or 0.0)
        }

    def _recalculate_auto_daily_hours_tx(self, cursor, operator_ids, start_date, end_date):
        op_ids = sorted({int(v) for v in (operator_ids or []) if v is not None})
        if not op_ids:
            return {'updated_days': 0, 'aggregated_months': 0}

        start_date_obj = self._normalize_schedule_date(start_date)
        end_date_obj = self._normalize_schedule_date(end_date)
        if end_date_obj < start_date_obj:
            start_date_obj, end_date_obj = end_date_obj, start_date_obj

        # Для ночных смен (например, 20:00-08:00) досчитываем следующий календарный день,
        # чтобы хвост после полуночи попадал в "день+1".
        calc_start_date_obj, calc_end_date_obj = self._schedule_auto_clamp_range(
            start_date_obj,
            end_date_obj + timedelta(days=1)
        )
        if calc_start_date_obj is None or calc_end_date_obj is None:
            return {'updated_days': 0, 'aggregated_months': 0}

        # Для корректного расчета текущего дня учитываем смены, начавшиеся днем ранее
        # и перешедшие через полночь.
        shifts_source_start = calc_start_date_obj - timedelta(days=1)
        shifts_source_end = calc_end_date_obj

        cursor.execute("""
            SELECT id, operator_id, shift_date, start_time, end_time
            FROM work_shifts
            WHERE operator_id = ANY(%s)
              AND shift_date >= %s
              AND shift_date <= %s
            ORDER BY operator_id, shift_date, start_time, end_time
        """, (op_ids, shifts_source_start, shifts_source_end))
        shifts_rows = cursor.fetchall() or []

        shifts_by_start_day = {}
        shifts_segments_by_day = {}
        for shift_id, op_id, shift_date_value, start_time_value, end_time_value in shifts_rows:
            start_min, end_min = self._schedule_interval_minutes(start_time_value, end_time_value)
            start_sec = int(start_min) * 60
            end_sec = int(end_min) * 60
            item = {
                'id': int(shift_id),
                'start': int(start_sec),
                'end': int(end_sec)
            }
            start_key = (int(op_id), shift_date_value)
            shifts_by_start_day.setdefault(start_key, []).append(item)

            shift_start_abs = int(start_sec)
            shift_end_abs = int(end_sec)
            first_day_offset = int(shift_start_abs // 86400)
            last_day_offset = int((shift_end_abs - 1) // 86400)
            for day_offset in range(first_day_offset, last_day_offset + 1):
                day_floor = int(day_offset) * 86400
                day_ceil = day_floor + 86400
                segment_abs_start = max(shift_start_abs, day_floor)
                segment_abs_end = min(shift_end_abs, day_ceil)
                if segment_abs_end <= segment_abs_start:
                    continue

                day_value = shift_date_value + timedelta(days=int(day_offset))
                if day_value < calc_start_date_obj or day_value > calc_end_date_obj:
                    continue

                local_start = int(segment_abs_start - day_floor)
                local_end = int(segment_abs_end - day_floor)
                if local_end <= local_start:
                    continue
                shifts_segments_by_day.setdefault((int(op_id), day_value), []).append({
                    'start': local_start,
                    'end': local_end
                })

        status_start_for_query = calc_start_date_obj - timedelta(days=1)
        status_end_for_query = calc_end_date_obj + timedelta(days=1)
        cursor.execute("""
            SELECT
                operator_id,
                status_date,
                start_at,
                end_at,
                status_key
            FROM operator_status_segments
            WHERE operator_id = ANY(%s)
              AND status_date >= %s
              AND status_date <= %s
            ORDER BY operator_id, status_date, start_at, end_at, id
        """, (op_ids, status_start_for_query, status_end_for_query))
        status_rows = cursor.fetchall() or []
        statuses_by_day = {}
        for op_id, status_date_value, start_at_value, end_at_value, status_key in status_rows:
            day_chunks = self._schedule_auto_split_status_segment_by_day(
                status_date_value=status_date_value,
                start_at_value=start_at_value,
                end_at_value=end_at_value
            )
            if not day_chunks:
                continue
            for chunk in day_chunks:
                chunk_day = chunk.get('day')
                start_sec = int(chunk.get('start', 0))
                end_sec = int(chunk.get('end', 0))
                if not isinstance(chunk_day, date):
                    continue
                if chunk_day < calc_start_date_obj or chunk_day > status_end_for_query:
                    continue
                if end_sec <= start_sec:
                    continue
                day_key = (int(op_id), chunk_day)
                statuses_by_day.setdefault(day_key, []).append({
                    'start': int(start_sec),
                    'end': int(end_sec),
                    'status_key': str(status_key or '').strip().lower()
                })

        cursor.execute("""
            SELECT
                operator_id,
                day,
                COALESCE(auto_aggregated, FALSE),
                late_status,
                early_leave_status,
                training_status,
                technical_reason_status,
                late_fine_id,
                early_leave_fine_id,
                training_fine_id,
                technical_reason_fine_id
            FROM daily_hours
            WHERE operator_id = ANY(%s)
              AND day >= %s
              AND day <= %s
        """, (op_ids, calc_start_date_obj, calc_end_date_obj))
        existing_rows = cursor.fetchall() or []
        existing_by_day = {}
        for row in existing_rows:
            op_id = int(row[0])
            day_value = row[1]
            existing_by_day[(op_id, day_value)] = {
                'auto_aggregated': bool(row[2]),
                'late_status': self._schedule_auto_normalize_flag_status(row[3]),
                'early_leave_status': self._schedule_auto_normalize_flag_status(row[4]),
                'training_status': self._schedule_auto_normalize_flag_status(row[5]),
                'technical_reason_status': self._schedule_auto_normalize_flag_status(row[6]),
                'late_fine_id': int(row[7]) if row[7] is not None else None,
                'early_leave_fine_id': int(row[8]) if row[8] is not None else None,
                'training_fine_id': int(row[9]) if row[9] is not None else None,
                'technical_reason_fine_id': int(row[10]) if row[10] is not None else None
            }

        target_days = set(shifts_segments_by_day.keys())
        for key, item in existing_by_day.items():
            if (
                item.get('auto_aggregated')
                or item.get('late_status')
                or item.get('early_leave_status')
                or item.get('training_status')
                or item.get('technical_reason_status')
            ):
                target_days.add(key)

        if not target_days:
            return {'updated_days': 0, 'aggregated_months': 0}

        upsert_rows = []
        affected_months = set()

        for op_id, day_value in sorted(target_days, key=lambda x: (x[0], x[1])):
            day_statuses = statuses_by_day.get((op_id, day_value)) or []

            def pick_status_intervals(source_items, predicate_fn):
                return self._merge_break_intervals([
                    {'start': int(seg.get('start', 0)), 'end': int(seg.get('end', 0))}
                    for seg in (source_items or [])
                    if predicate_fn(seg)
                ])

            day_work_status = pick_status_intervals(
                day_statuses,
                lambda seg: str(seg.get('status_key') or '') in SCHEDULE_AUTO_WORK_STATUS_KEYS
            )
            day_talk_status = pick_status_intervals(
                day_statuses,
                lambda seg: str(seg.get('status_key') or '') in SCHEDULE_AUTO_TALK_STATUS_KEYS
            )
            day_break_status = pick_status_intervals(
                day_statuses,
                lambda seg: str(seg.get('status_key') or '') in SCHEDULE_AUTO_BREAK_STATUS_KEYS
            )
            day_training_status = pick_status_intervals(
                day_statuses,
                lambda seg: str(seg.get('status_key') or '') == SCHEDULE_AUTO_TRAINING_STATUS_KEY
            )
            day_technical_reason_status = pick_status_intervals(
                day_statuses,
                lambda seg: self._schedule_auto_is_tech_reason_status_key(seg.get('status_key'))
            )
            day_late_excused_status = pick_status_intervals(
                day_statuses,
                lambda seg: self._schedule_auto_is_late_excused_status_key(seg.get('status_key'))
            )

            day_shift_intervals = self._merge_break_intervals(
                shifts_segments_by_day.get((op_id, day_value)) or []
            )
            work_seconds = self._schedule_auto_overlap_minutes(day_shift_intervals, day_work_status)
            talk_seconds = self._schedule_auto_overlap_minutes(day_shift_intervals, day_talk_status)
            # Перерывы считаем по фактическим статусам в пределах смены, независимо от
            # запланированных break-интервалов графика.
            break_seconds = self._schedule_auto_overlap_minutes(day_shift_intervals, day_break_status)
            training_seconds_in_shift = self._schedule_auto_overlap_minutes(day_shift_intervals, day_training_status)
            technical_reason_seconds_in_shift = self._schedule_auto_overlap_minutes(day_shift_intervals, day_technical_reason_status)

            late_seconds_total = 0.0
            early_leave_seconds_total = 0.0
            shifts_started_today = shifts_by_start_day.get((op_id, day_value)) or []
            if shifts_started_today:
                next_day_statuses = statuses_by_day.get((op_id, day_value + timedelta(days=1))) or []
                next_day_work_status = pick_status_intervals(
                    next_day_statuses,
                    lambda seg: str(seg.get('status_key') or '') in SCHEDULE_AUTO_WORK_STATUS_KEYS
                )
                next_day_late_excused_status = pick_status_intervals(
                    next_day_statuses,
                    lambda seg: self._schedule_auto_is_late_excused_status_key(seg.get('status_key'))
                )
                shifted_next_day_work_status = [
                    {
                        'start': int(seg.get('start', 0)) + 86400,
                        'end': int(seg.get('end', 0)) + 86400
                    }
                    for seg in (next_day_work_status or [])
                    if int(seg.get('end', 0)) > int(seg.get('start', 0))
                ]
                shifted_next_day_late_excused_status = [
                    {
                        'start': int(seg.get('start', 0)) + 86400,
                        'end': int(seg.get('end', 0)) + 86400
                    }
                    for seg in (next_day_late_excused_status or [])
                    if int(seg.get('end', 0)) > int(seg.get('start', 0))
                ]
                work_status_for_shift = self._merge_break_intervals(
                    (day_work_status or []) + shifted_next_day_work_status
                )
                late_excused_status_for_shift = self._merge_break_intervals(
                    (day_late_excused_status or []) + shifted_next_day_late_excused_status
                )
                for shift in shifts_started_today:
                    shift_interval = {'start': int(shift['start']), 'end': int(shift['end'])}
                    shift_work_status = self._schedule_auto_intersect_interval_with_list(
                        shift_interval,
                        work_status_for_shift
                    )
                    if not shift_work_status:
                        continue
                    first_work_start = int(shift_work_status[0].get('start', shift_interval['start']))
                    last_work_end = int(shift_work_status[-1].get('end', shift_interval['end']))
                    shift_excused_status = self._schedule_auto_intersect_interval_with_list(
                        shift_interval,
                        late_excused_status_for_shift
                    )
                    late_excused_boundary = self._schedule_auto_shift_start_excused_boundary(
                        shift_interval,
                        shift_excused_status
                    )
                    early_excused_boundary = self._schedule_auto_shift_end_excused_boundary(
                        shift_interval,
                        shift_excused_status
                    )
                    late_seconds_total += max(0, first_work_start - int(late_excused_boundary))
                    early_leave_seconds_total += max(0, int(early_excused_boundary) - last_work_end)

            day_shift_intervals = self._merge_break_intervals(day_shift_intervals)
            work_inside_shift_day = self._schedule_auto_overlap_minutes(day_work_status, day_shift_intervals)
            overtime_seconds_total = max(0.0, self._schedule_auto_total_minutes(day_work_status) - work_inside_shift_day)

            work_time_hours = round(work_seconds / 3600.0, 4)
            break_time_hours = round(break_seconds / 3600.0, 4)
            talk_time_hours = round(talk_seconds / 3600.0, 4)
            efficiency_hours = round(talk_seconds / 3600.0, 4)
            training_time_hours = round(training_seconds_in_shift / 3600.0, 4)

            late_seconds_int = max(0, int(round(late_seconds_total)))
            early_leave_seconds_int = max(0, int(round(early_leave_seconds_total)))
            overtime_seconds_int = max(0, int(round(overtime_seconds_total)))
            training_seconds_int = max(0, int(round(training_seconds_in_shift)))
            technical_reason_seconds_int = max(0, int(round(technical_reason_seconds_in_shift)))

            late_minutes_int = self._schedule_auto_seconds_to_display_minutes(late_seconds_int)
            early_leave_minutes_int = self._schedule_auto_seconds_to_display_minutes(early_leave_seconds_int)
            overtime_minutes_int = self._schedule_auto_seconds_to_display_minutes(overtime_seconds_int)
            training_minutes_int = self._schedule_auto_seconds_to_display_minutes(training_seconds_int)
            technical_reason_minutes_int = self._schedule_auto_seconds_to_display_minutes(technical_reason_seconds_int)

            existing = existing_by_day.get((op_id, day_value)) or {}
            prev_late_status = self._schedule_auto_normalize_flag_status(existing.get('late_status'))
            prev_early_status = self._schedule_auto_normalize_flag_status(existing.get('early_leave_status'))
            prev_training_status = self._schedule_auto_normalize_flag_status(existing.get('training_status'))
            prev_technical_reason_status = self._schedule_auto_normalize_flag_status(existing.get('technical_reason_status'))

            late_status = self._schedule_auto_flag_status_for_minutes(prev_late_status, late_minutes_int)
            early_status = self._schedule_auto_flag_status_for_minutes(prev_early_status, early_leave_minutes_int)
            training_status = self._schedule_auto_flag_status_for_minutes(prev_training_status, training_minutes_int)
            technical_reason_status = self._schedule_auto_flag_status_for_minutes(prev_technical_reason_status, technical_reason_minutes_int)

            upsert_rows.append((
                int(op_id),
                day_value,
                float(work_time_hours),
                float(break_time_hours),
                float(talk_time_hours),
                int(0),
                float(efficiency_hours),
                float(training_time_hours),
                int(late_minutes_int),
                int(late_seconds_int),
                int(early_leave_minutes_int),
                int(early_leave_seconds_int),
                int(overtime_minutes_int),
                int(overtime_seconds_int),
                int(training_minutes_int),
                int(training_seconds_int),
                int(technical_reason_minutes_int),
                int(technical_reason_seconds_int),
                late_status,
                early_status,
                training_status,
                technical_reason_status,
                existing.get('late_fine_id'),
                existing.get('early_leave_fine_id'),
                existing.get('training_fine_id'),
                existing.get('technical_reason_fine_id')
            ))
            affected_months.add((int(op_id), day_value.strftime('%Y-%m')))

        if upsert_rows:
            execute_values(
                cursor,
                """
                INSERT INTO daily_hours (
                    operator_id, day, work_time, break_time, talk_time, calls, efficiency,
                    training_time,
                    late_minutes, late_seconds,
                    early_leave_minutes, early_leave_seconds,
                    overtime_minutes, overtime_seconds,
                    training_minutes, training_seconds,
                    technical_reason_minutes, technical_reason_seconds,
                    late_status, early_leave_status, training_status, technical_reason_status,
                    late_fine_id, early_leave_fine_id, training_fine_id, technical_reason_fine_id,
                    auto_aggregated, auto_aggregated_at
                )
                VALUES %s
                ON CONFLICT (operator_id, day)
                DO UPDATE SET
                    work_time = EXCLUDED.work_time,
                    break_time = EXCLUDED.break_time,
                    talk_time = EXCLUDED.talk_time,
                    calls = EXCLUDED.calls,
                    efficiency = EXCLUDED.efficiency,
                    training_time = EXCLUDED.training_time,
                    late_minutes = EXCLUDED.late_minutes,
                    late_seconds = EXCLUDED.late_seconds,
                    early_leave_minutes = EXCLUDED.early_leave_minutes,
                    early_leave_seconds = EXCLUDED.early_leave_seconds,
                    overtime_minutes = EXCLUDED.overtime_minutes,
                    overtime_seconds = EXCLUDED.overtime_seconds,
                    training_minutes = EXCLUDED.training_minutes,
                    training_seconds = EXCLUDED.training_seconds,
                    technical_reason_minutes = EXCLUDED.technical_reason_minutes,
                    technical_reason_seconds = EXCLUDED.technical_reason_seconds,
                    late_status = EXCLUDED.late_status,
                    early_leave_status = EXCLUDED.early_leave_status,
                    training_status = EXCLUDED.training_status,
                    technical_reason_status = EXCLUDED.technical_reason_status,
                    late_fine_id = EXCLUDED.late_fine_id,
                    early_leave_fine_id = EXCLUDED.early_leave_fine_id,
                    training_fine_id = EXCLUDED.training_fine_id,
                    technical_reason_fine_id = EXCLUDED.technical_reason_fine_id,
                    auto_aggregated = TRUE,
                    auto_aggregated_at = CURRENT_TIMESTAMP,
                    created_at = CURRENT_TIMESTAMP
                """,
                upsert_rows,
                template="""(
                    %s, %s, %s, %s, %s, %s, %s,
                    %s,
                    %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    TRUE, CURRENT_TIMESTAMP
                )""",
                page_size=2000
            )

        fine_sync_result = self._schedule_auto_sync_flag_fines_for_range_tx(
            cursor=cursor,
            operator_ids=op_ids,
            start_date=calc_start_date_obj,
            end_date=calc_end_date_obj
        )

        for op_id, month_key in sorted(affected_months):
            self._aggregate_month_from_daily_tx(cursor, op_id, month_key)

        return {
            'updated_days': len(upsert_rows),
            'aggregated_months': len(affected_months),
            'auto_flag_fines': fine_sync_result
        }

    def recalculate_auto_daily_hours(self, operator_ids, start_date, end_date):
        with self._get_cursor() as cursor:
            return self._recalculate_auto_daily_hours_tx(
                cursor=cursor,
                operator_ids=operator_ids,
                start_date=start_date,
                end_date=end_date
            )

    def _schedule_auto_flag_columns(self, flag_type):
        flag_key = str(flag_type or '').strip().lower()
        if flag_key == 'late':
            return ('late_minutes', 'late_status', 'late_fine_id')
        if flag_key in ('early', 'early_leave'):
            return ('early_leave_minutes', 'early_leave_status', 'early_leave_fine_id')
        if flag_key == 'training':
            return ('training_minutes', 'training_status', 'training_fine_id')
        if flag_key in ('technical_reason', 'tech_reason', 'tech', 'technical'):
            return ('technical_reason_minutes', 'technical_reason_status', 'technical_reason_fine_id')
        raise ValueError("Unsupported flag_type. Allowed: late, early_leave, training, technical_reason")

    def resolve_auto_schedule_flag(self, operator_id, day, flag_type, action):
        operator_id_int = int(operator_id)
        day_obj = self._normalize_schedule_date(day)
        _, status_col, _ = self._schedule_auto_flag_columns(flag_type)

        action_norm = str(action or '').strip().lower()
        if action_norm in ('confirm', 'confirmed'):
            target_status = SCHEDULE_AUTO_FLAG_CONFIRMED
        elif action_norm in ('reject', 'rejected'):
            target_status = SCHEDULE_AUTO_FLAG_REJECTED
        elif action_norm in ('pending', 'reset'):
            target_status = SCHEDULE_AUTO_FLAG_PENDING
        else:
            raise ValueError("Unsupported action. Allowed: confirm, reject, pending")

        with self._get_cursor() as cursor:
            self._recalculate_auto_daily_hours_tx(
                cursor=cursor,
                operator_ids=[operator_id_int],
                start_date=day_obj,
                end_date=day_obj
            )

            cursor.execute("""
                SELECT
                    id,
                    COALESCE(late_minutes, 0),
                    COALESCE(late_seconds, 0),
                    COALESCE(early_leave_minutes, 0),
                    COALESCE(early_leave_seconds, 0),
                    COALESCE(training_minutes, 0),
                    COALESCE(training_seconds, 0),
                    COALESCE(technical_reason_minutes, 0),
                    COALESCE(technical_reason_seconds, 0),
                    late_status,
                    early_leave_status,
                    training_status,
                    technical_reason_status
                FROM daily_hours
                WHERE operator_id = %s
                  AND day = %s
                FOR UPDATE
            """, (operator_id_int, day_obj))
            row = cursor.fetchone()
            if not row:
                raise ValueError("Нет агрегированных данных за выбранный день")

            (
                daily_id,
                late_minutes,
                late_seconds,
                early_leave_minutes,
                early_leave_seconds,
                training_minutes,
                training_seconds,
                technical_reason_minutes,
                technical_reason_seconds,
                late_status,
                early_status,
                training_status,
                technical_reason_status
            ) = row

            late_seconds_int = max(0, int(late_seconds or 0))
            early_seconds_int = max(0, int(early_leave_seconds or 0))
            training_seconds_int = max(0, int(training_seconds or 0))
            technical_reason_seconds_int = max(0, int(technical_reason_seconds or 0))

            late_minutes_int = max(
                max(0, int(late_minutes or 0)),
                self._schedule_auto_seconds_to_display_minutes(late_seconds_int)
            )
            early_minutes_int = max(
                max(0, int(early_leave_minutes or 0)),
                self._schedule_auto_seconds_to_display_minutes(early_seconds_int)
            )
            training_minutes_int = max(
                max(0, int(training_minutes or 0)),
                self._schedule_auto_seconds_to_display_minutes(training_seconds_int)
            )
            technical_reason_minutes_int = max(
                max(0, int(technical_reason_minutes or 0)),
                self._schedule_auto_seconds_to_display_minutes(technical_reason_seconds_int)
            )

            status_map = {
                'late_status': self._schedule_auto_flag_status_for_minutes(late_status, late_minutes_int),
                'early_leave_status': self._schedule_auto_flag_status_for_minutes(early_status, early_minutes_int),
                'training_status': self._schedule_auto_flag_status_for_minutes(training_status, training_minutes_int),
                'technical_reason_status': self._schedule_auto_flag_status_for_minutes(technical_reason_status, technical_reason_minutes_int)
            }
            seconds_by_status_col = {
                'late_status': late_seconds_int,
                'early_leave_status': early_seconds_int,
                'training_status': training_seconds_int,
                'technical_reason_status': technical_reason_seconds_int
            }

            current_seconds = seconds_by_status_col.get(status_col, 0)
            if target_status == SCHEDULE_AUTO_FLAG_CONFIRMED and current_seconds <= 0:
                raise ValueError("Нельзя подтвердить флаг без рассчитанного времени")

            if current_seconds <= 0:
                status_map[status_col] = None
            else:
                status_map[status_col] = target_status

            cursor.execute("""
                UPDATE daily_hours
                SET late_status = %s,
                    early_leave_status = %s,
                    training_status = %s,
                    technical_reason_status = %s,
                    created_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (
                status_map['late_status'],
                status_map['early_leave_status'],
                status_map['training_status'],
                status_map['technical_reason_status'],
                daily_id
            ))

            fine_sync_result = self._schedule_auto_sync_flag_fines_for_range_tx(
                cursor=cursor,
                operator_ids=[operator_id_int],
                start_date=day_obj,
                end_date=day_obj
            )
            self._aggregate_month_from_daily_tx(cursor, operator_id_int, day_obj.strftime('%Y-%m'))

            flags_map = self._load_auto_schedule_flags_for_operators(
                cursor=cursor,
                operator_ids=[operator_id_int],
                start_date_obj=day_obj,
                end_date_obj=day_obj
            )
            day_key = day_obj.strftime('%Y-%m-%d')
            day_flags = (flags_map.get(operator_id_int) or {}).get(day_key, {})

            resolved_minutes = int(day_flags.get({
                'late_status': 'lateMinutes',
                'early_leave_status': 'earlyLeaveMinutes',
                'training_status': 'trainingMinutes',
                'technical_reason_status': 'technicalReasonMinutes'
            }.get(status_col, ''), 0))

            return {
                'operator_id': operator_id_int,
                'day': day_key,
                'flag_type': str(flag_type),
                'status': day_flags.get({
                    'late_status': 'lateStatus',
                    'early_leave_status': 'earlyLeaveStatus',
                    'training_status': 'trainingStatus',
                    'technical_reason_status': 'technicalReasonStatus'
                }.get(status_col, ''), None),
                'minutes': resolved_minutes,
                'day_flags': day_flags,
                'fine_sync': fine_sync_result
            }

    def add_manual_fine_for_day(self, operator_id, day, amount, reason, comment=None):
        operator_id_int = int(operator_id)
        day_obj = self._normalize_schedule_date(day)

        amount_value = round(float(amount or 0.0), 2)
        if amount_value <= 0:
            raise ValueError("amount must be greater than 0")

        reason_norm = str(reason or '').strip()
        if not reason_norm:
            raise ValueError("reason is required")
        if reason_norm.lower() in ('тренинг', 'training'):
            raise ValueError("Training must be saved in trainings table")

        comment_norm = str(comment or '').strip() or None

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id
                FROM daily_hours
                WHERE operator_id = %s AND day = %s
                FOR UPDATE
            """, (operator_id_int, day_obj))
            row = cursor.fetchone()

            if row:
                daily_id = row[0]
            else:
                cursor.execute("""
                    INSERT INTO daily_hours (
                        operator_id,
                        day,
                        work_time,
                        break_time,
                        talk_time,
                        calls,
                        efficiency,
                        training_time,
                        late_minutes,
                        early_leave_minutes,
                        overtime_minutes,
                        training_minutes,
                        auto_aggregated,
                        auto_aggregated_at
                    )
                    VALUES (
                        %s, %s,
                        0, 0, 0, 0, 0,
                        0, 0, 0, 0, 0,
                        TRUE, CURRENT_TIMESTAMP
                    )
                    ON CONFLICT (operator_id, day)
                    DO UPDATE SET operator_id = EXCLUDED.operator_id
                    RETURNING id
                """, (operator_id_int, day_obj))
                daily_id = cursor.fetchone()[0]

            cursor.execute("""
                INSERT INTO daily_fines (daily_hours_id, operator_id, day, amount, reason, comment)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                daily_id,
                operator_id_int,
                day_obj,
                amount_value,
                reason_norm,
                comment_norm
            ))
            fine_id = int(cursor.fetchone()[0])

            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0)
                FROM daily_fines
                WHERE daily_hours_id = %s
            """, (daily_id,))
            total_fine_amount = float(cursor.fetchone()[0] or 0.0)

            cursor.execute("""
                SELECT reason, comment
                FROM daily_fines
                WHERE daily_hours_id = %s
                ORDER BY id DESC
                LIMIT 1
            """, (daily_id,))
            latest_fine_row = cursor.fetchone() or (None, None)
            latest_reason = latest_fine_row[0]
            latest_comment = latest_fine_row[1]

            cursor.execute("""
                UPDATE daily_hours
                SET fine_amount = %s,
                    fine_reason = %s,
                    fine_comment = %s,
                    created_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (
                total_fine_amount,
                latest_reason,
                latest_comment,
                daily_id
            ))

            self._aggregate_month_from_daily_tx(cursor, operator_id_int, day_obj.strftime('%Y-%m'))

            return {
                'operator_id': operator_id_int,
                'day': day_obj.strftime('%Y-%m-%d'),
                'daily_hours_id': str(daily_id),
                'fine': {
                    'id': fine_id,
                    'amount': float(amount_value),
                    'reason': reason_norm,
                    'comment': comment_norm
                },
                'fine_total_for_day': float(total_fine_amount)
            }

    def _resolve_breaks_for_day_replacement_tx(self, cursor, operator_id, shift_date, start_time, end_time, breaks):
        """
        Для replace-сценариев (bulk set_shift): если клиент прислал старые автоперерывы
        и смена стала короче, переводим на серверную автогенерацию (возвращаем None).
        """
        if breaks is None:
            return None

        incoming_breaks_norm = self._normalize_shift_breaks(breaks)
        cursor.execute(
            """
            SELECT id, start_time, end_time
            FROM work_shifts
            WHERE operator_id = %s AND shift_date = %s
            ORDER BY start_time
            """,
            (int(operator_id), self._normalize_schedule_date(shift_date))
        )
        existing_rows = cursor.fetchall() or []
        if len(existing_rows) != 1:
            return incoming_breaks_norm

        direction_name = self._get_operator_direction_name_tx(cursor, operator_id)
        direction_rules = self._get_work_schedule_break_rules_for_direction_tx(cursor, direction_name)
        shift_id, existing_start_value, existing_end_value = existing_rows[0]
        existing_breaks = self._load_shift_breaks_tx(cursor, shift_id)
        if not existing_breaks:
            return incoming_breaks_norm
        if incoming_breaks_norm != existing_breaks:
            return incoming_breaks_norm

        existing_start_min, existing_end_min = self._schedule_interval_minutes(existing_start_value, existing_end_value)
        existing_duration = max(0, existing_end_min - existing_start_min)
        existing_auto_breaks = self._compute_auto_shift_breaks_minutes(
            existing_start_min,
            existing_end_min,
            direction_name=direction_name,
            direction_rules=direction_rules
        )
        if existing_breaks != existing_auto_breaks:
            return incoming_breaks_norm

        new_start_min, new_end_min = self._schedule_interval_minutes(start_time, end_time)
        new_duration = max(0, new_end_min - new_start_min)
        if new_duration < existing_duration:
            return None

        return incoming_breaks_norm

    def _save_shift_tx(
        self,
        cursor,
        operator_id,
        shift_date,
        start_time,
        end_time,
        breaks=None,
        previous_start_time=None,
        previous_end_time=None
    ):
        new_start_min, new_end_min = self._schedule_interval_minutes(start_time, end_time)
        new_duration_min = max(0, new_end_min - new_start_min)
        direction_name = self._get_operator_direction_name_tx(cursor, operator_id)
        direction_rules = self._get_work_schedule_break_rules_for_direction_tx(cursor, direction_name)

        prev_start_obj = None
        prev_end_obj = None
        if previous_start_time is not None and previous_end_time is not None:
            prev_start_obj = self._normalize_schedule_time(previous_start_time, 'previous_start_time')
            prev_end_obj = self._normalize_schedule_time(previous_end_time, 'previous_end_time')

        comparison_start_obj = prev_start_obj if prev_start_obj is not None else start_time
        comparison_end_obj = prev_end_obj if prev_end_obj is not None else end_time
        cursor.execute(
            """
            SELECT id, start_time, end_time
            FROM work_shifts
            WHERE operator_id = %s AND shift_date = %s
              AND start_time = %s AND end_time = %s
            LIMIT 1
            """,
            (operator_id, shift_date, comparison_start_obj, comparison_end_obj)
        )
        previous_shift_row = cursor.fetchone()
        previous_shift_breaks = []
        previous_shift_breaks_were_auto = False
        previous_duration_min = 0
        if previous_shift_row:
            previous_shift_id, previous_start_value, previous_end_value = previous_shift_row
            previous_start_min, previous_end_min = self._schedule_interval_minutes(previous_start_value, previous_end_value)
            previous_duration_min = max(0, previous_end_min - previous_start_min)
            previous_shift_breaks = self._load_shift_breaks_tx(cursor, previous_shift_id)
            previous_auto_breaks = self._compute_auto_shift_breaks_minutes(
                previous_start_min,
                previous_end_min,
                direction_name=direction_name,
                direction_rules=direction_rules
            )
            previous_shift_breaks_were_auto = (previous_shift_breaks == previous_auto_breaks)

        if breaks is None:
            breaks_norm = self._compute_auto_shift_breaks_minutes(
                new_start_min,
                new_end_min,
                direction_name=direction_name,
                direction_rules=direction_rules
            )
        else:
            incoming_breaks_norm = self._normalize_shift_breaks(breaks)
            should_regenerate_auto_breaks = (
                previous_shift_row is not None
                and previous_shift_breaks_were_auto
                and incoming_breaks_norm == previous_shift_breaks
                and new_duration_min < previous_duration_min
            )
            breaks_norm = (
                self._compute_auto_shift_breaks_minutes(
                    new_start_min,
                    new_end_min,
                    direction_name=direction_name,
                    direction_rules=direction_rules
                )
                if should_regenerate_auto_breaks
                else incoming_breaks_norm
            )

        # Приводим перерывы к правилам анти-пересечения между операторами одного направления.
        breaks_norm = self._adjust_shift_breaks_against_occupied_tx(
            cursor=cursor,
            operator_id=operator_id,
            shift_date=shift_date,
            start_time=start_time,
            end_time=end_time,
            breaks=breaks_norm
        )

        # Если редактировали существующую смену по previous_* и время изменилось,
        # удаляем старую запись, чтобы не оставлять дубль.
        if prev_start_obj is not None and prev_end_obj is not None:
            if (
                prev_start_obj != start_time or prev_end_obj != end_time
            ):
                cursor.execute(
                    """
                    DELETE FROM work_shifts
                    WHERE operator_id = %s AND shift_date = %s
                      AND start_time = %s AND end_time = %s
                    """,
                    (operator_id, shift_date, prev_start_obj, prev_end_obj)
                )

        cursor.execute(
            """
            SELECT id, start_time, end_time
            FROM work_shifts
            WHERE operator_id = %s AND shift_date = %s
            """,
            (operator_id, shift_date)
        )
        existing_rows = cursor.fetchall()

        overlap_ids = []
        for shift_id, existing_start, existing_end in existing_rows:
            if existing_start == start_time and existing_end == end_time:
                continue
            ex_start_min, ex_end_min = self._schedule_interval_minutes(existing_start, existing_end)
            if new_start_min < ex_end_min and ex_start_min < new_end_min:
                overlap_ids.append(shift_id)

        if overlap_ids:
            cursor.execute(
                "DELETE FROM work_shifts WHERE id = ANY(%s)",
                (overlap_ids,)
            )

        cursor.execute(
            """
            INSERT INTO work_shifts (operator_id, shift_date, start_time, end_time, updated_at)
            VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (operator_id, shift_date, start_time, end_time)
            DO UPDATE SET updated_at = CURRENT_TIMESTAMP
            RETURNING id
            """,
            (operator_id, shift_date, start_time, end_time)
        )
        shift_id = cursor.fetchone()[0]
        self._insert_shift_breaks(cursor, shift_id, breaks_norm)

        cursor.execute(
            """
            DELETE FROM days_off
            WHERE operator_id = %s AND day_off_date = %s
            """,
            (operator_id, shift_date)
        )

        # Если на дату сохранённой смены приходился период "увольнение",
        # считаем увольнение прерванным (не продолжаем его после рабочего дня).
        self._interrupt_dismissal_period_by_work_day_tx(cursor, operator_id, shift_date)
        self._sync_user_statuses_from_schedule_periods_tx(cursor, operator_ids=[operator_id])

        return shift_id

    def _normalize_schedule_status_code(self, status_code):
        code = str(status_code or '').strip()
        if code not in SCHEDULE_SPECIAL_STATUS_META:
            raise ValueError("Invalid status_code")
        return code

    def _serialize_schedule_status_period(self, row):
        period_id, operator_id, status_code, start_date_value, end_date_value, dismissal_reason, comment = row[:7]
        is_blacklist = bool(row[7]) if len(row) > 7 else False
        meta = SCHEDULE_SPECIAL_STATUS_META.get(status_code, {})
        return {
            'id': int(period_id),
            'operatorId': int(operator_id),
            'statusCode': status_code,
            'label': meta.get('label') or status_code,
            'kind': meta.get('kind') or 'absence',
            'startDate': start_date_value.strftime('%Y-%m-%d') if start_date_value else None,
            'endDate': end_date_value.strftime('%Y-%m-%d') if end_date_value else None,
            'dismissalReason': dismissal_reason,
            'comment': comment or '',
            'isBlacklist': is_blacklist,
            'is_blacklist': is_blacklist
        }

    def _load_schedule_status_periods_for_operators(self, cursor, operator_ids, start_date_obj=None, end_date_obj=None):
        operator_ids = [int(v) for v in (operator_ids or []) if v is not None]
        result = {
            op_id: {
                'scheduleStatusPeriods': [],
                'scheduleStatusDays': {}
            }
            for op_id in operator_ids
        }
        if not operator_ids:
            return result

        query = """
            SELECT id, operator_id, status_code, start_date, end_date, dismissal_reason, comment, COALESCE(is_blacklist, FALSE)
            FROM operator_schedule_status_periods
            WHERE operator_id = ANY(%s)
        """
        params = [operator_ids]

        if start_date_obj and end_date_obj:
            query += " AND start_date <= %s AND COALESCE(end_date, DATE '9999-12-31') >= %s"
            params.extend([end_date_obj, start_date_obj])
        elif start_date_obj:
            query += " AND COALESCE(end_date, DATE '9999-12-31') >= %s"
            params.append(start_date_obj)
        elif end_date_obj:
            query += " AND start_date <= %s"
            params.append(end_date_obj)

        query += " ORDER BY operator_id, start_date, id"
        cursor.execute(query, params)
        rows = cursor.fetchall()

        should_expand_days = bool(start_date_obj and end_date_obj)
        for row in rows:
            op_id = int(row[1])
            bucket = result.get(op_id)
            if bucket is None:
                continue

            period_payload = self._serialize_schedule_status_period(row)
            bucket['scheduleStatusPeriods'].append(period_payload)

            if not should_expand_days:
                continue

            period_start = row[3]
            period_end = row[4] or end_date_obj
            overlap_start = max(period_start, start_date_obj)
            overlap_end = min(period_end, end_date_obj)
            if overlap_end < overlap_start:
                continue

            day_value = {
                'id': period_payload['id'],
                'statusCode': period_payload['statusCode'],
                'label': period_payload['label'],
                'kind': period_payload['kind'],
                'startDate': period_payload['startDate'],
                'endDate': period_payload['endDate'],
                'dismissalReason': period_payload['dismissalReason'],
                'comment': period_payload['comment']
            }
            cur_day = overlap_start
            while cur_day <= overlap_end:
                bucket['scheduleStatusDays'][cur_day.strftime('%Y-%m-%d')] = day_value
                cur_day += timedelta(days=1)

        return result

    def _attach_schedule_status_periods_to_operators(self, cursor, operators_map, operator_ids, start_date_obj=None, end_date_obj=None):
        statuses_map = self._load_schedule_status_periods_for_operators(
            cursor=cursor,
            operator_ids=operator_ids,
            start_date_obj=start_date_obj,
            end_date_obj=end_date_obj
        )
        for op_id in (operator_ids or []):
            target = operators_map.get(op_id)
            if target is None:
                continue
            data = statuses_map.get(op_id) or {}
            target['scheduleStatusPeriods'] = data.get('scheduleStatusPeriods', [])
            target['scheduleStatusDays'] = data.get('scheduleStatusDays', {})

    def _load_auto_schedule_flags_for_operators(self, cursor, operator_ids, start_date_obj=None, end_date_obj=None):
        op_ids = [int(v) for v in (operator_ids or []) if v is not None]
        result = {op_id: {} for op_id in op_ids}
        if not op_ids:
            return result

        query = """
            SELECT
                operator_id,
                day,
                COALESCE(late_minutes, 0),
                COALESCE(late_seconds, 0),
                COALESCE(early_leave_minutes, 0),
                COALESCE(early_leave_seconds, 0),
                COALESCE(overtime_minutes, 0),
                COALESCE(overtime_seconds, 0),
                COALESCE(training_minutes, 0),
                COALESCE(training_seconds, 0),
                COALESCE(technical_reason_minutes, 0),
                COALESCE(technical_reason_seconds, 0),
                late_status,
                early_leave_status,
                training_status,
                technical_reason_status,
                late_fine_id,
                early_leave_fine_id,
                training_fine_id,
                technical_reason_fine_id
            FROM daily_hours
            WHERE operator_id = ANY(%s)
              AND (
                    COALESCE(late_minutes, 0) > 0
                 OR COALESCE(late_seconds, 0) > 0
                 OR COALESCE(early_leave_minutes, 0) > 0
                 OR COALESCE(early_leave_seconds, 0) > 0
                 OR COALESCE(overtime_minutes, 0) > 0
                 OR COALESCE(overtime_seconds, 0) > 0
                 OR COALESCE(training_minutes, 0) > 0
                 OR COALESCE(training_seconds, 0) > 0
                 OR COALESCE(technical_reason_minutes, 0) > 0
                 OR COALESCE(technical_reason_seconds, 0) > 0
                 OR late_status IS NOT NULL
                 OR early_leave_status IS NOT NULL
                 OR training_status IS NOT NULL
                 OR technical_reason_status IS NOT NULL
              )
        """
        params = [op_ids]

        if start_date_obj and end_date_obj:
            query += " AND day >= %s AND day <= %s"
            params.extend([start_date_obj, end_date_obj])
        elif start_date_obj:
            query += " AND day >= %s"
            params.append(start_date_obj)
        elif end_date_obj:
            query += " AND day <= %s"
            params.append(end_date_obj)

        query += " ORDER BY operator_id, day"
        cursor.execute(query, params)
        rows = cursor.fetchall() or []

        for row in rows:
            (
                operator_id,
                day_value,
                late_minutes,
                late_seconds,
                early_leave_minutes,
                early_leave_seconds,
                overtime_minutes,
                overtime_seconds,
                training_minutes,
                training_seconds,
                technical_reason_minutes,
                technical_reason_seconds,
                late_status,
                early_status,
                training_status,
                technical_reason_status,
                late_fine_id,
                early_fine_id,
                training_fine_id,
                technical_reason_fine_id
            ) = row
            op_id = int(operator_id)
            day_key = day_value.strftime('%Y-%m-%d') if hasattr(day_value, 'strftime') else str(day_value or '')
            if not day_key:
                continue

            late_seconds_int = max(0, int(late_seconds or 0))
            early_leave_seconds_int = max(0, int(early_leave_seconds or 0))
            overtime_seconds_int = max(0, int(overtime_seconds or 0))
            training_seconds_int = max(0, int(training_seconds or 0))
            technical_reason_seconds_int = max(0, int(technical_reason_seconds or 0))

            # Для совместимости с историческими данными берем максимум из старых минут
            # и минут, рассчитанных по секундам.
            late_minutes_int = max(
                max(0, int(late_minutes or 0)),
                self._schedule_auto_seconds_to_display_minutes(late_seconds_int)
            )
            early_leave_minutes_int = max(
                max(0, int(early_leave_minutes or 0)),
                self._schedule_auto_seconds_to_display_minutes(early_leave_seconds_int)
            )
            overtime_minutes_int = max(
                max(0, int(overtime_minutes or 0)),
                self._schedule_auto_seconds_to_display_minutes(overtime_seconds_int)
            )
            training_minutes_int = max(
                max(0, int(training_minutes or 0)),
                self._schedule_auto_seconds_to_display_minutes(training_seconds_int)
            )
            technical_reason_minutes_int = max(
                max(0, int(technical_reason_minutes or 0)),
                self._schedule_auto_seconds_to_display_minutes(technical_reason_seconds_int)
            )

            late_status_norm = self._schedule_auto_flag_status_for_minutes(late_status, late_minutes_int)
            early_status_norm = self._schedule_auto_flag_status_for_minutes(early_status, early_leave_minutes_int)
            training_status_norm = self._schedule_auto_flag_status_for_minutes(training_status, training_minutes_int)
            technical_reason_status_norm = self._schedule_auto_flag_status_for_minutes(technical_reason_status, technical_reason_minutes_int)

            has_late = late_minutes_int > 0 and late_status_norm != SCHEDULE_AUTO_FLAG_REJECTED
            has_early_leave = early_leave_minutes_int > 0 and early_status_norm != SCHEDULE_AUTO_FLAG_REJECTED
            has_training = training_minutes_int > 0 and training_status_norm != SCHEDULE_AUTO_FLAG_REJECTED
            has_technical_reason = technical_reason_minutes_int > 0 and technical_reason_status_norm != SCHEDULE_AUTO_FLAG_REJECTED
            has_pending = (
                (late_status_norm == SCHEDULE_AUTO_FLAG_PENDING and late_minutes_int > 0)
                or (early_status_norm == SCHEDULE_AUTO_FLAG_PENDING and early_leave_minutes_int > 0)
                or (training_status_norm == SCHEDULE_AUTO_FLAG_PENDING and training_minutes_int > 0)
                or (technical_reason_status_norm == SCHEDULE_AUTO_FLAG_PENDING and technical_reason_minutes_int > 0)
            )

            result.setdefault(op_id, {})[day_key] = {
                'lateMinutes': late_minutes_int,
                'lateSeconds': late_seconds_int,
                'earlyLeaveMinutes': early_leave_minutes_int,
                'earlyLeaveSeconds': early_leave_seconds_int,
                'overtimeMinutes': overtime_minutes_int,
                'overtimeSeconds': overtime_seconds_int,
                'trainingMinutes': training_minutes_int,
                'trainingSeconds': training_seconds_int,
                'technicalReasonMinutes': technical_reason_minutes_int,
                'technicalReasonSeconds': technical_reason_seconds_int,
                'lateStatus': late_status_norm,
                'earlyLeaveStatus': early_status_norm,
                'trainingStatus': training_status_norm,
                'technicalReasonStatus': technical_reason_status_norm,
                'lateFineId': int(late_fine_id) if late_fine_id is not None else None,
                'earlyLeaveFineId': int(early_fine_id) if early_fine_id is not None else None,
                'trainingFineId': int(training_fine_id) if training_fine_id is not None else None,
                'technicalReasonFineId': int(technical_reason_fine_id) if technical_reason_fine_id is not None else None,
                'hasLate': bool(has_late),
                'hasEarlyLeave': bool(has_early_leave),
                'hasTraining': bool(has_training),
                'hasTechnicalReason': bool(has_technical_reason),
                'hasDefect': bool(has_late or has_early_leave or has_training or has_technical_reason),
                'hasOvertime': bool(overtime_minutes_int > 0),
                'hasPending': bool(has_pending)
            }

        return result

    def _attach_auto_schedule_flags_to_operators(self, cursor, operators_map, operator_ids, start_date_obj=None, end_date_obj=None):
        flags_map = self._load_auto_schedule_flags_for_operators(
            cursor=cursor,
            operator_ids=operator_ids,
            start_date_obj=start_date_obj,
            end_date_obj=end_date_obj
        )
        for op_id in (operator_ids or []):
            target = operators_map.get(op_id)
            if target is None:
                continue
            target['aggregatedScheduleFlagsDays'] = flags_map.get(int(op_id), {})

    def _load_imported_status_segments_for_operators(self, cursor, operator_ids, start_date_obj=None, end_date_obj=None):
        operator_ids = [int(v) for v in (operator_ids or []) if v is not None]
        result = {op_id: {} for op_id in operator_ids}
        if not operator_ids:
            return result

        min_day = start_date_obj - timedelta(days=1) if start_date_obj else None
        max_day = end_date_obj + timedelta(days=1) if end_date_obj else None

        def _append_segment(op_id, status_date_value, start_at_value, end_at_value, duration_sec, status_key, state_note):
            target = result.get(int(op_id))
            if target is None:
                return

            day_key = status_date_value.strftime('%Y-%m-%d') if hasattr(status_date_value, 'strftime') else str(status_date_value or '')
            if not day_key:
                return

            status_key_norm = self._normalize_import_status_key(status_key)
            if not status_key_norm:
                return
            state_label_value = self._status_label_from_key(status_key_norm)
            is_work = status_key_norm in SCHEDULE_AUTO_WORK_STATUS_KEYS
            is_break = status_key_norm in SCHEDULE_AUTO_BREAK_STATUS_KEYS
            is_no_phone = status_key_norm == SCHEDULE_AUTO_NO_PHONE_STATUS_KEY

            target.setdefault(day_key, []).append({
                'statusDate': day_key,
                'start': start_at_value.isoformat() if hasattr(start_at_value, 'isoformat') else str(start_at_value or ''),
                'end': end_at_value.isoformat() if hasattr(end_at_value, 'isoformat') else str(end_at_value or ''),
                'durationSec': int(duration_sec or 0),
                'stateName': str(state_label_value or ''),
                'stateKey': str(status_key_norm or ''),
                'stateNote': str(state_note or ''),
                'isWork': bool(is_work),
                'isBreak': bool(is_break),
                'isNoPhone': bool(is_no_phone)
            })

        query = """
            SELECT
                operator_id,
                status_date,
                start_at,
                end_at,
                duration_sec,
                status_key,
                state_note
            FROM operator_status_segments
            WHERE operator_id = ANY(%s)
        """
        params = [operator_ids]

        if min_day and max_day:
            query += " AND status_date >= %s AND status_date <= %s"
            params.extend([min_day, max_day])
        elif min_day:
            query += " AND status_date >= %s"
            params.append(min_day)
        elif max_day:
            query += " AND status_date <= %s"
            params.append(max_day)

        query += " ORDER BY operator_id, status_date, start_at, end_at, id"
        cursor.execute(query, params)
        rows = cursor.fetchall() or []

        for row in rows:
            (
                operator_id,
                status_date_value,
                start_at_value,
                end_at_value,
                duration_sec,
                status_key,
                state_note
            ) = row
            op_id = int(operator_id)
            _append_segment(
                op_id=op_id,
                status_date_value=status_date_value,
                start_at_value=start_at_value,
                end_at_value=end_at_value,
                duration_sec=duration_sec,
                status_key=status_key,
                state_note=state_note
            )

        for op_days in result.values():
            for segments in op_days.values():
                segments.sort(key=lambda item: (
                    str(item.get('start') or ''),
                    str(item.get('end') or ''),
                    str(item.get('stateKey') or '')
                ))

        return result

    def _attach_imported_status_segments_to_operators(self, cursor, operators_map, operator_ids, start_date_obj=None, end_date_obj=None):
        timeline_map = self._load_imported_status_segments_for_operators(
            cursor=cursor,
            operator_ids=operator_ids,
            start_date_obj=start_date_obj,
            end_date_obj=end_date_obj
        )
        for op_id in (operator_ids or []):
            target = operators_map.get(op_id)
            if target is None:
                continue
            target['importedStatusTimelineDays'] = timeline_map.get(int(op_id), {})

    def _load_technical_issue_segments_for_operators(self, cursor, operator_ids, start_date_obj=None, end_date_obj=None):
        op_ids = [int(v) for v in (operator_ids or []) if v is not None]
        result = {op_id: {} for op_id in op_ids}
        if not op_ids:
            return result

        query = """
            SELECT
                id,
                operator_id,
                batch_id,
                issue_date,
                start_time,
                end_time,
                reason,
                comment
            FROM operator_technical_issues
            WHERE operator_id = ANY(%s)
        """
        params = [op_ids]
        if start_date_obj:
            query += " AND issue_date >= %s"
            params.append(start_date_obj - timedelta(days=1))
        if end_date_obj:
            query += " AND issue_date <= %s"
            params.append(end_date_obj + timedelta(days=1))

        query += " ORDER BY operator_id, issue_date, start_time, end_time, id"
        cursor.execute(query, params)
        rows = cursor.fetchall() or []

        for row in rows:
            issue_id, operator_id, batch_id, issue_date_value, start_time_value, end_time_value, reason, comment = row
            op_id = int(operator_id)
            if op_id not in result or issue_date_value is None:
                continue

            start_min, end_min = self._schedule_interval_minutes(start_time_value, end_time_value)
            day_chunks = self._split_absolute_minutes_intervals_by_day(
                [{'start': int(start_min), 'end': int(end_min)}],
                issue_date_value
            )
            for chunk in day_chunks:
                chunk_date = chunk.get('date')
                if not chunk_date:
                    continue
                if start_date_obj and chunk_date < start_date_obj:
                    continue
                if end_date_obj and chunk_date > end_date_obj:
                    continue

                day_key = chunk_date.strftime('%Y-%m-%d')
                start_min_local = int(chunk.get('start_min') or 0)
                end_min_local = int(chunk.get('end_min') or 0)
                if end_min_local <= start_min_local:
                    continue

                result[op_id].setdefault(day_key, []).append({
                    'id': int(issue_id),
                    'batch_id': str(batch_id) if batch_id is not None else None,
                    'date': day_key,
                    'start': chunk.get('start_time') or _minutes_to_time(start_min_local),
                    'end': chunk.get('end_time') or _minutes_to_time(end_min_local),
                    'startMin': start_min_local,
                    'endMin': end_min_local,
                    'reason': str(reason or ''),
                    'comment': str(comment or '')
                })

        for op_id, days_map in result.items():
            for day_key, items in list(days_map.items()):
                items_sorted = sorted(
                    (items or []),
                    key=lambda seg: (int(seg.get('startMin', 0)), int(seg.get('endMin', 0)), int(seg.get('id', 0)))
                )
                days_map[day_key] = items_sorted

        return result

    def _attach_technical_issue_segments_to_operators(self, cursor, operators_map, operator_ids, start_date_obj=None, end_date_obj=None):
        timeline_map = self._load_technical_issue_segments_for_operators(
            cursor=cursor,
            operator_ids=operator_ids,
            start_date_obj=start_date_obj,
            end_date_obj=end_date_obj
        )
        for op_id in (operator_ids or []):
            target = operators_map.get(op_id)
            if target is None:
                continue
            target['technicalIssueTimelineDays'] = timeline_map.get(int(op_id), {})

    def _load_offline_activity_segments_for_operators(self, cursor, operator_ids, start_date_obj=None, end_date_obj=None):
        op_ids = [int(v) for v in (operator_ids or []) if v is not None]
        result = {op_id: {} for op_id in op_ids}
        if not op_ids:
            return result

        query = """
            SELECT
                id,
                operator_id,
                batch_id,
                activity_date,
                start_time,
                end_time,
                comment
            FROM operator_offline_activities
            WHERE operator_id = ANY(%s)
        """
        params = [op_ids]
        if start_date_obj:
            query += " AND activity_date >= %s"
            params.append(start_date_obj - timedelta(days=1))
        if end_date_obj:
            query += " AND activity_date <= %s"
            params.append(end_date_obj + timedelta(days=1))

        query += " ORDER BY operator_id, activity_date, start_time, end_time, id"
        cursor.execute(query, params)
        rows = cursor.fetchall() or []

        for row in rows:
            activity_id, operator_id, batch_id, activity_date_value, start_time_value, end_time_value, comment = row
            op_id = int(operator_id)
            if op_id not in result or activity_date_value is None:
                continue

            start_min, end_min = self._schedule_interval_minutes(start_time_value, end_time_value)
            day_chunks = self._split_absolute_minutes_intervals_by_day(
                [{'start': int(start_min), 'end': int(end_min)}],
                activity_date_value
            )
            for chunk in day_chunks:
                chunk_date = chunk.get('date')
                if not chunk_date:
                    continue
                if start_date_obj and chunk_date < start_date_obj:
                    continue
                if end_date_obj and chunk_date > end_date_obj:
                    continue

                day_key = chunk_date.strftime('%Y-%m-%d')
                start_min_local = int(chunk.get('start_min') or 0)
                end_min_local = int(chunk.get('end_min') or 0)
                if end_min_local <= start_min_local:
                    continue

                comment_text = str(comment or '').strip()
                result[op_id].setdefault(day_key, []).append({
                    'id': int(activity_id),
                    'batch_id': str(batch_id) if batch_id is not None else None,
                    'date': day_key,
                    'start': chunk.get('start_time') or _minutes_to_time(start_min_local),
                    'end': chunk.get('end_time') or _minutes_to_time(end_min_local),
                    'startMin': start_min_local,
                    'endMin': end_min_local,
                    'comment': comment_text
                })

        for op_id, days_map in result.items():
            for day_key, items in list(days_map.items()):
                items_sorted = sorted(
                    (items or []),
                    key=lambda seg: (int(seg.get('startMin', 0)), int(seg.get('endMin', 0)), int(seg.get('id', 0)))
                )
                days_map[day_key] = items_sorted

        return result

    def _attach_offline_activity_segments_to_operators(self, cursor, operators_map, operator_ids, start_date_obj=None, end_date_obj=None):
        timeline_map = self._load_offline_activity_segments_for_operators(
            cursor=cursor,
            operator_ids=operator_ids,
            start_date_obj=start_date_obj,
            end_date_obj=end_date_obj
        )
        for op_id in (operator_ids or []):
            target = operators_map.get(op_id)
            if target is None:
                continue
            target['offlineActivityTimelineDays'] = timeline_map.get(int(op_id), {})

    def _sync_user_statuses_from_schedule_periods_tx(self, cursor, operator_ids=None, as_of_date=None):
        """
        Синхронизировать users.status по активным периодным статусам на дату.
        Активные периодные статусы имеют приоритет; при их отсутствии временные статусные значения
        (bs/sick_leave/annual_leave) сбрасываются в working.
        """
        as_of_date_obj = self._normalize_schedule_date(as_of_date) if as_of_date else datetime.now().date()

        if operator_ids is None:
            cursor.execute("""
                SELECT id
                FROM users
                WHERE role = 'operator'
            """)
            target_operator_ids = [int(row[0]) for row in cursor.fetchall()]
        else:
            target_operator_ids = [int(v) for v in operator_ids if v is not None]

        if not target_operator_ids:
            return 0

        cursor.execute("""
            SELECT id, status
            FROM users
            WHERE id = ANY(%s)
        """, (target_operator_ids,))
        current_status_map = {int(row[0]): (row[1] or 'working') for row in cursor.fetchall()}
        if not current_status_map:
            return 0

        cursor.execute("""
            SELECT DISTINCT ON (operator_id) operator_id, status_code
            FROM operator_schedule_status_periods
            WHERE operator_id = ANY(%s)
              AND start_date <= %s
              AND COALESCE(end_date, DATE '9999-12-31') >= %s
            ORDER BY operator_id, start_date DESC, id DESC
        """, (target_operator_ids, as_of_date_obj, as_of_date_obj))
        active_status_by_operator = {int(row[0]): str(row[1]) for row in cursor.fetchall()}

        cursor.execute("""
            SELECT DISTINCT operator_id
            FROM operator_schedule_status_periods
            WHERE operator_id = ANY(%s)
              AND status_code = 'dismissal'
        """, (target_operator_ids,))
        has_dismissal_history = {int(row[0]) for row in cursor.fetchall()}

        updated_count = 0
        temporary_schedule_statuses = {'bs', 'sick_leave', 'annual_leave', 'dismissal'}
        for op_id, current_status in current_status_map.items():
            current_status_norm = str(current_status or 'working')
            active_schedule_status = active_status_by_operator.get(op_id)
            desired_status = None

            if active_schedule_status:
                desired_status = SCHEDULE_STATUS_TO_USER_STATUS.get(active_schedule_status, active_schedule_status)
            else:
                if current_status_norm in temporary_schedule_statuses:
                    desired_status = 'working'
                elif current_status_norm == 'fired' and op_id in has_dismissal_history:
                    desired_status = 'working'

            if not desired_status or desired_status == current_status_norm:
                continue

            cursor.execute("""
                UPDATE users
                SET status = %s
                WHERE id = %s
            """, (desired_status, op_id))
            cursor.execute("""
                INSERT INTO user_history (user_id, changed_by, field_changed, old_value, new_value)
                VALUES (%s, %s, %s, %s, %s)
            """, (op_id, None, 'status', current_status_norm, desired_status))
            updated_count += 1

        return updated_count

    def sync_user_statuses_from_schedule_periods(self, operator_ids=None, as_of_date=None):
        """
        Публичный wrapper для фонового/ручного запуска синхронизации users.status
        по периодным статусам графика.
        """
        with self._get_cursor() as cursor:
            return self._sync_user_statuses_from_schedule_periods_tx(
                cursor=cursor,
                operator_ids=operator_ids,
                as_of_date=as_of_date
            )

    def get_operators_with_shifts(
        self,
        start_date=None,
        end_date=None,
        direction_name=None,
        include_imported_statuses=False,
        include_technical_issues=False,
        include_offline_activities=False
    ):
        """
        Получить всех операторов со сменами и выходными днями за период.
        Если direction_name задан — только операторы этого направления.
        Возвращает список операторов с их сменами и выходными.
        """
        start_date_obj = self._normalize_schedule_date(start_date) if start_date else None
        end_date_obj = self._normalize_schedule_date(end_date) if end_date else None

        with self._get_cursor() as cursor:
            self._sync_user_statuses_from_schedule_periods_tx(cursor)
            if direction_name:
                cursor.execute("""
                    SELECT u.id, u.name, u.supervisor_id, s.name as supervisor_name,
                           d.name as direction, u.status, u.rate
                    FROM users u
                    LEFT JOIN users s ON u.supervisor_id = s.id
                    LEFT JOIN directions d ON u.direction_id = d.id
                    WHERE u.role = 'operator' AND LOWER(d.name) = LOWER(%s) AND (u.status IS NULL OR u.status != 'fired')
                    ORDER BY u.name
                """, (direction_name,))
            else:
                cursor.execute("""
                    SELECT u.id, u.name, u.supervisor_id, s.name as supervisor_name,
                           d.name as direction, u.status, u.rate
                    FROM users u
                    LEFT JOIN users s ON u.supervisor_id = s.id
                    LEFT JOIN directions d ON u.direction_id = d.id
                    WHERE u.role = 'operator'
                    ORDER BY d.name, u.name
                """)
            operators = cursor.fetchall()
            if not operators:
                return []

            operator_ids = [row[0] for row in operators]
            result_map = {}
            for op_id, name, supervisor_id, supervisor_name, direction, status, rate in operators:
                result_map[op_id] = {
                    'id': op_id,
                    'name': name,
                    'supervisor_id': supervisor_id,
                    'supervisor_name': supervisor_name,
                    'direction': direction,
                    'status': status,
                    'rate': float(rate) if rate else 1.0,
                    'shifts': {},
                    'daysOff': [],
                    'scheduleStatusPeriods': [],
                    'scheduleStatusDays': {},
                    'aggregatedScheduleFlagsDays': {},
                    'importedStatusTimelineDays': {},
                    'technicalIssueTimelineDays': {},
                    'offlineActivityTimelineDays': {}
                }

            shifts_query = """
                SELECT ws.id, ws.operator_id, ws.shift_date, ws.start_time, ws.end_time
                FROM work_shifts ws
                WHERE ws.operator_id = ANY(%s)
            """
            shifts_params = [operator_ids]
            if start_date_obj:
                shifts_query += " AND ws.shift_date >= %s"
                shifts_params.append(start_date_obj)
            if end_date_obj:
                shifts_query += " AND ws.shift_date <= %s"
                shifts_params.append(end_date_obj)
            shifts_query += " ORDER BY ws.operator_id, ws.shift_date, ws.start_time"
            cursor.execute(shifts_query, shifts_params)
            shifts_rows = cursor.fetchall()

            shift_ids = []
            shift_ref = {}
            for shift_id, operator_id, shift_date, start_time_value, end_time_value in shifts_rows:
                op_entry = result_map.get(operator_id)
                if not op_entry:
                    continue
                date_str = shift_date.strftime('%Y-%m-%d')
                shift_item = {
                    'id': shift_id,
                    'start': start_time_value.strftime('%H:%M'),
                    'end': end_time_value.strftime('%H:%M'),
                    'breaks': []
                }
                op_entry['shifts'].setdefault(date_str, []).append(shift_item)
                shift_ids.append(shift_id)
                shift_ref[shift_id] = shift_item

            if shift_ids:
                cursor.execute(
                    """
                    SELECT shift_id, start_minutes, end_minutes
                    FROM shift_breaks
                    WHERE shift_id = ANY(%s)
                    ORDER BY shift_id, start_minutes, end_minutes
                    """,
                    (shift_ids,)
                )
                for shift_id, br_start, br_end in cursor.fetchall():
                    item = shift_ref.get(shift_id)
                    if item is None:
                        continue
                    item['breaks'].append({'start': int(br_start), 'end': int(br_end)})

            # Keep backward-compatible read behavior: frontend expects already merged segments on load.
            for op_entry in result_map.values():
                for day_key, day_shifts in list(op_entry['shifts'].items()):
                    op_entry['shifts'][day_key] = _merge_shifts_for_date(day_shifts)

            days_off_query = """
                SELECT operator_id, day_off_date
                FROM days_off
                WHERE operator_id = ANY(%s)
            """
            days_off_params = [operator_ids]
            if start_date_obj:
                days_off_query += " AND day_off_date >= %s"
                days_off_params.append(start_date_obj)
            if end_date_obj:
                days_off_query += " AND day_off_date <= %s"
                days_off_params.append(end_date_obj)
            days_off_query += " ORDER BY operator_id, day_off_date"
            cursor.execute(days_off_query, days_off_params)
            for operator_id, day_off_date in cursor.fetchall():
                op_entry = result_map.get(operator_id)
                if op_entry is None:
                    continue
                op_entry['daysOff'].append(day_off_date.strftime('%Y-%m-%d'))

            self._attach_schedule_status_periods_to_operators(
                cursor=cursor,
                operators_map=result_map,
                operator_ids=operator_ids,
                start_date_obj=start_date_obj,
                end_date_obj=end_date_obj
            )
            self._attach_auto_schedule_flags_to_operators(
                cursor=cursor,
                operators_map=result_map,
                operator_ids=operator_ids,
                start_date_obj=start_date_obj,
                end_date_obj=end_date_obj
            )
            if include_imported_statuses:
                self._attach_imported_status_segments_to_operators(
                    cursor=cursor,
                    operators_map=result_map,
                    operator_ids=operator_ids,
                    start_date_obj=start_date_obj,
                    end_date_obj=end_date_obj
                )
            if include_technical_issues:
                self._attach_technical_issue_segments_to_operators(
                    cursor=cursor,
                    operators_map=result_map,
                    operator_ids=operator_ids,
                    start_date_obj=start_date_obj,
                    end_date_obj=end_date_obj
                )
            if include_offline_activities:
                self._attach_offline_activity_segments_to_operators(
                    cursor=cursor,
                    operators_map=result_map,
                    operator_ids=operator_ids,
                    start_date_obj=start_date_obj,
                    end_date_obj=end_date_obj
                )

            return [result_map[row[0]] for row in operators]

    def get_operator_with_shifts(
        self,
        operator_id,
        start_date=None,
        end_date=None,
        include_imported_statuses=False,
        include_technical_issues=False,
        include_offline_activities=False
    ):
        """
        Получить одного оператора с его сменами/перерывами/выходными за период.
        Возвращает структуру, совместимую с get_operators_with_shifts()[i].
        """
        operator_id = int(operator_id)
        start_date_obj = self._normalize_schedule_date(start_date) if start_date else None
        end_date_obj = self._normalize_schedule_date(end_date) if end_date else None

        with self._get_cursor() as cursor:
            self._sync_user_statuses_from_schedule_periods_tx(cursor, operator_ids=[operator_id])
            cursor.execute("""
                SELECT u.id, u.name, u.supervisor_id, s.name as supervisor_name,
                       d.name as direction, u.status, u.rate
                FROM users u
                LEFT JOIN users s ON u.supervisor_id = s.id
                LEFT JOIN directions d ON u.direction_id = d.id
                WHERE u.id = %s AND u.role = 'operator'
                LIMIT 1
            """, (operator_id,))
            row = cursor.fetchone()
            if not row:
                return None

            op_id, name, supervisor_id, supervisor_name, direction, status, rate = row
            result = {
                'id': op_id,
                'name': name,
                'supervisor_id': supervisor_id,
                'supervisor_name': supervisor_name,
                'direction': direction,
                'status': status,
                'rate': float(rate) if rate else 1.0,
                'shifts': {},
                'daysOff': [],
                'scheduleStatusPeriods': [],
                'scheduleStatusDays': {},
                'aggregatedScheduleFlagsDays': {},
                'importedStatusTimelineDays': {},
                'technicalIssueTimelineDays': {},
                'offlineActivityTimelineDays': {}
            }

            shifts_query = """
                SELECT ws.id, ws.shift_date, ws.start_time, ws.end_time
                FROM work_shifts ws
                WHERE ws.operator_id = %s
            """
            shifts_params = [operator_id]
            if start_date_obj:
                shifts_query += " AND ws.shift_date >= %s"
                shifts_params.append(start_date_obj)
            if end_date_obj:
                shifts_query += " AND ws.shift_date <= %s"
                shifts_params.append(end_date_obj)
            shifts_query += " ORDER BY ws.shift_date, ws.start_time"
            cursor.execute(shifts_query, shifts_params)
            shifts_rows = cursor.fetchall()

            shift_ids = []
            shift_ref = {}
            for shift_id, shift_date, start_time_value, end_time_value in shifts_rows:
                date_str = shift_date.strftime('%Y-%m-%d')
                shift_item = {
                    'id': shift_id,
                    'start': start_time_value.strftime('%H:%M'),
                    'end': end_time_value.strftime('%H:%M'),
                    'breaks': []
                }
                result['shifts'].setdefault(date_str, []).append(shift_item)
                shift_ids.append(shift_id)
                shift_ref[shift_id] = shift_item

            if shift_ids:
                cursor.execute(
                    """
                    SELECT shift_id, start_minutes, end_minutes
                    FROM shift_breaks
                    WHERE shift_id = ANY(%s)
                    ORDER BY shift_id, start_minutes, end_minutes
                    """,
                    (shift_ids,)
                )
                for shift_id, br_start, br_end in cursor.fetchall():
                    item = shift_ref.get(shift_id)
                    if item is None:
                        continue
                    item['breaks'].append({'start': int(br_start), 'end': int(br_end)})

            for day_key, day_shifts in list(result['shifts'].items()):
                result['shifts'][day_key] = _merge_shifts_for_date(day_shifts)

            days_off_query = """
                SELECT day_off_date
                FROM days_off
                WHERE operator_id = %s
            """
            days_off_params = [operator_id]
            if start_date_obj:
                days_off_query += " AND day_off_date >= %s"
                days_off_params.append(start_date_obj)
            if end_date_obj:
                days_off_query += " AND day_off_date <= %s"
                days_off_params.append(end_date_obj)
            days_off_query += " ORDER BY day_off_date"
            cursor.execute(days_off_query, days_off_params)
            for (day_off_date,) in cursor.fetchall():
                result['daysOff'].append(day_off_date.strftime('%Y-%m-%d'))

            self._attach_schedule_status_periods_to_operators(
                cursor=cursor,
                operators_map={operator_id: result},
                operator_ids=[operator_id],
                start_date_obj=start_date_obj,
                end_date_obj=end_date_obj
            )
            self._attach_auto_schedule_flags_to_operators(
                cursor=cursor,
                operators_map={operator_id: result},
                operator_ids=[operator_id],
                start_date_obj=start_date_obj,
                end_date_obj=end_date_obj
            )
            if include_imported_statuses:
                self._attach_imported_status_segments_to_operators(
                    cursor=cursor,
                    operators_map={operator_id: result},
                    operator_ids=[operator_id],
                    start_date_obj=start_date_obj,
                    end_date_obj=end_date_obj
                )
            if include_technical_issues:
                self._attach_technical_issue_segments_to_operators(
                    cursor=cursor,
                    operators_map={operator_id: result},
                    operator_ids=[operator_id],
                    start_date_obj=start_date_obj,
                    end_date_obj=end_date_obj
                )
            if include_offline_activities:
                self._attach_offline_activity_segments_to_operators(
                    cursor=cursor,
                    operators_map={operator_id: result},
                    operator_ids=[operator_id],
                    start_date_obj=start_date_obj,
                    end_date_obj=end_date_obj
                )

            return result

    def _iter_schedule_dates_inclusive(self, start_date_obj, end_date_obj):
        current = start_date_obj
        while current <= end_date_obj:
            yield current
            current = current + timedelta(days=1)

    def _load_operator_shift_map_for_period_tx(self, cursor, operator_id, start_date_obj, end_date_obj):
        operator_id = int(operator_id)
        start_date_obj = self._normalize_schedule_date(start_date_obj)
        end_date_obj = self._normalize_schedule_date(end_date_obj)
        if end_date_obj < start_date_obj:
            raise ValueError("end_date must be >= start_date")

        cursor.execute(
            """
            SELECT ws.id, ws.shift_date, ws.start_time, ws.end_time
            FROM work_shifts ws
            WHERE ws.operator_id = %s
              AND ws.shift_date >= %s
              AND ws.shift_date <= %s
            ORDER BY ws.shift_date, ws.start_time
            """,
            (operator_id, start_date_obj, end_date_obj)
        )
        rows = cursor.fetchall() or []

        shifts_by_date = {}
        shift_ref = {}
        shift_ids = []
        for shift_id, shift_date, start_time_value, end_time_value in rows:
            day_key = shift_date.strftime('%Y-%m-%d')
            item = {
                'start': start_time_value.strftime('%H:%M'),
                'end': end_time_value.strftime('%H:%M'),
                'breaks': []
            }
            shifts_by_date.setdefault(day_key, []).append(item)
            shift_ref[int(shift_id)] = item
            shift_ids.append(int(shift_id))

        if shift_ids:
            cursor.execute(
                """
                SELECT shift_id, start_minutes, end_minutes
                FROM shift_breaks
                WHERE shift_id = ANY(%s)
                ORDER BY shift_id, start_minutes, end_minutes
                """,
                (shift_ids,)
            )
            for shift_id, br_start, br_end in cursor.fetchall() or []:
                target = shift_ref.get(int(shift_id))
                if target is None:
                    continue
                target['breaks'].append({
                    'start': int(br_start),
                    'end': int(br_end)
                })

        normalized = {}
        for day_key, segs in shifts_by_date.items():
            merged = _merge_shifts_for_date(segs or [])
            if merged:
                normalized[day_key] = merged
        return dict(sorted(normalized.items(), key=lambda x: x[0]))

    def _normalize_shift_snapshot_map(self, raw_snapshot):
        if raw_snapshot is None:
            return {}

        source = raw_snapshot
        if isinstance(source, str):
            text = source.strip()
            if not text:
                return {}
            try:
                source = json.loads(text)
            except Exception as exc:
                raise ValueError("Invalid requested_shifts_json format") from exc

        if not isinstance(source, dict):
            raise ValueError("requested_shifts_json must be an object")

        normalized = {}
        for raw_day, raw_shifts in source.items():
            try:
                day_obj = self._normalize_schedule_date(raw_day)
            except Exception as exc:
                raise ValueError("Invalid date in requested_shifts_json") from exc
            day_key = day_obj.strftime('%Y-%m-%d')

            if not isinstance(raw_shifts, list):
                continue

            parsed_shifts = []
            for seg in raw_shifts:
                if not isinstance(seg, dict):
                    continue
                start_raw = seg.get('start') or seg.get('start_time')
                end_raw = seg.get('end') or seg.get('end_time')
                if not start_raw or not end_raw:
                    continue
                start_obj = self._normalize_schedule_time(start_raw, 'start_time')
                end_obj = self._normalize_schedule_time(end_raw, 'end_time')
                breaks_raw = seg.get('breaks') if isinstance(seg.get('breaks'), list) else []
                breaks_norm = self._normalize_shift_breaks(breaks_raw)
                parsed_shifts.append({
                    'start': start_obj.strftime('%H:%M'),
                    'end': end_obj.strftime('%H:%M'),
                    'breaks': breaks_norm
                })

            merged = _merge_shifts_for_date(parsed_shifts)
            if merged:
                normalized[day_key] = merged

        return dict(sorted(normalized.items(), key=lambda x: x[0]))

    def _snapshot_shift_totals(self, snapshot):
        days_count = 0
        shifts_count = 0
        total_minutes = 0
        for day_key, shifts in (snapshot or {}).items():
            if not isinstance(shifts, list) or not shifts:
                continue
            days_count += 1
            for seg in shifts:
                start_val = seg.get('start')
                end_val = seg.get('end')
                if not start_val or not end_val:
                    continue
                start_min, end_min = self._schedule_interval_minutes(start_val, end_val)
                shifts_count += 1
                total_minutes += max(0, int(end_min) - int(start_min))
        return {
            'daysCount': int(days_count),
            'shiftsCount': int(shifts_count),
            'totalMinutes': int(total_minutes)
        }

    def _normalize_swap_time_range(self, start_time, end_time):
        start_obj = self._normalize_schedule_time(start_time, 'start_time')
        end_obj = self._normalize_schedule_time(end_time, 'end_time')
        start_str = start_obj.strftime('%H:%M')
        end_str = end_obj.strftime('%H:%M')
        start_min = _time_to_minutes(start_str)
        end_min = _time_to_minutes(end_str)
        if end_min == start_min:
            raise ValueError("end_time must be different from start_time")
        if end_min < start_min:
            # Интервал через полночь, например 22:00 -> 02:00.
            end_min += 24 * 60
        if (end_min - start_min) > 24 * 60:
            raise ValueError("swap interval cannot be longer than 24 hours")
        return {
            'startTime': start_str,
            'endTime': end_str,
            'startMin': int(start_min),
            'endMin': int(end_min)
        }

    def _normalize_swap_time_range_for_date(self, swap_date, start_time, end_time, end_date=None):
        """
        Нормализует интервал запроса замены относительно стартовой даты swap_date.
        Возвращает start/end в минутах в окне [0..2880) и итоговую дату окончания.
        """
        swap_date_obj = self._normalize_schedule_date(swap_date)
        base_interval = self._normalize_swap_time_range(start_time, end_time)
        start_raw_min = _time_to_minutes(base_interval['startTime'])
        end_raw_min = _time_to_minutes(base_interval['endTime'])
        start_raw_min = int(start_raw_min)
        end_raw_min = int(end_raw_min)

        def _build_valid_candidate(candidate_start, candidate_end):
            c_start = int(candidate_start)
            c_end = int(candidate_end)
            if c_end <= c_start:
                return None
            duration = c_end - c_start
            if duration <= 0 or duration > (24 * 60):
                return None
            if c_start < 0 or c_end > (2 * 24 * 60):
                return None
            return {'start': c_start, 'end': c_end, 'duration': duration}

        if end_date is None or str(end_date).strip() == '':
            resolved_end_date_obj = swap_date_obj + timedelta(days=1) if int(base_interval['endMin']) > 1440 else swap_date_obj
            return {
                'startTime': base_interval['startTime'],
                'endTime': base_interval['endTime'],
                'startMin': int(base_interval['startMin']),
                'endMin': int(base_interval['endMin']),
                'swapDateObj': swap_date_obj,
                'endDateObj': resolved_end_date_obj
            }

        end_date_obj = self._normalize_schedule_date(end_date)
        day_diff = int((end_date_obj - swap_date_obj).days)
        if day_diff < 0:
            raise ValueError("end_date must be >= swap_date")
        if day_diff > 1:
            raise ValueError("end_date can be only swap_date or next day")

        candidates = []
        # Вариант A: интервал начинается в swap_date.
        anchored_end = day_diff * 1440 + end_raw_min
        if day_diff == 0 and end_raw_min < start_raw_min:
            anchored_end += 1440
        candidate_a = _build_valid_candidate(start_raw_min, anchored_end)
        if candidate_a:
            candidate_a['kind'] = 'anchored'
            candidates.append(candidate_a)

        # Вариант B: при end_date=next day считаем, что интервал может быть полностью на следующем дне.
        if day_diff > 0:
            shifted_start = day_diff * 1440 + start_raw_min
            shifted_end = day_diff * 1440 + end_raw_min
            if shifted_end <= shifted_start:
                shifted_end += 1440
            candidate_b = _build_valid_candidate(shifted_start, shifted_end)
            if candidate_b:
                candidate_b['kind'] = 'shifted'
                candidates.append(candidate_b)

        if not candidates:
            raise ValueError("Invalid swap interval for selected dates")

        # Предпочитаем минимальную длительность; при равенстве — anchored.
        candidates.sort(key=lambda x: (int(x['duration']), 0 if x.get('kind') == 'anchored' else 1))
        chosen = candidates[0]
        start_min = int(chosen['start'])
        end_min = int(chosen['end'])

        end_day_offset = int(end_min // 1440)
        resolved_end_date_obj = swap_date_obj + timedelta(days=end_day_offset)

        return {
            'startTime': base_interval['startTime'],
            'endTime': base_interval['endTime'],
            'startMin': int(start_min),
            'endMin': int(end_min),
            'swapDateObj': swap_date_obj,
            'endDateObj': resolved_end_date_obj
        }

    def _normalize_swap_request_payload(self, raw_payload, fallback_date=None):
        payload = raw_payload
        if payload is None:
            payload = {}
        if isinstance(payload, str):
            text = payload.strip()
            if not text:
                payload = {}
            else:
                try:
                    payload = json.loads(text)
                except Exception as exc:
                    raise ValueError("Invalid swap payload format") from exc
        if not isinstance(payload, dict):
            payload = {}

        swap_date = payload.get('swapDate') or payload.get('swap_date') or fallback_date
        if swap_date:
            try:
                swap_date = self._normalize_schedule_date(swap_date).strftime('%Y-%m-%d')
            except Exception:
                swap_date = None
        end_date = payload.get('endDate') or payload.get('end_date')
        if end_date:
            try:
                end_date = self._normalize_schedule_date(end_date).strftime('%Y-%m-%d')
            except Exception:
                end_date = None

        interval_raw = payload.get('interval') if isinstance(payload.get('interval'), dict) else {}
        start_time = interval_raw.get('start') or interval_raw.get('start_time')
        end_time = interval_raw.get('end') or interval_raw.get('end_time')
        start_min = interval_raw.get('startMin')
        end_min = interval_raw.get('endMin')

        if start_time and end_time:
            try:
                if swap_date:
                    normalized_interval = self._normalize_swap_time_range_for_date(
                        swap_date=swap_date,
                        start_time=start_time,
                        end_time=end_time,
                        end_date=end_date
                    )
                    start_time = normalized_interval['startTime']
                    end_time = normalized_interval['endTime']
                    start_min = normalized_interval['startMin']
                    end_min = normalized_interval['endMin']
                    end_date_obj = normalized_interval.get('endDateObj')
                    if isinstance(end_date_obj, date):
                        end_date = end_date_obj.strftime('%Y-%m-%d')
                else:
                    normalized_interval = self._normalize_swap_time_range(start_time, end_time)
                    start_time = normalized_interval['startTime']
                    end_time = normalized_interval['endTime']
                    start_min = normalized_interval['startMin']
                    end_min = normalized_interval['endMin']
            except Exception:
                start_time = None
                end_time = None
                start_min = None
                end_min = None

        if start_min is not None and end_min is not None:
            try:
                start_min = int(start_min)
                end_min = int(end_min)
            except Exception:
                start_min = None
                end_min = None
            if start_min is not None and end_min is not None:
                if end_min == start_min:
                    start_min = None
                    end_min = None
                elif end_min < start_min:
                    end_min += 24 * 60
                if start_min is not None and end_min is not None and (end_min - start_min) > (24 * 60):
                    start_min = None
                    end_min = None
                if swap_date and start_min is not None and end_min is not None:
                    try:
                        base_date_obj = self._normalize_schedule_date(swap_date)
                        end_offset_days = int(end_min // (24 * 60))
                        end_date = (base_date_obj + timedelta(days=end_offset_days)).strftime('%Y-%m-%d')
                    except Exception:
                        pass

        requested_segments_raw = payload.get('requestedSegments')
        if not isinstance(requested_segments_raw, list):
            requested_segments_raw = payload.get('requested_segments')
        requested_segments_intervals = self._swap_parse_payload_segments(
            requested_segments_raw,
            swap_date=swap_date
        )
        requested_segments = self._swap_serialize_intervals_with_day_offset(requested_segments_intervals)

        target_segments_raw = payload.get('targetSegments')
        if not isinstance(target_segments_raw, list):
            target_segments_raw = payload.get('target_segments')
        target_segments_intervals = self._swap_parse_payload_segments(
            target_segments_raw,
            swap_date=swap_date
        )
        target_segments = self._swap_serialize_intervals_with_day_offset(target_segments_intervals)

        mode_raw = payload.get('mode')
        mode = str(mode_raw).strip() if mode_raw is not None else ''
        if not mode:
            mode = 'shift_exchange_v2' if target_segments else 'shift_replacement_v1'

        return {
            'mode': mode,
            'swapDate': swap_date,
            'endDate': end_date,
            'startTime': start_time,
            'endTime': end_time,
            'startMin': start_min if start_min is not None else None,
            'endMin': end_min if end_min is not None else None,
            'requestedSegments': requested_segments,
            'targetSegments': target_segments
        }

    def _swap_parse_payload_segments(self, raw_segments, swap_date=None):
        if not isinstance(raw_segments, list):
            return []

        swap_date_obj = None
        if swap_date:
            try:
                swap_date_obj = self._normalize_schedule_date(swap_date)
            except Exception:
                swap_date_obj = None

        intervals = []
        for seg in raw_segments:
            if not isinstance(seg, dict):
                continue

            seg_start = seg.get('start') or seg.get('start_time')
            seg_end = seg.get('end') or seg.get('end_time')
            if not seg_start or not seg_end:
                continue
            try:
                seg_start_obj = self._normalize_schedule_time(seg_start, 'segment_start')
                seg_end_obj = self._normalize_schedule_time(seg_end, 'segment_end')
            except Exception:
                continue

            day_offset_raw = seg.get('dayOffset')
            if day_offset_raw is None:
                day_offset_raw = seg.get('day_offset')
            if day_offset_raw is None and swap_date_obj is not None:
                seg_date_raw = seg.get('date') or seg.get('shiftDate') or seg.get('shift_date')
                if seg_date_raw:
                    try:
                        seg_date_obj = self._normalize_schedule_date(seg_date_raw)
                        day_offset_raw = int((seg_date_obj - swap_date_obj).days)
                    except Exception:
                        day_offset_raw = 0

            try:
                day_offset = int(day_offset_raw) if day_offset_raw is not None else 0
            except Exception:
                day_offset = 0
            if day_offset < -1 or day_offset > 1:
                continue

            seg_start_min = _time_to_minutes(seg_start_obj.strftime('%H:%M')) + (day_offset * 1440)
            seg_end_min = _time_to_minutes(seg_end_obj.strftime('%H:%M')) + (day_offset * 1440)
            if seg_end_min <= seg_start_min:
                seg_end_min += 1440

            duration = int(seg_end_min) - int(seg_start_min)
            if duration <= 0 or duration > (24 * 60):
                continue
            if int(seg_start_min) < -1440 or int(seg_end_min) > 2880:
                continue

            intervals.append({
                'start': int(seg_start_min),
                'end': int(seg_end_min)
            })

        return self._merge_break_intervals(intervals)

    def _swap_serialize_intervals_with_day_offset(self, intervals):
        result = []
        for seg in self._merge_break_intervals(intervals or []):
            seg_start = int(seg.get('start', 0))
            seg_end = int(seg.get('end', 0))
            if seg_end <= seg_start:
                continue

            cursor = seg_start
            while cursor < seg_end:
                day_offset = int(cursor // 1440)
                day_start_abs = day_offset * 1440
                day_end_abs = day_start_abs + 1440
                chunk_end = min(seg_end, day_end_abs)
                if chunk_end <= cursor:
                    break

                local_start = int(cursor - day_start_abs)
                local_end = int(chunk_end - day_start_abs)
                if local_end > local_start:
                    result.append({
                        'start': _minutes_to_time(local_start),
                        'end': _minutes_to_time(local_end),
                        'dayOffset': int(day_offset),
                        'breaks': []
                    })
                cursor = chunk_end
        return result

    def _swap_payload_segments_to_day_map(self, segments, swap_date):
        if not swap_date:
            return {}
        try:
            swap_date_obj = self._normalize_schedule_date(swap_date)
        except Exception:
            return {}
        intervals = self._swap_parse_payload_segments(segments, swap_date=swap_date_obj.strftime('%Y-%m-%d'))
        if not intervals:
            return {}
        return self._swap_serialize_intervals_to_day_map(
            intervals,
            swap_date_obj,
            allowed_day_offsets=[-1, 0, 1]
        )

    def _build_swap_window_intervals(self, prev_day_shifts=None, day_shifts=None, next_day_shifts=None):
        intervals = []

        def append_shifts(shifts, day_offset_minutes):
            for seg in (shifts or []):
                if not isinstance(seg, dict):
                    continue
                start_val = seg.get('start')
                end_val = seg.get('end')
                if not start_val or not end_val:
                    continue
                seg_start, seg_end = self._schedule_interval_minutes(start_val, end_val)
                seg_start = int(seg_start) + int(day_offset_minutes)
                seg_end = int(seg_end) + int(day_offset_minutes)
                if seg_end <= seg_start:
                    continue
                intervals.append({'start': seg_start, 'end': seg_end})

        append_shifts(prev_day_shifts, -1440)
        append_shifts(day_shifts, 0)
        append_shifts(next_day_shifts, 1440)
        return self._merge_break_intervals(intervals)

    def _swap_extract_interval(self, intervals, interval_start_min, interval_end_min):
        interval_start_min = int(interval_start_min)
        interval_end_min = int(interval_end_min)
        if interval_end_min <= interval_start_min:
            return []

        overlaps = []
        for seg in (intervals or []):
            seg_start = int(seg.get('start', 0))
            seg_end = int(seg.get('end', 0))
            if seg_end <= seg_start:
                continue
            overlap_start = max(seg_start, interval_start_min)
            overlap_end = min(seg_end, interval_end_min)
            if overlap_end <= overlap_start:
                continue
            overlaps.append({'start': overlap_start, 'end': overlap_end})
        return self._merge_break_intervals(overlaps)

    def _swap_intervals_overlap(self, intervals, interval_start_min, interval_end_min):
        return bool(self._swap_extract_interval(intervals, interval_start_min, interval_end_min))

    def _swap_is_interval_fully_covered(self, intervals, interval_start_min, interval_end_min):
        interval_start_min = int(interval_start_min)
        interval_end_min = int(interval_end_min)
        if interval_end_min <= interval_start_min:
            return False
        extracted = self._swap_extract_interval(intervals, interval_start_min, interval_end_min)
        if not extracted:
            return False
        covered_until = interval_start_min
        for seg in extracted:
            seg_start = int(seg.get('start', 0))
            seg_end = int(seg.get('end', 0))
            if seg_end <= covered_until:
                continue
            if seg_start > covered_until:
                return False
            covered_until = max(covered_until, seg_end)
            if covered_until >= interval_end_min:
                return True
        return covered_until >= interval_end_min

    def _swap_subtract_interval(self, intervals, interval_start_min, interval_end_min):
        interval_start_min = int(interval_start_min)
        interval_end_min = int(interval_end_min)
        if interval_end_min <= interval_start_min:
            return self._merge_break_intervals(intervals or [])

        result = []
        for seg in (intervals or []):
            seg_start = int(seg.get('start', 0))
            seg_end = int(seg.get('end', 0))
            if seg_end <= seg_start:
                continue
            if seg_end <= interval_start_min or seg_start >= interval_end_min:
                result.append({'start': seg_start, 'end': seg_end})
                continue
            if seg_start < interval_start_min:
                result.append({'start': seg_start, 'end': interval_start_min})
            if seg_end > interval_end_min:
                result.append({'start': interval_end_min, 'end': seg_end})
        return self._merge_break_intervals(result)

    def _swap_subtract_intervals(self, intervals, subtracted_intervals):
        result = self._merge_break_intervals(intervals or [])
        for seg in self._merge_break_intervals(subtracted_intervals or []):
            result = self._swap_subtract_interval(
                result,
                int(seg.get('start', 0)),
                int(seg.get('end', 0))
            )
            if not result:
                return []
        return self._merge_break_intervals(result)

    def _swap_interval_sets_overlap(self, left_intervals, right_intervals):
        left = self._merge_break_intervals(left_intervals or [])
        right = self._merge_break_intervals(right_intervals or [])
        if not left or not right:
            return False
        for left_seg in left:
            l_start = int(left_seg.get('start', 0))
            l_end = int(left_seg.get('end', 0))
            if l_end <= l_start:
                continue
            for right_seg in right:
                r_start = int(right_seg.get('start', 0))
                r_end = int(right_seg.get('end', 0))
                if r_end <= r_start:
                    continue
                if l_start < r_end and r_start < l_end:
                    return True
        return False

    def _swap_serialize_intervals_for_payload(self, intervals):
        return self._swap_serialize_intervals_with_day_offset(intervals)

    def _swap_intervals_signature(self, segments):
        sign = []
        for seg in (segments or []):
            if not isinstance(seg, dict):
                continue
            start_val = seg.get('start')
            end_val = seg.get('end')
            if not start_val or not end_val:
                continue
            day_offset_raw = seg.get('dayOffset')
            if day_offset_raw is None:
                day_offset_raw = seg.get('day_offset')
            try:
                day_offset = int(day_offset_raw) if day_offset_raw is not None else 0
            except Exception:
                day_offset = 0
            sign.append(f"{day_offset}|{start_val}|{end_val}")
        sign.sort()
        return sign

    def _swap_serialize_intervals_to_day_map(self, intervals, swap_date_obj, allowed_day_offsets):
        if not isinstance(allowed_day_offsets, (list, tuple, set)):
            return {}
        allowed_offsets = {int(v) for v in allowed_day_offsets}
        if not allowed_offsets:
            return {}

        day_map = {}
        for seg in self._merge_break_intervals(intervals or []):
            seg_start = int(seg.get('start', 0))
            seg_end = int(seg.get('end', 0))
            if seg_end <= seg_start:
                continue

            cursor = seg_start
            while cursor < seg_end:
                day_offset = int(cursor // 1440)
                chunk_end = min(seg_end, cursor + 1440)
                if chunk_end <= cursor:
                    break
                if day_offset in allowed_offsets:
                    day_start_abs = day_offset * 1440
                    local_start = int(cursor - day_start_abs)
                    local_end = int(chunk_end - day_start_abs)
                    if local_end > local_start:
                        day_key = (swap_date_obj + timedelta(days=day_offset)).strftime('%Y-%m-%d')
                        day_map.setdefault(day_key, []).append({
                            'start': _minutes_to_time(local_start),
                            'end': _minutes_to_time(local_end)
                        })
                cursor = chunk_end

        normalized = {}
        for day_key, shifts in day_map.items():
            merged = _merge_shifts_for_date(shifts or [])
            if merged:
                normalized[day_key] = merged
        return normalized

    def _extract_day_interval_segments_from_shifts(self, day_shifts, interval_start_min, interval_end_min):
        interval_start_min = int(interval_start_min)
        interval_end_min = int(interval_end_min)
        if interval_end_min <= interval_start_min:
            return []

        result = []
        for seg in (day_shifts or []):
            start_val = seg.get('start')
            end_val = seg.get('end')
            if not start_val or not end_val:
                continue
            seg_start, seg_end = self._schedule_interval_minutes(start_val, end_val)
            seg_start = max(0, min(1440, int(seg_start)))
            seg_end = max(0, min(1440, int(seg_end)))
            if seg_end <= seg_start:
                continue

            overlap_start = max(seg_start, interval_start_min)
            overlap_end = min(seg_end, interval_end_min)
            if overlap_end <= overlap_start:
                continue

            raw_breaks = self._normalize_shift_breaks(seg.get('breaks') if isinstance(seg.get('breaks'), list) else [])
            clipped_breaks = []
            for b in raw_breaks:
                b_start = max(overlap_start, int(b.get('start', 0)))
                b_end = min(overlap_end, int(b.get('end', 0)))
                if b_end > b_start:
                    clipped_breaks.append({'start': b_start, 'end': b_end})

            result.append({
                'start': _minutes_to_time(overlap_start),
                'end': _minutes_to_time(overlap_end),
                'breaks': clipped_breaks
            })

        if not result:
            return []
        return _merge_shifts_for_date(result)

    def _is_interval_fully_covered_by_day_shifts(self, day_shifts, interval_start_min, interval_end_min):
        extracted = self._extract_day_interval_segments_from_shifts(day_shifts, interval_start_min, interval_end_min)
        if not extracted:
            return False
        intervals = []
        for seg in extracted:
            s, e = self._schedule_interval_minutes(seg.get('start'), seg.get('end'))
            s = max(0, min(1440, int(s)))
            e = max(0, min(1440, int(e)))
            if e > s:
                intervals.append({'start': s, 'end': e})
        merged = self._merge_break_intervals(intervals)
        if len(merged) != 1:
            return False
        return int(merged[0]['start']) <= int(interval_start_min) and int(merged[0]['end']) >= int(interval_end_min)

    def _day_shifts_overlap_interval(self, day_shifts, interval_start_min, interval_end_min):
        interval_start_min = int(interval_start_min)
        interval_end_min = int(interval_end_min)
        if interval_end_min <= interval_start_min:
            return False

        for seg in (day_shifts or []):
            start_val = seg.get('start')
            end_val = seg.get('end')
            if not start_val or not end_val:
                continue
            seg_start, seg_end = self._schedule_interval_minutes(start_val, end_val)
            seg_start = max(0, min(1440, int(seg_start)))
            seg_end = max(0, min(1440, int(seg_end)))
            if seg_end <= seg_start:
                continue
            if seg_start < interval_end_min and interval_start_min < seg_end:
                return True
        return False

    def _subtract_interval_from_day_shifts(self, day_shifts, interval_start_min, interval_end_min):
        interval_start_min = int(interval_start_min)
        interval_end_min = int(interval_end_min)
        if interval_end_min <= interval_start_min:
            return _merge_shifts_for_date(day_shifts or [])

        result = []
        for seg in (day_shifts or []):
            start_val = seg.get('start')
            end_val = seg.get('end')
            if not start_val or not end_val:
                continue
            seg_start, seg_end = self._schedule_interval_minutes(start_val, end_val)
            seg_start = int(seg_start)
            seg_end = int(seg_end)
            if seg_end <= seg_start:
                continue

            # Для overlap-проверки учитываем только часть, принадлежащую выбранной дате (00:00-24:00).
            seg_day_start = max(0, min(1440, seg_start))
            seg_day_end = max(0, min(1440, seg_end))

            if seg_day_end <= interval_start_min or seg_day_start >= interval_end_min:
                result.append({
                    'start': _minutes_to_time(seg_start),
                    'end': _minutes_to_time(seg_end),
                    'breaks': self._normalize_shift_breaks(seg.get('breaks') if isinstance(seg.get('breaks'), list) else [])
                })
                continue

            left_start = seg_start
            left_end = min(seg_end, interval_start_min)
            right_start = max(seg_start, interval_end_min)
            right_end = seg_end

            raw_breaks = self._normalize_shift_breaks(seg.get('breaks') if isinstance(seg.get('breaks'), list) else [])

            def append_piece(piece_start, piece_end):
                if piece_end <= piece_start:
                    return
                piece_breaks = []
                for b in raw_breaks:
                    b_start = max(piece_start, int(b.get('start', 0)))
                    b_end = min(piece_end, int(b.get('end', 0)))
                    if b_end > b_start:
                        piece_breaks.append({'start': b_start, 'end': b_end})
                result.append({
                    'start': _minutes_to_time(piece_start),
                    'end': _minutes_to_time(piece_end),
                    'breaks': piece_breaks
                })

            append_piece(left_start, left_end)
            append_piece(right_start, right_end)

        if not result:
            return []
        return _merge_shifts_for_date(result)

    def _serialize_shift_swap_request_row(self, row):
        if not row:
            return None
        (
            request_id,
            requester_id,
            requester_name,
            target_id,
            target_name,
            direction_id,
            direction_name,
            start_date,
            end_date,
            status,
            request_comment,
            response_comment,
            requested_shifts_json,
            created_at,
            updated_at,
            responded_at,
            accepted_at,
            responded_by,
            responded_by_name
        ) = row

        payload = self._normalize_swap_request_payload(
            requested_shifts_json,
            fallback_date=start_date.strftime('%Y-%m-%d') if isinstance(start_date, date) else None
        )
        requested_segments = payload.get('requestedSegments') or []
        target_segments = payload.get('targetSegments') or []
        swap_date_value = payload.get('swapDate') or (start_date.strftime('%Y-%m-%d') if isinstance(start_date, date) else '')
        requested_day_map = self._swap_payload_segments_to_day_map(requested_segments, swap_date_value)
        target_day_map = self._swap_payload_segments_to_day_map(target_segments, swap_date_value)
        requested_totals = self._snapshot_shift_totals(requested_day_map or {})
        target_totals = self._snapshot_shift_totals(target_day_map or {})

        def _fmt_dt(value):
            if isinstance(value, datetime):
                return value.isoformat()
            return None

        exchange_mode = str(payload.get('mode') or '').strip()
        if not exchange_mode:
            exchange_mode = 'shift_exchange_v2' if target_segments else 'shift_replacement_v1'

        return {
            'id': int(request_id),
            'requester': {
                'id': int(requester_id),
                'name': requester_name
            },
            'target': {
                'id': int(target_id),
                'name': target_name
            },
            'direction': {
                'id': int(direction_id) if direction_id is not None else None,
                'name': direction_name
            },
            'swapDate': payload.get('swapDate') or (start_date.strftime('%Y-%m-%d') if isinstance(start_date, date) else None),
            'startTime': payload.get('startTime'),
            'endTime': payload.get('endTime'),
            'startDate': start_date.strftime('%Y-%m-%d') if isinstance(start_date, date) else str(start_date),
            'endDate': end_date.strftime('%Y-%m-%d') if isinstance(end_date, date) else str(end_date),
            'status': str(status),
            'requestComment': request_comment,
            'responseComment': response_comment,
            'requestedShifts': requested_day_map,
            'requestedSegments': requested_segments,
            'targetShifts': target_day_map,
            'targetSegments': target_segments,
            'summary': requested_totals,
            'exchangeSummary': target_totals,
            'exchangeMode': exchange_mode,
            'createdAt': _fmt_dt(created_at),
            'updatedAt': _fmt_dt(updated_at),
            'respondedAt': _fmt_dt(responded_at),
            'acceptedAt': _fmt_dt(accepted_at),
            'respondedBy': {
                'id': int(responded_by) if responded_by is not None else None,
                'name': responded_by_name
            }
        }

    def _select_shift_swap_request_by_id_tx(self, cursor, request_id):
        cursor.execute(
            """
            SELECT
                r.id,
                r.requester_operator_id, req.name,
                r.target_operator_id, tgt.name,
                r.direction_id, d.name,
                r.start_date, r.end_date,
                r.status,
                r.request_comment,
                r.response_comment,
                r.requested_shifts_json,
                r.created_at,
                r.updated_at,
                r.responded_at,
                r.accepted_at,
                r.responded_by,
                resp.name
            FROM work_shift_swap_requests r
            JOIN users req ON req.id = r.requester_operator_id
            JOIN users tgt ON tgt.id = r.target_operator_id
            LEFT JOIN directions d ON d.id = r.direction_id
            LEFT JOIN users resp ON resp.id = r.responded_by
            WHERE r.id = %s
            LIMIT 1
            """,
            (int(request_id),)
        )
        return cursor.fetchone()

    def get_shift_swap_candidates(
        self,
        requester_operator_id,
        swap_date,
        start_time,
        end_time,
        end_date=None,
        overlap_only=False,
        non_overlap_only=False
    ):
        requester_operator_id = int(requester_operator_id)
        overlap_only = bool(overlap_only)
        non_overlap_only = bool(non_overlap_only)
        interval = self._normalize_swap_time_range_for_date(
            swap_date=swap_date,
            start_time=start_time,
            end_time=end_time,
            end_date=end_date
        )
        swap_date_obj = interval['swapDateObj']
        swap_date_str = swap_date_obj.strftime('%Y-%m-%d')
        prev_date_obj = swap_date_obj - timedelta(days=1)
        next_date_obj = swap_date_obj + timedelta(days=1)
        prev_date_str = prev_date_obj.strftime('%Y-%m-%d')
        next_date_str = next_date_obj.strftime('%Y-%m-%d')
        interval_start_min = interval['startMin']
        interval_end_min = interval['endMin']

        with self._get_cursor() as cursor:
            cursor.execute(
                """
                SELECT u.id, u.role, u.direction_id, COALESCE(d.name, '')
                FROM users u
                LEFT JOIN directions d ON d.id = u.direction_id
                WHERE u.id = %s
                LIMIT 1
                """,
                (requester_operator_id,)
            )
            requester_row = cursor.fetchone()
            if not requester_row:
                raise ValueError("Requester not found")
            if str(requester_row[1]) != 'operator':
                raise ValueError("Only operators can request swap candidates")
            requester_direction_id = requester_row[2]
            requester_direction_name = requester_row[3]
            if requester_direction_id is None:
                raise ValueError("У вашего профиля не задано направление")

            cursor.execute(
                """
                SELECT
                    u.id,
                    u.name,
                    COALESCE(d.name, ''),
                    COALESCE(u.status, 'working'),
                    u.rate,
                    s.name
                FROM users u
                LEFT JOIN directions d ON d.id = u.direction_id
                LEFT JOIN users s ON s.id = u.supervisor_id
                WHERE u.role = 'operator'
                  AND u.id <> %s
                  AND COALESCE(u.status, 'working') <> 'fired'
                ORDER BY LOWER(u.name), u.id
                """,
                (requester_operator_id,)
            )
            candidate_rows = cursor.fetchall() or []
            candidate_rows = [
                row for row in candidate_rows
                if self._are_swap_directions_compatible(requester_direction_name, row[2])
            ]
            if not candidate_rows:
                return []

            candidate_ids = [int(row[0]) for row in candidate_rows]
            day_shift_map_by_operator = {}
            day_off_map_by_operator = {
                int(op_id): {
                    prev_date_str: False,
                    swap_date_str: False,
                    next_date_str: False
                }
                for op_id in candidate_ids
            }

            cursor.execute(
                """
                SELECT operator_id, day_off_date
                FROM days_off
                WHERE operator_id = ANY(%s)
                  AND day_off_date >= %s
                  AND day_off_date <= %s
                """,
                (candidate_ids, prev_date_obj, next_date_obj)
            )
            for op_id_value, day_off_date in (cursor.fetchall() or []):
                op_id_int = int(op_id_value)
                bucket = day_off_map_by_operator.get(op_id_int)
                if bucket is None:
                    continue
                day_key = day_off_date.strftime('%Y-%m-%d')
                if day_key in bucket:
                    bucket[day_key] = True

            for op_id in candidate_ids:
                shift_map = self._load_operator_shift_map_for_period_tx(
                    cursor=cursor,
                    operator_id=op_id,
                    start_date_obj=prev_date_obj,
                    end_date_obj=next_date_obj
                )
                day_shift_map_by_operator[int(op_id)] = {
                    'prevDayShifts': shift_map.get(prev_date_str, []),
                    'dayShifts': shift_map.get(swap_date_str, []),
                    'nextDayShifts': shift_map.get(next_date_str, [])
                }

        result = []
        for row in candidate_rows:
            op_id = int(row[0])
            shift_payload = day_shift_map_by_operator.get(op_id) or {}
            day_off_payload = day_off_map_by_operator.get(op_id) or {}
            op_prev_day_shifts = shift_payload.get('prevDayShifts') or []
            op_day_shifts = shift_payload.get('dayShifts') or []
            op_next_day_shifts = shift_payload.get('nextDayShifts') or []
            op_is_day_off = bool(day_off_payload.get(swap_date_str))
            op_is_next_day_off = bool(day_off_payload.get(next_date_str))

            window_intervals = self._build_swap_window_intervals(
                prev_day_shifts=op_prev_day_shifts,
                day_shifts=op_day_shifts,
                next_day_shifts=op_next_day_shifts
            )
            overlap_intervals = self._swap_extract_interval(window_intervals, interval_start_min, interval_end_min)
            has_overlap = bool(overlap_intervals)
            if overlap_only and not has_overlap:
                continue
            if non_overlap_only and has_overlap:
                continue

            day_display_intervals = self._swap_extract_interval(window_intervals, 0, 1440)
            next_day_display_intervals_abs = self._swap_extract_interval(window_intervals, 1440, 2880)

            normalized_day_shifts = []
            for seg in (day_display_intervals or []):
                seg_start = int(seg.get('start', 0))
                seg_end = int(seg.get('end', 0))
                if seg_end <= seg_start:
                    continue
                continues_next_day = any(
                    int(iv.get('start', 0)) < 1440
                    and int(iv.get('end', 0)) > 1440
                    and int(iv.get('start', 0)) < seg_end
                    and seg_start < int(iv.get('end', 0))
                    for iv in (window_intervals or [])
                )
                normalized_day_shifts.append({
                    'start': _minutes_to_time(seg_start),
                    'end': _minutes_to_time(seg_end),
                    'continuesNextDay': bool(continues_next_day)
                })

            normalized_next_day_shifts = []
            for seg in (next_day_display_intervals_abs or []):
                seg_start = int(seg.get('start', 0))
                seg_end = int(seg.get('end', 0))
                if seg_end <= seg_start:
                    continue
                normalized_next_day_shifts.append({
                    'start': _minutes_to_time(seg_start - 1440),
                    'end': _minutes_to_time(seg_end - 1440)
                })

            match_starts_at_request_end = any(int(iv.get('start', -1)) == int(interval_end_min) for iv in (window_intervals or []))
            match_ends_at_request_start = any(int(iv.get('end', -1)) == int(interval_start_min) for iv in (window_intervals or []))
            overlap_minutes = sum(
                max(0, int(seg.get('end', 0)) - int(seg.get('start', 0)))
                for seg in (overlap_intervals or [])
            )

            priority_score = int(overlap_minutes)
            if match_starts_at_request_end:
                priority_score += 10
            if match_ends_at_request_start:
                priority_score += 10
            result.append({
                'id': op_id,
                'name': row[1],
                'direction': row[2],
                'status': row[3],
                'rate': float(row[4]) if row[4] is not None else None,
                'supervisorName': row[5],
                'dayShifts': normalized_day_shifts,
                'isDayOff': op_is_day_off,
                'is_day_off': op_is_day_off,
                'nextDayDate': next_date_str,
                'nextDayShifts': normalized_next_day_shifts,
                'isNextDayOff': op_is_next_day_off,
                'is_next_day_off': op_is_next_day_off,
                'matchStartsAtRequestEnd': bool(match_starts_at_request_end),
                'matchEndsAtRequestStart': bool(match_ends_at_request_start),
                'hasOverlap': bool(has_overlap),
                'overlapMinutes': int(overlap_minutes),
                'priorityScore': int(priority_score)
            })

        result.sort(
            key=lambda item: (
                -int(item.get('priorityScore') or 0),
                str(item.get('name') or '').lower(),
                int(item.get('id') or 0)
            )
        )
        return result

    def create_shift_swap_request(
        self,
        requester_operator_id,
        target_operator_id,
        swap_date,
        start_time,
        end_time,
        end_date=None,
        request_comment=None,
        target_segments=None
    ):
        requester_operator_id = int(requester_operator_id)
        target_operator_id = int(target_operator_id)
        if requester_operator_id == target_operator_id:
            raise ValueError("Нельзя отправить запрос самому себе")

        interval = self._normalize_swap_time_range_for_date(
            swap_date=swap_date,
            start_time=start_time,
            end_time=end_time,
            end_date=end_date
        )
        swap_date_obj = interval['swapDateObj']
        swap_date_str = swap_date_obj.strftime('%Y-%m-%d')
        prev_date_obj = swap_date_obj - timedelta(days=1)
        next_date_obj = swap_date_obj + timedelta(days=1)
        prev_date_str = prev_date_obj.strftime('%Y-%m-%d')
        next_date_str = next_date_obj.strftime('%Y-%m-%d')
        interval_start_min = interval['startMin']
        interval_end_min = interval['endMin']
        request_end_date_obj = interval['endDateObj']
        target_requested_intervals = self._swap_parse_payload_segments(
            target_segments,
            swap_date=swap_date_str
        )
        exchange_requested = bool(target_requested_intervals)

        request_comment_norm = None
        if request_comment is not None:
            text = str(request_comment).strip()
            request_comment_norm = text or None

        with self._get_cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    u.id,
                    u.role,
                    u.direction_id,
                    u.name,
                    COALESCE(u.status, 'working'),
                    COALESCE(d.name, '')
                FROM users u
                LEFT JOIN directions d ON d.id = u.direction_id
                WHERE u.id = ANY(%s)
                """,
                ([requester_operator_id, target_operator_id],)
            )
            users_rows = cursor.fetchall() or []
            users_map = {int(row[0]): row for row in users_rows}
            requester = users_map.get(requester_operator_id)
            target = users_map.get(target_operator_id)
            if not requester:
                raise ValueError("Requester not found")
            if not target:
                raise ValueError("Target operator not found")
            if str(requester[1]) != 'operator' or str(target[1]) != 'operator':
                raise ValueError("Swap is available only between operators")
            if str(target[4]) == 'fired':
                raise ValueError("Выбранный оператор уволен")

            requester_direction_id = requester[2]
            target_direction_id = target[2]
            requester_direction_name = requester[5]
            target_direction_name = target[5]
            if requester_direction_id is None or target_direction_id is None:
                raise ValueError("У операторов должно быть указано направление")
            if not self._are_swap_directions_compatible(requester_direction_name, target_direction_name):
                raise ValueError("Оператор для замены должен быть в том же направлении (СМЗ и Основа взаимозаменяемы)")

            cursor.execute(
                """
                SELECT requested_shifts_json
                FROM work_shift_swap_requests
                WHERE status = 'pending'
                  AND requester_operator_id = %s
                  AND target_operator_id = %s
                  AND start_date = %s
                """,
                (requester_operator_id, target_operator_id, swap_date_obj)
            )
            existing_pending_payloads = cursor.fetchall() or []
            for (existing_raw_payload,) in existing_pending_payloads:
                existing_payload = self._normalize_swap_request_payload(existing_raw_payload, fallback_date=swap_date_str)
                ex_start = existing_payload.get('startMin')
                ex_end = existing_payload.get('endMin')
                if ex_start is None or ex_end is None:
                    continue
                if int(ex_start) < interval_end_min and interval_start_min < int(ex_end):
                    raise ValueError("Уже есть активный запрос на пересекающийся интервал для этого оператора")

            requester_day_map = self._load_operator_shift_map_for_period_tx(
                cursor=cursor,
                operator_id=requester_operator_id,
                start_date_obj=prev_date_obj,
                end_date_obj=next_date_obj
            )
            target_day_map = self._load_operator_shift_map_for_period_tx(
                cursor=cursor,
                operator_id=target_operator_id,
                start_date_obj=prev_date_obj,
                end_date_obj=next_date_obj
            )

            requester_window_intervals = self._build_swap_window_intervals(
                prev_day_shifts=requester_day_map.get(prev_date_str, []),
                day_shifts=requester_day_map.get(swap_date_str, []),
                next_day_shifts=requester_day_map.get(next_date_str, [])
            )
            target_window_intervals = self._build_swap_window_intervals(
                prev_day_shifts=target_day_map.get(prev_date_str, []),
                day_shifts=target_day_map.get(swap_date_str, []),
                next_day_shifts=target_day_map.get(next_date_str, [])
            )

            if not self._swap_is_interval_fully_covered(requester_window_intervals, interval_start_min, interval_end_min):
                raise ValueError("Выбранный интервал должен полностью находиться внутри ваших смен")
            if exchange_requested:
                is_full_shift_interval = any(
                    int(seg.get('start', -1)) == int(interval_start_min)
                    and int(seg.get('end', -1)) == int(interval_end_min)
                    for seg in (requester_window_intervals or [])
                )
                if not is_full_shift_interval:
                    raise ValueError("Для обмена нужно выбрать полную смену без обрезки интервала")

            requested_intervals = self._swap_extract_interval(requester_window_intervals, interval_start_min, interval_end_min)
            requested_segments = self._swap_serialize_intervals_for_payload(requested_intervals)
            if not requested_segments:
                raise ValueError("Не удалось сформировать сегменты смен для замены")
            target_segments_norm = []
            if exchange_requested:
                for seg in target_requested_intervals:
                    seg_start = int(seg.get('start', 0))
                    seg_end = int(seg.get('end', 0))
                    if seg_start < 0 or seg_end > 1440:
                        raise ValueError("Для обмена можно выбрать только смены кандидата на ту же дату")
                    if not self._swap_is_interval_fully_covered(target_window_intervals, seg_start, seg_end):
                        raise ValueError("У выбранного оператора больше нет одной из выбранных смен")

                target_selected_intervals = self._merge_break_intervals(target_requested_intervals)
                target_segments_norm = self._swap_serialize_intervals_for_payload(target_selected_intervals)
                if not target_segments_norm:
                    raise ValueError("Не удалось обработать выбранные смены оператора")

                target_remaining_intervals = self._swap_subtract_intervals(
                    target_window_intervals,
                    target_selected_intervals
                )
                if self._swap_interval_sets_overlap(target_remaining_intervals, requested_intervals):
                    raise ValueError("Выбранные смены оператора не покрывают интервал обмена")

            payload = {
                'mode': 'shift_exchange_v2' if exchange_requested else 'shift_replacement_v1',
                'swapDate': swap_date_str,
                'endDate': request_end_date_obj.strftime('%Y-%m-%d'),
                'interval': {
                    'start': interval['startTime'],
                    'end': interval['endTime'],
                    'startMin': interval_start_min,
                    'endMin': interval_end_min
                },
                'requestedSegments': requested_segments,
                'targetSegments': target_segments_norm
            }

            cursor.execute(
                """
                INSERT INTO work_shift_swap_requests (
                    requester_operator_id,
                    target_operator_id,
                    direction_id,
                    start_date,
                    end_date,
                    status,
                    request_comment,
                    requested_shifts_json,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, %s, 'pending', %s, %s::jsonb, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id
                """,
                (
                    requester_operator_id,
                    target_operator_id,
                    int(requester_direction_id),
                    swap_date_obj,
                    request_end_date_obj,
                    request_comment_norm,
                    json.dumps(payload, ensure_ascii=False)
                )
            )
            request_id = int(cursor.fetchone()[0])
            row = self._select_shift_swap_request_by_id_tx(cursor, request_id)
            return self._serialize_shift_swap_request_row(row)

    def get_shift_swap_requests_for_operator(self, operator_id, limit=200):
        operator_id = int(operator_id)
        limit = max(1, min(int(limit), 500))

        with self._get_cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    r.id,
                    r.requester_operator_id, req.name,
                    r.target_operator_id, tgt.name,
                    r.direction_id, d.name,
                    r.start_date, r.end_date,
                    r.status,
                    r.request_comment,
                    r.response_comment,
                    r.requested_shifts_json,
                    r.created_at,
                    r.updated_at,
                    r.responded_at,
                    r.accepted_at,
                    r.responded_by,
                    resp.name
                FROM work_shift_swap_requests r
                JOIN users req ON req.id = r.requester_operator_id
                JOIN users tgt ON tgt.id = r.target_operator_id
                LEFT JOIN directions d ON d.id = r.direction_id
                LEFT JOIN users resp ON resp.id = r.responded_by
                WHERE r.requester_operator_id = %s
                   OR r.target_operator_id = %s
                ORDER BY
                    CASE WHEN r.status = 'pending' THEN 0 ELSE 1 END,
                    r.created_at DESC,
                    r.id DESC
                LIMIT %s
                """,
                (operator_id, operator_id, limit)
            )
            rows = cursor.fetchall() or []

        serialized = [self._serialize_shift_swap_request_row(row) for row in rows]
        incoming = [item for item in serialized if int(item['target']['id']) == operator_id]
        outgoing = [item for item in serialized if int(item['requester']['id']) == operator_id]
        return {
            'incoming': incoming,
            'outgoing': outgoing
        }

    def get_shift_swap_journal_for_month(self, month, requester_role, requester_id, limit=500):
        month_text = str(month or '').strip()
        if not month_text:
            raise ValueError("month is required (YYYY-MM)")
        try:
            month_start = datetime.strptime(month_text, '%Y-%m').date().replace(day=1)
        except Exception:
            raise ValueError("Invalid month format, expected YYYY-MM")

        if month_start.month == 12:
            month_end = date(month_start.year + 1, 1, 1)
        else:
            month_end = date(month_start.year, month_start.month + 1, 1)

        role = str(requester_role or '').strip().lower()
        if not (role_has_min(role, 'admin') or role == 'sv'):
            raise ValueError("Only admin and sv can view swap journal")

        int(requester_id)
        limit = max(1, min(int(limit), 1000))

        with self._get_cursor() as cursor:
            query = """
                SELECT
                    r.id,
                    r.requester_operator_id, req.name,
                    r.target_operator_id, tgt.name,
                    r.direction_id, d.name,
                    r.start_date, r.end_date,
                    r.status,
                    r.request_comment,
                    r.response_comment,
                    r.requested_shifts_json,
                    r.created_at,
                    r.updated_at,
                    r.responded_at,
                    r.accepted_at,
                    r.responded_by,
                    resp.name
                FROM work_shift_swap_requests r
                JOIN users req ON req.id = r.requester_operator_id
                JOIN users tgt ON tgt.id = r.target_operator_id
                LEFT JOIN directions d ON d.id = r.direction_id
                LEFT JOIN users resp ON resp.id = r.responded_by
                WHERE r.start_date >= %s
                  AND r.start_date < %s
            """
            params = [month_start, month_end]

            query += """
                ORDER BY r.start_date DESC, r.created_at DESC, r.id DESC
                LIMIT %s
            """
            params.append(limit)

            cursor.execute(query, params)
            rows = cursor.fetchall() or []

        items = [self._serialize_shift_swap_request_row(row) for row in rows]
        return {
            'month': month_start.strftime('%Y-%m'),
            'items': items
        }

    def respond_shift_swap_request(self, request_id, responder_operator_id, action, response_comment=None):
        request_id = int(request_id)
        responder_operator_id = int(responder_operator_id)
        action_norm = str(action or '').strip().lower()
        if action_norm not in ('accept', 'reject'):
            raise ValueError("action must be 'accept' or 'reject'")

        response_comment_norm = None
        if response_comment is not None:
            text = str(response_comment).strip()
            response_comment_norm = text or None

        with self._get_cursor() as cursor:
            cursor.execute(
                """
                SELECT id, role
                FROM users
                WHERE id = %s
                LIMIT 1
                """,
                (responder_operator_id,)
            )
            responder_row = cursor.fetchone()
            if not responder_row:
                raise ValueError("Responder not found")
            if str(responder_row[1]) != 'operator':
                raise ValueError("Only operators can respond to swap requests")

            cursor.execute(
                """
                SELECT
                    id,
                    requester_operator_id,
                    target_operator_id,
                    start_date,
                    end_date,
                    status,
                    requested_shifts_json
                FROM work_shift_swap_requests
                WHERE id = %s
                FOR UPDATE
                """,
                (request_id,)
            )
            locked_row = cursor.fetchone()
            if not locked_row:
                raise ValueError("Swap request not found")

            (
                _locked_id,
                requester_operator_id,
                target_operator_id,
                start_date_obj,
                end_date_obj,
                current_status,
                requested_shifts_raw
            ) = locked_row

            requester_operator_id = int(requester_operator_id)
            target_operator_id = int(target_operator_id)
            if responder_operator_id != target_operator_id:
                raise ValueError("Только получатель запроса может принять или отклонить замену")
            if str(current_status) != 'pending':
                raise ValueError(f"Запрос уже обработан (статус: {current_status})")

            if action_norm == 'reject':
                cursor.execute(
                    """
                    UPDATE work_shift_swap_requests
                    SET status = 'rejected',
                        response_comment = %s,
                        responded_by = %s,
                        responded_at = CURRENT_TIMESTAMP,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """,
                    (response_comment_norm, responder_operator_id, request_id)
                )
                row = self._select_shift_swap_request_by_id_tx(cursor, request_id)
                return self._serialize_shift_swap_request_row(row)

            fallback_date = start_date_obj.strftime('%Y-%m-%d') if isinstance(start_date_obj, date) else None
            payload = self._normalize_swap_request_payload(requested_shifts_raw, fallback_date=fallback_date)
            swap_date_str = payload.get('swapDate') or fallback_date
            if not swap_date_str:
                raise ValueError("Запрос не содержит дату замены")
            swap_date_obj = self._normalize_schedule_date(swap_date_str)
            requested_segments = payload.get('requestedSegments') or []
            if not requested_segments:
                raise ValueError("Запрос не содержит сегменты смен для переноса")
            target_segments = payload.get('targetSegments') or []

            requested_intervals_payload = self._swap_parse_payload_segments(
                requested_segments,
                swap_date=swap_date_str
            )
            if not requested_intervals_payload:
                raise ValueError("Запрос не содержит корректные сегменты смен инициатора")

            target_intervals_payload = self._swap_parse_payload_segments(
                target_segments,
                swap_date=swap_date_str
            )
            if target_segments and not target_intervals_payload:
                raise ValueError("Запрос не содержит корректные сегменты смен получателя")

            prev_date_obj = swap_date_obj - timedelta(days=1)
            next_date_obj = swap_date_obj + timedelta(days=1)
            prev_date_str = prev_date_obj.strftime('%Y-%m-%d')
            next_date_str = next_date_obj.strftime('%Y-%m-%d')

            requester_day_map = self._load_operator_shift_map_for_period_tx(
                cursor=cursor,
                operator_id=requester_operator_id,
                start_date_obj=prev_date_obj,
                end_date_obj=next_date_obj
            )
            target_day_map = self._load_operator_shift_map_for_period_tx(
                cursor=cursor,
                operator_id=target_operator_id,
                start_date_obj=prev_date_obj,
                end_date_obj=next_date_obj
            )

            requester_window_intervals = self._build_swap_window_intervals(
                prev_day_shifts=requester_day_map.get(prev_date_str, []),
                day_shifts=requester_day_map.get(swap_date_str, []),
                next_day_shifts=requester_day_map.get(next_date_str, [])
            )
            target_window_intervals = self._build_swap_window_intervals(
                prev_day_shifts=target_day_map.get(prev_date_str, []),
                day_shifts=target_day_map.get(swap_date_str, []),
                next_day_shifts=target_day_map.get(next_date_str, [])
            )
            current_requested_intervals = []
            for seg in requested_intervals_payload:
                seg_start = int(seg.get('start', 0))
                seg_end = int(seg.get('end', 0))
                if not self._swap_is_interval_fully_covered(requester_window_intervals, seg_start, seg_end):
                    raise ValueError("Нельзя принять запрос: у инициатора больше нет выбранной смены")
                current_requested_intervals.extend(
                    self._swap_extract_interval(requester_window_intervals, seg_start, seg_end)
                )
            current_requested_intervals = self._merge_break_intervals(current_requested_intervals)
            current_requested_segments = self._swap_serialize_intervals_for_payload(current_requested_intervals)
            if not current_requested_intervals:
                raise ValueError("Нельзя принять запрос: у инициатора больше нет смен в этом интервале")

            requested_sign = self._swap_intervals_signature(requested_segments)
            current_sign = self._swap_intervals_signature(current_requested_segments)
            if requested_sign and requested_sign != current_sign:
                raise ValueError("Нельзя принять запрос: интервалы инициатора изменились")

            current_target_intervals = []
            for seg in target_intervals_payload:
                seg_start = int(seg.get('start', 0))
                seg_end = int(seg.get('end', 0))
                if not self._swap_is_interval_fully_covered(target_window_intervals, seg_start, seg_end):
                    raise ValueError("Нельзя принять запрос: одна из ваших выбранных смен уже изменилась")
                current_target_intervals.extend(
                    self._swap_extract_interval(target_window_intervals, seg_start, seg_end)
                )
            current_target_intervals = self._merge_break_intervals(current_target_intervals)
            current_target_segments = self._swap_serialize_intervals_for_payload(current_target_intervals)
            target_sign = self._swap_intervals_signature(target_segments)
            current_target_sign = self._swap_intervals_signature(current_target_segments)
            if target_sign and target_sign != current_target_sign:
                raise ValueError("Нельзя принять запрос: выбранные вами интервалы изменились")
            if target_sign and not current_target_intervals:
                raise ValueError("Нельзя принять запрос: выбранные вами смены больше недоступны")

            requester_remaining_intervals = self._swap_subtract_intervals(
                requester_window_intervals,
                current_requested_intervals
            )
            target_remaining_intervals = self._swap_subtract_intervals(
                target_window_intervals,
                current_target_intervals
            )

            if self._swap_interval_sets_overlap(target_remaining_intervals, current_requested_intervals):
                if target_sign:
                    raise ValueError("Нельзя принять запрос: выбранные вами смены не покрывают интервал обмена")
                raise ValueError("Нельзя принять запрос: у вас уже есть смена в этом интервале")

            requester_next_intervals = self._merge_break_intervals([
                *(requester_remaining_intervals or []),
                *(current_target_intervals or [])
            ])
            target_next_intervals = self._merge_break_intervals([
                *(target_remaining_intervals or []),
                *(current_requested_intervals or [])
            ])

            # Важно: swap может затронуть "соседние" дни даже если сам интервал
            # заканчивается ровно в 00:00. Пример: смена 17:00-02:00 и передача 17:00-00:00.
            # В этом случае остаток 00:00-02:00 должен сохраниться на следующий день.
            # Раньше это терялось из-за определения "следующего дня" только по interval_end_min > 1440.

            window_offsets = [-1, 0, 1]
            requester_save_day_map = self._swap_serialize_intervals_to_day_map(
                requester_next_intervals,
                swap_date_obj,
                allowed_day_offsets=window_offsets
            )
            target_save_day_map = self._swap_serialize_intervals_to_day_map(
                target_next_intervals,
                swap_date_obj,
                allowed_day_offsets=window_offsets
            )

            def _shift_signature(shifts):
                sign = []
                for seg in (shifts or []):
                    if not isinstance(seg, dict):
                        continue
                    start_val = seg.get('start')
                    end_val = seg.get('end')
                    if not start_val or not end_val:
                        continue
                    sign.append(f"{start_val}|{end_val}")
                sign.sort()
                return sign

            day_candidates = [
                (prev_date_obj, prev_date_str),
                (swap_date_obj, swap_date_str),
                (next_date_obj, next_date_str)
            ]

            requester_affected_dates = []
            target_affected_dates = []
            for day_obj, day_key in day_candidates:
                before_req = _shift_signature(requester_day_map.get(day_key) or [])
                after_req = _shift_signature(requester_save_day_map.get(day_key) or [])
                if before_req != after_req:
                    requester_affected_dates.append(day_obj)

                before_tgt = _shift_signature(target_day_map.get(day_key) or [])
                after_tgt = _shift_signature(target_save_day_map.get(day_key) or [])
                if before_tgt != after_tgt:
                    target_affected_dates.append(day_obj)

            for day_obj in requester_affected_dates:
                self._clear_day_schedule_tx(cursor, requester_operator_id, day_obj)
            for day_obj in target_affected_dates:
                self._clear_day_schedule_tx(cursor, target_operator_id, day_obj)

            for day_obj in requester_affected_dates:
                day_key = day_obj.strftime('%Y-%m-%d')
                for seg in (requester_save_day_map.get(day_key) or []):
                    seg_start_obj = self._normalize_schedule_time(seg.get('start'), 'start_time')
                    seg_end_obj = self._normalize_schedule_time(seg.get('end'), 'end_time')
                    self._save_shift_tx(
                        cursor=cursor,
                        operator_id=requester_operator_id,
                        shift_date=day_obj,
                        start_time=seg_start_obj,
                        end_time=seg_end_obj,
                        # Перерывы пересчитываем по правилам после обмена, старые не переносим.
                        breaks=None
                    )

            for day_obj in target_affected_dates:
                day_key = day_obj.strftime('%Y-%m-%d')
                for seg in (target_save_day_map.get(day_key) or []):
                    seg_start_obj = self._normalize_schedule_time(seg.get('start'), 'start_time')
                    seg_end_obj = self._normalize_schedule_time(seg.get('end'), 'end_time')
                    self._save_shift_tx(
                        cursor=cursor,
                        operator_id=target_operator_id,
                        shift_date=day_obj,
                        start_time=seg_start_obj,
                        end_time=seg_end_obj,
                        # Перерывы пересчитываем по правилам после обмена, старые не переносим.
                        breaks=None
                    )

            cursor.execute(
                """
                UPDATE work_shift_swap_requests
                SET status = 'accepted',
                    response_comment = %s,
                    responded_by = %s,
                    responded_at = CURRENT_TIMESTAMP,
                    accepted_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (response_comment_norm, responder_operator_id, request_id)
            )
            all_affected_dates = sorted(set(requester_affected_dates + target_affected_dates))
            if all_affected_dates:
                self._recalculate_auto_daily_hours_tx(
                    cursor=cursor,
                    operator_ids=[requester_operator_id, target_operator_id],
                    start_date=all_affected_dates[0],
                    end_date=all_affected_dates[-1]
                )
            row = self._select_shift_swap_request_by_id_tx(cursor, request_id)
            return self._serialize_shift_swap_request_row(row)

    def save_shift(
        self,
        operator_id,
        shift_date,
        start_time,
        end_time,
        breaks=None,
        previous_start_time=None,
        previous_end_time=None
    ):
        """
        Сохранить смену для оператора и синхронизировать её с фронтовым merged-состоянием.
        - заменяет пересекающиеся смены на этой дате одной присланной сменой
        - обновляет/перезаписывает перерывы
        - снимает day off на эту дату, если он был
        """
        shift_date_obj = self._normalize_schedule_date(shift_date)
        start_time_obj = self._normalize_schedule_time(start_time, 'start_time')
        end_time_obj = self._normalize_schedule_time(end_time, 'end_time')

        with self._get_cursor() as cursor:
            shift_id = self._save_shift_tx(
                cursor=cursor,
                operator_id=int(operator_id),
                shift_date=shift_date_obj,
                start_time=start_time_obj,
                end_time=end_time_obj,
                breaks=breaks,
                previous_start_time=previous_start_time,
                previous_end_time=previous_end_time
            )
            self._recalculate_auto_daily_hours_tx(
                cursor=cursor,
                operator_ids=[int(operator_id)],
                start_date=shift_date_obj,
                end_date=shift_date_obj
            )
            return shift_id

    def delete_shift(self, operator_id, shift_date, start_time, end_time):
        """
        Удалить конкретную смену оператора.
        """
        shift_date_obj = self._normalize_schedule_date(shift_date)
        start_time_obj = self._normalize_schedule_time(start_time, 'start_time')
        end_time_obj = self._normalize_schedule_time(end_time, 'end_time')

        with self._get_cursor() as cursor:
            cursor.execute("""
                DELETE FROM work_shifts
                WHERE operator_id = %s AND shift_date = %s
                  AND start_time = %s AND end_time = %s
                RETURNING id
            """, (int(operator_id), shift_date_obj, start_time_obj, end_time_obj))
            deleted = cursor.fetchone() is not None
            if deleted:
                self._recalculate_auto_daily_hours_tx(
                    cursor=cursor,
                    operator_ids=[int(operator_id)],
                    start_date=shift_date_obj,
                    end_date=shift_date_obj
                )
            return deleted

    def toggle_day_off(self, operator_id, day_off_date):
        """
        Переключить выходной день для оператора.
        Если день уже выходной - убрать, иначе - добавить и удалить все смены в этот день.
        Возвращает True если день стал выходным, False если был убран.
        """
        day_off_date_obj = self._normalize_schedule_date(day_off_date)
        operator_id = int(operator_id)

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT 1
                FROM days_off
                WHERE operator_id = %s AND day_off_date = %s
            """, (operator_id, day_off_date_obj))
            existing = cursor.fetchone()

            if existing:
                cursor.execute("""
                    DELETE FROM days_off
                    WHERE operator_id = %s AND day_off_date = %s
                """, (operator_id, day_off_date_obj))
                self._recalculate_auto_daily_hours_tx(
                    cursor=cursor,
                    operator_ids=[operator_id],
                    start_date=day_off_date_obj,
                    end_date=day_off_date_obj
                )
                return False

            cursor.execute("""
                INSERT INTO days_off (operator_id, day_off_date)
                VALUES (%s, %s)
                ON CONFLICT (operator_id, day_off_date) DO NOTHING
            """, (operator_id, day_off_date_obj))
            cursor.execute("""
                DELETE FROM work_shifts
                WHERE operator_id = %s AND shift_date = %s
            """, (operator_id, day_off_date_obj))
            self._recalculate_auto_daily_hours_tx(
                cursor=cursor,
                operator_ids=[operator_id],
                start_date=day_off_date_obj,
                end_date=day_off_date_obj
            )
            return True

    def _delete_all_shifts_for_day_tx(self, cursor, operator_id, shift_date):
        operator_id = int(operator_id)
        shift_date_obj = self._normalize_schedule_date(shift_date)
        cursor.execute("""
            DELETE FROM work_shifts
            WHERE operator_id = %s AND shift_date = %s
            RETURNING id
        """, (operator_id, shift_date_obj))
        deleted_rows = cursor.fetchall() or []
        return len(deleted_rows)

    def _delete_shifts_for_period_tx(self, cursor, operator_id, start_date, end_date=None):
        """
        Удалить смены и выходные оператора в периоде [start_date, end_date] включительно.
        Если end_date не указан, удаляются записи начиная с start_date и далее.
        """
        operator_id = int(operator_id)
        start_date_obj = self._normalize_schedule_date(start_date)
        end_date_obj = self._normalize_schedule_date(end_date) if end_date is not None else None
        if end_date_obj is not None and end_date_obj < start_date_obj:
            raise ValueError("end_date must be >= start_date")

        if end_date_obj is None:
            cursor.execute("""
                DELETE FROM work_shifts
                WHERE operator_id = %s
                  AND shift_date >= %s
                RETURNING id, shift_date
            """, (operator_id, start_date_obj))
        else:
            cursor.execute("""
                DELETE FROM work_shifts
                WHERE operator_id = %s
                  AND shift_date >= %s
                  AND shift_date <= %s
                RETURNING id, shift_date
            """, (operator_id, start_date_obj, end_date_obj))

        deleted_rows = cursor.fetchall() or []
        deleted_shifts_count = len(deleted_rows)
        max_deleted_shift_date = None
        for _, shift_date_value in deleted_rows:
            if shift_date_value is None:
                continue
            if max_deleted_shift_date is None or shift_date_value > max_deleted_shift_date:
                max_deleted_shift_date = shift_date_value

        if end_date_obj is None:
            cursor.execute("""
                DELETE FROM days_off
                WHERE operator_id = %s
                  AND day_off_date >= %s
                RETURNING day_off_date
            """, (operator_id, start_date_obj))
        else:
            cursor.execute("""
                DELETE FROM days_off
                WHERE operator_id = %s
                  AND day_off_date >= %s
                  AND day_off_date <= %s
                RETURNING day_off_date
            """, (operator_id, start_date_obj, end_date_obj))

        deleted_day_off_rows = cursor.fetchall() or []
        deleted_days_off_count = len(deleted_day_off_rows)
        max_deleted_day_off_date = None
        for row in deleted_day_off_rows:
            day_off_date_value = row[0] if isinstance(row, tuple) and row else None
            if day_off_date_value is None:
                continue
            if max_deleted_day_off_date is None or day_off_date_value > max_deleted_day_off_date:
                max_deleted_day_off_date = day_off_date_value

        return {
            'deleted_shifts': deleted_shifts_count,
            'deleted_day_off_rows': deleted_days_off_count,
            'max_deleted_shift_date': max_deleted_shift_date,
            'max_deleted_day_off_date': max_deleted_day_off_date
        }

    def _clear_day_schedule_tx(self, cursor, operator_id, target_date):
        """
        Очистка дня: удаляет все смены и снимает выходной на указанную дату.
        """
        operator_id = int(operator_id)
        date_obj = self._normalize_schedule_date(target_date)
        deleted_shifts = self._delete_all_shifts_for_day_tx(cursor, operator_id, date_obj)
        cursor.execute("""
            DELETE FROM days_off
            WHERE operator_id = %s AND day_off_date = %s
        """, (operator_id, date_obj))
        deleted_day_off_rows = int(cursor.rowcount or 0)
        return {
            'operator_id': operator_id,
            'date': date_obj.strftime('%Y-%m-%d'),
            'deleted_shifts': deleted_shifts,
            'deleted_day_off_rows': deleted_day_off_rows
        }

    def _set_day_off_tx(self, cursor, operator_id, day_off_date):
        operator_id = int(operator_id)
        day_off_date_obj = self._normalize_schedule_date(day_off_date)
        cursor.execute("""
            INSERT INTO days_off (operator_id, day_off_date)
            VALUES (%s, %s)
            ON CONFLICT (operator_id, day_off_date) DO NOTHING
        """, (operator_id, day_off_date_obj))
        deleted_shifts = self._delete_all_shifts_for_day_tx(cursor, operator_id, day_off_date_obj)
        return {
            'operator_id': operator_id,
            'date': day_off_date_obj.strftime('%Y-%m-%d'),
            'deleted_shifts': deleted_shifts
        }

    def apply_work_schedule_bulk_actions(self, actions):
        """
        Атомарное выполнение массовых действий по графикам в одном запросе.
        Поддерживаемые actions:
        - {action: "set_shift", operator_id, date, start, end, breaks?}
        - {action: "set_day_off", operator_id, date}
        - {action: "delete_shifts", operator_id, date}  # очистка дня: удаляет смены и снимает выходной
        """
        if not isinstance(actions, list) or not actions:
            raise ValueError("actions must be a non-empty list")

        summary = {
            'total': 0,
            'set_shift': 0,
            'set_day_off': 0,
            'delete_shifts': 0,
            'deleted_shift_rows': 0,
            'deleted_day_off_rows': 0,
            'shift_ids': [],
            'affected_operator_ids': []
        }
        affected_operator_ids = set()
        affected_range_by_operator = {}

        def register_affected_day(op_id, day_obj):
            op_id_int = int(op_id)
            prev = affected_range_by_operator.get(op_id_int)
            if prev is None:
                affected_range_by_operator[op_id_int] = [day_obj, day_obj]
                return
            if day_obj < prev[0]:
                prev[0] = day_obj
            if day_obj > prev[1]:
                prev[1] = day_obj

        with self._get_cursor() as cursor:
            for item in actions:
                if not isinstance(item, dict):
                    raise ValueError("Each action must be an object")

                action_type = str(item.get('action') or item.get('type') or '').strip()
                operator_id = item.get('operator_id')
                date_value = item.get('date') or item.get('shift_date') or item.get('day_off_date')

                if not action_type:
                    raise ValueError("Missing action type")
                if operator_id is None:
                    raise ValueError("Missing operator_id")
                if not date_value:
                    raise ValueError("Missing date")

                operator_id = int(operator_id)
                date_obj = self._normalize_schedule_date(date_value)
                affected_operator_ids.add(operator_id)
                register_affected_day(operator_id, date_obj)

                if action_type == 'set_day_off':
                    result = self._set_day_off_tx(cursor, operator_id, date_obj)
                    summary['set_day_off'] += 1
                    summary['deleted_shift_rows'] += int(result.get('deleted_shifts') or 0)
                    summary['total'] += 1
                    continue

                if action_type == 'delete_shifts':
                    result = self._clear_day_schedule_tx(cursor, operator_id, date_obj)
                    summary['delete_shifts'] += 1
                    summary['deleted_shift_rows'] += int(result.get('deleted_shifts') or 0)
                    summary['deleted_day_off_rows'] += int(result.get('deleted_day_off_rows') or 0)
                    summary['total'] += 1
                    continue

                if action_type == 'set_shift':
                    start_time_obj = self._normalize_schedule_time(item.get('start') or item.get('start_time'), 'start')
                    end_time_obj = self._normalize_schedule_time(item.get('end') or item.get('end_time'), 'end')
                    resolved_breaks = self._resolve_breaks_for_day_replacement_tx(
                        cursor=cursor,
                        operator_id=operator_id,
                        shift_date=date_obj,
                        start_time=start_time_obj,
                        end_time=end_time_obj,
                        breaks=item.get('breaks')
                    )
                    deleted_count = self._delete_all_shifts_for_day_tx(cursor, operator_id, date_obj)
                    shift_id = self._save_shift_tx(
                        cursor=cursor,
                        operator_id=operator_id,
                        shift_date=date_obj,
                        start_time=start_time_obj,
                        end_time=end_time_obj,
                        breaks=resolved_breaks
                    )
                    summary['set_shift'] += 1
                    summary['deleted_shift_rows'] += deleted_count
                    summary['shift_ids'].append(int(shift_id))
                    summary['total'] += 1
                    continue

                raise ValueError(f"Unsupported action type: {action_type}")

            for op_id, bounds in affected_range_by_operator.items():
                self._recalculate_auto_daily_hours_tx(
                    cursor=cursor,
                    operator_ids=[op_id],
                    start_date=bounds[0],
                    end_date=bounds[1]
                )

        summary['affected_operator_ids'] = sorted(affected_operator_ids)
        return summary

    def import_work_schedule_excel_entries(self, entries):
        """
        Импорт расписания из Excel-матрицы (ФИО x Даты), уже после парсинга.
        Каждый entry:
          {
            "operator_id": int,
            "date": "YYYY-MM-DD",
            "is_day_off": bool,
            "shifts": [{"start": "HH:MM", "end": "HH:MM", "breaks": [...]?}, ...]
          }
        Поведение: день заменяется целиком (очистка смен/выходного), затем вставляются смены или выходной.
        """
        if not isinstance(entries, list) or not entries:
            raise ValueError("entries must be a non-empty list")

        summary = {
            'days_processed': 0,
            'set_day_off_days': 0,
            'set_shift_days': 0,
            'shift_rows_saved': 0,
            'deleted_shift_rows': 0,
            'deleted_day_off_rows': 0,
            'blacklist_skipped_entries': [],
            'blacklist_skipped_total': 0,
            'affected_operator_ids': []
        }
        affected_operator_ids = set()
        affected_range_by_operator = {}

        def register_affected_day(op_id, day_obj):
            op_id_int = int(op_id)
            prev = affected_range_by_operator.get(op_id_int)
            if prev is None:
                affected_range_by_operator[op_id_int] = [day_obj, day_obj]
                return
            if day_obj < prev[0]:
                prev[0] = day_obj
            if day_obj > prev[1]:
                prev[1] = day_obj

        def _is_blacklist_shift_block_error(error):
            text = str(error or '').strip().lower()
            return ('чс-увольнение' in text and 'прервать' in text and 'смен' in text)

        def _has_blacklist_dismissal_on_day_tx(cursor, operator_id, day_obj):
            cursor.execute(
                """
                SELECT 1
                FROM operator_schedule_status_periods
                WHERE operator_id = %s
                  AND status_code = 'dismissal'
                  AND COALESCE(is_blacklist, FALSE) = TRUE
                  AND start_date <= %s
                  AND COALESCE(end_date, DATE '9999-12-31') >= %s
                LIMIT 1
                """,
                (int(operator_id), day_obj, day_obj)
            )
            return cursor.fetchone() is not None

        def _append_blacklist_skipped_entry(operator_id, date_obj, operator_name=None):
            payload = {
                'operator_id': int(operator_id),
                'date': date_obj.strftime('%Y-%m-%d'),
                'reason': 'blacklist_dismissal'
            }
            name_text = str(operator_name or '').strip()
            if name_text:
                payload['name'] = name_text
            summary['blacklist_skipped_entries'].append(payload)
            summary['blacklist_skipped_total'] += 1

        with self._get_cursor() as cursor:
            for item in entries:
                if not isinstance(item, dict):
                    raise ValueError("Each entry must be an object")

                operator_id = item.get('operator_id')
                date_value = item.get('date')
                is_day_off = bool(item.get('is_day_off'))
                shifts = item.get('shifts') or []
                operator_name = item.get('operator_name') or item.get('name')

                if operator_id is None:
                    raise ValueError("Missing operator_id in entry")
                if not date_value:
                    raise ValueError("Missing date in entry")
                if not is_day_off and not isinstance(shifts, list):
                    raise ValueError("shifts must be a list")
                if is_day_off and shifts:
                    raise ValueError("Entry cannot have shifts when is_day_off is true")

                operator_id = int(operator_id)
                date_obj = self._normalize_schedule_date(date_value)
                affected_operator_ids.add(operator_id)
                register_affected_day(operator_id, date_obj)

                if (not is_day_off) and shifts and _has_blacklist_dismissal_on_day_tx(cursor, operator_id, date_obj):
                    _append_blacklist_skipped_entry(operator_id, date_obj, operator_name)
                    continue

                cleared = self._clear_day_schedule_tx(cursor, operator_id, date_obj)
                summary['deleted_shift_rows'] += int(cleared.get('deleted_shifts') or 0)
                summary['deleted_day_off_rows'] += int(cleared.get('deleted_day_off_rows') or 0)

                if is_day_off:
                    cursor.execute("""
                        INSERT INTO days_off (operator_id, day_off_date)
                        VALUES (%s, %s)
                        ON CONFLICT (operator_id, day_off_date) DO NOTHING
                    """, (operator_id, date_obj))
                    summary['set_day_off_days'] += 1
                    summary['days_processed'] += 1
                    continue

                if not shifts:
                    # Пустая очистка дня (если будет нужна) — считаем обработанным днем.
                    summary['days_processed'] += 1
                    continue

                # Сохраняем несколько смен в один день после полной очистки этого дня.
                blocked_by_blacklist = False
                for shift in shifts:
                    if not isinstance(shift, dict):
                        raise ValueError("Each shift must be an object")
                    start_time_obj = self._normalize_schedule_time(shift.get('start') or shift.get('start_time'), 'start')
                    end_time_obj = self._normalize_schedule_time(shift.get('end') or shift.get('end_time'), 'end')
                    try:
                        self._save_shift_tx(
                            cursor=cursor,
                            operator_id=operator_id,
                            shift_date=date_obj,
                            start_time=start_time_obj,
                            end_time=end_time_obj,
                            breaks=shift.get('breaks') if ('breaks' in shift) else None
                        )
                    except ValueError as e:
                        if _is_blacklist_shift_block_error(e):
                            _append_blacklist_skipped_entry(operator_id, date_obj, operator_name)
                            rollback_cleared = self._clear_day_schedule_tx(cursor, operator_id, date_obj)
                            summary['deleted_shift_rows'] += int(rollback_cleared.get('deleted_shifts') or 0)
                            summary['deleted_day_off_rows'] += int(rollback_cleared.get('deleted_day_off_rows') or 0)
                            blocked_by_blacklist = True
                            break
                        raise
                    summary['shift_rows_saved'] += 1

                if blocked_by_blacklist:
                    continue

                summary['set_shift_days'] += 1
                summary['days_processed'] += 1

            for op_id, bounds in affected_range_by_operator.items():
                self._recalculate_auto_daily_hours_tx(
                    cursor=cursor,
                    operator_ids=[op_id],
                    start_date=bounds[0],
                    end_date=bounds[1]
                )

        summary['affected_operator_ids'] = sorted(affected_operator_ids)
        return summary

    def save_operator_status_import(self, events, segments, imported_by=None, summary=None):
        """
        Сохранить импорт статусов операторов:
        - сырые события переключений (operator_status_events)
        - рассчитанные интервалы/сегменты для аналитики (operator_status_segments)

        Перед вставкой удаляет существующие данные по затронутым операторам и датам,
        чтобы импорт работал как "replace" в рамках импортированного диапазона.
        """
        if not isinstance(events, list):
            raise ValueError("events must be a list")
        if not isinstance(segments, list):
            raise ValueError("segments must be a list")

        summary_payload = summary if isinstance(summary, dict) else {}

        def _safe_int(value, default=0):
            try:
                return int(value)
            except Exception:
                return int(default)

        imported_by_id = int(imported_by) if imported_by is not None else None
        batch_id = uuid.uuid4()

        normalized_events = []
        normalized_segments = []

        event_ranges_map = {}
        segment_ranges_map = {}
        affected_operator_ids = set()

        def _extend_range(ranges_map, operator_id, day_value):
            prev = ranges_map.get(operator_id)
            if prev is None:
                ranges_map[operator_id] = [day_value, day_value]
                return
            if day_value < prev[0]:
                prev[0] = day_value
            if day_value > prev[1]:
                prev[1] = day_value

        for item in events:
            if not isinstance(item, dict):
                continue
            try:
                operator_id = int(item.get('operator_id'))
            except Exception:
                continue

            event_at_raw = item.get('event_at')
            if isinstance(event_at_raw, datetime):
                event_at_obj = event_at_raw
            elif isinstance(event_at_raw, str):
                text = str(event_at_raw).strip()
                if not text:
                    continue
                try:
                    event_at_obj = datetime.fromisoformat(text)
                except Exception:
                    continue
            else:
                continue

            event_date_obj = event_at_obj.date()
            state_name_raw = str(item.get('state_name') or '').strip()
            status_key = self._normalize_import_status_key(
                item.get('status_key')
                or item.get('state_key')
                or state_name_raw
            )
            if not status_key:
                continue
            state_note = str(item.get('state_note') or '').strip() or None

            normalized_events.append({
                'operator_id': operator_id,
                'event_at': event_at_obj,
                'event_date': event_date_obj,
                'status_key': status_key,
                'state_note': state_note,
                'imported_by': imported_by_id
            })
            affected_operator_ids.add(operator_id)
            _extend_range(event_ranges_map, operator_id, event_date_obj)

        for item in segments:
            if not isinstance(item, dict):
                continue
            try:
                operator_id = int(item.get('operator_id'))
            except Exception:
                continue

            status_date_raw = item.get('status_date')
            if isinstance(status_date_raw, date):
                status_date_obj = status_date_raw
            elif isinstance(status_date_raw, str):
                text = str(status_date_raw).strip()
                if not text:
                    continue
                try:
                    status_date_obj = datetime.strptime(text, '%Y-%m-%d').date()
                except Exception:
                    continue
            else:
                continue

            start_at_raw = item.get('start_at')
            end_at_raw = item.get('end_at')
            if isinstance(start_at_raw, datetime):
                start_at_obj = start_at_raw
            elif isinstance(start_at_raw, str):
                text = str(start_at_raw).strip()
                if not text:
                    continue
                try:
                    start_at_obj = datetime.fromisoformat(text)
                except Exception:
                    continue
            else:
                continue

            if isinstance(end_at_raw, datetime):
                end_at_obj = end_at_raw
            elif isinstance(end_at_raw, str):
                text = str(end_at_raw).strip()
                if not text:
                    continue
                try:
                    end_at_obj = datetime.fromisoformat(text)
                except Exception:
                    continue
            else:
                continue

            if end_at_obj <= start_at_obj:
                continue

            duration_sec = _safe_int(item.get('duration_sec'), default=0)
            if duration_sec <= 0:
                duration_sec = int((end_at_obj - start_at_obj).total_seconds())
            if duration_sec <= 0:
                continue

            state_name_raw = str(item.get('state_name') or '').strip()
            status_key = self._normalize_import_status_key(
                item.get('status_key')
                or item.get('state_key')
                or state_name_raw
            )
            if not status_key:
                continue
            state_note = str(item.get('state_note') or '').strip() or None

            normalized_segments.append({
                'operator_id': operator_id,
                'status_date': status_date_obj,
                'start_at': start_at_obj,
                'end_at': end_at_obj,
                'duration_sec': duration_sec,
                'status_key': status_key,
                'state_note': state_note,
                'imported_by': imported_by_id
            })
            affected_operator_ids.add(operator_id)
            _extend_range(segment_ranges_map, operator_id, status_date_obj)

        event_ranges = sorted(
            [(int(op_id), bounds[0], bounds[1]) for op_id, bounds in event_ranges_map.items()],
            key=lambda x: (x[0], x[1], x[2])
        )
        segment_ranges = sorted(
            [(int(op_id), bounds[0], bounds[1]) for op_id, bounds in segment_ranges_map.items()],
            key=lambda x: (x[0], x[1], x[2])
        )
        # Для segment-данных используем объединение диапазонов:
        # даже если после новых событий не получилось сегментов (например, одиночное событие),
        # старые сегменты за этот период нужно очистить, чтобы не оставались устаревшие данные.
        segment_delete_ranges_map = {}
        for op_id, bounds in segment_ranges_map.items():
            segment_delete_ranges_map[int(op_id)] = [bounds[0], bounds[1]]
        for op_id, bounds in event_ranges_map.items():
            op_id_int = int(op_id)
            prev = segment_delete_ranges_map.get(op_id_int)
            if prev is None:
                segment_delete_ranges_map[op_id_int] = [bounds[0], bounds[1]]
                continue
            if bounds[0] < prev[0]:
                prev[0] = bounds[0]
            if bounds[1] > prev[1]:
                prev[1] = bounds[1]
        segment_delete_ranges = sorted(
            [(int(op_id), bounds[0], bounds[1]) for op_id, bounds in segment_delete_ranges_map.items()],
            key=lambda x: (x[0], x[1], x[2])
        )

        range_source = segment_delete_ranges if segment_delete_ranges else event_ranges
        date_from_obj = min((r[1] for r in range_source), default=None)
        date_to_obj = max((r[2] for r in range_source), default=None)

        deleted_events = 0
        deleted_segments = 0
        auto_aggregation_summary = {
            'updated_days': 0,
            'aggregated_months': 0,
            'auto_flag_fines': {
                'inserted': 0,
                'updated': 0,
                'deleted': 0
            }
        }

        with self._get_cursor() as cursor:
            for operator_id, start_date_obj, end_date_obj in event_ranges:
                cursor.execute(
                    """
                    DELETE FROM operator_status_events
                    WHERE operator_id = %s
                      AND event_date >= %s
                      AND event_date <= %s
                    """,
                    (operator_id, start_date_obj, end_date_obj)
                )
                deleted_events += max(0, int(cursor.rowcount or 0))

            for operator_id, start_date_obj, end_date_obj in segment_delete_ranges:
                cursor.execute(
                    """
                    DELETE FROM operator_status_segments
                    WHERE operator_id = %s
                      AND status_date >= %s
                      AND status_date <= %s
                    """,
                    (operator_id, start_date_obj, end_date_obj)
                )
                deleted_segments += max(0, int(cursor.rowcount or 0))

            batch_meta = summary_payload.get('meta')
            if not isinstance(batch_meta, dict):
                batch_meta = {}
            batch_meta_json = json.dumps(batch_meta, ensure_ascii=False)

            cursor.execute(
                """
                INSERT INTO operator_status_import_batches (
                    id, imported_by,
                    source_rows, valid_events, matched_events, segments_saved,
                    deleted_events, deleted_segments,
                    invalid_rows_count, parse_errors_count,
                    operators_count, open_tail_events, zero_or_negative_transitions,
                    date_from, date_to, meta_json
                )
                VALUES (
                    %s, %s,
                    %s, %s, %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s, %s::jsonb
                )
                """,
                (
                    str(batch_id),
                    imported_by_id,
                    _safe_int(summary_payload.get('source_rows'), default=0),
                    _safe_int(summary_payload.get('valid_events'), default=0),
                    len(normalized_events),
                    len(normalized_segments),
                    deleted_events,
                    deleted_segments,
                    _safe_int(summary_payload.get('invalid_rows_count'), default=0),
                    _safe_int(summary_payload.get('parse_errors_count'), default=0),
                    len(affected_operator_ids),
                    _safe_int(summary_payload.get('open_tail_events'), default=0),
                    _safe_int(summary_payload.get('zero_or_negative_transitions'), default=0),
                    date_from_obj,
                    date_to_obj,
                    batch_meta_json
                )
            )

            if normalized_events:
                event_insert_columns = [
                    'operator_id',
                    'event_at',
                    'event_date',
                    'status_key',
                    'state_note',
                    'imported_by'
                ]
                event_values = []
                for item in normalized_events:
                    row_values = []
                    for col in event_insert_columns:
                        row_values.append(item.get(col))
                    event_values.append(tuple(row_values))
                event_columns_sql = ', '.join(event_insert_columns)
                execute_values(
                    cursor,
                    f"""
                    INSERT INTO operator_status_events (
                        {event_columns_sql}
                    )
                    VALUES %s
                    """,
                    event_values,
                    page_size=STATUS_IMPORT_INSERT_PAGE_SIZE
                )

            if normalized_segments:
                segment_insert_columns = [
                    'operator_id',
                    'status_date',
                    'start_at',
                    'end_at',
                    'duration_sec',
                    'status_key',
                    'state_note',
                    'imported_by'
                ]
                segment_values = []
                for item in normalized_segments:
                    row_values = []
                    for col in segment_insert_columns:
                        row_values.append(item.get(col))
                    segment_values.append(tuple(row_values))
                segment_columns_sql = ', '.join(segment_insert_columns)
                execute_values(
                    cursor,
                    f"""
                    INSERT INTO operator_status_segments (
                        {segment_columns_sql}
                    )
                    VALUES %s
                    """,
                    segment_values,
                    page_size=STATUS_IMPORT_INSERT_PAGE_SIZE
                )

            if affected_operator_ids and date_from_obj and date_to_obj:
                auto_aggregation_summary = self._recalculate_auto_daily_hours_tx(
                    cursor=cursor,
                    operator_ids=sorted(int(v) for v in affected_operator_ids),
                    start_date=(date_from_obj - timedelta(days=1)),
                    end_date=date_to_obj
                )

        return {
            'batch_id': str(batch_id),
            'source_rows': _safe_int(summary_payload.get('source_rows'), default=0),
            'valid_events': _safe_int(summary_payload.get('valid_events'), default=0),
            'matched_events': len(normalized_events),
            'segments_saved': len(normalized_segments),
            'deleted_events': int(deleted_events),
            'deleted_segments': int(deleted_segments),
            'invalid_rows_count': _safe_int(summary_payload.get('invalid_rows_count'), default=0),
            'parse_errors_count': _safe_int(summary_payload.get('parse_errors_count'), default=0),
            'operators_count': len(affected_operator_ids),
            'open_tail_events': _safe_int(summary_payload.get('open_tail_events'), default=0),
            'zero_or_negative_transitions': _safe_int(summary_payload.get('zero_or_negative_transitions'), default=0),
            'date_from': date_from_obj.strftime('%Y-%m-%d') if date_from_obj else None,
            'date_to': date_to_obj.strftime('%Y-%m-%d') if date_to_obj else None,
            'auto_aggregation': auto_aggregation_summary
        }

    def save_schedule_status_period(
        self,
        operator_id,
        status_code,
        start_date,
        end_date=None,
        dismissal_reason=None,
        is_blacklist=False,
        comment=None,
        created_by=None
    ):
        """
        Сохранить период специального статуса оператора.
        Новая запись замещает пересекающиеся периоды (с обрезкой/разделением при необходимости).
        """
        operator_id = int(operator_id)
        status_code_norm = self._normalize_schedule_status_code(status_code)
        start_date_obj = self._normalize_schedule_date(start_date)

        comment_norm = None
        if comment is not None:
            comment_text = str(comment).strip()
            comment_norm = comment_text or None

        if status_code_norm == 'dismissal':
            end_date_obj = self._normalize_schedule_date(end_date) if end_date else None
            if end_date_obj is not None and end_date_obj < start_date_obj:
                raise ValueError("end_date must be >= start_date")
            dismissal_reason_norm = str(dismissal_reason or '').strip()
            if dismissal_reason_norm not in set(SCHEDULE_DISMISSAL_REASONS):
                raise ValueError("Invalid dismissal_reason")
            if isinstance(is_blacklist, str):
                is_blacklist_norm = str(is_blacklist).strip().lower() in ('1', 'true', 'yes', 'on')
            else:
                is_blacklist_norm = bool(is_blacklist)
            if is_blacklist_norm and end_date_obj is not None:
                raise ValueError("Blacklisted dismissal cannot have end_date")
            if not comment_norm:
                raise ValueError("Comment is required for dismissal")
        else:
            end_date_obj = self._normalize_schedule_date(end_date or start_date_obj)
            if end_date_obj < start_date_obj:
                raise ValueError("end_date must be >= start_date")
            dismissal_reason_norm = None
            is_blacklist_norm = False

        created_by_id = int(created_by) if created_by is not None else None
        infinite_date = date(9999, 12, 31)
        new_end_cmp = end_date_obj or infinite_date

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, status_code, start_date, end_date, dismissal_reason, comment, created_by, COALESCE(is_blacklist, FALSE)
                FROM operator_schedule_status_periods
                WHERE operator_id = %s
                  AND start_date <= %s
                  AND COALESCE(end_date, DATE '9999-12-31') >= %s
                ORDER BY start_date, id
            """, (operator_id, new_end_cmp, start_date_obj))
            overlapping = cursor.fetchall()

            has_blacklist_dismissal_overlap = any(
                str(row[1]) == 'dismissal' and bool(row[7])
                for row in overlapping
            )
            if has_blacklist_dismissal_overlap:
                raise ValueError("ЧС-увольнение нельзя изменить или перекрыть другим статусом")

            for row in overlapping:
                (
                    existing_id,
                    existing_status_code,
                    existing_start,
                    existing_end,
                    existing_dismissal_reason,
                    existing_comment,
                    existing_created_by,
                    existing_is_blacklist
                ) = row
                existing_end_cmp = existing_end or infinite_date

                left_exists = existing_start < start_date_obj
                right_exists = (end_date_obj is not None) and (existing_end_cmp > end_date_obj)

                if left_exists and right_exists:
                    left_end = start_date_obj - timedelta(days=1)
                    right_start = end_date_obj + timedelta(days=1)

                    cursor.execute("""
                        UPDATE operator_schedule_status_periods
                        SET end_date = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (left_end, existing_id))

                    cursor.execute("""
                        INSERT INTO operator_schedule_status_periods (
                            operator_id, status_code, start_date, end_date,
                            dismissal_reason, is_blacklist, comment, created_by, created_at, updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    """, (
                        operator_id,
                        existing_status_code,
                        right_start,
                        existing_end,
                        existing_dismissal_reason,
                        bool(existing_is_blacklist),
                        existing_comment,
                        existing_created_by
                    ))
                    continue

                if left_exists:
                    left_end = start_date_obj - timedelta(days=1)
                    cursor.execute("""
                        UPDATE operator_schedule_status_periods
                        SET end_date = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (left_end, existing_id))
                    continue

                if right_exists:
                    right_start = end_date_obj + timedelta(days=1)
                    cursor.execute("""
                        UPDATE operator_schedule_status_periods
                        SET start_date = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (right_start, existing_id))
                    continue

                cursor.execute("""
                    DELETE FROM operator_schedule_status_periods
                    WHERE id = %s
                """, (existing_id,))

            cursor.execute("""
                INSERT INTO operator_schedule_status_periods (
                    operator_id, status_code, start_date, end_date,
                    dismissal_reason, is_blacklist, comment, created_by, created_at, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id, operator_id, status_code, start_date, end_date, dismissal_reason, comment, is_blacklist
            """, (
                operator_id,
                status_code_norm,
                start_date_obj,
                end_date_obj,
                dismissal_reason_norm,
                is_blacklist_norm,
                comment_norm,
                created_by_id
            ))
            saved_row = cursor.fetchone()
            deleted_schedule = self._delete_shifts_for_period_tx(
                cursor=cursor,
                operator_id=operator_id,
                start_date=start_date_obj,
                end_date=end_date_obj
            )
            recalc_end_date_obj = end_date_obj
            if recalc_end_date_obj is None:
                recalc_end_date_obj = start_date_obj
                if isinstance(deleted_schedule, dict):
                    max_deleted_shift_date = deleted_schedule.get('max_deleted_shift_date')
                    max_deleted_day_off_date = deleted_schedule.get('max_deleted_day_off_date')
                    if isinstance(max_deleted_shift_date, date) and max_deleted_shift_date > recalc_end_date_obj:
                        recalc_end_date_obj = max_deleted_shift_date
                    if isinstance(max_deleted_day_off_date, date) and max_deleted_day_off_date > recalc_end_date_obj:
                        recalc_end_date_obj = max_deleted_day_off_date
            self._recalculate_auto_daily_hours_tx(
                cursor=cursor,
                operator_ids=[operator_id],
                start_date=start_date_obj,
                end_date=recalc_end_date_obj
            )
            self._sync_user_statuses_from_schedule_periods_tx(cursor, operator_ids=[operator_id])
            return self._serialize_schedule_status_period(saved_row)

    def _interrupt_dismissal_period_by_work_day_tx(self, cursor, operator_id, work_date):
        """
        Если на рабочую дату приходится статус "увольнение", обрываем его:
        - если увольнение начинается в этот день -> удаляем период
        - если началось раньше -> ставим конец на день раньше
        """
        operator_id = int(operator_id)
        work_date_obj = self._normalize_schedule_date(work_date)

        cursor.execute("""
            SELECT id, start_date, end_date, COALESCE(is_blacklist, FALSE)
            FROM operator_schedule_status_periods
            WHERE operator_id = %s
              AND status_code = 'dismissal'
              AND start_date <= %s
              AND COALESCE(end_date, DATE '9999-12-31') >= %s
            ORDER BY start_date, id
        """, (operator_id, work_date_obj, work_date_obj))
        rows = cursor.fetchall()

        for period_id, start_date_value, end_date_value, is_blacklist_value in rows:
            if bool(is_blacklist_value):
                raise ValueError("ЧС-увольнение нельзя прервать сменой")
            if start_date_value >= work_date_obj:
                cursor.execute("""
                    DELETE FROM operator_schedule_status_periods
                    WHERE id = %s
                """, (period_id,))
                continue

            new_end_date = work_date_obj - timedelta(days=1)
            cursor.execute("""
                UPDATE operator_schedule_status_periods
                SET end_date = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (new_end_date, period_id))

    def delete_schedule_status_period(self, status_period_id, operator_id=None):
        """
        Удалить специальный статус-период по id.
        Если передан operator_id — удаление ограничивается этим оператором.
        """
        period_id = int(status_period_id)
        operator_id_norm = int(operator_id) if operator_id is not None else None

        with self._get_cursor() as cursor:
            if operator_id_norm is None:
                cursor.execute("""
                    SELECT id, operator_id, status_code, start_date, end_date, dismissal_reason, comment, COALESCE(is_blacklist, FALSE)
                    FROM operator_schedule_status_periods
                    WHERE id = %s
                """, (period_id,))
            else:
                cursor.execute("""
                    SELECT id, operator_id, status_code, start_date, end_date, dismissal_reason, comment, COALESCE(is_blacklist, FALSE)
                    FROM operator_schedule_status_periods
                    WHERE id = %s AND operator_id = %s
                """, (period_id, operator_id_norm))
            existing_row = cursor.fetchone()
            if not existing_row:
                return None
            if str(existing_row[2]) == 'dismissal' and bool(existing_row[7]):
                raise ValueError("ЧС-увольнение нельзя удалить")

            if operator_id_norm is None:
                cursor.execute("""
                    DELETE FROM operator_schedule_status_periods
                    WHERE id = %s
                    RETURNING id, operator_id, status_code, start_date, end_date, dismissal_reason, comment, is_blacklist
                """, (period_id,))
            else:
                cursor.execute("""
                    DELETE FROM operator_schedule_status_periods
                    WHERE id = %s AND operator_id = %s
                    RETURNING id, operator_id, status_code, start_date, end_date, dismissal_reason, comment, is_blacklist
                """, (period_id, operator_id_norm))

            row = cursor.fetchone()
            if not row:
                return None
            self._sync_user_statuses_from_schedule_periods_tx(cursor, operator_ids=[int(row[1])])
            return self._serialize_schedule_status_period(row)

    def save_shifts_bulk(self, operator_id, shifts_data):
        """
        Массовое сохранение смен для оператора.
        shifts_data: [{'date': 'YYYY-MM-DD', 'start': 'HH:MM', 'end': 'HH:MM', 'breaks': [...]}, ...]
        """
        operator_id = int(operator_id)
        if not isinstance(shifts_data, list):
            raise ValueError("shifts must be a list")

        result_ids = []
        min_date_obj = None
        max_date_obj = None
        with self._get_cursor() as cursor:
            for shift in shifts_data:
                if not isinstance(shift, dict):
                    raise ValueError("Each shift must be an object")
                shift_date_obj = self._normalize_schedule_date(shift.get('date'))
                if min_date_obj is None or shift_date_obj < min_date_obj:
                    min_date_obj = shift_date_obj
                if max_date_obj is None or shift_date_obj > max_date_obj:
                    max_date_obj = shift_date_obj
                start_time_obj = self._normalize_schedule_time(shift.get('start'), 'start')
                end_time_obj = self._normalize_schedule_time(shift.get('end'), 'end')
                shift_id = self._save_shift_tx(
                    cursor=cursor,
                    operator_id=operator_id,
                    shift_date=shift_date_obj,
                    start_time=start_time_obj,
                    end_time=end_time_obj,
                    breaks=shift.get('breaks')
                )
                result_ids.append(shift_id)

            if min_date_obj is not None and max_date_obj is not None:
                self._recalculate_auto_daily_hours_tx(
                    cursor=cursor,
                    operator_ids=[operator_id],
                    start_date=min_date_obj,
                    end_date=max_date_obj
                )

        return result_ids

    def _serialize_recruiting_run_row(self, row):
        if not row:
            return None
        return {
            "run_uuid": row[0],
            "source": row[1],
            "triggered_by": row[2],
            "status": row[3],
            "total_items": int(row[4] or 0),
            "error_message": row[5],
            "started_at": row[6].isoformat() if hasattr(row[6], "isoformat") else row[6],
            "finished_at": row[7].isoformat() if hasattr(row[7], "isoformat") else row[7],
        }

    def _normalize_recruiting_item(self, item):
        payload = item or {}

        def to_text(value):
            if value is None:
                return None
            text = str(value).strip()
            return text or None

        page_found = None
        raw_page = payload.get("page_found")
        if raw_page is not None and str(raw_page).strip() != "":
            try:
                page_found = int(raw_page)
            except Exception:
                page_found = None

        return {
            "keyword_group": to_text(payload.get("keyword_group")),
            "keyword_query": to_text(payload.get("keyword_query")),
            "page_found": page_found,
            "title": to_text(payload.get("title")),
            "category": to_text(payload.get("category")),
            "experience": to_text(payload.get("experience")),
            "location": to_text(payload.get("location")),
            "salary": to_text(payload.get("salary")),
            "education": to_text(payload.get("education")),
            "published_at": to_text(payload.get("published_at")),
            "detail_url": to_text(payload.get("detail_url")),
        }

    def save_recruiting_parse_success(self, items, source='enbek', triggered_by='scheduler'):
        source_norm = (str(source or 'enbek').strip().lower() or 'enbek')[:64]
        trigger_norm = (str(triggered_by or 'scheduler').strip().lower() or 'scheduler')[:32]
        run_uuid = str(uuid.uuid4())
        normalized_rows = []
        for raw in (items or []):
            if isinstance(raw, dict):
                normalized_rows.append(self._normalize_recruiting_item(raw))
        deduped_rows = []
        seen_links = set()
        for row in normalized_rows:
            detail_url = row.get("detail_url")
            if detail_url:
                detail_key = str(detail_url).strip().lower()
                if detail_key in seen_links:
                    continue
                seen_links.add(detail_key)
            deduped_rows.append(row)

        with self._get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO recruiting_parse_runs (
                    run_uuid, source, triggered_by, status, total_items, error_message, started_at, finished_at
                )
                VALUES (%s::uuid, %s, %s, 'success', %s, NULL, CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty', CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
            """, (run_uuid, source_norm, trigger_norm, len(deduped_rows)))

            if deduped_rows:
                values = [
                    (
                        run_uuid,
                        row["keyword_group"],
                        row["keyword_query"],
                        row["page_found"],
                        row["title"],
                        row["category"],
                        row["experience"],
                        row["location"],
                        row["salary"],
                        row["education"],
                        row["published_at"],
                        row["detail_url"],
                    )
                    for row in deduped_rows
                ]
                execute_values(
                    cursor,
                    """
                    INSERT INTO recruiting_resumes (
                        run_uuid, keyword_group, keyword_query, page_found,
                        title, category, experience, location, salary,
                        education, published_at, detail_url
                    )
                    VALUES %s
                    ON CONFLICT (run_uuid, detail_url) DO NOTHING
                    """,
                    values,
                    page_size=1000
                )

            cursor.execute("""
                SELECT run_uuid::text, source, triggered_by, status, total_items, error_message, started_at, finished_at
                FROM recruiting_parse_runs
                WHERE run_uuid = %s::uuid
                LIMIT 1
            """, (run_uuid,))
            run_row = cursor.fetchone()

        return {
            "run": self._serialize_recruiting_run_row(run_row),
            "inserted": len(deduped_rows),
            "duplicates_dropped": max(0, len(normalized_rows) - len(deduped_rows))
        }

    def save_recruiting_parse_failure(self, error_message, source='enbek', triggered_by='scheduler'):
        source_norm = (str(source or 'enbek').strip().lower() or 'enbek')[:64]
        trigger_norm = (str(triggered_by or 'scheduler').strip().lower() or 'scheduler')[:32]
        run_uuid = str(uuid.uuid4())
        error_text = (str(error_message or '').strip() or 'Parser execution failed')[:4000]

        with self._get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO recruiting_parse_runs (
                    run_uuid, source, triggered_by, status, total_items, error_message, started_at, finished_at
                )
                VALUES (%s::uuid, %s, %s, 'failed', 0, %s, CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty', CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
            """, (run_uuid, source_norm, trigger_norm, error_text))

            cursor.execute("""
                SELECT run_uuid::text, source, triggered_by, status, total_items, error_message, started_at, finished_at
                FROM recruiting_parse_runs
                WHERE run_uuid = %s::uuid
                LIMIT 1
            """, (run_uuid,))
            run_row = cursor.fetchone()

        return {"run": self._serialize_recruiting_run_row(run_row)}

    def get_latest_recruiting_parse_run(self, status='success'):
        status_norm = (str(status or '').strip().lower() or None)
        with self._get_cursor() as cursor:
            if status_norm:
                cursor.execute("""
                    SELECT run_uuid::text, source, triggered_by, status, total_items, error_message, started_at, finished_at
                    FROM recruiting_parse_runs
                    WHERE status = %s
                    ORDER BY finished_at DESC, id DESC
                    LIMIT 1
                """, (status_norm,))
            else:
                cursor.execute("""
                    SELECT run_uuid::text, source, triggered_by, status, total_items, error_message, started_at, finished_at
                    FROM recruiting_parse_runs
                    ORDER BY finished_at DESC, id DESC
                    LIMIT 1
                """)
            row = cursor.fetchone()
        return self._serialize_recruiting_run_row(row)

    def get_recruiting_resumes(self, run_uuid=None, limit=500, offset=0, search=None, keyword_group=None):
        try:
            limit_val = int(limit)
        except Exception:
            limit_val = 500
        limit_val = max(1, min(limit_val, 5000))

        try:
            offset_val = int(offset)
        except Exception:
            offset_val = 0
        offset_val = max(0, offset_val)

        run_uuid_text = (str(run_uuid or '').strip() or None)
        with self._get_cursor() as cursor:
            if not run_uuid_text:
                cursor.execute("""
                    SELECT run_uuid::text, source, triggered_by, status, total_items, error_message, started_at, finished_at
                    FROM recruiting_parse_runs
                    WHERE status = 'success'
                    ORDER BY finished_at DESC, id DESC
                    LIMIT 1
                """)
                run_row = cursor.fetchone()
                if not run_row:
                    return {"run": None, "total": 0, "items": []}
                run_uuid_text = str(run_row[0])
            else:
                cursor.execute("""
                    SELECT run_uuid::text, source, triggered_by, status, total_items, error_message, started_at, finished_at
                    FROM recruiting_parse_runs
                    WHERE run_uuid = %s::uuid
                    LIMIT 1
                """, (run_uuid_text,))
                run_row = cursor.fetchone()
                if not run_row:
                    return {"run": None, "total": 0, "items": []}

            where_clauses = ["run_uuid = %s::uuid"]
            params = [run_uuid_text]

            keyword_group_norm = (str(keyword_group or '').strip() or None)
            if keyword_group_norm:
                where_clauses.append("keyword_group = %s")
                params.append(keyword_group_norm)

            search_text = (str(search or '').strip() or None)
            if search_text:
                like = f"%{search_text}%"
                where_clauses.append("""
                    (
                        COALESCE(title, '') ILIKE %s
                        OR COALESCE(category, '') ILIKE %s
                        OR COALESCE(location, '') ILIKE %s
                        OR COALESCE(experience, '') ILIKE %s
                        OR COALESCE(keyword_query, '') ILIKE %s
                        OR COALESCE(education, '') ILIKE %s
                    )
                """)
                params.extend([like, like, like, like, like, like])

            where_sql = " AND ".join(where_clauses)

            cursor.execute(f"""
                SELECT COUNT(*)
                FROM recruiting_resumes
                WHERE {where_sql}
            """, tuple(params))
            total_count = int(cursor.fetchone()[0] or 0)

            cursor.execute(
                f"""
                SELECT
                    id,
                    run_uuid::text,
                    keyword_group,
                    keyword_query,
                    page_found,
                    title,
                    category,
                    experience,
                    location,
                    salary,
                    education,
                    published_at,
                    detail_url,
                    scraped_at
                FROM recruiting_resumes
                WHERE {where_sql}
                ORDER BY COALESCE(page_found, 999999), id DESC
                LIMIT %s OFFSET %s
                """,
                tuple(params) + (limit_val, offset_val),
            )
            rows = cursor.fetchall()

        items = []
        for row in rows:
            items.append({
                "id": int(row[0]),
                "run_uuid": row[1],
                "keyword_group": row[2],
                "keyword_query": row[3],
                "page_found": row[4],
                "title": row[5],
                "category": row[6],
                "experience": row[7],
                "location": row[8],
                "salary": row[9],
                "education": row[10],
                "published_at": row[11],
                "detail_url": row[12],
                "scraped_at": row[13].isoformat() if hasattr(row[13], "isoformat") else row[13],
            })

        return {
            "run": self._serialize_recruiting_run_row(run_row),
            "total": total_count,
            "items": items,
        }

    def get_ai_feedback_cache(self, operator_id: int, month: str):
        """Получить кэшированный AI фидбэк для оператора за месяц"""
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT feedback_data, created_at, updated_at
                FROM ai_feedback_cache
                WHERE operator_id = %s AND month = %s
            """, (operator_id, month))
            result = cursor.fetchone()
            if result:
                return {
                    'feedback_data': result[0],
                    'created_at': result[1],
                    'updated_at': result[2]
                }
            return None

    def get_ai_birthday_greeting_cache(self, user_id: int, date_key: str):
        """Получить кэшированное AI-поздравление с днем рождения"""
        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT greeting_data, created_at, updated_at
                FROM ai_birthday_greeting_cache
                WHERE user_id = %s AND greeting_date = %s
            """, (user_id, date_key))
            result = cursor.fetchone()
            if result:
                return {
                    'greeting_data': result[0],
                    'created_at': result[1],
                    'updated_at': result[2]
                }
            return None

    def find_operator_by_name_fuzzy(self, name: str):
        """Попытаться найти оператора по ФИО. Сначала точное совпадение (case-insensitive),
        затем попытка по фамилии (ILIKE '%surname%'). Возвращает запись пользователя или None.
        """
        if not name:
            return None
        name = name.strip()
        with self._get_cursor() as cursor:
            # exact case-insensitive match
            cursor.execute("""
                SELECT id, telegram_id, name, role, direction_id, hire_date, supervisor_id, login
                FROM users WHERE LOWER(name) = LOWER(%s) AND role = 'operator' LIMIT 1
            """, (name,))
            row = cursor.fetchone()
            if row:
                return row

            # try surname match: take first token as likely surname
            tokens = re.split(r'\s+', name)
            if tokens:
                surname = tokens[0]
                cursor.execute("""
                    SELECT id, telegram_id, name, role, direction_id, hire_date, supervisor_id, login
                    FROM users WHERE name ILIKE %s AND role = 'operator' LIMIT 1
                """, (f"%{surname}%",))
                row = cursor.fetchone()
                if row:
                    return row

        return None

    def replace_baiga_scores_for_day(self, day, scores_map: dict):
        """Replace all baiga_daily_scores for given day.
        scores_map: { operator_id (int) : points (int) }
        If operator_id keys are not ints, they will be skipped.
        """
        if isinstance(day, str):
            try:
                day_date = datetime.strptime(day, "%Y-%m-%d").date()
            except Exception:
                raise ValueError("Invalid day format, expected YYYY-MM-DD")
        else:
            day_date = day

        with self._get_cursor() as cursor:
            # remove all existing for day
            cursor.execute("DELETE FROM baiga_daily_scores WHERE day = %s", (day_date,))
            insert_vals = []
            for op_id, pts in (scores_map or {}).items():
                try:
                    oid = int(op_id)
                    pts_i = int(pts) if pts is not None else 0
                    insert_vals.append((oid, day_date, pts_i))
                except Exception:
                    continue

            if insert_vals:
                cursor.executemany(
                    "INSERT INTO baiga_daily_scores (operator_id, day, points) VALUES (%s, %s, %s)",
                    insert_vals
                )

    def get_baiga_scores_for_day(self, day):
        """Return list of {operator_id, name, points} for given day."""
        if isinstance(day, str):
            try:
                day_date = datetime.strptime(day, "%Y-%m-%d").date()
            except Exception:
                raise ValueError("Invalid day format, expected YYYY-MM-DD")
        else:
            day_date = day

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT b.operator_id, b.points, u.name
                FROM baiga_daily_scores b
                LEFT JOIN users u ON u.id = b.operator_id
                WHERE b.day = %s
            """, (day_date,))
            rows = cursor.fetchall()
        return [ { 'operator_id': r[0], 'points': int(r[1] or 0), 'name': r[2] } for r in rows ]

    def get_baiga_scores_for_range(self, start_day, end_day):
        """Return mapping of day (YYYY-MM-DD) -> list of {operator_id, points, name} for days in [start_day, end_day]."""
        if isinstance(start_day, str):
            try:
                start_date = datetime.strptime(start_day, "%Y-%m-%d").date()
            except Exception:
                raise ValueError("Invalid start_day format, expected YYYY-MM-DD")
        else:
            start_date = start_day

        if isinstance(end_day, str):
            try:
                end_date = datetime.strptime(end_day, "%Y-%m-%d").date()
            except Exception:
                raise ValueError("Invalid end_day format, expected YYYY-MM-DD")
        else:
            end_date = end_day

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT to_char(b.day, 'YYYY-MM-DD') as day_str, b.operator_id, b.points, u.name
                FROM baiga_daily_scores b
                LEFT JOIN users u ON u.id = b.operator_id
                WHERE b.day >= %s AND b.day <= %s
                ORDER BY b.day, b.operator_id
            """, (start_date, end_date))
            rows = cursor.fetchall()

        result = {}
        for day_str, op_id, pts, name in rows:
            entry = { 'operator_id': op_id, 'points': int(pts or 0), 'name': name }
            result.setdefault(day_str, []).append(entry)
        return result

    @staticmethod
    def _survey_dt_to_iso(value):
        return value.isoformat() if hasattr(value, 'isoformat') else value

    @staticmethod
    def _normalize_survey_role(role):
        return normalize_role_value(role)

    def _get_visible_operator_ids_for_requester_tx(self, cursor, requester_id, requester_role):
        role = self._normalize_survey_role(requester_role)
        requester_id = int(requester_id)

        if role_has_min(role, 'admin') or role == 'trainer':
            cursor.execute("""
                SELECT id
                FROM users
                WHERE LOWER(TRIM(COALESCE(role, ''))) = 'operator'
                  AND COALESCE(NULLIF(LOWER(TRIM(COALESCE(status, 'working'))), ''), 'working') <> 'fired'
            """)
            return [int(row[0]) for row in cursor.fetchall()]

        if role == 'sv':
            cursor.execute("""
                SELECT id
                FROM users
                WHERE LOWER(TRIM(COALESCE(role, ''))) = 'operator'
                  AND COALESCE(NULLIF(LOWER(TRIM(COALESCE(status, 'working'))), ''), 'working') NOT IN ('fired', 'dismissal')
            """)
            return [int(row[0]) for row in cursor.fetchall()]

        if role == 'operator':
            return [requester_id]

        return []

    def get_visible_operator_ids_for_requester(self, requester_id, requester_role):
        with self._get_cursor() as cursor:
            return self._get_visible_operator_ids_for_requester_tx(cursor, requester_id, requester_role)

    def _get_visible_lms_learner_ids_for_requester_tx(self, cursor, requester_id, requester_role):
        role = self._normalize_survey_role(requester_role)
        requester_id = int(requester_id)

        if role_has_min(role, 'admin') or role == 'trainer':
            cursor.execute("""
                SELECT id
                FROM users
                WHERE role IN ('operator', 'trainee')
                  AND COALESCE(status, 'working') <> 'fired'
            """)
            return [int(row[0]) for row in cursor.fetchall()]

        if role == 'sv':
            cursor.execute("""
                SELECT id
                FROM users
                WHERE role IN ('operator', 'trainee')
                  AND supervisor_id = %s
                  AND COALESCE(status, 'working') <> 'fired'
            """, (requester_id,))
            return [int(row[0]) for row in cursor.fetchall()]

        if role in ('operator', 'trainee'):
            return [requester_id]

        return []

    def get_visible_lms_learner_ids_for_requester(self, requester_id, requester_role):
        with self._get_cursor() as cursor:
            return self._get_visible_lms_learner_ids_for_requester_tx(cursor, requester_id, requester_role)

    def create_survey(self, title, description, created_by, assignment, questions, operator_ids, repeat_from_survey_id=None, is_test=False):
        title_norm = str(title or '').strip()
        description_norm = str(description or '').strip() or None
        created_by_id = int(created_by) if created_by is not None else None
        if isinstance(is_test, str):
            is_test_norm = is_test.strip().lower() in ('1', 'true', 'yes', 'on')
        else:
            is_test_norm = bool(is_test)

        if not title_norm:
            raise ValueError("SURVEY_TITLE_REQUIRED")
        if not isinstance(questions, list) or len(questions) == 0:
            raise ValueError("SURVEY_QUESTIONS_REQUIRED")

        operator_ids_norm = []
        for op_id in operator_ids or []:
            try:
                parsed = int(op_id)
            except Exception:
                continue
            if parsed > 0 and parsed not in operator_ids_norm:
                operator_ids_norm.append(parsed)
        if not operator_ids_norm:
            raise ValueError("SURVEY_OPERATORS_REQUIRED")

        assignment = assignment or {}
        direction_ids_norm = []
        for direction_id in assignment.get('direction_ids') or []:
            try:
                parsed = int(direction_id)
            except Exception:
                continue
            if parsed > 0 and parsed not in direction_ids_norm:
                direction_ids_norm.append(parsed)

        tenure_weeks_min = assignment.get('tenure_weeks_min')
        tenure_weeks_max = assignment.get('tenure_weeks_max')

        try:
            tenure_weeks_min = int(tenure_weeks_min) if tenure_weeks_min is not None else None
        except Exception:
            raise ValueError("SURVEY_INVALID_TENURE_MIN")
        try:
            tenure_weeks_max = int(tenure_weeks_max) if tenure_weeks_max is not None else None
        except Exception:
            raise ValueError("SURVEY_INVALID_TENURE_MAX")

        if tenure_weeks_min is not None and tenure_weeks_min < 0:
            raise ValueError("SURVEY_INVALID_TENURE_MIN")
        if tenure_weeks_max is not None and tenure_weeks_max < 0:
            raise ValueError("SURVEY_INVALID_TENURE_MAX")
        if tenure_weeks_min is not None and tenure_weeks_max is not None and tenure_weeks_min > tenure_weeks_max:
            raise ValueError("SURVEY_INVALID_TENURE_RANGE")

        normalized_questions = []
        for idx, raw_question in enumerate(questions):
            text = str((raw_question or {}).get('text') or '').strip()
            qtype = str((raw_question or {}).get('type') or 'single').strip().lower()
            required = bool((raw_question or {}).get('required', True))
            allow_other = bool((raw_question or {}).get('allow_other', False))

            if not text:
                raise ValueError(f"SURVEY_QUESTION_TEXT_REQUIRED_{idx + 1}")
            if qtype not in ('single', 'multiple', 'rating'):
                raise ValueError(f"SURVEY_INVALID_QUESTION_TYPE_{idx + 1}")
            if is_test_norm and qtype == 'rating':
                raise ValueError(f"SURVEY_TEST_RATING_NOT_ALLOWED_{idx + 1}")

            options_norm = []
            correct_options_norm = []
            if qtype != 'rating':
                if is_test_norm:
                    allow_other = False

                for option in (raw_question or {}).get('options') or []:
                    option_text = str(option or '').strip()
                    if option_text and option_text not in options_norm:
                        options_norm.append(option_text)

                is_other_only_question = (qtype == 'single' and allow_other and len(options_norm) == 0)
                if len(options_norm) < 2 and not is_other_only_question:
                    raise ValueError(f"SURVEY_OPTIONS_REQUIRED_{idx + 1}")

                for option in (raw_question or {}).get('correct_options') or []:
                    option_text = str(option or '').strip()
                    if option_text and option_text not in correct_options_norm:
                        correct_options_norm.append(option_text)

                if is_test_norm:
                    if not correct_options_norm:
                        raise ValueError(f"SURVEY_CORRECT_OPTIONS_REQUIRED_{idx + 1}")
                    invalid_correct = [item for item in correct_options_norm if item not in options_norm]
                    if invalid_correct:
                        raise ValueError(f"SURVEY_CORRECT_OPTION_INVALID_{idx + 1}")
                    if qtype == 'single' and len(correct_options_norm) != 1:
                        raise ValueError(f"SURVEY_SINGLE_CORRECT_OPTION_REQUIRED_{idx + 1}")
                elif is_other_only_question:
                    correct_options_norm = []
            else:
                allow_other = False

            normalized_questions.append({
                'position': idx + 1,
                'text': text,
                'type': qtype,
                'required': required,
                'allow_other': allow_other,
                'options': options_norm,
                'correct_options': correct_options_norm
            })

        with self._get_cursor() as cursor:
            repeat_root_id_to_use = None
            repeat_iteration_to_use = 1
            if repeat_from_survey_id is not None:
                try:
                    repeat_from_survey_id = int(repeat_from_survey_id)
                except Exception:
                    raise ValueError("SURVEY_REPEAT_SOURCE_NOT_FOUND")

                cursor.execute("""
                    SELECT id, COALESCE(repeat_root_id, id), COALESCE(repeat_iteration, 1)
                    FROM surveys
                    WHERE id = %s
                """, (repeat_from_survey_id,))
                repeat_source_row = cursor.fetchone()
                if not repeat_source_row:
                    raise ValueError("SURVEY_REPEAT_SOURCE_NOT_FOUND")

                repeat_root_id_to_use = int(repeat_source_row[1])
                cursor.execute("""
                    SELECT COALESCE(MAX(repeat_iteration), 1)
                    FROM surveys
                    WHERE repeat_root_id = %s OR id = %s
                """, (repeat_root_id_to_use, repeat_root_id_to_use))
                max_iteration_row = cursor.fetchone()
                current_max_iteration = int(max_iteration_row[0] or 1) if max_iteration_row else 1
                repeat_iteration_to_use = max(1, current_max_iteration + 1)

            cursor.execute("""
                INSERT INTO surveys (
                    title,
                    description,
                    created_by,
                    direction_ids,
                    tenure_weeks_min,
                    tenure_weeks_max,
                    repeat_root_id,
                    repeat_iteration,
                    is_test,
                    created_at,
                    updated_at
                )
                VALUES (
                    %s,
                    %s,
                    %s,
                    %s::jsonb,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                    (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                )
                RETURNING id, created_at
            """, (
                title_norm,
                description_norm,
                created_by_id,
                json.dumps(direction_ids_norm),
                tenure_weeks_min,
                tenure_weeks_max,
                repeat_root_id_to_use,
                repeat_iteration_to_use,
                is_test_norm
            ))
            survey_id, created_at = cursor.fetchone()

            if repeat_root_id_to_use is None:
                repeat_root_id_to_use = int(survey_id)
                cursor.execute("""
                    UPDATE surveys
                    SET repeat_root_id = %s
                    WHERE id = %s
                """, (repeat_root_id_to_use, survey_id))
                repeat_iteration_to_use = 1

            for question in normalized_questions:
                cursor.execute("""
                    INSERT INTO survey_questions (
                        survey_id,
                    position,
                    question_text,
                    question_type,
                    is_required,
                    allow_other,
                    options_json,
                    correct_options_json
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb)
                """, (
                    survey_id,
                    question['position'],
                    question['text'],
                    question['type'],
                    question['required'],
                    question['allow_other'],
                    json.dumps(question['options'], ensure_ascii=False),
                    json.dumps(question['correct_options'], ensure_ascii=False)
                ))

            for operator_id in operator_ids_norm:
                cursor.execute("""
                    INSERT INTO survey_assignments (
                        survey_id,
                        operator_id,
                        assigned_by,
                        status,
                        assigned_at
                    )
                    VALUES (
                        %s,
                        %s,
                        %s,
                        'assigned',
                        (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    )
                    ON CONFLICT (survey_id, operator_id) DO UPDATE
                    SET
                        assigned_by = EXCLUDED.assigned_by,
                        status = CASE
                            WHEN survey_assignments.status = 'completed' THEN survey_assignments.status
                            ELSE 'assigned'
                        END,
                        completed_at = CASE
                            WHEN survey_assignments.status = 'completed' THEN survey_assignments.completed_at
                            ELSE NULL
                        END,
                        assigned_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                """, (survey_id, operator_id, created_by_id))

            return {
                'id': int(survey_id),
                'created_at': self._survey_dt_to_iso(created_at),
                'repeat_root_id': int(repeat_root_id_to_use),
                'repeat_iteration': int(repeat_iteration_to_use),
                'is_test': bool(is_test_norm)
            }

    def delete_survey(self, survey_id, requester_id, requester_role):
        survey_id = int(survey_id)
        requester_id = int(requester_id)
        role = self._normalize_survey_role(requester_role)

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, created_by
                FROM surveys
                WHERE id = %s
            """, (survey_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("SURVEY_NOT_FOUND")

            created_by = row[1]
            if (not role_has_min(role, 'admin')) and (created_by is None or int(created_by) != requester_id):
                raise PermissionError("SURVEY_FORBIDDEN")

            cursor.execute("DELETE FROM surveys WHERE id = %s", (survey_id,))
            return True

    def _build_survey_question_stats(self, questions, answers_for_survey, respondents_total=0):
        try:
            survey_respondents_total = int(respondents_total or 0)
        except Exception:
            survey_respondents_total = 0
        survey_respondents_total = max(0, survey_respondents_total)

        answers_by_question = defaultdict(list)
        for answer in answers_for_survey:
            try:
                question_id = int(answer.get('question_id'))
            except Exception:
                continue
            answers_by_question[question_id].append(answer)

        stats = []
        for question in questions:
            question_id = int(question['id'])
            qtype = question.get('type')
            question_text = str(question.get('text') or '')
            correct_options = [str(item) for item in (question.get('correct_options') or [])]
            question_answers = answers_by_question.get(question_id, [])
            answered_count = 0

            if qtype == 'rating':
                values = []
                distribution = {str(i): 0 for i in range(1, 6)}
                for answer in question_answers:
                    try:
                        value = int(answer.get('rating_value'))
                    except Exception:
                        continue
                    if 1 <= value <= 5:
                        answered_count += 1
                        values.append(value)
                        distribution[str(value)] += 1

                sorted_values = sorted(values)
                median_rating = None
                if sorted_values:
                    middle = len(sorted_values) // 2
                    if len(sorted_values) % 2:
                        median_rating = float(sorted_values[middle])
                    else:
                        median_rating = round((sorted_values[middle - 1] + sorted_values[middle]) / 2, 2)

                detailed_distribution = []
                for score in range(1, 6):
                    count = int(distribution[str(score)])
                    detailed_distribution.append({
                        'value': score,
                        'count': count,
                        'percent_of_answers': round((count / answered_count) * 100, 1) if answered_count > 0 else 0.0,
                        'percent_of_respondents': round((count / survey_respondents_total) * 100, 1) if survey_respondents_total > 0 else 0.0
                    })

                stats.append({
                    'question_id': question_id,
                    'text': question_text,
                    'type': qtype,
                    'correct_options': correct_options,
                    'respondents_total': survey_respondents_total,
                    'question_respondents_total': answered_count,
                    'survey_respondents_total': survey_respondents_total,
                    'responses_with_answer': answered_count,
                    'skipped_count': max(0, survey_respondents_total - answered_count),
                    'response_rate': round((answered_count / survey_respondents_total) * 100, 1) if survey_respondents_total > 0 else 0.0,
                    'average_rating': round(sum(values) / len(values), 2) if values else None,
                    'median_rating': median_rating,
                    'min_rating': min(values) if values else None,
                    'max_rating': max(values) if values else None,
                    'ratings_distribution': distribution,
                    'ratings_distribution_detailed': detailed_distribution
                })
                continue

            option_counts = defaultdict(int)
            other_count = 0
            selections_total = 0

            for answer in question_answers:
                selected_values = []
                for selected in answer.get('selected_options') or []:
                    selected_text = str(selected or '').strip()
                    if selected_text:
                        selected_values.append(selected_text)

                answer_text = str(answer.get('answer_text') or '').strip()
                if selected_values or answer_text:
                    answered_count += 1

                for selected_value in selected_values:
                    option_counts[selected_value] += 1
                    selections_total += 1
                if answer_text:
                    other_count += 1

            options_stats = []
            for option in question.get('options') or []:
                count = int(option_counts.get(option, 0))
                options_stats.append({
                    'option': option,
                    'count': count,
                    'percent': round((count / answered_count) * 100, 1) if answered_count > 0 else 0.0,
                    'percent_of_answers': round((count / answered_count) * 100, 1) if answered_count > 0 else 0.0,
                    'percent_of_respondents': round((count / survey_respondents_total) * 100, 1) if survey_respondents_total > 0 else 0.0,
                    'percent_of_selections': round((count / selections_total) * 100, 1) if selections_total > 0 else 0.0
                })
            if question.get('allow_other'):
                options_stats.append({
                    'option': 'Другое',
                    'count': int(other_count),
                    'is_other': True,
                    'percent': round((other_count / answered_count) * 100, 1) if answered_count > 0 else 0.0,
                    'percent_of_answers': round((other_count / answered_count) * 100, 1) if answered_count > 0 else 0.0,
                    'percent_of_respondents': round((other_count / survey_respondents_total) * 100, 1) if survey_respondents_total > 0 else 0.0,
                    'percent_of_selections': None
                })

            top_options = [
                option for option in sorted(
                    [item for item in options_stats if int(item.get('count') or 0) > 0],
                    key=lambda item: (-int(item.get('count') or 0), str(item.get('option') or ''))
                )
            ][:3]

            stats.append({
                'question_id': question_id,
                'text': question_text,
                'type': qtype,
                'correct_options': correct_options,
                'respondents_total': survey_respondents_total,
                'question_respondents_total': answered_count,
                'survey_respondents_total': survey_respondents_total,
                'responses_with_answer': answered_count,
                'skipped_count': max(0, survey_respondents_total - answered_count),
                'response_rate': round((answered_count / survey_respondents_total) * 100, 1) if survey_respondents_total > 0 else 0.0,
                'selections_total': int(selections_total),
                'options': options_stats,
                'top_options': top_options
            })

        return stats

    def _serialize_survey_list(self, surveys_rows, questions_rows, assignments_rows, responses_rows, answers_rows):
        questions_by_survey = defaultdict(list)
        question_positions_by_survey = defaultdict(dict)
        for row in questions_rows:
            survey_id = int(row[1])
            question_id = int(row[0])
            position = int(row[2])
            options = row[7] if isinstance(row[7], list) else []
            correct_options = row[8] if len(row) > 8 and isinstance(row[8], list) else []
            question_obj = {
                'id': question_id,
                'position': position,
                'text': row[3],
                'type': row[4],
                'required': bool(row[5]),
                'allow_other': bool(row[6]),
                'options': [str(item) for item in options],
                'correct_options': [str(item) for item in correct_options]
            }
            questions_by_survey[survey_id].append(question_obj)
            question_positions_by_survey[survey_id][question_id] = position

        assignments_by_survey = defaultdict(list)
        for row in assignments_rows:
            assignments_by_survey[int(row[1])].append({
                'id': int(row[0]),
                'operator_id': int(row[2]),
                'operator_name': row[3] or f"#{row[2]}",
                'status': row[4],
                'assigned_at': self._survey_dt_to_iso(row[5]),
                'completed_at': self._survey_dt_to_iso(row[6])
            })

        responses_by_survey = defaultdict(list)
        for row in responses_rows:
            responses_by_survey[int(row[1])].append({
                'id': int(row[0]),
                'operator_id': int(row[2]),
                'submitted_at': self._survey_dt_to_iso(row[3])
            })

        answers_by_survey = defaultdict(list)
        answers_by_response = defaultdict(list)
        for row in answers_rows:
            response_id = None
            if len(row) >= 6:
                response_id = int(row[0])
                survey_id = int(row[1])
                question_id = int(row[2])
                selected_options = row[3] if isinstance(row[3], list) else []
                answer_text = row[4] or ''
                rating_value = row[5]
            else:
                survey_id = int(row[0])
                question_id = int(row[1])
                selected_options = row[2] if isinstance(row[2], list) else []
                answer_text = row[3] or ''
                rating_value = row[4]

            answer_obj = {
                'question_id': question_id,
                'selected_options': [str(item) for item in selected_options],
                'answer_text': answer_text,
                'rating_value': rating_value
            }
            answers_by_survey[survey_id].append(answer_obj)
            if response_id is not None:
                answers_by_response[response_id].append(answer_obj)

        serialized = []
        for row in surveys_rows:
            survey_id = int(row[0])
            repeat_root_id = survey_id
            repeat_iteration = 1
            is_test = False
            if len(row) > 12:
                try:
                    repeat_root_id = int(row[12]) if row[12] is not None else survey_id
                except Exception:
                    repeat_root_id = survey_id
            if len(row) > 13:
                try:
                    parsed_repeat_iteration = int(row[13]) if row[13] is not None else 1
                    repeat_iteration = parsed_repeat_iteration if parsed_repeat_iteration >= 1 else 1
                except Exception:
                    repeat_iteration = 1
            if len(row) > 14:
                is_test = bool(row[14])

            direction_ids = row[6] if isinstance(row[6], list) else []
            questions = sorted(questions_by_survey.get(survey_id, []), key=lambda item: (item['position'], item['id']))
            assignments = assignments_by_survey.get(survey_id, [])
            responses = responses_by_survey.get(survey_id, [])
            question_positions = question_positions_by_survey.get(survey_id, {})
            test_question_meta_by_id = {}
            test_total_questions = 0
            if is_test:
                for question in questions:
                    question_id = int(question.get('id') or 0)
                    question_type = str(question.get('type') or '')
                    correct_options = [str(item) for item in (question.get('correct_options') or [])]
                    test_question_meta_by_id[question_id] = {
                        'type': question_type,
                        'correct_options': correct_options
                    }
                    if question_type in ('single', 'multiple'):
                        test_total_questions += 1

            assigned_count = len(assignments)
            completed_count = sum(1 for assignment in assignments if assignment.get('status') == 'completed')
            pending_count = max(0, assigned_count - completed_count)
            completion_rate = round((completed_count / assigned_count) * 100, 1) if assigned_count > 0 else 0.0

            normalized_direction_ids = []
            for value in direction_ids:
                try:
                    parsed = int(value)
                except Exception:
                    continue
                if parsed not in normalized_direction_ids:
                    normalized_direction_ids.append(parsed)

            latest_response_by_operator = {}
            for response in responses:
                operator_id = int(response.get('operator_id'))
                if operator_id not in latest_response_by_operator:
                    latest_response_by_operator[operator_id] = response

            responses_detailed = []
            for assignment in assignments:
                operator_id = int(assignment.get('operator_id'))
                operator_response = latest_response_by_operator.get(operator_id)
                response_id = int(operator_response['id']) if operator_response else None
                response_answers = list(answers_by_response.get(response_id, [])) if response_id is not None else []
                response_answers.sort(
                    key=lambda item: (
                        question_positions.get(int(item.get('question_id') or 0), 10 ** 9),
                        int(item.get('question_id') or 0)
                    )
                )

                answers_by_question = {}
                normalized_answers = []
                test_answered_questions = 0
                test_correct_answers = 0

                for answer in response_answers:
                    question_id = int(answer.get('question_id'))
                    selected_options = list(answer.get('selected_options') or [])
                    answer_text = str(answer.get('answer_text') or '')
                    answer_payload = {
                        'question_id': question_id,
                        'selected_options': selected_options,
                        'answer_text': answer_text,
                        'rating_value': answer.get('rating_value')
                    }
                    if is_test:
                        question_meta = test_question_meta_by_id.get(question_id) or {}
                        expected_options = list(question_meta.get('correct_options') or [])
                        question_type = str(question_meta.get('type') or '')
                        has_answer = bool(selected_options or answer_text or answer.get('rating_value') is not None)
                        if has_answer:
                            test_answered_questions += 1

                        is_correct_answer = False
                        if question_type == 'single':
                            is_correct_answer = (
                                len(expected_options) == 1
                                and len(selected_options) == 1
                                and selected_options[0] == expected_options[0]
                                and not answer_text
                            )
                        elif question_type == 'multiple':
                            is_correct_answer = (
                                len(expected_options) > 0
                                and set(selected_options) == set(expected_options)
                                and not answer_text
                            )
                        if is_correct_answer:
                            test_correct_answers += 1

                        answer_payload['is_correct'] = bool(is_correct_answer)
                        answer_payload['expected_options'] = expected_options

                    answers_by_question[str(question_id)] = answer_payload
                    normalized_answers.append(answer_payload)

                test_summary = None
                if is_test:
                    score_percent = None
                    if response_id is not None:
                        score_percent = round((test_correct_answers / test_total_questions) * 100, 1) if test_total_questions > 0 else 0.0
                    test_summary = {
                        'total_questions': int(test_total_questions),
                        'answered_questions': int(test_answered_questions),
                        'correct_answers': int(test_correct_answers),
                        'score_percent': float(score_percent) if score_percent is not None else None
                    }

                responses_detailed.append({
                    'operator_id': operator_id,
                    'operator_name': assignment.get('operator_name') or f"#{operator_id}",
                    'status': assignment.get('status') or 'assigned',
                    'assigned_at': assignment.get('assigned_at'),
                    'completed_at': assignment.get('completed_at'),
                    'response_id': response_id,
                    'submitted_at': operator_response.get('submitted_at') if operator_response else None,
                    'repeat_iteration': int(repeat_iteration),
                    'repeat_root_id': int(repeat_root_id),
                    'repeat_survey_id': int(survey_id),
                    'answers': normalized_answers,
                    'answers_by_question': answers_by_question,
                    'test_summary': test_summary
                })

            serialized.append({
                'id': survey_id,
                'title': row[1],
                'description': row[2] or '',
                'created_at': self._survey_dt_to_iso(row[3]),
                'updated_at': self._survey_dt_to_iso(row[4]),
                'created_by': {
                    'id': row[9],
                    'name': row[10] or 'Система',
                    'role': row[11] or 'unknown'
                },
                'repeat': {
                    'root_id': int(repeat_root_id),
                    'iteration': int(repeat_iteration),
                    'is_repeat': int(repeat_iteration) > 1
                },
                'is_test': bool(is_test),
                'assignment': {
                    'direction_ids': normalized_direction_ids,
                    'tenure_weeks_min': row[7],
                    'tenure_weeks_max': row[8],
                    'operator_ids': [int(assignment['operator_id']) for assignment in assignments],
                    'operators': assignments
                },
                'questions': questions,
                'statistics': {
                    'assigned_count': assigned_count,
                    'completed_count': completed_count,
                    'pending_count': pending_count,
                    'responses_count': len(responses),
                    'completion_rate': completion_rate,
                    'question_stats': self._build_survey_question_stats(
                        questions,
                        answers_by_survey.get(survey_id, []),
                        respondents_total=len(responses)
                    ),
                    'responses_detailed': responses_detailed
                }
            })

        grouped_repetition_rows = defaultdict(list)
        for survey in serialized:
            repeat_info = survey.get('repeat') or {}
            try:
                root_id = int(repeat_info.get('root_id') or survey.get('id'))
            except Exception:
                root_id = int(survey.get('id'))
            repeat_iteration = int(repeat_info.get('iteration') or 1)

            for response_row in (survey.get('statistics') or {}).get('responses_detailed') or []:
                row_copy = dict(response_row)
                row_copy['repeat_iteration'] = int(repeat_iteration)
                row_copy['repeat_root_id'] = int(root_id)
                row_copy['repeat_survey_id'] = int(survey.get('id'))
                row_copy['repeat_survey_title'] = survey.get('title') or f"#{survey.get('id')}"
                grouped_repetition_rows[root_id].append(row_copy)

        for root_id, rows in grouped_repetition_rows.items():
            rows.sort(key=lambda item: (
                int(item.get('repeat_iteration') or 1),
                str(item.get('operator_name') or '').lower(),
                int(item.get('operator_id') or 0)
            ))

        for survey in serialized:
            repeat_info = survey.get('repeat') or {}
            try:
                root_id = int(repeat_info.get('root_id') or survey.get('id'))
            except Exception:
                root_id = int(survey.get('id'))
            survey.setdefault('statistics', {})['responses_detailed_all_repetitions'] = list(grouped_repetition_rows.get(root_id, []))

        return serialized

    def get_surveys_for_management(self, requester_id, requester_role):
        requester_id = int(requester_id)
        role = self._normalize_survey_role(requester_role)

        if not (role_has_min(role, 'admin') or role in ('sv', 'trainer')):
            return []

        with self._get_cursor() as cursor:
            if role_has_min(role, 'admin') or role == 'trainer':
                cursor.execute("""
                    SELECT
                        s.id,
                        s.title,
                        s.description,
                        s.created_at,
                        s.updated_at,
                        s.created_by,
                        s.direction_ids,
                        s.tenure_weeks_min,
                        s.tenure_weeks_max,
                        creator.id,
                        creator.name,
                        creator.role,
                        s.repeat_root_id,
                        s.repeat_iteration,
                        s.is_test
                    FROM surveys s
                    LEFT JOIN users creator ON creator.id = s.created_by
                    ORDER BY s.created_at DESC, s.id DESC
                """)
            else:
                cursor.execute("""
                    SELECT DISTINCT
                        s.id,
                        s.title,
                        s.description,
                        s.created_at,
                        s.updated_at,
                        s.created_by,
                        s.direction_ids,
                        s.tenure_weeks_min,
                        s.tenure_weeks_max,
                        creator.id,
                        creator.name,
                        creator.role,
                        s.repeat_root_id,
                        s.repeat_iteration,
                        s.is_test
                    FROM surveys s
                    LEFT JOIN users creator ON creator.id = s.created_by
                    LEFT JOIN survey_assignments sa ON sa.survey_id = s.id
                    LEFT JOIN users op ON op.id = sa.operator_id
                    WHERE s.created_by = %s OR op.supervisor_id = %s
                    ORDER BY s.created_at DESC, s.id DESC
                """, (requester_id, requester_id))

            surveys_rows = cursor.fetchall()
            if not surveys_rows:
                return []

            survey_ids = [int(row[0]) for row in surveys_rows]
            visible_operator_ids = None
            if role == 'sv':
                visible_operator_ids = set(self._get_visible_operator_ids_for_requester_tx(cursor, requester_id, role))

            cursor.execute("""
                SELECT
                    q.id,
                    q.survey_id,
                    q.position,
                    q.question_text,
                    q.question_type,
                    q.is_required,
                    q.allow_other,
                    q.options_json,
                    q.correct_options_json
                FROM survey_questions q
                WHERE q.survey_id = ANY(%s)
                ORDER BY q.survey_id, q.position, q.id
            """, (survey_ids,))
            questions_rows = cursor.fetchall()

            cursor.execute("""
                SELECT
                    sa.id,
                    sa.survey_id,
                    sa.operator_id,
                    u.name,
                    sa.status,
                    sa.assigned_at,
                    sa.completed_at
                FROM survey_assignments sa
                LEFT JOIN users u ON u.id = sa.operator_id
                WHERE sa.survey_id = ANY(%s)
                ORDER BY sa.survey_id, u.name, sa.id
            """, (survey_ids,))
            assignments_rows = cursor.fetchall()

            if visible_operator_ids is not None:
                assignments_rows = [row for row in assignments_rows if int(row[2]) in visible_operator_ids]

            cursor.execute("""
                SELECT
                    sr.id,
                    sr.survey_id,
                    sr.operator_id,
                    sr.submitted_at
                FROM survey_responses sr
                WHERE sr.survey_id = ANY(%s)
                ORDER BY sr.submitted_at DESC, sr.id DESC
            """, (survey_ids,))
            responses_rows = cursor.fetchall()

            if visible_operator_ids is not None:
                responses_rows = [row for row in responses_rows if int(row[2]) in visible_operator_ids]

            response_ids = [int(row[0]) for row in responses_rows]
            answers_rows = []
            if response_ids:
                cursor.execute("""
                    SELECT
                        a.response_id,
                        r.survey_id,
                        a.question_id,
                        a.selected_options_json,
                        a.answer_text,
                        a.rating_value
                    FROM survey_answers a
                    JOIN survey_responses r ON r.id = a.response_id
                    WHERE a.response_id = ANY(%s)
                """, (response_ids,))
                answers_rows = cursor.fetchall()

            return self._serialize_survey_list(
                surveys_rows=surveys_rows,
                questions_rows=questions_rows,
                assignments_rows=assignments_rows,
                responses_rows=responses_rows,
                answers_rows=answers_rows
            )

    def get_surveys_for_operator(self, operator_id):
        operator_id = int(operator_id)

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    s.id,
                    s.title,
                    s.description,
                    s.created_at,
                    s.updated_at,
                    s.created_by,
                    s.direction_ids,
                    s.tenure_weeks_min,
                    s.tenure_weeks_max,
                    creator.id,
                    creator.name,
                    creator.role,
                    sa.id,
                    sa.status,
                    sa.assigned_at,
                    sa.completed_at,
                    sr.id,
                    sr.submitted_at,
                    s.repeat_root_id,
                    s.repeat_iteration,
                    s.is_test
                FROM survey_assignments sa
                JOIN surveys s ON s.id = sa.survey_id
                LEFT JOIN users creator ON creator.id = s.created_by
                LEFT JOIN survey_responses sr
                    ON sr.survey_id = sa.survey_id
                   AND sr.operator_id = sa.operator_id
                WHERE sa.operator_id = %s
                ORDER BY
                    CASE WHEN sa.status = 'completed' THEN 1 ELSE 0 END,
                    sa.assigned_at DESC,
                    sa.id DESC
            """, (operator_id,))
            surveys_rows = cursor.fetchall()
            if not surveys_rows:
                return []

            survey_ids = [int(row[0]) for row in surveys_rows]
            cursor.execute("""
                SELECT
                    q.id,
                    q.survey_id,
                    q.position,
                    q.question_text,
                    q.question_type,
                    q.is_required,
                    q.allow_other,
                    q.options_json,
                    q.correct_options_json
                FROM survey_questions q
                WHERE q.survey_id = ANY(%s)
                ORDER BY q.survey_id, q.position, q.id
            """, (survey_ids,))
            question_rows = cursor.fetchall()

            questions_by_survey = defaultdict(list)
            question_positions_by_survey = defaultdict(dict)
            for row in question_rows:
                survey_id = int(row[1])
                question_id = int(row[0])
                position = int(row[2])
                options = row[7] if isinstance(row[7], list) else []
                correct_options = row[8] if len(row) > 8 and isinstance(row[8], list) else []
                questions_by_survey[survey_id].append({
                    'id': question_id,
                    'position': position,
                    'text': row[3],
                    'type': row[4],
                    'required': bool(row[5]),
                    'allow_other': bool(row[6]),
                    'options': [str(item) for item in options],
                    'correct_options': [str(item) for item in correct_options]
                })
                question_positions_by_survey[survey_id][question_id] = position

            response_ids = [int(row[16]) for row in surveys_rows if row[16] is not None]
            answers_by_response = defaultdict(list)
            if response_ids:
                cursor.execute("""
                    SELECT
                        a.response_id,
                        a.question_id,
                        a.selected_options_json,
                        a.answer_text,
                        a.rating_value
                    FROM survey_answers a
                    WHERE a.response_id = ANY(%s)
                """, (response_ids,))
                for answer_row in cursor.fetchall():
                    selected_options = answer_row[2] if isinstance(answer_row[2], list) else []
                    answers_by_response[int(answer_row[0])].append({
                        'question_id': int(answer_row[1]),
                        'selected_options': [str(item) for item in selected_options],
                        'answer_text': answer_row[3] or '',
                        'rating_value': answer_row[4]
                    })

            result = []
            for row in surveys_rows:
                survey_id = int(row[0])
                repeat_root_id = survey_id
                repeat_iteration = 1
                is_test = bool(row[20]) if len(row) > 20 else False
                if len(row) > 18:
                    try:
                        repeat_root_id = int(row[18]) if row[18] is not None else survey_id
                    except Exception:
                        repeat_root_id = survey_id
                if len(row) > 19:
                    try:
                        parsed_repeat_iteration = int(row[19]) if row[19] is not None else 1
                        repeat_iteration = parsed_repeat_iteration if parsed_repeat_iteration >= 1 else 1
                    except Exception:
                        repeat_iteration = 1

                questions_source = sorted(
                    questions_by_survey.get(survey_id, []),
                    key=lambda item: (item.get('position', 0), item.get('id', 0))
                )
                question_positions = question_positions_by_survey.get(survey_id, {})
                direction_ids = row[6] if isinstance(row[6], list) else []
                normalized_direction_ids = []
                for value in direction_ids:
                    try:
                        parsed = int(value)
                    except Exception:
                        continue
                    if parsed not in normalized_direction_ids:
                        normalized_direction_ids.append(parsed)

                assignment_status = str(row[13] or 'assigned')
                questions = []
                for question in questions_source:
                    question_copy = {
                        'id': int(question.get('id')),
                        'position': int(question.get('position') or 0),
                        'text': str(question.get('text') or ''),
                        'type': str(question.get('type') or ''),
                        'required': bool(question.get('required')),
                        'allow_other': bool(question.get('allow_other')),
                        'options': list(question.get('options') or [])
                    }
                    if is_test and assignment_status == 'completed':
                        question_copy['correct_options'] = list(question.get('correct_options') or [])
                    questions.append(question_copy)

                response_id = int(row[16]) if row[16] is not None else None
                my_answers = list(answers_by_response.get(response_id, [])) if response_id is not None else []
                my_answers.sort(
                    key=lambda item: (
                        question_positions.get(int(item.get('question_id') or 0), 10 ** 9),
                        int(item.get('question_id') or 0)
                    )
                )

                my_answers_by_question = {}
                test_total_questions = 0
                test_answered_questions = 0
                test_correct_answers = 0
                question_meta_by_id = {int(item.get('id')): item for item in questions_source}
                if is_test:
                    test_total_questions = len([item for item in questions_source if item.get('type') in ('single', 'multiple')])
                for answer in my_answers:
                    question_id = int(answer.get('question_id'))
                    answer_selected_options = list(answer.get('selected_options') or [])
                    answer_text_value = str(answer.get('answer_text') or '')
                    answer_payload = {
                        'question_id': question_id,
                        'selected_options': answer_selected_options,
                        'answer_text': answer_text_value,
                        'rating_value': answer.get('rating_value')
                    }
                    if is_test:
                        question_meta = question_meta_by_id.get(question_id) or {}
                        expected_options = list(question_meta.get('correct_options') or [])
                        question_type = str(question_meta.get('type') or '')
                        has_answer = bool(answer_selected_options or answer_text_value or answer.get('rating_value') is not None)
                        if has_answer:
                            test_answered_questions += 1

                        is_correct_answer = False
                        if question_type == 'single':
                            is_correct_answer = (
                                len(expected_options) == 1
                                and len(answer_selected_options) == 1
                                and answer_selected_options[0] == expected_options[0]
                                and not answer_text_value
                            )
                        elif question_type == 'multiple':
                            is_correct_answer = (
                                set(answer_selected_options) == set(expected_options)
                                and len(expected_options) > 0
                                and not answer_text_value
                            )
                        if is_correct_answer:
                            test_correct_answers += 1

                        answer_payload['is_correct'] = bool(is_correct_answer)
                        answer_payload['expected_options'] = expected_options

                    my_answers_by_question[str(question_id)] = {
                        **answer_payload
                    }

                my_response = None
                if response_id is not None:
                    my_response = {
                        'id': response_id,
                        'submitted_at': self._survey_dt_to_iso(row[17]),
                        'answers': my_answers,
                        'answers_by_question': my_answers_by_question
                    }
                    if is_test:
                        score_percent = round((test_correct_answers / test_total_questions) * 100, 1) if test_total_questions > 0 else 0.0
                        my_response['test_summary'] = {
                            'total_questions': int(test_total_questions),
                            'answered_questions': int(test_answered_questions),
                            'correct_answers': int(test_correct_answers),
                            'score_percent': float(score_percent)
                        }

                result.append({
                    'id': survey_id,
                    'title': row[1],
                    'description': row[2] or '',
                    'created_at': self._survey_dt_to_iso(row[3]),
                    'updated_at': self._survey_dt_to_iso(row[4]),
                    'created_by': {
                        'id': row[9],
                        'name': row[10] or 'Система',
                        'role': row[11] or 'unknown'
                    },
                    'repeat': {
                        'root_id': int(repeat_root_id),
                        'iteration': int(repeat_iteration),
                        'is_repeat': int(repeat_iteration) > 1
                    },
                    'is_test': bool(is_test),
                    'assignment': {
                        'direction_ids': normalized_direction_ids,
                        'tenure_weeks_min': row[7],
                        'tenure_weeks_max': row[8],
                        'operator_ids': [operator_id]
                    },
                    'questions': questions,
                    'my_assignment': {
                        'id': int(row[12]),
                        'status': assignment_status,
                        'assigned_at': self._survey_dt_to_iso(row[14]),
                        'completed_at': self._survey_dt_to_iso(row[15]),
                        'response_id': response_id,
                        'submitted_at': self._survey_dt_to_iso(row[17]),
                        'can_submit': assignment_status != 'completed'
                    },
                    'my_response': my_response,
                    'statistics': {
                        'assigned_count': 1,
                        'completed_count': 1 if assignment_status == 'completed' else 0,
                        'pending_count': 0 if assignment_status == 'completed' else 1,
                        'responses_count': 1 if response_id is not None else 0,
                        'completion_rate': 100.0 if assignment_status == 'completed' else 0.0,
                        'question_stats': []
                    }
                })

            return result

    def submit_survey_response(self, survey_id, operator_id, answers):
        survey_id = int(survey_id)
        operator_id = int(operator_id)

        if not isinstance(answers, list):
            raise ValueError("SURVEY_ANSWERS_REQUIRED")

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT id, COALESCE(is_test, FALSE)
                FROM surveys
                WHERE id = %s
            """, (survey_id,))
            survey_row = cursor.fetchone()
            if not survey_row:
                raise ValueError("SURVEY_NOT_FOUND")
            is_test = bool(survey_row[1])

            cursor.execute("""
                SELECT id, status
                FROM survey_assignments
                WHERE survey_id = %s AND operator_id = %s
            """, (survey_id, operator_id))
            assignment_row = cursor.fetchone()
            if not assignment_row:
                raise PermissionError("SURVEY_NOT_ASSIGNED")

            assignment_id = int(assignment_row[0])
            assignment_status = str(assignment_row[1] or 'assigned')
            if assignment_status == 'completed':
                raise ValueError("SURVEY_ALREADY_COMPLETED")

            cursor.execute("""
                SELECT
                    id,
                    question_text,
                    question_type,
                    is_required,
                    allow_other,
                    options_json,
                    correct_options_json
                FROM survey_questions
                WHERE survey_id = %s
                ORDER BY position, id
            """, (survey_id,))
            question_rows = cursor.fetchall()
            if not question_rows:
                raise ValueError("SURVEY_HAS_NO_QUESTIONS")

            questions_by_id = {}
            for row in question_rows:
                options = row[5] if isinstance(row[5], list) else []
                correct_options = row[6] if len(row) > 6 and isinstance(row[6], list) else []
                questions_by_id[int(row[0])] = {
                    'id': int(row[0]),
                    'text': row[1],
                    'type': row[2],
                    'required': bool(row[3]),
                    'allow_other': bool(row[4]),
                    'options': [str(item) for item in options],
                    'correct_options': [str(item) for item in correct_options]
                }

            answers_by_question = {}
            for raw_answer in answers:
                if not isinstance(raw_answer, dict):
                    continue
                raw_question_id = raw_answer.get('question_id')
                try:
                    question_id = int(raw_question_id)
                except Exception:
                    continue
                if question_id in questions_by_id:
                    answers_by_question[question_id] = raw_answer

            normalized_answers = []
            for question in questions_by_id.values():
                raw_answer = answers_by_question.get(question['id']) or {}
                qtype = question['type']
                if is_test and qtype not in ('single', 'multiple'):
                    raise ValueError(f"SURVEY_TEST_RATING_NOT_ALLOWED_{question['id']}")

                answer_text = str(raw_answer.get('answer_text') or '').strip()
                if is_test:
                    answer_text = ''
                if answer_text and not question.get('allow_other'):
                    answer_text = ''

                if qtype == 'rating':
                    rating_value = raw_answer.get('rating_value')
                    if rating_value in (None, ''):
                        if question['required']:
                            raise ValueError(f"SURVEY_REQUIRED_QUESTION_{question['id']}")
                        continue
                    try:
                        rating_value = int(rating_value)
                    except Exception:
                        raise ValueError(f"SURVEY_INVALID_RATING_{question['id']}")
                    if rating_value < 1 or rating_value > 5:
                        raise ValueError(f"SURVEY_INVALID_RATING_{question['id']}")

                    normalized_answers.append({
                        'question_id': question['id'],
                        'selected_options': [],
                        'answer_text': None,
                        'rating_value': rating_value
                    })
                    continue

                raw_selected_options = raw_answer.get('selected_options')
                if isinstance(raw_selected_options, list):
                    selected_options = [str(item or '').strip() for item in raw_selected_options if str(item or '').strip()]
                elif raw_selected_options is None:
                    selected_options = []
                else:
                    one = str(raw_selected_options).strip()
                    selected_options = [one] if one else []

                selected_unique = []
                for selected in selected_options:
                    if selected not in selected_unique:
                        selected_unique.append(selected)

                allowed_options = set(question.get('options') or [])
                invalid_selected = [item for item in selected_unique if item not in allowed_options]
                if invalid_selected and not question.get('allow_other'):
                    raise ValueError(f"SURVEY_INVALID_OPTION_{question['id']}")

                valid_selected = [item for item in selected_unique if item in allowed_options]
                if qtype == 'single' and len(valid_selected) > 1:
                    raise ValueError(f"SURVEY_TOO_MANY_OPTIONS_{question['id']}")

                if qtype == 'single' and answer_text:
                    valid_selected = []
                    invalid_selected = []

                if question['required'] and not valid_selected and not answer_text and not (question.get('allow_other') and invalid_selected):
                    raise ValueError(f"SURVEY_REQUIRED_QUESTION_{question['id']}")

                if question.get('allow_other') and invalid_selected and not answer_text:
                    answer_text = ', '.join(invalid_selected)

                if answer_text and len(answer_text) > self.SURVEY_OTHER_ANSWER_MAX_LENGTH:
                    raise ValueError(f"SURVEY_OTHER_TEXT_TOO_LONG_{question['id']}")

                if not valid_selected and not answer_text:
                    continue

                normalized_answers.append({
                    'question_id': question['id'],
                    'selected_options': valid_selected,
                    'answer_text': answer_text or None,
                    'rating_value': None
                })

            if not normalized_answers:
                raise ValueError("SURVEY_EMPTY_RESPONSE")

            cursor.execute("""
                SELECT id
                FROM survey_responses
                WHERE survey_id = %s AND operator_id = %s
            """, (survey_id, operator_id))
            response_row = cursor.fetchone()

            if response_row:
                response_id = int(response_row[0])
                cursor.execute("DELETE FROM survey_answers WHERE response_id = %s", (response_id,))
                cursor.execute("""
                    UPDATE survey_responses
                    SET
                        assignment_id = %s,
                        submitted_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    WHERE id = %s
                """, (assignment_id, response_id))
            else:
                cursor.execute("""
                    INSERT INTO survey_responses (
                        survey_id,
                        operator_id,
                        assignment_id,
                        submitted_at
                    )
                    VALUES (
                        %s,
                        %s,
                        %s,
                        (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    )
                    RETURNING id
                """, (survey_id, operator_id, assignment_id))
                response_id = int(cursor.fetchone()[0])

            for answer in normalized_answers:
                cursor.execute("""
                    INSERT INTO survey_answers (
                        response_id,
                        question_id,
                        answer_text,
                        selected_options_json,
                        rating_value,
                        created_at
                    )
                    VALUES (
                        %s,
                        %s,
                        %s,
                        %s::jsonb,
                        %s,
                        (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    )
                """, (
                    response_id,
                    answer['question_id'],
                    answer['answer_text'],
                    json.dumps(answer['selected_options'], ensure_ascii=False),
                    answer['rating_value']
                ))

            cursor.execute("""
                UPDATE survey_assignments
                SET
                    status = 'completed',
                    completed_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                WHERE id = %s
            """, (assignment_id,))

            cursor.execute("SELECT completed_at FROM survey_assignments WHERE id = %s", (assignment_id,))
            completed_row = cursor.fetchone()
            completed_at = completed_row[0] if completed_row else None

            return {
                'survey_id': survey_id,
                'assignment_id': assignment_id,
                'response_id': response_id,
                'completed_at': self._survey_dt_to_iso(completed_at)
            }

    def _task_visible_for_requester(self, requester_role, requester_id, created_by, assigned_to, assignee_role, assignee_supervisor_id):
        role = normalize_role_value(requester_role)
        requester_id = int(requester_id)
        if role_has_min(role, 'admin'):
            return True
        if created_by is not None and int(created_by) == requester_id:
            return True
        if assigned_to is not None and int(assigned_to) == requester_id:
            return True
        if role == 'sv' and assignee_role == 'operator' and assignee_supervisor_id is not None and int(assignee_supervisor_id) == requester_id:
            return True
        return False

    def get_task_recipients(self, requester_id, requester_role):
        requester_id = int(requester_id)
        role = normalize_role_value(requester_role)
        with self._get_cursor() as cursor:
            if role_has_min(role, 'admin') or role == 'sv':
                cursor.execute("""
                    SELECT u.id, u.name, u.role, u.supervisor_id, COALESCE(u.status, 'working')
                    FROM users u
                    WHERE u.role IN ('super_admin', 'admin', 'sv')
                      AND COALESCE(u.status, 'working') <> 'fired'
                    ORDER BY CASE WHEN u.role IN ('super_admin', 'admin') THEN 0 ELSE 1 END, u.name
                """)
            else:
                return []

            rows = cursor.fetchall()

        return [
            {
                "id": row[0],
                "name": row[1],
                "role": row[2],
                "supervisor_id": row[3],
                "status": row[4]
            }
            for row in rows
        ]

    def create_task(self, subject, description, tag, assigned_to, created_by, attachments=None):
        subject_norm = (subject or '').strip()
        description_norm = (description or '').strip() or None
        tag_norm = (tag or 'task').strip().lower() or 'task'
        assigned_to_id = int(assigned_to)
        created_by_id = int(created_by) if created_by is not None else None
        attachments = attachments or []

        if not subject_norm:
            raise ValueError("SUBJECT_REQUIRED")
        if tag_norm not in ('task', 'problem', 'suggestion'):
            raise ValueError("INVALID_TAG")

        with self._get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO tasks (subject, description, tag, status, assigned_to, created_by, created_at, updated_at)
                VALUES (%s, %s, %s, 'assigned', %s, %s, (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'), (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'))
                RETURNING id, created_at, updated_at
            """, (subject_norm, description_norm, tag_norm, assigned_to_id, created_by_id))
            task_id, created_at, updated_at = cursor.fetchone()

            cursor.execute("""
                INSERT INTO task_status_history (task_id, status_code, changed_by)
                VALUES (%s, 'assigned', %s)
            """, (task_id, created_by_id))

            for attachment in attachments:
                file_name = (attachment.get('file_name') or 'attachment').strip() or 'attachment'
                content_type = (attachment.get('content_type') or '').strip() or None
                storage_type = (attachment.get('storage_type') or 'gcs').strip().lower() or 'gcs'
                if storage_type != 'gcs':
                    continue
                gcs_bucket = (attachment.get('gcs_bucket') or '').strip() or None
                gcs_blob_path = (attachment.get('gcs_blob_path') or '').strip() or None
                file_size = int(attachment.get('file_size') or 0)
                if not gcs_bucket or not gcs_blob_path:
                    continue
                cursor.execute("""
                    INSERT INTO task_attachments (
                        task_id, file_name, content_type, file_size, file_data,
                        storage_type, gcs_bucket, gcs_blob_path, attachment_kind, uploaded_by
                    )
                    VALUES (%s, %s, %s, %s, NULL, 'gcs', %s, %s, 'initial', %s)
                """, (task_id, file_name, content_type, file_size, gcs_bucket, gcs_blob_path, created_by_id))

        return {
            "id": task_id,
            "created_at": created_at.isoformat() if hasattr(created_at, 'isoformat') else created_at,
            "updated_at": updated_at.isoformat() if hasattr(updated_at, 'isoformat') else updated_at
        }

    def get_tasks_for_requester(self, requester_id, requester_role):
        requester_id = int(requester_id)
        role = normalize_role_value(requester_role)

        with self._get_cursor() as cursor:
            if role_has_min(role, 'admin'):
                cursor.execute("""
                    SELECT
                        t.id, t.subject, t.description, t.tag, t.status, t.created_at, t.updated_at,
                        t.completion_summary, t.completed_at, t.completed_by, completed_user.name,
                        assignee.id, assignee.name, assignee.role, assignee.supervisor_id,
                        creator.id, creator.name, creator.role
                    FROM tasks t
                    LEFT JOIN users assignee ON assignee.id = t.assigned_to
                    LEFT JOIN users creator ON creator.id = t.created_by
                    LEFT JOIN users completed_user ON completed_user.id = t.completed_by
                    ORDER BY t.created_at DESC, t.id DESC
                """)
            elif role == 'sv':
                cursor.execute("""
                    SELECT
                        t.id, t.subject, t.description, t.tag, t.status, t.created_at, t.updated_at,
                        t.completion_summary, t.completed_at, t.completed_by, completed_user.name,
                        assignee.id, assignee.name, assignee.role, assignee.supervisor_id,
                        creator.id, creator.name, creator.role
                    FROM tasks t
                    LEFT JOIN users assignee ON assignee.id = t.assigned_to
                    LEFT JOIN users creator ON creator.id = t.created_by
                    LEFT JOIN users completed_user ON completed_user.id = t.completed_by
                    WHERE
                        t.created_by = %s
                        OR t.assigned_to = %s
                    ORDER BY t.created_at DESC, t.id DESC
                """, (requester_id, requester_id))
            else:
                return []

            task_rows = cursor.fetchall()
            if not task_rows:
                return []

            task_ids = [row[0] for row in task_rows]

            cursor.execute("""
                SELECT
                    h.id, h.task_id, h.status_code, h.comment, h.changed_at,
                    h.changed_by, u.name
                FROM task_status_history h
                LEFT JOIN users u ON u.id = h.changed_by
                WHERE h.task_id = ANY(%s)
                ORDER BY h.changed_at ASC, h.id ASC
            """, (task_ids,))
            history_rows = cursor.fetchall()

            cursor.execute("""
                SELECT
                    a.id, a.task_id, a.file_name, a.content_type, a.file_size, a.created_at,
                    COALESCE(a.storage_type, 'db'), a.gcs_bucket, a.gcs_blob_path,
                    COALESCE(a.attachment_kind, 'initial')
                FROM task_attachments a
                WHERE a.task_id = ANY(%s)
                ORDER BY a.id ASC
            """, (task_ids,))
            attachment_rows = cursor.fetchall()

        history_map = defaultdict(list)
        for row in history_rows:
            history_map[row[1]].append({
                "id": row[0],
                "status_code": row[2],
                "comment": row[3],
                "changed_at": row[4].isoformat() if hasattr(row[4], 'isoformat') else row[4],
                "changed_by": row[5],
                "changed_by_name": row[6]
            })

        attachment_map = defaultdict(list)
        for row in attachment_rows:
            attachment_map[row[1]].append({
                "id": row[0],
                "file_name": row[2],
                "content_type": row[3],
                "file_size": row[4],
                "created_at": row[5].isoformat() if hasattr(row[5], 'isoformat') else row[5],
                "storage_type": row[6],
                "gcs_bucket": row[7],
                "gcs_blob_path": row[8],
                "attachment_kind": row[9]
            })

        result = []
        for row in task_rows:
            task_id = row[0]
            all_attachments = attachment_map.get(task_id, [])
            initial_attachments = [a for a in all_attachments if a.get("attachment_kind") != "result"]
            result_attachments = [a for a in all_attachments if a.get("attachment_kind") == "result"]
            result.append({
                "id": task_id,
                "subject": row[1],
                "description": row[2],
                "tag": row[3],
                "status": row[4],
                "created_at": row[5].isoformat() if hasattr(row[5], 'isoformat') else row[5],
                "updated_at": row[6].isoformat() if hasattr(row[6], 'isoformat') else row[6],
                "completion_summary": row[7],
                "completed_at": row[8].isoformat() if hasattr(row[8], 'isoformat') else row[8],
                "completed_by": row[9],
                "completed_by_name": row[10],
                "assignee": {
                    "id": row[11],
                    "name": row[12],
                    "role": row[13],
                    "supervisor_id": row[14]
                } if row[11] else None,
                "creator": {
                    "id": row[15],
                    "name": row[16],
                    "role": row[17]
                } if row[15] else None,
                "history": history_map.get(task_id, []),
                "attachments": initial_attachments,
                "completion_attachments": result_attachments
            })
        return result

    def update_task_status(self, task_id, requester_id, requester_role, action, comment=None, completion_summary=None, completion_attachments=None):
        task_id = int(task_id)
        requester_id = int(requester_id)
        role = normalize_role_value(requester_role)
        action_norm = str(action or '').strip().lower()
        comment_norm = (comment or '').strip() or None
        completion_summary_norm = (completion_summary or '').strip() or None
        completion_attachments = completion_attachments or []

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    t.id, t.created_by, t.assigned_to, t.status,
                    assignee.role, assignee.supervisor_id
                FROM tasks t
                LEFT JOIN users assignee ON assignee.id = t.assigned_to
                WHERE t.id = %s
            """, (task_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("TASK_NOT_FOUND")

            created_by = row[1]
            assigned_to = row[2]
            current_status = row[3]
            assignee_role = row[4]
            assignee_supervisor_id = row[5]

            if not self._task_visible_for_requester(role, requester_id, created_by, assigned_to, assignee_role, assignee_supervisor_id):
                raise PermissionError("TASK_FORBIDDEN")

            is_assignee = assigned_to is not None and int(assigned_to) == requester_id
            is_reviewer = (
                role_has_min(role, 'admin')
                or (created_by is not None and int(created_by) == requester_id)
                or (role == 'sv' and not is_assignee)
            )

            target_status = None
            history_status = None

            if action_norm == 'in_progress':
                if not is_assignee:
                    raise PermissionError("ONLY_ASSIGNEE")
                if current_status not in ('assigned', 'returned', 'in_progress'):
                    raise ValueError("INVALID_TRANSITION")
                if current_status == 'in_progress':
                    return {
                        "task_id": task_id,
                        "status": current_status
                    }
                target_status = 'in_progress'
                history_status = 'in_progress'
            elif action_norm == 'completed':
                if not is_assignee:
                    raise PermissionError("ONLY_ASSIGNEE")
                if current_status not in ('in_progress', 'returned'):
                    raise ValueError("INVALID_TRANSITION")
                target_status = 'completed'
                history_status = 'completed'
            elif action_norm == 'accepted':
                if not is_reviewer:
                    raise PermissionError("ONLY_REVIEWER")
                if current_status != 'completed':
                    raise ValueError("INVALID_TRANSITION")
                target_status = 'accepted'
                history_status = 'accepted'
            elif action_norm == 'returned':
                if not is_reviewer:
                    raise PermissionError("ONLY_REVIEWER")
                if current_status not in ('completed', 'accepted'):
                    raise ValueError("INVALID_TRANSITION")
                target_status = 'returned'
                history_status = 'returned'
            elif action_norm == 'reopened':
                if not is_reviewer:
                    raise PermissionError("ONLY_REVIEWER")
                if current_status != 'accepted':
                    raise ValueError("INVALID_TRANSITION")
                target_status = 'in_progress'
                history_status = 'reopened'
            else:
                raise ValueError("INVALID_ACTION")

            if action_norm == 'completed':
                cursor.execute("""
                    UPDATE tasks
                    SET
                        status = %s,
                        completion_summary = %s,
                        completed_by = %s,
                        completed_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty'),
                        updated_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    WHERE id = %s
                    RETURNING updated_at
                """, (target_status, completion_summary_norm, requester_id, task_id))
            else:
                cursor.execute("""
                    UPDATE tasks
                    SET status = %s, updated_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Almaty')
                    WHERE id = %s
                    RETURNING updated_at
                """, (target_status, task_id))
            updated_at_row = cursor.fetchone()
            updated_at = updated_at_row[0] if updated_at_row else None

            history_comment = completion_summary_norm if action_norm == 'completed' else comment_norm
            cursor.execute("""
                INSERT INTO task_status_history (task_id, status_code, changed_by, comment)
                VALUES (%s, %s, %s, %s)
                RETURNING id, changed_at
            """, (task_id, history_status, requester_id, history_comment))
            history_row = cursor.fetchone()

            if action_norm in ('completed', 'returned', 'reopened'):
                attachment_kind = 'result' if action_norm == 'completed' else 'initial'
                for attachment in completion_attachments:
                    file_name = (attachment.get('file_name') or 'attachment').strip() or 'attachment'
                    content_type = (attachment.get('content_type') or '').strip() or None
                    storage_type = (attachment.get('storage_type') or 'gcs').strip().lower() or 'gcs'
                    if storage_type != 'gcs':
                        continue
                    gcs_bucket = (attachment.get('gcs_bucket') or '').strip() or None
                    gcs_blob_path = (attachment.get('gcs_blob_path') or '').strip() or None
                    file_size = int(attachment.get('file_size') or 0)
                    if not gcs_bucket or not gcs_blob_path:
                        continue
                    cursor.execute("""
                        INSERT INTO task_attachments (
                            task_id, file_name, content_type, file_size, file_data,
                            storage_type, gcs_bucket, gcs_blob_path, attachment_kind, uploaded_by
                        )
                        VALUES (%s, %s, %s, %s, NULL, 'gcs', %s, %s, %s, %s)
                    """, (task_id, file_name, content_type, file_size, gcs_bucket, gcs_blob_path, attachment_kind, requester_id))

            return {
                "task_id": task_id,
                "status": target_status,
                "updated_at": updated_at.isoformat() if hasattr(updated_at, 'isoformat') else updated_at,
                "history_id": history_row[0] if history_row else None,
                "history_changed_at": (history_row[1].isoformat() if history_row and hasattr(history_row[1], 'isoformat') else (history_row[1] if history_row else None))
            }

    def get_task_attachment_for_requester(self, attachment_id, requester_id, requester_role):
        attachment_id = int(attachment_id)
        requester_id = int(requester_id)
        role = normalize_role_value(requester_role)

        with self._get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    a.id, a.file_name, a.content_type, a.file_size, a.file_data, a.created_at,
                    COALESCE(a.storage_type, 'db'), a.gcs_bucket, a.gcs_blob_path,
                    t.id, t.created_by, t.assigned_to,
                    assignee.role, assignee.supervisor_id
                FROM task_attachments a
                JOIN tasks t ON t.id = a.task_id
                LEFT JOIN users assignee ON assignee.id = t.assigned_to
                WHERE a.id = %s
            """, (attachment_id,))
            row = cursor.fetchone()
            if not row:
                return None

            created_by = row[10]
            assigned_to = row[11]
            assignee_role = row[12]
            assignee_supervisor_id = row[13]

            if not self._task_visible_for_requester(role, requester_id, created_by, assigned_to, assignee_role, assignee_supervisor_id):
                raise PermissionError("TASK_FORBIDDEN")

            storage_type = row[6] or 'db'
            file_data = row[4]
            if storage_type != 'gcs':
                if isinstance(file_data, memoryview):
                    file_data = file_data.tobytes()
                elif isinstance(file_data, bytearray):
                    file_data = bytes(file_data)
                elif file_data is None:
                    file_data = b''

            return {
                "id": row[0],
                "file_name": row[1],
                "content_type": row[2] or 'application/octet-stream',
                "file_size": int(row[3] or 0),
                "file_data": file_data,
                "created_at": row[5].isoformat() if hasattr(row[5], 'isoformat') else row[5],
                "storage_type": storage_type,
                "gcs_bucket": row[7],
                "gcs_blob_path": row[8],
                "task_id": row[9]
            }

    def save_ai_feedback_cache(self, operator_id: int, month: str, feedback_data: dict):
        """Сохранить или обновить AI фидбэк в кэше"""
        with self._get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO ai_feedback_cache (operator_id, month, feedback_data, updated_at)
                VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
                ON CONFLICT (operator_id, month)
                DO UPDATE SET
                    feedback_data = EXCLUDED.feedback_data,
                    updated_at = CURRENT_TIMESTAMP
            """, (operator_id, month, json.dumps(feedback_data, ensure_ascii=False)))

    def save_ai_birthday_greeting_cache(self, user_id: int, date_key: str, greeting_data: dict):
        """Сохранить или обновить AI-поздравление с днем рождения"""
        with self._get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO ai_birthday_greeting_cache (user_id, greeting_date, greeting_data, updated_at)
                VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
                ON CONFLICT (user_id, greeting_date)
                DO UPDATE SET
                    greeting_data = EXCLUDED.greeting_data,
                    updated_at = CURRENT_TIMESTAMP
            """, (user_id, date_key, json.dumps(greeting_data, ensure_ascii=False)))


# Initialize database
db = Database()
