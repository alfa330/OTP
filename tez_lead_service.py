"""Оркестрация успешек TEZ ОП: загрузка базы лидов и ночная сверка.

Разделение обязанностей:
  - tez_op_leads.py     — чистые правила (нормализация номера, статус лида);
  - tez_first_orders.py — клиент TEZ APP (дата первой поездки);
  - tez_binotel_calls.py — клиент Binotel (история звонков по номеру клиента);
  - этот модуль         — склейка всего вместе поверх Database.

Модуль намеренно не импортирует bot_schedule2: резолвер «имя сотрудника ->
оператор ОП» передаётся снаружи колбэком (`resolve_operator`), иначе получился
бы циклический импорт, а логику нельзя было бы прогнать в тестах.

Почему пайплайн устроен именно так: выехал ли водитель — приходится спрашивать
TEZ APP регулярно, пока он не выедет, поэтому про поездки спрашиваем по всей
базе месяца. Звонки поднимаем с двух сторон:

  - зеркало по дням (`sync_calls_for_period`) — весь трафик компании за окно
    месяца одним запросом на сутки. Нужно, чтобы «Обзвонено» показывало ВСЕ
    попытки дозвона по базе, а не только по тем, кто в итоге выехал;
  - история по номеру (`sync_calls_for_converted`) — только по выехавшим, зато
    за всё время: звонок старше окна успешки не даёт, но отличает «Не засчитана»
    от «Уже работающий».
"""

import csv
import io
import logging
import threading
import time
from contextlib import contextmanager
from datetime import date, datetime, time as dt_time, timedelta

from tez_op_leads import (
    ALMATY_TZ,
    DEFAULT_MIN_BILLSEC,
    STATUS_SUCCESS,
    as_almaty,
    call_window_for_period,
    compute_lead_outcome,
    normalize_kz_phone,
    reactivation_gap_for_period,
)

log = logging.getLogger(__name__)

# Сколько последних дней окна перекачиваем заново при каждом прогоне: день, за
# который синхронизировались «сегодня», в тот момент ещё не закончился, а
# вчерашний мог дописаться поздними звонками.
CALL_MIRROR_REFRESH_DAYS = 2
# Потолок времени на добор пропущенных дней в одном прогоне. Ручная «Сверка»
# отвечает по HTTP, а полное окно месяца — это ~38 запросов к Binotel (~3 минуты
# с учётом зазора между ними). Недобранное закроет следующий прогон: зеркало
# идемпотентно и помнит, какие дни уже перекачаны.
CALL_MIRROR_TIME_BUDGET = 60.0
# Ночью спешить некуда — за один заход добираем хоть весь месяц.
CALL_MIRROR_NIGHTLY_BUDGET = 900.0

# Заголовки, которыми СВ подписывает колонки (файл приходит как fio/phone).
FIO_HEADERS = {'fio', 'фио', 'имя', 'name', 'full_name', 'водитель', 'driver'}
PHONE_HEADERS = {'phone', 'телефон', 'номер', 'phone_number', 'msisdn', 'тел'}
MAX_LEAD_ROWS = 50000

# Ночная джоба и кнопка «Сверка» делают одно и то же и ходят в одни и те же
# внешние API. Advisory-лок периода в БД защищает данные, но не лимиты: пока
# один прогон ждёт лок, второй уже потратил запросы Binotel и TEZ APP заново.
# Тот же приём, что в tez_op_productivity._SYNC_LOCK.
_SYNC_LOCK = threading.Lock()


class SyncBusy(RuntimeError):
    """Сверка уже идёт. Второй параллельный прогон только удвоит расход лимитов."""


@contextmanager
def sync_lock():
    """Пропускает внутрь только один прогон сверки на процесс."""
    if not _SYNC_LOCK.acquire(blocking=False):
        raise SyncBusy('Сверка уже выполняется — дождитесь её окончания')
    try:
        yield
    finally:
        _SYNC_LOCK.release()


def _norm_header(value):
    return str(value or '').strip().lower().replace(' ', '_')


def _rows_from_csv(raw_bytes):
    text = raw_bytes.decode('utf-8-sig', errors='replace')
    sample = text[:4096]
    delimiter = ';' if sample.count(';') > sample.count(',') else ','
    return [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _rows_from_xlsx(raw_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def parse_leads_file(raw_bytes, file_ext):
    """Разбирает файл базы лидов в строки (row_number, ФИО, сырой телефон, phone_norm).

    Ожидаемый формат — две колонки с шапкой `fio` и `phone`. Если шапки нет,
    считаем первую колонку именем, вторую телефоном. Телефон в Excel часто
    хранится числом, поэтому приводим к строке аккуратно (без экспоненты).
    """
    ext = str(file_ext or '').lower()
    if ext in ('.xlsx', '.xlsm'):
        raw_rows = _rows_from_xlsx(raw_bytes)
    elif ext == '.csv':
        raw_rows = _rows_from_csv(raw_bytes)
    else:
        raise ValueError('Поддерживаются только .csv, .xlsx и .xlsm')

    raw_rows = [r for r in raw_rows if any(str(c or '').strip() for c in r)]
    if not raw_rows:
        raise ValueError('Файл пустой')

    fio_idx, phone_idx, start_at = 0, 1, 0
    header = [_norm_header(c) for c in raw_rows[0]]
    if any(h in FIO_HEADERS for h in header) or any(h in PHONE_HEADERS for h in header):
        for idx, name in enumerate(header):
            if name in FIO_HEADERS:
                fio_idx = idx
            elif name in PHONE_HEADERS:
                phone_idx = idx
        start_at = 1

    rows = []
    for offset, raw in enumerate(raw_rows[start_at:], start=start_at + 1):
        fio = str(raw[fio_idx] or '').strip() if fio_idx < len(raw) else ''
        phone_cell = raw[phone_idx] if phone_idx < len(raw) else ''
        if isinstance(phone_cell, float):
            phone_cell = f"{phone_cell:.0f}"
        phone_raw = str(phone_cell or '').strip()
        if not fio and not phone_raw:
            continue
        if len(rows) >= MAX_LEAD_ROWS:
            raise ValueError(
                f'В файле больше {MAX_LEAD_ROWS} строк с данными. '
                'Разделите его на несколько файлов.'
            )
        rows.append((offset, fio, phone_raw, normalize_kz_phone(phone_raw)))
    if not rows:
        raise ValueError('В файле не нашлось ни одной строки с данными')
    return rows


def import_lead_batch(db, department_id, year, month, uploaded_by, file_name, rows):
    """Создаёт загрузку и накатывает строки на помесячную базу."""
    batch_id = db.create_tez_lead_batch(department_id, year, month, uploaded_by, file_name)
    counts = db.import_tez_lead_rows(batch_id, year, month, rows)
    counts['batch_id'] = batch_id
    return counts


def check_batch_already_working(db, batch_id, year, month, first_orders_client, min_billsec=DEFAULT_MIN_BILLSEC):
    """Фоновая проверка свежей загрузки: кто из базы уже выехал раньше нас.

    Владелец просил помечать таких сразу, чтобы операторы не тратили день на
    обзвон тех, кто и так на линии. Проверяем только номера этой загрузки.
    """
    db.set_tez_lead_batch_check_state(batch_id, 'running')
    try:
        phones = db.get_tez_phones_pending_first_order(year, month, batch_id=batch_id)
        if phones:
            first_orders = first_orders_client.fetch_first_orders(
                phones, month=f"{int(year)}-{int(month):02d}"
            )
            db.save_tez_first_orders(year, month, first_orders)
        outcomes = recompute_outcomes(db, year, month, min_billsec=min_billsec)
        db.set_tez_lead_batch_check_state(
            batch_id, 'done', already_working=outcomes.get('already_working', 0)
        )
        return outcomes
    except Exception as exc:
        log.error('Проверка базы лидов %s не удалась: %s', batch_id, exc, exc_info=True)
        db.set_tez_lead_batch_check_state(batch_id, 'error', error=str(exc))
        raise


def sync_first_orders(db, year, month, first_orders_client):
    """Шаг 1 ночной джобы: спрашиваем TEZ APP по всем ещё не выехавшим лидам месяца.

    TEZ APP считает окно в два месяца, поэтому одним запросом получаем и заказ в
    отчётном месяце, и заказ в предыдущем (последний нужен, чтобы отсечь тех,
    кто уже работал).
    """
    phones = db.get_tez_phones_pending_first_order(year, month)
    if not phones:
        return {'checked': 0, 'found': 0}
    first_orders = first_orders_client.fetch_first_orders(
        phones, month=f"{int(year)}-{int(month):02d}"
    )
    found = db.save_tez_first_orders(year, month, first_orders)
    return {'checked': len(phones), 'found': found}


def sync_carry_first_orders(db, year, month, first_orders_client):
    """Шаг 1б: поездки ПЕРЕНЕСЁННЫХ лидов в отчётном месяце.

    Лид прошлого месяца, не ставший успешкой, дорабатывается ещё один месяц:
    спрашиваем по нему TEZ APP за отчётный месяц (тем же запросом, что и своих —
    контракт одинаковый), а ответ кладём в carry_* колонки его строки.

    Цена на проде (июль -> август): 6 993 кандидата, то есть +64 запроса
    пачками по 100. Число убывает по мере того, как водители выезжают.
    """
    phones = db.get_tez_carry_phones_pending_first_order(year, month)
    if not phones:
        return {'checked': 0, 'found': 0}
    first_orders = first_orders_client.fetch_first_orders(
        phones, month=f"{int(year)}-{int(month):02d}"
    )
    found = db.save_tez_carry_first_orders(year, month, first_orders)
    return {'checked': len(phones), 'found': found}


def _prepare_calls(raw_calls, resolve_operator, min_billsec=DEFAULT_MIN_BILLSEC, wanted=None):
    """Нормализованные звонки Binotel -> строки для tez_lead_calls.

    resolve_operator(employee_name, call_date) должен вернуть id оператора ОП
    либо None. Именно тут отсекаются звонки ТП/линии: по решению владельца они
    не должны перехватывать успешку у отдела продаж (в базе они остаются, но
    без operator_id, поэтому ни в «Обзвонено», ни в успешку не попадут).

    wanted — необязательный набор номеров-фильтр; None означает «берём всё».
    """
    prepared = []
    for call in raw_calls or []:
        phone_norm = normalize_kz_phone(call.get('external_number'))
        if not phone_norm or (wanted is not None and phone_norm not in wanted):
            continue
        started_at = as_almaty(call.get('start_time'))
        if started_at is None:
            continue
        operator_id = resolve_operator(call.get('employee_name'), started_at.date())
        prepared.append({
            'general_call_id': call.get('general_call_id'),
            'phone_norm': phone_norm,
            'started_at': started_at,
            'call_type': call.get('call_type'),
            'billsec': call.get('billsec'),
            'waitsec': call.get('waitsec'),
            'disposition': call.get('disposition'),
            'internal_number': call.get('internal_number'),
            'employee_name': call.get('employee_name'),
            'employee_email': call.get('employee_email'),
            'operator_id': operator_id,
            'is_qualifying': (
                int(call.get('call_type', -1)) == 1
                and int(call.get('billsec') or 0) >= int(min_billsec)
                and operator_id is not None
            ),
        })
    return prepared


def reattribute_calls_without_operator(db, year, month, resolve_operator,
                                       min_billsec=DEFAULT_MIN_BILLSEC):
    """Возвращает звонкам окна привязку к оператору по внутреннему номеру.

    Зачем: Binotel удаляет сотрудника у себя (например, при увольнении) и
    перестаёт отдавать `employeeData` по всем его прошлым звонкам. Матчинг у
    нас идёт по имени, поэтому такие звонки становятся «ничьими» — и вместе с
    ними пропадают уже начисленные успешки. На проде 05.08.2026 так исчезло 8
    июльских успешек оператора с sip 927.

    Имя восстанавливаем по нашим же данным: кто звонил с этого номера, видно по
    звонкам, где имя ещё было. Номер берём только с однозначным владельцем.
    Шаг локальный (Binotel не дёргаем) и идемпотентный: он лишь заполняет
    пустое, поэтому его безопасно гонять каждую ночь — он же чинит и прошлое.
    """
    window_start, window_end = call_window_for_period(year, month)
    window_from = datetime.combine(window_start, dt_time.min, tzinfo=ALMATY_TZ)
    window_to = datetime.combine(window_end + timedelta(days=1), dt_time.min, tzinfo=ALMATY_TZ)

    pending = db.get_tez_calls_without_operator(window_from, window_to)
    if not pending:
        return {'checked': 0, 'restored': 0}

    owners = db.get_tez_call_internal_number_owners() or {}
    restored = []
    for call in pending:
        name = owners.get(call.get('internal_number'))
        if not name:
            continue
        started_at = as_almaty(call.get('started_at'))
        # Модель оператора проверяется на дату ЗВОНКА: тот же принцип, что и
        # при обычной привязке, иначе восстановленный звонок мог бы засчитаться
        # за период, когда человек работал в другом отделе.
        operator_id = resolve_operator(name, (started_at or window_from).date())
        if operator_id is None:
            continue
        restored.append({
            'general_call_id': call['general_call_id'],
            'employee_name': name,
            'operator_id': operator_id,
            'is_qualifying': (
                int(call.get('call_type', -1)) == 1
                and int(call.get('billsec') or 0) >= int(min_billsec)
            ),
        })
    saved = db.restore_tez_call_operators(restored)
    if saved:
        log.info('Успешки TEZ ОП %s-%02d: привязка по внутреннему номеру восстановлена '
                 'у %s звонков', int(year), int(month), saved)
    return {'checked': len(pending), 'restored': saved}


def sync_calls_for_period(db, year, month, binotel_client, resolve_operator,
                          today=None, min_billsec=DEFAULT_MIN_BILLSEC,
                          time_budget=CALL_MIRROR_TIME_BUDGET,
                          refresh_days=CALL_MIRROR_REFRESH_DAYS):
    """Зеркало звонков за окно месяца: ВСЕ попытки дозвона, а не только по выехавшим.

    Шаг 2 ниже поднимает историю лишь по тем, кто уже выехал, — этого хватает
    для начисления успешки, но воронка из-за этого показывала «Обзвонено» лишь
    по горстке номеров (на июльской базе: 464 из 7 195), хотя операторы звонили
    почти всей базе. Здесь качаем весь трафик компании по дням окна и решаем
    локально, к каким лидам он относится.

    Почему по дням, а не по номерам: history-by-external-number пришлось бы
    звать сотнями пачек на каждую базу, а list-of-calls-for-period закрывает
    сутки одним запросом (500–1200 звонков) — на месяц выходит ~38 запросов
    один раз и 2 запроса за ночь дальше.

    Звонки сохраняются даже по номерам, которых сейчас нет ни в одной базе:
    базы грузят в течение месяца, и лид, добавленный 20-го числа, должен сразу
    видеть свои попытки с 1-го, а не ждать повторной выкачки из Binotel.

    Прогон ограничен по времени (time_budget): недобранные дни закроет следующий
    запуск, журнал перекачанных дней делает шаг идемпотентным.
    """
    window_start, window_end = call_window_for_period(year, month)
    today = today or date.today()
    last_day = min(window_end, today)
    if last_day < window_start:
        return {'days': 0, 'days_left': 0, 'calls': 0}

    synced = db.get_tez_call_synced_days(window_start, last_day) or set()
    # Заново перекачиваем только дни рядом с СЕГОДНЯ: сегодняшний на момент
    # синка ещё не закончился, вчерашний мог дописаться поздними звонками.
    # Порог считается от today, а не от конца окна, иначе закрытый месяц
    # бесконечно перекачивал бы свои последние два дня.
    refresh_from = today - timedelta(days=max(int(refresh_days), 1) - 1)

    pending = []
    day = last_day
    while day >= window_start:
        if day >= refresh_from or day not in synced:
            pending.append(day)
        day -= timedelta(days=1)

    started_at = time.monotonic()
    done = 0
    saved = 0
    for day in pending:
        if done and time.monotonic() - started_at >= float(time_budget):
            break
        raw_calls = binotel_client.list_calls_for_day(day)
        prepared = _prepare_calls(raw_calls, resolve_operator, min_billsec=min_billsec)
        saved += db.save_tez_lead_calls(prepared)
        db.mark_tez_call_day_synced(day, len(prepared))
        done += 1
    return {'days': done, 'days_left': len(pending) - done, 'calls': saved}


def sync_calls_for_converted(db, year, month, binotel_client, resolve_operator,
                             min_billsec=DEFAULT_MIN_BILLSEC):
    """Шаг 2: полная история звонков по тем, кто уже выехал.

    Зеркало выше знает только окно месяца, а тут нужна история целиком: звонок
    раньше окна успешки не даёт, но именно он отличает «Не засчитана» (оператор
    работал, но не попал в окно — такие случаи операторы оспаривают) от «Уже
    работающий» (выехал сам). Номеров за ночь единицы, отсюда и один запрос.
    """
    phones = db.get_tez_phones_needing_calls(year, month)
    if not phones:
        return {'phones': 0, 'calls': 0}

    raw_calls = binotel_client.list_calls_by_external_numbers(phones)
    prepared = _prepare_calls(raw_calls, resolve_operator, min_billsec=min_billsec,
                              wanted=set(phones))
    saved = db.save_tez_lead_calls(prepared)
    # Помечаем ВСЕ запрошенные номера, а не только те, по которым нашлись звонки:
    # «звонков нет» — это тоже результат, иначе такие лиды переспрашивались бы
    # в Binotel каждую ночь без конца.
    db.mark_tez_leads_calls_synced(year, month, phones)
    return {'phones': len(phones), 'calls': saved}


def recompute_outcomes(db, year, month, min_billsec=DEFAULT_MIN_BILLSEC, month_closed_before=None):
    """Шаг 3: пересчёт статусов лидов и успешек. Идемпотентен.

    month_closed_before — необязательная дата закрытия расчётного периода:
    успешки, найденные после неё, помечаются is_late, чтобы поздняя загрузка
    базы задним числом была видна, а не всплывала в выплате молча.
    """
    # Порог отдаём в выборку: расчёт всё равно отбрасывает всё, что ему не
    # соответствует, а без фильтра запрос тянет весь трафик по номерам базы за
    # всю историю (на июле 25 766 строк против 6 668 нужных).
    leads = db.get_tez_leads_for_recompute(year, month, min_billsec=min_billsec)
    if not leads:
        return {'success': 0, 'already_working': 0, 'not_counted': 0, 'in_progress': 0, 'new': 0}

    # Успешки — это деньги оператора, поэтому их убыль не должна проходить
    # молча: 05.08.2026 восемь июльских успешек исчезли из-за того, что Binotel
    # забыл сотрудника, и заметили это только вручную.
    successes_before = db.count_tez_successes(year, month)

    names_by_call = {}
    for lead in leads:
        for call in lead['calls']:
            names_by_call[call.get('general_call_id')] = call.get('employee_name') or ''

    # Рубеж вступления новых правил в силу: закрытые месяцы считаются прежней
    # веткой, иначе пересчёт июля переписал бы уже выплаченное.
    gap_days = reactivation_gap_for_period(year, month)

    outcomes = []
    for lead in leads:
        carried = bool(lead.get('is_carried'))
        if carried:
            # У перенесённого лида поездку ищем в ЦЕЛЕВОМ месяце, а его
            # собственные даты не трогаем: они описывают месяц его базы и нужны
            # её воронке.
            trip_at = lead.get('carry_first_order_at')
            last_before = lead.get('carry_last_order_before_at')
            base_period = (lead.get('lead_year', year), lead.get('lead_month', month))
        else:
            trip_at = lead['month_first_order_at']
            last_before = lead.get('last_order_before_at')
            base_period = None
        outcome = compute_lead_outcome(
            trip_at,
            lead['prev_month_first_order_at'],
            lead['calls'],
            min_billsec=min_billsec,
            last_order_before_at=last_before,
            reactivation_gap_days=gap_days,
            base_period=base_period,
        )
        item = {
            'lead_id': lead['id'],
            # Optimistic version: delete/restore может вернуть тот же UUID, но
            # уже другое состояние. Database применит результат только к той
            # версии лида, на которой он был рассчитан.
            'lead_version': lead.get('version'),
            'lead_year': lead.get('lead_year', year),
            'lead_month': lead.get('lead_month', month),
            # Перенос имеет право только ДОБАВИТЬ успешку. Свой статус лид
            # получил в собственном месяце: перепиши его отсюда — и пересчёт
            # того месяца тут же вернёт всё назад, а статус начнёт мигать
            # между прогонами.
            'write_status': not carried,
            'is_carried': carried,
            'phone_norm': lead['phone_norm'],
            'status': outcome['status'],
            'rule': outcome['rule'],
            'operator_id': outcome['operator_id'],
        }
        if outcome['status'] == STATUS_SUCCESS:
            call = outcome['call'] or {}
            success_date = outcome['success_date']
            item.update({
                'operator_name': names_by_call.get(call.get('general_call_id'), ''),
                'call_general_id': call.get('general_call_id'),
                'call_at': outcome['call_at'],
                'first_order_at': outcome['first_order_at'],
                'success_date': success_date,
                # Успешка живёт в месяце ПОЕЗДКИ: звонок из последних 7 дней июня
                # даёт успешку июля, если водитель выехал в июле (любым днём).
                'success_year': success_date.year,
                'success_month': success_date.month,
                'is_late': bool(month_closed_before and success_date < month_closed_before),
            })
        outcomes.append(item)

    outcomes = _drop_duplicate_successes(outcomes)
    stats = db.apply_tez_lead_outcomes(year, month, outcomes)
    stats['successes_before'] = successes_before
    # Сравнивать надо с тем, что реально забронировано за период: успешка
    # перенесённого лида книжится в месяц поездки, а его статус остался в месяце
    # базы. Считай мы только stats['success'], сторож кричал бы каждую ночь.
    booked = stats.get('success', 0) + stats.get('carried_success', 0)
    if booked < successes_before:
        log.warning('Успешки TEZ ОП %s-%02d УМЕНЬШИЛИСЬ: было %s, стало %s. '
                    'Проверьте, не потеряли ли звонки привязку к оператору.',
                    int(year), int(month), successes_before, booked)
    return stats


def _drop_duplicate_successes(outcomes):
    """Одна поездка — одна успешка, даже если номер есть в базах двух месяцев.

    UNIQUE(lead_id) от этого не защищает: у номера отдельная строка в базе
    каждого месяца, и оба лида видят ОДИН И ТОТ ЖЕ список звонков. На проде
    11.08.2026 таких номеров 374, и у шести из них уже была августовская
    успешка при живом июльском лиде.

    Побеждает лид того месяца, в котором поездка: он «свой» для периода.
    На деньги выбор не влияет — оба лида выбирают один и тот же последний
    квалифицирующий звонок и того же оператора, — поэтому проигравшего просто
    не пишем, а не помечаем отдельным статусом.
    """
    own = {(i['phone_norm'], i['success_date']) for i in outcomes
           if i['status'] == STATUS_SUCCESS and not i.get('is_carried')}
    if not own:
        return outcomes
    kept = []
    for item in outcomes:
        if (item.get('is_carried') and item['status'] == STATUS_SUCCESS
                and (item['phone_norm'], item['success_date']) in own):
            # Перенос ничего не добавляет — этот же водитель уже засчитан по
            # строке своего месяца. Свой статус перенесённый лид сохраняет.
            continue
        kept.append(item)
    return kept


def run_nightly(db, first_orders_client, binotel_client, resolve_operator,
                today=None, min_billsec=DEFAULT_MIN_BILLSEC):
    """Полный ночной цикл для текущего месяца (и прошлого — в первую неделю).

    Прошлый месяц добираем 1–7 числа из-за задержки данных на стыке месяцев:
    поездка 30-го числа и звонки по ней доходят уже в новом месяце, а пересчёт
    идемпотентен. Правило успешки при этом окна на день поездки не накладывает
    (звонок из последних 7 дней прошлого месяца засчитывается при поездке любым
    днём отчётного) — успешки отчётного месяца считаются каждую ночь в его
    собственном периоде.
    """
    today = today or date.today()
    periods = [(today.year, today.month)]
    if today.day <= 7:
        prev_year, prev_month = (today.year - 1, 12) if today.month == 1 else (today.year, today.month - 1)
        periods.append((prev_year, prev_month))

    try:
        with sync_lock():
            return _run_nightly_periods(db, periods, first_orders_client, binotel_client,
                                        resolve_operator, today=today, min_billsec=min_billsec)
    except SyncBusy:
        # Ручную сверку не прерываем: она делает ровно то же самое, а джоба
        # идемпотентна и наверстает всё следующей ночью. Изнутри тела цикла
        # SyncBusy прилететь не может — каждый период там ловит свои ошибки сам.
        log.warning('Ночная сверка успешек TEZ ОП пропущена: сверка уже идёт')
        return {'skipped': 'locked'}


def _run_nightly_periods(db, periods, first_orders_client, binotel_client, resolve_operator,
                         today=None, min_billsec=DEFAULT_MIN_BILLSEC):
    """Тело ночного цикла: по периоду за раз, ошибка одного не роняет остальные."""
    report = {}
    for year, month in periods:
        key = f"{year}-{month:02d}"
        try:
            orders = sync_first_orders(db, year, month, first_orders_client)
            carry = sync_carry_first_orders(db, year, month, first_orders_client)
            mirror = sync_calls_for_period(
                db, year, month, binotel_client, resolve_operator, today=today,
                min_billsec=min_billsec, time_budget=CALL_MIRROR_NIGHTLY_BUDGET,
            )
            calls = sync_calls_for_converted(db, year, month, binotel_client,
                                             resolve_operator, min_billsec=min_billsec)
            restored = reattribute_calls_without_operator(
                db, year, month, resolve_operator, min_billsec=min_billsec
            )
            outcomes = recompute_outcomes(db, year, month, min_billsec=min_billsec)
            report[key] = {'first_orders': orders, 'carry_first_orders': carry,
                           'calls_mirror': mirror, 'calls': calls,
                           'restored_calls': restored, 'outcomes': outcomes}
            log.info('Успешки TEZ ОП %s: проверено %s, выехали %s, перенос %s/%s, '
                     'дней зеркала %s (осталось %s), звонков %s, успешек %s (+%s переносом)',
                     key, orders.get('checked'), orders.get('found'),
                     carry.get('found'), carry.get('checked'),
                     mirror.get('days'), mirror.get('days_left'),
                     calls.get('calls'), outcomes.get('success'),
                     outcomes.get('carried_success'))
        except Exception as exc:
            log.error('Ночная сверка успешек TEZ ОП за %s упала: %s', key, exc, exc_info=True)
            report[key] = {'error': str(exc)}
    return report
