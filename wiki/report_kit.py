# -*- coding: utf-8 -*-
"""Кирпичи Excel-выгрузок раздела «Вики»: шапка, таблица, оговорки, имя файла.

Жили внутри wiki/trainer_report.py. Со второй выгрузкой («Аналитика») выбор был
тот же, что на фронте, когда рядом с отчётом по тренажёрам появился отчёт по
вике: скопировать кирпичи или вынести. Скопировать значило бы завести две шапки
одного цвета, две заморозки строк и два автофильтра, которые начали бы
расходиться с первой же правки ширины колонки, — а замечает такое расхождение не
разработчик, а тот, кто открыл оба файла подряд. Поэтому вынесено сюда, ровно по
образцу src/components/wiki/reportKit.jsx.

Правило простое: правка этого файла меняет ОБЕ выгрузки раздела, и проверять
надо обе (tests/test_wiki_trainer_stats.py и tests/test_wiki_analytics.py).

ЧТО ЗДЕСЬ ЗАФИКСИРОВАНО РЕШЕНИЯМИ, А НЕ ВКУСОМ.

* ЗАМОРОЗКА И АВТОФИЛЬТР — НА КАЖДОМ ЛИСТЕ С ТАБЛИЦЕЙ. Без них выгрузку на
  тысячу строк листают вслепую, потеряв шапку на втором экране, а первое, что с
  таким файлом делают, — фильтруют по отделу.
* ДАТА ОБЯЗАНА БЫТЬ ДАТОЙ. Строка «22.08.2026» не сортируется и не попадает в
  фильтр по периоду; в ячейку кладётся datetime, а читаемый вид даёт формат.
* ВРЕМЯ НАИВНОЕ. Метки в базе лежат без зоны, но записаны алматинским временем.
  Excel про зоны не знает вовсе, и aware-объект превращается в текст.
* ДОЛЯ — ЧИСЛОМ ОТ НУЛЯ ДО ЕДИНИЦЫ, а не строкой «35%»: только так по колонке
  считается среднее и строится сводная.
"""

from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ALMATY = ZoneInfo('Asia/Almaty')

HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(color='6B7280')

DATE_FMT = 'DD.MM.YYYY HH:MM'
DAY_FMT = 'DD.MM.YYYY'
PCT_FMT = '0%'
# Доля, посчитанная сервером с десятой (охват, доля пустых запросов): на экране
# у неё видна десятая, и округлять её в файле до целых значило бы показывать
# рядом с экраном другое число.
PCT1_FMT = '0.0%'


def now_almaty():
    """«Сейчас» в том же виде, в каком лежат метки в базе, — без зоны."""
    return datetime.now(ALMATY).replace(tzinfo=None)


def parse(value):
    """ISO-строка из слоя запросов → datetime без таймзоны.

    Пустое и неразобранное отдаём как None: пустая ячейка честнее строки
    «None», а падать выгрузка из-за одной кривой метки не должна.
    """
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value))
    except (TypeError, ValueError):
        return None
    return parsed.replace(tzinfo=None)


def _spec(column):
    """Колонка — (ключ, заголовок, ширина[, формат]).

    Формат необязателен: у ключей на `_at` он подразумевается датой со
    временем. Явный формат нужен там, где дата без времени (день читается
    как «22.08.2026», а не «22.08.2026 00:00») или где значение — доля.
    """
    key, title, width = column[0], column[1], column[2]
    fmt = column[3] if len(column) > 3 else (DATE_FMT if key.endswith('_at') else None)
    return key, title, width, fmt


def header(sheet, columns, row=1):
    for index, column in enumerate(columns, start=1):
        _key, title, width, _fmt = _spec(column)
        cell = sheet.cell(row=row, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[row].height = 28
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)


def autofilter(sheet, columns, rows):
    if not rows:
        return
    sheet.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(columns)), rows + 1)


def put(sheet, row, column, value, *, date=False, fmt=None):
    cell = sheet.cell(row=row, column=column, value=value)
    if fmt is None and date:
        fmt = DATE_FMT
    # Формат вешаем только на заполненную ячейку: у пустой он не виден, зато
    # превращает «нет данных» в «ноль» при копировании столбца.
    if fmt and value not in (None, ''):
        cell.number_format = fmt
    return cell


def fill_table(sheet, columns, rows, values_of):
    """Лист-таблица целиком: шапка, строки, автофильтр.

    values_of(row) → словарь по ключам колонок. Ключ, которого в словаре нет,
    даёт пустую ячейку — так добавление колонки не роняет старую выгрузку.
    """
    header(sheet, columns)
    for index, row in enumerate(rows, start=2):
        values = values_of(row)
        for column, spec in enumerate(columns, start=1):
            key, _title, _width, fmt = _spec(spec)
            put(sheet, index, column, values.get(key), fmt=fmt)
    autofilter(sheet, columns, len(rows))
    return sheet


def note(sheet, text):
    """Оговорка серым в первой колонке — под последней заполненной строкой."""
    cell = sheet.cell(row=sheet.max_row + 1, column=1, value=text)
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    return cell


def period_words(since, until):
    """Период словами. Пустые границы — «за всё время», а не пустая строка:
    файл живёт своей жизнью, и через месяц по одному числу уже не сказать, за
    какие дни оно посчитано."""
    if since and until:
        return 'с %s по %s' % (since, until)
    if since:
        return 'с %s' % since
    if until:
        return 'по %s' % until
    return 'за всё время'


def report_filename(prefix, generated_at=None):
    """Имя файла: латиница, дата, никаких слэшей.

    Часть имени приходит из адреса запроса (ключ тренажёра), поэтому всё, кроме
    букв, цифр и дефиса с подчёркиванием, вырезается: «../../etc/passwd» в
    Content-Disposition — это не выгрузка, а находка для того, кто её искал.
    """
    generated_at = generated_at or datetime.now(ALMATY)
    safe = ''.join(ch for ch in str(prefix or 'report')
                   if ch.isalnum() or ch in '-_')[:48] or 'report'
    return '%s_%s.xlsx' % (safe, generated_at.strftime('%Y-%m-%d'))
