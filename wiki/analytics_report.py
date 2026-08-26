# -*- coding: utf-8 -*-
"""Выгрузка вкладки «Аналитика» в xlsx.

ФАЙЛ ПОВТОРЯЕТ ЭКРАН, А НЕ ПЕРЕСКАЗЫВАЕТ ЕГО. Лист на таблицу, и названия
листов — те же, что заголовки таблиц во вкладке: человек, открывший файл после
экрана, обязан узнавать, что где лежит, без объяснений. Порядок листов — тот же
порядок вопросов: пользуются ли викой, в каком состоянии содержимое, выполнено
ли обязательное, чего в базе не хватает.

ЧЕМ ФАЙЛ ОТЛИЧАЕТСЯ ОТ ЭКРАНА, И ЭТО РАЗНИЦА ПО ЗАМЫСЛУ:

1. В нём ВСЁ. Экран режет таблицы потолком в сотню строк и листает их по пять —
   там это правильно, читать отчёт из девяти лент невозможно. Файл открывают
   ровно затем, чтобы отфильтровать и пересчитать, поэтому потолок у выгрузки
   свой и высокий (routes_analytics.EXPORT_ROWS).
2. Оговорки собраны на «Контексте», а не разложены к своим числам. На экране
   они живут под «i» у каждого показателя; в книге подсказок нет, а серая
   строка под таблицей попадает в выделение и уезжает в сводную как данные.
   Поэтому все оговорки — первым листом, вместе с периодом и автором.
3. Проценты лежат ДОЛЕЙ, а даты — датами. «35%» строкой не усредняется, а по
   тексту «22.08.2026» не работает фильтр по периоду — а это первое, что с
   выгрузкой делают.

СЛОВА НА ЭКРАНЕ И В ФАЙЛЕ ОБЯЗАНЫ СОВПАДАТЬ. Подписи статусов ознакомления,
названия каналов и причин «без ответа» продублированы здесь из
src/components/wiki/WikiAnalytics.jsx намеренно: сервер отдаёт фронту коды, а не
подписи, и другого места, где эти же коды превращаются в русские слова, нет.
Разъедутся — человек, сверяющий файл с экраном, решит, что это разные отчёты.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from .report_kit import (
    DATE_FMT, DAY_FMT, PCT1_FMT, PCT_FMT, TITLE_FONT, fill_table, note,
    now_almaty, parse, period_words,
)
from .report_kit import report_filename as report_kit_filename

# Подписи кодов. Зеркало WikiAnalytics.jsx — см. шапку модуля.
ACK_STATUS = {
    'not_open': 'не открывал',
    'in_progress': 'открыл',
    'read_completed': 'дочитал',
    'overdue': 'просрочено',
    'requires_reacknowledgement': 'нужна переподпись',
}

CHANNEL = {'search': 'Поиск', 'assistant': 'Помощник'}

# Причина, по которой спрос остался без ответа, и что из неё следует делать.
# Расшифровка нужна: «Ответ придержан» сам по себе не подсказывает действия, а
# действия у причин разные — написать статью, выдать доступ, уточнить числа.
REASON = {
    'missing': ('Нет статьи',
                'по теме не нашлось текста ни поиском, ни помощником — это и есть '
                'дыра в базе знаний'),
    'rights': ('Не выдан доступ',
               'статья есть — тому же запросу она нашлась у других; спрашивавшему '
               'её не выдали'),
    'empty_perimeter': ('Доступа нет ни к чему',
                        'человеку не выдан ни один раздел, поэтому поиск пуст на '
                        'любой запрос'),
    'unverified': ('Ответ придержан',
                   'помощник нашёл текст, но не смог подтвердить числа фрагментами '
                   'и промолчал — статью стоит уточнить'),
    'refused': ('Помощник отказал',
                'модель не стала отвечать своими словами по найденному тексту'),
}

YES_NO = {True: 'да', False: 'нет'}

# ── Колонки листов ───────────────────────────────────────────────────────────
#
# Идентификаторов статей и людей в файле нет намеренно: в переписке по такой
# выгрузке ссылаются на название и на имя, а лишняя колонка чисел — это тот же
# визуальный шум, только в Excel.

DAYS_COLUMNS = (
    ('day', 'Дата', 14, DAY_FMT),
    ('reads', 'Прочтений', 12),
    ('readers', 'Читателей', 12),
)

DEPARTMENT_COLUMNS = (
    ('name', 'Отдел', 28),
    ('readers', 'Читателей', 12),
    ('headcount', 'В штате', 11),
    ('share', 'Доля отдела', 13, PCT_FMT),
    ('reads', 'Прочтений', 12),
    ('articles_read', 'Статей', 10),
)

TOP_COLUMNS = (
    ('title', 'Статья', 52),
    ('status', 'Состояние', 14),
    ('reads', 'Прочтений', 12),
    ('readers', 'Читателей', 12),
    ('updated_at', 'Обновлена', 14, DAY_FMT),
)

# «Читали хоть раз» — отдельной колонкой «да/нет», а не пустой ячейкой даты:
# по ней фильтруют в первую очередь, а фильтр «пусто» находится не сразу и не
# отличает «никогда» от «метку потеряли».
UNREAD_COLUMNS = (
    ('title', 'Статья', 52),
    ('ever_read', 'Читали хоть раз', 16),
    ('last_at', 'Последний раз читали', 20, DAY_FMT),
    ('days_ago', 'Дней с последнего чтения', 16),
    ('updated_at', 'Обновлена', 14, DAY_FMT),
)

READER_COLUMNS = (
    ('name', 'Человек', 30),
    ('department', 'Отдел', 24),
    ('reads', 'Прочтений', 12),
    ('articles', 'Статей', 10),
    ('last_at', 'Последний заход', 18),
)

SECTION_COLUMNS = (
    ('parent', 'Родительский раздел', 24),
    ('name', 'Раздел', 30),
    ('articles', 'Статей', 10),
    ('published', 'Опубликовано', 14),
    ('last_update', 'Последняя правка', 18, DAY_FMT),
    ('editors', 'Правили за период', 44),
)

STALE_COLUMNS = (
    ('title', 'Статья', 52),
    ('section', 'Раздел', 24),
    ('editor', 'Правил последним', 26),
    ('days', 'Не обновляли, дней', 16),
    ('updated_at', 'Последняя правка', 18, DAY_FMT),
    ('review_overdue', 'Просрочен пересмотр', 18),
)

ACK_DEPARTMENT_COLUMNS = (
    ('name', 'Отдел', 28),
    ('total', 'Назначено', 12),
    ('done', 'Подтверждено', 14),
    ('share', 'Доля подтверждения', 17, PCT_FMT),
    ('overdue', 'Просрочено', 12),
)

# Руководитель в файле есть, хотя на экране его нет: экран отвечает на вопрос
# «кто просрочил», а по файлу пишут письма — и адресат письма именно он.
ACK_OVERDUE_COLUMNS = (
    ('name', 'Человек', 30),
    ('department', 'Отдел', 24),
    ('team', 'Группа', 20),
    ('supervisor', 'Руководитель', 26),
    ('title', 'Статья', 44),
    ('status', 'Состояние', 18),
    ('due_at', 'Срок был', 14, DAY_FMT),
    ('days', 'Дней просрочки', 14),
)

DEMAND_COLUMNS = (
    ('text', 'Запрос или вопрос', 52),
    ('channel', 'Откуда', 12),
    ('reason', 'Почему без ответа', 22),
    ('times', 'Спрашивали', 12),
    ('people', 'Людей', 10),
    ('last_at', 'Последний раз', 16, DAY_FMT),
)


def build_workbook(*, report, space_name='', requested_by='', since=None,
                   until=None, generated_at=None, row_cap=None):
    """Возвращает BytesIO с готовой книгой.

    report — ответ роута /analytics целиком (reading / content /
    acknowledgements / demand / notes / scoped). Собирать книгу из готового
    ответа, а не из отдельных запросов, — решение: так на экране и в файле
    гарантированно одни и те же числа, посчитанные одним кодом.

    row_cap — потолок строк, с которым отчёт считался. Нужен, чтобы честно
    сказать на «Контексте», что список упёрся в потолок; молчаливый обрез
    читается как «больше и нет».
    """
    generated_at = generated_at or now_almaty()
    reading = report.get('reading') or {}
    content = report.get('content') or {}
    ack = report.get('acknowledgements') or {}
    demand = report.get('demand') or {}

    workbook = Workbook()
    _fill_context(workbook.active, report, space_name=space_name,
                  requested_by=requested_by, since=since, until=until,
                  generated_at=generated_at, row_cap=row_cap)

    fill_table(workbook.create_sheet('Прочтения по дням'), DAYS_COLUMNS,
               reading.get('days') or [], _day_values)
    fill_table(workbook.create_sheet('Кто читает по отделам'), DEPARTMENT_COLUMNS,
               reading.get('departments') or [], _department_values)
    fill_table(workbook.create_sheet('Что читают чаще всего'), TOP_COLUMNS,
               reading.get('top') or [], _top_values)
    fill_table(workbook.create_sheet('Не открывали ни разу'), UNREAD_COLUMNS,
               reading.get('unread') or [],
               lambda row: _unread_values(row, generated_at))
    fill_table(workbook.create_sheet('Кто пользовался вики'), READER_COLUMNS,
               reading.get('people') or [], _reader_values)

    fill_table(workbook.create_sheet('Разделы'), SECTION_COLUMNS,
               content.get('sections') or [], _section_values)
    fill_table(workbook.create_sheet('Давно не обновляли'), STALE_COLUMNS,
               content.get('stale') or [], _stale_values)

    fill_table(workbook.create_sheet('Ознакомления по отделам'),
               ACK_DEPARTMENT_COLUMNS, ack.get('departments') or [],
               _ack_department_values)
    fill_table(workbook.create_sheet('Просрочено поимённо'), ACK_OVERDUE_COLUMNS,
               ack.get('overdue') or [], _ack_overdue_values)

    fill_table(workbook.create_sheet('Темы без ответа'), DEMAND_COLUMNS,
               demand.get('items') or [], _demand_values)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def report_filename(generated_at=None):
    return report_kit_filename('wiki_analytics', generated_at)


# ── Лист «Контекст» ──────────────────────────────────────────────────────────

def _share(part, whole):
    """Доля числом от нуля до единицы. Ноль в знаменателе — пустая ячейка, а не
    ноль: «подтверждено 0%» при отсутствии назначений — это неправда."""
    if not whole:
        return None
    return round((part or 0) / whole, 4)


def _pct(value):
    """Процент, посчитанный сервером (42.5), — в долю для формата Excel."""
    return None if value is None else round(value / 100.0, 4)


def _fill_context(sheet, report, *, space_name, requested_by, since, until,
                  generated_at, row_cap):
    sheet.title = 'Контекст'
    sheet.column_dimensions['A'].width = 38
    sheet.column_dimensions['B'].width = 56

    sheet['A1'] = 'Аналитика вики'
    sheet['A1'].font = TITLE_FONT

    reading = report.get('reading') or {}
    content = report.get('content') or {}
    ack = report.get('acknowledgements') or {}
    demand = report.get('demand') or {}
    totals = reading.get('totals') or {}
    ack_totals = ack.get('totals') or {}
    search = demand.get('search') or {}
    assistant = demand.get('assistant') or {}
    scoped = bool(report.get('scoped'))

    rows = [
        ('Пространство', space_name or 'все доступные'),
        ('Период', period_words(since, until)),
        ('Охват выгрузки', 'только ваши отделы' if scoped else 'вся вики'),
        ('Выгрузку собрал', requested_by or ''),
        ('Дата выгрузки', generated_at, DATE_FMT),
        ('', ''),

        ('ЧТЕНИЕ И ОХВАТ', ''),
        ('Прочтений', totals.get('reads', 0)),
        ('Открытий', totals.get('opens', 0)),
        ('Читателей', totals.get('readers', 0)),
        ('Опубликовано статей', totals.get('published', 0)),
        ('Из них открывали за период', totals.get('articles_read', 0)),
        ('Охват статей', _pct(totals.get('coverage')), PCT1_FMT),
        ('Статей без чтений', totals.get('unread', 0)),
        ('', ''),

        ('СОДЕРЖИМОЕ БАЗЫ', ''),
        ('Разделов', len(content.get('sections') or [])),
        ('Порог устаревания, дней', content.get('stale_days')),
        ('Статей давно не обновляли', content.get('stale_total', 0)),
        ('', ''),

        ('ОЗНАКОМЛЕНИЯ', ''),
        ('Назначено', ack_totals.get('total', 0)),
        ('Подтверждено', ack_totals.get('done', 0)),
        ('Не открывали', ack_totals.get('not_open', 0)),
        ('Просрочено', ack_totals.get('overdue', 0)),
        ('Людей в назначениях', ack_totals.get('people', 0)),
        ('Статей в назначениях', ack_totals.get('articles', 0)),
        ('', ''),

        ('ЧЕГО НЕ ХВАТАЕТ В БАЗЕ', ''),
        ('Запросов в поиске', search.get('total', 0)),
        ('Ничего не нашли', search.get('empty', 0)),
        ('Доля пустых запросов', _pct(search.get('empty_share')), PCT1_FMT),
        ('Журнал поиска ведётся с', parse(search.get('logging_since')), DAY_FMT),
        ('Вопросов помощнику', assistant.get('total', 0)),
        ('Помощник не нашёл ответа', assistant.get('no_answer', 0)),
        ('Просил уточнить вопрос', assistant.get('clarify', 0)),
    ]

    for index, row in enumerate(rows, start=3):
        label, value = row[0], row[1]
        fmt = row[2] if len(row) > 2 else None
        sheet.cell(row=index, column=1, value=label).font = Font(bold=bool(label))
        cell = sheet.cell(row=index, column=2, value=value)
        if fmt and value not in (None, ''):
            cell.number_format = fmt
        elif isinstance(value, datetime):
            cell.number_format = DATE_FMT

    sheet.cell(row=sheet.max_row + 2, column=1)
    for text in _notes(report, row_cap):
        note(sheet, text)


def _notes(report, row_cap):
    """Оговорки к книге: сначала те, что пришли с сервера вместе с числами,
    потом те, что касаются только файла.

    Порядок ключей задан явно, а не взят из словаря: он повторяет порядок
    блоков в книге, и «оговорка про сужение» обязана стоять рядом с остальными
    про людей, а не там, куда её положил хеш.
    """
    server = report.get('notes') or {}
    texts = [server[key] for key in ('empty', 'read', 'ack_now', 'scoped')
             if server.get(key)]

    texts.append('Период не действует на листы «Ознакомления по отделам» и '
                 '«Просрочено поимённо»: там состояние на дату выгрузки.')
    texts.append('Пустая ячейка «Последний раз читали» на листе «Не открывали '
                 'ни разу» означает, что статью не открывали никогда, — для '
                 'фильтра рядом стоит колонка «Читали хоть раз».')
    texts.append('«Правили за период» на листе «Разделы» — тройка самых частых '
                 'авторов версий за выбранный период, а не все правившие.')

    reasons = {row.get('reason') for row in ((report.get('demand') or {}).get('items') or [])}
    legend = ' '.join('«%s» — %s.' % (REASON[key][0], REASON[key][1])
                      for key in REASON if key in reasons)
    if legend:
        texts.append('Колонка «Почему без ответа»: ' + legend)

    for title, rows in _capped(report, row_cap):
        texts.append('Лист «%s» упёрся в потолок выгрузки — %d строк; в вике их '
                     'больше.' % (title, rows))
    return texts


def _capped(report, row_cap):
    """Какие листы уперлись в потолок строк. Молчаливый обрез читается как
    «больше и нет», а это тот случай, когда файл врёт не ошибкой, а умолчанием."""
    if not row_cap:
        return []
    reading = report.get('reading') or {}
    content = report.get('content') or {}
    ack = report.get('acknowledgements') or {}
    demand = report.get('demand') or {}
    lists = (
        ('Что читают чаще всего', reading.get('top')),
        ('Разделы', content.get('sections')),
        ('Не открывали ни разу', reading.get('unread')),
        ('Кто пользовался вики', reading.get('people')),
        ('Давно не обновляли', content.get('stale')),
        ('Просрочено поимённо', ack.get('overdue')),
        ('Темы без ответа', demand.get('items')),
    )
    return [(title, len(rows)) for title, rows in lists
            if rows is not None and len(rows) >= row_cap]


# ── Значения строк ───────────────────────────────────────────────────────────

def _day_values(row):
    return {'day': parse(row.get('day')),
            'reads': row.get('reads') or 0,
            'readers': row.get('readers') or 0}


def _department_values(row):
    return {
        'name': row.get('name'),
        'readers': row.get('readers') or 0,
        # Штат — живой, а прочтения по снимку отдела: у «Без отдела»
        # знаменателя нет вовсе, и доля там пустая, а не нулевая.
        'headcount': row.get('headcount') or None,
        'share': _share(row.get('readers'), row.get('headcount')),
        'reads': row.get('reads') or 0,
        'articles_read': row.get('articles_read') or 0,
    }


def _top_values(row):
    return {
        'title': row.get('title'),
        'status': 'опубликована' if row.get('status') == 'published' else 'черновик',
        'reads': row.get('reads') or 0,
        'readers': row.get('readers') or 0,
        'updated_at': parse(row.get('updated_at')),
    }


def _unread_values(row, generated_at):
    last = parse(row.get('last_at'))
    return {
        'title': row.get('title'),
        'ever_read': YES_NO[bool(last)],
        'last_at': last,
        'days_ago': max(0, (generated_at - last).days) if last else None,
        'updated_at': parse(row.get('updated_at')),
    }


def _reader_values(row):
    return {
        'name': row.get('name'),
        'department': row.get('department') or '',
        'reads': row.get('reads') or 0,
        'articles': row.get('articles') or 0,
        'last_at': parse(row.get('last_at')),
    }


def _section_values(row):
    editors = row.get('editors') or []
    return {
        'parent': row.get('parent') or '',
        'name': row.get('name'),
        'articles': row.get('articles') or 0,
        'published': row.get('published') or 0,
        'last_update': parse(row.get('last_update')),
        'editors': ', '.join('%s · %s' % (e.get('name'), e.get('edits'))
                             for e in editors),
    }


def _stale_values(row):
    return {
        'title': row.get('title'),
        'section': row.get('section') or '',
        'editor': row.get('editor') or '',
        'days': row.get('days'),
        'updated_at': parse(row.get('updated_at')),
        'review_overdue': YES_NO[bool(row.get('review_overdue'))],
    }


def _ack_department_values(row):
    return {
        'name': row.get('name'),
        'total': row.get('total') or 0,
        'done': row.get('done') or 0,
        'share': _share(row.get('done'), row.get('total')),
        'overdue': row.get('overdue') or 0,
    }


def _ack_overdue_values(row):
    status = row.get('status')
    return {
        'name': row.get('name') or '—',
        'department': row.get('department') or '',
        'team': row.get('team') or '',
        'supervisor': row.get('supervisor') or '',
        'title': row.get('title'),
        'status': ACK_STATUS.get(status, status or ''),
        'due_at': parse(row.get('due_at')),
        'days': row.get('days'),
    }


def _demand_values(row):
    reason = row.get('reason')
    return {
        'text': row.get('text'),
        'channel': CHANNEL.get(row.get('channel'), row.get('channel') or ''),
        'reason': REASON.get(reason, (reason or '', ''))[0],
        'times': row.get('times') or 0,
        'people': row.get('people') or 0,
        'last_at': parse(row.get('last_at')),
    }
