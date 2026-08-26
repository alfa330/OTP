"""Выгрузка статистики одного тренажёра в xlsx.

Четыре листа, и порядок у них смысловой — тот же, что в выгрузке «Провайдер
ЭДО» (fleet_edm/report.py), с которой этот модуль намеренно похож:

  «Контекст»    — что за файл, за какой период, кем и когда собран. Первым,
                  потому что цифра без даты и без оговорок живёт своей жизнью:
                  через месяц по одному числу уже не сказать, за какие дни оно.
  «Прохождения» — по строке на попытку. Это исходник, из которого посчитаны
                  остальные листы, и именно его пересчитывают вручную, когда
                  сводке не верят.
  «По людям»    — кто и сколько раз. Тот самый вопрос, ради которого выгрузку
                  и просят.
  «По статьям»  — откуда заходили. Отвечает «в какой статье сколько раз».

Длительность лежит ЧИСЛОМ секунд, а не строкой «3 мин 20 с»: по строке нельзя
ни отсортировать, ни посчитать среднее, а сводную таблицу по такой выгрузке
собирают в первый же день. Читаемая подпись стоит рядом отдельной колонкой.

Дат в ячейках две формы по той же причине: настоящая дата (её понимает
сортировка и фильтр Excel) плюс формат, в котором её читают глазами.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from .report_kit import (
    DATE_FMT, TITLE_FONT, fill_table, note, now_almaty, parse, period_words,
)
from .report_kit import report_filename as report_kit_filename

STATUS_TITLES = {
    'finished': 'Прошёл',
    'started': 'Не завершил',
    'abandoned': 'Бросил',
}

SOURCE_TITLES = {
    'article': 'Из статьи',
    'catalog': 'Из вкладки «Тренажёры»',
}

ROLE_TITLES = {
    'super_admin': 'Супер-админ',
    'admin': 'Администратор',
    'sv': 'Супервайзер',
    'supervisor': 'Супервайзер',
    'trainer': 'Тренер',
    'operator': 'Оператор',
    'trainee': 'Стажёр',
}

RUN_COLUMNS = (
    ('started_at', 'Начал', 18),
    ('name', 'ФИО', 30),
    ('department', 'Отдел', 20),
    ('group', 'Группа', 20),
    ('role', 'Должность', 16),
    ('status', 'Результат', 15),
    ('progress', 'Дошёл до шага', 15),
    ('errors', 'Промахов', 11),
    ('hints', 'Подсказок', 11),
    ('restarts', 'Начинал заново', 15),
    ('seconds', 'Время, с', 11),
    ('human_time', 'Время', 11),
    ('source', 'Откуда запустил', 22),
    ('article_title', 'Статья', 40),
    ('finished_at', 'Закончил', 18),
)

PERSON_COLUMNS = (
    ('name', 'ФИО', 30),
    ('department', 'Отдел', 20),
    ('group', 'Группа', 20),
    ('role', 'Должность', 16),
    ('runs', 'Попыток', 10),
    ('finished', 'Прошёл', 10),
    ('errors', 'Промахов всего', 15),
    ('hints', 'Подсказок всего', 16),
    ('best_seconds', 'Лучшее время, с', 16),
    ('first_at', 'Первая попытка', 18),
    ('last_at', 'Последняя попытка', 18),
)

ARTICLE_COLUMNS = (
    ('title', 'Статья', 46),
    ('runs', 'Запусков', 11),
    ('finished', 'Прошли', 10),
    ('people', 'Человек', 10),
    ('last_at', 'Последний запуск', 18),
)


def _human_time(seconds):
    if seconds is None:
        return ''
    minutes, rest = divmod(int(seconds), 60)
    return '%d:%02d' % (minutes, rest)


def _seconds(duration_ms):
    return None if duration_ms is None else round(duration_ms / 1000)


def build_workbook(*, trainer, totals, runs, people, articles,
                   since=None, until=None, requested_by='', generated_at=None):
    """Возвращает BytesIO с готовой книгой.

    trainer — {'key', 'title', 'app'}: сервер сценариев не знает, название
    приходит из фронта вместе с запросом. Ключ пишем рядом с названием, потому
    что по ключу выгрузку и опознают в переписке.
    """
    generated_at = generated_at or now_almaty()

    workbook = Workbook()
    _fill_context(workbook.active, trainer, totals, since, until,
                  requested_by, generated_at)
    fill_table(workbook.create_sheet('Прохождения'), RUN_COLUMNS, runs, _run_values)
    fill_table(workbook.create_sheet('По людям'), PERSON_COLUMNS, people, _person_values)
    fill_table(workbook.create_sheet('По статьям'), ARTICLE_COLUMNS, articles,
               _article_values)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def report_filename(trainer_key, generated_at=None):
    """Имя файла узнают по ключу сценария, поэтому он в имени и стоит."""
    return report_kit_filename('trainer_%s' % (trainer_key or 'trainer'), generated_at)


# ── Листы ────────────────────────────────────────────────────────────────────

def _fill_context(sheet, trainer, totals, since, until, requested_by, generated_at):
    sheet.title = 'Контекст'
    sheet.column_dimensions['A'].width = 34
    sheet.column_dimensions['B'].width = 52

    sheet['A1'] = 'Статистика тренажёра'
    sheet['A1'].font = TITLE_FONT

    rows = [
        ('Тренажёр', trainer.get('title') or trainer.get('key')),
        ('Ключ сценария', trainer.get('key')),
        ('Приложение', trainer.get('app') or ''),
        ('Период', period_words(since, until)),
        ('Выгрузку собрал', requested_by or ''),
        ('Дата выгрузки', generated_at),
        ('', ''),
        ('Запусков', totals.get('runs', 0)),
        ('Из них дошли до конца', totals.get('finished', 0)),
        ('Человек садилось', totals.get('people', 0)),
        ('Человек прошло', totals.get('people_done', 0)),
        ('Медианное время прохождения, с', _seconds(totals.get('median_ms'))),
        ('Промахов на прохождение (среднее)', totals.get('avg_errors')),
        ('Подсказок на прохождение (среднее)', totals.get('avg_hints')),
        ('Начинали заново', totals.get('restarts', 0)),
        ('Первый запуск', parse(totals.get('first_at'))),
        ('Последний запуск', parse(totals.get('last_at'))),
    ]
    for index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=index, column=1, value=label).font = Font(bold=bool(label))
        cell = sheet.cell(row=index, column=2, value=value)
        if isinstance(value, datetime):
            cell.number_format = DATE_FMT

    sheet.cell(row=sheet.max_row + 2, column=1)
    note(sheet, '«Человек садилось» и «человек прошло» считаются по разным людям, а не '
         'по попыткам: один и тот же оператор мог пройти тренажёр трижды.')
    note(sheet, 'Медиана, а не среднее: одна попытка, оставленная открытой на двадцать '
         'минут, сдвигает среднее так, что цифра перестаёт что-либо значить.')
    note(sheet, 'Время и промахи считаются только по завершённым попыткам — у брошенной '
         'ни то, ни другое не окончательно.')
    note(sheet, 'Строка со статусом «Не завершил» — попытка, которую человек начал и '
         'закрыл, не дойдя до конца. Это не ошибка учёта, а рабочая цифра: по ней '
         'видно, на каком шаге инструкция теряет людей.')


def _run_values(run):
    seconds = _seconds(run.get('duration_ms'))
    return {
        'started_at': parse(run.get('started_at')),
        'name': run.get('name'),
        'department': run.get('department') or '',
        'group': run.get('group') or '',
        'role': ROLE_TITLES.get(run.get('role'), run.get('role') or ''),
        'status': STATUS_TITLES.get(run.get('status'), run.get('status')),
        'progress': '%d из %d' % (run.get('stages_done') or 0,
                                  run.get('stages_total') or 0),
        'errors': run.get('errors') or 0,
        'hints': run.get('hints') or 0,
        'restarts': run.get('restarts') or 0,
        'seconds': seconds,
        'human_time': _human_time(seconds),
        'source': SOURCE_TITLES.get(run.get('source'), run.get('source')),
        'article_title': run.get('article_title') or '',
        'finished_at': parse(run.get('finished_at')),
    }


def _person_values(person):
    return {
        'name': person.get('name'),
        'department': person.get('department') or '',
        'group': person.get('group') or '',
        'role': ROLE_TITLES.get(person.get('role'), person.get('role') or ''),
        'runs': person.get('runs') or 0,
        'finished': person.get('finished') or 0,
        'errors': person.get('errors') or 0,
        'hints': person.get('hints') or 0,
        'best_seconds': _seconds(person.get('best_ms')),
        'first_at': parse(person.get('first_at')),
        'last_at': parse(person.get('last_at')),
    }


def _article_values(article):
    return {
        'title': article.get('title'),
        'runs': article.get('runs') or 0,
        'finished': article.get('finished') or 0,
        'people': article.get('people') or 0,
        'last_at': parse(article.get('last_at')),
    }
