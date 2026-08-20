"""Импорт документов в статьи: DOCX, PDF, XLSX/CSV, TXT/MD.

Порт wiki2.0 services/parser.ts на наш стек. Соответствие библиотек:
    mammoth (JS)  -> mammoth (Python, тот же автор и тот же style_map)
    pdf-parse     -> pypdf
    SheetJS       -> openpyxl (прямого аналога sheet_to_html в Python нет,
                     таблицу собираем сами)
    cheerio       -> BeautifulSoup

Три отличия от оригинала, все осознанные:

1. Картинки из DOCX уезжают в GCS, а не на диск. В оригинале они писались в
   backend/uploads/, а диск на Render эфемерный — по дампу прода видно, что
   одна такая ссылка уже 404, а остальные картинки редакторы вставляли base64,
   обходя механизм.

2. Нет таблицы сессий импорта. В оригинале документ загружался, складывался в
   document_import_sessions, редактировался во внешнем ONLYOFFICE и только
   потом становился статьёй. ONLYOFFICE у нас нет (в проде вики он тоже
   выключен), а джобы очистки сессий не было вовсе — на момент дампа семь из
   восьми висели в статусе active с конца июля. Здесь файл разбирается в HTML
   и сразу отдаётся в редактор: сохраняет уже человек.

3. Результат прогоняется через тот же санитайзер, что и ручная правка. В
   оригинале HTML из DOCX не санитайзился вообще — документ Word мог принести
   в базу произвольную разметку.
"""

import io
import os
import re
import uuid

from .sanitize import sanitize_html, to_plain_text

MAX_FILE_BYTES = 25 * 1024 * 1024

SUPPORTED = {
    '.docx': 'Word', '.doc': 'Word',
    '.pdf': 'PDF',
    '.xlsx': 'Excel', '.xlsm': 'Excel', '.csv': 'CSV',
    '.txt': 'Текст', '.md': 'Markdown',
    '.html': 'HTML', '.htm': 'HTML',
}

# Заголовки Word -> наши. Список повторяет оригинал: без него Word-документ
# превращается в сплошную простыню абзацев, и оглавление статьи не строится.
_STYLE_MAP = [
    "p[style-name='Heading 1'] => h1:fresh",
    "p[style-name='Heading 2'] => h2:fresh",
    "p[style-name='Heading 3'] => h3:fresh",
    "p[style-name='Heading 4'] => h4:fresh",
    "p[style-name='Title'] => h1:fresh",
    "p[style-name='Subtitle'] => h2:fresh",
    "p[style-name='Заголовок 1'] => h1:fresh",
    "p[style-name='Заголовок 2'] => h2:fresh",
    "p[style-name='Заголовок 3'] => h3:fresh",
]


class ImportError_(Exception):
    """Ошибка разбора документа, которую можно показать человеку."""


def _paragraphs_to_html(text):
    """Текст -> абзацы. Пустая строка разделяет абзацы, одиночный перевод — <br>."""
    blocks = []
    for block in re.split(r'\n\s*\n', str(text or '')):
        block = block.strip()
        if not block:
            continue
        escaped = (block.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))
        blocks.append('<p>%s</p>' % escaped.replace('\n', '<br>'))
    return ''.join(blocks)


def title_from_filename(name):
    base = os.path.splitext(os.path.basename(str(name or '')))[0]
    base = re.sub(r'[_\-]+', ' ', base).strip()
    return base[:255] or 'Импортированный документ'


# ─────────────────────────────────────────────────────────────────────────────
# DOCX
# ─────────────────────────────────────────────────────────────────────────────

def _convert_docx(data, *, store_image):
    import mammoth

    images = []

    def convert_image(image):
        """Каждая картинка Word уходит в GCS и получает постоянный адрес.

        store_image возвращает URL вида /api/wiki/file/<uuid> — он не протухает,
        в отличие от подписанной ссылки, и проверяет доступ при каждом запросе.
        """
        try:
            with image.open() as stream:
                blob = stream.read()
            url = store_image(blob, image.content_type or 'image/png')
            if not url:
                return {'src': ''}
            images.append(url)
            return {'src': url}
        except Exception:
            return {'src': ''}

    result = mammoth.convert_to_html(
        io.BytesIO(data),
        style_map='\n'.join(_STYLE_MAP),
        convert_image=mammoth.images.img_element(convert_image),
    )
    raw_text = mammoth.extract_raw_text(io.BytesIO(data)).value
    warnings = [str(m) for m in (result.messages or [])][:20]
    return result.value, raw_text, images, warnings


# ─────────────────────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────────────────────

def _convert_pdf(data):
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    if getattr(reader, 'is_encrypted', False):
        raise ImportError_('PDF защищён паролем — снимите защиту и попробуйте снова')

    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or '')
        except Exception:
            pages.append('')

    text = '\n\n'.join(p.strip() for p in pages if p.strip())
    if not text.strip():
        raise ImportError_(
            'В PDF нет текстового слоя — похоже, это скан. '
            'Такой файл можно приложить к статье, но не превратить в текст'
        )
    return _paragraphs_to_html(text), text, [], []


# ─────────────────────────────────────────────────────────────────────────────
# Таблицы
# ─────────────────────────────────────────────────────────────────────────────

def _cell_to_text(value):
    if value is None:
        return ''
    text = str(value).strip()
    return (text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))


# Порог, после которого книга читается потоком. Потоковый режим openpyxl не
# отдаёт объединённые клетки вовсе (у ReadOnlyWorksheet нет merged_cells), то
# есть двухуровневая шапка в нём разваливается — поэтому обычные файлы читаем
# полностью, а потоком только большие, где иначе рискуем памятью процесса.
_XLSX_STREAM_BYTES = 5 * 1024 * 1024


def _merge_map(sheet):
    """Карта объединённых клеток: (row, col) -> ('anchor', colspan, rowspan) | 'covered'.

    Без неё объединённая шапка теряется молча и незаметно. Проверено на файле с
    «Комиссия, %» над двумя колонками: потоковое чтение отдаёт
    ('Парк', 'Комиссия, %', None, 'Аренда, тг'), в таблице появляется пустая
    колонка, а вторая строка («парк», «сервис») уезжает в данные. То есть
    двухуровневая шапка документа превращается в мусор, а именно её и надо было
    понять.
    """
    anchors, covered = {}, set()
    for merged in getattr(sheet, 'merged_cells', None) and sheet.merged_cells.ranges or ():
        anchors[(merged.min_row, merged.min_col)] = (
            merged.max_col - merged.min_col + 1, merged.max_row - merged.min_row + 1)
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                if (row, col) != (merged.min_row, merged.min_col):
                    covered.add((row, col))
    return anchors, covered


def _header_depth(rows, anchors):
    """Сколько первых строк — шапка. Двухуровневая шапка это норма, а не редкость.

    Признак второго уровня: в первой строке есть клетка на две колонки и больше,
    а во второй строке заполнено ровно под ней. Так выглядит «Комиссия, %» с
    подписями «парк» и «сервис» под ней — и без этого правила подписи уходят в
    данные, где выглядят строкой таблицы.
    """
    if len(rows) < 2:
        return 1
    wide = [(col, span) for (row, col), (span, _rspan) in anchors.items()
            if row == 1 and span >= 2]
    if not wide:
        return 1
    second = rows[1]
    under = [col for col, span in wide for col in range(col, col + span)]
    filled_under = any(second[col - 1] not in (None, '')
                       for col in under if col - 1 < len(second))
    filled_outside = any(value not in (None, '')
                         for index, value in enumerate(second, start=1)
                         if index not in under)
    return 2 if filled_under and not filled_outside else 1


def _sheet_to_html(sheet, rows, anchors, covered):
    depth = _header_depth(rows, anchors)
    html = ['<table><thead>']
    for row_index, row in enumerate(rows, start=1):
        if row_index == depth + 1:
            html.append('</thead><tbody>')
        cell_tag = 'th' if row_index <= depth else 'td'
        html.append('<tr>')
        for col_index, value in enumerate(row, start=1):
            if (row_index, col_index) in covered:
                continue          # клетку уже занял её якорь через colspan/rowspan
            span_col, span_row = anchors.get((row_index, col_index), (1, 1))
            attrs = ''
            if span_col > 1:
                attrs += ' colspan="%d"' % span_col
            if span_row > 1:
                attrs += ' rowspan="%d"' % span_row
            html.append('<%s%s>%s</%s>' % (cell_tag, attrs, _cell_to_text(value), cell_tag))
        html.append('</tr>')
    html.append('</tbody></table>' if len(rows) > depth else '</thead></table>')
    return ''.join(html)


def _convert_xlsx(data):
    import openpyxl

    streamed = len(data) > _XLSX_STREAM_BYTES
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True,
                                      read_only=streamed)
    parts, plain, warnings = [], [], []
    if streamed:
        warnings.append('Файл больше %d МБ — читаем потоком, объединённые клетки '
                        'шапки могут не сохраниться'
                        % (_XLSX_STREAM_BYTES // (1024 * 1024)))

    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        # Полностью пустые хвостовые строки Excel отдаёт охотно — в таблице они
        # выглядят пустыми строками.
        while rows and all(value in (None, '') for value in rows[-1]):
            rows.pop()
        if not rows:
            continue
        anchors, covered = ({}, set()) if streamed else _merge_map(sheet)
        parts.append('<h3>%s</h3>' % _cell_to_text(sheet.title))
        parts.append(_sheet_to_html(sheet, rows, anchors, covered))
        for row in rows:
            plain.append(' '.join(str(c) for c in row if c is not None))

    workbook.close()
    if not parts:
        raise ImportError_('Файл не содержит данных')
    return ''.join(parts), '\n'.join(plain), [], warnings


def _convert_csv(data):
    import csv

    for encoding in ('utf-8-sig', 'cp1251', 'utf-8'):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ImportError_('Не удалось определить кодировку файла')

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ';' if sample.count(';') > sample.count(',') else ','

    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        raise ImportError_('Файл пуст')

    html = ['<table><thead><tr>']
    html += ['<th>%s</th>' % _cell_to_text(c) for c in rows[0]]
    html.append('</tr></thead><tbody>')
    for row in rows[1:]:
        html.append('<tr>')
        html += ['<td>%s</td>' % _cell_to_text(c) for c in row]
        html.append('</tr>')
    html.append('</tbody></table>')

    plain = '\n'.join(' '.join(row) for row in rows)
    return ''.join(html), plain, [], []


# ─────────────────────────────────────────────────────────────────────────────
# HTML
# ─────────────────────────────────────────────────────────────────────────────

def _decode_html(data):
    """Байты HTML -> текст. Кодировка берётся из самого файла, а не угадывается.

    Выгрузки из Word и старых порталов до сих пор приходят в windows-1251, и
    декодирование их как utf-8 даёт не ошибку, а страницу из «Ð¿Ñ€Ð¸Ð²ÐµÑ‚»:
    статья импортируется «успешно» и оказывается нечитаемой.
    """
    if data[:3] == b'\xef\xbb\xbf':
        return data.decode('utf-8-sig', errors='replace')

    candidates = []
    declared = re.search(rb'charset=["\']?\s*([A-Za-z0-9_\-]+)', data[:4096], re.I)
    if declared:
        candidates.append(declared.group(1).decode('ascii', 'ignore'))
    candidates += ['utf-8', 'cp1251']
    for encoding in candidates:
        try:
            return data.decode(encoding)
        except (LookupError, UnicodeDecodeError, ValueError):
            continue
    return data.decode('utf-8', errors='replace')


def _convert_html(data):
    """HTML-файл -> тело статьи. Возвращает (html, plain, images, warnings, title).

    Разметку НЕ пересобираем: в HTML уже есть заголовки, списки и таблицы —
    ровно то, ради чего импорт и затевается. Работа здесь в том, чтобы убрать
    всё, что к содержанию не относится, и честно предупредить про картинки.

    Картинки — единственное, что импорт HTML принести не может. Файл ссылается
    на них снаружи: либо адресом чужого сайта (он однажды перестанет
    открываться), либо относительным путём к папке рядом с файлом (её нам никто
    не передавал). Скачивать их отсюда нельзя — это сетевые запросы из
    обработчика, который всё это время держит соединение из общего пула на 40.
    Поэтому не молчим, а пишем в замечания: человек увидит их сразу после
    разбора и решит сам.
    """
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(_decode_html(data), 'html.parser')

    title = ''
    if soup.title:
        title = soup.title.get_text(strip=True)

    # head целиком: без этого <title> и содержимое <style> попадают в текст
    # статьи простым текстом — теги санитайзер снимет, а буквы оставит.
    for tag in soup(['script', 'style', 'noscript', 'iframe', 'head', 'meta', 'link']):
        tag.decompose()

    body = soup.body or soup

    # Заголовок документа предпочитаем <h1>: в выгрузках Word <title> сплошь
    # «Microsoft Word - договор_2026.doc», а h1 — то, как документ называется
    # для человека.
    heading = body.find(['h1', 'h2'])
    if heading:
        heading_text = heading.get_text(strip=True)
        if heading_text:
            title = heading_text

    external, relative = 0, 0
    for img in body.find_all('img'):
        src = str(img.get('src') or '').strip()
        if src.startswith('data:'):
            continue                      # картинка лежит в самом файле — дойдёт
        if src.startswith(('http://', 'https://', '//')):
            external += 1
        elif src:
            relative += 1
            # Битому адресу в статье не место: он оставит рамку с крестиком.
            img.decompose()

    warnings = []
    if external:
        warnings.append('Картинок осталось ссылками на внешний сайт: %d. Если сайт '
                        'закроется, они пропадут — перезалейте их в статью.' % external)
    if relative:
        warnings.append('Картинок не перенеслось: %d. В файле они лежат отдельными '
                        'файлами рядом, добавьте их в статью вручную.' % relative)

    html = body.decode_contents() if hasattr(body, 'decode_contents') else str(body)
    return html, body.get_text('\n', strip=True), [], warnings, (title or '')[:255]


# ─────────────────────────────────────────────────────────────────────────────
# Точка входа
# ─────────────────────────────────────────────────────────────────────────────

def convert(filename, data, *, store_image=None):
    """Документ -> материал для редактора.

    Возвращает {title, content, summary, images, warnings, kind}.
    content уже прошёл санитизацию — тем же кодом, что и ручная правка.
    """
    if not data:
        raise ImportError_('Пустой файл')
    if len(data) > MAX_FILE_BYTES:
        raise ImportError_('Файл больше %d МБ' % (MAX_FILE_BYTES // (1024 * 1024)))

    ext = os.path.splitext(str(filename or ''))[1].lower()
    if ext not in SUPPORTED:
        raise ImportError_(
            'Формат не поддерживается. Можно: %s' % ', '.join(sorted(set(SUPPORTED.values())))
        )

    # Заголовок из САМОГО документа: его знает только HTML (у остальных форматов
    # его взять неоткуда, там остаётся имя файла).
    doc_title = None

    if ext in ('.docx', '.doc'):
        if store_image is None:
            raise ImportError_('Нет хранилища для картинок документа')
        html, plain, images, warnings = _convert_docx(data, store_image=store_image)
    elif ext == '.pdf':
        html, plain, images, warnings = _convert_pdf(data)
    elif ext in ('.xlsx', '.xlsm'):
        html, plain, images, warnings = _convert_xlsx(data)
    elif ext == '.csv':
        html, plain, images, warnings = _convert_csv(data)
    elif ext in ('.html', '.htm'):
        html, plain, images, warnings, doc_title = _convert_html(data)
    else:
        text = data.decode('utf-8', errors='replace')
        html, plain, images, warnings = _paragraphs_to_html(text), text, [], []

    # Тот же санитайзер, что и для ручной правки: документ Word вполне может
    # принести произвольную разметку, а в оригинале импорт не чистился вовсе.
    clean = sanitize_html(html)
    summary = to_plain_text(clean, limit=280)

    return {
        'title': doc_title or title_from_filename(filename),
        'content': clean,
        'summary': summary,
        'images': images,
        'warnings': warnings,
        'kind': SUPPORTED[ext],
        'plain_length': len(plain or ''),
    }


def blob_path_for(original_name, content_type=''):
    """Путь в бакете. Раскладываем по дате, как это делает LMS."""
    import datetime

    safe = re.sub(r'[^A-Za-z0-9._-]+', '_', os.path.basename(str(original_name or 'file')))[:80]
    if not safe or safe.startswith('.'):
        safe = 'file' + (safe or '')
    today = datetime.datetime.now().strftime('%Y/%m/%d')
    return 'wiki/files/%s/%s_%s' % (today, uuid.uuid4().hex, safe)


def pdf_links(data, limit=40):
    """Ссылки PDF: адрес + текст, на котором она стоит.

    Отдельная функция, потому что в PDF адрес НЕ ЧАСТЬ ТЕКСТА. Он лежит в
    аннотации страницы с прямоугольником, а на странице видно только слова «по
    ссылке» или «форму регистрации». Отсюда следствие, ради которого всё это и
    написано: модель, читающая PDF глазами, адрес увидеть не может физически —
    ни vision, ни pypdf-текст его не содержат.

    Замер на боевом файле «Акции 24.07.2026 - коррект.pdf»: 7 аннотаций, 3
    уникальных адреса (форма регистрации в акциях, форма-редактор, админка
    yataxi), и в собранной статье из них не оказалось НИ ОДНОГО — вместо ссылки
    стоял пустой href="#". Для реестра акций это половина пользы: «добавить в
    форму регистрации» без адреса формы делать нечего.

    Ярлык ищется по попаданию текстовых кусков в прямоугольник аннотации.
    Совпадение приблизительное (±4 пункта), потому что базовая линия строки и
    рамка ссылки в PDF не совпадают никогда.
    """
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    found, seen = [], set()
    for page_number, page in enumerate(reader.pages, start=1):
        annotations = []
        for raw in (page.get('/Annots') or []):
            try:
                annotation = raw.get_object()
            except Exception:
                continue
            uri = (annotation.get('/A') or {}).get('/URI')
            rect = annotation.get('/Rect')
            if uri and rect:
                try:
                    annotations.append((str(uri), [float(v) for v in rect]))
                except (TypeError, ValueError):
                    continue
        if not annotations:
            continue

        pieces = []

        def visitor(text, _cm, tm, _font, _size, pieces=pieces):
            value = (text or '').strip()
            if value:
                pieces.append((tm[4], tm[5], value))

        try:
            page.extract_text(visitor_text=visitor)
        except Exception:
            pieces = []

        for uri, (x0, y0, x1, y1) in annotations:
            key = uri.rstrip('/')
            if key in seen:
                continue
            seen.add(key)
            low_x, high_x = min(x0, x1), max(x0, x1)
            low_y, high_y = min(y0, y1), max(y0, y1)
            inside = [value for x, y, value in pieces
                      if low_x - 2 <= x <= high_x + 2 and low_y - 4 <= y <= high_y + 4]
            label = ' '.join(' '.join(inside).split()).strip(' ,;:')[:80]
            found.append({'url': uri, 'label': label, 'page': page_number})
            if len(found) >= limit:
                return found
    return found
