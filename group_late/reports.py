"""Excel-отчёт по посещаемости из данных Workpace.

Перенесено из прежнего сервиса group_late_bot без изменений в самом отчёте:
структура листов, подсчёты и цветовая индикация статусов те же — люди к этой
выгрузке привыкли. Отличия только в обвязке: клиент Workpace синхронный, а
показатели готового отчёта отдаются наружу для карточки в разделе на сайте.
"""

import logging
from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from group_late import config
from group_late.departments import (
    build_employee_department_lookup,
    clean_department_filters as _normalize_department_filters,
    department_matches as _department_matches,
    resolve_department_name,
)
from group_late.helpers import (
    employee_id as _employee_id,
    employee_name as _employee_name,
    is_archived as _is_archived,
    mark_date as _mark_date,
    mark_type as _mark_type,
    parse_dt as _parse_dt,
)
from group_late.workpace import workpace_client

logger = logging.getLogger(__name__)


def _format_department_filter(dept_filter) -> str:
    department_filters = _normalize_department_filters(dept_filter)
    return ", ".join(department_filters) if department_filters else "Все отделы"


def generate_report(
    start_date_str: str,
    end_date_str: Optional[str] = None,
    dept_filter=None,
    stats_out: Optional[dict] = None,
) -> Tuple[Optional[bytes], str, str]:
    """Excel-отчёт по посещаемости за день или период, с фильтром по отделам.

    Синхронная: крутится в пуле потоков, поэтому Workpace дёргаем напрямую.
    `stats_out` — необязательный словарь, куда складываются показатели готового
    отчёта (строки, сотрудники, опоздания, неявки) для карточки в разделе
    «Бот опозданий»; при ошибке остаётся пустым. Имя не `stats`: внутри так
    называется сводка по сотруднику, и параметр затирался бы ею.
    """
    # 1. Parse start and end dates
    try:
        start_date = datetime.strptime(start_date_str.strip(), "%Y-%m-%d")
    except ValueError:
        return None, "", "❌ Неверный формат даты начала. Используйте ГГГГ-ММ-ДД (например, 2026-05-18)."

    if end_date_str and end_date_str.strip():
        try:
            end_date = datetime.strptime(end_date_str.strip(), "%Y-%m-%d")
        except ValueError:
            return None, "", "❌ Неверный формат даты конца. Используйте ГГГГ-ММ-ДД (например, 2026-05-20)."
    else:
        end_date = start_date

    if start_date > end_date:
        return None, "", "❌ Дата начала не может быть позже даты окончания."

    diff_days = (end_date - start_date).days
    if diff_days > config.MAX_REPORT_DAYS:
        return None, "", f"❌ Период отчета не должен превышать {config.MAX_REPORT_DAYS} дн."

    threshold = config.LATE_THRESHOLD_MINUTES
    department_filters = _normalize_department_filters(dept_filter)
    department_filter_label = _format_department_filter(dept_filter)

    try:
        employee_lookup = build_employee_department_lookup(workpace_client.get_employees())
    except Exception as exc:
        logger.error("Failed to fetch employees for department lookup: %s", exc)
        return None, "", f"❌ Ошибка получения списка сотрудников Workpace для определения отделов:\n<code>{exc}</code>"

    # Colors and Fonts
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)

    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_even = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green
    fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")    # light red
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # light yellow

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    wb = Workbook()
    
    # We will accumulate period statistics per employee here
    emp_summary = {}

    # Accumulators for Telegram Caption Summary
    total_unique_employees = set()
    total_row_count = 0
    total_ontime_days = 0
    total_late_count = 0
    total_absent_count = 0
    total_early_out_count = 0
    total_suspicious_marks_count = 0

    is_period = start_date != end_date
    current_date = start_date

    # To maintain sheet insertion order correctly
    first_sheet = True

    while current_date <= end_date:
        date_iso = current_date.strftime("%Y-%m-%d")
        start_local = current_date.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
        end_local = current_date.replace(hour=23, minute=59, second=59, microsecond=999999, tzinfo=None)

        try:
            records = workpace_client.get_timetable_spans(start_local, end_local)
            marks = workpace_client.get_marks(start_local, end_local)
        except Exception as exc:
            logger.error("Failed to fetch data for %s: %s", date_iso, exc)
            return None, "", f"❌ Ошибка получения данных от Workpace API за {date_iso}:\n<code>{exc}</code>"

        # Group data
        emp_data = {}

        for rec in records:
            if _is_archived(rec):
                continue
            emp_id = _employee_id(rec)
            if not emp_id:
                continue
            emp_name = _employee_name(rec)
            dept_name = resolve_department_name(rec, employee_lookup)
            if emp_id not in emp_data:
                emp_data[emp_id] = {
                    "span": {**rec, "departmentName": dept_name},
                    "marks": [],
                    "name": emp_name,
                    "dept": dept_name,
                }
            else:
                emp_data[emp_id]["span"] = {**rec, "departmentName": dept_name}
                emp_data[emp_id]["dept"] = dept_name

        for m in marks:
            if m.get("status") == 0:
                total_suspicious_marks_count += 1
            if _is_archived(m):
                continue
            emp_id = _employee_id(m)
            if not emp_id:
                continue
            emp_name = _employee_name(m)
            dept_name = resolve_department_name(m, employee_lookup)
            m = {**m, "departmentName": dept_name}
            if emp_id not in emp_data:
                emp_data[emp_id] = {
                    "span": None,
                    "marks": [m],
                    "name": emp_name,
                    "dept": dept_name,
                }
            else:
                emp_data[emp_id]["marks"].append(m)
                emp_data[emp_id]["dept"] = dept_name

        # Apply department filtering if specified
        if department_filters:
            emp_data = {
                k: v for k, v in emp_data.items()
                if _department_matches(v["dept"], department_filters)
            }

        # If we have no data and it's a single day, or if it's a period we still create the sheet but with warning
        if not emp_data and not is_period:
            return None, "", f"ℹ️ Данные по отделу <b>{department_filter_label}</b> за {date_iso} отсутствуют в Workpace."

        # Create or fetch sheet for this day
        if first_sheet:
            ws = wb.active
            ws.title = date_iso
            first_sheet = False
        else:
            ws = wb.create_sheet(title=date_iso)

        # Set up daily sheet
        ws.append([f"Отчет по посещаемости за {date_iso}"])
        ws.cell(row=1, column=1).font = font_title
        ws.row_dimensions[1].height = 30
        ws.append([]) # Spacer

        headers = [
            "Отдел", "ФИО сотрудника", "График", "Все отметки за день",
            "Время прихода (план)", "Время прихода (факт)", "Опоздание (мин)",
            "Время ухода (план)", "Время ухода (факт)", "Ранний уход (мин)",
            "Отработано (факт)", "Отклонение (HH:MM)", "Статус"
        ]
        ws.append(headers)
        ws.row_dimensions[3].height = 25

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        # Group employees by department
        dept_groups = {}
        for emp_id, info in emp_data.items():
            dept = info["dept"]
            if dept not in dept_groups:
                dept_groups[dept] = []
            dept_groups[dept].append(info)

        current_row = 4
        for dept_name, emps in sorted(dept_groups.items()):
            for info in sorted(emps, key=lambda x: x["name"]):
                name = info["name"]
                span = info["span"]
                raw_marks = info["marks"]
                
                # Update unique employees set
                total_unique_employees.add(name)
                total_row_count += 1

                # Format raw marks list
                formatted_marks = []
                for m in raw_marks:
                    m_dt = _parse_dt(_mark_date(m))
                    m_time = m_dt.strftime("%H:%M") if m_dt else "??:??"
                    mtype_val = _mark_type(m)
                    mtype = "Вход" if mtype_val == 0 else "Выход" if mtype_val == 1 else "Отм"
                    is_suspicious = m.get("status") == 0
                    susp_prefix = "⚠️ " if is_suspicious else ""
                    formatted_marks.append(f"{susp_prefix}{m_time}({mtype})")
                marks_str = ", ".join(formatted_marks) if formatted_marks else "нет отметок"

                plan_in = "—"
                fact_in = "—"
                late_in = 0
                plan_out = "—"
                fact_out = "—"
                early_out = 0
                status_text = "Вне графика"
                status_fill = fill_yellow

                fact_in_dt = None
                fact_out_dt = None

                if span:
                    plan_in_dt = _parse_dt(span.get("workTimeStart"))
                    fact_in_dt = _parse_dt(span.get("inMark"))
                    
                    if not fact_in_dt and plan_in_dt:
                        in_marks = [m for m in raw_marks if _mark_type(m) == 0]
                        if in_marks:
                            in_marks.sort(key=lambda x: _mark_date(x) or "")
                            fact_in_dt = _parse_dt(_mark_date(in_marks[0]))

                    plan_in = plan_in_dt.strftime("%H:%M") if plan_in_dt else "—"
                    fact_in = fact_in_dt.strftime("%H:%M") if fact_in_dt else "—"
                    
                    if plan_in_dt and fact_in_dt:
                        diff_mins = (fact_in_dt - plan_in_dt).total_seconds() / 60
                        late_in = max(0, int(diff_mins))
                    else:
                        late_in = span.get("lateIn") or 0

                    plan_out_dt = _parse_dt(span.get("workTimeEnd"))
                    fact_out_dt = _parse_dt(span.get("outMark"))
                    
                    if not fact_out_dt and plan_out_dt:
                        out_marks = [m for m in raw_marks if _mark_type(m) == 1]
                        if out_marks:
                            out_marks.sort(key=lambda x: _mark_date(x) or "")
                            fact_out_dt = _parse_dt(_mark_date(out_marks[-1]))

                    plan_out = plan_out_dt.strftime("%H:%M") if plan_out_dt else "—"
                    fact_out = fact_out_dt.strftime("%H:%M") if fact_out_dt else "—"

                    if plan_out_dt and fact_out_dt:
                        diff_out_mins = (plan_out_dt - fact_out_dt).total_seconds() / 60
                        early_out = max(0, int(diff_out_mins))
                    else:
                        early_out = span.get("earlyOut") or 0

                    if not fact_in_dt:
                        status_text = "Отсутствует"
                        status_fill = fill_red
                        total_absent_count += 1
                    else:
                        status_parts = []
                        if late_in >= threshold:
                            status_parts.append(f"Опоздал ({late_in} м)")
                            status_fill = fill_red
                            total_late_count += 1
                        if early_out >= threshold:
                            status_parts.append(f"Ранний уход ({early_out} м)")
                            status_fill = fill_red
                            total_early_out_count += 1
                        
                        in_marks = [m for m in raw_marks if _mark_type(m) == 0]
                        is_susp = in_marks and in_marks[0].get("status") == 0
                        
                        if not status_parts:
                            if is_susp:
                                status_parts.append("Вовремя (⚠️)")
                                status_fill = fill_yellow
                            else:
                                status_parts.append("Вовремя")
                                status_fill = fill_green
                            total_ontime_days += 1
                        else:
                            if is_susp:
                                status_parts.append("(⚠️)")
                            status_fill = fill_red
                        status_text = ", ".join(status_parts)
                else:
                    if raw_marks:
                        in_marks = [m for m in raw_marks if _mark_type(m) == 0]
                        if in_marks:
                            in_marks.sort(key=lambda x: _mark_date(x) or "")
                            fact_in_dt = _parse_dt(_mark_date(in_marks[0]))
                            fact_in = fact_in_dt.strftime("%H:%M") if fact_in_dt else "—"

                        out_marks = [m for m in raw_marks if _mark_type(m) == 1]
                        if out_marks:
                            out_marks.sort(key=lambda x: _mark_date(x) or "")
                            fact_out_dt = _parse_dt(_mark_date(out_marks[-1]))
                            fact_out = fact_out_dt.strftime("%H:%M") if fact_out_dt else "—"

                        status_text = "Вне графика"
                        status_fill = fill_yellow
                    else:
                        status_text = "Не явился"
                        status_fill = fill_red
                        total_absent_count += 1

                work_time_str = "—"
                work_seconds = 0
                if fact_in_dt and fact_out_dt:
                    diff_sec = (fact_out_dt - fact_in_dt).total_seconds()
                    if diff_sec > 0:
                        work_seconds = diff_sec
                        h = int(diff_sec // 3600)
                        m = int((diff_sec % 3600) // 60)
                        work_time_str = f"{h:02d}:{m:02d}"

                deviation_str = "—"
                if span and fact_in_dt and fact_out_dt:
                    plan_in_dt = _parse_dt(span.get("workTimeStart"))
                    plan_out_dt = _parse_dt(span.get("workTimeEnd"))
                    if plan_in_dt and plan_out_dt:
                        norm_sec = (plan_out_dt - plan_in_dt).total_seconds()
                        if norm_sec > 0:
                            diff_sec = work_seconds - norm_sec
                            sign = "+" if diff_sec >= 0 else "-"
                            abs_diff = abs(diff_sec)
                            h = int(abs_diff // 3600)
                            m = int((abs_diff % 3600) // 60)
                            deviation_str = f"{sign}{h:02d}:{m:02d}"

                # Update Period Aggregated Statistics
                emp_key = (dept_name, name)
                if emp_key not in emp_summary:
                    emp_summary[emp_key] = {
                        "scheduled_days": 0,
                        "worked_days": 0,
                        "ontime_days": 0,
                        "late_count": 0,
                        "late_mins": 0,
                        "early_out_count": 0,
                        "early_out_mins": 0,
                        "absent_count": 0,
                        "total_work_seconds": 0,
                    }
                
                stats = emp_summary[emp_key]
                if span:
                    stats["scheduled_days"] += 1
                if fact_in_dt:
                    stats["worked_days"] += 1
                    stats["total_work_seconds"] += work_seconds
                if status_text in ("Вовремя", "Вовремя (⚠️)"):
                    stats["ontime_days"] += 1
                if late_in >= threshold and span and fact_in_dt:
                    stats["late_count"] += 1
                    stats["late_mins"] += late_in
                if early_out >= threshold and span and fact_out_dt:
                    stats["early_out_count"] += 1
                    stats["early_out_mins"] += early_out
                if status_text in ("Отсутствует", "Не явился"):
                    stats["absent_count"] += 1

                # Append Row Data
                row_data = [
                    dept_name, name, info.get("span", {}).get("scheduleName", "—") if info.get("span") else "—",
                    marks_str, plan_in, fact_in, late_in if late_in > 0 else "—",
                    plan_out, fact_out, early_out if early_out > 0 else "—",
                    work_time_str, deviation_str, status_text
                ]
                ws.append(row_data)
                ws.row_dimensions[current_row].height = 20

                # Style the row
                for col_idx in range(1, 14):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = font_regular
                    cell.border = thin_border
                    
                    if col_idx in [1, 2, 3, 4]:
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center
                    
                    if col_idx == 13:
                        cell.fill = status_fill
                        cell.font = font_bold
                    else:
                        if current_row % 2 == 0:
                            cell.fill = fill_even
                        else:
                            cell.fill = fill_odd

                current_row += 1

        # Auto-fit columns for daily sheet
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or "")
                if cell.row == 1:
                    continue
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        current_date += timedelta(days=1)

    # 3. If it's a period report, create the Summary Sheet and place it at the front!
    if is_period:
        sum_ws = wb.create_sheet(title="Сводная по сотрудникам", index=0)
        
        sum_ws.append([f"Сводный отчет по посещаемости за период: {start_date_str} - {end_date_str}"])
        sum_ws.cell(row=1, column=1).font = font_title
        sum_ws.row_dimensions[1].height = 30
        sum_ws.append([]) # Spacer

        sum_headers = [
            "Отдел", "ФИО сотрудника", "Дней по графику", "Дней отработано", "Дней вовремя",
            "Опозданий (кол-во)", "Опозданий (всего мин)", "Ранних уходов (кол-во)", "Ранних уходов (всего мин)",
            "Прогулов/Неявок (кол-во)", "Всего отработано (HH:MM)", "Среднее время работы (HH:MM)"
        ]
        sum_ws.append(sum_headers)
        sum_ws.row_dimensions[3].height = 25

        for col_idx, header in enumerate(sum_headers, 1):
            cell = sum_ws.cell(row=3, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        sum_row = 4
        for (dept_name, name), stats in sorted(emp_summary.items()):
            tot_sec = stats["total_work_seconds"]
            work_days = stats["worked_days"]
            
            tot_hours = int(tot_sec // 3600)
            tot_mins = int((tot_sec % 3600) // 60)
            total_worked_str = f"{tot_hours:02d}:{tot_mins:02d}"

            avg_worked_str = "—"
            if work_days > 0:
                avg_sec = tot_sec / work_days
                avg_h = int(avg_sec // 3600)
                avg_m = int((avg_sec % 3600) // 60)
                avg_worked_str = f"{avg_h:02d}:{avg_m:02d}"

            sum_row_data = [
                dept_name, name, stats["scheduled_days"], stats["worked_days"], stats["ontime_days"],
                stats["late_count"], stats["late_mins"] if stats["late_mins"] > 0 else "—",
                stats["early_out_count"], stats["early_out_mins"] if stats["early_out_mins"] > 0 else "—",
                stats["absent_count"], total_worked_str, avg_worked_str
            ]
            sum_ws.append(sum_row_data)
            sum_ws.row_dimensions[sum_row].height = 20

            # Style sum row
            for col_idx in range(1, 13):
                cell = sum_ws.cell(row=sum_row, column=col_idx)
                cell.font = font_regular
                cell.border = thin_border
                
                if col_idx in [1, 2]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center
                
                if sum_row % 2 == 0:
                    cell.fill = fill_even
                else:
                    cell.fill = fill_odd

            sum_row += 1

        # Auto-fit columns for Summary Sheet
        for col in sum_ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or "")
                if cell.row == 1:
                    continue
                if len(val) > max_len:
                    max_len = len(val)
            sum_ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save workbook to memory stream
    file_stream = BytesIO()
    wb.save(file_stream)
    excel_bytes = file_stream.getvalue()

    if stats_out is not None:
        stats_out.update({
            "rows_count": total_row_count,
            "employees_count": len(total_unique_employees),
            "late_count": total_late_count,
            "absent_count": total_absent_count,
        })

    # Filename
    if is_period:
        filename = f"Period_Attendance_Report_{start_date_str}_to_{end_date_str}.xlsx"
    else:
        filename = f"Attendance_Report_{start_date_str}.xlsx"

    # Polish summary caption
    summary_lines = []
    if is_period:
        summary_lines.append(f"📊 <b>Сводный отчет по посещаемости за период</b>\n")
        summary_lines.append(f"📅 Период: <b>{start_date_str}</b> по <b>{end_date_str}</b>")
        summary_lines.append(f"🏢 Отдел: <b>{department_filter_label}</b>")
        summary_lines.append(f"\n📈 <b>Итоги за {diff_days + 1} дн.:</b>")
        summary_lines.append(f"• Всего уникальных сотрудников: <b>{len(total_unique_employees)}</b>")
        summary_lines.append(f"• Человеко-дней вовремя: <b>{total_ontime_days}</b>")
        summary_lines.append(f"• Всего опозданий: <b>{total_late_count}</b>")
        summary_lines.append(f"• Всего неявок/пропусков: <b>{total_absent_count}</b>")
        if total_early_out_count > 0:
            summary_lines.append(f"• Всего ранних уходов: <b>{total_early_out_count}</b>")
        if total_suspicious_marks_count > 0:
            summary_lines.append(f"• Подозрительные отметки: <b>⚠️ {total_suspicious_marks_count}</b>")
        summary_lines.append("\n📂 <i>В прикрепленном файле: первый лист — общая сводная статистика за весь период, а последующие листы — подробные отчеты на каждый день.</i>")
    else:
        summary_lines.append(f"📊 <b>Отчет по посещаемости за {start_date_str}</b>\n")
        summary_lines.append(f"🏢 Отдел: <b>{department_filter_label}</b>")
        summary_lines.append(f"\n📈 <b>Итоги дня по компании:</b>")
        summary_lines.append(f"• Всего уникальных сотрудников: <b>{len(total_unique_employees)}</b>")
        summary_lines.append(f"• Пришли вовремя: <b>{total_ontime_days}</b>")
        summary_lines.append(f"• Опоздали: <b>{total_late_count}</b>")
        summary_lines.append(f"• Отсутствуют: <b>{total_absent_count}</b>")
        if total_early_out_count > 0:
            summary_lines.append(f"• Ранний уход: <b>{total_early_out_count}</b>")
        if total_suspicious_marks_count > 0:
            summary_lines.append(f"• Подозрительные отметки: <b>⚠️ {total_suspicious_marks_count}</b>")
        summary_lines.append("\n📂 <i>Детальный отчет с разбивкой по отделам и всеми отметками прикреплен в Excel-файле ниже.</i>")

    return excel_bytes, filename, "\n".join(summary_lines)
