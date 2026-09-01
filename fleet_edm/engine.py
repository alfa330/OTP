"""Обход кабинета: из списка ID — таблица с провайдером ЭДО.

Экономика обхода держится на одном измерении (20.08.2026): **фильтр принимает
минимум 10 000 идентификаторов за раз**, а сотня — это ограничение размера
СТРАНИЦЫ ответа, не запроса. Значит цена запроса зависит от числа НАЙДЕННЫХ, а не
спрошенных: диспетчерская, где из списка нет никого, отвечает одним запросом, будь
в файле десять строк или десять тысяч. Второе измерение: шесть параллельных
запросов кабинет держит без роста времени ответа (590/мин против 170 в один поток),
а 429 не приходил ни разу.

Порядок работы:

1. **Разбор файла.** Ищем колонку с ID водителя и, если есть, колонку с ID парка.
   Парк — не прихоть: список парко-зависим, те же ID с чужим `x-park-id` дают 200
   и пустоту. Обычная выгрузка из кабинета колонку «ID парка» содержит.

2. **Определение парка** для строк без него: каждой диспетчерской отдаём ВЕСЬ
   остаток списка разом, диспетчерские опрашиваем параллельно. Раньше это стоило
   86 × (строк/100) запросов и было самым дорогим местом раздела; теперь — 86
   запросов плюс страница на сотню найденных. Спрашиваем при этом ПОЛНЫЙ набор
   полей, а не только id: тип занятости приезжает теми же запросами бесплатно, а
   стоит он дорого — см. пункт 3.

3. **Сотрудников парка снимаем с дорогого пути сразу.** Провайдер ЭДО бывает
   только у ИП и самозанятых; у сотрудника парка поля нет вовсе (проверено
   карточками 24.08.2026 — пусто у всех, и ни один из 12 фильтров провайдера их
   не находит). На настоящем файле сорвавшегося прогона №10 таких оказалось 955
   строк из 15 738: раньше каждая из них проходила 12 бесплодных раундов и стоила
   отдельный запрос карточки в доборе — 955 запросов ради ответа, известного
   заранее.

4. **Провайдер раундами.** Раунд на провайдера, внутри раунда парки идут
   параллельно; найденные уходят из очереди, следующий раунд спрашивает про
   остаток. «Бумажный документооборот» первым — на нём три четверти водителей.
   Потом такие же раунды по архиву: это отдельный сегмент, по умолчанию список
   его не отдаёт вовсе. Парки, где остались один-два водителя, дешевле спросить
   карточками — они уходят туда сразу.

5. **Разбор остатка.** Если после раундов не определилось много строк и тип
   занятости у них неизвестен (так бывает, когда ID парка был в файле и перебора
   не было), спрашиваем по одному запросу на парк БЕЗ фильтра провайдера. Это
   стоит два запроса на парк вместо одного на каждую строку — арифметика в
   _needs_classification().

6. **Добор карточками.** Фильтр списка молча не возвращает часть действующих
   профилей — 286 строк из 147 238 в августе и 4 из 8 800 в повторном прогоне.
   Причина неизвестна до сих пор, поэтому остаток добираем поштучно из карточки,
   где провайдер лежит значением.

7. **Контрольная сверка.** Случайная выборка результата перепроверяется ДРУГИМ
   путём — карточкой. Это не перестраховка: в августе первый (неполный) индекс
   идеально сходился со счётчиками самого кабинета, потому что счётчик считал тот
   же урезанный срез. Дефект нашла только сверка карточками. В выборку намеренно
   попадают и сотрудники парка — тот самый ярлык «провайдера не бывает» из
   пункта 3 обязан проверяться каждым прогоном, а не один раз при разработке.

ПРО ПЕРЕЗАПУСКИ. Обход умеет продолжаться с середины: найденное отдаётся наружу
контрольной точкой (`checkpoint`), а прерванная выгрузка приходит обратно с
`resume` — уже найденными строками, уже определёнными парками и списком
пройденных раундов. Причина простая: приложение живёт на Render, где каждый пуш
в main перезапускает процесс (21.08.2026 — 61 деплой, медиана промежутка 10
минут), а выгрузка на 15 тысяч строк идёт дольше. Без продолжения она просто
не доходила до конца — так и погибли оба прогона того дня.
"""

import logging
import random
import re
import threading
import time
from collections import Counter, OrderedDict
from concurrent.futures import ThreadPoolExecutor, as_completed

from .client import MAX_FILTER_IDS, FleetClient, FleetError, FleetSessionExpired

# Ниже этого числа водителей в парке дешевле спросить карточки, чем ждать раундов
# по провайдерам. Арифметика: карточка — ровно один запрос на человека, а парк,
# висящий в очереди, стоит по запросу на каждый раунд, пока не дойдёт очередь до
# ЕГО провайдера (в среднем около двух, «бумажный» идёт первым). На одном-двух
# водителях карточки выигрывают и вдобавок дают ответ из первоисточника; с трёх
# выигрывает список. Порог намеренно осторожный — не «оптимизация ради цифры».
CARDS_CHEAPER_BELOW = 3

# Поля, которые просим у списка. Провайдера среди них нет и быть не может — он
# приходит тем, ПО КАКОМУ фильтру строка нашлась.
LIST_PROJECTION = ['id', 'full_name', 'phone', 'work_status', 'employment_type']

ID_RE = re.compile(r'^[0-9a-f]{32}$', re.I)

# Сколько строк перепроверяем карточкой в конце. 25 — это ~30 секунд на любом
# объёме и при этом достаточно, чтобы поймать системную ошибку вроде потерянного
# архива (она бьёт не по одной строке, а по каждой десятой).
CONTROL_SAMPLE = 25

# Источники, полученные фильтром списка. Значение из карточки в перепроверке не
# нуждается — она и есть первоисточник.
LIST_SOURCES = ('список', 'архив')

# Ярлыки строк, прошедших подтверждение карточкой. Их ДВА, и это не педантизм:
# «список сказал правду, мы проверили» и «список соврал, взяли из карточки» —
# разные ответы, а третий ярлык, 'карточка', означает «список вовсе не отдал
# строку». Свалить их в один — значит написать в отчёте «список их не отдал» про
# строки, которые он прекрасно отдал, просто неверно.
SOURCE_VERIFIED = 'список · сверено'
SOURCE_CORRECTED = 'карточка · расхождение'
CARD_SOURCES = ('карточка', SOURCE_VERIFIED, SOURCE_CORRECTED)

# Ярлык строки, про которую кабинет отказался отвечать. Отдельный от «не найден»:
# см. _entry_unverified.
SOURCE_UNVERIFIED = 'не проверено'

# ПРОВАЙДЕРЫ, ЧЬЁ ЗНАЧЕНИЕ ОБЯЗАНО ПОДТВЕРЖДАТЬСЯ КАРТОЧКОЙ.
#
# Измерено 01.09.2026 на живом кабинете. Водитель 273a5c22… в парке 24a94d62…
# (iTaxi/Караганда): карточка говорит edm_provider='Sapar', а фильтр списка
# находит его ТОЛЬКО под 'paperdo' и ни под одним из шести остальных. Кабинет
# противоречит сам себе, и наш обход честно повторяет за списком: все 178 строк
# этого парка из выгрузки №32 совпали с корзинами фильтра, 178 из 178.
#
# Почему нельзя было заметить раньше: контрольная сверка берёт 25 случайных строк
# на весь файл, а доля брака — доли процента (266 случайных карточек из корзины
# paperdo этого парка дали 0 расхождений). Такую редкость выборка в 25 строк почти
# никогда не поймает, зато заказчица находит её сразу — она смотрит конкретного
# человека, а не выборку.
#
# Почему проверяем ИМЕННО «бумажный»: расхождение одностороннее. Список отстаёт
# и держит человека в «бумажном» после перехода на реального провайдера; обратных
# переходов (был провайдер — стал «бумажный») в природе не бывает, парки переходят
# НА ЭДО, а не с него. Значит проверять остальные шесть корзин смысла нет, а
# «бумажная» — это ровно тот ответ, ради которого файл и заказывают («кто ещё не
# перешёл»), и ровно тот, за которым идут к живому человеку.
#
# Цена: карточка — один запрос на строку, пачкой кабинет её не отдаёт (проверено:
# driver_ids, список в driver_id, ids → 400). Значением в списке провайдера тоже
# нет: 37 имён в projection дают 400 invalid field. Другого пути к правде нет.
VERIFY_BY_CARD = ('paperdo',)

# Сколько «ничьих» строк (парк не определился) готовы искать карточкой по всем
# паркам. Каждая такая строка стоит до 86 запросов.
#
# Три, а не сорок — и это измерено 24.08.2026 на живой выгрузке менеджера
# (7 242 строки): перебор диспетчерских уже спросил про эти самые ID ВСЕ 86
# диспетчерских и в обоих сегментах. Если там их нет, карточки почти всегда тоже
# нет — человек просто ушёл из парков. Зато стоил этот шаг сорок строк × до 86
# запросов, и на придушенном кабинете (20–40 отказов в минуту) занял больше, чем
# весь остальной обход: выгрузка выглядела зависшей на «Добираем из карточек».
#
# Поэтому теперь это не спасение, а ВЫБОРОЧНЫЙ КОНТРОЛЬ: три строки проверяем
# карточкой, чтобы не пропустить известный дефект списка (он молча не отдаёт
# 0,05–0,2% действующих профилей), остальные честно помечаем «не найден ни в одной
# диспетчерской» — это ровно то, что перебор и установил. Сколько строк не стали
# перебирать, отчёт пишет отдельной строкой (skipped_orphans).
MAX_ORPHAN_CARD_LOOKUPS = 3

# До скольких строк в доборе разрешаем перебор диспетчерских. Больше — значит
# добирать нужно многих, и перебор из полезного превращается в час ожидания.
MAX_CARD_SCANS = 50

# Типы занятости, у которых провайдер ЭДО бывает. У остальных поля нет вовсе:
# проверено карточками 24.08.2026 — у шести сотрудников парка edm_provider пуст,
# и ни один из 12 раундов (6 провайдеров × обычный/архив) их не находит.
PROVIDER_BEARING = ('individual_entrepreneur', 'self_employed')
PARK_EMPLOYEE = 'park_employee'

# Ярлык источника для строк, где провайдера не бывает по природе занятости.
# Отдельный, а не пустой «список»: в отчёте «нет провайдера» и «не смогли узнать
# провайдера» — это два разных ответа, и путать их нельзя.
SOURCE_NO_PROVIDER = 'сотрудник парка'

# Сколько сотрудников парка перепроверяем карточкой в контрольной сверке.
# Ярлык «провайдера не бывает» — это наше утверждение о чужой системе, и оно
# обязано подтверждаться каждым прогоном.
CONTROL_NO_PROVIDER_SAMPLE = 10

# Меньше этого числа строк в остатке разбирать списком не имеет смысла — проще
# спросить карточки (см. _needs_classification).
CLASSIFY_LEFTOVERS_FLOOR = 20

# Сколько раз возвращаемся к диспетчерским, которые не ответили, и сколько ждём
# между заходами.
#
# ЭТО САМОЕ ДОРОГОЕ МЕСТО РАЗДЕЛА ПО ЦЕНЕ ОШИБКИ. До 24.08.2026 не ответивший парк
# просто выпадал из обхода (`except FleetError: return 0`), и ВСЕ его водители
# уезжали в отчёт как «не найден ни в одной диспетчерской». На выгрузке №13 так
# случилось с парком Jana Taxi: 1 250 строк объявлены ненайденными, а на самом
# деле 1 106 из них — работающие ИП с провайдером, и список отдаёт их за 14
# запросов. Причина — не редкость: кабинет в тот момент отвечал 20–40 отказами
# «помедленнее» в минуту, и семи попыток внутри клиента не хватило.
#
# Поэтому теперь: не ответившие диспетчерские спрашиваем заново, а если и после
# трёх заходов молчат — обход честно прерывается. Врать «не найден» он больше не
# имеет права, а прерваться ему теперь не страшно: контрольная точка на месте, и
# подхват продолжит с того же места (см. шапку модуля).
FAILED_PARK_RETRIES = 3
FAILED_PARK_PAUSE = 8.0

# Как часто выгрузка отмечается в базе, пока идут длинные однообразные шаги
# (добор карточками, разбор остатка). Сторож считает молчание смертью, а на
# прогоне 21.08.2026 добор молчал двадцать минут — и был убит живым.
#
# Десять, а не сорок: одна задача добора бывает очень долгой (строка без парка
# перебирает диспетчерские), и на живой выгрузке 24.08.2026 экран полчаса показывал
# «40 из 99» — выгрузка шла, но выглядела мёртвой.
PROGRESS_EVERY = 10

WORK_STATUS_LABELS = {
    'working': 'Работает',
    'not_working': 'Не работает',
    'fired': 'Уволен',
    'busy': 'Занят',
    'in_park': 'В парке',
}

EMPLOYMENT_LABELS = {
    'individual_entrepreneur': 'ИП / самозанятый',
    'self_employed': 'Самозанятый',
    'park_employee': 'Сотрудник парка',
    'private': 'Частное лицо',
}

HEADER_ALIASES = {
    'contractor_id': (
        'contractor id', 'contractor_id', 'contractorid', 'id водителя', 'ид водителя',
        'driver id', 'driver_id', 'id исполнителя', 'идентификатор водителя', 'id',
    ),
    'park_id': (
        'id парка', 'ид парка', 'park id', 'park_id', 'parkid', 'идентификатор парка',
    ),
    'park_name': ('название парка', 'парк', 'наименование парка', 'park', 'park name'),
    'full_name': ('фио', 'ф.и.о.', 'имя', 'водитель', 'full name', 'full_name'),
    'phone': ('телефон', 'номер телефона', 'phone', 'телефон водителя'),
}


class InputError(ValueError):
    """Файл не годится: об этом говорим человеку словами, а не 500-й ошибкой."""


def _normalize_header(value):
    return re.sub(r'\s+', ' ', str(value or '').strip().lower()).strip(' :')


# ── разбор входного файла ────────────────────────────────────────────────────

def parse_input(file_bytes, filename=''):
    """Возвращает (rows, meta). rows — список словарей с contractor_id и, если
    была колонка, park_id. Порядок строк исходного файла сохраняется: человек
    сверяет результат глазами по своему же файлу."""
    if not file_bytes:
        raise InputError('Файл пустой')

    name = str(filename or '').lower()
    if name.endswith('.csv'):
        table = _read_csv(file_bytes)
    else:
        table = _read_xlsx(file_bytes)

    if not table:
        raise InputError('В файле нет ни одной строки')

    header_index, columns = _detect_columns(table)
    if 'contractor_id' not in columns:
        raise InputError(
            'Не нашли колонку с ID водителя. Нужна колонка «Contractor ID» '
            '(подойдёт также «ID водителя» или «ID»).'
        )

    rows, seen_bad = [], 0
    for row_number, raw in enumerate(table[header_index + 1:], start=header_index + 2):
        contractor_id = _cell(raw, columns.get('contractor_id'))
        if not contractor_id:
            continue
        park_id = _cell(raw, columns.get('park_id'))
        entry = {
            'row_number': row_number,
            'contractor_id': contractor_id.lower(),
            'park_id': (park_id or '').lower(),
            'source_park_name': _cell(raw, columns.get('park_name')),
            'source_full_name': _cell(raw, columns.get('full_name')),
            'source_phone': _cell(raw, columns.get('phone')),
        }
        if not ID_RE.match(entry['contractor_id']):
            entry['error'] = 'ID не похож на идентификатор водителя Fleet'
            seen_bad += 1
        rows.append(entry)

    if not rows:
        raise InputError('В колонке с ID водителя нет ни одного значения')

    meta = {
        'columns': {key: value for key, value in columns.items()},
        'rows_total': len(rows),
        'rows_bad_id': seen_bad,
        'has_park_column': 'park_id' in columns,
        'unique_ids': len({row['contractor_id'] for row in rows}),
    }
    return rows, meta


def _read_xlsx(file_bytes):
    from io import BytesIO

    import openpyxl

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as error:
        raise InputError('Не удалось открыть файл как Excel: {}'.format(error))
    try:
        sheet = workbook[workbook.sheetnames[0]]
        # Читаем шапку и данные целиком: файлы этого раздела — десятки тысяч
        # строк в одну колонку идентификаторов, это единицы мегабайт в памяти.
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_csv(file_bytes):
    import csv
    import io

    for encoding in ('utf-8-sig', 'cp1251'):
        try:
            text = file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=';,\t')
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ';' if sample.count(';') > sample.count(',') else ','
        return [row for row in csv.reader(io.StringIO(text), dialect)]
    raise InputError('Не удалось прочитать CSV: неизвестная кодировка')


def _detect_columns(table):
    """Ищем шапку в первых пяти строках: у выгрузок из кабинета сверху бывает
    заголовок отчёта, а не сразу названия колонок."""
    for index, row in enumerate(table[:5]):
        columns = {}
        for position, value in enumerate(row or []):
            header = _normalize_header(value)
            if not header:
                continue
            for key, aliases in HEADER_ALIASES.items():
                if key in columns:
                    continue
                if header in aliases:
                    columns[key] = position
        if 'contractor_id' in columns:
            return index, columns
    # Шапки нет вовсе — файл «одна колонка ID», как и просили в задаче.
    first = table[0] if table else []
    if len(first) == 1 and ID_RE.match(str(first[0] or '').strip()):
        return -1, {'contractor_id': 0}
    return 0, {}


def _cell(row, position):
    if position is None or row is None or position >= len(row):
        return ''
    value = row[position]
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


# ── обход ────────────────────────────────────────────────────────────────────

class Cancelled(RuntimeError):
    """Выгрузку у нас забрали: карточку либо закрыли, либо подхватил другой
    процесс. Останавливаемся сразу — второй обход по тому же файлу удвоил бы темп
    запросов к чужому кабинету, а его лимит нам не принадлежит.

    Именно это и произошло 21.08.2026: сторож счёл живую выгрузку мёртвой, человек
    запустил её заново, а старый поток продолжал работать ещё двадцать минут —
    два обхода разом, и оба в 429."""


class Progress:
    """Прогресс наружу. Отдельным объектом, чтобы движок ничего не знал ни про
    базу, ни про Flask — в тестах сюда приходит список."""

    def __init__(self, callback=None):
        self._callback = callback

    def __call__(self, **payload):
        if self._callback:
            try:
                self._callback(**payload)
            except Exception:
                logging.exception('Провайдер ЭДО: не удалось записать прогресс')


class Checkpoint:
    """Контрольная точка наружу — то же разделение, что у Progress: движок знает
    про «сохрани найденное», но ничего не знает про базу.

    rows — [(contractor_id, park_id, payload|None)], stages — словарь этапов.
    Ошибка записи не роняет обход: не сохранённая контрольная точка означает
    лишний повтор после перезапуска, а упавший обход — потерянный час работы.
    """

    def __init__(self, callback=None):
        self._callback = callback

    def __call__(self, rows=None, stages=None):
        if not self._callback or (not rows and stages is None):
            return
        try:
            self._callback(rows=rows or (), stages=stages)
        except Exception:
            logging.exception('Провайдер ЭДО: контрольная точка не записана')


class Stopper:
    """Проверка «нас ещё ждут?». Спрашиваем не чаще раза в пять секунд: обход
    зовёт её из каждой задачи, а поход в базу на каждый чих — это соединение из
    общего пула."""

    def __init__(self, callback=None, interval=5.0):
        self._callback = callback
        self._interval = float(interval)
        self._checked_at = 0.0
        self._stopped = False
        self._lock = threading.Lock()

    def __call__(self):
        if not self._callback:
            return
        with self._lock:
            if self._stopped:
                raise Cancelled('Выгрузку остановили')
            now = time.time()
            if now - self._checked_at < self._interval:
                return
            self._checked_at = now
        try:
            stop = bool(self._callback())
        except Exception:
            logging.exception('Провайдер ЭДО: не удалось проверить, ждут ли выгрузку')
            return
        if stop:
            with self._lock:
                self._stopped = True
            raise Cancelled('Выгрузку остановили')


def resolve(rows, client: FleetClient, *, progress=None, control_sample=CONTROL_SAMPLE,
            max_orphan_lookups=MAX_ORPHAN_CARD_LOOKUPS, rng=None,
            checkpoint=None, resume=None, should_stop=None,
            verify_providers=VERIFY_BY_CARD, card_cache=None):
    """Главная функция: заполняет провайдера по строкам. Возвращает словарь с
    результатом по каждому уникальному ID и статистикой прогона.

    checkpoint, resume и should_stop — про перезапуски (см. шапку модуля):
    найденное отдаётся наружу по ходу дела, прерванная выгрузка приходит обратно
    с уже готовыми строками, а должна ли она вообще продолжаться — спрашивается
    у вызывающего.
    """
    progress = Progress(progress) if not isinstance(progress, Progress) else progress
    save = checkpoint if isinstance(checkpoint, Checkpoint) else Checkpoint(checkpoint)
    stop = should_stop if isinstance(should_stop, Stopper) else Stopper(should_stop)
    rng = rng or random.Random(20260820)
    resume = resume or {}
    stages = dict(resume.get('stages') or {})
    done_rounds = set(stages.get('rounds') or ())

    valid = [row for row in rows if not row.get('error')]
    known_parks = dict(resume.get('parks') or {})
    unique = OrderedDict()
    for row in valid:
        contractor_id = row['contractor_id']
        unique.setdefault(contractor_id,
                          row.get('park_id') or known_parks.get(contractor_id) or '')

    # Уже готовое с прошлой попытки. Берём только то, что есть в этом файле:
    # карточка одна на файл, но лишняя проверка дешевле, чем строка-призрак.
    results = {}
    for contractor_id, entry in (resume.get('results') or {}).items():
        if contractor_id in unique:
            results[contractor_id] = dict(entry)
            if not unique[contractor_id] and entry.get('park_id'):
                unique[contractor_id] = entry['park_id']

    # Тип занятости, если он уже известен: у сотрудника парка провайдера не
    # бывает вовсе, и знание об этом снимает с человека 12 раундов и карточку.
    kinds = {contractor_id: (entry.get('employment_type') or '')
             for contractor_id, entry in results.items()}

    if results:
        progress(percent=6, rows_total=len(rows), rows_resolved=len(results),
                 note='Продолжаем прерванную выгрузку: {} строк уже собрано'
                      .format(len(results)))
    else:
        progress(percent=4, rows_total=len(rows),
                 note='Читаем справочник парков и провайдеров')

    parks = client.parks()
    park_names = {str(park.get('id')): _park_label(park) for park in parks}
    if not parks:
        raise FleetError('Кабинет не отдал ни одного парка')

    first_park = next((park for park in unique.values() if park), '') \
        or str(parks[0].get('id'))
    providers = client.edm_providers(first_park)
    if not providers:
        raise FleetError('Кабинет не отдал справочник провайдеров ЭДО')

    stats = Counter()
    park_probe_requests = 0
    classify_requests = 0
    # Где человек лежит — в обычном сегменте или в архиве. Перебор парков говорит
    # это тем же ответом, которым сообщает парк; знание бесплатное, а экономит
    # шесть пустых запросов на каждой диспетчерской без архивных.
    segments = {}

    # ── шаг 1: у кого нет парка ──────────────────────────────────────────────
    orphans = [contractor_id for contractor_id, park in unique.items()
               if not park and contractor_id not in results]
    if orphans and not stages.get('probe_done'):
        progress(percent=8, note='В файле нет ID парка — ищем водителей по паркам '
                                '({} шт.)'.format(len(orphans)))
        before = client.requests_count
        found_parks = _probe_parks(client, orphans, parks, progress, stop=stop)
        park_probe_requests = client.requests_count - before
        checkpoint_rows = []
        for contractor_id, info in found_parks.items():
            unique[contractor_id] = info['park_id']
            kinds[contractor_id] = info.get('employment_type') or ''
            segments[contractor_id] = bool(info.get('archive'))
            if kinds[contractor_id] == PARK_EMPLOYEE:
                # Провайдера у сотрудника парка не бывает — это готовый ответ, а
                # не пропуск. Те же запросы уже принесли и ФИО, и телефон.
                results[contractor_id] = _entry_without_provider(contractor_id, info)
                checkpoint_rows.append((contractor_id, info['park_id'],
                                        results[contractor_id]))
            else:
                checkpoint_rows.append((contractor_id, info['park_id'], None))
        stats['parks_probed'] = len(found_parks)
        stages['probe_done'] = True
        save(rows=checkpoint_rows, stages=stages)

    # ── шаг 2: провайдер — раундами по провайдерам, парки параллельно ────────
    by_park = OrderedDict()
    for contractor_id, park_id in unique.items():
        if park_id and contractor_id not in results:
            by_park.setdefault(park_id, []).append(contractor_id)

    # Совсем мелкие парки дешевле спросить карточками: проход по провайдерам
    # стоит до семи запросов даже там, где в парке один человек.
    tiny = {park: ids for park, ids in by_park.items() if len(ids) < CARDS_CHEAPER_BELOW}
    big = {park: ids for park, ids in by_park.items() if len(ids) >= CARDS_CHEAPER_BELOW}

    seen_providers = Counter(stages.get('provider_counts') or {})
    leftovers = []
    round_leftovers = []
    if big:
        found, pending = _resolve_by_providers(
            client, big, providers, seen_providers, progress,
            total=len(unique), done_before=len(results),
            stop=stop, save=save, stages=stages, done_rounds=done_rounds,
            segments=segments,
        )
        results.update(found)
        round_leftovers.extend(pending)
        leftovers.extend(pending)
    for park_id, ids in tiny.items():
        leftovers.extend((park_id, contractor_id) for contractor_id in ids)

    # ── шаг 3: разбор остатка одним запросом на парк ─────────────────────────
    # Кто эти люди, мы часто уже знаем из перебора парков. Но когда ID парка был
    # в файле, перебора не было — и остаток приходится разбирать здесь, иначе
    # каждый сотрудник парка стоил бы отдельную карточку (на прогоне №10 это 955
    # запросов ради ответа «провайдера не бывает»).
    #
    # Берём только остаток ПОСЛЕ раундов. Мелкие парки (один-два человека) сюда
    # не идут: там карточка и дешевле (один запрос против двух), и честнее —
    # отвечает первоисточник, а не наш вывод из типа занятости.
    unknown = [(park_id, contractor_id) for park_id, contractor_id in round_leftovers
               if not kinds.get(contractor_id)]
    if _needs_classification(unknown):
        progress(percent=89, requests=client.requests_count,
                 note='Разбираем остаток: {} строк'.format(len(unknown)))
        before = client.requests_count
        classified = _classify_leftovers(client, unknown, progress, stop=stop)
        classify_requests = client.requests_count - before
        checkpoint_rows = []
        for contractor_id, info in classified.items():
            kinds[contractor_id] = info.get('employment_type') or ''
            if kinds[contractor_id] == PARK_EMPLOYEE:
                results[contractor_id] = _entry_without_provider(contractor_id, info)
                checkpoint_rows.append((contractor_id, info['park_id'],
                                        results[contractor_id]))
        save(rows=checkpoint_rows)
        stats['classified'] = len(classified)
        leftovers = [(park_id, contractor_id) for park_id, contractor_id in leftovers
                     if contractor_id not in results]

    # ── шаг 4: добор карточками ──────────────────────────────────────────────
    orphan_left = [contractor_id for contractor_id, park_id in unique.items()
                   if not park_id and contractor_id not in results]
    skipped_orphans = 0
    if len(orphan_left) > max_orphan_lookups:
        skipped_orphans = len(orphan_left) - max_orphan_lookups
        orphan_left = orphan_left[:max_orphan_lookups]

    to_card = [(park_id, contractor_id) for park_id, contractor_id in leftovers] + \
              [('', contractor_id) for contractor_id in orphan_left]
    if to_card:
        progress(percent=90, requests=client.requests_count,
                 note='Добираем из карточек: {} строк'.format(len(to_card)))
        # Карточки не зависят друг от друга — значит идут в те же потоки.
        # Перебор диспетчерских включаем, только когда добирать нужно немногих:
        # один такой водитель стоит до 86 запросов.
        allow_scan = len(to_card) <= MAX_CARD_SCANS

        def card_task(task):
            stop()
            try:
                return _card_lookup(client, task[1], task[0], parks, allow_scan=allow_scan)
            except FleetSessionExpired:
                raise
            except FleetError as error:
                # ЖИВОЙ КЛИНЧ, пойманный на проде 01.09.2026 (выгрузка №33).
                # Молчащие диспетчерские правильно НЕ дают сказать «не найден», и
                # раньше обход на этом прерывался целиком: контрольная точка есть,
                # подхват продолжит. Но у строки БЕЗ парка перебор идёт по всем 86
                # диспетчерским, и под нагрузкой хоть одна молчит всегда — значит
                # «продолжим позже» не наступает никогда. Шесть подхватов подряд
                # умерли на одном и том же водителе, ни один запрос сверки так и не
                # прошёл, а впереди был двенадцатый и закрытие «слишком много
                # перезапусков».
                #
                # Поэтому отказ по ОДНОЙ строке больше не роняет прогон. Врать при
                # этом не начинаем: у строки свой ответ — «не смогли проверить», и
                # это НЕ «не найден ни в одной диспетчерской».
                logging.warning('Провайдер ЭДО: карточку %s… добрать не удалось (%s)',
                                task[1][:8], str(error)[:100])
                return _entry_unverified(task[1], task[0])

        # Шаг длинный и однообразный — отмечаемся в базе по ходу, иначе сторож
        # сочтёт живую выгрузку мёртвой (так и вышло 21.08.2026 на 1068 строках).
        def card_progress(entry, done, total):
            if entry:
                save(rows=[(entry['contractor_id'], entry.get('park_id'), entry)])
            if done % PROGRESS_EVERY == 0 or done == total:
                progress(percent=90, requests=client.requests_count,
                         note='Добираем из карточек: {} из {}'.format(done, total))

        for entry in _run_parallel(client, to_card, card_task,
                                   stop=stop, on_result=card_progress):
            if not entry:
                stats['not_found'] += 1
                continue
            results[entry['contractor_id']] = entry
            if entry.get('source') == SOURCE_UNVERIFIED:
                # Не «не найден»: кабинет отказался отвечать, и в отчёте это
                # отдельная строка, а не молчаливое приписывание к пропавшим.
                stats['unverified'] += 1

    # ── шаг 4½: подтверждение «бумажных» карточками ──────────────────────────
    # Зачем это вообще есть — см. VERIFY_BY_CARD. Коротко: фильтр списка врёт про
    # часть людей, и врёт всегда в одну сторону.
    verify = _verify_by_card(
        client, results, providers, verify_providers, progress,
        stop=stop, save=save, cache=card_cache)
    stats['verified'] = verify['checked']
    stats['verify_fixed'] = len(verify['fixed'])

    # ── шаг 5: контрольная сверка ────────────────────────────────────────────
    check = {'checked': 0, 'matched': 0, 'mismatched': []}
    sample = []
    if control_sample:
        from_list = [contractor_id for contractor_id, entry in results.items()
                     if entry.get('provider_name')
                     and entry.get('source') not in CARD_SOURCES]
        sample += rng.sample(from_list, min(control_sample, len(from_list)))
        # И отдельно — те, кому мы САМИ поставили «провайдера не бывает». Это
        # наше утверждение о чужой системе, и проверять его надо каждый прогон.
        no_provider = [contractor_id for contractor_id, entry in results.items()
                       if entry.get('source') == SOURCE_NO_PROVIDER]
        sample += rng.sample(no_provider,
                             min(CONTROL_NO_PROVIDER_SAMPLE, len(no_provider)))
    if sample:
        progress(percent=95, requests=client.requests_count,
                 note='Контрольная сверка {} строк по карточкам'.format(len(sample)))

        def verify(contractor_id):
            stop()
            entry = results[contractor_id]
            try:
                profile = client.driver_card(entry.get('park_id'), contractor_id)
            except FleetSessionExpired:
                raise
            except FleetError:
                return None
            if profile is None:
                return None
            return (contractor_id, entry.get('provider_name') or '',
                    FleetClient.card_provider(profile))

        for outcome in _run_parallel(client, sample, verify, stop=stop):
            if not outcome:
                continue
            contractor_id, listed, card_value = outcome
            if not card_value and listed:
                # Карточка молчит там, где список ответил: это «поле не про
                # него», а не расхождение.
                continue
            check['checked'] += 1
            if card_value == listed:
                check['matched'] += 1
            else:
                check['mismatched'].append({
                    'contractor_id': contractor_id, 'list': listed, 'card': card_value,
                })

    # Итоговые числа считаем ПО РЕЗУЛЬТАТУ, а не счётчиками по ходу дела: после
    # перезапуска половина строк приезжает из контрольной точки, и счётчик той
    # попытки остался в умершем процессе. Отчёт же обязан говорить про файл, а не
    # про последнюю попытку его собрать.
    stats['from_card'] = sum(1 for entry in results.values()
                             if entry.get('source') == 'карточка')
    stats['no_provider_by_kind'] = sum(1 for entry in results.values()
                                       if entry.get('source') == SOURCE_NO_PROVIDER)

    return {
        'results': results,
        'providers': providers,
        'park_names': park_names,
        'parks_total': len(parks),
        'check': check,
        'requests': client.requests_count,
        'park_probe_requests': park_probe_requests,
        'classify_requests': classify_requests,
        'skipped_orphans': skipped_orphans,
        'stats': dict(stats),
        'provider_counts': dict(seen_providers),
        'verify': verify,
    }


def _verify_by_card(client, results, providers, verify_providers, progress,
                    *, stop=None, save=None, cache=None):
    """Подтверждает карточкой значения, которым фильтр списка доверять нельзя.

    Возвращает {'checked', 'fixed', 'silent', 'blank', 'requests'} — сколько
    строк проверили, что исправили, сколько карточек не ответило и у скольких
    поле оказалось пустым.

    ПОЧЕМУ ОТКАЗ КАРТОЧКИ НЕ ПЕРЕПИСЫВАЕТ СТРОКУ. Молчащая карточка — это «мы не
    спросили», а не «провайдера нет»; ровно на этой разнице раздел один раз уже
    погорел (1 250 ложных «не найден» в августе). Здесь безопасное поведение —
    оставить значение списка и честно посчитать непроверенные, а не сочинить.

    ПОЧЕМУ ПЕРЕЗАПУСК НЕ НАЧИНАЕТ ЗАНОВО. Отметку 'card_checked' кладём в саму
    строку и сразу отправляем в контрольную точку. Строки приезжают обратно
    вместе с отметкой, и вторая попытка спрашивает только неотмеченных: проход
    длинный (тысячи запросов), а деплой разрешения не спрашивает.
    """
    verify_ids = tuple(verify_providers or ())
    outcome = {'checked': 0, 'fixed': [], 'silent': 0, 'blank': 0,
               'requests': 0, 'from_cache': 0}
    if not verify_ids:
        return outcome
    save = save or Checkpoint(None)
    by_name = {provider['name']: provider['id'] for provider in providers}
    names = {provider['id']: provider['name'] for provider in providers}

    pending = [contractor_id for contractor_id, entry in results.items()
               if entry.get('provider_id') in verify_ids
               and entry.get('source') in LIST_SOURCES]
    if not pending:
        return outcome

    label = ', '.join(names.get(code, code) for code in verify_ids)
    before = client.requests_count

    def apply(contractor_id, card_value):
        """Кладёт ответ карточки в строку. Возвращает True, если значение
        пришлось поправить."""
        entry = results[contractor_id]
        if not card_value:
            # Пусто в карточке — «поле не про него». Списку тут верить не в чем,
            # но и переписывать нечем: оставляем как есть и считаем.
            outcome['blank'] += 1
            return False
        if card_value == entry.get('provider_name'):
            entry['source'] = SOURCE_VERIFIED
            return False
        outcome['fixed'].append({
            'contractor_id': contractor_id,
            'list': entry.get('provider_name') or '',
            'card': card_value,
        })
        entry['comment'] = (
            'Список кабинета отставал: показывал «{}», в карточке «{}»'
            .format(entry.get('provider_name') or '—', card_value))
        entry['provider_name'] = card_value
        entry['provider_id'] = by_name.get(card_value, '')
        entry['source'] = SOURCE_CORRECTED
        return True

    # Сперва то, что уже подтверждали раньше. Кеш общий для всех выгрузок и
    # переживает перезапуск — без него подтверждение стоило бы полную цену
    # КАЖДОМУ прогону (до 97 минут на файле в 15 738 строк), а раздел
    # запускают по нескольку раз в день.
    if cache is not None:
        known = cache.get(pending) or {}
        if known:
            rows = []
            for contractor_id, card_value in known.items():
                if contractor_id not in results:
                    continue
                apply(contractor_id, card_value)
                outcome['from_cache'] += 1
                rows.append((contractor_id, results[contractor_id].get('park_id'),
                             results[contractor_id]))
            save(rows=rows)
            pending = [contractor_id for contractor_id in pending
                       if contractor_id not in known]
            progress(percent=91, requests=client.requests_count,
                     note='Подтверждение «{}»: {} строк уже сверено раньше, '
                          'спрашиваем {}'.format(label, outcome['from_cache'], len(pending)))
    if not pending:
        outcome['requests'] = client.requests_count - before
        return outcome

    progress(percent=91, requests=client.requests_count,
             note='Подтверждаем «{}» карточками: {} строк'.format(label, len(pending)))

    def ask(contractor_id):
        if stop:
            stop()
        entry = results[contractor_id]
        try:
            profile = client.driver_card(entry.get('park_id'), contractor_id)
        except FleetSessionExpired:
            raise
        except FleetError as error:
            logging.warning('Провайдер ЭДО: карточка %s… не ответила при '
                            'подтверждении (%s)', contractor_id[:8], str(error)[:80])
            return contractor_id, None
        if profile is None:
            return contractor_id, None
        return contractor_id, FleetClient.card_provider(profile)

    fresh = []

    def tick(result, done, total):
        contractor_id, card_value = result
        if card_value is None:
            # Не спросили — значение списка остаётся, но ярлык НЕ меняем: пусть
            # следующая попытка вернётся к этой строке.
            outcome['silent'] += 1
            return
        outcome['checked'] += 1
        apply(contractor_id, card_value)
        entry = results[contractor_id]
        save(rows=[(contractor_id, entry.get('park_id'), entry)])
        fresh.append((contractor_id, entry.get('park_id'), card_value))
        if done % PROGRESS_EVERY == 0 or done == total:
            if cache is not None and fresh:
                # Кеш пополняем по ходу, а не в конце: прогон на тысячи строк
                # переживает деплой, и подтверждённое не должно умирать вместе
                # с процессом — иначе следующая попытка заплатит за него снова.
                cache.put(list(fresh))
                del fresh[:]
            progress(percent=_percent(done, total, low=91, high=94),
                     requests=client.requests_count,
                     note='Подтверждаем «{}» карточками: {} из {}, поправлено {}'
                          .format(label, done, total, len(outcome['fixed'])))

    _run_parallel(client, pending, ask, stop=stop, on_result=tick)
    if cache is not None and fresh:
        cache.put(list(fresh))
    outcome['requests'] = client.requests_count - before
    progress(percent=94, requests=client.requests_count,
             note='Подтверждение по карточкам: проверено {}, поправлено {}'
                  .format(outcome['checked'], len(outcome['fixed'])))
    return outcome


def _entry_without_provider(contractor_id, info):
    """Запись для того, кому провайдер ЭДО не положен по типу занятости.

    Пустой провайдер здесь — это ОТВЕТ, а не отсутствие ответа, поэтому у записи
    свой источник: в отчёте «ЭДО не применяется» и «не смогли узнать» обязаны
    читаться по-разному.
    """
    return {
        'contractor_id': contractor_id,
        'provider_id': '',
        'provider_name': '',
        'full_name': str(info.get('full_name') or '').strip(),
        'phone': str(info.get('phone') or '').strip(),
        'work_status': str(info.get('work_status') or '').strip(),
        'employment_type': str(info.get('employment_type') or PARK_EMPLOYEE).strip(),
        'park_id': info.get('park_id') or '',
        'source': SOURCE_NO_PROVIDER,
    }


def _entry_unverified(contractor_id, park_id):
    """Строка, про которую кабинет отказался отвечать.

    Третий возможный ответ про человека, и он обязан отличаться от двух других:
    «нашли провайдера», «водителя нет ни в одной диспетчерской» и вот это —
    «спросить не смогли». Свалить его в «не найден» значит повторить самую
    дорогую ошибку раздела (1 250 работающих ИП, объявленных ненайденными
    24.08.2026), только с другой стороны.
    """
    return {
        'contractor_id': contractor_id,
        'provider_id': '',
        'provider_name': '',
        'full_name': '',
        'phone': '',
        'work_status': '',
        'employment_type': '',
        'park_id': park_id or '',
        'source': SOURCE_UNVERIFIED,
        'comment': 'Не смогли проверить: диспетчерские не ответили',
    }


def _needs_classification(unknown):
    """Стоит ли разбирать остаток списком, а не карточками.

    Арифметика прямая: разбор стоит два запроса на парк (обычный сегмент и
    архив), карточка — один запрос на строку. Значит список выгоден, когда строк
    в остатке больше, чем по две на каждый задействованный парк. Нижний порог —
    чтобы на десятке строк не устраивать обход диспетчерских: там карточки и
    быстрее, и честнее (ответ из первоисточника).
    """
    if len(unknown) < CLASSIFY_LEFTOVERS_FLOOR:
        return False
    parks = {park_id for park_id, _contractor_id in unknown if park_id}
    return bool(parks) and len(unknown) > 2 * len(parks)


def _park_label(park):
    name = str(park.get('name') or '').strip()
    city = str(park.get('city') or '').strip()
    return '{} / {}'.format(name, city) if city else name


def _percent(done, total, low=10, high=90):
    if not total:
        return low
    return int(low + (high - low) * min(1.0, done / float(total)))


def _run_parallel(client, tasks, worker, *, stop=None, on_result=None):
    """Раскладывает задачи по потокам. Параллельность — не украшение: замерено
    20.08.2026, шесть потоков дают 590 запросов/мин против 170 в один, и медиана
    ответа при этом не растёт. Упавшая задача не роняет остальные — её строки
    просто уйдут в добор карточками.

    on_result зовётся по КАЖДОЙ доехавшей задаче, а не после всех: длинные шаги
    (добор карточками — до тысячи запросов) обязаны отмечаться в базе по ходу
    дела и складывать найденное в контрольную точку. Порядок результатов при
    этом теряется — всем вызывающим он безразличен, они кладут ответы в словари.
    """
    if not tasks:
        return []
    if stop:
        stop()
    workers = max(1, min(client.concurrency, len(tasks)))
    if workers == 1:
        outcome = worker(tasks[0])
        if on_result:
            on_result(outcome, 1, 1)
        return [outcome]
    collected = []
    with ThreadPoolExecutor(max_workers=workers, thread_name_prefix='fleet-edm-io') as pool:
        futures = [pool.submit(worker, task) for task in tasks]
        for done, future in enumerate(as_completed(futures), start=1):
            outcome = future.result()
            collected.append(outcome)
            if on_result:
                on_result(outcome, done, len(tasks))
    return collected


def _split_tasks(pending, concurrency, slice_min=1000):
    """(парк, срез ID) для одного раунда.

    Обычно на поток приходится свой парк. Но когда парков меньше, чем потоков —
    а это ровно случай «весь файл из одной диспетчерской», — режем список одного
    парка на срезы, иначе шесть потоков простаивали бы за одним.
    """
    parks = [(park, ids) for park, ids in pending.items() if ids]
    if len(parks) >= concurrency:
        return [(park, ids) for park, ids in parks]
    tasks = []
    for park, ids in parks:
        slices = max(1, min(concurrency, (len(ids) + slice_min - 1) // slice_min))
        step = (len(ids) + slices - 1) // slices
        for start in range(0, len(ids), step):
            tasks.append((park, ids[start:start + step]))
    return tasks


def _resolve_by_providers(client, by_park, providers, seen_providers, progress,
                          total, done_before=0, stop=None, save=None, stages=None,
                          done_rounds=(), segments=None):
    """Провайдер для всех сразу: раунд на провайдера, парки внутри раунда параллельно.

    Почему именно так. Спросить можно про десять тысяч ID за раз, и пустой ответ
    стоит один запрос. Значит цена раунда — это число НАЙДЕННЫХ, а не спрошенных:
    один запрос на парк плюс страница на каждую сотню совпадений. После раунда
    найденные уходят из очереди, и следующий провайдер спрашивается уже про
    остаток. «Бумажный документооборот» идёт первым не случайно — на нём три
    четверти водителей, и один раунд снимает большую часть работы.

    Раунд — это ещё и естественная граница контрольной точки: он заканчивается
    целиком или не заканчивается вовсе, и пройденные раунды после перезапуска
    можно честно не повторять (найденное ими уже в базе).

    segments — где человек лежит: в обычном сегменте или в архиве. Если перебор
    парков уже это выяснил (а он выясняет — тем же ответом, что и парк), спрашивать
    парк про архив, когда архивных в нём нет, незачем: это шесть пустых запросов
    на каждую такую диспетчерскую. Кого не знаем — спрашиваем в обоих сегментах,
    как раньше: терять архивных нельзя (в августе так потерялось 14 444 строки).
    """
    segments = segments or {}
    # Очередь ведём ПО СЕГМЕНТАМ: у обычного и архивного прохода она своя.
    pending = {False: {}, True: {}}
    for park, ids in by_park.items():
        for contractor_id in ids:
            known = segments.get(contractor_id)
            targets = (False, True) if known is None else (bool(known),)
            for target in targets:
                pending[target].setdefault(park, []).append(contractor_id)
    stages = stages if stages is not None else {}
    done_rounds = set(done_rounds or ())
    save = save or Checkpoint(None)
    found = {}
    lock = threading.Lock()

    def ask(task, provider, archive):
        park_id, ids = task
        if stop:
            stop()
        try:
            items = client.contractors_all(
                park_id, contractor_ids=ids, edm_provider=provider['id'],
                archive=archive, projection=LIST_PROJECTION,
            )
        except FleetSessionExpired:
            raise
        except FleetError as error:
            # НЕ «здесь никого нет», а «мы не спросили». Разница стоила 1 250
            # ложных «не найден» на выгрузке №13 — см. FAILED_PARK_RETRIES.
            logging.warning('Провайдер ЭДО: запрос по парку %s не удался (%s) — '
                            'вернёмся к нему', park_id, str(error)[:120])
            return park_id, None, task
        entries = {}
        for item in items:
            cid = str(item.get('id') or '')
            if cid:
                entry = _entry_from_list(item, provider, archive)
                entry['park_id'] = park_id
                entries[cid] = entry
        return park_id, entries, task

    rounds = [(archive, provider) for archive in (False, True) for provider in providers]
    for archive, provider in rounds:
        round_key = '{}|{}'.format(provider['id'], 1 if archive else 0)
        if round_key in done_rounds:
            # Раунд уже проходили в прошлой попытке, и всё найденное им лежит в
            # контрольной точке. Повторять — значит платить за это второй раз.
            continue
        tasks = _split_tasks(pending[archive], client.concurrency)
        if not tasks and not any(ids for queue in pending.values()
                                 for ids in queue.values()):
            break
        if not tasks:
            # В этом сегменте спрашивать некого — но в другом ещё есть.
            done_rounds.add(round_key)
            continue

        # Найденное кладём в контрольную точку по КАЖДОЙ доехавшей диспетчерской,
        # не дожидаясь конца раунда: раунд по «Бумажному документообороту» на
        # большом файле идёт минуты, а деплой не спрашивает разрешения. Отметку
        # «раунд пройден» ставим только после полного раунда — незаконченный
        # обязан повториться, но повторится он уже по остатку.
        def keep(outcome, _done, _total):
            if outcome and outcome[1]:
                park_id, entries, _task = outcome
                save(rows=[(cid, park_id, entry) for cid, entry in entries.items()])

        attempt = 0
        while tasks:
            results = _run_parallel(client, tasks,
                                    lambda task: ask(task, provider, archive),
                                    stop=stop, on_result=keep)
            retry = []
            with lock:
                for park_id, entries, task in results:
                    if entries is None:
                        retry.append(task)
                        continue
                    if not entries:
                        continue
                    found.update(entries)
                    seen_providers[provider['id']] += len(entries)
                    # Найденного вычёркиваем из ОБОИХ сегментов: человек с
                    # неизвестным сегментом стоит в двух очередях, и без этого его
                    # бы спросили второй раз уже после ответа.
                    for queue in pending.values():
                        if park_id in queue:
                            queue[park_id] = [cid for cid in queue[park_id]
                                              if cid not in entries]
            if not retry:
                break
            attempt += 1
            if attempt >= FAILED_PARK_RETRIES:
                # Молчат — прерываемся. Соврать «не найден» про целую
                # диспетчерскую нельзя, а продолжить позже теперь можно.
                raise FleetError(
                    'Диспетчерские не ответили после {} заходов ({} шт.) — '
                    'обход продолжится позже'.format(attempt, len(retry)))
            progress(note='Диспетчерские не ответили ({} шт.) — повторяем'
                          .format(len(retry)),
                     requests=client.requests_count)
            time.sleep(FAILED_PARK_PAUSE * attempt)
            tasks = retry
        done_rounds.add(round_key)
        stages['rounds'] = sorted(done_rounds)
        stages['provider_counts'] = dict(seen_providers)
        # Строки уже улетели в контрольную точку по ходу раунда — здесь только
        # отметка «раунд пройден целиком».
        save(stages=stages)
        done = done_before + len(found)
        progress(
            percent=_percent(done, total, low=12, high=88),
            note='{}{}: определено {} из {}'.format(
                provider['name'], ' (архив)' if archive else '', done, total),
            rows_resolved=done,
            requests=client.requests_count,
        )

    # Остаток — то, чего не нашёл ни один сегмент. Через множество, потому что
    # человек с неизвестным сегментом стоял в двух очередях сразу.
    leftovers, seen = [], set()
    for queue in pending.values():
        for park_id, ids in queue.items():
            for contractor_id in ids:
                if contractor_id in found or contractor_id in seen:
                    continue
                seen.add(contractor_id)
                leftovers.append((park_id, contractor_id))
    return found, leftovers


def _entry_from_list(item, provider, archive):
    return {
        'contractor_id': str(item.get('id') or ''),
        'provider_id': provider['id'],
        'provider_name': provider['name'],
        'full_name': str(item.get('full_name') or '').strip(),
        'phone': str(item.get('phone') or '').strip(),
        'work_status': str(item.get('work_status') or '').strip(),
        'employment_type': str(item.get('employment_type') or '').strip(),
        'source': 'архив' if archive else 'список',
    }


def _probe_parks(client, ids, parks, progress, stop=None):
    """Перебор диспетчерских для строк без парка — параллельно и одним списком.

    Здесь выигрыш от больших пачек самый большой. Раньше каждый парк спрашивали
    сотнями: восемь тысяч «ничьих» строк — это 88 запросов НА КАЖДЫЙ из 86 парков.
    Теперь парку отдают весь список сразу, и парк, где никого нет, отвечает одним
    запросом. Цена перебора перестала зависеть от длины файла.

    Два круга: обычные сегменты, затем архив — архивного водителя список по
    умолчанию не отдаёт вовсе.

    Возвращает не «ID → парк», а ID → всё, что кабинет отдал про человека. Просим
    полный набор полей вместо одного id, и это НЕ стоит ни одного лишнего запроса
    (замер 24.08.2026 на файле прогона №10: 326 запросов, 38 секунд — ровно как с
    projection=['id']). Зато среди полей есть тип занятости, а он снимает с
    сотрудников парка 12 бесплодных раундов и по запросу карточки на каждого: на
    том же файле — 955 строк из 15 738.
    """
    pending = set(ids)
    found = {}
    lock = threading.Lock()

    def probe(item):
        park, archive = item
        if stop:
            stop()
        park_id = str(park.get('id'))
        with lock:
            snapshot = list(pending)
        if not snapshot:
            return None
        try:
            items = client.contractors_all(
                park_id, contractor_ids=snapshot, archive=archive,
                projection=LIST_PROJECTION,
            )
        except FleetSessionExpired:
            raise
        except FleetError as error:
            # Молча пропустить парк здесь — это соврать про КАЖДОГО его водителя:
            # он уедет в отчёт как «не найден ни в одной диспетчерской». Именно так
            # выгрузка №13 объявила ненайденными 1 250 работающих людей из Jana
            # Taxi. Возвращаем задачу на повтор.
            logging.warning('Провайдер ЭДО: перебор парка %s не удался (%s) — '
                            'вернёмся к нему', park_id, str(error)[:120])
            return item
        with lock:
            for entry in items:
                cid = str(entry.get('id') or '')
                if cid in pending:
                    pending.discard(cid)
                    found[cid] = _info_from_list(entry, park_id, archive)
        return None

    for archive in (False, True):
        if not pending:
            break
        tasks = [(park, archive) for park in parks]
        attempt = 0
        while tasks:
            retry = [task for task in _run_parallel(client, tasks, probe, stop=stop)
                     if task]
            if not retry:
                break
            attempt += 1
            if attempt >= FAILED_PARK_RETRIES:
                raise FleetError(
                    'Диспетчерские не ответили после {} заходов ({} шт.) — '
                    'перебор продолжится позже'.format(attempt, len(retry)))
            progress(note='Диспетчерские не ответили ({} шт.) — повторяем'
                          .format(len(retry)),
                     requests=client.requests_count)
            time.sleep(FAILED_PARK_PAUSE * attempt)
            tasks = retry
        progress(
            note='Ищем диспетчерские: осталось найти {} из {}'.format(len(pending), len(ids)),
            requests=client.requests_count,
        )
    return found


def _info_from_list(item, park_id, archive):
    """Что кабинет рассказал о человеке в ответе списка — без провайдера: его в
    полях нет и быть не может (см. шапку client.py)."""
    return {
        'park_id': park_id,
        'full_name': str(item.get('full_name') or '').strip(),
        'phone': str(item.get('phone') or '').strip(),
        'work_status': str(item.get('work_status') or '').strip(),
        'employment_type': str(item.get('employment_type') or '').strip(),
        'archive': bool(archive),
    }


def _classify_leftovers(client, unknown, progress, stop=None):
    """Кто эти люди из остатка — одним запросом на парк, без фильтра провайдера.

    Нужно, когда ID парка пришёл в файле: перебора диспетчерских не было, значит
    тип занятости остатка неизвестен, и каждый сотрудник парка обошёлся бы в
    отдельный запрос карточки. Список же за один запрос отвечает про весь остаток
    парка сразу — цена зависит от числа НАЙДЕННЫХ, а не спрошенных.

    Ненайденные здесь — это не «нет провайдера», а «в этом парке такого нет»:
    они уходят в добор карточками, где парк перебирается заново.
    """
    by_park = OrderedDict()
    for park_id, contractor_id in unknown:
        if park_id:
            by_park.setdefault(park_id, []).append(contractor_id)
    found = {}
    lock = threading.Lock()

    def ask(task):
        park_id, archive = task
        if stop:
            stop()
        with lock:
            ids = [cid for cid in by_park.get(park_id, ()) if cid not in found]
        if not ids:
            return None
        try:
            items = client.contractors_all(
                park_id, contractor_ids=ids, archive=archive, projection=LIST_PROJECTION,
            )
        except FleetSessionExpired:
            raise
        except FleetError as error:
            logging.warning('Провайдер ЭДО: разбор остатка по парку %s не удался (%s) — '
                            'вернёмся к нему', park_id, str(error)[:120])
            return task
        with lock:
            for item in items:
                cid = str(item.get('id') or '')
                if cid:
                    found[cid] = _info_from_list(item, park_id, archive)
        return None

    tasks = [(park_id, archive) for archive in (False, True) for park_id in by_park]

    def tick(_outcome, done, total):
        if done % PROGRESS_EVERY == 0 or done == total:
            progress(note='Разбираем остаток: диспетчерская {} из {}'.format(done, total),
                     requests=client.requests_count)

    attempt = 0
    while tasks:
        retry = [task for task in _run_parallel(client, tasks, ask, stop=stop,
                                                on_result=tick) if task]
        if not retry:
            break
        attempt += 1
        if attempt >= FAILED_PARK_RETRIES:
            # Не разобранный остаток — это не «сотрудники парка» и не «нет
            # провайдера», это «мы не спросили». Прерываемся и продолжим позже.
            raise FleetError(
                'Диспетчерские не ответили при разборе остатка ({} шт.) — '
                'обход продолжится позже'.format(len(retry)))
        time.sleep(FAILED_PARK_PAUSE * attempt)
        tasks = retry
    return found


def _card_lookup(client, contractor_id, park_id, parks, allow_scan=True):
    """Карточка водителя.

    Парк из файла проверяем первым, но если карточки там нет — идём по остальным
    диспетчерским. Это не перестраховка: в файле заказчика парк у части строк
    указан неверно (в августовской выгрузке для этого была отдельная колонка
    «Парк совпал»), и без перебора такие строки возвращались бы пустыми.

    Перебор дорогой — до 86 запросов на одного человека, — поэтому вызывающий
    выключает его, когда добирать нужно многих.
    """
    candidates = []
    if park_id:
        candidates.append(park_id)
    if allow_scan or not park_id:
        candidates.extend(str(park.get('id')) for park in parks
                          if str(park.get('id')) != park_id)
    # Отказ кабинета — это НЕ «здесь такого нет». Если хоть одна диспетчерская не
    # ответила и водителя мы так и не нашли, честного ответа у нас нет: молчаливое
    # «не найден» здесь того же сорта, что потерянный парк в переборе.
    silent = 0
    for candidate in candidates:
        if not candidate:
            continue
        try:
            profile = client.driver_card(candidate, contractor_id)
        except FleetSessionExpired:
            raise
        except FleetError as error:
            silent += 1
            logging.warning('Провайдер ЭДО: карточка %s в парке %s не ответила (%s)',
                            contractor_id[:8], candidate[:8], str(error)[:80])
            continue
        if not profile:
            continue
        return {
            'contractor_id': contractor_id,
            'provider_id': '',
            'provider_name': FleetClient.card_provider(profile),
            'full_name': str(profile.get('full_name') or '').strip() or _card_name(profile),
            'phone': _card_phone(profile),
            'work_status': str(profile.get('work_status') or '').strip(),
            'employment_type': str(profile.get('employment_type') or '').strip(),
            'park_id': candidate,
            'source': 'карточка',
        }
    if silent:
        raise FleetError(
            'Карточка водителя {}…: {} диспетчерских не ответили, «не найден» '
            'сказать не можем'.format(contractor_id[:8], silent))
    return None


def _card_name(profile):
    name = profile.get('name') or {}
    if isinstance(name, dict):
        parts = [name.get('last'), name.get('first'), name.get('middle')]
        return ' '.join(str(part).strip() for part in parts if part)
    return ''


def _card_phone(profile):
    phones = profile.get('phones')
    if isinstance(phones, list) and phones:
        first = phones[0]
        if isinstance(first, dict):
            return str(first.get('number') or first.get('phone') or '').strip()
        return str(first).strip()
    return str(profile.get('phone') or '').strip()
