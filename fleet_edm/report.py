"""Сборка выгрузки в xlsx.

Четыре листа: «Контекст», «Водители», «Свод по провайдерам», «Провайдеры по
паркам». Первым идёт именно контекст — цифра без даты и без оговорок живёт своей
жизнью, а провайдера водителю меняют, и файл недельной давности уже не «список
провайдеров», а снимок на дату.

Телефоны и идентификаторы лежат ТЕКСТОМ: числом номер теряет ведущие нули и
уезжает в экспоненту, а 32-значный hex Excel и вовсе округлит. Зелёный уголок
«Число сохранено как текст» гасится тем же приёмом, что и в остальных выгрузках
портала (тег <ignoredErrors>) — сама функция приходит аргументом, потому что
живёт в монолите, а импортировать монолит из пакета нельзя.
"""

from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ALMATY = ZoneInfo('Asia/Almaty')

HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(color='6B7280')

COLUMNS = (
    ('park_name', 'Название парка', 34),
    ('park_id', 'ID парка', 34),
    ('contractor_id', 'ID водителя', 34),
    ('full_name', 'ФИО', 32),
    ('phone', 'Телефон', 16),
    ('work_status', 'Статус в Fleet', 16),
    ('employment_type', 'Тип сотрудничества', 20),
    ('provider_name', 'Провайдер ЭДО', 26),
    ('source', 'Источник', 12),
    ('comment', 'Комментарий', 40),
)

# Колонки, которые обязаны остаться текстом (см. шапку модуля).
TEXT_COLUMNS = ('park_id', 'contractor_id', 'phone')

NOT_FOUND = 'не найден'


def build_workbook(rows, resolution, *, source_name='', generated_at=None,
                   text_warning_patch=None):
    """rows — строки исходного файла в исходном порядке, resolution — то, что
    вернул engine.resolve(). Возвращает BytesIO с готовой книгой."""
    generated_at = generated_at or datetime.now(ALMATY)
    results = resolution.get('results') or {}
    park_names = resolution.get('park_names') or {}
    providers = resolution.get('providers') or []

    table = _rows_for_sheet(rows, results, park_names)
    counts = _provider_counts(table)
    by_park = _park_counts(table, providers)

    workbook = Workbook()
    context_sheet = workbook.active
    context_sheet.title = 'Контекст'
    _fill_context(context_sheet, table, resolution, source_name, generated_at)

    drivers_sheet = workbook.create_sheet('Водители')
    _fill_drivers(drivers_sheet, table)

    summary_sheet = workbook.create_sheet('Свод по провайдерам')
    _fill_summary(summary_sheet, counts, len(table))

    parks_sheet = workbook.create_sheet('Провайдеры по паркам')
    _fill_parks(parks_sheet, by_park, providers)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    if text_warning_patch and table:
        # Лист «Водители» — второй в книге, поэтому sheet2.xml.
        first_text_column = get_column_letter(_column_index('park_id'))
        last_text_column = get_column_letter(_column_index('phone'))
        sqref = '{}2:{}{}'.format(first_text_column, last_text_column, len(table) + 1)
        try:
            stream = text_warning_patch(stream, sqref, sheet_path='xl/worksheets/sheet2.xml')
        except Exception:
            # Значок в углу ячейки — досадно, но не повод не отдать файл.
            stream.seek(0)
    return stream


def report_filename(source_name='', generated_at=None):
    generated_at = generated_at or datetime.now(ALMATY)
    return 'Провайдер ЭДО — {}.xlsx'.format(generated_at.strftime('%d.%m.%Y %H-%M'))


# ── листы ────────────────────────────────────────────────────────────────────

def _column_index(key):
    for index, (name, _title, _width) in enumerate(COLUMNS, start=1):
        if name == key:
            return index
    return 1


def _rows_for_sheet(rows, results, park_names):
    from .engine import (EMPLOYMENT_LABELS, PARK_EMPLOYEE, SOURCE_NO_PROVIDER,
                         WORK_STATUS_LABELS)

    table = []
    for row in rows:
        contractor_id = row.get('contractor_id') or ''
        entry = results.get(contractor_id) or {}
        park_id = entry.get('park_id') or row.get('park_id') or ''
        comment = ''
        if row.get('error'):
            comment = row['error']
        elif entry.get('comment'):
            # Строку уже прокомментировал обход — например, подтверждение
            # карточкой поправило отставшее значение списка.
            comment = entry['comment']
        elif not entry:
            comment = 'Водитель не найден ни в одной диспетчерской'
        elif (entry.get('source') == SOURCE_NO_PROVIDER
              or entry.get('employment_type') == PARK_EMPLOYEE):
            # Это ОТВЕТ, а не пропуск: поле ЭДО есть только у ИП и самозанятых, а
            # сотрудник парка работает по трудовому договору — провайдера у него
            # не бывает. Формулировка отдельная, чтобы «не применяется» не читали
            # как «мы не смогли узнать».
            comment = 'Сотрудник парка — ЭДО не применяется'
        elif not entry.get('provider_name'):
            # Пусто в карточке — это не «нет провайдера», а «поле не про него»:
            # у сотрудников парка провайдера ЭДО не бывает вовсе.
            comment = 'Провайдер не указан в кабинете'
        table.append({
            'park_name': park_names.get(park_id) or row.get('source_park_name') or '',
            'park_id': park_id,
            'contractor_id': contractor_id,
            'full_name': entry.get('full_name') or row.get('source_full_name') or '',
            'phone': entry.get('phone') or row.get('source_phone') or '',
            'work_status': WORK_STATUS_LABELS.get(entry.get('work_status'),
                                                  entry.get('work_status') or ''),
            'employment_type': EMPLOYMENT_LABELS.get(entry.get('employment_type'),
                                                     entry.get('employment_type') or ''),
            'provider_name': entry.get('provider_name') or (NOT_FOUND if not entry else ''),
            'source': entry.get('source') or '',
            'comment': comment,
        })
    return table


def _provider_counts(table):
    counts = {}
    for row in table:
        name = row['provider_name'] or 'не указан'
        counts[name] = counts.get(name, 0) + 1
    return dict(sorted(counts.items(), key=lambda item: -item[1]))


def _park_counts(table, providers):
    parks = {}
    for row in table:
        key = (row['park_name'] or '—', row['park_id'])
        bucket = parks.setdefault(key, {'total': 0, 'providers': {}})
        bucket['total'] += 1
        name = row['provider_name'] or 'не указан'
        bucket['providers'][name] = bucket['providers'].get(name, 0) + 1
    return dict(sorted(parks.items(), key=lambda item: -item[1]['total']))


def _fill_context(sheet, table, resolution, source_name, generated_at):
    check = resolution.get('check') or {}
    stats = resolution.get('stats') or {}
    resolved = sum(1 for row in table if row['provider_name'] and row['provider_name'] != NOT_FOUND)
    lines = [
        ('Выгрузка «Провайдер ЭДО»', None),
        ('', None),
        ('1. Главное', None),
        ('Данные собраны из диспетчерских Яндекс.Fleet', None),
        ('Дата и время сборки', generated_at.strftime('%d.%m.%Y %H:%M') + ' (Алматы)'),
        ('Исходный файл', source_name or '—'),
        ('Строк в файле', len(table)),
        ('Провайдер определён', '{} ({:.3f}%)'.format(
            resolved, (resolved / len(table) * 100) if table else 0)),
        ('Не найдено в кабинете', sum(1 for row in table if row['provider_name'] == NOT_FOUND)),
        ('Запросов в кабинет', resolution.get('requests') or 0),
        ('', None),
        ('2. Как это проверялось', None),
        ('Контрольная сверка', 'по карточкам водителей — независимый путь, не тот, '
                               'которым собирались данные'),
        ('Сверено строк', check.get('checked') or 0),
        ('Совпало', check.get('matched') or 0),
        ('Расхождений', len(check.get('mismatched') or [])),
        ('', None),
        ('3. Что важно знать про эти данные', None),
        ('«Бумажный документооборот» — это выбранный провайдер, а не пропуск.',
         'Не путать с «не найден»: там водителя нет в кабинете.'),
        ('Провайдера можно сменить.', 'Значение верно на дату сборки, выше.'),
        ('Пустой провайдер у сотрудника парка — норма.',
         'Поле ЭДО есть только у ИП и самозанятых.'),
        ('Соединение по ID водителя, а не по ФИО.', 'Совпадение точное.'),
    ]
    verify = resolution.get('verify') or {}
    verified_total = (verify.get('checked') or 0) + (verify.get('from_cache') or 0)
    if verified_total:
        lines.append(('Подтверждено карточками',
                      '{} строк со значением «Бумажный документооборот». Фильтр '
                      'списка кабинета по такой строке иногда расходится с '
                      'карточкой того же кабинета, поэтому каждую спросили у '
                      'первоисточника'.format(verified_total)))
        if verify.get('from_cache'):
            lines.append(('Из них спрошено раньше',
                          '{} строк взяты из подтверждений прошлых выгрузок (срок '
                          'годности — неделя), спрошено заново {}'.format(
                              verify['from_cache'], verify.get('checked') or 0)))
        lines.append(('Исправлено после подтверждения',
                      '{} строк: список отставал, в карточке стоит другой '
                      'провайдер'.format(len(verify.get('fixed') or []))))
    if verify.get('silent'):
        lines.append(('НЕ подтверждено',
                      '{} строк: карточка не ответила, оставлено значение списка'
                      .format(verify['silent'])))
    for fixed in (verify.get('fixed') or [])[:10]:
        lines.append(('Поправлено по карточке',
                      '{}: список «{}», карточка «{}»'.format(
                          fixed['contractor_id'], fixed['list'], fixed['card'])))
    if stats.get('unverified'):
        lines.append(('НЕ СМОГЛИ ПРОВЕРИТЬ',
                      '{} строк: диспетчерские не ответили на запрос карточки. Это '
                      'не «водителя нет» — это «мы не смогли спросить»; повторите '
                      'выгрузку позже'.format(stats['unverified'])))
    if stats.get('no_provider_by_kind'):
        lines.append(('Сотрудников парка', '{} строк — ЭДО к ним не применяется '
                                           '(работают по трудовому договору)'
                                           .format(stats['no_provider_by_kind'])))
    if stats.get('from_card'):
        lines.append(('Добрано из карточек', '{} строк — список кабинета их не отдал'
                                             .format(stats['from_card'])))
    if resolution.get('park_probe_requests'):
        lines.append(('В файле не было ID парка',
                      'Поиск по паркам стоил {} запросов из {}'.format(
                          resolution['park_probe_requests'], resolution.get('requests') or 0)))
    if resolution.get('skipped_orphans'):
        lines.append(('НЕ ИСКАЛИ по всем паркам',
                      '{} строк: без ID парка каждая такая строка стоит до 86 запросов'
                      .format(resolution['skipped_orphans'])))
    for mismatch in (check.get('mismatched') or [])[:10]:
        lines.append(('Расхождение при сверке',
                      '{}: список «{}», карточка «{}»'.format(
                          mismatch['contractor_id'], mismatch['list'], mismatch['card'])))

    for index, (left, right) in enumerate(lines, start=1):
        sheet.cell(row=index, column=1, value=left)
        if right is not None:
            sheet.cell(row=index, column=2, value=right)
        if left and right is None and not left.startswith(' '):
            sheet.cell(row=index, column=1).font = TITLE_FONT
    sheet.column_dimensions['A'].width = 46
    sheet.column_dimensions['B'].width = 76
    for row in sheet.iter_rows(min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)


def _fill_drivers(sheet, table):
    for index, (_key, title, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row_index, row in enumerate(table, start=2):
        for col_index, (key, _title, _width) in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=row.get(key, ''))
            if key in TEXT_COLUMNS:
                cell.data_type = 's'
                cell.number_format = '@'
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = 'A1:{}{}'.format(
        get_column_letter(len(COLUMNS)), max(1, len(table) + 1))
    sheet.row_dimensions[1].height = 24


def _fill_summary(sheet, counts, total):
    headers = ('Провайдер ЭДО', 'Водителей', 'Доля, %')
    for index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_index, (name, count) in enumerate(counts.items(), start=2):
        sheet.cell(row=row_index, column=1, value=name)
        sheet.cell(row=row_index, column=2, value=count)
        sheet.cell(row=row_index, column=3,
                   value=round(count / total * 100, 2) if total else 0)
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 10
    sheet.freeze_panes = 'A2'


def _fill_parks(sheet, by_park, providers):
    names = [provider['name'] for provider in providers]
    headers = ['Парк', 'ID парка', 'Водителей'] + names + ['не указан', 'не найден']
    for index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row_index, ((park_name, park_id), bucket) in enumerate(by_park.items(), start=2):
        sheet.cell(row=row_index, column=1, value=park_name)
        cell = sheet.cell(row=row_index, column=2, value=park_id)
        cell.data_type = 's'
        cell.number_format = '@'
        sheet.cell(row=row_index, column=3, value=bucket['total'])
        for offset, name in enumerate(names + ['не указан', NOT_FOUND], start=4):
            value = bucket['providers'].get(name, 0)
            sheet.cell(row=row_index, column=offset, value=value or None)
    sheet.column_dimensions['A'].width = 32
    sheet.column_dimensions['B'].width = 34
    sheet.column_dimensions['C'].width = 12
    for index in range(4, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 15
    sheet.freeze_panes = 'C2'
    sheet.row_dimensions[1].height = 30
