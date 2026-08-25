# -*- coding: utf-8 -*-
"""Выгрузка касаний в xlsx.

Пять листов: «Контекст», «Касания», «Операторы», «По дням», «Сводка». Первым
идёт именно контекст — цифра без периода и без оговорок живёт своей жизнью, а
через неделю по файлу «Касания.xlsx» уже не восстановить, за какие даты он и
что в нём считалось разговором.

Соглашения те же, что у остальных выгрузок портала:

* телефоны и внутренние номера лежат ТЕКСТОМ — числом номер теряет ведущие нули
  и уезжает в экспоненту; зелёный уголок «Число сохранено как текст» гасится
  тегом `<ignoredErrors>`, сама функция приходит аргументом, потому что живёт в
  монолите, а импортировать монолит из пакета нельзя;
* длительность — числом секунд И отдельной читаемой колонкой: по секундам
  считают, по «7:12» читают;
* дата — настоящей датой с форматом, а не строкой, иначе не отсортируется.

ПОЧЕМУ КНИГА ПИШЕТСЯ В ПОТОКОВОМ РЕЖИМЕ (`write_only=True`)
------------------------------------------------------------
Месяц — это около ста тысяч касаний, квартал — триста. Обычный режим openpyxl
держит все ячейки объектами в памяти, а Render на это не рассчитан. В потоковом
режиме строки уходят во временный файл по мере добавления, и книга на 300 тысяч
строк стоит 30 МБ памяти вместо гигабайта.

Две ловушки режима, обе проверены и обе молчаливые:

1. **`freeze_panes` и ширины колонок задаются ДО первого `append`.** Шапка листа
   пишется вместе с первой строкой, и всё, что выставлено после, в файл уже не
   попадёт — причём без ошибки: просто книга откроется без закреплённой шапки.

2. **Объекты стиля создаются один раз, модульными константами.** `Font(...)`
   внутри цикла — это 8-кратное замедление: 37 секунд против 4,4 на тех же 30
   тысячах строк (замерено 25.08.2026). Дедупликация стилей у openpyxl есть, но
   платит за неё вызывающий.

Счётчики для «Контекста» и «Сводки» приходят готовыми из SQL, а не считаются по
ходу обхода: касания стримятся генератором, и второго прохода по ним нет.
"""

import math
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ALMATY = ZoneInfo('Asia/Almaty')

# Стили — модульные константы. Создавать их в цикле нельзя, см. шапку.
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
TITLE_FONT = Font(bold=True, size=13)
SECTION_FONT = Font(bold=True, size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
LINK_FONT = Font(color='0563C1', underline='single')

DATE_FORMAT = 'DD.MM.YYYY HH:MM:SS'
DAY_FORMAT = 'DD.MM.YYYY'

# key, заголовок, ширина. Порядок — как читают: сначала когда и кому, потом кто
# и чем кончилось, потом служебное.
COLUMNS = (
    ('started_at', 'Дата и время', 18),
    ('answered_at', 'Ответили в', 18),
    ('phone', 'Телефон клиента', 16),
    ('operator', 'ФИО', 28),
    ('ext', 'Вн. номер', 10),
    ('direction', 'Направление', 18),
    ('call_type', 'Тип', 20),
    ('result', 'Результат', 20),
    ('talk_seconds', 'Разговор, с', 11),
    ('talk_human', 'Разговор', 11),
    ('dial_seconds', 'Вызов всего, с', 14),
    ('queue', 'Очередь', 10),
    ('has_recording', 'Есть запись', 11),
    ('recording_url', 'Ссылка на запись', 20),
    ('linkedid', 'linkedid', 22),
    ('legs', 'Плеч в CDR', 10),
)

# Колонки, которые обязаны остаться текстом (см. шапку модуля). Не диапазон, а
# перечисление: они разбросаны по листу (C, E, L, O), и одним отрезком C:E, как
# было сперва, зелёный уголок гасился только у телефона и внутреннего номера, а
# у «Очереди» и «linkedid» оставался на каждой строке.
TEXT_COLUMNS = ('phone', 'ext', 'queue', 'linkedid')


def _column_index(key):
    for index, (name, _title, _width) in enumerate(COLUMNS, start=1):
        if name == key:
            return index
    return 1


def hms(seconds):
    """Секунды → «7:12» или «1:04:30». Ноль — прочерк: «0:00» читается как
    «разговор был и длился ноль», а его не было вовсе."""
    seconds = int(seconds or 0)
    if seconds <= 0:
        return '—'
    hours, rest = divmod(seconds, 3600)
    minutes, secs = divmod(rest, 60)
    if hours:
        return '%d:%02d:%02d' % (hours, minutes, secs)
    return '%d:%02d' % (minutes, secs)


def _as_datetime(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value)[:19], '%Y-%m-%d %H:%M:%S')
    except ValueError:
        return None


def _as_date(value):
    """'2026-08-24' → date. Настоящей датой, а не строкой: иначе лист «По дням»
    сортируется по алфавиту и по нему нельзя построить график."""
    try:
        return datetime.strptime(str(value)[:10], '%Y-%m-%d').date()
    except ValueError:
        return None


def _ru_date(value):
    parts = str(value)[:10].split('-')
    return '%s.%s.%s' % (parts[2], parts[1], parts[0]) if len(parts) == 3 else str(value)[:10]


def _hours(seconds):
    """Секунды → часы с одним знаком. Двойник `hours` в touchMeta.js: то же
    правило округления, что у процентов (половина вверх, см. ниже)."""
    return math.floor(float(seconds or 0) / 360.0 + 0.5) / 10.0


def _percent(part, whole):
    """Доля в процентах с одним знаком. Двойник `percent` в touchMeta.js.

    Считаем через floor(x + 0.5), а НЕ через round(): у Python round половина
    уходит к чётному (6.25 → 6.2), у JS Math.round — вверх (6.25 → 6.3). На
    точных четвертях процента экран и файл показывали бы разные числа — ровно то
    расхождение, ради отсутствия которого подписи и вынесены в общий модуль.
    """
    if not whole:
        return 0.0
    return math.floor(1000.0 * (part or 0) / whole + 0.5) / 10.0


def _setup(sheet, titles, widths, freeze=True):
    """Шапка, ширины и закрепление. Всё ДО первой строки данных — иначе не
    доедет до файла (ловушка потокового режима, см. шапку модуля)."""
    if freeze:
        sheet.freeze_panes = 'A2'
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    header = []
    for title in titles:
        cell = WriteOnlyCell(sheet, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        header.append(cell)
    sheet.append(header)


def build_workbook(touches, *, period_from, period_to, summary, operators, daily,
                   by_type, by_result, generated_at=None, generated_by='',
                   filters_note='', coverage=None, text_warning_patch=None):
    """touches — итерируемое касаний (можно генератор). Возвращает (BytesIO, строк)."""
    generated_at = generated_at or datetime.now(ALMATY)

    workbook = Workbook(write_only=True)
    context_sheet = workbook.create_sheet('Контекст')
    touches_sheet = workbook.create_sheet('Касания')
    operators_sheet = workbook.create_sheet('Операторы')
    daily_sheet = workbook.create_sheet('По дням')
    summary_sheet = workbook.create_sheet('Сводка')

    _fill_context(context_sheet, period_from, period_to, summary, generated_at,
                  generated_by, filters_note, coverage)

    _setup(touches_sheet, [title for _k, title, _w in COLUMNS],
           [width for _k, _t, width in COLUMNS])
    written = 0
    for touch in touches:
        touches_sheet.append(_touch_row(touches_sheet, touch))
        written += 1
    if written:
        touches_sheet.auto_filter.ref = 'A1:%s%d' % (
            get_column_letter(len(COLUMNS)), written + 1)

    _fill_operators(operators_sheet, operators)
    _fill_daily(daily_sheet, daily)
    _fill_summary(summary_sheet, summary, by_type, by_result, operators)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    if text_warning_patch and written:
        # «Касания» — второй лист книги, поэтому sheet2.xml. sqref допускает
        # несколько диапазонов через пробел — этим и перечисляем разбросанные
        # текстовые колонки.
        sqref = ' '.join(
            '{0}2:{0}{1}'.format(get_column_letter(_column_index(key)), written + 1)
            for key in TEXT_COLUMNS)
        try:
            stream = text_warning_patch(stream, sqref, sheet_path='xl/worksheets/sheet2.xml')
        except Exception:
            # Значок в углу ячейки — досадно, но не повод не отдать файл.
            stream.seek(0)
    return stream, written


def report_filename(period_from, period_to):
    if str(period_from)[:10] == str(period_to)[:10]:
        return 'Касания %s.xlsx' % _ru_date(period_from)
    return 'Касания %s — %s.xlsx' % (_ru_date(period_from), _ru_date(period_to))


# ── листы ────────────────────────────────────────────────────────────────────

def _touch_row(sheet, touch):
    started = WriteOnlyCell(sheet, value=_as_datetime(touch['started_at']))
    started.number_format = DATE_FORMAT
    answered_value = _as_datetime(touch.get('answered_at'))
    answered = WriteOnlyCell(sheet, value=answered_value)
    if answered_value is not None:
        answered.number_format = DATE_FORMAT

    url = touch.get('recording_url') or ''
    if url:
        link = WriteOnlyCell(sheet, value='запись')
        link.hyperlink = url
        link.font = LINK_FONT
    else:
        link = None

    return [
        started, answered,
        str(touch.get('phone') or ''),
        touch.get('operator') or '—',
        str(touch.get('ext') or ''),
        touch.get('direction') or '',
        touch.get('call_type') or '',
        touch.get('result') or '',
        int(touch.get('talk_seconds') or 0),
        hms(touch.get('talk_seconds')),
        int(touch.get('dial_seconds') or 0),
        str(touch.get('queue') or ''),
        'да' if touch.get('has_recording') else 'нет',
        link,
        str(touch.get('linkedid') or ''),
        int(touch.get('legs') or 0),
    ]


def _fill_context(sheet, period_from, period_to, summary, generated_at,
                  generated_by, filters_note, coverage):
    coverage = coverage or {}
    total = summary.get('total', 0) or 0
    sheet.column_dimensions['A'].width = 46
    sheet.column_dimensions['B'].width = 92
    lines = [
        ('Касания отдела продаж', None),
        ('', None),
        ('1. Главное', None),
        ('Период', '%s — %s' % (_ru_date(period_from), _ru_date(period_to))),
        ('Собрано', generated_at.strftime('%d.%m.%Y %H:%M') + ' (Алматы)'),
        ('Собрал', generated_by or '—'),
        ('Источник', 'CDR АТС FreePBX отдела продаж'),
        ('Фильтры', filters_note or 'без фильтров — весь период целиком'),
        ('Касаний', total),
        ('Из них закончились разговором', '%d (%.1f%%)' % (
            summary.get('talks', 0), _percent(summary.get('talks'), total))),
        ('Исходящих / входящих / входящих без ответа', '%d / %d / %d' % (
            summary.get('outgoing', 0), summary.get('incoming', 0),
            summary.get('incoming_missed', 0))),
        ('Внутренних номеров', summary.get('operators', 0)),
        ('Уникальных номеров клиентов', summary.get('phones', 0)),
        ('Общее время разговоров', hms(summary.get('talk_seconds', 0))),
        ('', None),
        ('2. Полнота данных', None),
        ('Суток в периоде', coverage.get('days_total', 0)),
        ('Суток выкачано со станции', coverage.get('days_done', 0)),
        ('Суток не хватает', coverage.get('days_missing', 0)),
        ('Строк CDR прочитано', coverage.get('rows_fetched', 0)),
        ('', None),
        ('3. Что важно знать про эти цифры', None),
        ('Касание — это ОДИН ВЫЗОВ, а не строка CDR.',
         'У входящего в очередь строк бывает до двух десятков — каждая попытка '
         'дозвониться до агента. Они склеены по linkedid, число склеенных плеч '
         'есть в последней колонке.'),
        ('«Разговор, с» — честное время разговора.',
         'Взято с плеча самого агента. У плеча очереди billsec включает ожидание '
         'в очереди и завышает длительность на минуты.'),
        ('«Дата и время» — начало вызова, а не момент ответа.',
         'В колонке «Ответили в» — начало того плеча, на котором состоялся '
         'разговор: для входящего через очередь это момент, когда вызов дошёл до '
         'оператора (медиана 16 секунд после начала, бывает и 11 минут). Секунда '
         'в секунду снятия трубки станция не сообщает.'),
        ('«Вызов всего, с» — вся длительность вызова, включая разговор.',
         'Это поле duration из CDR. Чистое время дозвона — это «Вызов всего» '
         'минус «Разговор»; отдельной колонкой не даём, чтобы цифры совпадали с '
         'выгрузками, которые собирались до этого раздела.'),
        ('«Сброс без разговора» — соединение без разговора.',
         'Чаще всего повторный набор автодозвонщика: disposition ANSWERED при '
         'billsec = 0. Разговором такое не считается.'),
        ('Оператор определён по имени файла записи.',
         'В полях src/dst внутреннего номера у автодозвона нет вовсе. Если номер '
         'ни разу не назвал себя, в колонке ФИО стоит прочерк.'),
        ('ФИО подставлено на дату звонка.',
         'Внутренний номер уволившегося отдают новому сотруднику, поэтому у таких '
         'номеров два владельца, и звонок подписан тем, кто владел номером тогда.'),
        ('Ссылки на записи открываются только из внутренней сети.',
         'Записи лежат на сервере станции, наружу он не выведен.'),
        ('Все времена — Алматы (UTC+5).', 'Так их отдаёт станция.'),
    ]
    for index, (label, value) in enumerate(lines, start=1):
        left = WriteOnlyCell(sheet, value=label)
        if value is None and label:
            left.font = TITLE_FONT if index == 1 else SECTION_FONT
            sheet.append([left])
            continue
        right = WriteOnlyCell(sheet, value=value)
        right.alignment = WRAP_TOP
        sheet.append([left, right])


def _fill_operators(sheet, operators):
    _setup(sheet, ('ФИО', 'Вн. номера', 'Направление', 'Касаний', 'Разговоров',
                   'Разговор, ч', 'Клиентов', 'Дозваниваемость, %'),
           (30, 16, 22, 11, 12, 12, 11, 18))
    for row in operators:
        touches = row.get('touches') or 0
        talks = row.get('talks') or 0
        sheet.append([
            row.get('operator') or '—',
            row.get('exts') or '',
            row.get('direction') or '',
            touches, talks,
            _hours(row.get('talk_seconds')),
            row.get('phones') or 0,
            _percent(talks, touches),
        ])


def _fill_daily(sheet, daily):
    _setup(sheet, ('День', 'Касаний', 'Разговоров', 'Разговор, ч', 'Дозваниваемость, %'),
           (14, 11, 12, 12, 18))
    for row in daily:
        touches = row.get('touches') or 0
        talks = row.get('talks') or 0
        day = WriteOnlyCell(sheet, value=_as_date(row.get('day')))
        day.number_format = DAY_FORMAT
        sheet.append([
            day, touches, talks,
            _hours(row.get('talk_seconds')),
            _percent(talks, touches),
        ])


def _fill_summary(sheet, summary, by_type, by_result, operators):
    sheet.column_dimensions['A'].width = 42
    sheet.column_dimensions['B'].width = 16
    total = summary.get('total', 0) or 0

    def block(title, rows):
        head = WriteOnlyCell(sheet, value=title)
        head.font = SECTION_FONT
        sheet.append([head])
        for label, value in rows:
            sheet.append([label, value])
        sheet.append([])

    block('ГЛАВНОЕ', [
        ('Касаний', total),
        ('Закончились разговором', summary.get('talks', 0)),
        ('Дозваниваемость, %', _percent(summary.get('talks'), total)),
        ('Общее время разговоров, ч', _hours(summary.get('talk_seconds', 0))),
        ('Внутренних номеров', summary.get('operators', 0)),
        ('Уникальных клиентов', summary.get('phones', 0)),
        ('С записью разговора', summary.get('with_recording', 0)),
    ])
    block('ПО ТИПУ', list(by_type or []))
    block('ПО РЕЗУЛЬТАТУ', list(by_result or []))
    block('ОПЕРАТОРЫ, ТОП-20',
          [(row.get('operator') or '—', row.get('touches') or 0) for row in operators[:20]])
