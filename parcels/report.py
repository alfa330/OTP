# -*- coding: utf-8 -*-
"""Выгрузка реестра посылок в xlsx (задача #257, поручил Ядигаров Руслан).

Три листа: «Контекст», «Посылки», «По офисам». Первым идёт именно контекст —
реестр живой, статусы в нём меняются каждый день, и через неделю по файлу
«Посылки.xlsx» уже не восстановить, на какую дату он собран и что в нём
отобрано. Тот же порядок, что у выгрузок «Касания» и «Провайдер ЭДО».

Соглашения те же, что у остальных выгрузок портала:

* телефон, позывной, ID во Флите и номер заказа лежат ТЕКСТОМ — числом номер
  теряет ведущие нули и уезжает в экспоненту, а 32-значный ID Excel и вовсе
  округлит; зелёный уголок «Число сохранено как текст» гасится тегом
  `<ignoredErrors>`, сама функция приходит аргументом, потому что живёт в
  монолите, а импортировать монолит из пакета нельзя;
* дата приёма — настоящей датой с форматом, а не строкой, иначе она не
  отсортируется и по ней не построить сводную;
* книга пишется в потоковом режиме (`write_only=True`) — ловушки режима те же,
  что описаны в `cdr/report.py`: `freeze_panes` и ширины задаются ДО первого
  `append`, объекты стиля создаются один раз модульными константами.

ЧИСЛА В ФАЙЛЕ ОБЯЗАНЫ СОВПАДАТЬ С ЭКРАНОМ
------------------------------------------
«Дней в офисе» и «Залежалась» считаются здесь по тем же правилам, что
`daysInOffice` и `isStale` в `src/components/parcels/parcelMeta.js`:

* срок считается только у лежащих в офисе — у переданных в разделе стоит
  прочерк, и в файле стоит он же (сколько такая посылка пролежала, видно по
  колонкам «Принята» и «Статус изменён»);
* «сегодня» берётся по Алматы, а не по часовому поясу сервера: дата приёма —
  это рабочий день офиса в Казахстане;
* порог «залежалась» — 30 дней, тот же `STALE_AFTER_DAYS`, которым раздел
  подсвечивает строку янтарным.

Разойдись эти правила — человек увидел бы на экране «залежались 12», а в файле
одиннадцать, и объяснить это было бы нечем.

Колонок в файле ровно столько, сколько показывает раздел. Номер водительского
удостоверения в реестре хранится, но ни в списке, ни в карточке не показан —
в файл он поэтому не идёт: выгрузка не должна выносить наружу больше, чем
человек видит на экране.
"""

from datetime import date, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ALMATY = ZoneInfo('Asia/Almaty')

# Стили — модульные константы. Создавать их в цикле нельзя: у openpyxl это
# восьмикратное замедление на тех же строках (замер 25.08.2026, cdr/report.py).
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
TITLE_FONT = Font(bold=True, size=13)
SECTION_FONT = Font(bold=True, size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
LINK_FONT = Font(color='0563C1', underline='single')

DAY_FORMAT = 'DD.MM.YYYY'
DATETIME_FORMAT = 'DD.MM.YYYY HH:MM'

# Подписи статусов и типов — ДОСЛОВНО как в parcelMeta.js. Второй словарь тех же
# кодов неизбежен (питон не читает js), поэтому его сторожит тест: разойдись
# подписи, человек получил бы в файле не то слово, что видит на экране.
STATUS_LABELS = {
    'in_office': 'В офисе',
    'given_to_recipient': 'Передали получателю',
    'given_to_sender': 'Вернули отправителю',
}

KIND_LABELS = {
    'parcel': 'Посылка',
    'document': 'Документ',
    'other': 'Другое',
}

# Тот же порог, что подсвечивает строку в разделе (STALE_AFTER_DAYS).
STALE_AFTER_DAYS = 30

# key, заголовок, ширина. Порядок — как читают: сначала когда и где, потом что
# за посылка, потом чья она, и только затем статус и служебное.
COLUMNS = (
    ('id', '№', 7),
    ('received_on', 'Принята', 12),
    ('days_in_office', 'Дней в офисе', 13),
    ('stale', 'Залежалась', 12),
    ('city', 'Город', 16),
    ('office_name', 'Офис', 24),
    ('office_address', 'Адрес офиса', 34),
    ('kind', 'Тип', 12),
    ('description', 'Описание', 40),
    ('sender', 'Отправитель', 22),
    ('recipient', 'Получатель', 22),
    ('order_url', 'Заказ', 38),
    ('order_number', 'Номер заказа', 14),
    ('comment', 'Комментарий', 32),
    ('driver_name', 'Водитель', 26),
    ('driver_phone', 'Телефон', 16),
    ('driver_park', 'Таксопарк', 22),
    ('driver_car', 'Машина', 22),
    ('driver_callsign', 'Позывной', 14),
    ('driver_account_id', 'ID во Флите', 34),
    ('status', 'Статус', 20),
    ('status_changed_at', 'Статус изменён', 17),
    ('status_changed_by_name', 'Кто изменил статус', 24),
    ('created_by_name', 'Кто завёл', 24),
    ('created_at', 'Заведена', 17),
)

# Колонки, которые обязаны остаться текстом (см. шапку модуля). Перечисление, а
# не отрезок: они разбросаны по листу, и одним диапазоном уголок гасился бы не
# у всех.
TEXT_COLUMNS = ('order_number', 'driver_phone', 'driver_callsign', 'driver_account_id')

# Потолок строк в одном файле. Реестр сегодня — сотни записей, но выгрузка «без
# фильтров» обязана иметь предел: 25 колонок на неограниченной выборке кладут
# инстанс по памяти. Если потолок сработал, об этом написано на «Контексте» —
# молча обрезанный файл читался бы как полный.
EXPORT_LIMIT = 20000

SHEET_PARCELS = 'Посылки'
SHEET_CONTEXT = 'Контекст'
SHEET_OFFICES = 'По офисам'


# ── общие мелочи ─────────────────────────────────────────────────────────────

def today_almaty():
    return datetime.now(ALMATY).date()


def _column_index(key):
    for index, (name, _title, _width) in enumerate(COLUMNS, start=1):
        if name == key:
            return index
    return 1


def _as_date(value):
    """'2026-08-24' → date. Настоящей датой, а не строкой: иначе колонка
    сортируется по алфавиту и по ней не собрать сводную."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value)[:10], '%Y-%m-%d').date()
    except (TypeError, ValueError):
        return None


def _as_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    try:
        return datetime.strptime(str(value)[:19], '%Y-%m-%dT%H:%M:%S')
    except (TypeError, ValueError):
        pass
    try:
        return datetime.strptime(str(value)[:19], '%Y-%m-%d %H:%M:%S')
    except (TypeError, ValueError):
        return None


def _ru_date(value):
    day = _as_date(value)
    return day.strftime('%d.%m.%Y') if day else '—'


def is_closed(status):
    return status in ('given_to_recipient', 'given_to_sender')


def days_in_office(parcel, today=None):
    """Сколько дней посылка лежит. Двойник `daysInOffice` из parcelMeta.js.

    У переданной посылки срок не считается — в разделе на её месте прочерк, и в
    файле должен стоять он же.
    """
    if not parcel or is_closed(parcel.get('status')):
        return None
    received = _as_date(parcel.get('received_on'))
    if received is None:
        return None
    return max(0, ((today or today_almaty()) - received).days)


def is_stale(parcel, today=None):
    days = days_in_office(parcel, today)
    return days is not None and days >= STALE_AFTER_DAYS


def status_label(code):
    return STATUS_LABELS.get(code, code or '—')


def kind_label(code):
    return KIND_LABELS.get(code, code or '—')


def report_filename(generated_at=None):
    generated_at = generated_at or datetime.now(ALMATY)
    return 'Посылки %s.xlsx' % generated_at.strftime('%d.%m.%Y')


# ── книга ────────────────────────────────────────────────────────────────────

def _setup(sheet, titles, widths, freeze='A2'):
    """Шапка, ширины и закрепление. Всё ДО первой строки данных — в потоковом
    режиме выставленное после в файл уже не попадёт, причём молча."""
    if freeze:
        sheet.freeze_panes = freeze
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    header = []
    for title in titles:
        cell = WriteOnlyCell(sheet, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        header.append(cell)
    sheet.append(header)


def build_workbook(parcels, *, generated_at=None, generated_by='', filters_note='',
                   total=None, truncated=False, today=None, text_warning_patch=None):
    """parcels — список карточек в том же виде, в каком их отдаёт список раздела.

    Возвращает (BytesIO, число строк). Список, а не генератор: по тем же строкам
    считается лист «По офисам», и второго прохода по базе ради сводки не будет.
    """
    generated_at = generated_at or datetime.now(ALMATY)
    today = today or today_almaty()
    rows = list(parcels or [])

    workbook = Workbook(write_only=True)
    context_sheet = workbook.create_sheet(SHEET_CONTEXT)
    parcels_sheet = workbook.create_sheet(SHEET_PARCELS)
    offices_sheet = workbook.create_sheet(SHEET_OFFICES)

    counters = _counters(rows, today)
    _fill_context(context_sheet, counters, generated_at, generated_by, filters_note,
                  total if total is not None else len(rows), truncated)

    _setup(parcels_sheet, [title for _k, title, _w in COLUMNS],
           [width for _k, _t, width in COLUMNS])
    for parcel in rows:
        parcels_sheet.append(_parcel_row(parcels_sheet, parcel, today))
    if rows:
        parcels_sheet.auto_filter.ref = 'A1:%s%d' % (
            get_column_letter(len(COLUMNS)), len(rows) + 1)

    _fill_offices(offices_sheet, rows, today)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    if text_warning_patch and rows:
        # «Посылки» — второй лист книги, поэтому sheet2.xml. sqref допускает
        # несколько диапазонов через пробел — этим и перечисляем разбросанные
        # текстовые колонки.
        sqref = ' '.join(
            '{0}2:{0}{1}'.format(get_column_letter(_column_index(key)), len(rows) + 1)
            for key in TEXT_COLUMNS)
        try:
            stream = text_warning_patch(stream, sqref, sheet_path='xl/worksheets/sheet2.xml')
        except Exception:
            # Значок в углу ячейки — досадно, но не повод не отдать файл.
            stream.seek(0)
    return stream, len(rows)


def _counters(rows, today):
    counters = {code: 0 for code in STATUS_LABELS}
    counters['stale'] = 0
    for parcel in rows:
        status = parcel.get('status')
        if status in counters:
            counters[status] += 1
        if is_stale(parcel, today):
            counters['stale'] += 1
    counters['all'] = len(rows)
    return counters


def clean(value):
    """Текст, который xlsx вообще способен унести.

    Управляющие символы (\x00–\x1f, кроме табуляции и переводов строки) в XML
    запрещены, и openpyxl роняет на них ВСЮ книгу — `IllegalCharacterError` в
    момент `save()`. Одна карточка с таким символом означала бы, что выгрузка
    перестала работать у всех и навсегда, а починить её мог бы только
    разработчик: описание и комментарий сотрудники вставляют из чатов, а оттуда
    такой символ приезжает невидимым.

    Поэтому символ вычищается, а не роняет файл: потеря невидимого знака в
    описании несопоставима с потерей всей выгрузки. Заменяем пробелом, а не
    пустотой — иначе «коробка\x07со звонком» склеилось бы в одно слово.
    """
    if value is None:
        return ''
    return ILLEGAL_CHARACTERS_RE.sub(' ', str(value))


def _text(sheet, value, *, keep_format=False):
    """Строковая ячейка, которая ГАРАНТИРОВАННО остаётся строкой.

    Единственная дверь для КАЖДОЙ строковой ячейки книги — включая листы
    «Контекст» и «По офисам». Причин три, и по одной не хватает ни одной:

    * openpyxl выводит тип из значения, и строка с ведущим «=» становится
      ФОРМУЛОЙ. Описание и комментарий люди пишут руками — «=коробка» там вполне
      возможно, и такая книга открывалась бы с ошибкой вычисления, а то и с
      предупреждением безопасности. Поэтому тип проставляется явно, а не
      подразумевается.
    * управляющий символ роняет сборку целиком — см. `clean`.
    * `number_format='@'` нужен телефону, позывному, ID и номеру заказа: без
      него Excel всё равно предложит «преобразовать в число», и ведущий ноль
      пропадёт при первой же правке файла руками.

    Формат вешается ТОЛЬКО на непустую ячейку: у пустой он превращает «нет
    данных» в ноль при протяжке столбца (та же ловушка, что в wiki/report_kit).
    """
    text = clean(value)
    cell = WriteOnlyCell(sheet, value=text)
    cell.data_type = 's'
    if keep_format and text:
        cell.number_format = '@'
    return cell


def _order_cell(sheet, parcel):
    """Ссылка на заказ. Значением кладём сам адрес, а не слово «заказ»: по id
    заказа в файле ищут поиском, и подменённый текст этот поиск бы сломал.
    Кликабельной делаем только http(s) — ссылка, ведущая непонятно куда, хуже
    её отсутствия (то же правило, что в карточке)."""
    url = str(parcel.get('order_url') or '').strip()
    cell = _text(sheet, url)
    if url.lower().startswith(('http://', 'https://')):
        cell.hyperlink = url
        cell.font = LINK_FONT
    return cell


def _parcel_row(sheet, parcel, today):
    received_value = _as_date(parcel.get('received_on'))
    received = WriteOnlyCell(sheet, value=received_value)
    if received_value is not None:
        received.number_format = DAY_FORMAT

    changed_value = _as_datetime(parcel.get('status_changed_at'))
    changed = WriteOnlyCell(sheet, value=changed_value)
    if changed_value is not None:
        changed.number_format = DATETIME_FORMAT

    created_value = _as_datetime(parcel.get('created_at'))
    created = WriteOnlyCell(sheet, value=created_value)
    if created_value is not None:
        created.number_format = DATETIME_FORMAT

    lying = days_in_office(parcel, today)
    return [
        int(parcel.get('id') or 0),
        received,
        lying,
        # Прочерк, а не «нет»: «нет» у переданной посылки читалось бы как «её
        # проверяли и она свежая», а срок у неё не считается вовсе.
        ('да' if is_stale(parcel, today) else 'нет') if lying is not None else None,
        _text(sheet, parcel.get('city')),
        _text(sheet, parcel.get('office_name')),
        _text(sheet, parcel.get('office_address')),
        _text(sheet, kind_label(parcel.get('kind'))),
        _text(sheet, parcel.get('description')),
        _text(sheet, parcel.get('sender')),
        _text(sheet, parcel.get('recipient')),
        _order_cell(sheet, parcel),
        _text(sheet, parcel.get('order_number'), keep_format=True),
        _text(sheet, parcel.get('comment')),
        _text(sheet, parcel.get('driver_name')),
        _text(sheet, parcel.get('driver_phone'), keep_format=True),
        _text(sheet, parcel.get('driver_park')),
        _text(sheet, parcel.get('driver_car')),
        _text(sheet, parcel.get('driver_callsign'), keep_format=True),
        _text(sheet, parcel.get('driver_account_id'), keep_format=True),
        _text(sheet, status_label(parcel.get('status'))),
        changed,
        _text(sheet, parcel.get('status_changed_by_name')),
        _text(sheet, parcel.get('created_by_name')),
        created,
    ]


def _fill_context(sheet, counters, generated_at, generated_by, filters_note,
                  total, truncated):
    sheet.column_dimensions['A'].width = 46
    sheet.column_dimensions['B'].width = 92
    lines = [
        ('Невостребованные посылки', None),
        ('', None),
        ('1. Главное', None),
        ('Собрано', generated_at.strftime('%d.%m.%Y %H:%M') + ' (Алматы)'),
        ('Собрал', generated_by or '—'),
        ('Источник', 'Реестр раздела «Посылки» в iCORE'),
        ('Отобрано', filters_note or 'без фильтров — весь реестр целиком'),
        ('Строк в файле', counters['all']),
    ]
    if truncated:
        # Предупреждение стоит ДО счётчиков, а не сноской после них: счётчики
        # ниже посчитаны по строкам ФАЙЛА, и прочитать их как сводку по всему
        # реестру — первое, что сделает человек. Отдельно назван перекос: файл
        # собирается от свежих к старым, значит потолок срезает САМЫЕ СТАРЫЕ —
        # ровно те, что и попадают в «залежались».
        lines.append(('В ФАЙЛ ПОПАЛО НЕ ВСЁ',
                      'По этому отбору в реестре %d записей, а в файл помещается '
                      'не больше %d. Обрезаны самые старые: файл собран от свежих '
                      'к старым. Поэтому и числа ниже — только по строкам этого '
                      'файла, а «залежались» в реестре БОЛЬШЕ, чем показано. '
                      'Сузьте отбор — городом, офисом или датой приёма — и '
                      'выгрузите частями.' % (total, EXPORT_LIMIT)))
    # Когда потолок сработал, подписи честно говорят, что считают: без этого
    # «В офисе 20 000» читается как ответ про весь реестр.
    scope = ' (в этом файле)' if truncated else ''
    lines += [
        ('В офисе' + scope, counters.get('in_office', 0)),
        ('Из них залежались (%d дней и больше)%s' % (STALE_AFTER_DAYS, scope),
         counters.get('stale', 0)),
        ('Передали получателю' + scope, counters.get('given_to_recipient', 0)),
        ('Вернули отправителю' + scope, counters.get('given_to_sender', 0)),
        ('', None),
        ('2. Что важно знать про эти данные', None),
        ('Файл — снимок на дату сборки.',
         'Реестр живой: посылку могут отдать завтра, и статус в файле останется '
         'вчерашним. Дата и время сборки — выше.'),
        ('«Дней в офисе» считается только у лежащих.',
         'У переданных стоит прочерк: сколько такая посылка пролежала, видно по '
         'колонкам «Принята» и «Статус изменён».'),
        ('«Залежалась» — это %d дней и больше в офисе.' % STALE_AFTER_DAYS,
         'Тот же порог, которым раздел подсвечивает строку: после месяца водитель '
         'почти наверняка сам уже не придёт, и офису пора звонить.'),
        ('Офис и данные водителя — снимок на день приёма.',
         'Адрес офиса могли поправить, а водитель — сменить номер телефона. '
         'Карточка хранит то, что было в день, когда вещь оставили.'),
        ('Телефон, позывной, ID и номер заказа лежат текстом.',
         'Числом номер теряет ведущий ноль и уезжает в экспоненту, а 32-значный '
         'ID во Флите Excel округляет.'),
        ('Все даты — Алматы (UTC+5).',
         'Дата приёма — рабочий день офиса в Казахстане, а не день по часовому '
         'поясу того, кто открыл файл.'),
    ]
    # Ячейки контекста идут через ту же дверь, что таблица: в «Отобрано»
    # приезжает поисковый запрос человека, а в «Собрал» — его имя, и оба
    # способны принести и ведущее «=», и управляющий символ из буфера обмена.
    for index, (label, value) in enumerate(lines, start=1):
        left = _text(sheet, label)
        if value is None and label:
            left.font = TITLE_FONT if index == 1 else SECTION_FONT
            sheet.append([left])
            continue
        # Числа остаются числами: иначе по «Строк в файле» ничего не посчитать.
        right = (WriteOnlyCell(sheet, value=value) if isinstance(value, int)
                 else _text(sheet, value))
        right.alignment = WRAP_TOP
        sheet.append([left, right])


OFFICE_COLUMNS = (
    ('Город', 18),
    ('Офис', 28),
    ('Всего', 10),
    ('В офисе', 10),
    ('Залежались', 12),
    ('Передали получателю', 20),
    ('ОТДАЛИ ОТПРАВИТЕЛЮ', 20),
)


def _fill_offices(sheet, rows, today):
    """Сводка по офисам: где сколько лежит и сколько из этого залежалось.

    Считается по тем же строкам, что попали на лист «Посылки», а не отдельным
    запросом: сводка, не сходящаяся с данными под ней, хуже её отсутствия.
    Сортировка по «В офисе», а не по «Всего» — раздел про невостребованное, и
    первым должен стоять офис, которому пора разбирать полку.
    """
    _setup(sheet, [title for title, _w in OFFICE_COLUMNS],
           [width for _t, width in OFFICE_COLUMNS])
    buckets = {}
    for parcel in rows:
        key = (parcel.get('city') or '—', parcel.get('office_name') or '—')
        bucket = buckets.setdefault(key, {'all': 0, 'in_office': 0, 'stale': 0,
                                          'given_to_recipient': 0, 'given_to_sender': 0})
        bucket['all'] += 1
        status = parcel.get('status')
        if status in bucket:
            bucket[status] += 1
        if is_stale(parcel, today):
            bucket['stale'] += 1
    ordered = sorted(buckets.items(),
                     key=lambda item: (-item[1]['in_office'], -item[1]['all'], item[0]))
    for (city, office), bucket in ordered:
        sheet.append([
            _text(sheet, city), _text(sheet, office),
            bucket['all'], bucket['in_office'], bucket['stale'],
            bucket['given_to_recipient'], bucket['given_to_sender'],
        ])
    if ordered:
        sheet.auto_filter.ref = 'A1:%s%d' % (
            get_column_letter(len(OFFICE_COLUMNS)), len(ordered) + 1)
