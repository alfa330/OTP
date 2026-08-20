"""Обзвон фронт-офиса: статистика CRM + проверка дневного плана.

Задача #159 (Турарбек Даурен): каждое утро сообщать в Telegram, кто из
менеджеров фронт-офиса не выполнил дневной план по обзвону.

Источник — та же CRM yataxi, что и у конкурса регистраций, поэтому токен
берём общий (см. reg_contest.get_config):
  - метод: POST {CRM_REGION_CALL_STATS_URL} (= /api/partners/region-call-stats)
  - авторизация: заголовок X-Integration-Token
  - тело: {"from": "YYYY-MM-DD", "to": "YYYY-MM-DD"} — обе границы включительно
  - ответ: {"total", "dialog_statuses": [...], "managers": [...]},
    строка менеджера = {manager_id, manager_name, manager_login, total_calls,
    by_status: [{status_id, status_code, status_title, count}]}

Проверено на живом API 17.08.2026:
  * границы включительно, а суммы по дням складываются в сумму за период
    (41 + 65 + 111 = 217 за 10–12.08);
  * будущая дата — законный пустой ответ {"total": 0, "managers": []}, а не
    ошибка; отсюда правило «нет звонков = молчим», а не «все провалили план»;
  * кривые параметры (перевёрнутый период, дата в формате ДД.ММ.ГГГГ) CRM
    отдаёт HTML-страницей с кодом 200 — тот же капкан, что у конкурса
    регистраций, поэтому разбор строгий: незнакомый формат = ошибка;
  * без "to" период открытый (до сегодня), без тела — вообще за всё время;
    оба ключа поэтому передаём всегда;
  * неверный токен — честный 401.

ГЛАВНОЕ ПРО ПЛАН. CRM присылает только тех, кто звонил, — менеджер с нулём
звонков в ответе просто отсутствует. А это ровно тот, кто план и не выполнил.
Поэтому реестр берём у себя (менеджеры отдела «Фронт офисы» в iCORE), а числа
CRM подставляем к нему; несопоставленные звонки CRM не выбрасываем молча, а
показываем отдельной строкой.

Смен фронт-офиса в iCORE нет, так что «был ли человек сегодня на работе»
проверить нечем: выходной выглядит как невыполненный план. Ограничение
осознанное — иначе пришлось бы заводить графики на весь отдел.

ENV (окружение или .env.codex.local):
    CRM_REGION_CALL_STATS_URL  — по умолчанию адрес ниже, менять не нужно
    X-Integration-Token        — общий токен CRM (см. reg_contest)
"""

import html
import logging
import os
import time
from datetime import date, datetime, timedelta
from io import BytesIO

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import reg_contest

log = logging.getLogger(__name__)

DEFAULT_API_URL = "https://backend.yataxi.kz/api/partners/region-call-stats"

HTTP_TIMEOUT = 60
MAX_RETRIES = 3
RETRY_PAUSE = 2.0

# Поденные запросы ходят с коротким терпением и без повторов. Причина
# арифметическая: три попытки по минуте с паузами — это до 186 секунд на ОДИН
# день, то есть до полутора часов на месяц, и всё это время занят поток общего
# пула бота (в нём 4 места на весь бот, включая db.get_user в начале каждого
# хендлера). День, который не ответил, штатно деградирует в «нет выгрузки»,
# так что бороться за него ценой доступности бота нельзя.
DAY_HTTP_TIMEOUT = 15
DAY_MAX_RETRIES = 1

# Сколько менеджеров показываем в полной сводке. Отдел — 22 человека, запас
# на вырост; список длиннее просто не читается в Telegram.
MAX_ROWS = 60


def get_config(env_file=".env.codex.local"):
    """Адрес эндпоинта + токен CRM.

    Токен общий с конкурсом регистраций (одна и та же CRM), поэтому его
    разрешение целиком отдано reg_contest: там же лежит фолбэк на «человеческое»
    имя ключа X-Integration-Token. Адрес свой и не секретный — держим значением
    по умолчанию, чтобы не заводить лишнюю переменную на Render.
    """
    token = (reg_contest.get_config(env_file) or {}).get("token")
    url = os.getenv("CRM_REGION_CALL_STATS_URL") or DEFAULT_API_URL
    return {"url": url, "token": token}


def is_configured(config=None):
    cfg = config or get_config()
    return bool(cfg.get("url") and cfg.get("token"))


class RegionCallStatsClient:
    """Минимальный клиент POST {CRM_REGION_CALL_STATS_URL}."""

    def __init__(self, url, token, timeout=HTTP_TIMEOUT, retries=MAX_RETRIES):
        if not url or not token:
            raise ValueError("CRM_REGION_CALL_STATS_URL / токен CRM не заданы")
        self.url = url
        self.token = token
        self.timeout = timeout
        self.retries = max(int(retries or 1), 1)
        self.session = requests.Session()

    @classmethod
    def from_config(cls, config=None, timeout=HTTP_TIMEOUT, retries=MAX_RETRIES):
        cfg = config or get_config()
        return cls(cfg.get("url"), cfg.get("token"), timeout=timeout,
                   retries=retries)

    def _post(self, payload):
        headers = {
            "X-Integration-Token": self.token,
            "Content-Type": "application/json",
        }
        last_error = None
        for attempt in range(1, self.retries + 1):
            try:
                resp = self.session.post(self.url, json=payload,
                                         headers=headers, timeout=self.timeout)
                if resp.status_code >= 500:
                    raise requests.RequestException(f"CRM HTTP {resp.status_code}")
                if resp.status_code != 200:
                    raise RuntimeError(
                        f"CRM HTTP {resp.status_code}: {resp.text[:300]}")
                # На кривой период CRM отвечает HTML-страницей с кодом 200 —
                # для нас это ошибка разбора, а не пустой ответ.
                return resp.json()
            except (requests.RequestException, ValueError) as exc:
                last_error = exc
                log.warning("front_office_calls: попытка %s/%s не удалась: %s",
                            attempt, self.retries, exc)
                if attempt < self.retries:
                    time.sleep(RETRY_PAUSE * attempt)
        raise RuntimeError(
            f"CRM недоступна после {self.retries} попыток: {last_error}")

    def fetch_managers(self, date_from, date_to):
        """Звонки менеджеров за период [date_from, date_to] включительно.

        Формат проверяем строго: пустой список менеджеров — законный ответ
        («в этот день не звонили»), а вот ответ без ключа managers означает,
        что CRM сменила контракт, и молча считать это нулём нельзя.
        """
        data = self._post({
            "from": _iso(date_from),
            "to": _iso(date_to),
        })
        if not isinstance(data, dict):
            raise RuntimeError(
                f"CRM отдала не объект, а {type(data).__name__}")
        managers = data.get("managers")
        if not isinstance(managers, list):
            raise RuntimeError(
                "CRM отдала незнакомый формат: нет списка managers, "
                f"ключи ответа = {sorted(data)[:10]}")
        for row in managers:
            if not isinstance(row, dict) or "total_calls" not in row:
                raise RuntimeError(
                    "CRM отдала менеджера без total_calls, "
                    f"поля = {sorted(row)[:10] if isinstance(row, dict) else row}")
        return managers


def total_calls(managers):
    """Сумма звонков в сыром ответе CRM.

    Нужна для контрольной сверки: сумма по дням обязана совпасть с запросом
    за весь период, и расхождение надо показать, а не замолчать.
    """
    return sum(_int_or_zero(row.get("total_calls")) for row in managers or [])


def _iso(value):
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def _int_or_zero(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


# ---------------------------------------------------------------------------
# Сборка отчёта
# ---------------------------------------------------------------------------

# Город приходит из кадровой карточки (users.city): CRM географию не отдаёт
# вовсе, ключа region/city в её ответе нет ни на одном уровне. У новичка поле
# может быть пустым — такие идут одной группой и всегда последними, чтобы
# пропуск в карточке было видно, а не растворился среди городов.
CITY_UNKNOWN = "Город не указан"

# Год в дате приёма кое-где записан с опечаткой (в отделе есть карточка с 0024
# вместо 2024). Такую дату нельзя пускать в знаменатель плана: получится
# процент от двух тысяч дней, и человек молча выпадет из «не выполнивших».
MIN_SANE_HIRE_YEAR = 2000


def _city_of(user):
    return str(user.get("city") or "").strip() or CITY_UNKNOWN


def _hire_date_of(user):
    value = user.get("hire_date")
    if isinstance(value, datetime):
        value = value.date()
    if not isinstance(value, date):
        return None
    return value if value.year >= MIN_SANE_HIRE_YEAR else None


def _directory(roster):
    """Справочник для reg_contest.match_operator — только поля матчинга."""
    return [{"id": u.get("id"), "name": u.get("name"), "email": u.get("email")}
            for u in (roster or [])]


def _staff_from(hire_date, date_from, first_call_day=None):
    """С какого дня периода с человека можно требовать норму.

    Обычно это дата приёма: у принятого 3-го августа нет плана на 1–2 августа.
    Но если человек ЗВОНИЛ раньше своей даты приёма, верить надо звонкам, а не
    карточке. Иначе окно плана оказывается уже, чем набор дней, за которые мы
    показываем числа, и получается сразу три вранья: клетки со звонками
    прячутся под «вне штата», процент нормы считается от меньшего знаменателя
    (легко выходит 400%), а недобравший норму уезжает из списка «не выполнили».
    Дата приёма может быть неверной обыденно, без опечатки: повторный приём
    перезаписывает её на новую, а логин CRM могли передать другому человеку.
    """
    start = date_from
    if hire_date and hire_date > date_from:
        start = hire_date
        if first_call_day is not None and first_call_day < start:
            start = max(first_call_day, date_from)
    return start


def _staff_days(staff_from, date_to, no_data_days=()):
    """Сколько дней окна реально идёт в знаменатель плана.

    Дни, за которые CRM не ответила, вычитаем: там ноль означает дыру в
    выгрузке, а не молчащий телефон.
    """
    if staff_from > date_to:
        return 0
    blind = sum(1 for day in no_data_days if staff_from <= day <= date_to)
    return max((date_to - staff_from).days + 1 - blind, 0)


# Надёжность методов матчинга: почта точнее ФИО, точное ФИО точнее префикса.
_MATCH_RANK = {"email": 0, "name": 1, "name_prefix": 2}


def _match_calls(directory, crm_managers):
    """Разложить звонки CRM на реестр.

    Возвращает ({user_id: звонки}, [строки вне реестра], {user_id: метод}).
    Один человек может иметь в CRM несколько учёток (у отдела уже есть пары id
    из разных наборов), поэтому звонки по сопоставленному пользователю
    складываем, а не берём первую строку. Метод при этом запоминаем ОДИН на
    человека, самый надёжный из сработавших: иначе две учётки дают два плюса
    в разные счётчики и «Сопоставлено 23 из 22».
    """
    calls_by_user = {}
    unmatched = {}
    methods = {}
    for index, row in enumerate(crm_managers or []):
        calls = _int_or_zero(row.get("total_calls"))
        user, method = reg_contest.match_operator(
            row.get("manager_login"), row.get("manager_name"), directory)
        if user:
            calls_by_user[user["id"]] = calls_by_user.get(user["id"], 0) + calls
            previous = methods.get(user["id"])
            if previous is None or _MATCH_RANK.get(method, 9) < _MATCH_RANK.get(previous, 9):
                methods[user["id"]] = method
        elif calls:
            # Нулевые чужие строки прятать можно — они ничего не меняют,
            # а вот звонки мимо реестра обязаны быть видны.
            # У строки без manager_id ключа не существует: сливать такие в одну
            # «учётку None» нельзя — звонки нескольких разных подписались бы
            # именем первой, а счётчик учёток занизился бы.
            key = row.get("manager_id")
            if key is None:
                key = ("crm-row", index)
            entry = unmatched.setdefault(key, {
                "crm_manager_id": key,
                "name": row.get("manager_name"),
                "login": row.get("manager_login"),
                "calls": 0,
                "days": 0,
            })
            entry["calls"] += calls
    return calls_by_user, list(unmatched.values()), methods


def build_report(roster, crm_managers, date_from, date_to=None, plan_per_day=None,
                 no_data_days=(), first_call_days=None):
    """Свести реестр менеджеров iCORE с числами CRM.

    roster — список {id, name, email, city, hire_date} из отдела «Фронт офисы»;
    порядок в отчёте задаём мы. crm_managers — сырые строки CRM.

    Считаем по реестру, а не по ответу CRM: тот, кто не сделал ни одного
    звонка, в ответе отсутствует, а в отчёте обязан быть — иначе отбивка
    покажет «все молодцы» ровно в тот день, когда никто не работал.

    План у каждого свой: общая норма умножается не на длину периода, а на его
    личные дни в штате (см. _staff_from). За один день это то же самое число,
    а на месяце — разница между «новичок провалил план» и «новичка ещё не было».

    first_call_days — {user_id: первый день со звонками}. Нужен, чтобы окно
    плана не оказалось уже набора дней, за которые мы показываем числа; его
    передаёт build_daily_report, у которого разбивка по дням на руках. Без
    него (одним куском за период) день звонка неизвестен, и тогда любые
    звонки раздвигают окно на весь период — то есть считаем как раньше, но
    ничего не прячем.
    """
    date_to = date_to or date_from
    days = (date_to - date_from).days + 1
    plan_per_day = plan_per_day if plan_per_day and plan_per_day > 0 else None
    plan_total = plan_per_day * days if plan_per_day else None
    no_data_days = sorted(set(no_data_days or ()))

    calls_by_user, unmatched, matched_by = _match_calls(
        _directory(roster), crm_managers)

    rows = []
    for user in roster or []:
        user_id = user.get("id")
        calls = calls_by_user.get(user_id, 0)
        hire_date = _hire_date_of(user)
        if first_call_days is not None:
            first_call_day = first_call_days.get(user_id)
        else:
            first_call_day = date_from if calls else None
        staff_from = _staff_from(hire_date, date_from, first_call_day)
        staff_days = _staff_days(staff_from, date_to, no_data_days)
        own_plan = plan_per_day * staff_days if plan_per_day else None
        rows.append({
            "user_id": user_id,
            "name": user.get("name"),
            "city": _city_of(user),
            "hire_date": hire_date,
            "staff_from": staff_from,
            "calls": calls,
            "staff_days": staff_days,
            "plan_total": own_plan,
            # Ноль дней в штате — не провал, а «его тут не было»: такой человек
            # не попадает ни в «выполнил», ни в «не выполнил».
            "met": (calls >= own_plan) if own_plan else None,
            "per_day": {},
        })
    # Больше звонков — выше; при равенстве по алфавиту, чтобы список не
    # перетасовывался между днями на одинаковых числах.
    rows.sort(key=lambda r: (-r["calls"], reg_contest.fold_name(r["name"])))
    unmatched.sort(key=lambda r: -r["calls"])

    missing = [r for r in rows if r["met"] is False] if plan_total else []
    return _summarize({
        "date_from": date_from,
        "date_to": date_to,
        "days": days,
        # Сетка дней всегда календарная: пропуск внутри периода — это дыра
        # в данных, а не более короткий период.
        "day_list": [date_from + timedelta(days=i) for i in range(days)],
        "no_data_days": no_data_days,
        "plan_per_day": plan_per_day,
        "plan_total": plan_total,
        "rows": rows,
        "missing": missing,
        "roster_size": len(rows),
        "called_count": sum(1 for r in rows if r["calls"] > 0),
        "total_calls": sum(r["calls"] for r in rows),
        "unmatched": unmatched,
        "unmatched_calls": sum(r["calls"] for r in unmatched),
        # Людей по методу, а не строк CRM: у одного человека может быть две
        # учётки, и складывать их в разные счётчики нельзя.
        "matched_by": _count_methods(matched_by),
        "matched_total": len(matched_by),
        "crm_period_total": None,
    })


def _count_methods(methods_by_user):
    counts = {}
    for method in methods_by_user.values():
        counts[method] = counts.get(method, 0) + 1
    return counts


def _summarize(report):
    """Досчитать средние, проценты и разрез по городам поверх строк реестра.

    Вызывается дважды: сразу в build_report и ещё раз, когда поверх строк лёг
    поденный разрез (build_daily_report). Функция идемпотентна — всё считается
    заново из rows, накоплений между вызовами нет.
    """
    rows = report["rows"]
    plan_per_day = report["plan_per_day"]
    for row in rows:
        row["avg_per_day"] = (row["calls"] / row["staff_days"]
                              if row["staff_days"] else None)
        row["plan_pct"] = (row["calls"] / row["plan_total"]
                           if row["plan_total"] else None)
        row["days_met"] = (sum(1 for value in row["per_day"].values()
                               if value >= plan_per_day) if plan_per_day else 0)

    cities = {}
    for row in rows:
        city = cities.setdefault(row["city"], {
            "city": row["city"], "roster_size": 0, "calls": 0, "staff_days": 0,
            "plan_total": 0, "missing_count": 0, "per_day": {},
        })
        city["roster_size"] += 1
        city["calls"] += row["calls"]
        city["staff_days"] += row["staff_days"]
        city["plan_total"] += row["plan_total"] or 0
        city["missing_count"] += 1 if row["met"] is False else 0
        for day, value in row["per_day"].items():
            city["per_day"][day] = city["per_day"].get(day, 0) + value
    for city in cities.values():
        city["avg_per_day"] = (city["calls"] / city["staff_days"]
                               if city["staff_days"] else None)
        city["plan_total"] = city["plan_total"] or None
        city["plan_pct"] = (city["calls"] / city["plan_total"]
                            if city["plan_total"] else None)
    # Хуже всех — сверху: город, где норму не тянут, и есть сообщение отчёта.
    # Без плана сравнивать процентами нечем, тогда сортируем фактом.
    # «Город не указан» всегда последним: это не результат, а пропуск в карточке.
    report["cities"] = sorted(
        cities.values(),
        key=lambda c: (c["city"] == CITY_UNKNOWN,
                       c["plan_pct"] if c["plan_pct"] is not None else 0,
                       -c["calls"], c["city"]))

    report["per_day_totals"] = {
        day: sum(row["per_day"].get(day, 0) for row in rows)
        for day in report["day_list"] if day not in report["no_data_days"]
    }
    report["staff_days_total"] = sum(r["staff_days"] for r in rows)
    report["plan_sum"] = sum(r["plan_total"] or 0 for r in rows) or None
    report["avg_per_manager_day"] = (
        report["total_calls"] / report["staff_days_total"]
        if report["staff_days_total"] else None)
    report["plan_pct"] = (report["total_calls"] / report["plan_sum"]
                          if report["plan_sum"] else None)
    report["cells_met"] = (sum(r["days_met"] for r in rows)
                           if plan_per_day else None)
    report["managers_met_once"] = sum(1 for r in rows if r["days_met"] > 0)
    report["zero_managers"] = sum(1 for r in rows if r["calls"] == 0)
    report["no_city_count"] = sum(1 for r in rows if r["city"] == CITY_UNKNOWN)
    return report


def build_daily_report(roster, managers_by_day, plan_per_day=None,
                       no_data_days=(), crm_period_total=None):
    """Отчёт с поденной сеткой: город × сотрудник × дни.

    managers_by_day — {день: сырые строки CRM за этот день}. Разбивки по датам
    в ответе CRM нет, и параметра-разреза тоже нет, поэтому каждый день
    спрашивается отдельным запросом. Суммы по дням в точности совпадают с
    запросом за период (сверено на живом API — и по итогу, и по каждому
    менеджеру), поэтому итоги считает тот же build_report по объединению дней,
    а не вторая формула рядом.

    no_data_days — дни, за которые CRM не ответила: в сетке это пустая клетка,
    а не ноль, и в знаменатель плана они не идут.
    """
    no_data_days = sorted(set(no_data_days or ()))
    answered = {day: rows for day, rows in (managers_by_day or {}).items()
                if day not in no_data_days}
    day_list = sorted(set(answered) | set(no_data_days))
    if not day_list:
        raise ValueError("Нужен хотя бы один день")

    # Дни разбираем ПЕРВЫМИ: из них известно, с какого дня человек реально
    # звонил, а без этого окно плана можно посчитать уже, чем набор показанных
    # клеток (см. _staff_from).
    directory = _directory(roster)
    calls_by_day = {}
    first_call_days = {}
    unmatched_days = {}
    for day in sorted(answered):
        calls_by_user, unmatched, _methods = _match_calls(directory, answered[day])
        calls_by_day[day] = calls_by_user
        for user_id, calls in calls_by_user.items():
            if calls and day < first_call_days.get(user_id, day + timedelta(days=1)):
                first_call_days[user_id] = day
        for entry in unmatched:
            key = entry["crm_manager_id"]
            unmatched_days[key] = unmatched_days.get(key, 0) + 1

    all_rows = [row for day in sorted(answered) for row in (answered[day] or [])]
    report = build_report(roster, all_rows, day_list[0], day_list[-1],
                          plan_per_day=plan_per_day, no_data_days=no_data_days,
                          first_call_days=first_call_days)

    rows_by_id = {row["user_id"]: row for row in report["rows"]}
    for day, calls_by_user in calls_by_day.items():
        for user_id, calls in calls_by_user.items():
            row = rows_by_id.get(user_id)
            if row is not None:
                row["per_day"][day] = calls
    for entry in report["unmatched"]:
        entry["days"] = unmatched_days.get(entry["crm_manager_id"], 0)

    # Контрольный запрос за весь период: расхождение с суммой по дням значит,
    # что CRM изменила поведение, и это надо показать, а не замолчать.
    report["crm_period_total"] = crm_period_total
    return _summarize(report)


def day_cell(report, row, day):
    """Что стоит в клетке «сотрудник × день»: ('data'|'no_data'|'off_staff', n).

    Три состояния вместо одного нуля. Ноль на дне без выгрузки — вранье, а ноль
    до приёма на работу — обвинение человека в том, что его ещё не наняли.

    Границу берём из окна плана (staff_from), а не прямо из даты приёма:
    только так набор скрытых клеток совпадает с тем, что вычтено из
    знаменателя, и сумма видимых клеток строки равна её «Факту».
    """
    if day in report["no_data_days"]:
        return "no_data", None
    staff_from = row.get("staff_from")
    if staff_from and day < staff_from:
        return "off_staff", None
    return "data", row["per_day"].get(day, 0)


def has_data(report):
    """Есть ли вообще о чём говорить: хоть один звонок за период."""
    return bool(report.get("total_calls") or report.get("unmatched_calls"))


# ---------------------------------------------------------------------------
# Текст для Telegram
# ---------------------------------------------------------------------------

def _period_label(report):
    date_from = report["date_from"]
    date_to = report["date_to"]
    if date_from == date_to:
        return date_from.strftime("%d.%m.%Y")
    return "%s — %s" % (date_from.strftime("%d.%m.%Y"), date_to.strftime("%d.%m.%Y"))


def _plural(count, one, few, many):
    tail = abs(count) % 100
    if 11 <= tail <= 14:
        return many
    tail %= 10
    if tail == 1:
        return one
    if 2 <= tail <= 4:
        return few
    return many


def _num(value, digits=1):
    """Число по-русски: запятая вместо точки, «—» вместо пустоты."""
    if value is None:
        return "—"
    return ("%.*f" % (digits, value)).replace(".", ",")


def _pct(value):
    return "—" if value is None else "%d%%" % round(value * 100)


def _plan_line(report):
    """Строка про план: за день так и пишем, за период раскрываем умножение."""
    if not report["plan_total"]:
        return None
    if report["days"] == 1:
        return "План: %d %s в день." % (
            report["plan_per_day"],
            _plural(report["plan_per_day"], "звонок", "звонка", "звонков"))
    return "План: %d в день, за %d %s — %d." % (
        report["plan_per_day"], report["days"],
        _plural(report["days"], "день", "дня", "дней"), report["plan_total"])


def _and_more(shown):
    """Хвост «…и ещё N» для списков, которые упёрлись в MAX_ROWS."""
    if len(shown) <= MAX_ROWS:
        return []
    return ["…и ещё %d." % (len(shown) - MAX_ROWS)]


def weakest(report, limit=3):
    """Кто ниже всех — по СРЕДНЕМУ за день, а не по факту.

    У новичка факт мал законно, поэтому «последние по факту» назвали бы не тех.
    Людей без дней в штате не берём вовсе: их не с чем сравнивать.
    """
    return sorted((row for row in report["rows"] if row["staff_days"]),
                  key=lambda row: (row["avg_per_day"],
                                   reg_contest.fold_name(row["name"])))[:limit]


def _city_table(report):
    """Моноширинная таблица по городам — её читают с телефона, не открывая файл.

    Строк ровно столько, сколько городов (сейчас 16), ширина около 24 символов
    — влезает в мобильный экран и отвечает на главный вопрос «какой город
    провалил». Матрицы «сотрудник × день» здесь быть не может: 31 день — это
    больше сотни символов в строке, перенос убьёт таблицу.
    """
    cities = report["cities"]
    if not cities:
        return []
    width = max([len(city["city"]) for city in cities] + [len("Всего")]) + 1
    lines = ["<pre>", html.escape("%-*s %6s %6s" % (width, "Город", "звон.", "ср/дн"))]
    for city in cities:
        lines.append(html.escape("%-*s %6d %6s" % (
            width, city["city"], city["calls"], _num(city["avg_per_day"]))))
    lines.append("-" * (width + 14))
    lines.append(html.escape("%-*s %6d %6s" % (
        width, "Всего", report["total_calls"],
        _num(report["avg_per_manager_day"]))))
    lines.append("</pre>")
    return lines


def render_report(report, only_missing=False):
    """Текст отбивки (HTML для Telegram).

    Три режима, и разница между ними — не украшение, а объём:
      * only_missing — утренняя отбивка: только те, кто план не вытянул;
      * полная сводка за ОДИН день — все менеджеры и черта по планке (к этому
        списку постановщик привык, это его список действий);
      * полная сводка за ПЕРИОД — итоги и таблица по городам: 22 человека на
        31 день в сообщение не влезают и живут в приложенном файле.
    """
    lines = ["<b>Обзвон фронт-офиса за %s</b>" % _period_label(report)]

    total = report["total_calls"]
    summary = "Всего %d %s у %d %s." % (
        total, _plural(total, "звонок", "звонка", "звонков"),
        report["roster_size"],
        _plural(report["roster_size"], "менеджера", "менеджеров", "менеджеров"))
    lines.append("")
    lines.append(summary)

    plan_line = _plan_line(report)
    if plan_line:
        lines.append(plan_line)
        if report["days"] > 1:
            # За месяц «не выполнили 22 из 22» ничего не измеряет — важно,
            # насколько не выполнили.
            lines.append("В среднем %s на человека в день — %s нормы." % (
                _num(report["avg_per_manager_day"]), _pct(report["plan_pct"])))
        lines.append("Не выполнили: %d из %d." % (
            len(report["missing"]), report["roster_size"]))

    lines.append("")
    if only_missing:
        # Сколько человек план выполнило, уже сказано строкой «Не выполнили:
        # N из M» — повторять это списком или отдельной строкой незачем.
        shown = report["missing"]
        for row in shown[:MAX_ROWS]:
            lines.append("%s — %d из %d" % (
                html.escape(row["name"] or "—"), row["calls"], row["plan_total"]))
        lines.extend(_and_more(shown))
    elif report["days"] == 1:
        shown = report["rows"]
        previous_met = None
        for row in shown[:MAX_ROWS]:
            # Одна черта вместо значка у каждой строки: где кончился план,
            # видно сразу, а список не рябит. Рисуем её только там, где
            # граница настоящая — то есть выше есть кто-то с выполненным
            # планом. Если план не вытянул никто, черте не место.
            if previous_met is True and row["met"] is False:
                lines.append("— ниже плана —")
            previous_met = row["met"]
            lines.append("%s — %d" % (html.escape(row["name"] or "—"), row["calls"]))
        lines.extend(_and_more(shown))
    else:
        # За период людей по именам не перечисляем вовсе — только города и
        # худших, так что и обрезать нечего.
        lines.extend(_city_table(report))
        worst = weakest(report)
        if worst:
            lines.append("")
            lines.append("Ниже всех: %s." % ", ".join(
                "%s (%s в день)" % (html.escape(row["name"] or "—"),
                                    _num(row["avg_per_day"]))
                for row in worst))

    if report["unmatched_calls"]:
        lines.append("")
        lines.append("Ещё %d %s у тех, кого нет в реестре отдела: %s%s." % (
            report["unmatched_calls"],
            _plural(report["unmatched_calls"], "звонок", "звонка", "звонков"),
            html.escape(_unmatched_label(report["unmatched"][:5])),
            # Точка ставится форматом, поэтому здесь её быть не должно —
            # иначе строка кончается «и др..».
            " и др" if len(report["unmatched"]) > 5 else ""))

    return "\n".join(lines)


def _unmatched_label(unmatched):
    """«Иванов Иван, id 583, 476» — имена, где они есть, остальные одним хвостом.

    Имён у части менеджеров CRM нет вовсе (живой API отдаёт null), и пять раз
    подряд написать «id» — это шум ради шума.
    """
    named = [r["name"] for r in unmatched if r.get("name")]
    ids = [str(r.get("crm_manager_id")) for r in unmatched if not r.get("name")]
    if ids:
        named.append("id " + ", ".join(ids))
    return ", ".join(named)


def render_no_plan_hint(command="/obzvon_plan"):
    """Что ответить, когда план ещё не задан."""
    return ("Дневной план по обзвону не задан, поэтому утренняя отбивка молчит.\n"
            "Задайте норму командой <code>%s 10</code> — "
            "столько звонков в день на менеджера." % html.escape(command))


def yesterday(today=None):
    """Прошлый день — то, за что отчитывается утренняя отбивка."""
    return (today or date.today()) - timedelta(days=1)


# Форматы, которыми люди пишут дату в чате. «17.08» без года — текущий год:
# спрашивают почти всегда про свежие дни.
_DATE_FORMATS = ("%d.%m.%Y", "%d.%m.%y", "%d.%m", "%Y-%m-%d")


def parse_period(argument, today):
    """«17.08.2026», «17.08», «10.08 12.08» или пусто (вчера) -> (с, по).

    None — разобрать не удалось; вызывающий показывает формат. Перевёрнутый
    период разворачиваем сами: CRM на нём отдаёт HTML вместо данных.
    """
    chunks = (argument or "").replace(",", " ").split()
    if not chunks:
        day = yesterday(today)
        return day, day
    if len(chunks) > 2:
        return None

    days = []
    for chunk in chunks:
        parsed = None
        for fmt in _DATE_FORMATS:
            try:
                value = datetime.strptime(chunk, fmt).date()
            except ValueError:
                continue
            parsed = value.replace(year=today.year) if fmt == "%d.%m" else value
            break
        if parsed is None:
            return None
        days.append(parsed)

    date_from, date_to = days[0], days[-1]
    return (date_to, date_from) if date_to < date_from else (date_from, date_to)


# ---------------------------------------------------------------------------
# Таблица (xlsx)
# ---------------------------------------------------------------------------

# Палитра — из отчёта по аукциону смен, чтобы файл не выглядел чужим: тёмный
# титул, голубой подзаголовок, синяя шапка, зелёная заливка «норма взята».
# Два цвета пришлось взять темнее, чем там: зебра F8FAFC на экране неразличима
# вовсе, а «вне штата» F1F5F9 не отличался от зебры (проверено картинкой).
# Плюс свой жёлтый для дня, за который CRM не ответила.
_TITLE_FILL = "0F172A"
_SUBTITLE_FILL = "E0F2FE"
_HEADER_FILL = "1E3A8A"
_MET_FILL = "DCFCE7"
_OFF_STAFF_FILL = "CBD5E1"
_NO_DATA_FILL = "FEF3C7"
_ZEBRA_FILL = "EFF6FF"
_MUTED_COLOR = "94A3B8"

# Ноль в клетке рисуем прочерком: значение остаётся числом (суммы, фильтры и
# сортировка работают), а сетка не рябит нулями — их тут большинство.
_DAY_FORMAT = '0;-0;"—"'

# Ширина дневной колонки: в 5,2 знака дата ДД.ММ не влезает и Excel рисует
# вместо неё ####. 6,6 — минимум, при котором заголовок читается.
_DAY_WIDTH = 6.6

_THIN = Side(style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

SHEET_NAMES = ("Сводка", "Обзвон", "Города", "Как читать")


def _fill(color):
    return PatternFill(fill_type="solid", fgColor=color)


def _put(ws, row, column, value, number_format=None, fill=None, bold=False,
         color=None, align="center"):
    cell = ws.cell(row=row, column=column, value=value)
    cell.border = _BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = _fill(fill)
    if bold or color:
        cell.font = Font(bold=bold, color=color or "000000")
    return cell


def _widths(ws, widths):
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _subtitle(report, generated_at=None):
    parts = []
    if report["plan_per_day"]:
        parts.append("норма %d %s в день на менеджера" % (
            report["plan_per_day"],
            _plural(report["plan_per_day"], "звонок", "звонка", "звонков")))
    else:
        parts.append("норма не задана")
    parts.append("%d %s" % (report["roster_size"],
                            _plural(report["roster_size"], "менеджер",
                                    "менеджера", "менеджеров")))
    parts.append("%d %s" % (len(report["cities"]),
                            _plural(len(report["cities"]), "город", "города",
                                    "городов")))
    if generated_at:
        parts.append("сформировано %s" % generated_at.strftime("%d.%m.%Y %H:%M"))
    parts.append("CRM yataxi")
    return " · ".join(parts)


def _title_block(ws, report, width, subtitle):
    """Титул и подзаголовок. Вызывать ПОСЛЕ _widths: ширины нужны, чтобы понять,
    влезает ли подзаголовок — в объединённой клетке текст не переливается в
    соседние, а обрезается по границе."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = ws.cell(row=1, column=1,
                   value="Обзвон фронт-офиса · %s" % _period_label(report))
    cell.fill = _fill(_TITLE_FILL)
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.fill = _fill(_SUBTITLE_FILL)
    cell.font = Font(color=_TITLE_FILL, size=10)
    available = sum(
        (ws.column_dimensions[get_column_letter(column)].width or 8.43)
        for column in range(1, width + 1))
    if len(subtitle) > available:
        # Узкий лист (например, отчёт за один день — там всего пять колонок):
        # переносим текст и отдаём строке столько высоты, сколько нужно.
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True)
        ws.row_dimensions[2].height = 14 * max(
            2, -(-len(subtitle) // max(int(available), 1)))
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 6


def _header_row(ws, row, values, no_data_days=()):
    for column, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=column, value=value)
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        if isinstance(value, date) and value in no_data_days:
            # День, за который CRM не ответила: клетки под ним пустые, и
            # заголовок обязан сказать почему — иначе это читается как ноль.
            cell.number_format = "DD.MM"
            cell.fill = _fill(_NO_DATA_FILL)
            cell.font = Font(bold=True, color=_TITLE_FILL)
            continue
        cell.fill = _fill(_HEADER_FILL)
        cell.font = Font(bold=True, color="FFFFFF")
        if isinstance(value, date):
            cell.number_format = "DD.MM"
    ws.row_dimensions[row].height = 30


def _print_wide(ws, title_cols="A:B"):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    # Без этого флага «вписать по ширине» молча не применяется — тот же приём,
    # что в отчёте по эффективности Oktell.
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:4"
    if title_cols:
        ws.print_title_cols = title_cols


def _matrix_columns(report):
    """Колонки листа «Обзвон» до дневных: (заголовок, ширина, как достать, формат, выравнивание).

    За один день дневных колонок нет вовсе — они были бы копией «Факта»,
    поэтому вместо среднего и процентов показываем норму и отклонение.
    """
    plan = report["plan_per_day"]
    columns = [
        ("Город", 16, lambda row: row["city"], None, "left"),
        ("Сотрудник", 28, lambda row: row["name"] or "—", None, "left"),
        ("Факт", 9, lambda row: row["calls"], "0", "center"),
    ]
    if report["days"] == 1:
        if plan:
            # Ноль дней в штате — пустая клетка, а не «норма 0»: у человека,
            # которого в этот день ещё не наняли, нормы не существует.
            columns.append(("Норма", 8, lambda row: row["plan_total"] or None,
                            "0", "center"))
            columns.append(("Отклонение", 13,
                            lambda row: (row["calls"] - row["plan_total"]
                                         if row["plan_total"] else None),
                            "0", "center"))
        return columns
    columns.append(("Ср/день", 11, lambda row: row["avg_per_day"], "0.0", "center"))
    if plan:
        columns.append(("% нормы", 11, lambda row: row["plan_pct"], "0%", "center"))
        columns.append(("Дней с нормой", 12, lambda row: row["days_met"], "0", "center"))
    columns.append(("Дней в штате", 12, lambda row: row["staff_days"], "0", "center"))
    return columns


def _matrix_totals(report):
    """Что писать в строку ИТОГО под каждой колонкой — по её заголовку."""
    plan_sum = report["plan_sum"]
    return {
        "Факт": (report["total_calls"], "0"),
        "Норма": (plan_sum, "0"),
        "Отклонение": ((report["total_calls"] - plan_sum) if plan_sum else None, "0"),
        "Ср/день": (report["avg_per_manager_day"], "0.0"),
        "% нормы": (report["plan_pct"], "0%"),
        "Дней с нормой": (report["cells_met"], "0"),
        "Дней в штате": (report["staff_days_total"], "0"),
    }


def _sheet_matrix(ws, report, generated_at=None):
    """Главный лист: город × сотрудник × дни одной сеткой.

    Город печатается в каждой строке, а не объединённой клеткой: merge ломает
    и автофильтр, и сортировку. Границы городов видно зеброй, включаемой
    группами, — промежуточные итоги по городу были бы дублем строки, ведь
    13 городов из 16 представлены одним человеком.
    """
    columns = _matrix_columns(report)
    day_list = report["day_list"] if report["days"] > 1 else []
    width = len(columns) + len(day_list)
    _widths(ws, [column[1] for column in columns] + [_DAY_WIDTH] * len(day_list))
    _title_block(ws, report, width, _subtitle(report, generated_at))
    _header_row(ws, 4, [column[0] for column in columns] + list(day_list),
                no_data_days=set(report["no_data_days"]))

    rows = sorted(report["rows"],
                  key=lambda row: (row["city"] == CITY_UNKNOWN, row["city"],
                                   -row["calls"],
                                   reg_contest.fold_name(row["name"])))
    plan = report["plan_per_day"]
    index = 5
    zebra = False
    previous_city = None
    for row in rows:
        if previous_city is not None and row["city"] != previous_city:
            zebra = not zebra
        previous_city = row["city"]
        # Кто за весь период ни дня не был в штате — серой строкой целиком:
        # его пустые числа иначе читаются как провал.
        base_fill = (_OFF_STAFF_FILL if not row["staff_days"]
                     else (_ZEBRA_FILL if zebra else None))
        if not day_list and row["met"]:
            # За один день дневных клеток нет, а красить «норма взята» надо:
            # это и есть ежедневный отчёт, который читают чаще остальных.
            base_fill = _MET_FILL
        for column, (_title, _width, getter, number_format, align) in enumerate(
                columns, start=1):
            _put(ws, index, column, getter(row), number_format=number_format,
                 fill=base_fill, align=align)
        for offset, day in enumerate(day_list):
            state, value = day_cell(report, row, day)
            column = len(columns) + 1 + offset
            if state == "data":
                cell = _put(ws, index, column, value, number_format=_DAY_FORMAT,
                            fill=(_MET_FILL if plan and value >= plan else base_fill))
                if not value:
                    cell.font = Font(color=_MUTED_COLOR)
            else:
                _put(ws, index, column, None,
                     fill=_OFF_STAFF_FILL if state == "off_staff" else _NO_DATA_FILL)
        index += 1

    # Пустая строка перед итогом — не для красоты. Excel считает таблицей все
    # смежные строки и расширяет автофильтр на ИТОГО: тогда фильтр по городу
    # скрывает итоги, «ИТОГО» появляется в списке значений колонки «Город», а
    # сортировка по шапке утаскивает строку итога внутрь данных. Разрыв
    # схлопывает диапазон обратно (проверено живым Excel).
    index += 1
    totals = _matrix_totals(report)
    _put(ws, index, 1, "ИТОГО", bold=True, align="left")
    _put(ws, index, 2, "%d %s" % (report["roster_size"],
                                  _plural(report["roster_size"], "менеджер",
                                          "менеджера", "менеджеров")),
         bold=True, align="left")
    for column, (title, _width, _getter, _format, _align) in enumerate(columns, start=1):
        if title in totals:
            value, number_format = totals[title]
            _put(ws, index, column, value, number_format=number_format, bold=True)
    for offset, day in enumerate(day_list):
        _put(ws, index, len(columns) + 1 + offset,
             report["per_day_totals"].get(day), number_format="0", bold=True)

    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(width), index - 2)
    ws.freeze_panes = "C5"
    _print_wide(ws)


def _cities_columns(report):
    """Колонки листа «Города»: (заголовок, ширина, из города, итог, формат).

    Одна спецификация на шапку, строки и ИТОГО — иначе итоги пишутся по
    номерам колонок и разъезжаются, стоит добавить или убрать колонку.
    Ширины с запасом под кнопку автофильтра: иначе заголовки читаются как
    «Менеджер ов» и «% норм».
    """
    columns = [
        ("Город", 20, lambda city: city["city"], "ИТОГО", None),
        ("Менеджеров", 13, lambda city: city["roster_size"],
         report["roster_size"], "0"),
        ("Звонков", 11, lambda city: city["calls"], report["total_calls"], "0"),
        ("Ср/день на человека", 14, lambda city: city["avg_per_day"],
         report["avg_per_manager_day"], "0.0"),
    ]
    if report["plan_per_day"]:
        columns.append(("% нормы", 10, lambda city: city["plan_pct"],
                        report["plan_pct"], "0%"))
        columns.append(("Не выполнили", 14, lambda city: city["missing_count"],
                        len(report["missing"]), "0"))
    return columns


def _sheet_cities(ws, report, generated_at=None):
    """Тот же разрез, свёрнутый до города: 16 строк, по ним и принимают решение."""
    day_list = report["day_list"] if report["days"] > 1 else []
    plan = report["plan_per_day"]
    columns = _cities_columns(report)
    header = [column[0] for column in columns]
    widths = [column[1] for column in columns]
    width = len(header) + len(day_list)
    _widths(ws, widths + [_DAY_WIDTH] * len(day_list))
    _title_block(ws, report, width, _subtitle(report, generated_at))
    _header_row(ws, 4, header + list(day_list),
                no_data_days=set(report["no_data_days"]))

    index = 5
    for city in report["cities"]:
        met_fill = (_MET_FILL if plan and (city["plan_pct"] or 0) >= 1 else None)
        for column, (_title, _width, getter, _total, number_format) in enumerate(
                columns, start=1):
            _put(ws, index, column, getter(city), number_format=number_format,
                 fill=met_fill, align="left" if column == 1 else "center")
        for offset, day in enumerate(day_list):
            column = len(header) + 1 + offset
            if day in report["no_data_days"]:
                _put(ws, index, column, None, fill=_NO_DATA_FILL)
                continue
            cell = _put(ws, index, column, city["per_day"].get(day, 0),
                        number_format=_DAY_FORMAT, fill=met_fill)
            if not cell.value:
                cell.font = Font(color=_MUTED_COLOR)
        index += 1

    # Разрыв перед итогом — по той же причине, что на листе «Обзвон».
    index += 1
    for column, (_title, _width, _getter, total, number_format) in enumerate(
            columns, start=1):
        _put(ws, index, column, total, number_format=number_format, bold=True,
             align="left" if column == 1 else "center")
    for offset, day in enumerate(day_list):
        _put(ws, index, len(header) + 1 + offset,
             report["per_day_totals"].get(day), number_format="0", bold=True)

    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(width), index - 2)
    ws.freeze_panes = "B5"
    _print_wide(ws, title_cols="A:A")


def _sheet_summary(ws, report, generated_at=None):
    """Первый лист: цифры периода, кому подтянуться, звонки вне реестра, качество данных."""
    _widths(ws, [38, 28, 12, 12, 16])
    _title_block(ws, report, 5, _subtitle(report, generated_at))
    row = _section(ws, 4, "ИТОГО")
    row = _pairs(ws, row, _summary_totals(report))

    row = _section(ws, row, "КОМУ ПОДТЯНУТЬСЯ")
    worst = weakest(report, limit=5)
    _header_row(ws, row, ["Город", "Сотрудник", "Звонков", "Ср/день", "% нормы"])
    row += 1
    for item in worst:
        _put(ws, row, 1, item["city"], align="left")
        _put(ws, row, 2, item["name"] or "—", align="left")
        _put(ws, row, 3, item["calls"], number_format="0")
        _put(ws, row, 4, item["avg_per_day"], number_format="0.0")
        _put(ws, row, 5, item["plan_pct"], number_format="0%")
        row += 1
    row += 1

    if report["unmatched"]:
        row = _section(ws, row, "ЗВОНКИ ВНЕ РЕЕСТРА")
        _header_row(ws, row, ["Учётка CRM", "Имя в CRM", "Звонков",
                              "Дней со звонками", ""])
        row += 1
        for entry in report["unmatched"]:
            _put(ws, row, 1, entry.get("crm_manager_id"), align="left")
            _put(ws, row, 2, entry.get("name") or "—", align="left")
            _put(ws, row, 3, entry["calls"], number_format="0")
            _put(ws, row, 4, entry.get("days") or None, number_format="0")
            row += 1
        cell = ws.cell(row=row, column=1,
                       value="Учётки CRM, которые не сопоставились с реестром "
                             "(обычно без имени и логина). В города и в план не "
                             "входят, в «Итого звонков в CRM» входят.")
        cell.font = Font(color=_MUTED_COLOR, size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        row += 2

    row = _section(ws, row, "КАЧЕСТВО ДАННЫХ")
    _pairs(ws, row, _summary_quality(report))


def _summary_totals(report):
    days_with_data = report["days"] - len(report["no_data_days"])
    pairs = [
        ("Менеджеров в реестре", report["roster_size"], "0"),
        ("Звонков по реестру", report["total_calls"], "0"),
        ("Дней в периоде", "%d, из них с выгрузкой %d" % (report["days"],
                                                          days_with_data), None),
    ]
    if report["days"] > 1:
        pairs.append(("Среднее на менеджера в день",
                      report["avg_per_manager_day"], "0.0"))
    if report["plan_per_day"]:
        pairs.append(("Норма звонков в день", report["plan_per_day"], "0"))
        pairs.append(("Норма за период на всех", report["plan_sum"], "0"))
        pairs.append(("Выполнение нормы", report["plan_pct"], "0%"))
        pairs.append(("Не выполнили норму", "%d из %d" % (
            len(report["missing"]), report["roster_size"]), None))
        if report["days"] > 1:
            pairs.append(("Дней-персон с нормой", "%d из %d" % (
                report["cells_met"], report["staff_days_total"]), None))
            pairs.append(("Взяли норму хотя бы раз", "%d из %d" % (
                report["managers_met_once"], report["roster_size"]), None))
    pairs.append(("Менеджеров без звонков за период", report["zero_managers"], "0"))
    pairs.append(("Звонков вне реестра", report["unmatched_calls"], "0"))
    pairs.append(("Итого звонков в CRM",
                  report["total_calls"] + report["unmatched_calls"], "0"))
    return pairs


def _summary_quality(report):
    """Строки, без которых числа читаются неверно: чем сошлись, где дыры."""
    matched = report["matched_by"]
    by_email = matched.get("email", 0)
    by_name = matched.get("name", 0) + matched.get("name_prefix", 0)
    matched_total = report.get("matched_total", by_email + by_name)
    control = report.get("crm_period_total")
    total_crm = report["total_calls"] + report["unmatched_calls"]
    pairs = [
        ("Сопоставлено с реестром", "%d из %d (по почте %d, по ФИО %d)" % (
            matched_total, report["roster_size"], by_email, by_name), None),
        ("Учёток CRM без сопоставления", len(report["unmatched"]), "0"),
        ("Без города в карточке", report["no_city_count"], "0"),
        ("Дней без выгрузки CRM", "%d из %d" % (len(report["no_data_days"]),
                                                report["days"]), None),
        ("Сверка: реестр + вне реестра", total_crm, "0"),
    ]
    if control is not None:
        # Суммы по дням обязаны совпасть с запросом за весь период. Если нет —
        # CRM изменила поведение, и это надо увидеть, а не замолчать.
        pairs.append(("Контрольный запрос за период",
                      "%d — %s" % (control,
                                   "сходится" if control == total_crm
                                   else "РАСХОЖДЕНИЕ %+d" % (total_crm - control)),
                      None))
    return pairs


def _section(ws, row, title, width=5):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = _fill(_HEADER_FILL)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for column in range(2, width + 1):
        ws.cell(row=row, column=column).fill = _fill(_HEADER_FILL)
    return row + 1


def _pairs(ws, row, pairs):
    for label, value, number_format in pairs:
        cell = ws.cell(row=row, column=1, value=label)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell = ws.cell(row=row, column=2, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = Font(bold=True)
        if number_format:
            cell.number_format = number_format
        row += 1
    return row + 1


# Полного листа «КОНТЕКСТ» по рецепту выгрузок «сразу для ИИ» здесь нет
# намеренно: тот рецепт — для файлов, уходящих без автора, а это ежедневный
# управленческий отчёт заказчику. Но ловушки перечислить обязательно, иначе
# нули в сетке прочитают как прогулы.
_NOTES = [
    "Считаем все звонки CRM (total_calls), а не только «Согласия».",
    "Норма — общая на менеджера, меняется в боте командой /obzvon_plan N.",
    "«—» в клетке значит, что звонков не было. Это НЕ «не работал»: графиков "
    "смен у отдела в iCORE нет, выходной и прогул выглядят одинаково.",
    "Серая клетка — в этот день человек ещё не был принят на работу.",
    "Жёлтый заголовок дня — CRM за этот день не ответила. Клетки под ним "
    "пустые, и в норму этот день никому не считается.",
    "Город — из карточки сотрудника в iCORE. CRM географию не присылает вовсе.",
    "«Вне реестра» — учётки CRM, которые не сопоставились с отделом (обычно "
    "без имени и логина). В города и в норму не входят.",
    "Один человек может иметь несколько учёток CRM — их звонки складываются.",
    "% нормы = звонки / (норма × дни в штате). У принятого посреди периода "
    "знаменатель меньше, поэтому его процент честный.",
    "Числа CRM за прошедший день стабилизируются к утру: отчёт, собранный "
    "вечером того же дня, может отличаться от утреннего.",
]


def _sheet_notes(ws, report):
    _widths(ws, [6, 110])
    _title_block(ws, report, 2, "Как читать эту таблицу")
    row = 4
    for number, text in enumerate(_NOTES, start=1):
        cell = ws.cell(row=row, column=1, value=number)
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell = ws.cell(row=row, column=2, value=text)
        cell.alignment = Alignment(horizontal="left", vertical="top",
                                   wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1


def build_workbook(report, generated_at=None):
    """Отчёт таблицей: город × сотрудник × дни. Возвращает байты .xlsx.

    Блокирующая по CPU (openpyxl) — из бота вызывать через executor.
    """
    wb = Workbook()
    summary, matrix, cities, notes = SHEET_NAMES
    wb.active.title = summary
    _sheet_summary(wb.active, report, generated_at)
    _sheet_matrix(wb.create_sheet(matrix), report, generated_at)
    _sheet_cities(wb.create_sheet(cities), report, generated_at)
    _sheet_notes(wb.create_sheet(notes), report)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def report_filename(report):
    """Имя файла латиницей: его видно в списке вложений чата."""
    if report["date_from"] == report["date_to"]:
        return "obzvon_front_office_%s.xlsx" % report["date_from"].isoformat()
    return "obzvon_front_office_%s_%s.xlsx" % (report["date_from"].isoformat(),
                                               report["date_to"].isoformat())


def document_caption(report):
    """Короткая подпись к файлу: что внутри и за какой период."""
    return ("<b>Обзвон фронт-офиса за %s</b>\nТаблица «город × сотрудник × дни». "
            "Листы: %s." % (_period_label(report), " · ".join(SHEET_NAMES)))
