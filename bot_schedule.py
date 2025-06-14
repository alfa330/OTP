import logging
import os
import threading
import asyncio
import requests
from datetime import datetime
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from io import BytesIO
from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.dispatcher import FSMContext
from flask import Flask, request, jsonify
from flask_cors import CORS
from functools import wraps
from openpyxl import load_workbook
import re
import xlsxwriter
import json
from concurrent.futures import ThreadPoolExecutor

# === Логирование =====================================================================================================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# === Переменные окружения =========================================================================================
API_TOKEN = os.getenv('BOT_TOKEN')
admin = int(os.getenv('ADMIN_ID', '0'))
FLASK_API_KEY = os.getenv('FLASK_API_KEY')

if not API_TOKEN:
    raise Exception("Переменная окружения BOT_TOKEN обязательна.")
if not FLASK_API_KEY:
    raise Exception("Переменная окружения FLASK_API_KEY обязательна.")

# === Инициализация бота и диспетчера =============================================================================
bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot=bot, storage=storage)

# === В роли ДБ ==================================================================================================
SVlist = {}
SVlist_lock = threading.Lock()
report_lock = threading.Lock()
executor_pool = ThreadPoolExecutor(max_workers=4)

class SV:
    def __init__(self, name, id):
        self.name = name
        self.id = id
        self.table = ''
        self.calls = {}
        self.lock = threading.Lock()

# === Flask-сервер ===============================================================================================
app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "https://alfa330.github.io"}})

def require_api_key(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        api_key = request.headers.get('X-API-Key')
        if api_key and api_key == FLASK_API_KEY:
            return f(*args, **kwargs)
        else:
            return jsonify({"error": "Invalid or missing API key"}), 401
    return decorated

@app.route('/')
def index():
    return "Bot is alive!", 200

@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        if not data or 'key' not in data:
            return jsonify({"error": "Missing key field"}), 400
        key = data['key']
        with SVlist_lock:
            if key == FLASK_API_KEY:
                return jsonify({"status": "success", "role": "admin", "id": admin, "name": "Admin"})
            for sv_id, sv in SVlist.items():
                if str(sv_id) == str(key):
                    return jsonify({"status": "success", "role": "sv", "id": sv_id, "name": sv.name})
            return jsonify({"error": "Invalid key or ID"}), 401
    except Exception as e:
        logging.error(f"Login error: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/sv_list', methods=['GET'])
@require_api_key
def get_sv_list():
    try:
        with SVlist_lock:
            sv_data = [{"id": sv_id, "name": sv.name, "table": sv.table} for sv_id, sv in SVlist.items()]
        return jsonify({"status": "success", "sv_list": sv_data})
    except Exception as e:
        logging.error(f"Error fetching SV list: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/change_sv_table', methods=['POST'])
@require_api_key
def change_sv_table():
    try:
        data = request.get_json()
        required_fields = ['id', 'table_url']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        sv_id = int(data['id'])
        table_url = data['table_url']
        
        with SVlist_lock:
            sv = SVlist.get(sv_id)
            if not sv:
                return jsonify({"error": "SV not found"}), 404
            
            sheet_name, operators, error = extract_fio_and_links(table_url)
            if error:
                return jsonify({"error": error}), 400
            
            sv.table = table_url
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
            payload = {
                "chat_id": sv_id,
                "text": f"Таблица успешно обновлена администратором <b>успешно✅</b>",
                "parse_mode": "HTML"
            }
            response = requests.post(telegram_url, json=payload, timeout=10)
            if response.status_code != 200:
                error_detail = response.json().get('description', 'Unknown error')
                logging.error(f"Telegram API error: {error_detail}")
                return jsonify({"error": f"Failed to send Telegram message: {error_detail}"}), 500
        
        return jsonify({
            "status": "success",
            "message": "Table updated",
            "sv_name": sv.name,
            "operators": operators
        })
    except Exception as e:
        logging.error(f"Error updating SV table: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/add_sv', methods=['POST'])
@require_api_key
def add_sv():
    try:
        data = request.get_json()
        required_fields = ['name', 'id']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        sv_id = int(data['id'])
        sv_name = data['name']
        with SVlist_lock:
            if sv_id in SVlist:
                return jsonify({"error": "SV with this ID already exists"}), 400
            SVlist[sv_id] = SV(sv_name, sv_id)
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
            payload = {
                "chat_id": sv_id,
                "text": f"Принятие в команду прошло успешно <b>успешно✅</b>",
                "parse_mode": "HTML"
            }
            response = requests.post(telegram_url, json=payload, timeout=10)
            if response.status_code != 200:
                error_detail = response.json().get('description', 'Unknown error')
                logging.error(f"Telegram API error: {error_detail}")
                return jsonify({"error": f"Failed to send Telegram message: {error_detail}"}), 500
        return jsonify({"status": "success", "message": f"SV {sv_name} added"})
    except Exception as e:
        logging.error(f"Error adding SV: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/remove_sv', methods=['POST'])
@require_api_key
def remove_sv():
    try:
        data = request.get_json()
        if not data or 'id' not in data:
            return jsonify({"error": "Missing ID field"}), 400

        sv_id = int(data['id'])
        with SVlist_lock:
            if sv_id not in SVlist:
                return jsonify({"error": "SV not found"}), 404
            sv_name = SVlist[sv_id].name
            del SVlist[sv_id]
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
            payload = {
                "chat_id": sv_id,
                "text": f"Вы были исключены из команды❌",
                "parse_mode": "HTML"
            }
            response = requests.post(telegram_url, json=payload, timeout=10)
            if response.status_code != 200:
                error_detail = response.json().get('description', 'Unknown error')
                logging.error(f"Telegram API error: {error_detail}")
                return jsonify({"error": f"Failed to send Telegram message: {error_detail}"}), 500
        return jsonify({"status": "success", "message": f"SV {sv_name} removed"})
    except Exception as e:
        logging.error(f"Error removing SV: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/sv/data', methods=['GET'])
def get_sv_data():
    try:
        sv_id = request.args.get('id')
        if not sv_id:
            return jsonify({"error": "Missing ID parameter"}), 400
        sv_id = int(sv_id)
        with SVlist_lock:
            sv = SVlist.get(sv_id)
            if not sv:
                return jsonify({"error": "SV not found"}), 404
            sheet_name, operators, error = extract_fio_and_links(sv.table) if sv.table else (None, [], "No table set")
            if error:
                return jsonify({"error": error}), 400
            return jsonify({
                "status": "success",
                "name": sv.name,
                "table": sv.table,
                "operators": operators,
                "calls": sv.calls
            })
    except Exception as e:
        logging.error(f"Error fetching SV data: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/notify_sv', methods=['POST'])
@require_api_key
def notify_supervisor():
    try:
        data = request.get_json()
        required_fields = ['sv_id', 'operator_name']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        sv_id = int(data['sv_id'])
        operator_name = data['operator_name']
        
        current_week = get_current_week_of_month()
        expected_calls = get_expected_calls(current_week)
        
        with SVlist_lock:
            sv = SVlist.get(sv_id)
            if not sv:
                return jsonify({"error": "SV not found"}), 404

            notification_text = (
                f"⚠️ <b>Требуется внимание!</b>\n\n"
                f"У оператора <b>{operator_name}</b> недостаточно прослушанных звонков.\n"
                f"Текущая неделя: {current_week}\n"
                f"Ожидается: {expected_calls} звонков (по 5 в неделю)\n\n"
                f"Пожалуйста, проверьте и прослушайте недостающие звонки."
            )
            
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
            payload = {
                "chat_id": sv_id,
                "text": notification_text,
                "parse_mode": "HTML"
            }
            response = requests.post(telegram_url, json=payload, timeout=10)
            
            if response.status_code != 200:
                error_detail = response.json().get('description', 'Unknown error')
                logging.error(f"Telegram API error: {error_detail}")
                return jsonify({"error": f"Failed to send Telegram message: {error_detail}"}), 500

        return jsonify({"status": "success", "message": f"Notification sent to {sv.name}"}), 200
    except Exception as e:
        logging.error(f"Error in notify_supervisor: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/sv_operators', methods=['GET'])
@require_api_key
def get_sv_operators():
    try:
        sv_id = request.args.get('sv_id')
        if not sv_id:
            return jsonify({"error": "Missing sv_id parameter"}), 400
        
        sv_id = int(sv_id)
        with SVlist_lock:
            sv = SVlist.get(sv_id)
            if not sv:
                return jsonify({"error": "SV not found"}), 404
            
            sheet_name, operators, error = extract_fio_and_links(sv.table) if sv.table else (None, [], "No table set")
            if error:
                return jsonify({"error": error}), 400
            
            current_week = get_current_week_of_month()
            expected_calls = get_expected_calls(current_week)
            
            operators_with_issues = []
            for op in operators:
                call_count = op.get('call_count', 0)
                if call_count in [None, "#DIV/0!"]:
                    call_count = 0
                else:
                    try:
                        call_count = int(call_count)
                    except (ValueError, TypeError):
                        call_count = 0
                
                if call_count < expected_calls:
                    operators_with_issues.append({
                        'name': op.get('name', ''),
                        'call_count': call_count,
                        'expected_calls': expected_calls,
                        'avg_score': op.get('avg_score', None)
                    })
            
            return jsonify({
                "status": "success",
                "operators": operators,
                "operators_with_issues": operators_with_issues,
                "current_week": current_week,
                "expected_calls": expected_calls
            })
    except Exception as e:
        logging.error(f"Error fetching SV operators: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/admin/generate_report', methods=['GET'])
@require_api_key
def handle_generate_report():
    try:
        # Проверяем, не запущена ли уже генерация
        if not report_lock.acquire(blocking=False):
            return jsonify({"error": "Report generation is already in progress"}), 429
            
        # Запускаем в отдельном потоке
        executor_pool.submit(sync_generate_weekly_report)
        return jsonify({"status": "success", "message": "Report generation started"})
    except Exception as e:
        logging.error(f"Error in generate_report: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            report_lock.release()
        except:
            pass

@app.route('/api/sv/update_table', methods=['POST'])
def update_sv_table():
    try:
        data = request.get_json()
        required_fields = ['id', 'table_url']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing required fields"}), 400

        sv_id = int(data['id'])
        table_url = data['table_url']
        with SVlist_lock:
            sv = SVlist.get(sv_id)
            if not sv:
                return jsonify({"error": "SV not found"}), 404
            sheet_name, operators, error = extract_fio_and_links(table_url)
            if error:
                return jsonify({"error": error}), 400
            sv.table = table_url
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
            payload = {
                "chat_id": sv_id,
                "text": f"Таблица успешно обновлена <b>успешно✅</b>",
                "parse_mode": "HTML"
            }
            response = requests.post(telegram_url, json=payload, timeout=10)
            if response.status_code != 200:
                error_detail = response.json().get('description', 'Unknown error')
                logging.error(f"Telegram API error: {error_detail}")
                return jsonify({"error": f"Failed to send Telegram message: {error_detail}"}), 500
        return jsonify({"status": "success", "message": "Table updated", "operators": operators})
    except Exception as e:
        logging.error(f"Error updating SV table: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@app.route('/api/call_evaluation', methods=['POST'])
@require_api_key
def receive_call_evaluation():
    try:
        data = request.get_json()
        required_fields = ['evaluator', 'operator', 'month', 'call_number', 'phone_number', 'score', 'comment']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"error": "Missing or invalid required fields"}), 400

        for field in required_fields:
            if not isinstance(data[field], (str, int, float)):
                return jsonify({"error": f"Invalid type for {field}"}), 400
        
        with SVlist_lock:
            b = 1
            hint = ""
            for sv_id in SVlist:
                sv = SVlist[sv_id]
                if sv.name == data['evaluator']:
                    b = 0
                    with sv.lock:
                        if data['month'] in sv.calls:
                            if data['call_number'] in sv.calls[data['month']]:
                                hint += " - Корректировка оценки!"
                            else:
                                sv.calls[data['month']][data['call_number']] = data
                        else:
                            sv.calls[data['month']] = {}
                            sv.calls[data['month']][data['call_number']] = data
                    break
        
            if b:
                hint += " Оценивающего нет в списке супервайзеров!"
                
        message = (
            f"📞 <b>Оценка звонка</b>\n" 
            f"👤 Оценивающий: <b>{data['evaluator']}</b>\n"
            f"📋 Оператор: <b>{data['operator']}</b>\n"
            f"📄 За месяц: <b>{data['month']}</b>\n"
            f"📞 Звонок: <b>№{data['call_number']}</b>\n"
            f"📱 Номер телефона: <b>{data['phone_number']}</b>\n"
            f"💯 Оценка: <b>{data['score']}</b>\n"
        )
        if data['score'] < 100 and data['comment']:
            message += f"\n💬 Комментарий: \n{data['comment']}\n"
        message += "\n" + hint

        telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendMessage"
        payload = {
            "chat_id": admin,
            "text": message,
            "parse_mode": "HTML"
        }
        response = requests.post(telegram_url, json=payload, timeout=10)
        
        if response.status_code != 200:
            error_detail = response.json().get('description', 'Unknown error')
            logging.error(f"Telegram API error: {error_detail}")
            return jsonify({"error": f"Failed to send Telegram message: {error_detail}"}), 500

        return jsonify({"status": "success"}), 200
    except requests.RequestException as re:
        logging.error(f"HTTP request error: {re}")
        return jsonify({"error": f"Failed to send Telegram message: {str(re)}"}), 500
    except Exception as e:
        logging.error(f"Error processing call evaluation: {e}")
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

def run_flask():
    app.run(host='0.0.0.0', port=int(os.getenv('PORT', 8080)), debug=False, use_reloader=False)

# === Классы =====================================================================================================
class new_sv(StatesGroup):
    svname = State()
    svid = State()

class sv(StatesGroup):
    crtable = State()
    delete = State()
    verify_table = State()
    view_evaluations = State()
    change_table = State()

def get_current_week_of_month():
    today = datetime.now()
    week_number = (today.day - 1) // 7 + 1
    return week_number

def get_expected_calls(week_number):
    return week_number * 5

def get_cancel_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(KeyboardButton('Отмена ❌'))
    return kb

def get_admin_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Редактор СВ📝'))
    kb.insert(KeyboardButton('Оценки📊'))
    return kb

def get_evaluations_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Отчет за месяц📅'))
    kb.add(KeyboardButton('Назад 🔙'))
    return kb

def get_verify_keyboard():
    ikb = InlineKeyboardMarkup(row_width=2)
    ikb.add(
        InlineKeyboardButton("Да ✅", callback_data="verify_yes"),
        InlineKeyboardButton("Нет ❌", callback_data="verify_no")
    )
    return ikb

def get_editor_keyboard():
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(KeyboardButton('Добавить СВ➕'))
    kb.insert(KeyboardButton('Убрать СВ❌'))
    kb.add(KeyboardButton('Изменить таблицу СВ🔄'))
    kb.add(KeyboardButton('Назад 🔙'))
    return kb

@dp.message_handler(regexp='Отмена ❌', state='*')
async def cancel_handler(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is None:
        return
    await state.finish()
    kb = get_admin_keyboard() if message.from_user.id == admin else ReplyKeyboardRemove()
    await bot.send_message(
        chat_id=message.from_user.id,
        text="Действие отменено.",
        parse_mode='HTML',
        reply_markup=kb
    )
    await message.delete()

# === Команды ====================================================================================================
@dp.message_handler(commands=['start'])
async def start_command(message: types.Message):
    await message.delete()
    with SVlist_lock:
        is_sv = message.from_user.id in SVlist
    if message.from_user.id == admin:
        await bot.send_message(
            chat_id=message.from_user.id,
            text="<b>Бобро пожаловать!</b>\nЭто бот для прослушки прослушек.",
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )
    else:
        kb = ReplyKeyboardMarkup(resize_keyboard=True)
        if is_sv:
            kb.add(KeyboardButton('Добавить таблицу📑'))
        await bot.send_message(
            chat_id=message.from_user.id,
            text=f"<b>Бобро пожаловать!</b>\nТвой <b>ID</b> что бы присоединиться к команде:\n\n<pre>{message.from_user.id}</pre>",
            parse_mode='HTML',
            reply_markup=kb
        )

# === Админка ===================================================================================================
@dp.message_handler(regexp='Редактор СВ📝')
async def editor_sv(message: types.Message):
    if message.from_user.id == admin:
        await bot.send_message(
            chat_id=message.from_user.id,
            text='<b>Редактор супервайзеров</b>',
            parse_mode='HTML',
            reply_markup=get_editor_keyboard()
        )
    await message.delete()

@dp.message_handler(regexp='Назад 🔙')
async def back_to_admin(message: types.Message):
    if message.from_user.id == admin:
        await bot.send_message(
            chat_id=message.from_user.id,
            text='<b>Главное меню</b>',
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )
    await message.delete()

@dp.message_handler(regexp='Добавить СВ➕')
async def newSv(message: types.Message):
    if message.from_user.id == admin:
        await bot.send_message(
            text='<b>Добавление СВ, этап</b>: 1 из 2📍\n\nФИО нового СВ🖊',
            chat_id=message.from_user.id,
            parse_mode='HTML',
            reply_markup=get_cancel_keyboard()
        )
        await new_sv.svname.set()
    await message.delete()

@dp.message_handler(state=new_sv.svname)
async def newSVname(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['svname'] = message.text
    await message.answer(
        text=f'Класс, ФИО - <b>{message.text}</b>\n\n<b>Добавление СВ, этап</b>: 2 из 2📍\n\nНапишите <b>ID</b> нового СВ🆔',
        parse_mode='HTML',
        reply_markup=get_cancel_keyboard()
    )
    await new_sv.next()
    await message.delete()

@dp.message_handler(state=new_sv.svid)
async def newSVid(message: types.Message, state: FSMContext):
    try:
        sv_id = int(message.text)
        async with state.proxy() as data:
            data['svid'] = sv_id
        kb_sv = ReplyKeyboardMarkup(resize_keyboard=True)
        kb_sv.add(KeyboardButton('Добавить таблицу📑'))
        await bot.send_message(
            chat_id=sv_id,
            text=f"Принятие в команду прошло успешно <b>успешно✅</b>\n\nОсталось отправить таблицу вашей группы. Нажмите <b>Добавить таблицу📑</b> что бы сделать это.",
            parse_mode='HTML',
            reply_markup=kb_sv
        )
        with SVlist_lock:
            SVlist[sv_id] = SV(data['svname'], sv_id)
        await message.answer(
            text=f'Класс, ID - <b>{message.text}</b>\n\nДобавление СВ прошло <b>успешно✅</b>. Новому супервайзеру осталось лишь отправить таблицу этого месяца👌🏼',
            parse_mode='HTML',
            reply_markup=get_editor_keyboard()
        )
        await state.finish()
    except:
        await message.answer(
            text='Ой, похоже вы отправили не тот <b>ID</b>❌\n\n<b>Пожалуйста повторите попытку!</b>',
            parse_mode='HTML',
            reply_markup=get_cancel_keyboard()
        )
    await message.delete()

@dp.message_handler(regexp='Убрать СВ❌')
async def delSv(message: types.Message):
    if message.from_user.id == admin:
        with SVlist_lock:
            if SVlist:
                await bot.send_message(
                    text='<b>Выберете СВ которого надо исключить🖊</b>',
                    chat_id=admin,
                    parse_mode='HTML',
                    reply_markup=get_cancel_keyboard()
                )
                ikb = InlineKeyboardMarkup(row_width=1)
                for i in SVlist:
                    ikb.insert(InlineKeyboardButton(text=SVlist[i].name, callback_data=str(i)))
                await bot.send_message(
                    text='<b>Лист СВ:</b>',
                    chat_id=admin,
                    parse_mode='HTML',
                    reply_markup=ikb
                )
                await sv.delete.set()
            else:
                await bot.send_message(
                    text='<b>В команде нет СВ🤥</b>',
                    chat_id=admin,
                    parse_mode='HTML',
                    reply_markup=get_editor_keyboard()
                )
    await message.delete()

@dp.callback_query_handler(state=sv.delete)
async def delSVcall(callback: types.CallbackQuery, state: FSMContext):
    with SVlist_lock:
        sv_id = int(callback.data)
        sv = SVlist.get(sv_id)
        if sv:
            del SVlist[sv_id]
            await bot.send_message(
                text=f"Супервайзер <b>{sv.name}</b> успешно исключен из вашей команды✅",
                chat_id=admin,
                parse_mode='HTML',
                reply_markup=get_editor_keyboard()
            )
            await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
            await bot.send_message(
                text=f"Вы были исключены из команды❌",
                chat_id=sv.id,
                parse_mode='HTML',
                reply_markup=ReplyKeyboardRemove()
            )
    await state.finish()

@dp.message_handler(regexp='Изменить таблицу СВ🔄')
async def change_sv_table(message: types.Message):
    if message.from_user.id == admin:
        with SVlist_lock:
            if SVlist:
                await bot.send_message(
                    text='<b>Выберите СВ, чью таблицу нужно изменить🖊</b>',
                    chat_id=admin,
                    parse_mode='HTML',
                    reply_markup=get_cancel_keyboard()
                )
                ikb = InlineKeyboardMarkup(row_width=1)
                for i in SVlist:
                    ikb.insert(InlineKeyboardButton(text=SVlist[i].name, callback_data=f"change_table_{i}"))
                await bot.send_message(
                    text='<b>Лист СВ:</b>',
                    chat_id=admin,
                    parse_mode='HTML',
                    reply_markup=ikb
                )
                await sv.change_table.set()
            else:
                await bot.send_message(
                    text='<b>В команде нет СВ🤥</b>',
                    chat_id=admin,
                    parse_mode='HTML',
                    reply_markup=get_editor_keyboard()
                )
    await message.delete()

@dp.callback_query_handler(state=sv.change_table)
async def select_sv_for_table_change(callback: types.CallbackQuery, state: FSMContext):
    sv_id = int(callback.data.split('_')[2])
    async with state.proxy() as data:
        data['sv_id'] = sv_id
    await bot.send_message(
        chat_id=admin,
        text=f'<b>Отправьте новую таблицу ОКК для {SVlist[sv_id].name}🖊</b>',
        parse_mode='HTML',
        reply_markup=get_cancel_keyboard()
    )
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
    await sv.crtable.set()

@dp.message_handler(regexp='Оценки📊')
async def view_evaluations(message: types.Message):
    if message.from_user.id == admin:
        with SVlist_lock:
            if SVlist:
                await bot.send_message(
                    text='<b>Выберите чьи оценки просмотреть или сгенерируйте отчет</b>',
                    chat_id=admin,
                    parse_mode='HTML',
                    reply_markup=get_evaluations_keyboard()
                )
                ikb = InlineKeyboardMarkup(row_width=1)
                for i in SVlist:
                    ikb.insert(InlineKeyboardButton(text=SVlist[i].name, callback_data=f"eval_{i}"))
                await bot.send_message(
                    text='<b>Лист СВ:</b>',
                    chat_id=admin,
                    parse_mode='HTML',
                    reply_markup=ikb
                )
                await sv.view_evaluations.set()
            else:
                await bot.send_message(
                    text='<b>В команде нет СВ🤥</b>',
                    chat_id=admin,
                    parse_mode='HTML',
                    reply_markup=get_admin_keyboard()
                )
    await message.delete()

@dp.message_handler(regexp='Отчет за месяц📅', state='*')
async def handle_generate_monthly_report(message: types.Message, state: FSMContext):
    if message.from_user.id == admin:
        try:
            await bot.send_message(
                chat_id=admin,
                text="📊 Генерирую отчет за месяц...",
                parse_mode='HTML'
            )
            await generate_weekly_report()
            await bot.send_message(
                chat_id=admin,
                text="Отчет успешно отправлен!",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
        except Exception as e:
            logging.error(f"Ошибка при генерации отчета: {e}")
            await bot.send_message(
                chat_id=admin,
                text=f"❌ Ошибка при генерации отчета: {str(e)}",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
        await state.finish()
    else:
        await bot.send_message(
            chat_id=message.from_user.id,
            text="Ошибка: Эта команда доступна только администратору.",
            parse_mode='HTML'
        )
    await message.delete()

@dp.callback_query_handler(state=sv.view_evaluations)
async def show_evaluations(callback: types.CallbackQuery, state: FSMContext):
    sv_id = int(callback.data.split('_')[1])
    with SVlist_lock:
        sv = SVlist.get(sv_id)
    
    if not sv:
        await bot.send_message(
            chat_id=admin,
            text="Ошибка: СВ не найден.",
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )
        await state.finish()
        return
    
    current_week = get_current_week_of_month()
    expected_calls = get_expected_calls(current_week)
    
    sheet_name, operators, error = extract_fio_and_links(sv.table) if sv.table else (None, [], "Таблица не найдена")
    
    if error:
        await bot.send_message(
            chat_id=admin,
            text=f"Ошибка: {error}",
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )
        await state.finish()
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
        return

    max_name_length = 32
    max_count_length = 5
    max_score_length = 5
    
    message_text = (
        f"<b>Оценки {sv.name} (неделя {current_week}):</b>\n"
        f"<i>Ожидается: {expected_calls} звонков (по 5 в неделю)</i>\n\n"
    )
    
    operators_with_issues = []
    
    if operators:
        for op in operators:
            name = op.get('name', '').strip()
            if not name:
                continue

            display_name = (name[:max_name_length - 1] + '…') if len(name) > max_name_length else name

            call_count = op.get('call_count')
            score = op.get('avg_score')

            if call_count in [None, "#DIV/0!"]:
                call_count = 0
            else:
                try:
                    call_count = int(call_count)
                except (ValueError, TypeError):
                    call_count = 0

            try:
                score_val = float(score) if score else None
            except (ValueError, TypeError):
                score_val = 0

            if call_count < expected_calls:
                if call_count == 0:
                    c_color_icon = "🔴"
                elif call_count < (expected_calls * 0.5):
                    c_color_icon = "🟠"
                else:
                    c_color_icon = "🟡"
                operators_with_issues.append({
                    'name': name,
                    'sv_id': sv_id,
                    'call_count': call_count,
                    'expected': expected_calls
                })
            else:
                c_color_icon = "🟢"

            if score_val is None:
                color_icon = "-"
                score_str = "-"
            elif score_val < 60:
                color_icon = "🔴"
                score_str = f"{score_val:.2f}"
            elif score_val < 90:
                color_icon = "🟡"
                score_str = f"{score_val:.2f}"
            else:
                color_icon = "🟢"
                score_str = f"{score_val:.2f}"

            message_text += f"👤 {display_name}\n"
            message_text += f"   {str(call_count).rjust(max_count_length)} {c_color_icon} звон. | {score_str.rjust(max_score_length)} {color_icon}\n\n"
    else:
        message_text += "Операторов в таблице нет\n"

    ikb = InlineKeyboardMarkup(row_width=1)
    if operators_with_issues:
        message_text += "\n<b>Операторы с недостаточным количеством звонков:</b>"
        for op in operators_with_issues:
            btn_text = f"{op['name']} ({op['call_count']}/{op['expected']})"
            op_name = op['name'].split(" ")[1]
            callback_data = f"notify_sv_{sv_id}_{op_name}"
            ikb.add(InlineKeyboardButton(
                text=btn_text,
                callback_data=callback_data
            ))

    await bot.send_message(
        chat_id=admin,
        text=message_text,
        parse_mode='HTML',
        reply_markup=ikb if operators_with_issues else get_admin_keyboard()
    )
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
    await state.finish()

@dp.callback_query_handler(lambda c: c.data == 'generate_monthly_report', state=sv.view_evaluations)
async def handle_generate_monthly_report(callback: types.CallbackQuery, state: FSMContext):
    try:
        await bot.send_message(
            chat_id=admin,
            text="📊 Генерирую недельный отчет...",
            parse_mode='HTML'
        )
        await generate_weekly_report()
        await bot.answer_callback_query(
            callback.id,
            text="Отчет успешно отправлен!",
            show_alert=False
        )
    except Exception as e:
        logging.error(f"Ошибка при генерации отчета: {e}")
        await bot.answer_callback_query(
            callback.id,
            text="Ошибка при генерации отчета",
            show_alert=True
        )
    await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
    await bot.send_message(
        chat_id=admin,
        text='<b>Главное меню</b>',
        parse_mode='HTML',
        reply_markup=get_admin_keyboard()
    )
    await state.finish()

@dp.callback_query_handler(lambda c: c.data.startswith('notify_sv_'))
async def notify_supervisor(callback: types.CallbackQuery):
    try:
        parts = callback.data.split('_')
        sv_id = int(parts[2])
        operator_name = ' '.join(parts[3:])
        
        current_week = get_current_week_of_month()
        expected_calls = get_expected_calls(current_week)
        
        with SVlist_lock:
            sv = SVlist.get(sv_id)
            if sv:
                notification_text = (
                    f"⚠️ <b>Требуется внимание!</b>\n\n"
                    f"У оператора <b>{operator_name}</b> недостаточно прослушанных звонков.\n"
                    f"Текущая неделя: {current_week}\n"
                    f"Ожидается: {expected_calls} звонков (по 5 в неделю)\n\n"
                    f"Пожалуйста, проверьте и прослушайте недостающие звонки."
                )
                
                await bot.send_message(
                    chat_id=sv_id,
                    text=notification_text,
                    parse_mode='HTML'
                )
                
                await bot.answer_callback_query(
                    callback.id,
                    text=f"Уведомление отправлено СВ {sv.name}",
                    show_alert=False
                )
            else:
                await bot.answer_callback_query(
                    callback.id,
                    text="Ошибка: СВ не найден",
                    show_alert=True
                )
    except Exception as e:
        logging.error(f"Ошибка в notify_supervisor: {e}")
        await bot.answer_callback_query(
            callback.id,
            text="Произошла ошибка при отправке уведомления",
            show_alert=True
        )

def extract_fio_and_links(spreadsheet_url):
    try:
        match = re.search(r"/d/([a-zA-Z0-9_-]+)", spreadsheet_url)
        if not match:
            return None, None, "Ошибка: Неверный формат ссылки на Google Sheets."
        file_id = match.group(1)
        export_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"

        response = requests.get(export_url)
        if response.status_code != 200:
            return None, None, "Ошибка: Не удалось скачать таблицу. Проверьте, доступна ли таблица публично."
        
        temp_file = f"temp_table_{threading.current_thread().ident}.xlsx"
        with open(temp_file, "wb") as f:
            f.write(response.content)

        wb = load_workbook(temp_file, data_only=True)
        ws = wb.worksheets[-1]
        sheet_name = ws.title

        fio_column = None
        score_columns = []
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value is not None:
                    value = str(cell.value).strip()
                    if "ФИО" in value:
                        fio_column = cell.column
                    else:
                        try:
                            num = float(value)
                            if 1 <= int(num) <= 20:
                                score_columns.append(cell.column)
                        except (ValueError, TypeError):
                            continue

        if not fio_column:
            os.remove(temp_file)
            return None, None, "Ошибка: Колонка ФИО не найдена на листе."
        if not score_columns:
            os.remove(temp_file)
            return None, None, "Ошибка: Столбцы с оценками (1-20) не найдены."

        operators = []
        for row in ws.iter_rows(min_row=2):
            fio_cell = row[fio_column - 1]
            if not fio_cell.value:
                break
            scores = []
            for col_idx in score_columns:
                score_cell = row[col_idx - 1]
                try:
                    score = float(score_cell.value) if score_cell.value else None
                    if score is not None:
                        scores.append(score)
                except (ValueError, TypeError):
                    continue
            call_count = len(scores)
            avg_score = sum(scores) / call_count if scores else None
            operator_info = {
                "name": str(fio_cell.value),
                "link": fio_cell.hyperlink.target if fio_cell.hyperlink else None,
                "call_count": call_count,
                "avg_score": avg_score
            }
            operators.append(operator_info)

        os.remove(temp_file)
        return sheet_name, operators, None
    except Exception as e:
        if 'temp_file' in locals() and os.path.exists(temp_file):
            os.remove(temp_file)
        logging.error(f"Ошибка при обработке таблицы: {str(e)}")
        return None, None, f"Ошибка при обработке таблицы: {str(e)}"

@dp.message_handler(regexp='Добавить таблицу📑')
async def crtablee(message: types.Message):
    await bot.send_message(
        text='<b>Отправьте вашу таблицу ОКК🖊</b>',
        chat_id=message.from_user.id,
        parse_mode='HTML',
        reply_markup=get_cancel_keyboard()
    )
    await sv.crtable.set()
    await message.delete()

@dp.message_handler(state=sv.crtable)
async def tableName(message: types.Message, state: FSMContext):
    try:
        user_id = message.from_user.id
        is_admin_changing = await state.get_state() == sv.crtable.state and user_id == admin
        with SVlist_lock:
            if not is_admin_changing and user_id not in SVlist:
                await bot.send_message(
                    chat_id=user_id,
                    text="Ошибка: Вы не зарегистрированы как супервайзер! Пожалуйста, добавьтесь через администратора.",
                    parse_mode="HTML",
                    reply_markup=ReplyKeyboardRemove()
                )
                await state.finish()
                return

        sheet_name, operators, error = extract_fio_and_links(message.text)
        
        if error:
            await bot.send_message(
                chat_id=user_id,
                text=f"{error}\n\n<b>Пожалуйста, отправьте корректную ссылку на таблицу.</b>",
                parse_mode="HTML",
                reply_markup=get_cancel_keyboard()
            )
            return

        async with state.proxy() as data:
            data['table_url'] = message.text
            if is_admin_changing:
                data.setdefault('sv_id', user_id)

        message_text = f"<b>Название листа:</b> {sheet_name}\n\n<b>ФИО операторов:</b>\n"
        for op in operators:
            if op['link']:
                message_text += f"👤 {op['name']} → <a href='{op['link']}'>Ссылка</a>\n"
            else:
                message_text += f"👤 {op['name']} → Ссылка отсутствует\n"
        message_text += "\n<b>Это все ваши операторы?</b>"

        await bot.send_message(
            chat_id=user_id,
            text=message_text,
            parse_mode="HTML",
            reply_markup=get_verify_keyboard(),
            disable_web_page_preview=True
        )
        await sv.verify_table.set()
        await message.delete()
    except Exception as e:
        logging.error(f"Ошибка в tableName: {e}")
        await bot.send_message(
            chat_id=message.from_user.id,
            text="Произошла ошибка при обработке таблицы. Попробуйте снова или свяжитесь с администратором.",
            parse_mode="HTML",
            reply_markup=get_cancel_keyboard()
        )

@dp.callback_query_handler(state=sv.verify_table)
async def verify_table(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        table_url = data.get('table_url')
        sv_id = data.get('sv_id', callback.from_user.id)
    
    if callback.data == "verify_yes":
        with SVlist_lock:
            SVlist[sv_id].table = table_url
        kb = ReplyKeyboardMarkup(resize_keyboard=True)
        kb.add(KeyboardButton('Добавить таблицу📑'))
        reply_markup = kb if sv_id == callback.from_user.id else get_editor_keyboard()
        target_id = callback.from_user.id if sv_id == callback.from_user.id else admin
        await bot.send_message(
            chat_id=target_id,
            text=f'<b>Таблица успешно подтверждена и сохранена для {SVlist[sv_id].name}✅</b>',
            parse_mode='HTML',
            reply_markup=reply_markup
        )
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
        await state.finish()
    elif callback.data == "verify_no":
        await bot.send_message(
            chat_id=callback.from_user.id,
            text=f'<b>Отправьте корректную таблицу ОКК для {SVlist[sv_id].name}🖊</b>',
            parse_mode='HTML',
            reply_markup=get_cancel_keyboard()
        )
        await bot.delete_message(chat_id=callback.from_user.id, message_id=callback.message.message_id)
        await sv.crtable.set()

def sync_generate_weekly_report():
    try:
        if not report_lock.acquire(blocking=False):
            logging.warning("Report generation skipped: already running")
            return False
            
        logging.info("Starting weekly report generation")
        
        current_week = get_current_week_of_month()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D3D3D3',
            'border': 1
        })
        cell_format_int = workbook.add_format({'border': 1, 'num_format': '0'})
        cell_format_float = workbook.add_format({'border': 1, 'num_format': '0.00'})
        
        with SVlist_lock:
            sv_list_copy = dict(SVlist)
        
        for sv_id, sv in sv_list_copy.items():
            if not sv.table:
                continue
            sheet_name, operators, error = extract_fio_and_links(sv.table)
            if error:
                logging.error(f"Error processing table for SV {sv.name}: {error}")
                continue
                
            safe_sheet_name = sv.name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
            worksheet = workbook.add_worksheet(safe_sheet_name)
            headers = ['ФИО', 'Количество звонков', 'Средний балл']
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
                
            for row, op in enumerate(operators, start=1):
                name = op.get('name', '')
                call_count = op.get('call_count', 0)
                avg_score = op.get('avg_score', None)
                
                if call_count in [None, "#DIV/0!"]:
                    call_count = 0
                else:
                    try:
                        call_count = int(call_count)
                    except (ValueError, TypeError):
                        call_count = 0
                
                try:
                    score_val = float(avg_score) if avg_score else None
                except (ValueError, TypeError):
                    score_val = ''
                
                worksheet.write(row, 0, name, cell_format_int)
                worksheet.write(row, 1, call_count, cell_format_int)
                worksheet.write(row, 2, score_val, cell_format_float)
            
            worksheet.set_column('A:A', 30)
            worksheet.set_column('B:B', 20)
            worksheet.set_column('C:C', 15)
        
        workbook.close()
        output.seek(0)
        
        if output.getvalue():
            filename = f"Weekly_Report_Week{current_week}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            telegram_url = f"https://api.telegram.org/bot{API_TOKEN}/sendDocument"
            files = {'document': (filename, output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {'chat_id': admin, 'caption': f"[{now}] 📊 Еженедельный отчет за {current_week}-ю неделю"}
            response = requests.post(telegram_url, files=files, data=data)
            
            if response.status_code != 200:
                logging.error(f"Ошибка отправки отчета: {response.text}")
                return False
            return True
        else:
            logging.error("Отчет пустой")
            return False
    except Exception as e:
        logging.error(f"Critical error in report generation: {e}")
        return False
    finally:
        try:
            report_lock.release()
        except:
            pass

async def generate_weekly_report():
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(executor_pool, sync_generate_weekly_report)

# === Главный запуск =============================================================================================
if __name__ == '__main__':
    # Запускаем Flask в отдельном потоке
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    
    # Настраиваем и запускаем планировщик
    scheduler = AsyncIOScheduler()
    scheduler.add_job(
        generate_weekly_report, 
        CronTrigger(day_of_week='mon', hour=9, minute=0),
        misfire_grace_time=3600
    )
    scheduler.start()
    
    logging.info("🔄 Планировщик запущен")
    logging.info("🤖 Бот запущен")
    
    # Запускаем бота
    executor.start_polling(dp, skip_updates=True)