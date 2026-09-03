# -*- coding: utf-8 -*-
"""Выгрузка журнала раздела «Чаты водителей» в xlsx (задача #271).

Два листа: «Контекст» и «Журнал». Первым идёт именно контекст — журнал живой, и
через неделю по файлу «Журнал чатов.xlsx» уже не восстановить, за какой период
он собран, кто его выгрузил и что в нём отобрано. Тот же порядок, что у выгрузок
«Посылки», «Касания» и «Провайдер ЭДО».

Соглашения те же, что у остальных выгрузок портала:

* телефон водителя и идентификаторы чата лежат ТЕКСТОМ — числом номер теряет
  ведущие нули и уезжает в экспоненту, а 9-значный id заявки Excel показал бы
  как 7,58112E+07; зелёный уголок «Число сохранено как текст» гасится тегом
  `<ignoredErrors>`, сама функция приходит аргументом, потому что живёт в
  монолите, а импортировать монолит из пакета нельзя;
* время события — настоящей датой-временем с форматом, а не строкой, иначе оно
  не отсортируется и по нему не построить сводную;
* книга пишется в потоковом режиме (`write_only=True`): `freeze_panes` и ширины
  задаются ДО первого `append`, объекты стиля создаются один раз модульными
  константами (в цикле openpyxl замедляется восьмикратно).

Подписи видов события ДОСЛОВНО совпадают с интерфейсом раздела
(`src/components/driver_chats/journalMeta.js`). Второй словарь тех же кодов
неизбежен — питон не читает js, — поэтому его сторожит тест: разойдись подписи,
человек получил бы в файле не то слово, что видит на экране.
"""

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
TITLE_FONT = Font(bold=True, size=13)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')

DATETIME_FORMAT = 'DD.MM.YYYY HH:MM:SS'

# Подписи — как в интерфейсе. «Открыл переписку», а не «Скриншот»: система видит
# факт открытия чата, а не нажатие Cmd+Shift+4. Называть одно другим в документе,
# по которому разбирают утечку, нельзя.
KIND_LABELS = {
    'search': 'Искал номер',
    'open': 'Открыл переписку',
    'handoff': 'Передал чат-менеджеру',
}

ROLE_LABELS = {
    'operator': 'Оператор',
    'trainee': 'Стажёр',
    'sv': 'Супервайзер',
    'supervisor': 'Супервайзер',
    'admin': 'Админ',
    'super_admin': 'Супер-админ',
    'trainer': 'Тренер',
}

# key, заголовок, ширина. Порядок — как читают: когда, кто, что сделал, по
# какому водителю, в какой чат, и только затем служебное.
COLUMNS = (
    ('created_at', 'Когда', 19),
    ('user_name', 'Сотрудник', 28),
    ('user_role', 'Должность', 14),
    ('kind', 'Действие', 22),
    ('phone', 'Телефон водителя', 17),
    ('channel_name', 'Таксопарк', 22),
    ('request_id', 'Обращение', 13),
    ('dialog_id', 'Диалог', 13),
    ('messages_count', 'Сообщений', 11),
    ('comment_text', 'Текст комментария', 52),
    ('c2d_message_id', 'ID заметки', 13),
    ('ip_address', 'Адрес', 16),
)

# Колонки, которые обязаны лежать текстом (иначе Excel испортит значение).
_TEXT_COLUMNS = ('phone', 'request_id', 'dialog_id', 'c2d_message_id')


def _clean(value):
    """Excel не принимает управляющие символы — они прилетают из текста чата."""
    if value is None:
        return ''
    return ILLEGAL_CHARACTERS_RE.sub('', str(value))


def _ru_date(value):
    if not value:
        return '—'
    if isinstance(value, str):
        return value
    return value.strftime('%d.%m.%Y')


def export_file_name(period_from=None, period_to=None):
    """Имя файла. Двойник во фронте — journalMeta.js: заголовок
    Content-Disposition до него не доходит (его нет в
    Access-Control-Expose-Headers), поэтому имя собирается с двух сторон и
    обязано совпадать."""
    left, right = _ru_date(period_from), _ru_date(period_to)
    if left == '—' and right == '—':
        return 'Журнал чатов водителей.xlsx'
    if left == right:
        return 'Журнал чатов водителей %s.xlsx' % left
    return 'Журнал чатов водителей %s — %s.xlsx' % (left, right)


def _setup(sheet, titles, widths, freeze='A2'):
    """Шапка, ширины и закрепление. Всё ДО первой строки данных — в потоковом
    режиме выставленное после в файл уже не попадёт, причём молча."""
    if freeze:
        sheet.freeze_panes = freeze
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    header = []
    for title in titles:
        cell = WriteOnlyCell(sheet, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        header.append(cell)
    sheet.append(header)


def _context_sheet(sheet, *, period_from, period_to, generated_at, generated_by,
                   filters_note, total, truncated):
    sheet.column_dimensions['A'].width = 34
    sheet.column_dimensions['B'].width = 86
    title = WriteOnlyCell(sheet, value='Журнал раздела «Чаты водителей»')
    title.font = TITLE_FONT
    sheet.append([title])
    sheet.append([])

    def line(label, value):
        left = WriteOnlyCell(sheet, value=label)
        left.font = Font(bold=True)
        right = WriteOnlyCell(sheet, value=_clean(value))
        right.alignment = WRAP_TOP
        sheet.append([left, right])

    line('Период', '%s — %s' % (_ru_date(period_from), _ru_date(period_to)))
    line('Выгрузил', generated_by or '—')
    line('Собран', generated_at.strftime('%d.%m.%Y %H:%M') if generated_at else '—')
    line('Отбор', filters_note or 'без дополнительных фильтров')
    line('Строк в файле', str(total if total is not None else 0))
    if truncated:
        line('Внимание', 'Показаны не все строки: выгрузка ограничена сверху. '
                         'Сузьте период или фильтры, чтобы получить остаток.')
    sheet.append([])
    line('Что означает «Открыл переписку»',
         'Сотрудник открыл чат водителя на экране. Снимок экрана делается '
         'средствами операционной системы и системе не виден — журнал '
         'фиксирует доступ к переписке, а не сам факт снимка.')
    line('Что означает «Передал чат-менеджеру»',
         'Сотрудник нажал «Передан»: в чат водителя ушёл внутренний '
         'комментарий Chat2Desk. Водитель его не видит, чат-менеджер видит '
         'у себя в рабочем окне. Отозвать такой комментарий нельзя.')


def build_workbook(events, *, period_from=None, period_to=None, generated_at=None,
                   generated_by='', filters_note='', truncated=False,
                   text_warning_patch=None):
    """events — строки журнала в том же виде, в каком их отдаёт раздел.

    Возвращает (BytesIO, число строк).
    """
    rows = list(events or [])
    workbook = Workbook(write_only=True)

    _context_sheet(
        workbook.create_sheet('Контекст'),
        period_from=period_from, period_to=period_to, generated_at=generated_at,
        generated_by=generated_by, filters_note=filters_note,
        total=len(rows), truncated=truncated)

    sheet = workbook.create_sheet('Журнал')
    _setup(sheet, [title for _key, title, _w in COLUMNS],
           [width for _key, _t, width in COLUMNS])

    for item in rows:
        line = []
        for key, _title, _width in COLUMNS:
            value = item.get(key)
            if key == 'created_at' and value:
                try:
                    parsed = datetime.fromisoformat(str(value))
                except ValueError:
                    parsed = None
                cell = WriteOnlyCell(sheet, value=parsed or _clean(value))
                if parsed:
                    cell.number_format = DATETIME_FORMAT
                line.append(cell)
                continue
            if key == 'kind':
                line.append(WriteOnlyCell(sheet, value=KIND_LABELS.get(value, value or '')))
                continue
            if key == 'user_role':
                line.append(WriteOnlyCell(sheet, value=ROLE_LABELS.get(value, value or '')))
                continue
            if key == 'messages_count':
                line.append(WriteOnlyCell(sheet, value=value if value is not None else ''))
                continue
            if key in _TEXT_COLUMNS:
                line.append(WriteOnlyCell(sheet, value=_clean(value) if value is not None else ''))
                continue
            cell = WriteOnlyCell(sheet, value=_clean(value))
            if key == 'comment_text':
                cell.alignment = WRAP_TOP
            line.append(cell)
        sheet.append(line)

    if rows:
        sheet.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(COLUMNS)), len(rows) + 1)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    # Гасим «Число сохранено как текст» на колонках, которые ОБЯЗАНЫ быть
    # текстом. Диапазон собираем по буквам колонок, а не по всей ширине листа:
    # у «Сообщений» значок должен остаться, там действительно число.
    if text_warning_patch and rows:
        letters = [get_column_letter(index + 1)
                   for index, (key, _t, _w) in enumerate(COLUMNS) if key in _TEXT_COLUMNS]
        sqref = ' '.join('%s2:%s%d' % (letter, letter, len(rows) + 1) for letter in letters)
        try:
            stream = text_warning_patch(stream, sqref, sheet_path='xl/worksheets/sheet2.xml')
        except Exception:  # noqa: BLE001
            # Косметика. Книга уже собрана — ронять из-за уголка выгрузку нельзя.
            stream.seek(0)

    return stream, len(rows)
